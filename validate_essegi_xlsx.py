"""
validate_essegi_xlsx.py
Sprint 0 Kit Preparation: valida il tracciato dei file XLSX generati da Essegi
(lista prelievo "Reels traceability") prima dell'implementazione del parser.

Tracciato atteso (spec docs/PlanRespect_KitPreparation_Spec_v1.2.md §5.1.1):
  - intestazione (prime righe): ordini in formato compatto, es. 'PR554/553/552/551'
  - colonna A: Unique Number / Reel Code (es. HU000004744, HU000013367_01, 000229)
  - colonna B: MaterialCode (= Components.ComponentCode)
  - colonna E: Quantity

Uso:
  .venv\\Scripts\\python.exe validate_essegi_xlsx.py [file.xlsx ...] [--db]
  Senza argomenti valida tutti i .xlsx in T:\\KITTING.
  --db: verifica anche ordini e codici materiale sul DB Traceability_RS.
"""
import sys, io, os, re, argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

KITTING_DIR = r"T:\KITTING"

# Pattern osservati sui campioni reali (PDF "Reels traceability" del 12.06.2026)
RE_UNIQUE_HU = re.compile(r'^HU\d{9}(_\d{2})?$')
RE_UNIQUE_SHORT = re.compile(r'^\d{6}$')
RE_ORDERS_COMPACT = re.compile(r'^PR\d+(?:/\d+)+$|^PR\d+$')

ORDER_TOTAL_LEN = 9  # 'PR' + zeri di padding + numero = 9 caratteri


def normalize_order(num: str) -> str:
    """'554' -> 'PR0000554' (lunghezza totale 9, PR incluso)."""
    return 'PR' + num.zfill(ORDER_TOTAL_LEN - 2)


def expand_orders(compact: str):
    """'PR554/553/552/551' -> ['PR0000554', 'PR0000553', 'PR0000552', 'PR0000551']"""
    parts = compact.replace(' ', '').split('/')
    nums = [parts[0][2:]] + parts[1:]   # il primo segmento include 'PR'
    return [normalize_order(n) for n in nums if n.isdigit()]


def find_orders_header(ws, max_rows=10):
    """Cerca la riga ordini nelle prime righe del foglio (qualsiasi colonna)."""
    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        for cell in row:
            v = str(cell.value).strip() if cell.value is not None else ''
            if RE_ORDERS_COMPACT.match(v):
                return v, cell.row
    return None, None


def validate_file(path, db_cursor=None):
    from openpyxl import load_workbook
    print(f"\n=== {os.path.basename(path)} ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    ok = True

    # 1. Intestazione ordini
    compact, header_row = find_orders_header(ws)
    if not compact:
        print("  ERRORE: nessuna riga ordini 'PRnnn/nnn/...' trovata nelle prime 10 righe")
        ok = False
        orders = []
    else:
        orders = expand_orders(compact)
        print(f"  Ordini (riga {header_row}): {compact}  ->  {', '.join(orders)}")

    # 2. Trova la riga di intestazione colonne ('REEL CODE') e ricava da essa
    #    le posizioni delle colonne: il report Essegi puo' essere spostato
    #    (osservato: dati da col. B nel file reale, da col. A nella spec)
    data_start = (header_row or 0) + 1
    col_unique, col_code, col_qty = 0, 1, 4   # fallback: A, B, E (spec)
    for row in ws.iter_rows(min_row=1, max_row=20):
        headers = {str(c.value).strip().upper(): c.column - 1
                   for c in row if c.value is not None}
        if 'REEL CODE' in headers:
            data_start = row[0].row + 1
            col_unique = headers['REEL CODE']
            col_code = headers.get('ITEM CODE', col_unique + 1)
            col_qty = headers.get('QT', headers.get('QUANTITY', col_unique + 3))
            print(f"  Colonne rilevate: unique={chr(65+col_unique)}, "
                  f"code={chr(65+col_code)}, qty={chr(65+col_qty)} (riga {row[0].row})")
            break
    else:
        print("  ATTENZIONE: intestazione 'REEL CODE' non trovata, uso colonne A/B/E da spec")

    def cell(row, idx):
        return row[idx].value if len(row) > idx else None

    # 3. Righe dati
    n_rows = 0
    bad_unique, bad_qty, empty_code = [], [], []
    materials = set()
    for row in ws.iter_rows(min_row=data_start):
        a = str(cell(row, col_unique)).strip() if cell(row, col_unique) is not None else ''
        if not a:
            continue
        # salta eventuali ripetizioni di testata (cambio pagina nel report)
        if not (RE_UNIQUE_HU.match(a) or RE_UNIQUE_SHORT.match(a)):
            if a.upper() not in ('REEL CODE',) and not RE_ORDERS_COMPACT.match(a):
                bad_unique.append((row[0].row, a))
            continue
        n_rows += 1
        b = str(cell(row, col_code)).strip() if cell(row, col_code) is not None else ''
        if not b:
            empty_code.append(row[0].row)
        else:
            materials.add(b)
        e = cell(row, col_qty)
        try:
            q = float(str(e).replace('.', '').replace(',', '.')) if isinstance(e, str) else float(e)
            if q < 0:
                raise ValueError
        except (TypeError, ValueError):
            bad_qty.append((row[0].row, e))

    print(f"  Righe materiale valide: {n_rows} | codici materiale distinti: {len(materials)}")
    if bad_unique:
        ok = False
        print(f"  ERRORE: {len(bad_unique)} valori col.A non riconosciuti come unique number, es.: "
              f"{bad_unique[:5]}")
    if empty_code:
        ok = False
        print(f"  ERRORE: {len(empty_code)} righe senza MaterialCode (col.B), righe: {empty_code[:10]}")
    if bad_qty:
        ok = False
        print(f"  ERRORE: {len(bad_qty)} quantita' non valide (col.E), es.: {bad_qty[:5]}")
    if n_rows == 0:
        ok = False
        print("  ERRORE: nessuna riga dati riconosciuta")

    # 4. Validazione DB opzionale
    if db_cursor is not None and orders:
        placeholders = ','.join('?' * len(orders))
        db_cursor.execute(
            f"SELECT OrderNumber FROM Traceability_RS.dbo.Orders WHERE OrderNumber IN ({placeholders})",
            orders
        )
        found = {r[0] for r in db_cursor.fetchall()}
        missing = [o for o in orders if o not in found]
        print(f"  DB Orders: {len(found)}/{len(orders)} ordini trovati"
              + (f" — MANCANTI: {missing}" if missing else ""))
        if missing:
            ok = False

        if materials:
            mat_list = sorted(materials)
            found_mat = set()
            for i in range(0, len(mat_list), 500):
                chunk = mat_list[i:i + 500]
                ph = ','.join('?' * len(chunk))
                db_cursor.execute(
                    f"SELECT ComponentCode FROM Traceability_RS.dbo.Components "
                    f"WHERE ComponentCode IN ({ph})",
                    chunk
                )
                found_mat.update(r[0] for r in db_cursor.fetchall())
            missing_mat = [m for m in mat_list if m not in found_mat]
            # I codici a DB possono avere un suffisso '|n' (variante) assente
            # nel report Essegi: riprova il match come prefisso 'codice|%'
            suffix_matched = []
            still_missing = []
            for m in missing_mat:
                db_cursor.execute(
                    "SELECT TOP 1 ComponentCode FROM Traceability_RS.dbo.Components "
                    "WHERE ComponentCode LIKE ? + '|%'", (m,)
                )
                r = db_cursor.fetchone()
                if r:
                    suffix_matched.append((m, r[0]))
                else:
                    still_missing.append(m)
            print(f"  DB Components: {len(found_mat)}/{len(mat_list)} codici esatti"
                  + (f", {len(suffix_matched)} con suffisso variante (es. {suffix_matched[:3]})"
                     if suffix_matched else ""))
            if still_missing:
                ok = False
                print(f"  ERRORE: codici NON trovati a DB (primi 10): {still_missing[:10]}")

    print(f"  ESITO: {'TRACCIATO OK' if ok else 'NON CONFORME'}")
    return ok


def main():
    parser = argparse.ArgumentParser(description='Valida tracciato XLSX lista prelievo Essegi')
    parser.add_argument('files', nargs='*', help='file .xlsx da validare (default: tutti in T:\\KITTING)')
    parser.add_argument('--db', action='store_true', help='valida ordini/codici anche sul DB')
    args = parser.parse_args()

    files = args.files
    if not files:
        if not os.path.isdir(KITTING_DIR):
            print(f"ERRORE: {KITTING_DIR} non raggiungibile e nessun file indicato")
            sys.exit(2)
        files = [os.path.join(KITTING_DIR, f) for f in os.listdir(KITTING_DIR)
                 if f.lower().endswith('.xlsx')]
        if not files:
            print(f"Nessun file .xlsx in {KITTING_DIR}: indicare un file come argomento")
            sys.exit(2)

    db_cursor = None
    if args.db:
        import pyodbc
        from config_manager import ConfigManager
        cfg = ConfigManager(key_file='encryption_key.key', config_file='db_config.enc').load_config()
        conn = pyodbc.connect(
            f"DRIVER={cfg['driver']};SERVER={cfg['server']};DATABASE={cfg['database']};"
            f"UID={cfg['username']};PWD={cfg['password']};MARS_Connection=Yes;TrustServerCertificate=Yes"
        )
        db_cursor = conn.cursor()
        print("Validazione DB attiva (Traceability_RS).")

    results = [validate_file(f, db_cursor) for f in files]
    print(f"\nTotale: {sum(results)}/{len(results)} file conformi")
    sys.exit(0 if all(results) else 1)


if __name__ == '__main__':
    main()
