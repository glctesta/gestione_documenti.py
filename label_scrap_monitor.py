# -*- coding: utf-8 -*-
"""
label_scrap_monitor.py — stampa+email di fine turno per gli scarti etichette.

Attivo SOLO sul PC designato (labelscrap_print_host.json). Poll ogni 60s (stile
ShiftHandoverMonitor). A 15:15 / 23:15 / 07:15 (15 min prima della fine dei turni
15:30 / 23:30 / 07:30) stampa i riepiloghi per operatore delle dichiarazioni del
turno non ancora stampate (Printed IS NULL), le segna stampate, e invia una email
(Excel + PDF) ai destinatari 'sys_email_labelScrap' con dedup cross-PC.
Il venerdì pomeriggio invia il resoconto settimanale (rolling mese + YTD).
"""
import os
import socket
import logging
import tempfile
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)

POLL_INTERVAL_MS = 60_000

# finestra (h, m_min, m_max) -> etichetta turno che finisce (valore in labelscrap.Shift)
SHIFT_END_WINDOWS = [
    (15, 13, 17, '07:30'),   # 15:15 → fine turno mattutino
    (23, 13, 17, '15:30'),   # 23:15 → fine turno pomeridiano
    (7, 13, 17, '23:30'),    # 07:15 → fine turno notturno
]
FRIDAY_HOUR, FRIDAY_MMIN, FRIDAY_MMAX = 14, 0, 6   # venerdì ~14:00


def _claim_send_slot(conn, setting_key):
    """True se questo PC vince la corsa all'invio (INSERT WHERE NOT EXISTS atomico)."""
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO traceability_rs.dbo.settings (atribute, [value])
            SELECT ?, ?
            WHERE NOT EXISTS (
                SELECT 1 FROM traceability_rs.dbo.settings
                WITH (UPDLOCK, HOLDLOCK) WHERE atribute = ?)
        """, (setting_key, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), setting_key))
        claimed = cur.rowcount > 0
        conn.commit()
        cur.close()
        return claimed
    except Exception as e:
        logger.error(f"LabelScrapMonitor _claim_send_slot: {e}")
        return False


class LabelScrapMonitor:
    def __init__(self, master, db, lang):
        self.master = master
        self.db = db
        self.lang = lang
        self.hostname = socket.gethostname()
        self._running = True
        self._done = set()   # chiavi finestra già processate in memoria
        self._poll()

    def stop(self):
        self._running = False

    def _poll(self):
        if not self._running:
            return
        try:
            self._check()
        except Exception as e:
            logger.error(f"LabelScrapMonitor poll error: {e}", exc_info=True)
        finally:
            if self._running:
                self.master.after(POLL_INTERVAL_MS, self._poll)

    def _check(self):
        # ricontrolla il marker a runtime: se rimosso, si ferma
        try:
            from label_scrap_workstation_config import is_labelscrap_print_workstation
            if not is_labelscrap_print_workstation():
                self._running = False
                logger.info("LabelScrapMonitor: marker rimosso — monitor fermato")
                return
        except Exception:
            return

        now = datetime.now()
        h, m = now.hour, now.minute

        for wh, mlo, mhi, shift_label in SHIFT_END_WINDOWS:
            if h == wh and mlo <= m <= mhi:
                key = f"shift_{now:%Y%m%d}_{wh}"
                if key not in self._done:
                    self._process_shift_end(shift_label, now)
                    self._done.add(key)

        # venerdì pomeriggio: resoconto settimanale
        if now.weekday() == 4 and h == FRIDAY_HOUR and FRIDAY_MMIN <= m <= FRIDAY_MMAX:
            iso = now.isocalendar()
            wkey = f"week_{iso[0]}W{iso[1]:02d}"
            if wkey not in self._done:
                self._process_weekly(now)
                self._done.add(wkey)

    # ── Fine turno ────────────────────────────────────────────────────────
    def _fetch_shift_rows(self, shift_label):
        """Righe del turno non ancora stampate (ultime ~12h), con motivo."""
        cur = self.db.conn.cursor()
        cur.execute("""
            SELECT ls.LabelScrapId, ls.Operator, ls.LabelCode, r.Reason, ls.Category,
                   ls.ScrapDate, ls.DateIn,
                   ISNULL(ls.Qty, 1) AS Qty, ISNULL(ls.CodiceMateriale, '') AS CodiceMateriale
            FROM traceability_rs.dbo.labelscrap ls
            INNER JOIN traceability_rs.dbo.LabelScrapReasons r
                ON r.LabelScrapReasonId = ls.LabelScrapReasonId
            WHERE ls.Shift = ? AND ls.Printed IS NULL
              AND ls.DateIn >= DATEADD(HOUR, -12, GETDATE())
            ORDER BY ls.Operator, ls.DateIn
        """, (shift_label,))
        rows = cur.fetchall()
        cur.close()
        return rows

    def _process_shift_end(self, shift_label, now):
        rows = self._fetch_shift_rows(shift_label)
        if not rows:
            logger.info(f"LabelScrapMonitor: nessuna dichiarazione non stampata per turno {shift_label}")
            return
        logger.info(f"LabelScrapMonitor: fine turno {shift_label} — {len(rows)} righe, stampa per operatore")

        # 1) stampa un PDF per ogni operatore (salta chi ha già stampato = non è nella lista)
        import label_scrap_pdf
        wr = label_scrap_pdf.get_warehouse_responsible(self.db.conn)
        by_op = {}
        for r in rows:
            by_op.setdefault(r.Operator, []).append(r)
        printed_ids = []
        for operator, orows in by_op.items():
            try:
                fd, path = tempfile.mkstemp(suffix='.pdf', prefix=f'ScartiEtichette_{operator[:16]}_')
                os.close(fd)
                pdf_rows = [{'label': x.LabelCode, 'reason': x.Reason, 'category': x.Category,
                             'time': x.DateIn.strftime('%H:%M:%S') if x.DateIn else ''} for x in orows]
                label_scrap_pdf.generate_declaration_pdf(path, operator, now.date(), pdf_rows,
                                                         warehouse_responsible=wr)
                label_scrap_pdf.print_pdf(path)
                printed_ids.extend([x.LabelScrapId for x in orows])
            except Exception as e:
                logger.error(f"LabelScrapMonitor: stampa operatore {operator}: {e}", exc_info=True)

        # 2) segna stampate
        if printed_ids:
            try:
                cur = self.db.conn.cursor()
                ph = ','.join(['?'] * len(printed_ids))
                cur.execute(f"UPDATE traceability_rs.dbo.labelscrap SET Printed=GETDATE() "
                            f"WHERE LabelScrapId IN ({ph})", printed_ids)
                self.db.conn.commit()
                cur.close()
            except Exception as e:
                logger.error(f"LabelScrapMonitor: mark printed: {e}", exc_info=True)

        # 3) email (dedup cross-PC)
        detail = [(r.ScrapDate, r.Operator, r.LabelCode, r.Reason, r.Category, shift_label,
                   int(r.Qty or 1), r.CodiceMateriale) for r in rows]
        key = f"SentLabelScrap_{now:%Y%m%d}_{shift_label.replace(':', '')}"
        if _claim_send_slot(self.db.conn, key):
            self._send_email(
                subject=f"Scarti etichette — fine turno {shift_label} del {now:%d/%m/%Y}",
                intro=f"Riepilogo scarti etichette del turno terminato ({shift_label}).",
                detail=detail, date_from=now.date(), date_to=now.date())

    # ── Settimanale (venerdì) ─────────────────────────────────────────────
    def _process_weekly(self, now):
        key = f"SentLabelScrapWeek_{now.isocalendar()[0]}W{now.isocalendar()[1]:02d}"
        if not _claim_send_slot(self.db.conn, key):
            return
        monday = now.date() - timedelta(days=now.weekday())
        month_start = now.date().replace(day=1)
        year_start = now.date().replace(month=1, day=1)
        cur = self.db.conn.cursor()
        cur.execute("""
            SELECT ls.ScrapDate, ls.Operator, ls.LabelCode, r.Reason, ls.Category, ls.Shift,
                   ISNULL(ls.Qty, 1) AS Qty, ISNULL(ls.CodiceMateriale, '') AS CodiceMateriale
            FROM traceability_rs.dbo.labelscrap ls
            INNER JOIN traceability_rs.dbo.LabelScrapReasons r ON r.LabelScrapReasonId = ls.LabelScrapReasonId
            WHERE ls.ScrapDate BETWEEN ? AND ? ORDER BY ls.ScrapDate, ls.Operator
        """, (year_start, now.date()))
        allrows = cur.fetchall()
        cur.close()
        wk = [r for r in allrows if r.ScrapDate >= monday]
        mo = [r for r in allrows if r.ScrapDate >= month_start]

        def qty(rr):
            return sum(int(x.Qty or 1) for x in rr)

        # Etichette, non righe: una riga puo' valerne 500.
        intro = (f"Resoconto settimanale scarti etichette (etichette, non righe).\n"
                 f"Settimana (dal {monday:%d/%m}): {qty(wk)}  |  "
                 f"Mese (dal {month_start:%d/%m}): {qty(mo)}  |  YTD: {qty(allrows)}")

        def tup(r):
            return (r.ScrapDate, r.Operator, r.LabelCode, r.Reason, r.Category, r.Shift,
                    int(r.Qty or 1), r.CodiceMateriale)

        self._send_email(
            subject=f"Scarti etichette — resoconto settimanale {now:%d/%m/%Y}",
            intro=intro, detail=[tup(r) for r in wk], date_from=monday, date_to=now.date(),
            extra_rows_for_excel=[tup(r) for r in allrows])

    # ── Email + allegati ──────────────────────────────────────────────────
    def _send_email(self, subject, intro, detail, date_from, date_to, extra_rows_for_excel=None):
        try:
            import utils
            recipients = utils.get_email_recipients(self.db.conn, 'sys_email_labelScrap')
        except Exception as e:
            logger.error(f"LabelScrapMonitor: destinatari: {e}")
            recipients = []
        if not recipients:
            logger.warning("LabelScrapMonitor: nessun destinatario 'sys_email_labelScrap' — email non inviata")
            return

        # Riepiloghi in ETICHETTE (somma delle quantita'), non in righe
        def agg(idx):
            acc = {}
            for d in detail:
                acc[d[idx]] = acc.get(d[idx], 0) + int(d[6])
            return sorted(acc.items(), key=lambda kv: (-kv[1], str(kv[0])))

        by_reason, by_category, by_operator = agg(3), agg(4), agg(1)
        total_qty = sum(int(d[6]) for d in detail)

        attachments = []
        tmp = []
        try:
            import label_scrap_pdf
            fd, pdf_path = tempfile.mkstemp(suffix='.pdf', prefix='ScartiEtichette_')
            os.close(fd)
            wr = label_scrap_pdf.get_warehouse_responsible(self.db.conn)
            # Scarti a fronte del prelevato dal magazzino nello stesso periodo
            try:
                vs_rows, vs_totals = label_scrap_pdf.fetch_labels_vs_withdrawn(
                    self.db.conn, date_from, date_to)
            except Exception as e:
                logger.warning(f"LabelScrapMonitor: confronto con il prelevato non disponibile: {e}")
                vs_rows, vs_totals = [], (0, 0, 0.0)
            label_scrap_pdf.generate_report_pdf(pdf_path, date_from, date_to, None, detail,
                                                by_reason, by_category, by_operator,
                                                warehouse_responsible=wr,
                                                vs_rows=vs_rows, vs_totals=vs_totals)
            attachments.append(pdf_path); tmp.append(pdf_path)

            xls_path = self._build_excel(extra_rows_for_excel or detail)
            if xls_path:
                attachments.append(xls_path); tmp.append(xls_path)

            body = self._html_body(intro, by_reason, by_category, by_operator,
                                   len(detail), total_qty, vs_totals)
            import utils
            utils.send_email(recipients, subject, body, is_html=True, attachments=attachments or None)
            logger.info(f"LabelScrapMonitor: email inviata a {recipients}")
        except Exception as e:
            logger.error(f"LabelScrapMonitor: invio email: {e}", exc_info=True)
        finally:
            for p in tmp:
                try:
                    os.remove(p)
                except Exception:
                    pass

    def _build_excel(self, rows):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return None
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scarti Etichette'
        HF = PatternFill('solid', fgColor='1F3864')
        HFONT = Font(bold=True, color='FFFFFF')
        THIN = Border(*(Side(style='thin'),) * 4)
        for c, h in enumerate(['Data', 'Operatore', 'Etichetta', 'Q.tà', 'Materiale',
                               'Motivo', 'Categoria', 'Turno'], 1):
            cell = ws.cell(1, c, h)
            cell.fill, cell.font, cell.alignment, cell.border = HF, HFONT, Alignment(horizontal='center'), THIN
        for ri, r in enumerate(rows, 2):
            d = r[0].strftime('%d/%m/%Y') if hasattr(r[0], 'strftime') else str(r[0])
            for ci, v in enumerate((d, r[1], r[2], r[6], r[7], r[3], r[4], r[5] or ''), 1):
                ws.cell(ri, ci, v).border = THIN
        tot_row = len(rows) + 2
        ws.cell(tot_row, 3, 'Totale').font = Font(bold=True)
        ws.cell(tot_row, 4, sum(int(r[6]) for r in rows)).font = Font(bold=True)
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, 50)
        fd, path = tempfile.mkstemp(suffix='.xlsx', prefix='ScartiEtichette_')
        os.close(fd)
        wb.save(path)
        return path

    def _html_body(self, intro, by_reason, by_category, by_operator, total,
                   total_qty=0, vs_totals=None):
        def tbl(title, items):
            rows = ''.join(f"<tr><td style='padding:3px 8px;border:1px solid #ddd'>{k}</td>"
                           f"<td style='padding:3px 8px;border:1px solid #ddd;text-align:center'>{v}</td></tr>"
                           for k, v in items)
            return (f"<b style='color:#1F3864'>{title} (etichette)</b>"
                    f"<table style='border-collapse:collapse;font-size:12px;margin:4px 0 12px'>{rows}</table>")

        # Il rapporto con il prelevato dice se gli scarti sono tanti o pochi:
        # il totale da solo non lo dice.
        vs_html = ''
        if vs_totals and vs_totals[0]:
            taken, scrapped, rate = vs_totals
            vs_html = (f"<p style='background:#EAF0F6;padding:8px 12px;border-left:4px solid #1F3864'>"
                       f"Prelevate dal magazzino nel periodo: <b>{taken}</b> &nbsp;|&nbsp; "
                       f"scartate: <b>{scrapped}</b> &nbsp;|&nbsp; "
                       f"quota a scarto: <b>{rate:.2f}%</b></p>")
        return f"""<html><body style="font-family:Arial,sans-serif;font-size:13px;color:#333">
<div style="background:#1F3864;color:#fff;padding:14px 18px"><b>Scarti Etichette</b></div>
<p style="white-space:pre-line">{intro}</p>
<p>Nel dettaglio allegato: <b>{total}</b> righe, <b>{total_qty}</b> etichette scartate.</p>
{vs_html}
{tbl('Per motivo', by_reason)}{tbl('Per categoria', by_category)}{tbl('Per operatore', by_operator)}
<p style="color:#888;font-size:11px">In allegato: PDF con logo ed Excel. Email automatica TraceabilityRS.</p>
</body></html>"""
