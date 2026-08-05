# -*- coding: utf-8 -*-
"""
label_scrap_report_gui.py — Report scarti etichette.

Form con filtri data da/a e operatore; genera Excel e PDF (con logo) del dettaglio
e dei riepiloghi (per motivo, categoria, operatore). Legge traceability_rs.dbo.labelscrap.

I riepiloghi sono per QUANTITA' di etichette (labelscrap.Qty), non per numero di
righe: una riga puo' dichiarare 500 etichette e contarla come 1 rendeva i totali
inutilizzabili.

Il report include anche il confronto con le etichette ritirate dal magazzino nel
periodo (richieste materiali indiretti in stato PRELEVATA), cioe' quanta parte
del prelevato e' finita a scarto.
"""
import os
import logging
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta

try:
    from tkcalendar import DateEntry
except Exception:
    DateEntry = None

logger = logging.getLogger(__name__)

# Famiglia materiali indiretti che raccoglie le etichette (ind.FamigliaMateriali).
# Stesso valore usato dalla dichiarazione scarti in label_scrap_gui.py.
LABEL_FAMILY_ID = 1


def open_label_scrap_report(parent, db, lang):
    LabelScrapReportWindow(parent, db, lang)


class LabelScrapReportWindow(tk.Toplevel):
    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        L = self.lang.get
        self.title(L('lsr_title', 'Report Scarti Etichette'))
        self.geometry('1060x680')
        self.transient(parent)
        self._rows = []
        self._vs_rows = []       # (codice, descrizione, prelevate, scartate, perc)
        self._vs_totals = (0, 0, 0.0)
        self._build()
        self._load_operators()
        self._generate()

    def _build(self):
        L = self.lang.get
        flt = ttk.LabelFrame(self, text=L('filters', 'Filtri'), padding=8)
        flt.pack(fill=tk.X, padx=10, pady=8)

        ttk.Label(flt, text=L('lsr_from', 'Da') + ':').pack(side=tk.LEFT, padx=(0, 4))
        if DateEntry:
            self.d_from = DateEntry(flt, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
            self.d_from.set_date(datetime.now() - timedelta(days=30))
        else:
            self.d_from = ttk.Entry(flt, width=12)
            self.d_from.insert(0, (datetime.now() - timedelta(days=30)).strftime('%d/%m/%Y'))
        self.d_from.pack(side=tk.LEFT, padx=(0, 12))

        ttk.Label(flt, text=L('lsr_to', 'A') + ':').pack(side=tk.LEFT, padx=(0, 4))
        if DateEntry:
            self.d_to = DateEntry(flt, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
        else:
            self.d_to = ttk.Entry(flt, width=12)
            self.d_to.insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.d_to.pack(side=tk.LEFT, padx=(0, 12))

        ttk.Label(flt, text=L('lsc_operator', 'Operatore') + ':').pack(side=tk.LEFT, padx=(0, 4))
        self.op_combo = ttk.Combobox(flt, width=24, state='readonly')
        self.op_combo.pack(side=tk.LEFT, padx=(0, 12))

        ttk.Button(flt, text=L('lsr_generate', 'Genera'), command=self._generate).pack(side=tk.LEFT, padx=4)
        ttk.Button(flt, text=L('lsr_excel', '📊 Excel'), command=self._export_excel).pack(side=tk.LEFT, padx=4)
        ttk.Button(flt, text=L('lsr_pdf', '📄 PDF'), command=self._export_pdf).pack(side=tk.LEFT, padx=4)

        wrap = ttk.Frame(self)
        wrap.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        cols = ('date', 'operator', 'label', 'qty', 'material', 'reason', 'category', 'shift')
        self.tree = ttk.Treeview(wrap, columns=cols, show='headings', selectmode='browse')
        for c, t, w, a in (('date', L('lsc_date', 'Data'), 90, 'center'),
                           ('operator', L('lsc_operator', 'Operatore'), 170, 'w'),
                           ('label', L('lsc_scan', 'Etichetta'), 160, 'w'),
                           ('qty', L('lsc_qty', 'Q.tà'), 70, 'e'),
                           ('material', L('lsc_material', 'Materiale'), 130, 'w'),
                           ('reason', L('lsc_reason', 'Motivo'), 200, 'w'),
                           ('category', L('lsc_category', 'Categoria'), 90, 'center'),
                           ('shift', L('lsc_shift', 'Turno'), 60, 'center')):
            self.tree.heading(c, text=t)
            self.tree.column(c, width=w, anchor=a)
        vsb = ttk.Scrollbar(wrap, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        self.count_lbl = ttk.Label(self, text='', foreground='#555')
        self.count_lbl.pack(anchor='w', padx=12, pady=(2, 6))

        # ── Scarti a fronte del prelevato dal magazzino ──────────────────────
        # Le quantita' prelevate non sono per operatore: questo riquadro guarda
        # sempre tutto il periodo, anche quando il dettaglio sopra e' filtrato.
        vs = ttk.LabelFrame(
            self, padding=6,
            text=L('lsr_vs_title', 'Scarti a fronte del prelevato dal magazzino'))
        vs.pack(fill=tk.X, padx=10, pady=(0, 4))
        vcols = ('code', 'descr', 'taken', 'scrapped', 'rate')
        self.vs_tree = ttk.Treeview(vs, columns=vcols, show='headings',
                                    selectmode='none', height=5)
        for c, t, w, a in (('code', L('lsc_material', 'Materiale'), 130, 'w'),
                           ('descr', L('lsr_descr', 'Descrizione'), 330, 'w'),
                           ('taken', L('lsr_withdrawn', 'Prelevate'), 110, 'e'),
                           ('scrapped', L('lsr_scrapped', 'Scartate'), 110, 'e'),
                           ('rate', L('lsr_rate', '% scarto'), 90, 'e')):
            self.vs_tree.heading(c, text=t)
            self.vs_tree.column(c, width=w, anchor=a)
        vsb2 = ttk.Scrollbar(vs, orient='vertical', command=self.vs_tree.yview)
        self.vs_tree.configure(yscrollcommand=vsb2.set)
        self.vs_tree.grid(row=0, column=0, sticky='nsew')
        vsb2.grid(row=0, column=1, sticky='ns')
        vs.columnconfigure(0, weight=1)
        self.vs_tree.tag_configure('total', font=('Segoe UI', 9, 'bold'),
                                   background='#EAF0F6')
        self.vs_lbl = ttk.Label(vs, text='', foreground='#555')
        self.vs_lbl.grid(row=1, column=0, columnspan=2, sticky='w', pady=(4, 0))

    def _load_operators(self):
        vals = [self.lang.get('all_operators', 'Tutti')]
        try:
            cur = self.db.conn.cursor()
            cur.execute("SELECT DISTINCT Operator FROM traceability_rs.dbo.labelscrap ORDER BY Operator")
            vals.extend([r.Operator for r in cur.fetchall() if r.Operator])
            cur.close()
        except Exception as e:
            logger.error(f"Load operators: {e}")
        self.op_combo['values'] = vals
        self.op_combo.current(0)

    def _dates(self):
        def gd(w):
            if DateEntry and hasattr(w, 'get_date'):
                return w.get_date()
            return datetime.strptime(w.get().strip(), '%d/%m/%Y').date()
        return gd(self.d_from), gd(self.d_to)

    def _operator(self):
        op = self.op_combo.get()
        return None if op == self.lang.get('all_operators', 'Tutti') else op

    def _generate(self):
        L = self.lang.get
        try:
            df, dt = self._dates()
        except Exception:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('lsr_bad_dates', 'Date non valide.'), parent=self)
            return
        op = self._operator()
        sql = """
            SELECT ls.ScrapDate, ls.Operator, ls.LabelCode, r.Reason, ls.Category, ls.Shift,
                   ISNULL(ls.Qty, 1) AS Qty, ISNULL(ls.CodiceMateriale, '') AS CodiceMateriale
            FROM traceability_rs.dbo.labelscrap ls
            INNER JOIN traceability_rs.dbo.LabelScrapReasons r
                ON r.LabelScrapReasonId = ls.LabelScrapReasonId
            WHERE ls.ScrapDate BETWEEN ? AND ?
        """
        params = [df, dt]
        if op:
            sql += " AND ls.Operator = ?"
            params.append(op)
        sql += " ORDER BY ls.ScrapDate, ls.Operator"
        try:
            cur = self.db.conn.cursor()
            cur.execute(sql, params)
            self._rows = [(r.ScrapDate, r.Operator, r.LabelCode, r.Reason, r.Category,
                           r.Shift, int(r.Qty or 1), r.CodiceMateriale)
                          for r in cur.fetchall()]
            cur.close()
        except Exception as e:
            logger.error(f"Genera report: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)
            return
        self.tree.delete(*self.tree.get_children())
        for r in self._rows:
            d = r[0].strftime('%d/%m/%Y') if hasattr(r[0], 'strftime') else str(r[0])
            self.tree.insert('', 'end',
                             values=(d, r[1], r[2], r[6], r[7], r[3], r[4], r[5] or ''))
        # Righe E quantita': erano 13 "scarti" che valevano 1233 etichette.
        self.count_lbl.config(text=L('lsr_count_qty', '{0} righe ({1} etichette) nel periodo')
                              .format(len(self._rows), self._total_qty()))
        self._load_vs_warehouse(df, dt)

    def _total_qty(self):
        return sum(r[6] for r in self._rows)

    def _load_vs_warehouse(self, df, dt):
        """Etichette prelevate dal magazzino nel periodo, per materiale, con la
        quota finita a scarto.

        Prelevato = richieste materiali indiretti in stato PRELEVATA con
        DataPrelievo nel periodo. Si usa ind.MaterialiRichieste e non i movimenti
        di SCARICO perche' il ledger dei movimenti e' partito dopo ed e' vuoto
        per i mesi precedenti, mentre le richieste coprono tutto lo storico.

        Ignora il filtro operatore: i prelievi non sono attribuiti a un
        operatore, quindi il rapporto ha senso solo sul totale del periodo.

        La query sta in label_scrap_pdf.fetch_labels_vs_withdrawn perche' la usa
        anche il monitor che manda i report automatici.
        """
        L = self.lang.get
        self._vs_rows = []
        self._vs_totals = (0, 0, 0.0)
        try:
            import label_scrap_pdf
            self._vs_rows, self._vs_totals = label_scrap_pdf.fetch_labels_vs_withdrawn(
                self.db.conn, df, dt, LABEL_FAMILY_ID)
        except Exception as e:
            logger.error(f"Scarti vs prelevato: {e}", exc_info=True)
            self.vs_lbl.config(text=str(e), foreground='#B00020')
        tot_taken, tot_scrap, tot_rate = self._vs_totals

        self.vs_tree.delete(*self.vs_tree.get_children())
        for code, descr, taken, scrapped, rate in self._vs_rows:
            self.vs_tree.insert('', 'end', values=(
                code, descr, f"{taken:,}".replace(',', '.'),
                f"{scrapped:,}".replace(',', '.'),
                f"{rate:.2f}%" if taken else '—'))
        if self._vs_rows:
            self.vs_tree.insert('', 'end', tags=('total',), values=(
                L('lsr_total', 'Totale'), '',
                f"{tot_taken:,}".replace(',', '.'),
                f"{tot_scrap:,}".replace(',', '.'),
                f"{tot_rate:.2f}%" if tot_taken else '—'))
            self.vs_lbl.config(
                text=L('lsr_vs_note',
                       'Prelievi da richieste materiali indiretti (stato PRELEVATA) '
                       'nel periodo — tutti gli operatori'),
                foreground='#555')
        else:
            self.vs_lbl.config(
                text=L('lsr_vs_none', 'Nessun prelievo di etichette dal magazzino nel periodo'),
                foreground='#555')

    def _aggregates(self):
        """Riepiloghi per QUANTITA' di etichette, non per numero di righe: una
        riga puo' dichiararne 500 e contarla come 1 falsava ogni confronto."""
        from collections import defaultdict

        def agg(idx):
            acc = defaultdict(int)
            for r in self._rows:
                acc[r[idx]] += r[6]
            return sorted(acc.items(), key=lambda kv: (-kv[1], str(kv[0])))

        return agg(3), agg(4), agg(1)

    def _export_excel(self):
        L = self.lang.get
        if not self._rows:
            messagebox.showinfo(L('info', 'Info'), L('lsr_no_data', 'Nessun dato da esportare.'), parent=self)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(L('error', 'Errore'), 'openpyxl non disponibile', parent=self)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scarti Etichette'
        H_FILL = PatternFill('solid', fgColor='1F3864')
        H_FONT = Font(bold=True, color='FFFFFF')
        THIN = Border(*(Side(style='thin'),) * 4)
        headers = ['Data', 'Operatore', 'Etichetta', 'Q.tà', 'Materiale',
                   'Motivo', 'Categoria', 'Turno']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill, cell.font, cell.alignment, cell.border = H_FILL, H_FONT, Alignment(horizontal='center'), THIN
        for ri, r in enumerate(self._rows, 2):
            d = r[0].strftime('%d/%m/%Y') if hasattr(r[0], 'strftime') else str(r[0])
            for ci, v in enumerate((d, r[1], r[2], r[6], r[7], r[3], r[4], r[5] or ''), 1):
                ws.cell(ri, ci, v).border = THIN
        # Totale quantita' in coda: e' il numero che conta, non il conteggio righe
        tot_row = len(self._rows) + 2
        ws.cell(tot_row, 3, 'Totale').font = Font(bold=True)
        tc = ws.cell(tot_row, 4, self._total_qty())
        tc.font = Font(bold=True)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:H{len(self._rows) + 1}'
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, 50)

        # foglio riepiloghi (per quantita' di etichette)
        by_reason, by_category, by_operator = self._aggregates()
        ws2 = wb.create_sheet('Riepiloghi')
        r0 = 1
        for title, items in (('Per motivo', by_reason), ('Per categoria', by_category),
                             ('Per operatore', by_operator)):
            ws2.cell(r0, 1, title).font = Font(bold=True, size=12, color='1F3864')
            ws2.cell(r0, 2, 'Etichette').font = Font(bold=True, size=10, color='1F3864')
            r0 += 1
            for lbl, cnt in items:
                ws2.cell(r0, 1, lbl)
                ws2.cell(r0, 2, cnt)
                r0 += 1
            r0 += 1
        for col in ws2.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(w + 3, 50)

        # foglio scarti a fronte del prelevato dal magazzino
        ws3 = wb.create_sheet('Scarti vs prelevato')
        vs_headers = ['Materiale', 'Descrizione', 'Prelevate', 'Scartate', '% scarto']
        for c, h in enumerate(vs_headers, 1):
            cell = ws3.cell(1, c, h)
            cell.fill, cell.font, cell.alignment, cell.border = H_FILL, H_FONT, Alignment(horizontal='center'), THIN
        for ri, (code, descr, taken, scrapped, rate) in enumerate(self._vs_rows, 2):
            for ci, v in enumerate((code, descr, taken, scrapped), 1):
                ws3.cell(ri, ci, v).border = THIN
            cell = ws3.cell(ri, 5, (rate / 100.0) if taken else None)
            cell.number_format = '0.00%'
            cell.border = THIN
        if self._vs_rows:
            tot_taken, tot_scrap, tot_rate = self._vs_totals
            rr = len(self._vs_rows) + 2
            for ci, v in enumerate(('Totale', '', tot_taken, tot_scrap), 1):
                cell = ws3.cell(rr, ci, v)
                cell.font, cell.border = Font(bold=True), THIN
            cell = ws3.cell(rr, 5, (tot_rate / 100.0) if tot_taken else None)
            cell.number_format = '0.00%'
            cell.font, cell.border = Font(bold=True), THIN
        ws3.cell(len(self._vs_rows) + 4, 1,
                 'Prelievi = richieste materiali indiretti in stato PRELEVATA nel periodo '
                 '(tutti gli operatori).').font = Font(italic=True, size=9, color='777777')
        for col in ws3.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws3.column_dimensions[col[0].column_letter].width = min(w + 3, 60)

        temp_dir = r'c:\Temp'
        os.makedirs(temp_dir, exist_ok=True)
        path = os.path.join(temp_dir, f"ScartiEtichette_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(path)
        self._offer_open(path)

    def _export_pdf(self):
        L = self.lang.get
        if not self._rows:
            messagebox.showinfo(L('info', 'Info'), L('lsr_no_data', 'Nessun dato da esportare.'), parent=self)
            return
        try:
            import label_scrap_pdf
            df, dt = self._dates()
            by_reason, by_category, by_operator = self._aggregates()
            wr = label_scrap_pdf.get_warehouse_responsible(self.db.conn)
            temp_dir = r'c:\Temp'
            os.makedirs(temp_dir, exist_ok=True)
            path = os.path.join(temp_dir, f"ScartiEtichette_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
            label_scrap_pdf.generate_report_pdf(path, df, dt, self._operator(), self._rows,
                                                by_reason, by_category, by_operator,
                                                warehouse_responsible=wr,
                                                vs_rows=self._vs_rows,
                                                vs_totals=self._vs_totals)
            self._offer_open(path)
        except Exception as e:
            logger.error(f"Export PDF: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)

    def _offer_open(self, path):
        L = self.lang.get
        if messagebox.askyesno(L('success', 'Successo'),
                               f"{L('file_saved', 'File salvato in')}: {path}\n\n{L('open_file_question', 'Aprire il file?')}",
                               parent=self):
            try:
                os.startfile(path)
            except Exception as e:
                messagebox.showerror(L('error', 'Errore'), str(e), parent=self)
