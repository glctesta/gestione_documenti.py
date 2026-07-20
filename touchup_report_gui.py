# -*- coding: utf-8 -*-
"""
touchup_report_gui.py

Report Touch-up (menu Touch-up -> "Rapporti"). Analizza le segnalazioni problemi
schede registrate nelle tabelle dbo.TouchUp* per periodo, cliente, prodotto e tipo
di difetto, con reattivita' (tempo prima risposta), riaperture ed escalation.

Viste:
  - Dettaglio: una riga per report x difetto (filtrabile);
  - Sintesi: aggregati per difetto / cliente / prodotto / reparto / periodo + KPI.

Export: Excel (multi-foglio) e PDF. Accesso libero (sola lettura).

Catena cliente: TouchUpReportLabels.IDProduct -> Products.IDClient -> Clients.ClientName
(stessa usata dal report FQC).
"""
from __future__ import annotations

import datetime
import logging
import os
import tkinter as tk
from tkinter import messagebox, ttk

logger = logging.getLogger("TraceabilityRS")

STATUS_OPEN = ('NEW', 'REOPENED')

# Soglia oltre la quale una coppia (prodotto, errore) è considerata RICORRENTE:
# se ricorre nonostante le soluzioni dichiarate, le azioni correttive vanno riviste.
RECUR_THRESHOLD = 3


# ─── SQL ──────────────────────────────────────────────────────────────────────

def _build_report_id_subquery(filters: dict):
    """Costruisce (sql, params) di una SELECT TouchUpReportId filtrata, riusabile
    come `... IN (<subquery>)` da tutte le query di dettaglio."""
    conds = ["r.CreatedAt >= ? AND r.CreatedAt < ?"]
    params = [filters['start'], filters['end']]

    status = filters.get('status')
    if status == 'open':
        conds.append("r.Status IN ('NEW','REOPENED')")
    elif status == 'closed':
        conds.append("r.Status = 'CLOSED'")
    elif status == 'reopened':
        conds.append("r.ReopenCount > 0")

    if filters.get('boss_only'):
        conds.append("r.BossEscalated = 1")

    if filters.get('client'):
        conds.append(
            "EXISTS (SELECT 1 FROM dbo.TouchUpReportLabels rl "
            "  LEFT JOIN dbo.Products p ON p.IDProduct = rl.IDProduct "
            "  LEFT JOIN dbo.Clients c ON c.IDClient = p.IDClient "
            "  WHERE rl.TouchUpReportId = r.TouchUpReportId AND c.ClientName = ?)")
        params.append(filters['client'])

    if filters.get('product'):
        conds.append(
            "EXISTS (SELECT 1 FROM dbo.TouchUpReportLabels rl "
            "  WHERE rl.TouchUpReportId = r.TouchUpReportId AND rl.ProductCode LIKE ?)")
        params.append(f"%{filters['product']}%")

    if filters.get('problem_id'):
        conds.append(
            "EXISTS (SELECT 1 FROM dbo.TouchUpReportProblems rp "
            "  WHERE rp.TouchUpReportId = r.TouchUpReportId AND rp.TouchUpProblemId = ?)")
        params.append(filters['problem_id'])

    sql = ("SELECT r.TouchUpReportId FROM dbo.TouchUpReports r WHERE "
           + " AND ".join(conds))
    return sql, params


_Q_REPORTS = """
SELECT r.TouchUpReportId, r.CreatedAt, r.Status, r.EscalationLevel, r.ReopenCount,
       r.BossEscalated, r.EmailSentCount, r.CreatedByUser, r.ComputerSrc,
       r.FirstResponseAt, r.ClosedAt,
       CASE WHEN r.FirstResponseAt IS NOT NULL
            THEN DATEDIFF(SECOND, r.CreatedAt, r.FirstResponseAt) END AS FirstResponseSeconds
FROM dbo.TouchUpReports r
WHERE r.TouchUpReportId IN ({subq})
ORDER BY r.CreatedAt DESC
"""

_Q_LABELS = """
SELECT rl.TouchUpReportId, rl.LabelCod, rl.OrderNumber, rl.ProductCode,
       c.ClientName
FROM dbo.TouchUpReportLabels rl
  LEFT JOIN dbo.Products p ON p.IDProduct = rl.IDProduct
  LEFT JOIN dbo.Clients  c ON c.IDClient  = p.IDClient
WHERE rl.TouchUpReportId IN ({subq})
"""

_Q_PROBLEMS = """
SELECT rp.TouchUpReportId, rp.TouchUpProblemId, pr.ProblemCode,
       pr.ProblemDescription, pr.Severity,
       STUFF((SELECT DISTINCT ', ' + ISNULL(cs.SubCdcDescription, c.CdcDescription)
              FROM dbo.TouchUpProblemRouting rt
                LEFT JOIN employee.dbo.costcenters c ON c.CdcId = rt.CdcId
                LEFT JOIN employee.dbo.cdcsub     cs ON cs.SubCdcId = rt.SubCdcId
              WHERE rt.TouchUpProblemId = rp.TouchUpProblemId AND rt.DateOut IS NULL
              FOR XML PATH('')), 1, 2, '') AS Dept
FROM dbo.TouchUpReportProblems rp
  JOIN dbo.TouchUpProblems pr ON pr.TouchUpProblemId = rp.TouchUpProblemId
WHERE rp.TouchUpReportId IN ({subq})
"""

_Q_RESPONSES = """
SELECT resp.TouchUpReportId, resp.RespondedByUser, resp.RespondedAt,
       resp.ReactionSeconds, resp.ActionsTaken
FROM dbo.TouchUpResponses resp
WHERE resp.TouchUpReportId IN ({subq})
ORDER BY resp.RespondedAt
"""

_Q_FILTER_CLIENTS = """
SELECT DISTINCT c.ClientName
FROM dbo.TouchUpReportLabels rl
  JOIN dbo.Products p ON p.IDProduct = rl.IDProduct
  JOIN dbo.Clients  c ON c.IDClient  = p.IDClient
WHERE c.ClientName IS NOT NULL
ORDER BY c.ClientName
"""

_Q_FILTER_PROBLEMS = """
SELECT TouchUpProblemId, ProblemDescription
FROM dbo.TouchUpProblems
WHERE DateOut IS NULL
ORDER BY ProblemDescription
"""

_Q_THRESHOLD = "SELECT NoResponseEscalationMinutes FROM dbo.TouchUpConfig WHERE Id = 1"


def _rows(conn, sql, params=None):
    cur = conn.cursor()
    cur.execute(sql, params or [])
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def run_report(conn, filters: dict) -> dict:
    """Esegue tutte le query e calcola aggregati e KPI. Ritorna un dict completo."""
    subq, params = _build_report_id_subquery(filters)

    reports = _rows(conn, _Q_REPORTS.format(subq=subq), params)
    labels = _rows(conn, _Q_LABELS.format(subq=subq), params)
    problems = _rows(conn, _Q_PROBLEMS.format(subq=subq), params)
    responses = _rows(conn, _Q_RESPONSES.format(subq=subq), params)

    try:
        thr_min = int(_rows(conn, _Q_THRESHOLD)[0]['NoResponseEscalationMinutes'])
    except Exception:
        thr_min = 30

    # Indici per report
    labels_by_rep: dict = {}
    for l in labels:
        labels_by_rep.setdefault(l['TouchUpReportId'], []).append(l)
    problems_by_rep: dict = {}
    for p in problems:
        problems_by_rep.setdefault(p['TouchUpReportId'], []).append(p)
    resp_by_rep: dict = {}
    for r in responses:
        resp_by_rep.setdefault(r['TouchUpReportId'], []).append(r)

    def _join(vals):
        seen, out = set(), []
        for v in vals:
            if v and v not in seen:
                seen.add(v)
                out.append(v)
        return ', '.join(out)

    for rep in reports:
        rid = rep['TouchUpReportId']
        ls = labels_by_rep.get(rid, [])
        rep['ClientsStr'] = _join([x.get('ClientName') for x in ls])
        rep['ProductsStr'] = _join([x.get('ProductCode') for x in ls])
        rep['OrdersStr'] = _join([x.get('OrderNumber') for x in ls])
        rep['LabelsStr'] = _join([x.get('LabelCod') for x in ls])
        rep['ActionsStr'] = ' | '.join(
            [x['ActionsTaken'] for x in resp_by_rep.get(rid, []) if x.get('ActionsTaken')])

    # ── KPI ──
    total = len(reports)
    open_n = sum(1 for r in reports if r['Status'] in STATUS_OPEN)
    closed_n = sum(1 for r in reports if r['Status'] == 'CLOSED')
    reopened_n = sum(1 for r in reports if (r['ReopenCount'] or 0) > 0)
    boss_n = sum(1 for r in reports if r['BossEscalated'])
    resp_secs = [r['FirstResponseSeconds'] for r in reports
                 if r['FirstResponseSeconds'] is not None]
    avg_resp_min = round(sum(resp_secs) / len(resp_secs) / 60.0, 1) if resp_secs else None
    within = sum(1 for s in resp_secs if s <= thr_min * 60)
    pct_within = round(within * 100.0 / len(resp_secs), 1) if resp_secs else None

    kpi = {
        'total': total, 'open': open_n, 'closed': closed_n,
        'reopened': reopened_n, 'boss': boss_n,
        'avg_resp_min': avg_resp_min, 'pct_within': pct_within,
        'threshold_min': thr_min,
    }

    # ── Aggregati ──
    # per difetto: conteggio report distinti + tempo medio risposta dei loro report
    resp_secs_by_rep = {r['TouchUpReportId']: r['FirstResponseSeconds'] for r in reports}
    reopen_by_rep = {r['TouchUpReportId']: (r['ReopenCount'] or 0) for r in reports}

    def _agg_defect():
        acc = {}
        for p in problems:
            key = p['ProblemDescription'] or (p['ProblemCode'] or '—')
            d = acc.setdefault(key, {'reports': set(), 'secs': [], 'reopen': 0})
            d['reports'].add(p['TouchUpReportId'])
        out = []
        for key, d in acc.items():
            reps = d['reports']
            secs = [resp_secs_by_rep.get(x) for x in reps if resp_secs_by_rep.get(x) is not None]
            reopen = sum(1 for x in reps if reopen_by_rep.get(x, 0) > 0)
            out.append({
                'key': key, 'count': len(reps),
                'pct': round(len(reps) * 100.0 / total, 1) if total else 0,
                'avg_resp_min': round(sum(secs) / len(secs) / 60.0, 1) if secs else None,
                'reopened': reopen,
            })
        out.sort(key=lambda x: x['count'], reverse=True)
        return out

    def _agg_by(fetch_key):
        acc = {}
        for l in labels:
            key = fetch_key(l)
            if not key:
                continue
            acc.setdefault(key, set()).add(l['TouchUpReportId'])
        out = [{'key': k, 'count': len(v),
                'pct': round(len(v) * 100.0 / total, 1) if total else 0}
               for k, v in acc.items()]
        out.sort(key=lambda x: x['count'], reverse=True)
        return out

    def _agg_dept():
        acc = {}
        for p in problems:
            for dep in [x.strip() for x in (p.get('Dept') or '').split(',') if x.strip()] or ['—']:
                acc.setdefault(dep, set()).add(p['TouchUpReportId'])
        out = [{'key': k, 'count': len(v),
                'pct': round(len(v) * 100.0 / total, 1) if total else 0}
               for k, v in acc.items()]
        out.sort(key=lambda x: x['count'], reverse=True)
        return out

    def _agg_period(mode):
        acc = {}
        for r in reports:
            dt = r['CreatedAt']
            if not dt:
                continue
            if mode == 'week':
                iso = dt.isocalendar()
                key = f"{iso[0]}-W{iso[1]:02d}"
            elif mode == 'month':
                key = dt.strftime('%Y-%m')
            else:
                key = dt.strftime('%Y-%m-%d')
            acc[key] = acc.get(key, 0) + 1
        out = [{'key': k, 'count': v} for k, v in acc.items()]
        out.sort(key=lambda x: x['key'])
        return out

    def _agg_product_problem():
        """Ricorrenza per (prodotto, errore) con le soluzioni dichiarate dai tecnici.
        Se la stessa coppia prodotto+errore ricorre (>= soglia) e continuano a
        comparire le stesse segnalazioni nonostante le soluzioni, e' segnale che le
        azioni correttive non sono state valide o non effettivamente attuate."""
        acc = {}
        for rep in reports:
            rid = rep['TouchUpReportId']
            prods = [x.get('ProductCode') for x in labels_by_rep.get(rid, [])
                     if x.get('ProductCode')] or ['—']
            probs = [(p.get('ProblemDescription') or p.get('ProblemCode') or '—')
                     for p in problems_by_rep.get(rid, [])] or ['—']
            actions = rep.get('ActionsStr') or ''
            created = rep['CreatedAt']
            reopened = (rep['ReopenCount'] or 0) > 0
            for pc in dict.fromkeys(prods):
                for pb in dict.fromkeys(probs):
                    d = acc.setdefault((pc, pb), {'reports': set(), 'reopened': set(),
                                                  'solutions': [], 'first': None, 'last': None})
                    d['reports'].add(rid)
                    if reopened:
                        d['reopened'].add(rid)
                    if actions:
                        d['solutions'].append(actions)
                    if created:
                        d['first'] = created if d['first'] is None else min(d['first'], created)
                        d['last'] = created if d['last'] is None else max(d['last'], created)
        out = []
        for (pc, pb), d in acc.items():
            n = len(d['reports'])
            sols, seen = [], set()
            for s in d['solutions']:
                if s not in seen:
                    seen.add(s)
                    sols.append(s)
            recurring = n >= RECUR_THRESHOLD
            if recurring and sols:
                flag = 'RECURRING despite solutions'
            elif recurring and not sols:
                flag = 'RECURRING without solutions'
            else:
                flag = ''
            out.append({
                'product': pc, 'problem': pb, 'count': n,
                'reopened': len(d['reopened']), 'n_solutions': len(sols),
                'solutions': ' | '.join(sols),
                'first': d['first'], 'last': d['last'],
                'recurring': recurring, 'flag': flag,
            })
        out.sort(key=lambda x: (x['count'], x['reopened']), reverse=True)
        return out

    return {
        'reports': reports, 'labels': labels, 'problems': problems,
        'responses': responses, 'problems_by_rep': problems_by_rep,
        'kpi': kpi,
        'by_defect': _agg_defect(),
        'by_client': _agg_by(lambda l: l.get('ClientName')),
        'by_product': _agg_by(lambda l: l.get('ProductCode')),
        'by_product_problem': _agg_product_problem(),
        'by_dept': _agg_dept(),
        'by_period_day': _agg_period('day'),
        'by_period_week': _agg_period('week'),
        'by_period_month': _agg_period('month'),
        'filters': filters,
    }


def _fmt_dt(v) -> str:
    if v is None:
        return ''
    try:
        return v.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return str(v)


def _mins(secs) -> str:
    if secs is None:
        return ''
    return f"{secs / 60.0:.1f}"


# ─── Export Excel ──────────────────────────────────────────────────────────────

def _write_data_sheets(wb, data: dict, det_ws) -> None:
    """Popola il workbook con i fogli dati (Detail + sintesi + KPI).

    ``det_ws`` è il foglio da usare come 'Detail' (di norma ``wb.active``, ma per
    l'export dell'analisi AI è un foglio creato dopo quello dell'analisi).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    H_FILL = PatternFill('solid', fgColor='1F3864')
    H_FONT = Font(bold=True, color='FFFFFF', size=10)
    ALT = PatternFill('solid', fgColor='F4F6F8')
    side = Side(style='thin', color='C0C0C0')
    THIN = Border(left=side, right=side, top=side, bottom=side)

    def _sheet(ws, title, headers, rows_iter, widths=None):
        ws.title = title
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.fill, c.font, c.border = H_FILL, H_FONT, THIN
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        ri = 2
        for row in rows_iter:
            for ci, v in enumerate(row, 1):
                cell = ws.cell(ri, ci, v)
                cell.border = THIN
                cell.font = Font(size=9)
                if ri % 2 == 0:
                    cell.fill = ALT
            ri += 1
        for ci, w in enumerate(widths or [], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.freeze_panes = 'A2'

    # Dettaglio
    det = det_ws
    det_rows = []
    for rep in data['reports']:
        probs = data['problems_by_rep'].get(rep['TouchUpReportId'], []) or [{}]
        for p in probs:
            det_rows.append([
                _fmt_dt(rep['CreatedAt']), rep['TouchUpReportId'], rep['Status'],
                rep['ClientsStr'], rep['ProductsStr'], rep['OrdersStr'], rep['LabelsStr'],
                p.get('ProblemDescription', ''), p.get('Severity', ''), p.get('Dept', ''),
                rep['CreatedByUser'] or '', _mins(rep['FirstResponseSeconds']),
                rep['ReopenCount'] or 0, 'Sì' if rep['BossEscalated'] else '',
                rep['ActionsStr'],
            ])
    _sheet(det, 'Detail',
           ['Date/time', 'No.', 'Status', 'Client', 'Product', 'Order', 'Board',
            'Defect', 'Severity', 'Department', 'Operator', '1st response (min)',
            'Reopenings', 'Boss escal.', 'Actions'],
           det_rows, [16, 6, 10, 22, 20, 16, 18, 28, 8, 20, 16, 14, 10, 10, 40])

    _sheet(wb.create_sheet(), 'By defect',
           ['Defect', 'Reports', '%', 'Avg time (min)', 'Reopened'],
           [[d['key'], d['count'], d['pct'], d['avg_resp_min'], d['reopened']]
            for d in data['by_defect']], [30, 12, 8, 16, 10])
    _sheet(wb.create_sheet(), 'By client',
           ['Client', 'Reports', '%'],
           [[d['key'], d['count'], d['pct']] for d in data['by_client']], [30, 12, 8])
    _sheet(wb.create_sheet(), 'By product',
           ['Product', 'Reports', '%'],
           [[d['key'], d['count'], d['pct']] for d in data['by_product']], [30, 12, 8])
    _sheet(wb.create_sheet(), 'Product-error recurrence',
           ['Product', 'Error', 'Occurrences', 'Reopenings', 'No. solutions',
            'Declared solutions', 'First', 'Last', 'Flag'],
           [[d['product'], d['problem'], d['count'], d['reopened'], d['n_solutions'],
             d['solutions'], _fmt_dt(d['first']), _fmt_dt(d['last']), d['flag']]
            for d in data.get('by_product_problem', [])],
           [22, 30, 12, 12, 12, 45, 16, 16, 28])
    _sheet(wb.create_sheet(), 'By department',
           ['Department', 'Reports', '%'],
           [[d['key'], d['count'], d['pct']] for d in data['by_dept']], [30, 12, 8])
    _sheet(wb.create_sheet(), 'By period (month)',
           ['Period', 'Reports'],
           [[d['key'], d['count']] for d in data['by_period_month']], [16, 12])

    k = data['kpi']
    _sheet(wb.create_sheet(), 'KPI',
           ['Indicator', 'Value'],
           [['Total reports', k['total']], ['Open', k['open']],
            ['Closed', k['closed']], ['Reopened', k['reopened']],
            ['Boss escalation', k['boss']],
            ['Avg 1st response (min)', k['avg_resp_min']],
            [f"% responses within {k['threshold_min']} min", k['pct_within']]],
           [34, 14])


def export_excel(data: dict, meta: dict) -> str:
    import openpyxl

    out_dir = r'C:\temp'
    os.makedirs(out_dir, exist_ok=True)
    fn = (f"Report_TouchUp_{meta['date_from']:%Y%m%d}-{meta['date_to']:%Y%m%d}.xlsx")
    path = os.path.join(out_dir, fn)

    wb = openpyxl.Workbook()
    _write_data_sheets(wb, data, wb.active)
    wb.save(path)
    return path


def export_ai_excel(ai_text: str, data: dict, meta: dict) -> str:
    """Workbook con l'analisi AI nel primo foglio e i dati grezzi nei fogli
    successivi (stessa struttura dell'export standard)."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    out_dir = r'C:\temp'
    os.makedirs(out_dir, exist_ok=True)
    fn = (f"AnalisiAI_TouchUp_{meta['date_from']:%Y%m%d}-{meta['date_to']:%Y%m%d}.xlsx")
    path = os.path.join(out_dir, fn)

    wb = openpyxl.Workbook()
    ai_ws = wb.active
    ai_ws.title = 'AI Analysis'
    ai_ws.column_dimensions['A'].width = 120

    title = ai_ws.cell(1, 1, 'Touch-up — AI solution-validity analysis')
    title.font = Font(bold=True, color='FFFFFF', size=12)
    title.fill = PatternFill('solid', fgColor='1F3864')
    title.alignment = Alignment(horizontal='left', vertical='center')
    ai_ws.row_dimensions[1].height = 22

    sub = ai_ws.cell(2, 1, f"Period: {meta['date_from']:%d/%m/%Y} → {meta['date_to']:%d/%m/%Y}")
    sub.font = Font(italic=True, color='555555', size=9)

    ri = 4
    for raw in (ai_text or '').split('\n'):
        line = raw.rstrip()
        cell = ai_ws.cell(ri, 1, line)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = Font(size=10)
        ri += 1

    # Dati grezzi nei fogli successivi (Detail su un foglio creato dopo l'analisi)
    _write_data_sheets(wb, data, wb.create_sheet())
    wb.save(path)
    return path


def _ai_text_to_html(ai_text: str, meta: dict, user_name: str = None) -> str:
    """Converte il testo dell'analisi AI (plain/markdown leggero) in un'email
    HTML ben formattata, in stile coerente con le altre notifiche del sistema."""
    import html
    import re

    def _inline(s: str) -> str:
        s = html.escape(s)
        s = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)
        s = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'<em>\1</em>', s)
        return s

    blocks = []
    for raw in (ai_text or '').split('\n'):
        line = raw.rstrip()
        stripped = line.strip()
        if not stripped:
            continue
        # Titoli markdown (#, ##, ...)
        m_h = re.match(r'^#{1,6}\s+(.*)$', stripped)
        if m_h:
            blocks.append(
                f'<h3 style="color:#1F3864; margin:16px 0 6px 0; font-size:15px;">'
                f'{_inline(m_h.group(1))}</h3>')
            continue
        # Voci numerate: 1. ... / 1) ...
        m_num = re.match(r'^(\d+)[.)]\s+(.*)$', stripped)
        if m_num:
            blocks.append(
                f'<p style="margin:8px 0; line-height:1.55;">'
                f'<strong style="color:#1F3864;">{m_num.group(1)}.</strong> '
                f'{_inline(m_num.group(2))}</p>')
            continue
        # Elenchi puntati: - ... / * ... / • ...
        m_bul = re.match(r'^[-*•]\s+(.*)$', stripped)
        if m_bul:
            blocks.append(
                f'<p style="margin:4px 0 4px 18px; line-height:1.5;">'
                f'&bull; {_inline(m_bul.group(1))}</p>')
            continue
        blocks.append(
            f'<p style="margin:8px 0; line-height:1.55;">{_inline(stripped)}</p>')

    body_html = '\n'.join(blocks)
    period = f"{meta['date_from']:%d/%m/%Y} &rarr; {meta['date_to']:%d/%m/%Y}"
    greeting = f"Hi {html.escape(user_name)}," if user_name else "Hi,"
    year = meta['date_to'].year

    return f"""\
<html>
<body style="margin:0; padding:0; background:#f4f6f8; font-family:'Segoe UI',Arial,sans-serif; color:#333;">
  <div style="max-width:760px; margin:20px auto; background:#ffffff; border-radius:8px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,0.1);">
    <div style="background:#1F3864; padding:18px 24px;">
      <div style="color:#ffffff; font-size:18px; font-weight:bold;">🧠 Touch-up — AI Solution-Validity Analysis</div>
      <div style="color:#c9d4ea; font-size:12px; margin-top:4px;">Period: {period}</div>
    </div>
    <div style="padding:22px 24px;">
      <p style="margin:0 0 14px 0; line-height:1.55;">{greeting}</p>
      <p style="margin:0 0 16px 0; line-height:1.55; color:#555;">
        Below is the AI analysis of the recurrence of product&times;error pairs and the validity
        of the declared solutions. The same data — plus a separate tab with the raw records —
        is attached as an Excel file.
      </p>
      <div style="border-top:1px solid #e0e4ea; padding-top:14px;">
        {body_html}
      </div>
      <div style="margin-top:26px; padding-top:14px; border-top:1px solid #dee2e6;">
        <p style="font-size:11px; color:#888; line-height:1.5; margin:0;">
          This analysis was generated by a local AI model on plant-internal data.
          Figures are computed deterministically; the model only interprets them.<br/>
          &copy; {year} Vandewiele Romania — automated notification, please do not reply.
        </p>
      </div>
    </div>
  </div>
</body>
</html>"""


# ─── Export PDF ────────────────────────────────────────────────────────────────

def export_pdf(data: dict, meta: dict) -> str:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    out_dir = r'C:\temp'
    os.makedirs(out_dir, exist_ok=True)
    fn = (f"Report_TouchUp_{meta['date_from']:%Y%m%d}-{meta['date_to']:%Y%m%d}.pdf")
    path = os.path.join(out_dir, fn)

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=1 * cm, rightMargin=1 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle('t', parent=styles['Title'], fontSize=15, textColor=colors.HexColor('#1F3864'))
    h2 = ParagraphStyle('h2', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#1F3864'))
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=7)
    story = []

    story.append(Paragraph("Touch-up Report", title_st))
    story.append(Paragraph(
        f"Period: {meta['date_from']:%d/%m/%Y} → {meta['date_to']:%d/%m/%Y}",
        styles['Normal']))
    k = data['kpi']
    story.append(Paragraph(
        f"Total: <b>{k['total']}</b> &nbsp; Open: {k['open']} &nbsp; Closed: {k['closed']} "
        f"&nbsp; Reopened: {k['reopened']} &nbsp; Boss escalation: {k['boss']} &nbsp; "
        f"Avg 1st response: {k['avg_resp_min'] if k['avg_resp_min'] is not None else '-'} min "
        f"&nbsp; % within {k['threshold_min']}min: {k['pct_within'] if k['pct_within'] is not None else '-'}",
        styles['Normal']))
    story.append(Spacer(1, 8))

    def _tbl(headers, rows, col_widths, header_bg='#1F3864'):
        wrapped = [[Paragraph(str(h), ParagraphStyle('th', parent=small, textColor=colors.white)) for h in headers]]
        for r in rows:
            wrapped.append([Paragraph(str(v if v is not None else ''), small) for v in r])
        t = Table(wrapped, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_bg)),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#C0C0C0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F8')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        return t

    # Summary by defect
    story.append(Paragraph("By defect type", h2))
    story.append(_tbl(
        ['Defect', 'Reports', '%', 'Avg time (min)', 'Reopened'],
        [[d['key'], d['count'], d['pct'], d['avg_resp_min'], d['reopened']] for d in data['by_defect']],
        [10 * cm, 2 * cm, 1.6 * cm, 3.5 * cm, 2 * cm]))
    story.append(Spacer(1, 8))

    # Client + Product
    story.append(Paragraph("By client", h2))
    story.append(_tbl(['Client', 'Reports', '%'],
                      [[d['key'], d['count'], d['pct']] for d in data['by_client']],
                      [14 * cm, 2 * cm, 1.6 * cm]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("By product", h2))
    story.append(_tbl(['Product', 'Reports', '%'],
                      [[d['key'], d['count'], d['pct']] for d in data['by_product']],
                      [14 * cm, 2 * cm, 1.6 * cm]))
    story.append(Spacer(1, 10))

    # Product × error recurrence (solution validity)
    story.append(Paragraph("Product × error recurrence — solution validity", h2))
    story.append(Paragraph(
        f"Recurring product+error pairs (≥{RECUR_THRESHOLD} reports): if they persist despite the "
        f"solutions declared by the technicians, the corrective actions must be verified.",
        small))
    story.append(Spacer(1, 3))
    story.append(_tbl(
        ['Product', 'Error', 'Occ.', 'Reop.', 'No.sol', 'Declared solutions', 'First', 'Last', 'Flag'],
        [[d['product'], d['problem'], d['count'], d['reopened'], d['n_solutions'],
          d['solutions'], _fmt_dt(d['first']), _fmt_dt(d['last']), d['flag']]
         for d in data.get('by_product_problem', [])],
        [3.0 * cm, 4.5 * cm, 1.2 * cm, 1.2 * cm, 1.2 * cm, 6.0 * cm, 2.2 * cm, 2.2 * cm, 4.0 * cm]))
    story.append(Spacer(1, 10))

    # Detail
    story.append(Paragraph("Report detail", h2))
    det_rows = []
    for rep in data['reports']:
        probs = data['problems_by_rep'].get(rep['TouchUpReportId'], []) or [{}]
        for p in probs:
            det_rows.append([
                _fmt_dt(rep['CreatedAt']), rep['TouchUpReportId'], rep['Status'],
                rep['ClientsStr'], rep['ProductsStr'], p.get('ProblemDescription', ''),
                p.get('Dept', ''), _mins(rep['FirstResponseSeconds']),
                rep['ReopenCount'] or 0, 'Sì' if rep['BossEscalated'] else '',
            ])
    story.append(_tbl(
        ['Date/time', 'No.', 'Status', 'Client', 'Product', 'Defect', 'Department',
         '1st resp.(min)', 'Reop.', 'Boss'],
        det_rows,
        [2.6 * cm, 1.1 * cm, 1.6 * cm, 3.2 * cm, 2.8 * cm, 5.5 * cm, 3.2 * cm,
         1.8 * cm, 1.2 * cm, 1.1 * cm]))

    doc.build(story)
    return path


# ─── GUI ───────────────────────────────────────────────────────────────────────

class TouchUpReportWindow(tk.Toplevel):
    def __init__(self, master, db, lang, user_email=None, user_name=None):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_email = user_email
        self.user_name = user_name
        self._data = None
        self._ai_result_text = None
        L = self.lang.get
        self.title(L('touchup_report_title', 'Report Touch-up'))
        self.geometry('1180x740')
        self.minsize(960, 560)
        self.resizable(True, True)
        self._build_ui()
        self.grab_set()

    # ── UI ──
    def _build_ui(self):
        L = self.lang.get
        header = tk.Frame(self, bg='#1F3864')
        header.pack(fill=tk.X)
        tk.Label(header, text=L('touchup_report_title', 'Report Touch-up'),
                 bg='#1F3864', fg='white', font=('Helvetica', 13, 'bold')).pack(
            side=tk.LEFT, padx=12, pady=10)

        f = ttk.LabelFrame(self, text=L('touchup_report_filters', 'Filtri'))
        f.pack(fill=tk.X, padx=10, pady=6)

        today = datetime.date.today()
        d_from = today - datetime.timedelta(days=30)
        ttk.Label(f, text=L('touchup_report_from', 'Da:')).grid(row=0, column=0, sticky='w', padx=6, pady=5)
        self._v_from = tk.StringVar(value=d_from.strftime('%d/%m/%Y'))
        ttk.Entry(f, textvariable=self._v_from, width=12).grid(row=0, column=1, padx=4, pady=5)
        ttk.Label(f, text=L('touchup_report_to', 'A:')).grid(row=0, column=2, sticky='w', padx=6, pady=5)
        self._v_to = tk.StringVar(value=today.strftime('%d/%m/%Y'))
        ttk.Entry(f, textvariable=self._v_to, width=12).grid(row=0, column=3, padx=4, pady=5)

        ttk.Label(f, text=L('touchup_report_client', 'Cliente:')).grid(row=0, column=4, sticky='w', padx=6, pady=5)
        self._v_client = tk.StringVar()
        self._cb_client = ttk.Combobox(f, textvariable=self._v_client, width=22, state='readonly')
        self._cb_client.grid(row=0, column=5, padx=4, pady=5)

        ttk.Label(f, text=L('touchup_report_product', 'Prodotto:')).grid(row=0, column=6, sticky='w', padx=6, pady=5)
        self._v_product = tk.StringVar()
        ttk.Entry(f, textvariable=self._v_product, width=18).grid(row=0, column=7, padx=4, pady=5)

        ttk.Label(f, text=L('touchup_report_defect', 'Difetto:')).grid(row=1, column=0, sticky='w', padx=6, pady=5)
        self._v_problem = tk.StringVar()
        self._cb_problem = ttk.Combobox(f, textvariable=self._v_problem, width=30, state='readonly')
        self._cb_problem.grid(row=1, column=1, columnspan=3, sticky='w', padx=4, pady=5)

        ttk.Label(f, text=L('touchup_report_status', 'Stato:')).grid(row=1, column=4, sticky='w', padx=6, pady=5)
        self._v_status = tk.StringVar()
        self._status_map = {
            L('touchup_report_status_all', 'Tutti'): '',
            L('touchup_report_status_open', 'Aperti'): 'open',
            L('touchup_report_status_closed', 'Chiusi'): 'closed',
            L('touchup_report_status_reopened', 'Riaperti'): 'reopened',
        }
        cb_status = ttk.Combobox(f, textvariable=self._v_status, width=14, state='readonly',
                                 values=list(self._status_map.keys()))
        cb_status.current(0)
        cb_status.grid(row=1, column=5, padx=4, pady=5)

        self._v_boss = tk.BooleanVar(value=False)
        ttk.Checkbutton(f, text=L('touchup_report_boss_only', 'Solo escalation capo'),
                        variable=self._v_boss).grid(row=1, column=6, columnspan=2, sticky='w', padx=6, pady=5)

        ttk.Button(f, text=L('touchup_report_search', '🔍 Cerca'),
                   command=self._do_search).grid(row=0, column=8, rowspan=2, padx=12, pady=5)

        # KPI bar
        self._kpi = tk.Label(self, bg='#E8F0FE', fg='#1F3864', anchor='w',
                             font=('Helvetica', 9, 'bold'), justify='left', padx=10, pady=6)
        self._kpi.pack(fill=tk.X, padx=10, pady=(0, 4))

        # Notebook: Dettaglio / Sintesi
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        self._build_detail_tab()
        self._build_summary_tab()

        # Bottoni
        bar = ttk.Frame(self)
        bar.pack(fill=tk.X, padx=10, pady=6)
        self._btn_xls = ttk.Button(bar, text=L('touchup_report_export_xls', '📊 Esporta Excel'),
                                   command=lambda: self._export('xls'), state='disabled')
        self._btn_xls.pack(side=tk.LEFT, padx=4)
        self._btn_pdf = ttk.Button(bar, text=L('touchup_report_export_pdf', '📄 Esporta PDF'),
                                   command=lambda: self._export('pdf'), state='disabled')
        self._btn_pdf.pack(side=tk.LEFT, padx=4)
        self._btn_ai = ttk.Button(bar, text=L('touchup_report_ai', '🧠 Analisi AI'),
                                   command=self._run_ai_analysis, state='disabled')
        self._btn_ai.pack(side=tk.LEFT, padx=12)
        ttk.Button(bar, text=L('btn_close', 'Chiudi'), command=self.destroy).pack(side=tk.RIGHT, padx=4)

        self._load_filter_options()

    def _build_detail_tab(self):
        L = self.lang.get
        tab = ttk.Frame(self._nb)
        self._nb.add(tab, text=L('touchup_report_tab_detail', 'Dettaglio'))
        cols = ('date', 'id', 'status', 'client', 'product', 'order', 'label',
                'defect', 'sev', 'dept', 'user', 'resp', 'reopen', 'boss', 'actions')
        heads = (L('tur_c_date', 'Data/ora'), L('tur_c_id', 'N°'), L('tur_c_status', 'Stato'),
                 L('tur_c_client', 'Cliente'), L('tur_c_product', 'Prodotto'),
                 L('tur_c_order', 'Ordine'), L('tur_c_label', 'Scheda'),
                 L('tur_c_defect', 'Difetto'), L('tur_c_sev', 'Sev.'), L('tur_c_dept', 'Reparto'),
                 L('tur_c_user', 'Operatore'), L('tur_c_resp', '1ª risp.(min)'),
                 L('tur_c_reopen', 'Riap.'), L('tur_c_boss', 'Capo'), L('tur_c_actions', 'Azioni'))
        widths = (120, 45, 75, 140, 120, 110, 120, 200, 45, 130, 110, 90, 55, 45, 240)
        wrap = ttk.Frame(tab)
        wrap.pack(fill=tk.BOTH, expand=True)
        self._t_det = ttk.Treeview(wrap, columns=cols, show='headings')
        for c, h, w in zip(cols, heads, widths):
            self._t_det.heading(c, text=h)
            anchor = 'w' if c in ('client', 'product', 'defect', 'dept', 'actions') else 'center'
            self._t_det.column(c, width=w, anchor=anchor,
                               stretch=(c in ('defect', 'actions')))
        self._t_det.tag_configure('reopen', foreground='#B71C1C')
        self._t_det.tag_configure('boss', background='#FFF3E0')
        vsb = ttk.Scrollbar(wrap, orient='vertical', command=self._t_det.yview)
        hsb = ttk.Scrollbar(wrap, orient='horizontal', command=self._t_det.xview)
        self._t_det.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._t_det.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

    def _build_summary_tab(self):
        L = self.lang.get
        tab = ttk.Frame(self._nb)
        self._nb.add(tab, text=L('touchup_report_tab_summary', 'Sintesi'))
        sub = ttk.Notebook(tab)
        sub.pack(fill=tk.BOTH, expand=True)
        self._sum_trees = {}

        def _mk(name, label, cols, heads, widths):
            fr = ttk.Frame(sub)
            sub.add(fr, text=label)
            tv = ttk.Treeview(fr, columns=cols, show='headings')
            for c, h, w in zip(cols, heads, widths):
                tv.heading(c, text=h)
                tv.column(c, width=w, anchor='w' if c == 'key' else 'center')
            sb = ttk.Scrollbar(fr, orient='vertical', command=tv.yview)
            tv.configure(yscrollcommand=sb.set)
            tv.pack(side='left', fill=tk.BOTH, expand=True)
            sb.pack(side='right', fill='y')
            self._sum_trees[name] = tv

        _mk('defect', L('tur_sum_defect', 'Per difetto'),
            ('key', 'count', 'pct', 'resp', 'reopen'),
            (L('tur_c_defect', 'Difetto'), L('tur_sc_count', 'Segnalazioni'), '%',
             L('tur_sc_avgresp', 'Tempo medio (min)'), L('tur_sc_reopen', 'Riaperti')),
            (320, 100, 60, 130, 90))
        _mk('client', L('tur_sum_client', 'Per cliente'),
            ('key', 'count', 'pct'),
            (L('tur_c_client', 'Cliente'), L('tur_sc_count', 'Segnalazioni'), '%'),
            (340, 110, 70))
        _mk('product', L('tur_sum_product', 'Per prodotto'),
            ('key', 'count', 'pct'),
            (L('tur_c_product', 'Prodotto'), L('tur_sc_count', 'Segnalazioni'), '%'),
            (340, 110, 70))
        _mk('dept', L('tur_sum_dept', 'Per reparto'),
            ('key', 'count', 'pct'),
            (L('tur_c_dept', 'Reparto'), L('tur_sc_count', 'Segnalazioni'), '%'),
            (340, 110, 70))

        # Ricorrenze prodotto × errore (validità soluzioni tecnici)
        fr_rec = ttk.Frame(sub)
        sub.add(fr_rec, text=L('tur_sum_recurr', 'Ricorrenze prodotto×errore'))
        tk.Label(fr_rec, bg='#FFF8E1', fg='#7B4B00', anchor='w', justify='left', padx=8, pady=4,
                 text=L('tur_recurr_hint',
                        'Coppie prodotto+errore che ricorrono (≥{0} segnalazioni): se persistono '
                        'nonostante le soluzioni dichiarate, le azioni correttive vanno verificate.'
                        ).format(RECUR_THRESHOLD)).pack(fill=tk.X)
        rc_cols = ('product', 'problem', 'count', 'reopened', 'nsol', 'solutions', 'first', 'last', 'flag')
        rc_heads = (L('tur_c_product', 'Prodotto'), L('tur_c_defect', 'Errore'),
                    L('tur_rc_count', 'Occorrenze'), L('tur_rc_reopen', 'Riaperture'),
                    L('tur_rc_nsol', 'N. sol.'), L('tur_rc_solutions', 'Soluzioni dichiarate'),
                    L('tur_rc_first', 'Prima'), L('tur_rc_last', 'Ultima'), L('tur_rc_flag', 'Segnale'))
        rc_w = (140, 210, 80, 80, 55, 300, 90, 90, 210)
        rc_wrap = ttk.Frame(fr_rec)
        rc_wrap.pack(fill=tk.BOTH, expand=True)
        tv_rec = ttk.Treeview(rc_wrap, columns=rc_cols, show='headings')
        for c, h, w in zip(rc_cols, rc_heads, rc_w):
            tv_rec.heading(c, text=h)
            tv_rec.column(c, width=w,
                          anchor='w' if c in ('product', 'problem', 'solutions', 'flag') else 'center')
        tv_rec.tag_configure('recurring', background='#FFEBEE', foreground='#B71C1C')
        rvsb = ttk.Scrollbar(rc_wrap, orient='vertical', command=tv_rec.yview)
        rhsb = ttk.Scrollbar(rc_wrap, orient='horizontal', command=tv_rec.xview)
        tv_rec.configure(yscrollcommand=rvsb.set, xscrollcommand=rhsb.set)
        tv_rec.grid(row=0, column=0, sticky='nsew')
        rvsb.grid(row=0, column=1, sticky='ns')
        rhsb.grid(row=1, column=0, sticky='ew')
        rc_wrap.rowconfigure(0, weight=1)
        rc_wrap.columnconfigure(0, weight=1)
        self._sum_trees['recurr'] = tv_rec

        # periodo con selettore granularità
        fr = ttk.Frame(sub)
        sub.add(fr, text=L('tur_sum_period', 'Per periodo'))
        top = ttk.Frame(fr)
        top.pack(fill=tk.X, pady=3)
        ttk.Label(top, text=L('tur_group_by', 'Raggruppa per:')).pack(side='left', padx=6)
        self._v_period = tk.StringVar(value='month')
        self._period_labels = {L('tur_gb_day', 'Giorno'): 'day',
                               L('tur_gb_week', 'Settimana'): 'week',
                               L('tur_gb_month', 'Mese'): 'month'}
        cb = ttk.Combobox(top, width=14, state='readonly', values=list(self._period_labels.keys()))
        cb.current(2)
        cb.pack(side='left')
        cb.bind('<<ComboboxSelected>>',
                lambda e: self._fill_period(self._period_labels.get(cb.get(), 'month')))
        tv = ttk.Treeview(fr, columns=('key', 'count'), show='headings')
        tv.heading('key', text=L('tur_c_period', 'Periodo'))
        tv.heading('count', text=L('tur_sc_count', 'Segnalazioni'))
        tv.column('key', width=200, anchor='w')
        tv.column('count', width=120, anchor='center')
        tv.pack(fill=tk.BOTH, expand=True)
        self._sum_trees['period'] = tv

    # ── Dati ──
    def _load_filter_options(self):
        try:
            clients = [r['ClientName'] for r in _rows(self.db.conn, _Q_FILTER_CLIENTS)]
            self._cb_client['values'] = [''] + clients
            self._cb_client.current(0)
            probs = _rows(self.db.conn, _Q_FILTER_PROBLEMS)
            self._problem_map = {'': None}
            vals = ['']
            for p in probs:
                self._problem_map[p['ProblemDescription']] = p['TouchUpProblemId']
                vals.append(p['ProblemDescription'])
            self._cb_problem['values'] = vals
            self._cb_problem.current(0)
        except Exception as exc:
            logger.error(f"TouchUp report filter options: {exc}", exc_info=True)

    def _parse_date(self, text):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text.strip(), fmt).date()
            except ValueError:
                pass
        return None

    def _do_search(self):
        L = self.lang.get
        d_from = self._parse_date(self._v_from.get())
        d_to = self._parse_date(self._v_to.get())
        if not d_from or not d_to:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('touchup_report_bad_dates', 'Date non valide (gg/mm/aaaa).'),
                                   parent=self)
            return
        if d_from > d_to:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('touchup_report_date_order', 'La data "Da" deve precedere "A".'),
                                   parent=self)
            return
        filters = {
            'start': datetime.datetime.combine(d_from, datetime.time.min),
            'end': datetime.datetime.combine(d_to + datetime.timedelta(days=1), datetime.time.min),
            'client': self._v_client.get().strip() or None,
            'product': self._v_product.get().strip() or None,
            'problem_id': self._problem_map.get(self._v_problem.get()),
            'status': self._status_map.get(self._v_status.get()) or None,
            'boss_only': bool(self._v_boss.get()),
        }
        try:
            self._data = run_report(self.db.conn, filters)
        except Exception as exc:
            logger.error(f"TouchUp report query: {exc}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), f"{exc}", parent=self)
            return
        self._meta = {'date_from': d_from, 'date_to': d_to}
        self._render()

    def _render(self):
        self._fill_kpi()
        self._fill_detail()
        self._fill_summary()
        state = 'normal' if self._data and self._data['reports'] else 'disabled'
        self._btn_xls.config(state=state)
        self._btn_pdf.config(state=state)
        self._btn_ai.config(state=state)

    def _fill_kpi(self):
        L = self.lang.get
        k = self._data['kpi']
        avg = k['avg_resp_min'] if k['avg_resp_min'] is not None else '-'
        pct = k['pct_within'] if k['pct_within'] is not None else '-'
        self._kpi.config(text=L(
            'touchup_report_kpi',
            'Totale: {total}   Aperte: {open}   Chiuse: {closed}   Riaperte: {reopened}   '
            'Escalation capo: {boss}   Tempo medio 1ª risposta: {avg} min   '
            '% entro {thr} min: {pct}').format(
            total=k['total'], open=k['open'], closed=k['closed'], reopened=k['reopened'],
            boss=k['boss'], avg=avg, thr=k['threshold_min'], pct=pct))

    def _fill_detail(self):
        self._t_det.delete(*self._t_det.get_children())
        for rep in self._data['reports']:
            probs = self._data['problems_by_rep'].get(rep['TouchUpReportId'], []) or [{}]
            for p in probs:
                tags = []
                if (rep['ReopenCount'] or 0) > 0:
                    tags.append('reopen')
                if rep['BossEscalated']:
                    tags.append('boss')
                self._t_det.insert('', 'end', values=(
                    _fmt_dt(rep['CreatedAt']), rep['TouchUpReportId'], rep['Status'],
                    rep['ClientsStr'], rep['ProductsStr'], rep['OrdersStr'], rep['LabelsStr'],
                    p.get('ProblemDescription', ''), p.get('Severity', '') or '',
                    p.get('Dept', '') or '', rep['CreatedByUser'] or '',
                    _mins(rep['FirstResponseSeconds']), rep['ReopenCount'] or 0,
                    'Sì' if rep['BossEscalated'] else '', rep['ActionsStr'],
                ), tags=tuple(tags))

    def _fill_summary(self):
        d = self._data
        for name, rows in (('defect', d['by_defect']), ('client', d['by_client']),
                           ('product', d['by_product']), ('dept', d['by_dept'])):
            tv = self._sum_trees[name]
            tv.delete(*tv.get_children())
            for r in rows:
                if name == 'defect':
                    tv.insert('', 'end', values=(r['key'], r['count'], r['pct'],
                                                 r['avg_resp_min'] if r['avg_resp_min'] is not None else '-',
                                                 r['reopened']))
                else:
                    tv.insert('', 'end', values=(r['key'], r['count'], r['pct']))
        # Ricorrenze prodotto × errore
        tvr = self._sum_trees.get('recurr')
        if tvr is not None:
            tvr.delete(*tvr.get_children())
            for r in d.get('by_product_problem', []):
                tvr.insert('', 'end', values=(
                    r['product'], r['problem'], r['count'], r['reopened'], r['n_solutions'],
                    r['solutions'], _fmt_dt(r['first']), _fmt_dt(r['last']), r['flag']),
                    tags=('recurring',) if r['recurring'] else ())
        self._fill_period(self._v_period.get())

    def _fill_period(self, mode):
        if not self._data:
            return
        self._v_period.set(mode)
        tv = self._sum_trees['period']
        tv.delete(*tv.get_children())
        rows = self._data.get(f'by_period_{mode}', [])
        for r in rows:
            tv.insert('', 'end', values=(r['key'], r['count']))

    def _export(self, kind):
        L = self.lang.get
        if not self._data or not self._data['reports']:
            messagebox.showinfo(L('info', 'Info'),
                                L('touchup_report_no_data', 'Nessun dato da esportare. Esegui una ricerca.'),
                                parent=self)
            return
        try:
            path = export_excel(self._data, self._meta) if kind == 'xls' \
                else export_pdf(self._data, self._meta)
        except Exception as exc:
            logger.error(f"TouchUp report export {kind}: {exc}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), f"{exc}", parent=self)
            return
        messagebox.showinfo(L('success', 'Fatto'),
                            L('touchup_report_saved', 'File salvato:\n{0}').format(path),
                            parent=self)
        try:
            os.startfile(path)
        except Exception:
            pass

    def _run_ai_analysis(self):
        """Analisi AI (Ollama locale) della validità delle soluzioni. I numeri sono
        già calcolati; il modello interpreta i dati esatti in un thread separato."""
        L = self.lang.get
        if not self._data or not self._data['reports']:
            return
        import touchup_ai
        import threading
        url, model = touchup_ai.get_config(self.db.conn)

        win = tk.Toplevel(self)
        win.title(L('touchup_ai_title', 'Analisi AI — validità soluzioni Touch-up'))
        win.geometry('840x620')
        win.transient(self)
        info = tk.Label(win, anchor='w', fg='#1F3864', padx=10, pady=8,
                        font=('Helvetica', 10, 'bold'),
                        text=L('touchup_ai_running',
                               'Analisi in corso sul modello locale {0} ({1})...').format(model, url))
        info.pack(fill=tk.X)
        txt = tk.Text(win, wrap='word', font=('Segoe UI', 10))
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        txt.insert('1.0', L('touchup_ai_wait',
                            'Attendere: il modello sta analizzando i dati (può richiedere qualche minuto).'))
        txt.config(state='disabled')
        self._ai_result_text = None
        bar = ttk.Frame(win)
        bar.pack(fill=tk.X, padx=8, pady=6)
        btn_email = ttk.Button(
            bar, text=L('touchup_ai_email', '📧 Invia a me via email'),
            command=lambda: self._email_ai_analysis(win, btn_email), state='disabled')
        btn_email.pack(side=tk.LEFT)
        ttk.Button(bar, text=L('btn_close', 'Chiudi'), command=win.destroy).pack(side=tk.RIGHT)

        def _done(result, err):
            if not win.winfo_exists():
                return
            info.config(text=L('touchup_ai_done', 'Analisi completata — {0} ({1})').format(model, url))
            txt.config(state='normal')
            txt.delete('1.0', 'end')
            txt.insert('1.0', f"⚠️ {err}" if err else result)
            txt.config(state='disabled')
            if not err and result:
                self._ai_result_text = result
                btn_email.config(state='normal')

        def worker():
            try:
                res = touchup_ai.analyze_recurrence(self._data, self._meta, url, model)
                self.after(0, lambda: _done(res, None))
            except Exception as e:
                emsg = str(e)
                self.after(0, lambda: _done(None, emsg))

        threading.Thread(target=worker, daemon=True).start()

    def _email_ai_analysis(self, win, btn):
        """Invia all'utente loggato l'analisi AI: corpo HTML formattato +
        allegato Excel (analisi + dati grezzi in tab separati)."""
        L = self.lang.get
        if not self._ai_result_text:
            return
        if not self.user_email:
            messagebox.showwarning(
                L('warning', 'Attenzione'),
                L('touchup_ai_no_email',
                  'Email utente non disponibile: impossibile inviare l\'analisi.'),
                parent=win)
            return

        btn.config(state='disabled')
        import threading
        ai_text = self._ai_result_text
        data, meta = self._data, self._meta
        user_email = self.user_email
        user_name = self.user_name

        def _restore():
            if win.winfo_exists():
                btn.config(state='normal')

        def _ok():
            _restore()
            messagebox.showinfo(
                L('success', 'Fatto'),
                L('touchup_ai_email_sent', 'Email inviata a {0}').format(user_email),
                parent=win if win.winfo_exists() else self)

        def _fail(msg):
            _restore()
            messagebox.showerror(
                L('error', 'Errore'),
                L('touchup_ai_email_err', 'Invio email fallito:\n{0}').format(msg),
                parent=win if win.winfo_exists() else self)

        def worker():
            try:
                xls_path = export_ai_excel(ai_text, data, meta)
                html_body = _ai_text_to_html(ai_text, meta, user_name)
                import utils
                subject = L('touchup_ai_email_subject',
                            'Analisi AI Touch-up {0} → {1}').format(
                    meta['date_from'].strftime('%d/%m/%Y'),
                    meta['date_to'].strftime('%d/%m/%Y'))
                utils.send_email(recipients=[user_email], subject=subject,
                                 body=html_body, is_html=True, attachments=[xls_path])
                self.after(0, _ok)
            except Exception as e:
                emsg = str(e)
                logger.error(f"TouchUp AI email: {emsg}", exc_info=True)
                self.after(0, lambda: _fail(emsg))

        threading.Thread(target=worker, daemon=True).start()


def open_touchup_report(master, db, lang, user_email=None, user_name=None):
    """Entry point: apre la finestra Report Touch-up."""
    TouchUpReportWindow(master, db, lang, user_email=user_email, user_name=user_name)
