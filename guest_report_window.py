# -*- coding: utf-8 -*-
"""
Guest Report Window — Filter form + professional Excel/PDF generation.

Provides:
  - Date range filter (From / To)
  - Company filter (ALL or specific)
  - Visitor filter (ALL or specific)
  - Generate Excel button (grouped by company, period, visitor with services)
  - Generate PDF button (professional English report with groupings)
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from tkcalendar import DateEntry
import logging
import os

logger = logging.getLogger("TraceabilityRS")

# ================================================================
# SQL QUERY — visitors with shuttle/hotel services
# ================================================================
VISITORS_QUERY = """
    SELECT
        v.VisitorId,
        v.GuestName,
        v.CompanyName,
        v.StartVisit,
        v.EndVisit,
        v.SponsorGuy,
        v.Pourpose,
        CASE
            WHEN EXISTS (
                SELECT 1 FROM Employee.dbo.VisitorArrivalDetails vad
                WHERE vad.VisitorId = v.VisitorId
            ) THEN 1 ELSE 0
        END AS HasPickup,
        (
            SELECT TOP 1 vsd.Name
            FROM Employee.dbo.VisitorArrivalDetails vad
            INNER JOIN Employee.dbo.VisitorSupportersData vsd
                ON vsd.SupporterDataId = vad.HotelId
                AND vsd.SupporterTypeID = 1
            WHERE vad.VisitorId = v.VisitorId
        ) AS HotelName,
        (
            SELECT TOP 1 vsd.Name
            FROM Employee.dbo.VisitorArrivalDetails vad
            INNER JOIN Employee.dbo.VisitorSupportersData vsd
                ON vsd.SupporterDataId = vad.HotelId
                AND vsd.SupporterTypeID = 2
            WHERE vad.VisitorId = v.VisitorId
        ) AS ShuttleName
    FROM Employee.dbo.Visitors v
    WHERE CAST(v.StartVisit AS DATE) <= ?
      AND CAST(v.EndVisit AS DATE) >= ?
"""


class GuestReportWindow(tk.Toplevel):
    """Filter dialog for generating guest Excel/PDF reports."""

    # Color palette
    BG_MAIN = "#F0F4F8"
    BG_HEADER = "#1F4E79"
    BG_CARD = "#FFFFFF"
    FG_TITLE = "#FFFFFF"
    FG_SUBTITLE = "#B8D4E8"
    FG_LABEL = "#2C3E50"
    FG_HINT = "#7F8C8D"
    ACCENT = "#2E75B6"
    ACCENT_HOVER = "#1A5F9E"
    BTN_EXCEL_BG = "#217346"
    BTN_EXCEL_HOVER = "#1B5E38"
    BTN_PDF_BG = "#C0392B"
    BTN_PDF_HOVER = "#A93226"
    BTN_CLOSE_BG = "#5D6D7E"
    BTN_CLOSE_HOVER = "#4A5A6A"
    BORDER_COLOR = "#D5DFE9"

    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.title("Visitor Report")
        self.geometry("560x420")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.configure(bg=self.BG_MAIN)

        self._companies = []
        self._visitors = []

        self._build_ui()
        self._load_filter_data()

        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ================================================================
    # UI
    # ================================================================
    def _make_button(self, parent, text, bg, hover_bg, command, icon_char=""):
        """Create a flat, modern button with hover effect."""
        btn = tk.Button(
            parent, text=f"{icon_char}  {text}" if icon_char else text,
            font=("Segoe UI Semibold", 10), fg="#FFFFFF", bg=bg,
            activebackground=hover_bg, activeforeground="#FFFFFF",
            relief="flat", cursor="hand2", bd=0,
            padx=16, pady=7, command=command
        )
        btn.bind("<Enter>", lambda e: btn.configure(bg=hover_bg))
        btn.bind("<Leave>", lambda e: btn.configure(bg=bg))
        return btn

    def _build_ui(self):
        # ── Header bar ──
        header = tk.Frame(self, bg=self.BG_HEADER, height=72)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header, text="📋  Visitor Report", font=("Segoe UI Semibold", 16),
            fg=self.FG_TITLE, bg=self.BG_HEADER, anchor="w"
        ).pack(side="left", padx=20, pady=(14, 0))

        tk.Label(
            header, text="Select filters and generate a report",
            font=("Segoe UI", 9), fg=self.FG_SUBTITLE, bg=self.BG_HEADER, anchor="w"
        ).pack(side="left", padx=(0, 20), pady=(18, 0))

        # ── Card ──
        card = tk.Frame(self, bg=self.BG_CARD, highlightbackground=self.BORDER_COLOR,
                        highlightthickness=1, padx=20, pady=18)
        card.pack(fill="both", expand=True, padx=18, pady=(14, 10))

        # Section title
        tk.Label(
            card, text="FILTERS", font=("Segoe UI Semibold", 9),
            fg=self.ACCENT, bg=self.BG_CARD, anchor="w"
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 8))

        # Thin accent line under section title
        accent_line = tk.Frame(card, bg=self.ACCENT, height=2)
        accent_line.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(0, 12))

        # Row 2 — Date From / To
        lbl_font = ("Segoe UI", 10)
        lbl_kwargs = dict(font=lbl_font, fg=self.FG_LABEL, bg=self.BG_CARD)

        tk.Label(card, text="From:", **lbl_kwargs).grid(row=2, column=0, sticky="w", padx=(0, 8), pady=6)
        self.date_from = DateEntry(
            card, width=14, font=("Segoe UI", 10),
            background=self.ACCENT, foreground="white", headersbackground=self.BG_HEADER,
            headersforeground="white", selectbackground=self.ACCENT,
            borderwidth=1, date_pattern="yyyy-mm-dd"
        )
        first_of_month = datetime.now().replace(day=1).date()
        self.date_from.set_date(first_of_month)
        self.date_from.grid(row=2, column=1, sticky="w", padx=(0, 15), pady=6)

        tk.Label(card, text="To:", **lbl_kwargs).grid(row=2, column=2, sticky="w", padx=(0, 8), pady=6)
        self.date_to = DateEntry(
            card, width=14, font=("Segoe UI", 10),
            background=self.ACCENT, foreground="white", headersbackground=self.BG_HEADER,
            headersforeground="white", selectbackground=self.ACCENT,
            borderwidth=1, date_pattern="yyyy-mm-dd"
        )
        self.date_to.set_date(datetime.now().date())
        self.date_to.grid(row=2, column=3, sticky="w", pady=6)

        # Row 3 — Company
        tk.Label(card, text="Company:", **lbl_kwargs).grid(row=3, column=0, sticky="w", padx=(0, 8), pady=6)
        self.company_var = tk.StringVar(value="ALL")
        self.company_combo = ttk.Combobox(
            card, textvariable=self.company_var, state="readonly",
            width=38, font=("Segoe UI", 10)
        )
        self.company_combo.grid(row=3, column=1, columnspan=3, sticky="ew", pady=6)
        self.company_combo.bind("<<ComboboxSelected>>", self._on_company_changed)

        # Row 4 — Visitor
        tk.Label(card, text="Visitor:", **lbl_kwargs).grid(row=4, column=0, sticky="w", padx=(0, 8), pady=6)
        self.visitor_var = tk.StringVar(value="ALL")
        self.visitor_combo = ttk.Combobox(
            card, textvariable=self.visitor_var, state="readonly",
            width=38, font=("Segoe UI", 10)
        )
        self.visitor_combo.grid(row=4, column=1, columnspan=3, sticky="ew", pady=6)

        card.columnconfigure(1, weight=1)
        card.columnconfigure(3, weight=1)

        # ── Button bar ──
        btn_bar = tk.Frame(self, bg=self.BG_MAIN)
        btn_bar.pack(fill="x", padx=18, pady=(0, 14))

        self._make_button(
            btn_bar, "Generate Excel", self.BTN_EXCEL_BG, self.BTN_EXCEL_HOVER,
            self._generate_excel, "📊"
        ).pack(side="left", padx=(0, 8))

        self._make_button(
            btn_bar, "Generate PDF", self.BTN_PDF_BG, self.BTN_PDF_HOVER,
            self._generate_pdf, "📄"
        ).pack(side="left", padx=(0, 8))

        self._make_button(
            btn_bar, "Close", self.BTN_CLOSE_BG, self.BTN_CLOSE_HOVER,
            self.destroy
        ).pack(side="right")

    # ================================================================
    # DATA LOADING
    # ================================================================
    def _load_filter_data(self):
        """Load distinct companies and visitors for combo boxes."""
        try:
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT DISTINCT CompanyName
                FROM Employee.dbo.Visitors
                WHERE CompanyName IS NOT NULL AND CompanyName != ''
                ORDER BY CompanyName
            """)
            self._companies = [r[0] for r in cursor.fetchall()]
            self.company_combo["values"] = ["ALL"] + self._companies

            cursor.execute("""
                SELECT DISTINCT GuestName
                FROM Employee.dbo.Visitors
                WHERE GuestName IS NOT NULL AND GuestName != ''
                ORDER BY GuestName
            """)
            self._visitors = [r[0] for r in cursor.fetchall()]
            self.visitor_combo["values"] = ["ALL"] + self._visitors
            cursor.close()
        except Exception as e:
            logger.error(f"GuestReport: errore caricamento filtri: {e}")

    def _on_company_changed(self, event=None):
        """When company changes, reload visitor list filtered."""
        company = self.company_var.get()
        try:
            cursor = self.db.conn.cursor()
            if company == "ALL":
                cursor.execute("""
                    SELECT DISTINCT GuestName FROM Employee.dbo.Visitors
                    WHERE GuestName IS NOT NULL AND GuestName != ''
                    ORDER BY GuestName
                """)
            else:
                cursor.execute("""
                    SELECT DISTINCT GuestName FROM Employee.dbo.Visitors
                    WHERE CompanyName = ? AND GuestName IS NOT NULL AND GuestName != ''
                    ORDER BY GuestName
                """, (company,))
            visitors = [r[0] for r in cursor.fetchall()]
            self.visitor_combo["values"] = ["ALL"] + visitors
            self.visitor_var.set("ALL")
            cursor.close()
        except Exception as e:
            logger.error(f"GuestReport: errore filtro visitatori: {e}")

    def _fetch_data(self):
        """Fetch visitors based on current filters. Returns list of dicts."""
        date_from = self.date_from.get_date()
        date_to = self.date_to.get_date()
        company = self.company_var.get()
        visitor = self.visitor_var.get()

        query = VISITORS_QUERY
        params = [date_to, date_from]

        if company != "ALL":
            query += " AND v.CompanyName = ?"
            params.append(company)
        if visitor != "ALL":
            query += " AND v.GuestName = ?"
            params.append(visitor)

        query += " ORDER BY v.CompanyName, v.StartVisit, v.GuestName"

        results = []
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, tuple(params))
            for r in cursor.fetchall():
                results.append({
                    "VisitorId": r.VisitorId,
                    "GuestName": r.GuestName or "",
                    "CompanyName": r.CompanyName or "",
                    "StartVisit": r.StartVisit,
                    "EndVisit": r.EndVisit,
                    "Sponsor": r.SponsorGuy or "",
                    "Purpose": r.Pourpose or "",
                    "HasPickup": bool(r.HasPickup),
                    "HotelName": r.HotelName or "",
                    "ShuttleName": r.ShuttleName or "",
                })
            cursor.close()
        except Exception as e:
            logger.error(f"GuestReport: errore fetch dati: {e}")
            messagebox.showerror("Error", f"Database error: {e}")
        return results

    # ================================================================
    # EXCEL GENERATION
    # ================================================================
    def _generate_excel(self):
        """Generate a professional Excel report grouped by company/period/visitor."""
        data = self._fetch_data()
        if not data:
            messagebox.showinfo("Info", "No visitors found for the selected filters.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error", "openpyxl not installed.\nRun: pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Visitor Report"

        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        group_font = Font(name="Arial", size=11, bold=True, color="1F4E79")
        group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        data_font = Font(name="Arial", size=10)
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )
        center = Alignment(horizontal="center", vertical="center")
        wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Column widths
        col_widths = [32, 28, 14, 14, 26, 22, 22, 14]
        columns = [
            "Company", "Visitor Name", "Arrival", "Departure",
            "Purpose", "Hotel", "Shuttle", "Pickup"
        ]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Title row
        date_from_str = self.date_from.get_date().strftime("%d/%m/%Y")
        date_to_str = self.date_to.get_date().strftime("%d/%m/%Y")
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"Visitor Report — {date_from_str} to {date_to_str}"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Subtitle
        ws.merge_cells("A2:H2")
        sub = ws["A2"]
        sub.value = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')} — Vandewiele Romania"
        sub.font = Font(name="Arial", size=9, italic=True, color="666666")
        sub.alignment = Alignment(horizontal="center")

        row = 4

        # Group by company
        from collections import OrderedDict
        companies = OrderedDict()
        for v in data:
            companies.setdefault(v["CompanyName"], []).append(v)

        for company, visitors in companies.items():
            # Company header
            ws.merge_cells(f"A{row}:H{row}")
            cell = ws.cell(row=row, column=1, value=f"🏢  {company} ({len(visitors)} visitors)")
            cell.font = group_font
            cell.fill = group_fill
            cell.alignment = wrap
            ws.row_dimensions[row].height = 24
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = group_fill
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            # Column headers
            for ci, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row, column=ci, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
            ws.row_dimensions[row].height = 22
            row += 1

            # Data rows
            for idx, v in enumerate(visitors):
                arr = v["StartVisit"].strftime("%d/%m/%Y") if v["StartVisit"] else ""
                dep = v["EndVisit"].strftime("%d/%m/%Y") if v["EndVisit"] else ""
                pickup = "Yes" if v["HasPickup"] else "No"
                values = [
                    v["CompanyName"], v["GuestName"], arr, dep,
                    v["Purpose"], v["HotelName"] or "—", v["ShuttleName"] or "—", pickup
                ]
                fill = alt_fill if idx % 2 == 0 else None
                for ci, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center if ci in (3, 4, 8) else wrap
                    if fill:
                        cell.fill = fill
                ws.row_dimensions[row].height = 20
                row += 1

            row += 1  # spacing

        # Summary row
        ws.merge_cells(f"A{row}:H{row}")
        summary = ws.cell(row=row, column=1,
                          value=f"Total: {len(data)} visitors from {len(companies)} companies")
        summary.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        summary.alignment = Alignment(horizontal="right")

        # Print settings
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Save
        temp_dir = r"C:\Temp"
        os.makedirs(temp_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(temp_dir, f"Visitor_Report_{ts}.xlsx")
        wb.save(path)
        logger.info(f"GuestReport: Excel salvato: {path}")
        os.startfile(path)
        messagebox.showinfo("Success", f"Excel report generated:\n{path}")

    # ================================================================
    # PDF GENERATION
    # ================================================================
    def _generate_pdf(self):
        """Generate a professional PDF report in English."""
        data = self._fetch_data()
        if not data:
            messagebox.showinfo("Info", "No visitors found for the selected filters.")
            return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import cm, mm
            from reportlab.lib.colors import HexColor, white, black
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            messagebox.showerror("Error", "reportlab not installed.\nRun: pip install reportlab")
            return

        # Register Unicode font
        font_name = "Helvetica"
        font_bold = "Helvetica-Bold"
        arial_path = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts", "arial.ttf")
        arialbd_path = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts", "arialbd.ttf")
        if os.path.isfile(arial_path) and os.path.isfile(arialbd_path):
            try:
                pdfmetrics.registerFont(TTFont("ArialPDF", arial_path))
                pdfmetrics.registerFont(TTFont("ArialPDF-Bold", arialbd_path))
                font_name = "ArialPDF"
                font_bold = "ArialPDF-Bold"
            except Exception:
                pass

        # File setup
        temp_dir = r"C:\Temp"
        os.makedirs(temp_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = os.path.join(temp_dir, f"Visitor_Report_{ts}.pdf")

        doc = SimpleDocTemplate(
            pdf_path, pagesize=landscape(A4),
            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
            topMargin=1.5 * cm, bottomMargin=1.5 * cm
        )

        # Colors
        DARK_BLUE = HexColor("#1F4E79")
        MED_BLUE = HexColor("#2E75B6")
        LIGHT_BLUE = HexColor("#D6E4F0")
        ALT_ROW = HexColor("#F2F7FB")
        DARK_GRAY = HexColor("#333333")

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"],
            fontName=font_bold, fontSize=18, textColor=DARK_BLUE,
            spaceAfter=6, alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle", parent=styles["Normal"],
            fontName=font_name, fontSize=9, textColor=DARK_GRAY,
            alignment=TA_CENTER, spaceAfter=12
        )
        group_style = ParagraphStyle(
            "GroupHeader", parent=styles["Heading2"],
            fontName=font_bold, fontSize=12, textColor=DARK_BLUE,
            spaceBefore=14, spaceAfter=6
        )
        cell_style = ParagraphStyle(
            "CellText", parent=styles["Normal"],
            fontName=font_name, fontSize=8, leading=10
        )
        cell_bold = ParagraphStyle(
            "CellBold", parent=cell_style,
            fontName=font_bold
        )
        footer_style = ParagraphStyle(
            "Footer", parent=styles["Normal"],
            fontName=font_name, fontSize=7, textColor=DARK_GRAY,
            alignment=TA_CENTER, spaceBefore=20
        )

        elements = []

        # Logo
        logo_path = os.path.join(os.path.dirname(__file__), "Logo.png")
        if os.path.isfile(logo_path):
            try:
                elements.append(Image(logo_path, width=4 * cm, height=1.5 * cm))
                elements.append(Spacer(1, 6))
            except Exception:
                pass

        # Title
        date_from_str = self.date_from.get_date().strftime("%d/%m/%Y")
        date_to_str = self.date_to.get_date().strftime("%d/%m/%Y")
        elements.append(Paragraph(
            f"Visitor Report — {date_from_str} to {date_to_str}", title_style
        ))
        elements.append(Paragraph(
            f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} — "
            f"Vandewiele Romania SRL", subtitle_style
        ))

        # Group by company
        from collections import OrderedDict
        companies = OrderedDict()
        for v in data:
            companies.setdefault(v["CompanyName"], []).append(v)

        col_widths = [5.5 * cm, 3.2 * cm, 3.2 * cm, 4.5 * cm, 3.8 * cm, 3.8 * cm, 2.2 * cm]
        headers = ["Visitor", "Arrival", "Departure", "Purpose", "Hotel", "Shuttle", "Pickup"]

        for company, visitors in companies.items():
            elements.append(Paragraph(
                f"{company}  ({len(visitors)} visitor{'s' if len(visitors) != 1 else ''})",
                group_style
            ))

            # Build table data
            header_row = [Paragraph(h, cell_bold) for h in headers]
            table_data = [header_row]

            for v in visitors:
                arr = v["StartVisit"].strftime("%d/%m/%Y") if v["StartVisit"] else ""
                dep = v["EndVisit"].strftime("%d/%m/%Y") if v["EndVisit"] else ""
                pickup = "Yes" if v["HasPickup"] else "No"
                row_data = [
                    Paragraph(v["GuestName"], cell_style),
                    Paragraph(arr, cell_style),
                    Paragraph(dep, cell_style),
                    Paragraph(v["Purpose"], cell_style),
                    Paragraph(v["HotelName"] or "—", cell_style),
                    Paragraph(v["ShuttleName"] or "—", cell_style),
                    Paragraph(pickup, cell_style),
                ]
                table_data.append(row_data)

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), MED_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#B4C6E7")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
            # Alternating row colors
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), ALT_ROW))
            tbl.setStyle(TableStyle(style_cmds))
            elements.append(tbl)
            elements.append(Spacer(1, 8))

        # Summary
        elements.append(Spacer(1, 10))
        summary_style = ParagraphStyle(
            "Summary", parent=styles["Normal"],
            fontName=font_bold, fontSize=10, textColor=DARK_BLUE,
            alignment=TA_RIGHT
        )
        elements.append(Paragraph(
            f"Total: {len(data)} visitors from {len(companies)} companies",
            summary_style
        ))

        # Footer
        elements.append(Paragraph(
            "This report was automatically generated by TraceabilityRS — Vandewiele Romania",
            footer_style
        ))

        doc.build(elements)
        logger.info(f"GuestReport: PDF salvato: {pdf_path}")
        os.startfile(pdf_path)
        messagebox.showinfo("Success", f"PDF report generated:\n{pdf_path}")
