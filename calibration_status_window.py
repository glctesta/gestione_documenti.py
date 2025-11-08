import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
from datetime import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.comments import Comment
import subprocess
import platform

logger = logging.getLogger(__name__)


class CalibrationStatusWindow(tk.Toplevel):
    def __init__(self, parent, db_object, language_manager):
        super().__init__(parent)
        self.parent = parent
        self.db = db_object
        self.lang = language_manager

        self.title(self.lang.get('calibration_status_title', "Stato Calibrazioni"))
        self.geometry("1000x700")
        self.transient(parent)
        self.grab_set()

        self._create_widgets()
        self._load_data()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Notebook per le tre categorie
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Tab 1: Calibrazioni scadute o mancanti
        self.expired_frame = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(self.expired_frame, text=self.lang.get('calibration_expired', "Scadute/Mancanti"))

        # Tab 2: Calibrazioni in scadenza entro 30 giorni
        self.expiring_30_frame = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(self.expiring_30_frame, text=self.lang.get('calibration_expiring_30', "In Scadenza (≤30 gg)"))

        # Tab 3: Calibrazioni in scadenza oltre 30 giorni
        self.expiring_over_30_frame = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(self.expiring_over_30_frame,
                          text=self.lang.get('calibration_expiring_over_30', "In Scadenza (>30 gg)"))

        # Crea le treeview per ogni tab
        self._create_treeview(self.expired_frame, 'expired')
        self._create_treeview(self.expiring_30_frame, 'expiring_30')
        self._create_treeview(self.expiring_over_30_frame, 'expiring_over_30')

        # Pulsanti
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(
            button_frame,
            text=self.lang.get('generate_excel_report', "Genera Report Excel"),
            command=self._generate_excel_report
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text=self.lang.get('send_email_report', "Invia Report Email"),
            command=self._send_email_report
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text=self.lang.get('refresh', "Aggiorna"),
            command=self._load_data
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text=self.lang.get('close', "Chiudi"),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)

    def _create_treeview(self, parent_frame, tree_type):
        # Treeview con scrollbar
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        columns = ('nr', 'equipment', 'calibrated_on', 'expire_on', 'days_to_expire')
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=scrollbar.set
        )

        # Configurazione colonne
        tree.heading('nr', text='#')
        tree.heading('equipment', text=self.lang.get('equipment', 'Equipaggiamento'))
        tree.heading('calibrated_on', text=self.lang.get('calibration_date', 'Data Calibrazione'))
        tree.heading('expire_on', text=self.lang.get('expiry_date', 'Data Scadenza'))
        tree.heading('days_to_expire', text=self.lang.get('days_to_expire', 'Giorni alla Scadenza'))

        tree.column('nr', width=50, anchor='center')
        tree.column('equipment', width=400)
        tree.column('calibrated_on', width=120, anchor='center')
        tree.column('expire_on', width=120, anchor='center')
        tree.column('days_to_expire', width=120, anchor='center')

        tree.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=tree.yview)

        # Salva riferimento alla treeview
        if tree_type == 'expired':
            self.expired_tree = tree
        elif tree_type == 'expiring_30':
            self.expiring_30_tree = tree
        else:
            self.expiring_over_30_tree = tree

    def _load_data(self):
        """Carica i dati nelle tre categorie"""
        try:
            # Carica dati scaduti/mancanti
            expired_data = self.db.get_calibration_expired()
            self._populate_treeview(self.expired_tree, expired_data)

            # Carica dati in scadenza entro 30 giorni
            expiring_30_data = self.db.get_calibration_expiring_30days()
            self._populate_treeview(self.expiring_30_tree, expiring_30_data)

            # Carica dati in scadenza oltre 30 giorni
            expiring_over_30_data = self.db.get_calibration_expiring_over_30days()
            self._populate_treeview(self.expiring_over_30_tree, expiring_over_30_data)

            # Aggiorna i titoli delle tab con i conteggi
            self.notebook.tab(0,
                              text=f"{self.lang.get('calibration_expired', 'Scadute/Mancanti')} ({len(expired_data)})")
            self.notebook.tab(1,
                              text=f"{self.lang.get('calibration_expiring_30', 'In Scadenza (≤30 gg)')} ({len(expiring_30_data)})")
            self.notebook.tab(2,
                              text=f"{self.lang.get('calibration_expiring_over_30', 'In Scadenza (>30 gg)')} ({len(expiring_over_30_data)})")

        except Exception as e:
            logger.error(f"Errore caricamento dati calibrazioni: {e}")
            messagebox.showerror("Errore", f"Impossibile caricare i dati: {e}", parent=self)

    def _populate_treeview(self, treeview, data):
        """Popola una treeview con i dati"""
        # Pulisci treeview
        for item in treeview.get_children():
            treeview.delete(item)

        # Popola con nuovi dati
        for row in data:
            calibrated_on = row.CalibratedOn.strftime('%d/%m/%Y') if row.CalibratedOn else 'N/A'
            expire_on = row.ExpireOn.strftime('%d/%m/%Y') if row.ExpireOn else 'N/A'

            treeview.insert('', 'end', values=(
                row.Nr,
                row.Equipment,
                calibrated_on,
                expire_on,
                row.DaysToExpire
            ))

    def _generate_excel_report(self):
        """Genera il report Excel con formattazione professionale"""
        try:
            # Crea directory se non esiste
            output_dir = r"C:\Temp"
            os.makedirs(output_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Calibration_Report_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Calibrazioni scadute/mancanti
                expired_data = self.db.get_calibration_expired()
                df_expired = self._prepare_dataframe(expired_data)
                df_expired.to_excel(writer, sheet_name='Expired', index=False, startrow=4)

                # Sheet 2: Calibrazioni in scadenza entro 30 giorni
                expiring_30_data = self.db.get_calibration_expiring_30days()
                df_expiring_30 = self._prepare_dataframe(expiring_30_data)
                df_expiring_30.to_excel(writer, sheet_name='Expiring ≤30 days', index=False, startrow=4)

                # Sheet 3: Calibrazioni in scadenza oltre 30 giorni
                expiring_over_30_data = self.db.get_calibration_expiring_over_30days()
                df_expiring_over_30 = self._prepare_dataframe(expiring_over_30_data)
                df_expiring_over_30.to_excel(writer, sheet_name='Expiring >30 days', index=False, startrow=4)

            # Applica formattazione
            self._format_excel(filepath, expired_data)

            messagebox.showinfo(
                "Successo",
                f"Report Excel generato con successo:\n{filepath}",
                parent=self
            )

            # Apri il file Excel automaticamente
            self._open_file(filepath)

        except Exception as e:
            logger.error(f"Errore generazione report Excel: {e}")
            messagebox.showerror("Errore", f"Impossibile generare il report: {e}", parent=self)

    def _format_excel(self, filepath, expired_data):
        """Applica formattazione professionale al file Excel"""
        try:
            wb = load_workbook(filepath)
            current_date = datetime.now().strftime("%d-%m-%Y")

            # Definisci gli stili
            title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell_font = Font(name='Calibri', size=10)

            title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

            # Colori per le righe alternate
            light_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            # Colori per i giorni alla scadenza
            red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            orange_fill = PatternFill(start_color='FFB84D', end_color='FFB84D', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFF176', end_color='FFF176', fill_type='solid')
            green_fill = PatternFill(start_color='81C784', end_color='81C784', fill_type='solid')

            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')

            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Logo path
            logo_path = 'logo.png'
            logo_exists = os.path.exists(logo_path)

            # Crea un dizionario per identificare le righe con calibrazione N/A nel tab Expired
            na_calibration_rows = {}
            if expired_data:
                for idx, row in enumerate(expired_data, 1):
                    if not row.CalibratedOn:
                        na_calibration_rows[idx] = True

            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]

                # ========================================
                # LAYOUT:
                # Righe 1-2: Riservate per il logo (colonne A-B)
                # Riga 3: Titolo (colonne A-F)
                # Riga 4: Vuota (spazio)
                # Riga 5: Headers
                # Riga 6+: Dati
                # ========================================

                # Aggiungi logo se esiste (area A1:B2)
                if logo_exists:
                    try:
                        img = XLImage(logo_path)
                        # Ridimensiona il logo per farlo stare nell'area A1:B2
                        # Altezza massima: circa 40 pixel (2 righe)
                        img.height = 40
                        # Mantieni le proporzioni
                        aspect_ratio = img.width / img.height
                        img.width = int(40 * aspect_ratio)

                        # Posiziona il logo in A1
                        img.anchor = 'A1'
                        ws.add_image(img)
                    except Exception as e:
                        logger.warning(f"Impossibile aggiungere logo: {e}")

                # Imposta altezza righe per il logo
                ws.row_dimensions[1].height = 20
                ws.row_dimensions[2].height = 20

                # Imposta larghezza colonne A e B (area logo)
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 10

                # Aggiungi titolo nella riga 3 (da A3 a F3)
                title_cell = ws['A3']
                title_cell.value = f"Calibration Warning List - {current_date}"
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = center_alignment

                # Unisci celle per il titolo (A3:F3)
                ws.merge_cells('A3:F3')

                # Applica bordo al titolo
                for cell in ws['A3:F3'][0]:
                    cell.border = thin_border

                # Imposta altezza riga del titolo
                ws.row_dimensions[3].height = 25

                # Riga 4 è vuota (spazio)
                ws.row_dimensions[4].height = 5

                # Formatta intestazioni (riga 5)
                header_row = 5
                for col_num, cell in enumerate(ws[header_row], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # Applica filtro automatico
                if ws.max_row > header_row:
                    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

                # Formatta dati (a partire dalla riga 6)
                for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), 1):
                    for col_num, cell in enumerate(row, 1):
                        cell.font = cell_font
                        cell.border = thin_border

                        # Allineamento
                        if col_num == 1 or col_num >= 3:  # Nr, Date, Days
                            cell.alignment = center_alignment
                        else:  # Equipment
                            cell.alignment = left_alignment

                        # Colore di sfondo alternato
                        if row_num % 2 == 0:
                            cell.fill = light_fill

                        # Aggiungi nota per calibrazioni N/A nel tab "Expired" (colonna C = Calibration Date)
                        if sheet_index == 0 and col_num == 3 and cell.value == 'N/A':
                            comment = Comment("This equipment has not been calibrated", "System")
                            cell.comment = comment

                        # Colore basato sui giorni alla scadenza (ultima colonna)
                        if col_num == ws.max_column and cell.value is not None:
                            try:
                                days = int(cell.value) if isinstance(cell.value, (int, float)) else None
                                if days is not None:
                                    if days < 0:  # Scaduto
                                        cell.fill = red_fill
                                        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                                    elif days <= 7:  # Entro 7 giorni
                                        cell.fill = orange_fill
                                        cell.font = Font(name='Calibri', size=10, bold=True)
                                    elif days <= 30:  # Entro 30 giorni
                                        cell.fill = yellow_fill
                                    else:  # Oltre 30 giorni
                                        cell.fill = green_fill
                            except (ValueError, TypeError):
                                pass

                # Imposta larghezza colonne dati
                ws.column_dimensions['C'].width = 50  # Equipment
                ws.column_dimensions['D'].width = 18  # Calibration Date
                ws.column_dimensions['E'].width = 18  # Expiry Date
                ws.column_dimensions['F'].width = 18  # Days to Expire

                # Congela la riga delle intestazioni
                ws.freeze_panes = f'A{header_row + 1}'

            wb.save(filepath)
            logger.info(f"Formattazione Excel completata: {filepath}")

        except Exception as e:
            logger.error(f"Errore durante la formattazione Excel: {e}")
            raise

    def _open_file(self, filepath):
        """Apre il file Excel con l'applicazione predefinita del sistema"""
        try:
            system = platform.system()

            if system == 'Windows':
                os.startfile(filepath)
            elif system == 'Darwin':  # macOS
                subprocess.run(['open', filepath])
            else:  # Linux
                subprocess.run(['xdg-open', filepath])

            logger.info(f"File aperto: {filepath}")

        except Exception as e:
            logger.warning(f"Impossibile aprire automaticamente il file: {e}")
            # Non è un errore critico, il file è comunque stato generato

    def _prepare_dataframe(self, data):
        """Prepara un DataFrame per l'export Excel"""
        if not data:
            return pd.DataFrame()

        records = []
        for row in data:
            records.append({
                'Nr': row.Nr,
                'Equipment': row.Equipment,
                'Calibration Date': row.CalibratedOn.strftime('%d/%m/%Y') if row.CalibratedOn else 'N/A',
                'Expiry Date': row.ExpireOn.strftime('%d/%m/%Y') if row.ExpireOn else 'N/A',
                'Days to Expire': row.DaysToExpire
            })

        df = pd.DataFrame(records)
        return df

    def _send_email_report(self):
        """Invia il report via email"""
        try:
            from utils import send_email, get_email_recipients

            # Recupera destinatari
            recipients = self.db.get_calibration_emails()
            if not recipients:
                messagebox.showwarning(
                    "Attenzione",
                    "Nessun destinatario configurato per le notifiche calibrazioni.",
                    parent=self
                )
                return

            # Genera il report in memoria
            excel_data = self._generate_excel_in_memory()
            if not excel_data:
                messagebox.showerror("Errore", "Impossibile generare il report per l'email.", parent=self)
                return

            # Prepara email
            subject = f"Calibration Status Report - {datetime.now().strftime('%Y-%m-%d')}"
            body = self._get_email_body()

            # Qui dovresti implementare l'invio email con allegato
            # Per ora mostriamo solo un messaggio di conferma
            messagebox.showinfo(
                "Email Pronta",
                f"Report preparato per l'invio a {len(recipients)} destinatari:\n{', '.join(recipients)}",
                parent=self
            )

        except Exception as e:
            logger.error(f"Errore preparazione email: {e}")
            messagebox.showerror("Errore", f"Impossibile preparare l'email: {e}", parent=self)

    def _generate_excel_in_memory(self):
        """Genera il report Excel in memoria (placeholder)"""
        # Implementazione per generare Excel in memoria
        return None

    def _get_email_body(self):
        """Restituisce il corpo dell'email in inglese"""
        return """
Dear Team,

Please find attached the latest calibration status report.

This report includes:
- Equipment with expired or missing calibrations (URGENT ATTENTION REQUIRED)
- Equipment with calibrations expiring within 30 days (PLANNING REQUIRED)
- Equipment with calibrations expiring after 30 days (FOR INFORMATION)

Please review the report and take appropriate actions to ensure all equipment remains properly calibrated.

Best regards,
Calibration Management System
"""
