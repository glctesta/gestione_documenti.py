"""
Viewer per visualizzare e aprire i report FAI salvati nel database
"""
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import tempfile
import os
import subprocess
import logging

logger = logging.getLogger(__name__)


class FaiReportsViewerWindow(tk.Toplevel):
    """Finestra per visualizzare e aprire report FAI salvati"""
    
    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        
        self.db = db
        self.lang = lang
        self.user_name = user_name
        
        self.title(lang.get('storico_validazioni_fai', "Storico Validazioni FAI"))
        self.geometry("1200x700")
        
        # Centra finestra
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.winfo_screenheight() // 2) - (700 // 2)
        self.geometry(f"1200x700+{x}+{y}")
        
        self._build_ui()
        self._load_data()
    
    def _build_ui(self):
        """Costruisce l'interfaccia"""
        # Frame filtri
        filters_frame = ttk.LabelFrame(self, text=self.lang.get('fai_filters_search', "Filtri Ricerca"), padding="10")
        filters_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Prima riga filtri
        row1 = ttk.Frame(filters_frame)
        row1.pack(fill=tk.X, pady=5)
        
        # Data Da
        ttk.Label(row1, text=self.lang.get('fai_date_from', "Data Da:")).pack(side=tk.LEFT, padx=5)
        from tkcalendar import DateEntry
        self.date_from_picker = DateEntry(
            row1,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        # Imposta data di default (30 giorni fa)
        default_from = datetime.now() - timedelta(days=30)
        self.date_from_picker.set_date(default_from)
        self.date_from_picker.pack(side=tk.LEFT, padx=5)
        
        # Data A
        ttk.Label(row1, text=self.lang.get('fai_date_to', "Data A:")).pack(side=tk.LEFT, padx=5)
        self.date_to_picker = DateEntry(
            row1,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        self.date_to_picker.set_date(datetime.now())
        self.date_to_picker.pack(side=tk.LEFT, padx=5)
        
        # Codice Prodotto
        ttk.Label(row1, text=self.lang.get('fai_product_code', "Codice Prodotto:")).pack(side=tk.LEFT, padx=5)
        self.product_var = tk.StringVar()
        self.product_entry = ttk.Entry(row1, textvariable=self.product_var, width=25)
        self.product_entry.pack(side=tk.LEFT, padx=5)
        
        # Operatore
        ttk.Label(row1, text=self.lang.get('fai_operator', "Operatore:")).pack(side=tk.LEFT, padx=5)
        self.operator_var = tk.StringVar()
        self.operator_entry = ttk.Entry(row1, textvariable=self.operator_var, width=20)
        self.operator_entry.pack(side=tk.LEFT, padx=5)
        
        # Pulsante Cerca
        ttk.Button(row1, text="🔍 " + self.lang.get('search', "Cerca"), command=self._load_data).pack(side=tk.LEFT, padx=10)
        ttk.Button(row1, text="🔄 " + self.lang.get('reset', "Reset"), command=self._reset_filters).pack(side=tk.LEFT, padx=5)
        
        # Treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview - FaiLogId is hidden but still stored
        columns = ('FaiLogId', 'FAI_Document', 'Data', 'Ordine', 'Prodotto', 'Operatore', 'Risultato', 'HasPDF')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set, displaycolumns=('FAI_Document', 'Data', 'Ordine', 'Prodotto', 'Operatore', 'Risultato', 'HasPDF'))
        
        # Configura colonne
        self.tree.heading('FaiLogId', text='ID')  # Hidden
        self.tree.heading('FAI_Document', text='FAI Document')
        self.tree.heading('Data', text='Data/Ora')
        self.tree.heading('Ordine', text='N. Ordine')
        self.tree.heading('Prodotto', text='Codice Prodotto')
        self.tree.heading('Operatore', text='Operatore')
        self.tree.heading('Risultato', text='Risultato')
        self.tree.heading('HasPDF', text='PDF')
        
        self.tree.column('FaiLogId', width=0, stretch=False)  # Hidden
        self.tree.column('FAI_Document', width=300)
        self.tree.column('Data', width=150, anchor='center')
        self.tree.column('Ordine', width=120, anchor='center')
        self.tree.column('Prodotto', width=200)
        self.tree.column('Operatore', width=150)
        self.tree.column('Risultato', width=100, anchor='center')
        self.tree.column('HasPDF', width=60, anchor='center')
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Grid
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Doppio click per aprire
        self.tree.bind('<Double-1>', lambda e: self._open_pdf())
        
        # Pulsanti azioni
        buttons_frame = ttk.Frame(self)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(buttons_frame, text=self.lang.get('fai_open_pdf', "📄 Apri PDF"), 
                  command=self._open_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text=self.lang.get('fai_delete_validation', "🗑️ Elimina Validazione"), 
                  command=self._delete_validation).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="📊 " + self.lang.get('export_excel', "Esporta Excel"), 
                  command=self._export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="❌ " + self.lang.get('close', "Chiudi"), 
                  command=self.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Status bar
        self.status_label = ttk.Label(self, text=self.lang.get('ready', "Pronto"), relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM)
    
    def _reset_filters(self):
        """Reset filtri"""
        default_from = datetime.now() - timedelta(days=30)
        self.date_from_picker.set_date(default_from)
        self.date_to_picker.set_date(datetime.now())
        self.product_var.set('')
        self.operator_var.set('')
        self._load_data()
    
    def _load_data(self):
        """Carica dati validazioni"""
        try:
            # Pulisci treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Query
            query = """
            SELECT DISTINCT 
                t.NrDocument + ' Rev.' + cast(t.Revision as nvarchar(4))+ ' issued on ' + format(t.RevisionDate,'d','it-it') as FAI_Document,
                l.FaiLogId,
                l.DateIn,
                o.Ordernumber as OrderId,
                p.productcode,
                l.Operator,
                l.IsOk,
                CASE WHEN l.DocVerification IS NOT NULL THEN 1 ELSE 0 END AS HasPDF
            FROM [Traceability_RS].[fai].[FaiLogs] l
            LEFT JOIN [Traceability_RS].[dbo].[orders] o ON l.OrderId = o.IDOrder
            LEFT JOIN [Traceability_RS].[dbo].[Products] p ON o.IDProduct = p.IDProduct
            INNER JOIN traceability_rs.fai.FaiStepDetails D on l.FaiStepDetailId=d.FaiStepDetailId
            INNER JOIN traceability_rs.fai.FaiSteps S on s.FatStepId=D.FatStepId
            INNER JOIN traceability_rs.fai.FaiTemplates t on t.FaiTemplateId=s.FaiTemplateId 
            WHERE l.DateIn >= ? AND l.DateIn <= ?
            """
            
            params = [
                self.date_from_picker.get_date().strftime('%Y-%m-%d') + ' 00:00:00',
                self.date_to_picker.get_date().strftime('%Y-%m-%d') + ' 23:59:59'
            ]
            
            # Filtro prodotto (ricerca parziale)
            if self.product_var.get().strip():
                query += " AND p.productcode LIKE ?"
                params.append(f"%{self.product_var.get().strip()}%")
            
            # Filtro operatore (ricerca parziale)
            if self.operator_var.get().strip():
                query += " AND l.Operator LIKE ?"
                params.append(f"%{self.operator_var.get().strip()}%")
            
            query += " ORDER BY l.DateIn DESC"
            
            rows = self.db.fetch_all(query, tuple(params))
            
            # Popola treeview
            for row in rows:
                fai_document = row.FAI_Document or 'N/A'
                fai_log_id = row.FaiLogId
                date_in = row.DateIn.strftime('%d/%m/%Y %H:%M') if row.DateIn else ''
                order_id = row.OrderId or ''
                product_code = row.productcode or 'N/A'
                operator = row.Operator or 'N/A'
                result = '✓ OK' if row.IsOk else '✗ NOT OK'
                has_pdf = '✓' if row.HasPDF else '✗'
                
                # Colore in base al risultato
                tag = 'ok' if row.IsOk else 'not_ok'
                
                self.tree.insert('', 'end', values=(
                    fai_log_id, fai_document, date_in, order_id, product_code, 
                    operator, result, has_pdf
                ), tags=(tag,))
            
            # Tags colori
            self.tree.tag_configure('ok', foreground='green')
            self.tree.tag_configure('not_ok', foreground='red')
            
            self.status_label.config(text=f"Trovate {len(rows)} validazioni")
            
        except Exception as e:
            logger.error(f"Errore caricamento validazioni: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile caricare i dati:\n{e}")
    
    def _open_pdf(self):
        """Apre il PDF della validazione selezionata"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Attenzione", "Seleziona una validazione")
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        fai_log_id = values[0]  # FaiLogId is first column
        has_pdf = values[7]  # HasPDF is now at index 7
        
        if has_pdf != '✓':
            messagebox.showwarning("Attenzione", "Nessun PDF disponibile per questa validazione")
            return
        
        try:
            # Recupera PDF dal database
            query = """
            SELECT DocVerification 
            FROM [Traceability_RS].[fai].[FaiLogs]
            WHERE FaiLogId = ?
            """
            
            row = self.db.fetch_one(query, (fai_log_id,))
            
            if not row or not row.DocVerification:
                messagebox.showerror("Errore", "PDF non trovato nel database")
                return
            
            # Salva PDF in temp
            pdf_path = os.path.join(tempfile.gettempdir(), f"FAI_Report_{fai_log_id}.pdf")
            with open(pdf_path, 'wb') as f:
                f.write(row.DocVerification)
            
            # Apri PDF
            os.startfile(pdf_path)
            logger.info(f"PDF aperto: {pdf_path}")
            self.status_label.config(text=f"PDF aperto: FAI Report {fai_log_id}")
            
        except Exception as e:
            logger.error(f"Errore apertura PDF: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile aprire il PDF:\n{e}")
    
    def _delete_validation(self):
        """Elimina validazione selezionata"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Attenzione", "Seleziona una validazione")
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        fai_log_id = values[0]
        
        response = messagebox.askyesno(
            "Conferma Eliminazione",
            f"Sei sicuro di voler eliminare la validazione FAI ID {fai_log_id}?\n\n"
            "Questa azione è irreversibile!",
            icon='warning'
        )
        
        if not response:
            return
        
        try:
            # Elimina header se esiste
            delete_header = """
            DELETE FROM [Traceability_RS].[fai].[FaiLogHeathers]
            WHERE FaiLogId = ?
            """
            self.db.execute_query(delete_header, (fai_log_id,))
            
            # Elimina tutti i log con questo FaiLogId
            delete_logs = """
            DELETE FROM [Traceability_RS].[fai].[FaiLogs]
            WHERE FaiLogId = ?
            """
            self.db.execute_query(delete_logs, (fai_log_id,))
            
            messagebox.showinfo("Successo", "Validazione eliminata con successo")
            self._load_data()
            
        except Exception as e:
            logger.error(f"Errore eliminazione validazione: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile eliminare la validazione:\n{e}")
    
    def _export_to_excel(self):
        """Esporta i risultati della ricerca in Excel"""
        try:
            # Verifica che ci siano dati
            if not self.tree.get_children():
                messagebox.showwarning("Attenzione", "Nessun dato da esportare")
                return
            
            # Importa openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                messagebox.showerror("Errore", "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl")
                return
            
            # Crea workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "FAI Reports"
            
            # Headers
            headers = ['FAI Document', 'Data/Ora', 'N. Ordine', 'Codice Prodotto', 'Operatore', 'Risultato', 'PDF']
            ws.append(headers)
            
            # Formatta header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aggiungi dati (skip FaiLogId che è nascosto)
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                # values[0] è FaiLogId (nascosto), quindi skippiamo
                row_data = [
                    values[1],  # FAI_Document
                    values[2],  # Data
                    values[3],  # Ordine
                    values[4],  # Prodotto
                    values[5],  # Operatore
                    values[6],  # Risultato
                    values[7]   # HasPDF
                ]
                ws.append(row_data)
                
                # Colora riga in base al risultato
                last_row = ws.max_row
                if '✓ OK' in str(values[6]):
                    fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                for cell in ws[last_row]:
                    cell.fill = fill_color
            
            # Adatta larghezza colonne
            ws.column_dimensions['A'].width = 40  # FAI Document
            ws.column_dimensions['B'].width = 18  # Data
            ws.column_dimensions['C'].width = 15  # Ordine
            ws.column_dimensions['D'].width = 25  # Prodotto
            ws.column_dimensions['E'].width = 20  # Operatore
            ws.column_dimensions['F'].width = 12  # Risultato
            ws.column_dimensions['G'].width = 8   # PDF
            
            # Salva file
            import os
            from tkinter import filedialog
            
            default_filename = f"FAI_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Successo", f"File Excel salvato:\n{filepath}")
                logger.info(f"Excel esportato: {filepath}")
                
                # Chiedi se aprire il file
                if messagebox.askyesno("Apri File", "Vuoi aprire il file Excel?"):
                    os.startfile(filepath)
        
        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile esportare in Excel:\n{e}")
