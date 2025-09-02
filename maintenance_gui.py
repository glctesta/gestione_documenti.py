# maintenance_gui.py
from datetime import datetime, date
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import filedialog
import os
import reportlab
import io
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkcalendar import DateEntry
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill

import richieste_intervento
import openpyxl
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.platypus import Image as ReportLabImage, Table, TableStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from tkinter import messagebox
import richieste_intervento
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Image as ReportLabImage

# Import per ReportLab (necessari per MachineDetailsWindow)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as ReportLabImage
except ImportError:
    print("Warning: reportlab not installed. PDF generation might be affected.")

try:
    from PIL import Image, ImageTk
    # Necessario per MachineDetailsWindow se reportlab è installato
    from reportlab.platypus import Image as ReportLabImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False



class AddMachineWindow(tk.Toplevel):
    """Finestra per aggiungere una nuova macchina."""

    def __init__(self, parent, db, lang):
        # CORREZIONE: Passa solo 'parent' alla classe superiore
        super().__init__(parent)
        self.db = db
        self.lang = lang

        self.title(self.lang.get('submenu_add_machine'))
        self.geometry("550x450")
        self.transient(parent)
        self.grab_set()

        self.brands_data, self.types_data, self.phases_data = {}, {}, {}
        self.brand_var, self.type_var, self.phase_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
        self.serial_var, self.internal_name_var = tk.StringVar(), tk.StringVar()
        self.year_var, self.inventory_var = tk.StringVar(), tk.StringVar()

        self._create_widgets()
        self._load_combobox_data()

    def _create_widgets(self):
        """Crea e posiziona i widget nella finestra."""
        frame = ttk.Frame(self, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        frame.columnconfigure(1, weight=1)

        # --- Campi del form ---
        ttk.Label(frame, text=self.lang.get('brand_label')).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.brand_combo = ttk.Combobox(frame, textvariable=self.brand_var, state='readonly')
        self.brand_combo.grid(row=0, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('type_label')).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.type_combo = ttk.Combobox(frame, textvariable=self.type_var, state='readonly')
        self.type_combo.grid(row=1, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('phase_label')).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.phase_combo = ttk.Combobox(frame, textvariable=self.phase_var, state='readonly')
        self.phase_combo.grid(row=2, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('serial_number_label')).grid(row=3, column=0, sticky=tk.W, pady=5)
        self.serial_entry = ttk.Entry(frame, textvariable=self.serial_var)
        self.serial_entry.grid(row=3, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('internal_name_label')).grid(row=4, column=0, sticky=tk.W, pady=5)
        self.internal_name_entry = ttk.Entry(frame, textvariable=self.internal_name_var)
        self.internal_name_entry.grid(row=4, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('production_year_label')).grid(row=5, column=0, sticky=tk.W, pady=5)
        self.year_entry = ttk.Entry(frame, textvariable=self.year_var)
        self.year_entry.grid(row=5, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('inventory_number_label')).grid(row=6, column=0, sticky=tk.W, pady=5)
        self.inventory_entry = ttk.Entry(frame, textvariable=self.inventory_var)
        self.inventory_entry.grid(row=6, column=1, sticky=tk.EW, pady=5)

        # --- Pulsanti ---
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=7, column=1, sticky=tk.E, pady=(20, 0))

        self.save_button = ttk.Button(button_frame, text=self.lang.get('save_button'), command=self._save_machine)
        self.save_button.pack(side=tk.LEFT, padx=5)

        self.cancel_button = ttk.Button(button_frame, text=self.lang.get('cancel_button'), command=self.destroy)
        self.cancel_button.pack(side=tk.LEFT)

        # --- NUOVO: Aggiunta logo in basso a destra ---
        if PIL_AVAILABLE:
            try:
                # Carica l'immagine
                image = Image.open("logo.png")
                image.thumbnail((100, 100))  # Rimpicciolisce l'immagine se necessario
                # IMPORTANTE: conservare un riferimento all'immagine per evitare che venga eliminata
                self.logo_image = ImageTk.PhotoImage(image)

                # Crea e posiziona la label con il logo
                logo_label = ttk.Label(frame, image=self.logo_image)
                logo_label.grid(row=8, column=1, sticky=tk.SE, pady=(10, 0), padx=5)
            except FileNotFoundError:
                print("logo.png non trovato per la finestra di inserimento.")
            except Exception as e:
                print(f"Errore caricamento logo: {e}")

        # Rende l'ultima riga (quella del logo) espandibile per spingerlo in basso
        frame.rowconfigure(8, weight=1)

    def _load_combobox_data(self):
        # ... (questo metodo rimane invariato) ...
        """Carica i dati per i menu a tendina dal database."""
        # Carica Brand
        brands = self.db.fetch_brands()
        if brands:
            self.brands_data = {row.Brand: row.EquipmentBrandId for row in brands}
            self.brand_combo['values'] = list(self.brands_data.keys())

        # Carica Tipi Macchina
        types = self.db.fetch_equipment_types()
        if types:
            self.types_data = {row.EquipmentType: row.EquipmentTypeId for row in types}
            self.type_combo['values'] = list(self.types_data.keys())

        # Carica Fasi
        phases = self.db.fetch_parent_phases_for_maintenance()
        if phases:
            self.phases_data = {row.ParentPhaseName: row.IDParentPhase for row in phases}
            self.phase_combo['values'] = list(self.phases_data.keys())

    def _save_machine(self):
        # ... (questo metodo rimane invariato) ...
        """Valida i dati e li salva nel database."""
        # Validazione input
        if not all([self.brand_var.get(), self.type_var.get(), self.phase_var.get(), self.serial_var.get()]):
            messagebox.showerror(self.lang.get('error_title'), self.lang.get('error_required_fields'), parent=self)
            return

        try:
            # Converte l'anno in numero, se inserito
            prod_year = int(self.year_var.get()) if self.year_var.get() else None
        except ValueError:
            messagebox.showerror(self.lang.get('error_title'), self.lang.get('error_invalid_year'), parent=self)
            return

        # Recupera gli ID dai valori selezionati nei combobox
        brand_id = self.brands_data.get(self.brand_var.get())
        type_id = self.types_data.get(self.type_var.get())
        phase_id = self.phases_data.get(self.phase_var.get())

        success = self.db.add_new_equipment(
            brand_id=brand_id,
            type_id=type_id,
            phase_id=phase_id,
            serial_number=self.serial_var.get(),
            internal_name=self.internal_name_var.get(),
            prod_year=prod_year,
            inv_number=self.inventory_var.get()
        )

        if success:
            messagebox.showinfo(self.lang.get('success_title'), self.lang.get('info_machine_saved'), parent=self)
            self.destroy()
        else:
            error_msg = self.lang.get('error_saving_machine') + f"\n\n{self.db.last_error_details}"
            messagebox.showerror(self.lang.get('error_title'), error_msg, parent=self)

class SelectMachineToEditWindow(tk.Toplevel):
    """Finestra per selezionare quale macchina modificare."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.parent_app = parent
        self.user_name = user_name

        self.title(self.lang.get('select_machine_to_edit_title'))
        self.geometry("500x150")
        self.transient(parent)
        self.grab_set()

        self.equipments_data = {}
        self.selected_machine_var = tk.StringVar()

        frame = ttk.Frame(self, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=1)

        ttk.Label(frame, text=self.lang.get('select_machine_label')).pack(fill=tk.X)
        self.equipment_combo = ttk.Combobox(frame, textvariable=self.selected_machine_var, state='readonly', height=10)
        self.equipment_combo.pack(fill=tk.X, pady=5)

        ttk.Button(frame, text=self.lang.get('edit_button'), command=self._open_edit_window).pack(pady=10)

        self._load_equipments()

    def _load_equipments(self):
        """Carica la lista di tutte le macchine."""
        equipments = self.db.fetch_all_equipments()
        if equipments:
            # Crea un dizionario che mappa il testo visualizzato all'ID
            self.equipments_data = {f"{row.InternalName or 'N/D'} [{row.SerialNumber}]": row.EquipmentId for row in
                                    equipments}
            self.equipment_combo['values'] = list(self.equipments_data.keys())

    def _open_edit_window(self):
        """Apre la finestra di modifica per la macchina selezionata."""
        selection = self.selected_machine_var.get()
        if not selection:
            messagebox.showwarning(self.lang.get('warning_title'), self.lang.get('warning_no_machine_selected'),
                                   parent=self)
            return

        equipment_id = self.equipments_data.get(selection)

        # Apre la finestra di modifica passando tutti i parametri necessari
        EditMachineWindow(self.parent_app, self.db, self.lang, equipment_id, self.user_name)
        self.destroy()

class EditMachineWindow(tk.Toplevel):
    """Finestra per modificare i dati di una macchina specifica."""

    def __init__(self, parent, db, lang, equipment_id, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.equipment_id = equipment_id
        self.user_name = user_name

        # Per confrontare i valori prima/dopo
        self.original_data = {}
        self.phases_data = {}

        # Variabili di controllo
        self.phase_var = tk.StringVar()
        self.serial_var = tk.StringVar()
        self.internal_name_var = tk.StringVar()

        self.title(self.lang.get('submenu_edit_machine'))
        self.geometry("550x300")
        self.transient(parent)
        self.grab_set()

        self._create_widgets()
        self._load_data()

    def _create_widgets(self):
        frame = ttk.Frame(self, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text=self.lang.get('phase_label')).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.phase_combo = ttk.Combobox(frame, textvariable=self.phase_var, state='readonly')
        self.phase_combo.grid(row=0, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('internal_name_label')).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.internal_name_entry = ttk.Entry(frame, textvariable=self.internal_name_var)
        self.internal_name_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)

        ttk.Label(frame, text=self.lang.get('serial_number_label')).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.serial_entry = ttk.Entry(frame, textvariable=self.serial_var)
        self.serial_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=3, column=1, sticky=tk.E, pady=(20, 0))
        ttk.Button(button_frame, text=self.lang.get('save_button'), command=self._save_changes).pack(side=tk.LEFT,
                                                                                                     padx=5)
        ttk.Button(button_frame, text=self.lang.get('cancel_button'), command=self.destroy).pack(side=tk.LEFT)

    def _load_data(self):
        """Carica i dati della macchina selezionata e li inserisce nei campi."""
        # Carica le fasi di produzione per il combobox
        phases = self.db.fetch_parent_phases_for_maintenance()
        if phases:
            self.phases_data = {row.IDParentPhase: row.ParentPhaseName for row in phases}
            self.phase_combo['values'] = list(self.phases_data.values())

        # Carica i dettagli della macchina specifica
        details = self.db.fetch_equipment_details(self.equipment_id)
        if details:
            self.original_data = details

            # Imposta i valori nei widget
            self.phase_var.set(self.phases_data.get(details.ParentPhaseId, ""))
            self.internal_name_var.set(details.InternalName or "")
            self.serial_var.set(details.SerialNumber or "")
        else:
            messagebox.showerror(self.lang.get('error_title'), self.lang.get('error_loading_machine_details'),
                                 parent=self)
            self.destroy()

    def _save_changes(self):
        """Costruisce il log delle modifiche e salva nel database."""
        # Recupera i nuovi valori
        new_phase_name = self.phase_var.get()
        new_phase_id = [k for k, v in self.phases_data.items() if v == new_phase_name][0]
        new_name = self.internal_name_var.get()
        new_serial = self.serial_var.get()

        # Costruisce la stringa di log confrontando i valori vecchi e nuovi
        change_log = []
        if new_phase_id != self.original_data.ParentPhaseId:
            original_phase_name = self.phases_data.get(self.original_data.ParentPhaseId, 'N/D')
            change_log.append(f"Fase cambiata da '{original_phase_name}' a '{new_phase_name}'.")

        if new_name != self.original_data.InternalName:
            change_log.append(f"Nome Interno cambiato da '{self.original_data.InternalName}' a '{new_name}'.")

        if new_serial != self.original_data.SerialNumber:
            change_log.append(f"Numero di Serie cambiato da '{self.original_data.SerialNumber}' a '{new_serial}'.")

        if not change_log:
            messagebox.showinfo(self.lang.get('info_title'), self.lang.get('info_no_changes'), parent=self)
            return

        # Unisce le modifiche in un'unica stringa
        change_log_string = " ".join(change_log)

        success = self.db.update_and_log_equipment_changes(
            self.equipment_id,
            new_phase_id,
            new_name,
            new_serial,
            change_log_string,
            self.user_name
        )

        if success:
            messagebox.showinfo(self.lang.get('success_title'), self.lang.get('info_machine_updated'), parent=self)
            self.destroy()
        else:
            error_msg = self.lang.get('error_updating_machine') + f"\n\n{self.db.last_error_details}"
            messagebox.showerror(self.lang.get('error_title'), error_msg, parent=self)

class ViewMachineWindow(tk.Toplevel):
    """Finestra di ricerca macchine con filtri."""

    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.parent_app = parent

        self.title(self.lang.get('submenu_view_machines'))
        self.geometry("800x600")
        self.transient(parent)
        self.grab_set()

        self._create_filter_widgets()
        self._create_results_view()
        self._load_filter_data()

    def _create_filter_widgets(self):
        """Crea la sezione superiore con i filtri di ricerca."""
        filter_frame = ttk.LabelFrame(self, text=self.lang.get('search_filters_label'), padding="10")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # Inizializzazione dati
        self.brands_data, self.types_data, self.phases_data = {}, {}, {}
        self.brand_var, self.type_var, self.phase_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
        self.search_text_var = tk.StringVar()

        # Creazione widget
        ttk.Label(filter_frame, text=self.lang.get('brand_label')).grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.brand_combo = ttk.Combobox(filter_frame, textvariable=self.brand_var, state='readonly', width=20)
        self.brand_combo.grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(filter_frame, text=self.lang.get('type_label')).grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)
        self.type_combo = ttk.Combobox(filter_frame, textvariable=self.type_var, state='readonly', width=20)
        self.type_combo.grid(row=0, column=3, padx=5, pady=2)

        ttk.Label(filter_frame, text=self.lang.get('phase_label')).grid(row=0, column=4, padx=5, pady=2, sticky=tk.W)
        self.phase_combo = ttk.Combobox(filter_frame, textvariable=self.phase_var, state='readonly', width=20)
        self.phase_combo.grid(row=0, column=5, padx=5, pady=2)

        ttk.Label(filter_frame, text=self.lang.get('search_text_label')).grid(row=1, column=0, padx=5, pady=2,
                                                                              sticky=tk.W)
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_text_var, width=50)
        self.search_entry.grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky=tk.W)

        search_button = ttk.Button(filter_frame, text=self.lang.get('search_button'), command=self._perform_search)
        search_button.grid(row=1, column=4, padx=5, pady=5)

        clear_button = ttk.Button(filter_frame, text=self.lang.get('clear_filters_button'), command=self._clear_filters)
        clear_button.grid(row=1, column=5, padx=5, pady=5)

    def _create_results_view(self):
        """Crea la Treeview per mostrare i risultati della ricerca."""
        results_frame = ttk.Frame(self, padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ('name', 'serial', 'brand', 'type', 'phase')
        self.tree = ttk.Treeview(results_frame, columns=cols, show='headings', selectmode='browse')

        # Definisci le intestazioni
        self.tree.heading('name', text=self.lang.get('header_internal_name'))
        self.tree.heading('serial', text=self.lang.get('header_serial_number'))
        self.tree.heading('brand', text=self.lang.get('header_brand'))
        self.tree.heading('type', text=self.lang.get('header_type'))
        self.tree.heading('phase', text=self.lang.get('header_phase'))

        self.tree.column('name', width=150)
        self.tree.column('serial', width=150)
        self.tree.column('brand', width=100)
        self.tree.column('type', width=120)
        self.tree.column('phase', width=120)

        # Aggiungi scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)

        # Evento doppio click per aprire i dettagli
        self.tree.bind('<Double-1>', self._open_details_window)

    def _load_filter_data(self):
        """Carica i dati per i combobox dei filtri."""
        brands = self.db.fetch_brands()
        self.brands_data = {row.Brand: row.EquipmentBrandId for row in brands}
        self.brand_combo['values'] = [''] + list(self.brands_data.keys())

        types = self.db.fetch_equipment_types()
        self.types_data = {row.EquipmentType: row.EquipmentTypeId for row in types}
        self.type_combo['values'] = [''] + list(self.types_data.keys())

        phases = self.db.fetch_parent_phases_for_maintenance()
        self.phases_data = {row.ParentPhaseName: row.IDParentPhase for row in phases}
        self.phase_combo['values'] = [''] + list(self.phases_data.keys())

    def _clear_filters(self):
        """Pulisce tutti i filtri di ricerca."""
        self.brand_var.set('')
        self.type_var.set('')
        self.phase_var.set('')
        self.search_text_var.set('')
        self._perform_search()

    def _perform_search(self):
        """Raccoglie i filtri, esegue la ricerca e popola la Treeview."""
        # Pulisci risultati precedenti
        for i in self.tree.get_children():
            self.tree.delete(i)

        filters = {
            'brand_id': self.brands_data.get(self.brand_var.get()),
            'type_id': self.types_data.get(self.type_var.get()),
            'phase_id': self.phases_data.get(self.phase_var.get()),
            'search_text': self.search_text_var.get()
        }

        results = self.db.search_equipments(filters)

        for row in results:
            # L'ID viene salvato come primo elemento della riga nella treeview
            self.tree.insert('', tk.END, iid=row.EquipmentId,
                             values=(row.InternalName, row.SerialNumber, row.Brand, row.EquipmentType,
                                     row.ParentPhaseName))

    def _open_details_window(self, event=None):
        """Apre la finestra di dettaglio per la macchina selezionata."""
        selected_item = self.tree.focus()
        if not selected_item:
            return

        equipment_id = int(selected_item)
        MachineDetailsWindow(self.parent_app, self.db, self.lang, equipment_id)

class MachineDetailsWindow(tk.Toplevel):
    """Finestra che mostra tutte le informazioni di una singola macchina."""

    def __init__(self, parent, db, lang, equipment_id):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.equipment_id = equipment_id
        self.details = None

        self.title(self.lang.get('machine_details_title'))
        self.geometry("900x700")
        self.transient(parent)
        self.grab_set()

        self.main_frame = ttk.Frame(self, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self._create_widgets()
        self._load_and_display_data()

    def _create_widgets(self):
        """Crea i widget per visualizzare i dettagli."""
        # Frame Anagrafica
        self.master_frame = ttk.LabelFrame(self.main_frame, text=self.lang.get('master_data_label'), padding="10")
        self.master_frame.pack(fill=tk.X, pady=5)

        # Frame Modifiche
        changes_frame = ttk.LabelFrame(self.main_frame, text=self.lang.get('changes_log_label'), padding="10")
        changes_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.changes_tree = ttk.Treeview(changes_frame, columns=('date', 'user', 'change'), show='headings')
        self.changes_tree.heading('date', text=self.lang.get('header_date'))
        self.changes_tree.heading('user', text=self.lang.get('header_user'))
        self.changes_tree.heading('change', text=self.lang.get('header_change'))
        self.changes_tree.pack(fill=tk.BOTH, expand=True)

        # Frame Documenti e Log (usiamo un Notebook)
        notebook = ttk.Notebook(self.main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)

        docs_frame = ttk.Frame(notebook, padding="10")
        logs_frame = ttk.Frame(notebook, padding="10")
        notebook.add(docs_frame, text=self.lang.get('maintenance_docs_label'))
        notebook.add(logs_frame, text=self.lang.get('maintenance_logs_label'))

        self.docs_tree = ttk.Treeview(docs_frame, columns=('doc', 'user', 'date'), show='headings')
        self.docs_tree.heading('doc', text=self.lang.get('header_document'))
        self.docs_tree.heading('user', text=self.lang.get('header_user'))
        self.docs_tree.heading('date', text=self.lang.get('header_date'))
        self.docs_tree.pack(fill=tk.BOTH, expand=True)

        self.logs_tree = ttk.Treeview(logs_frame, columns=('date', 'user', 'notes'), show='headings')
        self.logs_tree.heading('date', text=self.lang.get('header_date'))
        self.logs_tree.heading('user', text=self.lang.get('header_user'))
        self.logs_tree.heading('notes', text=self.lang.get('header_notes'))
        self.logs_tree.pack(fill=tk.BOTH, expand=True)

        # Pulsante PDF
        pdf_button = ttk.Button(self.main_frame, text=self.lang.get('generate_pdf_button'), command=self._generate_pdf)
        pdf_button.pack(pady=10)

    def _load_and_display_data(self):
        """Carica tutti i dati e li visualizza nei widget appropriati."""
        self.details = self.db.fetch_full_equipment_details(self.equipment_id)
        if not self.details or not self.details.get('master'):
            messagebox.showerror(self.lang.get('error_title'), self.lang.get('error_loading_machine_details'),
                                 parent=self)
            self.destroy()
            return

        # Popola anagrafica
        master = self.details['master']
        for i, (key, value) in enumerate(master.items()):
            ttk.Label(self.master_frame, text=f"{key}:", font="Helvetica 9 bold").grid(row=i, column=0, sticky=tk.W,
                                                                                       padx=5)
            ttk.Label(self.master_frame, text=value or "N/D").grid(row=i, column=1, sticky=tk.W, padx=5)

        # Popola log modifiche
        for item in self.details.get('changes', []):
            self.changes_tree.insert('', tk.END, values=(item.DateChange, item.WhoChange, item.Changed))

        # Popola documenti
        for item in self.details.get('docs', []):
            self.docs_tree.insert('', tk.END, values=(item.DocumentSource, item.UploadedBy, item.DateSys))

        # Popola log manutenzioni
        for item in self.details.get('logs', []):
            self.logs_tree.insert('', tk.END, values=(item.DataEsecuzione, item.IdManutentore, item.NoteGenerali))

    def _generate_pdf(self):
        """Genera un report PDF con i dettagli della macchina."""
        if not self.details: return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Documents", "*.pdf")],
            title=self.lang.get('save_pdf_title')
        )
        if not file_path: return

        master = self.details['master']
        c = canvas.Canvas(file_path, pagesize=A4)
        width, height = A4

        # Funzione ausiliaria per scrivere testo
        def draw_text(y, text, size=10, bold=False):
            font = "Helvetica-Bold" if bold else "Helvetica"
            c.setFont(font, size)
            c.drawString(2 * cm, y, text)

        try:
            # Logo e Titolo
            if os.path.exists("logo.png"):
                logo = ReportLabImage("logo.png", width=3 * cm, height=3 * cm)
                logo.drawOn(c, width - 4.5 * cm, height - 3.5 * cm)

            draw_text(height - 2 * cm, self.lang.get('pdf_report_title'), 18, True)
            draw_text(height - 2.5 * cm,
                      f"{self.lang.get('pdf_print_date')}: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 8)

            y_pos = height - 4 * cm
            # Dati Anagrafici
            draw_text(y_pos, self.lang.get('master_data_label'), 14, True)
            y_pos -= cm
            for key, value in master.items():
                draw_text(y_pos, f"{key}: {value or 'N/D'}")
                y_pos -= 0.5 * cm
                if y_pos < 3 * cm:  # Vai a pagina nuova
                    c.showPage()
                    y_pos = height - 3 * cm

            # Qui si potrebbero aggiungere le altre sezioni (modifiche, log, etc.)
            # con logica simile per gestire il cambio pagina.

            c.save()
            messagebox.showinfo(self.lang.get('success_title'), self.lang.get('pdf_generated_success'), parent=self)
        except Exception as e:
            messagebox.showerror(self.lang.get('error_title'), f"{self.lang.get('pdf_generated_error')}\n\n{e}",
                                 parent=self)

class MaintenanceDocsWindow(tk.Toplevel):
    """Finestra per la gestione dei documenti di manutenzione."""

    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.title(lang.get('submenu_maintenance_docs'))
        self.geometry("600x400")
        ttk.Label(self, text="Finestra Documenti Manutenzione - In Sviluppo", font=("Helvetica", 16)).pack(pady=50,
                                                                                                           padx=20)
        self.transient(parent)
        self.grab_set()


    # In maintenance_gui.py

class FillTemplateWindow(tk.Toplevel):
    """Finestra per la compilazione delle schede di manutenzione."""

    def __init__(self, parent, db, lang, user_name):
        # Passa solo 'parent' alla classe superiore
        super().__init__(parent)

        # Le altre variabili vengono assegnate a questa classe
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.start_time = None

        self.title(lang.get('submenu_fill_templates'))
        self.geometry("950x600")
        self.transient(parent)
        self.grab_set()

        self.equipments_data = {}
        self.plans_data = {}
        self.completed_tasks = set()
        self.all_equipment_names = []

        self.equipment_var = tk.StringVar()
        self.plan_var = tk.StringVar()

        self._create_widgets()
        self._load_equipments()

    def _create_widgets(self):
        frame = ttk.Frame(self, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        # --- Sezione Selezione (Top) ---
        selection_frame = ttk.Frame(frame)
        selection_frame.pack(fill=tk.X, pady=10)

        ttk.Label(selection_frame, text=self.lang.get('select_machine_label')).pack(side=tk.LEFT, padx=5)
        self.equipment_combo = ttk.Combobox(selection_frame, textvariable=self.equipment_var, state='normal', width=40)
        self.equipment_combo.pack(side=tk.LEFT, padx=5)
        self.equipment_combo.bind("<<ComboboxSelected>>", self._on_equipment_select)
        self.equipment_combo.bind('<KeyRelease>', self._filter_equipment_combo)

        ttk.Label(selection_frame, text=self.lang.get('select_maintenance_plan', "Seleziona Piano:")).pack(side=tk.LEFT, padx=5)
        self.plan_combo = ttk.Combobox(selection_frame, textvariable=self.plan_var, state='disabled', width=40)
        self.plan_combo.pack(side=tk.LEFT, padx=5)
        self.plan_combo.bind("<<ComboboxSelected>>", self._on_plan_select)

        # --- Sezione Compiti (Center) ---
        tasks_frame = ttk.LabelFrame(frame, text=self.lang.get('maintenance_tasks_label', "Compiti da Eseguire"), padding="10")
        tasks_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        cols = ('check', 'name', 'category', 'description')
        self.tasks_tree = ttk.Treeview(tasks_frame, columns=cols, show='headings')
        self.tasks_tree.heading('check', text='☑')
        self.tasks_tree.heading('name', text=self.lang.get('header_task_name', 'Compito'))
        self.tasks_tree.heading('category', text=self.lang.get('header_category', 'Categoria'))
        self.tasks_tree.heading('description', text=self.lang.get('header_description', 'Descrizione'))
        self.tasks_tree.column('check', width=30, anchor='center', stretch=tk.NO)
        self.tasks_tree.column('name', width=250)
        self.tasks_tree.column('category', width=120)
        self.tasks_tree.column('description', width=450)

        scrollbar = ttk.Scrollbar(tasks_frame, orient=tk.VERTICAL, command=self.tasks_tree.yview)
        self.tasks_tree.configure(yscroll=scrollbar.set)
        self.tasks_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.tasks_tree.bind('<Button-1>', self._on_tree_click)
        self.tasks_tree.bind('<Double-1>', self._on_tree_double_click)

        # --- Sezione Note (Bottom-Center) ---
        notes_frame = ttk.LabelFrame(frame, text=self.lang.get('maintenance_notes_label', "Note Generali / Osservazioni"), padding="10")
        notes_frame.pack(fill=tk.X, pady=5)
        self.notes_text = tk.Text(notes_frame, height=4, wrap=tk.WORD)
        self.notes_text.pack(fill=tk.X, expand=True, padx=5)
        self.notes_text.config(state='disabled')

        # --- Sezione Azioni (Bottom) ---
        action_frame = ttk.Frame(frame)
        action_frame.pack(fill=tk.X, pady=10)

        request_button_text = self.lang.get('request_button', "Crea Richiesta (Parti/Intervento)")
        self.request_button = ttk.Button(action_frame, text=request_button_text, command=self._open_request_window, state='disabled')
        self.request_button.pack(side=tk.LEFT, padx=5)

        self.save_button = ttk.Button(action_frame, text=self.lang.get('save_completed_tasks', "Salva Compiti Eseguiti"), command=self._save_logs, state='disabled')
        self.save_button.pack(side=tk.RIGHT, padx=5)

    def _load_equipments(self):
        equipments = self.db.fetch_all_equipments()
        if equipments:
            self.equipments_data = {f"{row.InternalName or 'N/D'} [{row.SerialNumber}]": row.EquipmentId for row in equipments}
            self.all_equipment_names = sorted(list(self.equipments_data.keys()))
            self.equipment_combo['values'] = self.all_equipment_names

    def _filter_equipment_combo(self, event):
        typed_text = self.equipment_var.get().lower()
        if not typed_text:
            self.equipment_combo['values'] = self.all_equipment_names
            return
        filtered_list = [name for name in self.all_equipment_names if typed_text in name.lower()]
        self.equipment_combo['values'] = filtered_list

    def _reset_plan_and_tasks(self):
        self.plan_var.set("")
        if self.plan_combo:
            self.plan_combo.config(state='disabled', values=[])
        self.plans_data = {}
        self._reset_tasks()

    def _reset_tasks(self):
        if self.tasks_tree:
            for i in self.tasks_tree.get_children():
                self.tasks_tree.delete(i)
        self.completed_tasks = set()
        self.start_time = None
        if self.save_button:
            self.save_button.config(state='disabled')
        if self.notes_text:
            self.notes_text.config(state='normal')
            self.notes_text.delete('1.0', tk.END)
            self.notes_text.config(state='disabled')
        if self.request_button:
            self.request_button.config(state='disabled')

    def _on_equipment_select(self, event=None):
        self._reset_plan_and_tasks()
        equipment_selection = self.equipment_var.get()
        equipment_id = self.equipments_data.get(equipment_selection)

        if equipment_id:
            plans = self.db.fetch_available_maintenance_plans(equipment_id)
            if plans:
                self.plans_data = {row.TimingDescriprion: (getattr(row, 'PianoManutenzioneId', None), row.ProgrammedInterventionId) for row in plans}
                self.plan_combo['values'] = list(self.plans_data.keys())
                self.plan_combo.config(state='readonly')
            else:
                messagebox.showinfo(self.lang.get('info_title', "Info"), self.lang.get('info_no_plans_available', "Nessun piano di manutenzione disponibile..."), parent=self)

    def _on_plan_select(self, event=None):
        self._reset_tasks()
        plan_selection = self.plan_var.get()
        equipment_selection = self.equipment_var.get()
        equipment_id = self.equipments_data.get(equipment_selection)

        if plan_selection and equipment_id and plan_selection in self.plans_data:
            _, programmed_intervention_id = self.plans_data[plan_selection]
            self.start_time = datetime.now()
            tasks = self.db.fetch_maintenance_tasks(programmed_intervention_id, equipment_id)
            if tasks:
                for task in tasks:
                    self.tasks_tree.insert('', tk.END, iid=task.CompitoId, values=("", task.NomeCompito, task.Categoria, task.DescrizioneCompito))
                self.save_button.config(state='normal')
                self.notes_text.config(state='normal')
                self.request_button.config(state='normal')
            else:
                messagebox.showwarning(self.lang.get('warning_title', "Attenzione"), self.lang.get('warn_no_tasks_found', "Nessun compito trovato..."), parent=self)

    def _on_tree_click(self, event):
        region = self.tasks_tree.identify("region", event.x, event.y)
        column = self.tasks_tree.identify_column(event.x)
        if region == "cell" and column == "#1":
            item_iid = self.tasks_tree.identify_row(event.y)
            if item_iid:
                try:
                    task_id = int(item_iid)
                except ValueError: return
                if task_id in self.completed_tasks:
                    self.completed_tasks.remove(task_id)
                    self.tasks_tree.set(item_iid, 'check', "")
                else:
                    self.completed_tasks.add(task_id)
                    self.tasks_tree.set(item_iid, 'check', "✔")

    def _on_tree_double_click(self, event):
        item_iid = self.tasks_tree.focus()
        if not item_iid: return
        column = self.tasks_tree.identify_column(event.x)
        if column == "#1": return
        try:
            task_id = int(item_iid)
        except ValueError: return
        success = self.db.fetch_and_open_document_by_task_id(task_id)
        if not success:
            error_msg = self.lang.get('error_opening_task_document', "Impossibile aprire il documento...")
            if hasattr(self.db, 'last_error_details') and self.db.last_error_details:
                error_msg += f"\n\n{self.db.last_error_details}"
            messagebox.showwarning(self.lang.get('warning_title'), error_msg, parent=self)

    def _open_request_window(self):
        equipment_name = self.equipment_var.get()
        equipment_id = self.equipments_data.get(equipment_name)
        if not equipment_id:
            messagebox.showerror(self.lang.get('error_title'), "Errore interno: ID macchina non trovato.", parent=self)
            return
        richieste_intervento.open_request_window(parent=self, db=self.db, lang=self.lang, user_name=self.user_name, equipment_id=equipment_id, equipment_name=equipment_name)

    def _save_logs(self):
        if not self.completed_tasks:
            messagebox.showwarning(self.lang.get('warning_title', "Nessun compito selezionato..."), parent=self)
            return
        equipment_id = self.equipments_data.get(self.equipment_var.get())
        if not equipment_id or not self.start_time:
            messagebox.showerror(self.lang.get('error_title'), "Errore interno: Dati macchina o ora di inizio mancanti.", parent=self)
            return
        notes_content = self.notes_text.get("1.0", tk.END).strip()
        confirm_msg = self.lang.get('confirm_save_logs_message', "Salvare i log per {0} compiti?").format(len(self.completed_tasks))
        if messagebox.askyesno(self.lang.get('confirm_save_title', "Conferma Salvataggio"), confirm_msg, parent=self):
            success = self.db.log_completed_tasks(equipment_id=equipment_id, user_name=self.user_name, completed_task_ids=list(self.completed_tasks), start_time=self.start_time, notes=notes_content)
            if success:
                messagebox.showinfo(self.lang.get('success_title', "Successo"), self.lang.get('info_logs_saved_success', "Log manutenzione salvati..."), parent=self)
                self._on_equipment_select()
            else:
                messagebox.showerror(self.lang.get('error_title'), self.db.last_error_details, parent=self)

class MaintenanceReportWindow(tk.Toplevel):
    """Finestra avanzata per la generazione di report di manutenzione."""

    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.title(lang.get('submenu_reports', "Report Manutenzione"))
        self.geometry("1100x700")

        self.report_data = []
        self.equipments_data = {}
        self.cycles_data = {}

        self.from_date_var = tk.StringVar(value=date.today().strftime('%d/%m/%Y'))
        self.to_date_var = tk.StringVar(value=date.today().strftime('%d/%m/%Y'))
        self.equipment_var = tk.StringVar()
        self.cycle_var = tk.StringVar()

        self._create_widgets()
        self._load_filters()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)

        filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get('report_filters_label', "Filtri Report"),
                                      padding="10")
        filter_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        ttk.Label(filter_frame, text=self.lang.get('from_date_label', "Da:")).grid(row=0, column=0, padx=5)
        DateEntry(filter_frame, width=12, date_pattern='dd/MM/yyyy', textvariable=self.from_date_var).grid(row=0,
                                                                                                           column=1)

        ttk.Label(filter_frame, text=self.lang.get('to_date_label', "A:")).grid(row=0, column=2, padx=5)
        DateEntry(filter_frame, width=12, date_pattern='dd/MM/yyyy', textvariable=self.to_date_var).grid(row=0,
                                                                                                         column=3)

        ttk.Label(filter_frame, text=self.lang.get('machine_label', "Macchina:")).grid(row=0, column=4, padx=5)
        self.equipment_combo = ttk.Combobox(filter_frame, textvariable=self.equipment_var, state="readonly", width=30)
        self.equipment_combo.grid(row=0, column=5, padx=5)

        ttk.Label(filter_frame, text=self.lang.get('cycle_label', "Ciclo:")).grid(row=0, column=6, padx=5)
        self.cycle_combo = ttk.Combobox(filter_frame, textvariable=self.cycle_var, state="readonly", width=20)
        self.cycle_combo.grid(row=0, column=7, padx=5)

        ttk.Button(filter_frame, text=self.lang.get('generate_report_button', "Genera Report"),
                   command=self._generate_report).grid(row=0, column=8, padx=20)
        self.export_button = ttk.Button(filter_frame, text=self.lang.get('export_pdf_button', "Esporta PDF"),
                                        command=self._export_to_pdf, state="disabled")
        self.export_button.grid(row=0, column=9)

        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky="nsew")

        summary_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(summary_frame, text=self.lang.get('summary_chart_tab', "Riepilogo e Grafico"))
        summary_frame.columnconfigure(1, weight=3)
        summary_frame.columnconfigure(0, weight=1)
        summary_frame.rowconfigure(0, weight=1)

        kpi_frame = ttk.LabelFrame(summary_frame, text=self.lang.get('kpi_label', "Statistiche Chiave (KPI)"))
        kpi_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.kpi_total_tasks = ttk.Label(kpi_frame, text="", font=("Helvetica", 11))
        self.kpi_total_tasks.pack(pady=5, anchor="w", padx=10)
        self.kpi_total_time = ttk.Label(kpi_frame, text="", font=("Helvetica", 11))
        self.kpi_total_time.pack(pady=5, anchor="w", padx=10)
        self.kpi_top_machine = ttk.Label(kpi_frame, text="", font=("Helvetica", 11), wraplength=200)
        self.kpi_top_machine.pack(pady=5, anchor="w", padx=10)
        self._update_kpis()

        self.chart_frame = ttk.LabelFrame(summary_frame, text=self.lang.get('chart_label', "Interventi per Macchina"))
        self.chart_frame.grid(row=0, column=1, sticky="nsew")

        details_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(details_frame, text=self.lang.get('details_tab', "Dati di Dettaglio"))
        details_frame.rowconfigure(0, weight=1)
        details_frame.columnconfigure(0, weight=1)

        cols = ('machine', 'intervention', 'task', 'user', 'start', 'stop', 'duration')
        self.tree = ttk.Treeview(details_frame, columns=cols, show="headings")
        self.tree.heading('machine', text=self.lang.get('header_machine', "Macchina"))
        self.tree.heading('intervention', text=self.lang.get('header_intervention', "Tipo Intervento"))
        self.tree.heading('task', text=self.lang.get('header_task', "Compito"))
        self.tree.heading('user', text=self.lang.get('header_user', "Eseguito Da"))
        self.tree.heading('start', text=self.lang.get('header_start', "Inizio"))
        self.tree.heading('stop', text=self.lang.get('header_stop', "Fine"))
        self.tree.heading('duration', text=self.lang.get('header_duration', "Durata (min)"))
        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(details_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    def _load_filters(self):
        all_text = self.lang.get('all_filter_option', "TUTTI")

        equipments = self.db.fetch_all_equipments()
        self.equipments_data = {f"{r.InternalName or ''} [{r.SerialNumber}]": r.EquipmentId for r in equipments}
        self.equipment_combo['values'] = [all_text] + sorted(list(self.equipments_data.keys()))
        self.equipment_var.set(all_text)

        cycles = self.db.fetch_maintenance_cycles()
        self.cycles_data = {c.TimingDescriprion: c.ProgrammedInterventionId for c in cycles}
        self.cycle_combo['values'] = [all_text] + sorted(list(self.cycles_data.keys()))
        self.cycle_var.set(all_text)

    def _generate_report(self):
        from_date_str = self.from_date_var.get()
        to_date_str = self.to_date_var.get()

        if not from_date_str or not to_date_str:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('error_date_range', "Selezionare un intervallo di date valido."))
            return

        # --- CORREZIONE: Converti le stringhe in oggetti data ---
        try:
            # Il formato '%d/%m/%Y' deve corrispondere a quello del DateEntry
            from_date_obj = datetime.strptime(from_date_str, '%d/%m/%Y').date()
            to_date_obj = datetime.strptime(to_date_str, '%d/%m/%Y').date()
        except ValueError:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('error_invalid_date_format', "Il formato della data non è valido."))
            return
        # --- FINE CORREZIONE ---

        all_text = self.lang.get('all_filter_option', "TUTTI")
        equipment_id = self.equipments_data.get(
            self.equipment_var.get()) if self.equipment_var.get() != all_text else None
        cycle_id = self.cycles_data.get(self.cycle_var.get()) if self.cycle_var.get() != all_text else None

        # Passa gli oggetti data, non più le stringhe
        self.report_data = self.db.fetch_report_data(from_date_obj, to_date_obj, equipment_id, cycle_id)

        if not self.report_data:
            messagebox.showinfo(self.lang.get('info_title', "Info"),
                                self.lang.get('info_no_data_found', "Nessun dato trovato per i filtri selezionati."))
            self.export_button.config(state="disabled")
        else:
            self.export_button.config(state="normal")

        self._update_kpis()
        self._update_details_table()
        self._update_chart()

    def _update_kpis(self):
        kpi1_label = self.lang.get('kpi_total_interventions', "Interventi totali:")
        kpi2_label = self.lang.get('kpi_total_time', "Tempo totale (ore):")
        kpi3_label = self.lang.get('kpi_top_machine', "Macchina con più interventi:")

        if not self.report_data:
            self.kpi_total_tasks.config(text=f"{kpi1_label} 0")
            self.kpi_total_time.config(text=f"{kpi2_label} 0.0")
            self.kpi_top_machine.config(text=f"{kpi3_label} N/D")
            return

        total_tasks = len(self.report_data)
        total_minutes = sum(row.DurationInMinutes for row in self.report_data)
        total_hours = round(total_minutes / 60.0, 2)
        machine_counts = Counter(row.EquipmentName for row in self.report_data)
        top_machine = machine_counts.most_common(1)[0][0] if machine_counts else "N/D"

        self.kpi_total_tasks.config(text=f"{kpi1_label} {total_tasks}")
        self.kpi_total_time.config(text=f"{kpi2_label} {total_hours}")
        self.kpi_top_machine.config(text=f"{kpi3_label} {top_machine}")

    def _update_details_table(self):
        self.tree.delete(*self.tree.get_children())
        for row in self.report_data:
            self.tree.insert("", "end", values=(
                row.EquipmentName, row.InterventionType, row.TaskName, row.UserName,
                row.DateStart.strftime('%Y-%m-%d %H:%M'),
                row.DateStop.strftime('%Y-%m-%d %H:%M'),
                row.DurationInMinutes
            ))

    def _update_chart(self):
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        if not self.report_data: return
        machine_counts = Counter(row.EquipmentName for row in self.report_data)
        labels = list(machine_counts.keys())
        values = list(machine_counts.values())
        fig = plt.figure(figsize=(7, 5))
        ax = fig.add_subplot(111)
        ax.bar(labels, values, color='skyblue')
        ax.set_ylabel(self.lang.get('chart_y_axis', 'Numero di Interventi'))
        ax.set_title(self.lang.get('chart_title', 'Interventi di Manutenzione per Macchina'))
        plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def _export_to_pdf(self):
        messagebox.showinfo("Info", "Funzionalità di esportazione PDF in sviluppo.")

class AddTaskRowDialog(tk.Toplevel):
    """Dialogo modale per inserire o MODIFICARE i dettagli di un singolo task."""

    def __init__(self, parent, lang, initial_data=None):
        super().__init__(parent)
        self.lang = lang
        self.transient(parent)
        self.grab_set()

        self.is_edit_mode = initial_data is not None

        title = lang.get('edit_task_row_title', "Modifica Dettagli Compito") if self.is_edit_mode else lang.get(
            'add_task_row_title', "Aggiungi Dettagli Compito")
        self.title(title)

        self.task_name = ""
        self.description = ""
        self.document_data = None
        self.document_name = ""

        self.name_var = tk.StringVar()
        self.doc_path_var = tk.StringVar()

        frame = ttk.Frame(self, padding="15")
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text=lang.get('task_name_optional', "Nome Compito (Opzionale):")).grid(row=0, column=0,
                                                                                                sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.name_var).grid(row=0, column=1, sticky="ew", pady=5)

        ttk.Label(frame, text=lang.get('task_description_required', "Descrizione (*):")).grid(row=1, column=0,
                                                                                              sticky="nw", pady=5)
        self.desc_text = tk.Text(frame, height=5, wrap=tk.WORD)
        self.desc_text.grid(row=1, column=1, sticky="nsew", pady=5)
        frame.rowconfigure(1, weight=1)

        ttk.Label(frame, text=lang.get('task_document_optional', "Documento (Opzionale):")).grid(row=2, column=0,
                                                                                                 sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.doc_path_var, state="readonly").grid(row=2, column=1, sticky="ew", pady=5)
        ttk.Button(frame, text=lang.get('button_browse', "Sfoglia..."), command=self._browse_file).grid(row=2, column=2,
                                                                                                        padx=5)

        ttk.Button(frame, text="OK", command=self._on_ok).grid(row=3, column=1, columnspan=2, sticky="e", pady=10)
        self.desc_text.focus_set()

        if self.is_edit_mode:
            self._populate_fields(initial_data)

    def _populate_fields(self, data):
        """Carica i dati esistenti nei campi del dialogo."""
        self.name_var.set(data.get("name", ""))
        self.desc_text.insert("1.0", data.get("desc", ""))
        self.doc_path_var.set(data.get("doc_name", ""))
        self.document_data = data.get("doc_data")
        self.document_name = data.get("doc_name", "")

    def _browse_file(self):
        file_path = filedialog.askopenfilename(parent=self)
        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    self.document_data = f.read()
                self.document_name = os.path.basename(file_path)
                self.doc_path_var.set(self.document_name)
            except Exception as e:
                messagebox.showerror("Errore Lettura File", str(e), parent=self)

    def _on_ok(self):
        self.description = self.desc_text.get("1.0", tk.END).strip()
        if not self.description:
            messagebox.showwarning("Dato Mancante", "La descrizione è un campo obbligatorio.", parent=self)
            return
        self.task_name = self.name_var.get().strip()
        self.destroy()

class BrandManagerWindow(tk.Toplevel):
    """Finestra per la gestione dei Brand delle macchine."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(lang.get('brand_management_title', "Gestione Brand"))
        self.geometry("800x500")

        self.suppliers_data = {}
        self.current_brand_id = None
        self.logo_binary_data = None
        self.brands_list = []

        self.brand_name_var = tk.StringVar()
        self.supplier_var = tk.StringVar()

        self._create_widgets()
        self._load_suppliers()
        self._load_brands()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=3)
        main_frame.columnconfigure(1, weight=2)

        list_frame = ttk.LabelFrame(main_frame, text=self.lang.get('brands_list_label', "Lista Brand"))
        list_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        cols = ('brand', 'company')
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings")
        self.tree.heading('brand', text=self.lang.get('header_brand_name', "Nome Brand"))
        self.tree.heading('company', text=self.lang.get('header_supplier_name', "Produttore"))
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<<TreeviewSelect>>", self._on_brand_select)

        form_frame = ttk.LabelFrame(main_frame, text=self.lang.get('details_label', "Dettagli"))
        form_frame.grid(row=0, column=1, sticky="nsew")
        form_frame.columnconfigure(1, weight=1)

        ttk.Label(form_frame, text=self.lang.get('supplier_label', "Produttore (*):")).grid(row=0, column=0, sticky="w",
                                                                                            padx=5, pady=5)
        self.supplier_combo = ttk.Combobox(form_frame, textvariable=self.supplier_var, state="readonly")
        self.supplier_combo.grid(row=0, column=1, sticky="ew", padx=5)

        ttk.Label(form_frame, text=self.lang.get('brand_name_label', "Nome Brand (*):")).grid(row=1, column=0,
                                                                                              sticky="w", padx=5,
                                                                                              pady=5)
        self.brand_entry = ttk.Entry(form_frame, textvariable=self.brand_name_var)
        self.brand_entry.grid(row=1, column=1, sticky="ew", padx=5)

        ttk.Label(form_frame, text=self.lang.get('logo_label', "Logo:")).grid(row=2, column=0, sticky="w", padx=5,
                                                                              pady=5)
        ttk.Button(form_frame, text=self.lang.get('load_logo_button', "Carica..."), command=self._load_logo).grid(row=2,
                                                                                                                  column=1,
                                                                                                                  sticky="w",
                                                                                                                  padx=5)

        self.logo_label = ttk.Label(form_frame, background="lightgrey",
                                    text=self.lang.get('no_logo_text', "Nessun logo"))
        self.logo_label.grid(row=3, column=0, columnspan=2, pady=10, sticky="ew", ipady=20)

        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text=self.lang.get('new_button_short', "Nuovo"), command=self._clear_form).pack(
            side="left", padx=5)
        ttk.Button(btn_frame, text=self.lang.get('save_button', "Salva"), command=self._save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text=self.lang.get('delete_button', "Cancella"), command=self._delete).pack(side="left",
                                                                                                          padx=5)

    def _load_suppliers(self):
        suppliers = self.db.fetch_supplier_sites()
        if suppliers:
            self.suppliers_data = {s.SiteName: s.IDSite for s in suppliers}
            self.supplier_combo['values'] = sorted(list(self.suppliers_data.keys()))

    def _load_brands(self):
        self.tree.delete(*self.tree.get_children())
        self.brands_list = self.db.fetch_brands_with_company_name()
        for brand in self.brands_list:
            # --- CORREZIONE QUI: Usa 'CompanyName' invece di 'SiteName' ---
            company_name = brand.CompanyName or "N/A"
            self.tree.insert("", "end", iid=brand.EquipmentBrandId, values=(brand.Brand, company_name))

    def _on_brand_select(self, event=None):
        selected_item = self.tree.focus()
        if not selected_item: return
        self.current_brand_id = int(selected_item)

        selected_brand = next((b for b in self.brands_list if b.EquipmentBrandId == self.current_brand_id), None)
        if not selected_brand: return

        self.brand_name_var.set(selected_brand.Brand)
        supplier_name = next((name for name, id in self.suppliers_data.items() if id == selected_brand.CompanyId), "")
        self.supplier_var.set(supplier_name)
        self.logo_binary_data = selected_brand.BrandLogo
        self._display_logo()

    def _clear_form(self):
        self.tree.selection_set([])
        self.current_brand_id = None
        self.brand_name_var.set("")
        self.supplier_var.set("")
        self.logo_binary_data = None
        self._display_logo()
        self.brand_entry.focus_set()

    def _load_logo(self):
        """Apre una finestra di dialogo per selezionare un file immagine."""
        if not PIL_AVAILABLE:
            messagebox.showwarning("Libreria Mancante",
                                   "La libreria Pillow non è installata. Impossibile caricare immagini.")
            return

        file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.bmp")])
        if file_path:
            with open(file_path, 'rb') as f:
                self.logo_binary_data = f.read()
            self._display_logo()

    def _display_logo(self):
        """Mostra un'anteprima del logo caricato."""
        if self.logo_binary_data and PIL_AVAILABLE:
            try:
                image = Image.open(io.BytesIO(self.logo_binary_data))
                image = image.convert("RGB")
                image.thumbnail((150, 150))
                self.logo_photo = ImageTk.PhotoImage(image)
                self.logo_label.config(image=self.logo_photo, text="")
            except Exception as e:
                print(f"ERRORE caricamento logo: {e}")
                self.logo_label.config(image="", text=self.lang.get('logo_error_text', "Errore logo"))
        else:
            self.logo_label.config(image="", text=self.lang.get('no_logo_text', "Nessun logo"))

    def _save(self):
        brand_name = self.brand_name_var.get().strip()
        supplier_name = self.supplier_var.get()
        if not brand_name or not supplier_name:
            messagebox.showerror("Dati Mancanti", "Produttore e Nome Brand sono obbligatori.", parent=self)
            return

        company_id = self.suppliers_data.get(supplier_name)
        if company_id is None:
            messagebox.showerror("Errore", "Selezionare un produttore valido dalla lista.", parent=self)
            return

        if self.current_brand_id:
            success, message = self.db.update_brand(self.current_brand_id, brand_name, company_id,
                                                    self.logo_binary_data)
        else:
            success, message = self.db.add_new_brand(brand_name, company_id, self.logo_binary_data)

        if success:
            messagebox.showinfo("Successo", message, parent=self)
            self._load_brands()
            self._clear_form()
        else:
            messagebox.showerror("Errore", message, parent=self)

    def _delete(self):
        if not self.current_brand_id:
            messagebox.showwarning("Nessuna Selezione", "Selezionare un brand da cancellare.", parent=self)
            return

        if self.db.check_if_brand_is_used(self.current_brand_id):
            messagebox.showerror("Operazione non permessa",
                                 "Impossibile cancellare un brand già associato a una o più macchine.", parent=self)
            return

        if messagebox.askyesno("Conferma Cancellazione", "Sei sicuro di voler cancellare questo brand?"):
            success, message = self.db.delete_brand(self.current_brand_id)
            if success:
                messagebox.showinfo("Successo", message, parent=self)
                self._load_brands()
                self._clear_form()
            else:
                messagebox.showerror("Errore", message, parent=self)

class AddMaintenanceTasksWindow(tk.Toplevel):
    """
    Finestra per la GESTIONE (Aggiunta, Modifica, Rimozione) dei task.
    VERSIONE CORRETTA E COMPLETA.
    """

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(lang.get('submenu_add_maint_task', "Gestione Task di Manutenzione"))
        self.geometry("950x600")

        # 1. Inizializzazione delle variabili di stato
        self.equipments_data = {}
        self.interventions_data = {}
        self.current_tasks = []
        self.initial_task_ids = set()

        # 2. Inizializzazione delle variabili Tkinter
        self.equipment_var = tk.StringVar()
        self.intervention_var = tk.StringVar()
        self.category_var = tk.StringVar()

        # 3. Creazione dei widget e caricamento dati
        self._create_widgets()
        self._load_filters()

    def _create_widgets(self):
        """Crea tutti i componenti grafici della finestra."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)

        filter_frame = ttk.LabelFrame(main_frame, text="Seleziona e Carica", padding="10")
        filter_frame.grid(row=0, column=0, sticky="ew", pady=5)
        filter_frame.columnconfigure(1, weight=1)
        filter_frame.columnconfigure(3, weight=1)

        # Creazione dei widget Combobox (vengono creati PRIMA di essere popolati)
        ttk.Label(filter_frame, text="Macchina:").grid(row=0, column=0, sticky="w", padx=5)
        self.equipment_combo = ttk.Combobox(filter_frame, textvariable=self.equipment_var, state='readonly')
        self.equipment_combo.grid(row=0, column=1, sticky="ew", padx=(0, 15))
        self.equipment_combo.bind("<<ComboboxSelected>>", self._load_existing_tasks)

        ttk.Label(filter_frame, text="Tipo Intervento:").grid(row=0, column=2, sticky="w", padx=5)
        self.intervention_combo = ttk.Combobox(filter_frame, textvariable=self.intervention_var, state='readonly')
        self.intervention_combo.grid(row=0, column=3, sticky="ew")
        self.intervention_combo.bind("<<ComboboxSelected>>", self._load_existing_tasks)

        ttk.Label(filter_frame, text="Categoria/Titolo Generale:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.category_entry = ttk.Entry(filter_frame, textvariable=self.category_var)
        self.category_entry.grid(row=1, column=1, columnspan=3, sticky="ew", pady=5)

        # Creazione della Treeview e dei pulsanti
        tasks_frame = ttk.LabelFrame(main_frame, text="Compiti", padding="10")
        tasks_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        tasks_frame.rowconfigure(0, weight=1)
        tasks_frame.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tasks_frame, columns=("ordine", "nome", "desc", "doc"), show="headings")
        self.tree.heading("ordine", text="#")
        self.tree.heading("nome", text="Nome Compito")
        self.tree.heading("desc", text="Descrizione")
        self.tree.heading("doc", text="Documento Allegato")
        self.tree.column("ordine", width=40, stretch=tk.NO)
        self.tree.column("nome", width=200)
        self.tree.column("desc", width=400)
        self.tree.column("doc", width=150)
        self.tree.grid(row=0, column=0, sticky="nsew")

        action_frame = ttk.Frame(tasks_frame)
        action_frame.grid(row=0, column=1, sticky="ns", padx=5)
        ttk.Button(action_frame, text=self.lang.get('add_button_text', "Aggiungi..."), command=self._add_task_row).pack(
            pady=2, fill='x')
        ttk.Button(action_frame, text=self.lang.get('edit_button_text', "Modifica..."),
                   command=self._edit_task_row).pack(pady=2, fill='x')
        ttk.Button(action_frame, text=self.lang.get('remove_button_text', "Rimuovi"),
                   command=self._remove_selected_row).pack(pady=2, fill='x')

        save_button = ttk.Button(main_frame, text="Salva Modifiche", command=self._save_all_tasks)
        save_button.grid(row=2, column=0, sticky="e", pady=10)

    def _load_filters(self):
        """Carica i dati dal DB e li inserisce nei combobox."""
        # Carica Macchine
        equipments = self.db.fetch_all_equipments()
        if equipments:
            self.equipments_data = {f"{r.InternalName or ''} [{r.SerialNumber}]": r.EquipmentId for r in equipments}
            self.equipment_combo['values'] = sorted(list(self.equipments_data.keys()))

        # Carica Tipi di Intervento
        interventions = self.db.fetch_maintenance_interventions()
        if interventions:
            self.interventions_data = {r.TimingDescriprion: r.ProgrammedInterventionId for r in interventions}
            self.intervention_combo['values'] = sorted(list(self.interventions_data.keys()))

    def _load_existing_tasks(self, event=None):
        """Carica i task esistenti SOLO SE sia la macchina che l'intervento sono selezionati."""
        equipment_sel = self.equipment_var.get()
        intervention_sel = self.intervention_var.get()

        # Pulisce sempre la lista prima di un'azione
        self.current_tasks.clear()
        self.initial_task_ids.clear()
        self.category_var.set("")  # Pulisce anche la categoria
        self._refresh_treeview()

        # Procede solo se entrambi i filtri sono stati selezionati
        if not equipment_sel or not intervention_sel:
            return

        equipment_id = self.equipments_data.get(equipment_sel)
        intervention_id = self.interventions_data.get(intervention_sel)

        if not equipment_id or not intervention_id: return

        existing_tasks = self.db.fetch_tasks_for_editing(intervention_id, equipment_id)

        if existing_tasks:
            for task in existing_tasks:
                doc_name = "File presente" if task.LinkedDocument else ""
                self.current_tasks.append({
                    "id": task.CompitoId, "name": task.NomeCompito, "desc": task.DescrizioneCompito,
                    "doc_name": doc_name, "doc_data": task.LinkedDocument
                })
            self.initial_task_ids = {task['id'] for task in self.current_tasks}
            self.category_var.set(existing_tasks[0].Categoria)

        self._refresh_treeview()

    def _add_task_row(self):
        dialog = AddTaskRowDialog(self, self.lang)
        self.wait_window(dialog)
        if dialog.description:
            task_data = {"id": None, "name": dialog.task_name, "desc": dialog.description,
                         "doc_name": dialog.document_name, "doc_data": dialog.document_data}
            self.current_tasks.append(task_data)
            self._refresh_treeview()

    def _edit_task_row(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Nessuna Selezione", "Selezionare una riga da modificare.", parent=self)
            return
        item_index = self.tree.index(selected_item)
        existing_data = self.current_tasks[item_index]
        dialog = AddTaskRowDialog(self, self.lang, initial_data=existing_data)
        self.wait_window(dialog)
        if dialog.description:
            updated_data = {"id": existing_data['id'], "name": dialog.task_name, "desc": dialog.description,
                            "doc_name": dialog.document_name, "doc_data": dialog.document_data}
            self.current_tasks[item_index] = updated_data
            self._refresh_treeview()

    def _remove_selected_row(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Nessuna Selezione", "Selezionare una riga da rimuovere.", parent=self)
            return
        if messagebox.askyesno("Conferma Rimozione", "Sei sicuro di voler rimuovere questa riga?"):
            index_to_remove = self.tree.index(selected_item)
            del self.current_tasks[index_to_remove]
            self._refresh_treeview()

    def _refresh_treeview(self):
        self.tree.delete(*self.tree.get_children())
        for i, task in enumerate(self.current_tasks, 1):
            self.tree.insert("", "end", values=(i, task["name"], task["desc"], task["doc_name"]))

    def _save_all_tasks(self):
        equipment_sel = self.equipment_var.get()
        intervention_sel = self.intervention_var.get()
        category = self.category_var.get().strip()

        if not equipment_sel or not intervention_sel or not category:
            messagebox.showerror("Dati Mancanti",
                                 "Selezionare Macchina, Tipo di Intervento e inserire una Categoria prima di salvare.")
            return

        equipment_id = self.equipments_data[equipment_sel]
        intervention_id = self.interventions_data[intervention_sel]

        current_task_ids = {task['id'] for task in self.current_tasks if task['id'] is not None}
        ids_to_delete = list(self.initial_task_ids - current_task_ids)
        tasks_to_update = [t for t in self.current_tasks if t['id'] is not None]
        tasks_to_insert = [t for t in self.current_tasks if t['id'] is None]

        msg = (
            f"Stai per:\n- AGGIORNARE {len(tasks_to_update)} compiti.\n- INSERIRE {len(tasks_to_insert)} nuovi compiti.\n- CANCELLARE {len(ids_to_delete)} compiti.\n\nProcedere?")
        if not messagebox.askyesno("Riepilogo e Conferma", msg):
            return

        errors = []
        if ids_to_delete:
            if not self.db.delete_maintenance_tasks(ids_to_delete):
                errors.append(f"Cancellazione fallita: {self.db.last_error_details}")

        for i, task in enumerate(self.current_tasks, 1):
            task_name = task["name"] if task["name"] else category

            if task['id'] is not None:  # UPDATE
                if not self.db.update_maintenance_task(task['id'], equipment_id, category, task_name, task['desc'],
                                                       task['doc_data']):
                    errors.append(f"Aggiornamento fallito per ID {task['id']}: {self.db.last_error_details}")
            else:  # INSERT
                if not self.db.insert_new_maintenance_task(intervention_id, equipment_id, category, task_name,
                                                           task['desc'], i, task['doc_data']):
                    errors.append(f"Inserimento fallito per '{task_name}': {self.db.last_error_details}")

        if not errors:
            messagebox.showinfo("Successo", "Tutte le modifiche sono state salvate con successo.")
            self.destroy()
        else:
            messagebox.showerror("Errori durante il Salvataggio", "\n".join(errors))

def open_fill_templates(parent, db, lang, user_name=None):
    # Controlla se user_name è stato fornito (dovrebbe esserlo se chiamato tramite login da App)
    if user_name:
        FillTemplateWindow(parent, db, lang, user_name)  # Passa user_name alla classe
    else:
        # Fallback di sicurezza nel caso venga chiamata erroneamente senza utente
        print("Errore: Tentativo di aprire Compilazione Schede senza autenticazione.")

def open_add_machine(parent, db, lang):
    AddMachineWindow(parent, db, lang)

def open_edit_machine(parent, db, lang):
    # L'utente autenticato viene passato qui
    user_name = parent.authenticated_user_for_maintenance # Dovremo aggiungere questa variabile
    SelectMachineToEditWindow(parent, db, lang, user_name)

def open_view_machines(parent, db, lang):
    ViewMachineWindow(parent, db, lang)

def open_maintenance_docs(parent, db, lang):
    MaintenanceDocsWindow(parent, db, lang)

def open_fill_templates(parent, db, lang, user_name=None):
    # Controlla se user_name è stato fornito (dovrebbe esserlo se chiamato tramite login da App)
    if user_name:
        FillTemplateWindow(parent, db, lang, user_name) # Passa user_name alla classe
    else:
        # Fallback di sicurezza nel caso venga chiamata erroneamente senza utente
        print("Errore: Tentativo di aprire Compilazione Schede senza autenticazione.")

def open_add_maintenance_tasks(parent, db, lang, user_name):
    """Launcher function to create and show the AddMaintenanceTasksWindow."""
    AddMaintenanceTasksWindow(parent, db, lang, user_name)

def open_add_maintenance_doc(parent, db, lang, user_name):
    MaintenanceDocsWindow(parent, db, lang, user_name)

def open_reports(parent, db, lang):
    # Sostituisce la vecchia finestra segnaposto con quella nuova
    MaintenanceReportWindow(parent, db, lang)

def open_brand_manager(parent, db, lang, user_name):
    BrandManagerWindow(parent, db, lang, user_name)

def generate_missing_action_report(parent, db, lang):
    """
    Esegue la stored procedure e genera un report Excel formattato,
    salvandolo automaticamente in C:\Temp e aprendolo.
    """
    # 1. Esegui la stored procedure
    results, headers, error = db.execute_missing_action_report()

    if error:
        messagebox.showerror(lang.get('error_title', "Errore"), error, parent=parent)
        return

    if not results:
        messagebox.showinfo(lang.get('info_title', "Info"),
                            lang.get('no_missing_actions_found', "Nessuna manutenzione mancante trovata."),
                            parent=parent)
        return

    # --- NUOVA LOGICA DI SALVATAGGIO ---
    try:
        # 2. Definisci il percorso e crea la cartella se non esiste
        temp_dir = "C:\\Temp"
        os.makedirs(temp_dir, exist_ok=True)

        # 3. Crea un nome file univoco con data e ora
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Report_Azioni_Mancanti_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, file_name)

        # 4. Crea e formatta il file Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Azioni Mancanti"

        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row_data in results:
            ws.append(list(row_data))

        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 5. Salva il file
        wb.save(file_path)

        # 6. Avvisa l'utente e apri il file
        messagebox.showinfo(
            lang.get('success_title', "Successo"),
            f"{lang.get('report_saved_successfully', 'Report salvato con successo in:')}\n{file_path}",
            parent=parent
        )
        os.startfile(file_path)

    except PermissionError:
        messagebox.showerror(lang.get('error_title', "Errore"),
                             f"Errore di permesso. Impossibile salvare il file in {file_path}.\nAssicurati che il file non sia già aperto o di avere i permessi di scrittura.",
                             parent=parent)
    except Exception as e:
        messagebox.showerror(lang.get('error_title', "Errore"),
                             f"{lang.get('error_generating_excel', 'Impossibile generare il file Excel:')}\n{e}",
                             parent=parent)


