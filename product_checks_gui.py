"""
Modulo per la gestione delle verifiche periodiche sui prodotti.
Gestisce:
- Configurazione periodicità verifiche per prodotto
- Gestione task di verifica (generici e specifici)
- Esecuzione verifiche con checklist
- Report verifiche (da completare)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tempfile
import os
import subprocess
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger("TraceabilityRS")  # usa la config fatta in main.py

class ProductChecksManagementWindow(tk.Toplevel):
    """Finestra per la gestione della periodicità delle verifiche prodotti"""

    def __init__(self, parent, db_handler, lang_manager, user_id):
        super().__init__(parent)
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_id

        self.title(self.lang.get('product_checks_mgmt_title', 'Gestione Verifiche Prodotti'))
        self.geometry('900x600')
        self.transient(parent)

        self._current_check_id = None
        self._build_ui()
        self._load_products()
        self._load_checks()

    def _build_ui(self):
        """Costruisce l'interfaccia"""
        # Header con nome utente
        header = ttk.Frame(self)
        header.pack(fill='x', padx=10, pady=5)
        ttk.Label(header, text=f"{self.lang.get('logged_user', 'Utente')}: {self.user_name}",
                  font=('Arial', 10, 'bold')).pack(side='left')

        # Frame principale diviso in due colonne
        main_frame = ttk.Frame(self)
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Colonna sinistra: Form inserimento
        left_frame = ttk.LabelFrame(main_frame, text=self.lang.get('product_check_form', 'Configurazione Verifica'))
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

        # Combo prodotti
        ttk.Label(left_frame, text=self.lang.get('product', 'Prodotto') + ' *').grid(row=0, column=0, sticky='w',
                                                                                     padx=5, pady=5)
        self.product_combo = ttk.Combobox(left_frame, state='readonly', width=40)
        self.product_combo.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Periodicità in quantità
        ttk.Label(left_frame, text=self.lang.get('periodicity_qty', 'Periodicità (Quantità)') + ' *').grid(row=1,
                                                                                                           column=0,
                                                                                                           sticky='w',
                                                                                                           padx=5,
                                                                                                           pady=5)
        self.periodicity_var = tk.StringVar()
        ttk.Entry(left_frame, textvariable=self.periodicity_var, width=20).grid(row=1, column=1, sticky='w', padx=5,
                                                                                pady=5)

        # Pulsanti azione
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text=self.lang.get('btn_new', 'Nuovo'), command=self._on_new).pack(side='left', padx=2)
        ttk.Button(btn_frame, text=self.lang.get('btn_save', 'Salva'), command=self._on_save).pack(side='left', padx=2)
        ttk.Button(btn_frame, text=self.lang.get('btn_delete', 'Elimina'), command=self._on_delete).pack(side='left',
                                                                                                         padx=2)

        left_frame.columnconfigure(1, weight=1)

        # Colonna destra: Lista verifiche configurate
        right_frame = ttk.LabelFrame(main_frame, text=self.lang.get('configured_checks', 'Verifiche Configurate'))
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        # Treeview
        columns = ('id', 'product_code', 'product_name', 'periodicity')
        self.tree = ttk.Treeview(right_frame, columns=columns, show='headings', height=20)
        self.tree.heading('id', text='ID')
        self.tree.heading('product_code', text=self.lang.get('product_code', 'Codice'))
        self.tree.heading('product_name', text=self.lang.get('product_name', 'Nome Prodotto'))
        self.tree.heading('periodicity', text=self.lang.get('periodicity', 'Periodicità'))

        self.tree.column('id', width=50)
        self.tree.column('product_code', width=120)
        self.tree.column('product_name', width=250)
        self.tree.column('periodicity', width=100)

        scrollbar = ttk.Scrollbar(right_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.tree.bind('<<TreeviewSelect>>', self._on_select)

    def _load_products(self):
        """Carica la lista prodotti nel combo"""
        products = self.db.fetch_products_for_checks()
        self.products_dict = {f"{p.ProductCode} - {p.ProductName}": p.IDProduct for p in products}
        self.product_combo['values'] = list(self.products_dict.keys())

    def _load_checks(self):
        """Carica le verifiche configurate"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        checks = self.db.fetch_all_product_checks()
        for check in checks:
            self.tree.insert('', 'end', values=(
                check.PeriodicalProductCheckId,
                check.ProductCode,
                check.ProductName,
                check.PeriodicityInQty
            ))

    def _on_select(self, event):
        """Gestisce la selezione di una riga"""
        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.item(selection[0])
        values = item['values']

        self._current_check_id = values[0]
        # Trova il prodotto nel combo
        product_key = f"{values[1]} - {values[2]}"
        if product_key in self.products_dict:
            self.product_combo.set(product_key)
        self.periodicity_var.set(str(values[3]))

    def _on_new(self):
        """Pulisce il form per nuovo inserimento"""
        self._current_check_id = None
        self.product_combo.set('')
        self.periodicity_var.set('')

    def _on_save(self):
        """Salva o aggiorna una verifica"""
        # Validazione
        if not self.product_combo.get():
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('select_product', 'Selezionare un prodotto'))
            return

        try:
            periodicity = int(self.periodicity_var.get())
            if periodicity <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('invalid_periodicity', 'Inserire una periodicità valida (numero > 0)'))
            return

        product_id = self.products_dict[self.product_combo.get()]

        if self._current_check_id:
            # Update
            success = self.db.update_product_check(self._current_check_id, product_id, periodicity)
        else:
            # Insert
            success = self.db.insert_product_check(product_id, periodicity)

        if success:
            messagebox.showinfo(self.lang.get('success', 'Successo'),
                                self.lang.get('check_saved', 'Verifica salvata con successo'))
            self._load_checks()
            self._on_new()
        else:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('save_error', 'Errore durante il salvataggio'))

    def _on_delete(self):
        """Elimina una verifica"""
        if not self._current_check_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('select_check', 'Selezionare una verifica da eliminare'))
            return

        if messagebox.askyesno(self.lang.get('confirm', 'Conferma'),
                               self.lang.get('confirm_delete_check', 'Eliminare la verifica selezionata?')):
            if self.db.delete_product_check(self._current_check_id):
                messagebox.showinfo(self.lang.get('success', 'Successo'),
                                    self.lang.get('check_deleted', 'Verifica eliminata'))
                self._load_checks()
                self._on_new()
            else:
                messagebox.showerror(self.lang.get('error', 'Errore'),
                                     self.lang.get('delete_error', 'Errore durante l\'eliminazione'))

class CheckTasksManagementWindow(tk.Toplevel):
    """Finestra per la gestione dei task di verifica (generici e specifici)"""

    def __init__(self, parent, db_handler, lang_manager, user_id):
        super().__init__(parent)
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_id

        self.title(self.lang.get('check_tasks_mgmt_title', 'Gestione Task di Verifica'))
        self.geometry('1000x700')
        self.transient(parent)

        self._current_task_id = None
        self._doc_data = None
        self._build_ui()
        self._load_products()
        self._load_tasks()

    def _build_ui(self):
        """Costruisce l'interfaccia"""
        # Header
        header = ttk.Frame(self)
        header.pack(fill='x', padx=10, pady=5)
        ttk.Label(header, text=f"{self.lang.get('logged_user', 'Utente')}: {self.user_name}",
                  font=('Arial', 10, 'bold')).pack(side='left')

        # Frame principale
        main_frame = ttk.Frame(self)
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Colonna sinistra: Form
        left_frame = ttk.LabelFrame(main_frame, text=self.lang.get('task_form', 'Task di Verifica'))
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

        # Azione da eseguire
        ttk.Label(left_frame, text=self.lang.get('item_to_check', 'Azione da Eseguire') + ' *').grid(row=0, column=0,
                                                                                                     sticky='w', padx=5,
                                                                                                     pady=5)
        self.item_var = tk.StringVar()
        ttk.Entry(left_frame, textvariable=self.item_var, width=50).grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Checkbox IsGeneric
        self.is_generic_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(left_frame, text=self.lang.get('is_generic', 'Task Generico'),
                        variable=self.is_generic_var, command=self._on_generic_changed).grid(row=1, column=0,
                                                                                             columnspan=2, sticky='w',
                                                                                             padx=5, pady=5)

        # Combo prodotti (disabilitato se generico)
        ttk.Label(left_frame, text=self.lang.get('specific_product', 'Prodotto Specifico')).grid(row=2, column=0,
                                                                                                 sticky='w', padx=5,
                                                                                                 pady=5)
        self.product_combo = ttk.Combobox(left_frame, state='disabled', width=47)
        self.product_combo.grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        # Documento allegato
        doc_frame = ttk.Frame(left_frame)
        doc_frame.grid(row=3, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        ttk.Label(doc_frame, text=self.lang.get('document', 'Documento')).pack(side='left')
        ttk.Button(doc_frame, text=self.lang.get('btn_upload', 'Carica'), command=self._upload_doc).pack(side='left',
                                                                                                         padx=5)
        self.doc_label = ttk.Label(doc_frame, text='', foreground='blue')
        self.doc_label.pack(side='left', padx=5)

        # Pulsanti azione
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text=self.lang.get('btn_new', 'Nuovo'), command=self._on_new).pack(side='left', padx=2)
        ttk.Button(btn_frame, text=self.lang.get('btn_save', 'Salva'), command=self._on_save).pack(side='left', padx=2)
        ttk.Button(btn_frame, text=self.lang.get('btn_delete', 'Elimina'), command=self._on_delete).pack(side='left',
                                                                                                         padx=2)

        left_frame.columnconfigure(1, weight=1)

        # Colonna destra: Lista task
        right_frame = ttk.LabelFrame(main_frame, text=self.lang.get('configured_tasks', 'Task Configurati'))
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        columns = ('id', 'item', 'is_generic', 'product', 'has_doc', 'user', 'date')
        self.tree = ttk.Treeview(right_frame, columns=columns, show='headings', height=25)
        self.tree.heading('id', text='ID')
        self.tree.heading('item', text=self.lang.get('item', 'Azione'))
        self.tree.heading('is_generic', text=self.lang.get('type', 'Tipo'))
        self.tree.heading('product', text=self.lang.get('product', 'Prodotto'))
        self.tree.heading('has_doc', text=self.lang.get('doc', 'Doc'))
        self.tree.heading('user', text=self.lang.get('user', 'Utente'))
        self.tree.heading('date', text=self.lang.get('date', 'Data'))

        self.tree.column('id', width=50)
        self.tree.column('item', width=250)
        self.tree.column('is_generic', width=80)
        self.tree.column('product', width=150)
        self.tree.column('has_doc', width=50)
        self.tree.column('user', width=100)
        self.tree.column('date', width=100)

        scrollbar = ttk.Scrollbar(right_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.tree.bind('<<TreeviewSelect>>', self._on_select)
        self.tree.bind('<Double-Button-1>', self._on_double_click)

    def _on_generic_changed(self):
        """Gestisce il cambio dello stato del checkbox IsGeneric"""
        if self.is_generic_var.get():
            self.product_combo.set('')
            self.product_combo.config(state='disabled')
        else:
            self.product_combo.config(state='readonly')

    def _load_products(self):
        """Carica i prodotti con verifiche configurate"""
        products = self.db.fetch_products_with_checks()
        self.products_dict = {f"{p.ProductCode} - {p.ProductName}": p.PeriodicalProductCheckLogId for p in products}
        self.product_combo['values'] = list(self.products_dict.keys())

    def _load_tasks(self):
        """Carica i task configurati"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        tasks = self.db.fetch_all_check_tasks()
        for task in tasks:
            self.tree.insert('', 'end', values=(
                task.PriodicalProductCheckListId,
                task.ItemToCheck,
                self.lang.get('generic', 'Generico') if task.IsGeneric else self.lang.get('specific', 'Specifico'),
                task.ProductCode or '',
                '✓' if task.Doc else '',
                task.UserType,
                task.DateIn.strftime('%d/%m/%Y') if task.DateIn else ''
            ))

    def _upload_doc(self):
        """Carica un documento"""
        file_path = filedialog.askopenfilename(
            title=self.lang.get('select_document', 'Seleziona Documento'),
            filetypes=[
                ('PDF', '*.pdf'),
                ('Word', '*.doc;*.docx'),
                ('Excel', '*.xls;*.xlsx'),
                ('All files', '*.*')
            ]
        )

        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    self._doc_data = f.read()
                self.doc_label.config(text=os.path.basename(file_path))
            except Exception as e:
                messagebox.showerror(self.lang.get('error', 'Errore'),
                                     f"{self.lang.get('upload_error', 'Errore caricamento')}: {str(e)}")

    def _on_select(self, event):
        """Gestisce la selezione di una riga"""
        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.item(selection[0])
        values = item['values']

        self._current_task_id = values[0]

        # Carica i dettagli del task
        task = self.db.fetch_check_task_by_id(self._current_task_id)
        if task:
            self.item_var.set(task.ItemToCheck)
            self.is_generic_var.set(task.IsGeneric)
            self._on_generic_changed()

            if not task.IsGeneric and task.ProductCode:
                product_key = f"{task.ProductCode} - {task.ProductName}"
                if product_key in self.products_dict:
                    self.product_combo.set(product_key)

            self._doc_data = task.Doc
            if task.Doc:
                self.doc_label.config(text=self.lang.get('doc_present', 'Documento presente'))
            else:
                self.doc_label.config(text='')

    def _on_double_click(self, event):
        """Apre il documento se presente"""
        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.item(selection[0])
        task_id = item['values'][0]

        task = self.db.fetch_check_task_by_id(task_id)
        if task and task.Doc:
            self._open_document(task.Doc)

    def _open_document(self, doc_data):
        """Apre un documento binario"""
        try:
            # Crea un file temporaneo
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(doc_data)
                tmp_path = tmp_file.name

            # Apre il file con l'applicazione predefinita
            if os.name == 'nt':  # Windows
                os.startfile(tmp_path)
            elif os.name == 'posix':  # Linux/Mac
                subprocess.call(['xdg-open', tmp_path])
        except Exception as e:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 f"{self.lang.get('open_error', 'Errore apertura documento')}: {str(e)}")

    def _on_new(self):
        """Pulisce il form"""
        self._current_task_id = None
        self.item_var.set('')
        self.is_generic_var.set(True)
        self.product_combo.set('')
        self._doc_data = None
        self.doc_label.config(text='')
        self._on_generic_changed()

    def _on_save(self):
        """Salva o aggiorna un task"""
        # Validazione
        if not self.item_var.get().strip():
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('enter_item', 'Inserire l\'azione da eseguire'))
            return

        is_generic = self.is_generic_var.get()
        product_check_id = None

        if not is_generic:
            if not self.product_combo.get():
                messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                       self.lang.get('select_product', 'Selezionare un prodotto'))
                return
            product_check_id = self.products_dict[self.product_combo.get()]

        if self._current_task_id:
            # Update
            success = self.db.update_check_task(
                self._current_task_id,
                self.item_var.get().strip(),
                is_generic,
                self.user_name,
                self._doc_data,
                product_check_id
            )
        else:
            # Insert
            success = self.db.insert_check_task(
                self.item_var.get().strip(),
                is_generic,
                self.user_name,
                self._doc_data,
                product_check_id
            )

        if success:
            messagebox.showinfo(self.lang.get('success', 'Successo'),
                                self.lang.get('task_saved', 'Task salvato con successo'))
            self._load_tasks()
            self._on_new()
        else:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('save_error', 'Errore durante il salvataggio'))

    def _on_delete(self):
        """Elimina un task (soft delete)"""
        if not self._current_task_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('select_task', 'Selezionare un task da eliminare'))
            return

        if messagebox.askyesno(self.lang.get('confirm', 'Conferma'),
                               self.lang.get('confirm_delete_task', 'Eliminare il task selezionato?')):
            if self.db.delete_check_task(self._current_task_id):
                messagebox.showinfo(self.lang.get('success', 'Successo'),
                                    self.lang.get('task_deleted', 'Task eliminato'))
                self._load_tasks()
                self._on_new()
            else:
                messagebox.showerror(self.lang.get('error', 'Errore'),
                                     self.lang.get('delete_error', 'Errore durante l\'eliminazione'))

class ProductVerificationWindow(tk.Toplevel):
    """Finestra per l'esecuzione delle verifiche sui prodotti"""

    def __init__(self, parent, db_handler, lang_manager, user_id):
        super().__init__(parent)
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_id

        self.title(self.lang.get('product_verification_title', 'Esecuzione Verifiche Prodotti'))
        self.geometry('1200x750')
        self.transient(parent)

        self._current_must_check_id = None
        self._current_label_code_id = None  # Nuovo campo per memorizzare l'IDLabelCode
        self._check_items = []
        self._build_ui()
        self._load_products_to_check()
        # FORZA IL CARICAMENTO DEI TASK GENERICI ALL'APERTURA
        self.after(100, self._load_generic_tasks_only)

    def _load_generic_tasks_only(self):
        """Carica solo i task generici all'apertura della finestra"""
        self._load_checklist()

    def _on_label_code_changed(self, *args):
        """Verifica il label code quando viene modificato"""
        label_code = self.label_code_var.get().strip()

        if not label_code or not self._current_must_check_id:
            self.label_code_status_label.config(text='')
            self._current_label_code_id = None  # Reset IDLabelCode
            return

        # Esegui verifica dopo un breve delay per evitare troppe query
        self.after(500, self._perform_label_code_check, label_code)

    def _perform_label_code_check(self, label_code):
        """Esegue la verifica effettiva del label code"""
        if label_code != self.label_code_var.get().strip():
            return  # Il testo è cambiato durante il delay

        try:
            label_code_id = self.db.check_label_code_exists(label_code, self._current_must_check_id)

            if label_code_id:
                self.label_code_status_label.config(text='✓', foreground='green')
                self._current_label_code_id = label_code_id  # Memorizza l'IDLabelCode
            else:
                self.label_code_status_label.config(text='✗', foreground='red')
                self._current_label_code_id = None  # Reset IDLabelCode
                # Mostra messaggio di avviso
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('Label_Not_Found', 'Etichetta non trovata o non relativa al prodotto selezionato'),
                    parent=self
                )

        except Exception as e:
            self.label_code_status_label.config(text='', foreground='black')
            self._current_label_code_id = None  # Reset IDLabelCode
            #logger.error(f"Error during label code verification: {e}")

    def _on_result_changed(self, event=None):
        """Gestisce il cambio del risultato (mostra/nasconde campo commento)"""
        if self.result_var.get() == 'FAIL':
            self.comment_label.grid()
            self.comment_text.grid()
            self.comment_label.configure(text=self.lang.get('Comments', 'Commenti') + ' *')
        else:
            self.comment_label.grid_remove()
            self.comment_text.grid_remove()

    def _load_products_to_check(self):
        """Carica i prodotti che necessitano verifica"""
        for item in self.products_tree.get_children():
            self.products_tree.delete(item)

        products = self.db.fetch_products_must_check()
        for product in products:
            self.products_tree.insert('', 'end', values=(
                product.PeriodicalProductCheckMustListId,
                product.ordernumber,
                product.productcode,
                product.PhaseName,
                product.Date,
                product.Ora
            ))

    def _build_ui(self):

        """Costruisce l'interfaccia"""
        # Header
        header = ttk.Frame(self)
        header.pack(fill='x', padx=10, pady=5)
        ttk.Label(header, text=f"{self.lang.get('logged_user', 'Utente')}: {self.user_name}",
                  font=('Arial', 10, 'bold')).pack(side='left')

        # Frame principale con due listbox affiancate
        main_frame = ttk.Frame(self)
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Listbox sinistra: Prodotti da verificare
        left_frame = ttk.LabelFrame(main_frame, text=self.lang.get('products_to_check', 'Prodotti da Verificare'))
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

        columns_left = ('id', 'order', 'product', 'phase', 'date', 'time')
        self.products_tree = ttk.Treeview(left_frame, columns=columns_left, show='headings', height=25)
        self.products_tree.heading('id', text='ID')
        self.products_tree.heading('order', text=self.lang.get('order', 'Ordine'))
        self.products_tree.heading('product', text=self.lang.get('product', 'Prodotto'))
        self.products_tree.heading('phase', text=self.lang.get('phase', 'Fase'))
        self.products_tree.heading('date', text=self.lang.get('date', 'Data'))
        self.products_tree.heading('time', text=self.lang.get('time', 'Ora'))

        self.products_tree.column('id', width=50)
        self.products_tree.column('order', width=120)
        self.products_tree.column('product', width=120)
        self.products_tree.column('phase', width=150)
        self.products_tree.column('date', width=100)
        self.products_tree.column('time', width=80)

        scroll_left = ttk.Scrollbar(left_frame, orient='vertical', command=self.products_tree.yview)
        self.products_tree.configure(yscrollcommand=scroll_left.set)

        self.products_tree.pack(side='left', fill='both', expand=True)
        scroll_left.pack(side='right', fill='y')

        self.products_tree.bind('<<TreeviewSelect>>', self._on_product_select)

        # Listbox destra: Checklist
        right_frame = ttk.LabelFrame(main_frame, text=self.lang.get('checklist', 'Checklist Verifica'))
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        # Frame per input dati verifica
        input_frame = ttk.Frame(right_frame)
        input_frame.pack(fill='x', padx=5, pady=5)

        # Frame per label code con verifica
        label_frame = ttk.Frame(input_frame)
        label_frame.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=2)

        ttk.Label(label_frame, text=self.lang.get('label_code', 'Codice Label') + ' *').pack(side='left')

        # Frame per entry e icona verifica
        entry_frame = ttk.Frame(label_frame)
        entry_frame.pack(side='left', fill='x', expand=True, padx=(10, 5))

        self.label_code_var = tk.StringVar()
        self.label_code_entry = ttk.Entry(entry_frame, textvariable=self.label_code_var, width=20)
        self.label_code_entry.pack(side='left', fill='x', expand=True)

        # Icona verifica label code
        self.label_code_status_label = ttk.Label(entry_frame, text='', font=('Arial', 12))
        self.label_code_status_label.pack(side='left', padx=5)

        # Bind per verifica in tempo reale
        self.label_code_var.trace('w', self._on_label_code_changed)

        ttk.Label(input_frame, text=self.lang.get('result', 'Risultato') + ' *').grid(row=1, column=0, sticky='w',
                                                                                      padx=5, pady=2)
        self.result_var = tk.StringVar()
        self.result_combo = ttk.Combobox(input_frame, textvariable=self.result_var, state='readonly', width=18)
        self.result_combo['values'] = ['PASS', 'FAIL']
        self.result_combo.grid(row=1, column=1, sticky='w', padx=5, pady=2)
        self.result_combo.bind('<<ComboboxSelected>>', self._on_result_changed)

        # Campo commento (visibile solo quando risultato è FAIL)
        self.comment_label = ttk.Label(input_frame, text=self.lang.get('Comments', 'Commenti'))
        self.comment_label.grid(row=2, column=0, sticky='nw', padx=5, pady=2)

        self.comment_text = tk.Text(input_frame, width=40, height=4)
        self.comment_text.grid(row=2, column=1, sticky='ew', padx=5, pady=2)

        # Nascondi inizialmente il campo commento
        self.comment_label.grid_remove()
        self.comment_text.grid_remove()

        input_frame.columnconfigure(1, weight=1)

        # Treeview per checklist con checkbox
        columns_right = ('check', 'id', 'item', 'doc')
        self.checklist_tree = ttk.Treeview(right_frame, columns=columns_right, show='tree headings', height=15)
        self.checklist_tree.heading('#0', text='✓')
        self.checklist_tree.heading('check', text='')
        self.checklist_tree.heading('id', text='ID')
        self.checklist_tree.heading('item', text=self.lang.get('item', 'Azione'))
        self.checklist_tree.heading('doc', text=self.lang.get('doc', 'Doc'))

        self.checklist_tree.column('#0', width=30)
        self.checklist_tree.column('check', width=0, stretch=False)
        self.checklist_tree.column('id', width=50)
        self.checklist_tree.column('item', width=350)
        self.checklist_tree.column('doc', width=50)

        scroll_right = ttk.Scrollbar(right_frame, orient='vertical', command=self.checklist_tree.yview)
        self.checklist_tree.configure(yscrollcommand=scroll_right.set)

        self.checklist_tree.pack(side='top', fill='both', expand=True, padx=5)

        # Frame per i pulsanti
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(fill='x', padx=5, pady=10)

        ttk.Button(btn_frame, text=self.lang.get('Save_Verification', 'Salva Verifica'),
                   command=self._on_save_verification, style='Accent.TButton').pack(side='left', padx=5)

        ttk.Button(btn_frame, text=self.lang.get('Clear', 'Pulisci'),
                   command=self._clear_verification_form).pack(side='left', padx=5)

        scroll_right.pack(side='right', fill='y')

        self.checklist_tree.bind('<Button-1>', self._on_checklist_click)
        self.checklist_tree.bind('<Double-Button-1>', self._on_checklist_double_click)

    def _on_product_select(self, event):
        """Gestisce la selezione di un prodotto da verificare"""
        selection = self.products_tree.selection()
        if not selection:
            return

        item = self.products_tree.item(selection[0])
        self._current_must_check_id = item['values'][0]

        # Reset status label code quando cambia prodotto
        self.label_code_status_label.config(text='')
        self._current_label_code_id = None  # Reset IDLabelCode

        # Carica la checklist per questo prodotto
        self._load_checklist()

    def _load_checklist(self):
        """Carica la checklist per il prodotto selezionato"""
        for item in self.checklist_tree.get_children():
            self.checklist_tree.delete(item)

        self._check_items = []

        print(f"=== DEBUG CHECKLIST LOADING ===")
        print(f"Current must check ID: {self._current_must_check_id}")

        # CARICA SEMPRE I TASK GENERICI (anche senza prodotto selezionato)
        generic_tasks = self.db.fetch_generic_check_tasks()
        print(f"Generic tasks type: {type(generic_tasks)}")
        print(f"Generic tasks length: {len(generic_tasks)}")

        for task in generic_tasks:
            print(f"Generic task: {task.ItemToCheck}")
            item_id = self.checklist_tree.insert('', 'end', text='☐', values=(
                False,
                task.PriodicalProductCheckListId,
                task.ItemToCheck,
                '✓' if task.Doc else ''
            ))
            self._check_items.append({
                'tree_id': item_id,
                'task_id': task.PriodicalProductCheckListId,
                'checked': False,
                'doc': task.Doc
            })

        # CARICA TASK SPECIFICI solo se c'è un prodotto selezionato
        if self._current_must_check_id:
            specific_tasks = self.db.fetch_specific_check_tasks(self._current_must_check_id)
            print(f"Found {len(specific_tasks)} specific tasks")
            print(f"Specific tasks length: {len(specific_tasks)}")

            for task in specific_tasks:
                print(f"Specific task: {task.ItemToCheck}")
                item_id = self.checklist_tree.insert('', 'end', text='☐', values=(
                    False,
                    task.PriodicalProductCheckListId,
                    task.ItemToCheck,
                    '✓' if task.Doc else ''
                ))
                self._check_items.append({
                    'tree_id': item_id,
                    'task_id': task.PriodicalProductCheckListId,
                    'checked': False,
                    'doc': task.Doc
                })

        print(f"Total checklist items: {len(self._check_items)}")
        print(f"Tree children count: {len(self.checklist_tree.get_children())}")

    def _on_checklist_click(self, event):
        """Gestisce il click sulla checklist (toggle checkbox)"""
        region = self.checklist_tree.identify('region', event.x, event.y)
        if region == 'tree':
            item_id = self.checklist_tree.identify_row(event.y)
            if item_id:
                # Toggle checkbox
                for check_item in self._check_items:
                    if check_item['tree_id'] == item_id:
                        check_item['checked'] = not check_item['checked']
                        new_symbol = '☑' if check_item['checked'] else '☐'
                        self.checklist_tree.item(item_id, text=new_symbol)
                        break

    def _on_checklist_double_click(self, event):
        """Apre il documento se presente"""
        item_id = self.checklist_tree.identify_row(event.y)
        if item_id:
            for check_item in self._check_items:
                if check_item['tree_id'] == item_id and check_item['doc']:
                    task = self.db.fetch_check_task_by_id(check_item['task_id'])
                    if task and task.Doc:
                        self._open_document(task.Doc)
                    break

    def _open_document(self, doc_data):
        """Apre un documento binario"""
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(doc_data)
                tmp_path = tmp_file.name

            if os.name == 'nt':
                os.startfile(tmp_path)
            elif os.name == 'posix':
                subprocess.call(['xdg-open', tmp_path])
        except Exception as e:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 f"{self.lang.get('open_error', 'Errore apertura')}: {str(e)}")

    def _on_save_verification(self):
        """Salva la verifica completata"""
        # Validazione
        if not self._current_must_check_id:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_product_to_check', 'Selezionare un prodotto da verificare'),
                parent=self
            )
            return

        if not self.label_code_var.get().strip():
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('enter_label_code', 'Inserire il codice label'),
                parent=self
            )
            return

        if not self._current_label_code_id:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('Label_Not_Found', 'Etichetta non trovata o non relativa al prodotto selezionato'),
                parent=self
            )
            return

        if not self.result_var.get():
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_result', 'Selezionare il risultato (PASS/FAIL)'),
                parent=self
            )
            return

        # Validazione commento obbligatorio per FAIL
        comments = self.comment_text.get('1.0', 'end-1c').strip()
        if self.result_var.get() == 'FAIL' and not comments:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('Comments_Required_For_Fail', 'Il campo commenti è obbligatorio per risultato FAIL'),
                parent=self
            )
            return

        # Verifica che tutti i checkbox siano spuntati
        all_checked = all(item['checked'] for item in self._check_items)
        if not all_checked:
            if not messagebox.askyesno(
                    self.lang.get('confirm', 'Conferma'),
                    self.lang.get('not_all_checked', 'Non tutti gli item sono stati verificati. Continuare?'),
                    parent=self
            ):
                return

        # Salva nel database usando l'IDLabelCode
        success = self.db.save_product_verification(
            must_check_id=self._current_must_check_id,
            user_name=self.user_name,
            label_code_id=self._current_label_code_id,  # Usa l'IDLabelCode invece del codice testo
            status=self.result_var.get(),
            comments=comments if self.result_var.get() == 'FAIL' else None
        )

        if success:
            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                self.lang.get('verification_saved', 'Verifica salvata con successo'),
                parent=self
            )
            # Ricarica la lista prodotti
            self._load_products_to_check()
            # Pulisci il form
            self._clear_verification_form()
        else:
            # Gestione errori specifici
            error_msg = self.lang.get('save_error', 'Errore durante il salvataggio')
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"{error_msg}: {self.db.last_error_details}",
                parent=self
            )

    def _clear_verification_form(self):
        """Pulisce il form di verifica"""
        self._current_must_check_id = None
        self._current_label_code_id = None
        self.label_code_var.set('')
        self.label_code_status_label.config(text='')
        self.result_var.set('')
        self.comment_text.delete('1.0', 'end')
        self._on_result_changed()
        self._check_items = []
        for item in self.checklist_tree.get_children():
            self.checklist_tree.delete(item)

class VerificationReportsWindow(tk.Toplevel):
    """Finestra per i rapporti di verifica e statistiche"""

    def __init__(self, parent, db_handler, lang_manager, user_id):
        super().__init__(parent)
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_id

        self.title(self.lang.get('verification_reports_title', 'Rapporti Verifiche'))
        self.geometry('1300x850')
        self.transient(parent)

        self._build_ui()
        self._load_users()
        self._update_stats()

    def _build_ui(self):
        logger.info("Building UI")
        # Main container
        main_frame = ttk.Frame(self)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # --- Filters Frame ---
        filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get('filters', 'Filtri'))
        filter_frame.pack(fill='x', pady=(0, 10))

        # CheckUser Filter
        ttk.Label(filter_frame, text=self.lang.get('check_user_label', "Utente Controllo:")).pack(side='left', padx=5)
        self.user_var = tk.StringVar()
        self.user_combo = ttk.Combobox(filter_frame, textvariable=self.user_var, state='readonly', width=30)
        self.user_combo.pack(side='left', padx=5)
        self.user_combo.bind("<<ComboboxSelected>>", lambda e: self._on_search())
        # LabelCode Filter
        ttk.Label(filter_frame, text=self.lang.get('label_code_label', "Codice Etichetta:")).pack(side='left', padx=5)
        self.label_code_var = tk.StringVar(value=None) # Default from request
        self.label_code_entry = ttk.Entry(filter_frame, textvariable=self.label_code_var, width=20)
        self.label_code_entry.pack(side='left', padx=5)
        self.label_code_entry.bind("<KeyRelease>", lambda e: self._on_search())

        # IsAnalized Filter
        self.is_analyzed_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(filter_frame, text=self.lang.get('is_analyzed_label', "Analizzato"), variable=self.is_analyzed_var, command=self._on_search).pack(side='left', padx=10)

        # Buttons
        ttk.Button(filter_frame, text=self.lang.get('search', 'Cerca'), command=self._on_search).pack(side='left', padx=10)
        ttk.Button(filter_frame, text=self.lang.get('export_excel', 'Export Excel'), command=self._on_export).pack(side='left', padx=10)
        
        # Mark as Analyzed Button (Right aligned)
        #ttk.Button(filter_frame, text=self.lang.get('mark_as_analyzed_btn', "Segna come Analizzato"), command=self._on_mark_analyzed).pack(side='right', padx=10)

        # --- Results Treeview ---
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True)

        columns = ('CheckUser', 'LabelCod', 'ProductCode', 'ResultRepair', 'Minute', 'TimeDefectAfterCheck', 
                   'CodRiferimento', 'ComponentType', 'ComponentCode', 'Defect', 'BoxCode', 'ShipmentStatus')
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        # Configure columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        self.tree.column('CheckUser', width=150)
        self.tree.column('Defect', width=200)
        self.tree.column('TimeDefectAfterCheck', width=150)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')

        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        # --- Statistics Panel ---
        logger.info("Building Statistics Panel")
        stats_frame = ttk.LabelFrame(main_frame, text=self.lang.get('statistics', 'Statistiche Utenti (Top 1000)'))
        stats_frame.pack(fill='x', pady=(10, 0))

        self.stats_tree = ttk.Treeview(stats_frame, columns=('User', 'Count', 'Pass', 'Fail', 'FailPercent', 'Not_True', 'Not_True_Percent'), show='headings', height=5)
        self.stats_tree.heading('User', text=self.lang.get('stats_user', 'Utente'))
        self.stats_tree.heading('Count', text=self.lang.get('stats_total', 'Totale'))
        self.stats_tree.heading('Pass', text='Pass')
        self.stats_tree.heading('Fail', text='Fail')
        self.stats_tree.heading('FailPercent', text='Fail %')
        self.stats_tree.heading('Not_True', text='Not True')
        self.stats_tree.heading('Not_True_Percent', text='Not True %')
        
        self.stats_tree.column('User', width=200)
        self.stats_tree.column('Count', width=80, anchor='center')
        self.stats_tree.column('Pass', width=80, anchor='center')
        self.stats_tree.column('Fail', width=80, anchor='center')
        self.stats_tree.column('FailPercent', width=80, anchor='center')
        self.stats_tree.column('Not_True', width=80, anchor='center')
        self.stats_tree.column('Not_True_Percent', width=80, anchor='center')
        
        self.stats_tree.tag_configure('red_bold', foreground='red', font=('Segoe UI', 9, 'bold'))
        self.stats_tree.pack(fill='x', padx=5, pady=5)

    def _load_users(self):
        """Carica la lista utenti per il filtro"""
        sql = """
        SELECT Distinct         
            UPPER(e.EmployeeSurname + ' ' +e.EmployeeName) AS FullName              
        FROM employee.dbo.employees e 
        INNER JOIN resetservices.dbo.tbuserkey u ON e.employeeid = u.idanga
        inner join employee.dbo.employeehirehistory h on e.employeeid = h.employeeid and h.EndWorkDate is null and h.employeerid =2
        inner join Traceability_rs.dbo.PeriodicalProductCheckLogs pl on u.NomeUser collate database_default =pl.UserCheck
        order by  UPPER(e.EmployeeSurname + ' ' +e.EmployeeName)  ;
        """
        try:
            cursor = self.db.conn.cursor()
            try:
                cursor.execute(sql)
                users = [row[0] for row in cursor.fetchall()]
                self.user_combo['values'] = [''] + users
            finally:
                cursor.close()
        except Exception as e:
            messagebox.showerror("Errore", f"Errore caricamento utenti: {e}")

    def _on_search(self):
        """Esegue la query principale"""
        logger.info("Esecuzione query principale lancio verifica fail dopo verifica dei capiturno.")
        is_analyzed = 1 if self.is_analyzed_var.get() else 0
        label_code = self.label_code_var.get().strip() or None
        check_user_filter = self.user_var.get().strip()

        sql = """
        DECLARE @IsAnalized bit = ?;
        DECLARE @LabelCode as nvarchar(230) = ?;
        DECLARE @CheckUserFilter as nvarchar(255) = ?;

        WITH EmployeeMapping AS (
            SELECT 
                u.nomeuser,
                UPPER( e.EmployeeSurname + ' ' +e.EmployeeName) AS FullName,
                ROW_NUMBER() OVER (PARTITION BY u.nomeuser ORDER BY e.employeeid) AS rn
            FROM employee.dbo.employees e 
            INNER JOIN resetservices.dbo.tbuserkey u ON e.employeeid = u.idanga
        ),
        ComponentInfo AS (
            SELECT 
                ProductRiferiments.CodRiferimento,
                ProductComponentsErp.IDProduct,
                ParentPhases.ParentPhaseName,
                Components.ComponentCode,
                ROW_NUMBER() OVER (PARTITION BY ProductRiferiments.CodRiferimento, ProductComponentsErp.IDProduct 
                                  ORDER BY (SELECT NULL)) AS rn
            FROM ProductRiferiments 
            INNER JOIN ProductComponentsErp ON ProductComponentsErp.IDProductCompErp = ProductRiferiments.IDProductCompErp
            INNER JOIN ParentPhases ON ParentPhases.IDParentPhase = ProductRiferiments.IDParentPhase
            LEFT JOIN Components ON Components.IDComponent = ProductComponentsErp.IDComponent
        )
        SELECT * from (
        SELECT distinct   
            EmployeeMapping.FullName AS CheckUser,
            labelcodes.LabelCod,
            Products.ProductCode,                
            CASE WHEN ScanDefects.IsPass = 1 THEN 'REPAIRED' ELSE 'SCRAP' END AS ResultRepair,  
            DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) as [Minute], 
            CASE 
                WHEN DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) < 60 
                    THEN CAST(DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' MINUTE'
                WHEN DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) >= 24 
                    THEN CAST(DATEDIFF(DAY, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' DAYS'
                ELSE CAST(DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' HOURS'
            END AS TimeDefectAfterCheck,       
            
            Riferiments.CodRiferimento,
            ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') AS ComponentType,  
            ISNULL(ComponentInfo.ComponentCode, '#N/D') AS ComponentCode,
            Defects.DefectNameRO AS Defect,        
            IIF(CAST(Boxes.BoxCode AS NVARCHAR(12)) is not null ,'IN BOX', 'NOT IN A BOX') AS BoxCode,
            IIF(PackingLists.CodePack IS NULL, 'NOT SHIPPED YET', 'SHIPPED ALREADY') AS ShipmentStatus
            
        FROM ScanDefects 
        INNER JOIN ScanDefectDetails ON ScanDefects.IDScanDefect = ScanDefectDetails.IDScanDefect
        INNER JOIN DefectsRiferiments ON DefectsRiferiments.IDScanDefectDet = ScanDefectDetails.IDScanDefectDet
        INNER JOIN Riferiments ON Riferiments.IDDibaRiferimento = DefectsRiferiments.IDDibaRiferimento
        INNER JOIN Defects ON ScanDefectDetails.IDDefect = Defects.IDDefect
        INNER JOIN Scannings ON Scannings.IDScan = ScanDefects.IDScan
        INNER JOIN OrderPhases ON OrderPhases.IDOrderPhase = Scannings.IDOrderPhase
        INNER JOIN Orders ON OrderPhases.IDOrder = Orders.IDOrder
        INNER JOIN Products ON Products.IDProduct = Orders.IDProduct
        INNER JOIN Phases ON OrderPhases.IDPhase = Phases.IDPhase
        INNER JOIN Clients ON Clients.IDClient = Products.IDClient
        INNER JOIN Boards ON Scannings.IDBoard = Boards.IDBoard
        INNER JOIN Teams ON Teams.IDTeam = ScanDefects.IdTeam
        INNER JOIN WorkLines ON WorkLines.IDWorkLine = Teams.IDWorkLine
        INNER JOIN LabelCodes ON Boards.IDBoard = LabelCodes.IDBoard
        INNER JOIN PeriodicalProductCheckLogs PC ON LabelCodes.IDLabelCode = PC.IDLabelCode and isnull(pc.isanalized,0) = @IsAnalized 

        LEFT JOIN EmployeeMapping ON EmployeeMapping.nomeuser COLLATE database_default = PC.UserCheck 
            AND EmployeeMapping.rn = 1
        LEFT JOIN ComponentInfo ON ComponentInfo.CodRiferimento = Riferiments.CodRiferimento 
            AND ComponentInfo.IDProduct = Orders.IDProduct 
            AND ComponentInfo.rn = 1

        LEFT JOIN Areas ON Areas.IDArea = ScanDefectDetails.IDArea
        LEFT JOIN BoxDetails ON BoxDetails.IDBoard = Boards.IDBoard
        LEFT JOIN Boxes ON Boxes.IDBox = BoxDetails.IDBox
        LEFT JOIN BoxPKs ON BoxPKs.IDBoxPK = Boxes.IDBoxPK
        LEFT JOIN PalletPKs ON PalletPKs.IDPalletPK = BoxPKs.IDPalletPK
        LEFT JOIN PackingLists ON PackingLists.IDPackingList = PalletPKs.IDPackingList

        WHERE 
            Phases.IDPhase IN (102, 103, 107)
            AND PC.Status = 'PASS'
            AND ScanDefects.StopTime > PC.CheckTime
            and ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') = 'PTHM'
            AND NOT defects.DefectNameRO IN ('Schimbare pin fixture test','Componenta iesita din tolerante')
            and labelcodes.LabelCod = iif(@LabelCode is null,labelcodes.LabelCod,@LabelCode)
        ) as H 
        WHERE (@CheckUserFilter IS NULL OR @CheckUserFilter = '' OR H.CheckUser = @CheckUserFilter)
        Order By [Minute];
        """
        
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            cursor = self.db.conn.cursor()
            try:
                cursor.execute(sql, (is_analyzed, label_code, check_user_filter))
                rows = cursor.fetchall()
                
                for row in rows:
                    self.tree.insert('', 'end', values=[str(x) if x is not None else '' for x in row])
            finally:
                cursor.close()
                
        except Exception as e:
            messagebox.showerror("Errore Query", f"Errore esecuzione ricerca: {e}")

    def _on_export(self):
        """Esporta i dati in Excel"""
        if not self.tree.get_children():
            messagebox.showwarning("Attenzione", "Nessun dato da esportare")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Verification Report"

            # Headers
            columns = self.tree['columns']
            ws.append(columns)

            # Data
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                ws.append(values)

            wb.save(file_path)
            messagebox.showinfo("Successo", "Esportazione completata")
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Errore Export", f"Errore durante l'esportazione: {e}")

    def _update_stats(self):
        """Calcola e mostra le statistiche"""
        sql = """
        WITH EmployeeMapping AS (
             SELECT 
                 u.nomeuser,
                 UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS FullName,
                 ROW_NUMBER() OVER (PARTITION BY u.nomeuser ORDER BY e.employeeid) AS rn
             FROM employee.dbo.employees e 
             INNER JOIN resetservices.dbo.tbuserkey u ON e.employeeid = u.idanga
        ),
        Stats AS (
            SELECT 
                EmployeeMapping.FullName,
                COUNT(*) as CheckCount,
                SUM(CASE WHEN PC.Status = 'PASS' THEN 1 ELSE 0 END) as PassCount,
                SUM(CASE WHEN PC.Status <> 'PASS' OR PC.Status IS NULL THEN 1 ELSE 0 END) as FailCount
            FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLogs] PC
            LEFT JOIN EmployeeMapping ON EmployeeMapping.nomeuser COLLATE database_default = PC.UserCheck 
                AND EmployeeMapping.rn = 1
            GROUP BY EmployeeMapping.FullName
        ),
        ComponentInfo AS (
            SELECT 
                ProductRiferiments.CodRiferimento,
                ProductComponentsErp.IDProduct,
                ParentPhases.ParentPhaseName,
                Components.ComponentCode,
                ROW_NUMBER() OVER (PARTITION BY ProductRiferiments.CodRiferimento, ProductComponentsErp.IDProduct 
                                  ORDER BY (SELECT NULL)) AS rn
            FROM ProductRiferiments 
            INNER JOIN ProductComponentsErp ON ProductComponentsErp.IDProductCompErp = ProductRiferiments.IDProductCompErp
            INNER JOIN ParentPhases ON ParentPhases.IDParentPhase = ProductRiferiments.IDParentPhase
            LEFT JOIN Components ON Components.IDComponent = ProductComponentsErp.IDComponent
        ),
        DefectsStats AS (
            SELECT 
                EmployeeMapping.FullName AS CheckUser,
                COUNT(*) AS TotalDefects
            FROM ScanDefects 
            INNER JOIN ScanDefectDetails ON ScanDefects.IDScanDefect = ScanDefectDetails.IDScanDefect
            INNER JOIN DefectsRiferiments ON DefectsRiferiments.IDScanDefectDet = ScanDefectDetails.IDScanDefectDet
            INNER JOIN Riferiments ON Riferiments.IDDibaRiferimento = DefectsRiferiments.IDDibaRiferimento
            INNER JOIN Defects ON ScanDefectDetails.IDDefect = Defects.IDDefect
            INNER JOIN Scannings ON Scannings.IDScan = ScanDefects.IDScan
            INNER JOIN OrderPhases ON OrderPhases.IDOrderPhase = Scannings.IDOrderPhase
            INNER JOIN Orders ON OrderPhases.IDOrder = Orders.IDOrder
            INNER JOIN Products ON Products.IDProduct = Orders.IDProduct
            INNER JOIN Phases ON OrderPhases.IDPhase = Phases.IDPhase
            INNER JOIN Clients ON Clients.IDClient = Products.IDClient
            INNER JOIN Boards ON Scannings.IDBoard = Boards.IDBoard
            INNER JOIN Teams ON Teams.IDTeam = ScanDefects.IdTeam
            INNER JOIN WorkLines ON WorkLines.IDWorkLine = Teams.IDWorkLine
            INNER JOIN LabelCodes ON Boards.IDBoard = LabelCodes.IDBoard
            INNER JOIN PeriodicalProductCheckLogs PC ON LabelCodes.IDLabelCode = PC.IDLabelCode 
            
            LEFT JOIN EmployeeMapping ON EmployeeMapping.nomeuser COLLATE database_default = PC.UserCheck 
                AND EmployeeMapping.rn = 1
            LEFT JOIN ComponentInfo ON ComponentInfo.CodRiferimento = Riferiments.CodRiferimento 
                AND ComponentInfo.IDProduct = Orders.IDProduct 
                AND ComponentInfo.rn = 1

            LEFT JOIN Areas ON Areas.IDArea = ScanDefectDetails.IDArea
            LEFT JOIN BoxDetails ON BoxDetails.IDBoard = Boards.IDBoard
            LEFT JOIN Boxes ON Boxes.IDBox = BoxDetails.IDBox
            LEFT JOIN BoxPKs ON BoxPKs.IDBoxPK = Boxes.IDBoxPK
            LEFT JOIN PalletPKs ON PalletPKs.IDPalletPK = BoxPKs.IDPalletPK
            LEFT JOIN PackingLists ON PackingLists.IDPackingList = PalletPKs.IDPackingList

            WHERE 
                Phases.IDPhase IN (102, 103, 107)
                AND PC.Status = 'PASS'
                AND ScanDefects.StopTime > PC.CheckTime
                AND ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') = 'PTHM'
                AND NOT defects.DefectNameRO IN ('Schimbare pin fixture test','Componenta iesita din tolerante')
            
            GROUP BY EmployeeMapping.FullName
        )
        SELECT TOP (1000)
            S.FullName,
            S.CheckCount,
            S.PassCount,
            S.FailCount,
            ISNULL(D.TotalDefects, 0) as TotalDefects
        FROM Stats S
        LEFT JOIN DefectsStats D ON S.FullName = D.CheckUser
        ORDER BY S.CheckCount DESC
        """
        try:
            # Clear stats tree
            for item in self.stats_tree.get_children():
                self.stats_tree.delete(item)

            cursor = self.db.conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
                
                for row in rows:
                    total = row[1]
                    pass_count = row[2]
                    fail_count = row[3]
                    total_defects = row[4]
                    
                    fail_percent = (fail_count / total * 100) if total > 0 else 0
                    defect_percent = (total_defects / pass_count * 100) if pass_count > 0 else 0
                    
                    tags = ('red_bold',) if total_defects > 0 else ()
                    
                    self.stats_tree.insert('', 'end', values=(
                        row[0] or 'Unknown', 
                        total, 
                        pass_count, 
                        fail_count, 
                        f"{fail_percent:.2f}%",
                        total_defects,
                        f"{defect_percent:.2f}%"
                    ), tags=tags)
            finally:
                cursor.close()
                
        except Exception as e:
            print(f"Errore statistiche: {e}")

    def _on_mark_analyzed(self):
        """Marca i record visualizzati come analizzati"""
        pass
        # if not self.tree.get_children():
        #     return

        # if not messagebox.askyesno("Conferma", "Vuoi marcare tutti i record visualizzati come Analizzati?"):
        #     return

        # label_codes = []
        # for item in self.tree.get_children():
        #     vals = self.tree.item(item)['values']
        #     if len(vals) > 1:
        #         label_codes.append(vals[1])
        
        # if not label_codes:
        #     return

        # try:
        #     cursor = self.db.conn.cursor()
        #     try:
        #         placeholders = ','.join(['?'] * len(label_codes))
        #         sql_update = f"""
        #         UPDATE PC
        #         SET IsAnalized = 1
        #         FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLogs] PC
        #         INNER JOIN LabelCodes LC ON PC.IDLabelCode = LC.IDLabelCode
        #         WHERE LC.LabelCod IN ({placeholders})
        #         """
                
        #         cursor.execute(sql_update, label_codes)
        #         self.db.conn.commit()
                
        #         messagebox.showinfo("Successo", f"Aggiornati {cursor.rowcount} record.")
        #     finally:
        #         cursor.close()

        #     self._on_search() # Refresh
            
        # except Exception as e:
        #     self.db.conn.rollback()
        #     messagebox.showerror("Errore Update", f"Errore durante l'aggiornamento: {e}")

def check_and_notify_verification_discrepancies(db_handler):
    """
    Esegue una verifica in background per identificare discrepanze tra controlli operatore e difetti di produzione.
    Se vengono trovati dati, invia un'email di notifica con allegato Excel.
    """
    import utils
    import logging
    import os
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    logger = logging.getLogger('TraceabilityRS')
    
    # SQL Query fornita dall'utente
    logger.info('Background check: Starting query execution...')
    query = """
        WITH EmployeeMapping AS (
            SELECT 
                u.nomeuser,
                UPPER( e.EmployeeSurname + ' ' +e.EmployeeName) AS FullName,
                ROW_NUMBER() OVER (PARTITION BY u.nomeuser ORDER BY e.employeeid) AS rn
            FROM employee.dbo.employees e 
            INNER JOIN resetservices.dbo.tbuserkey u ON e.employeeid = u.idanga
        ),
        ComponentInfo AS (
            SELECT 
                ProductRiferiments.CodRiferimento,
                ProductComponentsErp.IDProduct,
                ParentPhases.ParentPhaseName,
                Components.ComponentCode,
                ROW_NUMBER() OVER (PARTITION BY ProductRiferiments.CodRiferimento, ProductComponentsErp.IDProduct 
                                  ORDER BY (SELECT NULL)) AS rn
            FROM ProductRiferiments 
            INNER JOIN ProductComponentsErp ON ProductComponentsErp.IDProductCompErp = ProductRiferiments.IDProductCompErp
            INNER JOIN ParentPhases ON ParentPhases.IDParentPhase = ProductRiferiments.IDParentPhase
            LEFT JOIN Components ON Components.IDComponent = ProductComponentsErp.IDComponent
        )
        SELECT * from (
        SELECT distinct   
            PC.PeriodicalProductCheckLogId, -- Aggiunto per poter aggiornare il record
            EmployeeMapping.FullName AS CheckUser,
            labelcodes.LabelCod,
            Products.ProductCode,                
            CASE WHEN ScanDefects.IsPass = 1 THEN 'REPAIRED' ELSE 'SCRAP' END AS ResultRepair,  
            DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) as [Minute], 
            CASE 
                WHEN DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) < 60 
                    THEN CAST(DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' MINUTE'
                WHEN DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) >= 24 
                    THEN CAST(DATEDIFF(DAY, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' DAYS'
                ELSE CAST(DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' HOURS'
            END AS TimeDefectAfterCheck,       
            
            Riferiments.CodRiferimento,
            ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') AS ComponentType,  
            ISNULL(ComponentInfo.ComponentCode, '#N/D') AS ComponentCode,
            Defects.DefectNameRO AS Defect,        
            IIF(CAST(Boxes.BoxCode AS NVARCHAR(12)) is not null ,'IN BOX', 'NOT IN A BOX') AS BoxCode,
            IIF(PackingLists.CodePack IS NULL, 'NOT SHIPPED YET', 'SHIPPED ALREADY') AS ShipmentStatus
            
        FROM ScanDefects 
        INNER JOIN Traceability_rs.dbo.ScanDefectDetails ON ScanDefects.IDScanDefect = ScanDefectDetails.IDScanDefect
        INNER JOIN Traceability_rs.dbo.DefectsRiferiments ON DefectsRiferiments.IDScanDefectDet = ScanDefectDetails.IDScanDefectDet
        INNER JOIN Traceability_rs.dbo.Riferiments ON Riferiments.IDDibaRiferimento = DefectsRiferiments.IDDibaRiferimento
        INNER JOIN Traceability_rs.dbo.Defects ON ScanDefectDetails.IDDefect = Defects.IDDefect
        INNER JOIN Traceability_rs.dbo.Scannings ON Scannings.IDScan = ScanDefects.IDScan
        INNER JOIN Traceability_rs.dbo.OrderPhases ON OrderPhases.IDOrderPhase = Scannings.IDOrderPhase
        INNER JOIN Traceability_rs.dbo.Orders ON OrderPhases.IDOrder = Orders.IDOrder
        INNER JOIN Traceability_rs.dbo.Products ON Products.IDProduct = Orders.IDProduct
        INNER JOIN Traceability_rs.dbo.Phases ON OrderPhases.IDPhase = Phases.IDPhase
        INNER JOIN Traceability_rs.dbo.Clients ON Clients.IDClient = Products.IDClient
        INNER JOIN Traceability_rs.dbo.Boards ON Scannings.IDBoard = Boards.IDBoard
        INNER JOIN Traceability_rs.dbo.Teams ON Teams.IDTeam = ScanDefects.IdTeam
        INNER JOIN Traceability_rs.dbo.WorkLines ON WorkLines.IDWorkLine = Teams.IDWorkLine
        INNER JOIN Traceability_rs.dbo.LabelCodes ON Boards.IDBoard = LabelCodes.IDBoard
        INNER JOIN Traceability_rs.dbo.PeriodicalProductCheckLogs PC ON LabelCodes.IDLabelCode = PC.IDLabelCode and isnull(pc.isanalized,0) = 0

        LEFT JOIN EmployeeMapping ON EmployeeMapping.nomeuser COLLATE database_default = PC.UserCheck 
            AND EmployeeMapping.rn = 1
        LEFT JOIN ComponentInfo ON ComponentInfo.CodRiferimento = Riferiments.CodRiferimento 
            AND ComponentInfo.IDProduct = Orders.IDProduct 
            AND ComponentInfo.rn = 1

        LEFT JOIN Areas ON Areas.IDArea = ScanDefectDetails.IDArea
        LEFT JOIN BoxDetails ON BoxDetails.IDBoard = Boards.IDBoard
        LEFT JOIN Boxes ON Boxes.IDBox = BoxDetails.IDBox
        LEFT JOIN BoxPKs ON BoxPKs.IDBoxPK = Boxes.IDBoxPK
        LEFT JOIN PalletPKs ON PalletPKs.IDPalletPK = BoxPKs.IDPalletPK
        LEFT JOIN PackingLists ON PackingLists.IDPackingList = PalletPKs.IDPackingList

        WHERE 
            Phases.IDPhase IN (102, 103, 107)
            AND PC.Status = 'PASS'
            AND ScanDefects.StopTime > PC.CheckTime
            and ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') = 'PTHM'
            AND NOT defects.DefectNameRO IN ('Schimbare pin fixture test','Componenta iesita din tolerante')          
        ) as H 
        Order By [checkuser],[Minute];
    """
    
    try:
        results = db_handler.fetch_all(query)
        logger.info(f'Background check: Found {len(results) if results else 0} discrepancies.')
        
        if not results:
            logger.info('Background check: No discrepancies found.')
            return

        # Se ci sono risultati, prepara l'email
        logger.info(f'Background check: Found {len(results)} discrepancies. Sending notification...')
        
        # Recupera destinatari
        recipients = utils.get_email_recipients(db_handler.conn, 'Sys_mail_verifiche')
        if not recipients:
            logger.warning('Background check: No recipients configured for Sys_mail_verifiche.')
            return

        # Costruisci corpo email
        # Rimuoviamo la colonna PeriodicalProductCheckLogId dalla visualizzazione
        display_results = []
        ids_to_update = []
        
        for row in results:
            # Assumiamo che PeriodicalProductCheckLogId sia la prima colonna (indice 0)
            ids_to_update.append(row[0])
            display_results.append(row[1:]) # Tutte le colonne tranne la prima
        
        # Genera file Excel
        excel_file = None
        try:
            # Crea workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Discrepancies Report"
            
            # Intestazioni (salta la prima colonna ID)
            if results and hasattr(results[0], 'cursor_description'):
                headers = [col[0] for col in results[0].cursor_description][1:]
            else:
                # Fallback headers
                headers = ['CheckUser', 'LabelCod', 'ProductCode', 'ResultRepair', 'Minute', 
                          'TimeDefectAfterCheck', 'CodRiferimento', 'ComponentType', 'ComponentCode',
                          'Defect', 'BoxCode', 'ShipmentStatus']
            
            # Scrivi header
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Scrivi dati
            for row_idx, row_data in enumerate(display_results, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size colonne
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salva in file temporaneo
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(temp_dir, f"Verification_Discrepancies_{timestamp}.xlsx")
            wb.save(excel_file)
            logger.info(f'Background check: Excel file created at {excel_file}')
            
        except Exception as ex:
            logger.error(f'Background check: Error creating Excel file: {ex}')
            excel_file = None
            
        # Costruisci HTML table per email body
        html_table = '<table border="1" style="border-collapse: collapse;">'
        # Header
        if results and hasattr(results[0], 'cursor_description'):
            headers = [col[0] for col in results[0].cursor_description][1:]
            html_table += '<tr>' + ''.join(f'<th style="padding: 5px; background-color: #f2f2f2;">{h}</th>' for h in headers) + '</tr>'
        
        for row in display_results:
            html_table += '<tr>' + ''.join(f'<td style="padding: 5px;">{str(cell)}</td>' for cell in row) + '</tr>'
        html_table += '</table>'

        email_body = f"""
        <html>
        <body>
            <p>Dear users,</p>
            <p>Discrepancies have been found between operator statements and production defects.</p>
            <p>It means that the hourly check control, done on one board, was not completed or was not completed correctly.</p>
            <p>It is important to note that this is not a production defect, but a discrepancy between the operator's statement and the production defect.</p>
            <p>Please check the production defects and the operator's statements to ensure that the production defects are correct.</p>
            <p>On top of that, ensure that the operator follows the correct procedure for the hourly check control.</p>
            <p>Note that the column 'TimeDefectAfterCheck' is the column that shows the time of the defect after the check. 
            Some times, the defect declaration has been posted very few minutes after the 'PASS' check response.</p>
            <p>Last point, please consider to enforce the correct procedure for the hourly check control, and the operator's preparation
            for the check.</p>
            <p>Below are the details:</p>
            <br>
            {html_table}
            <br>
            <p>The data can also be viewed in the Management Document program under the Reports section of the Verifications menu.</p>
            <p>Cordially,<br>Traceability System</p>
        </body>
        </html>
        """

        # Invia email con allegato Excel
        attachments_list = [excel_file] if excel_file and os.path.exists(excel_file) else []
        
        if utils.send_email(recipients, 'Notice on verify boards discrepancies', email_body, is_html=True, attachments=attachments_list):
            logger.info('Background check: Email sent successfully.')
            
            # Pulisci file temporaneo
            if excel_file and os.path.exists(excel_file):
                try:
                    os.remove(excel_file)
                    logger.info('Background check: Temporary Excel file removed.')
                except Exception as ex:
                    logger.warning(f'Background check: Could not remove temp file: {ex}')
            
            # Aggiorna database (IsAnalized = 1)
            if ids_to_update:
                try:
                    placeholders = ','.join(['?'] * len(ids_to_update))
                    update_query = f"UPDATE [Traceability_RS].[dbo].[PeriodicalProductCheckLogs] SET IsAnalized = 1 WHERE PeriodicalProductCheckLogId IN ({placeholders})"
                    
                    cursor = db_handler.conn.cursor()
                    cursor.execute(update_query, ids_to_update)
                    db_handler.conn.commit()
                    cursor.close()
                    logger.info(f'Background check: Updated {len(ids_to_update)} records as analyzed.')
                except Exception as ex:
                    logger.error(f'Background check: Error updating database: {ex}')
        else:
            logger.error('Background check: Error sending email.')
            # Pulisci file temporaneo anche in caso di errore
            if excel_file and os.path.exists(excel_file):
                try:
                    os.remove(excel_file)
                except:
                    pass

    except Exception as e:
        logger.error(f'Background check error: {e}')
