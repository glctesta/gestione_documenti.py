# -*- coding: utf-8 -*-
"""
material_consumption_report.py
Alloy consumption reporting for PTHM phase products.

Responsibilities:
  1. Daily scheduled check (08:05) — detects products with missing alloy data
     and sends a professional warning email (anti-duplication via settings table).
  2. On-demand report generation (PDF + Excel) triggered by the menu button.
  3. Query: single SQL joining Scannings → Products → ProductConsumptions.

Anti-duplication key: 'SentMissingAlloy_YYYYMMDD'  (24 chars, fits VARCHAR(30))
Recipients setting : 'Sys_missing_data_alloy'
"""
from __future__ import annotations

import base64
import datetime
import io
import logging
import os
import tempfile
from typing import Any

logger = logging.getLogger("TraceabilityRS")

# ─── Time window helpers ───────────────────────────────────────────────────────

def _production_window(reference: datetime.date | None = None):
    """
    Returns (start, end) for the production day window:
      start = reference-1 @ 07:30:00
      end   = reference   @ 07:30:00
    reference defaults to today.
    """
    ref = reference or datetime.date.today()
    start = datetime.datetime.combine(ref - datetime.timedelta(days=1),
                                      datetime.time(7, 30, 0))
    end   = datetime.datetime.combine(ref, datetime.time(7, 30, 0))
    return start, end


# ─── SQL ──────────────────────────────────────────────────────────────────────

_Q_PTHM_CONSUMPTION = """
SELECT
    o.IDProduct,
    p.ProductCode,
    ISNULL(SUM(CASE WHEN s.IsPass = 1 THEN 1 ELSE 0 END), 0) AS QtyProcessed,
    pc.ProductConsumptionId,
    pc.MaterialConsumption,
    pc.MaterialConsumptionGR,
    CASE WHEN pc.ProductConsumptionId IS NULL THEN 1 ELSE 0 END AS MissingAlloyConsumption
FROM Traceability_RS.dbo.Scannings s
INNER JOIN Traceability_RS.dbo.OrderPhases op ON s.IDOrderPhase  = op.IDOrderPhase
INNER JOIN Traceability_RS.dbo.Orders      o  ON op.IDOrder      = o.IDOrder
INNER JOIN Traceability_RS.dbo.Phases      ph ON op.IDPhase      = ph.IDPhase
INNER JOIN Traceability_RS.dbo.Products    p  ON o.IDProduct     = p.IDProduct
LEFT  JOIN Traceability_RS.dbo.ProductConsumptions pc
       ON  pc.IdProduct          = o.IDProduct
      AND  pc.MaterialConsumption = 'Alloy'
      AND  pc.DateOut             IS NULL
WHERE ph.IDPhase         = 107
  AND s.ScanTimeFinish  >= ?
  AND s.ScanTimeFinish   < ?
GROUP BY
    o.IDProduct, p.ProductCode,
    pc.ProductConsumptionId, pc.MaterialConsumption, pc.MaterialConsumptionGR
ORDER BY p.ProductCode
"""

_Q_CLAIM_SLOT = """
INSERT INTO traceability_rs.dbo.settings (atribute, [value])
SELECT ?, ?
WHERE NOT EXISTS (
    SELECT 1 FROM traceability_rs.dbo.settings WHERE atribute = ?
)
"""

_Q_GET_RECIPIENTS = """
SELECT [value]
FROM   traceability_rs.dbo.settings
WHERE  atribute = ?
"""

# ─── Data helpers ──────────────────────────────────────────────────────────────

def _fetch_pthm_rows(conn, start: datetime.datetime, end: datetime.datetime) -> list[dict]:
    try:
        cur = conn.cursor()
        cur.execute(_Q_PTHM_CONSUMPTION, (start, end))
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"material_consumption_report _fetch_pthm_rows: {e}", exc_info=True)
        return []


def _claim_send_slot(conn, setting_key: str) -> bool:
    try:
        cur = conn.cursor()
        cur.execute(_Q_CLAIM_SLOT, (setting_key,
                                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    setting_key))
        claimed = cur.rowcount > 0
        conn.commit()
        return claimed
    except Exception as e:
        logger.error(f"material_consumption_report _claim_send_slot: {e}")
        return False


def _get_recipients(conn, key: str) -> list[str]:
    try:
        cur = conn.cursor()
        cur.execute(_Q_GET_RECIPIENTS, (key,))
        row = cur.fetchone()
        if row and row[0]:
            return [a.strip() for a in row[0].replace(';', ',').split(',') if a.strip()]
    except Exception as e:
        logger.error(f"material_consumption_report _get_recipients: {e}")
    return []


# ─── HTML email builder ────────────────────────────────────────────────────────

def _logo_base64() -> str:
    logo_path = os.path.join(os.path.dirname(__file__), 'Logo.png')
    if not os.path.isfile(logo_path):
        return ''
    try:
        with open(logo_path, 'rb') as f:
            return 'data:image/png;base64,' + base64.b64encode(f.read()).decode()
    except Exception:
        return ''


def _build_missing_alloy_email_html(
    missing_rows: list[dict],
    date_start: datetime.datetime,
    date_end: datetime.datetime,
) -> str:
    logo_uri = _logo_base64()
    logo_html = (f'<img src="{logo_uri}" alt="Logo" style="height:50px">'
                 if logo_uri else '')

    period = (f'{date_start.strftime("%d/%m/%Y %H:%M")} → '
              f'{date_end.strftime("%d/%m/%Y %H:%M")}')

    rows_html = ''.join(
        f'<tr style="background:{"#f9f9f9" if i % 2 == 0 else "#ffffff"}">'
        f'<td style="padding:8px 12px;border:1px solid #e0e0e0">{r["ProductCode"]}</td>'
        f'<td style="padding:8px 12px;border:1px solid #e0e0e0;text-align:center">'
        f'{r["QtyProcessed"]}</td>'
        f'</tr>'
        for i, r in enumerate(missing_rows)
    )

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f6f8;
             font-family:Arial,Helvetica,sans-serif">
<div style="max-width:760px;margin:24px auto;background:#fff;border-radius:10px;
            box-shadow:0 2px 12px rgba(0,0,0,.10);overflow:hidden">

  <!-- Header -->
  <div style="background:#1f3864;padding:18px 28px;display:flex;align-items:center">
    {logo_html}
    <div style="margin-left:16px">
      <div style="color:#fff;font-size:18px;font-weight:bold">
        ⚠️ Missing Alloy Consumption Data
      </div>
      <div style="color:#a8c4e0;font-size:12px">
        PTHM Phase — Production period: {period}
      </div>
    </div>
  </div>

  <!-- Body -->
  <div style="padding:24px 28px;color:#2c3e50;font-size:14px;line-height:1.6">
    <p>Dear Team,</p>
    <p>
      The system detected that <strong>{len(missing_rows)} product code(s)</strong>
      processed in the <strong>PTHM phase</strong> during the previous production day
      have <strong>no active alloy consumption data</strong> configured in the product
      consumption master data.
    </p>
    <p>
      In order to calculate alloy consumption correctly, it is <strong>imperative</strong>
      that all products in the list below have a valid alloy consumption value configured.
    </p>
    <div style="background:#fff3cd;border-left:4px solid #f39c12;
                padding:12px 16px;border-radius:4px;margin:16px 0">
      <strong>⚖️ How to measure alloy consumption:</strong><br>
      Weigh one <em>unsoldered PCB</em>, then weigh the <em>same PCB immediately after
      the Wave soldering process</em>. Record the weight difference (in grams) in
      the system using the <em>Material Consumption → Data Management</em> function.
    </div>

    <p><strong>The following product codes are currently missing the required data:</strong></p>

    <table style="border-collapse:collapse;width:100%;font-size:13px">
      <tr style="background:#1f3864;color:#fff">
        <th style="padding:8px 12px;text-align:left;border:1px solid #16304f">
          Product Code
        </th>
        <th style="padding:8px 12px;text-align:center;border:1px solid #16304f">
          Qty Processed
        </th>
      </tr>
      {rows_html}
    </table>

    <p style="margin-top:20px">
      Please ensure these values are entered as soon as possible to allow accurate
      alloy consumption reporting.
    </p>
    <p>Thank you for your cooperation.</p>
    <p style="color:#7f8c8d;font-size:12px">
      This is an automated notification — do not reply to this message.<br>
      Generated by <strong>TraceabilityRS</strong> —
      {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}
    </p>
  </div>
</div>
</body>
</html>"""


# ─── PDF report builder ────────────────────────────────────────────────────────

def _build_pdf(
    report_rows: list[dict],
    total_day_gr: float,
    date_start: datetime.datetime,
    date_end: datetime.datetime,
    output_path: str,
) -> bool:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable, Image)
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        logger.error("reportlab not available — PDF skipped")
        return False

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm,   bottomMargin=15*mm,
    )
    styles = getSampleStyleSheet()
    story  = []

    # Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'Logo.png')
    if os.path.isfile(logo_path):
        try:
            story.append(Image(logo_path, width=80*mm, height=25*mm,
                               hAlign='LEFT'))
            story.append(Spacer(1, 4*mm))
        except Exception:
            pass

    # Title
    title_style = ParagraphStyle(
        'title', parent=styles['Title'],
        fontSize=15, textColor=colors.HexColor('#1f3864'),
        spaceAfter=2*mm,
    )
    sub_style = ParagraphStyle(
        'sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=4*mm,
    )
    story.append(Paragraph('Alloy Consumption Report — PTHM Phase', title_style))
    period = (f'Production window: {date_start.strftime("%d/%m/%Y %H:%M")} → '
              f'{date_end.strftime("%d/%m/%Y %H:%M")}')
    story.append(Paragraph(period, sub_style))
    story.append(HRFlowable(width='100%', color=colors.HexColor('#1f3864'),
                             thickness=1, spaceAfter=4*mm))

    # Table
    header = ['Product Code', 'Qty Processed', 'Unit GR (Alloy)',
              'Partial Total (gr)']
    data   = [header]
    for r in report_rows:
        data.append([
            r['ProductCode'],
            str(r['QtyProcessed']),
            f"{r['MaterialConsumptionGR']:.2f}" if r['MaterialConsumptionGR'] else '—',
            f"{r['PartialTotalGR']:.2f}",
        ])

    # Total row
    data.append(['', '', 'TOTAL DAY (gr):', f'{total_day_gr:.2f}'])

    col_w = [75*mm, 35*mm, 40*mm, 35*mm]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header
        ('BACKGROUND',  (0, 0), (-1,  0), colors.HexColor('#1f3864')),
        ('TEXTCOLOR',   (0, 0), (-1,  0), colors.white),
        ('FONTNAME',    (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1,  0), 9),
        ('ALIGN',       (0, 0), (-1,  0), 'CENTER'),
        # Data rows
        ('FONTSIZE',    (0, 1), (-1, -2), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2),
         [colors.HexColor('#f4f6f8'), colors.white]),
        # Total row
        ('BACKGROUND',  (0, -1), (-1, -1), colors.HexColor('#fff3cd')),
        ('FONTNAME',    (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, -1), (-1, -1), 9),
        ('ALIGN',       (2, -1), (3,  -1), 'RIGHT'),
        # Borders
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#e0e0e0')),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)

    # Footer
    story.append(Spacer(1, 8*mm))
    footer_style = ParagraphStyle(
        'footer', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#7f8c8d'),
    )
    story.append(Paragraph(
        f'Generated: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")} — '
        f'TraceabilityRS',
        footer_style
    ))

    doc.build(story)
    return True


# ─── Excel report builder ─────────────────────────────────────────────────────

def _build_excel(
    report_rows: list[dict],
    total_day_gr: float,
    date_start: datetime.datetime,
    date_end: datetime.datetime,
    output_path: str,
) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not available — Excel skipped")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alloy Consumption'

    H_FILL  = PatternFill('solid', fgColor='1F3864')
    H_FONT  = Font(bold=True, color='FFFFFF')
    TOT_FILL = PatternFill('solid', fgColor='FFF3CD')
    TOT_FONT = Font(bold=True)
    THIN     = Border(*(Side(style='thin'),) * 4)
    ALT_FILL = PatternFill('solid', fgColor='F4F6F8')

    # Period header
    ws['A1'] = 'Alloy Consumption Report — PTHM Phase'
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:D1')
    ws['A2'] = (f'Production window: '
                f'{date_start.strftime("%d/%m/%Y %H:%M")} → '
                f'{date_end.strftime("%d/%m/%Y %H:%M")}')
    ws['A2'].font = Font(size=9, color='7F8C8D')
    ws.merge_cells('A2:D2')

    # Column headers (row 4)
    headers = ['Product Code', 'Qty Processed', 'Unit GR (Alloy)', 'Partial Total (gr)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(4, ci, h)
        c.fill, c.font, c.border = H_FILL, H_FONT, THIN
        c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[4].height = 18

    # Data
    for ri, r in enumerate(report_rows, 5):
        vals = [
            r['ProductCode'],
            r['QtyProcessed'],
            r['MaterialConsumptionGR'],
            r['PartialTotalGR'],
        ]
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = THIN
            if fill:
                cell.fill = fill
            if ci in (2, 3, 4):
                cell.alignment = Alignment(horizontal='center')
            if ci in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = '0.00'

    # Total row
    tot_row = 5 + len(report_rows)
    ws.cell(tot_row, 3, 'TOTAL DAY (gr):').font  = TOT_FONT
    ws.cell(tot_row, 3).fill = TOT_FILL
    ws.cell(tot_row, 4, total_day_gr).font  = TOT_FONT
    ws.cell(tot_row, 4).fill   = TOT_FILL
    ws.cell(tot_row, 4).number_format = '0.00'
    ws.cell(tot_row, 4).border = THIN
    ws.cell(tot_row, 3).border = THIN

    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.freeze_panes = 'A5'

    wb.save(output_path)
    return True


# ─── Public API ───────────────────────────────────────────────────────────────

def get_pthm_rows(conn, reference_date: datetime.date | None = None) -> list[dict]:
    """Fetch PTHM rows for the production window of the given reference date."""
    start, end = _production_window(reference_date)
    return _fetch_pthm_rows(conn, start, end)


def generate_reports(
    conn,
    report_rows: list[dict],
    total_day_gr: float,
    date_start: datetime.datetime,
    date_end: datetime.datetime,
    output_dir: str | None = None,
) -> tuple[str | None, str | None]:
    """
    Generates PDF and Excel reports.
    Returns (pdf_path, excel_path) — None if generation failed.
    """
    out_dir = output_dir or tempfile.gettempdir()
    date_str = date_start.strftime('%Y%m%d')
    pdf_path   = os.path.join(out_dir, f'AlloyConsumption_{date_str}.pdf')
    excel_path = os.path.join(out_dir, f'AlloyConsumption_{date_str}.xlsx')

    start, end = date_start, date_end

    ok_pdf = _build_pdf(report_rows, total_day_gr, start, end, pdf_path)
    ok_xl  = _build_excel(report_rows, total_day_gr, start, end, excel_path)

    return (pdf_path if ok_pdf else None, excel_path if ok_xl else None)


def run_missing_alloy_check(db: Any) -> None:
    """
    Daily scheduled job: checks for products with missing alloy data
    and sends one warning email per day (anti-duplication).
    Entry point called from main.py thread.
    """
    import utils

    try:
        today  = datetime.date.today()
        # Mon=0 … Sat=5, Sun=6 → skip Sunday
        if today.weekday() == 6:
            return

        setting_key = f'SentMissingAlloy_{today.strftime("%Y%m%d")}'  # 24 chars
        conn = db.conn

        rows = get_pthm_rows(conn)
        missing = [r for r in rows if r['MissingAlloyConsumption'] == 1]

        if not missing:
            logger.info("material_consumption_report: nessun codice mancante — skip email")
            return

        if not _claim_send_slot(conn, setting_key):
            logger.info(f"material_consumption_report: email già inviata oggi ({today})")
            return

        recipients = _get_recipients(conn, 'Sys_missing_data_alloy')
        if not recipients:
            logger.warning("material_consumption_report: Sys_missing_data_alloy non configurata")
            return

        start, end = _production_window()
        html = _build_missing_alloy_email_html(missing, start, end)
        subject = (f"⚠️ Missing Alloy Data — {len(missing)} product(s) — "
                   f"{(today - datetime.timedelta(days=1)).strftime('%d/%m/%Y')}")

        utils.send_email(recipients, subject, html, is_html=True)
        logger.info(f"material_consumption_report: email inviata a {recipients}, "
                    f"{len(missing)} codici mancanti")

    except Exception as e:
        logger.error(f"material_consumption_report run_missing_alloy_check: {e}",
                     exc_info=True)
