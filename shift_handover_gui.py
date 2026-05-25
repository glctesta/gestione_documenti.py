# -*- coding: utf-8 -*-
"""
shift_handover_gui.py
Modulo Cambio Turno (Predare Schimb).

Finestra con 2 tab:
  - Compila Consegna  : capo turno uscente compila e salva
  - Storico / Conferma: capo turno entrante legge e conferma
"""
import tkinter as tk
from tkinter import ttk, messagebox
import logging
import socket
import os
from datetime import datetime, date

logger = logging.getLogger('TraceabilityRS')

# ─── Costanti turni ──────────────────────────────────────────────────────────
SHIFTS = {
    1: {'label': 'Turno 1  (fine 15:30)', 'end_h': 15, 'end_m': 30},
    2: {'label': 'Turno 2  (fine 23:30)', 'end_h': 23, 'end_m': 30},
    3: {'label': 'Turno 3  (fine 07:30)', 'end_h':  7, 'end_m': 30},
}


def _current_shift():
    """Indovina il turno corrente in base all'ora."""
    h = datetime.now().hour
    if 7 <= h < 15:
        return 1
    if 15 <= h < 23:
        return 2
    return 3


def _fetch_departments(db):
    """Carica reparti da employee.dbo.CdcSub."""
    try:
        sql = "SELECT SubCdcId, SubCdcDescription FROM Employee.dbo.CdcSub where SubCdcId in (1,9,17) ORDER BY SubCdcDescription"
        rows = None
        if hasattr(db, 'fetch_all'):
            rows = db.fetch_all(sql)
        else:
            with db._lock:
                db.cursor.execute(sql)
                rows = db.cursor.fetchall()
        return [(r[0], r[1]) for r in rows] if rows else []
    except Exception as e:
        logger.error(f"Errore caricamento reparti: {e}")
        return [(None, 'SMT'), (None, 'PTHH'), (None, 'WAREHOUSE')]


def _execute(db, sql, params=()):
    """Esegui query di scrittura."""
    try:
        if hasattr(db, 'execute_query'):
            return db.execute_query(sql, params)
        with db._lock:
            db.cursor.execute(sql, params)
            db.conn.commit()
        return True
    except Exception as e:
        logger.error(f"Errore SQL: {e}", exc_info=True)
        return False


def _fetch_one(db, sql, params=()):
    try:
        if hasattr(db, 'fetch_one'):
            return db.fetch_one(sql, params)
        with db._lock:
            db.cursor.execute(sql, params)
            return db.cursor.fetchone()
    except Exception as e:
        logger.error(f"fetch_one error: {e}")
        return None


def _fetch_all(db, sql, params=()):
    try:
        if hasattr(db, 'fetch_all'):
            return db.fetch_all(sql, params) or []
        with db._lock:
            db.cursor.execute(sql, params)
            return db.cursor.fetchall() or []
    except Exception as e:
        logger.error(f"fetch_all error: {e}")
        return []


# ─── Finestra principale ─────────────────────────────────────────────────────
class ShiftHandoverWindow(tk.Toplevel):

    def __init__(self, master, db, lang, current_user, preselect_dept=None, preselect_shift=None):
        super().__init__(master)
        self.db           = db
        self.lang         = lang
        self.current_user = current_user or ''
        self.hostname     = socket.gethostname()
        self.preselect_dept  = preselect_dept
        self.preselect_shift = preselect_shift
        self._depts = []  # [(SubCdcId, name)]

        self.title(lang.get('shift_handover_title', 'Cambio Turno — Predare Schimb'))
        self.geometry('900x750')
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()

        self._build_ui()
        self.protocol('WM_DELETE_WINDOW', self.destroy)

    # ── UI ───────────────────────────────────────────────────────────────────
    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        self._tab_compile = ttk.Frame(nb)
        self._tab_history  = ttk.Frame(nb)

        nb.add(self._tab_compile, text=self.lang.get('sh_tab_compile', '✏️  Compila Consegna'))
        nb.add(self._tab_history,  text=self.lang.get('sh_tab_history', '📋  Storico / Conferma'))

        self._build_compile_tab()
        self._build_history_tab()

    # ── Tab Compilazione ─────────────────────────────────────────────────────
    def _build_compile_tab(self):
        f = self._tab_compile
        self._depts = _fetch_departments(self.db)
        dept_names = [d[1] for d in self._depts]

        # Riga header filtri
        hdr = ttk.LabelFrame(f, text=self.lang.get('sh_header', 'Intestazione turno'), padding=8)
        hdr.pack(fill='x', padx=8, pady=(8, 4))

        row0 = ttk.Frame(hdr); row0.pack(fill='x')

        ttk.Label(row0, text=self.lang.get('sh_date', 'Data:')).pack(side='left')
        self.date_var = tk.StringVar(value=date.today().strftime('%d/%m/%Y'))
        ttk.Entry(row0, textvariable=self.date_var, width=12, state='readonly').pack(side='left', padx=(4, 20))

        ttk.Label(row0, text=self.lang.get('sh_shift', 'Turno:')).pack(side='left')
        self.shift_var = tk.IntVar(value=self.preselect_shift or _current_shift())
        for sn, sd in SHIFTS.items():
            ttk.Radiobutton(row0, text=sd['label'], variable=self.shift_var, value=sn).pack(side='left', padx=4)

        row1 = ttk.Frame(hdr); row1.pack(fill='x', pady=(6, 0))
        ttk.Label(row1, text=self.lang.get('sh_dept', 'Reparto:')).pack(side='left')
        self.dept_var = tk.StringVar()
        dept_cb = ttk.Combobox(row1, textvariable=self.dept_var, values=dept_names, width=25, state='readonly')
        dept_cb.pack(side='left', padx=4)
        if self.preselect_dept and self.preselect_dept in dept_names:
            self.dept_var.set(self.preselect_dept)
        elif dept_names:
            self.dept_var.set(dept_names[0])

        ttk.Label(row1, text=self.lang.get('sh_compiled_by', 'Compilato da:'), foreground='#555').pack(side='left', padx=(20, 4))
        ttk.Label(row1, text=self.current_user, font=('Segoe UI', 9, 'bold')).pack(side='left')

        # Sezioni report
        body = ttk.Frame(f, padding=(8, 0))
        body.pack(fill='both', expand=True)

        def _section(parent, label_key, label_default, height=3):
            lf = ttk.LabelFrame(parent, text=self.lang.get(label_key, label_default), padding=4)
            lf.pack(fill='x', pady=3)
            t = tk.Text(lf, height=height, wrap='word', font=('Segoe UI', 9))
            sb = ttk.Scrollbar(lf, command=t.yview)
            t.configure(yscrollcommand=sb.set)
            sb.pack(side='right', fill='y')
            t.pack(fill='both', expand=True)
            return t

        self.txt_prod_status  = _section(body, 'sh_field_prod_status',  'Stare producție (plan / deviații)', 3)
        self.txt_lines        = _section(body, 'sh_field_lines',         'Linii / echipamente',               3)

        qty_frame = ttk.LabelFrame(body, text=self.lang.get('sh_field_qty', 'Cantități'), padding=4)
        qty_frame.pack(fill='x', pady=3)
        ttk.Label(qty_frame, text=self.lang.get('sh_qty_plan', 'Planificat:')).pack(side='left')
        self.qty_plan_var = tk.StringVar()
        ttk.Entry(qty_frame, textvariable=self.qty_plan_var, width=10).pack(side='left', padx=(4, 20))
        ttk.Label(qty_frame, text=self.lang.get('sh_qty_prod', 'Realizat:')).pack(side='left')
        self.qty_prod_var = tk.StringVar()
        ttk.Entry(qty_frame, textvariable=self.qty_prod_var, width=10).pack(side='left', padx=4)

        self.txt_quality      = _section(body, 'sh_field_quality',  'Calitate — Probleme + Acțiuni', 3)
        self.txt_materials    = _section(body, 'sh_field_materials', 'Materiale (disponibile / lipsă)',  2)
        self.txt_open_issues  = _section(body, 'sh_field_open',      'Probleme deschise — de urmărit',  3)
        self.txt_notes        = _section(body, 'sh_field_notes',     'Note libere',                     2)

        # Pulsante salva
        btn_frame = ttk.Frame(f, padding=8)
        btn_frame.pack(fill='x')
        ttk.Button(
            btn_frame,
            text=self.lang.get('sh_btn_save', '💾  Salva Consegna Turno'),
            command=self._save_report
        ).pack(side='right')

    def _save_report(self):
        dept_name = self.dept_var.get().strip()
        if not dept_name:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('sh_err_dept', 'Selezionare un reparto.'), parent=self)
            return

        def _txt(widget): return widget.get('1.0', 'end').strip()
        def _qty(var):
            try: return float(var.get().replace(',', '.'))
            except: return None

        # Cerca SubCdcId corrispondente
        sub_cdc_id = next((d[0] for d in self._depts if d[1] == dept_name), None)

        sql = """
            INSERT INTO Employee.dbo.ShiftHandoverReports
                (ShiftDate, ShiftNumber, Department, SubCdcId,
                 CompiledBy, ComputerName,
                 ProductionStatus, LinesStatus,
                 QtyPlanned, QtyProduced,
                 QualityIssues, MaterialStatus,
                 OpenIssues, FreeNotes)
            VALUES (CAST(GETDATE() AS DATE), ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?)
        """
        params = (
            self.shift_var.get(), dept_name, sub_cdc_id,
            self.current_user, self.hostname,
            _txt(self.txt_prod_status), _txt(self.txt_lines),
            _qty(self.qty_plan_var), _qty(self.qty_prod_var),
            _txt(self.txt_quality), _txt(self.txt_materials),
            _txt(self.txt_open_issues), _txt(self.txt_notes),
        )
        if _execute(self.db, sql, params):
            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                self.lang.get('sh_saved_ok', 'Consegna turno salvata con successo.\nIl prossimo capo turno dovrà confermare la lettura.'),
                parent=self
            )
            # Ricarica storico
            self._load_history()
            logger.info(f"ShiftHandover salvato: turno={self.shift_var.get()} dept={dept_name} by={self.current_user}")
        else:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('sh_saved_err', 'Errore durante il salvataggio.'), parent=self)

    # ── Tab Storico / Conferma ───────────────────────────────────────────────
    def _build_history_tab(self):
        f = self._tab_history

        # Filtri
        flt = ttk.Frame(f, padding=(8, 8, 8, 4))
        flt.pack(fill='x')

        ttk.Label(flt, text=self.lang.get('sh_filter_date', 'Data:')).pack(side='left')
        self.hist_date_var = tk.StringVar(value=date.today().strftime('%d/%m/%Y'))
        ttk.Entry(flt, textvariable=self.hist_date_var, width=12).pack(side='left', padx=(4, 16))

        ttk.Label(flt, text=self.lang.get('sh_filter_dept', 'Reparto:')).pack(side='left')
        all_depts = [''] + [d[1] for d in (_fetch_departments(self.db) if not self._depts else self._depts)]
        self.hist_dept_var = tk.StringVar()
        ttk.Combobox(flt, textvariable=self.hist_dept_var, values=all_depts, width=20, state='readonly').pack(side='left', padx=4)

        ttk.Button(flt, text=self.lang.get('btn_search', '🔍 Cerca'), command=self._load_history).pack(side='left', padx=8)

        # Treeview lista
        cols = ('ID', 'Data', 'Turno', 'Reparto', 'Compilato da', 'Ora', 'Confermato', 'Da')
        self.hist_tree = ttk.Treeview(f, columns=cols, show='headings', height=8, selectmode='browse')
        widths = [50, 90, 70, 120, 140, 70, 90, 120]
        for c, w in zip(cols, widths):
            self.hist_tree.heading(c, text=c)
            self.hist_tree.column(c, width=w, anchor='w')
        sb = ttk.Scrollbar(f, orient='vertical', command=self.hist_tree.yview)
        self.hist_tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y', padx=(0, 8))
        self.hist_tree.pack(fill='x', padx=(8, 0), pady=4)
        self.hist_tree.tag_configure('confirmed', foreground='#27ae60')
        self.hist_tree.tag_configure('pending', foreground='#e74c3c')
        self.hist_tree.bind('<<TreeviewSelect>>', self._on_hist_select)

        # Dettaglio
        det = ttk.LabelFrame(f, text=self.lang.get('sh_detail', 'Dettaglio consegna selezionata'), padding=6)
        det.pack(fill='both', expand=True, padx=8, pady=(0, 4))

        self.detail_text = tk.Text(det, wrap='word', font=('Segoe UI', 9), state='disabled', height=12)
        sb2 = ttk.Scrollbar(det, command=self.detail_text.yview)
        self.detail_text.configure(yscrollcommand=sb2.set)
        sb2.pack(side='right', fill='y')
        self.detail_text.pack(fill='both', expand=True)

        # Pulsante conferma
        btn_f = ttk.Frame(f, padding=8)
        btn_f.pack(fill='x')
        self.confirm_btn = ttk.Button(
            btn_f,
            text=self.lang.get('sh_btn_confirm', '✅  Ho letto e preso nota della consegna'),
            command=self._confirm_reading,
            state='disabled'
        )
        self.confirm_btn.pack(side='right')

        self._selected_report_id = None
        self._load_history()

    def _load_history(self):
        self.hist_tree.delete(*self.hist_tree.get_children())
        self._selected_report_id = None
        self.confirm_btn.config(state='disabled')

        # Converti data filtro
        date_filter = None
        try:
            date_filter = datetime.strptime(self.hist_date_var.get().strip(), '%d/%m/%Y').date()
        except Exception:
            date_filter = date.today()

        dept_filter = self.hist_dept_var.get() if hasattr(self, 'hist_dept_var') else ''

        sql = """
            SELECT HandoverReportId, ShiftDate, ShiftNumber, Department,
                   CompiledBy, CompiledAt, IsConfirmed, ConfirmedBy
            FROM Employee.dbo.ShiftHandoverReports
            WHERE ShiftDate = ?
            {dept_clause}
            ORDER BY ShiftNumber, CompiledAt DESC
        """.format(dept_clause=f"AND Department = ?" if dept_filter else "")

        params = (date_filter, dept_filter) if dept_filter else (date_filter,)
        rows = _fetch_all(self.db, sql, params)

        for r in rows:
            conf_str = '✅ ' + (r[7] or '') if r[6] else '❌ In attesa'
            tag = 'confirmed' if r[6] else 'pending'
            self.hist_tree.insert('', 'end', iid=str(r[0]), tags=(tag,), values=(
                r[0],
                r[1].strftime('%d/%m/%Y') if r[1] else '',
                f"Turno {r[2]}",
                r[3] or '',
                r[4] or '',
                r[5].strftime('%H:%M') if r[5] else '',
                conf_str,
                r[7] or ''
            ))

    def _on_hist_select(self, _event=None):
        sel = self.hist_tree.selection()
        if not sel:
            return
        rid = int(sel[0])
        self._selected_report_id = rid

        sql = """
            SELECT ShiftDate, ShiftNumber, Department, CompiledBy, CompiledAt,
                   ProductionStatus, LinesStatus, QtyPlanned, QtyProduced,
                   QualityIssues, MaterialStatus, OpenIssues, FreeNotes,
                   IsConfirmed, ConfirmedBy, ConfirmedAt
            FROM Employee.dbo.ShiftHandoverReports
            WHERE HandoverReportId = ?
        """
        r = _fetch_one(self.db, sql, (rid,))
        if not r:
            return

        lines = [
            f"📅  {r[0].strftime('%d/%m/%Y')}  |  Turno {r[1]}  |  Reparto: {r[2]}",
            f"👤  Compilato da: {r[3]}  ore {r[4].strftime('%H:%M') if r[4] else ''}",
            "",
            "─── Stare producție ──────────────────────────",
            r[5] or '(nessuna indicazione)',
            "",
            "─── Linii / echipamente ──────────────────────",
            r[6] or '(nessuna indicazione)',
            "",
            f"📦  Cantitate — Planificat: {r[7] or '—'}  |  Realizat: {r[8] or '—'}",
            "",
            "─── Calitate ─────────────────────────────────",
            r[9] or '(nessuna indicazione)',
            "",
            "─── Materiale ────────────────────────────────",
            r[10] or '(nessuna indicazione)',
            "",
            "─── Probleme deschise ────────────────────────",
            r[11] or '(nessun problema aperto)',
            "",
            "─── Note libere ──────────────────────────────",
            r[12] or '',
        ]
        if r[13]:
            lines += ["", f"✅  Confermat de: {r[14]}  la {r[15].strftime('%d/%m/%Y %H:%M') if r[15] else ''}"]

        self.detail_text.config(state='normal')
        self.detail_text.delete('1.0', 'end')
        self.detail_text.insert('end', '\n'.join(lines))
        self.detail_text.config(state='disabled')

        # Abilita conferma solo se:
        #   1) non ancora confermato
        #   2) l'utente corrente NON è chi ha compilato (no auto-conferma)
        compiled_by = (r[3] or '').strip().lower()
        current     = self.current_user.strip().lower()
        self._compiled_by = r[3] or ''   # salva per guard in _confirm_reading

        if r[13]:
            # già confermato
            self.confirm_btn.config(state='disabled')
        elif compiled_by == current:
            # stessa persona — blocca con tooltip testuale
            self.confirm_btn.config(state='disabled')
            # mostra avviso nella status bar del dettaglio
            self.detail_text.config(state='normal')
            self.detail_text.insert('end',
                '\n\n⚠️  ' + self.lang.get(
                    'sh_no_self_confirm',
                    'Non è consentito confermare una propria consegna turno.'
                )
            )
            self.detail_text.config(state='disabled')
        else:
            self.confirm_btn.config(state='normal')

    def _confirm_reading(self):
        if not self._selected_report_id:
            return

        # Guard di sicurezza: l'autore non può confermare la propria consegna
        compiled_by = getattr(self, '_compiled_by', '').strip().lower()
        if compiled_by and compiled_by == self.current_user.strip().lower():
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get(
                    'sh_no_self_confirm',
                    'Non è consentito confermare una propria consegna turno.'
                ),
                parent=self
            )
            return

        dlg = ShiftHandoverConfirmDialog(self, self.lang, self.current_user)
        self.wait_window(dlg)
        if not dlg.confirmed:
            return

        # Aggiorna report principale
        _execute(self.db,
            """UPDATE Employee.dbo.ShiftHandoverReports
               SET IsConfirmed=1, ConfirmedBy=?, ConfirmedAt=GETDATE()
               WHERE HandoverReportId=?""",
            (self.current_user, self._selected_report_id))

        # Inserisci riga di conferma
        _execute(self.db,
            """INSERT INTO Employee.dbo.ShiftHandoverConfirmations
                   (HandoverReportId, ConfirmedBy, ComputerName, Notes)
               VALUES (?, ?, ?, ?)""",
            (self._selected_report_id, self.current_user, self.hostname, dlg.notes))

        messagebox.showinfo(
            self.lang.get('success', 'Successo'),
            self.lang.get('sh_confirmed_ok', 'Conferma registrata con successo.'),
            parent=self
        )
        logger.info(f"ShiftHandover {self._selected_report_id} confermato da {self.current_user}")
        self._load_history()


# ─── Dialog conferma lettura ─────────────────────────────────────────────────
class ShiftHandoverConfirmDialog(tk.Toplevel):

    def __init__(self, parent, lang, user):
        super().__init__(parent)
        self.lang      = lang
        self.confirmed = False
        self.notes     = ''

        self.title(lang.get('sh_confirm_title', 'Conferma Lettura Consegna'))
        self.geometry('480x300')
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        ttk.Label(self,
            text=lang.get('sh_confirm_msg',
                 f'Confermo di aver letto e preso nota\ndella consegna del turno precedente.\n\nUtente: {user}'),
            font=('Segoe UI', 11), justify='center'
        ).pack(pady=20)

        ttk.Label(self, text=lang.get('sh_confirm_notes', 'Note aggiuntive (facoltativo):')).pack(anchor='w', padx=20)
        self.notes_text = tk.Text(self, height=4, wrap='word', font=('Segoe UI', 9))
        self.notes_text.pack(fill='x', padx=20, pady=(4, 12))

        btn = ttk.Frame(self)
        btn.pack(fill='x', padx=20, pady=8)
        ttk.Button(btn, text=lang.get('sh_btn_confirm_ok', '✅ Confermo'),
                   command=self._on_confirm).pack(side='left', expand=True, fill='x', padx=(0, 4))
        ttk.Button(btn, text=lang.get('cancel', 'Annulla'),
                   command=self.destroy).pack(side='left', expand=True, fill='x', padx=(4, 0))

        self.protocol('WM_DELETE_WINDOW', self.destroy)

    def _on_confirm(self):
        self.notes     = self.notes_text.get('1.0', 'end').strip()
        self.confirmed = True
        self.destroy()


# ─── Entry point principale ───────────────────────────────────────────────────
def open_shift_handover(master, db, lang, current_user,
                         preselect_dept=None, preselect_shift=None):
    ShiftHandoverWindow(master, db, lang, current_user,
                        preselect_dept=preselect_dept,
                        preselect_shift=preselect_shift)


# ─── Finestra Report ──────────────────────────────────────────────────────────
class ShiftHandoverReportWindow(tk.Toplevel):
    """
    Report consegne turno con filtri periodo e reparto.
    Mostra:
      - Riepilogo tabellare (data, turno, reparto, compilato da, confermato)
      - KPI: % conferme, media Qty piano vs prodotto
    Export Excel in C:\\Temp.
    """

    def __init__(self, master, db, lang, current_user):
        super().__init__(master)
        self.db           = db
        self.lang         = lang
        self.current_user = current_user or ''
        self._rows        = []

        self.title(lang.get('sh_report_title', 'Report Cambio Turno'))
        self.geometry('1050x620')
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()

        self._build_ui()
        self._load()
        self.protocol('WM_DELETE_WINDOW', self.destroy)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Filtri ────────────────────────────────────────────────────────────
        flt = ttk.LabelFrame(self, text=self.lang.get('sh_rep_filters', 'Filtri'), padding=8)
        flt.pack(fill='x', padx=8, pady=(8, 4))

        # Date dal/al
        ttk.Label(flt, text=self.lang.get('sh_rep_from', 'Dal:')).pack(side='left')
        self.from_var = tk.StringVar(value=_first_day_of_month())
        ttk.Entry(flt, textvariable=self.from_var, width=11).pack(side='left', padx=(4, 14))

        ttk.Label(flt, text=self.lang.get('sh_rep_to', 'Al:')).pack(side='left')
        self.to_var = tk.StringVar(value=date.today().strftime('%d/%m/%Y'))
        ttk.Entry(flt, textvariable=self.to_var, width=11).pack(side='left', padx=(4, 14))

        # Reparto
        ttk.Label(flt, text=self.lang.get('sh_filter_dept', 'Reparto:')).pack(side='left')
        all_depts = [''] + [d[1] for d in _fetch_departments(self.db)]
        self.dept_var = tk.StringVar()
        ttk.Combobox(flt, textvariable=self.dept_var, values=all_depts, width=20, state='readonly').pack(side='left', padx=4)

        # Turno
        ttk.Label(flt, text=self.lang.get('sh_shift', 'Turno:')).pack(side='left', padx=(12, 2))
        self.shift_var = tk.StringVar(value='')
        ttk.Combobox(flt, textvariable=self.shift_var,
                     values=['', '1', '2', '3'], width=4, state='readonly').pack(side='left', padx=4)

        ttk.Button(flt, text=self.lang.get('btn_search', '🔍 Cerca'),
                   command=self._load).pack(side='left', padx=10)
        ttk.Button(flt, text=self.lang.get('sh_rep_btn_excel', '📥 Export Excel'),
                   command=self._export_excel).pack(side='left')

        # ── KPI strip ─────────────────────────────────────────────────────────
        kpi_frame = ttk.Frame(self, padding=(8, 2))
        kpi_frame.pack(fill='x')
        self.kpi_total  = ttk.Label(kpi_frame, text='', font=('Segoe UI', 9))
        self.kpi_conf   = ttk.Label(kpi_frame, text='', font=('Segoe UI', 9, 'bold'))
        self.kpi_qty    = ttk.Label(kpi_frame, text='', font=('Segoe UI', 9))
        for w in (self.kpi_total, self.kpi_conf, self.kpi_qty):
            w.pack(side='left', padx=12)

        # ── Treeview ──────────────────────────────────────────────────────────
        cols = ('Data', 'Turno', 'Reparto', 'Compilato da', 'Ora', 'Qty Piano', 'Qty Prod.', 'Confermato', 'Da', 'Ora conf.')
        self.tree = ttk.Treeview(self, columns=cols, show='headings', height=16, selectmode='browse')
        widths = [90, 65, 120, 140, 60, 80, 80, 100, 130, 80]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor='w')
        sb = ttk.Scrollbar(self, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y', padx=(0, 8))
        self.tree.pack(fill='both', expand=True, padx=(8, 0), pady=4)
        self.tree.tag_configure('confirmed', foreground='#27ae60')
        self.tree.tag_configure('pending',   foreground='#c0392b')

        # ── Dettaglio problemi aperti ─────────────────────────────────────────
        det = ttk.LabelFrame(self, text=self.lang.get('sh_field_open', 'Probleme deschise'), padding=4)
        det.pack(fill='x', padx=8, pady=(0, 8))
        self.open_issues_text = tk.Text(det, height=3, wrap='word', font=('Segoe UI', 9), state='disabled')
        self.open_issues_text.pack(fill='x')
        self.tree.bind('<<TreeviewSelect>>', self._on_select)

    # ── Dati ─────────────────────────────────────────────────────────────────
    def _parse_date(self, s):
        try:
            return datetime.strptime(s.strip(), '%d/%m/%Y').date()
        except Exception:
            return date.today()

    def _load(self):
        self.tree.delete(*self.tree.get_children())
        self._rows = []

        d_from = self._parse_date(self.from_var.get())
        d_to   = self._parse_date(self.to_var.get())
        dept   = self.dept_var.get().strip()
        shift  = self.shift_var.get().strip()

        where_extra = ''
        params = [d_from, d_to]
        if dept:
            where_extra += ' AND Department = ?'
            params.append(dept)
        if shift:
            where_extra += ' AND ShiftNumber = ?'
            params.append(int(shift))

        sql = f"""
            SELECT HandoverReportId, ShiftDate, ShiftNumber, Department,
                   CompiledBy, CompiledAt,
                   QtyPlanned, QtyProduced,
                   IsConfirmed, ConfirmedBy, ConfirmedAt,
                   OpenIssues
            FROM Employee.dbo.ShiftHandoverReports
            WHERE ShiftDate BETWEEN ? AND ?
            {where_extra}
            ORDER BY ShiftDate DESC, ShiftNumber, Department
        """
        rows = _fetch_all(self.db, sql, tuple(params))
        self._rows = rows

        total = len(rows)
        conf  = sum(1 for r in rows if r[8])
        pct   = f"{conf/total*100:.0f}%" if total else '—'
        qty_pairs = [(r[6], r[7]) for r in rows if r[6] is not None and r[7] is not None]
        avg_diff  = (sum(p - q for p, q in qty_pairs) / len(qty_pairs)) if qty_pairs else None
        diff_str  = f"{avg_diff:+.1f}" if avg_diff is not None else '—'

        self.kpi_total.config(text=f"Totale: {total}")
        self.kpi_conf.config(text=f"Confermate: {conf}/{total} ({pct})",
                             foreground='#27ae60' if conf == total else '#c0392b')
        self.kpi_qty.config(text=f"Δ Qty medio (Prod-Piano): {diff_str}")

        for r in rows:
            tag = 'confirmed' if r[8] else 'pending'
            conf_str = r[9] or '—'
            self.tree.insert('', 'end', iid=str(r[0]), tags=(tag,), values=(
                r[1].strftime('%d/%m/%Y') if r[1] else '',
                f"T{r[2]}",
                r[3] or '',
                r[4] or '',
                r[5].strftime('%H:%M') if r[5] else '',
                f"{r[6]:.0f}" if r[6] is not None else '',
                f"{r[7]:.0f}" if r[7] is not None else '',
                '✅' if r[8] else '❌',
                conf_str,
                r[10].strftime('%H:%M') if r[10] else '',
            ))

    def _on_select(self, _event=None):
        sel = self.tree.selection()
        if not sel:
            return
        rid = int(sel[0])
        row = next((r for r in self._rows if r[0] == rid), None)
        text = (row[11] or '(nessun problema aperto)') if row else ''
        self.open_issues_text.config(state='normal')
        self.open_issues_text.delete('1.0', 'end')
        self.open_issues_text.insert('end', text)
        self.open_issues_text.config(state='disabled')

    # ── Export Excel ──────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self._rows:
            messagebox.showinfo(self.lang.get('info', 'Info'),
                                self.lang.get('sh_rep_no_data', 'Nessun dato da esportare.'),
                                parent=self)
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = 'Cambio Turno'

            headers = ['Data', 'Turno', 'Reparto', 'Compilato da', 'Ora',
                       'Qty Pianificato', 'Qty Prodotto', 'Δ Qty',
                       'Confermato', 'Confermato da', 'Ora conferma',
                       'Problemi aperti']
            hdr_fill = PatternFill('solid', fgColor='1F497D')
            hdr_font = Font(bold=True, color='FFFFFF')
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')

            for ri, r in enumerate(self._rows, 2):
                qty_p = r[6]
                qty_r = r[7]
                delta = (qty_r - qty_p) if (qty_p is not None and qty_r is not None) else None
                ws.append([
                    r[1].strftime('%d/%m/%Y') if r[1] else '',
                    r[2],
                    r[3] or '',
                    r[4] or '',
                    r[5].strftime('%H:%M') if r[5] else '',
                    qty_p,
                    qty_r,
                    delta,
                    'Sì' if r[8] else 'No',
                    r[9] or '',
                    r[10].strftime('%H:%M') if r[10] else '',
                    r[11] or '',
                ])
                # Colora righe non confermate
                if not r[8]:
                    for ci in range(1, len(headers) + 1):
                        ws.cell(row=ri, column=ci).fill = PatternFill('solid', fgColor='FFE0E0')

            # Autosize
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

            os.makedirs(r'C:\Temp', exist_ok=True)
            ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
            path = rf'C:\Temp\CambioTurno_{ts}.xlsx'
            wb.save(path)
            logger.info(f"Report Cambio Turno esportato: {path}")

            if messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                self.lang.get('sh_rep_open_file', f'File salvato:\n{path}\n\nAprire il file?'),
                parent=self
            ):
                os.startfile(path)

        except Exception as e:
            logger.error(f"Errore export Excel Cambio Turno: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)


def _first_day_of_month():
    today = date.today()
    return today.replace(day=1).strftime('%d/%m/%Y')


def open_shift_handover_report(master, db, lang, current_user):
    ShiftHandoverReportWindow(master, db, lang, current_user)
