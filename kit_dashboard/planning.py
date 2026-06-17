# -*- coding: utf-8 -*-
"""
planning.py — Ora pianificata della prima fase PTHM per ordine, da T:\\Planning.

Riusa percorso/tab/colonne di `fai_autocheck.py` (sorgente già in uso per gli
ordini urgenti). A differenza di `fai_autocheck.read_planning_excel()` (che
filtra alla finestra +4h), qui leggiamo TUTTE le righe e teniamo, per ciascun
ordine, il **primo** `PlannedStart` con fase == 'PTHM' (match esatto, escluso
'PTHM SELECTIVE').
"""
import logging
from datetime import datetime

import fai_autocheck as fa

logger = logging.getLogger("KitDashboard")

_DATE_FORMATS = ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M',
                 '%Y-%m-%d %H:%M', '%m/%d/%Y %H:%M:%S')


def _parse_dt(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def pthm_planned_starts(phase_name: str = "PTHM", planning_path: str = None) -> dict:
    """Ritorna {order_number: planned_start} (il più antico) per la fase PTHM.

    `planning_path` (UNC consigliato) sovrascrive il default di fai_autocheck
    (`T:\\Planning`): necessario quando il server gira come SYSTEM (senza T:).

    Ritorna {} se il file non è raggiungibile/in lock: chi chiama deve
    trattare l'assenza come 'nessuna data pianificata' (nessun flag di ritardo).
    """
    folder = planning_path or fa.PLANNING_PATH
    filepath = fa._find_latest_excel(folder)
    if not filepath:
        logger.warning("Planning: nessun file Excel in %s", folder)
        return {}

    target = (phase_name or "PTHM").strip().upper()
    result = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if fa.PLANNING_TAB not in wb.sheetnames:
            logger.error("Planning: tab '%s' assente in %s", fa.PLANNING_TAB, filepath)
            wb.close()
            return {}
        ws = wb[fa.PLANNING_TAB]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= fa.COL_PLANNED_START:
                continue
            phase_raw = row[fa.COL_PHASE]
            order_raw = row[fa.COL_ORDER_NUMBER]
            start_raw = row[fa.COL_PLANNED_START]
            if not phase_raw or not order_raw or not start_raw:
                continue
            if str(phase_raw).strip().upper() != target:
                continue
            dt = _parse_dt(start_raw)
            if not dt:
                continue
            order = str(order_raw).strip()
            if order not in result or dt < result[order]:
                result[order] = dt
        wb.close()
    except PermissionError as e:
        logger.warning("Planning: file in lock (%s), riprovo al prossimo ciclo", e)
        return {}
    except Exception as e:
        logger.error("Planning: errore lettura Excel: %s", e, exc_info=True)
        return {}

    logger.info("Planning: %d ordini con PlannedStart fase %s", len(result), target)
    return result
