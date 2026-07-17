# File: orders/orders_reports_window.py
"""
Finestra per la gestione delle spedizioni dinamiche
"""
import logging
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

logger = logging.getLogger(__name__)


class DynamicShippingWindow(tk.Toplevel):
    """Finestra per la gestione delle spedizioni dinamiche"""
    
    def __init__(self, master, db, lang, user_name):
        """
        Inizializza la finestra di gestione spedizioni
        
        Args:
            master: Finestra padre
            db: Connessione database
            lang: Gestore traduzioni
            user_name: Nome dell'utente loggato
        """
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        
        self.current_order_id = None  # IDOrder selezionato
        self.current_production_order_id = None  # DynamicProductionOrderID
        self._warehouse_rows = []  # FASE 1: ordini finiti a magazzino non spediti
        self.current_remain = 0  # QuantitÃ  rimanente
        
        self.title(self.lang.get('dynamic_shipping_title', 'Gestione Spedizioni Dinamiche'))
        self.geometry('1600x900')
        self.transient(master)
        
        self._create_widgets()
        self._load_filter_data()
        self._load_order_data()  # Carica automaticamente i dati all'apertura
    
    def _create_widgets(self):
        """Crea i widget della finestra"""
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame filtri
        self._create_filters_frame(main_frame)
        
        # PanedWindow per dividere in due parti
        paned = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Parte superiore: Dati ordini
        self._create_orders_section(paned)
        
        # Parte inferiore: Regole spedizione
        self._create_shipping_rules_section(paned)
    
    def _create_filters_frame(self, parent):
        """Crea il frame dei filtri"""
        filters_frame = ttk.LabelFrame(parent, text=self.lang.get('filters', 'Filtri'), padding=10)
        filters_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Prima riga filtri
        row1 = ttk.Frame(filters_frame)
        row1.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(row1, text=self.lang.get('customer', 'Cliente:')).pack(side=tk.LEFT, padx=(0, 5))
        self.customer_filter = ttk.Combobox(row1, width=25, state='readonly')
        self.customer_filter.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text=self.lang.get('sale_order', 'Ordine Vendita:')).pack(side=tk.LEFT, padx=(0, 5))
        self.so_filter = ttk.Entry(row1, width=15)
        self.so_filter.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text=self.lang.get('production_order', 'Ordine Produzione:')).pack(side=tk.LEFT, padx=(0, 5))
        self.po_filter = ttk.Entry(row1, width=15)
        self.po_filter.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text=self.lang.get('product', 'Prodotto:')).pack(side=tk.LEFT, padx=(0, 5))
        self.product_filter = ttk.Combobox(row1, width=25, state='readonly')
        self.product_filter.pack(side=tk.LEFT)
        
        # Seconda riga filtri
        row2 = ttk.Frame(filters_frame)
        row2.pack(fill=tk.X)
        
        ttk.Label(row2, text=self.lang.get('ship_date_from', 'Data Spedizione Da:')).pack(side=tk.LEFT, padx=(0, 5))
        self.date_from = DateEntry(
            row2,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='it_IT'
        )
        self.date_from.set_date(datetime.now() - timedelta(days=180))
        self.date_from.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row2, text=self.lang.get('ship_date_to', 'A:')).pack(side=tk.LEFT, padx=(0, 5))
        self.date_to = DateEntry(
            row2,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='it_IT'
        )
        self.date_to.set_date(datetime.now() + timedelta(days=180))
        self.date_to.pack(side=tk.LEFT, padx=(0, 15))
        
        # Checkbox filtro rimanenti
        self.remain_filter_var = tk.BooleanVar(value=False)
        self.remain_checkbox = ttk.Checkbutton(
            row2, 
            text=self.lang.get('only_remaining', 'Solo rimanenti'),
            variable=self.remain_filter_var,
            command=self._apply_filters
        )
        self.remain_checkbox.pack(side=tk.LEFT, padx=15)
        
        ttk.Button(row2, text=self.lang.get('btn_filter', 'Filtra'), command=self._apply_filters).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text=self.lang.get('btn_reset_filters', 'Reset Filtri'), command=self._reset_filters).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text=self.lang.get('btn_export_excel', 'Esporta Excel'), command=self._export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def _create_orders_section(self, parent):
        """Crea la sezione superiore con i dati ordini"""
        orders_frame = ttk.LabelFrame(parent, text=self.lang.get('orders_data', 'Dati Ordini'), padding=10)
        parent.add(orders_frame, weight=2)

        # Barra superiore: legenda colori + bottone abbina/correggi magazzino.
        top_bar = ttk.Frame(orders_frame)
        top_bar.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 4))
        tk.Label(top_bar, text='  ', background='#ffe066').pack(side=tk.LEFT)
        ttk.Label(top_bar, text=self.lang.get(
            'wh_legend_yellow',
            'Giallo = da abbinare (magazzino, senza ordine di vendita)')).pack(
            side=tk.LEFT, padx=(4, 12))
        tk.Label(top_bar, text='  ', background='#f5c6cb').pack(side=tk.LEFT)
        ttk.Label(top_bar, text=self.lang.get(
            'wh_legend_red',
            'Rosso = urgente, merce già a magazzino')).pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(top_bar, text=self.lang.get(
            'wh_hint_dblclick',
            '— Doppio click su una riga gialla per abbinarla a un ordine di vendita '
            'e inserire i dati, oppure usa il bottone:'),
            foreground='#555').pack(side=tk.LEFT, padx=(0, 8))
        self.wh_match_btn = ttk.Button(
            top_bar, text=self.lang.get('wh_btn_open_match', '🔗 Abbina / Correggi'),
            command=self._open_warehouse_match, state=tk.DISABLED)
        self.wh_match_btn.pack(side=tk.LEFT)

        # Treeview ordini
        columns = ('IDOrder', 'Customer', 'SaleOrder', 'ProductionOrder', 'ItemCode', 'ItemName',
                  'ShipDate', 'QtyOrder', 'QtyAssigned', 'Associate', 'SMT', 'PTHM', 'ICT',
                  'FCT', 'Coating', 'CoatingBottom', 'OutOfBox', 'Remain')

        self.orders_tree = ttk.Treeview(orders_frame, columns=columns, show='headings', selectmode='browse')
        
        # Nascondi IDOrder ma usalo per selezione
        self.orders_tree.column('IDOrder', width=0, stretch=False)
        self.orders_tree.heading('IDOrder', text='')
        
        # Configura colonne visibili
        visible_columns = {
            'Customer': (self.lang.get('col_customer', 'Cliente'), 120),
            'SaleOrder': (self.lang.get('col_sale_order', 'Ord. Vendita'), 100),
            'ProductionOrder': (self.lang.get('col_production_order', 'Ord. Produzione'), 110),
            'ItemCode': (self.lang.get('col_item_code', 'Codice'), 100),
            'ItemName': (self.lang.get('col_item_name', 'Prodotto'), 150),
            'ShipDate': (self.lang.get('col_ship_date', 'Data Sped.'), 90),
            'QtyOrder': (self.lang.get('col_qty_order', 'QtÃ  Ord.'), 70),
            'QtyAssigned': (self.lang.get('col_qty_assigned', 'Assegnata'), 70),
            'Associate': (self.lang.get('col_associate', 'Associate'), 70),
            'SMT': (self.lang.get('col_smt', 'SMT'), 60),
            'PTHM': (self.lang.get('col_pthm', 'PTHM'), 60),
            'ICT': (self.lang.get('col_ict', 'ICT'), 60),
            'FCT': (self.lang.get('col_fct', 'FCT'), 60),
            'Coating': (self.lang.get('col_coating', 'Coating'), 60),
            'CoatingBottom': (self.lang.get('col_coating_bottom', 'Coating Bottom'), 90),
            'OutOfBox': (self.lang.get('col_outofbox', 'OutOfBox'), 70),
            'Remain': (self.lang.get('col_remain', 'Rimanenti'), 70)
        }
        
        for col, (text, width) in visible_columns.items():
            self.orders_tree.heading(col, text=text)
            # Centra colonne numeriche e date, allinea a sinistra testo
            if col in ['Customer', 'ItemName']:
                self.orders_tree.column(col, width=width, anchor=tk.W)
            else:
                self.orders_tree.column(col, width=width, anchor=tk.CENTER)
        
        # Scrollbars
        vsb = ttk.Scrollbar(orders_frame, orient=tk.VERTICAL, command=self.orders_tree.yview)
        hsb = ttk.Scrollbar(orders_frame, orient=tk.HORIZONTAL, command=self.orders_tree.xview)
        self.orders_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.orders_tree.grid(row=1, column=0, sticky='nsew')
        vsb.grid(row=1, column=1, sticky='ns')
        hsb.grid(row=2, column=0, sticky='ew')

        orders_frame.grid_rowconfigure(1, weight=1)
        orders_frame.grid_columnconfigure(0, weight=1)
        
        # Bind selezione
        self.orders_tree.bind('<<TreeviewSelect>>', self._on_order_select)
        # Doppio click su una riga "da abbinare" → abbinamento SO + destinazione
        # oppure correzione "già spedito" (FASE 2).
        self.orders_tree.bind('<Double-1>', self._on_order_double_click)
        
        # Tag per colorazione
        self.orders_tree.tag_configure('completed', background='#d4edda')
        self.orders_tree.tag_configure('partial', background='#fff3cd')
        # FASE 2: righe "da abbinare" (magazzino, da abbinare a un SO) = giallo
        # pieno; urgenti la cui merce è già a magazzino = rosso.
        self.orders_tree.tag_configure('da_abbinare', background='#ffe066')
        self.orders_tree.tag_configure('urgent_wh', background='#f5c6cb')
    
    def _create_shipping_rules_section(self, parent):
        """Crea la sezione inferiore per le regole di spedizione"""
        rules_frame = ttk.LabelFrame(parent, text=self.lang.get('shipping_rules', 'Regole di Spedizione'), padding=10)
        parent.add(rules_frame, weight=1)
        
        # Toolbar
        toolbar = ttk.Frame(rules_frame)
        toolbar.pack(fill=tk.X, pady=(0, 5))
        
        self.add_rule_btn = ttk.Button(toolbar, text=self.lang.get('btn_add_rule', 'Aggiungi Regola'), 
                                       command=self._add_shipping_rule, state=tk.DISABLED)
        self.add_rule_btn.pack(side=tk.LEFT, padx=5)
        
        self.edit_rule_btn = ttk.Button(toolbar, text=self.lang.get('btn_edit_rule', 'Modifica'), 
                                        command=self._edit_shipping_rule, state=tk.DISABLED)
        self.edit_rule_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_rule_btn = ttk.Button(toolbar, text=self.lang.get('btn_delete_rule', 'Elimina'), 
                                          command=self._delete_shipping_rule, state=tk.DISABLED)
        self.delete_rule_btn.pack(side=tk.LEFT, padx=5)
        
        # ðŸ†• Bottone esporta Excel
        ttk.Button(toolbar, text=self.lang.get('btn_export_rules_excel', 'Esporta Regole Excel'), 
                  command=self._export_rules_to_excel).pack(side=tk.LEFT, padx=5)
        
        # Label info
        self.rules_info_label = ttk.Label(toolbar, text=self.lang.get('select_order_first', 'Seleziona un ordine sopra'))
        self.rules_info_label.pack(side=tk.LEFT, padx=20)
        
        # Treeview regole
        columns = ('RuleId', 'ProdOrder', 'RequestedOn', 'RequestDateToShip', 'RequestQty', 
                   'RemainOverPO', 'RemainOverRequest', 'ShipTo', 'AddedBy')
        
        self.rules_tree = ttk.Treeview(rules_frame, columns=columns, show='headings', selectmode='browse')
        
        # Nascondi RuleId
        self.rules_tree.column('RuleId', width=0, stretch=False)
        self.rules_tree.heading('RuleId', text='')
        
        
        # Intestazioni colonne
        self.rules_tree.heading('ProdOrder', text=self.lang.get('col_prod_order', 'Ordine Prod.'))
        self.rules_tree.heading('RequestedOn', text=self.lang.get('col_requested_on', 'Richiesto Il'))
        self.rules_tree.heading('RequestDateToShip', text=self.lang.get('col_request_date_to_ship', 'Data Spedizione'))
        self.rules_tree.heading('RequestQty', text=self.lang.get('col_request_qty', 'Qta Richiesta'))
        self.rules_tree.heading('RemainOverPO', text=self.lang.get('col_remain_over_po', 'Rim. su PO'))
        self.rules_tree.heading('RemainOverRequest', text=self.lang.get('col_remain_over_request', 'Rim. su Richiesta'))
        self.rules_tree.heading('ShipTo', text=self.lang.get('col_ship_to', 'Destinazione'))
        self.rules_tree.heading('AddedBy', text=self.lang.get('col_added_by', 'Aggiunto Da'))
        
        # Larghezze colonne (centrate per migliore leggibilità)
        self.rules_tree.column('ProdOrder', width=120, anchor=tk.CENTER)
        self.rules_tree.column('RequestedOn', width=100, anchor=tk.CENTER)
        self.rules_tree.column('RequestDateToShip', width=120, anchor=tk.CENTER)
        self.rules_tree.column('RequestQty', width=90, anchor=tk.CENTER)
        self.rules_tree.column('RemainOverPO', width=90, anchor=tk.CENTER)
        self.rules_tree.column('RemainOverRequest', width=120, anchor=tk.CENTER)
        self.rules_tree.column('ShipTo', width=180, anchor=tk.CENTER)
        self.rules_tree.column('AddedBy', width=120, anchor=tk.CENTER)
        
        scrollbar = ttk.Scrollbar(rules_frame, orient=tk.VERTICAL, command=self.rules_tree.yview)
        self.rules_tree.configure(yscrollcommand=scrollbar.set)
        
        self.rules_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def _load_filter_data(self):
        """Carica i dati per i filtri (clienti e prodotti)"""
        try:
            # Carica clienti
            query_customers = """
            SELECT DISTINCT CustomerName 
            FROM [Traceability_RS].[dyn].[DynamicSaleOrders]
            WHERE CustomerName IS NOT NULL
            ORDER BY CustomerName
            """
            self.db.cursor.execute(query_customers)
            customers = [self.lang.get('all_customers', 'Tutti i Clienti')]
            customers.extend([row.CustomerName for row in self.db.cursor.fetchall()])
            self.customer_filter['values'] = customers
            self.customer_filter.current(0)
            
            # Carica prodotti
            query_products = """
            SELECT DISTINCT ItemCode + ' - ' + ItemName as Product
            FROM [Traceability_RS].[dyn].[DynamicSaleOrders]
            WHERE ItemCode IS NOT NULL
            ORDER BY Product
            """
            self.db.cursor.execute(query_products)
            products = [self.lang.get('all_products', 'Tutti i Prodotti')]
            products.extend([row.Product for row in self.db.cursor.fetchall()])
            self.product_filter['values'] = products
            self.product_filter.current(0)
            
        except Exception as e:
            logger.error(f"Errore caricamento dati filtri: {e}", exc_info=True)
    
    def _get_current_filters(self):
        """Restituisce i filtri correnti"""
        filters = {}
        
        customer = self.customer_filter.get()
        if customer and customer != self.lang.get('all_customers', 'Tutti i Clienti'):
            filters['customer'] = customer
        
        so = self.so_filter.get().strip()
        if so:
            filters['so_number'] = so
        
        po = self.po_filter.get().strip()
        if po:
            filters['production_order'] = po
        
        product = self.product_filter.get()
        if product and product != self.lang.get('all_products', 'Tutti i Prodotti'):
            filters['product'] = product
        
        # Filtro rimanenti
        if self.remain_filter_var.get():
            filters['only_remaining'] = True
        
        # ðŸ†• Date non piÃ¹ incluse automaticamente - solo quando utente applica filtri
        # filters['date_from'] = self.date_from.get_date()
        # filters['date_to'] = self.date_to.get_date()
        
        return filters
    
    
    def _apply_filters(self):
        """Applica i filtri e carica i dati"""
        # ðŸ†• Quando utente clicca "Filtra", includi anche le date
        filters = self._get_current_filters()
        filters['date_from'] = self.date_from.get_date()
        filters['date_to'] = self.date_to.get_date()
        self._load_order_data_with_filters(filters)
    
    def _reset_filters(self):
        """Reset tutti i filtri"""
        self.customer_filter.current(0)
        self.so_filter.delete(0, tk.END)
        self.po_filter.delete(0, tk.END)
        self.product_filter.current(0)
        self.remain_filter_var.set(False)
        self.date_from.set_date(datetime.now() - timedelta(days=180))
        self.date_to.set_date(datetime.now() + timedelta(days=180))
        self._load_order_data()
    
    
    def _load_order_data(self):
        """Carica i dati ordini senza filtri (caricamento iniziale)"""
        self._load_order_data_with_filters({})
    
    def _load_order_data_with_filters(self, filters=None):
        """Carica i dati ordini con la query principale"""
        try:
            if filters is None:
                filters = self._get_current_filters()
            
            query = """
            WITH OrderData AS (
                SELECT o.IDOrder, 
                    d.[SONumber], 
                    o.ordernumber,
                    d.[CustomerName], 
                    d.[ItemCode], 
                    d.[ItemName],
                    d.[ShipDateRequest], 
                    d.[QtyOrder], 
                    po.Qty AS QtyAssigned,
                    sub.NoBoards AS Associate,
                    sub2.NoBoards AS SMT,
                    sub3.NoBoards AS PTHM,
                    sub4.NoBoards AS ICT,
                    sub5.NoBoards AS FCT,
                    sub6.NoBoards AS Coating,
                    sub7.NoBoards AS [Coating Bottom],
                    sub8.NoBoards AS OutOfBox,
                    po.DynamicProductionOrderID
                FROM 
                    [Traceability_RS].[dyn].[DynamicSaleOrders] d
                INNER JOIN 
                    [Traceability_RS].[dyn].[DynamicProductionOrders] po ON d.DynamicSaleOrderId = po.DynamicSaleOrderId 
                INNER JOIN    
                    traceability_rs.dbo.orders o ON po.IdOrder = o.IDOrder 
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 0)) sub
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 1)) sub2
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 4)) sub3
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 102)) sub4
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 103)) sub5
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 132)) sub6
                OUTER APPLY 
                    (SELECT NoBoards FROM traceability_rs.[dbo].[QuantitaProdottaPerFase](o.IDOrder, 135)) sub7
                OUTER APPLY 
                    (select NoBoards from traceability_rs.[dbo].[GetOrderPhaseStatus](o.IDOrder,9)) as sub8
            """
            
            # Aggiungi filtri dinamici
            params = []
            where_clauses = []
            
            # Filtro data (opzionale)
            if 'date_from' in filters and 'date_to' in filters:
                where_clauses.append("d.ShipDateRequest BETWEEN ? AND ?")
                params.extend([filters['date_from'], filters['date_to']])
            
            if 'customer' in filters:
                where_clauses.append("d.CustomerName = ?")
                params.append(filters['customer'])
            
            if 'so_number' in filters:
                where_clauses.append("d.SONumber LIKE ?")
                params.append(f"%{filters['so_number']}%")
            
            if 'production_order' in filters:
                where_clauses.append("o.ordernumber LIKE ?")
                params.append(f"%{filters['production_order']}%")
            
            if 'product' in filters:
                where_clauses.append("(d.ItemCode + ' - ' + d.ItemName) LIKE ?")
                params.append(f"%{filters['product']}%")
            
            # Aggiungi WHERE solo se ci sono condizioni
            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)
            
            query += """
            )
            
            SELECT  
                IDOrder,
                [CustomerName],
                [SONumber] as SaleOrder, 
                ordernumber As ProductionOrder,
                [ItemCode], 
                [ItemName],
                [ShipDateRequest], 
                [QtyOrder], 
                QtyAssigned,
                Associate,
                SMT,
                PTHM,
                ICT,
                FCT,      
                Coating,
                [Coating Bottom],
                OutOfBox,
                [QtyOrder] - ISNULL(OutOfBox, 0) AS Remain,
                DynamicProductionOrderID
            FROM 
                OrderData
            """
            
            # Filtro rimanenti (applicato dopo la CTE)
            if 'only_remaining' in filters:
                query += " WHERE ([QtyOrder] - ISNULL(OutOfBox, 0)) > 0"
            
            query += """
            ORDER BY 
                SONumber
            """
            
            self.db.cursor.execute(query, params)
            rows = self.db.cursor.fetchall()

            # Ordini finiti a magazzino non ancora spediti (FASE 1).
            # Calcolati PRIMA del ciclo per: (a) colorare di rosso gli ordini
            # esistenti che sono urgenti e già a magazzino; (b) iniettare in coda
            # le righe gialle "da abbinare" con la quota residua.
            self._warehouse_rows = self._fetch_warehouse_available()
            red_orders = {str(r['IDOrder']) for r in self._warehouse_rows
                          if r['IsUrgent']}

            # Pulisci treeview
            for item in self.orders_tree.get_children():
                self.orders_tree.delete(item)

            # Popola treeview
            for row in rows:
                ship_date = row.ShipDateRequest.strftime('%d/%m/%Y') if row.ShipDateRequest else ''
                remain = row.Remain if row.Remain else 0

                # Rosso: ordine urgente la cui merce è già a magazzino (§4bis);
                # altrimenti verde (completo) / giallo tenue (parziale).
                if str(row.IDOrder) in red_orders:
                    tag = 'urgent_wh'
                else:
                    tag = 'completed' if remain == 0 else 'partial'

                self.orders_tree.insert('', tk.END, values=(
                    row.IDOrder,
                    row.CustomerName,
                    row.SaleOrder,
                    row.ProductionOrder,
                    row.ItemCode,
                    row.ItemName,
                    ship_date,
                    row.QtyOrder,
                    row.QtyAssigned,
                    row.Associate,
                    row.SMT,
                    row.PTHM,
                    row.ICT,
                    row.FCT,
                    row.Coating,
                    getattr(row, 'Coating Bottom', 0),  # Handle space in column name
                    row.OutOfBox,
                    remain
                ), tags=(tag,))

            # Inietta le righe "da abbinare" (gialle): ordini a magazzino con
            # quota residua > 0 non abbinata. SO/cliente vuoti, editabili
            # dall'operatore (FASE 2 interazione).
            n_inject = 0
            for w in self._warehouse_rows:
                if w['Residua'] <= 0:
                    continue
                n_inject += 1
                last_wh = w['LastWhDate'].strftime('%d/%m/%Y') if w['LastWhDate'] else ''
                self.orders_tree.insert('', tk.END, values=(
                    w['IDOrder'],
                    '',                      # Customer (da abbinare)
                    '',                      # SaleOrder (vuoto, editabile)
                    w['OrderNumber'],        # ProductionOrder
                    w['ProductCode'],        # ItemCode = ProductCode
                    '',                      # ItemName
                    last_wh,                 # ShipDate = ultima data a magazzino
                    '', '', '', '', '', '', '', '', '',  # QtyOrder..OutOfBox n/a
                    w['Residua']             # Remain = disponibile da abbinare
                ), tags=('da_abbinare',))

            logger.info(
                "Caricati %d ordini + %d righe da magazzino (%d gialle 'da abbinare', "
                "%d rosse urgenti-a-magazzino)",
                len(rows), len(self._warehouse_rows), n_inject, len(red_orders))

        except Exception as e:
            logger.error(f"Errore caricamento dati ordini: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore caricamento dati: {e}",
                parent=self
            )

    def _fetch_warehouse_available(self):
        """FASE 1 — Ordini finiti e versati a magazzino non ancora spediti.

        Riconcilia i dati di magazzino (LogApiDynamics + WarehouseFinish.Packing)
        con l'abbinamento dei planner e con lo spedito v2, sulla chiave ordine di
        produzione `IDOrder` (deterministica: un ordine → un prodotto). Per ogni
        ordine, nella finestra `MAX(data_inizio_spedizioni, date_from) .. date_to`:

            Residua = versato a magazzino − già abbinato (DynamicProductionOrders)
                      − già spedito (ShipmentPallets nella finestra)

        Restituisce una lista di dict con: IDOrder, ProductCode, OrderNumber,
        WhQty, MatchedQty, ShippedQty, AdjQty, Residua, IsUrgent, LastWhDate.
        Residua > 0 = riga "da abbinare" (gialla). IsUrgent + a magazzino = rossa.
        Il floor `data_inizio_spedizioni` è il punto zero: prima si assume spedito.
        AdjQty = correzione manuale "già spedito" (ledger di allineamento
        transitorio dyn.WarehouseShippedAdjustments), sottratta dal residuo.
        La parte correzione è inclusa solo se la tabella esiste (resiliente).
        """
        has_adj = self._adjustments_table_exists()
        adj_cte = ("""        ,ADJ AS (
            SELECT IDOrder, SUM(Qty) AS AdjQty
            FROM traceability_rs.dyn.WarehouseShippedAdjustments GROUP BY IDOrder
        )""" if has_adj else "")
        adj_join = ("        LEFT JOIN ADJ a ON a.IDOrder = wh.IDOrder\n" if has_adj else "")
        adj_sel = ("ISNULL(a.AdjQty, 0)" if has_adj else "0")

        sql = f"""
        DECLARE @from date = ?, @to date = ?;
        DECLARE @floor date;
        SELECT @floor = TRY_CAST(Value AS date)
        FROM traceability_rs.dbo.Settings WHERE atribute = 'data_inizio_spedizioni';
        DECLARE @start date = CASE WHEN @floor IS NULL OR @from > @floor THEN @from ELSE @floor END;

        ;WITH WH AS (
            SELECT o.IDOrder, p.ProductCode, o.OrderNumber,
                   SUM(CAST(JSON_VALUE(j.value, '$.RealValue') AS int)) AS WhQty,
                   CAST(MAX(L.CurrentDate) AS date) AS LastWhDate
            FROM traceability_rs.dbo.LogApiDynamics L
            INNER JOIN traceability_rs.dbo.Orders o
                ON o.OrderNumber = JSON_VALUE(L.MessageSend, '$.Message.Reference')
            INNER JOIN traceability_rs.dbo.Products p ON p.IDProduct = o.IDProduct
            CROSS APPLY OPENJSON(L.MessageSend,
                '$.Message.KeyValue.ListValue[0].ListValue[0].ListValue') j
            WHERE L.EndPointName = 'ProdFinishedGoods'
              AND CAST(L.CurrentDate AS date) BETWEEN @start AND @to
              AND JSON_VALUE(j.value, '$.Key') = 'GoodQty'
            GROUP BY o.IDOrder, p.ProductCode, o.OrderNumber
        ),
        MATCHED AS (
            SELECT IdOrder, SUM(Qty) AS MatchedQty
            FROM traceability_rs.dyn.DynamicProductionOrders GROUP BY IdOrder
        ),
        SHIPPED AS (
            SELECT po.IdOrder, SUM(sp.ConfirmedQty) AS ShippedQty
            FROM traceability_rs.dyn.ShipmentPallets sp
            INNER JOIN traceability_rs.dyn.DynamicProductionOrders po
                ON po.DynamicProductionOrderID = sp.DynamicProductionOrderID
            INNER JOIN traceability_rs.dyn.Shipments s ON s.ShipmentId = sp.ShipmentId
            WHERE s.ShipmentDate >= @start GROUP BY po.IdOrder
        ),
        URGENT AS (
            SELECT DISTINCT po.IdOrder
            FROM traceability_rs.dyn.DynamicShippingRules r
            INNER JOIN traceability_rs.dyn.DynamicProductionOrders po
                ON po.DynamicProductionOrderID = r.DynamicProductionOrderID
        ){adj_cte}
        SELECT wh.IDOrder, wh.ProductCode, wh.OrderNumber, wh.WhQty, wh.LastWhDate,
               ISNULL(m.MatchedQty, 0) AS MatchedQty,
               ISNULL(sh.ShippedQty, 0) AS ShippedQty,
               {adj_sel} AS AdjQty,
               wh.WhQty - ISNULL(m.MatchedQty, 0) - ISNULL(sh.ShippedQty, 0)
                        - {adj_sel} AS Residua,
               CASE WHEN u.IdOrder IS NOT NULL THEN 1 ELSE 0 END AS IsUrgent
        FROM WH wh
        LEFT JOIN MATCHED m ON m.IdOrder = wh.IDOrder
        LEFT JOIN SHIPPED sh ON sh.IdOrder = wh.IDOrder
        LEFT JOIN URGENT u ON u.IdOrder = wh.IDOrder
{adj_join}        ORDER BY IsUrgent DESC, Residua DESC
        """
        try:
            date_from = self.date_from.get_date()
            date_to = self.date_to.get_date()
            self.db.cursor.execute(sql, (date_from, date_to))
            out = []
            for r in self.db.cursor.fetchall():
                out.append({
                    'IDOrder': r.IDOrder,
                    'ProductCode': r.ProductCode,
                    'OrderNumber': r.OrderNumber,
                    'WhQty': int(r.WhQty or 0),
                    'MatchedQty': int(r.MatchedQty or 0),
                    'ShippedQty': int(r.ShippedQty or 0),
                    'AdjQty': int(r.AdjQty or 0),
                    'Residua': int(r.Residua or 0),
                    'IsUrgent': bool(r.IsUrgent),
                    'LastWhDate': r.LastWhDate,
                })
            return out
        except Exception as e:
            logger.error(f"Errore lettura ordini a magazzino (FASE 1): {e}", exc_info=True)
            return []

    def _adjustments_table_exists(self):
        """True se il ledger dyn.WarehouseShippedAdjustments è già stato creato."""
        try:
            self.db.cursor.execute(
                "SELECT OBJECT_ID('Traceability_RS.dyn.WarehouseShippedAdjustments')")
            return self.db.cursor.fetchone()[0] is not None
        except Exception:
            return False

    def _add_shipped_adjustment(self, id_order, product_code, qty, note):
        """Registra una correzione 'già spedito' per un ordine di produzione.

        Append-only: la quantità dichiarata si somma alle correzioni esistenti e
        viene sottratta dalla disponibilità a magazzino. Richiede la tabella
        (creata da add_warehouse_shipped_adjustments.sql)."""
        self.db.cursor.execute(
            """INSERT INTO Traceability_RS.dyn.WarehouseShippedAdjustments
               (IDOrder, ProductCode, Qty, Note, AdjustedByUser, AdjustedAt)
               VALUES (?, ?, ?, ?, ?, GETDATE())""",
            (id_order, product_code, qty, note, self.user_name))
        self.db.conn.commit()
        logger.info("Correzione 'già spedito' IDOrder=%s qty=%s utente=%s",
                    id_order, qty, self.user_name)

    def _on_order_select(self, event):
        """Gestisce la selezione di un ordine"""
        selection = self.orders_tree.selection()
        if not selection:
            self.current_order_id = None
            self.current_production_order_id = None
            self.current_remain = 0
            self.add_rule_btn.config(state=tk.DISABLED)
            self.wh_match_btn.config(state=tk.DISABLED)
            self.rules_info_label.config(text=self.lang.get('select_order_first', 'Seleziona un ordine sopra'))
            # Pulisci regole
            for item in self.rules_tree.get_children():
                self.rules_tree.delete(item)
            return
        
        
        tags = self.orders_tree.item(selection[0], 'tags')
        values = self.orders_tree.item(selection[0], 'values')
        self.current_order_id = values[0]

        # Riga "da abbinare" (ordine a magazzino senza SO): non ha regole di
        # spedizione né DynamicProductionOrderID. Si abbina con doppio click.
        if 'da_abbinare' in tags:
            self.current_production_order_id = None
            self.current_remain = int(values[17]) if values[17] else 0
            self.add_rule_btn.config(state=tk.DISABLED)
            self.wh_match_btn.config(state=tk.NORMAL)  # abilita bottone abbina
            for item in self.rules_tree.get_children():
                self.rules_tree.delete(item)
            self.rules_info_label.config(
                text=self.lang.get(
                    'wh_to_match_hint',
                    'Ordine a magazzino da abbinare — doppio click (o bottone Abbina) '
                    'per abbinare a un ordine di vendita o correggere il già spedito'))
            return

        self.wh_match_btn.config(state=tk.DISABLED)  # non è una riga magazzino
        # ðŸ†• Usa valore assoluto per rimanenti (può essere negativo nella form)
        self.current_remain = abs(int(values[17])) if values[16] else 0

        # Ottieni DynamicProductionOrderID
        try:
            query = """
            SELECT po.DynamicProductionOrderID
            FROM [Traceability_RS].[dyn].[DynamicProductionOrders] po
            WHERE po.IdOrder = ?
            """
            self.db.cursor.execute(query, (self.current_order_id,))
            row = self.db.cursor.fetchone()
            if row:
                self.current_production_order_id = row.DynamicProductionOrderID
            else:
                self.current_production_order_id = None
        except Exception as e:
            logger.error(f"Errore recupero DynamicProductionOrderID: {e}")
            self.current_production_order_id = None
        
        # Abilita bottone aggiungi
        self.add_rule_btn.config(state=tk.NORMAL)
        
        # Aggiorna label info
        self.rules_info_label.config(
            text=f"{self.lang.get('selected_order', 'Ordine selezionato')}: {values[3]} | {self.lang.get('remain', 'Rimanenti')}: {self.current_remain}"
        )
        
        # Carica regole
        self._load_shipping_rules()

    def _on_order_double_click(self, event):
        """Doppio click su una riga 'da abbinare' → abbinamento/correzione."""
        selection = self.orders_tree.selection()
        if selection and 'da_abbinare' in self.orders_tree.item(selection[0], 'tags'):
            self._open_warehouse_match()

    def _open_warehouse_match(self):
        """Apre l'abbinamento SO + destinazione / correzione 'già spedito' per la
        riga gialla selezionata. Condiviso da doppio click e bottone."""
        selection = self.orders_tree.selection()
        if not selection:
            return
        if 'da_abbinare' not in self.orders_tree.item(selection[0], 'tags'):
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('wh_select_yellow',
                              'Selezionare una riga gialla "da abbinare".'),
                parent=self)
            return
        id_order = self.orders_tree.item(selection[0], 'values')[0]
        wh = next((w for w in self._warehouse_rows
                   if str(w['IDOrder']) == str(id_order)), None)
        if not wh:
            return
        WarehouseMatchDialog(self, self.db, self.lang, self.user_name, wh,
                             callback=self._load_order_data)

    def _load_shipping_rules(self):
        """Carica le regole di spedizione per l'ordine selezionato"""
        if not self.current_production_order_id:
            return
        
        try:
            query = """
            SELECT R.DybamicShippingRuleId, po.ordernumber as Prod_Order, r.datesys as RequestedOn, 
                   FORMAT(R.DateToship, 'dd.MM.yyyy ''at'' hh:mm') as RequestDateToShip, R.QtyToShip AS RequestQty, 
                   S.QtyOrder - K.Packet AS RemainOverPO,
                   r.QtyToShip - p.Noboards As RemainOverRequest, R.ShipTo, R.AddBayUser
            FROM [Traceability_RS].[dyn].[DynamicShippingRules] R
            INNER JOIN Traceability_RS.dyn.DynamicProductionOrders O on O.DynamicProductionOrderID = R.DynamicProductionOrderID
            INNER JOIN TRACEABILITY_RS.DYN.DynamicSaleOrders S on s.DynamicSaleOrderId=o.DynamicSaleOrderId
            OUTER APPLY
                (select NoBoards as Packet from traceability_rs.[dbo].[GetOrderPhaseStatus](o.idorder,9)) as K
            OUTER APPLY     
                Traceability_rs.dbo.da_eusta_fn_GetPackedBoards(o.idorder, 920, NULL,0) AS P
            inner join traceability_rs.dbo.Orders PO on o.idorder=po.idorder
            WHERE R.DynamicProductionOrderID = ?
            ORDER BY R.DateToship
            """
            
            self.db.cursor.execute(query, (self.current_production_order_id,))
            rows = self.db.cursor.fetchall()
            
            # Pulisci treeview
            for item in self.rules_tree.get_children():
                self.rules_tree.delete(item)
            
            
            # Popola treeview
            for row in rows:
                # Formatta date (RequestedOn è datetime, RequestDateToShip è già formattata da SQL)
                requested_on = row.RequestedOn.strftime('%d/%m/%Y') if row.RequestedOn else ''
                request_date_to_ship = row.RequestDateToShip or ''  # Già formattata da SQL con FORMAT()
                ship_to = row.ShipTo or ''
                
                self.rules_tree.insert('', tk.END, values=(
                    row.DybamicShippingRuleId,
                    row.Prod_Order or '',
                    requested_on,
                    request_date_to_ship,
                    row.RequestQty or 0,
                    row.RemainOverPO or 0,
                    row.RemainOverRequest or 0,
                    ship_to,
                    row.AddBayUser or ''
                ))
            
            # Abilita/disabilita bottoni edit/delete
            self.edit_rule_btn.config(state=tk.NORMAL if rows else tk.DISABLED)
            self.delete_rule_btn.config(state=tk.NORMAL if rows else tk.DISABLED)
            
        except Exception as e:
            logger.error(f"Errore caricamento regole spedizione: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore: {e}",
                parent=self
            )
    
    def _add_shipping_rule(self):
        """Apre dialog per aggiungere una nuova regola di spedizione"""
        if not self.current_production_order_id:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_order_first', 'Seleziona un ordine'),
                parent=self
            )
            return
        
        ShippingRuleDialog(self, self.db, self.lang, self.user_name, 
                          self.current_production_order_id, self.current_remain, 
                          callback=self._load_shipping_rules)
    
    def _edit_shipping_rule(self):
        """Modifica la regola selezionata"""
        selection = self.rules_tree.selection()
        if not selection:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_rule_first', 'Seleziona una regola'),
                parent=self
            )
            return
        
        values = self.rules_tree.item(selection[0], 'values')
        rule_id = values[0]
        
        ShippingRuleDialog(self, self.db, self.lang, self.user_name,
                          self.current_production_order_id, self.current_remain,
                          rule_id=rule_id, callback=self._load_shipping_rules)
    
    def _delete_shipping_rule(self):
        """Elimina la regola selezionata"""
        selection = self.rules_tree.selection()
        if not selection:
            return
        
        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            self.lang.get('confirm_delete_rule', 'Eliminare la regola di spedizione?'),
            parent=self
        ):
            return
        
        values = self.rules_tree.item(selection[0], 'values')
        rule_id = values[0]
        
        try:
            query = "DELETE FROM [Traceability_RS].[dyn].[DynamicShippingRules] WHERE DybamicShippingRuleId = ?"
            self.db.cursor.execute(query, (rule_id,))
            self.db.conn.commit()
            
            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                self.lang.get('rule_deleted', 'Regola eliminata con successo'),
                parent=self
            )
            
            self._load_shipping_rules()
            
        except Exception as e:
            logger.error(f"Errore eliminazione regola: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore: {e}",
                parent=self
            )
    
    def _export_to_excel(self):
        """Esporta i dati in Excel con vista gerarchica: ordini + regole indentate"""
        try:
            # Crea directory se non esiste
            temp_dir = r'c:\Temp'
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            # Nome file con timestamp
            filename = f"DynamicShipping_Complete_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            # Crea workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vista Completa"
            
            # Stili
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            order_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            rule_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Ottieni colonne ordini (escludi IDOrder)
            order_columns = self.orders_tree['columns'][1:]
            # Ottieni colonne regole (escludi RuleId)
            rule_columns = self.rules_tree['columns'][1:]
            
            # Crea intestazioni combinate
            current_col = 1
            
            # Intestazioni ordini
            for col in order_columns:
                cell = ws.cell(row=1, column=current_col)
                cell.value = self.orders_tree.heading(col)['text']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                current_col += 1
            
            # Intestazioni regole (con prefisso per chiarezza)
            for col in rule_columns:
                cell = ws.cell(row=1, column=current_col)
                cell.value = f"↳ {self.rules_tree.heading(col)['text']}"
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                current_col += 1
            
            # Scrivi dati gerarchici
            excel_row = 2
            
            for item_id in self.orders_tree.get_children():
                # Ottieni dati ordine
                order_values = self.orders_tree.item(item_id, 'values')
                order_id = order_values[0]  # IDOrder
                order_data = order_values[1:]  # Dati visibili
                
                # Scrivi riga ordine
                current_col = 1
                for value in order_data:
                    cell = ws.cell(row=excel_row, column=current_col)
                    cell.value = value
                    cell.border = border
                    cell.fill = order_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_col += 1
                
                excel_row += 1
                
                # Carica regole per questo ordine
                try:
                    # Trova il DynamicProductionOrderID per questo ordine
                    query_po_id = """
                    SELECT DynamicProductionOrderID 
                    FROM [Traceability_RS].[dyn].[DynamicProductionOrders]
                    WHERE idorder = ?
                    """
                    self.db.cursor.execute(query_po_id, (order_id,))
                    po_result = self.db.cursor.fetchone()
                    
                    if po_result and po_result.DynamicProductionOrderID:
                        # Carica regole
                        query_rules = """
                        SELECT R.DybamicShippingRuleId, po.ordernumber as Prod_Order, r.datesys as RequestedOn, 
                               FORMAT(R.DateToship, 'dd.MM.yyyy ''at'' hh:mm') as RequestDateToShip, R.QtyToShip AS RequestQty, 
                               S.QtyOrder - K.Packet AS RemainOverPO,
                               r.QtyToShip - p.Noboards As RemainOverRequest, R.ShipTo, R.AddBayUser
                        FROM [Traceability_RS].[dyn].[DynamicShippingRules] R
                        INNER JOIN Traceability_RS.dyn.DynamicProductionOrders O on O.DynamicProductionOrderID = R.DynamicProductionOrderID
                        INNER JOIN TRACEABILITY_RS.DYN.DynamicSaleOrders S on s.DynamicSaleOrderId=o.DynamicSaleOrderId
                        OUTER APPLY (select NoBoards as Packet from traceability_rs.[dbo].[GetOrderPhaseStatus](o.idorder,9)) as K
                        OUTER APPLY Traceability_rs.dbo.da_eusta_fn_GetPackedBoards(o.idorder, 920, NULL,0) AS P
                        inner join traceability_rs.dbo.Orders PO on o.idorder=po.idorder
                        WHERE R.DynamicProductionOrderID = ?
                        ORDER BY R.DateToship
                        """
                        
                        self.db.cursor.execute(query_rules, (po_result.DynamicProductionOrderID,))
                        rules = self.db.cursor.fetchall()
                        
                        # Scrivi regole indentate
                        for rule in rules:
                            # Formatta date
                            requested_on = rule.RequestedOn.strftime('%d/%m/%Y') if rule.RequestedOn else ''
                            request_date_to_ship = rule.RequestDateToShip or ''
                            
                            rule_data = [
                                rule.Prod_Order or '',
                                requested_on,
                                request_date_to_ship,
                                rule.RequestQty or 0,
                                rule.RemainOverPO or 0,
                                rule.RemainOverRequest or 0,
                                rule.ShipTo or '',
                                rule.AddBayUser or ''
                            ]
                            
                            # Lascia vuote le colonne ordine, scrivi solo nelle colonne regole
                            current_col = 1
                            # Salta colonne ordine
                            for _ in order_data:
                                cell = ws.cell(row=excel_row, column=current_col)
                                cell.value = ""
                                cell.border = border
                                cell.fill = rule_fill
                                current_col += 1
                            
                            # Scrivi dati regola
                            for value in rule_data:
                                cell = ws.cell(row=excel_row, column=current_col)
                                cell.value = value
                                cell.border = border
                                cell.fill = rule_fill
                                cell.alignment = Alignment(horizontal='center', vertical='center', indent=1)
                                current_col += 1
                            
                            excel_row += 1
                
                except Exception as e:
                    logger.error(f"Errore caricamento regole per ordine {order_id}: {e}")
            
            # Auto-size colonne
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze prima riga
            ws.freeze_panes = 'A2'
            
            # Salva
            wb.save(filepath)
            
            # Chiedi se aprire il file
            if messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"{self.lang.get('file_saved', 'File salvato in')}: {filepath}\n\n{self.lang.get('open_file_question', 'Vuoi aprire il file?')}",
                parent=self
            ):
                try:
                    os.startfile(filepath)
                except Exception as e:
                    logger.error(f"Errore apertura file: {e}")
                    messagebox.showerror(
                        self.lang.get('error', 'Errore'),
                        f"Impossibile aprire il file: {e}",
                        parent=self
                    )
            
        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore esportazione: {e}",
                parent=self
            )
    
    def _export_rules_to_excel(self):
        "Esporta tutte le regole di spedizione in Excel"
        try:
            # Query per ottenere tutte le regole con informazioni ordini
            query = """
            SELECT 
                d.CustomerName,
                d.SONumber,
                o.ordernumber AS ProductionOrder,
                d.ItemCode,
                d.ItemName,
                d.QtyOrder,
                sr.QtyToShip,
                sr.DateToship,
                sr.ShipTo,
                sr.AddBayUser,
                sr.DateOut
            FROM 
                [Traceability_RS].[dyn].[DynamicShippingRules] sr
            INNER JOIN 
                [Traceability_RS].[dyn].[DynamicProductionOrders] po ON sr.DynamicProductionOrderID = po.DynamicProductionOrderID
            INNER JOIN 
                [Traceability_RS].[dyn].[DynamicSaleOrders] d ON po.DynamicSaleOrderId = d.DynamicSaleOrderId
            INNER JOIN 
                traceability_rs.dbo.orders o ON po.IdOrder = o.IDOrder
            ORDER BY 
                sr.DateToship, d.SONumber
            """
            
            self.db.cursor.execute(query)
            rows = self.db.cursor.fetchall()
            
            if not rows:
                messagebox.showinfo(
                    self.lang.get('info', 'Informazione'),
                    self.lang.get('no_rules_to_export', 'Nessuna regola di spedizione da esportare'),
                    parent=self
                )
                return
            
            # Crea directory se non esiste
            temp_dir = r'c:\Temp'
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            # Nome file con timestamp
            filename = f"ShippingRules_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            # Crea workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Regole Spedizione"
            
            # Stili
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Intestazioni
            headers = [
                'Cliente', 'Ordine Vendita', 'Ordine Produzione', 'Codice', 'Prodotto',
                'Qtà Ordine', 'Qtà da Spedire', 'Data Spedizione', 'Ora', 'Destinazione',
                'Aggiunto Da', 'Data Inserimento'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Dati
            for row_idx, row in enumerate(rows, 2):
                date_to_ship = row.DateToship.strftime('%d/%m/%Y') if row.DateToship else ''
                time_to_ship = row.DateToship.strftime('%H:%M') if row.DateToship else ''
                date_out = row.DateOut.strftime('%d/%m/%Y %H:%M') if row.DateOut else ''
                
                values = [
                    row.CustomerName,
                    row.SONumber,
                    row.ProductionOrder,
                    row.ItemCode,
                    row.ItemName,
                    row.QtyOrder,
                    row.QtyToShip,
                    date_to_ship,
                    time_to_ship,
                    row.ShipTo or '',
                    row.AddBayUser or '',
                    date_out
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-size colonne
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze prima riga
            ws.freeze_panes = 'A2'
            
            # Salva
            wb.save(filepath)
            
            # Chiedi se aprire il file
            if messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"{self.lang.get('file_saved', 'File salvato in')}: {filepath}\n\n{self.lang.get('open_file_question', 'Vuoi aprire il file?')}",
                parent=self
            ):
                try:
                    os.startfile(filepath)
                except Exception as e:
                    logger.error(f"Errore apertura file: {e}")
                    messagebox.showerror(
                        self.lang.get('error', 'Errore'),
                        f"Impossibile aprire il file: {e}",
                        parent=self
                    )
            
        except Exception as e:
            logger.error(f"Errore esportazione regole Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore esportazione: {e}",
                parent=self
            )


class ShippingRuleDialog(tk.Toplevel):
    """Dialog per aggiungere/modificare una regola di spedizione"""
    
    def __init__(self, master, db, lang, user_name, production_order_id, max_qty, rule_id=None, callback=None):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.production_order_id = production_order_id
        self.max_qty = max_qty
        self.rule_id = rule_id
        self.callback = callback
        
        self.title(self.lang.get('add_rule' if not rule_id else 'edit_rule', 
                                 'Aggiungi Regola' if not rule_id else 'Modifica Regola'))
        self.geometry('450x300')
        self.transient(master)
        self.grab_set()
        
        self._create_widgets()
        
        if rule_id:
            self._load_rule_data()
    
    def _create_widgets(self):
        """Crea i widget del dialog"""
        main_frame = ttk.Frame(self, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # QtÃ  da spedire
        ttk.Label(main_frame, text=self.lang.get('qty_to_ship', 'QuantitÃ  da Spedire:')).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.qty_entry = ttk.Entry(main_frame, width=15)
        self.qty_entry.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Label(main_frame, text=f"(Max: {self.max_qty})", foreground='gray').grid(row=0, column=2, sticky=tk.W, padx=(5, 0))
        
        # Data spedizione
        ttk.Label(main_frame, text=self.lang.get('date_to_ship', 'Data Spedizione:')).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.date_entry = DateEntry(
            main_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='it_IT'
        )
        self.date_entry.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        
        # ðŸ†• Ora spedizione
        ttk.Label(main_frame, text=self.lang.get('time_to_ship', 'Ora:')).grid(row=2, column=0, sticky=tk.W, pady=5)
        time_frame = ttk.Frame(main_frame)
        time_frame.grid(row=2, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        
        self.hour_spinbox = ttk.Spinbox(time_frame, from_=0, to=23, width=5, format='%02.0f')
        self.hour_spinbox.set('08')
        self.hour_spinbox.pack(side=tk.LEFT)
        
        ttk.Label(time_frame, text=':').pack(side=tk.LEFT, padx=2)
        
        self.minute_spinbox = ttk.Spinbox(time_frame, from_=0, to=59, width=5, format='%02.0f')
        self.minute_spinbox.set('00')
        self.minute_spinbox.pack(side=tk.LEFT)
        
        # ðŸ†• ShipTo
        ttk.Label(main_frame, text=self.lang.get('ship_to', 'Destinazione:')).grid(row=3, column=0, sticky=tk.W, pady=5)
        self.shipto_combo = ttk.Combobox(main_frame, width=25, state='readonly')
        self.shipto_combo['values'] = ['Normal Shipment', 'Direct to final Customer']
        self.shipto_combo.current(0)
        self.shipto_combo.grid(row=3, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        
        # Bottoni
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=(20, 0))
        
        ttk.Button(btn_frame, text=self.lang.get('btn_save', 'Salva'), command=self._save).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=self.lang.get('btn_cancel', 'Annulla'), command=self.destroy).pack(side=tk.LEFT, padx=5)
    
    def _load_rule_data(self):
        """Carica i dati della regola esistente"""
        try:
            query = "SELECT QtyToShip, DateToship, ShipTo FROM [Traceability_RS].[dyn].[DynamicShippingRules] WHERE DybamicShippingRuleId = ?"
            self.db.cursor.execute(query, (self.rule_id,))
            row = self.db.cursor.fetchone()
            
            if row:
                self.qty_entry.insert(0, str(row.QtyToShip))
                if row.DateToship:
                    # ðŸ†• Estrai data e ora
                    self.date_entry.set_date(row.DateToship.date())
                    self.hour_spinbox.set(f"{row.DateToship.hour:02d}")
                    self.minute_spinbox.set(f"{row.DateToship.minute:02d}")
                # ðŸ†• Carica ShipTo
                if row.ShipTo:
                    if row.ShipTo in self.shipto_combo['values']:
                        self.shipto_combo.set(row.ShipTo)
                    else:
                        self.shipto_combo.current(0)
        except Exception as e:
            logger.error(f"Errore caricamento regola: {e}", exc_info=True)
    
    def _save(self):
        """Salva la regola"""
        try:
            # Validazione
            qty_str = self.qty_entry.get().strip()
            if not qty_str:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('qty_required', 'Inserire la quantitÃ '),
                    parent=self
                )
                return
            
            try:
                qty = int(qty_str)
            except ValueError:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('qty_invalid', 'QuantitÃ  non valida'),
                    parent=self
                )
                return
            
            if qty <= 0:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('qty_positive', 'La quantitÃ  deve essere maggiore di zero'),
                    parent=self
                )
                return
            
            if qty > self.max_qty:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    f"{self.lang.get('qty_exceeds', 'La quantitÃ  supera il massimo disponibile')}: {self.max_qty}",
                    parent=self
                )
                return
            
            
            # ðŸ†• Combina data + ora
            date_to_ship = self.date_entry.get_date()
            try:
                hour = int(self.hour_spinbox.get())
                minute = int(self.minute_spinbox.get())
            except ValueError:
                hour = 8
                minute = 0
            
            # Crea datetime combinando data e ora
            from datetime import datetime, time
            datetime_to_ship = datetime.combine(date_to_ship, time(hour, minute))
            
            # ðŸ†• Ottieni ShipTo
            ship_to = self.shipto_combo.get()
            
            if self.rule_id:
                # UPDATE
                query = """
                UPDATE [Traceability_RS].[dyn].[DynamicShippingRules]
                SET QtyToShip = ?, DateToship = ?, ShipTo = ?
                WHERE DybamicShippingRuleId = ?
                """
                self.db.cursor.execute(query, (qty, datetime_to_ship, ship_to, self.rule_id))
            else:
                # INSERT
                query = """
                INSERT INTO [Traceability_RS].[dyn].[DynamicShippingRules]
                (DynamicProductionOrderID, QtyToShip, DateToship, ShipTo, AddBayUser, DateOut)
                VALUES (?, ?, ?, ?, ?, GETDATE())
                """
                self.db.cursor.execute(query, (self.production_order_id, qty, datetime_to_ship, ship_to, self.user_name))
            
            self.db.conn.commit()
            
            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                self.lang.get('rule_saved', 'Regola salvata con successo'),
                parent=self
            )
            
            if self.callback:
                self.callback()
            
            self.destroy()
            
        except Exception as e:
            logger.error(f"Errore salvataggio regola: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore: {e}",
                parent=self
            )


class WarehouseMatchDialog(tk.Toplevel):
    """FASE 2 — Abbina un ordine finito a magazzino a un ordine di vendita
    (stesso prodotto + quantità), imposta la destinazione e crea la regola di
    spedizione; oppure registra una correzione 'già spedito' (allineamento).

    L'ordine a magazzino non ha SO: l'operatore sceglie un SO con lo stesso
    prodotto (ItemCode = ProductCode senza suffisso |N) e residuo disponibile.
    """

    def __init__(self, master, db, lang, user_name, wh_row, callback=None):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.wh = wh_row
        self.callback = callback
        # ItemCode del sales order = ProductCode senza il suffisso |N (versione)
        self.item_code = (wh_row['ProductCode'] or '').split('|')[0]
        self._so_by_label = {}

        L = self.lang.get
        self.title(L('wh_match_title', 'Abbina ordine a magazzino'))
        self.geometry('620x420')
        self.transient(master)
        self.grab_set()
        self._build()
        self._load_candidate_sos()

    def _build(self):
        L = self.lang.get
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        info = (f"{L('production_order', 'Ordine Produzione')}: {self.wh['OrderNumber']}   |   "
                f"{L('product', 'Prodotto')}: {self.wh['ProductCode']}   |   "
                f"{L('wh_available', 'Disponibile a magazzino')}: {self.wh['Residua']}")
        ttk.Label(frm, text=info, font=('', 9, 'bold')).grid(
            row=0, column=0, columnspan=3, sticky='w', pady=(0, 10))

        # Sales order candidato
        ttk.Label(frm, text=L('sale_order', 'Ordine Vendita') + ' *').grid(
            row=1, column=0, sticky='w', pady=4)
        self.so_combo = ttk.Combobox(frm, width=55, state='readonly')
        self.so_combo.grid(row=1, column=1, columnspan=2, sticky='w', pady=4)

        # Quantità
        ttk.Label(frm, text=L('qty_to_ship', 'Quantità da Spedire') + ' *').grid(
            row=2, column=0, sticky='w', pady=4)
        self.qty_entry = ttk.Entry(frm, width=14)
        self.qty_entry.grid(row=2, column=1, sticky='w', pady=4)
        self.qty_hint = ttk.Label(frm, text='', foreground='gray')
        self.qty_hint.grid(row=2, column=2, sticky='w', padx=6)

        # Data + ora
        ttk.Label(frm, text=L('date_to_ship', 'Data Spedizione') + ' *').grid(
            row=3, column=0, sticky='w', pady=4)
        self.date_entry = DateEntry(frm, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
        self.date_entry.grid(row=3, column=1, sticky='w', pady=4)
        tf = ttk.Frame(frm)
        tf.grid(row=3, column=2, sticky='w')
        self.hour_sb = ttk.Spinbox(tf, from_=0, to=23, width=4, format='%02.0f')
        self.hour_sb.set('08'); self.hour_sb.pack(side=tk.LEFT)
        ttk.Label(tf, text=':').pack(side=tk.LEFT)
        self.min_sb = ttk.Spinbox(tf, from_=0, to=59, width=4, format='%02.0f')
        self.min_sb.set('00'); self.min_sb.pack(side=tk.LEFT)

        # Destinazione
        ttk.Label(frm, text=L('ship_to', 'Destinazione') + ' *').grid(
            row=4, column=0, sticky='w', pady=4)
        self.shipto_combo = ttk.Combobox(
            frm, width=28, state='readonly',
            values=['Normal Shipment', 'Direct to final Customer'])
        self.shipto_combo.current(0)
        self.shipto_combo.grid(row=4, column=1, columnspan=2, sticky='w', pady=4)

        # Bottoni
        bar = ttk.Frame(frm)
        bar.grid(row=6, column=0, columnspan=3, pady=(18, 0), sticky='w')
        ttk.Button(bar, text=L('wh_btn_match', '✅ Abbina e crea regola'),
                   command=self._match_and_rule).pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text=L('wh_btn_shipped', '📦 Già spedito (correzione)'),
                   command=self._already_shipped).pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text=L('btn_close', 'Chiudi'),
                   command=self.destroy).pack(side=tk.LEFT, padx=4)

    def _load_candidate_sos(self):
        """SO con lo stesso prodotto (ItemCode = ProductCode senza |N) e residuo > 0."""
        L = self.lang.get
        try:
            self.db.cursor.execute("""
                SELECT d.DynamicSaleOrderId, d.SONumber, d.CustomerName, d.QtyOrder,
                       d.QtyOrder - ISNULL((
                           SELECT SUM(po.Qty)
                           FROM Traceability_RS.dyn.DynamicProductionOrders po
                           WHERE po.DynamicSaleOrderId = d.DynamicSaleOrderId), 0) AS Residuo
                FROM Traceability_RS.dyn.DynamicSaleOrders d
                WHERE d.ItemCode = ?
                  AND d.QtyOrder - ISNULL((
                          SELECT SUM(po.Qty)
                          FROM Traceability_RS.dyn.DynamicProductionOrders po
                          WHERE po.DynamicSaleOrderId = d.DynamicSaleOrderId), 0) > 0
                ORDER BY d.ShipDateRequest
            """, (self.item_code,))
            rows = self.db.cursor.fetchall()
        except Exception as e:
            logger.error(f"Errore caricamento SO candidati: {e}", exc_info=True)
            rows = []
        self._so_by_label = {}
        labels = []
        for r in rows:
            label = f"{r.SONumber} — {r.CustomerName or ''} (residuo {int(r.Residuo)})"
            self._so_by_label[label] = {'id': r.DynamicSaleOrderId, 'residuo': int(r.Residuo)}
            labels.append(label)
        self.so_combo['values'] = labels
        if labels:
            self.so_combo.current(0)
            self._sync_qty_hint()
            self.so_combo.bind('<<ComboboxSelected>>', lambda _e: self._sync_qty_hint())
        else:
            self.qty_hint.config(
                text=L('wh_no_so', 'Nessun ordine di vendita con questo prodotto'),
                foreground='#c0392b')

    def _sync_qty_hint(self):
        sel = self._so_by_label.get(self.so_combo.get())
        if not sel:
            return
        cap = min(self.wh['Residua'], sel['residuo'])
        self.qty_hint.config(
            text=f"(max {cap} = min tra magazzino {self.wh['Residua']} e residuo SO {sel['residuo']})",
            foreground='gray')

    def _match_and_rule(self):
        L = self.lang.get
        sel = self._so_by_label.get(self.so_combo.get())
        if not sel:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('wh_pick_so', 'Selezionare un ordine di vendita'),
                                   parent=self)
            return
        try:
            qty = int(self.qty_entry.get().strip())
        except ValueError:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('qty_invalid', 'Quantità non valida'), parent=self)
            return
        cap = min(self.wh['Residua'], sel['residuo'])
        if qty <= 0 or qty > cap:
            messagebox.showwarning(
                L('warning', 'Attenzione'),
                f"{L('qty_exceeds', 'La quantità supera il massimo disponibile')}: {cap}",
                parent=self)
            return
        from datetime import datetime, time
        try:
            dt = datetime.combine(self.date_entry.get_date(),
                                  time(int(self.hour_sb.get()), int(self.min_sb.get())))
        except ValueError:
            dt = datetime.combine(self.date_entry.get_date(), time(8, 0))
        ship_to = self.shipto_combo.get()
        try:
            # 1) abbinamento SO↔PO (come i planner)
            self.db.cursor.execute("""
                INSERT INTO Traceability_RS.dyn.DynamicProductionOrders
                    (DynamicSaleOrderId, IdOrder, Qty, DateIn)
                OUTPUT INSERTED.DynamicProductionOrderID
                VALUES (?, ?, ?, GETDATE())
            """, (sel['id'], self.wh['IDOrder'], qty))
            po_id = self.db.cursor.fetchone()[0]
            # 2) regola di spedizione con destinazione
            self.db.cursor.execute("""
                INSERT INTO Traceability_RS.dyn.DynamicShippingRules
                    (DynamicProductionOrderID, QtyToShip, DateToship, ShipTo, AddBayUser, DateOut)
                VALUES (?, ?, ?, ?, ?, GETDATE())
            """, (po_id, qty, dt, ship_to, self.user_name))
            self.db.conn.commit()
            logger.info("Magazzino abbinato: IDOrder=%s → SO %s qty=%s ShipTo=%s (PO=%s)",
                        self.wh['IDOrder'], sel['id'], qty, ship_to, po_id)
            messagebox.showinfo(L('success', 'Successo'),
                                L('wh_matched', 'Ordine abbinato e regola creata'),
                                parent=self)
            if self.callback:
                self.callback()
            self.destroy()
        except Exception as e:
            logger.error(f"Errore abbinamento magazzino: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)

    def _already_shipped(self):
        """Correzione transitoria: dichiara quanto è già stato spedito fuori sistema."""
        L = self.lang.get
        if not self.master._adjustments_table_exists():
            messagebox.showwarning(
                L('warning', 'Attenzione'),
                L('wh_adj_missing',
                  "Tabella correzioni non presente. Eseguire "
                  "add_warehouse_shipped_adjustments.sql."),
                parent=self)
            return
        WarehouseShippedDialog(self, self.db, self.lang, self.user_name, self.wh,
                               callback=self._on_adjusted)

    def _on_adjusted(self):
        if self.callback:
            self.callback()
        self.destroy()


class WarehouseShippedDialog(tk.Toplevel):
    """Correzione 'già spedito' per un ordine a magazzino (allineamento transitorio)."""

    def __init__(self, master, db, lang, user_name, wh_row, callback=None):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.wh = wh_row
        self.callback = callback
        self.match_dialog = master  # per _add_shipped_adjustment via la window
        L = self.lang.get
        self.title(L('wh_shipped_title', 'Già spedito — correzione'))
        self.geometry('460x230')
        self.transient(master)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text=f"{L('production_order', 'Ordine Produzione')}: "
                            f"{self.wh['OrderNumber']}  ({self.wh['ProductCode']})",
                  font=('', 9, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 8))
        ttk.Label(frm, text=f"{L('wh_available', 'Disponibile a magazzino')}: "
                            f"{self.wh['Residua']}").grid(row=1, column=0, columnspan=2, sticky='w')
        ttk.Label(frm, text=L('wh_shipped_qty', 'Quantità già spedita') + ' *').grid(
            row=2, column=0, sticky='w', pady=(10, 4))
        self.qty_entry = ttk.Entry(frm, width=14)
        self.qty_entry.grid(row=2, column=1, sticky='w', pady=(10, 4))
        ttk.Label(frm, text=L('note', 'Nota')).grid(row=3, column=0, sticky='w', pady=4)
        self.note_entry = ttk.Entry(frm, width=40)
        self.note_entry.grid(row=3, column=1, sticky='w', pady=4)

        bar = ttk.Frame(frm)
        bar.grid(row=4, column=0, columnspan=2, pady=(16, 0))
        ttk.Button(bar, text=L('btn_save', 'Salva'), command=self._save).pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text=L('btn_cancel', 'Annulla'), command=self.destroy).pack(side=tk.LEFT, padx=4)

    def _save(self):
        L = self.lang.get
        try:
            qty = int(self.qty_entry.get().strip())
        except ValueError:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('qty_invalid', 'Quantità non valida'), parent=self)
            return
        if qty <= 0 or qty > self.wh['Residua']:
            messagebox.showwarning(
                L('warning', 'Attenzione'),
                f"{L('qty_exceeds', 'La quantità supera il massimo disponibile')}: {self.wh['Residua']}",
                parent=self)
            return
        try:
            # la window principale (master del WarehouseMatchDialog) espone il metodo
            window = self.match_dialog.master
            window._add_shipped_adjustment(
                self.wh['IDOrder'], self.wh['ProductCode'], qty,
                self.note_entry.get().strip() or None)
            messagebox.showinfo(L('success', 'Successo'),
                                L('wh_adj_saved', 'Correzione registrata'), parent=self)
            if self.callback:
                self.callback()
            self.destroy()
        except Exception as e:
            logger.error(f"Errore correzione già spedito: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)


def open_orders_reports_window(master, db, lang, user_name):
    """
    Apre la finestra di gestione spedizioni dinamiche
    
    Args:
        master: Finestra padre
        db: Connessione database
        lang: Gestore traduzioni
        user_name: Nome utente loggato
    """
    DynamicShippingWindow(master, db, lang, user_name)
    
    def _export_rules_to_excel(self):
        """Esporta tutte le regole di spedizione in Excel"""
        try:
            # Query per ottenere tutte le regole con informazioni ordini
            query = """
            SELECT 
                d.CustomerName,
                d.SONumber,
                o.ordernumber AS ProductionOrder,
                d.ItemCode,
                d.ItemName,
                d.QtyOrder,
                sr.QtyToShip,
                sr.DateToship,
                sr.ShipTo,
                sr.AddBayUser,
                sr.DateOut
            FROM 
                [Traceability_RS].[dyn].[DynamicShippingRules] sr
            INNER JOIN 
                [Traceability_RS].[dyn].[DynamicProductionOrders] po ON sr.DynamicProductionOrderID = po.DynamicProductionOrderID
            INNER JOIN 
                [Traceability_RS].[dyn].[DynamicSaleOrders] d ON po.DynamicSaleOrderId = d.DynamicSaleOrderId
            INNER JOIN 
                traceability_rs.dbo.orders o ON po.IdOrder = o.IDOrder
            ORDER BY 
                sr.DateToship, d.SONumber
            """
            
            self.db.cursor.execute(query)
            rows = self.db.cursor.fetchall()
            
            if not rows:
                messagebox.showinfo(
                    self.lang.get('info', 'Informazione'),
                    self.lang.get('no_rules_to_export', 'Nessuna regola di spedizione da esportare'),
                    parent=self
                )
                return
            
            # Crea directory se non esiste
            temp_dir = r'c:\Temp'
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            # Nome file con timestamp
            filename = f"ShippingRules_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            # Crea workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Regole Spedizione"
            
            # Stili
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Intestazioni
            headers = [
                'Cliente', 'Ordine Vendita', 'Ordine Produzione', 'Codice', 'Prodotto',
                'QtÃƒÂ  Ordine', 'QtÃƒÂ  da Spedire', 'Data Spedizione', 'Ora', 'Destinazione',
                'Aggiunto Da', 'Data Inserimento'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Dati
            for row_idx, row in enumerate(rows, 2):
                date_to_ship = row.DateToship.strftime('%d/%m/%Y') if row.DateToship else ''
                time_to_ship = row.DateToship.strftime('%H:%M') if row.DateToship else ''
                date_out = row.DateOut.strftime('%d/%m/%Y %H:%M') if row.DateOut else ''
                
                values = [
                    row.CustomerName,
                    row.SONumber,
                    row.ProductionOrder,
                    row.ItemCode,
                    row.ItemName,
                    row.QtyOrder,
                    row.QtyToShip,
                    date_to_ship,
                    time_to_ship,
                    row.ShipTo or '',
                    row.AddBayUser or '',
                    date_out
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-size colonne
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze prima riga
            ws.freeze_panes = 'A2'
            
            # Salva
            wb.save(filepath)
            
            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                f"{self.lang.get('file_saved', 'File salvato in')}: {filepath}",
                parent=self
            )
            
        except Exception as e:
            logger.error(f"Errore esportazione regole Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore esportazione: {e}",
                parent=self
            )


