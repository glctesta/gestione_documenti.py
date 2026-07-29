# -*- coding: utf-8 -*-
"""
finished_goods_transfers_window.py — Versamenti prodotto finito.

Mostra i versamenti di prodotto finito verso D365 (LogApiDynamics), con filtri
per periodo, codice prodotto, ordine di produzione e cliente finale. Gli ID
delle scatole del versamento sono aggregati in un'unica colonna (doppio click
per aprirli in un elenco copiabile).

Se un prodotto non ha il cliente finale associato (FinalClientName NULL), la
form lo segnala e offre di accoppiarlo subito: la scelta viene scritta in
Products.IdFinalClient.

Query e aggiornamenti stanno nel db handler:
  Database.fetch_finished_goods_transfers(...)
  Database.assign_final_client_to_product(...)
  Database.fetch_final_customers()
"""
import os
import logging
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, messagebox

try:
    from tkcalendar import DateEntry
except Exception:
    DateEntry = None

logger = logging.getLogger(__name__)

# Colonne della griglia: (id, chiave_lang, etichetta_default, larghezza, ancora)
# 'FinalClientName' resta all'indice 5: il doppio-click sull'associazione lo
# usa. 'IdBoxTrasb' va in fondo (indice 6) per non spostarlo.
_COLUMNS = (
    ('OrderNumber',    'fgt_col_order',    'Ordine Prod.',   130, 'w'),
    ('ProductCode',    'fgt_col_product',  'Codice Prodotto', 150, 'w'),
    ('QtySend',        'fgt_col_qty',      'Qta Versata',      90, 'center'),
    ('ResponseByD365', 'fgt_col_resp',     'Esito D365',      160, 'w'),
    ('DateTransfer',   'fgt_col_date',     'Data Transfer',   130, 'center'),
    ('FinalClientName', 'fgt_col_client',  'Cliente Finale',  170, 'w'),
    ('IdBoxTrasb',     'fgt_col_boxes',    'Scatole (ID)',    150, 'w'),
)
_NO_CLIENT = '⚠ (nessun cliente)'
_CLIENT_COL = 5   # indice di FinalClientName nei values della griglia
_BOX_COL = 6      # indice di IdBoxTrasb


class FinishedGoodsTransfersWindow(tk.Toplevel):
    def __init__(self, master, db, lang, user_name):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self._rows = []            # righe correnti (dopo i filtri)

        self.title(self.lang.get('fgt_title', 'Versamenti prodotto finito'))
        self.geometry('1150x680')
        self.minsize(900, 520)
        self.transient(master)

        self._build_ui()
        self._load_clients()
        self.grab_set()

    # ── UI ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        L = self.lang.get
        main = ttk.Frame(self, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        flt = ttk.LabelFrame(main, text=L('filters', 'Filtri'), padding=8)
        flt.pack(fill=tk.X, pady=(0, 8))

        r1 = ttk.Frame(flt); r1.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(r1, text=L('fgt_date_from', 'Data da:')).pack(side=tk.LEFT, padx=(0, 4))
        self.date_from = self._make_date(r1, datetime.now() - timedelta(days=7))
        ttk.Label(r1, text=L('fgt_date_to', 'a:')).pack(side=tk.LEFT, padx=(0, 4))
        self.date_to = self._make_date(r1, datetime.now())

        ttk.Label(r1, text=L('fgt_product', 'Prodotto:')).pack(side=tk.LEFT, padx=(12, 4))
        self.product_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self.product_var, width=18).pack(side=tk.LEFT)

        ttk.Label(r1, text=L('fgt_order', 'Ordine:')).pack(side=tk.LEFT, padx=(12, 4))
        self.order_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self.order_var, width=16).pack(side=tk.LEFT)

        r2 = ttk.Frame(flt); r2.pack(fill=tk.X)
        ttk.Label(r2, text=L('fgt_client', 'Cliente:')).pack(side=tk.LEFT, padx=(0, 4))
        self.client_combo = ttk.Combobox(r2, width=30, state='readonly')
        self.client_combo.pack(side=tk.LEFT)

        ttk.Button(r2, text=L('btn_filter', 'Filtra'),
                   command=self._load_data).pack(side=tk.LEFT, padx=(12, 4))
        ttk.Button(r2, text=L('btn_clear', 'Pulisci'),
                   command=self._clear_filters).pack(side=tk.LEFT, padx=4)
        ttk.Button(r2, text=L('btn_export_excel', 'Esporta Excel'),
                   command=self._export_excel).pack(side=tk.LEFT, padx=(16, 4))

        # ── Griglia ──────────────────────────────────────────────────────
        wrap = ttk.Frame(main); wrap.pack(fill=tk.BOTH, expand=True)
        cols = tuple(c[0] for c in _COLUMNS)
        self.tree = ttk.Treeview(wrap, columns=cols, show='headings', selectmode='browse')
        for cid, key, deflabel, width, anchor in _COLUMNS:
            self.tree.heading(cid, text=L(key, deflabel))
            self.tree.column(cid, width=width, anchor=anchor)
        # Righe senza cliente: evidenziate. Doppio click per associare.
        self.tree.tag_configure('no_client', background='#FFF3B0')
        self.tree.bind('<Double-1>', self._on_row_double_click)
        vsb = ttk.Scrollbar(wrap, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        self.status_var = tk.StringVar(value='')
        ttk.Label(main, textvariable=self.status_var, foreground='gray').pack(
            fill=tk.X, pady=(6, 0))

    def _make_date(self, parent, default_dt):
        if DateEntry:
            w = DateEntry(parent, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
            w.set_date(default_dt)
        else:
            w = ttk.Entry(parent, width=12)
            w.insert(0, default_dt.strftime('%d/%m/%Y'))
        w.pack(side=tk.LEFT, padx=(0, 8))
        return w

    def _get_date(self, widget):
        if DateEntry and hasattr(widget, 'get_date'):
            return widget.get_date()
        try:
            return datetime.strptime(widget.get().strip(), '%d/%m/%Y').date()
        except Exception:
            return None

    # ── Dati ──────────────────────────────────────────────────────────────
    def _load_clients(self):
        L = self.lang.get
        self._all_clients_label = L('fgt_all_clients', '(Tutti i clienti)')
        names = [self._all_clients_label]
        try:
            for r in (self.db.fetch_final_customers() or []):
                if r.FinalClientName:
                    names.append(r.FinalClientName)
        except Exception as e:
            logger.error(f"Load clienti finali: {e}", exc_info=True)
        self.client_combo['values'] = names
        self.client_combo.current(0)

    def _clear_filters(self):
        self.product_var.set('')
        self.order_var.set('')
        if self.client_combo['values']:
            self.client_combo.current(0)
        if DateEntry and hasattr(self.date_from, 'set_date'):
            self.date_from.set_date(datetime.now() - timedelta(days=7))
            self.date_to.set_date(datetime.now())

    def _selected_client(self):
        val = self.client_combo.get()
        if not val or val == getattr(self, '_all_clients_label', ''):
            return None
        return val

    def _load_data(self):
        L = self.lang.get
        d_from = self._get_date(self.date_from)
        d_to = self._get_date(self.date_to)
        if not d_from or not d_to:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('fgt_date_invalid', 'Date non valide.'), parent=self)
            return
        if d_from > d_to:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('fgt_date_order', 'La data iniziale supera quella finale.'),
                                   parent=self)
            return

        # La query filtra su BETWEEN @from AND @to: per includere l'intera
        # giornata finale spingo il limite a fine giornata.
        dt_from = datetime(d_from.year, d_from.month, d_from.day, 0, 0, 0)
        dt_to = datetime(d_to.year, d_to.month, d_to.day, 23, 59, 59)

        self.db.last_error_details = ""
        self._rows = self.db.fetch_finished_goods_transfers(
            dt_from, dt_to,
            product_code=self.product_var.get().strip() or None,
            production_order=self.order_var.get().strip() or None,
            client_name=self._selected_client())

        err = getattr(self.db, 'last_error_details', '')
        if err and not self._rows:
            messagebox.showerror(L('error', 'Errore'),
                                 f"{L('fgt_load_err', 'Errore nel caricamento')}:\n{err}",
                                 parent=self)

        self._populate_tree()
        self._check_missing_clients()

    def _populate_tree(self):
        L = self.lang.get
        for it in self.tree.get_children():
            self.tree.delete(it)
        n_missing = 0
        for r in self._rows:
            client = r.FinalClientName
            date_s = r.DateTransfer.strftime('%d/%m/%Y %H:%M') \
                if hasattr(r.DateTransfer, 'strftime') else (r.DateTransfer or '')
            tag = ()
            if not client:
                client = _NO_CLIENT
                tag = ('no_client',)
                n_missing += 1
            self.tree.insert('', tk.END, values=(
                r.OrderNumber or '', r.ProductCode or '', r.QtySend if r.QtySend is not None else '',
                r.ResponseByD365 or '', date_s, client, r.IdBoxTrasb or ''), tags=tag)
        self.status_var.set(
            L('fgt_count', '{n} versamenti — {m} senza cliente').format(
                n=len(self._rows), m=n_missing))

    def _missing_product_codes(self):
        """Codici prodotto distinti senza cliente associato, in ordine."""
        seen, out = set(), []
        for r in self._rows:
            if not r.FinalClientName and r.ProductCode and r.ProductCode not in seen:
                seen.add(r.ProductCode)
                out.append(r.ProductCode)
        return out

    def _check_missing_clients(self):
        L = self.lang.get
        codes = self._missing_product_codes()
        if not codes:
            return
        if messagebox.askyesno(
                L('fgt_missing_title', 'Prodotti senza cliente'),
                L('fgt_missing_q',
                  '{n} prodotti non hanno il cliente finale associato.\n'
                  'Vuoi associarli adesso?').format(n=len(codes)),
                parent=self):
            self._open_pairing(codes)

    # ── Doppio click: scatole o associazione cliente ──────────────────────
    def _on_row_double_click(self, event=None):
        row_id = self.tree.identify_row(event.y) if event else None
        if not row_id:
            sel = self.tree.selection()
            row_id = sel[0] if sel else None
        if not row_id:
            return
        vals = self.tree.item(row_id, 'values')
        # Colonna cliccata: '#1'.. -> indice 0-based
        col = self.tree.identify_column(event.x) if event else ''
        col_idx = (int(col[1:]) - 1) if col.startswith('#') else -1

        # Click sulla colonna Scatole: apre il visualizzatore per copiare gli ID
        if col_idx == _BOX_COL and len(vals) > _BOX_COL and vals[_BOX_COL]:
            self._show_box_ids(vals[0], vals[_BOX_COL])
            return

        # Altrimenti: se la riga e' senza cliente, apre l'associazione
        if len(vals) > _CLIENT_COL and vals[_CLIENT_COL] == _NO_CLIENT and vals[1]:
            self._open_pairing([vals[1]])

    def _show_box_ids(self, order, box_csv):
        """Apre un piccolo 'notepad' con gli ID scatola, selezionabili e
        copiabili (uno per riga, piu' comodo da leggere del CSV)."""
        L = self.lang.get
        ids = [x.strip() for x in str(box_csv).split(',') if x.strip()]
        top = tk.Toplevel(self)
        top.title(L('fgt_boxes_title', 'ID scatole') + (f' — {order}' if order else ''))
        top.geometry('320x420')
        top.transient(self)
        top.grab_set()

        ttk.Label(top, text=L('fgt_boxes_hint', '{n} scatole — seleziona e copia').format(n=len(ids)),
                  foreground='gray').pack(anchor='w', padx=10, pady=(10, 4))
        frm = ttk.Frame(top); frm.pack(fill=tk.BOTH, expand=True, padx=10)
        txt = tk.Text(frm, wrap='none', font=('Consolas', 11))
        txt.insert('1.0', '\n'.join(ids))
        txt.config(state='normal')
        vsb = ttk.Scrollbar(frm, orient='vertical', command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.focus_set()

        def copy_all():
            self.clipboard_clear()
            self.clipboard_append(', '.join(ids))
            self.status_var.set(L('fgt_boxes_copied', '{n} ID scatola copiati').format(n=len(ids)))

        bar = ttk.Frame(top); bar.pack(fill=tk.X, padx=10, pady=8)
        ttk.Button(bar, text=L('fgt_boxes_copy', 'Copia tutto'), command=copy_all).pack(side=tk.LEFT)
        ttk.Button(bar, text=L('btn_close', 'Chiudi'), command=top.destroy).pack(side=tk.RIGHT)

    def _open_pairing(self, product_codes):
        dlg = PairFinalClientDialog(self, self.db, self.lang, product_codes)
        self.wait_window(dlg)
        if getattr(dlg, 'changed', False):
            self._load_data()   # ricarica: le righe ora hanno il cliente

    # ── Export Excel ──────────────────────────────────────────────────────
    @staticmethod
    def _as_int(v):
        """Qta come intero per i totali; None se non convertibile."""
        try:
            return int(str(v).strip())
        except (ValueError, TypeError, AttributeError):
            return None

    def _export_excel(self):
        L = self.lang.get
        if not self._rows:
            messagebox.showinfo(L('info', 'Informazione'),
                                L('fgt_no_export', 'Nessun dato da esportare.'), parent=self)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            messagebox.showerror(L('error', 'Errore'),
                                 f"openpyxl non disponibile: {e}", parent=self)
            return

        # Colonne dell'export: Data e Ora separate (in griglia sono un'unica
        # colonna). L'export legge da self._rows, non dalla griglia, per avere
        # il datetime vero da spezzare e la qta gia' numerica.
        headers = [L('fgt_col_order',   'Ordine Prod.'),
                   L('fgt_col_product', 'Codice Prodotto'),
                   L('fgt_col_qty',     'Qta Versata'),
                   L('fgt_col_resp',    'Esito D365'),
                   L('fgt_col_only_date', 'Data'),
                   L('fgt_col_only_time', 'Ora'),
                   L('fgt_col_client',  'Cliente Finale'),
                   L('fgt_col_boxes',   'Scatole (ID)')]
        NUM_COL = 3       # Qta
        EMPTY_COLS = (4, 5, 6, 7, 8)   # colonne vuote nelle righe subtotale/totale

        # Righe ordinate per ordine e istante: cosi' gli stessi ordini sono
        # contigui e i subtotali hanno senso.
        def _sortkey(r):
            dt = r.DateTransfer
            return (r.OrderNumber or '', dt if hasattr(dt, 'strftime') else datetime.min)
        rows = sorted(self._rows, key=_sortkey)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = L('fgt_sheet', 'Versamenti')

            head_font = Font(bold=True, color='FFFFFF', size=11)
            head_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            sub_font = Font(bold=True)
            sub_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            tot_font = Font(bold=True, size=11)
            tot_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            thin = Side(style='thin', color='D0D0D0')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for c, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=title)
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            widths = [len(h) for h in headers]

            def put(r, c, value, *, num=False, center=False, font=None, fill=None):
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = border
                if center or num:
                    cell.alignment = Alignment(horizontal='center')
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                widths[c - 1] = max(widths[c - 1], len(str(value if value is not None else '')))
                return cell

            excel_row = 2
            grand_total = 0
            data_end_row = 1  # ultima riga "dati" (per l'autofiltro, che esclude i totali)

            def write_subtotal(order, qty):
                nonlocal excel_row
                put(excel_row, 1, order, font=sub_font, fill=sub_fill)
                put(excel_row, 2, L('fgt_subtotal', 'Subtotale'), font=sub_font, fill=sub_fill)
                put(excel_row, NUM_COL, qty, num=True, font=sub_font, fill=sub_fill)
                for c in EMPTY_COLS:
                    put(excel_row, c, '', fill=sub_fill)
                excel_row += 1

            cur_order = None
            group_qty = 0
            for r in rows:
                order = r.OrderNumber or ''
                if cur_order is not None and order != cur_order:
                    write_subtotal(cur_order, group_qty)
                    group_qty = 0
                cur_order = order

                dt = r.DateTransfer
                date_val = dt.date() if hasattr(dt, 'date') else None
                time_val = dt.strftime('%H:%M:%S') if hasattr(dt, 'strftime') else ''
                qty = self._as_int(r.QtySend)
                client = r.FinalClientName or _NO_CLIENT

                put(excel_row, 1, order)
                put(excel_row, 2, r.ProductCode or '')
                put(excel_row, NUM_COL, qty if qty is not None else (r.QtySend or ''), num=True)
                put(excel_row, 4, r.ResponseByD365 or '')
                dcell = put(excel_row, 5, date_val, center=True)
                if date_val is not None:
                    dcell.number_format = 'DD/MM/YYYY'
                put(excel_row, 6, time_val, center=True)
                put(excel_row, 7, client)
                put(excel_row, 8, r.IdBoxTrasb or '')

                if qty is not None:
                    group_qty += qty
                    grand_total += qty
                data_end_row = excel_row
                excel_row += 1

            if cur_order is not None:
                write_subtotal(cur_order, group_qty)

            # Totale generale dei pezzi
            put(excel_row, 1, L('fgt_grand_total', 'TOTALE GENERALE'), font=tot_font, fill=tot_fill)
            put(excel_row, 2, '', fill=tot_fill)
            put(excel_row, NUM_COL, grand_total, num=True, font=tot_font, fill=tot_fill)
            for c in EMPTY_COLS:
                put(excel_row, c, '', fill=tot_fill)

            # Autofiltro sulle sole righe dati (esclude subtotali e totale)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end_row}"
            ws.freeze_panes = 'A2'
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(w + 3, 10), 50)
            ws.row_dimensions[1].height = 26

            # Cartella temporanea: creata se non esiste
            temp_dir = r'c:\Temp'
            os.makedirs(temp_dir, exist_ok=True)
            fname = f"Versamenti_ProdottoFinito_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            path = os.path.join(temp_dir, fname)
            wb.save(path)
        except PermissionError:
            messagebox.showerror(L('error', 'Errore'),
                                 L('fgt_file_locked', 'File aperto in Excel? Impossibile scrivere.'),
                                 parent=self)
            return
        except Exception as e:
            logger.error(f"Export Excel versamenti: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'),
                                 f"{L('fgt_export_err', 'Errore export')}:\n{e}", parent=self)
            return

        self.status_var.set(L('fgt_exported', 'Esportato in {path}').format(path=path))
        try:
            os.startfile(path)
        except Exception:
            pass


class PairFinalClientDialog(tk.Toplevel):
    """Associa un cliente finale ai prodotti senza cliente, uno per volta."""

    def __init__(self, parent, db, lang, product_codes):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self._codes = list(product_codes)
        self.changed = False

        L = self.lang.get
        self.title(L('fgt_pair_title', 'Associa cliente finale'))
        self.geometry('520x220')
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self._clients = []   # (IDFinalClient, Name)
        self._build_ui()
        self._load_clients()

    def _build_ui(self):
        L = self.lang.get
        frm = ttk.Frame(self, padding=14)
        frm.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frm, text=L('fgt_pair_product', 'Codice prodotto:')).grid(
            row=0, column=0, sticky='w', pady=6)
        self.code_combo = ttk.Combobox(frm, width=32, state='readonly', values=self._codes)
        self.code_combo.grid(row=0, column=1, sticky='w', pady=6)
        if self._codes:
            self.code_combo.current(0)

        ttk.Label(frm, text=L('fgt_pair_client', 'Cliente finale:')).grid(
            row=1, column=0, sticky='w', pady=6)
        self.client_combo = ttk.Combobox(frm, width=32)
        self.client_combo.grid(row=1, column=1, sticky='w', pady=6)
        self.client_combo.bind('<KeyRelease>', self._on_client_typed)

        bar = ttk.Frame(frm)
        bar.grid(row=3, column=0, columnspan=2, sticky='e', pady=(16, 0))
        ttk.Button(bar, text=L('btn_close', 'Chiudi'),
                   command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(bar, text=L('fgt_pair_save', 'Associa'),
                   command=self._save).pack(side=tk.RIGHT, padx=(0, 8))
        frm.columnconfigure(1, weight=1)

    def _load_clients(self):
        try:
            for r in (self.db.fetch_final_customers() or []):
                if r.FinalClientName:
                    self._clients.append((r.IDFinalClient, r.FinalClientName))
        except Exception as e:
            logger.error(f"Pair: load clienti: {e}", exc_info=True)
        self._client_display = [n for _i, n in self._clients]
        self.client_combo['values'] = self._client_display

    def _on_client_typed(self, event=None):
        if event is not None and event.keysym in ('Up', 'Down', 'Return', 'Escape', 'Tab'):
            return
        txt = self.client_combo.get().strip().lower()
        self.client_combo['values'] = ([n for n in self._client_display if txt in n.lower()]
                                       if txt else self._client_display)

    def _current_client_id(self):
        txt = self.client_combo.get().strip()
        for cid, name in self._clients:
            if name == txt:
                return cid
        return None

    def _save(self):
        L = self.lang.get
        code = self.code_combo.get().strip()
        if not code:
            return
        client_id = self._current_client_id()
        if not client_id:
            messagebox.showwarning(L('warning', 'Attenzione'),
                                   L('fgt_pick_client', 'Selezionare un cliente valido.'),
                                   parent=self)
            self.client_combo.focus_set()
            return

        ok, res = self.db.assign_final_client_to_product(code, client_id)
        if not ok:
            messagebox.showerror(L('error', 'Errore'),
                                 f"{L('fgt_pair_err', 'Errore associazione')}:\n{res}",
                                 parent=self)
            return
        if res == 0:
            messagebox.showinfo(L('info', 'Informazione'),
                                L('fgt_pair_none', 'Nessun prodotto aggiornato per {code} '
                                  '(forse gia associato).').format(code=code), parent=self)
        else:
            self.changed = True
            messagebox.showinfo(L('info', 'Informazione'),
                                L('fgt_pair_ok', 'Associato {code} → {client} ({n} prodotti).').format(
                                    code=code, client=self.client_combo.get().strip(), n=res),
                                parent=self)

        # Toglie il codice appena gestito; se non ne restano, chiude
        self._codes = [c for c in self._codes if c != code]
        if not self._codes:
            self.destroy()
            return
        self.code_combo['values'] = self._codes
        self.code_combo.current(0)
        self.client_combo.set('')


def open_finished_goods_transfers_window(master, db, lang, user_name):
    """Apre la finestra 'Versamenti prodotto finito'."""
    FinishedGoodsTransfersWindow(master, db, lang, user_name)
