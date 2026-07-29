"""
orders/shipment_confirmation_window.py

Finestra "Conferma Shipping" v2.

Novita' rispetto alla v1:
 - lista raggruppata per Ordine di Produzione con somma delle quantita' da spedire
   e residuo che si scala man mano che si confermano i pallet;
 - filtri di ricerca per Ordine e per Prodotto;
 - una spedizione (dyn.Shipments) raggruppa piu' ordini su piu' pallet
   (dyn.ShipmentPallets); codice pallet inserito dall'operatore (univoco solo
   entro la spedizione) con suggerimento progressivo;
 - finalizzazione con generazione di 2 PDF (lista per pallet + riepilogo),
   archiviazione, ed email ai destinatari 'Sys_shipment_email' con PDF allegati;
 - recupero per data di una spedizione passata, correzione, email di correzione
   e ristampa documenti.

Il residuo per ordine = SUM(QtyToShip su regole ConfirmedAt IS NULL) -
SUM(ShipmentPallets.ConfirmedQty). Non si usa il roll-up di ConfirmedAt: cosi'
le correzioni ricalcolano tutto automaticamente. Vedi DESIGN_shipment_confirmation_v2.md
"""

import logging
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date, timedelta

from tkcalendar import DateEntry

import utils
from orders import shipment_pdf

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  Query
# ─────────────────────────────────────────────────────────────────────────────

# Ordini di produzione con residuo da spedire > 0 (filtri opzionali).
_QUERY_ORDERS = """
SELECT
    O.DynamicProductionOrderID                AS DPOID,
    PO.ordernumber                            AS ProductionOrder,
    MAX(D.CustomerName)                        AS CustomerName,
    MAX(D.SONumber)                            AS SONumber,
    MAX(D.ItemCode)                            AS ItemCode,
    MAX(D.ItemName)                            AS ItemName,
    MAX(R.ShipTo)                              AS ShipTo,
    FORMAT(MIN(R.DateToship), 'dd/MM/yyyy')    AS DateToShipFmt,
    SUM(R.QtyToShip)                           AS QtyToShipTotal,
    ISNULL((SELECT SUM(sp.ConfirmedQty)
            FROM [Traceability_RS].[dyn].[ShipmentPallets] sp
            WHERE sp.DynamicProductionOrderID = O.DynamicProductionOrderID), 0)
                                              AS QtyConfirmedTotal
FROM [Traceability_RS].[dyn].[DynamicShippingRules] R
INNER JOIN [Traceability_RS].[dyn].[DynamicProductionOrders] O
       ON O.DynamicProductionOrderID = R.DynamicProductionOrderID
INNER JOIN [Traceability_RS].[dyn].[DynamicSaleOrders] D
       ON D.DynamicSaleOrderId = O.DynamicSaleOrderId
INNER JOIN [Traceability_RS].[dbo].[Orders] PO
       ON PO.IDOrder = O.IdOrder
WHERE R.ConfirmedAt IS NULL
  {filters}
GROUP BY O.DynamicProductionOrderID, PO.ordernumber
HAVING SUM(R.QtyToShip) - ISNULL((SELECT SUM(sp.ConfirmedQty)
            FROM [Traceability_RS].[dyn].[ShipmentPallets] sp
            WHERE sp.DynamicProductionOrderID = O.DynamicProductionOrderID), 0) > 0
ORDER BY MIN(R.DateToship) ASC
"""

_ORDER_COLS = (
    "IDOrder",          # hidden (chiave riga = ordine di produzione)
    "ProductionOrder",
    "Product",
    "Versato",
    "Spedito",
    "Residual",
    "LastWhDate",
)
_ORDER_LABELS = {
    "ProductionOrder": ("Ordine Prod.",  110),
    "Product":         ("Prodotto",      230),
    "Versato":         ("Versato mag.",   95),
    "Spedito":         ("Già spedito",    95),
    "Residual":        ("Residuo",        85),
    "LastWhDate":      ("Ult. versam.",   95),
}

# ── Ordini versati a magazzino e NON ancora spediti (residuo = versato − spedito) ──
# Batch multi-istruzione: eseguire su un cursore DEDICATO (self.db.conn.cursor()),
# mai sul cursore condiviso (lascia result-set pendenti → 'Invalid cursor state').
# Finestra su LogApiDynamics.CurrentDate: MAX(data_inizio_spedizioni, @from) .. @to.
_QUERY_WAREHOUSE = """
SET NOCOUNT ON;
DECLARE @from date = ?, @to date = ?;
DECLARE @floor date;
SELECT @floor = TRY_CAST(Value AS date)
FROM traceability_rs.dbo.Settings WHERE atribute = 'data_inizio_spedizioni';
DECLARE @start date = CASE WHEN @floor IS NULL OR @from > @floor THEN @from ELSE @floor END;
;WITH WH AS (
    SELECT o.IDOrder, p.ProductCode, o.OrderNumber,
           SUM(CAST(JSON_VALUE(j.value, '$.RealValue') AS int)) AS WhQty,
           CAST(MAX(L.CurrentDate) AS date) AS LastWhDate
    FROM traceability_rs.dbo.LogApiDynamics L
    INNER JOIN traceability_rs.dbo.Orders o
        ON o.OrderNumber = JSON_VALUE(L.MessageSend, '$.Message.Reference')
    INNER JOIN traceability_rs.dbo.Products p ON p.IDProduct = o.IDProduct
    CROSS APPLY OPENJSON(L.MessageSend,
        '$.Message.KeyValue.ListValue[0].ListValue[0].ListValue') j
    WHERE L.EndPointName = 'ProdFinishedGoods'
      AND CAST(L.CurrentDate AS date) BETWEEN @start AND @to
      AND JSON_VALUE(j.value, '$.Key') = 'GoodQty'
    GROUP BY o.IDOrder, p.ProductCode, o.OrderNumber
),
MATCHED AS (
    SELECT IdOrder, SUM(Qty) AS MatchedQty
    FROM traceability_rs.dyn.DynamicProductionOrders GROUP BY IdOrder
),
SHIPPED AS (
    SELECT po.IdOrder, SUM(sp.ConfirmedQty) AS ShippedQty
    FROM traceability_rs.dyn.ShipmentPallets sp
    INNER JOIN traceability_rs.dyn.DynamicProductionOrders po
        ON po.DynamicProductionOrderID = sp.DynamicProductionOrderID
    INNER JOIN traceability_rs.dyn.Shipments s ON s.ShipmentId = sp.ShipmentId
    WHERE s.ShipmentDate >= @start GROUP BY po.IdOrder
),
URGENT AS (
    SELECT DISTINCT po.IdOrder
    FROM traceability_rs.dyn.DynamicShippingRules r
    INNER JOIN traceability_rs.dyn.DynamicProductionOrders po
        ON po.DynamicProductionOrderID = r.DynamicProductionOrderID
    WHERE r.ConfirmedAt IS NULL
)
SELECT wh.IDOrder, wh.ProductCode, wh.OrderNumber, wh.WhQty, wh.LastWhDate,
       ISNULL(m.MatchedQty, 0)  AS MatchedQty,
       ISNULL(sh.ShippedQty, 0) AS ShippedQty,
       wh.WhQty - ISNULL(sh.ShippedQty, 0) AS ResiduoDaSpedire,
       CASE WHEN u.IdOrder IS NOT NULL THEN 1 ELSE 0 END AS IsUrgent
FROM WH wh
LEFT JOIN MATCHED m ON m.IdOrder = wh.IDOrder
LEFT JOIN SHIPPED sh ON sh.IdOrder = wh.IDOrder
LEFT JOIN URGENT u ON u.IdOrder = wh.IDOrder
WHERE wh.WhQty - ISNULL(sh.ShippedQty, 0) > 0
ORDER BY IsUrgent DESC, wh.OrderNumber
"""

# SO candidati per un prodotto (chiave ESATTA: ItemCode = ProductCode senza suffisso |N),
# con residuo SO (QtyOrder − già assegnato) e l'eventuale DPO già esistente per QUESTO ordine.
_QUERY_SO_CANDIDATES = """
SELECT d.DynamicSaleOrderId, d.SONumber, d.CustomerName, d.ItemCode, d.ItemName,
       d.QtyOrder,
       ISNULL((SELECT SUM(po.Qty) FROM traceability_rs.dyn.DynamicProductionOrders po
               WHERE po.DynamicSaleOrderId = d.DynamicSaleOrderId), 0) AS Assigned,
       (SELECT TOP 1 po2.DynamicProductionOrderID
        FROM traceability_rs.dyn.DynamicProductionOrders po2
        WHERE po2.DynamicSaleOrderId = d.DynamicSaleOrderId AND po2.IdOrder = ?) AS DpoForThisOrder
FROM traceability_rs.dyn.DynamicSaleOrders d
WHERE d.ItemCode = ?
ORDER BY d.SONumber
"""

_PALLET_COLS = (
    "ShipmentPalletId",   # hidden
    "DPOID",              # hidden
    "PalletCode",
    "ProductionOrder",
    "ItemCode",
    "ItemName",
    "Qty",
)
_PALLET_LABELS = {
    "PalletCode":      ("Pallet",        90),
    "ProductionOrder": ("Ordine Prod.", 100),
    "ItemCode":        ("Codice",        90),
    "ItemName":        ("Prodotto",     200),
    "Qty":             ("Qta",           70),
}


class ShipmentConfirmationWindow(tk.Toplevel):
    """Finestra per confermare le spedizioni urgenti su pallet."""

    def __init__(self, master, db, lang, user_name: str):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name

        # Stato spedizione
        self.shipment_id = None
        self.shipment_status = None      # OPEN | CLOSED | CORRECTED
        self.mode = "current"            # current | recover
        # DPOID -> dict snapshot/totali dell'ordine selezionabile
        self._order_info = {}

        self.title(self.lang.get("shipment_confirm_title", "Conferma Spedizioni Urgenti"))
        self.geometry("1320x760")
        self.transient(master)

        self._build_ui()
        self._resume_or_create_shipment()
        self._refresh_all()

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ------------------------------------------------------------------ #
    #  UI
    # ------------------------------------------------------------------ #
    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        # --- Header + riga spedizione ---
        hdr = ttk.Frame(main)
        hdr.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(
            hdr,
            text=self.lang.get("shipment_confirm_header", "Spedizioni Urgenti - Conferma su Pallet"),
            font=("Segoe UI", 13, "bold"),
        ).pack(side=tk.LEFT)
        ttk.Button(hdr, text=self.lang.get("btn_refresh", "Aggiorna"),
                   command=self._refresh_all).pack(side=tk.RIGHT, padx=4)
        # In alto a destra: nella barra Filtri finiva spinto fuori dall'area
        # visibile con finestre strette (due calendari + due caselle prima).
        ttk.Button(hdr, text=self.lang.get("btn_export_excel", "Esporta Excel"),
                   command=self._export_orders_excel).pack(side=tk.RIGHT, padx=4)
        ttk.Button(hdr, text=self.lang.get("btn_recover_shipment", "Recupera spedizione..."),
                   command=self._open_recover_dialog).pack(side=tk.RIGHT, padx=4)

        ship_row = ttk.Frame(main)
        ship_row.pack(fill=tk.X, pady=(0, 6))
        self.lbl_shipment = ttk.Label(ship_row, text="", font=("Segoe UI", 10, "bold"))
        self.lbl_shipment.pack(side=tk.LEFT, padx=(0, 16))
        ttk.Label(ship_row, text=self.lang.get(
            "shipment_date_label",
            "Data spedizione (registra qta spedite):")).pack(side=tk.LEFT)
        self.date_entry = DateEntry(ship_row, width=12, date_pattern="dd/mm/yyyy")
        self.date_entry.pack(side=tk.LEFT, padx=(6, 16))
        self.lbl_mode = ttk.Label(ship_row, text="", foreground="#1a5276", font=("Segoe UI", 10, "bold"))
        self.lbl_mode.pack(side=tk.LEFT)

        # --- Filtri ---
        flt = ttk.LabelFrame(main, text=self.lang.get("filters", "Filtri"), padding=8)
        flt.pack(fill=tk.X, pady=(0, 6))

        # Filtro periodo Data da spedire (guida la query ordini)
        ttk.Label(flt, text=self.lang.get("filter_date_from", "Data da (da spedire):")).pack(side=tk.LEFT)
        _def_from, _def_to = self._default_date_range()
        self.filter_from_entry = DateEntry(flt, width=12, date_pattern="dd/mm/yyyy")
        self.filter_from_entry.set_date(_def_from)
        self.filter_from_entry.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(flt, text=self.lang.get("filter_date_to", "Data a:")).pack(side=tk.LEFT)
        self.filter_to_entry = DateEntry(flt, width=12, date_pattern="dd/mm/yyyy")
        self.filter_to_entry.set_date(_def_to)
        self.filter_to_entry.pack(side=tk.LEFT, padx=(4, 12))

        ttk.Label(flt, text=self.lang.get("filter_order", "Ordine:")).pack(side=tk.LEFT)
        self.filter_order_var = tk.StringVar()
        e1 = ttk.Entry(flt, textvariable=self.filter_order_var, width=18)
        e1.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(flt, text=self.lang.get("filter_product", "Prodotto:")).pack(side=tk.LEFT)
        self.filter_product_var = tk.StringVar()
        e2 = ttk.Entry(flt, textvariable=self.filter_product_var, width=22)
        e2.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Button(flt, text=self.lang.get("btn_filter", "Filtra"),
                   command=self._load_orders).pack(side=tk.LEFT, padx=4)
        ttk.Button(flt, text=self.lang.get("btn_clear", "Pulisci"),
                   command=self._clear_filters).pack(side=tk.LEFT, padx=4)
        e1.bind("<Return>", lambda _e: self._load_orders())
        e2.bind("<Return>", lambda _e: self._load_orders())
        self.filter_from_entry.bind("<<DateEntrySelected>>", lambda _e: self._load_orders())
        self.filter_to_entry.bind("<<DateEntrySelected>>", lambda _e: self._load_orders())

        # --- Griglia ordini (residuo > 0) ---
        og = ttk.LabelFrame(main, text=self.lang.get("orders_to_ship", "Ordini da spedire"), padding=6)
        og.pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        self.orders_tree = ttk.Treeview(og, columns=_ORDER_COLS, show="headings",
                                        selectmode="browse", height=8)
        self.orders_tree.column("IDOrder", width=0, stretch=False)
        self.orders_tree.heading("IDOrder", text="")
        for col in _ORDER_COLS[1:]:
            label, width = _ORDER_LABELS[col]
            self.orders_tree.heading(col, text=self.lang.get(f"ocol_{col.lower()}", label))
            anchor = tk.W if col in ("Product",) else tk.CENTER
            self.orders_tree.column(col, width=width, anchor=anchor)
        ovsb = ttk.Scrollbar(og, orient=tk.VERTICAL, command=self.orders_tree.yview)
        self.orders_tree.configure(yscrollcommand=ovsb.set)
        self.orders_tree.grid(row=0, column=0, sticky="nsew")
        ovsb.grid(row=0, column=1, sticky="ns")
        og.rowconfigure(0, weight=1)
        og.columnconfigure(0, weight=1)
        self.orders_tree.bind("<<TreeviewSelect>>", self._on_order_select)

        # --- Pannello aggiunta pallet ---
        addp = ttk.LabelFrame(main, text=self.lang.get("add_pallet", "Aggiungi pallet alla spedizione"), padding=8)
        addp.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(addp, text=self.lang.get("pallet_code", "Codice Pallet:")).grid(row=0, column=0, sticky=tk.W)
        self.pallet_var = tk.StringVar()
        self.pallet_entry = ttk.Entry(addp, textvariable=self.pallet_var, width=14)
        self.pallet_entry.grid(row=0, column=1, padx=(4, 16))
        ttk.Label(addp, text=self.lang.get("shipped_qty", "Qta spedita:")).grid(row=0, column=2, sticky=tk.W)
        self.add_qty_var = tk.StringVar()
        self.add_qty_entry = ttk.Entry(addp, textvariable=self.add_qty_var, width=10)
        self.add_qty_entry.grid(row=0, column=3, padx=(4, 16))
        self.btn_add = ttk.Button(addp, text=self.lang.get("btn_add_to_shipment", "Aggiungi a spedizione"),
                                  command=self._add_pallet, state=tk.DISABLED)
        self.btn_add.grid(row=0, column=4, padx=4)
        self.add_hint = ttk.Label(addp, text="", foreground="gray")
        self.add_hint.grid(row=0, column=5, sticky=tk.W, padx=(12, 0))

        # --- Griglia pallet della spedizione ---
        pg = ttk.LabelFrame(main, text=self.lang.get("shipment_pallets", "Pallet della spedizione"), padding=6)
        pg.pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        self.pallets_tree = ttk.Treeview(pg, columns=_PALLET_COLS, show="headings",
                                         selectmode="browse", height=7)
        for hidden in ("ShipmentPalletId", "DPOID"):
            self.pallets_tree.column(hidden, width=0, stretch=False)
            self.pallets_tree.heading(hidden, text="")
        for col in _PALLET_COLS[2:]:
            label, width = _PALLET_LABELS[col]
            self.pallets_tree.heading(col, text=self.lang.get(f"pcol_{col.lower()}", label))
            anchor = tk.W if col == "ItemName" else tk.CENTER
            self.pallets_tree.column(col, width=width, anchor=anchor)
        pvsb = ttk.Scrollbar(pg, orient=tk.VERTICAL, command=self.pallets_tree.yview)
        self.pallets_tree.configure(yscrollcommand=pvsb.set)
        self.pallets_tree.grid(row=0, column=0, sticky="nsew")
        pvsb.grid(row=0, column=1, sticky="ns")
        pg.rowconfigure(0, weight=1)
        pg.columnconfigure(0, weight=1)

        pbtns = ttk.Frame(pg)
        pbtns.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(pbtns, text=self.lang.get("btn_edit_pallet", "Modifica qta"),
                   command=self._edit_pallet).pack(side=tk.LEFT, padx=4)
        ttk.Button(pbtns, text=self.lang.get("btn_delete_pallet", "Elimina"),
                   command=self._delete_pallet).pack(side=tk.LEFT, padx=4)

        # --- Bottoni azione + status ---
        actions = ttk.Frame(main)
        actions.pack(fill=tk.X, pady=(2, 0))
        self.btn_finalize = ttk.Button(actions, text=self.lang.get("btn_finalize", "Finalizza spedizione"),
                                       command=self._finalize)
        self.btn_finalize.pack(side=tk.LEFT, padx=4)
        self.btn_save_corr = ttk.Button(actions, text=self.lang.get("btn_save_corrections", "Salva correzioni e ristampa"),
                                        command=self._save_corrections)
        # mostrato solo in modalita' recover

        self.status_var = tk.StringVar(value="")
        ttk.Label(main, textvariable=self.status_var, foreground="gray").pack(fill=tk.X, pady=(6, 0))

    # ------------------------------------------------------------------ #
    #  Spedizione corrente: resume / create
    # ------------------------------------------------------------------ #
    def _resume_or_create_shipment(self):
        """Riprende l'ultima spedizione OPEN o ne crea una nuova."""
        try:
            self.db.cursor.execute(
                """
                SELECT TOP 1 ShipmentId, ShipmentDate
                FROM [Traceability_RS].[dyn].[Shipments]
                WHERE Status = 'OPEN'
                ORDER BY CreatedAt DESC
                """
            )
            row = self.db.cursor.fetchone()
            if row:
                self.shipment_id = int(row.ShipmentId)
                self.shipment_status = "OPEN"
                if row.ShipmentDate:
                    self.date_entry.set_date(row.ShipmentDate)
                logger.info(f"Ripresa spedizione OPEN #{self.shipment_id}")
            else:
                self.db.cursor.execute(
                    """
                    INSERT INTO [Traceability_RS].[dyn].[Shipments]
                        (ShipmentDate, Status, CreatedByUser, CreatedAt)
                    OUTPUT INSERTED.ShipmentId
                    VALUES (?, 'OPEN', ?, GETDATE())
                    """,
                    (date.today(), self.user_name),
                )
                self.shipment_id = int(self.db.cursor.fetchone()[0])
                self.shipment_status = "OPEN"
                self.db.conn.commit()
                logger.info(f"Creata nuova spedizione OPEN #{self.shipment_id}")
        except Exception as e:
            logger.error(f"Errore apertura/creazione spedizione: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)

    def _set_mode(self, mode):
        self.mode = mode
        if mode == "current":
            self.lbl_mode.config(text="")
            self.btn_save_corr.pack_forget()
            self.btn_finalize.pack(side=tk.LEFT, padx=4)
        else:  # recover
            self.lbl_mode.config(
                text=self.lang.get("mode_correction", "MODALITA' CORREZIONE"))
            self.btn_finalize.pack_forget()
            self.btn_save_corr.pack(side=tk.LEFT, padx=4)
        self._update_shipment_label()

    def _update_shipment_label(self):
        self.lbl_shipment.config(
            text=self.lang.get("shipment_n", "Spedizione N. {0}  [{1}]").format(
                self.shipment_id or "-", self.shipment_status or "-")
        )

    # ------------------------------------------------------------------ #
    #  Caricamento dati
    # ------------------------------------------------------------------ #
    def _refresh_all(self):
        self._update_shipment_label()
        if self.mode == "current":
            self._set_mode("current")
        self._load_orders()
        self._load_pallets()

    def _default_date_range(self):
        """Intervallo predefinito del filtro 'Data da spedire': ultimi 30 giorni
        fino a +30 giorni (finestra pratica attorno a oggi)."""
        today = date.today()
        return today - timedelta(days=30), today + timedelta(days=30)

    def _clear_filters(self):
        self.filter_order_var.set("")
        self.filter_product_var.set("")
        _from, _to = self._default_date_range()
        self.filter_from_entry.set_date(_from)
        self.filter_to_entry.set_date(_to)
        self._load_orders()

    # Colonne esportate: tutte tranne DPOID, che e' tecnica e a video ha
    # larghezza 0. _ORDER_COLS[0] e' proprio DPOID.
    _EXPORT_COLS = _ORDER_COLS[1:]
    _EXPORT_NUMERIC = ("Versato", "Spedito", "Residual")

    def _export_orders_excel(self):
        """Esporta in Excel esattamente le righe attualmente visibili a video.

        Legge dal Treeview e non dal database: cosi' l'export corrisponde
        sempre a cio' che l'utente sta guardando, filtri compresi, senza
        rieseguire la query e rischiare di ottenere un risultato diverso.
        """
        import os
        from tkinter import filedialog
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            messagebox.showerror(self.lang.get("error", "Errore"),
                                 f"Modulo openpyxl non disponibile: {e}", parent=self)
            return

        iids = self.orders_tree.get_children("")
        if not iids:
            messagebox.showinfo(
                self.lang.get("info", "Informazione"),
                self.lang.get("no_data_to_export", "Nessuna riga da esportare."),
                parent=self)
            return

        # Intestazioni: quelle realmente a video, cosi' seguono la lingua scelta
        headers = [self.orders_tree.heading(c, "text") for c in self._EXPORT_COLS]

        default_name = f"Conferma_Spedizioni_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self,
            title=self.lang.get("btn_export_excel", "Esporta Excel"),
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.lang.get("sheet_shipments", "Spedizioni")

            head_font = Font(bold=True, color="FFFFFF", size=11)
            head_fill = PatternFill(start_color="366092", end_color="366092",
                                    fill_type="solid")
            thin = Side(style="thin", color="D0D0D0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Riga 1 = intestazioni, perche' e' li' che va l'autofiltro
            for c, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=title)
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border = border

            widths = [len(h) for h in headers]
            for r, iid in enumerate(iids, start=2):
                for c, col in enumerate(self._EXPORT_COLS, start=1):
                    # .set() ritorna sempre str: evita che Tcl converta un
                    # codice come "00123" nell'intero 123
                    raw = self.orders_tree.set(iid, col)
                    value = raw

                    if col in self._EXPORT_NUMERIC:
                        try:
                            value = int(str(raw).strip() or 0)
                        except ValueError:
                            value = raw
                    elif col == "DateToShip" and raw:
                        # A video e' gia' dd/MM/yyyy: la riconverto in data vera
                        # cosi' in Excel si ordina e si filtra come data.
                        try:
                            value = datetime.strptime(raw, "%d/%m/%Y").date()
                        except ValueError:
                            value = raw

                    cell = ws.cell(row=r, column=c, value=value)
                    cell.border = border
                    if col in self._EXPORT_NUMERIC:
                        cell.alignment = Alignment(horizontal="center")
                    elif col == "DateToShip":
                        cell.alignment = Alignment(horizontal="center")
                        if hasattr(value, "year"):
                            cell.number_format = "DD/MM/YYYY"
                    widths[c - 1] = max(widths[c - 1], len(str(raw)))

            # Autofiltro sulla riga delle intestazioni + blocco della stessa
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            ws.freeze_panes = "A2"

            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(w + 3, 10), 45)
            ws.row_dimensions[1].height = 28

            wb.save(path)
        except PermissionError:
            messagebox.showerror(
                self.lang.get("error", "Errore"),
                self.lang.get("export_file_locked",
                              "Impossibile scrivere il file: e' aperto in Excel?"),
                parent=self)
            return
        except Exception as e:
            logger.error("Errore export Excel conferma spedizioni: %s", e, exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"),
                                 f"{self.lang.get('export_error', 'Errore durante l export')}:\n{e}",
                                 parent=self)
            return

        self.status_var.set(
            self.lang.get("export_ok", "Esportate {n} righe in {path}").format(
                n=len(iids), path=path))
        try:
            os.startfile(path)
        except Exception:
            pass

    def _load_orders(self):
        """Carica TUTTI i codici versati a magazzino e non ancora spediti nel
        periodo (residuo = versato − già spedito). Gli ordini urgenti (con regola
        di spedizione aperta) sono evidenziati con il testo ROSSO."""
        for it in self.orders_tree.get_children():
            self.orders_tree.delete(it)
        self._order_info = {}
        self.orders_tree.tag_configure("urgent", foreground="#C0392B")  # testo rosso

        d_from = self.filter_from_entry.get_date()
        d_to = self.filter_to_entry.get_date()
        if d_from > d_to:
            messagebox.showwarning(
                self.lang.get("warning", "Attenzione"),
                self.lang.get("filter_date_order",
                              "La data iniziale deve precedere la data finale."),
                parent=self)
            return

        ofil = self.filter_order_var.get().strip().lower()
        pfil = self.filter_product_var.get().strip().lower()

        # Batch multi-istruzione → cursore DEDICATO, chiuso subito.
        cur = None
        try:
            cur = self.db.conn.cursor()
            cur.execute(_QUERY_WAREHOUSE, (d_from, d_to))
            rows = cur.fetchall()
        except Exception as e:
            logger.error(f"Errore caricamento ordini versati: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return
        finally:
            if cur is not None:
                try:
                    cur.close()
                except Exception:
                    pass

        n = 0
        for r in rows:
            order = r.OrderNumber or ""
            prod = r.ProductCode or ""
            if ofil and ofil not in order.lower():
                continue
            if pfil and pfil not in prod.lower():
                continue
            versato = int(r.WhQty or 0)
            spedito = int(r.ShippedQty or 0)
            residual = int(r.ResiduoDaSpedire or 0)
            idorder = str(r.IDOrder)
            self._order_info[idorder] = {
                "id_order": int(r.IDOrder),
                "production_order": order,
                "product_code": prod,
                "versato": versato,
                "spedito": spedito,
                "residual": residual,
                "matched_qty": int(r.MatchedQty or 0),
                "is_urgent": bool(r.IsUrgent),
            }
            last = r.LastWhDate.strftime("%d/%m/%Y") if r.LastWhDate else ""
            self.orders_tree.insert(
                "", tk.END,
                values=(r.IDOrder, order, prod, versato, spedito, residual, last),
                tags=(("urgent",) if r.IsUrgent else ()),
            )
            n += 1
        self.status_var.set(
            self.lang.get("wh_orders_count_period",
                          "{0} codici versati non spediti (periodo {1} - {2})").format(
                n, d_from.strftime("%d/%m/%Y"), d_to.strftime("%d/%m/%Y")))

    def _load_pallets(self):
        for it in self.pallets_tree.get_children():
            self.pallets_tree.delete(it)
        if not self.shipment_id:
            return
        try:
            self.db.cursor.execute(
                """
                SELECT ShipmentPalletId, DynamicProductionOrderID, PalletCode,
                       ProductionOrderNumber, ItemCode, ItemName, ConfirmedQty
                FROM [Traceability_RS].[dyn].[ShipmentPallets]
                WHERE ShipmentId = ?
                ORDER BY PalletCode, ProductionOrderNumber
                """,
                (self.shipment_id,),
            )
            for r in self.db.cursor.fetchall():
                self.pallets_tree.insert(
                    "", tk.END,
                    values=(r.ShipmentPalletId, r.DynamicProductionOrderID,
                            r.PalletCode or "", r.ProductionOrderNumber or "",
                            r.ItemCode or "", r.ItemName or "", int(r.ConfirmedQty or 0)),
                )
        except Exception as e:
            logger.error(f"Errore caricamento pallet: {e}", exc_info=True)

        self._suggest_pallet_code()

    def _suggest_pallet_code(self):
        """Propone il prossimo codice pallet progressivo entro la spedizione."""
        codes = [self.pallets_tree.set(it, "PalletCode")
                 for it in self.pallets_tree.get_children()]
        max_num = 0
        for c in codes:
            cs = str(c).strip()
            if cs.isdigit():
                max_num = max(max_num, int(cs))
        suggestion = str(max_num + 1)
        # Suggerisci solo se il campo e' vuoto (non sovrascrivere quanto digitato)
        if not self.pallet_var.get().strip():
            self.pallet_var.set(suggestion)

    # ------------------------------------------------------------------ #
    #  Selezione ordine
    # ------------------------------------------------------------------ #
    def _on_order_select(self, _event=None):
        sel = self.orders_tree.selection()
        if not sel:
            self.btn_add.config(state=tk.DISABLED)
            self.add_hint.config(text="")
            return
        idorder = str(self.orders_tree.set(sel[0], "IDOrder"))
        info = self._order_info.get(idorder, {})
        self.add_hint.config(
            text=self.lang.get(
                "wh_order_hint",
                "versato: {0} | già spedito: {1} | residuo: {2}",
            ).format(info.get("versato", 0), info.get("spedito", 0), info.get("residual", 0))
        )
        residual = info.get("residual", 0)
        self.add_qty_var.set(str(residual if residual > 0 else 0))
        self.btn_add.config(state=tk.NORMAL)

    # ------------------------------------------------------------------ #
    #  Aggiunta pallet
    # ------------------------------------------------------------------ #
    def _add_pallet(self):
        sel = self.orders_tree.selection()
        if not sel:
            return
        idorder = str(self.orders_tree.set(sel[0], "IDOrder"))
        info = self._order_info.get(idorder)
        if not info:
            return

        pallet_code = self.pallet_var.get().strip()
        if not pallet_code:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("pallet_required", "Inserire il codice pallet."),
                                   parent=self)
            return
        qty_str = self.add_qty_var.get().strip()
        try:
            qty = int(qty_str)
        except ValueError:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("qty_invalid", "Quantita' non valida."),
                                   parent=self)
            return
        if qty <= 0:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("qty_positive", "La quantita' deve essere maggiore di zero."),
                                   parent=self)
            return

        # Avviso eccesso vs residuo da spedire (versato − già spedito) — NON bloccante
        if qty > info["residual"]:
            diff = qty - info["residual"]
            if not messagebox.askyesno(
                self.lang.get("confirm", "Conferma"),
                self.lang.get(
                    "wh_excess_warn",
                    "Attenzione: la quantita' ({0}) supera il residuo da spedire ({1}) "
                    "di {2} pezzi.\n\nContinuare comunque?",
                ).format(qty, info["residual"], diff),
                parent=self,
            ):
                return

        # L'operatore assegna il SO (dello stesso prodotto) → risolve/crea il DPO.
        dpo = self._resolve_dpo(info, qty)
        if not dpo:
            return  # annullato
        dpoid = dpo["dpoid"]

        try:
            # Esiste gia' (stessa spedizione, stesso pallet, stesso DPO)? -> somma
            self.db.cursor.execute(
                """
                SELECT ShipmentPalletId, ConfirmedQty
                FROM [Traceability_RS].[dyn].[ShipmentPallets]
                WHERE ShipmentId = ? AND PalletCode = ? AND DynamicProductionOrderID = ?
                """,
                (self.shipment_id, pallet_code, dpoid),
            )
            existing = self.db.cursor.fetchone()
            if existing:
                self.db.cursor.execute(
                    """
                    UPDATE [Traceability_RS].[dyn].[ShipmentPallets]
                    SET ConfirmedQty = ConfirmedQty + ?,
                        ConfirmedByUser = ?, ConfirmedAt = GETDATE()
                    WHERE ShipmentPalletId = ?
                    """,
                    (qty, self.user_name, existing.ShipmentPalletId),
                )
            else:
                self.db.cursor.execute(
                    """
                    INSERT INTO [Traceability_RS].[dyn].[ShipmentPallets]
                        (ShipmentId, PalletCode, DynamicProductionOrderID, ConfirmedQty,
                         ConfirmedByUser, ConfirmedAt,
                         ProductionOrderNumber, SONumber, CustomerName, ItemCode, ItemName, ShipTo)
                    VALUES (?, ?, ?, ?, ?, GETDATE(), ?, ?, ?, ?, ?, ?)
                    """,
                    (self.shipment_id, pallet_code, dpoid, qty, self.user_name,
                     info["production_order"], dpo["so_number"], dpo["customer"],
                     info["product_code"], dpo["item_name"], dpo["ship_to"]),
                )
            self.db.conn.commit()
            logger.info(f"Pallet {pallet_code} +{qty} (ordine {info['production_order']}, "
                        f"SO {dpo['so_number']}) su spedizione #{self.shipment_id}")
        except Exception as e:
            logger.error(f"Errore aggiunta pallet: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return

        self.pallet_var.set("")          # azzera per riproporre il suggerimento
        self.add_qty_var.set("")
        self._load_orders()
        self._load_pallets()

    def _resolve_dpo(self, info, qty):
        """Fa scegliere all'operatore il SO (dello stesso prodotto) contro cui
        spedire questo ordine, e ritorna il DynamicProductionOrder (creandolo se
        non esiste). Ritorna dict {dpoid, so_number, customer, item_name, ship_to}
        oppure None se annullato."""
        dlg = _SOAssignDialog(self, self.db, self.lang, self.user_name, info, qty)
        self.wait_window(dlg)
        return getattr(dlg, "result", None)

    def _selected_pallet(self):
        sel = self.pallets_tree.selection()
        if not sel:
            messagebox.showinfo(self.lang.get("info", "Info"),
                                self.lang.get("select_pallet", "Seleziona un pallet nella lista."),
                                parent=self)
            return None
        return sel[0]

    def _edit_pallet(self):
        item = self._selected_pallet()
        if not item:
            return
        spid = self.pallets_tree.set(item, "ShipmentPalletId")
        current_qty = self.pallets_tree.set(item, "Qty")
        pallet_code = self.pallets_tree.set(item, "PalletCode")

        dlg = _QtyDialog(self, self.lang, pallet_code, current_qty)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        new_qty = dlg.result
        try:
            self.db.cursor.execute(
                """
                UPDATE [Traceability_RS].[dyn].[ShipmentPallets]
                SET ConfirmedQty = ?, ConfirmedByUser = ?, ConfirmedAt = GETDATE()
                WHERE ShipmentPalletId = ?
                """,
                (new_qty, self.user_name, spid),
            )
            self.db.conn.commit()
        except Exception as e:
            logger.error(f"Errore modifica pallet: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return
        self._load_orders()
        self._load_pallets()

    def _delete_pallet(self):
        item = self._selected_pallet()
        if not item:
            return
        spid = self.pallets_tree.set(item, "ShipmentPalletId")
        if not messagebox.askyesno(self.lang.get("confirm", "Conferma"),
                                   self.lang.get("delete_pallet_q", "Eliminare questa riga pallet?"),
                                   parent=self):
            return
        try:
            self.db.cursor.execute(
                "DELETE FROM [Traceability_RS].[dyn].[ShipmentPallets] WHERE ShipmentPalletId = ?",
                (spid,),
            )
            self.db.conn.commit()
        except Exception as e:
            logger.error(f"Errore eliminazione pallet: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return
        self._load_orders()
        self._load_pallets()

    # ------------------------------------------------------------------ #
    #  Conteggio righe pallet
    # ------------------------------------------------------------------ #
    def _pallet_count(self):
        return len(self.pallets_tree.get_children())

    def _get_shipment_date(self):
        try:
            return self.date_entry.get_date()
        except Exception:
            return date.today()

    # ------------------------------------------------------------------ #
    #  Finalizza (modalita' current)
    # ------------------------------------------------------------------ #
    def _finalize(self):
        if self._pallet_count() == 0:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("no_pallets", "Nessun pallet inserito nella spedizione."),
                                   parent=self)
            return
        if not messagebox.askyesno(
            self.lang.get("confirm", "Conferma"),
            self.lang.get("finalize_q", "Finalizzare la spedizione N. {0}?\nVerranno generati i documenti e inviata l'email.").format(self.shipment_id),
            parent=self,
        ):
            return

        ship_date = self._get_shipment_date()
        try:
            self.db.cursor.execute(
                """
                UPDATE [Traceability_RS].[dyn].[Shipments]
                SET Status = 'CLOSED', ShipmentDate = ?, ClosedByUser = ?, ClosedAt = GETDATE()
                WHERE ShipmentId = ?
                """,
                (ship_date, self.user_name, self.shipment_id),
            )
            self.db.conn.commit()
            self.shipment_status = "CLOSED"
        except Exception as e:
            logger.error(f"Errore finalizzazione: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return

        self._generate_and_dispatch(is_correction=False)
        messagebox.showinfo(self.lang.get("success", "Successo"),
                            self.lang.get("shipment_finalized", "Spedizione finalizzata. Documenti generati ed email in invio."),
                            parent=self)
        # Pronti per una nuova spedizione
        self.shipment_id = None
        self._resume_or_create_shipment()
        self._set_mode("current")
        self._refresh_all()

    # ------------------------------------------------------------------ #
    #  Correzione (modalita' recover)
    # ------------------------------------------------------------------ #
    def _save_corrections(self):
        if self._pallet_count() == 0:
            if not messagebox.askyesno(
                self.lang.get("confirm", "Conferma"),
                self.lang.get("corr_empty_q", "La spedizione non ha pallet. Salvare comunque la correzione?"),
                parent=self,
            ):
                return
        ship_date = self._get_shipment_date()
        try:
            self.db.cursor.execute(
                """
                UPDATE [Traceability_RS].[dyn].[Shipments]
                SET Status = 'CORRECTED', ShipmentDate = ?,
                    LastModifiedByUser = ?, LastModifiedAt = GETDATE()
                WHERE ShipmentId = ?
                """,
                (ship_date, self.user_name, self.shipment_id),
            )
            self.db.conn.commit()
            self.shipment_status = "CORRECTED"
        except Exception as e:
            logger.error(f"Errore salvataggio correzione: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return

        self._generate_and_dispatch(is_correction=True)
        messagebox.showinfo(self.lang.get("success", "Successo"),
                            self.lang.get("corr_saved", "Correzione salvata. Documenti rigenerati ed email di correzione in invio."),
                            parent=self)
        self._update_shipment_label()

    # ------------------------------------------------------------------ #
    #  PDF + email
    # ------------------------------------------------------------------ #
    def _generate_and_dispatch(self, is_correction: bool):
        """Genera i 2 PDF (main thread), salva i path, invia email (thread)."""
        pallet_pdf = summary_pdf = None
        try:
            pallet_pdf, summary_pdf = shipment_pdf.generate_shipment_documents(
                self.db, self.shipment_id, self.user_name)
            self.db.cursor.execute(
                """
                UPDATE [Traceability_RS].[dyn].[Shipments]
                SET PdfPalletPath = ?, PdfSummaryPath = ?
                WHERE ShipmentId = ?
                """,
                (pallet_pdf, summary_pdf, self.shipment_id),
            )
            self.db.conn.commit()
        except Exception as e:
            logger.error(f"Errore generazione documenti: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"),
                                 self.lang.get("pdf_error", "Errore nella generazione dei documenti:\n{0}").format(e),
                                 parent=self)

        # Stampa locale best-effort
        for p in (pallet_pdf, summary_pdf):
            if p:
                shipment_pdf.print_pdf(p)

        # Snapshot dati per l'email (nel main thread)
        ship_date = self._get_shipment_date()
        pallets = self._collect_pallets_for_email()

        threading.Thread(
            target=self._send_shipment_email,
            args=(self.shipment_id, ship_date, pallets, is_correction,
                  [p for p in (pallet_pdf, summary_pdf) if p]),
            daemon=True,
        ).start()

    def _collect_pallets_for_email(self):
        rows = []
        for it in self.pallets_tree.get_children():
            rows.append({
                "pallet": self.pallets_tree.set(it, "PalletCode"),
                "order": self.pallets_tree.set(it, "ProductionOrder"),
                "item_code": self.pallets_tree.set(it, "ItemCode"),
                "item_name": self.pallets_tree.set(it, "ItemName"),
                "qty": self.pallets_tree.set(it, "Qty"),
            })
        return rows

    def _shipto_map(self, shipment_id):
        """Mappa (PalletCode, ProductionOrderNumber) -> ShipTo per la spedizione.
        La destinazione (Normal Shipment / Direct to final Customer) è snapshot su
        dyn.ShipmentPallets; serve a mostrarla nel corpo email (FASE 3)."""
        out = {}
        try:
            cur = self.db.conn.cursor()
            cur.execute(
                "SELECT PalletCode, ProductionOrderNumber, ShipTo "
                "FROM [Traceability_RS].[dyn].[ShipmentPallets] WHERE ShipmentId = ?",
                (shipment_id,))
            for r in cur.fetchall():
                out[(str(r[0] or ''), str(r[1] or ''))] = (r[2] or '').strip()
            cur.close()
        except Exception as e:
            logger.error(f"_shipto_map spedizione {shipment_id}: {e}", exc_info=True)
        return out

    def _gather_recipients(self, shipment_id):
        """Ritorna (am_emails, cc_emails) per la spedizione in base ai clienti finali
        coinvolti e ai toggle per cliente (dbo.ClientShipmentEmailPrefs):
          - TO: email degli Account Manager dei clienti finali con SendToAccountManager=1;
          - CC: settings 'Sys_email_<FinalClientName>' dei clienti finali con SendToCc=1.
        Riga prefs assente => entrambi attivi (default)."""
        def _dedupe_ci(items):
            seen, out = set(), []
            for x in items:
                x = (x or '').strip()
                if x and x.lower() not in seen:
                    seen.add(x.lower())
                    out.append(x)
            return out

        am, cc = [], []
        try:
            cur = self.db.conn.cursor()
            cur.execute(
                "SELECT DISTINCT fc.IDFinalClient, fc.FinalClientName "
                "FROM [Traceability_RS].[dyn].[ShipmentPallets] sp "
                "JOIN [Traceability_RS].[dyn].[DynamicProductionOrders] dpo "
                "  ON dpo.DynamicProductionOrderID = sp.DynamicProductionOrderID "
                "JOIN [Traceability_RS].[dbo].[Orders] o ON o.IDOrder = dpo.IdOrder "
                "JOIN [Traceability_RS].[dbo].[Products] p ON p.IDProduct = o.IDProduct "
                "JOIN [Traceability_RS].[dbo].[FinalClients] fc ON fc.IDFinalClient = p.IdFinalClient "
                "WHERE sp.ShipmentId = ?", (shipment_id,))
            clients = [(r[0], r[1]) for r in cur.fetchall()]
            for idfc, name in clients:
                cur.execute("SELECT SendToAccountManager, SendToCc "
                            "FROM [Traceability_RS].[dbo].[ClientShipmentEmailPrefs] WHERE IDFinalClient = ?",
                            (idfc,))
                pr = cur.fetchone()
                send_am = True if pr is None else bool(pr[0])
                send_cc = True if pr is None else bool(pr[1])
                if send_am:
                    cur.execute(
                        "SELECT cam.WorkEmail FROM [Traceability_RS].[dbo].[ClientAccountLists] cal "
                        "JOIN [Traceability_RS].[dbo].[ClientAccountManagers] cam "
                        "  ON cam.ClientAccountManagerId = cal.ClientAccountManagerId "
                        "WHERE cal.IDFinalClient = ? AND cal.Dateout IS NULL AND cam.Dateout IS NULL "
                        "  AND cam.WorkEmail IS NOT NULL", (idfc,))
                    am.extend([r[0] for r in cur.fetchall()])
                if send_cc and name:
                    cur.execute("SELECT Value FROM [Traceability_RS].[dbo].[Settings] WHERE Atribute = ?",
                                (f"Sys_email_{name}",))
                    row = cur.fetchone()
                    if row and row[0]:
                        cc.extend(str(row[0]).replace(',', ';').split(';'))
            cur.close()
        except Exception as e:
            logger.error(f"_gather_recipients spedizione {shipment_id}: {e}", exc_info=True)
        return _dedupe_ci(am), _dedupe_ci(cc)

    def _send_shipment_email(self, shipment_id, ship_date, pallets, is_correction, attachments):
        try:
            base = utils.get_email_recipients(self.db.conn, "Sys_shipment_email") or []
            am_emails, cc_emails = self._gather_recipients(shipment_id)
            # TO = lista globale spedizioni + account manager; dedup case-insensitive
            recipients, seen = [], set()
            for r in list(base) + am_emails:
                r = (r or '').strip()
                if r and r.lower() not in seen:
                    seen.add(r.lower())
                    recipients.append(r)
            # CC = Sys_email_<cliente>, escludendo chi e' gia' in TO
            cc_emails = [c for c in cc_emails if c.lower() not in seen]
            if not recipients:
                logger.warning("Nessun destinatario (Sys_shipment_email + account manager) - email non inviata")
                return

            ship_date_str = ship_date.strftime("%d/%m/%Y") if hasattr(ship_date, "strftime") else str(ship_date)
            total_qty = sum(int(p["qty"] or 0) for p in pallets)
            n_pallets = len({p["pallet"] for p in pallets})
            n_orders = len({p["order"] for p in pallets})

            banner = ("CORREZIONE - " if is_correction else "")
            subject = f"{banner}Conferma Spedizione N. {shipment_id} del {ship_date_str}"

            # FASE 3: destinazione (direct/normal) per riga, da dyn.ShipmentPallets.
            shipto_map = self._shipto_map(shipment_id)

            def _shipto_cell(ship_to):
                if not ship_to:
                    return "<td></td>"
                # Evidenzia la spedizione diretta al cliente finale.
                if 'direct' in ship_to.lower():
                    return (f"<td style='color:#c0392b;font-weight:bold;'>{ship_to}</td>")
                return f"<td>{ship_to}</td>"

            rows_html = "".join(
                f"<tr><td>{p['pallet']}</td><td>{p['order']}</td><td>{p['item_code']}</td>"
                f"<td>{p['item_name']}</td><td style='text-align:center;'>{p['qty']}</td>"
                f"{_shipto_cell(shipto_map.get((str(p['pallet']), str(p['order'])), ''))}</tr>"
                for p in pallets
            )
            corr_note = ""
            if is_correction:
                corr_note = ("<p style='color:#c0392b;font-weight:bold;'>"
                             "Questa e' una versione CORRETTA della spedizione: "
                             "sostituisce la precedente.</p>")

            body = f"""
<!DOCTYPE html><html><head><meta charset="utf-8"/>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:13px;color:#333;}}
table{{border-collapse:collapse;width:100%;max-width:760px;}}
th{{background:#1a5276;color:#fff;padding:8px 10px;text-align:left;}}
td{{padding:6px 10px;border-bottom:1px solid #ddd;}}
tr:nth-child(even){{background:#f8f8f8;}}
.header-box{{background:#1a5276;color:#fff;padding:16px;border-radius:4px;margin-bottom:16px;}}
.footer{{font-size:11px;color:#888;margin-top:18px;}}
</style></head><body>
<div class="header-box">
  <h2 style="margin:0;">{banner}Conferma Spedizione N. {shipment_id}</h2>
  <p style="margin:4px 0 0;">Data spedizione: <strong>{ship_date_str}</strong> -
     Operatore: <strong>{self.user_name}</strong></p>
  <p style="margin:4px 0 0;">Generata il {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
</div>
{corr_note}
<p>Totali: <strong>{n_pallets}</strong> pallet - <strong>{n_orders}</strong> ordini -
   <strong>{total_qty}</strong> pezzi.</p>
<table>
  <tr><th>Pallet</th><th>Ordine Prod.</th><th>Codice</th><th>Prodotto</th><th>Qta</th><th>Destinazione</th></tr>
  {rows_html}
</table>
<p class="footer">In allegato: lista per pallet e riepilogo spedizione (PDF).<br/>
Email generata automaticamente da TraceabilityRS - non rispondere.</p>
</body></html>"""

            utils.send_email(recipients, subject, body, is_html=True,
                             attachments=attachments, cc_emails=cc_emails or None)
            logger.info(f"Email spedizione #{shipment_id} inviata - TO={recipients} CC={cc_emails} "
                        f"(correzione={is_correction})")
        except Exception as e:
            logger.error(f"Errore invio email spedizione: {e}", exc_info=True)

    # ------------------------------------------------------------------ #
    #  Recupero / correzione spedizione
    # ------------------------------------------------------------------ #
    def _open_recover_dialog(self):
        dlg = _RecoverDialog(self, self.db, self.lang)
        self.wait_window(dlg)
        if dlg.selected_shipment_id is None:
            return
        self._load_shipment_for_correction(dlg.selected_shipment_id)

    def _load_shipment_for_correction(self, shipment_id):
        try:
            self.db.cursor.execute(
                """
                SELECT ShipmentId, ShipmentDate, Status
                FROM [Traceability_RS].[dyn].[Shipments]
                WHERE ShipmentId = ?
                """,
                (shipment_id,),
            )
            row = self.db.cursor.fetchone()
            if not row:
                return
            self.shipment_id = int(row.ShipmentId)
            self.shipment_status = row.Status
            if row.ShipmentDate:
                self.date_entry.set_date(row.ShipmentDate)
        except Exception as e:
            logger.error(f"Errore caricamento spedizione {shipment_id}: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            return
        self._set_mode("recover")
        self._load_orders()
        self._load_pallets()
        self.status_var.set(
            self.lang.get("loaded_shipment", "Caricata spedizione N. {0} per correzione").format(shipment_id))

    # ------------------------------------------------------------------ #
    #  Chiusura finestra
    # ------------------------------------------------------------------ #
    def _on_close(self):
        # Pulisci la spedizione OPEN se vuota (evita header orfani)
        try:
            if (self.mode == "current" and self.shipment_id
                    and self.shipment_status == "OPEN" and self._pallet_count() == 0):
                self.db.cursor.execute(
                    "DELETE FROM [Traceability_RS].[dyn].[Shipments] WHERE ShipmentId = ? AND Status = 'OPEN'",
                    (self.shipment_id,),
                )
                self.db.conn.commit()
                logger.info(f"Rimossa spedizione OPEN vuota #{self.shipment_id}")
        except Exception as e:
            logger.warning(f"Cleanup spedizione vuota fallito: {e}")
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
#  Dialog: modifica quantita'
# ─────────────────────────────────────────────────────────────────────────────
class _SOAssignDialog(tk.Toplevel):
    """Assegna l'ordine di produzione (versato a magazzino) a un SO dello STESSO
    prodotto prima di confermare la spedizione. Elenca i SO candidati (chiave
    esatta ItemCode = ProductCode senza suffisso |N) con il loro residuo e
    l'eventuale abbinamento gia' esistente per questo ordine. Alla conferma crea
    (o riusa) il DynamicProductionOrder e ritorna il DPO su self.result."""

    def __init__(self, master, db, lang, user_name, info, qty):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.info = info
        self.qty = qty
        self.result = None

        L = self.lang.get
        self.title(L("so_assign_title", "Assegna SO e conferma"))
        self.geometry("780x430")
        self.transient(master)
        self.grab_set()

        self._rows = {}   # iid -> candidate dict
        self._build()
        self._load()

    def _product_key(self):
        return (self.info.get("product_code") or "").split("|")[0].strip()

    def _build(self):
        L = self.lang.get
        top = ttk.Frame(self, padding=10)
        top.pack(fill=tk.X)
        ttk.Label(top, font=("Segoe UI", 10, "bold"),
                  text=L("so_assign_head", "Ordine {0} — Prodotto {1}  |  da spedire ora: {2}").format(
                      self.info.get("production_order", ""), self.info.get("product_code", ""), self.qty)
                  ).pack(anchor=tk.W)
        ttk.Label(top, foreground="#555",
                  text=L("so_assign_hint",
                         "Seleziona il SO (stesso prodotto) contro cui spedire, poi conferma.")
                  ).pack(anchor=tk.W)

        cols = ("SO", "Customer", "ItemName", "Ordinato", "Assegnato", "ResiduoSO", "Linked")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", selectmode="browse", height=10)
        heads = {"SO": ("SO", 110), "Customer": ("Cliente", 150), "ItemName": ("Descrizione", 210),
                 "Ordinato": ("Ordinato", 80), "Assegnato": ("Assegnato", 80),
                 "ResiduoSO": ("Residuo SO", 85), "Linked": ("Abbinato", 80)}
        for c in cols:
            lbl, w = heads[c]
            self.tree.heading(c, text=L(f"soc_{c.lower()}", lbl))
            self.tree.column(c, width=w, anchor=(tk.W if c in ("Customer", "ItemName") else tk.CENTER))
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)

        bar = ttk.Frame(self, padding=10)
        bar.pack(fill=tk.X)
        ttk.Label(bar, text=L("so_assign_qty", "Qta da assegnare al SO:")).pack(side=tk.LEFT)
        self.qty_var = tk.StringVar(value=str(self.qty))
        ttk.Entry(bar, textvariable=self.qty_var, width=8).pack(side=tk.LEFT, padx=(4, 12))
        ttk.Button(bar, text=L("btn_confirm", "Conferma"), command=self._ok).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text=L("cancel", "Annulla"), command=self.destroy).pack(side=tk.RIGHT, padx=4)

    def _load(self):
        self.tree.delete(*self.tree.get_children())
        self._rows = {}
        key = self._product_key()
        try:
            self.db.cursor.execute(_QUERY_SO_CANDIDATES, (self.info["id_order"], key))
            rows = self.db.cursor.fetchall()
        except Exception as e:
            logger.error(f"Errore caricamento SO candidati: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)
            rows = []
        for r in rows:
            ordinato = int(r.QtyOrder or 0)
            assegnato = int(r.Assigned or 0)
            residuo_so = ordinato - assegnato
            linked = r.DpoForThisOrder is not None
            if residuo_so <= 0 and not linked:
                continue  # SO pieno e non gia' abbinato a questo ordine
            iid = self.tree.insert("", tk.END, values=(
                r.SONumber or "", r.CustomerName or "", r.ItemName or "",
                ordinato, assegnato, residuo_so,
                self.lang.get("yes", "Sì") if linked else ""))
            self._rows[iid] = {"so_id": r.DynamicSaleOrderId, "so_number": r.SONumber or "",
                               "customer": r.CustomerName or "", "item_name": r.ItemName or "",
                               "residuo_so": residuo_so, "dpo": r.DpoForThisOrder}
        if not self._rows:
            messagebox.showinfo(
                self.lang.get("info", "Info"),
                self.lang.get("so_none",
                              "Nessun SO disponibile per questo prodotto ({0}). "
                              "Verifica gli ordini di vendita importati.").format(key),
                parent=self)

    def _ok(self):
        L = self.lang.get
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(L("info", "Info"), L("so_pick", "Seleziona un SO."), parent=self)
            return
        cand = self._rows.get(sel[0])
        try:
            aqty = int(self.qty_var.get().strip())
        except ValueError:
            messagebox.showwarning(L("warning", "Attenzione"),
                                   L("qty_invalid", "Quantita' non valida."), parent=self)
            return
        if aqty <= 0:
            messagebox.showwarning(L("warning", "Attenzione"),
                                   L("qty_positive", "La quantita' deve essere maggiore di zero."), parent=self)
            return
        dpoid = cand["dpo"]
        try:
            if dpoid is None:
                if aqty > cand["residuo_so"]:
                    if not messagebox.askyesno(
                        L("confirm", "Conferma"),
                        L("so_over", "La quantita' ({0}) supera il residuo del SO ({1}). Continuare?").format(
                            aqty, cand["residuo_so"]), parent=self):
                        return
                self.db.cursor.execute(
                    "INSERT INTO Traceability_RS.dyn.DynamicProductionOrders "
                    "(DynamicSaleOrderId, IdOrder, Qty, DateIn) "
                    "OUTPUT INSERTED.DynamicProductionOrderID VALUES (?, ?, ?, GETDATE())",
                    (cand["so_id"], self.info["id_order"], aqty))
                dpoid = int(self.db.cursor.fetchone()[0])
                self.db.conn.commit()
                logger.info(f"Creato DPO {dpoid}: ordine {self.info['production_order']} "
                            f"↔ SO {cand['so_number']} qty {aqty}")
        except Exception as e:
            logger.error(f"Errore creazione abbinamento SO: {e}", exc_info=True)
            self.db.conn.rollback()
            messagebox.showerror(L("error", "Errore"), str(e), parent=self)
            return
        self.result = {"dpoid": str(dpoid), "so_number": cand["so_number"],
                       "customer": cand["customer"], "item_name": cand["item_name"],
                       "ship_to": "Normal Shipment"}
        self.destroy()


class _QtyDialog(tk.Toplevel):
    def __init__(self, master, lang, pallet_code, current_qty):
        super().__init__(master)
        self.lang = lang
        self.result = None
        self.title(lang.get("edit_qty_title", "Modifica quantita'"))
        self.transient(master)
        self.grab_set()
        self.resizable(False, False)

        frm = ttk.Frame(self, padding=14)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text=lang.get("pallet_code", "Codice Pallet:") + f" {pallet_code}").grid(
            row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 8))
        ttk.Label(frm, text=lang.get("new_qty", "Nuova quantita':")).grid(row=1, column=0, sticky=tk.W)
        self.var = tk.StringVar(value=str(current_qty))
        ent = ttk.Entry(frm, textvariable=self.var, width=10)
        ent.grid(row=1, column=1, padx=(6, 0))
        ent.focus_set()
        ent.select_range(0, tk.END)

        btns = ttk.Frame(frm)
        btns.grid(row=2, column=0, columnspan=2, pady=(12, 0))
        ttk.Button(btns, text=lang.get("ok", "OK"), command=self._ok).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text=lang.get("cancel", "Annulla"), command=self.destroy).pack(side=tk.LEFT, padx=4)
        ent.bind("<Return>", lambda _e: self._ok())

    def _ok(self):
        try:
            v = int(self.var.get().strip())
        except ValueError:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("qty_invalid", "Quantita' non valida."), parent=self)
            return
        if v <= 0:
            messagebox.showwarning(self.lang.get("warning", "Attenzione"),
                                   self.lang.get("qty_positive", "La quantita' deve essere maggiore di zero."), parent=self)
            return
        self.result = v
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
#  Dialog: recupero spedizione per data
# ─────────────────────────────────────────────────────────────────────────────
class _RecoverDialog(tk.Toplevel):
    def __init__(self, master, db, lang):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.selected_shipment_id = None

        self.title(lang.get("recover_title", "Recupera spedizione"))
        self.geometry("760x420")
        self.transient(master)
        self.grab_set()

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        flt = ttk.Frame(frm)
        flt.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(flt, text=lang.get("date_from", "Dal:")).pack(side=tk.LEFT)
        self.from_entry = DateEntry(flt, width=12, date_pattern="dd/mm/yyyy")
        self.from_entry.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(flt, text=lang.get("date_to", "Al:")).pack(side=tk.LEFT)
        self.to_entry = DateEntry(flt, width=12, date_pattern="dd/mm/yyyy")
        self.to_entry.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Button(flt, text=lang.get("btn_search", "Cerca"), command=self._search).pack(side=tk.LEFT, padx=4)

        cols = ("ShipmentId", "ShipmentDate", "Status", "Pallets", "Qty", "ClosedBy")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        _rcols = {
            "ShipmentId":   ("rcol_id", "N.", 60),
            "ShipmentDate": ("rcol_date", "Data", 100),
            "Status":       ("rcol_status", "Stato", 90),
            "Pallets":      ("rcol_pallets", "Pallet", 70),
            "Qty":          ("rcol_qty", "Pezzi", 70),
            "ClosedBy":     ("rcol_closedby", "Chiusa da", 160),
        }
        for c, (key, lbl, w) in _rcols.items():
            self.tree.heading(c, text=lang.get(key, lbl))
            self.tree.column(c, width=w, anchor=tk.CENTER if c != "ClosedBy" else tk.W)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", lambda _e: self._load())

        btns = ttk.Frame(frm)
        btns.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(btns, text=lang.get("btn_load", "Carica per correzione"), command=self._load).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text=lang.get("cancel", "Annulla"), command=self.destroy).pack(side=tk.RIGHT, padx=4)

        self._search()

    def _search(self):
        for it in self.tree.get_children():
            self.tree.delete(it)
        try:
            self.db.cursor.execute(
                """
                SELECT s.ShipmentId, s.ShipmentDate, s.Status, s.ClosedByUser,
                       (SELECT COUNT(DISTINCT sp.PalletCode)
                        FROM [Traceability_RS].[dyn].[ShipmentPallets] sp
                        WHERE sp.ShipmentId = s.ShipmentId) AS Pallets,
                       ISNULL((SELECT SUM(sp.ConfirmedQty)
                        FROM [Traceability_RS].[dyn].[ShipmentPallets] sp
                        WHERE sp.ShipmentId = s.ShipmentId), 0) AS Qty
                FROM [Traceability_RS].[dyn].[Shipments] s
                WHERE s.ShipmentDate >= ? AND s.ShipmentDate <= ?
                  AND s.Status <> 'OPEN'
                ORDER BY s.ShipmentDate DESC, s.ShipmentId DESC
                """,
                (self.from_entry.get_date(), self.to_entry.get_date()),
            )
            for r in self.db.cursor.fetchall():
                sd = r.ShipmentDate.strftime("%d/%m/%Y") if r.ShipmentDate else ""
                self.tree.insert("", tk.END, values=(
                    r.ShipmentId, sd, r.Status, r.Pallets, int(r.Qty or 0), r.ClosedByUser or ""))
        except Exception as e:
            logger.error(f"Errore ricerca spedizioni: {e}", exc_info=True)
            messagebox.showerror(self.lang.get("error", "Errore"), str(e), parent=self)

    def _load(self):
        sel = self.tree.selection()
        if not sel:
            return
        self.selected_shipment_id = int(self.tree.set(sel[0], "ShipmentId"))
        self.destroy()


def open_shipment_confirmation_window(master, db, lang, user_name: str):
    """Punto di ingresso pubblico."""
    ShipmentConfirmationWindow(master, db, lang, user_name)
