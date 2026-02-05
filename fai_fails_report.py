"""
FAI Fails Report Window - Rapporto FAI fails
Displays failure statistics by operator and labelcode with filtering and Excel export
"""
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import os
import logging

logger = logging.getLogger(__name__)


class FaiFailsReportWindow(tk.Toplevel):
    """Finestra per visualizzare statistiche di fallimento FAI per operatore e labelcode"""
    
    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        
        self.db = db
        self.lang = lang
        self.user_name = user_name
        
        self.title(lang.get('rapporto_fai_fails', "Rapporto FAI fails"))
        self.geometry("1200x700")
        
        # Centra finestra
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.winfo_screenheight() // 2) - (700 // 2)
        self.geometry(f"1200x700+{x}+{y}")
        
        self._build_ui()
        self._load_data()
    
    def _build_ui(self):
        """Costruisce l'interfaccia"""
        # Frame filtri
        filters_frame = ttk.LabelFrame(self, text=self.lang.get('fai_filters_search', "Filtri Ricerca"), padding="10")
        filters_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Riga filtri
        row1 = ttk.Frame(filters_frame)
        row1.pack(fill=tk.X, pady=5)
        
        # Data Da
        ttk.Label(row1, text=self.lang.get('fai_date_from', "Data Da:")).pack(side=tk.LEFT, padx=5)
        from tkcalendar import DateEntry
        self.date_from_picker = DateEntry(
            row1,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        # Imposta data di default (30 giorni fa)
        default_from = datetime.now() - timedelta(days=30)
        self.date_from_picker.set_date(default_from)
        self.date_from_picker.pack(side=tk.LEFT, padx=5)
        
        # Data A
        ttk.Label(row1, text=self.lang.get('fai_date_to', "Data A:")).pack(side=tk.LEFT, padx=5)
        self.date_to_picker = DateEntry(
            row1,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            width=12
        )
        self.date_to_picker.set_date(datetime.now())
        self.date_to_picker.pack(side=tk.LEFT, padx=5)
        
        # Operatore
        ttk.Label(row1, text=self.lang.get('fai_fails_operator_filter', "Operatore:")).pack(side=tk.LEFT, padx=5)
        self.operator_var = tk.StringVar()
        self.operator_entry = ttk.Entry(row1, textvariable=self.operator_var, width=20)
        self.operator_entry.pack(side=tk.LEFT, padx=5)
        
        # Checkbox Solo Fails
        self.only_fails_var = tk.BooleanVar(value=True)  # Default: solo fails
        ttk.Checkbutton(
            row1, 
            text=self.lang.get('fai_only_fails', "Solo Fails"),
            variable=self.only_fails_var,
            command=self._load_data
        ).pack(side=tk.LEFT, padx=10)
        
        # Pulsante Cerca
        ttk.Button(row1, text="🔍 " + self.lang.get('search', "Cerca"), command=self._load_data).pack(side=tk.LEFT, padx=10)
        ttk.Button(row1, text="🔄 " + self.lang.get('reset', "Reset"), command=self._reset_filters).pack(side=tk.LEFT, padx=5)
        
        # Treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        columns = ('Operator', 'LabelCode', 'TotalRecords', 'FailedRecords', 'FailureRate')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configura colonne
        self.tree.heading('Operator', text=self.lang.get('fai_fails_operator_filter', 'Operatore'))
        self.tree.heading('LabelCode', text=self.lang.get('product_code', 'Codice Prodotto'))
        self.tree.heading('TotalRecords', text=self.lang.get('fai_fails_total_records', 'Totale Record'))
        self.tree.heading('FailedRecords', text=self.lang.get('fai_fails_failed_records', 'Record Falliti'))
        self.tree.heading('FailureRate', text=self.lang.get('fai_fails_failure_rate', 'Tasso di Fallimento (%)'))
        
        self.tree.column('Operator', width=200)
        self.tree.column('LabelCode', width=250)
        self.tree.column('TotalRecords', width=150, anchor='center')
        self.tree.column('FailedRecords', width=150, anchor='center')
        self.tree.column('FailureRate', width=180, anchor='center')
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Grid
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Pulsanti azioni
        buttons_frame = ttk.Frame(self)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(buttons_frame, text="📊 " + self.lang.get('export_excel', "Esporta Excel"), 
                  command=self._export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="❌ " + self.lang.get('close', "Chiudi"), 
                  command=self.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Status bar
        self.status_label = ttk.Label(self, text=self.lang.get('ready', "Pronto"), relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM)
    
    def _reset_filters(self):
        """Reset filtri"""
        default_from = datetime.now() - timedelta(days=30)
        self.date_from_picker.set_date(default_from)
        self.date_to_picker.set_date(datetime.now())
        self.operator_var.set('')
        self.only_fails_var.set(True)  # Reset a solo fails
        self._load_data()
    
    def _load_data(self):
        """Carica dati statistiche FAI fails"""
        try:
            # Pulisci treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Query per statistiche FAI fails
            query = """
            SELECT 
                l.Operator,
                p.productcode,
                COUNT(DISTINCT l.FaiLogId) as TotalRecords,
                SUM(CASE WHEN l.IsOk = 0 THEN 1 ELSE 0 END) as FailedRecords,
                CAST(SUM(CASE WHEN l.IsOk = 0 THEN 1 ELSE 0 END) * 100.0 / 
                     NULLIF(COUNT(DISTINCT l.FaiLogId), 0) AS DECIMAL(5,2)) as FailureRate
            FROM [Traceability_RS].[fai].[FaiLogs] l
            LEFT JOIN [Traceability_RS].[dbo].[orders] o ON l.OrderId = o.IDOrder
            LEFT JOIN [Traceability_RS].[dbo].[Products] p ON o.IDProduct = p.IDProduct
            WHERE l.DateIn >= ? AND l.DateIn <= ?
            """
            
            params = [
                self.date_from_picker.get_date().strftime('%Y-%m-%d') + ' 00:00:00',
                self.date_to_picker.get_date().strftime('%Y-%m-%d') + ' 23:59:59'
            ]
            
            # Filtro operatore (ricerca parziale)
            if self.operator_var.get().strip():
                query += " AND l.Operator LIKE ?"
                params.append(f"%{self.operator_var.get().strip()}%")
            
            query += """
            GROUP BY l.Operator, p.productcode
            HAVING COUNT(DISTINCT l.FaiLogId) > 0
            """
            
            # Filtro solo fails
            if self.only_fails_var.get():
                query += " AND SUM(CASE WHEN l.IsOk = 0 THEN 1 ELSE 0 END) > 0"
            
            query += """
            ORDER BY FailureRate DESC, l.Operator, p.productcode
            """
            
            rows = self.db.fetch_all(query, tuple(params))
            
            # Popola treeview
            for row in rows:
                operator = row.Operator or 'N/A'
                product_code = row.productcode or 'N/A'
                total_records = row.TotalRecords or 0
                failed_records = row.FailedRecords or 0
                failure_rate = row.FailureRate or 0.0
                
                # Determina tag colore in base al tasso di fallimento
                if failure_rate < 5.0:
                    tag = 'low'
                elif failure_rate < 15.0:
                    tag = 'medium'
                else:
                    tag = 'high'
                
                self.tree.insert('', 'end', values=(
                    operator, product_code, total_records, failed_records, f"{failure_rate:.2f}%"
                ), tags=(tag,))
            
            # Tags colori
            self.tree.tag_configure('low', foreground='green')
            self.tree.tag_configure('medium', foreground='orange')
            self.tree.tag_configure('high', foreground='red')
            
            self.status_label.config(text=f"Trovati {len(rows)} record")
            
        except Exception as e:
            logger.error(f"Errore caricamento statistiche FAI fails: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile caricare i dati:\n{e}")
    
    def _export_to_excel(self):
        """Esporta i risultati in Excel nella cartella C:\Temp"""
        try:
            # Verifica che ci siano dati
            if not self.tree.get_children():
                messagebox.showwarning("Attenzione", "Nessun dato da esportare")
                return
            
            # Importa openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                messagebox.showerror("Errore", "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl")
                return
            
            # Crea directory C:\Temp se non esiste
            temp_dir = r"C:\Temp"
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
                logger.info(f"Creata directory: {temp_dir}")
            
            # Crea workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "FAI Fails Report"
            
            # Headers
            headers = [
                self.lang.get('fai_fails_operator_filter', 'Operatore'),
                self.lang.get('product_code', 'Codice Prodotto'),
                self.lang.get('fai_fails_total_records', 'Totale Record'),
                self.lang.get('fai_fails_failed_records', 'Record Falliti'),
                self.lang.get('fai_fails_failure_rate', 'Tasso di Fallimento (%)')
            ]
            ws.append(headers)
            
            # Formatta header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aggiungi dati
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                # Rimuovi il simbolo % dalla percentuale per Excel
                failure_rate_str = str(values[4]).replace('%', '')
                row_data = [
                    values[0],  # Operator
                    values[1],  # LabelCode
                    values[2],  # TotalRecords
                    values[3],  # FailedRecords
                    failure_rate_str  # FailureRate
                ]
                ws.append(row_data)
                
                # Colora riga in base al tasso di fallimento
                last_row = ws.max_row
                try:
                    failure_rate = float(failure_rate_str)
                    if failure_rate < 5.0:
                        fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
                    elif failure_rate < 15.0:
                        fill_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Giallo
                    else:
                        fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rosso
                    
                    for cell in ws[last_row]:
                        cell.fill = fill_color
                except ValueError:
                    pass  # Se non riesce a convertire, salta la colorazione
            
            # Adatta larghezza colonne
            ws.column_dimensions['A'].width = 25  # Operator
            ws.column_dimensions['B'].width = 30  # LabelCode
            ws.column_dimensions['C'].width = 18  # TotalRecords
            ws.column_dimensions['D'].width = 18  # FailedRecords
            ws.column_dimensions['E'].width = 25  # FailureRate
            
            # Salva file in C:\Temp
            filename = f"FAI_Fails_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            logger.info(f"Excel esportato: {filepath}")
            
            # Chiedi se aprire il file
            if messagebox.askyesno("File Creato", f"File Excel salvato in:\n{filepath}\n\nVuoi aprire il file?"):
                os.startfile(filepath)
            
            self.status_label.config(text=f"Excel esportato: {filename}")
        
        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile esportare in Excel:\n{e}")
