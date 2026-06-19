"""
Overtime Analysis GUI
Form per analisi straordinari con confronto presenza/approvazione
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date, timedelta
from tkcalendar import DateEntry
import logging
import os
import tempfile

logger = logging.getLogger(__name__)


def open_overtime_analysis_window(parent, db_handler, lang_manager, user_name):
    """
    Apre la finestra per analisi straordinari.
    
    Args:
        parent: Finestra parent
        db_handler: DatabaseHandler instance
        lang_manager: LanguageManager instance
        user_name: Nome utente loggato
    """
    OvertimeAnalysisWindow(parent, db_handler, lang_manager, user_name)


class OvertimeAnalysisWindow(tk.Toplevel):
    """
    Finestra per analisi straordinari con export PDF/Excel.
    """
    
    def __init__(self, parent, db_handler, lang_manager, user_name):
        super().__init__(parent)
        
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_name
        self.analysis_data = []
        self.economics = None          # risultato compute_economics (dict) o None
        self.sort_column = None
        self.sort_reverse = False
        
        # Setup finestra
        self.title(self.lang.get('overtime_analysis_title', 'Analisi Straordinari'))
        self.geometry("1400x700")
        self.transient(parent)
        self.grab_set()
        
        self._create_widgets()
        
        # Imposta date default (mese corrente)
        today = date.today()
        first_day = date(today.year, today.month, 1)
        self.start_date.set_date(first_day)
        self.end_date.set_date(today)
    
    def _create_widgets(self):
        """Crea i widget dell'interfaccia."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === FILTRI ===
        filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get('filters', 'Filtri'), padding="10")
        filter_frame.pack(fill=tk.X, pady=5)
        
        # Date
        ttk.Label(filter_frame, text=self.lang.get('from', 'Da:')).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_date = DateEntry(
            filter_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy'
        )
        self.start_date.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(filter_frame, text=self.lang.get('to', 'A:')).grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.end_date = DateEntry(
            filter_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy'
        )
        self.end_date.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        
        # Filtro tipo
        ttk.Label(filter_frame, text=self.lang.get('filter_type', 'Tipo Filtro:')).grid(row=0, column=4, padx=5, pady=5, sticky=tk.W)
        self.filter_type_var = tk.StringVar(value='ALL')
        filter_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.filter_type_var,
            values=['ALL', 'OVER APPROVED', 'Time approved = time presence'],
            state='readonly',
            width=30
        )
        filter_combo.grid(row=0, column=5, padx=5, pady=5, sticky=tk.W)
        
        # Pulsante genera
        ttk.Button(
            filter_frame,
            text=self.lang.get('generate_analysis', 'Genera Analisi'),
            command=self._generate_analysis
        ).grid(row=0, column=6, padx=5, pady=5)
        
        # === RISULTATI (Notebook con 2 schede) ===
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=5)

        # --- Scheda 1: dettaglio straordinari ---
        results_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(results_frame,
                          text=self.lang.get('overtime_detail_tab', 'Dettaglio Straordinari'))

        # Treeview
        columns = ('nr', 'name', 'date', 'min_done', 'min_approved', 'notes')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=20)

        # Intestazioni cliccabili per ordinamento
        self.tree.heading('nr', text='Nr', command=lambda: self._sort_by_column('nr'))
        self.tree.heading('name', text=self.lang.get('employee', 'Dipendente'), command=lambda: self._sort_by_column('name'))
        self.tree.heading('date', text=self.lang.get('date', 'Data'), command=lambda: self._sort_by_column('date'))
        self.tree.heading('min_done', text=self.lang.get('min_done', 'Min Presenza'), command=lambda: self._sort_by_column('min_done'))
        self.tree.heading('min_approved', text=self.lang.get('min_approved', 'Min Approvati'), command=lambda: self._sort_by_column('min_approved'))
        self.tree.heading('notes', text=self.lang.get('notes', 'Note'), command=lambda: self._sort_by_column('notes'))

        self.tree.column('nr', width=50, anchor=tk.CENTER)
        self.tree.column('name', width=250, anchor=tk.W)
        self.tree.column('date', width=100, anchor=tk.CENTER)
        self.tree.column('min_done', width=120, anchor=tk.CENTER)
        self.tree.column('min_approved', width=120, anchor=tk.CENTER)
        self.tree.column('notes', width=250, anchor=tk.W)

        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Scheda 2: convenienza economica ---
        self._build_economics_tab()
        
        # === PULSANTI EXPORT ===
        export_frame = ttk.Frame(main_frame)
        export_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('export_excel', 'Esporta Excel'),
            command=self._export_to_excel
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('export_pdf', 'Esporta PDF'),
            command=self._export_to_pdf
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('close', 'Chiudi'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)

    # ── Scheda Convenienza Economica ────────────────────────────────────────────
    def _build_economics_tab(self):
        """Costruisce la scheda 'Convenienza Economica' del notebook."""
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text=self.lang.get('economics_tab', 'Convenienza Economica'))

        # File prezzi D365 usato
        self._eco_file_var = tk.StringVar(value=self.lang.get(
            'economics_run_hint', "Premi 'Genera Analisi' per calcolare la convenienza economica."))
        ttk.Label(tab, textvariable=self._eco_file_var,
                  foreground='#1F3864', font=('Helvetica', 8)).pack(anchor=tk.W, pady=(0, 6))

        self._eco_vars = {}
        self._eco_labels = {}

        def add_section(parent, title, fields):
            lf = ttk.LabelFrame(parent, text=title, padding="8")
            for i, (key, label) in enumerate(fields):
                r, col = divmod(i, 2)
                cell = ttk.Frame(lf)
                cell.grid(row=r, column=col, sticky=tk.W, padx=10, pady=2)
                ttk.Label(cell, text=label + ':', width=32, anchor=tk.W).pack(side=tk.LEFT)
                var = tk.StringVar(value='—')
                self._eco_vars[key] = var
                lbl = ttk.Label(cell, textvariable=var,
                                font=('Helvetica', 10, 'bold'), foreground='#0056b3')
                lbl.pack(side=tk.LEFT)
                self._eco_labels[key] = lbl
            return lf

        # Sezione Produzione del periodo
        prod = add_section(
            tab, self.lang.get('economics_prod_section', 'Produzione del periodo (tutte le ore)'), [
                ('finalized',      self.lang.get('eco_finalized', 'Pezzi finalizzati')),
                ('finalized_value', self.lang.get('eco_finalized_value', 'Valore finalizzato')),
                ('wip',            self.lang.get('eco_wip', 'WIP (schede / pezzi-eq.)')),
                ('wip_value',      self.lang.get('eco_wip_value', 'Valore WIP')),
                ('total_value',    self.lang.get('eco_total_value', 'Valore prodotto totale')),
                ('labor_hours',    self.lang.get('eco_labor_hours', 'Ore lavorate (produzione)')),
                ('productivity',   self.lang.get('eco_productivity', 'Produttività media (valore/ora)')),
            ])
        prod.pack(fill=tk.X, pady=4)

        # Sezione Straordinario
        ot = add_section(
            tab, self.lang.get('economics_ot_section', 'Straordinario'), [
                ('people',          self.lang.get('eco_people', 'Persone in straordinario')),
                ('ot_hours',        self.lang.get('eco_ot_hours', 'Ore straordinario (svolte / appr.)')),
                ('ot_incidence',    self.lang.get('eco_ot_incidence', 'Incidenza ore straord. (%)')),
                ('ot_cost',         self.lang.get('eco_ot_cost', 'Costo straordinario')),
                ('ot_cost_per_hour', self.lang.get('eco_ot_cost_per_hour', 'Costo medio straord. (/h)')),
                ('ot_value',        self.lang.get('eco_ot_value', 'Valore attribuibile allo straord.')),
                ('ot_margin',       self.lang.get('eco_ot_margin', 'Margine straordinario (valore - costo)')),
            ])
        ot.pack(fill=tk.X, pady=4)

        # ROI evidenziato
        roi_frame = ttk.Frame(tab)
        roi_frame.pack(fill=tk.X, pady=(2, 4))
        ttk.Label(roi_frame,
                  text=self.lang.get('eco_ot_roi', 'ROI straordinario (valore/costo)') + ':',
                  font=('Helvetica', 12, 'bold')).pack(side=tk.LEFT, padx=(2, 8))
        self._eco_roi_var = tk.StringVar(value='—')
        self._eco_roi_label = tk.Label(roi_frame, textvariable=self._eco_roi_var,
                                       font=('Helvetica', 16, 'bold'), fg='#0056b3')
        self._eco_roi_label.pack(side=tk.LEFT)
        self._eco_roi_note = tk.Label(roi_frame, text='', font=('Helvetica', 9))
        self._eco_roi_note.pack(side=tk.LEFT, padx=12)

        # Avviso prezzi mancanti
        self._eco_missing_var = tk.StringVar(value='')
        ttk.Label(tab, textvariable=self._eco_missing_var,
                  foreground='#B71C1C', font=('Helvetica', 8)).pack(anchor=tk.W, pady=(4, 0))

        # Dettaglio per giorno
        day_frame = ttk.LabelFrame(
            tab, text=self.lang.get('economics_per_day', 'Dettaglio per giorno'), padding="6")
        day_frame.pack(fill=tk.BOTH, expand=True, pady=6)

        cols = ('day', 'people', 'hours', 'cost', 'finalized_value')
        self.eco_tree = ttk.Treeview(day_frame, columns=cols, show='headings', height=8)
        self.eco_tree.heading('day', text=self.lang.get('date', 'Data'))
        self.eco_tree.heading('people', text=self.lang.get('eco_people_short', 'Persone'))
        self.eco_tree.heading('hours', text=self.lang.get('eco_hours_short', 'Ore straord.'))
        self.eco_tree.heading('cost', text=self.lang.get('eco_cost_short', 'Costo straord.'))
        self.eco_tree.heading('finalized_value',
                              text=self.lang.get('eco_finalized_value_short', 'Valore finalizzato'))
        self.eco_tree.column('day', width=110, anchor=tk.CENTER)
        self.eco_tree.column('people', width=90, anchor=tk.CENTER)
        self.eco_tree.column('hours', width=110, anchor=tk.CENTER)
        self.eco_tree.column('cost', width=130, anchor=tk.E)
        self.eco_tree.column('finalized_value', width=150, anchor=tk.E)

        eco_sb = ttk.Scrollbar(day_frame, orient=tk.VERTICAL, command=self.eco_tree.yview)
        self.eco_tree.configure(yscroll=eco_sb.set)
        self.eco_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        eco_sb.pack(side=tk.RIGHT, fill=tk.Y)

    def _fmt_money(self, v):
        cur = ''
        try:
            cur = self.economics['summary']['rates'].get('currency', '') if self.economics else ''
        except Exception:
            cur = ''
        if v is None:
            return '—'
        return f"{v:,.2f} {cur}".strip()

    def _compute_and_show_economics(self, start_date, end_date):
        """Calcola i KPI di convenienza e popola la scheda dedicata."""
        try:
            from overtime import overtime_economics as eco
            self.economics = eco.compute_economics(self.db.conn, start_date, end_date)
        except Exception as e:
            logger.error(f"Errore calcolo convenienza economica: {e}", exc_info=True)
            self.economics = None
            self._eco_file_var.set(
                self.lang.get('economics_error', 'Errore calcolo convenienza economica: ') + str(e))
            return

        s = self.economics['summary']

        # File D365
        if s['d365_file']:
            self._eco_file_var.set(
                self.lang.get('economics_d365_file', 'Prezzi da: ') + str(s['d365_file']))
        else:
            self._eco_file_var.set(self.lang.get(
                'economics_no_d365', '⚠️ File prezzi D365 non trovato in T:\\D365 data — valori a 0.'))

        v = self._eco_vars
        # Produzione
        v['finalized'].set(f"{s['finalized_pieces']:,}")
        v['finalized_value'].set(self._fmt_money(s['finalized_value']))
        v['wip'].set(f"{s['wip_boards']:,} / {s['wip_pieces_equiv']:.1f}")
        v['wip_value'].set(self._fmt_money(s['wip_value']))
        v['total_value'].set(self._fmt_money(s['total_value']))
        v['labor_hours'].set(f"{s['total_labor_hours']:,.1f} h")
        v['productivity'].set(self._fmt_money(s['productivity']) + "/h" if s['productivity'] else '—')
        # Straordinario
        v['people'].set(str(s['people']))
        v['ot_hours'].set(f"{s['ot_hours_done']:.1f} / {s['ot_hours_approved']:.1f} h")
        v['ot_incidence'].set(f"{s['ot_incidence_pct']:.2f}%" if s['ot_incidence_pct'] is not None else '—')
        v['ot_cost'].set(self._fmt_money(s['ot_cost']))
        v['ot_cost_per_hour'].set(self._fmt_money(s['ot_cost_per_hour']) + "/h" if s['ot_cost_per_hour'] else '—')
        v['ot_value'].set(self._fmt_money(s['ot_value']))
        v['ot_margin'].set(self._fmt_money(s['ot_margin']))
        self._eco_labels['ot_margin'].config(
            foreground='#1B5E20' if s['ot_margin'] >= 0 else '#B71C1C')

        # ROI evidenziato
        roi = s['ot_roi']
        if roi is not None:
            self._eco_roi_var.set(f"{roi:.2f}")
            if roi >= 1:
                self._eco_roi_label.config(fg='#1B5E20')
                self._eco_roi_note.config(
                    text=self.lang.get('eco_roi_good', '✓ Straordinario conveniente'), fg='#1B5E20')
            else:
                self._eco_roi_label.config(fg='#B71C1C')
                self._eco_roi_note.config(
                    text=self.lang.get('eco_roi_bad', '✗ Straordinario non conveniente'), fg='#B71C1C')
        else:
            self._eco_roi_var.set('—')
            self._eco_roi_note.config(text='')

        # Prezzi mancanti
        miss = s['missing_price']
        if miss:
            self._eco_missing_var.set(
                self.lang.get('economics_missing_price',
                              '⚠️ {0} ordini/prodotti senza prezzo (valorizzati a 0): ').format(len(miss))
                + ', '.join(miss[:8]) + (' …' if len(miss) > 8 else ''))
        else:
            self._eco_missing_var.set('')

        # Dettaglio per giorno
        for item in self.eco_tree.get_children():
            self.eco_tree.delete(item)
        for d in self.economics['per_day']:
            self.eco_tree.insert('', tk.END, values=(
                d['day'].strftime('%d/%m/%Y'),
                d['people'],
                f"{d['hours_done']:.1f}",
                f"{d['cost']:,.2f}",
                f"{d['finalized_value']:,.2f}",
            ))

    def _generate_analysis(self):
        """Genera l'analisi in base ai filtri selezionati."""
        # Pulisci risultati precedenti
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.analysis_data = []
        
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()
        filter_type = self.filter_type_var.get()
        
        # Converti filtro per query
        filter_param = None if filter_type == 'ALL' else filter_type
        
        # Query di analisi
        query = """
        DECLARE @dateStart DATE = ?;
        DECLARE @DateStop DATE = ?;
        DECLARE @Filter AS NVARCHAR(30) = ?;

        WITH
        CTE_DailyState_Employee AS (
            SELECT 
                ds.IDDailyState,
                ds.DailyStateDate,
                e.IDEmployee,
                UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS Name, 
                e.UniqueID
            FROM Timeclocking.dbo.DailyState ds
            INNER JOIN Timeclocking.dbo.Employee e
                ON e.IDEmployee = ds.IDEmployee AND ds.DailyStateDate BETWEEN @dateStart AND @DateStop
        ),
        CTE_Done AS (
            SELECT
                fd.IDDailyState,
                fd.NoMin AS MinSuplimentarDone,
                r.RequestName
            FROM Timeclocking.dbo.EmployeeRequestFractionalDay fd
            INNER JOIN Timeclocking.dbo.RequestType r
                ON r.IDRequestType = fd.IDRequestType
            WHERE r.IDRequestType = 8
        ),
        CTE_HireHistory AS (
            SELECT
                h.EmployeeHireHistoryId AS EmployeeHireId,
                ee.EmployeeNID COLLATE DATABASE_DEFAULT AS UniqueID
            FROM employee.dbo.employees ee
            INNER JOIN employee.dbo.employeehirehistory h
                ON ee.EmployeeId = h.EmployeeId
                AND h.employeerid = 2
                AND h.EndWorkDate IS NULL
            LEFT JOIN Employee.dbo.EmployeeCdcStories cs
                ON cs.EmployeeHireHistoryId = h.EmployeeHireHistoryId
                AND cs.DateOut IS NULL
            LEFT JOIN Employee.dbo.Functions f
                ON cs.FunctionId = f.FunctionId
            WHERE ISNULL(f.FunctionCode, 0) <= 60
        ),
        CTE_ExtraTimeApprovalStory AS (
            SELECT
                es.IdEmployee AS EmployeeHireHistoryId,
                CAST(es.DateStart AS DATE) AS DateStart,
                ISNULL(DATEDIFF(MINUTE, es.DateStart, es.DateEnd), 0) AS MinExtraTimeApproved
            FROM [ResetServices].[dbo].ExtraTimeApprovalStory es
            WHERE CAST(es.DateStart AS DATE) BETWEEN @dateStart AND @dateStop
        ),
        CTE_Combined AS (
            SELECT 
                ROW_NUMBER() OVER (ORDER BY dse.DailyStateDate, dse.Name) AS Nr,
                dse.Name,
                dse.DailyStateDate AS OvertimeDate,
                req.MinSuplimentarDone,
                req.RequestName,
                ISNULL(eta.MinExtraTimeApproved, 0) AS MinExtraTimeApproved,
                CASE
                    WHEN ISNULL(req.MinSuplimentarDone, 0) > ISNULL(eta.MinExtraTimeApproved, 0) THEN
                        'OVER APPROVED'
                    ELSE
                        'Time approved = time presence'
                END AS Notes
            FROM CTE_DailyState_Employee dse
            INNER JOIN CTE_Done req ON dse.IDDailyState = req.IDDailyState
            INNER JOIN CTE_HireHistory hh ON dse.UniqueID COLLATE DATABASE_DEFAULT = hh.UniqueID
            LEFT JOIN CTE_ExtraTimeApprovalStory eta ON hh.EmployeeHireId = eta.EmployeeHireHistoryId
                AND eta.DateStart = dse.DailyStateDate
        )
        SELECT DISTINCT *
        FROM CTE_Combined
        WHERE Notes LIKE ISNULL(@Filter, Notes)
        ORDER BY OvertimeDate, Name;
        """
        
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, (start_date, end_date, filter_param))
            results = cursor.fetchall()
            cursor.close()
            
            # Popola tabella e salva dati
            for row in results:
                nr = row[0]
                name = row[1]
                overtime_date = row[2].strftime('%d/%m/%Y') if row[2] else 'N/D'
                min_done = row[3] if row[3] else 0
                min_approved = row[5] if row[5] else 0
                notes = row[6]
                
                # Colora righe OVER APPROVED in rosso
                tag = 'over_approved' if notes == 'OVER APPROVED' else ''
                
                self.tree.insert('', tk.END, values=(
                    nr, name, overtime_date, min_done, min_approved, notes
                ), tags=(tag,))
                
                self.analysis_data.append({
                    'nr': nr,
                    'name': name,
                    'date': overtime_date,
                    'min_done': min_done,
                    'min_approved': min_approved,
                    'notes': notes
                })
            
            # Configura tag per evidenziare
            self.tree.tag_configure('over_approved', background='#FFE0E0')

            # Calcola e mostra la convenienza economica (scheda dedicata)
            self._compute_and_show_economics(start_date, end_date)

            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                f"Analisi completata: {len(results)} record trovati",
                parent=self
            )
            
        except Exception as e:
            logger.error(f"Errore generazione analisi: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore generazione analisi:\n{str(e)}",
                parent=self
            )
    
    def _sort_by_column(self, col):
        """
        Ordina i dati della tabella in base alla colonna selezionata.
        
        Args:
            col: Nome della colonna da ordinare
        """
        if not self.analysis_data:
            return
        
        # Determina direzione ordinamento
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        
        # Mappa colonne ai campi dati
        col_map = {
            'nr': 'nr',
            'name': 'name',
            'date': 'date',
            'min_done': 'min_done',
            'min_approved': 'min_approved',
            'notes': 'notes'
        }
        
        sort_key = col_map.get(col, 'nr')
        
        # Funzione di ordinamento personalizzata per gestire date
        def sort_func(item):
            value = item[sort_key]
            # Converti date in formato ordinabile
            if sort_key == 'date' and isinstance(value, str):
                try:
                    # Formato: DD/MM/YYYY
                    parts = value.split('/')
                    if len(parts) == 3:
                        return f"{parts[2]}{parts[1]}{parts[0]}"  # YYYYMMDD
                except:
                    pass
            # Gestisci valori numerici
            if sort_key in ['nr', 'min_done', 'min_approved']:
                try:
                    return int(value) if value else 0
                except:
                    return 0
            # Default: ordinamento stringa
            return str(value).lower() if value else ''
        
        # Ordina i dati
        self.analysis_data.sort(key=sort_func, reverse=self.sort_reverse)
        
        # Ricarica la tabella
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for item_data in self.analysis_data:
            tag = 'over_approved' if item_data['notes'] == 'OVER APPROVED' else ''
            self.tree.insert('', tk.END, values=(
                item_data['nr'],
                item_data['name'],
                item_data['date'],
                item_data['min_done'],
                item_data['min_approved'],
                item_data['notes']
            ), tags=(tag,))
        
        # Configura tag
        self.tree.tag_configure('over_approved', background='#FFE0E0')
        
        # Aggiorna intestazioni con indicatore ordinamento
        self._update_column_headers()
    
    def _update_column_headers(self):
        """Aggiorna le intestazioni delle colonne con indicatori di ordinamento."""
        headers = {
            'nr': 'Nr',
            'name': self.lang.get('employee', 'Dipendente'),
            'date': self.lang.get('date', 'Data'),
            'min_done': self.lang.get('min_done', 'Min Presenza'),
            'min_approved': self.lang.get('min_approved', 'Min Approvati'),
            'notes': self.lang.get('notes', 'Note')
        }
        
        for col, header_text in headers.items():
            if col == self.sort_column:
                # Aggiungi indicatore direzione
                indicator = ' ▼' if self.sort_reverse else ' ▲'
                self.tree.heading(col, text=header_text + indicator)
            else:
                self.tree.heading(col, text=header_text)
    
    def _export_to_excel(self):
        """Esporta i risultati in Excel."""
        if not self.analysis_data:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare'),
                parent=self
            )
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crea directory C:\Temp se non esiste
            output_dir = r"C:\Temp"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Nome file
            filter_type = self.filter_type_var.get().replace(' ', '_')
            filename = f"ReportAnalysis_Overtime_{filter_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(output_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Overtime Analysis"
            
            # Header
            headers = ['Nr', 'Employee', 'Date', 'Min Presence', 'Min Approved', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Dati
            for row_idx, item in enumerate(self.analysis_data, 2):
                ws.cell(row=row_idx, column=1, value=item['nr']).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=2, value=item['name'])
                ws.cell(row=row_idx, column=3, value=item['date']).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=4, value=item['min_done']).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=5, value=item['min_approved']).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=6, value=item['notes'])
                
                # Evidenzia OVER APPROVED in rosso
                if item['notes'] == 'OVER APPROVED':
                    for col in range(1, 7):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                        )
            
            # Adatta larghezza colonne
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30

            # === Foglio Convenienza Economica ===
            if self.economics:
                self._write_economics_sheet(wb, Font, PatternFill, Alignment)

            wb.save(file_path)
            
            # Chiedi se aprire
            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File Excel salvato:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            
            if open_file:
                os.startfile(file_path)
            
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl",
                parent=self
            )
        except Exception as e:
            logger.error(f"Errore export Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export Excel:\n{str(e)}",
                parent=self
            )

    def _write_economics_sheet(self, wb, Font, PatternFill, Alignment):
        """Aggiunge al workbook il foglio 'Convenienza Economica'."""
        s = self.economics['summary']
        cur = s['rates'].get('currency', '')
        ws = wb.create_sheet(title="Convenienza Economica")

        title_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF")

        ws.cell(row=1, column=1, value="ANALISI DI CONVENIENZA ECONOMICA").font = Font(bold=True, size=13, color="1F3864")
        ws.cell(row=2, column=1, value=(
            f"Periodo: {self.start_date.get_date().strftime('%d/%m/%Y')} - "
            f"{self.end_date.get_date().strftime('%d/%m/%Y')}"))
        ws.cell(row=3, column=1, value=f"Prezzi da: {s['d365_file'] or 'N/D'}")
        ws.cell(row=4, column=1, value=(
            f"Tariffe: feriale {s['rates'].get('weekday_ot',0):.2f} {cur}/h  ·  "
            f"weekend {s['rates'].get('weekend',0):.2f} {cur}/h"))

        roi = s['ot_roi']
        kpis = [
            ("— PRODUZIONE DEL PERIODO (tutte le ore) —", None),
            ("Pezzi finalizzati", s['finalized_pieces']),
            (f"Valore finalizzato ({cur})", round(s['finalized_value'], 2)),
            ("Schede WIP", s['wip_boards']),
            ("WIP pezzi-equivalenti", round(s['wip_pieces_equiv'], 1)),
            (f"Valore WIP ({cur})", round(s['wip_value'], 2)),
            (f"Valore prodotto totale ({cur})", round(s['total_value'], 2)),
            ("Ore lavorate produzione", round(s['total_labor_hours'], 1)),
            (f"Produttività media (valore/ora) ({cur})",
             round(s['productivity'], 2) if s['productivity'] else "—"),
            ("— STRAORDINARIO —", None),
            ("Persone in straordinario", s['people']),
            ("Ore straordinario svolte", round(s['ot_hours_done'], 1)),
            ("Ore straordinario approvate", round(s['ot_hours_approved'], 1)),
            ("Incidenza ore straord. (%)",
             round(s['ot_incidence_pct'], 2) if s['ot_incidence_pct'] is not None else "—"),
            (f"Costo straordinario ({cur})", round(s['ot_cost'], 2)),
            (f"Costo medio straordinario ({cur}/h)",
             round(s['ot_cost_per_hour'], 2) if s['ot_cost_per_hour'] else "—"),
            (f"Valore attribuibile allo straordinario ({cur})", round(s['ot_value'], 2)),
            (f"Margine straordinario valore-costo ({cur})", round(s['ot_margin'], 2)),
            ("ROI straordinario (valore/costo)", round(roi, 2) if roi is not None else "—"),
        ]
        r0 = 6
        ws.cell(row=r0, column=1, value="Indicatore").font = title_font
        ws.cell(row=r0, column=1).fill = title_fill
        ws.cell(row=r0, column=2, value="Valore").font = title_font
        ws.cell(row=r0, column=2).fill = title_fill
        section_font = Font(bold=True, color="1F3864")
        for i, (label, value) in enumerate(kpis, start=r0 + 1):
            cell_l = ws.cell(row=i, column=1, value=label)
            if value is None:                       # riga di sezione
                cell_l.font = section_font
                continue
            ws.cell(row=i, column=2, value=value).alignment = Alignment(horizontal='right')

        # Dettaglio per giorno
        rstart = r0 + len(kpis) + 3
        ws.cell(row=rstart - 1, column=1, value="Dettaglio per giorno").font = Font(bold=True, size=11, color="1F3864")
        headers = ["Data", "Persone", "Ore straord.", f"Costo ({cur})", f"Valore finalizzato ({cur})"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=rstart, column=col, value=h)
            c.font = title_font
            c.fill = title_fill
            c.alignment = Alignment(horizontal='center')
        for i, d in enumerate(self.economics['per_day'], start=rstart + 1):
            ws.cell(row=i, column=1, value=d['day'].strftime('%d/%m/%Y')).alignment = Alignment(horizontal='center')
            ws.cell(row=i, column=2, value=d['people']).alignment = Alignment(horizontal='center')
            ws.cell(row=i, column=3, value=round(d['hours_done'], 1)).alignment = Alignment(horizontal='center')
            ws.cell(row=i, column=4, value=round(d['cost'], 2)).alignment = Alignment(horizontal='right')
            ws.cell(row=i, column=5, value=round(d['finalized_value'], 2)).alignment = Alignment(horizontal='right')

        # Prezzi mancanti
        if s['missing_price']:
            rmiss = rstart + len(self.economics['per_day']) + 2
            ws.cell(row=rmiss, column=1,
                    value=f"Ordini/prodotti senza prezzo (valorizzati a 0): {len(s['missing_price'])}").font = Font(bold=True, color="B71C1C")
            for j, m in enumerate(s['missing_price'], start=rmiss + 1):
                ws.cell(row=j, column=1, value=m)

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 24

    def _export_to_pdf(self):
        """Esporta i risultati in PDF."""
        if not self.analysis_data:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare'),
                parent=self
            )
            return
        
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import cm
            from reportlab.platypus import Table, TableStyle, Image as ReportLabImage
            from reportlab.lib import colors
            
            # Crea directory C:\Temp se non esiste
            output_dir = r"C:\Temp"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Nome file
            filter_type = self.filter_type_var.get().replace(' ', '_')
            filename = f"ReportAnalysis_Overtime_{filter_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = os.path.join(output_dir, filename)
            
            c = canvas.Canvas(file_path, pagesize=landscape(A4))
            width, height = landscape(A4)
            
            # Logo
            logo_path = "Logo.png"
            if os.path.exists(logo_path):
                try:
                    logo = ReportLabImage(logo_path, width=1.5 * cm, height=1.5 * cm)
                    logo.drawOn(c, width - 2.5 * cm, height - 2.5 * cm)
                except Exception as e:
                    logger.warning(f"Cannot load logo: {e}")
            
            # Titolo
            c.setFont("Helvetica-Bold", 18)
            c.drawString(2 * cm, height - 2 * cm, "OVERTIME ANALYSIS REPORT")
            
            c.setFont("Helvetica", 10)
            c.drawString(2 * cm, height - 2.5 * cm, 
                f"Period: {self.start_date.get_date().strftime('%d/%m/%Y')} - {self.end_date.get_date().strftime('%d/%m/%Y')}")
            c.drawString(2 * cm, height - 3 * cm, 
                f"Filter: {self.filter_type_var.get()}")
            c.drawString(2 * cm, height - 3.5 * cm, 
                f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            # Tabella
            table_data = [['Nr', 'Employee', 'Date', 'Min Presence', 'Min Approved', 'Notes']]
            over_approved_rows = []
            
            for idx, item in enumerate(self.analysis_data, 1):
                table_data.append([
                    item['nr'],
                    item['name'],
                    item['date'],
                    item['min_done'],
                    item['min_approved'],
                    item['notes']
                ])
                if item['notes'] == 'OVER APPROVED':
                    over_approved_rows.append(idx)
            
            table = Table(table_data, colWidths=[1.5*cm, 5*cm, 2.5*cm, 3*cm, 3*cm, 5*cm])
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
            ]
            
            # Evidenzia righe OVER APPROVED
            for row_idx in over_approved_rows:
                table_style.append(
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#FFE0E0"))
                )
            
            table.setStyle(TableStyle(table_style))
            
            table.wrapOn(c, width, height)
            table_height = table._height
            table.drawOn(c, 2 * cm, height - 5 * cm - table_height)
            
            # Footer
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2, 1.5 * cm,
                "Document automatically generated by TraceabilityRS system")

            # === Pagina Convenienza Economica ===
            if self.economics:
                self._draw_economics_pdf_page(c, width, height, cm, Table, TableStyle, colors)

            c.save()
            
            # Chiedi se aprire
            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File PDF salvato:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            
            if open_file:
                os.startfile(file_path)
            
        except Exception as e:
            logger.error(f"Errore export PDF: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export PDF:\n{str(e)}",
                parent=self
            )

    def _draw_economics_pdf_page(self, c, width, height, cm, Table, TableStyle, colors):
        """Disegna una nuova pagina PDF con i KPI di convenienza economica."""
        s = self.economics['summary']
        cur = s['rates'].get('currency', '')
        c.showPage()

        c.setFont("Helvetica-Bold", 16)
        c.drawString(2 * cm, height - 2 * cm, "ANALISI DI CONVENIENZA ECONOMICA")
        c.setFont("Helvetica", 10)
        c.drawString(2 * cm, height - 2.6 * cm,
                     f"Periodo: {self.start_date.get_date().strftime('%d/%m/%Y')} - "
                     f"{self.end_date.get_date().strftime('%d/%m/%Y')}")
        c.drawString(2 * cm, height - 3.1 * cm, f"Prezzi da: {s['d365_file'] or 'N/D'}")
        c.drawString(2 * cm, height - 3.6 * cm,
                     f"Tariffe: feriale {s['rates'].get('weekday_ot',0):.2f} {cur}/h  -  "
                     f"weekend {s['rates'].get('weekend',0):.2f} {cur}/h")

        roi = s['ot_roi']
        def m(v):
            return f"{v:,.2f} {cur}" if v is not None else "—"
        rows = [
            ["Indicatore", "Valore"],
            ["PRODUZIONE DEL PERIODO (tutte le ore)", ""],
            ["Pezzi finalizzati", f"{s['finalized_pieces']:,}"],
            ["Valore finalizzato", m(s['finalized_value'])],
            ["WIP (schede / pezzi-eq.)", f"{s['wip_boards']:,} / {s['wip_pieces_equiv']:.1f}"],
            ["Valore WIP", m(s['wip_value'])],
            ["Valore prodotto totale", m(s['total_value'])],
            ["Ore lavorate produzione", f"{s['total_labor_hours']:,.1f} h"],
            ["Produttività media (valore/ora)", m(s['productivity'])],
            ["STRAORDINARIO", ""],
            ["Persone in straordinario", str(s['people'])],
            ["Ore straordinario (svolte / appr.)", f"{s['ot_hours_done']:.1f} / {s['ot_hours_approved']:.1f}"],
            ["Incidenza ore straord. (%)", f"{s['ot_incidence_pct']:.2f}%" if s['ot_incidence_pct'] is not None else "—"],
            ["Costo straordinario", m(s['ot_cost'])],
            ["Costo medio straordinario (/h)", m(s['ot_cost_per_hour'])],
            ["Valore attribuibile allo straordinario", m(s['ot_value'])],
            ["Margine straordinario (valore - costo)", m(s['ot_margin'])],
            ["ROI straordinario (valore/costo)", f"{roi:.2f}" if roi is not None else "—"],
        ]
        section_rows = [1, 9]   # indici righe di sezione
        table = Table(rows, colWidths=[11 * cm, 7 * cm])
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E5090")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
        ]
        for sr in section_rows:
            style.append(('BACKGROUND', (0, sr), (-1, sr), colors.HexColor("#D6E4F0")))
            style.append(('FONTNAME', (0, sr), (-1, sr), 'Helvetica-Bold'))
            style.append(('TEXTCOLOR', (0, sr), (-1, sr), colors.HexColor("#1F3864")))
        # ROI: ultima riga, colore in base al valore
        roi_row = len(rows) - 1
        roi_color = colors.HexColor("#1B5E20") if (roi is not None and roi >= 1) else colors.HexColor("#B71C1C")
        style.append(('FONTNAME', (0, roi_row), (-1, roi_row), 'Helvetica-Bold'))
        style.append(('TEXTCOLOR', (0, roi_row), (-1, roi_row), roi_color))
        table.setStyle(TableStyle(style))
        table.wrapOn(c, width, height)
        table.drawOn(c, 2 * cm, height - 4.2 * cm - table._height)

        if s['missing_price']:
            c.setFont("Helvetica", 8)
            c.setFillColorRGB(0.72, 0.11, 0.11)
            c.drawString(2 * cm, 2.2 * cm,
                         f"Ordini/prodotti senza prezzo (valorizzati a 0): {len(s['missing_price'])}")
            c.setFillColorRGB(0, 0, 0)

        c.setFont("Helvetica", 8)
        c.drawCentredString(width / 2, 1.5 * cm,
                            "Document automatically generated by TraceabilityRS system")
