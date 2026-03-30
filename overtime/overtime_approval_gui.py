"""
Overtime Approval GUI
Form per l'autorizzazione delle richieste di straordinario
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import logging
from tkcalendar import DateEntry

logger = logging.getLogger(__name__)


def open_overtime_approval_window(parent, db_handler, lang_manager, user_name, user_id=0):
    """
    Apre la finestra per autorizzare richieste di straordinario.
    
    Args:
        parent: Finestra parent
        db_handler: DatabaseHandler instance
        lang_manager: LanguageManager instance
        user_name: Nome utente loggato
        user_id: ID utente (idanga da tbuserkey)
    """
    OvertimeApprovalWindow(parent, db_handler, lang_manager, user_name, user_id)


class OvertimeApprovalWindow(tk.Toplevel):
    """
    Finestra per autorizzare/rifiutare richieste di straordinario.
    """
    
    def __init__(self, parent, db_handler, lang_manager, user_name, user_id=0):
        super().__init__(parent)
        
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_name
        self.user_id = user_id  # ID utente (idanga)
        
        # Setup finestra
        self.title(self.lang.get('overtime_approval_title', 'Autorizzazione Straordinari'))
        self.geometry("1400x700")
        self.transient(parent)
        self.grab_set()
        
        self._create_widgets()
        self._load_pending_requests()
    
    def _create_widgets(self):
        """Crea i widget dell'interfaccia."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === FILTRI ===
        filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get('filters', 'Filtri'), padding="10")
        filter_frame.pack(fill=tk.X, pady=5)
        
        # Stato
        ttk.Label(filter_frame, text=self.lang.get('status', 'Stato:')).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.status_var = tk.StringVar(value='Pending')
        status_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.status_var,
            values=['All', 'Pending', 'Approved', 'Rejected'],
            state='readonly',
            width=15
        )
        status_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        status_combo.bind('<<ComboboxSelected>>', lambda e: self._load_pending_requests())
        
        # Data Da
        ttk.Label(filter_frame, text=self.lang.get('date_from', 'Da:')).grid(row=0, column=2, padx=(20, 5), pady=5, sticky=tk.W)
        self.date_from = DateEntry(
            filter_frame,
            width=12,
            date_pattern='dd/mm/yyyy',
            background='darkblue',
            foreground='white',
            borderwidth=2
        )
        self.date_from.set_date(datetime.now() - timedelta(days=30))
        self.date_from.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        
        # Data A
        ttk.Label(filter_frame, text=self.lang.get('date_to', 'A:')).grid(row=0, column=4, padx=(10, 5), pady=5, sticky=tk.W)
        self.date_to = DateEntry(
            filter_frame,
            width=12,
            date_pattern='dd/mm/yyyy',
            background='darkblue',
            foreground='white',
            borderwidth=2
        )
        self.date_to.set_date(datetime.now())
        self.date_to.grid(row=0, column=5, padx=5, pady=5, sticky=tk.W)
        
        ttk.Button(
            filter_frame,
            text=self.lang.get('refresh', 'Aggiorna'),
            command=self._load_pending_requests
        ).grid(row=0, column=6, padx=10, pady=5)
        
        # === LISTA RICHIESTE ===
        list_frame = ttk.LabelFrame(main_frame, text=self.lang.get('requests_list', 'Richieste'), padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ('id', 'number', 'date', 'supervisor', 'employees', 'total_hours', 'status')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15, selectmode='extended')
        
        self.tree.heading('id', text='ID')
        self.tree.heading('number', text=self.lang.get('request_number', 'Numero'))
        self.tree.heading('date', text=self.lang.get('date', 'Data'))
        self.tree.heading('supervisor', text=self.lang.get('supervisor', 'Supervisore'))
        self.tree.heading('employees', text=self.lang.get('employees_count', 'N. Dipendenti'))
        self.tree.heading('total_hours', text=self.lang.get('total_hours', 'Ore Totali'))
        self.tree.heading('status', text=self.lang.get('status', 'Stato'))
        
        self.tree.column('id', width=50)
        self.tree.column('number', width=120)
        self.tree.column('date', width=150)
        self.tree.column('supervisor', width=200)
        self.tree.column('employees', width=100)
        self.tree.column('total_hours', width=100)
        self.tree.column('status', width=100)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind('<Double-1>', self._show_request_details)
        
        # === PULSANTI AZIONE ===
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            action_frame,
            text=self.lang.get('view_details', 'Visualizza Dettagli'),
            command=self._show_request_details
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text=self.lang.get('approve', 'Approva'),
            command=lambda: self._approve_or_reject(True)
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text=self.lang.get('reject', 'Rifiuta'),
            command=lambda: self._approve_or_reject(False)
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text=self.lang.get('ask_question', 'Ask Question'),
            command=self._ask_question
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            action_frame,
            text=self.lang.get('generate_report', '📊 Genera Rapporto'),
            command=self._export_filtered_to_excel
        ).pack(side=tk.RIGHT, padx=5)

        ttk.Button(
            action_frame,
            text=self.lang.get('close', 'Chiudi'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)
    
    def _export_filtered_to_excel(self):
        """Esporta i dati filtrati del treeview in un file Excel formattato."""
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare'),
                parent=self
            )
            return

        import os

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl",
                parent=self
            )
            return

        try:
            # Prepara path file
            output_dir = r"C:\Temp"
            os.makedirs(output_dir, exist_ok=True)

            date_from_str = self.date_from.get_date().strftime('%d-%m-%Y')
            date_to_str = self.date_to.get_date().strftime('%d-%m-%Y')
            status = self.status_var.get()
            filename = f"OvertimeApproval_{status}_{date_from_str}_to_{date_to_str}.xlsx"
            file_path = os.path.join(output_dir, filename)

            # Se file bloccato, aggiungi timestamp
            if os.path.exists(file_path):
                try:
                    with open(file_path, 'ab'):
                        pass
                except PermissionError:
                    ts = datetime.now().strftime('%H%M%S')
                    filename = f"OvertimeApproval_{status}_{date_from_str}_to_{date_to_str}_{ts}.xlsx"
                    file_path = os.path.join(output_dir, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Autorizzazioni Straordinari"

            # Stili
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_align = Alignment(vertical='top', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Intestazioni
            headers = ['ID', 'Numero', 'Data', 'Supervisore', 'N. Dipendenti', 'Ore Totali', 'Stato']
            col_widths = [8, 15, 18, 25, 14, 12, 12]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                ws.column_dimensions[cell.column_letter].width = col_widths[col - 1]

            # Dati dal treeview
            for row_idx, item in enumerate(items, 2):
                values = self.tree.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_align
                    cell.border = thin_border

            last_row = len(items) + 1

            # Riga totale
            total_row = last_row + 1
            ws.cell(row=total_row, column=4, value="TOTALE:").font = Font(bold=True, size=10)
            sum_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{last_row})")
            sum_cell.font = Font(bold=True, size=10)
            sum_cell.alignment = Alignment(horizontal='center')

            # Titolo informativo
            ws.cell(row=total_row + 2, column=1,
                    value=f"Periodo: {date_from_str} - {date_to_str} | Stato: {status}")

            # Auto-filtro
            ws.auto_filter.ref = f"A1:G{last_row}"

            wb.save(file_path)

            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File Excel salvato:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            if open_file:
                os.startfile(file_path)

        except Exception as e:
            logger.error(f"Errore export Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export Excel:\n{str(e)}",
                parent=self
            )

    def _load_pending_requests(self):
        """Carica le richieste in attesa di approvazione, filtrate per gerarchia."""
        # Pulisci treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        status_filter = self.status_var.get()
        
        # Recupera date dal filtro
        date_from = self.date_from.get_date().strftime('%Y-%m-%d')
        date_to = self.date_to.get_date().strftime('%Y-%m-%d 23:59:59')

        # Determina filtro gerarchico sui dipendenti
        subordinate_filter = ""
        filter_ids = []

        if self.user_id:
            from .overtime_manager import OvertimeManager
            mgr = OvertimeManager(self.db)
            if not mgr.is_manager_admin(self.user_id):
                subordinates = mgr.fetch_subordinates(self.user_id)
                if subordinates:
                    placeholders = ", ".join(["?"] * len(subordinates))
                    subordinate_filter = f"""
                    AND a.ExtraHourApprovalId IN (
                        SELECT DISTINCT ExtraHourApprovalId
                        FROM ResetServices.dbo.ExtraTimeApprovalStory
                        WHERE IdEmployee IN ({placeholders})
                    )"""
                    filter_ids = list(subordinates)
                else:
                    # Non admin e nessun subalterno: non mostrare nulla
                    logger.info(f"Nessun subalterno per ID {self.user_id}, lista richieste vuota")
                    return

        # Filtro ApprovelDate: solo Pending richiede IS NULL
        approval_date_filter = ""
        if status_filter == 'Pending':
            approval_date_filter = "AND s.ApprovelDate IS NULL"

        # Query per recuperare richieste
        query = f"""
        WITH StoryAggregated AS (
    SELECT
        ExtraHourApprovalId,
        COUNT(DISTINCT IdEmployee) AS EmployeeCount,
        SUM(DATEDIFF(HOUR, DateStart, DateEnd)) AS TotalHours,
        CASE
            WHEN MAX(ApprovedId) IS NULL THEN 'Pending'
            WHEN MAX(CAST(Approved AS INT)) = 1 THEN 'Approved'
            ELSE 'Rejected'
        END AS Status,
        ApprovelDate,
        MAX(ApprovedId) AS ApprovedId,
        MAX(CAST(Approved AS INT)) AS Approved
    FROM ResetServices.dbo.ExtraTimeApprovalStory
    WHERE Datesys >= ?
      AND Datesys <= ?
    GROUP BY ExtraHourApprovalId, ApprovelDate
)
SELECT
    a.ExtraHourApprovalId,
    ISNULL(r.NumRegistro, CAST(a.IdRegistro AS VARCHAR)) AS RequestNumber,
    a.DateSys,
    ISNULL(u.NomeUser, 'N/A') AS Supervisor,
    ISNULL(s.EmployeeCount, 0) AS EmployeeCount,
    ISNULL(s.TotalHours, 0) AS TotalHours,
    ISNULL(s.Status, 'Pending') AS Status
FROM ResetServices.dbo.ExtraTimeApproval a
LEFT JOIN StoryAggregated s
    ON a.ExtraHourApprovalId = s.ExtraHourApprovalId
LEFT JOIN resetservices.dbo.tbuserkey u
    ON a.IdChief = u.idanga
LEFT JOIN ResetServices.dbo.TbRegistro r
    ON a.IdRegistro = r.Contatore
WHERE a.DateSys >= ? AND a.DateSys <= ?
  {approval_date_filter}
  {subordinate_filter}
        """

        # Parametri: date_from, date_to per CTE + date_from, date_to per WHERE + eventuali subordinate IDs
        query_params = [date_from, date_to, date_from, date_to] + filter_ids

        if status_filter == 'Approved':
            query += " AND s.Approved = 1"
        elif status_filter == 'Rejected':
            query += " AND s.Approved = 0 AND s.ApprovedId IS NOT NULL"

        query += """
        ORDER BY a.DateSys DESC
        """

        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            cursor.close()

            for row in results:
                date_str = row[2].strftime('%d/%m/%Y %H:%M') if row[2] else 'N/D'
                self.tree.insert('', tk.END, values=(
                    row[0],  # ID
                    row[1],  # Number
                    date_str,  # Date
                    row[3],  # Supervisor
                    row[4],  # Employee count
                    row[5] if row[5] else 0,  # Total hours
                    row[6]  # Status
                ))
        except Exception as e:
            logger.error(f"Errore caricamento richieste: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore caricamento richieste:\n{str(e)}",
                parent=self
            )

    
    def _show_request_details(self, event=None):
        """Mostra i dettagli della richiesta selezionata."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_request', 'Selezionare una richiesta'),
                parent=self
            )
            return
        
        request_id = self.tree.item(selected[0])['values'][0]
        
        # Apri finestra dettagli
        RequestDetailsWindow(self, self.db, self.lang, request_id)
    
    def _approve_or_reject(self, approve):
        """Approva o rifiuta le richieste selezionate (supporta selezione multipla).
        
        Ogni richiesta viene elaborata singolarmente:
        - DB update individuale
        - Generazione PDF/Excel individuale
        - Invio email individuale
        """
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_request', 'Selezionare una richiesta'),
                parent=self
            )
            return
        
        action = 'approvare' if approve else 'rifiutare'
        count = len(selected)
        
        # Conferma unica per tutte le richieste selezionate
        if count == 1:
            confirm_msg = self.lang.get('confirm_action', f'Confermare {action} la richiesta?')
        else:
            confirm_msg = self.lang.get(
                'confirm_action_multiple',
                f'Confermare {action} {count} richieste selezionate?'
            )
        
        confirm = messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            confirm_msg,
            parent=self
        )
        
        if not confirm:
            return
        
        # Elabora ogni richiesta singolarmente
        success_count = 0
        error_list = []
        
        for item in selected:
            request_id = self.tree.item(item)['values'][0]
            request_number = self.tree.item(item)['values'][1]
            
            try:
                self._process_single_approval(request_id, request_number, approve)
                success_count += 1
            except Exception as e:
                logger.error(f"Errore elaborazione richiesta {request_number}: {e}", exc_info=True)
                error_list.append(f"{request_number}: {str(e)}")
        
        # Riepilogo finale
        if error_list:
            error_summary = '\n'.join(error_list[:10])
            messagebox.showwarning(
                self.lang.get('partial_success', 'Elaborazione Parziale'),
                f"{success_count}/{count} richieste elaborate con successo.\n\n"
                f"Errori:\n{error_summary}",
                parent=self
            )
        else:
            if count == 1:
                messagebox.showinfo(
                    self.lang.get('success', 'Successo'),
                    self.lang.get('request_updated', 'Richiesta aggiornata con successo'),
                    parent=self
                )
            else:
                messagebox.showinfo(
                    self.lang.get('success', 'Successo'),
                    f"{count} richieste elaborate con successo.",
                    parent=self
                )
        
        # Ricarica lista
        self._load_pending_requests()
    
    def _process_single_approval(self, request_id, request_number, approve):
        """Elabora una singola approvazione/rifiuto: DB, PDF, email.
        
        Raises Exception se qualcosa va storto (gestita dal chiamante).
        """
        # Recupera IdChief per inviare email
        chief_query = """
        SELECT IdChief 
        FROM ResetServices.dbo.ExtraTimeApproval 
        WHERE ExtraHourApprovalId = ?
        """
        chief_result = self.db.fetch_one(chief_query, (request_id,))
        requester_id = chief_result[0] if chief_result else None
        
        cursor = self.db.conn.cursor()
        
        # Se approvato, chiama SP Registro per ottenere ID approvazione
        approval_id = None
        if approve:
            from datetime import datetime
            registro_query = """
            EXEC ResetServices.dbo.Registro 
                @tipo=192, 
                @anno=?, 
                @data=?, 
                @operatore=?, 
                @obj=?, 
                @chichiama=0
            """
            
            cursor.execute(
                registro_query,
                (datetime.now().year, datetime.now().strftime('%Y-%m-%d'), self.user_id, request_id)
            )
            
            # Recupera il risultato della SP
            result = cursor.fetchone()
            if result:
                approval_id = result[0]
                logger.info(f"Registro SP returned approval_id: {approval_id} for request {request_id}")
            else:
                logger.warning(f"Registro SP did not return a value for request {request_id}")
                approval_id = self.user_id  # Fallback
        
        # Aggiorna stato approvazione
        update_query = """
        UPDATE ResetServices.dbo.ExtraTimeApprovalStory
        SET ApprovedId = ?,
            Approved = ?,
            ApprovelDate = GETDATE()
        WHERE ExtraHourApprovalId = ?
        """
        
        try:
            cursor.execute(update_query, (approval_id if approve else self.user_id, 1 if approve else 0, request_id))
            self.db.conn.commit()
        except Exception:
            self.db.conn.rollback()
            raise
        finally:
            cursor.close()
        
        # Invia email di notifica al richiedente
        if requester_id:
            from .overtime_manager import OvertimeManager
            manager = OvertimeManager(self.db)
            extra_attachments = []

            # Genera PDF di conferma approvazione (solo se approvato)
            if approve:
                approval_pdf = manager.generate_approval_confirmation_pdf(
                    request_id=request_id,
                    approver_name=self.user_name
                )
                if approval_pdf:
                    logger.info(f"Approval confirmation PDF generated: {approval_pdf}")
                    extra_attachments.append(approval_pdf)

                # Genera Excel con lista dipendenti, ore e totale mensile
                approval_excel = manager.generate_approval_excel(request_id=request_id)
                if approval_excel:
                    logger.info(f"Approval Excel generated: {approval_excel}")
                    extra_attachments.append(approval_excel)

                # Genera PDF foglio firma dipendenti (rumeno)
                signature_pdf = manager.generate_employee_signature_pdf(request_id=request_id)
                if signature_pdf:
                    logger.info(f"Employee signature PDF generated: {signature_pdf}")
                    extra_attachments.append(signature_pdf)

            manager.send_approval_notification(
                request_id=request_id,
                request_number=request_number,
                requester_id=requester_id,
                approved=approve,
                approver_name=self.user_name,
                extra_attachments=extra_attachments if extra_attachments else None
            )

    def _ask_question(self):
        """Apre il dialogo per porre una domanda sulla richiesta selezionata."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('select_request', 'Selezionare una richiesta'),
                parent=self
            )
            return

        request_id = self.tree.item(selected[0])['values'][0]
        request_number = self.tree.item(selected[0])['values'][1]

        # Apri dialogo domanda
        dialog = AskQuestionDialog(
            self, self.db, self.lang,
            request_id, request_number, self.user_id, self.user_name
        )
        self.wait_window(dialog)


class AskQuestionDialog(tk.Toplevel):
    """Dialogo per porre una domanda su una richiesta di straordinario."""

    def __init__(self, parent, db_handler, lang_manager, request_id, request_number, user_id, user_name):
        super().__init__(parent)

        self.db = db_handler
        self.lang = lang_manager
        self.request_id = request_id
        self.request_number = request_number
        self.user_id = user_id
        self.user_name = user_name

        self.title(self.lang.get('ask_question_title', 'Question about Overtime Request'))
        self.geometry("650x500")
        self.transient(parent)
        self.grab_set()

        self._create_widgets()
        self._load_employees()

    def _create_widgets(self):
        """Crea i widget del dialogo."""
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Riferimento richiesta
        info_frame = ttk.LabelFrame(
            main_frame,
            text=self.lang.get('request_reference', 'Request Reference'),
            padding="10"
        )
        info_frame.pack(fill=tk.X, pady=5)

        ttk.Label(info_frame, text=f"{self.lang.get('request_number', 'Numero')}:",
                  font=('', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(info_frame, text=str(self.request_number)).grid(
            row=0, column=1, sticky=tk.W, padx=5, pady=2
        )

        # Selezione dipendente (opzionale)
        ttk.Label(info_frame, text=f"{self.lang.get('select_employee_optional', 'Select Employee (optional)')}:",
                  font=('', 9, 'bold')).grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(
            info_frame, textvariable=self.employee_var,
            state='readonly', width=40
        )
        self.employee_combo.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        self._employee_map = {}  # display_name -> StoryExtraHourApprovalId

        # Testo domanda
        question_frame = ttk.LabelFrame(
            main_frame,
            text=self.lang.get('your_question', 'Your Question'),
            padding="10"
        )
        question_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.question_text = tk.Text(question_frame, height=10, wrap=tk.WORD)
        self.question_text.pack(fill=tk.BOTH, expand=True)

        # Pulsanti
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            btn_frame,
            text=self.lang.get('send_question', 'Send Question'),
            command=self._send_question
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            btn_frame,
            text=self.lang.get('cancel', 'Annulla'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)

    def _load_employees(self):
        """Carica i dipendenti della richiesta per selezione opzionale."""
        query = """
        SELECT
            s.StoryExtraHourApprovalId,
            e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName
        FROM ResetServices.dbo.ExtraTimeApprovalStory s
        INNER JOIN Employee.dbo.EmployeeHireHistory h ON s.IdEmployee = h.EmployeeHireHistoryId
        INNER JOIN Employee.dbo.Employees e ON h.EmployeeId = e.EmployeeId
        WHERE s.ExtraHourApprovalId = ?
        ORDER BY e.EmployeeSurname, e.EmployeeName
        """
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, (self.request_id,))
            rows = cursor.fetchall()
            cursor.close()

            none_label = self.lang.get('none_all', '-- None (general) --')
            names = [none_label]
            self._employee_map = {none_label: None}

            for row in rows:
                name = row[1] or 'N/A'
                names.append(name)
                self._employee_map[name] = row[0]  # StoryExtraHourApprovalId

            self.employee_combo['values'] = names
            self.employee_combo.current(0)

        except Exception as e:
            logger.error(f"Errore caricamento dipendenti: {e}", exc_info=True)

    def _send_question(self):
        """Salva la domanda nel DB e invia email."""
        question = self.question_text.get('1.0', tk.END).strip()

        if not question:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('question_empty', 'Please enter your question text.'),
                parent=self
            )
            return

        # Ottieni StoryExtraHourApprovalId selezionato (opzionale)
        selected_employee = self.employee_var.get()
        story_id = self._employee_map.get(selected_employee)

        try:
            # Recupera IdChief (richiedente) per invio email
            chief_query = """
            SELECT IdChief FROM ResetServices.dbo.ExtraTimeApproval
            WHERE ExtraHourApprovalId = ?
            """
            chief_result = self.db.fetch_one(chief_query, (self.request_id,))
            requester_id = chief_result[0] if chief_result else None

            # Inserisci nel database
            insert_query = """
            INSERT INTO ResetServices.dbo.ExtraTimeApprovalQA
                (ExtraHourApprovalId, StoryExtraHourApprovalId, QuestionBy, QuestionText)
            OUTPUT INSERTED.QAId
            VALUES (?, ?, ?, ?)
            """
            cursor = self.db.conn.cursor()
            cursor.execute(insert_query, (self.request_id, story_id, self.user_id, question))
            qa_row = cursor.fetchone()
            qa_id = qa_row[0] if qa_row else None
            self.db.conn.commit()
            cursor.close()

            # Invia email al richiedente
            if requester_id and qa_id:
                try:
                    from .overtime_manager import OvertimeManager
                    manager = OvertimeManager(self.db)
                    manager.send_question_email(
                        qa_id=qa_id,
                        request_id=self.request_id,
                        requester_id=requester_id,
                        question_text=question,
                        asker_name=self.user_name
                    )
                except Exception as email_err:
                    logger.error(f"Errore invio email domanda: {email_err}", exc_info=True)

            messagebox.showinfo(
                self.lang.get('success', 'Successo'),
                self.lang.get('question_sent_ok', 'Question sent successfully.'),
                parent=self
            )
            self.destroy()

        except Exception as e:
            self.db.conn.rollback()
            logger.error(f"Errore salvataggio domanda: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"{self.lang.get('question_send_error', 'Error sending the question.')}:\n{str(e)}",
                parent=self
            )


class RequestDetailsWindow(tk.Toplevel):
    """Finestra dettagli richiesta."""
    
    def __init__(self, parent, db_handler, lang_manager, request_id):
        super().__init__(parent)
        
        self.db = db_handler
        self.lang = lang_manager
        self.request_id = request_id
        
        self.title(self.lang.get('request_details', 'Dettagli Richiesta'))
        self.geometry("900x600")
        self.transient(parent)
        self.grab_set()
        
        self._create_widgets()
        self._load_details()
    
    def _create_widgets(self):
        """Crea i widget."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Info richiesta
        info_frame = ttk.LabelFrame(main_frame, text=self.lang.get('request_info', 'Informazioni Richiesta'), padding="10")
        info_frame.pack(fill=tk.X, pady=5)
        
        self.info_labels = {}
        for i, key in enumerate(['number', 'date', 'supervisor', 'status']):
            ttk.Label(info_frame, text=f"{self.lang.get(key, key)}:", font=('', 9, 'bold')).grid(
                row=i, column=0, sticky=tk.W, padx=5, pady=2
            )
            self.info_labels[key] = ttk.Label(info_frame, text='')
            self.info_labels[key].grid(row=i, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Dipendenti
        emp_frame = ttk.LabelFrame(main_frame, text=self.lang.get('employees', 'Dipendenti'), padding="10")
        emp_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ('employee', 'reason', 'start', 'end', 'hours', 'justify')
        self.tree = ttk.Treeview(emp_frame, columns=columns, show='headings')
        
        self.tree.heading('employee', text=self.lang.get('employee', 'Dipendente'))
        self.tree.heading('reason', text=self.lang.get('reason', 'Motivo'))
        self.tree.heading('start', text=self.lang.get('start', 'Inizio'))
        self.tree.heading('end', text=self.lang.get('end', 'Fine'))
        self.tree.heading('hours', text=self.lang.get('hours', 'Ore'))
        self.tree.heading('justify', text=self.lang.get('justify', 'Giustificazione'))
        
        scrollbar = ttk.Scrollbar(emp_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Pulsante chiudi
        ttk.Button(main_frame, text=self.lang.get('close', 'Chiudi'), command=self.destroy).pack(pady=10)
    
    def _load_details(self):
        """Carica i dettagli della richiesta."""
        # Query info richiesta
        info_query = """
        SELECT CAST(a.IdRegistro AS VARCHAR), a.DateSys, ISNULL(u.NomeUser, 'N/A')
        FROM ResetServices.dbo.ExtraTimeApproval a
        LEFT JOIN resetservices.dbo.tbuserkey u ON a.IdChief = u.idanga
        WHERE a.ExtraHourApprovalId = ?
        """
        
        # Query dipendenti
        emp_query = """
        SELECT 
            e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName,
            s.Descriptionreasons,
            s.DateStart,
            s.DateEnd,
            DATEDIFF(HOUR, s.DateStart, s.DateEnd) AS Hours,
            s.Justify
        FROM ResetServices.dbo.ExtraTimeApprovalStory s
        INNER JOIN Employee.dbo.EmployeeHireHistory h ON s.IdEmployee = h.EmployeeHireHistoryId
        INNER JOIN Employee.dbo.Employees e ON h.EmployeeId = e.EmployeeId
        WHERE s.ExtraHourApprovalId = ?
        """
        
        try:
            # Carica info
            info_result = self.db.fetch_one(info_query, (self.request_id,))
            if info_result:
                self.info_labels['number'].config(text=info_result[0])
                self.info_labels['date'].config(text=info_result[1].strftime('%d/%m/%Y %H:%M') if info_result[1] else 'N/D')
                self.info_labels['supervisor'].config(text=info_result[2] or 'N/D')
                self.info_labels['status'].config(text='Pending')  # TODO: Calcolare stato reale
            
            # Carica dipendenti
            cursor = self.db.conn.cursor()
            cursor.execute(emp_query, (self.request_id,))
            employees = cursor.fetchall()
            cursor.close()
            
            for emp in employees:
                self.tree.insert('', tk.END, values=(
                    emp[0],  # Name
                    emp[1],  # Reason
                    emp[2].strftime('%d/%m/%Y %H:%M') if emp[2] else 'N/D',  # Start
                    emp[3].strftime('%d/%m/%Y %H:%M') if emp[3] else 'N/D',  # End
                    emp[4] if emp[4] else 0,  # Hours
                    emp[5] or 'N/A'  # Justify
                ))
                
        except Exception as e:
            logger.error(f"Errore caricamento dettagli: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore caricamento dettagli:\n{str(e)}",
                parent=self
            )
