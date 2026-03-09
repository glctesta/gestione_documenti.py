"""
Overtime Manager - Business Logic
Gestione logica business per straordinari
"""

import logging
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional
import os
import tempfile

logger = logging.getLogger(__name__)


class OvertimeManager:
    """
    Gestisce la logica business per le richieste di straordinario.
    """
    
    def __init__(self, db_handler):
        """
        Inizializza il manager.
        
        Args:
            db_handler: Istanza di DatabaseHandler
        """
        self.db = db_handler
    
    # ==================== METODI DIPENDENTI ====================

    def is_manager_admin(self, manager_hire_history_id):
        """
        Verifica se il manager ha il FunctionCode massimo (= amministratore).
        L'admin non ha filtri e vede tutti i dipendenti.

        Args:
            manager_hire_history_id: EmployeeHireHistoryId del responsabile loggato

        Returns:
            bool: True se admin, False altrimenti
        """
        query = """
        SELECT TOP 1
            CASE WHEN f.FunctionCode = (SELECT MAX(FunctionCode) FROM Employee.dbo.Functions)
                 THEN 1 ELSE 0 END AS IsAdmin
        FROM Employee.dbo.EmployeeCdcStories cs
        INNER JOIN Employee.dbo.Functions f ON cs.FunctionId = f.FunctionId
        WHERE cs.EmployeeHireHistoryId = ?
          AND cs.DateOut IS NULL
        """
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, (manager_hire_history_id,))
            row = cursor.fetchone()
            cursor.close()
            return bool(row and row[0] == 1)
        except Exception as e:
            logger.error(f"Errore verifica admin per ID {manager_hire_history_id}: {e}", exc_info=True)
            return False

    def fetch_subordinates(self, manager_hire_history_id):
        """
        Recupera i EmployeeHireHistoryId di tutti i subalterni diretti del responsabile.
        Usa la query CTE gerarchica basata su FunctionCode e CdcId.

        Args:
            manager_hire_history_id: EmployeeHireHistoryId del responsabile loggato

        Returns:
            set: Set di EmployeeHireHistoryId dei subalterni, vuoto se nessuno
        """
        query = """
       WITH Manager (SubCdcId, FunctionCode, MainCdcId) AS
            (
                SELECT cs.SubCdcId, f.FunctionCode, c.CdcId
                FROM employee.dbo.EmployeeCdcStories cs
                INNER JOIN employee.dbo.CdcSub c
                    ON c.SubCdcId = cs.SubCdcId
                    AND cs.DateOut IS NULL
                INNER JOIN Employee.dbo.Functions f
                    ON cs.FunctionId = f.FunctionId
                WHERE cs.EmployeeHireHistoryId = ?
                AND cs.DateOut IS NULL
            )
            SELECT
                h.EmployeeHireHistoryId    
            FROM employee.dbo.EmployeeHireHistory h
            INNER JOIN employee.dbo.EmployeeCdcStories css
                ON h.EmployeeHireHistoryId = css.EmployeeHireHistoryId
                AND css.DateOut IS NULL
                AND h.EndWorkDate IS NULL
                AND h.employeerid = 2
            INNER JOIN employee.dbo.Employees e
                ON h.EmployeeId = e.EmployeeId
            INNER JOIN employee.dbo.CdcSub s
                ON s.SubCdcId = css.SubCdcId
            INNER JOIN employee.dbo.Functions f
                ON f.FunctionId = css.FunctionId
            INNER JOIN employee.dbo.CostCenters c
                ON c.CdcId = s.CdcId
            LEFT JOIN Manager m
                ON m.MainCdcId = s.CdcId
            AND m.FunctionCode > f.FunctionCode
            WHERE
                (
                    -- caso normale: filtrato per struttura del manager (2086)
                    h.EmployeeHireHistoryId <> ?
                    AND m.MainCdcId IS NOT NULL
                    AND f.FunctionCode < m.FunctionCode
                )
                OR
                (
                    -- caso speciale: 100 vede tutti i dipendenti
                    EXISTS (
                        SELECT 1
                        FROM employee.dbo.EmployeeCdcStories ecs
                        INNER JOIN employee.dbo.Functions fn
                            ON ecs.FunctionId = fn.FunctionId
                        WHERE ecs.EmployeeHireHistoryId = ?
                        AND ecs.DateOut IS NULL
                        AND fn.FunctionCode = 100
                    )
                )
                ORDER BY UPPER(e.employeesurname + ' ' + e.EmployeeName);
        """
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, (manager_hire_history_id, manager_hire_history_id, manager_hire_history_id))
            rows = cursor.fetchall()
            cursor.close()
            return {row[0] for row in rows if row[0] is not None}
        except Exception as e:
            logger.error(f"Errore recupero subalterni per ID {manager_hire_history_id}: {e}", exc_info=True)
            return set()

    def fetch_eligible_employees(self, month=None, year=None, manager_hire_history_id=None):
        """
        Recupera TUTTI i dipendenti subalterni con le ore mensili attuali.
        Include un flag 'Exceeded' per quelli che hanno superato il limite.
        Se manager_hire_history_id e' fornito e non e' admin, filtra per subalterni.

        Args:
            month: Mese di riferimento (default: mese corrente)
            year: Anno di riferimento (default: anno corrente)
            manager_hire_history_id: EmployeeHireHistoryId del responsabile (per filtraggio)

        Returns:
            List di tuple: (EmployeeHireHistoryId, Cognome, Nome, OreMensili, MaxOre, Exceeded)
        """
        if month is None:
            month = datetime.now().month
        if year is None:
            year = datetime.now().year

        # Calcola primo e ultimo giorno del mese
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1)
        else:
            last_day = date(year, month + 1, 1)

        # Determina filtro gerarchia
        subordinate_filter = ""
        filter_params = [first_day, last_day]

        if manager_hire_history_id:
            if not self.is_manager_admin(manager_hire_history_id):
                subordinates = self.fetch_subordinates(manager_hire_history_id)
                logger.info(f"Subalterni trovati per manager {manager_hire_history_id}: {subordinates}")
                if subordinates:
                    placeholders = ", ".join(["?"] * len(subordinates))
                    subordinate_filter = f"AND h.EmployeeHireHistoryId IN ({placeholders})"
                    filter_params = [first_day, last_day] + list(subordinates)
                else:
                    # Nessun subalterno: ritorna lista vuota (non e' admin)
                    logger.info(f"Nessun subalterno trovato per ID {manager_hire_history_id}, lista vuota")
                    return []
            else:
                logger.info(f"Manager {manager_hire_history_id} è admin, nessun filtro subalterni")

        query = f"""
        SELECT
            h.EmployeeHireHistoryId,
            e.EmployeeSurname,
            e.EmployeeName,
            ISNULL(SUM(DATEDIFF(HOUR, s.DateStart, s.DateEnd)), 0) AS MonthlyHours,
            ISNULL((SELECT MaxHourPerMonth FROM Employee.dbo.OverTimeRules WHERE DateOut IS NULL), 999) AS MaxHourPerMonth,
            CASE
                WHEN ISNULL(SUM(DATEDIFF(HOUR, s.DateStart, s.DateEnd)), 0) >=
                     ISNULL((SELECT MaxHourPerMonth FROM Employee.dbo.OverTimeRules WHERE DateOut IS NULL), 999)
                THEN 1 ELSE 0
            END AS Exceeded
        FROM Employee.dbo.Employees e
        INNER JOIN Employee.dbo.EmployeeHireHistory h
            ON e.EmployeeId = h.EmployeeId
        LEFT JOIN ResetServices.dbo.ExtraTimeApprovalStory s
            ON h.EmployeeHireHistoryId = s.IdEmployee
            AND s.DateStart >= ?
            AND s.DateStart < ?
        WHERE h.EndWorkDate IS NULL
            AND h.EmployeerId = 2
            AND NOT e.employeeName = 'ANONYMOUS'
            {subordinate_filter}
        GROUP BY h.EmployeeHireHistoryId, e.EmployeeSurname, e.EmployeeName
        ORDER BY e.EmployeeSurname, e.EmployeeName
        """

        try:
            cursor = self.db.conn.cursor()
            logger.debug(f"fetch_eligible_employees: params={filter_params}")
            cursor.execute(query, filter_params)
            results = cursor.fetchall()
            cursor.close()
            logger.info(f"fetch_eligible_employees: {len(results)} dipendenti trovati")
            if results:
                exceeded_count = sum(1 for row in results if row[5] == 1)
                logger.info(f"fetch_eligible_employees: {exceeded_count} dipendenti hanno superato il limite ore")
            return results
        except Exception as e:
            logger.error(f"Errore recupero dipendenti eligibili: {e}", exc_info=True)
            return []


    def get_employee_monthly_hours(self, employee_id, month=None, year=None):
        """
        Calcola le ore di straordinario già effettuate da un dipendente nel mese.
        
        Args:
            employee_id: ID dipendente (EmployeeHireHistoryId)
            month: Mese di riferimento
            year: Anno di riferimento
            
        Returns:
            int: Ore totali nel mese
        """
        if month is None:
            month = datetime.now().month
        if year is None:
            year = datetime.now().year
        
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1)
        else:
            last_day = date(year, month + 1, 1)
        
        query = """
        SELECT ISNULL(SUM(DATEDIFF(HOUR, DateStart, DateEnd)), 0) AS TotalHours
        FROM ResetServices.dbo.ExtraTimeApprovalStory
        WHERE IdEmployee = ?
            AND DateStart >= ?
            AND DateStart < ?
        """
        
        try:
            result = self.db.fetch_one(query, (employee_id, first_day, last_day))
            return result[0] if result else 0
        except Exception as e:
            logger.error(f"Errore calcolo ore mensili: {e}", exc_info=True)
            return 0
    
    # ==================== METODI MOTIVI ====================
    
    def fetch_overtime_reasons(self):
        """
        Recupera tutti i motivi di straordinario ordinati.
        
        Returns:
            List di motivi con (IdMotivo, OverTime_Reason, ObbgligoOrdine, Justify)
        """
        query = """
        SELECT IdMotivo, OverTime_Reason, ObbgligoOrdine, Justify, MaxIterazioniMese
        FROM ResetServices.dbo.OverTime_Reason
        ORDER BY Ordine
        """
        
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query)
            results = cursor.fetchall()
            cursor.close()
            return results
        except Exception as e:
            logger.error(f"Errore recupero motivi straordinario: {e}", exc_info=True)
            return []
    
    # ==================== METODI ORDINI ====================
    
    def fetch_recent_orders(self):
        """
        Recupera ordini recenti per associazione a straordinario.
        
        Returns:
            List di ordini con (IdOrdine, PO, QtaStory)
        """
        query = """
        SELECT o.IDOrder as IdOrdine, o.OrderNumber as PO, o.OrderQuantity as QtaStory
        FROM traceability_rs.dbo.orders o
        inner join traceability_rs.dbo.products p on o.idproduct=p.IDProduct
        WHERE 
            o.OrderDate > DATEADD(DAY, -100, GETDATE())  -- Ultimi 100 giorni        
        ORDER BY o.OrderDate DESC
        """
        
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query)
            results = cursor.fetchall()
            cursor.close()
            return results
        except Exception as e:
            logger.error(f"Errore recupero ordini recenti: {e}", exc_info=True)
            return []
    
    # ==================== VALIDAZIONI ====================
    
    def validate_overtime_request(self, employee_id, start_datetime, end_datetime, reason_id):
        """
        Valida una richiesta di straordinario.
        
        Args:
            employee_id: ID dipendente
            start_datetime: Data/ora inizio
            end_datetime: Data/ora fine
            reason_id: ID motivo
            
        Returns:
            Tuple (is_valid: bool, error_message: str)
        """
        # Validazione 1: Data fine >= Data inizio
        if end_datetime <= start_datetime:
            return False, "Data/ora fine deve essere successiva a data/ora inizio"
        
        # Validazione 2: Ore massime mensili (solo log informativo, decisione in fase approvazione)
        hours_requested = (end_datetime - start_datetime).total_seconds() / 3600
        current_hours = self.get_employee_monthly_hours(
            employee_id, 
            start_datetime.month, 
            start_datetime.year
        )
        
        max_hours_query = "SELECT MaxHourPerMonth FROM Employee.dbo.OverTimeRules WHERE DateOut IS NULL"
        max_hours_result = self.db.fetch_one(max_hours_query)
        max_hours = max_hours_result[0] if max_hours_result else 32
        
        if current_hours + hours_requested > max_hours:
            logger.warning(
                f"Dipendente {employee_id}: superamento ore massime mensili "
                f"({max_hours}h). Ore attuali: {current_hours}h, richieste: {hours_requested}h. "
                f"Decisione demandata alla fase di approvazione."
            )
        
        # Validazione 3: Verifica sovrapposizioni
        overlap_query = """
        SELECT COUNT(*) 
        FROM ResetServices.dbo.ExtraTimeApprovalStory
        WHERE IdEmployee = ?
            AND (
                (DateStart <= ? AND DateEnd > ?)
                OR (DateStart < ? AND DateEnd >= ?)
                OR (DateStart >= ? AND DateEnd <= ?)
            )
        """
        overlap_result = self.db.fetch_one(
            overlap_query, 
            (employee_id, start_datetime, start_datetime, end_datetime, end_datetime, start_datetime, end_datetime)
        )
        
        if overlap_result and overlap_result[0] > 0:
            return False, "Esiste già una richiesta di straordinario per questo dipendente in questo periodo"
        
        return True, ""
    
    # ==================== CREAZIONE RICHIESTE ====================
    
    def create_overtime_request(self, supervisor_id, supervisor_name, employees_data, orders_data=None):
        """
        Crea una nuova richiesta di straordinario.
        
        Args:
            supervisor_id: ID supervisore richiedente
            supervisor_name: Nome supervisore
            employees_data: Lista di dict con dati dipendenti
                [{
                    'employee_id': int,
                    'reason_id': int,
                    'reason_text': str,
                    'start': datetime,
                    'end': datetime,
                    'justify': str,
                    'order_id': int (optional),
                    'qty_target': int (optional)
                }]
            orders_data: Lista di dict con ordini associati (optional)
                [{'order_id': int, 'quantity': int}]
                
        Returns:
            Tuple (success: bool, request_id: int, request_number: str, error_msg: str)
        """
        try:
            # 1. Chiama SP Registro per ottenere numero richiesta
            registro_query = """
            EXEC ResetServices.dbo.Registro 
                @tipo=190, 
                @anno=?, 
                @data=?, 
                @operatore=?, 
                @obj=?, 
                @chichiama=0
            """
            
            first_order_id = orders_data[0]['order_id'] if orders_data and len(orders_data) > 0 else 0
            
            cursor = self.db.conn.cursor()
            cursor.execute(
                registro_query,
                (datetime.now().year, datetime.now().strftime('%Y-%m-%d'), supervisor_id, first_order_id)
            )
            registro_result = cursor.fetchone()
            registro_id = registro_result[0]
            request_number = registro_result[1]
            
            # 2. Inserisci in ExtraTimeApproval (usa OUTPUT per recuperare ID)
            approval_query = """
            SET NOCOUNT ON;
            DECLARE @InsertedId TABLE (Id INT);
            INSERT INTO ResetServices.dbo.ExtraTimeApproval (IdChief, IdRegistro, DateSys)
            OUTPUT INSERTED.ExtraHourApprovalId INTO @InsertedId
            VALUES (?, ?, GETDATE());
            SELECT Id FROM @InsertedId;
            """
            cursor.execute(approval_query, (supervisor_id, registro_id))
            
            # Recupera ID generato
            result = cursor.fetchone()
            
            if result is None or result[0] is None:
                raise Exception("Impossibile recuperare ID della richiesta creata. Verifica la struttura della tabella ExtraTimeApproval.")
            
            approval_id = int(result[0])
            logger.info(f"ExtraTimeApproval creato con ID: {approval_id}")
            
            # 3. Inserisci dipendenti in ExtraTimeApprovalStory (usa struttura esistente)
            story_query = """
            INSERT INTO ResetServices.dbo.ExtraTimeApprovalStory 
            (ExtraHourApprovalId, IdEmployee, Descriptionreasons, DateStart, DateEnd, 
             SuperVisorId, Justify, IdOrder, QtyTarget, DateSys)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
            """
            
            for emp in employees_data:
                cursor.execute(story_query, (
                    approval_id,
                    emp['employee_id'],
                    emp['reason_text'],
                    emp['start'],
                    emp['end'],
                    supervisor_id,
                    emp.get('justify', 'N/A'),
                    emp.get('order_id'),
                    emp.get('qty_target')
                ))
            
            # 4. Inserisci ordini associati (se presenti) - usa struttura esistente
            if orders_data and len(orders_data) > 0:
                orders_query = """
                INSERT INTO ResetServices.dbo.ExtraTimeOrders 
                (IdOrder, NoQuantityToAchieve, ExtraHourApprovalId, DateSys)
                VALUES (?, ?, ?, GETDATE())
                """
                for order in orders_data:
                    cursor.execute(orders_query, (
                        order['order_id'],
                        order['quantity'],
                        approval_id
                    ))
            
            self.db.conn.commit()
            cursor.close()
            
            logger.info(f"Richiesta straordinario creata: ID={approval_id}, Numero={request_number}")
            return True, approval_id, request_number, ""
            
        except Exception as e:
            self.db.conn.rollback()
            error_msg = f"Errore creazione richiesta: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return False, None, None, error_msg
    
    # ==================== GENERAZIONE PDF ====================
    
    def generate_overtime_pdf(self, request_id, request_number, supervisor_name, employees_data, orders_data=None):
        """
        Genera PDF accordo straordinari in inglese.
        
        Args:
            request_id: ID richiesta
            request_number: Numero richiesta
            supervisor_name: Nome supervisore
            employees_data: Dati dipendenti (deve includere 'employee_id' per calcolare ore)
            orders_data: Dati ordini (optional)
            
        Returns:
            str: Path del PDF generato, None se errore
        """
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.platypus import Image as ReportLabImage, Table, TableStyle, Paragraph
            from reportlab.lib import colors
            from reportlab.lib.styles import ParagraphStyle
            
            # Recupera numero richiesta formattato da TbRegistro
            request_number_query = """
            SELECT r.NumRegistro + ' on ' + FORMAT(r.datareg, 'd', 'ro-ro') AS RequestNumber
            FROM ResetServices.dbo.TbRegistro r 
            INNER JOIN ResetServices.dbo.ExtraTimeApproval e ON r.Contatore = e.IdRegistro
            WHERE e.ExtraHourApprovalId = ?
            """
            
            result = self.db.fetch_one(request_number_query, (request_id,))
            formatted_request_number = result[0] if result else request_number
            
            # Crea file temporaneo
            temp_file = tempfile.NamedTemporaryFile(
                prefix=f"OverTimeRequest_{formatted_request_number.replace(' ', '_').replace('/', '-')}_",
                suffix=".pdf",
                delete=False,
                dir="C:\\Temp"
            )
            file_path = temp_file.name
            temp_file.close()
            
            c = canvas.Canvas(file_path, pagesize=A4)
            width, height = A4
            
            # Helper per testo
            def draw_text(y, text, size=10, bold=False):
                font = "Helvetica-Bold" if bold else "Helvetica"
                c.setFont(font, size)
                c.drawString(2 * cm, y, text)
            
            def resolve_logo_path():
                base_dir = os.path.dirname(os.path.dirname(__file__))
                for candidate in (
                    os.path.join(base_dir, "Logo.png"),
                    os.path.join(base_dir, "logo.png"),
                    "Logo.png",
                    "logo.png",
                ):
                    if os.path.exists(candidate):
                        return candidate
                return None

            # Logo aziendale
            logo_path = resolve_logo_path()
            if logo_path and os.path.exists(logo_path):
                try:
                    logo = ReportLabImage(logo_path, width=1.8 * cm, height=1.8 * cm)
                    logo.drawOn(c, width - 3.0 * cm, height - 2.4 * cm)
                except Exception as e:
                    logger.warning(f"Cannot load logo: {e}")
            
            # Header professionale
            c.setFillColor(colors.HexColor("#1F3A5F"))
            c.rect(0, height - 3 * cm, width, 3 * cm, stroke=0, fill=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 18)
            c.drawString(2 * cm, height - 1.8 * cm, "OVERTIME AGREEMENT")
            c.setFont("Helvetica", 8.5)
            c.drawString(2 * cm, height - 2.35 * cm, f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            c.setFillColor(colors.black)
            
            y_pos = height - 4.0 * cm
            
            # Dettagli richiesta in riquadro con wrapping
            c.setFillColor(colors.HexColor("#F3F6FB"))
            c.setStrokeColor(colors.HexColor("#C5D1E0"))
            c.roundRect(1.8 * cm, y_pos - 2.2 * cm, width - 3.6 * cm, 2.0 * cm, 6, stroke=1, fill=1)
            c.setFillColor(colors.HexColor("#1F3A5F"))
            c.setFont("Helvetica-Bold", 11)
            c.drawString(2.2 * cm, y_pos - 0.65 * cm, "REQUEST DETAILS")
            detail_style = ParagraphStyle(
                name='overtime_detail_style',
                fontName='Helvetica',
                fontSize=9,
                leading=10.8,
                wordWrap='CJK'
            )
            req_line = Paragraph(f"<b>Request Number:</b> {formatted_request_number}", detail_style)
            req_line.wrapOn(c, width - 4.4 * cm, 0.5 * cm)
            req_line.drawOn(c, 2.2 * cm, y_pos - 1.3 * cm)
            by_line = Paragraph(f"<b>Requested by:</b> {supervisor_name}", detail_style)
            by_line.wrapOn(c, width - 4.4 * cm, 0.5 * cm)
            by_line.drawOn(c, 2.2 * cm, y_pos - 1.8 * cm)
            y_pos -= 2.8 * cm
            
            # Disclaimer box
            c.setStrokeColor(colors.red)
            c.setLineWidth(2)
            c.rect(2 * cm, y_pos - 1.5 * cm, width - 4 * cm, 1.2 * cm)
            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.red)
            c.drawCentredString(width / 2, y_pos - 0.6 * cm, "IMPORTANT NOTICE")
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.black)
            notice_style = ParagraphStyle(
                name='overtime_notice_style',
                fontName='Helvetica',
                fontSize=8,
                leading=9.2,
                alignment=1,
                wordWrap='CJK'
            )
            notice = Paragraph(
                "This is a REQUEST. Overtime work cannot be performed until explicitly approved.",
                notice_style
            )
            notice_width = width - 4.6 * cm
            _, notice_height = notice.wrap(notice_width, 0.8 * cm)
            notice.drawOn(c, 2.3 * cm, y_pos - 0.8 * cm - notice_height)
            y_pos -= 2 * cm
            
            # Tabella dipendenti con ore già effettuate
            draw_text(y_pos, "EMPLOYEES INVOLVED", 14, True)
            y_pos -= 0.8 * cm
            
            # Header con word-wrap reale tramite Paragraph
            cell_style = ParagraphStyle(
                name='overtime_cell_style',
                fontName='Helvetica',
                fontSize=7,
                leading=8.2,
                alignment=1,
                wordWrap='CJK'
            )
            table_data = [[
                Paragraph('Employee', cell_style),
                Paragraph('Start Date/Time', cell_style),
                Paragraph('End Date/Time', cell_style),
                Paragraph('Hours', cell_style),
                Paragraph('Current Month Hrs', cell_style),
                Paragraph('Reason', cell_style),
                Paragraph('Order (Target Qty)', cell_style)
            ]]

            for emp in employees_data:
                hours = (emp['end'] - emp['start']).total_seconds() / 3600
                
                # Recupera ore già effettuate nel mese
                employee_id = emp.get('employee_id')
                current_hours = 0
                if employee_id:
                    current_hours = self.get_employee_monthly_hours(
                        employee_id,
                        emp['start'].month,
                        emp['start'].year
                    )
                
                # Formatta ordine se presente
                order_info = 'N/A'
                if emp.get('order_id') and emp.get('order_number'):
                    qty = emp.get('qty_target', 0)
                    order_info = f"{emp['order_number']} ({qty} pcs)"
                
                table_data.append([
                    Paragraph(str(emp.get('name', '')), cell_style),
                    Paragraph(emp['start'].strftime('%d/%m/%Y %H:%M'), cell_style),
                    Paragraph(emp['end'].strftime('%d/%m/%Y %H:%M'), cell_style),
                    Paragraph(f"{hours:.1f}", cell_style),
                    Paragraph(f"{current_hours:.1f}", cell_style),
                    Paragraph(str(emp.get('reason', '')), cell_style),
                    Paragraph(order_info, cell_style)
                ])
            
            # Tabella con larghezze adattate
            table = Table(table_data, colWidths=[3.0*cm, 2.4*cm, 2.4*cm, 1.3*cm, 1.8*cm, 3.8*cm, 2.6*cm], repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor("#A9B9CF")),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FBFD")]),
            ]))
            
            table.wrapOn(c, width, height)
            table_height = table._height
            table.drawOn(c, 1.4 * cm, y_pos - table_height)
            y_pos -= (table_height + 1 * cm)
            
            # Footer in inglese
            c.setStrokeColor(colors.HexColor("#D7DFEB"))
            c.line(1.5 * cm, 2.1 * cm, width - 1.5 * cm, 2.1 * cm)
            c.setFillColor(colors.HexColor("#4D5E73"))
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2, 1.6 * cm, 
                "Document automatically generated by TraceabilityRS system")
            
            c.save()
            logger.info(f"PDF generated: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}", exc_info=True)
            return None
    
    def generate_approval_confirmation_pdf(self, request_id, approver_name):
        """
        Genera PDF di conferma approvazione con RequestNumber e ApprovalNumber.
        
        Args:
            request_id: ID richiesta
            approver_name: Nome di chi ha approvato
            
        Returns:
            str: Path del PDF generato, None se errore
        """
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.platypus import Image as ReportLabImage, Table, TableStyle, Paragraph
            from reportlab.lib import colors
            from reportlab.lib.styles import ParagraphStyle
            
            # Recupera numeri richiesta e approvazione da TbRegistro
            numbers_query = """
            SELECT DISTINCT r.NumRegistro + ' on ' + FORMAT(r.datareg, 'd', 'ro-ro') AS RequestNumber,
                   r1.NumRegistro + ' on ' + FORMAT(r1.datareg, 'd', 'ro-ro') AS ApprovalNumber
            FROM ResetServices.dbo.TbRegistro r 
            INNER JOIN ResetServices.dbo.ExtraTimeApproval e ON r.Contatore = e.IdRegistro
            INNER JOIN ResetServices.dbo.ExtraTimeApprovalStory es ON es.ExtraHourApprovalId = e.ExtraHourApprovalId
            INNER JOIN ResetServices.dbo.TbRegistro r1 ON r1.Contatore = es.ApprovedId
            WHERE e.ExtraHourApprovalId = ?
            """
            
            result = self.db.fetch_one(numbers_query, (request_id,))
            if not result:
                logger.error(f"Could not retrieve numbers for request_id: {request_id}")
                return None
            
            request_number, approval_number = result[0], result[1]

            # Recupera dettaglio dipendenti/ore autorizzate e ordini collegati
            employees_query = """
            SELECT
                e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName,
                s.DateStart,
                s.DateEnd,
                CAST(DATEDIFF(MINUTE, s.DateStart, s.DateEnd) / 60.0 AS DECIMAL(10,2)) AS AuthorizedHours,
                s.IdOrder,
                o.OrderNumber,
                s.QtyTarget
            FROM ResetServices.dbo.ExtraTimeApprovalStory s
            INNER JOIN Employee.dbo.EmployeeHireHistory h
                ON s.IdEmployee = h.EmployeeHireHistoryId
            INNER JOIN Employee.dbo.Employees e
                ON h.EmployeeId = e.EmployeeId
            LEFT JOIN traceability_rs.dbo.orders o
                ON o.IDOrder = s.IdOrder
            WHERE s.ExtraHourApprovalId = ?
            ORDER BY e.EmployeeSurname, e.EmployeeName, s.DateStart
            """
            cursor = self.db.conn.cursor()
            cursor.execute(employees_query, (request_id,))
            employee_rows = cursor.fetchall()
            cursor.close()
            
            # Crea file temporaneo
            temp_file = tempfile.NamedTemporaryFile(
                prefix=f"ApprovalConfirmation_{approval_number.replace(' ', '_').replace('/', '-')}_",
                suffix=".pdf",
                delete=False,
                dir="C:\\Temp"
            )
            file_path = temp_file.name
            temp_file.close()
            
            c = canvas.Canvas(file_path, pagesize=A4)
            width, height = A4
            
            # Helper per testo
            def draw_text(y, text, size=10, bold=False):
                font = "Helvetica-Bold" if bold else "Helvetica"
                c.setFont(font, size)
                c.drawString(2 * cm, y, text)
            
            def resolve_logo_path():
                base_dir = os.path.dirname(os.path.dirname(__file__))
                for candidate in (
                    os.path.join(base_dir, "Logo.png"),
                    os.path.join(base_dir, "logo.png"),
                    "Logo.png",
                    "logo.png",
                ):
                    if os.path.exists(candidate):
                        return candidate
                return None

            # Logo aziendale
            logo_path = resolve_logo_path()
            if logo_path and os.path.exists(logo_path):
                try:
                    logo = ReportLabImage(logo_path, width=1.8 * cm, height=1.8 * cm)
                    logo.drawOn(c, width - 3.0 * cm, height - 2.4 * cm)
                except Exception as e:
                    logger.warning(f"Cannot load logo: {e}")
            
            # Header
            c.setFillColor(colors.HexColor("#1F3A5F"))
            c.rect(0, height - 3 * cm, width, 3 * cm, stroke=0, fill=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 18)
            c.drawString(2 * cm, height - 1.8 * cm, "OVERTIME APPROVAL CONFIRMATION")
            c.setFont("Helvetica", 8.5)
            c.drawString(2 * cm, height - 2.35 * cm, f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            c.setFillColor(colors.black)
            
            y_pos = height - 4.0 * cm
            
            # Box di conferma verde
            c.setFillColor(colors.HexColor("#28A745"))
            c.setStrokeColor(colors.HexColor("#28A745"))
            c.setLineWidth(2)
            c.rect(2 * cm, y_pos - 2 * cm, width - 4 * cm, 1.5 * cm, fill=0)
            
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(width / 2, y_pos - 0.8 * cm, "APPROVED")
            c.setFont("Helvetica", 10)
            
            y_pos -= 2.5 * cm
            
            # Dettagli con tabella (word-wrap sicuro)
            draw_text(y_pos, "APPROVAL DETAILS", 14, True)
            y_pos -= 0.8 * cm

            details_style = ParagraphStyle(
                name='approval_details_style',
                fontName='Helvetica',
                fontSize=9,
                leading=11,
                wordWrap='CJK'
            )

            details_data = [
                [Paragraph('<b>Request Number</b>', details_style), Paragraph(str(request_number), details_style)],
                [Paragraph('<b>Approval Number</b>', details_style), Paragraph(str(approval_number), details_style)],
                [Paragraph('<b>Approved by</b>', details_style), Paragraph(str(approver_name), details_style)],
                [Paragraph('<b>Approval Date</b>', details_style), Paragraph(datetime.now().strftime('%d/%m/%Y %H:%M'), details_style)],
            ]

            details_table = Table(details_data, colWidths=[5.0 * cm, width - 9.0 * cm])
            details_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.7, colors.HexColor('#A9B9CF')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F9FBFD')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            details_table.wrapOn(c, width, height)
            details_table.drawOn(c, 2 * cm, y_pos - details_table._height)

            y_pos = y_pos - details_table._height - 0.9 * cm

            # Sezione dipendenti/ore autorizzate + ordini (se presenti)
            c.setFillColor(colors.HexColor("#1F3A5F"))
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2 * cm, y_pos, "AUTHORIZED EMPLOYEES")
            y_pos -= 0.6 * cm

            employee_cell_style = ParagraphStyle(
                name='approval_employee_cell_style',
                fontName='Helvetica',
                fontSize=8,
                leading=9.4,
                alignment=1,
                wordWrap='CJK'
            )

            has_orders = any((row[4] is not None) or (row[5] is not None) for row in employee_rows)

            if has_orders:
                employee_table_data = [[
                    Paragraph("Employee", employee_cell_style),
                    Paragraph("Time Window", employee_cell_style),
                    Paragraph("Authorized Hours", employee_cell_style),
                    Paragraph("Order / Target Qty", employee_cell_style),
                ]]
            else:
                employee_table_data = [[
                    Paragraph("Employee", employee_cell_style),
                    Paragraph("Time Window", employee_cell_style),
                    Paragraph("Authorized Hours", employee_cell_style),
                ]]

            for row in employee_rows:
                employee_name = row[0] or "N/A"
                date_start = row[1]
                date_end = row[2]
                authorized_hours = row[3] if row[3] is not None else 0
                order_number = row[5]
                qty_target = row[6]

                time_window = "N/A"
                if date_start and date_end:
                    time_window = f"{date_start.strftime('%d/%m/%Y %H:%M')} - {date_end.strftime('%d/%m/%Y %H:%M')}"

                if has_orders:
                    if order_number:
                        order_info = f"{order_number} / {int(qty_target) if qty_target is not None else 0}"
                    else:
                        order_info = "N/A"
                    employee_table_data.append([
                        Paragraph(str(employee_name), employee_cell_style),
                        Paragraph(time_window, employee_cell_style),
                        Paragraph(f"{float(authorized_hours):.2f}", employee_cell_style),
                        Paragraph(order_info, employee_cell_style),
                    ])
                else:
                    employee_table_data.append([
                        Paragraph(str(employee_name), employee_cell_style),
                        Paragraph(time_window, employee_cell_style),
                        Paragraph(f"{float(authorized_hours):.2f}", employee_cell_style),
                    ])

            if has_orders:
                employee_col_widths = [5.0 * cm, 4.0 * cm, 2.8 * cm, width - 2 * cm - 2 * cm - 5.0 * cm - 4.0 * cm - 2.8 * cm]
            else:
                employee_col_widths = [5.5 * cm, 5.0 * cm, width - 2 * cm - 2 * cm - 5.5 * cm - 5.0 * cm]

            employee_table = Table(employee_table_data, colWidths=employee_col_widths, repeatRows=1)
            employee_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.7, colors.HexColor('#A9B9CF')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FBFD')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            employee_table.wrapOn(c, width, height)

            # Evita overflow verticale
            if y_pos - employee_table._height < 2.5 * cm:
                c.showPage()
                c.setFillColor(colors.HexColor("#1F3A5F"))
                c.rect(0, height - 2.5 * cm, width, 2.5 * cm, stroke=0, fill=1)
                c.setFillColor(colors.white)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2 * cm, height - 1.6 * cm, "OVERTIME APPROVAL CONFIRMATION - DETAILS")
                c.setFillColor(colors.black)
                y_pos = height - 3.2 * cm

            employee_table.drawOn(c, 2 * cm, y_pos - employee_table._height)
            
            # Footer uniforme al report richiesta
            c.setStrokeColor(colors.HexColor("#D7DFEB"))
            c.line(1.5 * cm, 2.1 * cm, width - 1.5 * cm, 2.1 * cm)
            c.setFillColor(colors.HexColor("#4D5E73"))
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2, 1.6 * cm, 
                "Document automatically generated by TraceabilityRS system")
            
            c.save()
            logger.info(f"Approval confirmation PDF generated: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error generating approval PDF: {e}", exc_info=True)
            return None

    def generate_approval_excel(self, request_id):
        """
        Genera un file Excel con il dettaglio dei dipendenti autorizzati allo
        straordinario: nome, data, ore autorizzate e totale ore già effettuate
        nel mese corrente da ciascun dipendente.

        Args:
            request_id: ID richiesta (ExtraHourApprovalId)

        Returns:
            str: Path del file Excel generato, None in caso di errore
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # ── Query dati dipendenti della richiesta ──────────────────────
            query = """
            SELECT
                h.EmployeeHireHistoryId,
                e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName,
                s.DateStart,
                s.DateEnd,
                CAST(DATEDIFF(MINUTE, s.DateStart, s.DateEnd) / 60.0 AS DECIMAL(10,2)) AS AuthorizedHours
            FROM ResetServices.dbo.ExtraTimeApprovalStory s
            INNER JOIN Employee.dbo.EmployeeHireHistory h
                ON s.IdEmployee = h.EmployeeHireHistoryId
            INNER JOIN Employee.dbo.Employees e
                ON h.EmployeeId = e.EmployeeId
            WHERE s.ExtraHourApprovalId = ?
            ORDER BY e.EmployeeSurname, e.EmployeeName, s.DateStart
            """
            cursor = self.db.conn.cursor()
            cursor.execute(query, (request_id,))
            rows = cursor.fetchall()
            cursor.close()

            if not rows:
                logger.warning(f"generate_approval_excel: no employees found for request_id={request_id}")
                return None

            # ── Query costo per dipendente/data ────────────────────────────
            cost_query = """
            SELECT s.IdEmployee,
                   SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 AS OvertimeHours,
                   SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 * x.ValueITem AS OvertimeCost,
                   x.Currency
            FROM [ResetServices].[dbo].[ExtraTimeApprovalStory] s
            INNER JOIN [ResetServices].[dbo].[ExtraTimeApproval] a
                ON a.ExtraHourApprovalId = s.ExtraHourApprovalId
            OUTER APPLY (
                SELECT ot.ValueITem, t.[DESC] AS Currency
                FROM [ResetServices].[dbo].[OverTimeDefaults] ot
                INNER JOIN [ResetServices].[dbo].[OverTimeDescriptions] od
                    ON ot.DescriptionId = od.DescpriptionId
                INNER JOIN ResetServices.dbo.TbValute t
                    ON t.IdValuta = ot.CurrencyId
                WHERE ot.DateOut IS NULL
                  AND od.DescpriptionId = 1
            ) AS x
            WHERE CAST(s.DateStart AS DATE) = ?
            GROUP BY s.IdEmployee, x.ValueITem, x.Currency
            """
            # Dict lookup: (hire_id, date_str) -> (cost, currency)
            unique_dates = list({r[2].date() for r in rows if r[2]})
            cost_lookup = {}
            for udate in unique_dates:
                try:
                    cur2 = self.db.conn.cursor()
                    cur2.execute(cost_query, (udate.strftime('%Y-%m-%d'),))
                    for crow in cur2.fetchall():
                        key = (crow[0], udate.strftime('%Y-%m-%d'))
                        cost_lookup[key] = (
                            float(crow[2]) if crow[2] is not None else 0.0,
                            crow[3] or ''
                        )
                    cur2.close()
                except Exception as _ce:
                    logger.warning(f"Cost query failed for date {udate}: {_ce}")

            header_fill   = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
            header_font   = Font(bold=True, color="FFFFFF", size=10)
            alt_fill      = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
            center_align  = Alignment(horizontal="center", vertical="center")
            left_align    = Alignment(horizontal="left",   vertical="center")
            thin_border   = Border(
                left=Side(style="thin", color="A9B9CF"),
                right=Side(style="thin", color="A9B9CF"),
                top=Side(style="thin", color="A9B9CF"),
                bottom=Side(style="thin", color="A9B9CF")
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Overtime Approved"

            # ── Intestazione colonne ──────────────────────────────────────
            headers = [
                "Employee",
                "Date",
                "Start Time",
                "End Time",
                "Authorized Hours",
                "Monthly Hours (already done)",
                "Overtime Cost",
                "Currency",
            ]
            col_widths = [32, 14, 12, 12, 18, 28, 16, 10]

            for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[1].height = 20

            # ── Righe dati ────────────────────────────────────────────────
            for row_idx, row in enumerate(rows, start=2):
                hire_id        = row[0]
                emp_name       = row[1]
                date_start     = row[2]   # datetime
                date_end       = row[3]   # datetime
                auth_hours     = float(row[4]) if row[4] is not None else 0.0

                # Ore mensili già effettuate (mese del giorno di straordinario)
                ref_month = date_start.month if date_start else datetime.now().month
                ref_year  = date_start.year  if date_start else datetime.now().year
                monthly_hours = self.get_employee_monthly_hours(hire_id, ref_month, ref_year)

                # Costo da lookup
                date_key = date_start.date().strftime('%Y-%m-%d') if date_start else ''
                ot_cost, ot_currency = cost_lookup.get((hire_id, date_key), (0.0, ''))

                fill = alt_fill if row_idx % 2 == 0 else None

                def _cell(col, value, align=center_align, _fill=fill):
                    c = ws.cell(row=row_idx, column=col, value=value)
                    c.border    = thin_border
                    c.alignment = align
                    if _fill:
                        c.fill = _fill
                    return c

                _cell(1, emp_name,    align=left_align)
                _cell(2, date_start.strftime("%d/%m/%Y")    if date_start else "N/A")
                _cell(3, date_start.strftime("%H:%M")       if date_start else "N/A")
                _cell(4, date_end.strftime("%H:%M")         if date_end   else "N/A")
                _cell(5, round(auth_hours, 2))
                _cell(6, round(float(monthly_hours), 2))
                _cell(7, round(ot_cost, 2))
                _cell(8, ot_currency, align=left_align)

            # ── Totale ────────────────────────────────────────────────────
            total_row = len(rows) + 2
            total_font = Font(bold=True, size=10)
            for col in range(1, 9):
                c = ws.cell(row=total_row, column=col)
                c.border    = thin_border
                c.font      = total_font
                c.fill      = PatternFill(start_color="D7DFEB", end_color="D7DFEB", fill_type="solid")
                c.alignment = center_align
            ws.cell(row=total_row, column=1, value="TOTAL").alignment = left_align
            ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row - 1})")
            ws.cell(row=total_row, column=6).value = ""
            ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row - 1})")
            ws.cell(row=total_row, column=8).value = ""

            # ── Salva file temporaneo ─────────────────────────────────────
            temp_file = tempfile.NamedTemporaryFile(
                prefix=f"OvertimeApproval_{request_id}_",
                suffix=".xlsx",
                delete=False,
                dir="C:\\Temp"
            )
            file_path = temp_file.name
            temp_file.close()
            wb.save(file_path)

            logger.info(f"Approval Excel generated: {file_path}")
            return file_path

        except Exception as e:
            logger.error(f"Error generating approval Excel: {e}", exc_info=True)
            return None

    # ==================== SIGNATURE PDF (Romanian) ====================

    def generate_employee_signature_pdf(self, request_id):
        """
        Genera un PDF in rumeno con la lista dei dipendenti autorizzati,
        completo di colonna firma per accettazione individuale.

        Args:
            request_id: ID richiesta (ExtraHourApprovalId)

        Returns:
            str: Path del PDF generato, None in caso di errore
        """
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm, mm
            from reportlab.platypus import (
                Image as ReportLabImage, Table, TableStyle, Paragraph
            )
            from reportlab.lib import colors
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # ── Register Arial TTF for Romanian diacritics ──────────────
            FONT = 'Arial'
            FONT_BOLD = 'Arial-Bold'
            try:
                pdfmetrics.registerFont(
                    TTFont(FONT, r'C:\Windows\Fonts\arial.ttf'))
                pdfmetrics.registerFont(
                    TTFont(FONT_BOLD, r'C:\Windows\Fonts\arialbd.ttf'))
            except Exception as fe:
                logger.warning(f"Cannot register Arial TTF: {fe}, "
                               f"falling back to Helvetica")
                FONT = 'Helvetica'
                FONT_BOLD = 'Helvetica-Bold'

            # ── Romanian weekday names ──────────────────────────────────
            RO_DAYS = {
                0: 'Luni',
                1: 'Mar\u021bi',
                2: 'Miercuri',
                3: 'Joi',
                4: 'Vineri',
                5: 'S\u00e2mb\u0103t\u0103',
                6: 'Duminic\u0103',
            }

            # ── Query authorization number ───────────────────────────────
            auth_query = """
            SELECT DISTINCT
                r.NumRegistro + '/' + FORMAT(r.DataReg, 'd', 'ro-ro') AS OvertimeRequest,
                r1.NumRegistro + '/' + FORMAT(r1.DataReg, 'd', 'ro-ro') AS OvertimeApproval
            FROM ResetServices.dbo.ExtraTimeApproval a
            INNER JOIN ResetServices.dbo.TbRegistro r
                ON a.IdRegistro = r.Contatore
            INNER JOIN ResetServices.dbo.ExtraTimeApprovalStory es
                ON a.ExtraHourApprovalId = es.ExtraHourApprovalId
            INNER JOIN ResetServices.dbo.TbRegistro r1
                ON r1.Contatore = es.ApprovedId
            WHERE a.ExtraHourApprovalId = ?
            """
            cursor = self.db.conn.cursor()
            cursor.execute(auth_query, (request_id,))
            auth_row = cursor.fetchone()
            cursor.close()
            overtime_request = ''
            overtime_approval = ''
            if auth_row:
                overtime_request = auth_row[0] or ''
                overtime_approval = auth_row[1] or ''

            # ── Query dati dipendenti ───────────────────────────────────
            emp_query = """
            SELECT
                e.EmployeeSurname,
                e.EmployeeName,
                s.DateStart,
                s.DateEnd,
                CAST(DATEDIFF(MINUTE, s.DateStart, s.DateEnd) / 60.0
                     AS DECIMAL(10,2)) AS AuthorizedHours
            FROM ResetServices.dbo.ExtraTimeApprovalStory s
            INNER JOIN Employee.dbo.EmployeeHireHistory h
                ON s.IdEmployee = h.EmployeeHireHistoryId
            INNER JOIN Employee.dbo.Employees e
                ON h.EmployeeId = e.EmployeeId
            WHERE s.ExtraHourApprovalId = ?
            ORDER BY e.EmployeeSurname, e.EmployeeName, s.DateStart
            """
            cursor = self.db.conn.cursor()
            cursor.execute(emp_query, (request_id,))
            rows = cursor.fetchall()
            cursor.close()

            if not rows:
                logger.warning(
                    f"generate_employee_signature_pdf: no employees for "
                    f"request_id={request_id}"
                )
                return None

            # ── Resolve logo ────────────────────────────────────────────
            def resolve_logo_path():
                base_dir = os.path.dirname(os.path.dirname(__file__))
                for candidate in (
                    os.path.join(base_dir, "Logo.png"),
                    os.path.join(base_dir, "logo.png"),
                    r"c:\Users\gtesta\PythonProjetcs\Python\PrductionDocumentation\Logo.png",
                    "Logo.png",
                    "logo.png",
                ):
                    if os.path.exists(candidate):
                        return candidate
                return None

            # ── Temp file ───────────────────────────────────────────────
            temp_file = tempfile.NamedTemporaryFile(
                prefix=f"SignatureSheet_{request_id}_",
                suffix=".pdf",
                delete=False,
                dir="C:\\Temp"
            )
            file_path = temp_file.name
            temp_file.close()

            c = canvas.Canvas(file_path, pagesize=A4)
            width, height = A4

            # ── Header bar (draw FIRST) ─────────────────────────────────
            NAVY = colors.HexColor("#1F3A5F")
            c.setFillColor(NAVY)
            c.rect(0, height - 3 * cm, width, 3 * cm, stroke=0, fill=1)
            c.setFillColor(colors.white)
            c.setFont(FONT_BOLD, 16)
            c.drawString(
                2 * cm, height - 1.8 * cm,
                "ACORD ORE SUPLIMENTARE"
            )
            # Authorization number subtitle
            auth_label = ''
            if overtime_request and overtime_approval:
                auth_label = f"Cerere: {overtime_request}  |  Autoriza\u021bie: {overtime_approval}"
            elif overtime_approval:
                auth_label = f"Autoriza\u021bie: {overtime_approval}"
            c.setFont(FONT_BOLD, 9)
            c.drawString(
                2 * cm, height - 2.25 * cm,
                auth_label
            )
            c.setFont(FONT, 7.5)
            c.drawString(
                2 * cm, height - 2.65 * cm,
                f"Generat la {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            )
            c.setFillColor(colors.black)

            # ── Logo (draw AFTER header so it appears on top) ───────────
            logo_path = resolve_logo_path()
            if logo_path:
                try:
                    logo = ReportLabImage(
                        logo_path, width=1.8 * cm, height=1.8 * cm
                    )
                    logo.drawOn(c, width - 3.0 * cm, height - 2.4 * cm)
                except Exception as le:
                    logger.warning(f"Cannot load logo: {le}")

            y_pos = height - 4.0 * cm

            # ── Romanian disclaimer text ────────────────────────────────
            disclaimer_style = ParagraphStyle(
                name='ro_disclaimer',
                fontName=FONT,
                fontSize=10,
                leading=13,
                alignment=TA_JUSTIFY,
                wordWrap='CJK',
            )
            disclaimer_text = (
                "Pentru persoanele de mai jos a fost colectat\u0103 "
                "aprobarea individual\u0103 "
                "(fiecare angajat semneaz\u0103 pentru acceptarea "
                "orelor suplimentare care \u00eel privesc)."
            )
            disclaimer_para = Paragraph(disclaimer_text, disclaimer_style)
            dw, dh = disclaimer_para.wrap(width - 4 * cm, 3 * cm)
            disclaimer_para.drawOn(c, 2 * cm, y_pos - dh)
            y_pos -= (dh + 0.8 * cm)

            # ── Cell style for table ────────────────────────────────────
            cell_style = ParagraphStyle(
                name='sig_cell',
                fontName=FONT,
                fontSize=8,
                leading=9.5,
                alignment=TA_CENTER,
                wordWrap='CJK',
            )
            cell_left = ParagraphStyle(
                name='sig_cell_left',
                fontName=FONT,
                fontSize=8,
                leading=9.5,
                alignment=TA_LEFT,
                wordWrap='CJK',
            )

            # ── Build table data ────────────────────────────────────────
            header_style = ParagraphStyle(
                name='sig_header',
                fontName=FONT_BOLD,
                fontSize=8,
                leading=9.5,
                alignment=TA_CENTER,
                textColor=colors.white,
                wordWrap='CJK',
            )
            table_data = [[
                Paragraph('Nr.', header_style),
                Paragraph('Nume', header_style),
                Paragraph('Prenume', header_style),
                Paragraph('Ore', header_style),
                Paragraph('Data', header_style),
                Paragraph('De la', header_style),
                Paragraph('P\u00e2n\u0103 la', header_style),
                Paragraph('Semn\u0103tur\u0103', header_style),
            ]]

            total_hours = 0.0
            for idx, row in enumerate(rows, start=1):
                surname = row[0] or ''
                firstname = row[1] or ''
                date_start = row[2]
                date_end = row[3]
                hours = float(row[4]) if row[4] is not None else 0.0
                total_hours += hours

                # Date + weekday in Romanian
                if date_start:
                    weekday = RO_DAYS.get(date_start.weekday(), '')
                    date_str = (
                        f"{date_start.strftime('%d/%m/%Y')}<br/>({weekday})"
                    )
                    from_str = date_start.strftime('%H:%M')
                else:
                    date_str = 'N/A'
                    from_str = 'N/A'

                to_str = date_end.strftime('%H:%M') if date_end else 'N/A'

                table_data.append([
                    Paragraph(str(idx), cell_style),
                    Paragraph(surname, cell_left),
                    Paragraph(firstname, cell_left),
                    Paragraph(f"{hours:.1f}", cell_style),
                    Paragraph(date_str, cell_style),
                    Paragraph(from_str, cell_style),
                    Paragraph(to_str, cell_style),
                    Paragraph('', cell_style),  # empty for signature
                ])

            # ── Totals row ──────────────────────────────────────────────
            total_style = ParagraphStyle(
                name='sig_total',
                fontName=FONT_BOLD,
                fontSize=9,
                leading=10,
                alignment=TA_CENTER,
                wordWrap='CJK',
            )
            total_left = ParagraphStyle(
                name='sig_total_left',
                fontName=FONT_BOLD,
                fontSize=9,
                leading=10,
                alignment=TA_LEFT,
                wordWrap='CJK',
            )
            num_employees = len(rows)
            table_data.append([
                Paragraph('', total_style),
                Paragraph('TOTAL', total_left),
                Paragraph(f'{num_employees} angaja\u021bi', total_style),
                Paragraph(f'{total_hours:.1f}', total_style),
                Paragraph('', total_style),
                Paragraph('', total_style),
                Paragraph('', total_style),
                Paragraph('', total_style),
            ])

            # ── Draw table ──────────────────────────────────────────────
            col_widths = [
                1.0 * cm,   # Nr.
                3.2 * cm,   # Nume
                3.2 * cm,   # Prenume
                1.4 * cm,   # Ore
                3.0 * cm,   # Data
                1.5 * cm,   # De la
                1.5 * cm,   # Pana la
                3.2 * cm,   # Semnatura
            ]

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), NAVY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                # Totals row
                ('BACKGROUND', (0, -1), (-1, -1),
                 colors.HexColor("#D7DFEB")),
                ('FONTNAME', (0, -1), (-1, -1), FONT_BOLD),
                # Body
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.6,
                 colors.HexColor("#A9B9CF")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2),
                 [colors.white, colors.HexColor("#F9FBFD")]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                # Signature column min height for actual signature
                ('BOTTOMPADDING', (7, 1), (7, -2), 12),
            ]))

            table.wrapOn(c, width, height)
            table_height = table._height

            # If table overflows page, start new page
            if y_pos - table_height < 2.5 * cm:
                c.showPage()
                c.setFillColor(NAVY)
                c.rect(0, height - 2.5 * cm, width, 2.5 * cm,
                       stroke=0, fill=1)
                c.setFillColor(colors.white)
                c.setFont(FONT_BOLD, 14)
                c.drawString(
                    2 * cm, height - 1.6 * cm,
                    "ACORD ORE SUPLIMENTARE - continuare"
                )
                c.setFillColor(colors.black)
                y_pos = height - 3.2 * cm

            table.drawOn(c, 1.4 * cm, y_pos - table_height)
            y_pos_after_table = y_pos - table_height - 0.6 * cm

            # ── Obligation text (Romanian) after table ──────────────────
            obligation_style = ParagraphStyle(
                name='ro_obligation',
                fontName=FONT,
                fontSize=9,
                leading=12,
                alignment=TA_JUSTIFY,
                wordWrap='CJK',
            )
            obligation_text = (
                "Acest document, odat\u0103 colectate toate semn\u0103turile "
                "angaja\u021bilor din prezenta list\u0103, de c\u0103tre "
                "responsabilul care a solicitat aprobarea orelor "
                "suplimentare, trebuie predat obligatoriu biroului "
                "de personal."
            )
            obligation_para = Paragraph(obligation_text, obligation_style)
            ow, oh = obligation_para.wrap(width - 4 * cm, 3 * cm)
            # Check if enough space; if not, draw on next line area
            if y_pos_after_table - oh < 2.5 * cm:
                # Not enough space, skip to footer area
                pass
            else:
                obligation_para.drawOn(c, 2 * cm, y_pos_after_table - oh)

            # ── Footer ──────────────────────────────────────────────────
            c.setStrokeColor(colors.HexColor("#D7DFEB"))
            c.line(1.5 * cm, 2.1 * cm, width - 1.5 * cm, 2.1 * cm)
            c.setFillColor(colors.HexColor("#4D5E73"))
            c.setFont(FONT, 8)
            c.drawCentredString(
                width / 2, 1.6 * cm,
                "Document generat automat de sistemul TraceabilityRS"
            )

            c.save()
            logger.info(
                f"Employee signature PDF generated: {file_path}"
            )
            return file_path

        except Exception as e:
            logger.error(
                f"Error generating employee signature PDF: {e}",
                exc_info=True
            )
            return None

    # ==================== EMAIL ====================

    
    def send_overtime_notification(self, request_number, supervisor_name, employees_count, pdf_path):
        """
        Invia email di notifica ai responsabili.
        
        Args:
            request_number: Numero richiesta
            supervisor_name: Nome supervisore
            employees_count: Numero dipendenti coinvolti
            pdf_path: Path del PDF da allegare
            
        Returns:
            bool: True se invio riuscito
        """
        try:
            # Recupera destinatari da Settings
            recipients_str = self.db.fetch_setting('overtime_request')
            if not recipients_str:
                logger.error("No recipients configured for overtime_request in Settings")
                return False
            
            # Converti stringa in lista (separati da virgola o punto e virgola)
            recipients = [r.strip() for r in recipients_str.replace(';', ',').split(',') if r.strip()]
            
            if not recipients:
                logger.error("No valid recipients found in overtime_request setting")
                return False
            
            # Prepara email in inglese con HTML
            subject = f"Overtime Request Pending Approval - {request_number}"
            
            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #2E5090;">Overtime Request Notification</h2>
                <p>Dear Manager,</p>
                <p>A new overtime request has been submitted and requires your approval.</p>
                <table style="border-collapse: collapse; margin: 20px 0;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 180px;">Request Number:</td>
                        <td style="padding: 8px;">{request_number}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Requested by:</td>
                        <td style="padding: 8px;">{supervisor_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Employees Involved:</td>
                        <td style="padding: 8px;">{employees_count}</td>
                    </tr>
                </table>
                <p style="background-color: #FFF3CD; border-left: 4px solid #FFC107; padding: 12px; margin: 20px 0;">
                    <strong>⚠️ Action Required:</strong> Please review and approve/reject this request via:<br>
                    <strong>ERP → Operations → Personnel → Overtime → Authorization</strong>
                </p>
                <p>The overtime agreement document signed by the employees is attached to this email.</p>
                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """
            
            # Importa funzione send_email da utils
            import sys
            import os
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import send_email
            
            # Invia email con allegato
            send_email(
                recipients=recipients,
                subject=subject,
                body=body,
                is_html=True,
                attachments=[pdf_path] if pdf_path and os.path.exists(pdf_path) else None
            )
            
            logger.info(f"Email sent successfully to: {', '.join(recipients)}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending email: {e}", exc_info=True)
            return False
    
    def send_approval_notification(self, request_id, request_number, requester_id, approved, approver_name, extra_attachments=None, attachment_path=None):
        """
        Invia email di notifica approvazione/rifiuto al richiedente.

        Args:
            request_id: ID richiesta
            request_number: Numero richiesta (non usato, recuperato da DB)
            requester_id: ID richiedente (EmployeeHireHistoryId)
            approved: True se approvata, False se rifiutata
            approver_name: Nome di chi ha approvato/rifiutato
            extra_attachments: Lista di path file da allegare (PDF + Excel, opzionale)
            attachment_path: (deprecato) singolo path allegato per compatibilità

        Returns:
            bool: True se invio riuscito
        """
        try:
            # Recupera email richiedente
            email_query = """
            SELECT DISTINCT e.employeename + ' '+ e.employeesurname as Employee,  
                    ea.WorkEmail AS Email
            FROM  Employee.dbo.EmployeeHireHistory h
            LEFT JOIN Employee.dbo.Employees e 
                ON e.EmployeeId = h.EmployeeId and h.employeerid =2 and h.EndWorkDate is null
            LEFT JOIN Employee.dbo.EmployeeAddress ea
                ON ea.EmployeeId = e.EmployeeId
               AND ea.DateOut IS NULL
            WHERE h.EmployeeHireHistoryId = ?
            """
            
            result = self.db.fetch_one(email_query, (requester_id,))
            
            if not result or not result[1]:
                logger.error(f"No email found for requester ID: {requester_id}")
                return False
            
            requester_name = result[0]
            requester_email = result[1]
            
            # Recupera numeri formattati da TbRegistro
            if approved:
                # Se approvato, recupera sia RequestNumber che ApprovalNumber
                numbers_query = """
                SELECT DISTINCT r.NumRegistro + ' on ' + FORMAT(r.datareg, 'd', 'ro-ro') AS RequestNumber,
                       r1.NumRegistro + ' on ' + FORMAT(r1.datareg, 'd', 'ro-ro') AS ApprovalNumber
                FROM ResetServices.dbo.TbRegistro r 
                INNER JOIN ResetServices.dbo.ExtraTimeApproval e ON r.Contatore = e.IdRegistro
                INNER JOIN ResetServices.dbo.ExtraTimeApprovalStory es ON es.ExtraHourApprovalId = e.ExtraHourApprovalId
                INNER JOIN ResetServices.dbo.TbRegistro r1 ON r1.Contatore = es.ApprovedId
                WHERE e.ExtraHourApprovalId = ?
                """
                numbers_result = self.db.fetch_one(numbers_query, (request_id,))
                formatted_request_number = numbers_result[0] if numbers_result else request_number
                formatted_approval_number = numbers_result[1] if numbers_result else "N/A"
            else:
                # Se rifiutato, recupera solo RequestNumber
                request_query = """
                SELECT r.NumRegistro + ' on ' + FORMAT(r.datareg, 'd', 'ro-ro') AS RequestNumber
                FROM ResetServices.dbo.TbRegistro r 
                INNER JOIN ResetServices.dbo.ExtraTimeApproval e ON r.Contatore = e.IdRegistro
                WHERE e.ExtraHourApprovalId = ?
                """
                request_result = self.db.fetch_one(request_query, (request_id,))
                formatted_request_number = request_result[0] if request_result else request_number
                formatted_approval_number = None
            
            # Prepara email
            status = "APPROVED" if approved else "REJECTED"
            status_color = "#28A745" if approved else "#DC3545"
            
            subject = f"Overtime Request {status} - {formatted_request_number}"
            
            # Costruisci tabella dettagli
            details_rows = f"""
                <tr>
                    <td style="padding: 8px; font-weight: bold; width: 180px;">Request Number:</td>
                    <td style="padding: 8px;">{formatted_request_number}</td>
                </tr>
            """
            
            if approved and formatted_approval_number:
                details_rows += f"""
                <tr>
                    <td style="padding: 8px; font-weight: bold;">Approval Number:</td>
                    <td style="padding: 8px;">{formatted_approval_number}</td>
                </tr>
                """
            
            details_rows += f"""
                <tr>
                    <td style="padding: 8px; font-weight: bold;">Status:</td>
                    <td style="padding: 8px; color: {status_color}; font-weight: bold;">{status}</td>
                </tr>
                <tr>
                    <td style="padding: 8px; font-weight: bold;">Reviewed by:</td>
                    <td style="padding: 8px;">{approver_name}</td>
                </tr>
            """
            
            # ── Sezione costi nell'email (solo se approvato) ──────────────
            cost_section = ""
            if approved:
                try:
                    # Recupera la data della richiesta (DateStart del primo story record)
                    date_row = self.db.fetch_one(
                        "SELECT TOP 1 CAST(DateStart AS DATE) FROM ResetServices.dbo.ExtraTimeApprovalStory "
                        "WHERE ExtraHourApprovalId = ? ORDER BY DateStart",
                        (request_id,)
                    )
                    req_date = date_row[0].strftime('%Y-%m-%d') if date_row and date_row[0] else None

                    cost_query_day = """
                    SELECT CAST(s.DateStart AS DATE) AS OverTimeDate,
                           SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 AS OvertimeHours,
                           SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 * x.ValueITem AS OvertimeCost,
                           x.Currency
                    FROM [ResetServices].[dbo].[ExtraTimeApprovalStory] s
                    INNER JOIN [ResetServices].[dbo].[ExtraTimeApproval] a
                        ON a.ExtraHourApprovalId = s.ExtraHourApprovalId
                    OUTER APPLY (
                        SELECT ot.ValueITem, t.[DESC] AS Currency
                        FROM [ResetServices].[dbo].[OverTimeDefaults] ot
                        INNER JOIN [ResetServices].[dbo].[OverTimeDescriptions] od
                            ON ot.DescriptionId = od.DescpriptionId
                        INNER JOIN ResetServices.dbo.TbValute t
                            ON t.IdValuta = ot.CurrencyId
                        WHERE ot.DateOut IS NULL AND od.DescpriptionId = 1
                    ) AS x
                    WHERE CAST(s.DateStart AS DATE) = ?
                    GROUP BY CAST(s.DateStart AS DATE), x.ValueITem, x.Currency
                    """
                    cost_monthly_query = """
                    SELECT MONTH(s.DateStart) AS Month, YEAR(s.DateStart) AS Year,
                           SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 AS OvertimeHours,
                           SUM(DATEDIFF(MINUTE, s.DateStart, s.DateEnd)) / 60.0 * x.ValueITem AS OvertimeCost,
                           x.Currency
                    FROM [ResetServices].[dbo].[ExtraTimeApprovalStory] s
                    INNER JOIN [ResetServices].[dbo].[ExtraTimeApproval] a
                        ON a.ExtraHourApprovalId = s.ExtraHourApprovalId
                    OUTER APPLY (
                        SELECT ot.ValueITem, t.[DESC] AS Currency
                        FROM [ResetServices].[dbo].[OverTimeDefaults] ot
                        INNER JOIN [ResetServices].[dbo].[OverTimeDescriptions] od
                            ON ot.DescriptionId = od.DescpriptionId
                        INNER JOIN ResetServices.dbo.TbValute t
                            ON t.IdValuta = ot.CurrencyId
                        WHERE ot.DateOut IS NULL AND od.DescpriptionId = 1
                    ) AS x
                    WHERE YEAR(s.DateStart) = YEAR(GETDATE())
                    GROUP BY MONTH(s.DateStart), YEAR(s.DateStart), x.ValueITem, x.Currency
                    ORDER BY YEAR(s.DateStart), MONTH(s.DateStart)
                    """

                    _MONTHS = ['Jan','Feb','Mar','Apr','May','Jun',
                               'Jul','Aug','Sep','Oct','Nov','Dec']

                    # Totale del giorno
                    day_result = self.db.fetch_one(cost_query_day, (req_date,)) if req_date else None
                    if day_result:
                        day_hours    = round(float(day_result[1]), 2) if day_result[1] else 0.0
                        day_cost     = round(float(day_result[2]), 2) if day_result[2] else 0.0
                        day_currency = day_result[3] or ''
                        cost_section += f"""
                        <h3 style="color:#1F3A5F; margin-top:24px;">&#128200; Request Cost Summary ({req_date})</h3>
                        <table style="border-collapse:collapse; margin:10px 0;">
                          <tr style="background:#1F3A5F; color:#fff;">
                            <th style="padding:8px 14px;">Date</th>
                            <th style="padding:8px 14px;">Total Hours</th>
                            <th style="padding:8px 14px;">Total Cost</th>
                            <th style="padding:8px 14px;">Currency</th>
                          </tr>
                          <tr>
                            <td style="padding:8px 14px;">{req_date}</td>
                            <td style="padding:8px 14px; text-align:center;">{day_hours}</td>
                            <td style="padding:8px 14px; text-align:right; font-weight:bold;">{day_cost:,.2f}</td>
                            <td style="padding:8px 14px;">{day_currency}</td>
                          </tr>
                        </table>
                        """

                    # Tabella mensile YTD
                    cur3 = self.db.conn.cursor()
                    cur3.execute(cost_monthly_query)
                    monthly_rows = cur3.fetchall()
                    cur3.close()
                    if monthly_rows:
                        ytd_currency = monthly_rows[0][4] if monthly_rows[0][4] else ''
                        ytd_rows_html = ""
                        ytd_total_cost = 0.0
                        ytd_total_hours = 0.0
                        for i, mr in enumerate(monthly_rows):
                            m_name = _MONTHS[int(mr[0]) - 1] if mr[0] else '?'
                            m_hours = round(float(mr[2]), 2) if mr[2] else 0.0
                            m_cost  = round(float(mr[3]), 2) if mr[3] else 0.0
                            ytd_total_hours += m_hours
                            ytd_total_cost  += m_cost
                            bg = ' background:#F0F4FA;' if i % 2 == 0 else ''
                            ytd_rows_html += f"""
                          <tr style="{bg}">
                            <td style="padding:6px 14px;">{m_name} {mr[1]}</td>
                            <td style="padding:6px 14px; text-align:center;">{m_hours}</td>
                            <td style="padding:6px 14px; text-align:right;">{m_cost:,.2f}</td>
                            <td style="padding:6px 14px;">{ytd_currency}</td>
                          </tr>"""
                        ytd_rows_html += f"""
                          <tr style="background:#D7DFEB; font-weight:bold;">
                            <td style="padding:6px 14px;">TOTAL YTD</td>
                            <td style="padding:6px 14px; text-align:center;">{round(ytd_total_hours,2)}</td>
                            <td style="padding:6px 14px; text-align:right;">{ytd_total_cost:,.2f}</td>
                            <td style="padding:6px 14px;">{ytd_currency}</td>
                          </tr>"""
                        cost_section += f"""
                        <h3 style="color:#1F3A5F; margin-top:24px;">&#128197; Monthly YTD Overtime Cost ({int(datetime.now().year)})</h3>
                        <table style="border-collapse:collapse; margin:10px 0;">
                          <tr style="background:#1F3A5F; color:#fff;">
                            <th style="padding:8px 14px;">Month</th>
                            <th style="padding:8px 14px;">Total Hours</th>
                            <th style="padding:8px 14px;">Total Cost</th>
                            <th style="padding:8px 14px;">Currency</th>
                          </tr>
                          {ytd_rows_html}
                        </table>
                        """
                except Exception as _cost_exc:
                    logger.warning(f"Could not build cost section for email: {_cost_exc}")

            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: {status_color};">Overtime Request {status}</h2>
                <p>Dear {requester_name},</p>
                <p>Your overtime request has been <strong style="color: {status_color};">{status}</strong>.</p>
                <table style="border-collapse: collapse; margin: 20px 0;">
                    {details_rows}
                </table>
                {cost_section}
                {'<div style="background-color: #FFF3CD; border-left: 4px solid #FFC107; padding: 14px; margin: 24px 0; font-family: Arial, sans-serif;">'
                 '<strong>&#9888; Aten&#539;ie / Attention:</strong><br>'
                 'Documentul ata&#537;at trebuie predat obligatoriu biroului de personal (HR) '
                 'cel t&#226;rziu &#238;n ziua anterioar&#259; desf&#259;&#537;ur&#259;rii orelor suplimentare, '
                 'de c&#259;tre responsabilul care a solicitat aprobarea.'
                 '</div>' if approved else ''}
                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """
            
            # Importa funzione send_email da utils
            import sys
            import os
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import send_email
            
            attachments = []
            # Supporto per extra_attachments (nuovo) e attachment_path (legacy)
            if extra_attachments:
                attachments.extend([p for p in extra_attachments if p and os.path.exists(p)])
            elif attachment_path and os.path.exists(attachment_path):
                attachments.append(attachment_path)
            if not attachments:
                attachments = None

            # Recupera indirizzi CC da Settings
            cc_emails = None
            try:
                cc_result = self.db.fetch_one(
                    "SELECT [Value] FROM traceability_rs.dbo.Settings WHERE Atribute = 'Sys_email_Overtami_responce'"
                )
                if cc_result and cc_result[0]:
                    cc_emails = [e.strip() for e in cc_result[0].replace(';', ',').split(',') if e.strip()]
                    logger.info(f"Approval notification CC: {cc_emails}")
                else:
                    logger.warning("Sys_email_Overtami_responce not configured in Settings – sending without CC")
            except Exception as _cc_exc:
                logger.warning(f"Could not read CC setting: {_cc_exc}")

            # Invia email: TO = richiedente, CC = indirizzi da Settings
            send_email(
                recipients=[requester_email],
                subject=subject,
                body=body,
                is_html=True,
                attachments=attachments,
                cc_emails=cc_emails
            )

            logger.info(f"Approval notification sent to: {requester_email} | CC: {cc_emails}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending approval notification: {e}", exc_info=True)
            return False

    def send_question_email(self, qa_id, request_id, requester_id, question_text, asker_name):
        """
        Invia email con domanda al richiedente dello straordinario.

        Args:
            qa_id: ID della domanda (ExtraTimeApprovalQA.QAId)
            request_id: ExtraHourApprovalId
            requester_id: idanga del richiedente (IdChief)
            question_text: Testo della domanda
            asker_name: Nome di chi chiede

        Returns:
            bool: True se invio riuscito
        """
        try:
            # Recupera email richiedente tramite tbuserkey -> Employee
            email_query = """
            SELECT DISTINCT e.employeename + ' ' + e.employeesurname AS Employee,
                   ea.WorkEmail AS Email
            FROM resetservices.dbo.tbuserkey u
            INNER JOIN Employee.dbo.EmployeeHireHistory h
                ON h.EmployeeHireHistoryId = u.employeehirehistoryid
                AND h.employeerid = 2 AND h.EndWorkDate IS NULL
            INNER JOIN Employee.dbo.Employees e
                ON e.EmployeeId = h.EmployeeId
            LEFT JOIN Employee.dbo.EmployeeAddress ea
                ON ea.EmployeeId = e.EmployeeId AND ea.DateOut IS NULL
            WHERE u.idanga = ?
            """
            result = self.db.fetch_one(email_query, (requester_id,))

            if not result or not result[1]:
                logger.error(f"No email found for requester idanga: {requester_id}")
                return False

            requester_name = result[0]
            requester_email = result[1]

            # Recupera RequestNumber da TbRegistro
            req_query = """
            SELECT ISNULL(r.NumRegistro, CAST(a.IdRegistro AS VARCHAR)) AS RequestNumber
            FROM ResetServices.dbo.ExtraTimeApproval a
            LEFT JOIN ResetServices.dbo.TbRegistro r ON a.IdRegistro = r.Contatore
            WHERE a.ExtraHourApprovalId = ?
            """
            req_result = self.db.fetch_one(req_query, (request_id,))
            request_number = req_result[0] if req_result else str(request_id)

            subject = f"Question about Overtime Request {request_number}"

            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #2E5090;">Overtime Request - Question</h2>
                <p>Dear {requester_name},</p>
                <p>A question has been raised regarding your overtime request before a decision can be made.</p>
                <table style="border-collapse: collapse; margin: 20px 0;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 180px;">Request Number:</td>
                        <td style="padding: 8px;">{request_number}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Asked by:</td>
                        <td style="padding: 8px;">{asker_name}</td>
                    </tr>
                </table>
                <div style="background-color: #E8F4FD; border-left: 4px solid #2E5090; padding: 14px; margin: 20px 0;">
                    <strong>Question:</strong><br>
                    <p style="white-space: pre-wrap;">{question_text}</p>
                </div>
                <p style="background-color: #FFF3CD; border-left: 4px solid #FFC107; padding: 12px; margin: 20px 0;">
                    <strong>&#9888; Action Required:</strong> Please reply to this question via:<br>
                    <strong>ERP &rarr; Operations &rarr; Personnel &rarr; Overtime &rarr; Responses</strong>
                </p>
                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """

            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import send_email

            send_email(
                recipients=[requester_email],
                subject=subject,
                body=body,
                is_html=True
            )

            logger.info(f"Question email sent to: {requester_email} for request {request_number}")
            return True

        except Exception as e:
            logger.error(f"Error sending question email: {e}", exc_info=True)
            return False

    def send_answer_email(self, qa_id, asker_id, answer_text, responder_name, request_number):
        """
        Invia email con risposta al mittente della domanda (approver).

        Args:
            qa_id: ID della domanda
            asker_id: idanga di chi ha posto la domanda
            answer_text: Testo della risposta
            responder_name: Nome di chi risponde
            request_number: Numero richiesta per riferimento

        Returns:
            bool: True se invio riuscito
        """
        try:
            # Recupera email dell'approver tramite tbuserkey -> Employee
            email_query = """
            SELECT DISTINCT e.employeename + ' ' + e.employeesurname AS Employee,
                   ea.WorkEmail AS Email
            FROM resetservices.dbo.tbuserkey u
            INNER JOIN Employee.dbo.EmployeeHireHistory h
                ON h.EmployeeHireHistoryId = u.employeehirehistoryid
                AND h.employeerid = 2 AND h.EndWorkDate IS NULL
            INNER JOIN Employee.dbo.Employees e
                ON e.EmployeeId = h.EmployeeId
            LEFT JOIN Employee.dbo.EmployeeAddress ea
                ON ea.EmployeeId = e.EmployeeId AND ea.DateOut IS NULL
            WHERE u.idanga = ?
            """
            result = self.db.fetch_one(email_query, (asker_id,))

            if not result or not result[1]:
                logger.error(f"No email found for asker idanga: {asker_id}")
                return False

            asker_name = result[0]
            asker_email = result[1]

            # Recupera la domanda originale
            q_query = """
            SELECT q.QuestionText
            FROM ResetServices.dbo.ExtraTimeApprovalQA q
            WHERE q.QAId = ?
            """
            q_result = self.db.fetch_one(q_query, (qa_id,))
            original_question = q_result[0] if q_result else "N/A"

            subject = f"Answer to Overtime Question - Request {request_number}"

            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #28A745;">Overtime Request - Answer Received</h2>
                <p>Dear {asker_name},</p>
                <p>Your question regarding overtime request <strong>{request_number}</strong> has been answered.</p>
                <div style="background-color: #F0F0F0; border-left: 4px solid #6C757D; padding: 14px; margin: 20px 0;">
                    <strong>Your Question:</strong><br>
                    <p style="white-space: pre-wrap;">{original_question}</p>
                </div>
                <div style="background-color: #D4EDDA; border-left: 4px solid #28A745; padding: 14px; margin: 20px 0;">
                    <strong>Answer from {responder_name}:</strong><br>
                    <p style="white-space: pre-wrap;">{answer_text}</p>
                </div>
                <p style="background-color: #FFF3CD; border-left: 4px solid #FFC107; padding: 12px; margin: 20px 0;">
                    <strong>&#9888; Next Steps:</strong> You can now proceed to approve or reject this request, or ask another question via:<br>
                    <strong>ERP &rarr; Operations &rarr; Personnel &rarr; Overtime &rarr; Authorization</strong>
                </p>
                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """

            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import send_email

            send_email(
                recipients=[asker_email],
                subject=subject,
                body=body,
                is_html=True
            )

            logger.info(f"Answer email sent to: {asker_email} for request {request_number}")
            return True

        except Exception as e:
            logger.error(f"Error sending answer email: {e}", exc_info=True)
            return False

    def send_weekly_overtime_analysis_email(self):
        """
        Invia email settimanale con analisi straordinari non autorizzati.
        Da chiamare ogni lunedì per il periodo della settimana precedente.
        
        Returns:
            bool: True se email inviata con successo
        """
        try:
            from datetime import timedelta
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Calcola date settimana precedente (lunedì-domenica)
            today = date.today()
            days_since_monday = today.weekday()  # 0 = Monday
            last_monday = today - timedelta(days=days_since_monday + 7)
            last_sunday = last_monday + timedelta(days=6)
            
            logger.info(f"Generating weekly overtime analysis for period: {last_monday} to {last_sunday}")
            
            # Query analisi con filtro OVER APPROVED
            query = """
            DECLARE @dateStart DATE = ?;
            DECLARE @DateStop DATE = ?;
            DECLARE @Filter AS NVARCHAR(30) = 'OVER APPROVED';

            WITH
            CTE_DailyState_Employee AS (
                SELECT 
                    ds.IDDailyState,
                    ds.DailyStateDate,
                    e.IDEmployee,
                    UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS Name, 
                    e.UniqueID
                FROM Timeclocking.dbo.DailyState ds
                INNER JOIN Timeclocking.dbo.Employee e
                    ON e.IDEmployee = ds.IDEmployee AND ds.DailyStateDate BETWEEN @dateStart AND @DateStop
            ),
            CTE_Done AS (
                SELECT
                    fd.IDDailyState,
                    fd.NoMin AS MinSuplimentarDone,
                    r.RequestName
                FROM Timeclocking.dbo.EmployeeRequestFractionalDay fd
                INNER JOIN Timeclocking.dbo.RequestType r
                    ON r.IDRequestType = fd.IDRequestType
                WHERE r.IDRequestType = 8
            ),
            CTE_HireHistory AS (
                SELECT
                    h.EmployeeHireHistoryId,
                    ee.EmployeeNID COLLATE DATABASE_DEFAULT AS UniqueID
                FROM employee.dbo.employees ee
                INNER JOIN employee.dbo.employeehirehistory h
                    ON ee.EmployeeId = h.EmployeeId
                    AND h.employeerid = 2
                    AND h.EndWorkDate IS NULL
            ),
            CTE_ExtraTimeApprovalStory AS (
                SELECT
                    es.IdEmployee AS EmployeeHireHistoryId,
                    CAST(es.DateStart AS DATE) AS DateStart,
                    ISNULL(DATEDIFF(MINUTE, es.DateStart, es.DateEnd), 0) AS MinExtraTimeApproved
                FROM [ResetServices].[dbo].ExtraTimeApprovalStory es
                WHERE CAST(es.DateStart AS DATE) BETWEEN @dateStart AND @dateStop
            ),
            CTE_Combined AS (
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY dse.DailyStateDate, dse.Name) AS Nr,
                    dse.Name,
                    dse.DailyStateDate AS OvertimeDate,
                    req.MinSuplimentarDone,
                    req.RequestName,
                    ISNULL(eta.MinExtraTimeApproved, 0) AS MinExtraTimeApproved,
                    CASE
                        WHEN ISNULL(req.MinSuplimentarDone, 0) > ISNULL(eta.MinExtraTimeApproved, 0) THEN
                            'OVER APPROVED'
                        ELSE
                            'Time approved = time presence'
                    END AS Notes
                FROM CTE_DailyState_Employee dse
                INNER JOIN CTE_Done req ON dse.IDDailyState = req.IDDailyState
                INNER JOIN CTE_HireHistory hh ON dse.UniqueID COLLATE DATABASE_DEFAULT = hh.UniqueID
                LEFT JOIN CTE_ExtraTimeApprovalStory eta ON hh.EmployeeHireId = eta.EmployeeHireHistoryId
                    AND eta.DateStart = dse.DailyStateDate
            )
            SELECT DISTINCT *
            FROM CTE_Combined
            WHERE Notes LIKE @Filter
            ORDER BY OvertimeDate, Name;
            """
            
            cursor = self.db.conn.cursor()
            cursor.execute(query, (last_monday, last_sunday))
            results = cursor.fetchall()
            cursor.close()
            
            if not results:
                logger.info("No unauthorized overtime found for last week. Email not sent.")
                return True
            
            # Crea Excel
            output_dir = tempfile.gettempdir()
            filename = f"Unauthorized_Overtime_{last_monday.strftime('%Y%m%d')}_{last_sunday.strftime('%Y%m%d')}.xlsx"
            file_path = os.path.join(output_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Unauthorized Overtime"
            
            # Header
            headers = ['Nr', 'Employee', 'Date', 'Min Presence', 'Min Approved', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Dati
            for row_idx, row in enumerate(results, 2):
                ws.cell(row=row_idx, column=1, value=row[0]).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=2, value=row[1])
                ws.cell(row=row_idx, column=3, value=row[2].strftime('%d/%m/%Y') if row[2] else 'N/D').alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=4, value=row[3]).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=5, value=row[5]).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=6, value=row[6])
                
                # Evidenzia in rosso
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                    )
            
            # Adatta larghezza colonne
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
            
            wb.save(file_path)
            
            # Recupera destinatari
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT AttributeValue 
                FROM Traceability_RS.dbo.Settings 
                WHERE AttributeName = 'Sys_email_overtime_issues'
            """)
            result = cursor.fetchone()
            cursor.close()
            
            if not result or not result[0]:
                logger.warning("No email recipients configured in Settings.Sys_email_overtime_issues")
                return False
            
            recipients = [email.strip() for email in result[0].split(',')]
            
            # Prepara email
            subject = f"⚠️ Unauthorized Overtime Report - Week {last_monday.strftime('%d/%m/%Y')} to {last_sunday.strftime('%d/%m/%Y')}"
            
            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #C00000;">⚠️ Unauthorized Overtime Alert</h2>
                <p>Dear Team,</p>
                <p>The attached report summarizes <strong style="color: #C00000;">{len(results)} unauthorized overtime entries</strong> 
                for the period <strong>{last_monday.strftime('%d/%m/%Y')} to {last_sunday.strftime('%d/%m/%Y')}</strong>.</p>
                
                <p>These entries indicate cases where the actual overtime presence exceeded the approved overtime hours.</p>
                
                <p style="margin-top: 20px;">
                    <strong>Action Required:</strong><br>
                    Please review the attached Excel file and take appropriate action.
                </p>
                
                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """
            
            # Importa funzione send_email
            import sys
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import send_email
            
            # Invia email con allegato
            send_email(
                recipients=recipients,
                subject=subject,
                body=body,
                is_html=True,
                attachments=[file_path]
            )
            
            logger.info(f"Weekly overtime analysis email sent to: {recipients}")
            
            # Rimuovi file temporaneo
            try:
                os.remove(file_path)
            except:
                pass
            
            return True
            
        except Exception as e:
            logger.error(f"Error sending weekly overtime analysis email: {e}", exc_info=True)
            return False

    def send_weekly_unauthorized_overtime_email(self):
        """
        Invia email settimanale (ogni lunedì) con i dipendenti (FunctionCode <= 60) che hanno:
        - Timbrature durante il weekend (sabato/domenica)
        - Ore lavorate > 8 durante i giorni feriali
        Evidenzia quelli NON autorizzati.

        Destinatari da: settings.atribute = 'Sys_email_overtimeNotAuth'
        Include Logo.png e allegato Excel.

        Returns:
            bool: True se email inviata con successo
        """
        try:
            from datetime import timedelta
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import base64

            # Calcola date settimana precedente (lunedì-domenica)
            today = date.today()
            days_since_monday = today.weekday()  # 0 = Monday
            last_monday = today - timedelta(days=days_since_monday + 7)
            last_sunday = last_monday + timedelta(days=6)

            logger.info(f"Generating weekly unauthorized overtime report for: {last_monday} to {last_sunday}")

            # === DEDUP: verifica se email già inviata per questa settimana ===
            dedup_attribute = 'Sys_email_overtimeNotAuth'
            try:
                dedup_cursor = self.db.conn.cursor()
                dedup_cursor.execute("""
                    SELECT TOP 1 1 FROM [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog]
                    WHERE WeekStartDate = ? AND Attribute = ?
                """, (last_monday, dedup_attribute))
                already_sent = dedup_cursor.fetchone() is not None
                dedup_cursor.close()

                if already_sent:
                    logger.info(f"Weekly unauthorized overtime email already sent for week {last_monday}. Skipping.")
                    return True
            except Exception as dedup_err:
                logger.warning(f"Dedup check failed (fail-safe: skip send): {dedup_err}")
                return True  # Fail-safe: evita duplicati

            # Query: dipendenti con FunctionCode <= 60, timbrature weekend o > 8h feriali
            query = """
            DECLARE @dateStart DATE = ?;
            DECLARE @DateStop DATE = ?;

            WITH
            CTE_DailyState_Employee AS (
                SELECT 
                    ds.IDDailyState,
                    ds.DailyStateDate,
                    e.IDEmployee,
                    UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS Name, 
                    e.UniqueID,
                    DATEPART(WEEKDAY, ds.DailyStateDate) AS DayOfWeek  -- 1=Sun, 7=Sat
                FROM Timeclocking.dbo.DailyState ds
                INNER JOIN Timeclocking.dbo.Employee e
                    ON e.IDEmployee = ds.IDEmployee 
                    AND ds.DailyStateDate BETWEEN @dateStart AND @DateStop
            ),
            CTE_Presence AS (
                SELECT
                    fd.IDDailyState,
                    fd.NoMin AS MinWorked,
                    r.RequestName,
                    r.IDRequestType
                FROM Timeclocking.dbo.EmployeeRequestFractionalDay fd
                INNER JOIN Timeclocking.dbo.RequestType r
                    ON r.IDRequestType = fd.IDRequestType
                WHERE r.IDRequestType = 8
            ),
            CTE_HireHistory AS (
                SELECT
                    h.EmployeeHireHistoryId AS EmployeeHireId,
                    ee.EmployeeNID COLLATE DATABASE_DEFAULT AS UniqueID,
                    ISNULL(f.FunctionCode, 0) AS FunctionCode,
                    ISNULL(f.FunctionName, 'N/A') AS FunctionName
                FROM employee.dbo.employees ee
                INNER JOIN employee.dbo.employeehirehistory h
                    ON ee.EmployeeId = h.EmployeeId
                    AND h.employeerid = 2
                    AND h.EndWorkDate IS NULL
                LEFT JOIN Employee.dbo.EmployeeCdcStories cs
                    ON cs.EmployeeHireHistoryId = h.EmployeeHireHistoryId
                    AND cs.DateOut IS NULL
                LEFT JOIN Employee.dbo.Functions f
                    ON cs.FunctionId = f.FunctionId
                WHERE ISNULL(f.FunctionCode, 0) <= 60
            ),
            CTE_ExtraTimeApprovalStory AS (
                SELECT
                    es.IdEmployee AS EmployeeHireHistoryId,
                    CAST(es.DateStart AS DATE) AS DateStart,
                    ISNULL(DATEDIFF(MINUTE, es.DateStart, es.DateEnd), 0) AS MinExtraTimeApproved
                FROM [ResetServices].[dbo].ExtraTimeApprovalStory es
                WHERE CAST(es.DateStart AS DATE) BETWEEN @dateStart AND @DateStop
            ),
            CTE_Combined AS (
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY dse.DailyStateDate, dse.Name) AS Nr,
                    dse.Name,
                    dse.DailyStateDate AS OvertimeDate,
                    dse.DayOfWeek,
                    pres.MinWorked,
                    ISNULL(eta.MinExtraTimeApproved, 0) AS MinApproved,
                    hh.FunctionCode,
                    hh.FunctionName,
                    CASE
                        WHEN dse.DayOfWeek IN (1, 7) THEN 'WEEKEND'
                        ELSE 'WEEKDAY > 8H'
                    END AS OvertimeType,
                    CASE
                        WHEN ISNULL(eta.MinExtraTimeApproved, 0) > 0 THEN 'AUTHORIZED'
                        ELSE 'NOT AUTHORIZED'
                    END AS AuthorizationStatus
                FROM CTE_DailyState_Employee dse
                INNER JOIN CTE_Presence pres ON dse.IDDailyState = pres.IDDailyState
                INNER JOIN CTE_HireHistory hh ON dse.UniqueID COLLATE DATABASE_DEFAULT = hh.UniqueID
                LEFT JOIN CTE_ExtraTimeApprovalStory eta ON hh.EmployeeHireId = eta.EmployeeHireHistoryId
                    AND eta.DateStart = dse.DailyStateDate
                WHERE 
                    dse.DayOfWeek IN (1, 7)                     -- Weekend (Sun=1, Sat=7)
                    OR (dse.DayOfWeek NOT IN (1, 7) AND pres.MinWorked > 480)  -- Weekday > 8h (480 min)
            )
            SELECT DISTINCT *
            FROM CTE_Combined
            ORDER BY OvertimeDate, Name;
            """

            cursor = self.db.conn.cursor()
            cursor.execute(query, (last_monday, last_sunday))
            results = cursor.fetchall()
            cursor.close()

            if not results:
                logger.info("No weekend/over-8h overtime found for last week. Email not sent.")
                return True

            # Conta autorizzati / non autorizzati
            total = len(results)
            not_auth_count = sum(1 for r in results if r[9] == 'NOT AUTHORIZED')
            auth_count = total - not_auth_count

            # === CREA EXCEL ===
            output_dir = tempfile.gettempdir()
            filename = f"Overtime_Report_{last_monday.strftime('%Y%m%d')}_{last_sunday.strftime('%Y%m%d')}.xlsx"
            file_path = os.path.join(output_dir, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Overtime Report"

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title row
            ws.merge_cells('A1:H1')
            title_cell = ws.cell(row=1, column=1, value="OVERTIME REPORT - Weekend & Over 8h Weekday Presence")
            title_cell.font = Font(bold=True, size=14, color="2E5090")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Period row
            ws.merge_cells('A2:H2')
            period_cell = ws.cell(row=2, column=1,
                value=f"Period: {last_monday.strftime('%d/%m/%Y')} to {last_sunday.strftime('%d/%m/%Y')}  |  "
                      f"Total: {total}  |  Authorized: {auth_count}  |  NOT Authorized: {not_auth_count}")
            period_cell.font = Font(size=10, italic=True)
            period_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 20

            # Headers row 3
            headers = ['Nr', 'Employee', 'Date', 'Day', 'Min Worked', 'Min Approved',
                        'Overtime Type', 'Authorization']
            header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            ws.row_dimensions[3].height = 22

            # Data rows
            day_names = {1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday',
                         5: 'Thursday', 6: 'Friday', 7: 'Saturday'}
            not_auth_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            auth_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            weekend_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            for row_idx, row in enumerate(results, 4):
                nr = row[0]
                name = row[1]
                ot_date = row[2].strftime('%d/%m/%Y') if row[2] else 'N/D'
                day_of_week = day_names.get(row[3], '?')
                min_worked = row[4] if row[4] else 0
                min_approved = row[5] if row[5] else 0
                ot_type = row[8]
                auth_status = row[9]

                is_not_auth = auth_status == 'NOT AUTHORIZED'
                is_weekend = ot_type == 'WEEKEND'

                ws.cell(row=row_idx, column=1, value=nr).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=ot_date).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=4, value=day_of_week).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=5, value=min_worked).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=6, value=min_approved).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=7, value=ot_type).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=8, value=auth_status).alignment = Alignment(horizontal='center')

                # Formattazione condizionale
                row_fill = None
                if is_not_auth:
                    row_fill = not_auth_fill
                elif is_weekend:
                    row_fill = weekend_fill
                else:
                    row_fill = auth_fill

                for c in range(1, 9):
                    ws.cell(row=row_idx, column=c).border = thin_border
                    if row_fill:
                        ws.cell(row=row_idx, column=c).fill = row_fill

                # Bold e rosso per NOT AUTHORIZED
                if is_not_auth:
                    ws.cell(row=row_idx, column=8).font = Font(bold=True, color="C00000")

            # Adatta larghezza colonne
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 18

            wb.save(file_path)
            logger.info(f"Excel report saved: {file_path}")

            # === RECUPERA DESTINATARI ===
            import sys
            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from utils import get_email_recipients, send_email

            recipients = get_email_recipients(self.db.conn, attribute='Sys_email_overtimeNotAuth')

            if not recipients:
                logger.warning("No email recipients configured for 'Sys_email_overtimeNotAuth'")
                return False

            # === LOGO BASE64 ===
            logo_html = ""
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Logo.png")
            if os.path.exists(logo_path):
                try:
                    with open(logo_path, "rb") as f:
                        logo_data = base64.b64encode(f.read()).decode("utf-8")
                    logo_html = f'<img src="data:image/png;base64,{logo_data}" style="height: 50px; margin-bottom: 10px;" /><br>'
                except Exception as logo_err:
                    logger.warning(f"Cannot embed logo: {logo_err}")

            # === TABELLA RIASSUNTIVA IN EMAIL ===
            table_rows = ""
            for row in results:
                ot_date = row[2].strftime('%d/%m/%Y') if row[2] else 'N/D'
                day_name = day_names.get(row[3], '?')
                min_w = row[4] if row[4] else 0
                min_a = row[5] if row[5] else 0
                ot_type = row[8]
                auth = row[9]

                bg_color = "#FFE0E0" if auth == "NOT AUTHORIZED" else ("#FFF3CD" if ot_type == "WEEKEND" else "#D4EDDA")
                font_weight = "bold" if auth == "NOT AUTHORIZED" else "normal"
                font_color = "#C00000" if auth == "NOT AUTHORIZED" else "#333333"

                table_rows += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{row[0]}</td>
                    <td style="padding: 6px; border: 1px solid #ddd;">{row[1]}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{ot_date}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{day_name}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{min_w}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{min_a}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center;">{ot_type}</td>
                    <td style="padding: 6px; border: 1px solid #ddd; text-align: center; font-weight: {font_weight}; color: {font_color};">{auth}</td>
                </tr>
                """

            subject = f"⚠️ Weekly Overtime Report - {last_monday.strftime('%d/%m/%Y')} to {last_sunday.strftime('%d/%m/%Y')}"

            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                {logo_html}
                <h2 style="color: #2E5090;">Weekly Overtime Report</h2>
                <p>Dear Team,</p>
                <p>Below is the overtime report for the week <strong>{last_monday.strftime('%d/%m/%Y')}</strong> to
                <strong>{last_sunday.strftime('%d/%m/%Y')}</strong>.</p>

                <p>This report includes employees (FunctionCode &le; 60) who either worked on weekends
                or exceeded 8 hours on weekdays.</p>

                <table style="margin: 10px 0; font-size: 13px;">
                    <tr>
                        <td style="padding: 4px 12px;"><strong>Total entries:</strong></td>
                        <td style="padding: 4px 12px;"><strong>{total}</strong></td>
                    </tr>
                    <tr>
                        <td style="padding: 4px 12px;">✅ Authorized:</td>
                        <td style="padding: 4px 12px; color: #28A745;"><strong>{auth_count}</strong></td>
                    </tr>
                    <tr>
                        <td style="padding: 4px 12px;">❌ NOT Authorized:</td>
                        <td style="padding: 4px 12px; color: #C00000;"><strong>{not_auth_count}</strong></td>
                    </tr>
                </table>

                <table style="border-collapse: collapse; width: 100%; margin: 20px 0; font-size: 12px;">
                    <thead>
                        <tr style="background-color: #2E5090; color: white;">
                            <th style="padding: 8px; border: 1px solid #ddd;">Nr</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Employee</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Date</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Day</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Min Worked</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Min Approved</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Type</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Authorization</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>

                <p style="background-color: #FFF3CD; border-left: 4px solid #FFC107; padding: 12px; margin: 20px 0;">
                    <strong>&#9888; Legend:</strong><br>
                    <span style="background-color: #FFE0E0; padding: 2px 8px;">Red</span> = NOT Authorized &nbsp;|&nbsp;
                    <span style="background-color: #FFF3CD; padding: 2px 8px;">Yellow</span> = Weekend (Authorized) &nbsp;|&nbsp;
                    <span style="background-color: #D4EDDA; padding: 2px 8px;">Green</span> = Weekday &gt; 8h (Authorized)
                </p>

                <p>The detailed Excel report is attached to this email.</p>

                <p style="margin-top: 30px;">
                    Best regards,<br>
                    <strong>TraceabilityRS System</strong>
                </p>
            </body>
            </html>
            """

            # Invia email con allegato
            send_email(
                recipients=recipients,
                subject=subject,
                body=body,
                is_html=True,
                attachments=[file_path]
            )

            logger.info(f"Weekly unauthorized overtime email sent to: {recipients}")

            # === DEDUP: registra invio riuscito ===
            try:
                log_cursor = self.db.conn.cursor()
                log_cursor.execute("""
                    INSERT INTO [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog]
                    (WeekStartDate, Attribute)
                    VALUES (?, ?)
                """, (last_monday, dedup_attribute))
                self.db.conn.commit()
                log_cursor.close()
                logger.info(f"Dedup log saved for week {last_monday}")
            except Exception as log_err:
                logger.warning(f"Failed to write dedup log: {log_err}")

            # Rimuovi file temporaneo
            try:
                os.remove(file_path)
            except:
                pass

            return True

        except Exception as e:
            logger.error(f"Error sending weekly unauthorized overtime email: {e}", exc_info=True)
            return False
