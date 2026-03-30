"""
Overtime Monitoring GUI
Dashboard monitoraggio 48h/4 mesi con raccomandazione automatica e export Excel.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date
import logging
import os
import tempfile

logger = logging.getLogger(__name__)


def open_overtime_monitoring_window(parent, db_handler, lang_manager, user_name, user_id=0):
    """
    Entry-point richiamabile da main.py.

    Args:
        parent: Finestra parent
        db_handler: DatabaseHandler instance
        lang_manager: LanguageManager instance
        user_name: Nome utente loggato
        user_id: ID utente (EmployeeHireHistoryId)
    """
    OvertimeMonitoringWindow(parent, db_handler, lang_manager, user_name, user_id)


class OvertimeMonitoringWindow(tk.Toplevel):
    """Dashboard monitoraggio straordinari 48h/4 mesi."""

    def __init__(self, parent, db_handler, lang_manager, user_name, user_id=0):
        super().__init__(parent)

        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_name
        self.user_id = user_id
        self.dashboard_data = []
        self.sort_column = None
        self.sort_reverse = False

        from overtime.overtime_manager import OvertimeManager
        self.manager = OvertimeManager(db_handler)

        # Setup finestra
        self.title(self.lang.get('overtime_monitoring_title', 'Monitoraggio Ore Supplementari (48h / 4 Mesi)'))
        self.geometry("1400x700")
        self.transient(parent)
        self.grab_set()

        self._create_widgets()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _create_widgets(self):
        """Crea i widget dell'interfaccia."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # === HEADER ===
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(
            header_frame,
            text=self.lang.get('overtime_monitoring_header', '📊 Monitoraggio Ore Supplementari'),
            font=("Segoe UI", 14, "bold")
        ).pack(side=tk.LEFT)

        # === FILTRI ===
        filter_frame = ttk.LabelFrame(
            main_frame,
            text=self.lang.get('overtime_mon_filters', 'Filtri'),
            padding="10"
        )
        filter_frame.pack(fill=tk.X, pady=5)

        # Mese
        ttk.Label(filter_frame, text=self.lang.get('month', 'Mese:')).grid(row=0, column=0, padx=5, sticky=tk.W)
        self.month_var = tk.IntVar(value=datetime.now().month)
        month_spin = ttk.Spinbox(filter_frame, from_=1, to=12, textvariable=self.month_var, width=5)
        month_spin.grid(row=0, column=1, padx=5, sticky=tk.W)

        # Anno
        ttk.Label(filter_frame, text=self.lang.get('year', 'Anno:')).grid(row=0, column=2, padx=5, sticky=tk.W)
        self.year_var = tk.IntVar(value=datetime.now().year)
        year_spin = ttk.Spinbox(filter_frame, from_=2020, to=2100, textvariable=self.year_var, width=8)
        year_spin.grid(row=0, column=3, padx=5, sticky=tk.W)

        # Filtro stato
        ttk.Label(
            filter_frame,
            text=self.lang.get('overtime_mon_filter_status', 'Stato:')
        ).grid(row=0, column=4, padx=5, sticky=tk.W)
        self.status_filter_var = tk.StringVar(value='ALL')
        status_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.status_filter_var,
            values=['ALL', 'CRITICAL', 'WARNING', 'OK'],
            state='readonly',
            width=12
        )
        status_combo.grid(row=0, column=5, padx=5, sticky=tk.W)

        # Pulsante genera
        ttk.Button(
            filter_frame,
            text=self.lang.get('overtime_mon_generate', '🔍 Genera Analisi'),
            command=self._generate_dashboard
        ).grid(row=0, column=6, padx=15, pady=5)

        # === SOGLIE INFO ===
        self.thresholds_label_var = tk.StringVar(value="")
        ttk.Label(
            filter_frame,
            textvariable=self.thresholds_label_var,
            foreground="gray",
            font=("Segoe UI", 8)
        ).grid(row=1, column=0, columnspan=7, sticky=tk.W, pady=(5, 0))

        # === RISULTATI ===
        results_frame = ttk.LabelFrame(
            main_frame,
            text=self.lang.get('overtime_mon_results', 'Risultati'),
            padding="5"
        )
        results_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Treeview
        columns = ('name', 'monthly_hrs', 'avg_4m', 'weekly_hrs', 'weekend', 'status', 'decision', 'reason')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=20)

        headers = {
            'name': self.lang.get('employee', 'Dipendente'),
            'monthly_hrs': self.lang.get('overtime_mon_monthly', 'Ore/Mese'),
            'avg_4m': self.lang.get('overtime_mon_avg4m', 'Media 4M (sett.)'),
            'weekly_hrs': self.lang.get('overtime_mon_weekly', 'Ore Settimana'),
            'weekend': self.lang.get('overtime_mon_weekend', 'Weekend'),
            'status': self.lang.get('overtime_mon_status', 'Stato'),
            'decision': self.lang.get('overtime_mon_decision', 'Decisione'),
            'reason': self.lang.get('overtime_mon_reason', 'Motivazione'),
        }

        for col, text in headers.items():
            self.tree.heading(col, text=text, command=lambda c=col: self._sort_by_column(c))

        self.tree.column('name', width=220, anchor=tk.W)
        self.tree.column('monthly_hrs', width=90, anchor=tk.CENTER)
        self.tree.column('avg_4m', width=120, anchor=tk.CENTER)
        self.tree.column('weekly_hrs', width=110, anchor=tk.CENTER)
        self.tree.column('weekend', width=70, anchor=tk.CENTER)
        self.tree.column('status', width=90, anchor=tk.CENTER)
        self.tree.column('decision', width=130, anchor=tk.CENTER)
        self.tree.column('reason', width=350, anchor=tk.W)

        scrollbar_v = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_h = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_v.grid(row=0, column=1, sticky='ns')
        scrollbar_h.grid(row=1, column=0, sticky='ew')

        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)

        # === CONTATORI ===
        counter_frame = ttk.Frame(main_frame)
        counter_frame.pack(fill=tk.X, pady=(5, 0))

        self.counter_label_var = tk.StringVar(value="")
        ttk.Label(
            counter_frame,
            textvariable=self.counter_label_var,
            font=("Segoe UI", 9)
        ).pack(side=tk.LEFT)

        # === PULSANTI ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            btn_frame,
            text=self.lang.get('overtime_mon_export_excel', '📥 Esporta Excel'),
            command=self._export_to_excel
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            btn_frame,
            text=self.lang.get('close', 'Chiudi'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)

        # Configura tag colori
        self.tree.tag_configure('critical', background='#FFD6D6')
        self.tree.tag_configure('warning', background='#FFF3CD')
        self.tree.tag_configure('ok', background='#D4EDDA')

    def _generate_dashboard(self):
        """Genera l'analisi dashboard."""
        # Pulisci
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.dashboard_data = []

        month = self.month_var.get()
        year = self.year_var.get()
        status_filter = self.status_filter_var.get()

        # Mostra soglie
        try:
            thresholds = self.manager.get_monitoring_thresholds()
            self.thresholds_label_var.set(
                f"Soglie: Limite settim. = {thresholds['weekly_limit']}h | "
                f"Warning = {thresholds['warning_threshold']}h | "
                f"Critical = {thresholds['critical_threshold']}h | "
                f"Periodo = {thresholds['monitoring_months']} mesi | "
                f"Max giorn. = {thresholds['max_daily_hours']}h"
            )
        except Exception:
            pass

        # Cursore attesa
        self.config(cursor="wait")
        self.update_idletasks()

        try:
            data = self.manager.get_monitoring_dashboard_data(
                manager_hire_history_id=self.user_id if self.user_id else None,
                month=month,
                year=year
            )

            counts = {'OK': 0, 'WARNING': 0, 'CRITICAL': 0}

            for row in data:
                status = row['status']
                # Filtra per stato
                if status_filter != 'ALL' and status != status_filter:
                    continue

                weekend_text = '✅' if row['has_weekend'] else '—'
                status_text = {
                    'OK': '🟢 OK',
                    'WARNING': '🟡 WARNING',
                    'CRITICAL': '🔴 CRITICAL'
                }.get(status, status)

                decision_text = {
                    'SUPPLEMENTARI': '📋 Supplementari',
                    'PREMI': '💰 Premi',
                    'SPLIT': '🔀 Split',
                }.get(row['decision'], row['decision'])

                tag = status.lower()
                counts[status] = counts.get(status, 0) + 1

                self.tree.insert('', tk.END, values=(
                    row['name'],
                    f"{row['monthly_hours']:.1f}",
                    f"{row['avg_4months']:.1f}",
                    f"{row['weekly_hours']:.1f}",
                    weekend_text,
                    status_text,
                    decision_text,
                    row['reason']
                ), tags=(tag,))

                self.dashboard_data.append(row)

            total = len(self.dashboard_data)
            self.counter_label_var.set(
                f"Totale: {total}  |  "
                f"🔴 Critical: {counts.get('CRITICAL', 0)}  |  "
                f"🟡 Warning: {counts.get('WARNING', 0)}  |  "
                f"🟢 OK: {counts.get('OK', 0)}"
            )

            if total == 0:
                messagebox.showinfo(
                    self.lang.get('info', 'Info'),
                    self.lang.get('overtime_mon_no_data', 'Nessun dipendente trovato con i filtri selezionati.'),
                    parent=self
                )

        except Exception as e:
            logger.error(f"Errore generazione dashboard monitoraggio: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore generazione analisi:\n{str(e)}",
                parent=self
            )
        finally:
            self.config(cursor="")

    def _sort_by_column(self, col):
        """Ordina la tabella per colonna."""
        if not self.dashboard_data:
            return

        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        col_map = {
            'name': 'name',
            'monthly_hrs': 'monthly_hours',
            'avg_4m': 'avg_4months',
            'weekly_hrs': 'weekly_hours',
            'weekend': 'has_weekend',
            'status': 'status',
            'decision': 'decision',
            'reason': 'reason',
        }
        sort_key = col_map.get(col, 'name')

        status_order = {'CRITICAL': 0, 'WARNING': 1, 'OK': 2}

        def sort_func(item):
            val = item.get(sort_key, '')
            if sort_key == 'status':
                return status_order.get(val, 3)
            if isinstance(val, (int, float)):
                return val
            if isinstance(val, bool):
                return 0 if val else 1
            return str(val).lower()

        self.dashboard_data.sort(key=sort_func, reverse=self.sort_reverse)

        # Rigenera treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in self.dashboard_data:
            status = row['status']
            weekend_text = '✅' if row['has_weekend'] else '—'
            status_text = {'OK': '🟢 OK', 'WARNING': '🟡 WARNING', 'CRITICAL': '🔴 CRITICAL'}.get(status, status)
            decision_text = {'SUPPLEMENTARI': '📋 Supplementari', 'PREMI': '💰 Premi', 'SPLIT': '🔀 Split'}.get(row['decision'], row['decision'])
            tag = status.lower()

            self.tree.insert('', tk.END, values=(
                row['name'],
                f"{row['monthly_hours']:.1f}",
                f"{row['avg_4months']:.1f}",
                f"{row['weekly_hours']:.1f}",
                weekend_text,
                status_text,
                decision_text,
                row['reason']
            ), tags=(tag,))

    def _export_to_excel(self):
        """Esporta i risultati in Excel con situazione reale + soluzione proposta nello stesso tab."""
        if not self.dashboard_data:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare. Generare prima l\'analisi.'),
                parent=self
            )
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Directory output
            output_dir = r"C:\Temp"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            month = self.month_var.get()
            year = self.year_var.get()
            filename = f"OvertimeMonitoring_{year}_{month:02d}_{datetime.now().strftime('%H%M%S')}.xlsx"
            file_path = os.path.join(output_dir, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Monitoraggio Straordinari"

            # Stili
            header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            section_font = Font(bold=True, color="FFFFFF", size=10)
            critical_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            ok_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="A9B9CF"),
                right=Side(style="thin", color="A9B9CF"),
                top=Side(style="thin", color="A9B9CF"),
                bottom=Side(style="thin", color="A9B9CF")
            )

            # ── Titolo ──
            ws.merge_cells('A1:J1')
            title_cell = ws.cell(row=1, column=1,
                value=f"MONITORAGGIO ORE SUPPLEMENTARI — {month:02d}/{year}")
            title_cell.font = Font(bold=True, size=14, color="1F3A5F")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # ── Soglie info ──
            try:
                thresholds = self.manager.get_monitoring_thresholds()
                ws.merge_cells('A2:J2')
                info_cell = ws.cell(row=2, column=1,
                    value=f"Limite settimanale: {thresholds['weekly_limit']}h  |  "
                          f"Warning: {thresholds['warning_threshold']}h  |  "
                          f"Critical: {thresholds['critical_threshold']}h  |  "
                          f"Periodo monitoraggio: {thresholds['monitoring_months']} mesi  |  "
                          f"Max giornaliero: {thresholds['max_daily_hours']}h")
                info_cell.font = Font(size=9, italic=True)
                info_cell.alignment = Alignment(horizontal="center")
            except Exception:
                pass

            # ── SEZIONE: Situazione Reale ──
            ws.merge_cells('A4:F4')
            sec1 = ws.cell(row=4, column=1, value="SITUAZIONE REALE")
            sec1.font = section_font
            sec1.fill = section_fill
            sec1.alignment = center_align
            for c in range(2, 7):
                ws.cell(row=4, column=c).fill = section_fill

            # ── SEZIONE: Soluzione Proposta ──
            ws.merge_cells('G4:J4')
            sec2 = ws.cell(row=4, column=7, value="SOLUZIONE PROPOSTA")
            sec2.font = section_font
            sec2.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            sec2.alignment = center_align
            for c in range(8, 11):
                ws.cell(row=4, column=c).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

            ws.row_dimensions[4].height = 22

            # ── Intestazioni ──
            headers = [
                ('Dipendente', 30),
                ('Ore/Mese', 12),
                ('Media 4M (sett.)', 16),
                ('Ore Settimana', 14),
                ('Weekend', 10),
                ('Stato', 14),
                ('Decisione', 18),
                ('Motivazione', 50),
                ('Ore Suppl.', 12),
                ('Ore Premio', 12),
            ]

            for col_idx, (header, width) in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else ''].width = width

            # Correzione larghezza colonne per lettere oltre I
            from openpyxl.utils import get_column_letter
            for col_idx, (_, width) in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[5].height = 22

            # ── Dati ──
            for row_idx, row_data in enumerate(self.dashboard_data, start=6):
                status = row_data['status']
                decision = row_data['decision']

                # Calcola ore per colonne "soluzione"
                supplementari_hours = 0
                premio_hours = 0
                if decision == 'SUPPLEMENTARI' or decision == 'SPLIT':
                    supplementari_hours = row_data['monthly_hours']
                elif decision == 'PREMI':
                    premio_hours = row_data['monthly_hours']

                row_fill = {'CRITICAL': critical_fill, 'WARNING': warning_fill, 'OK': ok_fill}.get(status)

                values = [
                    (row_data['name'], left_align),
                    (round(row_data['monthly_hours'], 1), center_align),
                    (round(row_data['avg_4months'], 1), center_align),
                    (round(row_data['weekly_hours'], 1), center_align),
                    ('Sì' if row_data['has_weekend'] else 'No', center_align),
                    (status, center_align),
                    (decision, center_align),
                    (row_data['reason'], left_align),
                    (round(supplementari_hours, 1), center_align),
                    (round(premio_hours, 1), center_align),
                ]

                for col_idx, (val, align) in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = align
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill

                # Bold + colore per status
                status_cell = ws.cell(row=row_idx, column=6)
                if status == 'CRITICAL':
                    status_cell.font = Font(bold=True, color="C00000")
                elif status == 'WARNING':
                    status_cell.font = Font(bold=True, color="856404")

            # ── Riga totale ──
            total_row = len(self.dashboard_data) + 6
            total_font = Font(bold=True, size=10)
            summary_fill = PatternFill(start_color="D7DFEB", end_color="D7DFEB", fill_type="solid")

            for col in range(1, 11):
                c = ws.cell(row=total_row, column=col)
                c.border = thin_border
                c.font = total_font
                c.fill = summary_fill
                c.alignment = center_align

            ws.cell(row=total_row, column=1, value="TOTALE").alignment = left_align
            ws.cell(row=total_row, column=2, value=f"=SUM(B6:B{total_row - 1})")
            ws.cell(row=total_row, column=9, value=f"=SUM(I6:I{total_row - 1})")
            ws.cell(row=total_row, column=10, value=f"=SUM(J6:J{total_row - 1})")

            # Conteggi per stato
            crit_count = sum(1 for d in self.dashboard_data if d['status'] == 'CRITICAL')
            warn_count = sum(1 for d in self.dashboard_data if d['status'] == 'WARNING')
            ok_count = sum(1 for d in self.dashboard_data if d['status'] == 'OK')
            ws.cell(row=total_row, column=6,
                value=f"C:{crit_count} W:{warn_count} OK:{ok_count}")

            # ── Legenda ──
            legend_row = total_row + 2
            ws.cell(row=legend_row, column=1, value="LEGENDA:").font = Font(bold=True, size=9)
            ws.cell(row=legend_row + 1, column=1, value="SUPPLEMENTARI = Ore mantenute come straordinario").font = Font(size=9)
            ws.cell(row=legend_row + 2, column=1, value="PREMI = Ore convertite in premi (rischio superamento 48h)").font = Font(size=9)
            ws.cell(row=legend_row + 3, column=1, value="SPLIT = Ore redistribuite su più giorni").font = Font(size=9)

            ws.merge_cells(f'A{legend_row + 1}:J{legend_row + 1}')
            ws.merge_cells(f'A{legend_row + 2}:J{legend_row + 2}')
            ws.merge_cells(f'A{legend_row + 3}:J{legend_row + 3}')

            # Salva
            wb.save(file_path)

            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File Excel salvato:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            if open_file:
                os.startfile(file_path)

        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl",
                parent=self
            )
        except Exception as e:
            logger.error(f"Errore export Excel monitoraggio: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export Excel:\n{str(e)}",
                parent=self
            )
