"""
Overtime Reports GUI
Form per la generazione di rapporti e statistiche straordinari
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
from tkcalendar import DateEntry
import logging

logger = logging.getLogger(__name__)


def open_overtime_reports_window(parent, db_handler, lang_manager, user_name):
    """
    Apre la finestra per generare rapporti straordinari.
    
    Args:
        parent: Finestra parent
        db_handler: DatabaseHandler instance
        lang_manager: LanguageManager instance
        user_name: Nome utente loggato
    """
    OvertimeReportsWindow(parent, db_handler, lang_manager, user_name)


class OvertimeReportsWindow(tk.Toplevel):
    """
    Finestra per generare rapporti e statistiche straordinari.
    """
    
    def __init__(self, parent, db_handler, lang_manager, user_name):
        super().__init__(parent)
        
        self.db = db_handler
        self.lang = lang_manager
        self.user_name = user_name
        
        # Setup finestra
        self.title(self.lang.get('overtime_reports_title', 'Rapporti Straordinari'))
        self.geometry("1200x700")
        self.transient(parent)
        self.grab_set()
        
        self._create_widgets()
    
    def _create_widgets(self):
        """Crea i widget dell'interfaccia."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === FILTRI ===
        filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get('filters', 'Filtri'), padding="10")
        filter_frame.pack(fill=tk.X, pady=5)
        
        # Periodo
        ttk.Label(filter_frame, text=self.lang.get('period', 'Periodo:')).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(filter_frame, text=self.lang.get('from', 'Da:')).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        self.start_date = DateEntry(
            filter_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy'
        )
        self.start_date.grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(filter_frame, text=self.lang.get('to', 'A:')).grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        self.end_date = DateEntry(
            filter_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy'
        )
        self.end_date.grid(row=0, column=4, padx=5, pady=5, sticky=tk.W)
        
        # Tipo rapporto
        ttk.Label(filter_frame, text=self.lang.get('report_type', 'Tipo Rapporto:')).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.report_type_var = tk.StringVar(value='summary')
        report_types = [
            ('summary', self.lang.get('summary_report', 'Riepilogo Generale')),
            ('by_employee', self.lang.get('by_employee', 'Per Dipendente')),
            ('by_reason', self.lang.get('by_reason', 'Per Motivo')),
            ('by_department', self.lang.get('by_department', 'Per Reparto'))
        ]
        
        report_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.report_type_var,
            values=[r[1] for r in report_types],
            state='readonly',
            width=25
        )
        report_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W)
        
        # Pulsanti
        ttk.Button(
            filter_frame,
            text=self.lang.get('generate_report', 'Genera Rapporto'),
            command=self._generate_report
        ).grid(row=1, column=3, columnspan=2, padx=5, pady=5)
        
        # === RISULTATI ===
        results_frame = ttk.LabelFrame(main_frame, text=self.lang.get('results', 'Risultati'), padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Notebook per diverse viste
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Tab Tabella
        table_frame = ttk.Frame(self.notebook)
        self.notebook.add(table_frame, text=self.lang.get('table_view', 'Vista Tabella'))
        
        columns = ('employee', 'reason', 'date', 'day', 'hours', 'status')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings')
        
        self.tree.heading('employee', text=self.lang.get('employee', 'Dipendente'))
        self.tree.heading('reason', text=self.lang.get('reason', 'Motivo'))
        self.tree.heading('date', text=self.lang.get('date', 'Data'))
        self.tree.heading('day', text=self.lang.get('day_of_week', 'Giorno'))
        self.tree.heading('hours', text=self.lang.get('hours', 'Ore'))
        self.tree.heading('status', text=self.lang.get('status', 'Stato'))
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Tab Statistiche
        stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(stats_frame, text=self.lang.get('statistics', 'Statistiche'))
        
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, height=20)
        self.stats_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # === PULSANTI EXPORT ===
        export_frame = ttk.Frame(main_frame)
        export_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('export_excel', 'Esporta Excel'),
            command=self._export_to_excel
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('export_pdf', 'Esporta PDF'),
            command=self._export_to_pdf
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text=self.lang.get('close', 'Chiudi'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)
    
    def _generate_report(self):
        """Genera il rapporto in base ai filtri selezionati."""
        # Pulisci risultati precedenti
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.stats_text.delete('1.0', tk.END)
        
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()
        
        # Query base
        # Mappa nomi giorni settimana
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        query = """
        SELECT DISTINCT
            e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName,
            s.Descriptionreasons,
            s.DateStart,
            DATEDIFF(HOUR, s.DateStart, s.DateEnd) AS Hours,
            CASE 
                WHEN s.SuperVisorId IS NULL THEN 'Pending'
                WHEN s.SuperVisorId > 0 THEN 'Approved'
                ELSE 'Rejected'
            END AS Status
        FROM ResetServices.dbo.ExtraTimeApprovalStory s
        INNER JOIN Employee.dbo.EmployeeHireHistory h ON s.IdEmployee = h.EmployeeHireHistoryId
        INNER JOIN Employee.dbo.Employees e ON h.EmployeeId = e.EmployeeId
        WHERE s.DateStart >= ? AND s.DateStart <= ?
        ORDER BY s.DateStart DESC
        """
        
        try:
            cursor = self.db.conn.cursor()
            cursor.execute(query, (start_date, end_date))
            results = cursor.fetchall()
            cursor.close()
            
            # Popola tabella
            total_hours = 0
            approved_hours = 0
            pending_hours = 0
            rejected_hours = 0
            
            for row in results:
                date_str = row[2].strftime('%d/%m/%Y %H:%M') if row[2] else 'N/D'
                day_name = day_names[row[2].weekday()] if row[2] else ''
                hours = row[3] if row[3] else 0
                status = row[4]
                
                self.tree.insert('', tk.END, values=(
                    row[0],  # Employee
                    row[1],  # Reason
                    date_str,  # Date
                    day_name,  # Day of week
                    hours,  # Hours
                    status  # Status
                ))
                
                total_hours += hours
                if status == 'Approved':
                    approved_hours += hours
                elif status == 'Pending':
                    pending_hours += hours
                elif status == 'Rejected':
                    rejected_hours += hours
            
            # Genera statistiche
            stats = f"""
STATISTICHE STRAORDINARI
Periodo: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}

RIEPILOGO GENERALE:
- Totale richieste: {len(results)}
- Ore totali richieste: {total_hours}

PER STATO:
- Ore approvate: {approved_hours}
- Ore in attesa: {pending_hours}
- Ore rifiutate: {rejected_hours}

PERCENTUALI:
- Approvate: {(approved_hours/total_hours*100) if total_hours > 0 else 0:.1f}%
- In attesa: {(pending_hours/total_hours*100) if total_hours > 0 else 0:.1f}%
- Rifiutate: {(rejected_hours/total_hours*100) if total_hours > 0 else 0:.1f}%
            """
            
            self.stats_text.insert('1.0', stats)
            
        except Exception as e:
            logger.error(f"Errore generazione rapporto: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore generazione rapporto:\n{str(e)}",
                parent=self
            )
    
    def _export_to_excel(self):
        """Esporta i risultati in Excel."""
        if not self.tree.get_children():
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare'),
                parent=self
            )
            return
        
        import os
        
        # Crea directory C:\Temp se non esiste
        output_dir = r"C:\Temp"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()
        filename = f"OvertimeReport_{start_date.strftime('%d-%m-%Y')}_to_{end_date.strftime('%d-%m-%Y')}.xlsx"
        file_path = os.path.join(output_dir, filename)
        # Se il file è già aperto/bloccato, aggiungi timestamp
        if os.path.exists(file_path):
            try:
                with open(file_path, 'ab') as _test:
                    pass
            except PermissionError:
                from datetime import datetime as dt_now
                ts = dt_now.now().strftime('%H%M%S')
                filename = f"OvertimeReport_{start_date.strftime('%d-%m-%Y')}_to_{end_date.strftime('%d-%m-%Y')}_{ts}.xlsx"
                file_path = os.path.join(output_dir, filename)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Straordinari"
            
            # Header
            headers = ['Dipendente', 'Motivo', 'Data', 'Giorno', 'Ore', 'Stato']
            header_widths = [30, 35, 18, 12, 8, 12]
            wrap_align = Alignment(wrap_text=True, vertical='top')
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.column_dimensions[cell.column_letter].width = header_widths[col - 1]
            
            # Dati
            last_row = 1
            for row_idx, item in enumerate(self.tree.get_children(), 2):
                values = self.tree.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = wrap_align
                last_row = row_idx
            
            # Riga somma ore (colonna 5 = Ore)
            sum_row = last_row + 1
            sum_cell = ws.cell(row=sum_row, column=5, value=f"=SUM(E2:E{last_row})")
            sum_cell.font = Font(bold=True)
            sum_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=sum_row, column=4, value="TOTAL:").font = Font(bold=True)
            
            # Auto-filtro sulle intestazioni
            ws.auto_filter.ref = f"A1:F{last_row}"
            
            wb.save(file_path)
            
            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File Excel salvato con successo:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            if open_file:
                os.startfile(file_path)
            
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "Libreria openpyxl non installata.\nInstallare con: pip install openpyxl",
                parent=self
            )
        except Exception as e:
            logger.error(f"Errore export Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export Excel:\n{str(e)}",
                parent=self
            )
    
    def _export_to_pdf(self):
        """Esporta i risultati in PDF con calcolo costi."""
        if not self.tree.get_children():
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('no_data_to_export', 'Nessun dato da esportare'),
                parent=self
            )
            return
        
        try:
            import os
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import cm
            from reportlab.platypus import Table, TableStyle, Image as ReportLabImage, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from collections import defaultdict
            from datetime import datetime
            
            # Recupera prezzi straordinari
            pricing_query = """
            SELECT t.[Description], d.ValueITem [Value], v.[DESC] Currency,
                  case
                    when t.[Description] = 'WeekendCost' then
                        1
                    else
                        0
                end
                as IsWeekend            
              FROM 
              ResetServices.dbo.OverTimeDefaults D inner join 
              [ResetServices].[dbo].[OverTimeDescriptions] T on t.DescpriptionId=d.DescriptionId inner join 
              ResetServices.dbo.TbValute v on v.IdValuta =d.CurrencyId
              where t.DateOut is null and d.DateOut is null
            """
            
            cursor = self.db.conn.cursor()
            cursor.execute(pricing_query)
            pricing_results = cursor.fetchall()
            
            # Estrai prezzi
            weekday_cost = 0
            weekend_cost = 0
            currency = "EUR"
            
            for row in pricing_results:
                if row[3] == 1:  # IsWeekend
                    weekend_cost = float(row[1])
                else:
                    weekday_cost = float(row[1])
                currency = row[2] if row[2] else "EUR"
            
            # Recupera dati straordinari con date complete
            overtime_query = """
            SELECT 
                e.EmployeeSurname + ' ' + e.EmployeeName AS EmployeeName,
                s.Descriptionreasons,
                s.DateStart,
                s.DateEnd,
                DATEDIFF(HOUR, s.DateStart, s.DateEnd) AS Hours,
                CASE 
                    WHEN s.SuperVisorId IS NULL THEN 'Pending'
                    WHEN s.SuperVisorId > 0 THEN 'Approved'
                    ELSE 'Rejected'
                END AS Status
            FROM ResetServices.dbo.ExtraTimeApprovalStory s
            INNER JOIN Employee.dbo.EmployeeHireHistory h ON s.IdEmployee = h.EmployeeHireHistoryId
            INNER JOIN Employee.dbo.Employees e ON h.EmployeeId = e.EmployeeId
            WHERE s.DateStart >= ? AND s.DateStart <= ?
            ORDER BY s.DateStart DESC
            """
            
            start_date = self.start_date.get_date()
            end_date = self.end_date.get_date()
            
            cursor.execute(overtime_query, (start_date, end_date))
            overtime_results = cursor.fetchall()
            cursor.close()
            
            # Calcola costi per mese e anno
            monthly_costs = defaultdict(lambda: {'hours': 0, 'cost': 0})
            yearly_costs = defaultdict(lambda: {'hours': 0, 'cost': 0})
            total_cost = 0
            total_hours = 0
            
            overtime_data = []
            
            for row in overtime_results:
                employee = row[0]
                reason = row[1]
                date_start = row[2]
                date_end = row[3]
                hours = row[4] if row[4] else 0
                status = row[5]
                
                # Determina se è weekend (sabato=5, domenica=6)
                is_weekend = date_start.weekday() >= 5
                cost_per_hour = weekend_cost if is_weekend else weekday_cost
                cost = hours * cost_per_hour
                
                # Aggrega per mese e anno
                month_key = date_start.strftime('%Y-%m')
                year_key = date_start.strftime('%Y')
                
                monthly_costs[month_key]['hours'] += hours
                monthly_costs[month_key]['cost'] += cost
                yearly_costs[year_key]['hours'] += hours
                yearly_costs[year_key]['cost'] += cost
                
                total_hours += hours
                total_cost += cost
                
                # Giorno della settimana
                day_names = ['Lun', 'Mar', 'Mie', 'Joi', 'Vin', 'Sâm', 'Dum']
                day_name = day_names[date_start.weekday()]

                # Ore inizio e fine
                time_from = date_start.strftime('%H:%M') if date_start else ''
                time_to = date_end.strftime('%H:%M') if date_end else ''

                overtime_data.append({
                    'employee': employee,
                    'reason': reason,
                    'date': date_start.strftime('%d/%m/%Y'),
                    'time_from': time_from,
                    'time_to': time_to,
                    'day': day_name,
                    'date_obj': date_start,
                    'month_key': month_key,
                    'hours': hours,
                    'status': status,
                    'is_weekend': is_weekend,
                    'cost': cost
                })
            
            # Ordina per data
            overtime_data.sort(key=lambda x: x['date_obj'])
            
            from reportlab.platypus import SimpleDocTemplate, Spacer, KeepTogether
            
            # Crea directory C:\Temp se non esiste
            output_dir = r"C:\Temp"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Genera nome file automatico basato sulle date
            filename = f"report_overtime_from_{start_date.strftime('%d-%m-%Y')}_to_{end_date.strftime('%d-%m-%Y')}.pdf"
            file_path = os.path.join(output_dir, filename)
            # Se il file è già aperto/bloccato, aggiungi timestamp
            if os.path.exists(file_path):
                try:
                    with open(file_path, 'ab') as _test:
                        pass
                except PermissionError:
                    ts = datetime.now().strftime('%H%M%S')
                    filename = f"report_overtime_from_{start_date.strftime('%d-%m-%Y')}_to_{end_date.strftime('%d-%m-%Y')}_{ts}.pdf"
                    file_path = os.path.join(output_dir, filename)
            
            page_w, page_h = A4  # Portrait
            
            # Footer callback — disegnato su ogni pagina, mai sopra il contenuto
            def _draw_footer(canvas_obj, doc):
                canvas_obj.saveState()
                # Logo in alto a destra
                logo_path = "Logo.png"
                if os.path.exists(logo_path):
                    try:
                        logo = ReportLabImage(logo_path, width=1.5 * cm, height=1.5 * cm)
                        logo.drawOn(canvas_obj, page_w - 2.5 * cm, page_h - 2.2 * cm)
                    except Exception:
                        pass
                # Footer text
                canvas_obj.setFont("Helvetica", 7)
                canvas_obj.drawCentredString(page_w / 2, 1.2 * cm,
                    "Document generat automat de sistemul TraceabilityRS")
                # Numero pagina
                canvas_obj.drawRightString(page_w - 1.5 * cm, 0.7 * cm,
                    f"Page {doc.page}")
                canvas_obj.restoreState()
            
            doc = SimpleDocTemplate(
                file_path, pagesize=A4,
                leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                topMargin=2 * cm, bottomMargin=2 * cm
            )
            
            elements = []
            
            # === TITOLO ===
            title_style = ParagraphStyle(
                'Title', fontName='Helvetica-Bold', fontSize=16,
                spaceAfter=6, alignment=0
            )
            subtitle_style = ParagraphStyle(
                'Subtitle', fontName='Helvetica', fontSize=9,
                spaceAfter=2, alignment=0
            )
            elements.append(Paragraph("RAPORT ORE SUPLIMENTARE", title_style))
            elements.append(Paragraph(
                f"Perioada: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}", subtitle_style))
            elements.append(Paragraph(
                f"Data generării: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
            elements.append(Spacer(1, 0.5 * cm))
            
            # === RIEPILOGO COSTI PER MESE + ANNO side-by-side ===
            section_title = ParagraphStyle(
                'SectionTitle', fontName='Helvetica-Bold', fontSize=11,
                spaceAfter=4, spaceBefore=8, alignment=0
            )
            
            # Monthly summary
            elements.append(Paragraph("REZUMAT COSTURI PE LUNĂ", section_title))
            monthly_table_data = [['Luna', 'Ore', f'Cost ({currency})']]
            for month in sorted(monthly_costs.keys()):
                month_name = datetime.strptime(month, '%Y-%m').strftime('%B %Y')
                monthly_table_data.append([
                    month_name,
                    f"{monthly_costs[month]['hours']:.1f}",
                    f"{monthly_costs[month]['cost']:.2f}"
                ])
            
            monthly_table = Table(monthly_table_data, colWidths=[4*cm, 2*cm, 3*cm])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(monthly_table)
            elements.append(Spacer(1, 0.5 * cm))
            
            # Yearly summary
            elements.append(Paragraph("REZUMAT COSTURI PE AN", section_title))
            yearly_table_data = [['An', 'Ore', f'Cost ({currency})']]
            for year in sorted(yearly_costs.keys()):
                yearly_table_data.append([
                    year,
                    f"{yearly_costs[year]['hours']:.1f}",
                    f"{yearly_costs[year]['cost']:.2f}"
                ])
            yearly_table_data.append([
                'TOTAL',
                f"{total_hours:.1f}",
                f"{total_cost:.2f}"
            ])
            
            yearly_table = Table(yearly_table_data, colWidths=[3*cm, 2*cm, 3*cm])
            yearly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FFD700")),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(yearly_table)
            elements.append(Spacer(1, 0.3 * cm))
            
            # === IPOTEZE DI CALCOLO ===
            assumption_style = ParagraphStyle(
                'Assumption', fontName='Helvetica-Oblique', fontSize=8,
                spaceAfter=2, spaceBefore=2, alignment=0,
                textColor=colors.HexColor("#555555")
            )
            elements.append(Paragraph(
                f"<b>Ipoteze de calcul:</b> Tarif zi lucr\u0103toare = {weekday_cost:.2f} {currency}/or\u0103 | "
                f"Tarif weekend = {weekend_cost:.2f} {currency}/or\u0103",
                assumption_style))
            elements.append(Spacer(1, 0.5 * cm))
            
            # === TABELLA DETTAGLIO ORE SUPLIMENTARE ===
            elements.append(Paragraph("DETALIU ORE SUPLIMENTARE", section_title))
            
            # Stili Paragraph per word-wrap nelle celle PDF
            cell_style = ParagraphStyle(
                'CellWrap', fontName='Helvetica',
                fontSize=6, leading=7, alignment=0
            )
            cell_center = ParagraphStyle(
                'CellCenter', parent=cell_style,
                alignment=1
            )
            cell_bold = ParagraphStyle(
                'CellBold', parent=cell_style,
                fontName='Helvetica-Bold', fontSize=6
            )
            header_style = ParagraphStyle(
                'HeaderStyle', fontName='Helvetica-Bold', fontSize=6,
                textColor=colors.whitesmoke, alignment=1, leading=8
            )
            
            # 10 colonne per A4 portrait (18.6 cm disponibili)
            col_widths = [2.8*cm, 2.8*cm, 1.8*cm, 1.1*cm, 1.1*cm, 1.0*cm, 1.0*cm, 1.5*cm, 1.5*cm, 1.8*cm]
            
            detail_table_data = [[
                Paragraph('Angajat', header_style),
                Paragraph('Motiv', header_style),
                Paragraph('Data', header_style),
                Paragraph('De la ora', header_style),
                Paragraph('Până la ora', header_style),
                Paragraph('Ziua', header_style),
                Paragraph('Ore', header_style),
                Paragraph('Tip', header_style),
                Paragraph('Stare', header_style),
                Paragraph(f'Cost ({currency})', header_style)
            ]]
            
            # Raggruppa per mese e aggiungi subtotali
            current_month = None
            month_hours = 0
            month_cost = 0
            subtotal_rows = []
            month_header_rows = []
            
            for item in overtime_data:
                item_month = item['month_key']
                
                if current_month is not None and item_month != current_month:
                    month_name = datetime.strptime(current_month, '%Y-%m').strftime('%B %Y')
                    subtotal_rows.append(len(detail_table_data))
                    detail_table_data.append([
                        Paragraph(f'Subtotal {month_name}', cell_bold), '', '', '', '', '',
                        Paragraph(f'{month_hours:.1f}', cell_bold), '', '',
                        Paragraph(f'{month_cost:.2f}', cell_bold)
                    ])
                    month_hours = 0
                    month_cost = 0
                
                if item_month != current_month:
                    month_name = datetime.strptime(item_month, '%Y-%m').strftime('%B %Y')
                    month_header_rows.append(len(detail_table_data))
                    detail_table_data.append([
                        month_name.upper(), '', '', '', '', '', '', '', '', ''
                    ])
                    current_month = item_month
                
                status_ro = {'Pending': 'În așteptare', 'Approved': 'Aprobat', 'Rejected': 'Respins'}
                detail_table_data.append([
                    Paragraph(item['employee'] or '', cell_style),
                    Paragraph(item['reason'] or '', cell_style),
                    Paragraph(item['date'], cell_center),
                    Paragraph(item['time_from'], cell_center),
                    Paragraph(item['time_to'], cell_center),
                    Paragraph(item['day'], cell_center),
                    Paragraph(f"{item['hours']:.1f}", cell_center),
                    Paragraph('Weekend' if item['is_weekend'] else 'Lucr.', cell_center),
                    Paragraph(status_ro.get(item['status'], item['status']), cell_center),
                    Paragraph(f"{item['cost']:.2f}", cell_center)
                ])
                
                month_hours += item['hours']
                month_cost += item['cost']
            
            if current_month is not None:
                month_name = datetime.strptime(current_month, '%Y-%m').strftime('%B %Y')
                subtotal_rows.append(len(detail_table_data))
                detail_table_data.append([
                    Paragraph(f'Subtotal {month_name}', cell_bold), '', '', '', '', '',
                    Paragraph(f'{month_hours:.1f}', cell_bold), '', '',
                    Paragraph(f'{month_cost:.2f}', cell_bold)
                ])
            
            grand_total_row = len(detail_table_data)
            detail_table_data.append([
                Paragraph('TOTAL GENERAL', cell_bold), '', '', '', '', '',
                Paragraph(f'{total_hours:.1f}', cell_bold), '', '',
                Paragraph(f'{total_cost:.2f}', cell_bold)
            ])
            
            detail_table = Table(
                detail_table_data, colWidths=col_widths,
                repeatRows=1  # Repeat header row on each page
            )
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
            
            for row_idx in month_header_rows:
                table_style.extend([
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#2E5090")),
                    ('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.whitesmoke),
                    ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, row_idx), (-1, row_idx), 8),
                    ('SPAN', (0, row_idx), (-1, row_idx)),
                ])
            
            for row_idx in subtotal_rows:
                table_style.extend([
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor("#E0E0E0")),
                    ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, row_idx), (-1, row_idx), 7),
                ])
            
            table_style.extend([
                ('BACKGROUND', (0, grand_total_row), (-1, grand_total_row), colors.HexColor("#FFD700")),
                ('FONTNAME', (0, grand_total_row), (-1, grand_total_row), 'Helvetica-Bold'),
                ('FONTSIZE', (0, grand_total_row), (-1, grand_total_row), 8),
            ])
            
            detail_table.setStyle(TableStyle(table_style))
            elements.append(detail_table)
            
            # Build PDF with automatic pagination
            doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
            
            # Chiedi se aprire il file
            open_file = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"File PDF salvato con successo:\n{file_path}\n\nVuoi aprire il file?",
                parent=self
            )
            
            if open_file:
                os.startfile(file_path)
            
        except Exception as e:
            logger.error(f"Errore export PDF: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export PDF:\n{str(e)}",
                parent=self
            )

