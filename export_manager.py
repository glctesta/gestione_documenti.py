"""
Export Manager per Report Controllo Viscosità
Gestisce l'export in PDF e Excel con formattazione professionale
"""

import pandas as pd
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

logger = logging.getLogger("TraceabilityRS")


class ViscosityReportExporter:
    """Gestisce l'export dei report di controllo viscosità"""

    @staticmethod
    def export_to_pdf(df: pd.DataFrame, filename: str, start_date: datetime,
                      end_date: datetime, translations: dict):
        """
        Esporta il DataFrame in formato PDF

        Args:
            df: DataFrame con i dati
            filename: Nome del file di output
            start_date: Data inizio periodo
            end_date: Data fine periodo
            translations: Dizionario traduzioni
        """
        try:
            logger.info(f"Generazione PDF: {filename}")

            # Crea il documento PDF in formato landscape
            doc = SimpleDocTemplate(
                filename,
                pagesize=landscape(A4),
                rightMargin=1 * cm,
                leftMargin=1 * cm,
                topMargin=1.5 * cm,
                bottomMargin=1.5 * cm
            )

            elements = []
            styles = getSampleStyleSheet()

            # Stile titolo
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=12,
                alignment=1  # Center
            )

            # Titolo
            title_text = translations.get('coating_report_pdf_title', 'REPORT CONTROLLO VISCOSITÀ')
            title = Paragraph(title_text, title_style)
            elements.append(title)

            # Periodo
            period_style = ParagraphStyle(
                'Period',
                parent=styles['Normal'],
                fontSize=11,
                alignment=1,
                spaceAfter=20
            )
            period_template = translations.get('coating_report_pdf_period', 'Periodo: {0} - {1}')
            period_text = period_template.format(
                start_date.strftime('%d/%m/%Y'),
                end_date.strftime('%d/%m/%Y')
            )
            period = Paragraph(period_text, period_style)
            elements.append(period)

            # Prepara i dati per la tabella
            headers = [
                translations.get('coating_report_col_id', 'ID'),
                translations.get('coating_report_col_doc_date', 'Data Doc'),
                translations.get('coating_report_col_operator', 'Operatore'),
                translations.get('coating_report_col_reg_date', 'Data Reg.'),
                translations.get('coating_report_col_qty_mat', 'Qtà Mat.'),
                translations.get('coating_report_col_qty_dil', 'Qtà Dil.'),
                translations.get('coating_report_col_material', 'Materiale')
            ]

            data = [headers]

            for _, row in df.iterrows():
                data.append([
                    str(row['ViscosityControlId']),
                    str(row['DocumentDate']),
                    str(row['UserName']),
                    str(row['RegDate']),
                    str(row['QtyMatUsed']),
                    str(row['QtyDiluantUsed']),
                    str(row['Material'])
                ])

            # Crea la tabella
            table = Table(data, colWidths=[2 * cm, 3 * cm, 4 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 6 * cm])

            # Stile della tabella
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                # Dati
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

                # Righe alternate
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

                # Bordi
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ]))

            elements.append(table)

            # Footer
            elements.append(Spacer(1, 20))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=9,
                alignment=2  # Right
            )
            footer_template = translations.get('coating_report_pdf_footer',
                                               'Generato il: {0} | Totale record: {1}')
            footer_text = footer_template.format(
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                len(df)
            )
            footer = Paragraph(footer_text, footer_style)
            elements.append(footer)

            # Genera il PDF
            doc.build(elements)

            logger.info(f"PDF generato con successo: {filename}")
            return True

        except Exception as e:
            logger.error(f"Errore generazione PDF: {str(e)}", exc_info=True)
            raise Exception(f"Errore durante la generazione del PDF: {str(e)}")

    @staticmethod
    def export_to_excel(df: pd.DataFrame, filename: str, start_date: datetime,
                        end_date: datetime, translations: dict):
        """
        Esporta il DataFrame in formato Excel

        Args:
            df: DataFrame con i dati
            filename: Nome del file di output
            start_date: Data inizio periodo
            end_date: Data fine periodo
            translations: Dizionario traduzioni
        """
        try:
            logger.info(f"Generazione Excel: {filename}")

            # Crea il workbook
            wb = Workbook()
            ws = wb.active
            ws.title = translations.get('coating_report_excel_sheet', 'Controllo Viscosità')

            # Stili
            title_font = Font(name='Calibri', size=16, bold=True, color='1F4788')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            normal_font = Font(name='Calibri', size=10)

            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Titolo
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = translations.get('coating_report_pdf_title', 'REPORT CONTROLLO VISCOSITÀ')
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Periodo
            ws.merge_cells('A2:G2')
            period_cell = ws['A2']
            period_template = translations.get('coating_report_pdf_period', 'Periodo: {0} - {1}')
            period_cell.value = period_template.format(
                start_date.strftime('%d/%m/%Y'),
                end_date.strftime('%d/%m/%Y')
            )
            period_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 20

            # Headers (riga 4)
            headers = [
                translations.get('coating_report_col_id', 'ID Controllo'),
                translations.get('coating_report_col_doc_date', 'Data Documento'),
                translations.get('coating_report_col_operator', 'Operatore'),
                translations.get('coating_report_col_reg_date', 'Data Registrazione'),
                translations.get('coating_report_col_qty_mat', 'Qtà Materiale'),
                translations.get('coating_report_col_qty_dil', 'Qtà Diluente'),
                translations.get('coating_report_col_material', 'Materiale')
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

            ws.row_dimensions[4].height = 25

            # Dati
            for row_num, (_, row_data) in enumerate(df.iterrows(), 5):
                ws.cell(row=row_num, column=1, value=row_data['ViscosityControlId'])
                ws.cell(row=row_num, column=2, value=row_data['DocumentDate'])
                ws.cell(row=row_num, column=3, value=row_data['UserName'])
                ws.cell(row=row_num, column=4, value=row_data['RegDate'])
                ws.cell(row=row_num, column=5, value=row_data['QtyMatUsed'])
                ws.cell(row=row_num, column=6, value=row_data['QtyDiluantUsed'])
                ws.cell(row=row_num, column=7, value=row_data['Material'])

                # Applica stile
                for col_num in range(1, 8):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = normal_font
                    cell.border = border_style
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # Footer
            footer_row = len(df) + 6
            ws.merge_cells(f'A{footer_row}:E{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            footer_cell.font = Font(name='Calibri', size=9, italic=True)

            total_label = translations.get('coating_report_excel_total', 'Totale record:')
            ws[f'F{footer_row}'] = total_label
            ws[f'F{footer_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'F{footer_row}'].alignment = Alignment(horizontal='right')

            ws[f'G{footer_row}'] = len(df)
            ws[f'G{footer_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'G{footer_row}'].alignment = Alignment(horizontal='center')

            # Larghezza colonne
            column_widths = [15, 18, 25, 25, 15, 15, 40]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width

            # Salva il file
            wb.save(filename)

            logger.info(f"Excel generato con successo: {filename}")
            return True

        except Exception as e:
            logger.error(f"Errore generazione Excel: {str(e)}", exc_info=True)
            raise Exception(f"Errore durante la generazione dell'Excel: {str(e)}")
