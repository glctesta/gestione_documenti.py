# -*- coding: utf-8 -*-
"""
fqc_products_report_gui.py

Report interattivo delle schede (LabelCode) validate durante il processo di
esecuzione controlli FQC Prodotti (_open_fqc_products_with_login).

Fonte dati: [Traceability_RS].[chk].[ProductCheckListDataLabelLogs]
- Una "scheda validata" = un LabelCode con i relativi item di checklist salvati.
- Filtri: periodo (preimpostato al giorno precedente), codice prodotto, LabelCode.
- Esito OK se tutti gli item sono OK, altrimenti NOK (evidenziato in rosso).
- Export Excel in C:\\temp.
"""
from __future__ import annotations

import datetime
import logging
import os
import tkinter as tk
from tkinter import messagebox, ttk

logger = logging.getLogger("TraceabilityRS")

# ─── SQL ──────────────────────────────────────────────────────────────────────

_Q_VALIDATED_CARDS = """
SELECT
    ll.IDLabelCode,
    l.LabelCod,
    o.OrderNumber,
    p.ProductCode,
    c.ClientName,
    COUNT(*)                                       AS ItemsTotal,
    SUM(CASE WHEN ll.IsOK = 1 THEN 1 ELSE 0 END)   AS ItemsOK,
    SUM(CASE WHEN ll.IsOK = 0 THEN 1 ELSE 0 END)   AS ItemsNOK,
    MIN(ll.DateCheckList)                          AS FirstCheck,
    MAX(ll.DateCheckList)                          AS LastCheck,
    MAX(ll.[User])                                 AS LastUser
FROM [Traceability_RS].[chk].[ProductCheckListDataLabelLogs] ll
INNER JOIN Traceability_RS.dbo.LabelCodes l ON l.IDLabelCode = ll.IDLabelCode
INNER JOIN Traceability_RS.dbo.Boards     b ON b.IDBoard     = l.IDBoard
INNER JOIN Traceability_RS.dbo.Orders     o ON o.IDOrder     = b.IDOrder
INNER JOIN Traceability_RS.dbo.Products   p ON p.IDProduct   = o.IDProduct
LEFT  JOIN Traceability_RS.dbo.Clients    c ON c.IDClient    = p.IDClient
WHERE ll.DateCheckList >= ? AND ll.DateCheckList < ?
{code_filter}
{label_filter}
GROUP BY ll.IDLabelCode, l.LabelCod, o.OrderNumber, p.ProductCode, c.ClientName
ORDER BY MAX(ll.DateCheckList) DESC, l.LabelCod
"""


def _run_query(conn, date_from: datetime.date, date_to: datetime.date,
               product_code: str = '', labelcode: str = '') -> list[dict]:
    """Esegue la query nel periodo [date_from, date_to] inclusi (giorni interi)."""
    start = datetime.datetime.combine(date_from, datetime.time(0, 0, 0))
    end   = datetime.datetime.combine(date_to + datetime.timedelta(days=1),
                                      datetime.time(0, 0, 0))
    params: list = [start, end]

    if product_code.strip():
        code_filter = "AND p.ProductCode LIKE ?"
        params.append(f'%{product_code.strip()}%')
    else:
        code_filter = ''

    if labelcode.strip():
        label_filter = "AND l.LabelCod LIKE ?"
        params.append(f'%{labelcode.strip()}%')
    else:
        label_filter = ''

    sql = _Q_VALIDATED_CARDS.format(code_filter=code_filter,
                                    label_filter=label_filter)
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _fmt_dt(value) -> str:
    """Formatta un datetime in dd/mm/yyyy HH:MM."""
    if value is None:
        return ''
    try:
        return value.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return str(value)


# ─── Excel export ─────────────────────────────────────────────────────────────

def _export_excel(rows: list[dict], date_from: datetime.date,
                  date_to: datetime.date) -> str:
    """Crea il file Excel, lo salva in C:\\temp e ne restituisce il percorso."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export.") from exc

    out_dir = r'C:\temp'
    os.makedirs(out_dir, exist_ok=True)
    filename = (
        f"Raport_scheda_validate_FQC_"
        f"{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    )
    path = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FQC Validated Cards'

    H_FILL  = PatternFill('solid', fgColor='1F3864')
    H_FONT  = Font(bold=True, color='FFFFFF', size=10)
    NOK_FONT = Font(color='CC0000', size=9, bold=True)
    ALT_FILL = PatternFill('solid', fgColor='F4F6F8')
    THIN_SIDE = Side(style='thin', color='C0C0C0')
    THIN = Border(left=THIN_SIDE, right=THIN_SIDE,
                  top=THIN_SIDE,  bottom=THIN_SIDE)

    ws['A1'] = 'FQC Products — Validated Cards Report'
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:I1')
    ws['A2'] = (f'Period: {date_from.strftime("%d/%m/%Y")} -> '
                f'{date_to.strftime("%d/%m/%Y")}')
    ws['A2'].font = Font(size=9, color='7F8C8D')
    ws.merge_cells('A2:I2')

    headers = ['Validated at', 'LabelCode', 'Product Code', 'Client',
               'Order', 'Items', 'OK', 'NOK', 'Result']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(4, ci, h)
        c.fill, c.font, c.border = H_FILL, H_FONT, THIN
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[4].height = 22

    excel_row = 5
    for i, r in enumerate(rows):
        nok = int(r.get('ItemsNOK') or 0)
        result = 'NOK' if nok > 0 else 'OK'
        row_font = NOK_FONT if nok > 0 else Font(size=9)
        fill = ALT_FILL if i % 2 == 0 else None

        vals = [
            _fmt_dt(r.get('LastCheck')),
            r.get('LabelCod') or '',
            r.get('ProductCode') or '',
            r.get('ClientName') or '',
            r.get('OrderNumber') or '',
            int(r.get('ItemsTotal') or 0),
            int(r.get('ItemsOK') or 0),
            nok,
            result,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(excel_row, ci, v)
            cell.font   = row_font
            cell.border = THIN
            if fill:
                cell.fill = fill
            if ci in (6, 7, 8, 9):
                cell.alignment = Alignment(horizontal='center')
        excel_row += 1

    widths = (18, 22, 24, 24, 16, 8, 8, 8, 10)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'

    wb.save(path)
    return path


# ─── GUI ──────────────────────────────────────────────────────────────────────

class FqcProductsReportWindow(tk.Toplevel):
    """Form di ricerca delle schede FQC validate."""

    _COL_IDS   = ('date', 'labelcode', 'product', 'client',
                  'order', 'items', 'ok', 'nok', 'result')
    _COL_HEADS = ('Validated at', 'LabelCode', 'Product Code', 'Client',
                  'Order', 'Items', 'OK', 'NOK', 'Result')
    _COL_WIDTHS = (130, 150, 180, 170, 110, 60, 50, 50, 70)

    def __init__(self, master, db, lang):
        super().__init__(master)
        self.db   = db
        self.lang = lang
        self._rows: list[dict] = []

        self.title(self.lang.get('fqc_report_title',
                                 'FQC Prodotti — Schede Validate'))
        self.resizable(True, True)
        self.geometry('1020x680')
        self.minsize(860, 500)

        self._build_ui()
        self.grab_set()

    # ── UI ─────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Header
        header = tk.Frame(self, bg='#1F3864')
        header.pack(fill=tk.X)
        logo_path = os.path.join(os.path.dirname(__file__), 'Logo.png')
        try:
            from PIL import Image as PILImage, ImageTk
            img = PILImage.open(logo_path)
            img.thumbnail((120, 40))
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(header, image=self._logo_img,
                     bg='#1F3864').pack(side=tk.LEFT, padx=14, pady=8)
        except Exception:
            pass
        tk.Label(
            header,
            text=self.lang.get('fqc_report_title', 'FQC Prodotti — Schede Validate'),
            bg='#1F3864', fg='white',
            font=('Helvetica', 13, 'bold')
        ).pack(side=tk.LEFT, padx=8, pady=12)

        # Filtri
        filter_frame = ttk.LabelFrame(
            self, text=self.lang.get('fqc_report_filter_title', 'Filtri di ricerca')
        )
        filter_frame.pack(fill=tk.X, padx=10, pady=6)

        yesterday = datetime.date.today() - datetime.timedelta(days=1)

        ttk.Label(filter_frame,
                  text=self.lang.get('fqc_report_date_from', 'Da (gg/mm/aaaa):')).grid(
            row=0, column=0, sticky='w', padx=8, pady=6)
        self._var_date_from = tk.StringVar(value=yesterday.strftime('%d/%m/%Y'))
        ttk.Entry(filter_frame, textvariable=self._var_date_from, width=13).grid(
            row=0, column=1, padx=4, pady=6)

        ttk.Label(filter_frame,
                  text=self.lang.get('fqc_report_date_to', 'A (gg/mm/aaaa):')).grid(
            row=0, column=2, sticky='w', padx=8, pady=6)
        self._var_date_to = tk.StringVar(value=yesterday.strftime('%d/%m/%Y'))
        ttk.Entry(filter_frame, textvariable=self._var_date_to, width=13).grid(
            row=0, column=3, padx=4, pady=6)

        ttk.Label(filter_frame,
                  text=self.lang.get('fqc_report_product_code',
                                     'Codice prodotto:')).grid(
            row=0, column=4, sticky='w', padx=8, pady=6)
        self._var_code = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self._var_code, width=18).grid(
            row=0, column=5, padx=4, pady=6)

        ttk.Label(filter_frame,
                  text=self.lang.get('fqc_report_labelcode', 'LabelCode:')).grid(
            row=0, column=6, sticky='w', padx=8, pady=6)
        self._var_label = tk.StringVar()
        label_entry = ttk.Entry(filter_frame, textvariable=self._var_label, width=18)
        label_entry.grid(row=0, column=7, padx=4, pady=6)

        ttk.Button(
            filter_frame,
            text=self.lang.get('fqc_report_btn_search', '🔍 Cerca'),
            command=self._do_search
        ).grid(row=0, column=8, padx=12, pady=6)

        # Invio = cerca
        for w in (label_entry,):
            w.bind('<Return>', lambda e: self._do_search())

        # Treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        self._tree = ttk.Treeview(
            tree_frame, columns=self._COL_IDS, show='headings', selectmode='browse'
        )
        for cid, head, width in zip(self._COL_IDS, self._COL_HEADS, self._COL_WIDTHS):
            self._tree.heading(cid, text=head)
            anchor = 'w' if cid in ('product', 'client') else 'center'
            self._tree.column(cid, width=width, anchor=anchor,
                              stretch=(cid in ('product', 'client')))

        self._tree.tag_configure('nok', foreground='#CC0000')
        self._tree.tag_configure('ok', foreground='#1B5E20')

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # Riepilogo
        self._summary_frame = tk.Frame(self, bg='#E8F5E9', bd=1, relief='groove')
        self._summary_frame.pack(fill=tk.X, padx=10, pady=2)
        self._lbl_total = tk.Label(self._summary_frame, bg='#E8F5E9',
                                   font=('Helvetica', 9, 'bold'), fg='#1F3864',
                                   anchor='w')
        self._lbl_ok = tk.Label(self._summary_frame, bg='#E8F5E9',
                                font=('Helvetica', 9, 'bold'), fg='#1B5E20',
                                anchor='w')
        self._lbl_nok = tk.Label(self._summary_frame, bg='#E8F5E9',
                                 font=('Helvetica', 9, 'bold'), fg='#CC0000',
                                 anchor='w')
        for lbl in (self._lbl_total, self._lbl_ok, self._lbl_nok):
            lbl.pack(side=tk.LEFT, padx=18, pady=4)

        # Pulsanti
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill=tk.X, padx=10, pady=6)
        self._btn_export = ttk.Button(
            btn_bar,
            text=self.lang.get('fqc_report_btn_export', '📊 Esporta Excel'),
            command=self._do_export, state='disabled'
        )
        self._btn_export.pack(side=tk.LEFT, padx=4)
        ttk.Button(
            btn_bar, text=self.lang.get('btn_close', 'Chiudi'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=4)

    # ── Ricerca ─────────────────────────────────────────────────────────────────

    def _parse_date(self, text: str):
        text = text.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        return None

    def _do_search(self):
        date_from = self._parse_date(self._var_date_from.get())
        date_to   = self._parse_date(self._var_date_to.get())

        if not date_from or not date_to:
            messagebox.showwarning(
                self.lang.get('warning', 'Warning'),
                self.lang.get('fqc_report_invalid_dates',
                              'Inserire date valide nel formato gg/mm/aaaa.'),
                parent=self)
            return
        if date_from > date_to:
            messagebox.showwarning(
                self.lang.get('warning', 'Warning'),
                self.lang.get('fqc_report_date_order',
                              'La data "Da" deve essere <= alla data "A".'),
                parent=self)
            return

        try:
            self._rows = _run_query(
                self.db.conn, date_from, date_to,
                self._var_code.get(), self._var_label.get()
            )
        except Exception as exc:
            logger.error(f'FqcProductsReportWindow search: {exc}', exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Error'),
                f'Query error:\n{exc}', parent=self)
            return

        self._date_from = date_from
        self._date_to   = date_to
        self._populate_tree()

    # ── Popolamento ─────────────────────────────────────────────────────────────

    def _populate_tree(self):
        for item in self._tree.get_children():
            self._tree.delete(item)

        if not self._rows:
            self._tree.insert('', 'end', values=(
                '', self.lang.get('fqc_report_no_results',
                                  'Nessuna scheda validata nel periodo selezionato.'),
                '', '', '', '', '', '', ''
            ))
            self._update_summary(0, 0, 0)
            self._btn_export.config(state='disabled')
            return

        total_ok = 0
        total_nok = 0
        for r in self._rows:
            nok = int(r.get('ItemsNOK') or 0)
            result = 'NOK' if nok > 0 else 'OK'
            tag = 'nok' if nok > 0 else 'ok'
            if nok > 0:
                total_nok += 1
            else:
                total_ok += 1

            self._tree.insert('', 'end', values=(
                _fmt_dt(r.get('LastCheck')),
                r.get('LabelCod') or '',
                r.get('ProductCode') or '',
                r.get('ClientName') or '',
                r.get('OrderNumber') or '',
                int(r.get('ItemsTotal') or 0),
                int(r.get('ItemsOK') or 0),
                nok,
                result,
            ), tags=(tag,))

        self._update_summary(len(self._rows), total_ok, total_nok)
        self._btn_export.config(state='normal')

    def _update_summary(self, total: int, ok: int, nok: int):
        self._lbl_total.config(
            text=self.lang.get('fqc_report_summary_total',
                               'Schede validate: {0}').format(total))
        self._lbl_ok.config(
            text=self.lang.get('fqc_report_summary_ok', 'OK: {0}').format(ok))
        self._lbl_nok.config(
            text=self.lang.get('fqc_report_summary_nok', 'NOK: {0}').format(nok))

    # ── Export ──────────────────────────────────────────────────────────────────

    def _do_export(self):
        if not self._rows:
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('fqc_report_no_data_export',
                              'Nessun dato da esportare. Esegui prima una ricerca.'),
                parent=self)
            return
        try:
            path = _export_excel(self._rows, self._date_from, self._date_to)
        except Exception as exc:
            logger.error(f'FqcProductsReportWindow export: {exc}', exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Error'),
                f'Export error:\n{exc}', parent=self)
            return

        messagebox.showinfo(
            self.lang.get('success', 'Success'),
            self.lang.get('fqc_report_export_ok', 'File salvato:\n{0}').format(path),
            parent=self)
        try:
            os.startfile(path)
        except Exception:
            pass


def open_fqc_products_report(parent, db, lang):
    """Entry point: apre la finestra del report schede FQC validate."""
    FqcProductsReportWindow(parent, db, lang)
