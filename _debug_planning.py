# -*- coding: utf-8 -*-
"""Test diagnostico per leggere il file Excel PlanningMachine"""
import os
import glob
from datetime import datetime

PLANNING_PATH = r"T:\Planning"
PLANNING_TAB = "PlanningMachine"

# Trova file più recente
patterns = [os.path.join(PLANNING_PATH, '*.xlsx'), os.path.join(PLANNING_PATH, '*.xls')]
files = []
for p in patterns:
    files.extend(glob.glob(p))

if not files:
    print("NESSUN FILE TROVATO")
    exit()

latest = max(files, key=os.path.getmtime)
print(f"File: {latest}")
print(f"Mod: {datetime.fromtimestamp(os.path.getmtime(latest))}")

import openpyxl
wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

if PLANNING_TAB not in wb.sheetnames:
    print(f"TAB '{PLANNING_TAB}' NON TROVATO!")
    wb.close()
    exit()

ws = wb[PLANNING_TAB]

# Leggi header (prima riga)
headers = []
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = list(row)
    break

print(f"\nTotale colonne: {len(headers)}")
print("=" * 80)
for i, h in enumerate(headers):
    col_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
    print(f"  Col {col_letter} (idx {i}): {h}")

# Cerca righe con PTHM
COL_PHASE = 4       # E
COL_ORDER = 10      # K
COL_START = 14      # O

print(f"\n{'='*80}")
print(f"Colonna PHASE (E, idx {COL_PHASE}): {headers[COL_PHASE] if COL_PHASE < len(headers) else 'FUORI RANGE'}")
print(f"Colonna ORDER (K, idx {COL_ORDER}): {headers[COL_ORDER] if COL_ORDER < len(headers) else 'FUORI RANGE'}")
print(f"Colonna START (O, idx {COL_START}): {headers[COL_START] if COL_START < len(headers) else 'FUORI RANGE'}")

print(f"\n{'='*80}")
print("RIGHE CON PTHM (prime 20):")
count = 0
now = datetime.now()
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row or len(row) <= COL_START:
        continue
    
    phase_raw = row[COL_PHASE]
    if not phase_raw:
        continue
    
    phase_str = str(phase_raw).strip().upper()
    if 'PTHM' not in phase_str:
        continue
    
    order_raw = row[COL_ORDER]
    start_raw = row[COL_START]
    
    print(f"\n  Row {row_idx}:")
    print(f"    PHASE: '{phase_raw}' (type={type(phase_raw).__name__})")
    print(f"    ORDER: '{order_raw}' (type={type(order_raw).__name__})")
    print(f"    START: '{start_raw}' (type={type(start_raw).__name__})")
    
    # Prova parsing
    if isinstance(start_raw, datetime):
        planned = start_raw
        diff = (planned - now).total_seconds() / 3600
        print(f"    PARSED: {planned} (in {diff:.1f}h)")
        in_window = 0 <= diff <= 4
        print(f"    IN WINDOW 4h: {in_window} (now={now.strftime('%H:%M')}, cutoff={4}h)")
    else:
        print(f"    ★ Non è datetime! Tentativo parsing stringa...")
        for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M',
                    '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %I:%M:%S %p', '%d/%m/%Y %H:%M:%S'):
            try:
                planned = datetime.strptime(str(start_raw).strip(), fmt)
                diff = (planned - now).total_seconds() / 3600
                print(f"    PARSED ({fmt}): {planned} (in {diff:.1f}h)")
                break
            except ValueError:
                continue
        else:
            print(f"    ❌ NESSUN FORMATO MATCH!")
    
    count += 1
    if count >= 20:
        break

wb.close()
print(f"\nTotale righe PTHM trovate: {count}")
