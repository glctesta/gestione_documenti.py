"""
indirect_materials_report.py
Report mensile richieste e consegne materiali indiretti.

- Tab "Dettaglio":   ogni riga = una singola richiesta, filtrata per anno/mese/tipo/stato
- Tab "Riepilogo":   aggregato per mese → totale richieste / totale consegnate / qty
- Tab "Verifica acquisti": per ogni codice sollecitato al riordino, se l'acquisto
                     e' poi arrivato davvero (confronto fra la giacenza al momento
                     della richiesta e i carichi Excel/D365 successivi)
- Export Excel:      salva i tre tab in un file .xlsx
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  Finestra principale
# ─────────────────────────────────────────────────────────────────────────────
class IndirectMaterialsReportWindow(tk.Toplevel):

    def __init__(self, master, db, lang):
        super().__init__(master)
        self.db   = db
        self.lang = lang

        self.title(lang.get('ind_rep_title', 'Report Materiali Indiretti per Mese'))
        self.geometry("1100x700")
        self.resizable(True, True)
        self.transient(master)

        self._build_ui()
        self._populate_filters()
        self._load_data()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ------------------------------------------------------------------ #
    #  UI                                                                  #
    # ------------------------------------------------------------------ #
    def _build_ui(self):
        # ── Toolbar filtri ──────────────────────────────────────────────
        toolbar = ttk.LabelFrame(self, text=self.lang.get('ind_rep_filters', 'Filtri'), padding=8)
        toolbar.pack(fill='x', padx=10, pady=(10, 0))

        # Anno
        ttk.Label(toolbar, text=self.lang.get('ind_rep_year', 'Anno:')).grid(row=0, column=0, padx=(0, 4))
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(toolbar, textvariable=self.year_var, width=8, state='readonly')
        self.year_combo.grid(row=0, column=1, padx=(0, 12))
        self.year_combo.bind('<<ComboboxSelected>>', lambda e: self._load_data())

        # Mese
        ttk.Label(toolbar, text=self.lang.get('ind_rep_month', 'Mese:')).grid(row=0, column=2, padx=(0, 4))
        self.month_var = tk.StringVar()
        months = [self.lang.get('ind_rep_all_months', 'Tutti i mesi')] + [
            f"{i:02d} - {self._month_name(i)}" for i in range(1, 13)
        ]
        self.month_combo = ttk.Combobox(toolbar, textvariable=self.month_var, values=months, width=18, state='readonly')
        self.month_combo.current(0)
        self.month_combo.grid(row=0, column=3, padx=(0, 12))
        self.month_combo.bind('<<ComboboxSelected>>', lambda e: self._load_data())

        # Tipo materiale
        ttk.Label(toolbar, text=self.lang.get('ind_rep_type', 'Tipo:')).grid(row=0, column=4, padx=(0, 4))
        self.type_var = tk.StringVar()
        self.type_combo = ttk.Combobox(toolbar, textvariable=self.type_var, width=20, state='readonly')
        self.type_combo.grid(row=0, column=5, padx=(0, 12))
        self.type_combo.bind('<<ComboboxSelected>>', lambda e: self._load_data())

        # Stato
        ttk.Label(toolbar, text=self.lang.get('ind_rep_status', 'Stato:')).grid(row=0, column=6, padx=(0, 4))
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(toolbar, textvariable=self.status_var, width=16, state='readonly')
        self.status_combo.grid(row=0, column=7, padx=(0, 12))
        self.status_combo.bind('<<ComboboxSelected>>', lambda e: self._load_data())

        # Pulsanti
        ttk.Button(toolbar, text=self.lang.get('btn_refresh', '🔄 Aggiorna'),
                   command=self._load_data).grid(row=0, column=8, padx=6)
        ttk.Button(toolbar, text=self.lang.get('ind_rep_export_excel', '📊 Export Excel'),
                   command=self._export_excel).grid(row=0, column=9, padx=6)

        # ── Notebook (Dettaglio / Riepilogo) ────────────────────────────
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Tab Dettaglio
        detail_frame = ttk.Frame(self.notebook)
        self.notebook.add(detail_frame, text=self.lang.get('ind_rep_tab_detail', 'Dettaglio Richieste'))

        det_cols = ('id', 'data', 'codice', 'descrizione', 'tipo',
                    'qty_rich', 'stato', 'richiedente', 'preparatore', 'data_consegna')
        self.detail_tree = ttk.Treeview(detail_frame, columns=det_cols, show='headings')
        headers = {
            'id':           self.lang.get('ind_rep_col_id', 'ID'),
            'data':         self.lang.get('ind_rep_col_date', 'Data Richiesta'),
            'codice':       self.lang.get('ind_import_col_code', 'Codice'),
            'descrizione':  self.lang.get('ind_import_col_desc', 'Descrizione'),
            'tipo':         self.lang.get('ind_req_col_type', 'Tipo'),
            'qty_rich':     self.lang.get('ind_rep_col_qty_req', 'Qty Richiesta'),
            'stato':        self.lang.get('ind_req_col_status', 'Stato'),
            'richiedente':  self.lang.get('ind_req_col_requester', 'Richiedente'),
            'preparatore':  self.lang.get('ind_req_col_preparer', 'Preparatore'),
            'data_consegna':self.lang.get('ind_rep_col_delivery_date', 'Data Consegna'),
        }
        widths = {'id': 50, 'data': 130, 'codice': 110, 'descrizione': 250, 'tipo': 100,
                  'qty_rich': 80, 'stato': 100, 'richiedente': 110, 'preparatore': 110, 'data_consegna': 130}
        for c in det_cols:
            self.detail_tree.heading(c, text=headers[c])
            self.detail_tree.column(c, width=widths[c], anchor='e' if c == 'qty_rich' else 'w')

        det_sb_v = ttk.Scrollbar(detail_frame, orient='vertical', command=self.detail_tree.yview)
        det_sb_h = ttk.Scrollbar(detail_frame, orient='horizontal', command=self.detail_tree.xview)
        self.detail_tree.configure(yscrollcommand=det_sb_v.set, xscrollcommand=det_sb_h.set)
        det_sb_v.pack(side='right', fill='y')
        det_sb_h.pack(side='bottom', fill='x')
        self.detail_tree.pack(fill='both', expand=True)

        # Tag colori stato
        self.detail_tree.tag_configure('PRELEVATA', background='#d4edda')
        self.detail_tree.tag_configure('ANNULLATA',  background='#f8d7da')

        # Totale dettaglio
        self.detail_total_var = tk.StringVar()
        ttk.Label(detail_frame, textvariable=self.detail_total_var,
                  font=('Segoe UI', 9, 'italic')).pack(anchor='e', padx=6)

        # Tab Riepilogo mensile
        summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(summary_frame, text=self.lang.get('ind_rep_tab_summary', 'Riepilogo Mensile'))

        sum_cols = ('anno_mese', 'tipo', 'n_richieste', 'n_consegnate', 'qty_richiesta', 'qty_consegnata', 'pct')
        self.summary_tree = ttk.Treeview(summary_frame, columns=sum_cols, show='headings')
        sum_headers = {
            'anno_mese':     self.lang.get('ind_rep_col_month', 'Anno/Mese'),
            'tipo':          self.lang.get('ind_req_col_type', 'Tipo'),
            'n_richieste':   self.lang.get('ind_rep_col_n_req', '# Richieste'),
            'n_consegnate':  self.lang.get('ind_rep_col_n_del', '# Consegnate'),
            'qty_richiesta': self.lang.get('ind_rep_col_qty_req', 'Qty Richiesta'),
            'qty_consegnata':self.lang.get('ind_rep_col_qty_del', 'Qty Consegnata'),
            'pct':           self.lang.get('ind_rep_col_pct', '% Consegnato'),
        }
        sum_widths = {'anno_mese': 100, 'tipo': 130, 'n_richieste': 100,
                      'n_consegnate': 110, 'qty_richiesta': 110, 'qty_consegnata': 120, 'pct': 100}
        for c in sum_cols:
            self.summary_tree.heading(c, text=sum_headers[c])
            self.summary_tree.column(c, width=sum_widths[c],
                                     anchor='e' if c != 'anno_mese' and c != 'tipo' else 'w')

        sum_sb = ttk.Scrollbar(summary_frame, orient='vertical', command=self.summary_tree.yview)
        self.summary_tree.configure(yscrollcommand=sum_sb.set)
        sum_sb.pack(side='right', fill='y')
        self.summary_tree.pack(fill='both', expand=True)

        self.summary_total_var = tk.StringVar()
        ttk.Label(summary_frame, textvariable=self.summary_total_var,
                  font=('Segoe UI', 9, 'italic')).pack(anchor='e', padx=6)

        # ── Tab Verifica acquisti ───────────────────────────────────────
        purch_frame = ttk.Frame(self.notebook)
        self.notebook.add(purch_frame,
                          text=self.lang.get('ind_rep_tab_purchases', 'Verifica Acquisti'))

        pur_cols = ('codice', 'descrizione', 'n_rich', 'ultima_rich', 'giac_rich',
                    'minimo', 'giac_carico', 'data_carico', 'differenza',
                    'entrata', 'data_entrata', 'esito')
        self.purch_tree = ttk.Treeview(purch_frame, columns=pur_cols, show='headings')
        pur_headers = {
            'codice':       self.lang.get('ind_import_col_code', 'Codice'),
            'descrizione':  self.lang.get('ind_import_col_desc', 'Descrizione'),
            'n_rich':       self.lang.get('ind_rep_col_n_reorders', '# Richieste'),
            'ultima_rich':  self.lang.get('ind_rep_col_last_reorder', 'Ultima richiesta'),
            'giac_rich':    self.lang.get('ind_rep_col_stock_at_req', 'Giacenza alla richiesta'),
            'minimo':       self.lang.get('ind_stock_col_min', 'Scorta minima'),
            'giac_carico':  self.lang.get('ind_rep_col_stock_last_load', 'Giacenza ultimo carico'),
            'data_carico':  self.lang.get('ind_rep_col_last_load', 'Data ultimo carico'),
            'differenza':   self.lang.get('ind_rep_col_difference', 'Differenza'),
            'entrata':      self.lang.get('ind_rep_col_incoming', 'Entrata rilevata'),
            'data_entrata': self.lang.get('ind_rep_col_incoming_date', 'Data entrata'),
            'esito':        self.lang.get('ind_rep_col_outcome', 'Esito'),
        }
        pur_widths = {'codice': 100, 'descrizione': 220, 'n_rich': 80, 'ultima_rich': 130,
                      'giac_rich': 130, 'minimo': 100, 'giac_carico': 140, 'data_carico': 130,
                      'differenza': 90, 'entrata': 110, 'data_entrata': 130, 'esito': 170}
        num_cols = ('n_rich', 'giac_rich', 'minimo', 'giac_carico', 'differenza', 'entrata')
        for c in pur_cols:
            self.purch_tree.heading(c, text=pur_headers[c])
            self.purch_tree.column(c, width=pur_widths[c],
                                   anchor='e' if c in num_cols else 'w')

        pur_sb_v = ttk.Scrollbar(purch_frame, orient='vertical', command=self.purch_tree.yview)
        pur_sb_h = ttk.Scrollbar(purch_frame, orient='horizontal', command=self.purch_tree.xview)
        self.purch_tree.configure(yscrollcommand=pur_sb_v.set, xscrollcommand=pur_sb_h.set)
        pur_sb_v.pack(side='right', fill='y')
        pur_sb_h.pack(side='bottom', fill='x')
        self.purch_tree.pack(fill='both', expand=True)

        self.purch_tree.tag_configure('ARRIVATO', background='#d4edda')
        self.purch_tree.tag_configure('MANCANTE', background='#f8d7da')
        self.purch_tree.tag_configure('ATTESA', background='#fff3cd')

        self.purch_total_var = tk.StringVar()
        ttk.Label(purch_frame, textvariable=self.purch_total_var,
                  font=('Segoe UI', 9, 'italic')).pack(anchor='e', padx=6)
        ttk.Label(purch_frame,
                  text=self.lang.get(
                      'ind_rep_purch_note',
                      "L'entrata e' misurata sugli aumenti di giacenza fra un carico Excel "
                      "(stock D365) e il successivo, dopo la data della richiesta."),
                  font=('Segoe UI', 8), foreground='#666').pack(anchor='w', padx=6, pady=(0, 4))

    # ------------------------------------------------------------------ #
    #  Helpers                                                             #
    # ------------------------------------------------------------------ #
    @staticmethod
    def _month_name(m: int) -> str:
        names = ['Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                 'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
        return names[m - 1] if 1 <= m <= 12 else ''

    def _fetch(self, query, params=None):
        try:
            if hasattr(self.db, 'fetch_all'):
                return self.db.fetch_all(query, params) or []
            self.db._ensure_connection()
            with self.db._lock:
                self.db.cursor.execute(query, params or [])
                return self.db.cursor.fetchall() or []
        except Exception as e:
            logger.error(f"[IndReport] Query error: {e}", exc_info=True)
            return []

    # ------------------------------------------------------------------ #
    #  Popolamento filtri                                                  #
    # ------------------------------------------------------------------ #
    def _populate_filters(self):
        # Anni disponibili
        rows = self._fetch("""
            SELECT DISTINCT YEAR(DataRichiesta) AS yr
            FROM ind.MaterialiRichieste
            ORDER BY yr DESC
        """)
        years = [str(r[0]) for r in rows]
        all_label = self.lang.get('ind_rep_all_years', 'Tutti gli anni')
        self.year_combo['values'] = [all_label] + years
        self.year_combo.current(0)

        # Tipi materiale
        rows_t = self._fetch("""
            SELECT DISTINCT ISNULL(t.Tipo, 'Generico')
            FROM ind.Materiali m
            LEFT JOIN ind.TipoMateriali t ON m.TipoMaterialeId = t.TipoMaterialeId
            ORDER BY 1
        """)
        all_types = self.lang.get('ind_rep_all_types', 'Tutti i tipi')
        self.type_combo['values'] = [all_types] + [r[0] for r in rows_t]
        self.type_combo.current(0)

        # Stati
        rows_s = self._fetch("""
            SELECT DISTINCT Stato FROM ind.MaterialiRichieste ORDER BY Stato
        """)
        all_status = self.lang.get('ind_rep_all_status', 'Tutti gli stati')
        self.status_combo['values'] = [all_status] + [r[0] for r in rows_s]
        self.status_combo.current(0)

    # ------------------------------------------------------------------ #
    #  Caricamento dati                                                    #
    # ------------------------------------------------------------------ #
    def _build_where(self):
        """Costruisce clausola WHERE + parametri in base ai filtri UI."""
        where_parts = ['1=1']
        params = []

        # Anno
        year_val = self.year_var.get()
        all_years = self.lang.get('ind_rep_all_years', 'Tutti gli anni')
        if year_val and year_val != all_years:
            where_parts.append('YEAR(r.DataRichiesta) = ?')
            params.append(int(year_val))

        # Mese
        month_val = self.month_var.get()
        all_months = self.lang.get('ind_rep_all_months', 'Tutti i mesi')
        if month_val and month_val != all_months:
            month_num = int(month_val.split(' - ')[0])
            where_parts.append('MONTH(r.DataRichiesta) = ?')
            params.append(month_num)

        # Tipo
        type_val = self.type_var.get()
        all_types = self.lang.get('ind_rep_all_types', 'Tutti i tipi')
        if type_val and type_val != all_types:
            where_parts.append("ISNULL(t.Tipo, 'Generico') = ?")
            params.append(type_val)

        # Stato
        status_val = self.status_var.get()
        all_status = self.lang.get('ind_rep_all_status', 'Tutti gli stati')
        if status_val and status_val != all_status:
            where_parts.append('r.Stato = ?')
            params.append(status_val)

        return ' AND '.join(where_parts), params

    def _load_data(self):
        where, params = self._build_where()

        # ── Dettaglio ───────────────────────────────────────────────────
        detail_query = f"""
            SELECT r.RichiestaId,
                   r.DataRichiesta,
                   m.CodiceMateriale,
                   m.DescrizioneMateriale,
                   ISNULL(t.Tipo, 'Generico') AS Tipo,
                   r.QtaRichiesta,
                   r.Stato,
                   r.RichiestoDa,
                   ISNULL(r.PreparatoDa, '—') AS PreparatoDa,
                   r.DataPrelievo
            FROM ind.MaterialiRichieste r
            JOIN ind.Materiali m ON r.MaterialeId = m.MaterialeId
            LEFT JOIN ind.TipoMateriali t ON m.TipoMaterialeId = t.TipoMaterialeId
            WHERE {where}
            ORDER BY r.DataRichiesta DESC
        """
        detail_rows = self._fetch(detail_query, params or None)

        self.detail_tree.delete(*self.detail_tree.get_children())
        self._detail_data = []

        # DataPrelievo mancava del tutto: la colonna "Data Consegna" mostrava '—'
        # su ogni riga e l'export Excel si aspettava comunque 10 campi, quindi
        # falliva sempre con "not enough values to unpack".
        for row in detail_rows:
            rid, data_r, cod, desc, tipo, qty, stato, richiedente, prep, data_c = row
            data_str = data_r.strftime('%d/%m/%Y %H:%M') if data_r else ''
            qty_str  = f"{float(qty):.2f}" if qty is not None else '0.00'

            tag = stato or ''
            self.detail_tree.insert('', 'end', values=(
                rid, data_str, cod or '', desc or '', tipo,
                qty_str, stato or '', richiedente or '', prep or '',
                data_c.strftime('%d/%m/%Y %H:%M') if data_c else '—'
            ), tags=(tag,))
            self._detail_data.append(row)

        n_det = len(detail_rows)
        qty_tot = sum(float(r[5] or 0) for r in detail_rows)
        self.detail_total_var.set(
            self.lang.get('ind_rep_total_detail', 'Totale: {0} righe  |  Qty totale: {1:.2f}').format(n_det, qty_tot)
        )

        # ── Riepilogo mensile ────────────────────────────────────────────
        # (ri-esegui senza filtro stato per avere sempre tutti gli stati nel riepilogo)
        where_s, params_s = self._build_where_no_status()
        summary_query = f"""
            SELECT FORMAT(r.DataRichiesta, 'yyyy-MM') AS AnnoMese,
                   ISNULL(t.Tipo, 'Generico') AS Tipo,
                   COUNT(*) AS NRichieste,
                   SUM(CASE WHEN r.Stato = 'PRELEVATA' THEN 1 ELSE 0 END) AS NConsegnate,
                   SUM(r.QtaRichiesta) AS QtyRichiesta,
                   SUM(CASE WHEN r.Stato = 'PRELEVATA' THEN r.QtaRichiesta ELSE 0 END) AS QtyConsegnata
            FROM ind.MaterialiRichieste r
            JOIN ind.Materiali m ON r.MaterialeId = m.MaterialeId
            LEFT JOIN ind.TipoMateriali t ON m.TipoMaterialeId = t.TipoMaterialeId
            WHERE {where_s}
            GROUP BY FORMAT(r.DataRichiesta, 'yyyy-MM'), ISNULL(t.Tipo, 'Generico')
            ORDER BY AnnoMese DESC, Tipo
        """
        summary_rows = self._fetch(summary_query, params_s or None)

        self.summary_tree.delete(*self.summary_tree.get_children())
        self._summary_data = []

        tot_req = tot_del = tot_qty_req = tot_qty_del = 0
        for row in summary_rows:
            anno_mese, tipo, n_req, n_del, qty_req, qty_del = row
            qty_req = float(qty_req or 0)
            qty_del = float(qty_del or 0)
            pct = f"{(n_del / n_req * 100):.1f}%" if n_req else '0.0%'
            self.summary_tree.insert('', 'end', values=(
                anno_mese, tipo, n_req, n_del,
                f"{qty_req:.2f}", f"{qty_del:.2f}", pct
            ))
            self._summary_data.append(row)
            tot_req += n_req; tot_del += n_del
            tot_qty_req += qty_req; tot_qty_del += qty_del

        overall_pct = f"{(tot_del / tot_req * 100):.1f}%" if tot_req else '0.0%'
        self.summary_total_var.set(
            self.lang.get(
                'ind_rep_total_summary',
                'Totale: {0} richieste  |  {1} consegnate ({2})  |  Qty: {3:.2f} / {4:.2f}'
            ).format(tot_req, tot_del, overall_pct, tot_qty_req, tot_qty_del)
        )

        # ── Verifica acquisti ────────────────────────────────────────────
        self._load_purchases()

    # ------------------------------------------------------------------ #
    #  Verifica acquisti                                                   #
    # ------------------------------------------------------------------ #
    PURCHASES_SQL = """
        WITH ultima AS (
            SELECT MaterialeId, MAX(DataInvio) AS UltimaRichiesta
            FROM ind.RiordineEmailLog
            GROUP BY MaterialeId
        ), richiesta AS (
            SELECT l.MaterialeId, l.DataInvio, l.GiacenzaRilevata, l.LivelloMinimo
            FROM ind.RiordineEmailLog l
            JOIN ultima u ON u.MaterialeId = l.MaterialeId AND u.UltimaRichiesta = l.DataInvio
        ), conteggio AS (
            SELECT MaterialeId, COUNT(*) AS NRichieste, MIN(DataInvio) AS PrimaRichiesta
            FROM ind.RiordineEmailLog
            GROUP BY MaterialeId
        ), snapshot AS (
            -- una riga per materiale e per carico: il file Excel puo' contenere
            -- piu' righe per lo stesso codice, si sommano
            SELECT MaterialeId, DateIn, SUM(Qty) AS Qty
            FROM ind.MaterialiStock
            GROUP BY MaterialeId, DateIn
        ), serie AS (
            SELECT s.MaterialeId, s.DateIn, s.Qty,
                   LAG(s.Qty) OVER (PARTITION BY s.MaterialeId ORDER BY s.DateIn) AS QtyPrec
            FROM snapshot s
        ), ultimo_carico AS (
            SELECT MaterialeId, DateIn, Qty,
                   ROW_NUMBER() OVER (PARTITION BY MaterialeId ORDER BY DateIn DESC) AS rn
            FROM snapshot
        ), carichi_dopo AS (
            -- Solo gli AUMENTI fra un carico e il precedente: e' la merce entrata.
            -- La differenza secca fra due istanti mescolerebbe acquisti e consumi
            -- e darebbe per non arrivato un acquisto poi consumato.
            SELECT se.MaterialeId,
                   COUNT(*) AS NCarichiDopo,
                   SUM(CASE WHEN se.QtyPrec IS NOT NULL AND se.Qty > se.QtyPrec
                            THEN se.Qty - se.QtyPrec ELSE 0 END) AS QtaEntrata,
                   MAX(CASE WHEN se.QtyPrec IS NOT NULL AND se.Qty > se.QtyPrec
                            THEN se.DateIn END) AS DataEntrata
            FROM serie se
            JOIN richiesta r ON r.MaterialeId = se.MaterialeId
            WHERE se.DateIn > r.DataInvio
            GROUP BY se.MaterialeId
        )
        SELECT m.CodiceMateriale, m.DescrizioneMateriale,
               c.NRichieste, r.DataInvio, r.GiacenzaRilevata, r.LivelloMinimo,
               uc.Qty AS GiacenzaUltimoCarico, uc.DateIn AS DataUltimoCarico,
               ISNULL(cd.QtaEntrata, 0) AS QtaEntrata, cd.DataEntrata,
               ISNULL(cd.NCarichiDopo, 0) AS NCarichiDopo,
               ISNULL(t.Tipo, 'Generico') AS Tipo, c.PrimaRichiesta
        FROM richiesta r
        JOIN ind.Materiali m ON m.MaterialeId = r.MaterialeId
        JOIN conteggio c ON c.MaterialeId = r.MaterialeId
        LEFT JOIN ind.TipoMateriali t ON t.TipoMaterialeId = m.TipoMaterialeId
        LEFT JOIN ultimo_carico uc ON uc.MaterialeId = r.MaterialeId AND uc.rn = 1
        LEFT JOIN carichi_dopo cd ON cd.MaterialeId = r.MaterialeId
        WHERE {where}
        ORDER BY ISNULL(cd.QtaEntrata, 0), r.DataInvio DESC, m.CodiceMateriale
    """

    def _build_where_reorders(self):
        """Filtri della toolbar applicati alla DATA DELLA RICHIESTA di riordino.
        Lo stato (della richiesta di prelievo) qui non ha significato."""
        parts = ['1=1']
        params = []

        year_val = self.year_var.get()
        if year_val and year_val != self.lang.get('ind_rep_all_years', 'Tutti gli anni'):
            parts.append('YEAR(r.DataInvio) = ?')
            params.append(int(year_val))

        month_val = self.month_var.get()
        if month_val and month_val != self.lang.get('ind_rep_all_months', 'Tutti i mesi'):
            parts.append('MONTH(r.DataInvio) = ?')
            params.append(int(month_val.split(' - ')[0]))

        type_val = self.type_var.get()
        if type_val and type_val != self.lang.get('ind_rep_all_types', 'Tutti i tipi'):
            parts.append("ISNULL(t.Tipo, 'Generico') = ?")
            params.append(type_val)

        return ' AND '.join(parts), params

    @staticmethod
    def classify_purchase(qta_entrata, n_carichi_dopo):
        """Esito della verifica: 'ARRIVATO', 'ATTESA' o 'MANCANTE'.

        'ATTESA' non e' un fallimento: se dopo la richiesta non e' ancora stato
        caricato nessun file di stock non c'e' proprio nulla da verificare, e
        marcarlo come mancante darebbe un falso allarme ogni mattina (le email
        di riordino partono prima del carico giornaliero)."""
        if (qta_entrata or 0) > 0:
            return 'ARRIVATO'
        if not n_carichi_dopo:
            return 'ATTESA'
        return 'MANCANTE'

    def _load_purchases(self):
        where, params = self._build_where_reorders()
        rows = self._fetch(self.PURCHASES_SQL.format(where=where), params or None)

        self.purch_tree.delete(*self.purch_tree.get_children())
        self._purchases_data = []
        n_arr = n_att = n_man = 0

        for row in rows:
            (cod, desc, n_rich, data_r, giac_r, minimo, giac_c, data_c,
             entrata, data_e, n_carichi, _tipo, _prima) = row
            giac_r = float(giac_r or 0)
            giac_c = float(giac_c or 0) if giac_c is not None else None
            entrata = float(entrata or 0)
            # La differenza chiesta: giacenza dell'ultimo carico meno quella
            # registrata al momento della richiesta. Ha senso solo se l'ultimo
            # carico e' successivo alla richiesta: altrimenti si confronterebbe
            # una lettura piu' VECCHIA della richiesta e verrebbe fuori un
            # aumento inesistente (es. codice fermo da settimane nel file Excel).
            if giac_c is not None and data_c and data_r and data_c >= data_r:
                diff = giac_c - giac_r
            else:
                diff = None

            esito = self.classify_purchase(entrata, n_carichi)
            if esito == 'ARRIVATO':
                n_arr += 1
                esito_txt = self.lang.get('ind_rep_outcome_ok', 'Acquisto rilevato')
            elif esito == 'ATTESA':
                n_att += 1
                esito_txt = self.lang.get('ind_rep_outcome_wait', 'In attesa di carico')
            else:
                n_man += 1
                esito_txt = self.lang.get('ind_rep_outcome_ko', 'Nessun acquisto')

            self.purch_tree.insert('', 'end', values=(
                cod or '', desc or '', n_rich,
                data_r.strftime('%d/%m/%Y %H:%M') if data_r else '',
                f"{giac_r:.2f}", f"{float(minimo or 0):.2f}",
                f"{giac_c:.2f}" if giac_c is not None else '—',
                data_c.strftime('%d/%m/%Y %H:%M') if data_c else '—',
                f"{diff:+.2f}" if diff is not None else '—',
                f"{entrata:.2f}" if entrata else '—',
                data_e.strftime('%d/%m/%Y %H:%M') if data_e else '—',
                esito_txt,
            ), tags=(esito,))
            self._purchases_data.append(tuple(row) + (diff, esito, esito_txt))

        tot = len(rows)
        self.purch_total_var.set(
            self.lang.get(
                'ind_rep_total_purchases',
                'Codici sollecitati: {0}  |  arrivati: {1}  |  in attesa: {2}  |  senza acquisto: {3}'
            ).format(tot, n_arr, n_att, n_man)
        )

    def _build_where_no_status(self):
        """Stessa logica di _build_where ma senza filtro Stato (per il riepilogo)."""
        where_parts = ['1=1']
        params = []

        year_val = self.year_var.get()
        all_years = self.lang.get('ind_rep_all_years', 'Tutti gli anni')
        if year_val and year_val != all_years:
            where_parts.append('YEAR(r.DataRichiesta) = ?')
            params.append(int(year_val))

        month_val = self.month_var.get()
        all_months = self.lang.get('ind_rep_all_months', 'Tutti i mesi')
        if month_val and month_val != all_months:
            month_num = int(month_val.split(' - ')[0])
            where_parts.append('MONTH(r.DataRichiesta) = ?')
            params.append(month_num)

        type_val = self.type_var.get()
        all_types = self.lang.get('ind_rep_all_types', 'Tutti i tipi')
        if type_val and type_val != all_types:
            where_parts.append("ISNULL(t.Tipo, 'Generico') = ?")
            params.append(type_val)

        return ' AND '.join(where_parts), params

    # ------------------------------------------------------------------ #
    #  Export Excel                                                         #
    # ------------------------------------------------------------------ #
    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "La libreria 'openpyxl' non è installata.\nEsegui: pip install openpyxl",
                parent=self
            )
            return

        default_name = f"Report_Materiali_Indiretti_{datetime.now().strftime('%Y%m')}.xlsx"
        import os
        default_dir = r'C:\Temp'
        os.makedirs(default_dir, exist_ok=True)
        path = filedialog.asksaveasfilename(
            parent=self,
            title=self.lang.get('ind_rep_save_excel', 'Salva Report Excel'),
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialdir=default_dir,
            initialfile=default_name
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()

            # ── Foglio Dettaglio ────────────────────────────────────────
            ws_det = wb.active
            ws_det.title = 'Dettaglio'

            hdr_font = Font(bold=True, color='FFFFFF')
            hdr_fill = PatternFill('solid', fgColor='2F6DA4')

            det_headers = ['ID', 'Data Richiesta', 'Codice', 'Descrizione', 'Tipo',
                           'Qty Richiesta', 'Stato', 'Richiedente', 'Preparatore', 'Data Consegna']
            ws_det.append(det_headers)
            for cell in ws_det[1]:
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            green_fill = PatternFill('solid', fgColor='D4EDDA')
            red_fill   = PatternFill('solid', fgColor='F8D7DA')

            for row in (self._detail_data or []):
                rid, data_r, cod, desc, tipo, qty, stato, richiedente, prep, data_c = row
                ws_det.append([
                    rid,
                    data_r if data_r else '',
                    cod or '',
                    desc or '',
                    tipo or '',
                    float(qty or 0),
                    stato or '',
                    richiedente or '',
                    prep or '',
                    data_c if data_c else '',
                ])
                last = ws_det.max_row
                if stato == 'PRELEVATA':
                    for c in ws_det[last]:
                        c.fill = green_fill
                elif stato == 'ANNULLATA':
                    for c in ws_det[last]:
                        c.fill = red_fill

            # Auto-larghezza colonne
            col_widths_det = [8, 18, 14, 36, 16, 12, 14, 18, 18, 18]
            for i, w in enumerate(col_widths_det, 1):
                ws_det.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # ── Foglio Riepilogo ────────────────────────────────────────
            ws_sum = wb.create_sheet('Riepilogo Mensile')
            sum_headers = ['Anno/Mese', 'Tipo', '# Richieste', '# Consegnate',
                           'Qty Richiesta', 'Qty Consegnata', '% Consegnato']
            ws_sum.append(sum_headers)
            for cell in ws_sum[1]:
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            for row in (self._summary_data or []):
                anno_mese, tipo, n_req, n_del, qty_req, qty_del = row
                pct = round(n_del / n_req * 100, 1) if n_req else 0.0
                ws_sum.append([
                    anno_mese, tipo, n_req, n_del,
                    float(qty_req or 0), float(qty_del or 0), pct
                ])

            col_widths_sum = [12, 20, 14, 14, 16, 16, 14]
            for i, w in enumerate(col_widths_sum, 1):
                ws_sum.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # ── Foglio Verifica acquisti ────────────────────────────────
            ws_pur = wb.create_sheet('Verifica Acquisti')
            pur_headers = ['Codice', 'Descrizione', 'Tipo', '# Richieste',
                           'Prima richiesta', 'Ultima richiesta', 'Giacenza alla richiesta',
                           'Scorta minima', 'Giacenza ultimo carico', 'Data ultimo carico',
                           'Differenza', 'Entrata rilevata', 'Data entrata', 'Esito']
            ws_pur.append(pur_headers)
            for cell in ws_pur[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            yellow_fill = PatternFill('solid', fgColor='FFF3CD')
            for row in (getattr(self, '_purchases_data', None) or []):
                (cod, desc, n_rich, data_r, giac_r, minimo, giac_c, data_c,
                 entrata, data_e, _n_carichi, tipo, prima, diff, esito, esito_txt) = row
                ws_pur.append([
                    cod or '', desc or '', tipo or '', n_rich,
                    prima if prima else '', data_r if data_r else '',
                    float(giac_r or 0), float(minimo or 0),
                    float(giac_c) if giac_c is not None else None,
                    data_c if data_c else '',
                    float(diff) if diff is not None else None,
                    float(entrata or 0), data_e if data_e else '',
                    esito_txt,
                ])
                last = ws_pur.max_row
                fill = {'ARRIVATO': green_fill, 'MANCANTE': red_fill,
                        'ATTESA': yellow_fill}.get(esito)
                if fill:
                    for c in ws_pur[last]:
                        c.fill = fill

            col_widths_pur = [14, 36, 16, 12, 18, 18, 20, 14, 20, 18, 12, 16, 18, 20]
            for i, w in enumerate(col_widths_pur, 1):
                ws_pur.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws_pur.freeze_panes = 'A2'

            wb.save(path)
            logger.info(f"[IndReport] Export Excel: {path}")
            if messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                self.lang.get('ind_rep_export_ok', 'Report Excel salvato:\n{0}\n\nAprire il file?').format(path),
                parent=self
            ):
                import subprocess
                subprocess.Popen(['start', '', path], shell=True)

        except Exception as e:
            logger.error(f"[IndReport] Export Excel error: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore export Excel:\n{e}",
                parent=self
            )


# ─────────────────────────────────────────────────────────────────────────────
#  Entry-point
# ─────────────────────────────────────────────────────────────────────────────
def open_indirect_materials_report(master, db, lang):
    """Entry-point richiamabile da main.py."""
    IndirectMaterialsReportWindow(master, db, lang)
