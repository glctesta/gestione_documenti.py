#import configparser
# --- StdIO safeguard + Faulthandler sicuro per exe windowed ---
import shutil
import sys, os, atexit
import urllib.parse
from pathlib import Path
import tkinter as tk

#Verifica che non sia eseguito sul server
# --- Configurazione IP BLOCCATI ---
# Solo 1-2 IP da bloccare, niente configurazioni complesse
NOT_ALLOWED_IPS = [
    '192.168.10.110',  # IP specifico da bloccare
    '203.0.113.25'  # Altro IP da bloccare
]

# --- Verifica esecuzione locale ---
try:
    # Importa il modulo di verifica
    from local_execution_check import is_local_execution_simple

    # Esegui la verifica con IP bloccati
    is_local_execution_simple(not_allowed_ips=NOT_ALLOWED_IPS)

except ImportError:
    # Fallback alla versione semplice
    try:
        from simple_local_check import verify_simple

        verify_simple(not_allowed_ips=NOT_ALLOWED_IPS)
    except ImportError:
        print("AVVISO: Moduli di sicurezza non trovati")
        print("Continuo senza verifiche...")

except Exception as e:
    print(f"ERRORE nella verifica sicurezza: {e}")
    sys.exit(1)

#Fine controllo esecuzione verifica IP non permessi

if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')
_STDOUT_FILE = None
_STDERR_FILE = None


try:
    base = Path(os.getenv("LOCALAPPDATA", ".")) / "TraceabilityRS" / "logs"
    base.mkdir(parents=True, exist_ok=True)

    if getattr(sys, "stdout", None) is None:
        _STDOUT_FILE = open(base / "stdout.log", "w", buffering=1, encoding="utf-8")
        sys.stdout = _STDOUT_FILE

    if getattr(sys, "stderr", None) is None:
        _STDERR_FILE = open(base / "stderr.log", "w", buffering=1, encoding="utf-8")
        sys.stderr = _STDERR_FILE
except Exception:
    pass

@atexit.register
def _close_stdio_files():
    for f in (_STDOUT_FILE, _STDERR_FILE):
        try:
            if f: f.close()
        except Exception:
            pass

# Faulthandler: opzionale e sicuro (usa un file dedicato)
try:
    import faulthandler
    if os.getenv("ENABLE_FAULTHANDLER", "0") == "1":
        _FH = open(base / "faulthandler.log", "w", buffering=1, encoding="utf-8")
        faulthandler.enable(file=_FH)
        t = int(os.getenv("FH_DUMP_TIMEOUT", "0"))
        if t > 0:
            faulthandler.dump_traceback_later(t, repeat=True, file=_FH)

        @atexit.register
        def _close_fh():
            try:
                _FH.close()
            except Exception:
                pass
except Exception:
    pass

# --- Logging robusto (file + console solo se esiste) ---
import logging
from logging.handlers import RotatingFileHandler

def setup_logging(debug: bool = False,
                  log_dir: str | None = None,
                  logfile_name: str = "traceability_rs.log",
                  logger_name: str = "TraceabilityRS") -> str:
    root_logger = logging.getLogger()
    if root_logger.handlers:
        # giÃ  configurato
        for h in root_logger.handlers:
            if hasattr(h, "baseFilename"):
                return h.baseFilename
        return ""

    level = logging.DEBUG if debug or os.getenv("TRACE_RS_DEBUG") == "1" else logging.INFO
    root_logger.setLevel(level)
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s", "%Y-%m-%d %H:%M:%S")

    if not log_dir:
        log_dir = os.path.join(os.getenv("LOCALAPPDATA", "."), "TraceabilityRS", "logs")
    os.makedirs(log_dir, exist_ok=True)
    file_path = os.path.join(log_dir, logfile_name)

    fh = RotatingFileHandler(file_path, maxBytes=5*1024*1024, backupCount=5, encoding="utf-8")
    fh.setLevel(level)
    fh.setFormatter(fmt)
    root_logger.addHandler(fh)

    if getattr(sys, "stdout", None):
        ch = logging.StreamHandler(sys.stdout)
        ch.setLevel(level)
        ch.setFormatter(fmt)
        root_logger.addHandler(ch)

    app_logger = logging.getLogger(logger_name)
    app_logger.setLevel(level)
    app_logger.debug("Logging inizializzato. Livello=%s, file=%s", logging.getLevelName(level), file_path)
    return file_path

LOG_FILE_PATH = setup_logging(debug=False, logger_name="TraceabilityRS")
logger = logging.getLogger("TraceabilityRS")
logger.info("Logging avviato. File: %s", LOG_FILE_PATH)

import re
import subprocess
from sqlalchemy import create_engine, text
from sqlalchemy.pool import QueuePool
from collections import defaultdict
from datetime import datetime, timedelta
from tkinter import ttk, messagebox, filedialog, simpledialog
import random
import pyodbc
from PIL import ImageOps, ImageDraw, ImageFont
from packaging import version
from tkcalendar import DateEntry
import pandas as pd
import general_docs_gui
import maintenance_gui
import fixtures_report_window
import materials_gui
import operations_gui
import permissions_gui
import translations_manager
import printer_config_manager

import submissions_gui
import tools_gui
from traceability import TraceabilityManager
from calibration_gui import CalibrationsWindow
import fct_transfer
import collections.abc
import scarti_gui
import scrap_reports_gui
import coating_gui
import product_checks_gui
import guests_gui
import guests_report_generator
import tempfile
import assign_submissions_gui
import utils
import submissions_management_gui
import json, socket
import threading
import time
import scrap_validation_gui
from add_complaint import AddComplaintWindow
from business_days import should_send_notification
from npi.npi_manager import GestoreNPI
from npi.windows.dashboard_window import NpiDashboardWindow
from npi.windows.gantt_window import NpiGanttWindow
from npi.windows.config_window import NpiConfigWindow
from npi.windows.project_window import ProjectWindow
from typing import Optional, List, Dict, Tuple

def _detect_log_file_path(logger_name: str) -> str:
    """
    Ritorna il percorso del primo FileHandler/RotatingFileHandler trovato
    sul logger specificato o, in fallback, sul root logger.
    """
    def _first_file_handler_path(lg: logging.Logger) -> str:
        for h in lg.handlers:
            if hasattr(h, "baseFilename"):
                try:
                    return str(Path(h.baseFilename).resolve())
                except Exception:
                    return str(h.baseFilename)
        return ""

    lg_app = logging.getLogger(logger_name)
    p = _first_file_handler_path(lg_app)
    if p:
        return p
    return _first_file_handler_path(logging.getLogger())


def setup_logging(debug: bool = False,
                  log_dir: str | None = None,
                  logfile_name: str = "traceability_rs.log",
                  logger_name: str = "TraceabilityRS") -> str:
  
    root_logger = logging.getLogger()

    # Se giÃ  configurato, non duplicare; prova a restituire il file esistente
    if root_logger.handlers:
        existing = _detect_log_file_path(logger_name)
        if existing:
            return existing
        # se non troviamo un file handler, proseguiamo a configurarne uno

    level = logging.DEBUG if debug or os.getenv("TRACE_RS_DEBUG") == "1" else logging.INFO
    root_logger.setLevel(level)

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # File handler
    if not log_dir:
        log_dir = os.path.join(os.getenv("LOCALAPPDATA", "."), "TraceabilityRS", "logs")
    os.makedirs(log_dir, exist_ok=True)
    file_path = os.path.join(log_dir, logfile_name)

    fh = RotatingFileHandler(file_path, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8")
    fh.setLevel(level)
    fh.setFormatter(formatter)
    root_logger.addHandler(fh)

    # Console handler (solo se câ€™Ã¨ stdout)
    stdout = getattr(sys, "stdout", None)
    if stdout:
        ch = logging.StreamHandler(stdout)
        ch.setLevel(level)
        ch.setFormatter(formatter)
        root_logger.addHandler(ch)

    # Imposta anche il logger applicativo al livello scelto
    app_logger = logging.getLogger(logger_name)
    app_logger.setLevel(level)
    app_logger.debug("Logging inizializzato. Livello=%s, file=%s",
                     logging.getLevelName(level), file_path)

    return str(Path(file_path).resolve())


LOG_FILE_PATH = setup_logging(debug=False, logger_name="TraceabilityRS")
logger = logging.getLogger("TraceabilityRS")
logger.info("Logging avviato. File: %s", LOG_FILE_PATH)

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE: bool = True
except ImportError:
    PIL_AVAILABLE = False


# --- CONFIGURAZIONE APPLICAZIONE ---
APP_VERSION = '2.3.2.7'  # Versione aggiornata
APP_DEVELOPER = 'Gianluca Testa'

# # --- CONFIGURAZIONE DATABASE ---
DB_DRIVER = '{SQL Server Native Client 11.0}'
DB_SERVER = 'roghipsql01.vandewiele.local\\emsreset'
DB_DATABASE = 'Traceability_rs'
DB_UID = "emsreset"
DB_PWD = 'E6QhqKUxHFXTbkB7eA8c9ya'
DB_CONN_STR = (f'DRIVER={DB_DRIVER};SERVER={DB_SERVER};DATABASE={DB_DATABASE};'
               f'UID={DB_UID};PWD={DB_PWD};MARS_Connection=Yes;TrustServerCertificate=Yes')


def is_update_needed(current_ver_str, db_ver_str):
    """Confronta due stringhe di versione (es. '1.4.0') in modo sicuro."""
    try:
        return version.parse(db_ver_str) > version.parse(current_ver_str)
    except Exception:
        # Fallback a un confronto di stringhe semplice in caso di errore
        return db_ver_str > current_ver_str


def get_update_skip_file_path():
    """Restituisce il percorso del file JSON per tracciare i rinvii dell'update."""
    # Salva il file nella directory di lavoro del programma
    work_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(work_dir, "update_skip_count.json")


def load_update_skip_count():
    """Carica il conteggio dei rinvii dell'update dal file JSON."""
    skip_file = get_update_skip_file_path()
    try:
        if os.path.exists(skip_file):
            with open(skip_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('skip_count', 0), data.get('last_version', '')
        return 0, ''
    except Exception as e:
        logger.warning(f"Errore nel caricamento del file di rinvio update: {e}")
        return 0, ''


def save_update_skip_count(skip_count, version_str):
    """Salva il conteggio dei rinvii dell'update nel file JSON."""
    skip_file = get_update_skip_file_path()
    try:
        data = {
            'skip_count': skip_count,
            'last_version': version_str,
            'last_skip_date': datetime.now().isoformat()
        }
        with open(skip_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        logger.info(f"Salvato conteggio rinvii update: {skip_count} per versione {version_str}")
    except Exception as e:
        logger.error(f"Errore nel salvataggio del file di rinvio update: {e}")


def reset_update_skip_count():
    """Resetta il conteggio dei rinvii dell'update eliminando il file JSON."""
    skip_file = get_update_skip_file_path()
    try:
        if os.path.exists(skip_file):
            os.remove(skip_file)
            logger.info("File di rinvio update eliminato (reset conteggio)")
    except Exception as e:
        logger.error(f"Errore nell'eliminazione del file di rinvio update: {e}")


def fetch_working_areas(self):
    """Recupera le Aree di Lavoro principali."""
    query = "SELECT [WorkingAreaID], [AreaName] FROM [ResetServices].[BreakDown].[WorkingAreas] ORDER BY AreaName;"
    try:
        self.cursor.execute(query)
        return self.cursor.fetchall()
    except pyodbc.Error as e:
        self.last_error_details = str(e); return []

def fetch_working_sub_areas(self, working_area_id):
    """Recupera le Sotto-Aree in base all'Area di Lavoro selezionata."""
    query = """
        SELECT [WorkingSubAreaID], [AreaSubName] 
        FROM [ResetServices].[BreakDown].[WorkingSubAreas]
        WHERE [WorkingAreaID] = ? ORDER BY AreaSubName;
    """
    try:
        self.cursor.execute(query, working_area_id)
        return self.cursor.fetchall()
    except pyodbc.Error as e:
        self.last_error_details = str(e); return []

def fetch_working_lines(self, working_area_id, sub_area_id):
    """Recupera le Linee in base all'Area e Sotto-Area selezionate."""
    query = """
        SELECT DISTINCT wl.WorkingLineID, WL.WorkingLineName 
        FROM [ResetServices].[BreakDown].[WorkingAreas] AS WA 
        INNER JOIN [ResetServices].[BreakDown].WorkingSubAreas WSA ON WA.WorkingAreaID = WSA.WorkingAreaID 
        INNER JOIN [ResetServices].[BreakDown].WorkingLines AS WL ON WSA.WorkingSubAreaID = WL.WorkingSubAreaID 
        WHERE WA.WorkingAreaID = ? AND WSA.workingsubareaid = ? 
        ORDER BY wl.WorkingLineName;
    """
    try:
        self.cursor.execute(query, working_area_id, sub_area_id)
        return self.cursor.fetchall()
    except pyodbc.Error as e:
        self.last_error_details = str(e); return []

def fetch_production_orders_for_breakdown(self):
    """Recupera gli ordini di produzione per la selezione."""
    query = """
        SELECT DISTINCT 
            o.idorder as idordine, 
            o.OrderNumber + ' [' + pf.ProductCode +']' as OrderNumber 
        FROM Traceability_RS.dbo.orders as o 
        INNER JOIN traceability_rs.dbo.products as pf ON pf.IDProduct = O.IDProduct
        LEFT JOIN ResetServices.DBO.TBORDINI RO ON ro.IdPOTrace = o.IDOrder
        LEFT JOIN resetservices.dbo.tbregistro r ON ro.idregistro = r.contatore 
            AND r.idregistro IN (21,26)
        WHERE CAST(O.DataInserted as date) >= '2025-08-01' 
            AND (ro.IdOrdine IS NULL 
                OR NOT EXISTS (
                    SELECT 1
                    FROM resetservices.dbo.TbSubOrdine s 
                    LEFT JOIN resetservices.dbo.TbFattStory fs ON fs.IdPoSub = s.IdOrdStori
                    WHERE s.idordine = ro.IdOrdine
                    GROUP BY s.IdOrdStori, s.QtaStory
                    HAVING s.QtaStory > ISNULL(SUM(fs.QtaFaturata), 0)
                )
            and (ro.idordine is NULL 
                OR NOT EXISTS (
                    SELECT 1 
                    FROM resetservices.dbo.TbSubOrdine inner join resetservices.dbo.tbprodfin on tbsubordine.idpf = 
                    tbprodfin.idpf inner join resetservices.dbo.TbProdFinStuff on TbProdFinStuff.Idpf =tbprodfin.idpf
                    )
            ))
    """
    try:
        self.cursor.execute(query)
        return self.cursor.fetchall()
    except pyodbc.Error as e:
        self.last_error_details = str(e); return []

def fetch_issue_areas(self):
    """Recupera le aree problematiche (es. Meccanica, Elettrica)."""
    query = "SELECT [IssueAreaId], [IssueArea] FROM [ResetServices].[BreakDown].[IssuesAreas] ORDER BY [IssueArea];"
    try:
        self.cursor.execute(query)
        return self.cursor.fetchall()
    except pyodbc.Error as e:
        self.last_error_details = str(e); return []

class KanbanRulesManagementForm(tk.Toplevel):
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager

        self.title(self.lang.get('rules_mgmt_title', "KanBan - Gestione regole"))
        self.geometry("640x460")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # Stato form
        self._current_rule_id = None
        self.rule_type_var = tk.StringVar(value="percent")  # 'percent' | 'qty'
        self.rule_value_var = tk.StringVar()

        self._logo_imgtk = None

        self._build_ui()
        self._load_rules()

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # Lista regole
        list_frame = ttk.Labelframe(root, text=self.lang.get('rules_list', "Regole"))
        list_frame.pack(fill="both", expand=True)

        cols = ("id", "type", "value", "status", "dateout")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=10)
        self.tree.heading("id", text="ID")
        self.tree.column("id", width=60, anchor="center")
        self.tree.heading("type", text=self.lang.get('rule_type', "Tipo"))
        self.tree.column("type", width=140)
        self.tree.heading("value", text=self.lang.get('rule_value', "Valore"))
        self.tree.column("value", width=100, anchor="e")
        self.tree.heading("status", text=self.lang.get('rule_status', "Stato"))
        self.tree.column("status", width=100)
        self.tree.heading("dateout", text="DateOut")
        self.tree.column("dateout", width=140)
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # Editor regola
        edit = ttk.Labelframe(root, text=self.lang.get('rule_editor', "Editor regola"))
        edit.pack(fill="x", pady=(10, 0))

        r1 = ttk.Frame(edit); r1.pack(fill="x", padx=8, pady=6)
        ttk.Label(r1, text=self.lang.get('rule_type', "Tipo")).pack(side="left")

        ttk.Radiobutton(r1, text=self.lang.get('rule_percent', "Percentuale"), value="percent",
                        variable=self.rule_type_var).pack(side="left", padx=(8, 8))
        ttk.Radiobutton(r1, text=self.lang.get('rule_quantity', "QuantitÃ "), value="qty",
                        variable=self.rule_type_var).pack(side="left", padx=(0, 8))

        r2 = ttk.Frame(edit); r2.pack(fill="x", padx=8, pady=6)
        ttk.Label(r2, text=self.lang.get('rule_value', "Valore")).pack(side="left")
        self.value_entry = ttk.Entry(r2, textvariable=self.rule_value_var, width=12)
        self.value_entry.pack(side="left", padx=(8, 8))
        ttk.Label(r2, text=self.lang.get('rule_value_hint', "(% intero 1..100 o quantitÃ  > 0)")).pack(side="left")

        # Pulsanti azione
        btns = ttk.Frame(root); btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text=self.lang.get('button_new', "Nuova"), command=self._on_new).pack(side="left")
        ttk.Button(btns, text=self.lang.get('button_save', "Salva"), command=self._on_save).pack(side="left", padx=(8,0))
        ttk.Button(btns, text=self.lang.get('button_delete', "Elimina"), command=self._on_delete).pack(side="left", padx=(8,0))
        ttk.Button(btns, text=self.lang.get('button_clear', "Pulisci"), command=self._on_clear).pack(side="left", padx=(8,0))

        # Logo in basso a destra
        bottom = ttk.Frame(root); bottom.pack(fill="both", expand=True)
        self.logo_lbl = ttk.Label(bottom)
        self.logo_lbl.pack(side="right", anchor="se", padx=4, pady=4)
        self._load_logo()

    def _load_logo(self):
        try:
            if os.path.exists("logo.png"):
                img = Image.open("logo.png")
                img.thumbnail((160, 80))
                self._logo_imgtk = ImageTk.PhotoImage(img)
                self.logo_lbl.configure(image=self._logo_imgtk)
        except Exception:
            pass

    def _load_rules(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        rows = self.db.fetch_all_kanban_rules()
        for r in rows:
            is_pct = getattr(r, "MinimumProcent", None) is not None
            value = int(getattr(r, "MinimumProcent", 0) if is_pct else getattr(r, "MinimumQty", 0))
            type_txt = self.lang.get('rule_percent', "Percentuale") if is_pct else self.lang.get('rule_quantity', "QuantitÃ ")
            status_txt = self.lang.get('rule_status_active', "Attiva") if getattr(r, "DateOut", None) is None else self.lang.get('rule_status_closed', "Chiusa")
            dateout_txt = "" if getattr(r, "DateOut", None) is None else str(getattr(r, "DateOut"))
            self.tree.insert("", "end", iid=str(r.KanBanRuleID),
                             values=(r.KanBanRuleID, type_txt, value, status_txt, dateout_txt))

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        rid = int(sel[0])
        # recupera record dalla lista corrente
        rows = self.db.fetch_all_kanban_rules()
        the = next((x for x in rows if x.KanBanRuleID == rid), None)
        if not the:
            return
        self._current_rule_id = rid
        is_pct = getattr(the, "MinimumProcent", None) is not None
        self.rule_type_var.set("percent" if is_pct else "qty")
        val = int(getattr(the, "MinimumProcent", 0) if is_pct else getattr(the, "MinimumQty", 0))
        self.rule_value_var.set(str(val))

    def _validate(self):
        typ = self.rule_type_var.get()
        sval = self.rule_value_var.get().strip()
        if not sval.isdigit():
            return False, self.lang.get('rules_err_numeric', "Inserire un numero intero positivo.")
        n = int(sval)
        if typ == "percent":
            if n < 1 or n > 100:
                return False, self.lang.get('rules_err_percent_range', "La percentuale deve essere tra 1 e 100.")
            return True, ("percent", n)
        else:
            if n <= 0:
                return False, self.lang.get('rules_err_qty_positive', "La quantitÃ  deve essere > 0.")
            return True, ("qty", n)

    def _on_new(self):
        self._current_rule_id = None
        self.rule_type_var.set("percent")
        self.rule_value_var.set("")
        self.tree.selection_remove(self.tree.selection())

    def _on_clear(self):
        self._on_new()

    def _on_save(self):
        ok, res = self._validate()
        if not ok:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), res, parent=self)
            return
        typ, val = res
        min_pct = val if typ == "percent" else None
        min_qty = val if typ == "qty" else None

        if self._current_rule_id is None:
            ok, err = self.db.add_kanban_rule(min_pct, min_qty)
        else:
            ok, err = self.db.update_kanban_rule(self._current_rule_id, min_pct, min_qty)

        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('rules_saved', "Regola salvata."), parent=self)
            self._load_rules()
            if self._current_rule_id is None:
                self._on_new()
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('rules_save_err', f"Errore salvataggio: {err or self.db.last_error_details}"),
                                 parent=self)

    def _on_delete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('rules_select_row', "Seleziona una regola."), parent=self)
            return
        rid = int(sel[0])
        if not messagebox.askyesno(self.lang.get('confirm_title', "Conferma"),
                                   self.lang.get('confirm_delete_rule', "Eliminare (chiudere) la regola selezionata?"),
                                   parent=self):
            return
        ok, err = self.db.soft_delete_kanban_rule(rid)
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('rules_deleted', "Regola eliminata."), parent=self)
            self._load_rules()
            self._on_new()
        else:
            # giÃ  chiusa o altro errore
            msg = self.lang.get('rules_already_closed', "La regola Ã¨ giÃ  chiusa o non esiste.") if err == "already_closed_or_not_found" else (err or self.db.last_error_details)
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('rules_delete_err', f"Errore eliminazione: {msg}"),
                                 parent=self)

class LanguageManager:
    """Gestisce le traduzioni e la lingua corrente dell'applicazione."""

    def __init__(self, db_handler):
        self.db = db_handler
        self.translations = defaultdict(dict)
        self.current_language = 'it'  # Lingua predefinita
        self.load_translations()

    def load_translations(self):
        """Carica le traduzioni dal database."""
        records = self.db.fetch_translations()
        if not records:
            # Usiamo print inizialmente, messagebox potrebbe non essere disponibile se la GUI principale fallisce
            print("âš ï¸ Traduzioni Mancanti: Nessuna traduzione trovata nel database. VerrÃ  usato il testo di default.")
            logger.warning("Nessuna traduzione trovata nel database")
            return
        
        # DEBUG: Conta traduzioni per lingua
        lang_counts = {}
        for lang_code, key, value in records:
            lang_lower = lang_code.lower()
            self.translations[lang_lower][key] = value
            lang_counts[lang_lower] = lang_counts.get(lang_lower, 0) + 1
        
        # DEBUG: Stampa riepilogo
        print(f"\nâœ… Traduzioni caricate dal database:")
        for lang, count in sorted(lang_counts.items()):
            print(f"   {lang.upper()}: {count} traduzioni")
        print(f"   Lingua corrente: {self.current_language.upper()}")
        
        # DEBUG: Mostra alcune chiavi IT per verifica
        if 'it' in self.translations:
            sample_keys = ['project_window_title', 'col_task', 'status_todo']
            print(f"\nðŸ” Verifica chiavi IT:")
            for key in sample_keys:
                value = self.translations['it'].get(key, 'NON TROVATA')
                print(f"   {key}: {value}")
        
        logger.info(f"Caricate {sum(lang_counts.values())} traduzioni per {len(lang_counts)} lingue")

    def get(self, key, *args):
        """Restituisce la traduzione per una data chiave nella lingua corrente."""
        translated_text = self.translations[self.current_language].get(key, key)
        
        # DEBUG: Log se la traduzione non viene trovata
        if translated_text == key and key not in ['', ' ']:
            logger.debug(f"Traduzione non trovata per '{key}' in lingua '{self.current_language}'")
            # Verifica se esiste in altre lingue
            found_in = [lang for lang, trans in self.translations.items() if key in trans]
            if found_in:
                logger.debug(f"  -> Chiave '{key}' trovata in: {found_in}")
        
        if args:
            try:
                return translated_text.format(*args)
            except (IndexError, KeyError):
                return translated_text
        return translated_text

    def get_raw(self, key):
        """Restituisce il template di traduzione senza formattazione."""
        return self.translations[self.current_language].get(key, key)

    def set_language(self, lang_code):
        """Imposta la lingua corrente."""
        self.current_language = lang_code.lower()

class Database:
    """Gestisce la connessione e le operazioni sul database."""

    def __init__(self, conn_str):
        self.conn_str = conn_str
        self.conn = None
        self.cursor = None
        self.engine = None
        self.last_error_details = ""
        self._lock = threading.RLock()

    def connect(self):
        with self._lock:
            try:
                # Usiamo autocommit=False per gestire le transazioni manualmente (commit/rollback)
                self.conn = pyodbc.connect(self.conn_str, autocommit=False)
                self.cursor = self.conn.cursor()

                def get_existing_connection():
                    return self.conn

                # Creiamo l'Engine SQLAlchemy
                self.engine = create_engine(
                    "mssql+pyodbc://",
                    creator=get_existing_connection
                )

                self.engine.connect().close()
                self.npi_engine = self._create_npi_engine()

                return True
            except pyodbc.Error as ex:
                self.last_error_details = str(ex)
                logger.error(f"Database Connection Error: {ex}")
                return False

    def _create_npi_engine(self):
        """
        Crea un engine SQLAlchemy separato con connection pooling
        per il NPI Manager.
        """
        try:
            # Costruisci la connection string per SQLAlchemy
            params = urllib.parse.quote_plus(self.conn_str)
            connection_url = f"mssql+pyodbc:///?odbc_connect={params}"

            # Crea engine con pooling e pre_ping
            npi_engine = create_engine(
                connection_url,
                poolclass=QueuePool,
                pool_size=5,  # Numero di connessioni nel pool
                max_overflow=10,  # Connessioni extra se necessario
                pool_recycle=3600,  # Ricicla connessioni dopo 1 ora
                pool_pre_ping=True,  # â­ Testa connessioni prima dell'uso
                pool_timeout=30,  # Timeout per ottenere connessione
                echo=False  # Non loggare SQL (cambia in True per debug)
            )

            # Test della connessione
            with npi_engine.connect() as conn:
                conn.execute(text("SELECT 1"))

            logger.info("NPI Engine con pooling creato con successo")
            return npi_engine

        except Exception as e:
            logger.error(f"Errore nella creazione del NPI Engine: {e}")
            raise

    def disconnect(self):
        """Closes the cursor and connection safely, preventing errors if called multiple times."""
        if self.cursor:
            self.cursor.close()
            self.cursor = None  # Set to None after closing
        if self.conn:
            self.conn.close()
            self.conn = None  # Set to None after closing
        if self.npi_engine:
            self.npi_engine.dispose()
            self.npi_engine = None

    def _clear_cursor_state(self):
        """
        Pulisce lo stato del cursore consumando tutti i result set pendenti.
        Previene errori 'Function sequence error' (HY010).
        """
        with self._lock:
            if not self.cursor or not self.conn:
                return False
            
            try:
                # Consuma tutti i result set pendenti
                while self.cursor.nextset():
                    pass
                return True
            except Exception as e:
                # Se fallisce, prova a ricreare il cursore
                try:
                    logger.debug(f"Ricreazione cursore dopo errore: {e}")
                    self.cursor.close()
                    self.cursor = self.conn.cursor()
                    return True
                except Exception as e2:
                    logger.error(f"Impossibile ricreare il cursore: {e2}")
                    return False

        # NUOVO METODO: Cerca documenti esistenti attivi che corrispondono ai parametri

    def mark_document_out_of_validation(self, document_id):
        """Marca un documento come fuori validazione aggiornando DateOutOfValidation"""
        try:
            from datetime import datetime

            query = """
                    UPDATE [Traceability_RS].[dbo].[ProductDocuments]
                    SET
                        DateOutOfValidation = ?, Validated = 0
                    WHERE DocumentProductionID = ? \
                    """

            params = (datetime.now(), document_id)

            self.cursor.execute(query, params)
            self.conn.commit()

            logger.info(
                "Document %s marked out of validation",
                document_id
            )
            return True

        except Exception as e:
            self.last_error_details = str(e)
            logger.error("Error marking document out of validation: %s", e)
            return False

    def get_calibration_expired(self):
        """Recupera equipaggiamenti con calibrazione scaduta o senza calibrazione"""
        query = """
                SELECT ROW_NUMBER()                            OVER (ORDER BY c.ExpireOn DESC) as Nr, e.InternalName + \
                                                                                                      ' [#SN: ' + \
                                                                                                      ISNULL(e.SerialNumber, '') + \
                                                                                                      '] ' + \
                                                                                                      p.ParentPhaseName as Equipment, \
                       c.CalibratedOn, \
                       c.ExpireOn, \
                       DATEDIFF(DAY, GETDATE(), c.ExpireOn) AS DaysToExpire, \
                       e.MustCalibrated
                FROM eqp.Equipments e
                         LEFT JOIN eqp.Calibrations c ON c.EquipmentID = e.EquipmentID
                         INNER JOIN parentphases p ON p.IDParentPhase = e.ParentPhaseId
                WHERE e.MustCalibrated = 1
                  AND (
                    DATEDIFF(DAY, GETDATE(), c.ExpireOn) <= 0 -- Calibrazioni scadute
                        OR c.EquipmentID IS NULL -- Equipaggiamenti senza calibrazione
                    )
                ORDER BY c.ExpireOn DESC, e.InternalName \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch calibrazioni scadute: {e}")
            return []

    def get_calibration_expiring_30days(self):
        """Recupera equipaggiamenti con calibrazione in scadenza entro 30 giorni"""
        query = """
                SELECT ROW_NUMBER()                            OVER (ORDER BY c.ExpireOn) as Nr, e.InternalName + \
                                                                                                 ' [#SN: ' + \
                                                                                                 ISNULL(e.SerialNumber, '') + \
                                                                                                 '] ' + \
                                                                                                 p.ParentPhaseName as Equipment, \
                       c.CalibratedOn, \
                       c.ExpireOn, \
                       DATEDIFF(DAY, GETDATE(), c.ExpireOn) AS DaysToExpire
                FROM eqp.Equipments e
                         INNER JOIN eqp.Calibrations c ON c.EquipmentID = e.EquipmentID
                         INNER JOIN parentphases p ON p.IDParentPhase = e.ParentPhaseId
                WHERE e.MustCalibrated = 1
                  AND DATEDIFF(DAY, GETDATE(), c.ExpireOn) BETWEEN 1 AND 30
                ORDER BY c.ExpireOn, e.InternalName \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch calibrazioni in scadenza 30 giorni: {e}")
            return []

    def get_calibration_expiring_over_30days(self):
        """Recupera equipaggiamenti con calibrazione in scadenza oltre 30 giorni"""
        query = """
                SELECT ROW_NUMBER()                            OVER (ORDER BY c.ExpireOn) as Nr, e.InternalName + \
                                                                                                 ' [#SN: ' + \
                                                                                                 ISNULL(e.SerialNumber, '') + \
                                                                                                 '] ' + \
                                                                                                 p.ParentPhaseName as Equipment, \
                       c.CalibratedOn, \
                       c.ExpireOn, \
                       DATEDIFF(DAY, GETDATE(), c.ExpireOn) AS DaysToExpire
                FROM eqp.Equipments e
                         INNER JOIN eqp.Calibrations c ON c.EquipmentID = e.EquipmentID
                         INNER JOIN parentphases p ON p.IDParentPhase = e.ParentPhaseId
                WHERE e.MustCalibrated = 1
                  AND DATEDIFF(DAY, GETDATE(), c.ExpireOn) > 30
                ORDER BY c.ExpireOn, e.InternalName \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch calibrazioni oltre 30 giorni: {e}")
            return []

    def get_calibration_check_interval(self):
        """Recupera l'intervallo di controllo calibrazioni dalle impostazioni"""
        query = "SELECT Value FROM settings WHERE atribute = 'Sys_Check_Calibration_time'"
        try:
            self.cursor.execute(query)
            result = self.cursor.fetchone()
            return int(result[0]) if result and result[0] else 7  # Default 7 giorni
        except pyodbc.Error as e:
            logger.error(f"Errore fetch intervallo calibrazioni: {e}")
            return 7

    def get_calibration_emails(self):
        """Recupera gli indirizzi email per le notifiche calibrazioni"""
        query = "SELECT Value FROM settings WHERE atribute = 'Sys_Check_Calibration_Emails'"
        try:
            self.cursor.execute(query)
            result = self.cursor.fetchone()
            if result and result[0]:
                emails = [email.strip() for email in result[0].split(';') if email.strip()]
                return emails
            return []
        except pyodbc.Error as e:
            logger.error(f"Errore fetch email calibrazioni: {e}")
            return []

    def fetch_products_for_checks(self):
        """Recupera prodotti per combo gestione verifiche"""
        query = """
                SELECT p.IDProduct, productcode, productname
                FROM traceability_rs.dbo.products AS P
                         LEFT JOIN traceability_rs.dbo.PeriodicalProductChecks AS PC
                                   ON p.idproduct = pc.idproduct AND pc.datestop IS NULL
                WHERE CHARINDEX('MICR', ProductName, 1) = 0
                ORDER BY ProductCode; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_scrap_declarations_pending(self):
        """Recupera le dichiarazioni di scrap in attesa di validazione"""
        query = """
                 SELECT s.ScrapDeclarationId,
                   s.[User] as DECLAREDBY,
                   FORMAT(s.DateIn, 'dd/MM/yyyy') as [Date],
                o.OrderNumber,
                l.labelcod,
                p.productCode as Produc,
                A.AreaName,
                d.Reason as Defect,
                s.Riferiments,  -- Campo aggiunto per la stored procedure
                1 as Qty,
                s.Picture,  -- Aggiunto il campo Picture
                de.IDDefect  -- Campo IDDefect per la stored procedure (NON VISUALIZZATO NELLA GUI)
            FROM [Traceability_RS].[dbo].ScarpDeclarations S
                INNER JOIN Traceability_RS.dbo.LabelCodes L
            ON l.IDLabelCode = s.IdLabelCode
                INNER JOIN [Traceability_RS].dbo.Areas A ON a.IDArea = s.IDParentPhase
                INNER JOIN [Traceability_RS].dbo.ScrapResons D ON d.ScrapReasonId = s.ScrapReasonId
                INNER JOIN [Traceability_RS].dbo.defects de ON de.IDDefect = d.ScrapDeclarationId  -- CORREZIONE: de.IDDefect = d.IDDefect
                INNER JOIN [Traceability_RS].dbo.boards B ON l.IDBoard = b.IDBoard
                INNER JOIN [Traceability_RS].dbo.orders o ON o.idorder = b.IDOrder
                inner join traceability_rs.dbo.products P on p.idproduct=o.idproduct
            WHERE (s.Accepted IS NULL
               OR s.Accepted = 0)
              AND (s.Refuzed IS NULL
               OR s.Refuzed = 0)
              AND s.ScrapDeclarationId
                > 28
            ORDER BY S.DateIn;
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch dichiarazioni scrap: {e}")
            return []

    def fetch_scrap_declaration_by_id(self, declaration_id):
        """Recupera una singola dichiarazione di scrap per ID"""
        query = """
                SELECT s.ScrapDeclarationId, 
                       s.Datein as DeclarationDate, 
                       o.idorder as ProductionOrderId, 
                       o.OrderNumber, 
                       p.ProductCode, 
                       p.ProductName, 
                       1 as ScrapQuantity, 
                       sr.ScrapReasonId, 
                       sr.Reason as RasonDescription, 
                       s.[Note] as Notes, 
                       iif(s.refuzed = 0 and s.accepted = 0 , 'Waiting acceptance',iif(s.refuzed=1 and s.accepted =0,'Refuzed',
                       iif(s.refuzed=0 and s.accepted=1,'Accepted',''))) as  ValidationStatus, 
                       getdate() as ValidationDate, 
                       '' as ValidatorNotes, 
                       s.[user]  AS DeclaredBy
                FROM [Traceability_RS].[dbo].ScarpDeclarations S
                    INNER JOIN Traceability_RS.dbo.LabelCodes L 
                ON l.IDLabelCode = s.IdLabelCode
                    INNER JOIN [Traceability_RS].dbo.Areas A ON a.IDArea = s.IDParentPhase
                    INNER JOIN [Traceability_RS].dbo.defects D ON d.IDDefect = s.ScrapReasonId
                    INNER JOIN [Traceability_RS].dbo.boards B ON l.IDBoard = b.IDBoard
                    INNER JOIN [Traceability_RS].dbo.orders o ON o.idorder = b.IDOrder
                    inner join traceability_rs.dbo.products P on p.idproduct=o.idproduct
                    inner join [traceability_rs].dbo.ScrapResons sr on s.ScrapReasonId=sr.ScrapReasonId
                WHERE s.ScrapDeclarationId = ?; \
                """
        try:
            self.cursor.execute(query, (declaration_id,))
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch dichiarazione scrap {declaration_id}: {e}")
            return None

    def validate_scrap_declaration(self, declaration_id, validation_status, validator_notes, validator_name):
        """Valida o rifiuta una dichiarazione di scrap ed esegue la stored procedure se approvata"""
        try:
            if validation_status == 'Approved':
                # Imposta Accepted = 1, Refuzed = 0
                query = """
                        UPDATE [Traceability_RS].[dbo].[ScarpDeclarations]
                        SET Accepted = 1, Refuzed = 0, ValidationDate = GETDATE(), ValidatorNotes = ?, ValidatedBy = ?
                        WHERE ScrapDeclarationId = ?;
                        """
            else:  # Rejected
                # Imposta Accepted = 0, Refuzed = 1
                query = """
                        UPDATE [Traceability_RS].[dbo].[ScarpDeclarations]
                        SET Accepted = 0, Refuzed = 1, ValidationDate = GETDATE(), ValidatorNotes = ?, ValidatedBy = ?
                        WHERE ScrapDeclarationId = ?;
                        """

            # Esegui l'update
            self.cursor.execute(query, (validator_notes, validator_name, declaration_id))

            # Se Ã¨ approvato, recupera i dati ed esegui la stored procedure
            if validation_status == 'Approved':
                # Recupera i dati per la stored procedure - MODIFICA: ora recupera IDDefect invece di ScrapReasonId
                data_query = """
                                 SELECT l.labelcod as LabelCode, 
                    de.IDDefect     as IdDefect,   
                    s.Riferiments   as Riferiments, 
                    s.IDParentPhase as IdAreaDefect
             FROM [Traceability_RS].[dbo].ScarpDeclarations S
                 INNER JOIN [Traceability_RS].dbo.LabelCodes L on L.IDLabelCode=S.IdLabelCode
                 INNER JOIN [Traceability_RS].dbo.ScrapResons D ON d.ScrapReasonId = S.ScrapReasonId
                 INNER JOIN [Traceability_RS].dbo.defects de ON de.IDDefect = d.ScrapDeclarationId                         
            WHERE s.ScrapDeclarationId = ? \
                             """

                self.cursor.execute(data_query, (declaration_id,))
                row = self.cursor.fetchone()

                if row:
                    # Esegui la stored procedure
                    procedure_query = """
                    EXEC [Traceability_RS].dbo.[DeclareSCRAP] 
                        @LabelCode = ?,
                        @IdDefect = ?,
                        @riferiments = ?,
                        @idAreaDefect = ?
                    """
                    self.cursor.execute(procedure_query,
                                        (row.LabelCode, row.IdDefect, row.Riferiments or '', row.IdAreaDefect))
                    logger.info(f"Stored procedure DeclareSCRAP eseguita per LabelCode: {row.LabelCode}")

            self.conn.commit()
            logger.info(f"Dichiarazione scrap {declaration_id} validata: {validation_status}")
            return True, None

        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore validazione dichiarazione scrap {declaration_id}: {e}")
            return False, str(e)

    def fetch_scrap_reasons(self):
        """Recupera le causali di scarto"""
        query = """
                SELECT ScrapReasonId, ReasonCode, ReasonDescription
                FROM [Traceability_RS].[dbo].[ScrapReasons]
                WHERE DateStop IS NULL
                ORDER BY ReasonCode; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch causali scrap: {e}")
            return []

    def fetch_all_product_checks(self):
        """Recupera tutte le verifiche configurate"""
        query = """
                SELECT pc.PeriodicalProductCheckId, \
                       pc.IdProduct, \
                       pc.PeriodicityInQty,
                       p.ProductCode, \
                       p.ProductName
                FROM [Traceability_RS].[dbo].[PeriodicalProductChecks] pc
                    INNER JOIN dbo.products p \
                ON p.IDProduct = pc.IdProduct
                WHERE pc.datestop IS NULL
                ORDER BY p.ProductCode; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def insert_product_check(self, product_id, periodicity):
        """Inserisce una nuova verifica prodotto"""
        query = """
                INSERT INTO [Traceability_RS].[dbo].[PeriodicalProductChecks]
                    (IdProduct, PeriodicityInQty)
                VALUES (?, ?); \
                """
        try:
            self.cursor.execute(query, product_id, periodicity)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def update_product_check(self, check_id, product_id, periodicity):
        """Aggiorna una verifica prodotto"""
        query = """
                UPDATE [Traceability_RS].[dbo].[PeriodicalProductChecks]
                SET IdProduct = ?, PeriodicityInQty = ?
                WHERE PeriodicalProductCheckId = ?; \
                """
        try:
            self.cursor.execute(query, product_id, periodicity, check_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def delete_product_check(self, check_id):
        """Elimina (soft delete) una verifica prodotto"""
        query = """
                UPDATE [Traceability_RS].[dbo].[PeriodicalProductChecks]
                SET datestop = GETDATE()
                WHERE PeriodicalProductCheckId = ?; \
                """
        try:
            self.cursor.execute(query, check_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def fetch_products_with_checks(self):
        """Recupera prodotti con verifiche configurate"""
        query = """
                SELECT pc.PeriodicalProductCheckId, p.productcode, p.productname
                FROM [Traceability_RS].[dbo].[PeriodicalProductChecks] pc
                    INNER JOIN dbo.products p 
                ON p.IDProduct = pc.IdProduct
                WHERE pc.datestop IS NULL
                ORDER BY p.productcode; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_all_check_tasks(self):
        """Recupera tutti i task di verifica"""
        query = """
                SELECT ppl.PriodicalProductCheckListId, \
                       ppl.ItemToCheck, \
                       ppl.IsGeneric,
                       ppl.UserType, \
                       ppl.Doc, \
                       ppl.DateIn,
                       p.ProductCode, \
                       p.ProductName
                FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLists] ppl
                    LEFT JOIN [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics] ps
                ON ps.PriodicalProductCheckListId = ppl.PriodicalProductCheckListId
                    AND ps.dateout IS NULL
                    LEFT JOIN [Traceability_RS].[dbo].[PeriodicalProductChecks] pc
                    ON pc.PeriodicalProductCheckId = ps.PeriodicalProductCheckId
                    LEFT JOIN dbo.products p ON p.IDProduct = pc.IdProduct
                WHERE ppl.dateout IS NULL
                ORDER BY ppl.IsGeneric DESC, p.ProductCode, ppl.DateIn DESC; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_check_task_by_id(self, task_id):
        """Recupera un task specifico per ID"""
        query = """
                SELECT ppl.PriodicalProductCheckListId, \
                       ppl.ItemToCheck, \
                       ppl.IsGeneric,
                       ppl.UserType, \
                       ppl.Doc, \
                       ppl.DateIn,
                       p.ProductCode, \
                       p.ProductName
                FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLists] ppl
                    LEFT JOIN [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics] ps
                ON ps.PriodicalProductCheckListId = ppl.PriodicalProductCheckListId
                    LEFT JOIN [Traceability_RS].[dbo].[PeriodicalProductChecks] pc
                    ON pc.PeriodicalProductCheckId = ps.PeriodicalProductCheckId
                    LEFT JOIN dbo.products p ON p.IDProduct = pc.IdProduct
                WHERE ppl.PriodicalProductCheckListId = ?; \
                """
        try:
            self.cursor.execute(query, task_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def insert_check_task(self, item_to_check, is_generic, user_type, doc_data, product_check_id=None):
        """Inserisce un nuovo task di verifica"""
        try:
            self.conn.autocommit = False

            # Inserisci in PeriodicalProductCheckLists
            query1 = """
                     SET NOCOUNT ON;
                     INSERT INTO [Traceability_RS].[dbo].[PeriodicalProductCheckLists]
                         (ItemToCheck, IsGeneric, UserType, Doc, DateIn)
                     VALUES (?, ?, ?, ?, GETDATE());
                     SELECT CAST(SCOPE_IDENTITY() AS INT) AS NewID;
                     """
            self.cursor.execute(query1, item_to_check, 1 if is_generic else 0, user_type, doc_data)
            result = self.cursor.fetchone()

            if not result:
                raise Exception("Impossibile recuperare l'ID del task inserito")

            task_id = result[0]

            # Se non Ã¨ generico, inserisci anche in PeriodicProductCheckListSpecifics
            if not is_generic and product_check_id:
                query2 = """
                         INSERT INTO [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics]
                             (PeriodicalProductCheckId, PriodicalProductCheckListId)
                         VALUES (?, ?);
                         """
                self.cursor.execute(query2, product_check_id, task_id)

            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            print(f"âŒ Errore insert_check_task: {e}")
            return False
        finally:
            self.conn.autocommit = True

    def delete_check_task(self, task_id):
        """Elimina (soft delete) un task di verifica"""
        try:
            self.conn.autocommit = False

            # Soft delete in PeriodicalProductCheckLists
            query1 = """
                     UPDATE [Traceability_RS].[dbo].[PeriodicalProductCheckLists]
                     SET dateout = GETDATE()
                     WHERE PriodicalProductCheckListId = ?; \
                     """
            self.cursor.execute(query1, task_id)

            # Soft delete anche in PeriodicProductCheckListSpecifics
            query2 = """
                     UPDATE [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics]
                     SET dateout = GETDATE()
                     WHERE PriodicalProductCheckListId = ? AND dateout IS NULL; \
                     """
            self.cursor.execute(query2, task_id)

            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            self.conn.autocommit = True

    def check_label_code_exists(self, label_code, must_check_id):
        """Verifica se il label code esiste ed Ã¨ relativo all'ordine selezionato e restituisce l'IDLabelCode"""
        try:
            query = """
                    SELECT l.IDLabelCode
                    FROM traceability_rs.dbo.orders O
                             INNER JOIN traceability_rs.dbo.boards b ON o.IDOrder = b.IDOrder
                             INNER JOIN traceability_rs.dbo.LabelCodes l ON l.IDBoard = b.IDBoard
                             INNER JOIN traceability_rs.dbo.PeriodicalProductCheckMustLists M ON m.idorder = o.IDOrder
                             INNER JOIN traceability_rs.dbo.products p ON p.idproduct = o.IDProduct
                    WHERE l.LabelCod = ?
                      AND m.PeriodicalProductCheckMustListId = ?; \
                    """

            self.cursor.execute(query, label_code, must_check_id)
            result = self.cursor.fetchone()

            if result:
                return result.IDLabelCode  # Restituisce l'IDLabelCode se trovato
            else:
                return None  # Restituisce None se non trovato
        except pyodbc.Error as e:
            logger.error(f"Error checking label code: {e}")
            return None

    def fetch_products_must_check(self):
        """Recupera prodotti che necessitano verifica"""
        query = """
                SELECT M.PeriodicalProductCheckMustListId, \
                       o.ordernumber, \
                       p.productcode, \
                       pa.PhaseName,
                       FORMAT([Date], 'd', 'it-it') AS [Date], [Ora]
                FROM [Traceability_RS].[dbo].[PeriodicalProductCheckMustLists] M
                    INNER JOIN [Traceability_RS].[dbo].Orders O \
                ON M.[IdOrder] = o.idorder
                    INNER JOIN [Traceability_RS].[dbo].products P ON p.idproduct = M.idproduct
                    INNER JOIN traceability_rs.dbo.Phases pa ON pa.IDPhase = m.idphase
                    LEFT JOIN [Traceability_RS].[dbo].PeriodicalProductChecks PP ON PP.IdProduct = M.idproduct
                    LEFT JOIN [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics] PS
                    ON PS.PeriodicalProductCheckId = pp.PeriodicalProductCheckId
                    LEFT JOIN [Traceability_RS].[dbo].PeriodicalProductCheckLogs L
                    ON l.PeriodicalProductCheckMustListId = m.PeriodicalProductCheckMustListId
                WHERE L.PeriodicalProductCheckMustListId IS NULL
                ORDER BY [date], [ora]; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_generic_check_tasks(self):
        """Recupera task generici"""
        try:
            query = """
                SELECT ppl.PriodicalProductCheckListId, ppl.ItemToCheck, ppl.Doc
                FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLists] AS ppl
                WHERE isgeneric = 1 AND ppl.dateout IS NULL
                ORDER BY ppl.DateIn; \
                """

            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_specific_check_tasks(self, must_check_id):
        """Recupera task specifici per un prodotto"""
        # Prima recupera il PeriodicalProductCheckId dal must_check_id
        try:
            query1 = """
                 SELECT pp.PeriodicalProductCheckId
                 FROM [Traceability_RS].[dbo].[PeriodicalProductCheckMustLists] M
                     INNER JOIN [Traceability_RS].[dbo].PeriodicalProductChecks PP \
                 ON PP.IdProduct = M.idproduct
                 WHERE M.PeriodicalProductCheckMustListId = ?; \
                 """

            self.cursor.execute(query1, must_check_id)
            row = self.cursor.fetchone()
            if not row:
                return []

            product_check_id = row.PeriodicalProductCheckId

            # Ora recupera i task specifici
            query2 = """
                     SELECT ppl.PriodicalProductCheckListId, ppl.ItemToCheck, ppl.Doc
                     FROM [Traceability_RS].[dbo].[PeriodicalProductCheckLists] AS ppl
                         INNER JOIN [Traceability_RS].[dbo].[PeriodicProductCheckListSpecifics] AS ps
                     ON ps.PriodicalProductCheckListId = ppl.PriodicalProductCheckListId
                     WHERE isgeneric = 0
                       AND ppl.dateout IS NULL
                       AND ps.dateout IS NULL
                       AND ps.PeriodicalProductCheckId = ?
                     ORDER BY ppl.DateIn; \
                     """
            self.cursor.execute(query2, product_check_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def save_product_verification(self, must_check_id, user_name, label_code_id, status, comments=None):
        """Salva una verifica prodotto completata usando l'IDLabelCode"""
        try:
            # Verifica che la connessione e il cursor esistano
            if self.conn is None or self.cursor is None:
                logger.warning("Database connection or cursor is None, attempting to reconnect...")
                # Tenta di riconnettersi
                if not self.connect():
                    self.last_error_details = "Cannot reconnect to database"
                    logger.error("Failed to reconnect to database")
                    return False
                logger.info("Successfully reconnected to database")
            
            self.conn.autocommit = False

            # Inserisci in PeriodicalProductCheckLogs usando IDLabelCode
            query1 = """
                     INSERT INTO [Traceability_RS].[dbo].[PeriodicalProductCheckLogs]
                     (PeriodicalProductCheckMustListId, CheckTime, UserCheck, IDLabelCode, Status, Comments)
                     VALUES (?, GETDATE(), ?, ?, ?, ?);
                     """
            self.cursor.execute(query1, must_check_id, user_name, label_code_id, status, comments)

            # Aggiorna AllertActiveted in PeriodicalProductCheckMustLists
            query2 = """
                     UPDATE [dbo].[PeriodicalProductCheckMustLists]
                     SET AllertActiveted = 0
                     WHERE PeriodicalProductCheckMustListId = ?;
                     """
            self.cursor.execute(query2, must_check_id)

            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Error saving product verification: {e}")
            return False
        finally:
            self.conn.autocommit = True

    def execute_product_check_sp(self):
        """Esegue la stored procedure InsertProductToCheck"""
        with self._lock:
            try:
                self.cursor.execute("{CALL Traceability_rs.dbo.InsertProductToCheck}")
                self.conn.commit()
                return True
            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.error(f"Error executing InsertProductToCheck SP: {e}")
                return False

    def get_product_check_interval(self):
        """Recupera l'intervallo di controllo prodotti (in minuti)"""
        query = "SELECT [value] FROM traceability_rs.dbo.settings WHERE atribute = 'Sys_CheckTimeProduct'"
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query)
                row = self.cursor.fetchone()
                if row and row[0]:
                    try:
                        return int(row[0])
                    except:
                        return 30
                return 30
            except Exception as e:
                logger.error(f"Error fetching product check interval: {e}")
                return 30

    def check_monthly_report_sent(self):
        """
        Verifica se il report mensile Ã¨ giÃ  stato inviato questo mese.
        
        Returns:
            bool: True se giÃ  inviato questo mese, False altrimenti
        """
        query = """
        SELECT IDSettings 
        FROM traceability_rs.dbo.settings 
        WHERE atribute = 'Sys_Verify_check_fail' 
        AND lastcheck IS NOT NULL
        AND MONTH(lastcheck) = MONTH(GETDATE())
        AND YEAR(lastcheck) = YEAR(GETDATE())
        """
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query)
                row = self.cursor.fetchone()
                return row is not None  # True se trovato record (giÃ  inviato)
            except Exception as e:
                logger.error(f"Error checking monthly report status: {e}")
                return True  # In caso di errore, assume giÃ  inviato per sicurezza

    def update_monthly_report_timestamp(self):
        """
        Aggiorna il timestamp lastcheck dopo l'invio del report mensile.
        
        Returns:
            bool: True se aggiornamento riuscito, False altrimenti
        """
        query = """
        UPDATE traceability_rs.dbo.settings 
        SET lastcheck = GETDATE() 
        WHERE atribute = 'Sys_Verify_check_fail'
        """
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query)
                self.conn.commit()
                logger.info("Timestamp report mensile aggiornato con successo")
                return True
            except Exception as e:
                logger.error(f"Error updating monthly report timestamp: {e}")
                self.conn.rollback()
                return False

    def generate_monthly_report_data(self):
        """
        Genera i dati per il report mensile (tutti gli utenti).
        
        Returns:
            list: Lista di tuple con i dati del report
        """
        is_analyzed = 0  # Solo non analizzati
        label_code = None  # Tutti i label code
        
        sql = """
        DECLARE @IsAnalized bit = ?;
        DECLARE @LabelCode as nvarchar(230) = ?;

        WITH EmployeeMapping AS (
            SELECT 
                u.nomeuser,
                UPPER( e.EmployeeSurname + ' ' +e.EmployeeName) AS FullName,
                ROW_NUMBER() OVER (PARTITION BY u.nomeuser ORDER BY e.employeeid) AS rn
            FROM employee.dbo.employees e 
            INNER JOIN resetservices.dbo.tbuserkey u ON e.employeeid = u.idanga
        ),
        ComponentInfo AS (
            SELECT 
                ProductRiferiments.CodRiferimento,
                ProductComponentsErp.IDProduct,
                ParentPhases.ParentPhaseName,
                Components.ComponentCode,
                ROW_NUMBER() OVER (PARTITION BY ProductRiferiments.CodRiferimento, ProductComponentsErp.IDProduct 
                                  ORDER BY (SELECT NULL)) AS rn
            FROM ProductRiferiments 
            INNER JOIN ProductComponentsErp ON ProductComponentsErp.IDProductCompErp = ProductRiferiments.IDProductCompErp
            INNER JOIN ParentPhases ON ParentPhases.IDParentPhase = ProductRiferiments.IDParentPhase
            LEFT JOIN Components ON Components.IDComponent = ProductComponentsErp.IDComponent
        )
        SELECT distinct   
            EmployeeMapping.FullName AS CheckUser,
            labelcodes.LabelCod,
            Products.ProductCode,                
            CASE WHEN ScanDefects.IsPass = 1 THEN 'REPAIRED' ELSE 'SCRAP' END AS ResultRepair,  
            DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) as [Minute], 
            CASE 
                WHEN DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) < 60 
                    THEN CAST(DATEDIFF(MINUTE, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' MINUTE'
                WHEN DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) >= 24 
                    THEN CAST(DATEDIFF(DAY, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' DAYS'
                ELSE CAST(DATEDIFF(HOUR, PC.CheckTime, ScanDefects.StopTime) AS NVARCHAR(10)) + ' HOURS'
            END AS TimeDefectAfterCheck,       
            
            Riferiments.CodRiferimento,
            ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') AS ComponentType,  
            ISNULL(ComponentInfo.ComponentCode, '#N/D') AS ComponentCode,
            Defects.DefectNameRO AS Defect,        
            IIF(CAST(Boxes.BoxCode AS NVARCHAR(12)) is not null ,'IN BOX', 'NOT IN A BOX') AS BoxCode,
            IIF(PackingLists.CodePack IS NULL, 'NOT SHIPPED YET', 'SHIPPED ALREADY') AS ShipmentStatus
            
        FROM ScanDefects 
        INNER JOIN ScanDefectDetails ON ScanDefects.IDScanDefect = ScanDefectDetails.IDScanDefect
        INNER JOIN DefectsRiferiments ON DefectsRiferiments.IDScanDefectDet = ScanDefectDetails.IDScanDefectDet
        INNER JOIN Riferiments ON Riferiments.IDDibaRiferimento = DefectsRiferiments.IDDibaRiferimento
        INNER JOIN Defects ON ScanDefectDetails.IDDefect = Defects.IDDefect
        INNER JOIN Scannings ON Scannings.IDScan = ScanDefects.IDScan
        INNER JOIN OrderPhases ON OrderPhases.IDOrderPhase = Scannings.IDOrderPhase
        INNER JOIN Orders ON OrderPhases.IDOrder = Orders.IDOrder
        INNER JOIN Products ON Products.IDProduct = Orders.IDProduct
        INNER JOIN Phases ON OrderPhases.IDPhase = Phases.IDPhase
        INNER JOIN Clients ON Clients.IDClient = Products.IDClient
        INNER JOIN Boards ON Scannings.IDBoard = Boards.IDBoard
        INNER JOIN Teams ON Teams.IDTeam = ScanDefects.IdTeam
        INNER JOIN WorkLines ON WorkLines.IDWorkLine = Teams.IDWorkLine
        INNER JOIN LabelCodes ON Boards.IDBoard = LabelCodes.IDBoard
        INNER JOIN PeriodicalProductCheckLogs PC ON LabelCodes.IDLabelCode = PC.IDLabelCode and isnull(pc.isanalized,0) = @IsAnalized 

        LEFT JOIN EmployeeMapping ON EmployeeMapping.nomeuser COLLATE database_default = PC.UserCheck 
            AND EmployeeMapping.rn = 1
        LEFT JOIN ComponentInfo ON ComponentInfo.CodRiferimento = Riferiments.CodRiferimento 
            AND ComponentInfo.IDProduct = Orders.IDProduct 
            AND ComponentInfo.rn = 1

        LEFT JOIN Areas ON Areas.IDArea = ScanDefectDetails.IDArea
        LEFT JOIN BoxDetails ON BoxDetails.IDBoard = Boards.IDBoard
        LEFT JOIN Boxes ON Boxes.IDBox = BoxDetails.IDBox
        LEFT JOIN BoxPKs ON BoxPKs.IDBoxPK = Boxes.IDBoxPK
        LEFT JOIN PalletPKs ON PalletPKs.IDPalletPK = BoxPKs.IDPalletPK
        LEFT JOIN PackingLists ON PackingLists.IDPackingList = PalletPKs.IDPackingList

        WHERE 
            Phases.IDPhase IN (102, 103, 107)
            AND PC.Status = 'PASS'
            AND ScanDefects.StopTime > PC.CheckTime
            and ISNULL(ComponentInfo.ParentPhaseName, 'PTHM') = 'PTHM'
            AND NOT defects.DefectNameRO IN ('Schimbare pin fixture test','Componenta iesita din tolerante')
            and labelcodes.LabelCod = iif(@LabelCode is null,labelcodes.LabelCod,@LabelCode)
        Order By [Minute];
        """
        
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(sql, (is_analyzed, label_code))
                rows = self.cursor.fetchall()
                logger.info(f"Dati report mensile recuperati: {len(rows)} record")
                return rows
            except Exception as e:
                logger.error(f"Error generating monthly report data: {e}")
                return []

    def ensure_npi_weekly_email_log_table(self):
        """
        Crea la tabella di log per l'email settimanale NPI se non esiste.
        """
        query = """
        IF NOT EXISTS (
            SELECT 1
            FROM sys.objects
            WHERE object_id = OBJECT_ID(N'[Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog]')
              AND type in (N'U')
        )
        BEGIN
            CREATE TABLE [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog](
                [Id] INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
                [WeekStartDate] DATE NOT NULL,
                [Attribute] NVARCHAR(100) NOT NULL,
                [SentOn] DATETIME NOT NULL DEFAULT GETDATE()
            );
            CREATE UNIQUE INDEX UX_NpiWeeklyGeneralEmailLog_WeekAttribute
            ON [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog] ([WeekStartDate], [Attribute]);
        END
        """
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query)
                self.conn.commit()
                return True
            except Exception as e:
                logger.error(f"Errore creazione tabella NpiWeeklyGeneralEmailLog: {e}")
                self.conn.rollback()
                return False

    def check_weekly_npi_email_sent(self, week_start_date, attribute):
        """
        Verifica se l'email settimanale NPI è già stata inviata per la settimana.
        """
        query = """
        SELECT TOP 1 1
        FROM [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog]
        WHERE WeekStartDate = ? AND Attribute = ?
        """
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query, (week_start_date, attribute))
                return self.cursor.fetchone() is not None
            except Exception as e:
                logger.error(f"Errore verifica invio email settimanale NPI: {e}")
                return True  # Fail-safe: evita duplicati

    def log_weekly_npi_email_sent(self, week_start_date, attribute):
        """
        Registra l'invio dell'email settimanale NPI.
        """
        query = """
        INSERT INTO [Traceability_RS].[dbo].[NpiWeeklyGeneralEmailLog] (WeekStartDate, Attribute)
        VALUES (?, ?)
        """
        with self._lock:
            try:
                self._clear_cursor_state()
                self.cursor.execute(query, (week_start_date, attribute))
                self.conn.commit()
                return True
            except Exception as e:
                logger.error(f"Errore log invio email settimanale NPI: {e}")
                self.conn.rollback()
                return False

    def fetch_assigned_submissions(self, employee_hire_history_id: int):
        """Carica segnalazioni assegnate all'utente"""
        sql = """
              SELECT s.SegnalazioneId, \
                     ts.NomeTipo, \
                     s.Titolo, \
                     s.Descrizione,
                     e.EmployeeName + ' ' + e.EmployeeSurname   as Segnalatore,
                     sst.TipoStato, \
                     e1.EmployeeName + ' ' + e1.EmployeeSurname as Assegnatario,
                     ass.NomeFile, \
                     ass.DatiFile
              FROM employee.dbo.Segnalazioni s
                       INNER JOIN employee.dbo.EmployeeHireHistory h ON s.IdDipendente = h.EmployeeHireHistoryId
                       INNER JOIN employee.dbo.employees e ON h.EmployeeId = e.EmployeeId
                       INNER JOIN employee.dbo.SegnalazioneStati ss ON s.SegnalazioneId = ss.SegnalazioneId
                       INNER JOIN employee.dbo.SegnalazioniTipoStati sst \
                                  ON sst.SegnalazioniTipoStatoId = ss.SegnalazioniTipoStatoId
                       INNER JOIN employee.dbo.TipiSegnalazione ts ON ts.TipoSegnalazioneId = s.TipoSegnalazioneId
                       LEFT JOIN employee.dbo.SegnalazioneAllegati ass ON ass.SegnalazioneId = s.SegnalazioneId
                       INNER JOIN employee.dbo.SegnalazioneAssegnazioni sa ON sa.SegnalazioneId = s.SegnalazioneId
                       INNER JOIN employee.dbo.employeehirehistory h1 \
                                  ON h1.EmployeeHireHistoryId = sa.EmployeeHireHistoryId
                       INNER JOIN employee.dbo.employees e1 ON e1.employeeid = h1.EmployeeId
              WHERE sa.EmployeeHireHistoryId = ?
                AND NOT EXISTS (SELECT 1 \
                                FROM [Employee].[dbo].[SegnalazioneStati] ss2 
            INNER JOIN employee.dbo.segnalazioniTipoStati sst2 
                ON sst2.SegnalazioniTipoStatoId = ss2.SegnalazioniTipoStatoId
            WHERE ss2.SegnalazioneId = s.SegnalazioneId AND sst2.statochiuso = 1)
                AND sst.SegnalazioniTipoStatoId IN (1, 2)
              ORDER BY s.SegnalazioneId, sa.DateSys; \
              """
        try:
            self.cursor.execute(sql, employee_hire_history_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Error fetching assigned submissions: {e}")
            return []

    def fetch_submission_activities(self, segnalazione_id: int):
        """Carica attivitÃ  svolte per una segnalazione"""
        sql = """
              SELECT sv.[SegnalazioneSvolgimentoId], \
                     sa.Note, \
                     sv.[DescrizioneAttivita],
                     sv.[Documentazione], \
                     sv.DateSys as DataAzione
              FROM [Employee].[dbo].[SegnalazioneSvolgimenti] sv
                  INNER JOIN employee.dbo.SegnalazioneAssegnazioni sa
              ON sv.SegnalazioneAssegnazioneId=sa.SegnalazioniAssegnazioniID
                  INNER JOIN employee.dbo.Segnalazioni s ON s.SegnalazioneId=sa.SegnalazioneId
              WHERE s.SegnalazioneId=?
              ORDER BY sv.DateSys DESC; \
              """
        try:
            self.cursor.execute(sql, segnalazione_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Error fetching activities: {e}")
            return []

    def fetch_submission_status_types(self):
        """Carica tipi stato per combo"""
        sql = """
              SELECT [SegnalazioniTipoStatoId], [TipoStato], Statochiuso
              FROM [Employee].[dbo].[SegnalazioniTipoStati]
              WHERE SegnalazioniTipoStatoId IN (3, 2, 5, 4)
              ORDER BY [TipoStato]; \
              """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def insert_submission_activity(self, assegnazione_id: int, descrizione: str):
        """Inserisce nuova attivitÃ """
        sql = """
              INSERT INTO [Employee].[dbo].[SegnalazioneSvolgimenti]
                  ([SegnalazioneAssegnazioneId], [DescrizioneAttivita], [DateSys])
              VALUES (?, ?, GETDATE());
              SELECT SCOPE_IDENTITY(); \
              """
        try:
            self.conn.autocommit = False
            self.cursor.execute(sql, assegnazione_id, descrizione)
            svolgimento_id = int(self.cursor.fetchone()[0])
            self.conn.commit()
            return True, svolgimento_id
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Error inserting activity: {e}")
            return False, None
        finally:
            self.conn.autocommit = True

    def insert_activity_attachment(self, svolgimento_id: int, descrizione: str, file_data: bytes):
        """Inserisce allegato per attivitÃ """
        sql = """
              INSERT INTO [Employee].[dbo].[SegnalazioniStatiAllegati]
              ([SegnalazioneSvolgimentoId], [DescrizioneDocumento], [Allegato], [DateIn])
              VALUES (?, ?, ?, GETDATE()); \
              """
        try:
            self.cursor.execute(sql, svolgimento_id, descrizione, file_data)
            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, str(e)

    def update_submission_status(self, segnalazione_id: int, tipo_stato_id: int,
                                 nota: str, operato_da: str):
        """Aggiorna stato segnalazione"""
        sql = """
              INSERT INTO [Employee].[dbo].[SegnalazioneStati]
              ([SegnalazioneId], [SegnalazioniTipoStatoId], [DataAttivita], [Nota], [OperatoDa])
              VALUES (?, ?, GETDATE(), ?, ?); \
              """
        try:
            self.conn.autocommit = False
            self.cursor.execute(sql, segnalazione_id, tipo_stato_id, nota, operato_da)
            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def get_submission_assignment_id(self, segnalazione_id: int, employee_hire_history_id: int):
        """Recupera ID assegnazione per inserire attivitÃ """
        sql = """
              SELECT SegnalazioniAssegnazioniID
              FROM employee.dbo.SegnalazioneAssegnazioni
              WHERE SegnalazioneId = ? \
                AND EmployeeHireHistoryId = ?; \
              """
        try:
            self.cursor.execute(sql, segnalazione_id, employee_hire_history_id)
            row = self.cursor.fetchone()
            return row.SegnalazioniAssegnazioniID if row else None
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def _ensure_connection(self):
        """Verifica e ripristina la connessione se necessario"""
        with self._lock:
            try:
                # Se non c'Ã¨ proprio il cursore o la connessione, prova a connettere
                if not self.conn or not self.cursor:
                    return self.connect()

                # Prova una query semplice per testare la connessione
                self.cursor.execute("SELECT 1")
                self.cursor.fetchone()
                return True
            except Exception as e:
                logger.warning(f"Connessione persa, tentativo di riconnessione: {e}")
                try:
                    if self.conn:
                        self.conn.close()
                except:
                    pass

                try:
                    # Riconnessione completa
                    self.conn = pyodbc.connect(self.conn_str, autocommit=False)
                    self.cursor = self.conn.cursor()  # Nuovo cursore
                    logger.info("âœ“ Connessione ripristinata")
                    return True
                except Exception as reconnect_error:
                    logger.error(f"âœ— Impossibile ripristinare la connessione: {reconnect_error}")
                    return False

    def fetch_kanban_current_stock_by_component(self):
        """
        Recupera lo stock corrente di tutti i componenti dal Kanban.
        """
        sql = """
              SELECT IdComponent, COALESCE(SUM(Quantity), 0) AS Stock
              FROM Traceability_rs.knb.KanBanRecords
              WHERE DateOut IS NULL
              GROUP BY IdComponent
              """
        with self._lock:
            try:
                if not self._ensure_connection():
                    logger.error("âœ— Connessione non disponibile in fetch_kanban_current_stock_by_component")
                    return None

                self._clean_cursor()
                self.cursor.execute(sql)
                rows = self.cursor.fetchall()
                # If rows are returned as tuples, access by index
                return {int(row[0]): int(row[1]) for row in rows}

            except Exception as e:
                self.last_error_details = str(e)
                logger.exception(f"Error in fetch_kanban_current_stock_by_component: {e}")
                return None

    def _clean_cursor(self):
        """Pulisce lo stato del cursore"""
        with self._lock:
            try:
                # Prova a chiudere il cursore corrente
                if self.cursor:
                    try:
                        # Tenta di consumare eventuali risultati pendenti
                        while self.cursor.nextset():
                            pass
                    except:
                        pass

                    # Chiudi il cursore e creane uno nuovo
                    self.cursor.close()
                    self.cursor = self.conn.cursor()
                    logger.debug("âœ“ Cursore pulito")
            except Exception as e:
                logger.warning(f"Errore durante la pulizia del cursore: {e}")
                # Se fallisce, crea un nuovo cursore
                try:
                    self.cursor = self.conn.cursor()
                except:
                    pass


    def fetch_active_rules_by_component(self):
        """
        Ritorna {KanBanRuleID: {'min_qty':..., 'min_pct':...}} per le regole attive
        """
        sql = """
              SELECT KanBanRuleID, MinimumProcent, MinimumQty
              FROM [Traceability_RS].[knb].[KanBanRules]
              WHERE DateOut IS NULL
              """
        with self._lock:
            try:
                if not self._ensure_connection(): return None
                self.cursor.execute(sql)
                rows = self.cursor.fetchall()
                result = {}
                for row in rows:
                    rule_id = int(row[0])
                    result[rule_id] = {
                        'min_pct': row[1] if row[1] is not None else None,
                        'min_qty': row[2] if row[2] is not None else None
                    }
                return result

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.error(f"Error in fetch_active_rules_by_component: {e}")
                return None

    def fetch_first_load_qty_by_component(self, component_ids):
        """
        Ritorna {IdComponent: FirstLoadQty} per i componenti specificati.
        """
        if not component_ids:
            return {}

        placeholders = ','.join(['?'] * len(component_ids))
        sql = f"""
              SELECT k1.IdComponent, k1.Quantity as FirstLoadQty
              FROM knb.KanBanRecords k1
              INNER JOIN (
                  SELECT IdComponent, MIN(DateIn) as FirstDate
                  FROM knb.KanBanRecords 
                  WHERE IdComponent IN ({placeholders})
                  GROUP BY IdComponent
              ) k2 ON k1.IdComponent = k2.IdComponent AND k1.DateIn = k2.FirstDate
              """
        with self._lock:
            try:
                if not self._ensure_connection(): return None
                self.cursor.execute(sql, component_ids)
                rows = self.cursor.fetchall()
                return {int(row[0]): int(row[1]) for row in rows}

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.error(f"Error in fetch_first_load_qty_by_component: {e}")
                return None

    def fetch_max_single_load_by_component(self, component_ids):
        """
        Ritorna {IdComponent: {'max_qty': X, 'record_id': Y}}
        """
        if not component_ids:
            return {}

        placeholders = ','.join(['?'] * len(component_ids))
        sql = f"""
              SELECT IdComponent, MAX(Quantity) as MaxQty, MIN(KanBanRecordId) as RecordId
              FROM knb.KanBanRecords 
              WHERE IdComponent IN ({placeholders})
              GROUP BY IdComponent
              """
        with self._lock:
            try:
                if not self._ensure_connection(): return None
                self.cursor.execute(sql, component_ids)
                rows = self.cursor.fetchall()
                return {
                    int(row[0]): {
                        'max_qty': int(row[1]),
                        'record_id': int(row[2])
                    } for row in rows
                }

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.error(f"Error in fetch_max_single_load_by_component: {e}")
                return None

    def fetch_components_master(self, component_ids):
        """
        Recupera i dati master dei componenti.
        """
        if not component_ids:
            return {}

        placeholders = ','.join(['?' for _ in component_ids])
        sql = f"""
            SELECT IdComponent, ComponentCode, ComponentDescription 
            FROM Traceability_rs.dbo.Components 
            WHERE IdComponent IN ({placeholders})
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return None
                self.cursor.execute(sql, component_ids)
                rows = self.cursor.fetchall()
                master_map = {}
                for row in rows:
                    master_map[getattr(row, 'IdComponent', row[0])] = {
                        'code': getattr(row, 'ComponentCode', row[1]),
                        'desc': getattr(row, 'ComponentDescription', row[2])
                    }
                return master_map

            except Exception as e:
                self.last_error_details = str(e)
                logger.error(f"Error in fetch_components_master: {e}")
                self._clear_cursor_state()
                return None

    def has_refill_request_today(self, kanban_record_id: int) -> bool:
        """
        True se esiste giÃ  una richiesta per la stessa KanBanRecordId oggi.
        """
        sql = """
        SELECT 1
        FROM knb.KanBanMaterialRequestes
        WHERE KanBanRecordId = ?
          AND CAST(RequestedOn AS date) = CAST(GETDATE() AS date);
        """
        with self._lock:
            try:
                self.cursor.execute(sql, kanban_record_id)
                row = self.cursor.fetchone()
                return row is not None
            except Exception as e:
                logger.error(f"Error in has_refill_request_today: {e}")
                return False

    def insert_refill_request(self, kanban_record_id: int, qty_to_refill: int) -> bool:
        """
        Inserisce la richiesta di refill.
        """
        sql = """
        INSERT INTO knb.KanBanMaterialRequestes (KanBanRecordId, QtyToRefill, RequestedOn)
        VALUES (?, ?, GETDATE());
        """
        with self._lock:
            try:
                self.cursor.execute(sql, (kanban_record_id, qty_to_refill))
                self.conn.commit()
                logger.info('Richiesta materiali KanBan, registrata in DB')
                return True
            except Exception as e:
                self.conn.rollback()
                self.last_error_details = str(e)
                logger.error(f"Error in insert_refill_request: {e}")
                return False

    def get_total_stock_component(self, id_component: int) -> int:
        """
        Stock totale del componente su tutte le locazioni (DateOut IS NULL).
        """
        sql = """
        SELECT COALESCE(SUM(Quantity), 0) AS Qty
        FROM knb.KanBanRecords
        WHERE IdComponent = ? AND DateOut IS NULL;
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return 0
                self.cursor.execute(sql, id_component)
                row = self.cursor.fetchone()
                return int(row[0] if row and row[0] is not None else 0)
            except Exception as e:
                logger.error(f"Error in get_total_stock_component: {e}")
                return 0

    def get_component_locations_with_stock(self, id_component: int) -> dict[int, int]:
        """
        Restituisce {LocationId: Qty} per le locazioni dove il componente ha stock > 0 (DateOut IS NULL).
        """
        sql = """
        SELECT LocationId, SUM(Quantity) AS Qty
        FROM knb.KanBanRecords
        WHERE IdComponent = ? AND DateOut IS NULL
        GROUP BY LocationId
        HAVING SUM(Quantity) > 0;
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return {}
                self.cursor.execute(sql, id_component)
                rows = self.cursor.fetchall()
                return {int(row[0]): int(row[1]) for row in rows}
            except Exception as e:
                logger.error(f"Error in get_component_locations_with_stock: {e}")
                return {}

    def fetch_all_components_for_combo(self):
        """
        Recupera tutti i componenti (Id, Code, Desc) per popolare le combo.
        """
        sql = """
        SELECT IdComponent, ComponentCode, ComponentDescription
        FROM Traceability_rs.dbo.ComponentsMaster
        WHERE dateout IS NULL
        ORDER BY ComponentCode;
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return []
                self.cursor.execute(sql)
                rows = self.cursor.fetchall()
                return [(int(row[0]), str(row[1]), str(row[2])) for row in rows]
            except Exception as e:
                logger.error(f"Error in fetch_all_components_for_combo: {e}")
                return []

    def fetch_all_locations_for_combo(self):
        """
        Elenco per combo: LocationId, LocationCode, KanBanLocation (area)
        """
        sql = """
        SELECT l.LocationId, l.LocationCode, kbl.KanBanLocation
        FROM knb.Locations AS l
        LEFT JOIN knb.KanBanLocations AS kbl ON kbl.KanBanLocationId = l.KanBanLocationId
        ORDER BY kbl.KanBanLocation, l.LocationCode;
        """
        self.cursor.execute(sql)
        return self.cursor.fetchall()

    def get_component_id_by_code(self, component_code: str):
        sql = "SELECT IdComponent FROM dbo.Components WHERE ComponentCode = ?;"
        self.cursor.execute(sql, component_code)
        row = self.cursor.fetchone()
        return row.IdComponent if row else None

    def get_location_id_by_code(self, location_code: str):
        sql = "SELECT LocationId FROM knb.Locations WHERE LocationCode = ?;"
        with self._lock:
            try:
                if not self._ensure_connection(): return None
                self.cursor.execute(sql, location_code)
                row = self.cursor.fetchone()
                return row[0] if row else None
            except Exception as e:
                logger.error(f"Error in get_location_id_by_code: {e}")
                return None

    def get_current_stock(self, id_component: int, location_id: int) -> int:
        """
        Stock corrente = somma dei movimenti (Quantity) per componente+locazione con DateOut IS NULL.
        """
        sql = """
        SELECT COALESCE(SUM(Quantity), 0) AS Qty
        FROM knb.KanBanRecords
        WHERE IdComponent = ? AND LocationId = ? AND DateOut IS NULL;
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return 0
                self.cursor.execute(sql, (id_component, location_id))
                row = self.cursor.fetchone()
                return int(row[0] if row and row[0] is not None else 0)
            except Exception as e:
                logger.error(f"Error in get_current_stock: {e}")
                return 0

    def insert_kanban_movement(self, location_id: int, id_component: int, quantity: int, user_name: str | None = None):
        """
        Inserisce un movimento: quantity >0 carico, <0 prelievo.
        """
        if not isinstance(quantity, int) or quantity == 0:
            return False, "invalid_quantity"

        sql = """
        INSERT INTO knb.KanBanRecords (LocationId, IdComponent, Quantity, DateIn, DateOut, [User])
        VALUES (?, ?, ?, GETDATE(), NULL, ?);
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return False, "connection_error"
                self.cursor.execute(sql, (location_id, id_component, quantity, user_name))
                self.conn.commit()
                return True, None
            except Exception as e:
                if self.conn: self.conn.rollback()
                self.last_error_details = str(e)
                logger.error(f"Error in insert_kanban_movement: {e}")
                return False, "db_error"

    def fetch_components_locations_report(self):
        """
        Elenca componenti attivi con la loro locazione e area (solo record aperti: DateOut IS NULL).
        Colonne: ComponentCode, ComponentDescription, LocationCode, KanBanLocation
        """
        sql = """
        SELECT c.ComponentCode,
               c.ComponentDescription,
               l.LocationCode,
               kbl.KanBanLocation
        FROM dbo.Components AS c
        INNER JOIN knb.KanBanRecords AS kr ON kr.IdComponent = c.IdComponent
        INNER JOIN knb.Locations     AS l  ON l.LocationId   = kr.LocationId
        INNER JOIN knb.KanBanLocations AS kbl ON kbl.KanBanLocationId = l.KanBanLocationId
        WHERE kr.DateOut IS NULL
        ORDER BY kbl.KanBanLocation, l.LocationCode, c.ComponentCode;
        """
        with self._lock:
            try:
                if not self._ensure_connection(): return []
                self.cursor.execute(sql)
                return self.cursor.fetchall()
            except Exception as e:
                logger.error(f"Error in fetch_components_locations_report: {e}")
                return []

    def fetch_component_id_by_code(self, component_code: str):
        """
        Ritorna IdComponent dato il ComponentCode.
        Ritorna None se non trovato.
        """
        sql = "SELECT IdComponent FROM dbo.Components WHERE ComponentCode = ?;"
        try:
            self.cursor.execute(sql, component_code)
            row = self.cursor.fetchone()
            return row.IdComponent if row else None
        except Exception as e:
            self.last_error_details = str(e)
            return None

    def fetch_location_id_by_code(self, location_code: str):
        """
        Ritorna LocationId dato il LocationCode.
        Ritorna None se non trovato.
        """
        sql = "SELECT LocationId FROM knb.Locations WHERE LocationCode = ?;"
        try:
            self.cursor.execute(sql, location_code)
            row = self.cursor.fetchone()
            return row.LocationId if row else None
        except Exception as e:
            self.last_error_details = str(e)
            return None

    def fetch_all_kanban_rules(self):
        """
        Ritorna tutte le regole (attive e chiuse).
        """
        sql = """
        SELECT KanBanRuleID, MinimumProcent, MinimumQty, DateOut
        FROM knb.KanBanRules
        ORDER BY CASE WHEN DateOut IS NULL THEN 0 ELSE 1 END,
                 CASE WHEN MinimumProcent IS NOT NULL THEN 0 ELSE 1 END,
                 MinimumProcent, MinimumQty, KanBanRuleID;
        """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def add_kanban_rule(self, min_percent: int | None, min_qty: int | None):
        """
        Crea una nuova regola. Esattamente uno tra min_percent e min_qty deve essere valorizzato.
        """
        if (min_percent is None) == (min_qty is None):
            return False, "exactly_one_value_required"
        try:
            self.conn.autocommit = False
            self.cursor.execute(
                "INSERT INTO knb.KanBanRules (MinimumProcent, MinimumQty) VALUES (?, ?);",
                min_percent, min_qty
            )
            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def update_kanban_rule(self, rule_id: int, min_percent: int | None, min_qty: int | None):
        """
        Modifica una regola esistente. Esattamente uno tra min_percent e min_qty deve essere valorizzato.
        Non tocca DateOut.
        """
        if (min_percent is None) == (min_qty is None):
            return False, "exactly_one_value_required"
        try:
            self.conn.autocommit = False
            self.cursor.execute(
                "UPDATE knb.KanBanRules SET MinimumProcent = ?, MinimumQty = ? WHERE KanBanRuleID = ?;",
                min_percent, min_qty, rule_id
            )
            if self.cursor.rowcount == 0:
                self.conn.rollback()
                return False, "not_found"
            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def soft_delete_kanban_rule(self, rule_id: int):
        """
        Cancellazione logica: setta DateOut = GETDATE() se non Ã¨ giÃ  valorizzata.
        """
        try:
            self.conn.autocommit = False
            self.cursor.execute(
                "UPDATE knb.KanBanRules SET DateOut = GETDATE() WHERE KanBanRuleID = ? AND DateOut IS NULL;",
                rule_id
            )
            if self.cursor.rowcount == 0:
                # giÃ  chiusa o non trovata
                self.conn.rollback()
                return False, "already_closed_or_not_found"
            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def fetch_component_types(self):
        """
        Select IdComponentType, ComponentTypeName from ComponentTypes order by ComponentTypeName;
        """
        sql = """
        SELECT IdComponentType, ComponentTypeName
        FROM dbo.ComponentTypes
        ORDER BY ComponentTypeName;
        """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_components(self, component_type_id: int | None = None):
        """
        Carica i componenti (len(componentcode) > 5), opzionalmente filtrati per tipo.
        select idcomponent, c.ComponentCode, ComponentDescription, ct.ComponentTypeName
        from components c
        left join ComponentTypes ct on c.IDComponentType = ct.IDComponentType
        where len(c.componentcode) > 5
          [and c.IDComponentType = ?]
        order by c.componentcode;
        """
        base = """
        SELECT c.IdComponent, c.ComponentCode, c.ComponentDescription, ct.ComponentTypeName
        FROM dbo.Components c
        LEFT JOIN dbo.ComponentTypes ct ON c.IDComponentType = ct.IDComponentType
        WHERE LEN(c.ComponentCode) > 5
        """
        params = []
        if component_type_id:
            base += " AND c.IDComponentType = ?"
            params.append(component_type_id)
        base += " ORDER BY c.ComponentCode;"
        try:
            self.cursor.execute(base, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_kanban_rules(self):
        """
        SELECT KanBanRuleID, MinimumProcent, MinimumQty FROM knb.KanBanRules
        Ordina con percentuali prima, poi quantitÃ .
        """
        sql = """
        SELECT KanBanRuleID, MinimumProcent, MinimumQty
        FROM knb.KanBanRules
        ORDER BY CASE WHEN MinimumProcent IS NOT NULL THEN 0 ELSE 1 END,
                 MinimumProcent, MinimumQty;
        """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_active_rule_for_component(self, id_component: int):
        """
        Ritorna la regola attiva (DateOut IS NULL) per il componente, se presente.
        """
        sql = """
        SELECT TOP 1 KanBanRuleId
        FROM knb.KanBanRecodRules
        WHERE IdComponent = ? AND DateOut IS NULL
        ORDER BY DateIn DESC;
        """
        try:
            self.cursor.execute(sql, id_component)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def assign_rule_to_component(self, id_component: int, kanban_rule_id: int | None):
        """
        Associa una regola a un componente:
        - Chiude eventuale regola attiva con DateOut = GETDATE()
        - Se kanban_rule_id Ã¨ None: solo chiusura (rimozione regola)
        - Altrimenti inserisce un nuovo record con DateIn = GETDATE(), DateOut = NULL

        Usa transazione.
        """
        try:
            self.conn.autocommit = False

            # Chiudi eventuali regole attive
            self.cursor.execute(
                "UPDATE knb.KanBanRecodRules SET DateOut = GETDATE() WHERE IdComponent = ? AND DateOut IS NULL;",
                id_component
            )

            if kanban_rule_id is not None:
                self.cursor.execute(
                    "INSERT INTO knb.KanBanRecodRules (IdComponent, KanBanRuleId, DateIn) VALUES (?, ?, GETDATE());",
                    id_component, kanban_rule_id
                )

            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def _table_has_column(self, schema: str, table: str, column: str) -> bool:
        sql = """
        SELECT 1
        FROM sys.columns
        WHERE object_id = OBJECT_ID(?)
          AND name = ?;
        """
        try:
            self.cursor.execute(sql, f"{schema}.{table}", column)
            return self.cursor.fetchone() is not None
        except pyodbc.Error:
            return False

    def fetch_locations_for_combo(self):
        """
        Ritorna elenco locazioni per combo destinazione.
        Colonne: LocationId, KanBanLocationId, LocationCode, KanBanLocation
        """
        sql = """
        SELECT l.LocationId, k.KanBanLocationId, l.LocationCode, k.KanBanLocation
        FROM knb.Locations l
        INNER JOIN knb.KanBanLocations k ON k.KanBanLocationId = l.KanBanLocationId
        ORDER BY k.KanBanLocation, l.LocationCode;
        """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def search_component_open_records(self, component_code: str):
        """
        Ricerca componenti attivi (DateOut IS NULL) per codice componente.
        Aggiungo KanBanRecordId per identificare il record da chiudere.
        """
        sql = """
        SELECT k.KanBanRecordId, c.IdComponent, l.LocationId,
               c.ComponentCode, c.ComponentDescription,
               l.LocationCode + ' (' +kl.KanBanLocation +')' as LocationCode, k.Quantity
        FROM dbo.Components AS c
        INNER JOIN knb.KanBanRecords AS k ON c.IdComponent = k.IdComponent
        INNER JOIN knb.Locations     AS l ON k.LocationId  = l.LocationId
        inner join knb.KanBanLocations as KL on l.KanBanLocationId=Kl.KanBanLocationId
        WHERE c.ComponentCode = ? AND k.DateOut IS NULL
        ORDER BY l.LocationCode;
        """
        try:
            self.cursor.execute(sql, component_code)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def move_record_to_location(self, record_id: int, to_location_id: int, to_kanban_location_id: int | None = None):
        """
        Sposta UN record (singolo componente in una specifica locazione) in un'altra locazione.
        - Chiude il record sorgente (DateOut = GETDATE()).
        - Inserisce nuovo record con stessa Quantity/IdComponent in locazione destinazione (DateIn = GETDATE()).
        Se la tabella ha la colonna KanBanLocationId, la valorizza.
        """
        has_kbl = self._table_has_column('knb', 'KanBanRecords', 'KanBanLocationId')
        try:
            self.conn.autocommit = False

            # Carico dati del record
            self.cursor.execute(
                "SELECT IdComponent, Quantity FROM knb.KanBanRecords WHERE KanBanRecordId = ?;", record_id
            )
            row = self.cursor.fetchone()
            if not row:
                self.conn.rollback()
                return False, "not_found"
            id_component, qty = row

            # Chiudi il record sorgente (solo se ancora aperto)
            self.cursor.execute(
                "UPDATE knb.KanBanRecords SET DateOut = GETDATE() WHERE KanBanRecordId = ? AND DateOut IS NULL;",
                record_id
            )
            if self.cursor.rowcount == 0:
                self.conn.rollback()
                return False, "already_closed"

            # Inserisci nuovo record in destinazione
            if has_kbl and to_kanban_location_id is not None:
                sql_ins = """
                INSERT INTO knb.KanBanRecords (LocationId, IdComponent, Quantity, DateIn, KanBanLocationId)
                VALUES (?, ?, ?, GETDATE(), ?);
                """
                self.cursor.execute(sql_ins, to_location_id, id_component, qty, to_kanban_location_id)
            else:
                sql_ins = """
                INSERT INTO knb.KanBanRecords (LocationId, IdComponent, Quantity, DateIn)
                VALUES (?, ?, ?, GETDATE());
                """
                self.cursor.execute(sql_ins, to_location_id, id_component, qty)

            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def move_all_from_location(self, from_location_id: int, to_location_id: int,
                               to_kanban_location_id: int | None = None):
        """
        Sposta TUTTI i record aperti (DateOut IS NULL) da una locazione a un'altra.
        Chiude i record sorgenti e inserisce nuovi record in blocco.
        """
        has_kbl = self._table_has_column('knb', 'KanBanRecords', 'KanBanLocationId')
        try:
            self.conn.autocommit = False

            # Conta quanti record aperti ci sono nella sorgente
            self.cursor.execute(
                "SELECT COUNT(*) FROM knb.KanBanRecords WHERE LocationId = ? AND DateOut IS NULL;",
                from_location_id
            )
            n = self.cursor.fetchone()[0]
            if n == 0:
                self.conn.rollback()
                return False, "nothing_to_move"

            # Chiudi i record sorgenti
            self.cursor.execute(
                "UPDATE knb.KanBanRecords SET DateOut = GETDATE() WHERE LocationId = ? AND DateOut IS NULL;",
                from_location_id
            )

            # Inserimento con INSERT ... SELECT
            if has_kbl and to_kanban_location_id is not None:
                sql_ins = """
                INSERT INTO knb.KanBanRecords (LocationId, IdComponent, Quantity, DateIn, KanBanLocationId)
                SELECT ?, IdComponent, Quantity, GETDATE(), ?
                FROM knb.KanBanRecords
                WHERE LocationId = ? AND DateOut = (SELECT MAX(DateOut) FROM knb.KanBanRecords kk WHERE kk.KanBanRecordId = knb.KanBanRecords.KanBanRecordId)
                   OR (LocationId = ? AND DateOut IS NULL) -- per sicurezza, ma dopo l'update sopra non ci saranno piÃ¹ NULL
                ;
                """
                # Notare: la riga sopra replica quantity e component dagli ultimi record; in pratica, dopo l'UPDATE, le righe appena chiuse vengono reinserite.
                # Passo due volte from_location_id per la OR.
                self.cursor.execute(sql_ins, to_location_id, to_kanban_location_id, from_location_id, from_location_id)
            else:
                sql_ins = """
                INSERT INTO knb.KanBanRecords (LocationId, IdComponent, Quantity, DateIn)
                SELECT ?, IdComponent, Quantity, GETDATE()
                FROM knb.KanBanRecords
                WHERE LocationId = ? AND DateOut = (SELECT MAX(DateOut) FROM knb.KanBanRecords kk WHERE kk.KanBanRecordId = knb.KanBanRecords.KanBanRecordId)
                   OR (LocationId = ? AND DateOut IS NULL);
                """
                self.cursor.execute(sql_ins, to_location_id, from_location_id, from_location_id)

            self.conn.commit()
            return True, None
        except pyodbc.Error as e:
            try:
                self.conn.rollback()
            except Exception:
                pass
            self.last_error_details = str(e)
            return False, str(e)
        finally:
            self.conn.autocommit = True

    def fetch_kanban_locations_all(self):
        """
        Restituisce tutte le locazioni KanBan (LocationCode) ordinate.
        """
        sql = """
            SELECT L.LocationCode, L.KanBanLocationId
            FROM knb.Locations AS L
            ORDER BY L.LocationCode;
        """
        try:
            self.cursor.execute(sql)
            return self.cursor.fetchall()  # .LocationCode, .KanBanLocationId
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def count_kanban_locations(self) -> int | None:
        """
        Ritorna il numero totale di locazioni KanBan (knb.Locations).
        """
        try:
            return self.cursor.execute("SELECT COUNT(*) FROM knb.Locations;").fetchval()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def fetch_kanban_areas(self):
        """
        Restituisce le aree KanBan (IDParentPhase, ParentPhaseName) filtrate per CodCDC in (10,30,90).
        """
        query = """            
            select [KanBanLocationId] AS IDParentPhase,
            KanBanLocation AS ParentPhaseName
             FROM  KNB.KANBANLOCATIONS order by KanBanLocation;
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def insert_kanban_location(self, kanban_location_id: int, location_code: str):
        """
        Inserisce una nuova locazione KanBan.
        Ritorna (True, None) su successo, oppure (False, 'duplicate') se violazione chiave/unicitÃ ,
        oppure (False, 'messaggio_errore') per altri errori.
        """
        sql = """
            INSERT INTO knb.Locations (KanBanLocationId, LocationCode)
            VALUES (?, ?);
        """
        try:
            self.cursor.execute(sql, int(kanban_location_id), location_code)
            self.conn.commit()
            return True, None
        except pyodbc.IntegrityError as e:
            # SQL Server: 2627 (unique constraint), 2601 (duplicate key)
            msg = str(e)
            if any(code in msg for code in ("2627", "2601")) or "UNIQUE" in msg.upper() or "duplicate" in msg.lower():
                self.conn.rollback()
                return False, "duplicate"
            self.conn.rollback()
            return False, msg
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, str(e)

    def fetch_card_referiments(self, label_code):
        """
        Carica i riferimenti scheda per una label (LabelCod).
        Ritorna una lista di stringhe (Referiment).
        """
        query = """
        select 
            ProductRiferiments.CodRiferimento As Referiment
        from LabelCodes 
        inner join Boards on LabelCodes.IDBoard=Boards.IDBoard
        inner join Orders on Orders.IDOrder = Boards.IDOrder
        inner join Products on Products.IDProduct = Orders.IDProduct
        inner join ProductComponentsErp on Products.IDProduct = ProductComponentsErp.IDProduct
        inner join ProductRiferiments on ProductComponentsErp.IDProductCompErp = ProductRiferiments.IDProductCompErp
        inner join Components on Components.IDComponent = ProductComponentsErp.IDComponent
        inner join ParentPhases on ParentPhases.IDParentPhase = ProductRiferiments.IDParentPhase
        where LabelCodes.LabelCod = ?
        order by ParentPhases.ParentPhaseName,
                 ProductRiferiments.CodRiferimento + ' [' + ParentPhases.ParentPhaseName + ']';
        """
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query, label_code)
            rows = cur.fetchall()
            return [getattr(r, 'Referiment', r[0]) for r in rows] if rows else []
        except Exception as e:
            self.last_error_details = str(e)
            return []
        finally:
            try:
                if cur: cur.close()
            except Exception:
                pass

    def get_calibration_details(self, calibration_id):
        """Recupera i dettagli completi di una calibrazione specifica"""
        try:
            query = """
                    SELECT [CalibrationID]
                            , e.[EquipmentID]
                            , e.InternalName
                            , s.SiteName
                            , [CalibratedOn]
                            , [ExpireOn]
                            , [NrCertificate]
                            , [User]
                            , cast (c.[DateSys] as date) as DateIn
                            , [WarningSentOn]
                            , [IsValid]
                    FROM [Traceability_RS].[eqp].[Calibrations] AS C
                        inner join traceability_rs.eqp.equipments as E \
                    on e.EquipmentId= c.EquipmentID
                        inner join traceability_rs.eqp.EquipmentBrands eb on eb.EquipmentBrandId=e.BrandId
                        inner join traceability_rs.dbo.sites as s on s.IDSite = eb.CompanyId
                    WHERE [CalibrationID] = ? \
                    """

            self.cursor.execute(query, (calibration_id,))
            row = self.cursor.fetchone()

            if row:
                # Crea un namedtuple o oggetto con attributi per accesso facilitato
                CalibrationDetails = collections.namedtuple(
                    'CalibrationDetails',
                    ['CalibrationID', 'EquipmentID', 'InternalName', 'SiteName',
                     'CalibratedOn', 'ExpireOn', 'NrCertificate', 'User',
                     'DateIn', 'WarningSentOn', 'IsValid']
                )
                return CalibrationDetails(*row)
            else:
                return None

        except Exception as e:
            logger.error(f"Errore nel recupero dettagli calibrazione {calibration_id}: {e}")
            raise

    def update_calibration(self, calibration_id, expiry_date, supplier_id, pdf_bytes, username):
        """Aggiorna una calibrazione esistente"""
        try:
            query = """
                    UPDATE [Traceability_RS].[eqp].[Calibrations]
                    SET [ExpireOn] = ?, [SupplierId] = ?, [NrCertificate] = ?, [User] = ?, [DateSys] = GETDATE()
                    WHERE [CalibrationID] = ? \
                    """

            self.cursor.execute(query, (expiry_date, supplier_id, pdf_bytes, username, calibration_id))
            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            logger.error(f"Errore nell'aggiornamento calibrazione {calibration_id}: {e}")
            raise

    def invalidate_previous_calibrations(self, equipment_id):
        """Imposta IsValid = 0 per tutte le calibrazioni precedenti di un'attrezzatura"""
        try:
            query = """
                    UPDATE [Traceability_RS].[eqp].[Calibrations]
                    SET [IsValid] = 0, [DateSys] = GETDATE()
                    WHERE [EquipmentID] = ?
                      AND [IsValid] = 1
                      AND [ExpireOn] \
                        < GETDATE() \
                    """

            self.cursor.execute(query, (equipment_id,))
            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            logger.error(f"Errore nell'invalidazione calibrazioni per equipment {equipment_id}: {e}")
            raise

    def fetch_calibration_warnings(self):
        """
        Elenco attrezzature con calibrazione mancante o in scadenza (<=7 giorni)
        """
        query = """
        SELECT 
            e.EquipmentId,
            e.InternalName + ' [Inventory: ' + ISNULL(e.InventoryNumber, '#N/DD') + ']' AS Equipment,
            CAST(c1.CalibratedOn AS date) AS LastCalibrationDate,
            s1.SiteName AS CalibratedBy,
            '' as NrCertificate,
            CAST(c1.ExpireOn AS date) AS ExpireOn,
            IIF(c1.CalibratedOn IS NULL, 'No calibration record!', '') AS [Note]
        FROM eqp.Equipments e
        OUTER APPLY (
            SELECT TOP 1 c.*
            FROM eqp.Calibrations c
            WHERE c.EquipmentID = e.EquipmentId
            ORDER BY c.CalibratedOn DESC
        ) c1
        LEFT JOIN dbo.Sites s1 ON s1.IDSite = c1.[SupplierId]
        LEFT JOIN eqp.EquipmentBrands eb ON eb.EquipmentBrandId = e.BrandId
        INNER JOIN dbo.Sites s ON s.IDSite = eb.CompanyId
        OUTER APPLY (
            SELECT TOP 1 w.WarningSentOn AS LastWarn
            FROM eqp.CalibrationWarnings w
            WHERE w.EquipmentId = e.EquipmentId
            ORDER BY w.WarningSentOn DESC
        ) w1
        WHERE e.MustCalibrated = 1
          AND (
                c1.ExpireOn IS NULL
                OR DATEDIFF(day, GETDATE(), c1.ExpireOn) <= 7
              )
          AND (
                w1.LastWarn IS NULL
                OR DATEDIFF(day, w1.LastWarn, GETDATE()) >= 1
              )
        ORDER BY e.InternalName + ' [Inventory: ' + ISNULL(e.InventoryNumber,'#N/DD') + ']', c1.ExpireOn;
        """
        with self._lock:
            try:
                self.cursor.execute(query)
                return self.cursor.fetchall()
            except Exception as e:
                self.last_error_details = str(e)
                logger.error(f"Error fetching calibration warnings: {e}")
                return []

    def mark_calibration_warning_sent(self, equipment_ids):
        """
        Registra l'invio avviso per gli EquipmentId passati, inserendo una riga in eqp.CalibrationWarnings
        con WarningSentOn = GETDATE() per ciascun equipment.

        Ritorna True/False.
        """
        if not equipment_ids:
            return True

        sql = """INSERT INTO eqp.CalibrationWarnings (EquipmentId, WarningSentOn)
            SELECT ?, GETDATE()
            WHERE NOT EXISTS (
               SELECT 1 FROM eqp.CalibrationWarnings
               WHERE EquipmentId = ? AND CAST(WarningSentOn AS date) = CAST(GETDATE() AS date)
            );
            """
        cur = None
        try:
            cur = self.conn.cursor()
            # normalizza a int e crea tuples per executemany
            params = [(int(eid), int(eid)) for eid in equipment_ids]
            cur.executemany(sql, params)
            self.conn.commit()
            return True
        except Exception as e:
            if self.conn:
                self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            try:
                if cur: cur.close()
            except Exception:
                pass

    def fetch_unassigned_submissions(self):
        query = """
        SELECT
            CAST(se.SegnalazioneId AS int) AS SegID,
            CAST(se.DataSegnalazione AS date) AS DataSegnalazione,
            e.EmployeeName + ' ' + e.EmployeeSurname AS InputFrom,
            se.Titolo,
            se.Descrizione,
            '' Documents,
            sts.TipoStato,
            IIF(sts.StatoChiuso = 0, 'OPEN', 'CLOSED') AS StatoType,
            eaFrom.WorkEmail AS Email,
            '' as DescrizioneDocumento
        FROM Employee.dbo.Segnalazioni se
        INNER JOIN Employee.dbo.SegnalazioneStati ss
            ON ss.SegnalazioneId = se.SegnalazioneId              -- join corretto sull'ID segnalazione
        INNER JOIN Employee.dbo.SegnalazioniTipoStati sts
            ON sts.SegnalazioniTipoStatoId = ss.SegnalazioniTipoStatoId
        LEFT JOIN Employee.dbo.EmployeeHireHistory h
            ON h.EmployeeHireHistoryId = se.IdDipendente
        LEFT JOIN Employee.dbo.Employees e
            ON e.EmployeeId = h.EmployeeId
        LEFT JOIN Employee.dbo.EmployeeAddress eaFrom
            ON eaFrom.EmployeeId = e.EmployeeId
           AND eaFrom.DateOut IS NULL
        
        WHERE sts.StatoChiuso = 0
          AND NOT EXISTS (
              SELECT 1
              FROM Employee.dbo.SegnalazioneAssegnazioni sea
              WHERE sea.SegnalazioneId = se.SegnalazioneId
          )
        ORDER BY se.DataSegnalazione DESC;

        select * from Employee.dbo.SegnalazioniStatiAllegati
        """
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query)
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
            # ritorna lista di dict con chiavi: SegID, DataSegnalazione, InputFrom, Titolo, Descrizione, Documents, TipoStato, StatoType, Email
            return [dict(zip(cols, row)) for row in rows]
        except Exception as e:
            self.last_error_details = str(e)
            return []
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def fetch_assignable_employees(self):
        query = """
        SELECT h.EmployeeHireHistoryId,
               UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS Employee,
               ea.WorkEmail
        FROM employee.dbo.employees e
        INNER JOIN employee.dbo.EmployeeHireHistory h ON e.EmployeeId = h.EmployeeId
        INNER JOIN employee.dbo.EmployeeCdcStories ec ON ec.EmployeeHireHistoryId = h.EmployeeHireHistoryId
                                                      AND ec.DateOut IS NULL
                                                      AND h.employeerid = 2
                                                      AND h.EndWorkDate IS NULL 
        INNER JOIN employee.dbo.Functions f ON ec.FunctionId = f.FunctionId
        INNER JOIN employee.dbo.EmployeeAddress ea ON ea.EmployeeId = e.EmployeeId
                                                 AND ea.DateOut IS NULL
        WHERE f.IsStructure = 1
        ORDER BY UPPER(e.EmployeeSurname + ' ' + e.EmployeeName);
        """
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query)
            return cur.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            return []
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def assign_submission(self, segnalazione_id: int, employee_hire_history_id: int,
                          note: str = None):
        """
        Registra l'assegnazione:
          - Inserisce in SegnalazioneAssegnazioni (marca come assegnata)
          - Inserisce la nota in SegnalazioneSvolgimento
        """
        q_assign = """
            INSERT INTO Employee.dbo.SegnalazioneAssegnazioni
                (SegnalazioneId, EmployeeHireHistoryId, [Note])
            VALUES (?, ?, ?);
        """
        q_progress = """
            INSERT INTO Employee.dbo.SegnalazioneStati
                (SegnalazioneId, SegnalazioniTipoStatoId, DataAttivita,[Nota], [OperatoDa])
            VALUES (?, 1, GetDate(),'Assigned', 'System');
        """
        cur = None
        try:
            cur = self.conn.cursor()
            # Cast difensivo per evitare clash di tipi
            seg_id = int(segnalazione_id)
            ehh_id = int(employee_hire_history_id)

            cur.execute(q_assign, seg_id, ehh_id, note if note else None)
            cur.execute(q_progress, seg_id )# , ehh_id)

            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def insert_scrap_reason(self, reason):
        query = "INSERT INTO Traceability_RS.dbo.ScrapResons ([Reason]) VALUES (?);"
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query, reason.strip())
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def update_scrap_reason(self, scrap_reason_id, reason):
        query = "UPDATE Traceability_RS.dbo.ScrapResons SET [Reason] = ? WHERE ScrapReasonId = ?;"
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query, reason.strip(), scrap_reason_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def delete_scrap_reason(self, scrap_reason_id):
        query = "DELETE FROM Traceability_RS.dbo.ScrapResons WHERE ScrapReasonId = ?;"
        cur = None
        try:
            cur = self.conn.cursor()
            cur.execute(query, scrap_reason_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False
        finally:
            try:
                if cur: cur.close()
            except:
                pass

    def get_scrap_label_info(self, label_code):
        """
        Verifica il codice etichetta (scheda scrap).
        Ritorna la riga con: IDLabelCode (se esiste), OrderNumber, OrderDate (formattata), OrderQuantity, ProductCode, IDBoard.
        """
        query = """
            SELECT 
                l.IDLabelCode,           -- se non esiste la colonna, verrÃ  None nell'accesso attributo
                b.IDBoard,
                o.OrderNumber,
                FORMAT(o.OrderDate,'d','ro-ro') AS OrderDate,
                o.OrderQuantity,
                p.ProductCode
            FROM Traceability_RS.dbo.LabelCodes L
            INNER JOIN Traceability_RS.dbo.boards b ON b.idboard = l.IDBoard
            INNER JOIN Traceability_RS.dbo.orders o ON o.IDOrder = b.IDOrder
            INNER JOIN Traceability_RS.dbo.products p ON p.idproduct = o.idproduct
            WHERE l.labelcod = ?
        """
        try:
            cur = self.conn.cursor()
            cur.execute(query, label_code)
            return cur.fetchone()
        except Exception as e:
            self.last_error_details = str(e)
            return None
        finally:
            try:
                cur.close()
            except:
                pass

    def get_label_id_by_code(self, label_code):
        """
        Recupera l'ID della label se la colonna IDLabelCode esiste, altrimenti prova a dedurre dall'IDBoard.
        """
        try:
            cur = self.conn.cursor()
            # Primo tentativo: IDLabelCode
            cur.execute("SELECT IDLabelCode FROM Traceability_RS.dbo.LabelCodes WHERE LabelCod = ?", label_code)
            row = cur.fetchone()
            if row and hasattr(row, 'IDLabelCode') and row.IDLabelCode is not None:
                return row.IDLabelCode

            # Fallback (se non esiste IDLabelCode, NON lo forziamo a IDBoard: lasciamo None)
            return None
        except Exception as e:
            self.last_error_details = str(e)
            return None
        finally:
            try:
                cur.close()
            except:
                pass

    def fetch_origin_areas_for_scrap(self):
        """
        Carica la combo 'Area di provenienza' dalle ParentPhases (filtrate).
        """
        query = """
            SELECT idArea as idParentPhase, AreaName as ParentPhaseName
            FROM Traceability_RS.dbo.Areas           
            ORDER BY AreaName;
        """
        try:
            cur = self.conn.cursor()
            cur.execute(query)
            return cur.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            return []
        finally:
            try:
                cur.close()
            except:
                pass

    def fetch_scrap_reasons(self):
        """
        Carica la combo 'Motivo' da dbo.ScrapResons (nome tabella fornito).
        """
        query = """
            SELECT ScrapReasonId, [Reason]
            FROM Traceability_RS.dbo.ScrapResons
            ORDER BY [Reason];
        """
        try:
            cur = self.conn.cursor()
            cur.execute(query)
            return cur.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            return []
        finally:
            try:
                cur.close()
            except:
                pass

    def insert_scrap_declaration(self, user_name, id_label_code, id_parent_phase, scrap_reason_id,
                                 note, picture_bytes, riferiments=None):
        """
        Salva la dichiarazione nella tabella dbo.ScarpDeclarations.
        Tenta prima l'inserimento con la nuova colonna [Riferiments];
        se la colonna non esiste, fa fallback al vecchio INSERT.

        riferiments: stringa con riferimenti separati da ';' (puÃ² essere None o '')
        """
        if riferiments is None:
            riferiments = ""

        cur = None
        try:
            cur = self.conn.cursor()

            # Nuovo insert con Riferiments
            query_new = """
                INSERT INTO dbo.ScarpDeclarations
                    ([User], [IdLabelCode], [IDParentPhase], [ScrapReasonId], [Note], [Riferiments], [DateIn], [Picture])
                VALUES (?, ?, ?, ?, ?, ?, GETDATE(), ?);
            """
            cur.execute(query_new, user_name, id_label_code, id_parent_phase, scrap_reason_id,
                        note, riferiments, picture_bytes)
            self.conn.commit()
            return True

        except Exception as e:
            # Se la colonna non esiste, fallback al vecchio schema
            self.conn.rollback()
            msg = str(e).lower()
            if "invalid column" in msg and "riferiments" in msg:
                try:
                    cur = self.conn.cursor()
                    query_old = """
                        INSERT INTO dbo.ScarpDeclarations
                            ([User], [IdLabelCode], [IDParentPhase], [ScrapReasonId], [Note], [DateIn], [Picture])
                        VALUES (?, ?, ?, ?, ?, GETDATE(), ?);
                    """
                    cur.execute(query_old, user_name, id_label_code, id_parent_phase, scrap_reason_id,
                                note, picture_bytes)
                    self.conn.commit()
                    return True
                except Exception as e2:
                    self.conn.rollback()
                    self.last_error_details = str(e2)
                    return False
            else:
                self.last_error_details = str(e)
                return False
        finally:
            try:
                if cur: cur.close()
            except Exception:
                pass

    def authenticate_and_get_user(self, user_id, password):
        """
        Autentica un utente e, se ha successo, recupera i suoi dati e permessi.
        :return: Un oggetto User in caso di successo, altrimenti None.
        """
        # 1. Eseguiamo l'autenticazione esistente per verificare le credenziali
        employee_name = self.authenticate_user(user_id, password)
        logger.info("authenticate_user result for user_id=%r: %s", user_id,
                     "OK" if employee_name else "FAILED")

        if not employee_name:
            return None  # Login fallito

        # 2. Se l'autenticazione ha successo, carichiamo i permessi

        permissions_query = """
        select a.TranslationKey as PermissionKey, k.NomeUser as [user] 
            from AutorizedUsers a inner join employee.dbo.employeehirehistory h on a.EmployeeHireHistoryId=h.EmployeeHireHistoryId
            inner join employee.dbo.employees e on e.employeeid=h.EmployeeId inner join resetservices.dbo.tbuserkey k on e.employeeid=k.IdAnga
            where k.nomeuser=?

        """
        #      SELECT p.PermissionKey , ep.[user]
        #     FROM dbo.EmployeePermissions ep
        #     inner JOIN dbo.Permissions p ON ep.PermissionId = p.PermissionId
        #    WHERE ep.[user] = ?            
        # """
        permissions = set()
        cursor = None
        try:
            cursor = self.conn.cursor()
            logger.info("Executing permissions query with param user_id=%r", user_id)
            # Usiamo lo user_id numerico per la ricerca dei permessi
            cursor.execute(permissions_query, user_id)
            # Creiamo un set con tutte le chiavi di permesso trovate
            permissions = {row.PermissionKey for row in cursor.fetchall()}
            logger.info(f"Permissions for user_id={user_id}: {permissions}")
        except Exception as e:
            logger.error(f"ATTENZIONE: Impossibile caricare i permessi per l'utente {user_id}. Errore: {e}")
            # L'utente potrÃ  accedere ma non avrÃ  permessi speciali
        finally:
            if cursor:
                cursor.close()

        # 3. Creiamo e restituiamo l'oggetto User completo
        return User(name=employee_name, permissions=permissions)

    def authenticate_and_get_user_simple(self, user_id, password):
            """
            Autentica un utente e, se ha successo, recupera i suoi dati e permessi.
            :return: Un oggetto User in caso di successo, altrimenti None.
            """
            # 1. Eseguiamo l'autenticazione esistente per verificare le credenziali
            employee_name = self.authenticate_user(user_id, password)
            logger.info("authenticate_user result for user_id=%r: %s", user_id,
                        "OK" if employee_name else "FAILED")

            if not employee_name:
                return None  # Login fallito

            return User(name=employee_name)        

    def get_calibratable_equipment(self):
        """
        Recupera dal database la lista di tutte le attrezzature che
        richiedono una calibrazione.
        :return: Una lista di righe (oggetti row) o None in caso di errore.
        """
        query = """
            SELECT e.EquipmentId, e.InternalName, e.InventoryNumber, eb.Brand, s.SiteName
            FROM eqp.equipments e 
            INNER JOIN eqp.EquipmentBrands eb ON e.BrandId = eb.EquipmentBrandId
            INNER JOIN dbo.Sites s ON eb.CompanyId = s.IDSite
            INNER JOIN eqp.EquipmentTypes et ON e.EquipmentTypeId = et.EquipmentTypeId
            WHERE e.MustCalibrated = 1
            ORDER BY e.InternalName
        """
        cursor = None
        try:
            # Assumo che la tua connessione si chiami self.conn
            cursor = self.conn.cursor()
            cursor.execute(query)
            return cursor.fetchall()
        finally:
            if cursor:
                cursor.close()

    def get_last_calibration(self, equipment_id):
        """
        Recupera l'ultima registrazione di calibrazione per una specifica attrezzatura.
        :param equipment_id: L'ID dell'attrezzatura da cercare.
        :return: Una singola riga (oggetto row) o None se non trovata.
        """
        query = """
            SELECT TOP (1)
            c.CalibratedOn,
            c.ExpireOn,
            c.NrCertificate
            FROM eqp.Calibrations AS c
            WHERE c.EquipmentId = ?
        ORDER BY c.CalibratedOn DESC, c.CalibrationId DESC;
        """
        cursor = None
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, equipment_id)
            return cursor.fetchone()
        finally:
            if cursor:
                cursor.close()
    
    def get_all_calibrations_history(self, equipment_id):
        """
        Recupera tutte le calibrazioni per un equipment (incluse quelle con IsValid=0)
        Ordinate per data scadenza DESC per mostrare lo storico completo
        
        Args:
            equipment_id: ID dell'equipment
            
        Returns:
            Lista di tutte le calibrazioni per l'equipment
        """
        query = """
            SELECT 
                c.CalibrationID,
                c.EquipmentId,
                c.CalibratedOn,
                c.ExpireOn,
                c.SupplierId,
                c.NrCertificate,
                c.IsValid
            FROM eqp.Calibrations AS c
            WHERE c.EquipmentId = ?
            ORDER BY c.ExpireOn DESC
        """
        cursor = None
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, equipment_id)
            return cursor.fetchall()
        finally:
            if cursor:
                cursor.close()

    def get_suppliers(self):
        """
        NUOVO: Recupera la lista dei fornitori dal database.
        :return: Una lista di righe (oggetti row) con IDSite e SiteName.
        """
        query = "SELECT IDSite, SiteName FROM Sites WHERE IsSupplier = 1 ORDER BY SiteName;"
        cursor = None
        try:
            cursor = self.conn.cursor()
            cursor.execute(query)
            return cursor.fetchall()
        finally:
            if cursor:
                cursor.close()

    def add_new_calibration(self, equipment_id, expiry_date, supplier_id, pdf_bytes, username):
        """Inserisce una nuova calibrazione"""
        try:
            query = """
                    INSERT INTO [Traceability_RS].[eqp].[Calibrations]
                    ([EquipmentID], [SupplierId], [CalibratedOn], [ExpireOn], [NrCertificate], [User], [DateSys], [IsValid])
                    VALUES (?, ?, GETDATE(), ?, ?, ?, GETDATE(), 1) \
                    """

            self.cursor.execute(query, (equipment_id, supplier_id, expiry_date, pdf_bytes, username))
            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            logger.error(f"Errore nell'inserimento calibrazione per equipment {equipment_id}: {e}")
            raise

    def update_calibration(self, calibration_id, expiry_date, supplier_id, pdf_bytes, username):
        """Aggiorna una calibrazione esistente"""
        try:
            query = """
                    UPDATE [Traceability_RS].[eqp].[Calibrations]
                    SET [ExpireOn] = ?, [SupplierId] = ?, [NrCertificate] = ?, [User] = ?, [DateSys] = GETDATE()
                    WHERE [CalibrationID] = ? \
                    """

            self.cursor.execute(query, (expiry_date, supplier_id, pdf_bytes, username, calibration_id))
            self.conn.commit()

        except Exception as e:
            self.conn.rollback()
            logger.error(f"Errore nell'aggiornamento calibrazione {calibration_id}: {e}")
            raise

    # --- Room Booking Methods ---
    def fetch_meeting_rooms(self):
        """Recupera tutte le sale riunioni."""
        query = "SELECT MeetingRoomId, MeetingRoomName FROM [Employee].[dbo].[MeetingRooms] ORDER BY MeetingRoomName"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch meeting rooms: {e}")
            return []

    def add_meeting_room(self, name):
        """Aggiunge una nuova sala riunioni."""
        query = "INSERT INTO [Employee].[dbo].[MeetingRooms] (MeetingRoomName) VALUES (?)"
        try:
            self.cursor.execute(query, name)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore add meeting room: {e}")
            return False

    def update_meeting_room(self, room_id, name):
        """Aggiorna il nome di una sala riunioni."""
        query = "UPDATE [Employee].[dbo].[MeetingRooms] SET MeetingRoomName = ? WHERE MeetingRoomId = ?"
        try:
            self.cursor.execute(query, name, room_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update meeting room: {e}")
            return False

    def delete_meeting_room(self, room_id):
        """Elimina una sala riunioni."""
        query = "DELETE FROM [Employee].[dbo].[MeetingRooms] WHERE MeetingRoomId = ?"
        try:
            self.cursor.execute(query, room_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore delete meeting room: {e}")
            return False

    def fetch_room_bookings(self, start_date, end_date, room_name=None):
        """Recupera le prenotazioni in un intervallo di date."""
        query = """
            SELECT BookingID, RoomName, MeetingTitle, Organizer, StartTime, EndTime, BookingStatus
            FROM [Employee].[dbo].[RoomBookings]
            WHERE StartTime < ? AND EndTime > ?
            AND BookingStatus <> 'Cancelled'
        """
        params = [end_date, start_date]
        if room_name:
            query += " AND RoomName = ?"
            params.append(room_name)
        
        query += " ORDER BY StartTime"
        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch room bookings: {e}")
            return []

    def check_room_availability(self, room_name, start_time, end_time, exclude_booking_id=None):
        """Verifica se una sala Ã¨ libera."""
        query = """
            SELECT COUNT(*) FROM [Employee].[dbo].[RoomBookings]
            WHERE RoomName = ? 
            AND StartTime < ? AND EndTime > ?
            AND BookingStatus <> 'Cancelled'
        """
        params = [room_name, end_time, start_time]
        if exclude_booking_id:
            query += " AND BookingID <> ?"
            params.append(exclude_booking_id)
            
        try:
            self.cursor.execute(query, params)
            count = self.cursor.fetchone()[0]
            return count == 0
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore check availability: {e}")
            return False

    def add_room_booking(self, room_name, title, organizer, start_time, end_time):
        """Aggiunge una prenotazione."""
        if not self.check_room_availability(room_name, start_time, end_time):
            return False, "Room not available"
            
        query = """
            INSERT INTO [Employee].[dbo].[RoomBookings]
            (RoomName, MeetingTitle, Organizer, StartTime, EndTime, BookingStatus)
            VALUES (?, ?, ?, ?, ?, 'Confirmed')
        """
        try:
            self.cursor.execute(query, room_name, title, organizer, start_time, end_time)
            self.conn.commit()
            return True, "Booking confirmed"
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore add booking: {e}")
            return False, str(e)

    def update_room_booking(self, booking_id, room_name, title, start_time, end_time):
        """Aggiorna una prenotazione."""
        if not self.check_room_availability(room_name, start_time, end_time, exclude_booking_id=booking_id):
             return False, "Room not available"

        query = """
            UPDATE [Employee].[dbo].[RoomBookings]
            SET RoomName = ?, MeetingTitle = ?, StartTime = ?, EndTime = ?
            WHERE BookingID = ?
        """
        try:
            self.cursor.execute(query, room_name, title, start_time, end_time, booking_id)
            self.conn.commit()
            return True, "Booking updated"
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update booking: {e}")
            return False, str(e)

    def cancel_room_booking(self, booking_id):
        """Cancella (o marca come cancellata) una prenotazione."""
        query = "UPDATE [Employee].[dbo].[RoomBookings] SET BookingStatus = 'Cancelled' WHERE BookingID = ?"
        try:
            self.cursor.execute(query, booking_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore cancel booking: {e}")
            return False

    # --- Guest Management Methods ---
    def fetch_guests(self, filters=None):
        """Recupera la lista degli ospiti con filtri opzionali."""
        query = """
            SELECT [VisitorId], [RegistryId], [CompanyName], [GuestName], 
                   [StartVisit], [EndVisit], [Pourpose], [WelcomeMessage],
                   [SponsorGuy], [Logo]
            FROM [Employee].[dbo].[Visitors]
            WHERE [IsActive] = 0
        """
        params = []
        
        if filters:
            if filters.get('start_date'):
                query += " AND [StartVisit] >= ?"
                params.append(filters['start_date'])
            if filters.get('end_date'):
                query += " AND [EndVisit] <= ?"
                params.append(filters['end_date'])
            if filters.get('company'):
                query += " AND [CompanyName] LIKE ?"
                params.append(f"%{filters['company']}%")
            if filters.get('guest_name'):
                query += " AND [GuestName] LIKE ?"
                params.append(f"%{filters['guest_name']}%")
                
        query += " ORDER BY [StartVisit] DESC"
        
        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch guests: {e}")
            return []

    def fetch_distinct_guest_info(self):
        """Recupera liste distinte di aziende e nomi ospiti per autocomplete."""
        info = {'companies': [], 'guests': []}
        try:
            # Companies
            self.cursor.execute("SELECT DISTINCT [CompanyName] FROM [Employee].[dbo].[Visitors] WHERE [CompanyName] IS NOT NULL ORDER BY [CompanyName]")
            info['companies'] = [row[0] for row in self.cursor.fetchall()]
            
            # Guests
            self.cursor.execute("SELECT DISTINCT [GuestName] FROM [Employee].[dbo].[Visitors] WHERE [GuestName] IS NOT NULL ORDER BY [GuestName]")
            info['guests'] = [row[0] for row in self.cursor.fetchall()]
            
            return info
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch distinct guest info: {e}")
            return info

    def fetch_sponsors(self):
        """Recupera la lista degli sponsor."""
        query = """
            select upper(e.employeesurname+ ' ' + e.employeename ) as Sponsor,
            c.CdcDescription + ' (' + f.FunctionDescription +')' As Department
            from employee.dbo.employees e 
            inner join employee.dbo.EmployeeHireHistory H on e.employeeid=h.employeeid and h.employeerid = 2 and h.EndWorkDate is null 
            inner join employee.dbo.EmployeeCdcStories ec on h.employeehirehistoryid = ec.employeehirehistoryid and ec.dateout is null
            inner join employee.dbo.Functions F on f.FunctionId=ec.FunctionId
            inner join employee.dbo.CdcSub cs on cs.SubCdcId=ec.SubCdcId
            inner join employee.dbo.CostCenters c on c.CdcId=cs.CdcId
            where f.FunctionCode >= 45
            order by e.employeename + ' ' + e.employeesurname;
        """
        try:
            self.cursor.execute(query)
            return [f"{row.Sponsor} - {row.Department}" for row in self.cursor.fetchall()]
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch sponsors: {e}")
            return []

    def fetch_guests_by_company(self, company_name):
        """Recupera i nomi degli ospiti associati a una specifica azienda."""
        query = "SELECT DISTINCT [GuestName] FROM [Employee].[dbo].[Visitors] WHERE [CompanyName] = ? ORDER BY [GuestName]"
        try:
            self.cursor.execute(query, company_name)
            return [row[0] for row in self.cursor.fetchall()]
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch guests by company: {e}")
            return []

    def get_registry_id(self, user_name):
        """Recupera un nuovo RegistryId tramite SP."""
        try:
            # Parametri per la SP Employee.dbo.Registro
            # @RegistryTypeId = 930
            # @IussedBy = user_name
            # @EmployeerId = 1
            # @RegistroId OUTPUT
            
            # Nota: pyodbc gestisce i parametri di output in modo specifico, 
            # ma spesso Ã¨ piÃ¹ semplice fare una SELECT dopo o usare sintassi specifica.
            # Qui proviamo con la sintassi standard T-SQL EXEC
            
            sql = """
                DECLARE @NewId INT;
                EXEC [Employee].[dbo].[Registro]
                    @RegistryTypeId = 930,
                    @anno = ?,
                    @DataDocumento = ?,
                    @IussedBy = ?,
                    @EmployeerId = 1,
                    @RegistroId = @NewId OUTPUT;
                SELECT @NewId;
            """
            
            now = datetime.now()
            anno = now.year
            data_doc = now.date()
            
            self.cursor.execute(sql, anno, data_doc, user_name)
            row = self.cursor.fetchone()
            if row:
                return row[0]
            return None
            
        except Exception as e:
            self.last_error_details = str(e)
            logger.error(f"Errore get_registry_id: {e}")
            return None

    def add_guest(self, registry_id, company, name, start, end, pourpose, welcome_msg, sponsor_guy, logo_data):
        """Aggiunge un nuovo ospite."""
        query = """
            INSERT INTO [Employee].[dbo].[Visitors]
            ([RegistryId], [CompanyName], [GuestName], [StartVisit], [EndVisit], 
             [Pourpose], [WelcomeMessage], [ShowFrom], [ShowUntil], [IsActive], [CreatedAt],
             [SponsorGuy], [Logo])
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, GETDATE(), ?, ?)
        """
        try:
            # ShowFrom/ShowUntil uguali a StartVisit/EndVisit
            # logo_data deve essere bytes o None
            self.cursor.execute(query, registry_id, company, name, start, end, pourpose, welcome_msg, start, end, sponsor_guy, logo_data)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore add guest: {e}")
            return False

    def update_guest(self, visitor_id, company, name, start, end, pourpose, welcome_msg, sponsor_guy, logo_data):
        """Aggiorna un ospite esistente."""
        query = """
            UPDATE [Employee].[dbo].[Visitors]
            SET [CompanyName] = ?, [GuestName] = ?, [StartVisit] = ?, [EndVisit] = ?,
                [Pourpose] = ?, [WelcomeMessage] = ?, [ShowFrom] = ?, [ShowUntil] = ?,
                [SponsorGuy] = ?, [Logo] = ?
            WHERE [VisitorId] = ?
        """
        try:
            self.cursor.execute(query, company, name, start, end, pourpose, welcome_msg, start, end, sponsor_guy, logo_data, visitor_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update guest: {e}")
            return False

    def delete_guest(self, visitor_id):
        """Elimina (o disattiva) un ospite."""
        # Soft delete
        query = "UPDATE [Employee].[dbo].[Visitors] SET [IsActive] = 1 WHERE [VisitorId] = ?"
        try:
            self.cursor.execute(query, visitor_id)
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore delete guest: {e}")
            return False

    def get_equipment_info(self, equipment_id):
        """Recupera le informazioni del macchinario dal database."""
        try:
            query = """
            SELECT eb.Brand + ' ' + 
                [InternalName] + '(' + [SerialNumber] + ')' as InternalName,[InventoryNumber] as SerialNumber
                FROM [Traceability_RS].[eqp].[Equipments] e inner join [eqp].[EquipmentBrands] eb on e.BrandId=eb.EquipmentBrandId 
                where e.equipmentid= ?
            """
            with self.conn.cursor() as cursor:
                cursor.execute(query, (equipment_id,))
                row = cursor.fetchone()
                return row
        except Exception as e:
            self.last_error_details = str(e)
            return None

    def fetch_clients_for_verification(self):
        """Recupera i clienti per la verifica associazione."""
        query = """
        select distinct p.idclient, fc.FinalClientName
        from traceability_rs.dbo.FinalClients fc 
        inner join traceability_rs.dbo.products p on p.idfinalclient = fc.idfinalclient 
        inner join traceability_rs.dbo.Clients c on c.IDClient = p.idclient 
        order by fc.FinalClientName, p.idclient
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_orders_by_year_and_client(self, year, client_id=None):
        """Recupera gli ordini per anno e cliente."""
        query = """
        select o.idorder, o.OrderNumber, p.productcode, fc.FinalClientName
        from traceability_rs.dbo.products p 
        inner join traceability_rs.dbo.Clients c on c.IDClient = p.idclient 
        inner join traceability_rs.dbo.orders o on p.idproduct = o.IDProduct 
        inner join traceability_rs.dbo.FinalClients fc on fc.idfinalclient = p.idfinalclient
        where year(o.OrderDate) = ?
        """
        params = [year]

        if client_id:
            query += " AND p.idclient = ?"
            params.append(client_id)

        query += " order by o.ordernumber"

        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def verify_label_association(self, label_code):
        """Verifica l'associazione di un'etichetta."""
        query = """
        Select distinct o.OrderNumber, p.ProductCode, l.LabelCod, b.IDBoard
        from traceability_rs.dbo.orders as O 
        inner join traceability_rs.dbo.products as P on p.idproduct = o.idproduct 
        inner join traceability_rs.dbo.boards as B on b.IDOrder = o.idorder 
        inner join traceability_rs.dbo.LabelCodes as L on l.IDBoard = b.IDBoard
        where l.IDBoard in (select IDBoard from LabelCodes where LabelCod = ?)
        """
        try:
            self.cursor.execute(query, label_code)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_final_clients_for_linking(self):
        """Recupera i clienti finali per il filtro."""
        query = """
        SELECT fc.IDFinalClient, fc.FinalClientName, fc.AcronimForCode
        FROM 
         FinalClients fc 
        ORDER BY fc.FinalClientName
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_final_products_by_client(self, client_id):
        """Recupera i prodotti finali per un cliente specifico."""
        query = """
        SELECT idproduct, productcode, productname, ProductCodClienteFinal, isnull(Version, '') as Version
        FROM traceability_rs.dbo.products
        WHERE idfinalclient = ? AND IsFinalProduct = 1
        ORDER BY productcode
        """
        try:
            self.cursor.execute(query, client_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_semi_products_by_client(self, client_id):
        """Recupera i semilavorati per un cliente specifico."""
        query = """
        SELECT idproduct, productcode, productname, isnull(Version, '') as Version
        FROM traceability_rs.dbo.products
        WHERE (idfinalclient = ? ) 
        AND (IsFinalProduct = 0 OR IsFinalProduct IS NULL)
        AND CHARINDEX('cipr', productcode, 1) = 0
        ORDER BY ProductCode
        """
        try:
            self.cursor.execute(query, client_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_existing_links(self, final_product_id=None, client_id=None):
        """Recupera i collegamenti esistenti."""
        query = """
        SELECT pl.ProductLInkedTableId, pl.IdProductFinal, pl.IdProductSemi, pl.Dateout,
               pf.ProductCode as FinalCode, pf.ProductName as FinalName,
               ps.ProductCode as SemiCode, ps.ProductName as SemiName,
               fc.FinalClientName,
               isnull(pf.Version, '') as FinalVersion,
               isnull(ps.Version, '') as SemiVersion
        FROM traceability_rs.dbo.ProductsLinked pl
        INNER JOIN traceability_rs.dbo.products pf ON pl.IdProductFinal = pf.idproduct
        INNER JOIN traceability_rs.dbo.products ps ON pl.IdProductSemi = ps.idproduct
        INNER JOIN traceability_rs.dbo.FinalClients fc ON pf.idfinalclient = fc.IDFinalClient
        WHERE pl.Dateout IS NULL
        """
        params = []

        if final_product_id:
            query += " AND pl.IdProductFinal = ?"
            params.append(final_product_id)
        elif client_id:
            query += " AND pf.idfinalclient = ?"
            params.append(client_id)

        query += " ORDER BY fc.FinalClientName, pf.ProductCode, ps.ProductCode"

        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def add_product_link(self, final_product_id, semi_product_id):
        """Aggiunge un nuovo collegamento."""
        # Prima verifica se esiste giÃ  un collegamento attivo
        check_query = """
        SELECT COUNT(*) FROM traceability_rs.dbo.ProductsLinked 
        WHERE IdProductFinal = ? AND IdProductSemi = ? AND Dateout IS NULL
        """
        try:
            count = self.cursor.execute(check_query, final_product_id, semi_product_id).fetchval()
            if count > 0:
                return False, "Esiste giÃ  un collegamento attivo tra questi prodotti."

            insert_query = """
            INSERT INTO dbo.ProductsLinked (IdProductFinal, IdProductSemi)
            VALUES (?, ?)
            """
            self.cursor.execute(insert_query, final_product_id, semi_product_id)
            self.conn.commit()
            return True, "Collegamento aggiunto con successo."

        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_product_link(self, link_id):
        """Esegue un soft delete del collegamento."""
        query = "UPDATE dbo.ProductsLinked SET Dateout = GETDATE() WHERE ProductLInkedTableId = ?"
        try:
            self.cursor.execute(query, link_id)
            self.conn.commit()
            return True, "Collegamento eliminato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_final_products(self):
        """Recupera tutti i prodotti con informazioni sui clienti finali."""
        query = """
           select idproduct, upper(ProductCode) as ProductCode, ProductName, 
                  isnull(ProductCodClienteFinal,'#ND') as ProductCodClienteFinal, 
                  p.IsFinalProduct, fc.FinalClientName, AcronimForCode,
                  p.idfinalclient, isnull(p.Version, '') as Version
           from dbo.products as P 
           left join FinalClients FC on fc.IDFinalClient = p.idfinalclient
           where charindex('cipr', p.productcode, 1) = 0
           order by p.ProductCode
           """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_final_clients_for_products(self):
        """Recupera i clienti finali per la selezione."""
        query = """
           SELECT [IDFinalClient], [FinalClientName], [AcronimForCode]
           FROM [Traceability_RS].[dbo].[FinalClients] 
           ORDER BY FinalClientName
           """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def update_product_final_info(self, product_id, is_final_product, final_client_id, customer_code, version=None):
        """Aggiorna le informazioni di prodotto finale."""
        query = """
           UPDATE dbo.products 
           SET IsFinalProduct = ?, idfinalclient = ?, ProductCodClienteFinal = ?, Version = ?
           WHERE idproduct = ?
           """
        try:
            # Converti i valori per il database
            is_final_int = 1 if is_final_product else 0
            final_client_id = final_client_id if final_client_id else None
            customer_code = customer_code if customer_code else None
            version = version if version else None

            self.cursor.execute(query, is_final_int, final_client_id, customer_code, version, product_id)
            self.conn.commit()
            return True, "Prodotto aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_final_customers(self):
        """Recupera tutti i clienti finali."""
        query = """
        SELECT [IDFinalClient], [FinalClientName], [FinalClientFullName], [AcronimForCode],
               [ClientAddress], [ClientCity], [ClientZIP], [ClientCountry], [VatCode]
        FROM [Traceability_RS].[dbo].[FinalClients] 
        ORDER BY [FinalClientName]
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def add_final_customer(self, name, full_name, acronim, address, city, zip_code, country, vat_code):
        """Aggiunge un nuovo cliente finale."""
        query = """
        INSERT INTO [Traceability_RS].[dbo].[FinalClients] 
        (FinalClientName, FinalClientFullName, AcronimForCode, ClientAddress, ClientCity, ClientZIP, ClientCountry, VatCode)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """
        try:
            self.cursor.execute(query, name, full_name, acronim, address, city, zip_code, country, vat_code)
            self.conn.commit()
            return True, "Cliente aggiunto con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def update_final_customer(self, customer_id, name, full_name, acronim, address, city, zip_code, country, vat_code):
        """Aggiorna un cliente finale esistente."""
        query = """
        UPDATE [Traceability_RS].[dbo].[FinalClients] 
        SET FinalClientName = ?, FinalClientFullName = ?, AcronimForCode = ?,
            ClientAddress = ?, ClientCity = ?, ClientZIP = ?, ClientCountry = ?, VatCode = ?
        WHERE IDFinalClient = ?
        """
        try:
            self.cursor.execute(query, name, full_name, acronim, address, city, zip_code, country, vat_code,
                                customer_id)
            self.conn.commit()
            return True, "Cliente aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_final_customer(self, customer_id):
        """Elimina un cliente finale."""
        query = "DELETE FROM [Traceability_RS].[dbo].[FinalClients] WHERE IDFinalClient = ?"
        try:
            self.cursor.execute(query, customer_id)
            self.conn.commit()
            return True, "Cliente eliminato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_kce_products(self):
        query = "SELECT idProduct, productcode FROM dbo.PRODUCTS WHERE PRODUCTCODE LIKE '%KCE%' ORDER BY productcode;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def fetch_supplier_sites(self):
        """Recupera i fornitori/compagnie dalla tabella Sites."""
        query = "SELECT [IDSite], [SiteName] FROM [Traceability_RS].[dbo].[Sites] WHERE isSupplier = 1 ORDER BY SiteName;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_all_sites(self):
        """Recupera tutti i siti/compagnie per la gestione."""
        query = "SELECT IDSite, SiteName, SiteAddress, SiteVat, SiteCountry, Logo FROM dbo.Sites ORDER BY SiteName;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    # --- PASTE MANAGEMENT METHODS ---
    def fetch_paste_producers(self):
        """Recupera i produttori per le paste."""
        query = "SELECT [ProducerId], [Producers] FROM [Traceability_RS].[dbo].[Producers] ORDER BY Producers;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_paste_producers: {e}")
            return []

    def fetch_all_pastas(self):
        """Recupera tutte le paste configurate."""
        query = """
        SELECT p.Pastaid, p.ProducerId, p.CreatedAt, p.CreatedBy, p.PastaCode, 
               p.PastaDataSheet, p.DateEntry, pr.Producers
        FROM [Traceability_RS].[pst].[Pastas] p
        LEFT JOIN [Traceability_RS].[dbo].[Producers] pr ON p.ProducerId = pr.ProducerId
        WHERE p.DateEntry IS NOT NULL
        ORDER BY p.PastaCode
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_all_pastas: {e}")
            return []

    def fetch_pasta_by_id(self, pasta_id):
        """Recupera i dettagli di una pasta specifica."""
        query = """
        SELECT p.Pastaid, p.ProducerId, p.CreatedAt, p.CreatedBy, p.PastaCode, 
               p.PastaDataSheet, p.DateEntry, s.SiteName
        FROM [Traceability_RS].[pst].[Pastas] p
        LEFT JOIN [Traceability_RS].[dbo].[Sites] s ON p.ProducerId = s.IDSite
        WHERE p.Pastaid = ?
        """
        try:
            self.cursor.execute(query, pasta_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_pasta_by_id: {e}")
            return None

    def insert_pasta(self, producer_id, pasta_code, datasheet_data, user_id):
        """Inserisce una nuova pasta."""
        query = """
        INSERT INTO [Traceability_RS].[pst].[Pastas] 
        (ProducerId, PastaCode, PastaDataSheet, CreatedBy, CreatedAt, DateEntry)
        VALUES (?, ?, ?, ?, GETDATE(), GETDATE())
        """
        try:
            self.cursor.execute(query, producer_id, pasta_code, datasheet_data, user_id)
            self.conn.commit()
            logger.info(f"Pasta inserita: {pasta_code} da user {user_id}")
            return True, "Pasta inserita con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore insert_pasta: {e}")
            return False, f"Errore database: {e}"

    def update_pasta(self, pasta_id, producer_id, pasta_code, datasheet_data=None):
        """Aggiorna una pasta esistente."""
        if datasheet_data is not None:
            query = """
            UPDATE [Traceability_RS].[pst].[Pastas]
            SET ProducerId = ?, PastaCode = ?, PastaDataSheet = ?
            WHERE Pastaid = ?
            """
            params = (producer_id, pasta_code, datasheet_data, pasta_id)
        else:
            query = """
            UPDATE [Traceability_RS].[pst].[Pastas]
            SET ProducerId = ?, PastaCode = ?
            WHERE Pastaid = ?
            """
            params = (producer_id, pasta_code, pasta_id)
        
        try:
            self.cursor.execute(query, params)
            self.conn.commit()
            logger.info(f"Pasta aggiornata: ID {pasta_id}")
            return True, "Pasta aggiornata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update_pasta: {e}")
            return False, f"Errore database: {e}"

    def delete_pasta(self, pasta_id):
        """Elimina una pasta (soft delete impostando DateEntry a NULL)."""
        query = """
        UPDATE [Traceability_RS].[pst].[Pastas]
        SET DateEntry = NULL
        WHERE Pastaid = ?
        """
        try:
            self.cursor.execute(query, pasta_id)
            self.conn.commit()
            logger.info(f"Pasta eliminata (soft delete): ID {pasta_id}")
            return True, "Pasta eliminata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore delete_pasta: {e}")
            return False, f"Errore database: {e}"

    def insert_pasta_config(self, pasta_id, valability, low_temp, high_temp):
        """Inserisce la configurazione temperatura per una pasta."""
        query = """
        INSERT INTO [Traceability_RS].[pst].[PastaConfigs]
        (PastaId, Valability, LowTemperature, HighTemperature, DateIn)
        VALUES (?, ?, ?, ?, GETDATE())
        """
        try:
            self.cursor.execute(query, pasta_id, valability, low_temp, high_temp)
            self.conn.commit()
            logger.info(f"Configurazione pasta inserita per PastaId: {pasta_id}")
            return True, "Configurazione salvata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore insert_pasta_config: {e}")
            return False, f"Errore database: {e}"

    def update_pasta_config(self, pasta_id, valability, low_temp, high_temp):
        """Aggiorna la configurazione temperatura per una pasta."""
        # Prima chiude la configurazione esistente
        close_query = """
        UPDATE [Traceability_RS].[pst].[PastaConfigs]
        SET DateOut = GETDATE()
        WHERE PastaId = ? AND DateOut IS NULL
        """
        # Poi inserisce la nuova configurazione
        insert_query = """
        INSERT INTO [Traceability_RS].[pst].[PastaConfigs]
        (PastaId, Valability, LowTemperature, HighTemperature, DateIn)
        VALUES (?, ?, ?, ?, GETDATE())
        """
        try:
            self.cursor.execute(close_query, pasta_id)
            self.cursor.execute(insert_query, pasta_id, valability, low_temp, high_temp)
            self.conn.commit()
            logger.info(f"Configurazione pasta aggiornata per PastaId: {pasta_id}")
            return True, "Configurazione aggiornata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update_pasta_config: {e}")
            return False, f"Errore database: {e}"

    def fetch_pasta_config(self, pasta_id):
        """Recupera la configurazione attiva di una pasta."""
        query = """
        SELECT PastaConfigId, PastaId, Valability, LowTemperature, HighTemperature, DateIn, DateOut
        FROM [Traceability_RS].[pst].[PastaConfigs]
        WHERE PastaId = ? AND DateOut IS NULL
        """
        try:
            self.cursor.execute(query, pasta_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_pasta_config: {e}")
            return None

    # --- PASTE RECEPTION METHODS ---
    def generate_label_code(self):
        """Genera un nuovo codice etichetta progressivo con 12 zeri iniziali."""
        query = """
        SELECT MAX(CAST(SUBSTRING(LabelCode, 1, LEN(LabelCode)) AS BIGINT)) AS MaxCode
        FROM [Traceability_RS].[pst].[PastaLabelCodes]
        WHERE LabelCode LIKE '0%'
        """
        try:
            self.cursor.execute(query)
            row = self.cursor.fetchone()
            max_code = row[0] if row and row[0] else 0
            new_code = max_code + 1
            # Formatta con 12 zeri iniziali
            label_code = str(new_code).zfill(13)  # 13 cifre totali (12 zeri + numero)
            return label_code
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore generate_label_code: {e}")
            return None

    def insert_label_code(self, label_code):
        """Inserisce un nuovo codice etichetta."""
        query = """
        SET NOCOUNT ON;
        INSERT INTO [Traceability_RS].[pst].[PastaLabelCodes]
        (LabelCode, LabeCreationDate)
        VALUES (?, GETDATE());
        SELECT CAST(SCOPE_IDENTITY() AS INT) AS NewID;
        """
        try:
            self.cursor.execute(query, label_code)
            row = self.cursor.fetchone()
            label_id = row[0] if row else None
            self.conn.commit()
            logger.info(f"Label code inserito: {label_code}, ID: {label_id}")
            return label_id
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore insert_label_code: {e}")
            return None

    def delete_label_code(self, label_code_id):
        """Elimina un codice etichetta non utilizzato."""
        query = """
        DELETE FROM [Traceability_RS].[pst].[PastaLabelCodes]
        WHERE LabelCodeId = ?
        """
        try:
            self.cursor.execute(query, label_code_id)
            self.conn.commit()
            logger.info(f"Label code eliminato: ID={label_code_id}")
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore delete_label_code: {e}")
            return False


    def fetch_all_pastas_for_reception(self):
        """Recupera tutte le paste attive per il ricevimento."""
        query = """
        SELECT p.Pastaid, p.PastaCode, s.SiteName as ProducerName
        FROM [Traceability_RS].[pst].[Pastas] p
        LEFT JOIN [Traceability_RS].[dbo].[Sites] s ON p.ProducerId = s.IDSite
        WHERE p.DateEntry IS NOT NULL
        ORDER BY p.PastaCode
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_all_pastas_for_reception: {e}")
            return []

    def insert_pasta_log(self, pasta_id, label_code_id, user_name, incoming_doc=None):
        """Inserisce un log di ricevimento pasta."""
        query = """
        INSERT INTO [Traceability_RS].[pst].[PastaLogs]
        (PastaId, LabeCodeId, GetIn, [User], IncomingDoc)
        VALUES (?, ?, GETDATE(), ?, ?)
        """
        try:
            self.cursor.execute(query, pasta_id, label_code_id, user_name, incoming_doc)
            self.conn.commit()
            logger.info(f"Pasta log inserito: PastaId={pasta_id}, LabelCodeId={label_code_id}, User={user_name}")
            return True, "Ricevimento registrato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore insert_pasta_log: {e}")
            return False, f"Errore database: {e}"

    def fetch_pasta_logs(self, limit=100):
        """Recupera i log di ricevimento paste."""
        query = """
        SELECT TOP (?) 
            pl.PastaLogId, 
            p.PastaCode, 
            s.SiteName as ProducerName,
            lc.LabelCode,
            pl.GetIn,
            pl.[User],
            pl.GetOut,
            CASE WHEN pl.IncomingDoc IS NULL THEN 0 ELSE 1 END as HasDoc
        FROM [Traceability_RS].[pst].[PastaLogs] pl
        INNER JOIN [Traceability_RS].[pst].[Pastas] p ON pl.PastaId = p.Pastaid
        LEFT JOIN [Traceability_RS].[dbo].[Sites] s ON p.ProducerId = s.IDSite
        INNER JOIN [Traceability_RS].[pst].[PastaLabelCodes] lc ON pl.LabeCodeId = lc.LabelCodeId
        ORDER BY pl.GetIn DESC
        """
        try:
            self.cursor.execute(query, limit)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_pasta_logs: {e}")
            return []

    def fetch_pasta_log_document(self, log_id):
        """Recupera il documento di un log."""
        query = """
        SELECT IncomingDoc
        FROM [Traceability_RS].[pst].[PastaLogs]
        WHERE PastaLogId = ?
        """
        try:
            self.cursor.execute(query, log_id)
            row = self.cursor.fetchone()
            return row[0] if row else None
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_pasta_log_document: {e}")
            return None

    def update_pasta_log_document(self, log_id, document_data):
        """Aggiorna il documento di un log."""
        query = """
        UPDATE [Traceability_RS].[pst].[PastaLogs]
        SET IncomingDoc = ?
        WHERE PastaLogId = ?
        """
        try:
            self.cursor.execute(query, document_data, log_id)
            self.conn.commit()
            logger.info(f"Documento aggiornato per PastaLogId: {log_id}")
            return True, "Documento aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update_pasta_log_document: {e}")
            return False, f"Errore database: {e}"


    # --- PASTE REFRIGERATORS MANAGEMENT METHODS ---
    def fetch_all_refrigerators(self):
        """Recupera tutti i frigoriferi per paste."""
        query = """
        SELECT PastaStoreFrigiderId, PastaStoreFrigiderName, PastaStoreFrigiderLocation, IsConnected
        FROM [Traceability_RS].[pst].[PastaStoreFrigiders]
        ORDER BY PastaStoreFrigiderName
        """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            logger.error(f"Errore fetch_all_refrigerators: {e}")
            return []

    def insert_refrigerator(self, name, location, is_connected):
        """Inserisce un nuovo frigorifero."""
        query = """
        INSERT INTO [Traceability_RS].[pst].[PastaStoreFrigiders]
        (PastaStoreFrigiderName, PastaStoreFrigiderLocation, IsConnected)
        VALUES (?, ?, ?)
        """
        try:
            is_connected_int = 1 if is_connected else 0
            self.cursor.execute(query, name, location, is_connected_int)
            self.conn.commit()
            logger.info(f"Frigorifero inserito: {name}")
            return True, "Frigorifero inserito con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore insert_refrigerator: {e}")
            return False, f"Errore database: {e}"

    def update_refrigerator(self, refrigerator_id, name, location, is_connected):
        """Aggiorna un frigorifero esistente."""
        query = """
        UPDATE [Traceability_RS].[pst].[PastaStoreFrigiders]
        SET PastaStoreFrigiderName = ?, PastaStoreFrigiderLocation = ?, IsConnected = ?
        WHERE PastaStoreFrigiderId = ?
        """
        try:
            is_connected_int = 1 if is_connected else 0
            self.cursor.execute(query, name, location, is_connected_int, refrigerator_id)
            self.conn.commit()
            logger.info(f"Frigorifero aggiornato: ID {refrigerator_id}")
            return True, "Frigorifero aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore update_refrigerator: {e}")
            return False, f"Errore database: {e}"

    def delete_refrigerator(self, refrigerator_id):
        """Elimina un frigorifero."""
        query = """
        DELETE FROM [Traceability_RS].[pst].[PastaStoreFrigiders]
        WHERE PastaStoreFrigiderId = ?
        """
        try:
            self.cursor.execute(query, refrigerator_id)
            self.conn.commit()
            logger.info(f"Frigorifero eliminato: ID {refrigerator_id}")
            return True, "Frigorifero eliminato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            logger.error(f"Errore delete_refrigerator: {e}")
            return False, f"Errore database: {e}"



    def fetch_all(self, query, params=None):
        """
        Esegue una query SELECT e restituisce tutti i risultati
        """
        with self._lock:
            try:
                if params:
                    self.cursor.execute(query, params)
                else:
                    self.cursor.execute(query)

                results = self.cursor.fetchall()
                logger.debug(f"[DATABASE] fetch_all eseguito - Righe recuperate: {len(results)}")
                return results

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.exception(f"[DATABASE] Errore fetch_all: {e}")
                return []
            except Exception as e:
                self.last_error_details = str(e)
                logger.exception(f"[DATABASE] Errore inaspettato fetch_all: {e}")
                return []

    def get_claim_document(self, claim_log_id: int, output_path: str) -> bool:
        """
        Recupera il documento allegato a un reclamo e lo salva su disco

        Args:
            claim_log_id: ID del reclamo
            output_path: Percorso dove salvare il file

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            query = """
                    SELECT
                        [TransportDocumentData]
                    FROM [Traceability_RS].[clm].[ClaimLogs]
                    WHERE [ClaimLogId] = ? \
                    """

            result = self.fetch_one(query, (claim_log_id,))

            if result and result[0]:
                file_data = result[0]

                with open(output_path, 'wb') as f:
                    f.write(file_data)

                logger.info(f"[DATABASE] File salvato: {output_path} ({len(file_data)} bytes)")
                return True
            else:
                logger.warning(f"[DATABASE] Nessun documento trovato per reclamo {claim_log_id}")
                return False

        except Exception as e:
            logger.exception(f"[DATABASE] Errore recupero documento: {e}")
            return False

    def delete_claim_document(self, claim_log_id: int) -> bool:
        """
        Elimina il documento allegato a un reclamo

        Args:
            claim_log_id: ID del reclamo

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            query = """
                    UPDATE [Traceability_RS].[clm].[ClaimLogs]
                    SET [TransportDocumentData] = NULL
                    WHERE [ClaimLogId] = ? \
                    """

            return self.execute_query(query, (claim_log_id,))

        except Exception as e:
            logger.exception(f"[DATABASE] Errore eliminazione documento: {e}")
            return False

    def fetch_one(self, query, params=None):
        """
        Esegue una query SELECT e restituisce il primo risultato
        """
        with self._lock:
            try:
                if params:
                    self.cursor.execute(query, params)
                else:
                    self.cursor.execute(query)

                result = self.cursor.fetchone()
                return result

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                logger.exception(f"[DATABASE] Errore fetch_one: {e}")
                return None
            except Exception as e:
                self.last_error_details = str(e)
                logger.exception(f"[DATABASE] Errore inaspettato fetch_one: {e}")
                return None

    def execute_query(self, query, params=None):
        """
        Esegue una query (INSERT, UPDATE, DELETE) senza restituire risultati
        """
        with self._lock:
            try:
                if params:
                    self.cursor.execute(query, params)
                else:
                    self.cursor.execute(query)

                self.conn.commit()
                return True

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                self.conn.rollback()
                logger.exception(f"[DATABASE] Errore execute_query: {e}")
                return False
            except Exception as e:
                self.last_error_details = str(e)
                self.conn.rollback()
                logger.exception(f"[DATABASE] Errore inaspettato execute_query: {e}")
                return False

    def execute_query_with_id(self, query, params=None):
        """
        Esegue una query INSERT e restituisce l'ID della riga inserita (SCOPE_IDENTITY())
        """
        with self._lock:
            try:
                if params:
                    self.cursor.execute(query, params)
                else:
                    self.cursor.execute(query)

                # Recupera l'ID della riga inserita
                self.cursor.execute("SELECT SCOPE_IDENTITY() as id")
                result = self.cursor.fetchone()
                self.conn.commit()

                inserted_id = result[0] if result and result[0] is not None else None
                return inserted_id

            except pyodbc.Error as e:
                self.last_error_details = str(e)
                self.conn.rollback()
                logger.exception(f"[DATABASE] Errore execute_query_with_id: {e}")
                return None
            except Exception as e:
                self.last_error_details = str(e)
                self.conn.rollback()
                logger.exception(f"[DATABASE] Errore inaspettato execute_query_with_id: {e}")
                return None

###sALVATAGGIO DATI COMPLAIN
    def insert_claim_header(self, header) -> Optional[int]:
        """
        Inserisce la testata del reclamo e restituisce l'ID generato usando OUTPUT INSERTED.
        """
        try:
            # Usiamo OUTPUT INSERTED.ClaimLogId per ricevere immediatamente l'ID,
            # funziona meglio di SCOPE_IDENTITY() con pyodbc.
            query = """
                    INSERT INTO [Traceability_RS].[clm].[ClaimLogs]
                    ([ClaimTypeId]
                        , [IdProduct]
                        , [DateClaim]
                        , [AWB]
                        , [TransportDocument]
                        , [DateSys]
                        , [CustomerClaimNumber]
                        , [InternalClaimNumber]
                        , [ShortClaimDescription]
                        , [TargetDate]
                        , [Quantity]
                        , [ClaimDecisionId]
                        , [USERName])
                        OUTPUT INSERTED.ClaimLogId
                    VALUES
                        (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) \
                    """

            # Preparazione parametri (TransportDocumentData rimosso)
            params = (
                header.ClaimTypeId,
                header.IdProduct,
                header.DateClaim,
                header.AWB,
                header.TransportDocument,
                header.DateSys,
                header.CustomerClaimNumber,
                header.InternalClaimNumber,
                header.ShortClaimDescription,
                header.TargetDate,
                header.Quantity,
                header.ClaimDecisionId,
                header.USERName
            )

            cursor = self.conn.cursor()

            logger.debug(f"[DATABASE] Esecuzione INSERT header con InternalClaimNumber: {header.InternalClaimNumber}")

            # Eseguiamo la query
            cursor.execute(query, params)

            # Recuperiamo subito l'ID restituito dalla clausola OUTPUT
            row = cursor.fetchone()

            if row:
                new_id = int(row[0])
                self.conn.commit()  # Confermiamo la transazione solo se abbiamo l'ID
                cursor.close()
                logger.info(f"[DATABASE] Header inserito con successo. ClaimLogId generato: {new_id}")
                return new_id
            else:
                logger.error("[DATABASE] INSERT eseguita ma nessun ID restituito da OUTPUT INSERTED.")
                self.conn.rollback()
                cursor.close()
                return None

        except Exception as e:
            logger.exception(f"[DATABASE] Errore critico in insert_claim_header: {e}")
            try:
                self.conn.rollback()
            except:
                pass
            return None

    def insert_claim_details(self, claim_log_id: int, claim_details: List) -> bool:
        """
        Inserisce i dettagli di un reclamo nel database

        Args:
            claim_log_id: ID della testata del reclamo
            claim_details: Lista di dettagli (ClaimDetail objects o dict)

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            query = """
                    INSERT INTO [Traceability_RS].[clm].[ClaimDataLogs]
                    ([ClaimLogId],
                        [FirstInspectionResultId],
                        [LabelCod],
                        [RootCause],
                        [SummaryCorrectiveAction],
                        [SummaryPreventiveAction],
                        [ClaimStatusId],
                    [ClaimDefectId])
                    VALUES
                        (?, ?, ?, ?, ?, ?, ?, ?) \
                    """

            for detail in claim_details:
                # Se Ã¨ un dict
                if isinstance(detail, dict):
                    params = (
                        claim_log_id,
                        detail.get('FirstInspectionResultId'),
                        detail.get('LabelCod'),
                        detail.get('RootCause'),
                        detail.get('SummaryCorrectiveAction'),
                        detail.get('SummaryPreventiveAction'),
                        detail.get('ClaimStatusId'),
                        detail.get('ClaimDefectId')
                    )
                else:
                    # Se Ã¨ un oggetto ClaimDetail
                    params = (
                        claim_log_id,
                        detail.FirstInspectionResultId,
                        detail.LabelCod,
                        detail.RootCause,
                        detail.SummaryCorrectiveAction,
                        detail.SummaryPreventiveAction,
                        detail.ClaimStatusId,
                        detail.ClaimDefectId
                    )

                success = self.execute_query(query, params)
                if not success:
                    return False

            logger.info(f"[DATABASE] {len(claim_details)} dettagli reclamo inseriti per ClaimLogId={claim_log_id}")
            return True

        except Exception as e:
            logger.exception(f"[DATABASE] Errore inserimento dettagli reclamo: {e}")
            return False

    def update_claim_detail(self, detail_id: int, detail_data: Dict) -> bool:
        """
        Aggiorna un dettaglio specifico di un reclamo

        Args:
            detail_id: ID del dettaglio (ClaimLogDataId)
            detail_data: Dizionario con i dati da aggiornare

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            query = """
                    UPDATE [Traceability_RS].[clm].[ClaimDataLogs]
                    SET [FirstInspectionResultId] = ?,
                        [LabelCod] = ?,
                        [RootCause] = ?,
                        [SummaryCorrectiveAction] = ?,
                        [SummaryPreventiveAction] = ?,
                        [ClaimStatusId] = ?,
                        [ClaimDefectId] = ?
                    WHERE [ClaimLogDataId] = ? \
                    """

            params = (
                detail_data.get('FirstInspectionResultId'),
                detail_data.get('LabelCod'),
                detail_data.get('RootCause', ''),
                detail_data.get('SummaryCorrectiveAction', ''),
                detail_data.get('SummaryPreventiveAction', ''),
                detail_data.get('ClaimStatusId'),
                detail_data.get('ClaimDefectId'),
                detail_id
            )

            success = self.execute_query(query, params)
            
            if success:
                logger.info(f"[DATABASE] Dettaglio reclamo aggiornato: ClaimLogDataId={detail_id}")
            else:
                logger.error(f"[DATABASE] Errore aggiornamento dettaglio reclamo: ClaimLogDataId={detail_id}")
            
            return success

        except Exception as e:
            logger.exception(f"[DATABASE] Errore aggiornamento dettaglio reclamo: {e}")
            return False

    def get_claim_by_id(self, claim_log_id: int) -> Optional[Dict]:
        """
        Recupera una testata di reclamo dal database

        Args:
            claim_log_id: ID del reclamo

        Returns:
            dict: Dati del reclamo, None se non trovato
        """
        try:
            query = """
                    SELECT
                        [ClaimLogId], [ClaimTypeId], [IdProduct], [DateClaim], [AWB], [TransportDocument], [TransportDocumentData], [DateSys], [CustomerClaimNumber], [InternalClaimNumber], [ShortClaimDescription], [TargetDate], [Quantity], [ClaimDecisionId], [USERName]
                    FROM [Traceability_RS].[clm].[ClaimLogs]
                    WHERE [ClaimLogId] = ? \
                    """

            result = self.fetch_one(query, (claim_log_id,))

            if result:
                return {
                    'ClaimLogId': result[0],
                    'ClaimTypeId': result[1],
                    'IdProduct': result[2],
                    'DateClaim': result[3],
                    'AWB': result[4],
                    'TransportDocument': result[5],
                    'TransportDocumentData': result[6],
                    'DateSys': result[7],
                    'CustomerClaimNumber': result[8],
                    'InternalClaimNumber': result[9],
                    'ShortClaimDescription': result[10],
                    'TargetDate': result[11],
                    'Quantity': result[12],
                    'ClaimDecisionId': result[13],
                    'USERName': result[14]
                }

            return None

        except Exception as e:
            logger.exception(f"[DATABASE] Errore recupero reclamo: {e}")
            return None

    def get_claim_details(self, claim_log_id: int) -> List[Dict]:
        """
        Recupera i dettagli di un reclamo

        Args:
            claim_log_id: ID della testata del reclamo

        Returns:
            List[dict]: Lista dei dettagli del reclamo
        """
        try:
            query = """
                    SELECT
                        [ClaimLogDataId], [ClaimLogId], [FirstInspectionResultId], [LabelCod], [RootCause], [SummaryCorrectiveAction], [SummaryPreventiveAction], [ClaimStatusId], [ClaimDefectId]
                    FROM [Traceability_RS].[clm].[ClaimDataLogs]
                    WHERE [ClaimLogId] = ?
                    ORDER BY [ClaimLogDataId] \
                    """

            results = self.fetch_all(query, (claim_log_id,))

            details = []
            for row in results:
                details.append({
                    'ClaimLogDataId': row[0],
                    'ClaimLogId': row[1],
                    'FirstInspectionResultId': row[2],
                    'LabelCod': row[3],
                    'RootCause': row[4],
                    'SummaryCorrectiveAction': row[5],
                    'SummaryPreventiveAction': row[6],
                    'ClaimStatusId': row[7],
                    'ClaimDefectId': row[8]
                })

            return details

        except Exception as e:
            logger.exception(f"[DATABASE] Errore recupero dettagli reclamo: {e}")
            return []

    def get_claims_by_date_range(self, start_date: str, end_date: str) -> List[Dict]:
        """
        Recupera i reclami in un intervallo di date

        Args:
            start_date: Data inizio (formato YYYY-MM-DD)
            end_date: Data fine (formato YYYY-MM-DD)

        Returns:
            List[dict]: Lista dei reclami
        """
        try:
            query = """
                    SELECT
                        [ClaimLogId], [ClaimTypeId], [InternalClaimNumber], [CustomerClaimNumber], [ShortClaimDescription], [DateClaim], [TargetDate], [Quantity], [USERName]
                    FROM [Traceability_RS].[clm].[ClaimLogs]
                    WHERE [DateClaim] BETWEEN ? AND ?
                    ORDER BY [DateClaim] DESC \
                    """

            results = self.fetch_all(query, (start_date, end_date))

            claims = []
            for row in results:
                claims.append({
                    'ClaimLogId': row[0],
                    'ClaimTypeId': row[1],
                    'InternalClaimNumber': row[2],
                    'CustomerClaimNumber': row[3],
                    'ShortClaimDescription': row[4],
                    'DateClaim': row[5],
                    'TargetDate': row[6],
                    'Quantity': row[7],
                    'USERName': row[8]
                })

            return claims

        except Exception as e:
            logger.exception(f"[DATABASE] Errore recupero reclami per data: {e}")
            return []

    def get_claims_by_client(self, client_id: int) -> List[Dict]:
        """
        Recupera i reclami di un cliente

        Args:
            client_id: ID del cliente (IDFinalClient)

        Returns:
            List[dict]: Lista dei reclami del cliente
        """
        try:
            query = """
                    SELECT
                        [ClaimLogId], [InternalClaimNumber], [CustomerClaimNumber], [ShortClaimDescription], [DateClaim], [Quantity], [USERName]
                    FROM [Traceability_RS].[clm].[ClaimLogs]
                    WHERE [IDFinalClient] = ?
                    ORDER BY [DateClaim] DESC \
                    """

            results = self.fetch_all(query, (client_id,))

            claims = []
            for row in results:
                claims.append({
                    'ClaimLogId': row[0],
                    'InternalClaimNumber': row[1],
                    'CustomerClaimNumber': row[2],
                    'ShortClaimDescription': row[3],
                    'DateClaim': row[4],
                    'Quantity': row[5],
                    'USERName': row[6]
                })

            return claims

        except Exception as e:
            logger.exception(f"[DATABASE] Errore recupero reclami per cliente: {e}")
            return []

    def search_claims(self, search_term: str) -> List[Dict]:
        """
        Cerca reclami per numero interno o numero cliente

        Args:
            search_term: Termine di ricerca

        Returns:
            List[dict]: Lista dei reclami trovati
        """
        try:
            query = """
                    SELECT TOP 100 [ClaimLogId],
                    [InternalClaimNumber],
                    [CustomerClaimNumber],
                    [ShortClaimDescription],
                    [DateClaim],
                    [Quantity],
                    [USERName]
                    FROM [Traceability_RS].[clm].[ClaimLogs]
                    WHERE [InternalClaimNumber] LIKE ?
                       OR [CustomerClaimNumber] LIKE ?
                       OR [ShortClaimDescription] LIKE ?
                    ORDER BY [DateClaim] DESC \
                    """

            search_pattern = f"%{search_term}%"
            results = self.fetch_all(query, (search_pattern, search_pattern, search_pattern))

            claims = []
            for row in results:
                claims.append({
                    'ClaimLogId': row[0],
                    'InternalClaimNumber': row[1],
                    'CustomerClaimNumber': row[2],
                    'ShortClaimDescription': row[3],
                    'DateClaim': row[4],
                    'Quantity': row[5],
                    'USERName': row[6]
                })

            return claims

        except Exception as e:
            logger.exception(f"[DATABASE] Errore ricerca reclami: {e}")
            return []

    # === NUOVI METODI PER GESTIONE DOCUMENTI RECLAMI ===
    
    def fetch_claim_doc_types(self):
        """
        Recupera i tipi di documento disponibili per i reclami
        
        Returns:
            List[tuple]: Lista di tuple (ClaimDocTypeId, ClaimDocType)
        """
        try:
            query = """
                    SELECT [ClaimDocTypeId], [ClaimDocType]
                    FROM [Traceability_RS].[dbo].[ClaimDocTypes]
                    ORDER BY [ClaimDocType] \
                    """
            
            results = self.fetch_all(query)
            logger.debug(f"[DATABASE] Caricati {len(results)} tipi di documento reclami")
            return results
            
        except Exception as e:
            logger.exception(f"[DATABASE] Errore caricamento tipi documento reclami: {e}")
            return []
    
    def save_claim_document(self, claim_log_id: int, doc_type_id: int, doc_data: bytes, doc_name: str) -> bool:
        """
        Salva un documento nella tabella ClaimLogDocs
        
        Args:
            claim_log_id: ID del reclamo
            doc_type_id: ID del tipo di documento
            doc_data: Dati binari del documento
            doc_name: Nome del file (non salvato - solo per logging)
            
        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            # NOTE: DocumentName column doesn't exist in ClaimLogDocs table
            # Only ClaimLogId, ClaimDocTypeId, and ClaimDoc are stored
            query = """
                    INSERT INTO [Traceability_RS].[dbo].[ClaimLogDocs]
                    ([ClaimLogId], [ClaimDocTypeId], [ClaimDoc])
                    VALUES (?, ?, ?)
                    """
            
            params = (claim_log_id, doc_type_id, doc_data)
            
            success = self.execute_query(query, params)
            
            if success:
                logger.info(f"[DATABASE] Documento '{doc_name}' salvato per ClaimLogId={claim_log_id}, Type={doc_type_id}")
            else:
                logger.error(f"[DATABASE] Errore salvataggio documento '{doc_name}' per ClaimLogId={claim_log_id}")
                
            return success
            
        except Exception as e:
            logger.exception(f"[DATABASE] Errore salvataggio documento reclamo: {e}")
            return False

    
    def get_claim_documents_count(self, claim_log_id: int) -> int:
        """
        Recupera il numero di documenti caricati per un reclamo
        
        Args:
            claim_log_id: ID del reclamo
            
        Returns:
            int: Numero di documenti caricati
        """
        try:
            query = """
                    SELECT COUNT(*) as DocCount
                    FROM [Traceability_RS].[dbo].[ClaimLogDocs]
                    WHERE [ClaimLogId] = ? \
                    """
            
            result = self.fetch_one(query, (claim_log_id,))
            
            if result:
                count = int(result[0])
                logger.debug(f"[DATABASE] ClaimLogId={claim_log_id} ha {count} documenti")
                return count
            else:
                return 0
                
        except Exception as e:
            logger.exception(f"[DATABASE] Errore conteggio documenti reclamo: {e}")
            return 0
    
    def send_claim_notification_email(self, claim_log_id: int, claim_header) -> bool:
        """
        Invia email di notifica per un reclamo completato
        
        Args:
            claim_log_id: ID del reclamo
            claim_header: Oggetto ClaimHeader con i dati del reclamo
            
        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            # Recupera gli indirizzi email dalla configurazione
            query_setting = """
                           SELECT [Value]
                           FROM [Traceability_RS].[dbo].[settings]
                           WHERE [atribute] = 'sys_claim_notice' \
                           """
            
            result = self.fetch_one(query_setting)
            
            if not result or not result[0]:
                logger.warning("[DATABASE] Impostazione 'sys_claim_notice' non trovata o vuota")
                return False
            
            email_addresses = result[0]
            logger.debug(f"[DATABASE] Email destinatari notifica reclamo: {email_addresses}")
            
            # Prepara il contenuto dell'email
            subject = f"Nuovo Reclamo: {claim_header.InternalClaimNumber}"
            
            body = f"""
Nuovo reclamo inserito nel sistema:

Numero Interno: {claim_header.InternalClaimNumber}
Numero Cliente: {claim_header.CustomerClaimNumber}
Descrizione: {claim_header.ShortClaimDescription}
Data Reclamo: {claim_header.DateClaim}
Data Target: {claim_header.TargetDate}
QuantitÃ : {claim_header.Quantity}
Inserito da: {claim_header.USERName}

Accedi al sistema per visualizzare i dettagli completi.
"""
            
            # Invia l'email usando il metodo esistente (se disponibile)
            # Nota: Questo richiede che esista un metodo send_email nel database o in utils
            try:
                # Prova a importare e usare il metodo di invio email
                import utils
                if hasattr(utils, 'send_email'):
                    success = utils.send_email(
                        recipients=email_addresses,
                        subject=subject,
                        body=body
                    )
                    if success:
                        logger.info(f"[DATABASE] Email notifica inviata per ClaimLogId={claim_log_id}")
                    else:
                        logger.error(f"[DATABASE] Errore invio email per ClaimLogId={claim_log_id}")
                    return success
                else:
                    logger.warning("[DATABASE] Metodo send_email non disponibile in utils")
                    # TODO: Implementare invio email alternativo
                    return False
                    
            except ImportError:
                logger.warning("[DATABASE] Modulo utils non disponibile per invio email")
                return False
                
        except Exception as e:
            logger.exception(f"[DATABASE] Errore invio notifica email reclamo: {e}")
            return False

    def update_claim_header(self, claim_log_id: int, updates: dict) -> bool:
        """
        Aggiorna una testata di reclamo

        Args:
            claim_log_id: ID della testata
            updates: Dizionario con i campi da aggiornare

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            if not updates:
                return False

            # Costruisci la query dinamicamente
            set_clause = ", ".join([f"[{k}] = %s" for k in updates.keys()])
            values = list(updates.values())
            values.append(claim_log_id)

            query = f"""
                UPDATE [Traceability_RS].[clm].[ClaimLogs]
                SET {set_clause}
                WHERE [ClaimLogId] = %s
            """

            cursor = self.conn.cursor()
            cursor.execute(query, values)
            self.conn.commit()
            cursor.close()

            logger.info(f"[DATABASE] Testata reclamo aggiornata: ClaimLogId={claim_log_id}")
            return True

        except Exception as e:
            logger.exception(f"[DATABASE] Errore aggiornamento testata: {e}")
            self.conn.rollback()
            return False

    def delete_claim(self, claim_log_id: int) -> bool:
        """
        Elimina un reclamo (testata e dettagli)

        Args:
            claim_log_id: ID della testata

        Returns:
            bool: True se successo, False altrimenti
        """
        try:
            cursor = self.conn.cursor()

            # Elimina prima i dettagli
            query_details = """
                            DELETE \
                            FROM [Traceability_RS].[clm].[ClaimDataLogs]
                            WHERE [ClaimLogId] = %s \
                            """
            cursor.execute(query_details, (claim_log_id,))

            # Poi elimina la testata
            query_header = """
                           DELETE \
                           FROM [Traceability_RS].[clm].[ClaimLogs]
                           WHERE [ClaimLogId] = %s \
                           """
            cursor.execute(query_header, (claim_log_id,))

            self.conn.commit()
            cursor.close()

            logger.info(f"[DATABASE] Reclamo eliminato: ClaimLogId={claim_log_id}")
            return True

        except Exception as e:
            logger.exception(f"[DATABASE] Errore eliminazione reclamo: {e}")
            self.conn.rollback()
            return False

    ###fine metodi db claims

    def add_new_site(self, name, address, vat, country, logo):
        """Aggiunge una nuova compagnia."""
        query = "INSERT INTO dbo.Sites (SiteName, SiteAddress, SiteVat, SiteCountry, Logo, IDLastPhase, IsSupplier, IsTempraryLeasingComp) VALUES (?, ?, ?, ?, ?, 139, 1, 0);"
        try:
            self.cursor.execute(query, name, address, vat, country, logo)
            self.conn.commit()
            return True, "Compagnia aggiunta con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            if 'UNIQUE' in str(e) or 'duplicate' in str(e):
                return False, "Errore: Esiste giÃ  una compagnia con questo nome o Partita IVA."
            return False, f"Errore database: {e}"

    def update_site(self, site_id, name, address, vat, country, logo):
        """Aggiorna una compagnia esistente."""
        query = "UPDATE dbo.Sites SET SiteName=?, SiteAddress=?, SiteVat=?, SiteCountry=?, Logo=? WHERE IDSite=?;"
        try:
            self.cursor.execute(query, name, address, vat, country, logo, site_id)
            self.conn.commit()
            return True, "Compagnia aggiornata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_site(self, site_id):
        """Cancella una compagnia (da usare con cautela)."""
        # Aggiungere un controllo per verificare se il sito Ã¨ usato in altre tabelle prima di cancellare
        query = "DELETE FROM dbo.Sites WHERE IDSite = ?;"
        try:
            self.cursor.execute(query, site_id)
            self.conn.commit()
            return True, "Compagnia cancellata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def check_if_brand_is_used(self, brand_id):
        """Controlla se un brand Ã¨ utilizzato in almeno una macchina."""
        query = "SELECT COUNT(*) FROM eqp.Equipments WHERE BrandId = ?;"
        try:
            count = self.cursor.execute(query, brand_id).fetchval()
            return count > 0
        except pyodbc.Error:
            return True  # Per sicurezza, in caso di errore, si assume che sia usato

    def update_brand(self, brand_id, brand_name, company_id, logo_data):
        """Aggiorna un brand esistente."""
        query = "UPDATE eqp.EquipmentBrands SET Brand = ?, CompanyId = ?, BrandLogo = ? WHERE EquipmentBrandId = ?;"
        try:
            company_id_to_save = int(company_id) if company_id is not None else None
            self.cursor.execute(query, brand_name, company_id_to_save, logo_data, brand_id)
            self.conn.commit()
            return True, "Brand aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_brand(self, brand_id):
        """Cancella un brand."""
        query = "DELETE FROM eqp.EquipmentBrands WHERE EquipmentBrandId = ?;"
        try:
            self.cursor.execute(query, brand_id)
            self.conn.commit()
            return True, "Brand cancellato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def execute_missing_action_report(self):
        """Esegue la stored procedure per il report delle manutenzioni mancanti."""
        try:
            # La sintassi per chiamare una stored procedure che restituisce risultati
            sql = "{CALL eqp.sp_CheckMaintenanceStatus}"
            self.cursor.execute(sql)
            results = self.cursor.fetchall()

            # Recupera anche le intestazioni delle colonne per l'Excel
            headers = [column[0] for column in self.cursor.description]

            return results, headers, None
        except pyodbc.Error as e:
            error_message = f"Errore durante l'esecuzione della stored procedure: {e}"
            self.last_error_details = str(e)
            return None, None, error_message

    def fetch_shipping_plan_items(self, shipping_date):
        """Recupera le righe del piano di spedizione per una data specifica."""
        # --- CORREZIONE QUI: Aggiunta la clausola WHERE ---
        query = "SELECT * FROM dbo.ShippingPlanItems WHERE ShippingDate = ?;"
        try:
            self.cursor.execute(query, shipping_date)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def update_xls_sheet_name(self, file_type_id, sheet_name):
        """Aggiorna il nome del foglio di lavoro per un dato tipo di file."""
        query = "UPDATE dbo.XlsFileTypes SET SheetName = ? WHERE FileTypeId = ?;"
        try:
            self.cursor.execute(query, sheet_name, file_type_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def fetch_xls_file_types(self):
        """Recupera i tipi di file configurabili per l'import."""
        query = "SELECT FileTypeId, FileTypeName, TranslationKey FROM dbo.XlsFileTypes ORDER BY FileTypeName;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def fetch_xls_mappings(self, file_type_id):
        """
        Recupera la mappatura delle colonne E il nome del foglio di lavoro
        per un dato tipo di file.
        """
        query = """
            SELECT 
                m.MappingId, m.FieldName, m.ColumnLetter, m.StartRow, t.SheetName
            FROM 
                dbo.XlsColumnMappings m
            INNER JOIN 
                dbo.XlsFileTypes t ON m.FileTypeId = t.FileTypeId
            WHERE 
                m.FileTypeId = ?;
        """
        try:
            results = self.cursor.execute(query, file_type_id).fetchall()
            if not results:
                return {}, None

            # Estrae il nome del foglio (sarÃ  lo stesso per tutte le righe)
            sheet_name = results[0].SheetName

            # Costruisce il dizionario della mappatura
            mappings = {}
            for row in results:
                mappings[row.FieldName] = {'id': row.MappingId, 'col': row.ColumnLetter, 'row': row.StartRow}

            return mappings, sheet_name
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return {}, None

    def update_xls_mapping(self, mapping_id, column_letter, start_row):
        """Aggiorna una singola riga di mappatura."""
        query = "UPDATE dbo.XlsColumnMappings SET ColumnLetter = ?, StartRow = ? WHERE MappingId = ?;"
        try:
            self.cursor.execute(query, column_letter, start_row, mapping_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def save_shipping_plan_items(self, plan_data_list):
        """Esegue un'operazione di UPSERT (Update/Insert) per i dati del piano."""
        if not plan_data_list: return True, "Nessun dato da salvare."
        try:
            # Prepara le query
            update_sql = """
                UPDATE dbo.ShippingPlanItems 
                SET OriginalQty = ?, ModifiedQty = ?, Note = ?, Status = ?, LastUpdated = GETDATE(), UpdatedBy = ?
                WHERE ItemId = ?;
            """
            insert_sql = """
                INSERT INTO dbo.ShippingPlanItems (ShippingDate, OrderNumber, ProductCode, OriginalQty, ModifiedQty, Note, Status, UpdatedBy)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?);
            """

            for item in plan_data_list:
                if item.get('id'):  # Se ha un ID, Ã¨ un UPDATE
                    params = (item['original_qty'], item['modified_qty'], item['note'], item['status'], item['user'],
                              item['id'])
                    self.cursor.execute(update_sql, params)
                else:  # Altrimenti Ã¨ un INSERT
                    params = (item['shipping_date'], item['order'], item['product'], item['original_qty'],
                              item['modified_qty'], item['note'], item['status'], item['user'])
                    self.cursor.execute(insert_sql, params)

            self.conn.commit()
            return True, "Piano di spedizione salvato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore durante il salvataggio: {e}"

    def fetch_shipping_settings(self):
        """Recupera tutte le impostazioni di spedizione attive."""
        query = "SELECT [ShippingSettingId], [DayOfWeek], [ShippingType] FROM [Traceability_RS].[dbo].[ShippingSettings] WHERE dateend IS NULL ORDER BY dayofweek;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def save_backlog_report(self, user_name, general_note, excel_data, notes_list):
        """Salva il report di backlog, il file excel e le note in una transazione."""
        try:
            # 1. Inserisce il log principale e ottiene il nuovo ID
            insert_log_sql = "INSERT INTO dbo.BackLogs (UserLog, Note) OUTPUT INSERTED.BackLogId VALUES (?, ?);"
            new_log_id = self.cursor.execute(insert_log_sql, user_name, general_note).fetchval()
            if not new_log_id: raise Exception("Creazione BackLog fallita.")

            # 2. Inserisce i dati binari del file Excel
            insert_data_sql = "INSERT INTO dbo.BackLogData (BackLogId, ExcelDataFile) VALUES (?, ?);"
            self.cursor.execute(insert_data_sql, new_log_id, excel_data)

            # 3. Inserisce tutte le note specifiche per riga
            if notes_list:
                insert_notes_sql = "INSERT INTO dbo.BackLogNotes (BackLogId, ExcelRowNumber, Note) VALUES (?, ?, ?);"
                params = [(new_log_id, row_num, note) for row_num, note in notes_list]
                self.cursor.executemany(insert_notes_sql, params)

            self.conn.commit()
            return True, "Report di backlog salvato con successo nel database."
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore durante il salvataggio nel database: {e}"

    def add_shipping_setting(self, day_of_week, shipping_type):
        """Aggiunge una nuova impostazione di spedizione dopo aver controllato i duplicati."""
        # 1. Controlla se esiste giÃ  un'impostazione identica e attiva
        check_query = """
            SELECT COUNT(*) FROM [dbo].[ShippingSettings]
            WHERE DayOfWeek = ? AND ShippingType = ? AND DateEnd IS NULL;
        """
        insert_query = "INSERT INTO [dbo].[ShippingSettings] (DayOfWeek, ShippingType) VALUES (?, ?);"
        try:
            count = self.cursor.execute(check_query, day_of_week, shipping_type).fetchval()
            if count > 0:
                return False, "Errore: Esiste giÃ  un''impostazione identica per questo giorno e tipo di spedizione."

            # 2. Se non esiste, procedi con l'inserimento
            self.cursor.execute(insert_query, day_of_week, shipping_type)
            self.conn.commit()
            return True, "Impostazione aggiunta con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_shipping_setting(self, setting_id):
        """Esegue un soft delete di un'impostazione di spedizione impostando DateEnd."""
        query = "UPDATE [dbo].[ShippingSettings] SET DateEnd = GETDATE() WHERE ShippingSettingId = ?;"
        try:
            self.cursor.execute(query, setting_id)
            self.conn.commit()
            return True, "Impostazione cancellata con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_all_active_employees_birthdays(self):
        """Recupera nome, cognome e data di nascita di tutti i dipendenti attivi."""
        query = """
                SELECT E.EmployeeName, E.EmployeeSurname, E.EmployeeBirthDate
                FROM employee.dbo.employees E
                         INNER JOIN employee.dbo.EmployeeHireHistory H ON E.EmployeeId = H.EmployeeId
                WHERE H.EMPLOYEERID = 2 \
                  AND H.EndWorkDate IS NULL \
                  AND E.EmployeeBirthDate IS NOT NULL; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_equipments_for_timing(self):
        """Recupera i macchinari con il loro brand."""
        query = """
                SELECT e.EquipmentId, e.InternalName + ' [' + b.Brand + ']' AS EquipmentName
                FROM eqp.Equipments e
                         INNER JOIN eqp.EquipmentBrands b ON e.BrandId = b.EquipmentBrandId
                ORDER BY e.InternalName + ' [' + b.Brand + ']'; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def fetch_tasks_for_timing(self, equipment_id, intervention_id):
        """Recupera i compiti di manutenzione con i relativi tempi."""
        # NOTA: Ho aggiunto cmt.CompitoManutenzioneTimingId alla SELECT perchÃ© Ã¨ necessario per l'UPDATE.
        query = """
                SELECT am.CompitoId, \
                       am.NomeCompito, \
                       am.DescrizioneCompito, \
                       am.PdfRiferiment, \
                       cmt.TimingMinutes, \
                       cmt.CompitoManutenzioneTimingId
                FROM [Traceability_RS].[eqp].[CompitiManutenzione] am
                    INNER JOIN eqp.ProgrammedInterventions pr \
                ON pr.ProgrammedInterventionId = am.ProgrammedInterventionId
                    INNER JOIN eqp.Equipments e ON e.EquipmentId = am.EquipmentId
                    LEFT JOIN eqp.CompitiManutenzioneTiming cmt ON cmt.compitoid = am.CompitoId AND cmt.dateend IS NULL
                WHERE e.EquipmentId = ? AND pr.ProgrammedInterventionId = ?
                ORDER BY pr.Ordineprn, am.CompitoId; \
                """
        try:
            self.cursor.execute(query, equipment_id, intervention_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def update_task_pdf_reference(self, task_id, pdf_reference):
        """Aggiorna solo il campo PdfRiferiment per un dato compito."""
        query = "UPDATE [Traceability_RS].[eqp].[CompitiManutenzione] SET PdfRiferiment = ? WHERE compitoid = ?;"
        try:
            self.cursor.execute(query, pdf_reference, task_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def update_task_timing(self, task_id, new_minutes, old_timing_id):
        """Aggiorna il tempo di manutenzione (chiude il vecchio record e inserisce il nuovo)."""
        try:
            # 1. Se esiste un vecchio record di timing, lo chiude impostando la DateEnd
            if old_timing_id:
                update_query = "UPDATE [Traceability_RS].[eqp].CompitiManutenzioneTiming SET Dateend = GETDATE() WHERE CompitoManutenzioneTimingId = ?;"
                self.cursor.execute(update_query, old_timing_id)

            # 2. Se Ã¨ stato fornito un nuovo valore in minuti, lo inserisce
            if new_minutes is not None and new_minutes != '':
                insert_query = "INSERT INTO EQP.CompitiManutenzioneTiming (compitoid, TimingMinutes) VALUES (?, ?);"
                self.cursor.execute(insert_query, task_id, new_minutes)

            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def fetch_working_areas(self):
        """Recupera le Aree di Lavoro principali."""
        query = "SELECT [WorkingAreaID], [AreaName] FROM [ResetServices].[BreakDown].[WorkingAreas] ORDER BY AreaName;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_working_areas: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_working_sub_areas(self, working_area_id):
        """Recupera le Sotto-Aree in base all'Area di Lavoro selezionata."""
        query = """
                SELECT [WorkingSubAreaID], [AreaSubName]
                FROM [ResetServices].[BreakDown].[WorkingSubAreas]
                WHERE [WorkingAreaID] = ? \
                ORDER BY AreaSubName; \
                """
        try:
            self.cursor.execute(query, working_area_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_working_sub_areas: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_working_lines(self, working_area_id, sub_area_id):
        """Recupera le Linee in base all'Area e Sotto-Area selezionate."""
        query = """
                SELECT DISTINCT wl.WorkingLineID, WL.WorkingLineName
                FROM [ResetServices].[BreakDown].[WorkingAreas] AS WA
                    INNER JOIN [ResetServices].[BreakDown].WorkingSubAreas WSA \
                ON WA.WorkingAreaID = WSA.WorkingAreaID
                    INNER JOIN [ResetServices].[BreakDown].WorkingLines AS WL ON WSA.WorkingSubAreaID = WL.WorkingSubAreaID
                WHERE WA.WorkingAreaID = ? AND WSA.workingsubareaid = ?
                ORDER BY wl.WorkingLineName; \
                """
        try:
            self.cursor.execute(query, working_area_id, sub_area_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_working_lines: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_production_orders_for_breakdown(self):
        """Recupera la lista degli ordini di produzione aperti."""
        query = """
        SELECT DISTINCT 
            o.idorder as IdOrdine, 
            o.OrderNumber + ' [' + pf.ProductCode +']' as OrderNumber 
        FROM Traceability_RS.dbo.orders as o 
        INNER JOIN traceability_rs.dbo.products as pf ON pf.IDProduct = O.IDProduct
        LEFT JOIN ResetServices.DBO.TBORDINI RO ON ro.IdPOTrace = o.IDOrder
        LEFT JOIN resetservices.dbo.tbregistro r ON ro.idregistro = r.contatore 
            AND r.idregistro IN (21,26)
        WHERE CAST(O.DataInserted as date) >= '2025-08-01' 
            AND (ro.IdOrdine IS NULL 
                OR NOT EXISTS (
                    SELECT 1
                    FROM resetservices.dbo.TbSubOrdine s 
                    LEFT JOIN resetservices.dbo.TbFattStory fs ON fs.IdPoSub = s.IdOrdStori
                    WHERE s.idordine = ro.IdOrdine
                    GROUP BY s.IdOrdStori, s.QtaStory
                    HAVING s.QtaStory > ISNULL(SUM(fs.QtaFaturata), 0)
                )
            and (ro.idordine is NULL 
                OR NOT EXISTS (
                    SELECT 1 
                    FROM resetservices.dbo.TbSubOrdine inner join resetservices.dbo.tbprodfin on tbsubordine.idpf = 
                    tbprodfin.idpf inner join resetservices.dbo.TbProdFinStuff on TbProdFinStuff.Idpf =tbprodfin.idpf
                    )
            ))
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            print(f"ERRORE SQL in fetch_production_orders_for_breakdown: {e}")
            return []

    def fetch_report_data(self, from_date, to_date, equipment_id=None, intervention_id=None):
        """Recupera i dati dei log di manutenzione per un dato periodo e filtri opzionali."""
        query = """
                SELECT eq.InternalName                             AS EquipmentName, \
                       pin.TimingDescriprion                       AS InterventionType, \
                       cm.NomeCompito                              AS TaskName, \
                       lm.UserName, \
                       lm.DateStart, \
                       lm.DateStop, \
                       DATEDIFF(MINUTE, lm.DateStart, lm.DateStop) AS DurationInMinutes
                FROM eqp.LogManutenzioni lm
                         JOIN eqp.CompitiManutenzione cm ON lm.CompitoId = cm.CompitoId
                         JOIN eqp.Equipments eq ON lm.EquipmentId = eq.EquipmentId
                         JOIN eqp.ProgrammedInterventions pin \
                              ON cm.ProgrammedInterventionId = pin.ProgrammedInterventionId
                WHERE lm.DateStart >= ? \
                  AND lm.DateStart < DATEADD(day, 1, ?) \
                """
        params = [from_date, to_date]

        if equipment_id:
            query += " AND eq.EquipmentId = ?"
            params.append(equipment_id)

        if intervention_id:
            query += " AND pin.ProgrammedInterventionId = ?"
            params.append(intervention_id)

        query += " ORDER BY lm.DateStart;"

        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_open_production_orders(self):
        """Recupera la lista degli ordini di produzione aperti."""
        query = """
                SELECT o.idordine, o.po + ' [' + pf.epiccode + ']' AS OrderNumber
                FROM ResetServices.dbo.tbordini o
                         INNER JOIN Resetservices.dbo.tbsubordine so ON o.IdOrdine = so.IdOrdine
                         INNER JOIN Resetservices.dbo.tbprodfin pf ON so.idpf = pf.idpf
                         INNER JOIN resetservices.dbo.TbRegistro r \
                                    ON o.idregistro = r.contatore AND r.idregistro IN (21, 26)
                         LEFT JOIN resetservices.dbo.TbFattStory fs ON fs.IdPoSub = so.IdOrdStori
                         LEFT JOIN resetservices.dbo.TbProdFinStuff Micro ON micro.Idpf = so.idpf
                WHERE YEAR (o.dataord) >= 2025 AND micro.idpf IS NULL
                GROUP BY o.idordine, o.po + ' [' + pf.epiccode +']', so.QtaStory, o.dataord, so.DataDeliSubOrdine
                HAVING so.QtaStory > ISNULL(SUM (fs.QtaFaturata), 0)
                ORDER BY o.dataord; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            print(f"Errore nel recupero degli ordini di produzione: {e}")
            return []

    def fetch_issue_problems_by_subarea(self, sub_area_id):
        """Recupera i problemi/eventi specifici per una Sotto-Area."""
        query = """
                SELECT ip.IssueProblemId, ip.DescriptionEN
                FROM ResetServices.BreakDown.IssueProblemsPerLines AS BDL
                         INNER JOIN ResetServices.BreakDown.IssueProblems AS ip ON ip.IssueProblemId = bdl.IusseProblemId
                WHERE BDL.WorkingSubAreaId = ?
                GROUP BY ip.IssueProblemId, ip.DescriptionEN; \
                """
        try:
            print(
                f"DEBUG: La query per 'Evento/Problema' viene eseguita con il parametro WorkingSubAreaId = {sub_area_id}")
            self.cursor.execute(query, sub_area_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_issue_problems_by_subarea: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_issue_areas(self):
        """Recupera le aree problematiche (es. Meccanica, Elettrica)."""
        query = "SELECT [IssueAreaId], [IssueArea] FROM [ResetServices].[BreakDown].[IssuesAreas] ORDER BY [IssueArea];"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_issue_areas: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_issue_problems(self, working_area_id, issue_area_id, sub_area_id):
        """
        Recupera i problemi/eventi seguendo ESATTAMENTE la logica VBA fornita.
        """
        # Query per contare i problemi specifici per la linea/sotto-area
        # Ho corretto i typo 'IusseProblemId' in 'IssueProblemId' e 'PerLines' in 'PerLine'
        count_query = """
                      SELECT COUNT(BDL.IssueProblemsPerLineId)
                      FROM ResetServices.BreakDown.IssueProblemsPerLines AS BDL
                      WHERE BDL.WorkingSubAreaId = ?; \
                      """
        try:
            count = self.cursor.execute(count_query, sub_area_id).fetchval()
            print(f"DEBUG: Conteggio problemi specifici per SubAreaID {sub_area_id}: {count}")

            if count == 0:
                # NESSUN problema specifico -> Esegui la query GENERICA
                print("DEBUG: Eseguo la query GENERICA.")
                final_query = """
                              SELECT IP.IssueProblemId, \
                                     '[' + IP.Code + '] ' + IP.DescriptionEN AS IssueDescription, \
                                     IP.MandatoryExplication
                              FROM ResetServices.BreakDown.IssueProblems AS IP
                                       INNER JOIN ResetServices.BreakDown.IssuesAreas AS IA ON IA.IssueAreaId = IP.IssueAreaId
                                       INNER JOIN ResetServices.BreakDown.WorkingAreas AS WA ON IP.WorkingAreaID = WA.WorkingAreaID
                              WHERE (WA.WorkingAreaID = ?) \
                                AND (IP.IssueAreaId = ?) \
                                AND (IP.DateOut IS NULL); \
                              """
                params = (working_area_id, issue_area_id)
            else:
                # TROVATI problemi specifici -> Esegui la query SPECIFICA
                print("DEBUG: Eseguo la query SPECIFICA.")
                final_query = """
                              SELECT IP.IssueProblemId, \
                                     '[' + IP.Code + '] ' + IP.DescriptionEN AS IssueDescription, \
                                     IP.MandatoryExplication
                              FROM ResetServices.BreakDown.IssueProblems AS IP
                                       INNER JOIN ResetServices.BreakDown.IssuesAreas AS IA ON IA.IssueAreaId = IP.IssueAreaId
                                       INNER JOIN ResetServices.BreakDown.WorkingAreas AS WA ON IP.WorkingAreaID = WA.WorkingAreaID
                                       INNER JOIN ResetServices.BreakDown.IssueProblemsPerLines AS BDL \
                                                  on ip.IssueProblemId = bdl.IssueProblemId
                              WHERE (WA.WorkingAreaID = ?) \
                                AND (IP.IssueAreaId = ?) \
                                AND (BDL.WorkingSubAreaId = ?) \
                                AND (IP.DateOut IS NULL); \
                              """
                params = (working_area_id, issue_area_id, sub_area_id)

            self.cursor.execute(final_query, params)
            return self.cursor.fetchall()

        except pyodbc.Error as e:
            print(f"ERRORE SQL in fetch_issue_problems: {e}")
            self.last_error_details = str(e)
            return []
    
    def add_production_interruption(self, params):
        """Salva un nuovo record di interruzione produzione in ReportIssueLogs."""
        query = """
            SET NOCOUNT ON;
            INSERT INTO ResetServices.[BreakDown].[ReportIssueLogs] ([DateReport], [HourReport], [UserName], [IssueAreaId], \
                                                        [WorkingAreaID],
                [WorkingLineID], [WorkingSubAreaID], [IssueProblemId], [FromHour], [ToHour],
                [Lost_OR_Gain], [Hours], [PoNumber], [ProductCode], [Note], [ActionPlan]) \
            VALUES (?, GETDATE(), ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            SELECT CAST(SCOPE_IDENTITY() AS INT) AS NewID;
            """
        try:
            self.cursor.execute(query,
                                params['report_date'], params['user_name'], params['issue_area_id'],
                                params['working_area_id'],
                                params['line_id'], params['sub_area_id'], params['problem_id'], params['from_hour'],
                                params['to_hour'],
                                params['lost_gain'], params['hours'], params['po_number'], params['product_code'],
                                params['note'], params['action_plan']
                                )
            
            # Recupera l'ID appena inserito
            row = self.cursor.fetchone()
            breakdown_id = row[0] if row else None
            
            self.conn.commit()
            return True, "Interruzione di produzione registrata con successo.", breakdown_id
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore durante il salvataggio: {e}", None
    
    def add_interruption_documents(self, breakdown_id, documents_list, user_name):
        """Salva i documenti allegati a un'interruzione di produzione."""
        query = """
            INSERT INTO ResetServices.[BreakDown].[ReportedIssueLogDocs] 
            ([BreakDownProblemLogId], [Document], [DocDescription], [AddBy])
            VALUES (?, ?, ?, ?);
            """
        try:
            for doc in documents_list:
                self.cursor.execute(query, 
                                   breakdown_id, 
                                   doc['binary_data'], 
                                   doc['description'], 
                                   user_name)
            
            self.conn.commit()
            return True, f"{len(documents_list)} documento/i salvato/i con successo."
        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore durante il salvataggio dei documenti: {e}"
    
    def check_if_material_exists(self, part_number):
        """Controlla se un materiale con un dato Codice Articolo esiste giÃ . Restituisce True se esiste, altrimenti False."""
        query = "SELECT COUNT(*) FROM eqp.SparePartMaterials WHERE MaterialPartNumber = ?;"
        try:
            count = self.cursor.execute(query, part_number).fetchval()
            return count > 0
        except pyodbc.Error as e:
            print(f"Errore nel controllo esistenza materiale: {e}")
            return False  # Per sicurezza, non blocchiamo l'utente in caso di errore


    def fetch_material_document(self, material_id):
        """Recupera i dati binari del documento per un singolo materiale."""
        query = "SELECT CatalogDetail FROM eqp.SparePartMaterials WHERE SparePartMaterialId = ?;"
        try:
            self.cursor.execute(query, material_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def search_materials(self, code_filter, desc_filter):
        """Cerca i materiali in base a filtri per codice/nome e descrizione."""
        query = """
                SELECT SparePartMaterialId, MaterialPartNumber, MaterialCode, MaterialDescription
                FROM eqp.SparePartMaterials
                WHERE (MaterialPartNumber LIKE ? OR MaterialCode LIKE ?)
                  AND MaterialDescription LIKE ?
                ORDER BY MaterialCode; \
                """
        try:
            # I parametri % per il LIKE vengono aggiunti qui
            code_param = f"%{code_filter}%"
            desc_param = f"%{desc_filter}%"

            self.cursor.execute(query, code_param, code_param, desc_param)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_single_material(self, material_id):
        """Recupera tutti i dati di un singolo materiale."""
        query = "SELECT * FROM eqp.SparePartMaterials WHERE SparePartMaterialId = ?;"
        try:
            self.cursor.execute(query, material_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return None

    def add_material(self, part_number, code, description, catalog_data, doc_name, user_name):
        """Aggiunge un nuovo materiale, includendo l'utente che ha eseguito l'operazione."""
        query = """
                INSERT INTO eqp.SparePartMaterials
                (MaterialPartNumber, MaterialCode, MaterialDescription, CatalogDetail, CatalogFileName, [User])
                    OUTPUT INSERTED.SparePartMaterialId
                VALUES (?, ?, ?, ?, ?, ?); \
                """
        try:
            new_id = self.cursor.execute(query, part_number, code, description, catalog_data, doc_name,
                                         user_name).fetchval()
            self.conn.commit()
            return True, new_id
        except pyodbc.IntegrityError as e:
            self.conn.rollback()
            if 'UQ_SparePartMaterials_MaterialPartNumber' in str(e):
                return False, "error_duplicate_material"
            else:
                self.last_error_details = str(e)
                return False, str(e)
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, str(e)

    def update_material(self, material_id, part_number, code, description, catalog_data, doc_name, user_name):
        """Aggiorna un materiale esistente, includendo l'utente che ha eseguito l'operazione."""
        query = """
                UPDATE eqp.SparePartMaterials
                SET MaterialPartNumber  = ?, \
                    MaterialCode        = ?, \
                    MaterialDescription = ?,
                    CatalogDetail       = ?, \
                    CatalogFileName     = ?, [User] = ?
                WHERE SparePartMaterialId = ?; \
                """
        try:
            self.cursor.execute(query, part_number, code, description, catalog_data, doc_name, user_name, material_id)
            self.conn.commit()
            return True, "Aggiornato."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, str(e)

    def delete_material(self, material_id):
        """Cancella un materiale. La cancellazione a cascata (ON DELETE CASCADE) rimuoverÃ  i link."""
        query = "DELETE FROM eqp.SparePartMaterials WHERE SparePartMaterialId = ?;"
        try:
            self.cursor.execute(query, material_id)
            self.conn.commit()
            return True, "Cancellato."
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, str(e)

    def fetch_linked_equipment(self, material_id):
        """Recupera gli ID delle macchine collegate a un materiale."""
        query = "SELECT EquipmentId FROM eqp.SparePartParents WHERE SparePartMaterialId = ?;"
        try:
            self.cursor.execute(query, material_id)
            return [row.EquipmentId for row in self.cursor.fetchall()]
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def update_material_links(self, material_id, equipment_ids, user_name):
        """Sincronizza i collegamenti tra un materiale e le macchine, includendo l'utente."""
        try:
            self.cursor.execute("DELETE FROM eqp.SparePartParents WHERE SparePartId = ?", material_id)

            if equipment_ids:
                insert_query = "INSERT INTO eqp.SparePartParents (SparePartId, EquipmentId, [User]) VALUES (?, ?, ?);"
                # Aggiunge l'utente a ogni tupla di parametri
                params = [(material_id, eq_id, user_name) for eq_id in equipment_ids]
                self.cursor.executemany(insert_query, params)

            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            print(f"Errore durante l'aggiornamento dei link materiali: {e}")
            return False

    def fetch_materials_for_equipment(self, equipment_id):
        """Recupera tutti i materiali collegati a una specifica macchina."""
        query = """
                SELECT m.MaterialPartNumber, m.MaterialCode, m.MaterialDescription
                FROM eqp.SparePartMaterials m
                         INNER JOIN eqp.SparePartParents l ON m.SparePartMaterialId = l.SparePartMaterialId
                WHERE l.EquipmentId = ?
                ORDER BY m.MaterialCode; \
                """
        try:
            self.cursor.execute(query, equipment_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e);
            return []

    def fetch_user_permissions(self, employee_hire_history_id):
        """Recupera i permessi attivi per un dato dipendente."""
        query = """
                SELECT a.AuthorizedUsedId, ap.TranslationValue + ' [' + ap.MenuValue + ']' AS MenuKey
                FROM Employee.dbo.employees e
                         INNER JOIN employee.dbo.EmployeeHireHistory h ON h.EmployeeId = e.EmployeeId
                         LEFT JOIN Traceability_rs.dbo.AutorizedUsers a \
                                   ON h.EmployeeHireHistoryId = a.EmployeeHireHistoryId
                         LEFT JOIN traceability_rs.dbo.AppTranslations ap ON ap.TranslationKey = a.TranslationKey
                WHERE h.EmployeeHireHistoryId = ? \
                  AND a.DateOut IS NULL
                  ORDER BY ap.TranslationValue + ' [' + ap.LanguageCode + ']'; \
                """
        try:
            self.cursor.execute(query, employee_hire_history_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_available_permissions(self, employee_hire_history_id):
        """Recupera i menu a cui un utente NON Ã¨ ancora abilitato."""
        query = """
        SELECT DISTINCT
               a.TranslationKey AS Translationkey,
               CAST(a.MenuValue AS nvarchar(255)) AS translationvalue
        FROM AppTranslations AS a
        WHERE a.MenuValue IS NOT NULL
          AND NOT EXISTS (
              SELECT 1
              FROM AutorizedUsers AS au
              WHERE au.TranslationKey = a.TranslationKey
                AND au.EmployeeHireHistoryId = ?
                AND au.DateOut IS NULL
          )
        ORDER BY translationvalue, Translationkey;
                """
        try:
            self.cursor.execute(query, employee_hire_history_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def grant_permission(self, employee_hire_history_id, translation_key):
        """Assegna un permesso a un utente."""
        query = "INSERT INTO [Traceability_RS].[dbo].AutorizedUsers (EmployeeHireHistoryId ,TranslationKey) VALUES (?, ?);"
        
        # Prima verifica se il PermissionKey esiste nella tabella apptranslations
        #check_permission_query = "SELECT [PermissionId] FROM [Traceability_RS].[dbo].[Permissions] WHERE [PermissionKey] = ?"
        check_permission_query ="SELECT ID FROM [Traceability_RS].[dbo].apptranslations where not menuvalue is null and languagecode ='it' and TranslationKey =?;"
        
       
        try:
            logger.info(f"HistoryHireId:{employee_hire_history_id} per argomento  {translation_key}")
            
            # Verifica se il PermissionKey esiste
            self.cursor.execute(check_permission_query, translation_key)
            permission_row = self.cursor.fetchone()
            
            if not permission_row or permission_row[0] is None:
                logger.error(f"PermissionKey '{translation_key}' non trovato nella tabella AppTranslations")
                return False
            
            permission_id = permission_row[0]
            logger.info(f"PermissionId trovato: {permission_id} per PermissionKey: {translation_key}")
            
            # Inserisci in AutorizedUsers
            self.cursor.execute(query, employee_hire_history_id, translation_key)
            
            # Inserisci in EmployeePermissions con il PermissionId verificato
            #self.cursor.execute(query2, employee_hire_history_id, employee_hire_history_id, permission_id)
            
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            logger.error(f"Errore durante l'assegnazione del permesso: {e} per {employee_hire_history_id} e {translation_key}")
            self.conn.rollback()
            return False

    def revoke_permission(self, authorized_user_id):
        """Revoca un permesso (soft delete)."""
        query = "UPDATE AutorizedUsers SET DateOut = GETDATE() WHERE AuthorizedUsedId = ?;"
        try:
            self.cursor.execute(query, authorized_user_id)
            self.conn.commit()
            return True
        except pyodbc.Error:
            self.conn.rollback()
            return False

    def check_if_doc_type_is_used(self, category_id):
        """Controlla se una categoria di documenti Ã¨ usata in almeno un documento."""
        query = "SELECT COUNT(DocumentoId) FROM dbo.DocumentiGenerali WHERE CategoriaId = ?;"
        try:
            count = self.cursor.execute(query, category_id).fetchval()
            return count > 0
        except pyodbc.Error as e:
            print(f"Errore nel controllo uso categoria: {e}")
            return True  # Per sicurezza, in caso di errore, si assume che sia usata

    def add_new_doc_type(self, name, key):
        """Aggiunge una nuova categoria di documenti."""
        query = "INSERT INTO dbo.DocCategorie (NomeCategoria, TranslationKey) VALUES (?, ?);"
        try:
            self.cursor.execute(query, name, key)
            self.conn.commit()
            return True, "Tipo documento aggiunto con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            # Gestisce l'errore di chiave duplicata
            if 'UNIQUE KEY' in str(e):
                return False, "Errore: Esiste giÃ  un tipo con questo nome o chiave di traduzione."
            return False, f"Errore database: {e}"

    def update_doc_type(self, category_id, name, key):
        """Aggiorna una categoria di documenti esistente."""
        query = "UPDATE dbo.DocCategorie SET NomeCategoria = ?, TranslationKey = ? WHERE CategoriaId = ?;"
        try:
            self.cursor.execute(query, name, key, category_id)
            self.conn.commit()
            return True, "Tipo documento aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            if 'UNIQUE KEY' in str(e):
                return False, "Errore: Esiste giÃ  un tipo con questo nome o chiave di traduzione."
            return False, f"Errore database: {e}"

    def delete_doc_type(self, category_id):
        """Cancella una categoria di documenti."""
        query = "DELETE FROM dbo.DocCategorie WHERE CategoriaId = ?;"
        try:
            self.cursor.execute(query, category_id)
            self.conn.commit()
            return True, "Tipo documento cancellato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_doc_categories(self):
        """Recupera tutte le categorie di documenti generali per il menu."""
        query = "SELECT CategoriaId, NomeCategoria, TranslationKey FROM dbo.DocCategorie ORDER BY CategoriaId;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore recupero categorie documenti: {e}")
            return []

    def fetch_general_documents(self, category_id):
        """Recupera i metadati dei documenti per una data categoria."""
        query = "SELECT DocumentoId, Titolo, Versione, DataCaricamento, CaricatoDa FROM dbo.DocumentiGenerali WHERE CategoriaId = ? ORDER BY Titolo;"
        try:
            self.cursor.execute(query, category_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return []

    def fetch_single_general_document(self, document_id):
        """Recupera tutti i dati di un singolo documento, inclusi i dati binari."""
        query = "SELECT * FROM dbo.DocumentiGenerali WHERE DocumentoId = ?;"
        try:
            self.cursor.execute(query, document_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            self.last_error_details = str(e)
            return None

    def add_general_document(self, category_id, title, desc, version, file_name, data, user):
        """Aggiunge un nuovo documento generale."""
        query = """
                INSERT INTO dbo.DocumentiGenerali (CategoriaId, Titolo, Descrizione, Versione, NomeFile, DatiFile, \
                                                   CaricatoDa)
                VALUES (?, ?, ?, ?, ?, ?, ?); \
                """
        try:
            self.cursor.execute(query, category_id, title, desc, version, file_name, data, user)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def update_general_document(self, doc_id, title, desc, version, file_name, data, user):
        """Aggiorna un documento generale esistente."""
        query = """
                UPDATE dbo.DocumentiGenerali
                SET Titolo          = ?, \
                    Descrizione     = ?, \
                    Versione        = ?, \
                    NomeFile        = ?, \
                    DatiFile        = ?, \
                    CaricatoDa      = ?, \
                    DataCaricamento = GETDATE()
                WHERE DocumentoId = ?; \
                """
        try:
            self.cursor.execute(query, title, desc, version, file_name, data, user, doc_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def delete_general_document(self, doc_id):
        """Cancella un documento generale."""
        query = "DELETE FROM dbo.DocumentiGenerali WHERE DocumentoId = ?;"
        try:
            self.cursor.execute(query, doc_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def authenticate_and_authorize(self, user_id, password, menu_translation_key):
        """
        Esegue l'autenticazione e controlla l'autorizzazione per una specifica funzione.
        Restituisce l'intera riga del DB se l'utente e la password sono corretti, altrimenti None.
        La riga conterrÃ  AuthorizedUsedId (che puÃ² essere NULL se non autorizzato).
        """
        query = """
                SELECT u.NomeUser, \
                       ISNULL(e.EmployeeName + ' ' + e.EmployeeSurname, '#ND') as EmployeeName, \
                       u.pass, \
                       a.AuthorizedUsedId
                FROM resetservices.dbo.tbuserkey as U
                         INNER JOIN employee.dbo.employees as e ON e.EmployeeId = u.idanga
                         INNER JOIN employee.dbo.EmployeeHireHistory as h ON e.EmployeeId = h.EmployeeId
                         LEFT JOIN dbo.AutorizedUsers as a \
                                   ON a.Employeehirehistoryid = h.EmployeeHireHistoryId AND a.TranslationKey = ?
                WHERE h.EndWorkDate IS NULL \
                  AND h.employeerid = 2 \
                  AND u.Nomeuser = ? \
                  AND u.Pass = ? \
                  AND a.DateOut IS NULL; \
                """
        try:
            # L'ordine dei parametri Ã¨ fondamentale: TranslationKey, Nomeuser, Pass
            logger.info(f"Chiave per accesso speciale:{menu_translation_key} per user:{user_id}")

            self.cursor.execute(query, menu_translation_key, user_id, password)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            print(f"Error during authentication/authorization: {e}")
            self.last_error_details = str(e)
            return None

    def fetch_maintenance_cycles(self):
        """Recupera tutti i cicli di manutenzione programmati (solo attivi)."""
        query = """SELECT ProgrammedInterventionId, TimingDescriprion, TimingValue, 
                          OrdinePrn, IsFixture, NoCycle, IsStensil, DateOut
                   FROM eqp.ProgrammedInterventions 
                   WHERE DateOut IS NULL
                   ORDER BY OrdinePrn, TimingDescriprion;"""
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei cicli di manutenzione: {e}")
            return []

    def check_if_cycle_is_used(self, intervention_id):
        """Controlla se un ciclo di manutenzione Ã¨ usato in almeno un log."""
        # Un ciclo Ã¨ usato se esiste un compito associato ad esso che Ã¨ stato registrato in LogManutenzioni
        query = """
                SELECT COUNT(lm.LogId)
                FROM eqp.LogManutenzioni lm
                         INNER JOIN eqp.CompitiManutenzione cm ON lm.CompitoId = cm.CompitoId
                WHERE cm.ProgrammedInterventionId = ?; \
                """
        try:
            count = self.cursor.execute(query, intervention_id).fetchval()
            return count > 0
        except pyodbc.Error as e:
            print(f"Errore nel controllo uso ciclo: {e}")
            return True  # Per sicurezza, in caso di errore, si assume che sia usato

    def add_new_maintenance_cycle(self, description, value, ordine_prn, is_fixture, no_cycle, is_stensil):
        """Aggiunge un nuovo ciclo di manutenzione."""
        query = """INSERT INTO eqp.ProgrammedInterventions 
                   (TimingDescriprion, TimingValue, OrdinePrn, IsFixture, NoCycle, IsStensil, DateOut) 
                   VALUES (?, ?, ?, ?, ?, ?, NULL);"""
        try:
            self.cursor.execute(query, description, value, ordine_prn, is_fixture, no_cycle, is_stensil)
            self.conn.commit()
            return True, "Ciclo aggiunto con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def update_maintenance_cycle(self, intervention_id, description, value, ordine_prn, is_fixture, no_cycle, is_stensil):
        """Aggiorna un ciclo di manutenzione esistente."""
        query = """UPDATE eqp.ProgrammedInterventions 
                   SET TimingDescriprion = ?, TimingValue = ?, OrdinePrn = ?, 
                       IsFixture = ?, NoCycle = ?, IsStensil = ?
                   WHERE ProgrammedInterventionId = ?;"""
        try:
            self.cursor.execute(query, description, value, ordine_prn, is_fixture, no_cycle, is_stensil, intervention_id)
            self.conn.commit()
            return True, "Ciclo aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def delete_maintenance_cycle(self, intervention_id):
        """Disattiva un ciclo di manutenzione (soft delete con DateOut)."""
        query = "UPDATE eqp.ProgrammedInterventions SET DateOut = GETDATE() WHERE ProgrammedInterventionId = ?;"
        try:
            self.cursor.execute(query, intervention_id)
            self.conn.commit()
            return True, "Ciclo disattivato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_authorized_employees(self):
        """Recupera la lista dei dipendenti autorizzati a fare segnalazioni."""
        query = """
                SELECT eh.EmployeeHireHistoryId, UPPER(e.EmployeeSurname + ' ' + e.EmployeeName) AS Employ
                FROM employee.dbo.employees e
                         INNER JOIN employee.dbo.EmployeeHireHistory eh \
                                    ON e.EmployeeId = eh.EmployeeId AND eh.EndWorkDate IS NULL
                         INNER JOIN employee.dbo.employeers er \
                                    ON eh.EmployeerId = er.EmployeerId AND er.EmployeerFiscalCode = 'RO35713341'
                ORDER BY UPPER(e.EmployeeSurname + ' ' + e.EmployeeName); \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore recupero dipendenti autorizzati: {e}")
            return []

    def fetch_submission_types(self, lang_code):
        """Recupera i tipi di segnalazione tradotti in base alla lingua selezionata."""
        # La tabella TipiSegnalazione deve essere quella corretta, es. dbo.TipiSegnalazione
        # Assicurati che i nomi delle tabelle e delle colonne siano corretti.
        query = """
                SELECT t.TipoSegnalazioneId, t.NomeTipo
                FROM employee.dbo.TipiSegnalazione AS t
                         INNER JOIN employee.dbo.Languages L ON t.LanguageID = l.LanguageID
                WHERE l.LanguageAcronim = ?
                ORDER BY t.NomeTipo; \
                """
        try:
            print(f"DEBUG: Sto cercando tipi di segnalazione con il codice lingua: '{lang_code}'")
            self.cursor.execute(query, lang_code)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore recupero tipi di segnalazione per lingua '{lang_code}': {e}")
            self.last_error_details = str(e)
            return []

    def add_new_submission(self, type_id, title, desc, location, employee_id, attachments):
        """Salva una nuova segnalazione e i suoi allegati in una transazione."""
        try:
            # 1. Inserisce la segnalazione principale e ottiene il suo ID
            insert_submission_sql = """
                                    INSERT INTO Employee.dbo.Segnalazioni (TipoSegnalazioneId, Titolo, Descrizione, Luogo, IdDipendente)
                                        OUTPUT INSERTED.SegnalazioneId
                                    VALUES (?, ?, ?, ?, ?); \
                                    """
            new_submission_id = self.cursor.execute(
                insert_submission_sql, type_id, title, desc, location, employee_id
            ).fetchval()

            if not new_submission_id:
                raise Exception("Creazione segnalazione fallita, ID non restituito.")

            # 2. Inserisce il record nella tabella delle verifiche con stato 5 - da analizzare
            insert_status_sql = """
                            INSERT INTO Employee.dbo.SegnalazioneStati
                                (SegnalazioneId, SegnalazioniTipoStatoId, Nota, OperatoDa)
                            VALUES (?, 5, 'Trigger after insert', 'System'); \
                             """
            self.cursor.execute(insert_status_sql, new_submission_id)

            # 3. Se ci sono allegati, li inserisce uno per uno
            if attachments:
                insert_attachment_sql = """
                                        INSERT INTO Employee.dbo.SegnalazioneAllegati (SegnalazioneId, NomeFile, DatiFile)
                                        VALUES (?, ?, ?); \
                                        """
                for attachment in attachments:
                    self.cursor.execute(insert_attachment_sql, new_submission_id, attachment['name'],
                                        attachment['data'])

            # 3. Se tutto Ã¨ andato bene, conferma la transazione
            self.conn.commit()
            return True, "Segnalazione registrata con successo."

        except Exception as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore durante il salvataggio: {e}"

    def fetch_brands_with_company_name(self):
        """Recupera tutti i brand con il nome del produttore associato."""
        query = """
                SELECT b.EquipmentBrandId, \
                       b.Brand, \
                       b.BrandLogo, \
                       s.acronimo AS CompanyName, \
                       s.idsoc    AS CompanyId
                FROM eqp.EquipmentBrands b \
                         INNER JOIN \
                     resetservices.dbo.tbsocieta s ON b.CompanyId = s.idsoc
                ORDER BY b.Brand; \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei brand: {e}")
            return []

    def add_new_brand(self, brand_name, company_id, logo_data):
        """Aggiunge un nuovo brand."""
        query = "INSERT INTO eqp.EquipmentBrands (Brand, CompanyId, BrandLogo) VALUES (?, ?, ?);"
        try:
            # Assicura che company_id sia None se non Ã¨ un numero valido
            company_id_to_save = int(company_id) if company_id is not None else None
            self.cursor.execute(query, brand_name, company_id_to_save, logo_data)
            self.conn.commit()
            return True, "Brand aggiunto con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            if 'UNIQUE' in str(e) or 'duplicate' in str(e):
                return False, "Errore: Esiste giÃ  un brand con questo nome."
            return False, f"Errore database: {e}"

    def update_brand(self, brand_id, brand_name, company_id, logo_data):
        """Aggiorna un brand esistente."""
        query = "UPDATE eqp.EquipmentBrands SET Brand = ?, CompanyId = ?, BrandLogo = ? WHERE EquipmentBrandId = ?;"
        try:
            company_id_to_save = int(company_id) if company_id is not None else None
            self.cursor.execute(query, brand_name, company_id_to_save, logo_data, brand_id)
            self.conn.commit()
            return True, "Brand aggiornato con successo."
        except pyodbc.Error as e:
            self.conn.rollback()
            return False, f"Errore database: {e}"

    def fetch_suppliers(self):
        """Recupera la lista dei fornitori."""
        query = "SELECT idsoc, acronimo, nazione FROM resetservices.dbo.tbsocieta WHERE acronimo IS NOT NULL ORDER BY acronimo;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero fornitori: {e}")
            return []

    def fetch_currencies(self):
        """Recupera la lista delle valute attive."""
        query = "SELECT IdValuta, [desc] FROM resetservices.dbo.TbValute WHERE loadexchange = 1 ORDER BY [desc];"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero valute: {e}")
            return []

    def add_new_supplier(self, denom_soc, nazione, cui, id_valuta):
        """Aggiunge un nuovo fornitore dopo aver controllato che la P.IVA non esista giÃ ."""
        # 1. Controlla se la Partita IVA (cui) esiste giÃ 
        check_query = "SELECT COUNT(*) FROM resetservices.dbo.tbsocieta WHERE cui = ?;"
        insert_query = """
                       INSERT INTO resetservices.dbo.tbsocieta (DenomSoc, Nazione, cui, IdValuta, Appruved)
                       VALUES (?, ?, ?, ?, 1); \
                       """
        try:
            count = self.cursor.execute(check_query, cui).fetchval()
            if count > 0:
                return False, "Errore: Partita IVA giÃ  presente nel database."

            # 2. Se non esiste, procedi con l'inserimento
            self.cursor.execute(insert_query, denom_soc, nazione, cui, id_valuta)
            self.conn.commit()
            return True, "Fornitore aggiunto con successo."

        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False, f"Errore database: {e}"

    def fetch_tasks_for_editing(self, intervention_id, equipment_id):
        """Recupera i compiti per un intervento e una macchina specifici."""
        query = """
                SELECT CompitoId, NomeCompito, Categoria, DescrizioneCompito, LinkedDocument
                FROM eqp.CompitiManutenzione
                WHERE ProgrammedInterventionId = ? \
                  AND EquipmentId = ?
                ORDER BY Ordine, CompitoId; \
                """
        try:
            self.cursor.execute(query, intervention_id, equipment_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei task per la modifica: {e}")
            self.last_error_details = str(e)
            return []

    def update_maintenance_task(self, task_id, equipment_id, category, task_name, description, document_data):
        """Aggiorna un compito di manutenzione esistente, incluso l'EquipmentId."""
        query = """
                UPDATE eqp.CompitiManutenzione
                SET EquipmentId        = ?, \
                    Categoria          = ?, \
                    NomeCompito        = ?, \
                    DescrizioneCompito = ?, \
                    LinkedDocument     = ?
                WHERE CompitoId = ?; \
                """
        try:
            self.cursor.execute(query, equipment_id, category, task_name, description, document_data, task_id)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def delete_maintenance_tasks(self, task_ids_to_delete):
        """Cancella una lista di compiti dal database."""
        if not task_ids_to_delete:
            return True  # Nessuna operazione da eseguire

        # Crea i segnaposto '?' per la clausola IN
        placeholders = ', '.join('?' for _ in task_ids_to_delete)
        query = f"DELETE FROM eqp.CompitiManutenzione WHERE CompitoId IN ({placeholders});"

        try:
            self.cursor.execute(query, task_ids_to_delete)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            return False

    def fetch_maintenance_interventions(self, equipment_type_filter=None):
        """Recupera i tipi di intervento di manutenzione per la selezione, con filtro opzionale per tipo."""
        query = """
                SELECT [ProgrammedInterventionId], [TimingDescriprion],
                       ISNULL(IsFixture, 0) AS IsFixture,
                       ISNULL(IsStensil, 0) AS IsStensil
                FROM [Traceability_RS].[eqp].[ProgrammedInterventions]
                WHERE DateOut IS NULL
        """
        
        # Aggiungi filtro per tipo di equipaggiamento
        if equipment_type_filter == 'fixture':
            query += " AND ISNULL(IsFixture, 0) = 1"
        elif equipment_type_filter == 'stensil':
            query += " AND ISNULL(IsStensil, 0) = 1"
        elif equipment_type_filter == 'all':
            # Per "all", mostra solo interventi che NON sono specifici per fixture o stencil
            query += " AND ISNULL(IsFixture, 0) = 0 AND ISNULL(IsStensil, 0) = 0"
        
        query += " ORDER BY TimingDescriprion;"
        
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei tipi di intervento: {e}")
            return []

    def insert_new_maintenance_task(self, intervention_id, equipment_id, category, task_name, description, order,
                                    document_data=None):
        """Inserisce un singolo nuovo compito, ora direttamente collegato a una macchina."""
        query = """
                INSERT INTO eqp.CompitiManutenzione
                (ProgrammedInterventionId, EquipmentId, Categoria, NomeCompito, DescrizioneCompito, Ordine, \
                 LinkedDocument)
                VALUES (?, ?, ?, ?, ?, ?, ?); \
                """
        try:
            self.cursor.execute(query, intervention_id, equipment_id, category, task_name, description, order,
                                document_data)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            print(f"Errore nell'inserimento del nuovo task: {e}")
            self.last_error_details = str(e)
            return False
    def fetch_report_maintainers(self):
        """Recupera la lista dei manutentori che hanno eseguito almeno una manutenzione."""
        query = """
                SELECT DISTINCT UPPER(e.EmployeeName + ' ' + e.EmployeeSurname) AS Manutentore
                FROM resetservices.dbo.tbuserkey AS U
                         INNER JOIN employee.dbo.employees AS e ON e.EmployeeId = u.idanga AND U.DataOut IS NULL
                         INNER JOIN employee.dbo.EmployeeHireHistory AS h \
                                    ON e.EmployeeId = h.EmployeeId AND h.EndWorkDate IS NULL AND h.employeerid = 2
                         INNER JOIN Traceability_rs.EQP.LogManutenzioni LM ON lm.UserName COLLATE DATABASE_DEFAULT = \
                                                                              UPPER(e.EmployeeName + ' ' + e.EmployeeSurname)
                ORDER BY UPPER(e.EmployeeName + ' ' + e.EmployeeSurname); \
                """
        try:
            self.cursor.execute(query)
            # Restituiamo una lista semplice di nomi
            return [row.Manutentore for row in self.cursor.fetchall()]
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei manutentori per report: {e}")
            return []

    def fetch_report_dates(self):
        """Recupera le date uniche in cui sono state fatte manutenzioni."""
        query = """
                SELECT DISTINCT CAST(lm.datestart AS DATE) AS DateMaintenance
                FROM Traceability_rs.EQP.LogManutenzioni lm
                ORDER BY CAST(lm.datestart AS DATE) DESC; \
                """
        try:
            self.cursor.execute(query)
            # Restituiamo una lista di oggetti data
            return [row.DateMaintenance for row in self.cursor.fetchall()]
        except pyodbc.Error as e:
            print(f"Errore nel recupero delle date per report: {e}")
            return []

    def search_maintenance_report(self, equipment_id=None, maintenance_date=None, maintainer_name=None):
        """Esegue la ricerca per il report di manutenzione con filtri opzionali."""
        # Query di base
        base_query = """
                     SELECT ROW_NUMBER() OVER (ORDER BY cm.ordine, UPPER(ISNULL(e.EmployeeName + ' ' + e.EmployeeSurname, '#ND'))) AS [Row],
                UPPER(ISNULL(e.EmployeeName + ' ' + e.EmployeeSurname, '#ND')) AS EmployeeName,
                cm.NomeCompito,
                eq.InternalName AS EquipmentName,
                cm.DescrizioneCompito,
                FORMAT(lm.datestart, 'd', 'ro-ro') AS DataIntervento,
                CONVERT(VARCHAR(8), lm.datestart, 108) AS InizioOraIntervento,
                CONVERT(VARCHAR(8), lm.datestop, 108) AS FineIntervento,
                DATEDIFF(MINUTE, lm.DateStart, lm.DateStop) AS [DurataInterventoInMin]
                     FROM Traceability_rs.EQP.LogManutenzioni LM
                         INNER JOIN employee.dbo.employees AS e \
                     ON lm.UserName COLLATE DATABASE_DEFAULT = UPPER (e.EmployeeName + ' ' + e.EmployeeSurname)
                         INNER JOIN resetservices.dbo.tbuserkey AS U ON e.EmployeeId = u.idanga AND U.DataOut IS NULL
                         INNER JOIN employee.dbo.EmployeeHireHistory AS h ON e.EmployeeId = h.EmployeeId AND h.EndWorkDate IS NULL AND h.employeerid = 2
                         INNER JOIN traceability_rs.eqp.CompitiManutenzione cm ON lm.CompitoId = cm.CompitoId
                         INNER JOIN traceability_rs.eqp.Equipments eq ON eq.EquipmentID = lm.EquipmentID \
                     """

        # Costruzione dinamica della clausola WHERE
        where_clauses = []
        params = []

        if equipment_id:
            where_clauses.append("eq.EquipmentId = ?")
            params.append(equipment_id)

        if maintenance_date:
            where_clauses.append("CAST(lm.dateStart AS DATE) = ?")
            params.append(maintenance_date)

        if maintainer_name:
            where_clauses.append("UPPER(e.EmployeeName + ' ' + e.EmployeeSurname) = ?")
            params.append(maintainer_name.upper())

        # Combina la query finale
        final_query = base_query
        if where_clauses:
            final_query += " WHERE " + " AND ".join(where_clauses)

        final_query += " ORDER BY lm.datestart, EmployeeName, cm.ordine;"

        try:
            self.cursor.execute(final_query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nella ricerca del report di manutenzione: {e}")
            self.last_error_details = str(e)
            return []

    def add_new_spare_part(self, material_part_number, material_code, material_description):
        """
        Inserisce una nuova parte di ricambio in eqp.SparePartMaterials,
        imposta toberevizited a 1 e restituisce il nuovo ID.
        """
        query = """
                INSERT INTO eqp.SparePartMaterials (MaterialPartNumber, MaterialCode, MaterialDescription, toberevizited)
                    OUTPUT INSERTED.SparePartMaterialId
                VALUES (?, ?, ?, 1);
                """
        try:
            # Usiamo fetchval() che Ã¨ perfetto per recuperare un singolo valore
            # da una query che restituisce una riga e una colonna, come il nostro OUTPUT.
            new_id = self.cursor.execute(query, material_part_number, material_code, material_description).fetchval()

            if new_id:
                self.conn.commit()
                return new_id
            else:
                # Questo caso Ã¨ improbabile con OUTPUT, ma Ã¨ una buona pratica gestirlo
                self.conn.rollback()
                self.last_error_details = "Inserimento nel DB riuscito ma impossibile recuperare il nuovo ID."
                return None

        except pyodbc.Error as e:
            self.conn.rollback()
            print(f"Errore nell'aggiunta di una nuova parte di ricambio: {e}")
            self.last_error_details = str(e)
            return None

    def fetch_setting(self, attribute_name):
        """Recupera un valore dalla tabella Settings."""
        query = "select [value] from traceability_rs.dbo.Settings where atribute = ?;"
        try:
            row = self.fetch_one(query, (attribute_name,))
            return row[0] if row else None
        except Exception as e:
            logger.error(f"Errore nel recupero impostazione '{attribute_name}': {e}")
            return None
        # NUOVO METODO: Aggiunge una nuova parte di ricambio al catalogo

    # def add_new_spare_part(self, material_part_number, material_code=None, material_description=None,
    #                        to_be_revizited=1):
    #     """Inserisce una nuova parte in eqp.SparePartMaterials e restituisce il nuovo ID."""
    #     # Assumiamo che la tabella abbia un IDENTITY ID (SparePartMaterialId)
    #     query = """
    #             INSERT INTO eqp.SparePartMaterials (MaterialPartNumber, MaterialCode, MaterialDescription,toberevizited)
    #             OUTPUT INSERTED.SparePartMaterialId
    #             VALUES (?, ?, ?, 1);
    #             """
    #     try:
    #         # Esegui l'INSERT
    #         self.cursor.execute(query, material_part_number, material_code, material_description)
    #
    #         # PoichÃ© abbiamo usato OUTPUT, l'INSERT ora restituisce un risultato che possiamo leggere con fetchval().
    #         new_id = self.cursor.fetchval()
    #
    #         if new_id:
    #             self.conn.commit()
    #             return new_id
    #         else:
    #             self.conn.rollback()
    #             self.last_error_details = "Inserimento riuscito ma impossibile recuperare il nuovo ID."
    #             return None

        # except pyodbc.Error as e:
        #     self.conn.rollback()
        #     print(f"Errore nell'aggiunta nuova parte di ricambio: {e}")
        #     self.last_error_details = str(e)
        #     return None

    def insert_spare_part_request(self, equipment_id, spare_part_id, quantity, notes, requested_by):
        """Inserisce una nuova richiesta di parti di ricambio o intervento."""
        query = """
                INSERT INTO eqp.RequestSpareParts
                (EquipmentId, SparePartMaterialId, Quantity, Note, RequestedBy, DateRequest, Solved)
                VALUES (?, ?, ?, ?, ?, GETDATE(), 0)
                """
        try:
            self.cursor.execute(query, equipment_id, spare_part_id, quantity, notes, requested_by)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            print(f"Errore nell'inserimento richiesta: {e}")
            self.last_error_details = str(e)
            return False

    def fetch_and_open_document_by_task_id(self, task_id):
        """Recupera e apre il documento associato a un CompitoId."""

        self.last_error_details = ""

        # Query fornita dall'utente. Aggiungiamo FileName e FileType necessari per l'apertura.
        # Aggiungiamo ORDER BY per assicurarci di prendere il documento piÃ¹ recente se la JOIN ne producesse piÃ¹ di uno.
        query = """
                select emd.DocumentSource, emd.FileName, emd.FileType
                from Traceability_rs.eqp.EquipmentMantainanceDocs emd
                         left join Traceability_rs.eqp.CompitiManutenzione CM \
                                   on cm.ProgrammedInterventionId = emd.ProgrammedInterventionId
                where cm.compitoid = ?
                -- Potresti voler aggiungere AND emd.DateOut IS NULL se vuoi mostrare solo documenti attivi
                ORDER BY emd.DateSys DESC
                """
        try:
            self.cursor.execute(query, task_id)
            # Usiamo fetchone() per prendere solo il primo risultato (il piÃ¹ recente)
            row = self.cursor.fetchone()

            if row and row.DocumentSource:
                binary_data = row.DocumentSource

                # --- Logica Gestione File Temporaneo (Robusta) ---

                # 1. Determina estensione e prefisso
                file_extension = row.FileType if row.FileType else os.path.splitext(row.FileName)[1]
                if not file_extension:
                    print(f"Attenzione: Estensione mancante per task ID {task_id}. Usando default .pdf")
                    file_extension = '.pdf'  # Default ragionevole

                if not file_extension.startswith('.'):
                    file_extension = '.' + file_extension

                temp_prefix = "task_doc_"
                if row.FileName:
                    # Pulisce il nome del file (rimuove caratteri non sicuri) per usarlo come prefisso
                    safe_name = re.sub(r'[^\w\-]', '_', os.path.splitext(row.FileName)[0])
                    temp_prefix = safe_name[:50] + "_"

                # 2. Crea e scrivi file temporaneo
                # delete=False Ã¨ necessario affinchÃ© il programma esterno possa aprirlo
                temp_file = tempfile.NamedTemporaryFile(prefix=temp_prefix, delete=False, suffix=file_extension)
                temp_file.write(binary_data)
                temp_file.close()  # Chiudi handle Python per permettere apertura esterna

                print(f"Apertura documento per compito ID {task_id}: {temp_file.name}")

                # 3. Apertura File (Cross-platform)
                try:
                    if sys.platform == "win32":
                        # Metodo per Windows (es. apre con Adobe Reader o Word/Excel predefinito)
                        os.startfile(temp_file.name)
                    else:
                        # Fallback per macOS e Linux
                        opener = "open" if sys.platform == "darwin" else "xdg-open"
                        subprocess.call([opener, temp_file.name])
                except Exception as open_e:
                    self.last_error_details = f"Errore OS nell'apertura del file: {open_e}"
                    return False

                return True
            else:
                # Caso in cui la query non restituisce risultati (nessun documento collegato)
                self.last_error_details = f"Nessun documento trovato associato al compito ID {task_id}."
                return False

        except pyodbc.Error as e:
            print(f"Errore durante il recupero del documento dal DB per task ID {task_id}: {e}")
            self.last_error_details = f"Errore Database: {e}"
            return False
        except Exception as e:
            # Gestione errori (es. permessi scrittura file temporaneo)
            print(f"Errore imprevisto durante la gestione del file temporaneo: {e}")
            self.last_error_details = f"Errore Applicazione (File System): {e}"
            return False

    def fetch_spare_parts(self):
        """Recupera la lista di parti di ricambio e servizi disponibili."""
        # ATTENZIONE: Assicurati che i nomi delle colonne corrispondano alla tua tabella eqp.SparePartMaterials.
        query = """
                SELECT SparePartMaterialId, MaterialPartNumber, MaterialCode, MaterialDescription
                FROM eqp.SparePartMaterials
                ORDER BY MaterialCode \
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero parti di ricambio: {e}")
            self.last_error_details = str(e)
            return []



    def fetch_and_open_maintenance_document(self, document_id):
        """Recupera i dati binari di un documento di manutenzione dal DB, lo salva temporaneamente e lo apre."""

        self.last_error_details = ""  # Resetta l'errore

        try:
            # Seleziona i dati binari (DocumentSource) e i metadati (FileType, FileName)
            sql_select = """
                         SELECT DocumentSource, FileName, FileType
                         FROM eqp.EquipmentMantainanceDocs
                         WHERE EquipmentDocumentationId = ? \
                         """
            self.cursor.execute(sql_select, document_id)
            row = self.cursor.fetchone()

            if row and row.DocumentSource:
                binary_data = row.DocumentSource

                # --- 1. Gestione Robusta dell'Estensione e del Nome File Temporaneo ---

                # Determina l'estensione: PrioritÃ  a FileType, fallback su FileName, default a .pdf
                file_extension = row.FileType if row.FileType else os.path.splitext(row.FileName)[1]
                if not file_extension:
                    print(f"Attenzione: Estensione mancante per doc ID {document_id}. Usando default .pdf")
                    file_extension = '.pdf'  # Default ragionevole

                # Assicurarsi che l'estensione inizi con '.'
                if not file_extension.startswith('.'):
                    file_extension = '.' + file_extension

                # Usa una versione pulita del nome file originale come prefisso per facilitare l'identificazione
                temp_prefix = "doc_"
                if row.FileName:
                    # Pulisce il nome del file (rimuove caratteri non sicuri)
                    safe_name = re.sub(r'[^\w\-]', '_', os.path.splitext(row.FileName)[0])
                    temp_prefix = safe_name[:50] + "_"  # Limita la lunghezza e aggiunge separatore

                # --- 2. Creazione File Temporaneo ---

                # delete=False Ã¨ necessario affinchÃ© il programma esterno possa aprirlo prima che Python lo elimini
                temp_file = tempfile.NamedTemporaryFile(prefix=temp_prefix, delete=False, suffix=file_extension)
                temp_file.write(binary_data)
                # Chiudi il file handle in Python affinchÃ© il sistema operativo possa aprirlo
                temp_file.close()

                print(f"Apertura del file temporaneo: {temp_file.name}")

                # --- 3. Apertura File (Cross-platform) ---
                try:
                    if sys.platform == "win32":
                        # Metodo specifico per Windows
                        os.startfile(temp_file.name)
                    else:
                        # Fallback per macOS e Linux
                        opener = "open" if sys.platform == "darwin" else "xdg-open"
                        subprocess.call([opener, temp_file.name])
                except Exception as open_e:
                    self.last_error_details = f"Errore OS nell'apertura del file: {open_e}"
                    return False

                return True
            else:
                self.last_error_details = f"Documento ID {document_id} non trovato o dati binari assenti nel database."
                return False

        except pyodbc.Error as e:
            print(f"Errore durante il recupero del documento dal DB: {e}")
            self.last_error_details = f"Errore Database: {e}"
            return False
        except Exception as e:
            # Gestione errori generici (es. permessi scrittura file temporaneo)
            print(f"Errore imprevisto durante la gestione del file temporaneo: {e}")
            self.last_error_details = f"Errore Applicazione (File System): {e}"
            return False

    def fetch_programmed_interventions(self):
        """Recupera tutti gli interventi programmati (Tipi di manutenzione)."""
        query = """
                SELECT [ProgrammedInterventionId]
                      ,[TimingDescriprion]
                      ,[TimingValue]
                FROM [Traceability_RS].[eqp].[ProgrammedInterventions]
                ORDER BY [TimingValue] -- Ordinamento per facilitare la lettura
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero degli interventi programmati: {e}")
            return []

        # Assicuriamoci che DateOut sia NULL per il nuovo documento.
        insert_query = """
                       INSERT INTO eqp.EquipmentMantainanceDocs
                       (EquipmentId, ProgrammedInterventionId, EquipmentMaintenanceDocTypeId, DocDescription, FileName, \
                        FileType, DocumentSource, UploadedBy, DateSys, DateOut)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), NULL) \
                       """
        try:
            # --- INIZIO TRANSAZIONE ATOMICA ---
            # (PoichÃ© pyodbc Ã¨ stato connesso con autocommit=False, siamo giÃ  in una transazione)

            # 1. Se richiesto, invalida i vecchi documenti (UPDATE)
            if invalidate_ids:
                if not self._invalidate_maintenance_docs(invalidate_ids):
                    # Se l'UPDATE fallisce, annulla tutto (ROLLBACK) e restituisci False
                    self.conn.rollback()
                    return False

            # 2. Inserisci il nuovo documento (INSERT)
            self.cursor.execute(insert_query, equipment_id, intervention_id, doc_type_id, description, file_name,
                                file_type, binary_data, user_name)

            # 3. Conferma la transazione (COMMIT)
            self.conn.commit()
            return True

            # --- FINE TRANSAZIONE ATOMICA ---

        except pyodbc.Error as e:
            self.conn.rollback()  # Annulla tutto in caso di errore nell'INSERT
            self.last_error_details = str(e)
            print(f"Errore SQL nella transazione replace_maintenance_document: {e}")
            return False

    def search_equipments(self, filters):
        """
        Esegue una ricerca dinamica delle macchine basata su un dizionario di filtri.
        """
        query = """
                SELECT eq.EquipmentId, eq.InternalName, eq.SerialNumber, b.Brand, et.EquipmentType, pp.ParentPhaseName
                FROM eqp.Equipments eq
                         LEFT JOIN eqp.EquipmentBrands b ON eq.BrandId = b.EquipmentBrandId
                         LEFT JOIN eqp.EquipmentTypes et ON eq.EquipmentTypeId = et.EquipmentTypeId
                         LEFT JOIN dbo.ParentPhases pp ON eq.ParentPhaseId = pp.IDParentPhase
                """
        where_clauses = []
        params = []

        if filters.get('brand_id'):
            where_clauses.append("eq.BrandId = ?")
            params.append(filters['brand_id'])

        if filters.get('type_id'):
            where_clauses.append("eq.EquipmentTypeId = ?")
            params.append(filters['type_id'])

        if filters.get('phase_id'):
            where_clauses.append("eq.ParentPhaseId = ?")
            params.append(filters['phase_id'])

        if filters.get('search_text'):
            # Cerca il testo nel nome, seriale o inventario
            where_clauses.append("(eq.InternalName LIKE ? OR eq.SerialNumber LIKE ? OR eq.InventoryNumber LIKE ?)")
            search_param = f"%{filters['search_text']}%"
            params.extend([search_param, search_param, search_param])

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        query += " ORDER BY eq.InternalName;"

        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nella ricerca macchine: {e}")
            return []

    def fetch_full_equipment_details(self, equipment_id):
        """
        Recupera tutte le informazioni correlate a una singola macchina.
        """
        details = {}
        try:
            # 1. Dati anagrafici
            master_query = """
                           SELECT eq.*, b.Brand, et.EquipmentType, pp.ParentPhaseName
                           FROM eqp.Equipments eq
                                    LEFT JOIN eqp.EquipmentBrands b ON eq.BrandId = b.EquipmentBrandId
                                    LEFT JOIN eqp.EquipmentTypes et ON eq.EquipmentTypeId = et.EquipmentTypeId
                                    LEFT JOIN dbo.ParentPhases pp ON eq.ParentPhaseId = pp.IDParentPhase
                           WHERE eq.EquipmentId = ?
                           """
            details['master'] = self.cursor.execute(master_query, equipment_id).fetchone()

            # 2. Log delle modifiche
            changes_query = "SELECT Changed, WhoChange, DateChange FROM eqp.EquipmentChanges WHERE EquipmentId = ? ORDER BY DateChange DESC"
            details['changes'] = self.cursor.execute(changes_query, equipment_id).fetchall()

            # 3. Documenti di manutenzione
            # NOTA: Selezioniamo FileName invece di DocumentSource (che Ã¨ binario e pesante) per l'elenco
            docs_query = "SELECT FileName, UploadedBy, DateSys FROM eqp.EquipmentMantainanceDocs WHERE EquipmentId = ? ORDER BY DateSys DESC"
            details['docs'] = self.cursor.execute(docs_query, equipment_id).fetchall()

            # 4. Schede di manutenzione compilate
            logs_query = "SELECT DataEsecuzione, IdManutentore, NoteGenerali FROM dbo.LogManutenzioni WHERE EquipmentId = ? ORDER BY DataEsecuzione DESC"
            details['logs'] = self.cursor.execute(logs_query, equipment_id).fetchall()

            return details
        except pyodbc.Error as e:
            print(f"Errore nel recupero dettagli completi macchina: {e}")
            return None

    def fetch_all_equipments(self, only_with_plan=False, phase_id=None, equipment_type_id=None, equipment_type_filter=None):
        """Recupera ID, Nome Interno, Seriale e tipo (Fixture/Stensil) di tutte le macchine per la selezione."""
        query = """
            SELECT DISTINCT e.EquipmentId, 
                   InternalName + IIF(cm.CompitoId IS NULL, '', ' (*) ') AS InternalName, 
                   SerialNumber,
                   ISNULL(e.IsFixture, 0) AS IsFixture,
                   ISNULL(e.IsStensil, 0) AS IsStensil
            FROM eqp.Equipments E 
            LEFT JOIN [eqp].[CompitiManutenzione] CM ON e.EquipmentId = cm.EquipmentId
        """
        
        where_clauses = []
        params = []
        
        if only_with_plan:
            where_clauses.append("cm.CompitoId IS NOT NULL")
        
        if phase_id is not None:
            where_clauses.append("e.ParentPhaseId = ?")
            params.append(phase_id)
        
        if equipment_type_id is not None:
            where_clauses.append("e.EquipmentTypeId = ?")
            params.append(equipment_type_id)
        
        # Nuovo filtro per tipo di equipaggiamento
        if equipment_type_filter == 'fixture':
            where_clauses.append("ISNULL(e.IsFixture, 0) = 1")
        elif equipment_type_filter == 'stensil':
            where_clauses.append("ISNULL(e.IsStensil, 0) = 1")
        # Se equipment_type_filter è 'all' o None, non aggiungiamo filtri
        
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        
        query += " ORDER BY InternalName, SerialNumber;"

        try:
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero delle macchine: {e}")
            return []

    def fetch_equipment_details(self, equipment_id):
        """Recupera i dettagli di una singola macchina per la modifica."""
        query = """SELECT ParentPhaseId, EquipmentTypeId, InternalName, SerialNumber, 
                          ProductionYear, IsFixture, IsStensil, MustCalibrated 
                   FROM eqp.Equipments 
                   WHERE EquipmentId = ?;"""
        try:
            self.cursor.execute(query, equipment_id)
            return self.cursor.fetchone()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei dettagli macchina: {e}")
            return None

    def update_and_log_equipment_changes(self, equipment_id, new_phase_id, new_internal_name, new_serial,
                                         change_log_string, user_name, new_equipment_type_id=None, new_production_year=None,
                                         new_is_fixture=None, new_is_stensil=None, new_must_calibrated=None):
        """Aggiorna la macchina e registra la modifica in una transazione."""
        try:
            # Costruisce la query dinamicamente in base ai parametri forniti
            set_clauses = [
                "ParentPhaseId = ?",
                "InternalName = ?",
                "SerialNumber = ?"
            ]
            params = [new_phase_id, new_internal_name, new_serial]
            
            # Aggiungi campi opzionali se forniti
            if new_equipment_type_id is not None:
                set_clauses.append("EquipmentTypeId = ?")
                params.append(new_equipment_type_id)
            
            if new_production_year is not None:
                set_clauses.append("ProductionYear = ?")
                params.append(new_production_year)
            
            if new_is_fixture is not None:
                set_clauses.append("IsFixture = ?")
                params.append(1 if new_is_fixture else 0)
            
            if new_is_stensil is not None:
                set_clauses.append("IsStensil = ?")
                params.append(1 if new_is_stensil else 0)
            
            if new_must_calibrated is not None:
                set_clauses.append("MustCalibrated = ?")
                params.append(1 if new_must_calibrated else 0)
            
            # Aggiungi equipment_id alla fine
            params.append(equipment_id)
            
            # Costruisci la query completa
            update_query = f"""
                UPDATE eqp.Equipments
                SET {', '.join(set_clauses)}
                WHERE EquipmentId = ?;
            """
            
            self.cursor.execute(update_query, *params)

            # 2. Inserisce il log delle modifiche
            log_query = """
                        INSERT INTO eqp.EquipmentChanges (EquipmentId, Changed, WhoChange, DateChange)
                        VALUES (?, ?, ?, GETDATE());
                        """
            self.cursor.execute(log_query, equipment_id, change_log_string, user_name)

            # 3. Se entrambe le operazioni vanno a buon fine, conferma la transazione
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            # Se una delle due query fallisce, annulla tutto
            self.conn.rollback()
            self.last_error_details = str(e)
            print(f"Errore durante l'aggiornamento della macchina: {e}")
            return False

    def fetch_brands(self):
        """Recupera tutti i brand delle macchine."""
        query = "SELECT EquipmentBrandId, Brand FROM eqp.EquipmentBrands ORDER BY Brand;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei brand: {e}")
            return []

    def fetch_equipment_types(self):
        """Recupera tutti i tipi di macchine."""
        query = "SELECT EquipmentTypeId, EquipmentType FROM eqp.EquipmentTypes ORDER BY EquipmentType;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei tipi di macchine: {e}")
            return []

    def fetch_parent_phases_for_maintenance(self):
        """Recupera tutte le fasi di produzione."""
        query = "SELECT IDParentPhase, ParentPhaseName FROM dbo.ParentPhases ORDER BY ParentPhaseName;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero delle fasi di produzione: {e}")
            return []

    def add_new_equipment(self, brand_id, type_id, phase_id, serial_number, internal_name, prod_year, inv_number, must_calibrated):
        """Salva una nuova macchina nel database."""
        query = """
                INSERT INTO eqp.Equipments
                (BrandId, EquipmentTypeId, ParentPhaseId, SerialNumber, InternalName, ProductionYear, InventoryNumber,
                 DateSys, MustCalibrated)
                VALUES (?, ?, ?, ?, ?, ?, ?, GETDATE(), ?)
                """
        try:
            self.cursor.execute(query, brand_id, type_id, phase_id, serial_number, internal_name, prod_year, inv_number, must_calibrated)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            self.last_error_details = str(e)
            print(f"Errore durante l'inserimento della macchina: {e}")
            return False

    def fetch_translations(self):
        query = "SELECT LanguageCode, TranslationKey, TranslationValue FROM Traceability_rs.dbo.AppTranslations;"
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Error fetching translations: {e}")
            return []

    def authenticate_user(self, user_id, password):
        query = """
                SELECT u.NomeUser, ISNULL(e.EmployeeName + ' ' + e.EmployeeSurname, '#ND') as EmployeeName, u.pass
                FROM resetservices.dbo.tbuserkey as U
                INNER JOIN employee.dbo.employees as e ON e.EmployeeId = u.idanga
                INNER JOIN employee.dbo.EmployeeHireHistory as h ON e.EmployeeId = h.EmployeeId
                WHERE h.EndWorkDate IS NULL AND h.employeerid = 2 AND u.Nomeuser = ? AND Pass = ?;
                """
        try:

            self.cursor.execute(query, user_id, password)
            row = self.cursor.fetchone()
            return row.EmployeeName if row else None

        except pyodbc.Error as e:
            print(f"Error during authentication: {e}")
            return None

    def fetch_products(self):
        query = """
                SELECT DISTINCT p.IDProduct, p.ProductCode + ' ['+ SUBSTRING(p.productcode, 3, 2) + ']' AS ProductCode
                FROM Traceability_RS.dbo.Products AS P
                INNER JOIN [Traceability_RS].[dbo].[ProductParentPhases] AS PP ON pp.IDProduct = p.IDProduct
                INNER JOIN Traceability_RS.dbo.ParentPhases AS PF ON pf.IDParentPhase = pp.IDParentPhase
                ORDER BY p.ProductCode + ' ['+ SUBSTRING(p.productcode, 3, 2) + ']';
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error:
            return []

    def fetch_products_with_documents(self):
        query = """
                SELECT p.IDProduct,
                       p.ProductCode + ' [' + CAST(doc_counts.DocCount AS NVARCHAR(10)) + ' docs]' AS ProductCode
                FROM Traceability_RS.dbo.Products AS p
                         INNER JOIN (SELECT ProductId, COUNT(*) AS DocCount
                                     FROM Traceability_RS.dbo.ProductDocuments
                                     WHERE DateOutOfValidation IS NULL
                                     GROUP BY ProductId) AS doc_counts ON p.IDProduct = doc_counts.ProductId
                ORDER BY ProductCode;
                """
        try:
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pyodbc.Error:
            return []

    def fetch_phases_with_documents_for_product(self, product_id):
        """
        Recupera solo le fasi di produzione che contengono almeno un documento
        per un dato ID prodotto.
        """
        # Query ottimizzata per restituire una lista pulita di fasi
        query = """
                SELECT DISTINCT pp.IDParentPhase, pp.ParentPhaseName
                FROM dbo.ProductDocuments p
                         INNER JOIN dbo.ParentPhases pp ON p.ParentPhaseId = pp.IDParentPhase
                WHERE p.ProductId = ?
                ORDER BY pp.ParentPhaseName; \
                """
        try:
            self.cursor.execute(query, product_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero delle fasi con documenti: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_parent_phases(self, id_product):
        query = """
                SELECT distinct pf.IDParentPhase, pf.ParentPhaseName + IIF(pp.IDProduct IS NULL, '*', '') AS Phase
                FROM Traceability_RS.dbo.ParentPhases AS pf
                LEFT JOIN Traceability_RS.dbo.ProductParentPhases AS pp ON pf.IDParentPhase = pp.IDParentPhase AND pp.IDProduct = ?
                ORDER BY Phase;
                """
        try:
            self.cursor.execute(query, id_product)
            return self.cursor.fetchall()
        except pyodbc.Error:
            return []

    def fetch_existing_documents(self, product_id, parent_phase_id):
        """Recupera i documenti di un prodotto per una fase (tutte le revisioni)"""
        try:
            query = """
                    SELECT DocumentProductionID, \
                           DocumentName, \
                           DocumentRevisionNumber, \
                           Validated, \
                           ApprovatoDa, \
                           DocumentData, \
                           DateOutOfValidation
                    FROM [Traceability_RS].[dbo].[ProductDocuments]
                    WHERE ProductId = ? AND ParentPhaseId = ?                       
                    ORDER BY DateSys DESC \
                    """

            self.cursor.execute(query, (product_id, parent_phase_id))
            rows = self.cursor.fetchall()

            return rows

        except Exception as e:
            logger.error("Error fetching existing documents: %s", e)
            return []

    def fetch_and_open_document(self, document_id):
        """Recupera il documento dal database e lo apre (solo se validato e non fuori validazione)"""
        try:
            query = """
                    SELECT DocumentName, \
                           DocumentData, \
                           Validated, \
                           ApprovatoDa, \
                           DateOutOfValidation
                    FROM [Traceability_RS].[dbo].[ProductDocuments]
                    WHERE DocumentProductionID = ? \
                    """
            
            query2 = """
                    INSERT INTO [Traceability_RS].[dbo].[DocumentProductViews]
                    ([DocumentProductionId])
                    VALUES
                    (?)
                    """
            
            logger.info("Fetching document: %s", document_id)

            # Prima esegui la SELECT e salva i risultati
            self.cursor.execute(query, (document_id,))
            row = self.cursor.fetchone()
            
            # Poi esegui l'INSERT per tracciare la visualizzazione
            if row:
                self.cursor.execute(query2, (document_id,))
                self.conn.commit()

            if not row:
                logger.error("Document not found: %s", document_id)
                return False

            # Controlla se il documento Ã¨ validato e non fuori validazione
            if not row.Validated or row.ApprovatoDa is None:
                logger.warning("Document %s is not validated (ApprovatoDa is NULL)", document_id)
                return False
            
            # NUOVO: Controlla anche DateOutOfValidation
            if row.DateOutOfValidation is not None:
                logger.warning("Document %s is out of validation (DateOutOfValidation is set)", document_id)
                return False

            doc_name = row.DocumentName
            doc_data = row.DocumentData

            if not doc_data:
                logger.error("Document has no binary data: %s", document_id)
                return False

            # Crea file temporaneo
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, doc_name)

            with open(temp_path, 'wb') as f:
                f.write(doc_data)

            # Apri il file
            if sys.platform == 'win32':
                os.startfile(temp_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{temp_path}"')
            else:
                os.system(f'xdg-open "{temp_path}"')

            logger.info("Document opened: %s", doc_name)
            return True

        except Exception as e:
            logger.error("Error fetching and opening document: %s", e)
            return False

    def save_document_to_db(
            self,
            product_id,
            parent_phase_id,
            doc_name,
            local_file_path,
            revision,
            user_name,
            validated_as_int,
            document_date=None,
            validator_info=None
    ):
        """Salva il documento nella tabella ProductDocuments con data e info validatore"""
        try:
            # Leggi il file
            with open(local_file_path, 'rb') as f:
                file_content = f.read()

            from datetime import datetime
            date_sys = datetime.now()
            data_caricamento = datetime.now()

            # Se nessuna data fornita, usa quella attuale
            if document_date is None:
                document_date = date_sys

            query = """
                    INSERT INTO [Traceability_RS].[dbo].[ProductDocuments]
                    (ProductId, \
                     ParentPhaseId, \
                     DocumentName, \
                     DocumentRevisionNumber, \
                     DocumentPath, \
                     Validated, \
                     DateIn, \
                     UserName, \
                     DateSys, \
                     DocumentData, \
                     IsGenericDocument, \
                     DataCaricamento, \
                     ApprovatoDa, \
                     ApprovatoOn)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) \
                    """

            # Prepara i valori
            approvato_on = datetime.now() if (validated_as_int == 1 and validator_info) else None
            document_path = local_file_path
            is_generic_document = 0

            params = (
                product_id,
                parent_phase_id,
                doc_name,
                revision,
                document_path,
                validated_as_int,
                document_date,
                user_name,
                date_sys,
                file_content,
                is_generic_document,
                data_caricamento,
                validator_info,
                approvato_on
            )

            self.cursor.execute(query, params)
            self.conn.commit()

            logger.info(
                "Document saved successfully to ProductDocuments: %s (validated=%s, validator=%s)",
                doc_name,
                validated_as_int,
                validator_info
            )
            return True

        except FileNotFoundError:
            self.last_error_details = f"File non trovato: {local_file_path}"
            logger.error("File not found: %s", local_file_path)
            return False

        except Exception as e:
            self.last_error_details = str(e)
            logger.error("Error saving document to ProductDocuments: %s", e)
            return False

    def update_document_validation(self, document_id, validator_name):
        """Aggiorna lo stato di validazione di un documento nella tabella ProductDocuments"""
        try:
            from datetime import datetime

            query = """
                    UPDATE [Traceability_RS].[dbo].[ProductDocuments]
                    SET
                        Validated = 1, ApprovatoDa = ?, ApprovatoOn = ?, DateOutOfValidation = NULL
                    WHERE DocumentProductionID = ? \
                    """

            params = (validator_name, datetime.now(), document_id)

            self.cursor.execute(query, params)
            self.conn.commit()

            logger.info(
                "Document %s validated by %s in ProductDocuments table",
                document_id,
                validator_name
            )
            return True

        except Exception as e:
            self.last_error_details = str(e)
            logger.error("Error updating document validation in ProductDocuments: %s", e)
            return False

    def fetch_latest_version_info(self, software_name):
        """
        Recupera la versione piÃ¹ recente, il percorso di aggiornamento e il flag Must per un dato software.
        """
        query = "SELECT Version, MainPath, ISNULL(Must, 0) as Must FROM traceability_rs.dbo.SwVersions WHERE NameProgram = ? AND dateout IS NULL"
        with self._lock:
            try:
                logger.info(f"fetch_latest_version_info: Esecuzione query per software_name='{software_name}'")
                self.cursor.execute(query, software_name)
                logger.info(f"fetch_latest_version_info: Query eseguita, recupero risultati...")
                result = self.cursor.fetchone()
                logger.info(f"fetch_latest_version_info: Risultato: {result}")
                return result
            except pyodbc.Error as e:
                logger.error(f"Errore durante il recupero della versione del software: {e}")
                return None

    def fetch_available_maintenance_plans(self, equipment_id):
        """Recupera i piani di manutenzione disponibili per una macchina, basandosi sui compiti assegnati."""
        # La logica per determinare se un piano Ã¨ "scaduto" Ã¨ complessa e la manteniamo,
        # ma la struttura della query Ã¨ piÃ¹ semplice senza join inutili.
        query = """
                WITH LatestLogs AS (SELECT CompitoId, MAX(DateStop) AS LastCompletionDate \
                                    FROM eqp.LogManutenzioni \
                                    WHERE EquipmentId = ? \
                                    GROUP BY CompitoId)
                SELECT DISTINCT pi.ProgrammedInterventionId, pi.TimingDescriprion
                FROM eqp.CompitiManutenzione cm
                         INNER JOIN eqp.ProgrammedInterventions pi \
                                    ON cm.ProgrammedInterventionId = pi.ProgrammedInterventionId
                         LEFT JOIN LatestLogs ll ON cm.CompitoId = ll.CompitoId
                WHERE cm.EquipmentId = ? \
                  AND (
                    ll.LastCompletionDate IS NULL OR
                    (CASE
                         WHEN pi.TimingValue < 1 THEN IIF(DATEDIFF(HOUR, ll.LastCompletionDate, GETDATE()) > 8, 1, 0)
                         WHEN pi.TimingValue >= 1 THEN IIF( \
                                 DATEDIFF(DAY, ll.LastCompletionDate, GETDATE()) >= pi.TimingValue, 1, 0)
                         ELSE 1
                        END) = 1
                    )
                ORDER BY pi.TimingDescriprion; \
                """
        try:
            self.cursor.execute(query, equipment_id, equipment_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero piani manutenzione: {e}")
            self.last_error_details = str(e)
            return []

    def fetch_maintenance_tasks(self, programmed_intervention_id, equipment_id):
        query = """
                WITH LatestLogs AS (SELECT CompitoId, MAX(DateStop) AS LastCompletionDate \
                                    FROM eqp.LogManutenzioni \
                                    WHERE EquipmentId = ? \
                                    GROUP BY CompitoId)
                SELECT cm.CompitoId, cm.NomeCompito, cm.Categoria, cm.DescrizioneCompito
                FROM eqp.CompitiManutenzione AS cm
                         INNER JOIN eqp.ProgrammedInterventions AS pin \
                                    ON cm.ProgrammedInterventionId = pin.ProgrammedInterventionId
                         LEFT JOIN LatestLogs AS ll ON cm.CompitoId = ll.CompitoId
                WHERE cm.ProgrammedInterventionId = ?
                  AND cm.EquipmentId = ?
                  AND (
                    ll.LastCompletionDate IS NULL OR
                    (CASE
                         WHEN pin.TimingValue = 0 THEN 1
                         WHEN pin.TimingValue < 1 THEN IIF(DATEDIFF(HOUR, ll.LastCompletionDate, GETDATE()) > 8, 1, 0)
                         WHEN pin.TimingValue >= 1 THEN IIF( \
                                 DATEDIFF(DAY, ll.LastCompletionDate, GETDATE()) >= pin.TimingValue, 1, 0)
                         ELSE 0
                        END) = 1
                    )
                ORDER BY cm.Ordine, cm.CompitoId; \
                """
        try:
            # I parametri sono: equipment_id (per WITH), intervention_id (per WHERE), equipment_id (per WHERE)
            self.cursor.execute(query, equipment_id, programmed_intervention_id, equipment_id)
            return self.cursor.fetchall()
        except pyodbc.Error as e:
            print(f"Errore nel recupero dei task da eseguire: {e}")
            self.last_error_details = str(e)
            return []

    def log_completed_tasks(self, equipment_id, user_name, completed_task_ids, start_time, notes=""):
        """Inserisce record in LogManutenzioni per i compiti completati in una transazione batch."""
        if not completed_task_ids:
            return True
        query = """
                INSERT INTO Traceability_rs.eqp.LogManutenzioni
                (CompitoId, EquipmentId, UserName, DateStop, DateStart, NoteGenerali)
                VALUES (?, ?, ?, GETDATE(), ?, ?)
                """
        try:
            # Prepariamo i dati per l'inserimento batch
            batch_data = []
            for task_id in completed_task_ids:
                # Ordine parametri: (CompitoId, EquipmentId, IdManutentore, StartTime, NoteGenerali)
                batch_data.append((task_id, equipment_id, user_name, start_time, notes))

            # Esegui l'inserimento batch (efficiente)
            self.cursor.executemany(query, batch_data)
            self.conn.commit()
            return True
        except pyodbc.Error as e:
            self.conn.rollback()
            print(f"Errore nel logging dei compiti completati: {e}")
            self.last_error_details = str(e)
            return False

    def fetch_and_open_document_by_task_id(self, task_id):
        """
        Recupera e apre il documento specifico per un compito, leggendo
        i dati binari dalla colonna [LinkedDocument] della tabella CompitiManutenzione.
        """
        self.last_error_details = ""

        # --- QUERY MODIFICATA ---
        # Seleziona direttamente il documento binario dalla riga del compito specifico.
        query = """
                SELECT LinkedDocument
                FROM eqp.CompitiManutenzione
                WHERE CompitoId = ?; \
                """
        try:
            self.cursor.execute(query, task_id)
            row = self.cursor.fetchone()

            # Controlla che la riga esista e che il campo LinkedDocument non sia vuoto
            if row and row.LinkedDocument:
                binary_data = row.LinkedDocument

                # --- Logica per il file temporaneo (con nome generico) ---
                # PoichÃ© non abbiamo il nome/tipo file, usiamo un default.
                file_extension = '.pdf'  # Assumiamo PDF come default
                temp_prefix = f"task_{task_id}_documento_"

                temp_file = tempfile.NamedTemporaryFile(prefix=temp_prefix, delete=False, suffix=file_extension)
                temp_file.write(binary_data)
                temp_file.close()

                print(f"Apertura documento specifico per compito ID {task_id}: {temp_file.name}")

                # Apertura del file cross-platform
                try:
                    if sys.platform == "win32":
                        os.startfile(temp_file.name)
                    else:
                        opener = "open" if sys.platform == "darwin" else "xdg-open"
                        subprocess.call([opener, temp_file.name])
                    return True
                except Exception as open_e:
                    self.last_error_details = f"Errore del sistema operativo nell'apertura del file: {open_e}"
                    return False
            else:
                self.last_error_details = f"Nessun documento specifico trovato per il compito ID {task_id}."
                return False

        except pyodbc.Error as e:
            print(f"Errore DB durante il recupero del documento specifico per il task ID {task_id}: {e}")
            self.last_error_details = f"Errore Database: {e}"
            return False
        except Exception as e:
            print(f"Errore imprevisto durante la gestione del file temporaneo: {e}")
            self.last_error_details = f"Errore Applicazione (File System): {e}"
            return False

    def fetch_equipment_cycle_status(self, equipment_id):
        """
        Verifica lo stato dei cicli per un equipaggiamento.
        Ritorna records solo se l'equipaggiamento ha raggiunto soglie critiche (>=95%).
        
        Returns:
            list: Lista di record con informazioni sui cicli se soglie critiche raggiunte, [] altrimenti
        """
        query = """
        WITH LastMaint AS (
            SELECT 
                LM.EquipmentID,
                MAX(LM.DateStop) AS LastDateStop
            FROM Traceability_RS.eqp.LogManutenzioni LM
            GROUP BY LM.EquipmentID
        ),
        ScanBase AS (
            SELECT 
                S.ScanTimeFinish,
                S.IsPass,
                P.ProductCode,
                Ph.PhaseName,
                E.EquipmentId,
                E.InternalName + ' S/N:' + E.SerialNumber AS Equipment,
                T.EquipmentType,
                ISNULL(T.EndOfLifeCycle, 100000) AS EndOfLifeCycle
            FROM Traceability_rs.dbo.Scannings S
            INNER JOIN Traceability_rs.dbo.OrderPhases OP   ON OP.IDOrderPhase = S.IDOrderPhase
            INNER JOIN Traceability_rs.dbo.Phases Ph        ON Ph.IDPhase      = OP.IDPhase
            INNER JOIN Traceability_rs.dbo.Boards  B        ON B.IDBoard       = S.IDBoard
            INNER JOIN Traceability_rs.dbo.Orders  O        ON O.IDOrder       = B.IDOrder
            INNER JOIN Traceability_rs.dbo.Products P       ON P.IDProduct     = O.IDProduct
            INNER JOIN Traceability_rs.eqp.EquipmentFixtures F ON F.IdProduct  = P.IDProduct
            INNER JOIN Traceability_rs.eqp.Equipments E     ON E.EquipmentId   = F.EquipmentId
            INNER JOIN Traceability_rs.eqp.EquipmentTypes T ON T.EquipmentTypeId = E.EquipmentTypeId
            INNER JOIN Traceability_RS.eqp.EquipmentFixtureRules R 
                ON R.EquipmentTypeId = T.EquipmentTypeId
            LEFT JOIN LastMaint LM 
                ON LM.EquipmentID = E.EquipmentId
            WHERE 
                S.ScanTimeFinish BETWEEN 
                    COALESCE(LM.LastDateStop, DATEFROMPARTS(E.ProductionYear,1,1))
                    AND GETDATE()
                AND Ph.IDPhase IN (102,103)
                AND E.EquipmentId = ?
        ),
        K AS (
            SELECT 
                A.EquipmentId,
                A.Equipment + ' [' + A.EquipmentType + ']' AS Equipment,
                A.EndOfLifeCycle,
                SUM(DISTINCT CASE WHEN A.IsPass = 0 THEN A.QtyBoards ELSE 0 END)
                + SUM(DISTINCT CASE WHEN A.IsPass = 1 THEN A.QtyBoards ELSE 0 END) AS QtyScan
            FROM (
                SELECT 
                    SB.ProductCode,
                    SB.PhaseName,
                    SB.IsPass,
                    COUNT(*) AS QtyBoards,
                    SB.Equipment,
                    SB.EquipmentType,
                    SB.EquipmentId,
                    SB.EndOfLifeCycle
                FROM ScanBase SB
                GROUP BY 
                    SB.ProductCode,
                    SB.PhaseName,
                    SB.IsPass,
                    SB.Equipment,
                    SB.EquipmentType,
                    SB.EquipmentId,
                    SB.EndOfLifeCycle
            ) A
            INNER JOIN Traceability_rs.eqp.CompitiManutenzione CM
                ON CM.EquipmentId = A.EquipmentId
            GROUP BY 
                A.Equipment,
                A.EquipmentType,
                A.EquipmentId,
                A.EndOfLifeCycle
        ),
        P AS (
            SELECT DISTINCT 
                e.EquipmentId,
                p.NoCycle
            FROM Traceability_RS.eqp.CompitiManutenzione C
            INNER JOIN eqp.Equipments e 
                ON e.EquipmentId = C.EquipmentId
            INNER JOIN eqp.EquipmentTypes t 
                ON t.EquipmentTypeId = e.EquipmentTypeId 
            INNER JOIN eqp.ProgrammedInterventions p 
                ON p.ProgrammedInterventionId = C.ProgrammedInterventionId
            WHERE t.istest = 1
        )
        SELECT *
        FROM (
            SELECT 
                (SELECT DISTINCT p1.ProgrammedInterventionId
                 FROM [Traceability_RS].[eqp].[CompitiManutenzione] C
                 INNER JOIN eqp.Equipments e ON e.EquipmentId=C.EquipmentId
                 INNER JOIN eqp.EquipmentTypes t ON t.EquipmentTypeId=e.EquipmentTypeId 
                 INNER JOIN eqp.ProgrammedInterventions p1 ON p1.ProgrammedInterventionId=C.ProgrammedInterventionId
                 WHERE t.istest=1 AND e.EquipmentId=k.EquipmentId AND p1.NoCycle=p.NoCycle
                ) AS ProgrammedInterventionId,
                k.equipmentid,
                k.Equipment,
                k.QtyScan,
                k.QtyScan - k.EndOfLifeCycle AS CylesOverEquipmentLife,
                p.NoCycle,
                CASE WHEN k.QtyScan >= 0.95 * p.NoCycle      THEN 1 ELSE 0 END AS IsAt95PercentTask,
                k.EndOfLifeCycle,
                CASE WHEN k.QtyScan >= 0.95 * k.EndOfLifeCycle THEN 1 ELSE 0 END AS IsAt95ProcentEndOfLIfe,
                CASE WHEN k.QtyScan >= k.EndOfLifeCycle        THEN 1 ELSE 0 END AS IsEndOfLife
            FROM K
            LEFT JOIN P
                ON P.EquipmentId = K.EquipmentId
        ) G
        WHERE 
            G.IsAt95PercentTask = 1 
            OR G.IsAt95ProcentEndOfLIfe = 1
            OR G.IsEndOfLife = 1
        ORDER BY 
            G.QtyScan DESC;
        """
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (equipment_id,))
            rows = cursor.fetchall()
            cursor.close()
            return rows if rows else []
        except Exception as e:
            logger.error(f"Error fetching equipment cycle status: {e}", exc_info=True)
            return []

    def is_test_equipment(self, equipment_id):
        """
        Verifica se l'equipaggiamento è di tipo test (IsTest=1).
        
        Args:
            equipment_id: ID dell'equipaggiamento da verificare
            
        Returns:
            bool: True se l'equipaggiamento è di tipo test, False altrimenti
        """
        query = """
        SELECT t.IsTest
        FROM [Traceability_RS].[eqp].[Equipments] e
        INNER JOIN [Traceability_RS].[eqp].[EquipmentTypes] t 
            ON t.EquipmentTypeId = e.EquipmentTypeId
        WHERE e.EquipmentId = ?
        """
        try:
            result = self.fetch_one(query, (equipment_id,))
            return bool(result[0]) if result else False
        except Exception as e:
            logger.error(f"Error checking if equipment is test type: {e}", exc_info=True)
            return False


class LoginWindow(tk.Toplevel):
    """Finestra per raccogliere le credenziali dell'utente."""

    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler  # Manteniamo db e lang per i testi tradotti
        self.lang = lang_manager

        # Attributi per restituire i risultati
        self.user_id = None
        self.password = None
        self.clicked_login = False

        self.protocol("WM_DELETE_WINDOW", self._on_cancel)
        self.transient(master)
        self.grab_set()

        self._create_widgets()
        self.update_texts()
        self.bind('<Return>', self._attempt_login_event)

    def _create_widgets(self):
        # ... questo metodo rimane ESATTAMENTE invariato ...
        self.geometry("350x200")
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        self.user_id_label = ttk.Label(frame)
        self.user_id_label.grid(row=0, column=0, padx=5, pady=10, sticky="w")
        self.user_id_entry = ttk.Entry(frame, width=30)
        self.user_id_entry.grid(row=0, column=1, padx=5, pady=10)
        self.password_label = ttk.Label(frame)
        self.password_label.grid(row=1, column=0, padx=5, pady=10, sticky="w")
        self.password_entry = ttk.Entry(frame, show="*", width=30)
        self.password_entry.grid(row=1, column=1, padx=5, pady=10)
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=(15, 0))
        self.login_button = ttk.Button(button_frame, command=self._attempt_login)
        self.login_button.pack(side=tk.LEFT, padx=10)
        self.cancel_button = ttk.Button(button_frame, command=self._on_cancel)
        self.cancel_button.pack(side=tk.LEFT, padx=10)
        self.user_id_entry.focus_set()

    def update_texts(self):
        # ... questo metodo rimane ESATTAMENTE invariato ...
        self.title(self.lang.get('login_title'))
        self.user_id_label.config(text=self.lang.get('login_user_id'))
        self.password_label.config(text=self.lang.get('login_password'))
        self.login_button.config(text=self.lang.get('login_button'))
        self.cancel_button.config(text=self.lang.get('login_cancel_button'))

    def _attempt_login_event(self, event=None):
        self._attempt_login()

    def _attempt_login(self):
        # --- LOGICA MODIFICATA ---
        # Ora la finestra si limita a raccogliere i dati
        user_id = self.user_id_entry.get()
        password = self.password_entry.get()
        if not user_id or not password:
            messagebox.showerror(self.lang.get('login_title'), self.lang.get('login_error_credentials'), parent=self)
            return

        self.user_id = user_id
        self.password = password
        self.clicked_login = True
        self.destroy()  # Chiude la finestra

    def _on_cancel(self):
        # L'utente ha chiuso la finestra senza premere login
        self.clicked_login = False
        self.destroy()

class InsertDocumentForm(tk.Toplevel):
    """Finestra di inserimento documenti (di Produzione)."""

    def __init__(self, master, db_handler, user_name, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.user_name = user_name
        self.lang = lang_manager
        self.master_window = master

        self.products_data = {}
        self.all_product_names = []
        self.parent_phases_data = {}

        self.product_var = tk.StringVar()
        self.parent_phase_var = tk.StringVar()
        self.file_name_var = tk.StringVar()
        self.document_date_var = tk.StringVar()
        self.revision_var = tk.StringVar()
        self.validated_var = tk.BooleanVar()

        self.validated_var.trace('w', self._on_validated_check_changed)

        self._create_widgets()
        self.update_texts()
        self._load_products()

    def _create_widgets(self):
        self.geometry("650x750")  # Altezza aumentata
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        self.product_label = ttk.Label(frame, font=("Helvetica", 10, "bold"))
        self.product_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        self.product_combo = ttk.Combobox(frame, textvariable=self.product_var)
        self.product_combo.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(0, 15))
        self.product_combo.bind("<<ComboboxSelected>>", self._on_product_select)
        self.product_combo.bind("<KeyRelease>", self._on_product_keyrelease)

        self.phase_label = ttk.Label(frame, font=("Helvetica", 10, "bold"))
        self.phase_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        self.parent_phase_combo = ttk.Combobox(frame, textvariable=self.parent_phase_var, state="disabled")
        self.parent_phase_combo.grid(row=3, column=0, columnspan=2, sticky=tk.EW, pady=(0, 15))
        self.parent_phase_combo.bind("<<ComboboxSelected>>", self._on_phase_select)

        self.details_frame = ttk.LabelFrame(frame, padding="10")
        self.details_frame.grid(row=4, column=0, columnspan=2, sticky=tk.EW, pady=(0, 15))
        self.details_frame.columnconfigure(1, weight=1)

        self.file_name_label = ttk.Label(self.details_frame)
        self.file_name_label.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.file_entry = ttk.Entry(self.details_frame, textvariable=self.file_name_var, state="disabled")
        self.file_entry.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=5)
        self.browse_button = ttk.Button(self.details_frame, command=self._browse_file, state="disabled")
        self.browse_button.grid(row=0, column=2, padx=(0, 5), pady=5)

        self.revision_label = ttk.Label(self.details_frame)
        self.revision_label.grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.revision_entry = ttk.Entry(self.details_frame, textvariable=self.revision_var, state="disabled")
        self.revision_entry.grid(row=1, column=1, columnspan=2, sticky=tk.EW, padx=5, pady=5)

        # NUOVO: Campo Data Documento
        self.document_date_label = ttk.Label(self.details_frame)
        self.document_date_label.grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.document_date_entry = ttk.Entry(self.details_frame, textvariable=self.document_date_var, state="disabled")
        self.document_date_entry.grid(row=2, column=1, columnspan=2, sticky=tk.EW, padx=5, pady=5)

        self.validated_check = ttk.Checkbutton(self.details_frame, variable=self.validated_var, state="disabled")
        self.validated_check.grid(row=3, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)

        self.docs_frame = ttk.LabelFrame(frame, padding="10")
        self.docs_frame.grid(row=5, column=0, columnspan=2, sticky=tk.EW, pady=(0, 15))
        self.docs_listbox = tk.Listbox(self.docs_frame, height=6, selectbackground="#a6a6a6")
        self.docs_listbox.pack(fill=tk.BOTH, expand=True)

        self.save_button = ttk.Button(frame, command=self._save_document, state="disabled")
        self.save_button.grid(row=6, column=1, sticky=tk.E, pady=10)

    def update_texts(self):
        """Aggiorna i testi della UI."""
        self.title(self.lang.get('insert_doc_title'))
        self.product_label.config(text=self.lang.get('label_select_product'))
        self.phase_label.config(text=self.lang.get('label_select_phase'))
        self.details_frame.config(text=self.lang.get('frame_new_doc_details'))
        self.file_name_label.config(text=self.lang.get('label_file_name'))
        self.browse_button.config(text=self.lang.get('button_browse'))
        self.revision_label.config(text=self.lang.get('label_revision'))
        self.document_date_label.config(text=self.lang.get('label_document_date'))
        self.validated_check.config(text=self.lang.get('check_validated'))
        self.docs_frame.config(text=self.lang.get('frame_active_docs'))
        self.save_button.config(text=self.lang.get('button_save'))
        self._refresh_document_list()

    def _on_product_keyrelease(self, event):
        """Filtra il combobox mentre l'utente digita"""
        typed_text = self.product_var.get()
        if not typed_text:
            self.product_combo['values'] = self.all_product_names
        else:
            filtered_list = [
                name for name in self.all_product_names
                if typed_text.lower() in name.lower()
            ]
            self.product_combo['values'] = filtered_list

    def _load_products(self):
        products = self.db.fetch_products()
        if products:
            self.products_data = {p.ProductCode: p.IDProduct for p in products}
            self.all_product_names = list(self.products_data.keys())
            self.product_combo['values'] = self.all_product_names
        else:
            messagebox.showwarning(self.lang.get('app_title'), self.lang.get('warn_no_products_found'))

    def _on_product_select(self, event=None):
        self._reset_phase_section()
        self._reset_details_section()
        product_id = self.products_data.get(self.product_var.get())

        if product_id:
            parent_phases = self.db.fetch_parent_phases(product_id)
            if parent_phases:
                self.parent_phases_data = {p.Phase: p.IDParentPhase for p in parent_phases}
                self.parent_phase_combo['values'] = list(self.parent_phases_data.keys())
                self.parent_phase_combo.config(state="readonly")
            else:
                messagebox.showerror(self.lang.get('app_title'), self.lang.get('error_no_phases_found'))
                self.product_combo.focus()

    def _on_phase_select(self, event=None):
        """Popola la lista dei documenti e abilita l'inserimento"""
        self.docs_listbox.delete(0, tk.END)
        self.documents_in_phase = []

        # Abilita i campi di input per un nuovo documento
        self.file_entry.config(state="readonly")
        self.browse_button.config(state="normal")
        self.revision_entry.config(state="normal")
        self.document_date_entry.config(state="normal")
        self.validated_check.config(state="normal")
        self.save_button.config(state="normal")

        product_id = self.products_data.get(self.product_var.get())
        parent_phase_id = self.parent_phases_data.get(self.parent_phase_var.get())

        if not (product_id and parent_phase_id):
            return

        # Recupera i documenti esistenti per questa fase
        existing_docs = self.db.fetch_existing_documents(product_id, parent_phase_id)

        if existing_docs:
            # Se ci sono documenti, mostrarli nella listbox
            self.documents_in_phase = existing_docs
            yes_text = self.lang.get('text_yes')
            no_text = self.lang.get('text_no')
            for i, doc in enumerate(existing_docs):
                is_valid_text = yes_text if doc.Validated else no_text
                display_text = f"{doc.DocumentName} (Rev: {doc.DocumentRevisionNumber}) - Validato: {is_valid_text}"
                self.docs_listbox.insert(tk.END, display_text)
                if doc.Validated:
                    self.docs_listbox.itemconfig(i, {'bg': '#c8e6c9'})
        else:
            # Se non ci sono documenti, mostra un messaggio informativo
            self.docs_listbox.insert(tk.END, self.lang.get('info_no_existing_docs',
                                                           "Nessun documento esistente. Inserisci un nuovo documento."))

    def _open_selected_document(self, event=None):
        """Apre il documento selezionato (solo se validato)"""
        selected_indices = self.docs_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning(
                self.lang.get('app_title'),
                self.lang.get('warn_no_document_selected', "Selezionare un documento da aprire"),
                parent=self
            )
            return

        selected_index = selected_indices[0]
        selected_doc = self.documents_in_phase[selected_index]

        # NUOVO: Controlla se il documento Ã¨ validato
        if not selected_doc.Validated or selected_doc.ApprovatoDa is None:
            messagebox.showerror(
                self.lang.get('app_title'),
                self.lang.get('error_doc_not_validated', "Il documento prescelto non Ã¨ ancora validato"),
                parent=self
            )
            return

        success = self.db.fetch_and_open_document(selected_doc.DocumentProductionID)
        if not success:
            messagebox.showerror(
                self.lang.get('error_title', "Errore"),
                "Impossibile aprire il documento.",
                parent=self)

    def _browse_file(self, event=None):
        file_path = filedialog.askopenfilename(title=self.lang.get('insert_doc_title'),
                                               filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if file_path:
            self.file_name_var.set(file_path)

    def _on_validated_check_changed(self, *args):
        """Gestisce il cambio dello stato del checkbox Validato"""
        if self.validated_var.get():
            # Annulla il cambio temporaneamente
            self.validated_var.set(False)

            # Verifica che master_window sia disponibile
            if not hasattr(self.master_window, '_execute_authorized_action'):
                messagebox.showerror(
                    self.lang.get('app_title'),
                    "Errore: finestra principale non disponibile"
                )
                return

            # Chiama la procedura di autorizzazione
            if self.master_window._execute_authorized_action(
                    'validatore_documenti',
                    self._apply_validation
            ):
                self.validated_var.set(True)
                messagebox.showinfo(
                    self.lang.get('app_title'),
                    self.lang.get('info_document_validated')
                )
            else:
                self.validated_var.set(False)

    def _apply_validation(self):
        """NUOVO: Callback da eseguire dopo autorizzazione riuscita"""
        # Questo metodo sarÃ  chiamato solo se l'utente Ã¨ autorizzato
        pass

    def _save_document(self):
        """Validazione aggiornata con data documento e gestione sostituzione"""
        # Validazione input
        required_fields = [
            self.product_var.get(),
            self.parent_phase_var.get(),
            self.file_name_var.get(),
            self.revision_var.get(),
            self.document_date_var.get()
        ]

        if not all(required_fields):
            messagebox.showerror(self.lang.get('app_title'), self.lang.get('error_input_all_fields'))
            return

        # Validazione formato data
        try:
            from datetime import datetime
            document_date = datetime.strptime(self.document_date_var.get(), "%d/%m/%Y")
        except ValueError:
            messagebox.showerror(
                self.lang.get('app_title'),
                self.lang.get('error_invalid_date_format', "Formato data non valido. Utilizzare DD/MM/YYYY")
            )
            return

        # Validazione lunghezza revisione
        revision = self.revision_var.get()
        if len(revision) > 10:
            msg_template = self.lang.get_raw('error_input_revision_length')
            msg = msg_template.replace('{revision}', revision).replace('{length}', str(len(revision)))
            messagebox.showerror(self.lang.get('app_title'), msg)
            return

        # Recupero dati per il salvataggio
        product_id = self.products_data.get(self.product_var.get())
        parent_phase_id = self.parent_phases_data.get(self.parent_phase_var.get())
        local_file_path = self.file_name_var.get()
        doc_name = os.path.basename(local_file_path)
        is_validated_bool = self.validated_var.get()
        validated_as_int = 1 if is_validated_bool else 0

        # NUOVO: Preparare informazioni di validazione
        validator_info = None
        if is_validated_bool and hasattr(self.master_window, 'last_authenticated_user_name'):
            validator_info = self.master_window.last_authenticated_user_name

        # NUOVO: Controlla se esiste giÃ  un documento per questa fase
        existing_docs = self.db.fetch_existing_documents(product_id, parent_phase_id)

        if existing_docs:
            # Chiedi conferma per sostituire il documento esistente
            existing_doc_names = ", ".join([doc.DocumentName for doc in existing_docs])
            response = messagebox.askyesno(
                self.lang.get('app_title'),
                self.lang.get(
                    'confirm_replace_document',
                    f"Esiste giÃ  un documento per questa fase: {existing_doc_names}.\nDesideri sostituirlo?"
                ),
                parent=self
            )

            if response:
                # Marca i documenti esistenti come "non piÃ¹ validi"
                for doc in existing_docs:
                    self.db.mark_document_out_of_validation(doc.DocumentProductionID)
            else:
                # Utente ha annullato
                return

        success = self.db.save_document_to_db(
            product_id,
            parent_phase_id,
            doc_name,
            local_file_path,
            revision,
            self.user_name,
            validated_as_int,
            document_date,
            validator_info
        )

        if success:
            messagebox.showinfo(self.lang.get('app_title'), self.lang.get('info_save_success'))
            self._reset_input_fields()
            self._refresh_document_list()
        else:
            msg_template = self.lang.get_raw('error_save_failed')
            msg = msg_template.replace('{e}', self.db.last_error_details)
            messagebox.showerror(self.lang.get('app_title'), msg)

    def _refresh_document_list(self):
        self.docs_listbox.delete(0, tk.END)
        product_id = self.products_data.get(self.product_var.get())
        parent_phase_id = self.parent_phases_data.get(self.parent_phase_var.get())

        if not (product_id and parent_phase_id):
            return

        existing_docs = self.db.fetch_existing_documents(product_id, parent_phase_id)
        yes_text = self.lang.get('text_yes')
        no_text = self.lang.get('text_no')
        for i, doc in enumerate(existing_docs):
            is_valid_text = yes_text if doc.Validated else no_text
            display_text = f"File: {doc.DocumentName} | Rev: {doc.DocumentRevisionNumber} | Validato: {is_valid_text}"
            self.docs_listbox.insert(tk.END, display_text)
            if doc.Validated:
                self.docs_listbox.itemconfig(i, {'bg': '#c8e6c9'})

    def _reset_phase_section(self):
        self.parent_phase_var.set("")
        self.parent_phase_combo.config(state="disabled", values=[])


    def _reset_details_section(self):
        self._reset_input_fields()
        self.file_entry.config(state="disabled")
        self.browse_button.config(state="disabled")
        self.revision_entry.config(state="disabled")
        self.document_date_entry.config(state="disabled")  # NUOVO
        self.validated_check.config(state="disabled")
        self.save_button.config(state="disabled")
        self.docs_listbox.delete(0, tk.END)

    def _reset_input_fields(self):
        self.file_name_var.set("")
        self.revision_var.set("")
        self.document_date_var.set("")  # NUOVO
        self.validated_var.set(False)
        self.file_entry.config(state="readonly")

class ViewDocumentForm(tk.Toplevel):
    """Finestra per visualizzare un documento (di Produzione)."""

    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager
        self.master_window = master  # NUOVO

        self.transient(master)
        self.grab_set()

        self.products_data = {}
        self.all_product_names = []
        self.parent_phases_data = {}
        self.documents_in_phase = []

        self.product_var = tk.StringVar()
        self.parent_phase_var = tk.StringVar()

        self._create_widgets()
        self.update_texts()
        self._load_products()

    def _create_widgets(self):
        self.geometry("700x450")
        frame = ttk.Frame(self, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=1)

        self.product_label = ttk.Label(frame, font=("Helvetica", 10, "bold"))
        self.product_label.pack(fill=tk.X, pady=(0, 5))
        self.product_combo = ttk.Combobox(frame, textvariable=self.product_var, width=50)
        self.product_combo.pack(fill=tk.X, pady=(0, 15))
        self.product_combo.bind("<<ComboboxSelected>>", self._on_product_select)
        self.product_combo.bind("<KeyRelease>", self._on_product_keyrelease)

        self.phase_label = ttk.Label(frame, font=("Helvetica", 10, "bold"))
        self.phase_label.pack(fill=tk.X, pady=(0, 5))
        self.parent_phase_combo = ttk.Combobox(
            frame,
            textvariable=self.parent_phase_var,
            state="disabled",
            width=50
        )
        self.parent_phase_combo.pack(fill=tk.X, pady=(0, 15))
        self.parent_phase_combo.bind("<<ComboboxSelected>>", self._on_phase_select)

        self.docs_listbox = tk.Listbox(frame, height=8)
        self.docs_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        self.docs_listbox.bind("<Double-1>", self._open_selected_document)

        # Frame per i pulsanti
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))

        # ASSICURATI CHE QUESTI TRE PULSANTI SIANO CREATI
        self.open_button = ttk.Button(
            button_frame,
            text="Apri",  # Testo temporaneo
            command=self._open_selected_document
        )
        self.open_button.pack(side="left", padx=(0, 5))

        self.validate_button = ttk.Button(
            button_frame,
            text="Valida",  # Testo temporaneo
            command=self._validate_selected_document
        )
        self.validate_button.pack(side="left", padx=(0, 5))

        self.close_button = ttk.Button(
            button_frame,
            text="Chiudi",  # Testo temporaneo
            command=self.destroy
        )
        self.close_button.pack(side="right")

    def update_texts(self):
        """Aggiorna i testi della UI."""
        self.title(self.lang.get('view_doc_title'))
        self.product_label.config(text=self.lang.get('label_select_product'))
        self.phase_label.config(text=self.lang.get('label_select_phase'))
        self.open_button.config(text=self.lang.get('button_open'))  # NUOVO
        self.validate_button.config(text=self.lang.get('button_validate_doc'))  # NUOVO
        self.close_button.config(text=self.lang.get('button_close'))

    def _load_products(self):
        products = self.db.fetch_products_with_documents()
        if products:
            self.products_data = {p.ProductCode: p.IDProduct for p in products}
            self.all_product_names = list(self.products_data.keys())
            self.product_combo['values'] = self.all_product_names
        else:
            messagebox.showwarning(self.lang.get('app_title'), self.lang.get('warn_no_products_found'), parent=self)

    def _on_product_keyrelease(self, event):
        typed_text = self.product_var.get()
        if not typed_text:
            self.product_combo['values'] = self.all_product_names
        else:
            filtered_list = [name for name in self.all_product_names if typed_text.lower() in name.lower()]
            self.product_combo['values'] = filtered_list

    def _on_product_select(self, event=None):
        self.parent_phase_var.set("")
        self.parent_phase_combo.config(state="disabled", values=[])
        self.docs_listbox.delete(0, tk.END)
        self.documents_in_phase = []

        product_id = self.products_data.get(self.product_var.get())
        if product_id:
            parent_phases = self.db.fetch_phases_with_documents_for_product(product_id)
            if parent_phases:
                self.parent_phases_data = {p.ParentPhaseName: p.IDParentPhase for p in parent_phases}
                self.parent_phase_combo.config(state="readonly", values=list(self.parent_phases_data.keys()))
            else:
                messagebox.showwarning(
                    self.lang.get('app_title'),
                    self.lang.get('warn_no_document_found_for_product',
                                  "Nessun documento trovato per il prodotto selezionato."),
                    parent=self
                )

    def _on_phase_select(self, event=None):
        """Popola la lista dei documenti (tutte le revisioni)."""
        self.docs_listbox.delete(0, tk.END)
        self.documents_in_phase = []

        product_id = self.products_data.get(self.product_var.get())
        parent_phase_id = self.parent_phases_data.get(self.parent_phase_var.get())

        if not (product_id and parent_phase_id):
            return

        self.documents_in_phase = self.db.fetch_existing_documents(product_id, parent_phase_id)

        if not self.documents_in_phase:
            messagebox.showwarning(self.lang.get('app_title'), self.lang.get('warn_no_document_found'), parent=self)
        else:
            for i, doc in enumerate(self.documents_in_phase):
                # Determina se il documento Ã¨ valido
                is_valid = doc.Validated == 1 and doc.DateOutOfValidation is None
                
                # Formatta il testo di visualizzazione
                approver_info = f" - Approvato da: {doc.ApprovatoDa}" if doc.ApprovatoDa else ""
                status_text = "âœ“ VALIDO" if is_valid else "âœ— Non valido"
                
                display_text = f"{doc.DocumentName} (Rev: {doc.DocumentRevisionNumber}){approver_info} [{status_text}]"
                self.docs_listbox.insert(tk.END, display_text)
                
                # Colora in base alla validitÃ 
                if is_valid:
                    self.docs_listbox.itemconfig(i, {'bg': '#c8e6c9'})  # Verde chiaro per validi
                else:
                    self.docs_listbox.itemconfig(i, {'bg': '#e0e0e0', 'fg': '#757575'})  # Grigio per non validi

    def _open_selected_document(self, event=None):
        """Apre il documento selezionato (solo se valido)"""
        selected_indices = self.docs_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning(
                self.lang.get('app_title'),
                self.lang.get('warn_no_document_selected', "Selezionare un documento da aprire"),
                parent=self
            )
            return

        selected_index = selected_indices[0]
        selected_doc = self.documents_in_phase[selected_index]

        # NUOVO: Verifica che il documento sia valido prima di aprirlo
        is_valid = selected_doc.Validated == 1 and selected_doc.DateOutOfValidation is None
        
        if not is_valid:
            messagebox.showwarning(
                self.lang.get('app_title'),
                self.lang.get('warn_cannot_open_invalid_doc', 
                             "Impossibile aprire questo documento. Solo i documenti validi possono essere aperti."),
                parent=self
            )
            return

        success = self.db.fetch_and_open_document(selected_doc.DocumentProductionID)
        if not success:
            messagebox.showerror(
                self.lang.get('error_title', "Errore"),
                "Impossibile aprire il documento.",
                parent=self
            )

    def _validate_selected_document(self):
        """NUOVO: Valida il documento selezionato con autorizzazione"""
        selected_indices = self.docs_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning(
                self.lang.get('app_title'),
                self.lang.get('warn_no_document_selected', "Selezionare un documento da validare"),
                parent=self
            )
            return

        selected_index = selected_indices[0]
        selected_doc = self.documents_in_phase[selected_index]

        # Se giÃ  validato, chiedi conferma
        if selected_doc.Validated:
            response = messagebox.askyesno(
                self.lang.get('app_title'),
                self.lang.get('confirm_revalidate_document',
                              "Il documento Ã¨ giÃ  validato. Desideri convalidarlo di nuovo?"),
                parent=self
            )
            if not response:
                return

        # Chiama la procedura di autorizzazione
        if self.master_window._execute_authorized_action(
                'validatore_documenti',
                lambda: self._apply_document_validation(selected_doc)
        ):
            messagebox.showinfo(
                self.lang.get('app_title'),
                self.lang.get('info_document_validated'),
                parent=self
            )
            self._on_phase_select()  # Aggiorna la lista
        else:
            messagebox.showerror(
                self.lang.get('app_title'),
                self.lang.get('error_validation_failed', "Validazione non riuscita"),
                parent=self
            )

    def _apply_document_validation(self, doc):
        """NUOVO: Callback per applicare la validazione dopo autorizzazione"""
        validator_name = getattr(self.master_window, 'last_authenticated_user_name', None)

        success = self.db.update_document_validation(
            doc.DocumentProductionID,
            validator_name
        )

        if not success:
            raise Exception("Errore durante l'aggiornamento del documento nel database")

class KanbanLocationCreateForm(tk.Toplevel):
    """
    Crea una locazione KanBan:
    - Combo aree (da dbo.ParentPhases, CodCDC in 10,30,90)
    - Campo locazione (max 8 char)
    - Bottone salva
    - Logo bottom-right
    """
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager

        self.title(self.lang.get('kanban_locations_title', "Crea locazione KanBan"))
        self.geometry("500x260")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # State
        self.area_map = {}  # {display_name: id}
        self.area_var = tk.StringVar()
        self.location_var = tk.StringVar()

        self._build_ui()
        self._load_areas()
        self._update_save_state()

    def _build_ui(self):
        container = ttk.Frame(self, padding=16)
        container.pack(fill=tk.BOTH, expand=True)

        # Area label + combo
        ttk.Label(container, text=self.lang.get('kanban_area_label', "Area KanBan"),
                  font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.area_combo = ttk.Combobox(container, textvariable=self.area_var, state="readonly", width=44)
        self.area_combo.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 12))
        self.area_combo.bind("<<ComboboxSelected>>", lambda e: self._update_save_state())

        # Location label + entry
        ttk.Label(container, text=self.lang.get('kanban_location_label', "Locazione"),
                  font=("Helvetica", 10, "bold")).grid(row=2, column=0, sticky="w")
        self.location_entry = ttk.Entry(container, textvariable=self.location_var, width=30)
        self.location_entry.grid(row=3, column=0, sticky="w", pady=(2, 12))
        self.location_entry.bind("<KeyRelease>", self._on_location_change)

        # Buttons (Save + Close)
        btns = ttk.Frame(container)
        btns.grid(row=4, column=0, columnspan=2, sticky="e", pady=(10, 0))

        self.save_btn = ttk.Button(btns, text=self.lang.get('button_save', "Salva"), command=self._on_save)
        self.save_btn.pack(side="right", padx=(6, 0))

        close_btn = ttk.Button(btns, text=self.lang.get('button_close', "Chiudi"), command=self.destroy)
        close_btn.pack(side="right")

        # Bottom bar: left -> total locations, right -> logo
        bottom = ttk.Frame(container)
        bottom.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        bottom.columnconfigure(0, weight=1)  # sinistra si espande

        # Totale locazioni (sinistra)
        self.total_locations_label = ttk.Label(bottom, text="")
        self.total_locations_label.grid(row=0, column=0, sticky="w")

        # Logo (destra)
        if PIL_AVAILABLE:
            try:
                from PIL import Image, ImageTk
                img = Image.open("logo.png")
                img.thumbnail((100, 100))
                self._logo_img = ImageTk.PhotoImage(img)
                ttk.Label(bottom, image=self._logo_img).grid(row=0, column=1, sticky="e")
            except Exception as e:
                print(f"Errore caricamento logo: {e}")

        # Grid weights
        container.columnconfigure(0, weight=1)

        # Carica il totale
        self._load_locations_count()

    def _load_locations_count(self):
        n = self.db.count_kanban_locations()
        n_str = str(n) if n is not None else "#N/A"
        # usa template traducibile con {n}
        template = self.lang.get('kanban_locations_total', "Totale locazioni: {n}")
        self.total_locations_label.config(text=template.replace("{n}", n_str))

    def _load_areas(self):
        rows = self.db.fetch_kanban_areas()
        if not rows:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('kanban_areas_load_error', "Impossibile caricare le aree KanBan."),
                                 parent=self)
            self.area_combo.config(state="disabled")
            self.save_btn.config(state="disabled")
            return

        # Mappa display -> id
        self.area_map = {row.ParentPhaseName: row.IDParentPhase for row in rows}
        self.area_combo['values'] = list(self.area_map.keys())
        if self.area_combo['values']:
            self.area_combo.current(0)

    def _on_location_change(self, event=None):
        # Uppercase e max 8 caratteri
        val = self.location_var.get().upper()
        if len(val) > 8:
            val = val[:8]
        if val != self.location_var.get():
            self.location_var.set(val)
            # riposiziona il cursore alla fine
            self.location_entry.icursor(tk.END)
        self._update_save_state()

    def _update_save_state(self):
        has_area = bool(self.area_var.get())
        loc = self.location_var.get().strip()
        enable = has_area and 1 <= len(loc) <= 8
        self.save_btn.config(state=("normal" if enable else "disabled"))

    def _on_save(self):
        # Validazioni
        area_name = self.area_var.get()
        if not area_name:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_area_required', "Selezionare un'area."), parent=self)
            return
        area_id = self.area_map.get(area_name)
        print(area_id)
        loc = self.location_var.get().strip().upper()

        if not loc:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_location_required', "Inserire una locazione."), parent=self)
            self.location_entry.focus_set()
            return
        if len(loc) > 8:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_location_len', "La locazione puÃ² avere al massimo 8 caratteri."),
                                   parent=self)
            self.location_entry.focus_set()
            return

        ok, err = self.db.insert_kanban_location(area_id, loc)
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"),
                                self.lang.get('kanban_location_saved', "Locazione creata con successo."),
                                parent=self)
            self.location_var.set("")
            self.location_entry.focus_set()
            self._update_save_state()
            self._load_locations_count()  # <-- aggiorna conteggio
            return

        if err == "duplicate":
            self.location_var.set("")
            self.location_entry.focus_set()
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_location_exists',
                                                 "La locazione esiste giÃ  per l'area selezionata. Inserire un valore diverso."),
                                   parent=self)
            return

        messagebox.showerror(self.lang.get('error_title', "Errore"),
                             self.lang.get('kanban_location_save_error',
                                           f"Errore durante il salvataggio della locazione: {err}"),
                             parent=self)

class KanbanLocationLabelsForm(tk.Toplevel):
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager
        self.title(self.lang.get('kanban_labels_title', "Stampa etichette locazioni"))
        self.geometry("520x260")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self.cfg = load_printer_config()
        if not self.cfg:
            # chiedi subito la configurazione
            dlg = PrinterSetupDialog(self, self.lang)
            self.wait_window(dlg)
            if not getattr(dlg, "result", None):
                messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                       self.lang.get('printer_not_configured', "Stampante non configurata."), parent=self)
                self.destroy()
                return
            self.cfg = dlg.result

        self.location_var = tk.StringVar()
        self.copies_var = tk.StringVar(value="1")

        self._build_ui()
        self._load_locations()

    def _build_ui(self):
        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)

        # Riga configurazione stampante + bottone setup
        top = ttk.Frame(frm)
        top.grid(row=0, column=0, columnspan=2, sticky="ew")
        ttk.Label(top, text=self.lang.get('printer_current', "Stampante:")).pack(side="left")
        self.prn_label = ttk.Label(top, text=self._printer_summary(), foreground="#006400")
        self.prn_label.pack(side="left", padx=(6, 0))
        ttk.Button(top, text=self.lang.get('printer_setup_btn', "Imposta..."), command=self._setup_printer).pack(side="right")

        # Selezione locazione
        ttk.Label(frm, text=self.lang.get('kanban_location_label', "Locazione"), font=("Helvetica", 10, "bold")).grid(row=1, column=0, sticky="w", pady=(12,2))
        self.location_combo = ttk.Combobox(frm, textvariable=self.location_var, state="readonly", width=40)
        self.location_combo.grid(row=2, column=0, sticky="w")

        # Copie
        ttk.Label(frm, text=self.lang.get('copies_label', "Copie")).grid(row=1, column=1, sticky="w", pady=(12,2))
        self.copies_entry = ttk.Spinbox(frm, from_=1, to=99, textvariable=self.copies_var, width=5)
        self.copies_entry.grid(row=2, column=1, sticky="w")

        # Pulsanti
        btns = ttk.Frame(frm)
        btns.grid(row=10, column=0, columnspan=2, sticky="e", pady=(16,0))
        ttk.Button(btns, text=self.lang.get('button_close', "Chiudi"), command=self.destroy).pack(side="right")
        self.print_btn = ttk.Button(btns, text=self.lang.get('button_print', "Stampa"), command=self._on_print)
        self.print_btn.pack(side="right", padx=(0,8))

        frm.columnconfigure(0, weight=1)

    def _printer_summary(self) -> str:
        c = self.cfg or {}
        return f"{c.get('name','?')} ({c.get('ip','?')}:{c.get('port','?')}, {c.get('dpi','203')} dpi)"

    def _setup_printer(self):
        dlg = PrinterSetupDialog(self, self.lang)
        self.wait_window(dlg)
        if getattr(dlg, "result", None):
            self.cfg = dlg.result
            self.prn_label.config(text=self._printer_summary())

    def _load_locations(self):
        rows = self.db.fetch_kanban_locations_all()
        if not rows:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('kanban_locations_load_error', "Impossibile caricare le locazioni."),
                                 parent=self)
            self.print_btn.config(state="disabled")
            return
        values = [r.LocationCode for r in rows]
        self.location_combo['values'] = values
        if values:
            self.location_combo.current(0)

    def _on_print(self):
        loc = self.location_var.get().strip().upper()
        if not loc:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_location_required', "Inserire una locazione."), parent=self)
            return

        # ConformitÃ  al limite 8 char se desideri mantenerlo anche in stampa
        if len(loc) > 8:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('kanban_location_len', "La locazione puÃ² avere al massimo 8 caratteri."),
                                   parent=self)
            return

        try:
            copies = int(self.copies_var.get())
        except ValueError:
            copies = 1

        zpl = build_zpl_label(loc, copies, self.cfg)
        ok, err = send_raw_to_printer(self.cfg["ip"], int(self.cfg["port"]), zpl.encode("utf-8"))
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"),
                                self.lang.get('print_ok', "Stampa inviata alla stampante."), parent=self)
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('print_error', f"Errore di stampa: {err}"), parent=self)

def get_printer_config_dir() -> Path:
    base = Path(os.getenv("LOCALAPPDATA", ".")) / "TraceabilityRS"
    base.mkdir(parents=True, exist_ok=True)
    return base

def get_printer_config_path() -> Path:
    return get_printer_config_dir() / "printer_config.json"

def migrate_printer_config(cfg: dict) -> dict:
    # Versioning e default
    cfg = dict(cfg or {})
    cfg.setdefault("version", 1)
    cfg.setdefault("name", "")
    cfg.setdefault("ip", "")
    cfg.setdefault("port", 9100)
    cfg.setdefault("dpi", 203)
    cfg.setdefault("label_cm", [5, 5])   # [Larghezza, Altezza] in cm
    cfg.setdefault("text_pt", 12)
    cfg.setdefault("language", "ZPL")
    # NUOVO: scheduling refill Kanban
    cfg.setdefault("kanban_refill_enabled", True)
    cfg.setdefault("kanban_refill_check_minutes", 60)
    return cfg

def load_printer_config() -> dict | None:
    path = get_printer_config_path()
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return migrate_printer_config(raw)
    except Exception:
        # JSON corrotto o non leggibile
        return None

def save_printer_config(cfg: dict) -> tuple[bool, str | None]:
    """
    Scrive il JSON con backup e sostituzione atomica.
    Ritorna (True, None) se ok, (False, err) se errore.
    """
    try:
        cfg = migrate_printer_config(cfg)
        path = get_printer_config_path()
        dir_ = path.parent
        dir_.mkdir(parents=True, exist_ok=True)

        # Backup del file esistente
        if path.exists():
            bak = dir_ / f"printer_config.bak"
            shutil.copy2(path, bak)

        # Scrittura atomica: tmp -> replace
        with tempfile.NamedTemporaryFile("w", delete=False, dir=str(dir_), encoding="utf-8") as tf:
            json.dump(cfg, tf, ensure_ascii=False, indent=2)
            tmp_name = tf.name
        os.replace(tmp_name, path)
        return True, None
    except Exception as e:
        return False, str(e)

def validate_ip(ip: str) -> bool:
    try:
        socket.inet_aton(ip)
        return True
    except OSError:
        return False

def open_config_folder():
    # Apre la cartella del file JSON nel file manager
    path = get_printer_config_dir()
    try:
        os.startfile(str(path))  # Windows
    except AttributeError:
        import subprocess, sys
        if sys.platform.startswith("darwin"):
            subprocess.call(["open", str(path)])
        else:
            subprocess.call(["xdg-open", str(path)])

def pt_to_dots(pt: int, dpi: int) -> int:
    # 1 pt = 1/72 inch
    return max(10, round(dpi * (pt / 72.0)))

def cm_to_dots(cm: float, dpi: int) -> int:
    return round(dpi * (cm / 2.54))

def build_zpl_label(location_code: str, copies: int, cfg: dict) -> str:
    """
    Costruisce ZPL per etichetta 5x5 cm:
    - QR centrato
    - Testo sotto, corpo 12pt
    """
    dpi = int(cfg.get("dpi", 203))
    w_cm, h_cm = cfg.get("label_cm", [5, 5])
    width = cm_to_dots(w_cm, dpi)
    height = cm_to_dots(h_cm, dpi)
    text_h = pt_to_dots(int(cfg.get("text_pt", 12)), dpi)
    text_w = text_h  # quadrato

    margin = max(12, round(0.1 * dpi))  # ~2.5mm a 203dpi -> ~20 dot
    available_h = height - text_h - margin * 3
    if available_h < 50:
        available_h = max(50, height - text_h - margin)

    # Magnification QR (1..10). Considero QR v1 21x21 moduli.
    modules = 21
    max_qr_dots = min(width - 2 * margin, available_h)
    mag = max(1, min(10, max_qr_dots // modules))
    qr_size = modules * mag
    qr_x = (width - qr_size) // 2
    qr_y = margin

    text_y = qr_y + qr_size + margin

    # ZPL
    data = location_code.upper()
    zpl = []
    zpl.append("^XA")
    zpl.append("^CI28")                           # UTF-8 (anche se non serve per A-Z0-9)
    zpl.append(f"^PW{width}")                    # print width
    zpl.append(f"^LL{height}")                   # label length
    zpl.append("^LH0,0")

    # QR code centrato
    zpl.append(f"^FO{qr_x},{qr_y}")
    zpl.append(f"^BQN,2,{mag}")
    zpl.append(f"^FDLA,{data}^FS")

    # Testo centrato
    # Uso ^FB per centrare rispetto alla larghezza disponibile
    zpl.append(f"^FO0,{text_y}")
    zpl.append(f"^A0N,{text_h},{text_w}")
    zpl.append(f"^FB{width},1,0,C,0")
    zpl.append(f"^FD{data}^FS")

    # Numero copie
    copies = max(1, int(copies))
    zpl.append(f"^PQ{copies},0,1,Y")

    zpl.append("^XZ")
    return "".join(zpl)

def send_raw_to_printer(ip: str, port: int, payload: bytes, timeout: float = 5.0) -> tuple[bool, str | None]:
    try:
        with socket.create_connection((ip, port), timeout=timeout) as s:
            s.sendall(payload)
        return True, None
    except Exception as e:
        return False, str(e)

class KanbanLocationModifyForm(tk.Toplevel):
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager
        self.title(self.lang.get('kanban_modify_title', "Modifica locazioni"))
        self.geometry("720x420")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self.component_var = tk.StringVar()
        self.dest_loc_var = tk.StringVar()
        self.src_loc_var = tk.StringVar()
        self.dest_loc_var2 = tk.StringVar()

        self._locations_cache = []  # rows for combo

        self._build_ui()
        self._load_locations()

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=12)

        # Tab 1: Sposta componente
        tab1 = ttk.Frame(nb)
        nb.add(tab1, text=self.lang.get('tab_move_component', "Sposta componente"))

        frm1 = ttk.Frame(tab1, padding=12)
        frm1.pack(fill="both", expand=True)

        # Ricerca componente
        row = ttk.Frame(frm1)
        row.pack(fill="x")
        ttk.Label(row, text=self.lang.get('component_code', "Codice componente")).pack(side="left")
        ent = ttk.Entry(row, textvariable=self.component_var, width=30)
        ent.pack(side="left", padx=(8,8))
        ttk.Button(row, text=self.lang.get('button_search', "Cerca"), command=self._on_search_component).pack(side="left")

        # Risultati
        ttk.Label(frm1, text=self.lang.get('results_label', "Risultati")).pack(anchor="w", pady=(12,4))
        cols = ("record_id", "loc_code", "qty", "desc")
        self.tree = ttk.Treeview(frm1, columns=cols, show="headings", height=8)
        self.tree.heading("record_id", text="ID")
        self.tree.column("record_id", width=60, anchor="center")
        self.tree.heading("loc_code", text=self.lang.get('location', "Locazione"))
        self.tree.column("loc_code", width=140)
        self.tree.heading("qty", text=self.lang.get('quantity', "Q.tÃ "))
        self.tree.column("qty", width=80, anchor="e")
        self.tree.heading("desc", text=self.lang.get('description', "Descrizione"))
        self.tree.column("desc", width=360)
        self.tree.pack(fill="both", expand=True)

        # Destinazione
        dest_row = ttk.Frame(frm1)
        dest_row.pack(fill="x", pady=(12,0))
        ttk.Label(dest_row, text=self.lang.get('destination', "Destinazione")).pack(side="left")
        self.dest_combo = ttk.Combobox(dest_row, textvariable=self.dest_loc_var, state="readonly", width=48)
        self.dest_combo.pack(side="left", padx=(8,8))

        ttk.Button(frm1, text=self.lang.get('button_move', "Sposta"), command=self._on_move_single).pack(anchor="e", pady=(10,0))

        # Tab 2: Sposta intera locazione
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text=self.lang.get('tab_move_location', "Sposta locazione"))

        frm2 = ttk.Frame(tab2, padding=12)
        frm2.pack(fill="both", expand=True)

        r2a = ttk.Frame(frm2); r2a.pack(fill="x")
        ttk.Label(r2a, text=self.lang.get('source_location', "Sorgente")).pack(side="left")
        self.src_combo = ttk.Combobox(r2a, textvariable=self.src_loc_var, state="readonly", width=40)
        self.src_combo.pack(side="left", padx=(8,16))
        ttk.Label(r2a, text=self.lang.get('destination', "Destinazione")).pack(side="left")
        self.dest_combo2 = ttk.Combobox(r2a, textvariable=self.dest_loc_var2, state="readonly", width=40)
        self.dest_combo2.pack(side="left", padx=(8,0))

        ttk.Button(frm2, text=self.lang.get('button_move_all', "Sposta tutto"), command=self._on_move_all).pack(anchor="e", pady=(12,0))

    def _load_locations(self):
        rows = self.db.fetch_locations_for_combo()
        self._locations_cache = rows
        values = [f"{r.KanBanLocation} â€¢ {r.LocationCode}" for r in rows]
        self.dest_combo['values'] = values
        self.src_combo['values'] = values
        self.dest_combo2['values'] = values
        if values:
            self.dest_combo.current(0)
            self.src_combo.current(0)
            if len(values) > 1:
                self.dest_combo2.current(1)
            else:
                self.dest_combo2.current(0)

    def _on_search_component(self):
        code = self.component_var.get().strip()
        if not code:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('component_required', "Inserire un codice componente."), parent=self)
            return
        rows = self.db.search_component_open_records(code)
        # pulisci
        for i in self.tree.get_children():
            self.tree.delete(i)
        for r in rows:
            self.tree.insert("", "end", iid=str(r.KanBanRecordId),
                             values=(r.KanBanRecordId, r.LocationCode, r.Quantity, r.ComponentDescription or ""))

    def _pick_combo_ids(self, combo_value: str):
        """
        Estrae LocationId e KanBanLocationId dal valore scelto in combo (usando cache).
        """
        try:
            idx = list(self.dest_combo['values']).index(combo_value)
        except ValueError:
            # prova con altre combo
            vals = list(self.src_combo['values'])
            idx = vals.index(combo_value) if combo_value in vals else -1
        if idx < 0:
            return None, None
        row = self._locations_cache[idx]
        return row.LocationId, row.KanBanLocationId

    def _on_move_single(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('select_row', "Seleziona una riga dai risultati."), parent=self)
            return
        record_id = int(sel[0])
        dest_val = self.dest_loc_var.get()
        to_loc_id, to_kbl_id = self._pick_combo_ids(dest_val)
        if not to_loc_id:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('destination_required', "Seleziona la destinazione."), parent=self)
            return

        # Evita movimento verso la stessa locazione (se deducibile)
        # Recupero LocationId della riga selezionata leggendo i dati dal DB (piÃ¹ sicuro)
        rows = self.db.search_component_open_records(self.component_var.get().strip())
        src_loc_id = None
        for r in rows:
            if r.KanBanRecordId == record_id:
                src_loc_id = r.LocationId
                break
        if src_loc_id == to_loc_id:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('same_location', "Sorgente e destinazione coincidono."), parent=self)
            return

        ok, err = self.db.move_record_to_location(record_id, to_loc_id, to_kbl_id)
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('move_ok', "Spostamento completato."), parent=self)
            self._on_search_component()  # refresh risultati
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('move_err', f"Errore durante lo spostamento: {err or self.db.last_error_details}"),
                                 parent=self)

    def _on_move_all(self):
        src_val = self.src_loc_var.get()
        dest_val = self.dest_loc_var2.get()
        if not src_val or not dest_val:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('source_dest_required', "Seleziona sorgente e destinazione."), parent=self)
            return
        src_loc_id, _ = self._pick_combo_ids(src_val)
        to_loc_id, to_kbl_id = self._pick_combo_ids(dest_val)
        if not src_loc_id or not to_loc_id:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('source_dest_required', "Seleziona sorgente e destinazione."), parent=self)
            return
        if src_loc_id == to_loc_id:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('same_location', "Sorgente e destinazione coincidono."), parent=self)
            return

        ok, err = self.db.move_all_from_location(src_loc_id, to_loc_id, to_kbl_id)
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('move_ok', "Spostamento completato."), parent=self)
        else:
            msg = self.lang.get('nothing_to_move', "Nessun record da spostare.") if err == "nothing_to_move" else (err or self.db.last_error_details)
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('move_err', f"Errore durante lo spostamento: {msg}"),
                                 parent=self)

class KanbanMaterialsManagementForm(tk.Toplevel):
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager

        self.title(self.lang.get('materials_mgmt_title', "KanBan - Materiali: Gestione"))
        self.geometry("760x420")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # State
        self.type_var = tk.StringVar()
        self.comp_var = tk.StringVar()
        self.rule_var = tk.StringVar()

        self._types = []       # cache tipi
        self._components = []  # cache componenti
        self._rules = []       # cache regole
        self._comp_index_by_display = {}  # mappa display->row
        self._rule_index_by_display = {}  # mappa display->row

        self._logo_imgtk = None

        self._build_ui()
        self._load_types()
        self._load_rules()
        self._load_components(None)  # all

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # Filtri riga 1: Tipo componente
        fr1 = ttk.Frame(root)
        fr1.pack(fill="x")
        ttk.Label(fr1, text=self.lang.get('material_type', "Tipo componente")).pack(side="left")
        self.type_combo = ttk.Combobox(fr1, textvariable=self.type_var, width=40, state="readonly")
        self.type_combo.pack(side="left", padx=(8, 0))
        self.type_combo.bind("<<ComboboxSelected>>", self._on_type_changed)

        # Filtri riga 2: Componente
        fr2 = ttk.Frame(root)
        fr2.pack(fill="x", pady=(8, 0))
        ttk.Label(fr2, text=self.lang.get('component', "Componente")).pack(side="left")
        self.comp_combo = ttk.Combobox(fr2, textvariable=self.comp_var, width=60, state="readonly")
        self.comp_combo.pack(side="left", padx=(8, 8))
        self.comp_combo.bind("<<ComboboxSelected>>", self._on_component_changed)

        # Stato/regola corrente
        state = ttk.Labelframe(root, text=self.lang.get('current_rule_group', "Regola attiva"))
        state.pack(fill="x", pady=(12, 0))
        self.current_rule_label = ttk.Label(state, text=self.lang.get('no_rule', "Nessuna regola associata"))
        self.current_rule_label.pack(anchor="w", padx=8, pady=6)

        # Assegnazione regola
        assign = ttk.Labelframe(root, text=self.lang.get('assign_rule_group', "Associa regola"))
        assign.pack(fill="x", pady=(12, 0))
        row = ttk.Frame(assign)
        row.pack(fill="x", padx=8, pady=6)
        ttk.Label(row, text=self.lang.get('select_rule', "Regola")).pack(side="left")
        self.rule_combo = ttk.Combobox(row, textvariable=self.rule_var, state="readonly", width=50)
        self.rule_combo.pack(side="left", padx=(8, 8))
        ttk.Button(row, text=self.lang.get('button_assign', "Associa"), command=self._on_assign).pack(side="left")
        ttk.Button(row, text=self.lang.get('button_remove', "Rimuovi"), command=self._on_remove).pack(side="left", padx=(8,0))

        # Logo in basso a destra
        bottom = ttk.Frame(root)
        bottom.pack(fill="both", expand=True)
        self.logo_lbl = ttk.Label(bottom)
        self.logo_lbl.pack(side="right", anchor="se", padx=4, pady=4)
        self._load_logo()

    def _load_logo(self):
        try:
            if os.path.exists("logo.png"):
                img = Image.open("logo.png")
                # Ridimensiona per adattare
                img.thumbnail((160, 80))
                self._logo_imgtk = ImageTk.PhotoImage(img)
                self.logo_lbl.configure(image=self._logo_imgtk)
        except Exception:
            # Fallback: nessun logo
            pass

    def _load_types(self):
        rows = self.db.fetch_component_types()
        self._types = rows
        values = [self.lang.get('all_types', "Tutti i tipi")] + [r.ComponentTypeName for r in rows]
        self.type_combo['values'] = values
        if values:
            self.type_combo.current(0)

    def _load_components(self, type_id: int | None):
        rows = self.db.fetch_components(type_id)
        self._components = rows
        values = []
        self._comp_index_by_display.clear()
        for r in rows:
            disp = f"{r.ComponentCode} â€¢ {r.ComponentDescription or ''}"
            values.append(disp)
            self._comp_index_by_display[disp] = r
        self.comp_combo['values'] = values
        if values:
            self.comp_combo.current(0)
            # Aggiorna stato per il primo elemento
            self._on_component_changed()

    def _load_rules(self):
        rows = self.db.fetch_kanban_rules()
        self._rules = rows
        values = []
        self._rule_index_by_display.clear()
        for r in rows:
            if getattr(r, "MinimumProcent", None) is not None:
                disp = f"Percentuale {int(r.MinimumProcent)}%"
            elif getattr(r, "MinimumQty", None) is not None:
                disp = f"QuantitÃ  {int(r.MinimumQty)}"
            else:
                disp = f"Regola {r.KanBanRuleID}"
            values.append(disp)
            self._rule_index_by_display[disp] = r
        self.rule_combo['values'] = values
        if values:
            self.rule_combo.current(0)

    def _on_type_changed(self, event=None):
        sel = self.type_var.get()
        if not sel or sel == self.lang.get('all_types', "Tutti i tipi"):
            self._load_components(None)
            return
        # Trova type_id selezionato
        type_id = None
        for t in self._types:
            if t.ComponentTypeName == sel:
                type_id = t.IdComponentType
                break
        self._load_components(type_id)

    def _get_selected_component_row(self):
        disp = self.comp_var.get()
        return self._comp_index_by_display.get(disp)

    def _on_component_changed(self, event=None):
        row = self._get_selected_component_row()
        if not row:
            self.current_rule_label.configure(text=self.lang.get('no_rule', "Nessuna regola associata"))
            return
        active = self.db.fetch_active_rule_for_component(row.IdComponent)
        if active and getattr(active, "KanBanRuleId", None) is not None:
            # Trova dettaglio regola per visualizzare percentuale/qty
            rule = None
            for r in self._rules:
                if r.KanBanRuleID == active.KanBanRuleId:
                    rule = r
                    break
            if rule:
                if getattr(rule, "MinimumProcent", None) is not None:
                    text = self.lang.get('current_rule_fmt_pct', "Regola attiva: {v}%").format(v=int(rule.MinimumProcent))
                elif getattr(rule, "MinimumQty", None) is not None:
                    text = self.lang.get('current_rule_fmt_qty', "Regola attiva: Q.tÃ  {v}").format(v=int(rule.MinimumQty))
                else:
                    text = self.lang.get('current_rule_fmt_id', "Regola attiva ID {id}").format(id=rule.KanBanRuleID)
                self.current_rule_label.configure(text=text)
                # Pre-seleziona la regola nel combo se matcha
                for disp, rr in self._rule_index_by_display.items():
                    if rr.KanBanRuleID == rule.KanBanRuleID:
                        self.rule_var.set(disp)
                        break
            else:
                self.current_rule_label.configure(text=self.lang.get('current_rule_unknown', "Regola attiva non trovata"))
        else:
            self.current_rule_label.configure(text=self.lang.get('no_rule', "Nessuna regola associata"))

    def _on_assign(self):
        comp = self._get_selected_component_row()
        if not comp:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('component_required', "Seleziona un componente."), parent=self)
            return
        disp = self.rule_var.get()
        rule = self._rule_index_by_display.get(disp)
        if not rule:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('rule_required', "Seleziona una regola."), parent=self)
            return
        ok, err = self.db.assign_rule_to_component(comp.IdComponent, rule.KanBanRuleID)
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('rule_assigned', "Regola associata."), parent=self)
            self._on_component_changed()
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('rule_assign_err', f"Errore associazione: {err or self.db.last_error_details}"),
                                 parent=self)

    def _on_remove(self):
        comp = self._get_selected_component_row()
        if not comp:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"), self.lang.get('component_required', "Seleziona un componente."), parent=self)
            return
        if not messagebox.askyesno(self.lang.get('confirm_title', "Conferma"), self.lang.get('confirm_remove_rule', "Rimuovere la regola attiva?"), parent=self):
            return
        ok, err = self.db.assign_rule_to_component(comp.IdComponent, None)  # solo chiusura
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"), self.lang.get('rule_removed', "Regola rimossa."), parent=self)
            self._on_component_changed()
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('rule_remove_err', f"Errore rimozione: {err or self.db.last_error_details}"),
                                 parent=self)

class KanbanMoveForm(tk.Toplevel):
    def __init__(self, master, db_handler, lang_manager):
        super().__init__(master)
        self.db = db_handler
        self.lang = lang_manager

        self.title(self.lang.get('kanban_move_title', 'KanBan - Movimenta'))
        self.geometry("720x420")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # Stato
        self.op_var = tk.StringVar(value="unload")  # default: Withdrawal
        self._session_user = self._get_app_username()  # utente loggato alla maschera (se esiste)
        self._load_user = None  # utente che fa login quando passa a Load
        self.qty_var = tk.StringVar()
        self.component_var = tk.StringVar()
        self.location_var = tk.StringVar()

        # Cache liste per filtro
        self._components = []   # list of (IdComponent, code, descr)
        self._locations = []    # list of (LocationId, code, area)
        self._in_use_location_ids = set()  # locazioni dove il componente ha stock > 0
        self._build_ui()
        self._load_lists()

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # Operazione
        opf = ttk.Labelframe(root, text=self.lang.get('move_operation', 'Operazione'))
        opf.pack(fill="x")
        ttk.Radiobutton(opf, text=self.lang.get('move_load', 'Carico'),
                        variable=self.op_var, value="load", command=self._on_op_changed).pack(side="left", padx=(10, 10), pady=6)
        ttk.Radiobutton(opf, text=self.lang.get('move_unload', 'Prelievo'),
                        variable=self.op_var, value="unload", command=self._on_op_changed).pack(side="left", padx=(0, 10), pady=6)

        # Selezioni
        sf = ttk.Labelframe(root, text=self.lang.get('move_selection', 'Selezione'))
        sf.pack(fill="x", pady=(10, 0))

        r1 = ttk.Frame(sf); r1.pack(fill="x", padx=6, pady=6)
        ttk.Label(r1, text=self.lang.get('component', 'Componente'), width=16).pack(side="left")
        self.cb_component = ttk.Combobox(r1, textvariable=self.component_var, width=50)
        self.cb_component.pack(side="left", fill="x", expand=True)
        self.cb_component.bind("<KeyRelease>", self._on_component_typed)

        # Location row (giÃ  presente)
        r2 = ttk.Frame(sf)
        r2.pack(fill="x", padx=6, pady=6)
        ttk.Label(r2, text=self.lang.get('location', 'Locazione'), width=16).pack(side="left")
        self.cb_location = ttk.Combobox(r2, textvariable=self.location_var, width=50)
        self.cb_location.pack(side="left", fill="x", expand=True)
        self.cb_location.bind("<KeyRelease>", self._on_location_typed)

        # HINT: *** in uso per questo componente
        self.lbl_loc_hint = ttk.Label(
            r2,
            text=self.lang.get('loc_in_use_hint', '*** = locazione in uso per questo componente'),
            foreground="gray"
        )
        self.lbl_loc_hint.pack(side="left", padx=(8, 0))

        # Quantity row (giÃ  presente)
        r3 = ttk.Frame(sf);
        r3.pack(fill="x", padx=6, pady=6)
        ttk.Label(r3, text=self.lang.get('quantity', 'QuantitÃ '), width=16).pack(side="left")
        ttk.Entry(r3, textvariable=self.qty_var, width=12).pack(side="left")

        # Etichette saldo: qui e altrove
        self.lbl_here_balance = ttk.Label(r3, text=self.lang.get('balance_here', 'Saldo qui: {qty}').format(qty='-'))
        self.lbl_here_balance.pack(side="left", padx=(16, 8))
        self.lbl_other_balance = ttk.Label(r3, text=self.lang.get('balance_other', 'Altrove: {qty}').format(qty='-'))
        self.lbl_other_balance.pack(side="left")

        # Pulsanti
        bf = ttk.Frame(root); bf.pack(fill="x", pady=(12, 0))
        ttk.Button(bf, text=self.lang.get('button_execute', 'Esegui'),
                   command=self._on_execute).pack(side="left")
        ttk.Button(bf, text=self.lang.get('button_import_excel', 'Importa da Excel'),
                   command=self._on_import_excel).pack(side="left", padx=(8,0))
        ttk.Button(bf, text=self.lang.get('button_close', 'Chiudi'),
                   command=self.destroy).pack(side="right")

        # Log/risultati
        self.log = tk.Text(root, height=10, state="disabled")
        self.log.pack(fill="both", expand=True, pady=(12, 0))

        # Component combobox bindings
        self.cb_component.bind("<<ComboboxSelected>>", self._on_component_selected)
        self.cb_component.bind("<FocusOut>", self._on_component_focus_out)

        # Location combobox bindings
        self.cb_location.bind("<<ComboboxSelected>>", self._on_location_selected)
        self.cb_location.bind("<FocusOut>", lambda e: self._update_balances())

    def _on_op_changed(self):
        if self.op_var.get() == "load":
            if not self._ensure_load_login():
                # login annullato/negato: torna a Withdrawal
                self.op_var.set("unload")
                messagebox.showwarning(
                    self.lang.get('warn_title', 'Attenzione'),
                    self.lang.get('load_requires_login', 'Per eseguire un carico Ã¨ necessario effettuare il login.'),
                    parent=self
                )
        self._refresh_component_dependent_ui()

    def _load_lists(self):
        # Componenti
        self._components = []
        for r in self.db.fetch_all_components_for_combo():
            self._components.append((r.IdComponent, r.ComponentCode, r.ComponentDescription or ""))
        # Stringhe combo: "CODE - Description"
        comp_items = [f"{c[1]} - {c[2]}" if c[2] else c[1] for c in self._components]
        self.cb_component["values"] = comp_items

        # Locazioni
        self._locations = []
        for r in self.db.fetch_all_locations_for_combo():
            self._locations.append((r.LocationId, r.LocationCode, r.KanBanLocation or ""))
        loc_items = [f"{x[1]} - {x[2]}" if x[2] else x[1] for x in self._locations]
        self.cb_location["values"] = loc_items

    def _on_component_typed(self, event):
        text = self.component_var.get().lower().strip()
        items = []
        for (_id, code, descr) in self._components:
            s = f"{code} - {descr}" if descr else code
            if text in code.lower() or text in (descr or "").lower():
                items.append(s)
        self.cb_component["values"] = items

    def _on_location_typed(self, event):
        text = self.location_var.get().lower().strip()
        items = []
        for (lid, code, area) in self._locations:
            if text in code.lower() or text in (area or "").lower():
                star = " (***)" if lid in getattr(self, "_in_use_location_ids", set()) else ""
                s = f"{code}{star} - {area}" if area else f"{code}{star}"
                items.append(s)
        self.cb_location["values"] = items
        # non aggiorniamo i saldi finchÃ© non c'Ã¨ una locazione valida; ci pensa <<ComboboxSelected>> o FocusOut

    def _resolve_component(self, text: str):
        if not text:
            return None, None, None
        code = text.split(" - ", 1)[0].strip()
        # Prova match veloce dal cache
        for (cid, ccode, cdescr) in self._components:
            if ccode == code:
                return cid, ccode, cdescr
        # Fallback DB
        cid = self.db.get_component_id_by_code(code)
        return (cid, code, None) if cid else (None, None, None)

    def _resolve_location(self, text: str):
        if not text:
            return None, None, None
        code = text.split(" - ", 1)[0].strip()
        code = code.replace(" (***)", "")  # rimuove marcatura
        # Rimuove eventuale (Qty: ...)
        code = re.sub(r"\s*\(Qty:.*?\)", "", code)

        for (lid, lcode, area) in self._locations:
            if lcode == code:
                return lid, lcode, area
        lid = self.db.get_location_id_by_code(code)
        return (lid, code, None) if lid else (None, None, None)

    def _append_log(self, msg: str):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.configure(state="disabled")
        self.log.see("end")

    def _parse_qty(self):
        s = self.qty_var.get().strip()
        if not s.isdigit():
            return None
        n = int(s)
        return n if n > 0 else None

    def _on_execute(self):
        # Validazioni
        comp_id, comp_code, _ = self._resolve_component(self.component_var.get())
        if not comp_id:
            messagebox.showwarning(self.lang.get('warn_title', 'Attenzione'),
                                   self.lang.get('move_err_select_component', 'Seleziona un componente valido.'),
                                   parent=self)
            return
        loc_id, loc_code, _ = self._resolve_location(self.location_var.get())
        if not loc_id:
            messagebox.showwarning(self.lang.get('warn_title', 'Attenzione'),
                                   self.lang.get('move_err_select_location', 'Seleziona una locazione valida.'),
                                   parent=self)
            return
        qty = self._parse_qty()
        if qty is None:
            messagebox.showwarning(self.lang.get('warn_title', 'Attenzione'),
                                   self.lang.get('move_err_qty_positive', 'La quantitÃ  deve essere un intero > 0.'),
                                   parent=self)
            return

        op = self.op_var.get()
        delta = qty if op == "load" else -qty

        # Se prelievo, verifica disponibilitÃ 
        if delta < 0:
            avail = self.db.get_current_stock(comp_id, loc_id)
            if avail + delta < 0:
                message = self.lang.get('move_err_stock_insufficient',
                                        'QuantitÃ  non disponibile. Disponibile: {avail}').format(avail=avail)
                messagebox.showerror(self.lang.get('error_title', 'Errore'), message, parent=self)
                return

        # Determina l'utente da salvare
        if delta < 0:
            user_to_save = self._session_user or self._get_app_username()
        else:
            if not self._ensure_load_login():
                messagebox.showwarning(
                    self.lang.get('warn_title', 'Attenzione'),
                    self.lang.get('load_requires_login', 'Per eseguire un carico Ã¨ necessario effettuare il login.'),
                    parent=self
                )
                return
            user_to_save = self._load_user

        ok, err = self.db.insert_kanban_movement(loc_id, comp_id, int(delta), user_to_save)

        if ok:
            msg = self.lang.get('move_ok_load', 'Carico registrato.') if delta > 0 else \
                  self.lang.get('move_ok_unload', 'Prelievo registrato.')
            messagebox.showinfo(self.lang.get('info_title', 'Informazione'), msg, parent=self)
            self._append_log(f"{'+' if delta>0 else ''}{delta} {comp_code} @ {self.location_var.get()} - OK")
            self.qty_var.set("")
            self._refresh_component_dependent_ui()  # riflette il nuovo stock
        else:
            messagebox.showerror(self.lang.get('error_title', 'Errore'),
                                 err or self.db.last_error_details, parent=self)

    def _on_import_excel(self):
        """
        Importa movimenti da file CSV o XLSX con colonne:
          - Component (o codici simili: ComponentCode, Codice, Code)
          - Location (o LocationCode, Locazione, Loc)
          - Quantity (o Qty, Quantita, Qta)
        Regole utente:
          - Prelievo (qty < 0): salva l'utente di sessione (login alla maschera).
          - Carico (qty > 0): al primo carico richiede login; usa quell'utente per tutti i carichi dell'import.
        """
        import os, csv
        from tkinter import filedialog, messagebox

        # Selezione file
        file_path = filedialog.askopenfilename(
            title=self.lang.get('import_select_file', 'Seleziona file movimenti'),
            filetypes=[('Excel/CSV', '*.xlsx *.csv'), ('Excel', '*.xlsx'), ('CSV', '*.csv'), ('Tutti i file', '*.*')]
        )
        if not file_path:
            return

        ext = os.path.splitext(file_path)[1].lower()

        # Lettura righe dal file in una lista di dict con chiavi canoniche: Component, Location, Quantity
        rows = []

        def _canon(key: str) -> str:
            k = (key or '').strip().lower().replace(' ', '').replace('_', '')
            if k in ('component', 'componentcode', 'code', 'codice', 'idcomponent', 'idcomp', 'comp'):
                return 'Component'
            if k in ('location', 'locationcode', 'loc', 'locazione', 'codicelocazione'):
                return 'Location'
            if k in ('quantity', 'qty', 'quantita', 'qta', 'quantitÃ '):
                return 'Quantity'
            return key

        try:
            if ext == '.csv':
                with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                    reader = csv.DictReader(f)
                    # Mappa header a canonici
                    field_map = {h: _canon(h) for h in reader.fieldnames or []}
                    logger.info(f"Import CSV Headers: {reader.fieldnames} -> Mapped: {field_map}")
                    for r in reader:
                        row = {field_map.get(k, k): (v if v is not None else '') for k, v in r.items()}
                        rows.append(row)
            elif ext == '.xlsx':
                try:
                    from openpyxl import load_workbook
                except Exception:
                    messagebox.showerror(
                        self.lang.get('error_title', 'Errore'),
                        self.lang.get('openpyxl_missing', 'Per importare file .xlsx Ã¨ necessario openpyxl.'),
                        parent=self
                    )
                    return
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = []
                for j, cell in enumerate(ws[1], start=1):
                    headers.append(_canon(str(cell.value) if cell.value is not None else f'Col{j}'))
                logger.info(f"Import XLSX Headers (canon): {headers}")
                for i, xrow in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    r = {}
                    for j, val in enumerate(xrow):
                        key = headers[j] if j < len(headers) else f'Col{j + 1}'
                        r[key] = '' if val is None else str(val)
                    rows.append(r)
            else:
                messagebox.showerror(
                    self.lang.get('error_title', 'Errore'),
                    self.lang.get('unsupported_file', 'Formato file non supportato. Usa .csv o .xlsx.'),
                    parent=self
                )
                return
        except Exception as e:
            messagebox.showerror(
                self.lang.get('error_title', 'Errore'),
                self.lang.get('import_read_error', f'Errore lettura file: {e}'),
                parent=self
            )
            return

        if not rows:
            messagebox.showinfo(
                self.lang.get('info_title', 'Informazione'),
                self.lang.get('import_no_rows', 'Nessuna riga da importare.'),
                parent=self
            )
            return

        # Stato utenti per salvataggio
        load_login_done = False
        session_user = self._session_user or self._get_app_username()

        ok_count = 0
        err_count = 0
        first_error_msg = None

        for idx, r in enumerate(rows, start=1):
            comp_text = str(r.get('Component', '') or '').strip()
            loc_text = str(r.get('Location', '') or '').strip()
            qty_text = str(r.get('Quantity', '') or '').strip()

            # Validazioni base
            if not comp_text or not loc_text or not qty_text:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_missing_fields',
                                                    f'Riga {idx}: campi mancanti (Component/Location/Quantity).')
                continue

            # Parse quantitÃ 
            try:
                qty = int(float(qty_text))  # accetta "10.0" -> 10
            except Exception:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_qty', f'Riga {idx}: QuantitÃ  non valida: {qty_text}')
                continue
            if qty == 0:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_qty_zero', f'Riga {idx}: QuantitÃ  zero non valida.')
                continue

            # Risolvi component e location
            comp_id, comp_code, _ = self._resolve_component(comp_text)
            if not comp_id:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_comp',
                                                    f'Riga {idx}: Componente non trovato: {comp_text}')
                continue

            loc_id, loc_code, _ = self._resolve_location(loc_text)
            if not loc_id:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_loc', f'Riga {idx}: Locazione non trovata: {loc_text}')
                continue

            # Determina l'utente da salvare, in base al segno della quantitÃ 
            if qty < 0:
                # Prelievo: richiede utente di sessione (maschera)
                if not session_user:
                    # L'app richiede che nei prelievi ci sia un utente loggato alla maschera
                    messagebox.showwarning(
                        self.lang.get('warn_title', 'Attenzione'),
                        self.lang.get('withdraw_requires_session_login',
                                      'Prelievo rilevato nell\'import: Ã¨ necessario essere loggati alla maschera.'),
                        parent=self
                    )
                    return  # interrompe l'import intero
                user_to_save = session_user
            else:
                # Carico: richiede login on-demand (una sola volta)
                if not load_login_done:
                    if not self._ensure_load_login():
                        messagebox.showwarning(
                            self.lang.get('warn_title', 'Attenzione'),
                            self.lang.get('import_load_requires_login',
                                          'Import annullato: login necessario per i carichi.'),
                            parent=self
                        )
                        return  # interrompe l'import intero
                    load_login_done = True
                user_to_save = self._load_user

            # Controllo disponibilitÃ  per prelievo (opzionale, ma utile)
            if qty < 0:
                try:
                    avail = self.db.get_current_stock(comp_id, loc_id)
                except Exception:
                    avail = None
                if avail is not None and avail + qty < 0:
                    err_count += 1
                    if first_error_msg is None:
                        first_error_msg = self.lang.get(
                            'import_err_insufficient',
                            f'Riga {idx}: QuantitÃ  non disponibile. Disponibile: {avail}'
                        )
                    continue

            # Inserimento movimento
            ok, err = self.db.insert_kanban_movement(loc_id, comp_id, qty, user_to_save)
            if ok:
                ok_count += 1
                if hasattr(self, '_append_log'):
                    sign = '+' if qty > 0 else ''
                    self._append_log(f"{sign}{qty} {comp_code} @ {loc_code} - {user_to_save} - OK")
                    logger.info(f"Riga {idx}: OK - {comp_code} @ {loc_code} Qty={qty} User={user_to_save}")
            else:
                err_count += 1
                if first_error_msg is None:
                    first_error_msg = self.lang.get('import_err_db', f'Riga {idx}: Errore DB: {err}')
                if hasattr(self, '_append_log'):
                    self._append_log(f"{qty} {comp_code} @ {loc_code} - {user_to_save} - ERR: {err}")
                    logger.error(f"Riga {idx}: ERR - {comp_code} @ {loc_code} Qty={qty} User={user_to_save} Error={err}")
        # Riepilogo
        if err_count == 0:
            msg_fmt = self.lang.get('import_ok', 'Import completato: {ok} movimenti importati.')
            msg = msg_fmt.format(ok=ok_count)
            messagebox.showinfo(
                self.lang.get('success_title', 'Successo'),
                msg,
                parent=self
            )
            logger.info(msg)
        else:
            msg_fmt = self.lang.get('import_summary',
                                'Import terminato: OK={ok}, Errori={err}.\n{msg}')
            try:
                msg = msg_fmt.format(ok=ok_count, err=err_count, msg=first_error_msg or "",
                                     ok_count=ok_count, err_count=err_count)
            except Exception:
                # Fallback se format fallisce (es. chiavi strane)
                msg = f"Import terminato: OK={ok_count}, Errori={err_count}.\n{first_error_msg or ''}"
            
            messagebox.showwarning(self.lang.get('warn_title', 'Attenzione'), msg, parent=self)
            logger.warning(msg)

        # Refresh UI/saldi dopo import
        if hasattr(self, '_refresh_component_dependent_ui'):
            try:
                self._refresh_component_dependent_ui()
            except Exception:
                pass

    def _on_component_selected(self, event=None):
        self._refresh_component_dependent_ui()

    def _on_component_focus_out(self, event=None):
        # Se il testo corrisponde a un componente valido, aggiorna UI
        comp_id, _, _ = self._resolve_component(self.component_var.get())
        if comp_id:
            self._refresh_component_dependent_ui()

    def _on_location_selected(self, event=None):
        self._update_balances()

    def _refresh_component_dependent_ui(self):
        """
        Dopo la scelta del componente: marca le locazioni in uso (***), aggiorna i saldi.
        Se 'unload' (Prelievo), filtra solo le locazioni con stock > 0 e auto-seleziona.
        """
        comp_id, _, _ = self._resolve_component(self.component_var.get())
        if not comp_id:
            # reset marcature e saldi
            self._in_use_location_ids = set()
            # ricostruisci la lista locazioni senza marcature
            items = []
            for (lid, code, area) in self._locations:
                s = f"{code} - {area}" if area else code
                items.append(s)
            self.cb_location["values"] = items
            self._update_balances()
            return

        # Ottieni locazioni in uso per il componente
        loc_map = self.db.get_component_locations_with_stock(comp_id)  # {LocationId: Qty}
        self._in_use_location_ids = set(loc_map.keys())

        op = self.op_var.get()
        items = []

        if op == "unload":
            # Filtra: solo locazioni con stock
            for (lid, code, area) in self._locations:
                if lid in self._in_use_location_ids:
                    qty = loc_map.get(lid, 0)
                    label = f"{code} (Qty: {qty}) - {area}" if area else f"{code} (Qty: {qty})"
                    items.append(label)
            self.cb_location["values"] = items

            # Auto-select prima locazione se presente
            if items:
                self.cb_location.current(0)
                self._on_location_selected()
            else:
                self.cb_location.set('')
                self._update_balances()
        else:
            # Load: mostra tutte, marca con (***)
            for (lid, code, area) in self._locations:
                star = " (***)" if lid in self._in_use_location_ids else ""
                label = f"{code}{star} - {area}" if area else f"{code}{star}"
                items.append(label)
            self.cb_location["values"] = items
            self._update_balances()

    def _update_balances(self):
        """
        Aggiorna le etichette 'Saldo qui' e 'Altrove' in base a componente/locazione selezionati.
        """
        comp_id, _, _ = self._resolve_component(self.component_var.get())
        if not comp_id:
            self.lbl_here_balance.config(text=self.lang.get('balance_here', 'Saldo qui: {qty}').format(qty='-'))
            self.lbl_other_balance.config(text=self.lang.get('balance_other', 'Altrove: {qty}').format(qty='-'))
            return

        loc_id, _, _ = self._resolve_location(self.location_var.get())
        total = self.db.get_total_stock_component(comp_id)
        here = self.db.get_current_stock(comp_id, loc_id) if loc_id else 0
        other = max(total - here, 0)

        self.lbl_here_balance.config(text=self.lang.get('balance_here', 'Saldo qui: {qty}').format(qty=here))
        self.lbl_other_balance.config(text=self.lang.get('balance_other', 'Altrove: {qty}').format(qty=other))

    def _get_app_username(self):
        app = self.master
        for attr in ('last_authenticated_user_name', 'current_user', 'logged_user',
                     'current_username', 'username', 'user_name', 'session_user'):
            if hasattr(app, attr):
                val = getattr(app, attr)
                if isinstance(val, dict):
                    if 'username' in val and val['username']:
                        return str(val['username'])
                    if 'name' in val and val['name']:
                        return str(val['name'])
                elif val:
                    return str(val)
        return None

    def _ensure_load_login(self) -> bool:
        """
        Richiede il login quando si seleziona 'Load'. Se ok, memorizza _load_user.
        Ritorna True se login effettuato (o giÃ  presente), False se annullato/negato.
        """
        if self._load_user:
            return True

        app = self.master
        did_login = {'ok': False}

        def _after_login():
            # Preferisci il nome impostato da _execute_authorized_action; fallback a _get_app_username
            user = getattr(app, 'last_authenticated_user_name', None) or self._get_app_username()
            if user:
                self._load_user = user
                did_login['ok'] = True

        # NB: non affidarti al valore di ritorno (None); usa il flag did_login
        app._execute_authorized_action(menu_translation_key='kanban_load', action_callback=_after_login)
        return did_login['ok']

        def _after_login():
            # Preferisci il nome impostato da _execute_authorized_action; fallback a _get_app_username
            user = getattr(app, 'last_authenticated_user_name', None) or self._get_app_username()
            if user:
                self._load_user = user
                did_login['ok'] = True

        # NB: non affidarti al valore di ritorno (None); usa il flag did_login
        app._execute_authorized_action(menu_translation_key='kanban_load', action_callback=_after_login)
        return did_login['ok']
        def _after_login():
            # Preferisci il nome impostato da _execute_authorized_action; fallback a _get_app_username
            user = getattr(app, 'last_authenticated_user_name', None) or self._get_app_username()
            if user:
                self._load_user = user
                did_login['ok'] = True

        # NB: non affidarti al valore di ritorno (None); usa il flag did_login
        app._execute_authorized_action(menu_translation_key='kanban_load', action_callback=_after_login)
        return did_login['ok']

class PrinterSetupDialog(tk.Toplevel):
    def __init__(self, master, lang):
        super().__init__(master)
        self.lang = lang
        self.title(self.lang.get('printer_setup_title', "Configurazione stampante"))
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self.name_var = tk.StringVar()
        self.ip_var = tk.StringVar()
        self.port_var = tk.StringVar(value="9100")
        self.dpi_var = tk.StringVar(value="203")
        self.textpt_var = tk.StringVar(value="12")

        frm = ttk.Frame(self, padding=16)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=self.lang.get('printer_name', "Nome stampante")).grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.name_var, width=36).grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(frm, text="IP").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.ip_var, width=36).grid(row=1, column=1, sticky="ew", pady=4)

        ttk.Label(frm, text=self.lang.get('printer_port', "Porta")).grid(row=2, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.port_var, width=10).grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(frm, text=self.lang.get('printer_dpi', "DPI")).grid(row=3, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.dpi_var, width=10).grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(frm, text=self.lang.get('printer_textpt', "Corpo testo")).grid(row=4, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.textpt_var, width=10).grid(row=4, column=1, sticky="w", pady=4)

        # Bottoni azione
        btns = ttk.Frame(frm)
        btns.grid(row=10, column=0, columnspan=2, sticky="e", pady=(10,0))

        # Extra: Apri cartella config
        ttk.Button(btns, text=self.lang.get('printer_open_folder', "Apri cartella"),
                   command=open_config_folder).pack(side="left")

        # Extra: Test stampa veloce
        ttk.Button(btns, text=self.lang.get('printer_test', "Test stampa"),
                   command=self._test_print).pack(side="left", padx=(8, 0))

        ttk.Button(btns, text=self.lang.get('button_cancel', "Annulla"),
                   command=self._cancel).pack(side="right")
        ttk.Button(btns, text=self.lang.get('button_save', "Salva"),
                   command=self._save).pack(side="right", padx=(0,6))

        frm.columnconfigure(1, weight=1)

        # Precarica config
        cur = load_printer_config()
        if cur:
            self.name_var.set(cur.get("name",""))
            self.ip_var.set(cur.get("ip",""))
            self.port_var.set(str(cur.get("port","9100")))
            self.dpi_var.set(str(cur.get("dpi","203")))
            self.textpt_var.set(str(cur.get("text_pt","12")))

    def _cancel(self):
        self.result = None
        self.destroy()

    def _save(self):
        name = self.name_var.get().strip()
        ip = self.ip_var.get().strip()
        port = self.port_var.get().strip()
        dpi = self.dpi_var.get().strip()
        textpt = self.textpt_var.get().strip()

        if not name or not ip or not port:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('printer_required', "Compila nome, IP e porta."), parent=self)
            return
        if not validate_ip(ip):
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('printer_ip_invalid', "IP non valido."), parent=self)
            return
        try:
            port_i = int(port)
            dpi_i = int(dpi)
            textpt_i = int(textpt)
        except ValueError:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('printer_numeric_required', "Porta, DPI e corpo testo devono essere numerici."), parent=self)
            return

        cfg = {
            "name": name,
            "ip": ip,
            "port": port_i,
            "dpi": dpi_i,
            "label_cm": [5, 5],
            "text_pt": textpt_i,
            "language": "ZPL",
            "version": 1
        }
        ok, err = save_printer_config(cfg)
        if not ok:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('printer_save_error', f"Impossibile salvare la configurazione. {err}"),
                                 parent=self)
            return

        self.result = cfg
        self.destroy()

    def _test_print(self):
        # Costruisce e invia una piccola etichetta di test
        try:
            ip = self.ip_var.get().strip()
            port = int(self.port_var.get().strip())
            dpi = int(self.dpi_var.get().strip() or "203")
        except ValueError:
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('printer_numeric_required', "Porta, DPI e corpo testo devono essere numerici."), parent=self)
            return
        if not validate_ip(ip):
            messagebox.showwarning(self.lang.get('warn_title', "Attenzione"),
                                   self.lang.get('printer_ip_invalid', "IP non valido."), parent=self)
            return
        zpl = "^XA^CI28^PW400^LL400^FO20,20^A0N,40,40^FDTest stampa^FS^FO20,80^GB350,1,1^FS^XZ"
        ok, err = send_raw_to_printer(ip, port, zpl.encode("utf-8"))
        if ok:
            messagebox.showinfo(self.lang.get('info_title', "Informazione"),
                                self.lang.get('print_ok', "Stampa inviata alla stampante."), parent=self)
        else:
            messagebox.showerror(self.lang.get('error_title', "Errore"),
                                 self.lang.get('print_error', f"Errore di stampa: {err}"), parent=self)

class LineStoppageReportForm(tk.Toplevel):
    def __init__(self, parent, db_handler, lang_manager):
        super().__init__(parent)
        self.db = db_handler
        self.lang = lang_manager

        self.title(self.lang.get('line_stoppage_report_title', "Report Fermi Linea"))
        self.geometry("400x300")

        # Crea il frame principale
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Date selectors
        date_frame = ttk.LabelFrame(main_frame, text=self.lang.get('date_range_label', "Intervallo Date"),
                                    padding="10")
        date_frame.pack(fill=tk.X, pady=(0, 20))

        # From Date
        ttk.Label(date_frame, text=self.lang.get('from_date_label', "Da:")).grid(row=0, column=0, padx=5, pady=5)
        self.from_date = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.from_date.grid(row=0, column=1, padx=5, pady=5)

        # To Date
        ttk.Label(date_frame, text=self.lang.get('to_date_label', "A:")).grid(row=1, column=0, padx=5, pady=5)
        self.to_date = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.to_date.grid(row=1, column=1, padx=5, pady=5)

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)

        ttk.Button(button_frame, text=self.lang.get('generate_report_button', "Genera Report"),
                   command=self._generate_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.lang.get('close_button', "Chiudi"),
                   command=self.destroy).pack(side=tk.RIGHT, padx=5)

    def _generate_report(self):
        """Genera il report Excel dei fermi linea."""
        try:
            # Esegue la query
            query = """
                SELECT [BreakDownProblemLogId] AS ID,
                       cast([DateReport] as DATE) AS ReportIussedOn,
                       [HourReport] as TimeIussed,
                       [FromHour] as FromTime,
                       [ToHour] as ToTime,
                       [UserName] AS IussedBy,
                       wa.AreaName As WorkingArea,
                       ia.IssueArea As Thema,
                       WS.AreaSubName As SubWorkingArea,
                       WL.WorkingLineName As Line,
                       [Lost_OR_Gain] AS BreakDownType,
                       [Hours] AS NrMinutesBreakDown,
                       [PoNumber],
                       [ProductCode],
                       [ip].DescriptionEN As BreakDownReason,
                       cast([Note] as text) as [Note],
                       CAST([DateSys] as datetime) As RealDataEntry 
                FROM [ResetServices].[BreakDown].[ReportIssueLogs] AS RI 
                inner join [ResetServices].[BreakDown].IssuesAreas AS IA on Ri.IssueAreaId=IA.IssueAreaId 
                inner join [ResetServices].[BreakDown].WorkingAreas as WA on Wa.WorkingAreaID=RI.WorkingAreaID 
                inner join [ResetServices].[BreakDown].WorkingLines AS WL on WL.WorkingLineID=RI.WorkingLineID 
                inner join [ResetServices].[BreakDown].WorkingSubAreas AS WS on WS.WorkingSubAreaID=Ri.WorkingSubAreaID 
                inner join [ResetServices].[BreakDown].IssueProblems AS [IP] on [IP].IssueProblemId=ri.IssueProblemId 
                Where Datereport between ? AND ?
                """

            # Verifica e ripristina la connessione se necessario
            if not self.db._ensure_connection():
                raise Exception("Impossibile connettersi al database")
            
            # Esegue la query con i parametri delle date
            logger.info("Esecuzione query con parametri: %s, %s", self.from_date.get_date(), self.to_date.get_date())
            self.db.cursor.execute(query, (self.from_date.get_date(), self.to_date.get_date()))
            results = self.db.cursor.fetchall()
            logger.info("Query eseguita con successo, righe recuperate: %d", len(results))

            if not results:
                messagebox.showinfo(self.lang.get('info_title', "Informazione"),
                                    self.lang.get('no_data_found',
                                                  "Nessun dato trovato per il periodo selezionato."),
                                    parent=self)
                return

            # Crea la cartella C:\temp se non esiste
            temp_dir = r"C:\temp"
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
                logger.info("Cartella %s creata", temp_dir)

            # Genera il nome del file con timestamp
            timestamp = datetime.now().strftime("%y%m%d%H%M%S")
            file_name = f"ReportBreakDown{timestamp}.xlsx"
            file_path = os.path.join(temp_dir, file_name)
            logger.info("Percorso file Excel: %s", file_path)

            # Usa pandas per creare un Excel formattato
            logger.info("Creazione DataFrame con %d righe", len(results))
            df = pd.DataFrame.from_records(results, columns=[x[0] for x in self.db.cursor.description])
            logger.info("DataFrame creato con successo, shape: %s", df.shape)

            logger.info("Inizio scrittura file Excel...")
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                logger.info("ExcelWriter creato, scrittura DataFrame...")
                df.to_excel(writer, sheet_name='Report Fermi Linea', index=False)
                logger.info("DataFrame scritto, inizio formattazione...")

                # Ottiene il foglio di lavoro
                worksheet = writer.sheets['Report Fermi Linea']

                # Formatta le intestazioni
                header_format = writer.book.add_format({
                    'bold': True,
                    'fg_color': '#D7E4BC',
                    'border': 1,
                    'align': 'center'
                })

                # Formatta le date
                date_format = writer.book.add_format({'num_format': 'dd/mm/yyyy'})
                time_format = writer.book.add_format({'num_format': 'hh:mm'})

                logger.info("Formattazione colonne...")
                # Applica la formattazione alle intestazioni e imposta le larghezze delle colonne
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                    # Imposta larghezze colonne specifiche
                    if 'Date' in value or 'Data' in value:
                        worksheet.set_column(col_num, col_num, 12)
                        # Applica formato data alle celle della colonna
                        worksheet.set_column(col_num, col_num, 12, date_format)
                    elif 'Time' in value or 'Hour' in value:
                        worksheet.set_column(col_num, col_num, 10)
                        # Applica formato ora alle celle della colonna
                        worksheet.set_column(col_num, col_num, 10, time_format)
                    elif value in ['Note', 'BreakDownReason']:
                        worksheet.set_column(col_num, col_num, 40)
                    else:
                        worksheet.set_column(col_num, col_num, 15)

                logger.info("Impostazione filtri e freeze panes...")
                # Imposta filtri automatici
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

                # Congela la prima riga
                worksheet.freeze_panes(1, 0)
                logger.info("Formattazione completata")


            logger.info("File Excel scritto con successo: %s", file_path)
            
            # Apre il file Excel direttamente
            logger.info("Apertura file Excel...")
            os.startfile(file_path)

            # Chiude la finestra del form
            logger.info("Chiusura finestra report")
            self.destroy()


        except Exception as e:
            error_msg = self.lang.get('error_generating_report', "Errore durante la generazione del report: {error}")
            logger.error(error_msg.format(error=str(e)), exc_info=True)
            messagebox.showerror(
                self.lang.get('error_title', "Errore"),
                error_msg.format(error=str(e)),
                parent=self
            )

def connect_to_database():
    """Stabilisce la connessione al database usando la configurazione sicura"""
    try:
        # Ottieni la stringa di connessione sicura
        from database_config import db_config
        conn_str = db_config.get_connection_string()

        # DEBUG: Controlla il tipo e il valore
        logger.info(f"Tipo conn_str: {type(conn_str)}")
        if not isinstance(conn_str, str):
            error_msg = f"ERRORE: conn_str non Ã¨ una stringa! Tipo: {type(conn_str)}, Valore: {conn_str}"
            logger.error(error_msg)
            raise TypeError(error_msg)

        logger.info("Tentativo di connessione al database...")

        # Connessione con timeout
        timeout = db_config.get_connection_params().get('timeout', 30)
        conn = pyodbc.connect(conn_str, timeout=timeout)

        logger.info("Connessione al database stabilita con successo")
        return conn

    except pyodbc.Error as e:
        logger.error(f"Errore di connessione al database: {e}")
        raise
    except Exception as e:
        logger.error(f"Errore imprevisto nella connessione: {e}")
        raise

class User:
    """Una semplice classe per contenere i dati dell'utente loggato."""
    def __init__(self, name, permissions=None):
        self.name = name
        # Usiamo un 'set' per ricerche di permessi super veloci
        self.permissions = set(permissions) if permissions else set()

    def has_permission(self, permission_key):
        """Verifica se l'utente ha un permesso specifico."""
        return permission_key in self.permissions

class App(tk.Tk):
    """Classe principale dell'applicazione."""

    def __init__(self):
        super().__init__()
        logger.debug("INIT: start __init__")
        self.should_exit = False  # Flag to control shutdown
        self.geometry("1024x768")
        #self.project_tree = None  # Placeholder per la lista progetti NPI
        # Variabili per lo slideshow
        self.slideshow_label = None
        self.slideshow_photo = None
        self.image_files = []
        self.current_image_index = 0
        self.slideshow_interval_ms = 60000
        self.slideshow_job_id = None

        # --- NUOVE VARIABILI PER IL FLASHING ---
        self.birthday_flash_job_id = None
        self.birthday_stop_job_id = None
        self.periodic_check_job_id = None

        # --- NUOVE VARIABILI PER LA SCRITTA SCORREVOLE ---
        self.scrolling_job_id = None
        self.scrolling_text = ""
        self.scrolling_position = 0
        self.flash_colors = ["#FFD700", "#FF4500", "#1E90FF", "#32CD32", "#FF69B4", "#9400D3"]
        # --- FINE ---

        # === 1. INIZIALIZZAZIONE DELLE DIPENDENZE FONDAMENTALI ===

        # Inizializza il database
        self.db = Database(DB_CONN_STR)
        logger.info("Tentativo di connessione al database...")
        if not self.db.connect():
            logger.error("INIT: Connessione al DB fallita. Dettagli: %s", self.db.last_error_details)
            messagebox.showerror("Database Error",
                                 f"Impossibile connettersi al database.\n\nDetails: {self.db.last_error_details}")
            self.destroy()
            self.should_exit = True
            return
        logger.info("Connessione al database stabilita con successo")

        # Carica la lingua salvata
        initial_lang = self._load_language_setting()
        logger.debug("INIT: language setting loaded -> %s", initial_lang)
        self.lang = LanguageManager(self.db)
        logger.debug("INIT: LanguageManager loaded")
        self.lang.set_language(initial_lang)

        # === 2. INIZIALIZZAZIONE DEI MODULI PRINCIPALI (INCLUSO NPI) ===
        # Questo Ã¨ il posto ideale per inizializzare i moduli che dipendono da DB e Lingua

        # --- INIZIALIZZAZIONE GESTORE NPI (POSIZIONE CORRETTA) ---
        try:
            self.npi_manager = GestoreNPI(engine=self.db.npi_engine)
            logger.info(f"NPI Manager inizializzato con engine: {self.db.npi_engine}")
            logger.info(f"Pool size: {self.db.npi_engine.pool.size()}")
            # Avvio notifiche automatiche NPI (task in ritardo/scadenza)
            try:
                from npi.npi_auto_notifications import start_notification_service, ensure_notification_config
                ensure_notification_config('npi_notifications_config.json')
                self._npi_notification_service = start_notification_service(
                    self.npi_manager, 'npi_notifications_config.json'
                )
            except Exception as e:
                logger.error(f"Errore avvio servizio notifiche automatiche NPI: {e}", exc_info=True)
        except Exception as e:
            logger.error("ERRORE CRITICO: Impossibile inizializzare il Gestore NPI: %s", e, exc_info=True)
            messagebox.showerror("Errore Modulo NPI",
                                 f"Impossibile avviare il modulo NPI.\nContattare l'assistenza.\n\nDettagli: {e}")
            self.npi_manager = None
            self.npi_menu = None
            self._npi_notification_service = None

        self.traceability_manager = TraceabilityManager(self, self.db, self.lang)
        logger.debug("INIT: traceability manager OK")

        # Aggiungi qui altri moduli principali se necessario...
        self.fct_config = fct_transfer.FCTTransferConfig()
        self.fct_manager = fct_transfer.FCTTransferManager(DB_CONN_STR, self.fct_config)
        self.fct_run_menu_index = None

        # === 3. CONTROLLI DI AVVIO E CREAZIONE UI ===

        # Controlla la versione (e se l'app deve chiudersi)
        if self.check_version() is False:
            # check_version giÃ  gestisce la chiusura, quindi fermiamo l'init
            return
        logger.debug("INIT: after check_version")

        self.doc_categories = self.db.fetch_doc_categories()
        self.logo_label = None
        self.authenticated_user_for_maintenance = None

        # Creazione dei widget e dei menu
        # Ora queste chiamate sono sicure perchÃ© self.npi_manager esiste
        self._create_widgets()
        logger.debug("INIT: widgets created")
        self._create_menu()
        logger.debug("INIT: menu created")

        # Aggiornamento testi e UI
        self.update_texts()
        logger.debug("INIT: texts updated")

        # === 4. COMPITI DI POST-AVVIO E BACKGROUND ===

        # Operazioni per rendere la finestra visibile e reattiva
        self.update_idletasks()
        self.deiconify()
        self.lift()
        self.focus_force()
        self.attributes('-topmost', True)
        self.after(500, lambda: self.attributes('-topmost', False))

        # Avvio dell'orologio e altre task post-avvio
        self._update_clock()
        self.after(100, self._post_startup_tasks)
        logger.debug("INIT: scheduled post_startup_tasks")

        # Inizializza il thread per il controllo periodico prodotti
        # NOTA: Spostato protocol handler qui per evitare duplicati
        self._product_check_thread = None
        self._product_check_stop_event = threading.Event()
        self._start_product_check_routine()
        self._start_product_check_background_task()
        
        # Inizializza il thread per il report mensile
        self._monthly_report_thread = None
        self._monthly_report_stop_event = threading.Event()
        self._start_monthly_report_background_task()
        
        # Inizializza il thread per l'invio automatico email FAI fails
        self._fai_fails_email_thread = None
        self._fai_fails_email_stop_event = threading.Event()
        self._fai_fails_email_last_sent = None  # Track last send date
        self._start_fai_fails_email_background_task()

        # Inizializza il thread per l'email settimanale NPI Overview
        self._weekly_npi_email_thread = None
        self._weekly_npi_email_stop_event = threading.Event()
        self._start_weekly_npi_email_background_task()
        
        logger.info("INIT: App initialization complete.")

        # Imposta la gestione della chiusura della finestra una sola volta
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def open_scrap_validation_with_login(self):
        """Apre la finestra di validazione scarti dopo login."""

        def action():
            user_name = getattr(self, 'last_authenticated_user_name', 'Unknown')
            scrap_validation_gui.open_scrap_validation(self, self.db, self.lang, user_name)

        self._execute_authorized_action('validate_scrap', action)

    def _start_product_check_background_task(self):
        """Avvia il thread per il controllo periodico dei prodotti"""
        from business_days import should_send_notification
        if not should_send_notification(country_code='IT'):
            logger.info("Report non inviato: oggi non Ã¨ un giorno lavorativo")
            return

        if self._product_check_thread is None or not self._product_check_thread.is_alive():
            self._product_check_stop_event.clear()
            self._product_check_thread = threading.Thread(
                target=self._product_check_worker,
                daemon=True,  # Il thread termina quando l'app si chiude
                name="ProductCheckWorker"
            )
            self._product_check_thread.start()
            logger.info("Background task per controllo prodotti avviato")

    def _is_working_hours(self):
        """Verifica se l'ora corrente Ã¨ nell'orario lavorativo (08:00 - 22:00)"""
        current_hour = datetime.now().hour
        return 8 <= current_hour <= 22

    def _product_check_worker(self):
        """
        Worker thread che esegue periodicamente la SP InsertProductToCheck.
        
        REGOLE OPERATIVE:
        - Orario: 08:00 - 22:00
        - Giorni: Solo giorni lavorativi (esclusi weekend e festivitÃ )
        - FestivitÃ  gestite: Natale, Capodanno, Epifania, Ferragosto, 
          Ognissanti, Pasqua (calcolata dinamicamente), ecc.
        """
        while not self._product_check_stop_event.is_set():
            try:
                # 1. Leggi l'intervallo in minuti dal database
                interval_minutes = self.db.get_product_check_interval()

                if interval_minutes <= 0:
                    logger.warning("Intervallo controllo prodotti non valido, uso default 30 min")
                    interval_minutes = 30

                logger.info(f"Prossimo controllo prodotti tra {interval_minutes} minuti")

                # 2. Attendi per l'intervallo specificato (con controllo stop ogni 10 secondi)
                wait_seconds = interval_minutes * 60
                elapsed = 0
                while elapsed < wait_seconds and not self._product_check_stop_event.is_set():
                    time.sleep(10)  # Check ogni 10 secondi per permettere stop rapido
                    elapsed += 10

                if self._product_check_stop_event.is_set():
                    break

                # 3. Verifica orario lavorativo (08:00 - 22:00)
                if not self._is_working_hours():
                    logger.debug("Fuori orario lavorativo (08:00-22:00), skip controllo prodotti")
                    continue

                # 4. Verifica giorni lavorativi (esclusi weekend e festivitÃ  italiane)
                # Il metodo should_send_notification() controlla:
                # - Weekend (Sabato, Domenica)
                # - FestivitÃ  fisse (Natale, Capodanno, Ferragosto, ecc.)
                # - Pasqua e LunedÃ¬ dell'Angelo (calcolati con algoritmo di Gauss)
                if not should_send_notification():
                    logger.debug("Giorno non lavorativo (weekend o festivitÃ ), skip controllo prodotti")
                    continue

                # 5. Esegui la stored procedure
                logger.info("Esecuzione SP InsertProductToCheck...")
                success = self.db.execute_product_check_sp()

                if success:
                    logger.info("âœ“ SP InsertProductToCheck eseguita con successo")

                    # 6. (Opzionale) Verifica se ci sono nuovi prodotti da controllare
                    # e mostra una notifica all'utente
                    self._check_and_notify_pending_verifications()
                else:
                    logger.error(f"âœ— Errore esecuzione SP: {self.db.last_error_details}")

            except Exception as e:
                logger.error(f"Errore nel worker controllo prodotti: {e}", exc_info=True)
                # In caso di errore, attendi 5 minuti prima di riprovare
                time.sleep(300)

        logger.info("Background task controllo prodotti terminato")

    def _start_monthly_report_background_task(self):
        """Avvia il thread per il report mensile"""
        try:
            if self._monthly_report_thread is None or not self._monthly_report_thread.is_alive():
                self._monthly_report_stop_event.clear()
                self._monthly_report_thread = threading.Thread(
                    target=self._monthly_report_worker,
                    daemon=True,
                    name="MonthlyReportWorker"
                )
                self._monthly_report_thread.start()
                logger.info("Background task per report mensile avviato")
        except Exception as e:
            logger.error(f"Errore nell'avvio del background task report mensile: {e}", exc_info=True)

    def _is_first_working_day_of_month(self):
        """
        Verifica se oggi è il primo giorno lavorativo del mese.
        Se il primo del mese cade in weekend o festività, considera il primo lunedì.
        
        Returns:
            bool: True se oggi è il primo giorno lavorativo del mese
        """
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        
        # Trova il primo giorno del mese corrente
        first_day_of_month = today.replace(day=1)
        
        # Cerca il primo giorno lavorativo del mese
        # Controlliamo i primi 10 giorni (sicuramente il primo giorno lavorativo è entro i primi 10)
        for day_offset in range(10):
            check_day = first_day_of_month + timedelta(days=day_offset)
            
            # Verifica se è weekend (sabato=5, domenica=6)
            if check_day.weekday() >= 5:
                continue
            
            # Verifica se è festività italiana
            is_holiday = False
            if check_day.month == 1 and check_day.day in [1, 6]:  # Capodanno, Epifania
                is_holiday = True
            elif check_day.month == 4 and check_day.day == 25:  # Liberazione
                is_holiday = True
            elif check_day.month == 5 and check_day.day == 1:  # Festa del lavoro
                is_holiday = True
            elif check_day.month == 6 and check_day.day == 2:  # Festa della Repubblica
                is_holiday = True
            elif check_day.month == 8 and check_day.day == 15:  # Ferragosto
                is_holiday = True
            elif check_day.month == 11 and check_day.day == 1:  # Ognissanti
                is_holiday = True
            elif check_day.month == 12 and check_day.day in [8, 25, 26]:  # Immacolata, Natale, S.Stefano
                is_holiday = True
            
            if not is_holiday:
                # Questo è il primo giorno lavorativo del mese
                return check_day == today
        
        return False
        
        # Trova il primo giorno del mese corrente
        first_day_of_month = today.replace(day=1)
        
        # Cerca il primo giorno lavorativo del mese
        current_day = first_day_of_month
        while current_day <= today:
            # Temporaneamente imposta la data per il controllo
            original_now = datetime.now
            try:
                datetime.now = lambda: datetime.combine(current_day, datetime.min.time())
                if should_send_notification(country_code='IT'):
                    # Trovato il primo giorno lavorativo
                    return current_day == today
            finally:
                datetime.now = original_now
            
            current_day += timedelta(days=1)
        
        return False

    def _monthly_report_worker(self):
        """
        Worker thread che invia il report mensile il primo giorno lavorativo del mese.
        
        REGOLE OPERATIVE:
        - Esecuzione: Primo giorno lavorativo del mese alle 09:00
        - Se il primo del mese Ã¨ weekend/festivitÃ , invia il primo lunedÃ¬
        - Controlla database per evitare invii duplicati
        """
        from business_days import should_send_notification
        from product_checks_gui import generate_monthly_excel_report
        from utils import get_email_recipients, send_monthly_report_email
        from datetime import datetime, timedelta
        import time
        
        first_run = True  # Flag per la prima esecuzione
        
        while not self._monthly_report_stop_event.is_set():
            try:
                # PRIMA ESECUZIONE: Controlla immediatamente al lancio dell'app
                if first_run:
                    first_run = False
                    logger.info("Prima esecuzione worker report mensile - controllo immediato")
                    
                    # Verifica se esiste già un record in database
                    query = """
                    SELECT IDSettings 
                    FROM traceability_rs.dbo.settings 
                    WHERE atribute = 'Sys_Verify_check_fail' 
                    AND lastcheck IS  NULL
                    """
                    
                    try:
                        cursor = self.db.conn.cursor()
                        cursor.execute(query)
                        row = cursor.fetchone()
                        cursor.close()
                        
                        if row is None:
                            # Nessun record trovato - prima esecuzione mai fatta
                            logger.info("Nessun record trovato - eseguo invio immediato del report mensile")
                            
                            # Verifica se è giorno lavorativo
                            if should_send_notification(country_code='IT'):
                                # Verifica se è il primo giorno lavorativo del mese
                                if self._is_first_working_day_of_month():
                                    # Salta l'attesa e procedi direttamente all'invio
                                    logger.info("Oggi è il primo giorno lavorativo - invio immediato")
                                else:
                                    logger.info("Oggi non è il primo giorno lavorativo del mese")
                            else:
                                logger.info("Oggi non è un giorno lavorativo")
                        else:
                            logger.info("Record trovato in database - skip invio immediato")
                    
                    except Exception as e:
                        logger.error(f"Errore nel controllo prima esecuzione: {e}")
                

                # Attendi fino alle 09:00 del giorno successivo
                now = datetime.now()
                target_hour = 9
                
                if now.hour >= target_hour:
                    # GiÃ  passate le 09:00, attendi domani
                    next_check = now.replace(hour=target_hour, minute=0, second=0, microsecond=0)
                    next_check = next_check + timedelta(days=1)
                else:
                    # Attendi le 09:00 di oggi
                    next_check = now.replace(hour=target_hour, minute=0, second=0, microsecond=0)
                
                wait_seconds = (next_check - now).total_seconds()
                logger.info(f"Prossimo controllo report mensile: {next_check.strftime('%Y-%m-%d %H:%M')}")
                
                # Attendi con controllo stop ogni 60 secondi
                elapsed = 0
                while elapsed < wait_seconds and not self._monthly_report_stop_event.is_set():
                    time.sleep(60)
                    elapsed += 60
                
                if self._monthly_report_stop_event.is_set():
                    break
                
                # Verifica se Ã¨ giorno lavorativo
                if not should_send_notification(country_code='IT'):
                    logger.debug("Oggi non Ã¨ un giorno lavorativo, skip report mensile")
                    continue
                
                # Verifica se Ã¨ il primo giorno lavorativo del mese
                if not self._is_first_working_day_of_month():
                    logger.debug("Oggi non Ã¨ il primo giorno lavorativo del mese")
                    continue
                
                # Verifica se il report Ã¨ giÃ  stato inviato questo mese
                if self.db.check_monthly_report_sent():
                    logger.info("Report mensile giÃ  inviato questo mese, skip")
                    continue
                
                logger.info("=== INVIO REPORT MENSILE ===")
                
                # Genera il file Excel
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                temp_dir = r'C:\Temp'
                if not os.path.exists(temp_dir):
                    os.makedirs(temp_dir)
                
                filename = f"Monthly_Report_{timestamp}.xlsx"
                file_path = os.path.join(temp_dir, filename)
                
                logger.info(f"Generazione report Excel: {file_path}")
                excel_file = generate_monthly_excel_report(self.db, file_path)
                
                if not excel_file:
                    logger.error("Errore nella generazione del report Excel")
                    continue
                
                # Ottieni destinatari email
                recipients = get_email_recipients(self.db.conn, 'Sys_Verify_check_fail')
                
                if not recipients:
                    logger.warning("Nessun destinatario configurato per Sys_Verify_check_fail")
                    # Pulisci file temporaneo
                    if os.path.exists(excel_file):
                        os.remove(excel_file)
                    continue
                
                logger.info(f"Destinatari email: {recipients}")
                
                # Invia email con allegato
                send_monthly_report_email(
                    recipients=recipients,
                    attachment_path=excel_file,
                    logo_path="logo.png"
                )
                
                # Aggiorna timestamp nel database
                if self.db.update_monthly_report_timestamp():
                    logger.info("âœ“ Report mensile inviato con successo")
                else:
                    logger.error("Errore nell'aggiornamento del timestamp")
                
                # Pulisci file temporaneo
                try:
                    if os.path.exists(excel_file):
                        os.remove(excel_file)
                        logger.info(f"File temporaneo rimosso: {excel_file}")
                except Exception as e:
                    logger.warning(f"Impossibile rimuovere file temporaneo: {e}")
                
            except Exception as e:
                logger.error(f"Errore nel worker report mensile: {e}", exc_info=True)
                # In caso di errore, attendi 1 ora prima di riprovare
                time.sleep(3600)
        
        logger.info("Background task report mensile terminato")

    def _start_fai_fails_email_background_task(self):
        """Avvia il thread per l'invio automatico email FAI fails"""
        try:
            if self._fai_fails_email_thread is None or not self._fai_fails_email_thread.is_alive():
                self._fai_fails_email_stop_event.clear()
                self._fai_fails_email_thread = threading.Thread(
                    target=self._fai_fails_email_worker,
                    daemon=True,
                    name="FaiFailsEmailWorker"
                )
                self._fai_fails_email_thread.start()
                logger.info("Background task per invio email FAI fails avviato")
        except Exception as e:
            logger.error(f"Errore nell'avvio del background task FAI fails email: {e}", exc_info=True)

    def _start_weekly_npi_email_background_task(self):
        """Avvia il thread per l'email settimanale NPI Overview"""
        try:
            if self._weekly_npi_email_thread is None or not self._weekly_npi_email_thread.is_alive():
                self._weekly_npi_email_stop_event.clear()
                self._weekly_npi_email_thread = threading.Thread(
                    target=self._weekly_npi_email_worker,
                    daemon=True,
                    name="WeeklyNpiOverviewWorker"
                )
                self._weekly_npi_email_thread.start()
                logger.info("Background task per email settimanale NPI avviato")
        except Exception as e:
            logger.error(f"Errore nell'avvio del background task email settimanale NPI: {e}", exc_info=True)

    def _weekly_npi_email_worker(self):
        """
        Worker thread che invia il report NPI Overview ogni lunedì lavorativo.
        """
        from business_days import should_send_notification
        from utils import get_email_recipients, send_npi_weekly_overview_email
        from datetime import datetime, timedelta
        import time

        attribute = 'Sys_email_general_napi'
        target_hour = 9
        first_run = True

        while not self._weekly_npi_email_stop_event.is_set():
            try:
                # Prima esecuzione: se oggi è lunedì lavorativo, invia subito (se non già inviato)
                if first_run:
                    first_run = False
                    today = datetime.now().date()
                    if today.weekday() == 0 and should_send_notification(country_code='IT'):
                        week_start = today - timedelta(days=today.weekday())
                        self.db.ensure_npi_weekly_email_log_table()
                        if not self.db.check_weekly_npi_email_sent(week_start, attribute):
                            recipients = get_email_recipients(self.db.conn, attribute=attribute)
                            if recipients:
                                report_data = self.npi_manager.get_npi_overview_report_data()
                                report_path = self.npi_manager.export_npi_overview_report()
                                chart_path = self._create_npi_overview_pie_chart(report_data, prefix="NPI_Overview_Pie")
                                send_npi_weekly_overview_email(
                                    recipients,
                                    report_path,
                                    summary=(report_data or {}).get('summary'),
                                    chart_path=chart_path
                                )
                                self.db.log_weekly_npi_email_sent(week_start, attribute)
                                logger.info("Email settimanale NPI inviata (prima esecuzione)")
                            else:
                                logger.warning("Email NPI settimanale: nessun destinatario configurato")

                # Attendi fino alle 09:00 del giorno successivo
                now = datetime.now()
                if now.hour >= target_hour:
                    next_check = now.replace(hour=target_hour, minute=0, second=0, microsecond=0) + timedelta(days=1)
                else:
                    next_check = now.replace(hour=target_hour, minute=0, second=0, microsecond=0)

                wait_seconds = (next_check - now).total_seconds()
                logger.info(f"Prossimo controllo email settimanale NPI: {next_check.strftime('%Y-%m-%d %H:%M')}")

                elapsed = 0
                while elapsed < wait_seconds and not self._weekly_npi_email_stop_event.is_set():
                    time.sleep(60)
                    elapsed += 60

                if self._weekly_npi_email_stop_event.is_set():
                    break

                today = datetime.now().date()
                if today.weekday() != 0:
                    continue
                if not should_send_notification(country_code='IT'):
                    logger.debug("Email NPI settimanale: oggi non è un giorno lavorativo")
                    continue

                week_start = today - timedelta(days=today.weekday())
                self.db.ensure_npi_weekly_email_log_table()
                if self.db.check_weekly_npi_email_sent(week_start, attribute):
                    logger.info("Email NPI settimanale già inviata per questa settimana, skip")
                    continue

                recipients = get_email_recipients(self.db.conn, attribute=attribute)
                if not recipients:
                    logger.warning("Email NPI settimanale: nessun destinatario configurato")
                    continue

                report_data = self.npi_manager.get_npi_overview_report_data()
                report_path = self.npi_manager.export_npi_overview_report()
                chart_path = self._create_npi_overview_pie_chart(report_data, prefix="NPI_Overview_Pie")
                send_npi_weekly_overview_email(
                    recipients,
                    report_path,
                    summary=(report_data or {}).get('summary'),
                    chart_path=chart_path
                )
                self.db.log_weekly_npi_email_sent(week_start, attribute)
                logger.info("Email NPI settimanale inviata con successo")

            except Exception as e:
                logger.error(f"Errore nel worker email settimanale NPI: {e}", exc_info=True)
                time.sleep(3600)

    def _create_npi_overview_pie_chart(self, report_data, prefix="NPI_Overview_Pie"):
        """
        Crea un grafico a torta riassuntivo e restituisce il path del file PNG.
        """
        if not report_data or not report_data.get('summary'):
            return None
        summary = report_data['summary']
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
        except Exception as e:
            logger.warning(f"Matplotlib non disponibile per grafico NPI: {e}")
            return None

        labels = ["Active", "In Completion", "Completed", "Overdue"]
        values = [
            summary.get('active', 0),
            summary.get('in_completion', 0),
            summary.get('completed', 0),
            summary.get('overdue', 0),
        ]

        if sum(values) == 0:
            return None

        colors = ["#4E79A7", "#F28E2B", "#59A14F", "#E15759"]
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors)
        ax.set_title("NPI Overview Summary", fontsize=12)
        ax.axis('equal')

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_dir = r"C:\Temp"
        os.makedirs(temp_dir, exist_ok=True)
        file_path = os.path.join(temp_dir, f"{prefix}_{timestamp}.png")
        fig.tight_layout()
        fig.savefig(file_path, dpi=150)
        plt.close(fig)
        return file_path

    def _fai_fails_email_worker(self):
        """
        Worker thread che invia email FAI fails automaticamente una volta al giorno.
        
        REGOLE OPERATIVE:
        - Orario: 07:00 (una volta al giorno)
        - Giorni: Solo giorni lavorativi
        - Invia solo se ci sono fails non analizzati (IsAnalized = 0)
        """
        import utils
        from business_days import should_send_notification
        from datetime import datetime, timedelta
        import time
        
        logger.info("Worker FAI fails email avviato")
        
        while not self._fai_fails_email_stop_event.is_set():
            try:
                now = datetime.now()
                current_date = now.date()
                current_hour = now.hour
                
                # Verifica se è giorno lavorativo
                if not should_send_notification(country_code='IT'):
                    logger.debug("FAI fails email: oggi non è un giorno lavorativo")
                    # Attendi fino a domani
                    time.sleep(3600)  # 1 ora
                    continue
                
                # Verifica se l'email è già stata inviata oggi
                if self._fai_fails_email_last_sent == current_date:
                    logger.debug(f"FAI fails email già inviata oggi ({current_date})")
                    # Attendi 1 ora prima di ricontrollare
                    time.sleep(3600)
                    continue
                
                # Verifica se è l'orario corretto (07:00)
                if current_hour == 7:
                    logger.info("FAI fails email: orario corretto (07:00) - tentativo invio")
                    
                    try:
                        # Invia email FAI fails
                        success = utils.send_fai_fails_notification(
                            self.db.conn, 
                            logo_path="logo.png"
                        )
                        
                        if success:
                            logger.info("✅ Email FAI fails inviata con successo")
                            self._fai_fails_email_last_sent = current_date
                        else:
                            logger.info("ℹ️ Nessun FAI fail non analizzato - email non inviata")
                            # Marca comunque come "inviata" per oggi per evitare retry continui
                            self._fai_fails_email_last_sent = current_date
                    
                    except Exception as e:
                        logger.error(f"Errore nell'invio email FAI fails: {e}", exc_info=True)
                        # Non marcare come inviata in caso di errore - riproverà
                    
                    # Attendi 1 ora per evitare invii multipli nella stessa ora
                    time.sleep(3600)
                
                else:
                    # Non è ancora l'orario giusto
                    # Calcola quanto tempo manca alle 07:00
                    target_time = now.replace(hour=7, minute=0, second=0, microsecond=0)
                    
                    if current_hour >= 7:
                        # Già passate le 07:00 oggi, attendi fino a domani
                        target_time += timedelta(days=1)
                    
                    wait_seconds = (target_time - now).total_seconds()
                    
                    # Attendi massimo 1 ora alla volta per permettere controlli intermedi
                    wait_seconds = min(wait_seconds, 3600)
                    
                    logger.debug(f"FAI fails email: attesa {wait_seconds/60:.1f} minuti fino alle 07:00")
                    time.sleep(wait_seconds)
            
            except Exception as e:
                logger.error(f"Errore nel worker FAI fails email: {e}", exc_info=True)
                # In caso di errore, attendi 1 ora prima di riprovare
                time.sleep(3600)
        
        logger.info("Background task FAI fails email terminato")



    def _show_verification_notification(self, count):
        """Mostra una notifica all'utente (eseguito nel thread principale)"""
        msg = f"âš ï¸ Ci sono {count} prodotti che necessitano verifica!"
        logger.info(msg)

        # Opzione 1: Messagebox (invasivo)
        # messagebox.showinfo(self.lang.get('app_title'), msg)

        # Opzione 2: Label temporanea nella status bar (meno invasivo)
        if hasattr(self, 'status_label'):
            original_text = self.status_label.cget('text')
            self.status_label.config(text=f"âš ï¸ {msg}", foreground='orange')
            # Ripristina dopo 10 secondi
            self.after(10000, lambda: self.status_label.config(
                text=original_text, foreground='black'))

    def _stop_product_check_background_task(self):
        """Ferma il thread di controllo prodotti (chiamato alla chiusura app)"""
        if self._product_check_thread and self._product_check_thread.is_alive():
            logger.info("Arresto background task controllo prodotti...")
            self._product_check_stop_event.set()
            self._product_check_thread.join(timeout=5)
        
        # Ferma anche il thread del report mensile
        if self._monthly_report_thread and self._monthly_report_thread.is_alive():
            logger.info("Arresto background task report mensile...")
            self._monthly_report_stop_event.set()
            self._monthly_report_thread.join(timeout=5)

        # Ferma anche il thread email settimanale NPI
        if self._weekly_npi_email_thread and self._weekly_npi_email_thread.is_alive():
            logger.info("Arresto background task email settimanale NPI...")
            self._weekly_npi_email_stop_event.set()
            self._weekly_npi_email_thread.join(timeout=5)

        # Ferma anche il servizio notifiche automatiche NPI
        try:
            from npi.npi_auto_notifications import stop_notification_service
            stop_notification_service()
        except Exception as e:
            logger.error(f"Errore arresto servizio notifiche automatiche NPI: {e}", exc_info=True)

    def _check_and_notify_pending_verifications(self):
        """Controlla se ci sono verifiche pendenti e notifica l'utente"""
        try:
            pending = self.db.fetch_products_must_check()

            if pending:
                count = len(pending)
                # Mostra notifica nella UI (thread-safe)
                self.after(0, lambda: self._show_verification_notification(count))
        except Exception as e:
            logger.error(f"Errore controllo verifiche pendenti: {e}")

    def _open_product_checks_management(self):
        """Apre la finestra di gestione periodicitÃ  verifiche prodotti"""
        import product_checks_gui
        product_checks_gui.ProductChecksManagementWindow(
            self, self.db, self.lang, self._temp_authorized_user_id
        )

    def _open_check_tasks_management(self):
        """Apre la finestra di gestione task di verifica"""
        import product_checks_gui
        product_checks_gui.CheckTasksManagementWindow(
            self, self.db, self.lang, self._temp_authorized_user_id
        )

    def _open_product_verification(self, user_id=None):
        """Apre la finestra di esecuzione verifiche prodotti"""
        import product_checks_gui
        # Se user_id non Ã¨ passato, usa _temp_authorized_user_id (fallback per vecchie chiamate)
        if user_id is None:
            user_id = getattr(self, '_temp_authorized_user_id', None)
        product_checks_gui.ProductVerificationWindow(
            self, self.db, self.lang, user_id
        )
    
    def open_product_verification_with_login(self):
        """Richiede il login e poi apre la finestra di verifica prodotti"""
        self._execute_simple_login(
            action_callback=lambda user_id: self._open_product_verification(user_id)
        )


    def _start_product_check_routine(self):
        """Avvia la routine periodica di controllo prodotti"""

    def open_coating_reports_direct(self):
        """Apre i report coating SENZA login"""
        try:
            from coating_gui import show_coating_reports_direct
            show_coating_reports_direct()
        except Exception as e:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"{self.lang.get('error_opening_coating_reports', 'Errore apertura report coating')}: {str(e)}"
            )

    def _schedule_kanban_refill_check(self, manual=False):
        """
        Pianifica il controllo refill Kanban su base configurabile (JSON stampante).
        """
        # Verifica se Ã¨ giorno lavorativo
        if not should_send_notification(country_code='IT'):
            logger.info("Report non inviato: oggi non Ã¨ un giorno lavorativo")
            return
        try:
            cfg = load_printer_config() or {}
            enabled = bool(cfg.get("kanban_refill_enabled", True))
            minutes = int(cfg.get("kanban_refill_check_minutes", 60))
        except Exception:
            enabled = True
            minutes = 60

        if not enabled or minutes <= 0:
            return

        # Esegui subito la prima volta in background
        self._kanban_refill_check_async(manual=manual)

        # Pianifica ripetizione
        interval_ms = max(60_000, minutes * 60 * 1000)
        # Salva l'id per eventuale cancel
        self.kanban_refill_job_id = self.after(interval_ms, self._schedule_kanban_refill_check)

    def _kanban_refill_check_async(self, manual=False):
        """
        Avvia il controllo in un thread per non bloccare la UI.
        """
        import threading
        threading.Thread(
            target=self._kanban_refill_check_worker,
            kwargs={'manual': manual},
            name="KanbanRefillCheck",
            daemon=True).start()

    def _kanban_refill_check_worker(self, manual=False):
        """
        Logica di controllo Kanban refill
        """
        log = logging.getLogger("TraceabilityRS")
        try:
            # Verifica preliminare della connessione
            if not self.db._ensure_connection():
                log.error("Connessione al database non disponibile per Kanban refill check")
                if manual:
                    self.after(0, lambda: messagebox.showerror(
                        self.lang.get('error_title', 'Errore'),
                        "Connessione al database non disponibile. Riprovare piÃ¹ tardi."
                    ))
                return
        
            # 1. Stock corrente per componente
            log.info("Recupero stock corrente...")
            stock_map = self.db.fetch_kanban_current_stock_by_component()
        
            # Gestione degli errori
            if stock_map is None:  # Errore di connessione o query fallita
                log.error("Errore nel recupero stock - connessione persa o query fallita")
                if manual:
                    self.after(0, lambda: messagebox.showerror(
                        self.lang.get('error_title', 'Errore'),
                        "Errore di connessione durante il recupero dei dati stock."
                    ))
                return
        
            if not stock_map:  # Nessun dato ma senza errori (tabella vuota)
                log.warning("Nessun dato stock recuperato (tabella vuota o tutti i componenti hanno DateOut)")
                if manual:
                    self.after(0, lambda: messagebox.showinfo(
                        self.lang.get('info_title', 'Informazione'),
                        "Nessun componente trovato con stock attivo (tutti i componenti hanno DateOut compilato)."
                    ))
                return
        
            comp_ids = list(stock_map.keys())
            log.info(f"Componenti da processare: {len(comp_ids)}")
        
            # 2. Regole attive per componente
            log.info("Recupero regole attive...")
            rules_map = self.db.fetch_active_rules_by_component()
            if rules_map is None:  # Errore di connessione
                log.error("Errore nel recupero regole attive")
                return
            log.info(f"Regole attive trovate: {len(rules_map)}")
        
            #3. Prima quantitÃ  (per chi ha regola percentuale)
            pct_comp_ids = [cid for cid, r in rules_map.items() if r.get('min_pct') is not None]
            first_qty_map = {}
            if pct_comp_ids:
                log.info(f"Recupero prime quantitÃ  per {len(pct_comp_ids)} componenti...")
                first_qty_map = self.db.fetch_first_load_qty_by_component(pct_comp_ids)
                if first_qty_map is None:  # Errore di connessione
                    log.error("Errore nel recupero prime quantitÃ ")
                    return
                log.info(f"Prime quantitÃ  recuperate: {len(first_qty_map)}")
        
            # 4. Max singolo carico + record id
            log.info("Recupero max carichi...")
            max_load_map = self.db.fetch_max_single_load_by_component(comp_ids)
            if max_load_map is None:  # Errore di connessione
                log.error("Errore nel recupero max carichi")
                return
            log.info(f"Max carichi recuperati: {len(max_load_map)}")
        
            # 5. Master component
            log.info("Recupero master component...")
            master_map = self.db.fetch_components_master(comp_ids)
            if master_map is None:  # Errore di connessione
                log.error("Errore nel recupero master component")
                return
            log.info(f"Master component recuperati: {len(master_map)}")
        
            # 6. Valuta richieste
            requests = []  # elementi: dict con info per Excel e per insert
            for cid, cur_stock in stock_map.items():
                rule = rules_map.get(cid)
                if rule:
                    if rule.get('min_qty') is not None:
                        threshold = int(rule['min_qty'])
                        rule_type = "ABS"
                        rule_value = threshold
                    else:
                        # percentuale: calcolo su prima quantitÃ 
                        base = int(first_qty_map.get(cid, 0))
                        pct = int(rule.get('min_pct') or 0)
                        # floor per non anticipare troppo
                        threshold = int((base * pct) / 100) if base > 0 and pct > 0 else 0
                        rule_type = "PCT"
                        rule_value = pct
                else:
                    threshold = 0
                    rule_type = "NA"
                    rule_value = None
        
                if cur_stock <= threshold:
                    ml = max_load_map.get(cid)
                    if not ml:
                        continue
                    qty_to_refill = int(ml['max_qty'])
                    krec_id = int(ml['record_id'])
                    # dedup: se giÃ  presente oggi per questa KanBanRecordId, salta
                    if self.db.has_refill_request_today(krec_id):
                        continue
                    # prepara riga
                    meta = master_map.get(cid, {'code': f"#{cid}", 'desc': ''})
                    requests.append({
                        'component_id': cid,
                        'component_code': meta['code'],
                        'component_desc': meta['desc'],
                        'current_stock': int(cur_stock),
                        'threshold': int(threshold),
                        'rule_type': rule_type,
                        'rule_value': rule_value,
                        'max_single_load': qty_to_refill,
                        'qty_to_refill': qty_to_refill,
                        'kanban_record_id': krec_id
                    })
        
            if not requests:
                if manual:
                    self.after(0, lambda: messagebox.showinfo(
                        self.lang.get('info_title', 'Informazione'),
                        self.lang.get('kanban_refill_none', 'Nessun componente da richiedere.')
                    ))
                return
        
            # 7. Genera Excel in memoria
            excel_bytes = self._build_kanban_refill_excel(requests)
        
            # 8. Destinatari mail
            try:
                recipients = utils.get_email_recipients(self.db.conn, attribute='Sys_email_KanBanRefill')
            except Exception as e:
                log.error("KanbanRefill: error reading recipients: %s", e)
                recipients = []
        
            if not recipients:
                log.warning("KanbanRefill: no recipients for Sys_email_KanBanRefill; skipping email.")
                return
        
            # 9. Invia email (con allegato Excel)
            subject = "[Kanban] Refill request - Repair Kanban"
            body = (
                "Dear Colleagues,\n\n"
                "Please find attached the refill request for dedicated materials of the repairers' KANBAN.\n"
                "The list includes all components whose current stock is at or below the configured reorder point.\n\n"
                "Regards,\nTraceability System"
            )
        
            try:
                # Salva Excel in file temporaneo
                import tempfile
                import os
                from datetime import datetime
                
                temp_dir = tempfile.gettempdir()
                excel_temp_file = os.path.join(temp_dir, f"KanbanRefill_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                
                with open(excel_temp_file, 'wb') as f:
                    f.write(excel_bytes)
                
                # Invia email con allegato
                success = utils.send_email(
                    recipients=recipients,
                    subject=subject,
                    body=body,
                    attachments=[excel_temp_file]
                )
                
                if success:
                    log.info("KanbanRefill: email sent to %d recipients.", len(recipients))
                else:
                    log.error("KanbanRefill: email send returned False")
                
                # Pulisci file temporaneo
                try:
                    os.remove(excel_temp_file)
                except:
                    pass
                    
            except Exception as e:
                log.error("KanbanRefill: email send failed: %s", e)
                return
        
            # 10. Registra richieste su DB
            for req in requests:
                ok = self.db.insert_refill_request(req['kanban_record_id'], req['qty_to_refill'])
                if not ok:
                    log.warning("KanbanRefill: insert failed for KanBanRecordId=%s: %s",
                                req['kanban_record_id'], self.db.last_error_details)
        
        except Exception as e:
            logging.getLogger("TraceabilityRS").exception("KanbanRefill job failed: %s", e)
            log.exception("KanbanRefill job failed completamente: %s", e)

    def _build_kanban_refill_excel(self, rows: list[dict]) -> bytes:
        """
        Crea un Excel ben formattato in memoria e ritorna i bytes.
        Colonne: Component, Description, Current stock, Threshold, Rule type, Rule value, Max single load, Qty requested.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kanban Refill"

        headers = ["Component", "Description", "Current stock", "Threshold",
                   "Rule type", "Rule value", "Max single load", "Qty requested"]
        ws.append(headers)

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style='thin', color='B7CCE1')
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = header_border

        # Rows with zebra stripes and borders
        alt_fill = PatternFill(fill_type="solid", start_color="EEF3F7", end_color="EEF3F7")
        data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for i, r in enumerate(rows, start=2):
            values = [
                r['component_code'],
                r['component_desc'],
                r['current_stock'],
                r['threshold'],
                r['rule_type'],
                r['rule_value'] if r['rule_type'] == 'PCT' else (r['rule_value'] if r['rule_type'] == 'ABS' else ""),
                r['max_single_load'],
                r['qty_to_refill']
            ]
            ws.append(values)
            if i % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=i, column=c).fill = alt_fill
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).border = data_border

        # Autofit columns
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                v = r[0].value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

        # Filters, freeze
        ws.auto_filter.ref = f"A1:H{ws.max_row}"
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        logger.info('File xls per richiesta materiali di kanban creato')
        return buf.getvalue()

    def open_kanban_move(self):
        ok = self._execute_authorized_action(
            menu_translation_key='kanban_move',
            action_callback=lambda: KanbanMoveForm(self, self.db, self.lang)
        )
        if not ok:
            return

    def open_kanban_materials_report(self):
        import os
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from tkinter import messagebox

        rows = self.db.fetch_components_locations_report()
        if not rows:
            messagebox.showinfo(self.lang.get('info_title', 'Informazione'),
                                self.lang.get('materials_report_no_data', 'Nessun dato da esportare.'))
            return

        # Workbook e foglio
        wb = Workbook()
        ws = wb.active
        ws.title = self.lang.get('materials_report_title', 'Materiali - Report')

        # Intestazioni con filtro e colore azzurro tenue
        headers = [
            self.lang.get('col_component_code', 'Codice componente'),
            self.lang.get('col_description', 'Descrizione'),
            self.lang.get('col_location', 'Locazione'),
            self.lang.get('col_area', 'Area')
        ]
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')  # azzurro tenue
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='B7CCE1')
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        # Dati
        for r in rows:
            ws.append([
                getattr(r, 'ComponentCode', ''),
                getattr(r, 'ComponentDescription', '') or '',
                getattr(r, 'LocationCode', ''),
                getattr(r, 'KanBanLocation', '')
            ])

        # Auto filtro e freeze pane
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

        # Zebra stripes leggere e bordi sottili
        alt_fill = PatternFill(fill_type='solid', start_color='F7FBFF', end_color='F7FBFF')
        data_border = Border(left=Side(style='thin', color='D9D9D9'),
                             right=Side(style='thin', color='D9D9D9'),
                             top=Side(style='thin', color='D9D9D9'),
                             bottom=Side(style='thin', color='D9D9D9'))
        for r in range(2, ws.max_row + 1):
            if r % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = alt_fill
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = data_border

        # Larghezza colonne auto-fit semplice
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1]))
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

            # Salva in C:\Temp e chiedi se aprire il file
            try:
                target_dir = r"C:\Temp"
                os.makedirs(target_dir, exist_ok=True)
                filename = f"KanBan_Materiali_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                fullpath = os.path.join(target_dir, filename)
                wb.save(fullpath)

                prompt = self.lang.get('materials_report_saved_open',
                                       'Report salvato in: {path}\nAprirlo ora?').format(path=fullpath)
                if messagebox.askyesno(self.lang.get('info_title', 'Informazione'), prompt):
                    try:
                        if hasattr(os, 'startfile'):  # Windows
                            os.startfile(fullpath)
                        else:
                            import subprocess, sys
                            if sys.platform == 'darwin':  # macOS
                                subprocess.Popen(['open', fullpath])
                            else:  # Linux
                                subprocess.Popen(['xdg-open', fullpath])
                    except Exception as e:
                        messagebox.showerror(self.lang.get('error_title', 'Errore'),
                                             self.lang.get('materials_report_open_error',
                                                           f'Impossibile aprire il file: {e}'))
            except Exception as e:
                messagebox.showerror(self.lang.get('error_title', 'Errore'),
                                     self.lang.get('materials_report_error',
                                                   f'Errore durante la creazione del report: {e}'))

    def open_kanban_rules_management(self):
        KanbanRulesManagementForm(self, self.db, self.lang)

    def open_printer_setup(self):
        # Apre la finestra di configurazione stampante.
        # Richiede che PrinterSetupDialog sia importata/definita.
        dlg = PrinterSetupDialog(self, self.lang)
        self.wait_window(dlg)
        if getattr(dlg, "result", None):
            # Se vuoi tenere un cache della config
            self.printer_cfg = dlg.result
            # TODO: notifica eventuali form aperte che usano la stampante
            # es. if hasattr(self, "labels_form") and self.labels_form.winfo_exists(): ...
            pass


    # Locazioni
    def open_kanban_locations_create(self):
        KanbanLocationCreateForm(self, self.db, self.lang)

    def open_kanban_locations_modify(self):
        KanbanLocationModifyForm(self, self.db, self.lang)

    def open_kanban_locations_edit(self):
        self._not_implemented('KanBan - Locazioni', 'Modifica')

    def open_kanban_locations_labels(self):
        KanbanLocationLabelsForm(self, self.db, self.lang)

    # Materiali
    def open_kanban_materials_management(self):
        ok = self._execute_authorized_action(
            menu_translation_key='kanban_move',
            action_callback=lambda: KanbanMaterialsManagementForm(self, self.db, self.lang)
        )
        if not ok:
            return

    def open_kanban_load(self):
        """Apre la maschera per il caricamento materiali KanBan (manuale o da Excel) con autorizzazione."""
        logger.info("open_kanban_load chiamata - richiesta apertura form caricamento KanBan")
        ok = self._execute_authorized_action(
            menu_translation_key='kanban_load',
            action_callback=lambda: KanbanMoveForm(self, self.db, self.lang)
        )
        if not ok:
            logger.warning("open_kanban_load - autorizzazione negata o annullata")
            return
        logger.info("open_kanban_load - form aperta con successo")

    # Alias per compatibilitÃ  con diverse configurazioni menu
    def open_kanban_refill(self):
        """Alias per open_kanban_load"""
        return self.open_kanban_load()
    
    def open_kanban_reload(self):
        """Alias per open_kanban_load"""
        return self.open_kanban_load()
    
    def kanban_load(self):
        """Alias per open_kanban_load"""
        return self.open_kanban_load()

    def open_kanban_manage(self):
        self._not_implemented('KanBan', 'Gestione')

    def open_assign_submissions_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='submenu_assign',
            action_callback=lambda: assign_submissions_gui.open_assign_submissions(self, self.db, self.lang)
        )

    def open_scrap_declaration_with_login(self):
        """Apre la finestra per la dichiarazione scarti con autenticazione e autorizzazione."""

        self._execute_authorized_action(
            menu_translation_key='submenu_scrap_declaration',
            action_callback=lambda:scarti_gui.open_scrap_declaration_window(self, self.db, self.lang)
        )

    def open_fai_template_manager_with_login(self):
        """Apre la finestra di gestione template FAI con autenticazione e autorizzazione."""
        self._execute_authorized_action(
            menu_translation_key='gestione_template_fai',
            action_callback=lambda: self._open_fai_template_manager()
        )

    def _open_fai_template_manager(self):
        """Apre la finestra di gestione template FAI."""
        try:
            import fai_template_manager
            fai_template_manager.open_fai_template_manager(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        except Exception as e:
            logger.error(f"Errore apertura gestione template FAI: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire la gestione template FAI:\n{e}"
            )

    def open_line_validations_with_login(self):
        """Apre la finestra validazioni linea con autenticazione e autorizzazione."""
        self._execute_authorized_action(
            menu_translation_key='validazione_line',
            action_callback=lambda: self._open_line_validations()
        )

    def _open_line_validations(self):
        """Apre la finestra validazioni linea."""
        try:
            import line_validation_gui
            line_validation_gui.open_line_validation_window(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        except Exception as e:
            logger.error(f"Errore apertura validazioni linea: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire le validazioni:\n{e}"
            )
    
    def open_fai_reports_viewer_with_login(self):
        """Apre il viewer storico validazioni FAI con autenticazione"""
        self._execute_simple_login(
            action_callback=lambda user_id: self._open_fai_reports_viewer(user_id)
        )
    
    def _open_fai_reports_viewer(self, user_name):
        """Apre il viewer storico validazioni FAI"""
        try:
            from fai_reports_viewer import FaiReportsViewerWindow
            FaiReportsViewerWindow(
                self, self.db, self.lang, user_name
            )
        except Exception as e:
            logger.error(f"Errore apertura viewer FAI: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire lo storico validazioni:\n{e}"
            )
    
    def open_fai_fails_report_with_login(self):
        """Apre il rapporto FAI fails con autenticazione"""
        self._execute_simple_login(
            action_callback=lambda user_id: self._open_fai_fails_report(user_id)
        )
    
    def _open_fai_fails_report(self, user_name):
        """Apre il rapporto FAI fails"""
        try:
            from fai_fails_report import FaiFailsReportWindow
            FaiFailsReportWindow(
                self, self.db, self.lang, user_name
            )
        except Exception as e:
            logger.error(f"Errore apertura rapporto FAI fails: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire il rapporto FAI fails:\n{e}"
            )
    
    def send_fai_fails_email_with_login(self):
        """Invia email automatica FAI fails con autenticazione"""
        self._execute_simple_login(
            action_callback=lambda user_id: self._send_fai_fails_email(user_id)
        )
    
    def _send_fai_fails_email(self, user_name):
        """Invia email automatica per FAI fails non analizzati"""
        try:
            import utils
            
            # Mostra messaggio di attesa
            self.status_label.config(text="Invio email FAI fails in corso...")
            self.update_idletasks()
            
            # Invia email
            success = utils.send_fai_fails_notification(self.db.conn, logo_path="logo.png")
            
            if success:
                messagebox.showinfo(
                    "Successo", 
                    "Email FAI fails inviata con successo!\n\n"
                    "I record sono stati marcati come analizzati."
                )
                logger.info(f"Email FAI fails inviata da {user_name}")
            else:
                messagebox.showwarning(
                    "Nessun invio", 
                    "Nessun FAI fail non analizzato trovato.\n\n"
                    "L'email non è stata inviata."
                )
            
            self.status_label.config(text="Pronto")
            
        except Exception as e:
            logger.error(f"Errore invio email FAI fails: {e}", exc_info=True)
            messagebox.showerror(
                "Errore", 
                f"Errore durante l'invio dell'email FAI fails:\n{e}"
            )
            self.status_label.config(text="Errore invio email")


    def open_calibrations_manager_with_login(self):
        logger = logging.getLogger("TraceabilityRS")
        required = 'calibration_management'
        logger.info("Request to open CalibrationsWindow; required_permission=%r", required)

        # Chiama il login in modalitÃ  "gatekeeper"
        user = self._execute_simple_login(required_permission=required)
        logger.info("Login result for calibrations: %s", "OK" if user else "FAILED")

        if user:
            print(f"Login riuscito per l'utente {user.name}. Apertura finestra...")
            try:
                # Salva l'utente corrente nell'istanza principale
                self.current_user = user

                # Crea la finestra senza memorizzarla
                logging.info('Tentativo apertura finestra di calibrazione')
                calibration_window = CalibrationsWindow(self, self.db, self.lang)

                # Focus sulla nuova finestra
                calibration_window.focus_force()

            except Exception as e:
                logging.getLogger("TraceabilityRS").exception("Error opening CalibrationsWindow: %s", e)
                messagebox.showerror("Errore", f"Impossibile aprire il modulo di calibrazione: {e}")

        else:
            print("Login fallito o annullato. Accesso al modulo negato.")

    def open_line_stoppage_report(self):
            """Apre la finestra per generare il report dei fermi linea."""
            LineStoppageReportForm(self, self.db, self.lang)

    def open_verification_association_with_login(self):
        """Apre la verifica associazione dopo il login"""
        self._execute_simple_login(
            action_callback=lambda user_name: self.traceability_manager.open_verification_association(user_name)
        )

    def open_manage_customers_with_login(self):
        """Apre la gestione clienti dopo il login"""
        self._execute_simple_login(
            action_callback=lambda user_name: self.traceability_manager.open_manage_customers(user_name)
        )

    def open_define_products_with_login(self):
        """Apre la definizione prodotti dopo il login"""
        self._execute_simple_login(
            action_callback=lambda user_name: self.traceability_manager.open_define_products(user_name)
        )

    def open_manage_links_with_login(self):
        """Apre la gestione collegamenti dopo il login"""
        self._execute_simple_login(
            action_callback=lambda user_name: self.traceability_manager.open_manage_links()
        )

    def _trigger_update(self, version_info):
        """Lancia l'updater e chiude l'applicazione."""
        source = version_info.MainPath
        destination = os.path.dirname(sys.executable)
        exe_name = os.path.basename(sys.executable)
        updater_path = os.path.join(destination, "updater.exe")

        if not os.path.exists(updater_path):
            messagebox.showerror("Errore Critico", "File updater.exe non trovato! Impossibile aggiornare.", parent=self)
            return

        subprocess.Popen([updater_path, source, destination, exe_name])
        self._on_closing(force_quit=True)  # Chiude l'app senza chiedere conferma

    def _periodic_version_check(self):
        """Controlla periodicamente la presenza di nuove versioni."""
        logger.info("Controllo periodico versione in corso...")

        app_name = os.path.basename(sys.executable)
        version_info = self.db.fetch_latest_version_info(app_name)

        # Se trova una nuova versione, mostra il dialogo
        if version_info and is_update_needed(APP_VERSION, version_info.Version):
            dialog = UpdateNotificationDialog(self, self.lang, version_info.Version, APP_VERSION)
            self.wait_window(dialog)

            if dialog.result == 'now':
                self._trigger_update(version_info)
                return  # Interrompe il ciclo di controllo
            elif dialog.result == 'later':
                # Ripianifica il controllo tra 15 minuti
                remind_interval_ms = 15 * 60 * 1000
                self.periodic_check_job_id = self.after(remind_interval_ms, self._periodic_version_check)
                return
            elif dialog.result == 'ignore':
                # Interrompe il controllo per questa sessione
                print("Controllo versione silenziato per questa sessione.")
                return

        # Se non ci sono aggiornamenti, ripianifica il controllo tra 120 minuti
        default_interval_ms = 120 * 60 * 1000
        self.periodic_check_job_id = self.after(default_interval_ms, self._periodic_version_check)

    def open_company_manager_with_login(self):
        """Apre la finestra di gestione delle compagnie dopo il login."""
        self._execute_simple_login(
            action_callback=lambda user_name: tools_gui.open_company_manager(self, self.db, self.lang, user_name)
        )

    def open_brand_manager_with_login(self):
        """Apre la finestra di gestione dei brand dopo il login."""
        self._execute_simple_login(
            action_callback=lambda user_name: maintenance_gui.open_brand_manager(self, self.db, self.lang, user_name)
        )

    def open_task_cycles_manager_with_login(self):
        """Apre la finestra di gestione voci task dopo autenticazione."""
        self._execute_authorized_action(
            'gestione_voci_task',
            lambda: maintenance_gui.open_task_cycles_manager(self, self.db, self.lang, self._temp_authorized_user_id)
        )

    def open_fixture_rules_with_login(self):
        """Apre la finestra di gestione regole fixture dopo autorizzazione."""
        import fixture_gui
        self._execute_authorized_action(
            menu_translation_key='submenu_fixture_rules',
            action_callback=lambda: fixture_gui.open_fixture_rules(self, self.db, self.lang, self._temp_authorized_user_id)
        )

    def open_assign_products_fixture_with_login(self):
        """Apre la finestra di assegnazione prodotti a fixture dopo autorizzazione."""
        import fixture_gui
        self._execute_authorized_action(
            menu_translation_key='submenu_assign_products_fixture',
            action_callback=lambda: fixture_gui.open_assign_products_to_fixtures(
                self, self.db, self.lang, self._temp_authorized_user_id
            )
        )

    def open_missing_action_report(self):
        """Esegue il report Missing Action e lo esporta in Excel."""
        # Chiama direttamente la funzione per generare il report, senza login
        maintenance_gui.generate_missing_action_report(self, self.db, self.lang)

    def open_xls_settings_with_login(self):
        """Apre la finestra per configurare la mappatura dei file Excel."""
        self._execute_authorized_action(  # Usiamo il login con permessi
            menu_translation_key='submenu_shipping_settings',  # Riusiamo questo permesso
            action_callback=lambda: operations_gui.open_xls_settings_window(self, self.db, self.lang, None)
        )

    def open_shipping_settings_with_login(self):
        """Apre la finestra per gestire le impostazioni di spedizione."""
        self._execute_authorized_action(  # Usiamo il login con permessi
            menu_translation_key='submenu_shipping_settings',
            action_callback=lambda: operations_gui.open_shipping_settings_window(self, self.db, self.lang, None)
        )

    def open_shipping_window_with_login(self):
        """Apre la finestra per il caricamento del report spedizioni."""
        self._execute_simple_login(
            action_callback=lambda user_name: operations_gui.open_shipping_report_window(self, self.db, self.lang,
                                                                                         user_name)
        )

    def open_shipping_report_window_with_login(self):
        """Apre la finestra per il caricamento del report spedizioni dopo un login semplice."""
        self._execute_simple_login(
            action_callback=lambda user_name: operations_gui.open_shipping_report_window(self, self.db, self.lang,
                                                                                         user_name)
        )

    def _flash_birthday_message(self, message):
        """Crea un effetto flashing con cambio colore per il messaggio di auguri."""
        # Se il testo Ã¨ attualmente visibile, lo nasconde
        if self.birthday_label.cget("text"):
            self.birthday_label.config(text="")
        else:
            # Altrimenti, lo mostra con un nuovo colore casuale
            random_color = random.choice(self.flash_colors)
            self.birthday_label.config(text=message, foreground=random_color)

        # Ripianifica se stessa dopo 750 millisecondi (circa 3/4 di secondo)
        self.birthday_flash_job_id = self.after(750, lambda: self._flash_birthday_message(message))

    def _check_for_birthdays(self):
        """Controlla i compleanni e avvia l'avviso appropriato (fisso o scorrevole)."""
        # Ferma sempre un eventuale avviso precedente
        self._stop_birthday_display()
        if hasattr(self, 'birthday_label'):
            self.birthday_label.config(text="")

        try:
            # ... (la logica per leggere le impostazioni e trovare i compleanni rimane invariata)
            pre_alert_days = int(self.db.fetch_setting('Sys_BirthDay_Prealler'))
            post_alert_days = int(self.db.fetch_setting('Sys_BirthDay_PostAllert'))
            image_path = os.path.join(
                self.db.fetch_setting('Sys_Pics_Directory'),
                self.db.fetch_setting('Sys_BirthDay_Allert')
            )
        except (ValueError, TypeError, AttributeError):
            return False
        if not image_path or not os.path.isfile(image_path): return False
        today = datetime.now().date()
        employees = self.db.fetch_all_active_employees_birthdays()
        celebrating = []
        for emp in employees:
            bday = emp.EmployeeBirthDate
            bday_this_year = bday.replace(year=today.year)
            if timedelta(days=0) <= bday_this_year - today <= timedelta(days=pre_alert_days):
                celebrating.append(emp)
            elif timedelta(days=0) <= today - bday_this_year <= timedelta(days=post_alert_days):
                celebrating.append(emp)

        # --- LOGICA MODIFICATA ---
        if len(celebrating) == 1:
            # Caso 1: Un solo festeggiato -> Testo lampeggiante
            employee = celebrating[0]
            message = f"LA MULÈšI ANI {employee.EmployeeName.upper()} ({employee.EmployeeSurname.upper()}) !!!"
            self._display_special_image(image_path, message)
            self._flash_birthday_message(message)
            duration_ms = 2 * 60 * 1000
            self.birthday_stop_job_id = self.after(duration_ms, self._stop_birthday_display)
            return True

        elif len(celebrating) > 1:
            # Caso 2: PiÃ¹ festeggiati -> Testo scorrevole
            messages = [f"LA MULÈšI ANI {emp.EmployeeName.upper()} ({emp.EmployeeSurname.upper()}) !!!" for emp in
                        celebrating]
            # Unisce i messaggi con un separatore visivo
            full_message = "    â€¢â€¢â€¢    ".join(messages)

            self._display_special_image(image_path, "Compleanni di Oggi!")  # Un messaggio generico sull'immagine
            self._start_scrolling_message(full_message)
            duration_ms = 2 * 60 * 1000
            self.birthday_stop_job_id = self.after(duration_ms, self._stop_birthday_display)
            return True

        return False

    def _start_scrolling_message(self, message):
        """Prepara e avvia l'animazione della scritta scorrevole."""
        # Aggiunge spazi alla fine per creare un loop fluido
        self.scrolling_text = message + " " * 20
        self.scrolling_position = 0

        # Cambia il colore del testo per una migliore visibilitÃ 
        self.birthday_label.config(foreground="#FFD700")  # Colore oro

        # Avvia il primo frame dell'animazione
        self._scroll_text()

    def _scroll_text(self):
        """Esegue un singolo passo dell'animazione di scrolling."""
        # Crea il testo da visualizzare ruotando la stringa originale
        display_text = self.scrolling_text[self.scrolling_position:] + self.scrolling_text[:self.scrolling_position]
        self.birthday_label.config(text=display_text)

        # Incrementa la posizione per il prossimo frame
        self.scrolling_position += 1
        if self.scrolling_position >= len(self.scrolling_text):
            self.scrolling_position = 0

        # Ripianifica il prossimo frame dopo 150 millisecondi
        self.scrolling_job_id = self.after(150, self._scroll_text)

    def _stop_birthday_display(self):

        if self.birthday_flash_job_id:
            self.after_cancel(self.birthday_flash_job_id)
            self.birthday_flash_job_id = None

        # Ferma il ciclo del testo scorrevole, se attivo
        if self.scrolling_job_id:
            self.after_cancel(self.scrolling_job_id)
            self.scrolling_job_id = None

        # Pulisce l'etichetta degli auguri e ripristina il colore
        if hasattr(self, 'birthday_label'):
            self.birthday_label.config(text="", foreground="black")

        # Rimuove l'immagine speciale e riavvia lo slideshow standard
        self._setup_slideshow()

    def _display_special_image(self, image_path, text):
        """Carica un'immagine, ci scrive sopra del testo e la visualizza."""
        try:
            if self.slideshow_job_id:
                self.after_cancel(self.slideshow_job_id)

            image = Image.open(image_path)
            w, h = self.slideshow_label.winfo_width(), self.slideshow_label.winfo_height()
            if w <= 1 or h <= 1:
                self.after(200, lambda: self._display_special_image(image_path, text))
                return

            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype("arial.ttf", 48)
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            position = ((image.width - text_width) / 2, (image.height - text_height) * 0.85)

            draw.text(position, text, font=font, fill="white", stroke_width=2, stroke_fill="black")
            resized_image = ImageOps.fit(image, (w, h), Image.Resampling.LANCZOS)
            self.slideshow_photo = ImageTk.PhotoImage(resized_image)
            self.slideshow_label.config(image=self.slideshow_photo)
        except Exception as e:
            print(f"ERRORE: Impossibile visualizzare l'immagine speciale: {e}")
            self._setup_slideshow()

    def open_add_interruption_window_with_login(self):
        """Apre la finestra per dichiarare un'interruzione di produzione."""
        self._execute_simple_login(
            action_callback=lambda user_name: operations_gui.open_add_interruption_window(self, self.db, self.lang,
                                                                                          user_name)
        )

    def open_maintenance_times_with_login(self):
        """Apre la finestra per gestire i tempi di manutenzione."""
        self._execute_simple_login(
            action_callback=lambda user_name: tools_gui.open_maintenance_times_manager(self, self.db, self.lang,
                                                                                       user_name)
        )

    def open_settings_email_with_login(self):
        """Apre la finestra di gestione impostazioni email dopo autorizzazione."""
        import settings_gui
        self._execute_authorized_action(
            menu_translation_key='gestisci_email_automatiche',
            action_callback=lambda: settings_gui.open_settings_email(
                self, self.db, self.lang, self._temp_authorized_user_id
            )
        )

    def open_label_print_with_login(self):
        """Apre la finestra di stampa etichette dopo login."""
        import label_printing_gui
        self._execute_simple_login(
            action_callback=lambda user_name: label_printing_gui.open_label_print_window(
                self, self.db, self.lang, user_name
            )
        )

    def open_printer_settings_with_login(self):
        """Apre la finestra impostazioni stampante dopo login."""
        import label_printing_gui
        self._execute_simple_login(
            action_callback=lambda user_name: label_printing_gui.open_printer_settings_window(
                self, self.db, self.lang, user_name
            )
        )

    def _open_label_config_placeholder(self):
        """Apre la finestra di configurazione etichetta."""
        def action():
            import label_printing_gui
            label_printing_gui.open_label_config_window(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        self._execute_authorized_action('organizza_label_setup', action)

    def open_add_interruption_window(self):
        """Apre la finestra per dichiarare un'interruzione di produzione."""
        self._execute_simple_login(
            action_callback=lambda user_name: operations_gui.open_add_interruption_window(self, self.db, self.lang,
                                                                                          user_name)
        )

    def _post_startup_tasks(self):
        """
        Esegue le operazioni post-avvio in modo scaglionato per evitare conflitti
        sulla connessione al database.
        """
        logger.info("Eseguo le operazioni post-avvio...")
        
        # âœ… Operazione 1: Orologio (immediato, non usa DB)
        self._update_clock()
        
        # âœ… Operazione 2: Controllo compleanni (immediato, operazione veloce)
        logger.info('Avviato controllo compleanni')
        is_birthday = self._check_for_birthdays()
        if not is_birthday:
            self._setup_slideshow()
        
        # â±ï¸ SCAGLIONAMENTO OPERAZIONI IN BACKGROUND
        # Ogni operazione viene ritardata progressivamente per evitare conflitti DB
        
        # Ritardo 3 secondi: Check versione programma
        self.after(3000, lambda: self._delayed_task(
            'check_versione',
            lambda: self.after(120000, self._periodic_version_check)
        ))
        
        # Ritardo 6 secondi: Controllo calibrazioni
        #self.after(6000, lambda: self._delayed_task(
        #    'controllo_calibrazioni',
        #    self._check_calibration_warnings_startup_async
        #))
        
        # Ritardo 9 secondi: Controllo quantitÃ  kanban
        self.after(9000, lambda: self._delayed_task(
            'controllo_kanban',
            self._schedule_kanban_refill_check
        ))
        
        # Ritardo 12 secondi: Verifica discrepanze
        self.after(12000, lambda: self._delayed_task(
            'verifica_discrepanze',
            self._schedule_verification_check
        ))
    
    def _delayed_task(self, task_name, task_function):
        """
        Esegue un task in background con gestione errori e logging.
        
        Args:
            task_name: Nome del task per il logging
            task_function: Funzione da eseguire
        """
        try:
            logger.info(f'Avviato {task_name}')
            task_function()
        except Exception as e:
            logger.error(f'Errore durante {task_name}: {e}', exc_info=True)

    def _stop_product_check_background_task(self):
        """Ferma/Pulisce il task di verifica discrepanze."""
        if hasattr(self, 'product_check_bg_job') and self.product_check_bg_job:
            try:
                self.after_cancel(self.product_check_bg_job)
            except Exception:
                pass
            self.product_check_bg_job = None

    def _run_verification_check_thread(self):
        """Esegue la logica di verifica in un thread separato."""
        try:
            logger.info("Background Check Thread: Inizio verifica...")
            product_checks_gui.check_and_notify_verification_discrepancies(self.db)
            logger.info("Background Check Thread: Verifica completata.")
        except Exception as e:
            logger.error(f"Errore nel thread di verifica background: {e}")

    def _schedule_verification_check(self):
        """
        Verifica periodicamente se Ã¨ il momento di eseguire il check.
        Viene chiamato ogni minuto per controllare orari configurati.
        """
        try:
            config = self._get_task_config('VerificationDiscrepanciesCheck')
            
            if not config or not config['IsEnabled']:
                logger.debug("Verification check: disabled or not configured")
                self.after(60000, self._schedule_verification_check)
                return
            
            if config['OnlyWorkdays'] and not self._is_workday():
                logger.debug("Verification check: skipped (not a workday)")
                self.after(60000, self._schedule_verification_check)
                return
            
            current_time = datetime.now().strftime('%H:%M')
            execution_times = [t.strip() for t in config['ExecutionTimes'].split(',')]
            
            if current_time in execution_times:
                last_exec = config.get('LastExecutionDate')
                if last_exec:
                    last_hour = last_exec.hour
                    current_hour = datetime.now().hour
                    if last_hour == current_hour:
                        logger.debug(f"Verification check: already executed at {current_time}")
                        self.after(60000, self._schedule_verification_check)
                        return
                
                logger.info(f"Verification check: executing scheduled check at {current_time}")
                self._execute_verification_check()
            
            self.after(60000, self._schedule_verification_check)
            
        except Exception as e:
            logger.error(f"Errore scheduling verification check: {e}", exc_info=True)
            self.after(60000, self._schedule_verification_check)

    def _get_task_config(self, task_name):
        """Recupera configurazione task da database."""
        query = """
        SELECT TaskName, IsEnabled, ExecutionTimes, OnlyWorkdays, 
               LastExecutionDate, LastExecutionStatus
        FROM Employee.[dbo].[ScheduledTasksConfig]
        WHERE TaskName = ?
        """
        try:
            result = self.db.fetch_one(query, (task_name,))
            if result:
                return {
                    'TaskName': result[0],
                    'IsEnabled': bool(result[1]),
                    'ExecutionTimes': result[2],
                    'OnlyWorkdays': bool(result[3]),
                    'LastExecutionDate': result[4],
                    'LastExecutionStatus': result[5]
                }
        except Exception as e:
            logger.error(f"Error fetching task config: {e}")
        return None

    def _is_workday(self):
        """
        Verifica se oggi Ã¨ un giorno lavorativo usando UF_GetWorkingDay.
        La funzione ritorna il prossimo giorno lavorativo.
        Se la data ritornata == oggi, allora oggi Ã¨ lavorativo.
        """
        query = "SELECT employee.[dbo].[UF_GetWorkingDay](?)"
        try:
            today = datetime.now().date()
            result = self.db.fetch_one(query, (today,))
            if result and result[0]:
                next_workday = result[0]
                # Se Ã¨ un datetime, converti a date
                if isinstance(next_workday, datetime):
                    next_workday = next_workday.date()
                # Se il prossimo giorno lavorativo Ã¨ oggi, allora oggi Ã¨ lavorativo
                return next_workday == today
        except Exception as e:
            logger.error(f"Error checking workday: {e}", exc_info=True)
            # In caso di errore, assume sia lavorativo per sicurezza
            return True
        return False

    def _execute_verification_check(self):
        """Esegue il check in un thread separato e aggiorna lo stato."""
        def run_check():
            try:
                logger.info("Scheduled Verification Check: Starting...")
                product_checks_gui.check_and_notify_verification_discrepancies(self.db)
                
                self._update_task_execution_status(
                    'VerificationDiscrepanciesCheck',
                    'Success',
                    'Check completed successfully'
                )
                logger.info("Scheduled Verification Check: Completed successfully")
                
            except Exception as e:
                logger.error(f"Scheduled Verification Check: Error - {e}", exc_info=True)
                self._update_task_execution_status(
                    'VerificationDiscrepanciesCheck',
                    'Error',
                    str(e)
                )
        
        threading.Thread(target=run_check, daemon=True).start()

    def _update_task_execution_status(self, task_name, status, message):
        """Aggiorna lo stato dell'ultima esecuzione."""
        query = """
        UPDATE Employee.[dbo].[ScheduledTasksConfig]
        SET LastExecutionDate = GETDATE(),
            LastExecutionStatus = ?,
            LastExecutionMessage = ?,
            ModifiedDate = GETDATE()
        WHERE TaskName = ?
        """
        try:
            self.db.execute_query(query, (status, message, task_name))
            logger.info(f"Task {task_name} status updated: {status}")
        except Exception as e:
            logger.error(f"Error updating task status: {e}")

    def _setup_slideshow(self):
        """Legge le impostazioni e avvia il ciclo dello slideshow."""
        folder_path = self.db.fetch_setting('SlideshowFolderPath')
        interval_min_str = self.db.fetch_setting('SlideshowIntervalMinutes')

        if not folder_path or not os.path.isdir(folder_path):
            self.slideshow_label.config(text="Percorso immagini non configurato o non valido.", foreground="white")
            return

        try:
            interval_min = int(interval_min_str)
            self.slideshow_interval_ms = interval_min * 60 * 1000
        except (ValueError, TypeError):
            pass

        valid_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
        try:
            all_images = [
                f for f in os.listdir(folder_path)
                if f.lower().endswith(valid_extensions)
            ]
            
            # --- Seasonal Logic ---
            from datetime import date
            current_date = datetime.now().date()
            
            def get_int(key):
                val = self.db.fetch_setting(key)
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return 0

            def get_list(key):
                val = self.db.fetch_setting(key)
                if not val: return []
                return [x.strip() for x in val.split(';') if x.strip()]

            special_images_to_add = []
            special_images_set = set()

            # 1. Christmas
            christmas_imgs = get_list('Sys_natale')
            if christmas_imgs:
                for img in christmas_imgs:
                    special_images_set.add(img)
                
                # Check both common typo variations
                christmas_pre = get_int('Sys_natale_preallert') or get_int('Sys_natale_prealler')
                christmas_post = get_int('Sys_natale_postallert')
                
                c_date = date(current_date.year, 12, 25)
                delta = (current_date - c_date).days
                
                if -christmas_pre <= delta <= christmas_post:
                    special_images_to_add.extend(christmas_imgs)

            # 2. New Year
            ny_imgs = get_list('Sys_HappyNewYear')
            if ny_imgs:
                for img in ny_imgs:
                    special_images_set.add(img)
                # Also ensure Sys_nuovoanno images are excluded normally (Rule 2)
                for extra_ny in get_list('Sys_nuovoanno'):
                    special_images_set.add(extra_ny)
                
                ny_pre = get_int('Sys_HappyNewYear_Preallert')
                ny_post = get_int('Sys_HappyNewYear_PostAllert')
                
                ny_date_curr = date(current_date.year, 1, 1)
                delta_curr = (current_date - ny_date_curr).days
                ny_date_next = date(current_date.year + 1, 1, 1)
                delta_next = (current_date - ny_date_next).days
                
                if (-ny_pre <= delta_curr <= ny_post) or (-ny_pre <= delta_next <= ny_post):
                    special_images_to_add.extend(ny_imgs)

            # 3. Easter (Pasqua)
            easter_imgs = get_list('Sys_pasqua') + get_list('Sys_HappyEastern')
            if easter_imgs:
                for img in easter_imgs:
                    special_images_set.add(img)
                
                # Fetch Easter date using Sys_eastern_data query from DB
                easter_query_tmpl = self.db.fetch_setting('Sys_eastern_data')
                if easter_query_tmpl:
                    try:
                        # Determine religion (heuristic: if no specific setting, derive from language)
                        religion = self.db.fetch_setting('Sys_Religion')
                        if not religion:
                            religion = 'Ortodossa' if (self.lang.current_language.lower() == 'ro') else 'Cattolica'
                        
                        row = self.db.fetch_one(easter_query_tmpl, (religion, 'Pasqua'))
                        if row and row[0]:
                            e_date = row[0]
                            if hasattr(e_date, 'date'): e_date = e_date.date()
                            
                            # Using keys as found in database settings (with typos)
                            easter_pre = get_int('Sys_HappyEastern_Preallert')
                            easter_post = get_int('Sys_HappyEasterh_PostAllert')
                            
                            delta = (current_date - e_date).days
                            if -easter_pre <= delta <= easter_post:
                                logger.info(f"SlideShow: Easter period ACTIVE ({religion})")
                                special_images_to_add.extend(easter_imgs)
                    except Exception as e:
                        logger.error(f"Error during Easter slideshow check: {e}")

            # --- Final image list construction ---
            if special_images_to_add:
                # Rule 1: During special holiday periods, show ONLY special images
                self.image_files = []
                for _ in range(10): # Multiply for frequency
                    for img in special_images_to_add:
                        if img in all_images:
                            self.image_files.append(os.path.join(folder_path, img))
            else:
                # Rule 2: During normal periods, exclude all images defined for special periods
                self.image_files = []
                for img in all_images:
                    if img not in special_images_set:
                        self.image_files.append(os.path.join(folder_path, img))


        except Exception as e:
            print(f"Errore durante la lettura della cartella immagini: {e}")

        if not self.image_files:
            self.slideshow_label.config(text="Nessuna immagine trovata nella cartella specificata.", foreground="white")
        else:
            self._cycle_image()

    def _cycle_image(self):
        """Cambia l'indice dell'immagine, la disegna e si riprogramma."""
        if not self.image_files:
            return

        self._draw_current_image()
        self.current_image_index = (self.current_image_index + 1) % len(self.image_files)

        # --- CORREZIONE QUI ---
        # Salva l'ID del prossimo ciclo per poterlo annullare
        self.slideshow_job_id = self.after(self.slideshow_interval_ms, self._cycle_image)

    def _draw_current_image(self, event=None):  # Aggiunto 'event=None' per la compatibilitÃ  con bind
        """Funzione dedicata a disegnare l'immagine corrente alla dimensione corretta."""
        if not self.image_files:
            return

        image_path = self.image_files[self.current_image_index]
        w, h = self.slideshow_label.winfo_width(), self.slideshow_label.winfo_height()

        if w <= 1 or h <= 1:
            return

        try:
            image = Image.open(image_path)
            resized_image = ImageOps.fit(image, (w, h), Image.Resampling.LANCZOS)
            self.slideshow_photo = ImageTk.PhotoImage(resized_image)
            self.slideshow_label.config(image=self.slideshow_photo)
        except Exception as e:
            print(f"Errore nel disegnare l'immagine {image_path}: {e}")

    def _update_clock(self):
        """Aggiorna l'etichetta dell'orologio ogni secondo."""
        # Formato: Giorno/Mese/Anno Ora:Minuti:Secondi
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.clock_label.config(text=now_str)

        # Richiama se stessa dopo 1000 millisecondi (1 secondo)
        self.clock_label.after(1000, self._update_clock)

    def open_add_maintenance_tasks_with_login(self):
        """Richiede il login e poi apre la finestra per aggiungere/gestire i task."""
        # This action modifies data, so it requires a simple login.
        self._execute_simple_login(
            action_callback=lambda user_name: maintenance_gui.open_add_maintenance_tasks(self, self.db, self.lang,
                                                                                         user_name)
        )

    def open_assign_responsibles_with_login(self):
        """Richiede il login con autorizzazione e apre la finestra per assegnare responsabili ai programmi di manutenzione."""
        self._execute_authorized_action(
            menu_translation_key='submenu_assign_responsibles',
            action_callback=lambda: maintenance_gui.open_assign_responsibles(self, self.db, self.lang)
        )

    def open_manage_permissions_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='submenu_permissions',  # Chiave per accedere alla gestione
            action_callback=lambda: permissions_gui.open_manage_permissions_window(self, self.db, self.lang)
        )

    def open_view_permissions_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='submenu_permissions',  # Stessa chiave, per ora
            action_callback=lambda: permissions_gui.open_view_permissions_window(self, self.db, self.lang)
        )

    def open_doc_types_manager_with_login(self):
        """Richiede il login e poi apre la finestra di gestione dei tipi di documento."""
        # Nota: usiamo una nuova chiave 'submenu_doc_types' per un eventuale permesso dedicato
        self._execute_authorized_action(
            menu_translation_key='submenu_doc_types',
            action_callback=lambda: tools_gui.open_doc_types_manager(self, self.db, self.lang)
        )
    
    def open_translations_manager_with_login(self):
        """Richiede il login e poi apre la finestra di gestione traduzioni."""
        self._execute_authorized_action(
            menu_translation_key='manage_translations',
            action_callback=lambda: translations_manager.open_translations_manager(self, self.db, self.lang)
        )


    def _open_general_docs_viewer(self, category_id, category_name):
        """Apre la finestra di visualizzazione dei documenti in modalitÃ  SOLA LETTURA (senza login)."""
        # L'utente non Ã¨ loggato, quindi passiamo None come user_name
        general_docs_gui.open_general_docs_viewer(
            self, self.db, self.lang, category_id, category_name, user_name=None, view_only=True
        )

    def _open_general_docs_viewer_with_login(self, category_id, category_name):
        """Richiede il login e poi apre la finestra di GESTIONE (lettura/scrittura)."""
        self._execute_simple_login(
            action_callback=lambda user_name: general_docs_gui.open_general_docs_viewer(
                self, self.db, self.lang, category_id, category_name, user_name, view_only=False
            )
        )

    def open_insert_form(self):
        """Apre la finestra di inserimento documenti dopo un login semplice."""

        def action(user_name):
            # Crea e mostra la finestra di inserimento
            form = InsertDocumentForm(self, self.db, user_name, self.lang)
            form.grab_set()

        self._execute_simple_login(action_callback=action)

    def open_view_form(self):
        """Apre la finestra per visualizzare i documenti di produzione."""
        view_form = ViewDocumentForm(self, self.db, self.lang)
        view_form.transient(self)
        view_form.grab_set()
        self.wait_window(view_form)

    def _execute_simple_login(self, action_callback=None, required_permission=None):
        import inspect
        # Get the caller information
        caller_frame = inspect.stack()[1]
        caller_function = caller_frame.function
        caller_filename = caller_frame.filename
        caller_lineno = caller_frame.lineno
        
        logger.info("_execute_simple_login called by function '%s' at %s:%d; required_permission=%r", 
                    caller_function, caller_filename, caller_lineno, required_permission)

        login_form = LoginWindow(self, self.db, self.lang)
        self.wait_window(login_form)

        if not login_form.clicked_login:
            logger.info("Login window closed without login.")
            return None

        user_id = login_form.user_id
        logger.info("LoginWindow returned user_id=%r", user_id)

        password = login_form.password
        user = self.db.authenticate_and_get_user_simple(user_id, password)

        if not user:
            logger.info("Authentication failed for user_id=%r", user_id)
            messagebox.showerror(self.lang.get('login_title'),
                                 self.lang.get('login_auth_failed'), parent=self)
            return None

        logger.debug("Authenticated as %r; permissions=%s", user.name,
                     sorted(user.permissions) if hasattr(user, 'permissions') else [])

        # âš ï¸ VERIFICA SCADENZA PASSWORD
        import change_password_gui
        expired, msg = change_password_gui.check_password_expiration(self.db, user_id)
        if expired:
            logger.warning(f"Password scaduta per utente {user_id}: {msg}")
            messagebox.showwarning(
                self.lang.get('password_expired_title', 'Password Scaduta'),
                self.lang.get('password_expired_message', 
                             f'La tua password Ã¨ scaduta.\n{msg}\n\nDevi cambiarla per continuare.'),
                parent=self
            )
            # Forza cambio password
            changed = change_password_gui.open_change_password_window(
                self, self.db, self.lang, user_id=user_id, force_change=True
            )
            if not changed:
                logger.info(f"Utente {user_id} non ha cambiato la password. Accesso negato.")
                return None
            logger.info(f"Utente {user_id} ha cambiato la password con successo.")

        # Permessi commentati - login semplice senza controllo autorizzazioni
        # if required_permission and not user.has_permission(required_permission):
        #     logger.info("Permission check FAILED for %r -> required=%r",
        #                  user.name, required_permission)
        #     messagebox.showwarning(
        #         self.lang.get('access_denied', "Accesso Negato"),
        #         self.lang.get('permission_missing', "Non si dispone delle autorizzazioni per questa operazione."),
        #         parent=self
        #     )
        #     return None

        if isinstance(action_callback, collections.abc.Callable):
            # Passa user_id invece di user.name per compatibilitÃ  con _open_product_verification
            action_callback(user_id)

        return user

    def _execute_authorized_action(self, menu_translation_key, action_callback):
        #logger.debug("_execute_authorized_action called; menu_translation_key=%r", menu_translation_key)
        import inspect
        # Get the caller information
        caller_frame = inspect.stack()[1]
        caller_function = caller_frame.function
        caller_filename = caller_frame.filename
        caller_lineno = caller_frame.lineno
        login_form = LoginWindow(self, self.db, self.lang)
        logger.info("_execute_authorized_action called by function '%s' at %s:%d; menu_translation_key=%r", 
                    caller_function, caller_filename, caller_lineno, menu_translation_key)
        self.wait_window(login_form)

        if not login_form.clicked_login:
            return False

        user_id = login_form.user_id
        logger.debug("LoginWindow returned user_id=%r for authorized action %r", user_id, menu_translation_key)
        password = login_form.password

        auth_result = self.db.authenticate_and_authorize(user_id, password, menu_translation_key)

        if auth_result is None:
            messagebox.showerror(self.lang.get('login_title'), self.lang.get('login_auth_failed'), parent=self)
            logger.debug("User %r authenticated but NOT authorized for %r", user_id, menu_translation_key)
            return False
        elif auth_result.AuthorizedUsedId is None:
            logger.debug("User %r authenticated but NOT authorized for %r", user_id, menu_translation_key)
            messagebox.showwarning(
                self.lang.get('auth_access_denied_title', "Accesso Negato"),
                self.lang.get('auth_access_denied_message',
                              "Non si dispone delle autorizzazioni necessarie per accedere a questa funzione."),
                parent=self
            )
            return False
        else:
            logger.debug("User %r authorized for %r; executing action.", user_id, menu_translation_key)
            
            # âš ï¸ VERIFICA SCADENZA PASSWORD
            import change_password_gui
            expired, msg = change_password_gui.check_password_expiration(self.db, user_id)
            if expired:
                logger.warning(f"Password scaduta per utente {user_id}: {msg}")
                messagebox.showwarning(
                    self.lang.get('password_expired_title', 'Password Scaduta'),
                    self.lang.get('password_expired_message', 
                                 f'La tua password Ã¨ scaduta.\n{msg}\n\nDevi cambiarla per continuare.'),
                    parent=self
                )
                # Forza cambio password
                changed = change_password_gui.open_change_password_window(
                    self, self.db, self.lang, user_id=user_id, force_change=True
                )
                if not changed:
                    logger.info(f"Utente {user_id} non ha cambiato la password. Accesso negato.")
                    return False
                logger.info(f"Utente {user_id} ha cambiato la password con successo.")
            
            try:
                self.last_authenticated_user_name = auth_result.EmployeeName
                self.last_authorized_user_id = auth_result.AuthorizedUsedId  # Salva l'EmployeeHireHistoryId
            except Exception:
                self.last_authenticated_user_name = None
                self.last_authorized_user_id = None

            self._temp_authorized_user_id = user_id

            action_callback()
            return True

    def open_maint_cycles_manager_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='submenu_maint_cycles',
            action_callback=lambda: tools_gui.open_maint_cycles_manager(self, self.db, self.lang)
        )

    def open_paste_configuration_with_login(self):
        """Apre la configurazione Paste con login"""
        def open_paste_config():
            import paste_manager
            paste_manager.open_paste_configuration(
                self, 
                self.db, 
                self.lang, 
                self._temp_authorized_user_id
            )
        
        self._execute_authorized_action(
            menu_translation_key='paste_form',
            action_callback=open_paste_config
        )

    def _open_paste_refrigerators(self):
        """Apre la gestione frigoriferi paste"""
        import paste_manager
        paste_manager.open_paste_refrigerators(
            self,
            self.db,
            self.lang,
            self._temp_authorized_user_id
        )


    def open_new_submission_form(self):
        """Apre la finestra di inserimento nuova segnalazione (senza login)."""
        submissions_gui.open_new_submission_form(self, self.db, self.lang)

    def open_brands_manager_with_login(self):
        """Richiede il login e poi apre la finestra di gestione dei brand."""
        login_form = LoginWindow(self, self.db, self.lang)
        self.wait_window(login_form)
        authenticated_user = login_form.authenticated_user_name
        if authenticated_user:
            tools_gui.open_brands_manager(self, self.db, self.lang)

    def open_suppliers_manager_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='submenu_suppliers',
            action_callback=lambda: tools_gui.open_suppliers_manager(self, self.db, self.lang)
        )

    def open_suppliers_manager(self):
        """Apre la finestra di gestione dei fornitori."""
        login_form = LoginWindow(self, self.db, self.lang)
        self.wait_window(login_form)
        authenticated_user = login_form.authenticated_user_name
        if authenticated_user:
            tools_gui.open_suppliers_manager(self, self.db, self.lang)

    def _save_language_setting(self, lang_code):
        """Salva la lingua corrente nel file di configurazione."""
        try:
            with open("lang.conf", "w") as f:
                f.write(lang_code)
        except Exception as e:
            print(f"Impossibile salvare le impostazioni della lingua: {e}")

    def _load_language_setting(self):
        """Carica la lingua dal file di configurazione, se esiste."""
        try:
            with open("lang.conf", "r") as f:
                return f.read().strip()
        except FileNotFoundError:
            return 'it'  # Lingua di default se il file non esiste
        except Exception as e:
            print(f"Impossibile leggere le impostazioni della lingua: {e}")
            return 'it'  # Ritorna al default in caso di errore

    def open_fill_templates_with_login(self):
        """Apre la finestra per compilare le schede dopo un login semplice."""
        logger.info("open_fill_templates_with_login called")
        self._execute_simple_login(
            action_callback=lambda user_name: maintenance_gui.open_fill_templates(self, self.db, self.lang, user_name)
        )

    def check_version(self):
        """
        Controlla se l'app Ã¨ eseguita dalla sorgente, poi verifica la versione
        e, se necessario, lancia l'updater.
        Permette fino a 3 rinvii dell'update a meno che il campo Must non sia True.
        Restituisce False se l'app deve chiudersi, altrimenti True.
        """
        try:
            app_name = os.path.basename(sys.executable)
            logger.info(f"check_version: Inizio controllo versione per: {app_name}")
            logger.info(f"check_version: Chiamata a fetch_latest_version_info...")
            version_info = self.db.fetch_latest_version_info(app_name)
            logger.info(f"check_version: fetch_latest_version_info completata. Result: {version_info}")

            if not version_info or not version_info.Version or not version_info.MainPath:
                logger.info("Informazioni di versione non trovate o incomplete nel DB. Controllo saltato.")
                return True

            source_path = os.path.normpath(version_info.MainPath)
            current_path = os.path.normpath(os.path.dirname(sys.executable))

            if source_path.lower() == current_path.lower():
                title = self.lang.get("error_running_from_source_title", "Esecuzione non Permessa")
                message = self.lang.get(
                    "error_running_from_source_message",
                    "L'applicazione non puÃ² essere eseguita direttamente dal percorso sorgente sul server.\n\n"
                    "Si prega di lanciare la copia installata localmente."
                )
                messagebox.showerror(title, message, parent=self)
                self.db.disconnect()
                self.destroy()
                self.should_exit = True  # Set the flag
                return False

            if is_update_needed(APP_VERSION, version_info.Version):
                # Recupera il flag Must (default False se non presente)
                is_mandatory = getattr(version_info, 'Must', False)
                
                # Carica il conteggio dei rinvii e l'ultima versione vista
                skip_count, last_version = load_update_skip_count()
                
                # Se la versione disponibile Ã¨ cambiata, resetta il conteggio
                if last_version != version_info.Version:
                    skip_count = 0
                    logger.info(f"Nuova versione disponibile ({version_info.Version}), reset conteggio rinvii")
                
                # Determina se l'update Ã¨ obbligatorio
                # L'update Ã¨ obbligatorio se:
                # 1. Il campo Must Ã¨ True, OPPURE
                # 2. L'utente ha giÃ  saltato l'update 3 volte
                force_update = is_mandatory or skip_count >= 3
                
                if force_update:
                    # Update obbligatorio
                    title = self.lang.get("upgrade_required_title")
                    
                    if is_mandatory:
                        message = self.lang.get(
                            "force_upgrade_message_mandatory",
                            version_info.Version, APP_VERSION
                        )
                    else:
                        message = self.lang.get(
                            "force_upgrade_message_max_skips",
                            version_info.Version, APP_VERSION
                        )
                    
                    messagebox.showinfo(title, message, parent=self)

                    destination = os.path.dirname(sys.executable)
                    exe_name = os.path.basename(sys.executable)
                    updater_path = os.path.join(destination, "updater.exe")

                    if not os.path.exists(updater_path):
                        messagebox.showerror("Errore Critico", "File updater.exe non trovato! Impossibile aggiornare.",
                                             parent=self)
                        self.db.disconnect()
                        self.destroy()
                        self.should_exit = True  # Set the flag
                        return False

                    # Reset del conteggio prima di aggiornare
                    reset_update_skip_count()
                    
                    subprocess.Popen([updater_path, source_path, destination, exe_name])
                    self.db.disconnect()
                    self.destroy()
                    self.should_exit = True  # Set the flag
                    return False
                else:
                    # Update opzionale - chiedi all'utente
                    title = self.lang.get("upgrade_available_title")
                    remaining_skips = 3 - skip_count
                    message = self.lang.get(
                        "optional_upgrade_message",
                        version_info.Version, APP_VERSION, remaining_skips
                    )
                    
                    response = messagebox.askyesno(title, message, parent=self)
                    
                    if response:
                        # L'utente vuole aggiornare
                        destination = os.path.dirname(sys.executable)
                        exe_name = os.path.basename(sys.executable)
                        updater_path = os.path.join(destination, "updater.exe")

                        if not os.path.exists(updater_path):
                            messagebox.showerror("Errore Critico", "File updater.exe non trovato! Impossibile aggiornare.",
                                                 parent=self)
                            self.db.disconnect()
                            self.destroy()
                            self.should_exit = True
                            return False

                        # Reset del conteggio prima di aggiornare
                        reset_update_skip_count()
                        
                        subprocess.Popen([updater_path, source_path, destination, exe_name])
                        self.db.disconnect()
                        self.destroy()
                        self.should_exit = True
                        return False
                    else:
                        # L'utente ha scelto di rinviare
                        skip_count += 1
                        save_update_skip_count(skip_count, version_info.Version)
                        logger.info(f"Update rinviato. Conteggio rinvii: {skip_count}/3")
                        
                        # Mostra un messaggio informativo
                        info_message = self.lang.get(
                            "update_skipped_message",
                            "Aggiornamento rinviato.\n\n"
                            "Rinvii rimanenti: {0}/3",
                            remaining_skips - 1
                        )
                        messagebox.showinfo(
                            self.lang.get("update_skipped_title", "Aggiornamento Rinviato"),
                            info_message,
                            parent=self
                        )
                        return True
            else:
                # Versione aggiornata - reset del conteggio se esiste
                reset_update_skip_count()
                print(f"Versione applicazione ({APP_VERSION}) aggiornata.")
                return True

        except Exception as e:
            logger.error(f"Errore imprevisto durante il controllo versione: {e}", exc_info=True)
            print(f"Errore imprevisto durante il controllo versione: {e}")
            return True


    def _create_widgets(self):
        # --- PRIMA: Crea la Barra di Stato Inferiore ---
        status_bar = ttk.Frame(self, style="Card.TFrame", padding=5)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Contenitore per Logo e Orologio, allineato a destra
        bottom_right_frame = ttk.Frame(status_bar)
        bottom_right_frame.pack(side=tk.RIGHT)

        # Logo (ora in basso a destra)
        if PIL_AVAILABLE:
            try:
                image = Image.open("logo.png")
                image.thumbnail((100, 100))
                self.logo_image = ImageTk.PhotoImage(image)
                self.logo_label = ttk.Label(bottom_right_frame, image=self.logo_image)
                self.logo_label.pack()
            except Exception as e:
                print(f"Errore caricamento logo: {e}")

        # Orologio (ora sotto il logo)
        self.clock_label = ttk.Label(bottom_right_frame, font=("Helvetica", 9), justify=tk.RIGHT)
        self.clock_label.pack()

        # --- NUOVA ETICHETTA PER GLI AUGURI (al centro) ---
        self.birthday_label = ttk.Label(status_bar, text="", font=("Helvetica", 10, "bold"), anchor="center")
        # Questo pack fa sÃ¬ che l'etichetta si espanda per riempire lo spazio centrale
        self.birthday_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- DOPO: Crea l'Area Centrale per lo Slideshow ---
        # Ora questo label si espanderÃ  per riempire tutto lo spazio RIMANENTE
        self.slideshow_label = ttk.Label(self, background="black")
        self.slideshow_label.pack(fill=tk.BOTH, expand=True)

        # Associa il ridimensionamento al disegno dell'immagine
        self.slideshow_label.bind('<Configure>', lambda e: self._draw_current_image())


    def _launch_specific_gantt_window(self, project_id):
        """Lancia la finestra Gantt per un ID progetto specifico."""
        if not project_id:
            return
        logger.info(f"Lancio della finestra Gantt per il progetto ID: {project_id}")
        NpiGanttWindow(
            master=self,
            npi_manager=self.npi_manager,
            lang=self.lang,
            progetto_id=int(project_id)
        )

    def _create_menu(self):
        """Crea la struttura completa dei menu con gerarchia organizzata"""
        self.menubar = tk.Menu(self)
        self.config(menu=self.menubar)

        # Inizializza tutti i menu principali
        self._init_main_menus()

        # Inizializza i sottomenu complessi
        self._init_production_submenus()
        #self._init_npi_submenus()

        #self._init_other_submenus()  # Metodo fittizio per gli altri sottomenu
        self._init_tools_submenus()
        self._init_help_submenus()

        # Aggiungi i menu principali alla barra
        self._add_main_menus_to_bar()

    def _init_main_menus(self):
        """Inizializza i menu principali"""
        self.document_menu = tk.Menu(self.menubar, tearoff=0)
        self.general_docs_menu = tk.Menu(self.menubar, tearoff=0)
        self.operations_menu = tk.Menu(self.menubar, tearoff=0)
        self.maintenance_menu = tk.Menu(self.menubar, tearoff=0)
        self.submissions_menu = tk.Menu(self.menubar, tearoff=0)
        self.tools_menu = tk.Menu(self.menubar, tearoff=0)
        self.help_menu = tk.Menu(self.menubar, tearoff=0)
        self.npi_menu = tk.Menu(self.operations_menu, tearoff=0)
        # Menu Gestione Reclami
        self.complaints_menu = tk.Menu(self.menubar, tearoff=False)
        self.complaints_submenu = tk.Menu(self.complaints_menu, tearoff=False)
        
        # Menu Personale
        self.personnel_menu = tk.Menu(self.operations_menu, tearoff=0)
        self.guests_submenu = tk.Menu(self.personnel_menu, tearoff=0)
        
        # Menu Assenze
        self.absences_submenu = tk.Menu(self.personnel_menu, tearoff=0)
        
        # Menu Ordini
        self.orders_menu = tk.Menu(self.operations_menu, tearoff=0)

    def _init_production_submenus(self):
        """Inizializza la gerarchia completa del menu Produzione"""
        # Menu principale Produzione
        self.production_submenu = tk.Menu(self.operations_menu, tearoff=0)

        # Sottomenu di Produzione
        self.declarations_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.line_validation_submenu = tk.Menu(self.production_submenu, tearoff=0)

        # Gerarchia KanBan
        self.kanban_root_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.kanban_locations_submenu = tk.Menu(self.kanban_root_submenu, tearoff=0)
        self.kanban_materials_submenu = tk.Menu(self.kanban_root_submenu, tearoff=0)
        self.kanban_core_submenu = tk.Menu(self.kanban_root_submenu, tearoff=0)

        # TracciabilitÃ 
        self.traceability_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.fct_transfer_submenu = tk.Menu(self.traceability_submenu, tearoff=0)

        # Calibrazioni
        self.calibrations_submenu = tk.Menu(self.production_submenu, tearoff=0)
        # Aggiungi le voci al menu Calibrazioni
        self.calibrations_submenu.add_command(
            label=self.lang.get('calibration_management', "Gestione Calibrazioni"),
            command=self._open_calibration_management
        )
        self.calibrations_submenu.add_command(
            label=self.lang.get('view_calibration_status', "Visualizza Situazione Calibrazioni"),
            command=self._open_calibration_status
        )
        self.calibrations_submenu.add_separator()
        self.calibrations_submenu.add_command(
            label=self.lang.get('generate_calibration_report', "Genera Report Calibrazioni"),
            command=self._generate_calibration_report
        )

        # Coating - Struttura organizzata
        self.coating_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.coating_materials_submenu = tk.Menu(self.coating_submenu, tearoff=0)

        # Paste - Gestione Paste
        self.paste_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.paste_config_submenu = tk.Menu(self.paste_submenu, tearoff=0)
        self.paste_transfer_submenu = tk.Menu(self.paste_submenu, tearoff=0)

        # Verifiche Prodotti
        self.product_checks_submenu = tk.Menu(self.production_submenu, tearoff=0)

        # Rapporti
        self.reports_submenu = tk.Menu(self.production_submenu, tearoff=0)
        self.operativity_submenu = tk.Menu(self.reports_submenu, tearoff=0)

    def _init_tools_submenus(self):
        """Inizializza i sottomenu di Strumenti"""
        self.permissions_submenu = tk.Menu(self.tools_menu, tearoff=0)
        self.materials_submenu = tk.Menu(self.tools_menu, tearoff=0)
        self.room_booking_submenu = tk.Menu(self.tools_menu, tearoff=0)

    def _init_help_submenus(self):
        """Inizializza i sottomenu di Aiuto"""
        self.language_menu = tk.Menu(self.help_menu, tearoff=0)
        self.language_menu.add_command(label="Italiano", command=lambda: self._change_language('it'))
        self.language_menu.add_command(label="English", command=lambda: self._change_language('en'))
        self.language_menu.add_command(label="RomÃ¢nÄƒ", command=lambda: self._change_language('ro'))
        self.language_menu.add_command(label="Deutsch", command=lambda: self._change_language('de'))
        self.language_menu.add_command(label="Svenska", command=lambda: self._change_language('sv'))

    def _add_main_menus_to_bar(self):
        """Aggiunge i menu principali alla barra dei menu"""
        self.menubar.add_cascade(label=self.lang.get('menu_documents', "Documenti di Produzione"),
                                 menu=self.document_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_general_docs', "Documenti Generali"),
                                 menu=self.general_docs_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_operations', "Operazioni"), menu=self.operations_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_maintenance', "Manutenzione"), menu=self.maintenance_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_submissions', "Segnalazioni"), menu=self.submissions_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_tools', "Strumenti"), menu=self.tools_menu)
        self.menubar.add_cascade(label=self.lang.get('menu_help', "Aiuto"), menu=self.help_menu)

    def update_texts(self):
        """Aggiorna tutti i testi della UI principale con struttura organizzata"""
        self.title(self.lang.get('app_title'))

        # Aggiorna tutti i menu in sequenza
        self._update_document_menu()
        self._update_general_docs_menu()
        self._update_operations_menu()
        self._update_maintenance_menu()
        self._update_submissions_menu()
        self._update_tools_menu()
        self._update_help_menu()

        # Aggiorna le etichette della barra principale
        self._update_main_menubar()

    def _update_document_menu(self):
        """Aggiorna il menu Documenti di Produzione"""
        self.document_menu.delete(0, 'end')
        self.document_menu.add_command(label=self.lang.get('menu_insert_doc'), command=self.open_insert_form)
        self.document_menu.add_command(label=self.lang.get('menu_view_doc'), command=self.open_view_form)
        self.document_menu.add_separator()
        self.document_menu.add_command(label=self.lang.get('menu_quit'), command=self._on_closing)

    def _update_general_docs_menu(self):
        """Aggiorna il menu Documenti Generali"""
        self.general_docs_menu.delete(0, 'end')
        if self.doc_categories:
            for category in self.doc_categories:
                category_submenu = tk.Menu(self.general_docs_menu, tearoff=0)
                category_name = self.lang.get(category.TranslationKey, category.NomeCategoria)
                self.general_docs_menu.add_cascade(label=category_name, menu=category_submenu)

                cmd_edit = lambda cid=category.CategoriaId, cname=category_name: \
                    self._open_general_docs_viewer_with_login(cid, cname)

                category_submenu.add_command(
                    label=self.lang.get('submenu_add_edit', "Aggiungi/Modifica"),
                    command=cmd_edit
                )

                cmd_view = lambda cid=category.CategoriaId, cname=category_name: \
                    self._open_general_docs_viewer(cid, cname)

                category_submenu.add_command(
                    label=self.lang.get('submenu_view', "Visualizza"),
                    command=cmd_view
                )

    def _update_operations_menu(self):
        """Aggiorna il menu Operazioni con tutta la gerarchia Produzione e NPI."""
        self.operations_menu.delete(0, 'end')

        # 1. Popola il sottomenu 'Produzione'
        self.operations_menu.add_cascade(
            label=self.lang.get('submenu_production_ops', "Produzione"),
            menu=self.production_submenu
        )
        self._update_production_submenu()  # Questo popola il menu produzione

        self.operations_menu.add_separator()

        # 2. Popola il sottomenu 'NPI Management'
        self.operations_menu.add_cascade(
            label=self.lang.get('menu_npi_management', 'NPI Management'),
            menu=self.npi_menu
        )
        self.npi_menu.delete(0, 'end')  # Pulisce prima di riempire

        self.operations_menu.add_separator()

        # 3. Popola il sottomenu 'Gestione Reclami'
        self.operations_menu.add_cascade(
            label=self.lang.get('menu_complaints_management', 'Gestione Reclami'),
            menu=self.complaints_menu
        )
        self._update_complaints_menu()

        self.operations_menu.add_separator()

        # 4. Popola il sottomenu 'Personale'
        self.operations_menu.add_cascade(
            label=self.lang.get('menu_personnel', 'Personale'),
            menu=self.personnel_menu
        )
        self.personnel_menu.delete(0, 'end')
        
        self.personnel_menu.add_cascade(
            label=self.lang.get('submenu_guests', 'Ospiti'),
            menu=self.guests_submenu
        )
        self.guests_submenu.delete(0, 'end')
        
        self.guests_submenu.add_command(
            label=self.lang.get('submenu_guest_registration', 'Registrazione Ospiti'),
            command=self.open_guest_registration_with_login
        )
        
        self.guests_submenu.add_command(
            label=self.lang.get('submenu_guest_report', 'Report Ospiti'),
            command=self.generate_guests_pdf_report_with_login #open_guest_report_with_login
        )
        
        # Sottomenu Assenze
        self.personnel_menu.add_separator()
        
        self.personnel_menu.add_cascade(
            label=self.lang.get('submenu_absences', 'Assenze'),
            menu=self.absences_submenu
        )
        self.absences_submenu.delete(0, 'end')
        
        # Autorizzazione Assenze (diretto, senza sottomenu)
        self.absences_submenu.add_command(
            label=self.lang.get('submenu_absence_authorization', 'Autorizzazione Assenze'),
            command=self.open_absence_authorization_with_login
        )
        
        # Regole Assenze
        self.absences_submenu.add_command(
            label=self.lang.get('submenu_absence_rules', 'Regole Assenze'),
            command=self.open_absence_rules_with_login
        )
        
        # Messaggi
        self.personnel_menu.add_separator()
        self.personnel_menu.add_command(
            label=self.lang.get('submenu_news', 'Messaggi'),
            command=self.open_news_management_with_login
        )

        self.operations_menu.add_separator()

        # 5. Popola il menu 'Ordini'
        self.operations_menu.add_cascade(
            label=self.lang.get('menu_orders', 'Ordini'),
            menu=self.orders_menu
        )
        self.orders_menu.delete(0, 'end')
        
        self.orders_menu.add_command(
            label=self.lang.get('submenu_load_orders', 'Carica Ordini'),
            command=self._load_orders_placeholder
        )
        
        
        self.orders_menu.add_command(
            label=self.lang.get('submenu_orders_reports', 'Urgenze'),
            command=self._orders_reports_placeholder
        )

        # Comandi del menu NPI

        self.npi_menu.add_command(
            label=self.lang.get('npi_project_management', 'Gestione Progetti NPI'),
            command=self.open_npi_project_management
        )
        self.npi_menu.add_command(
            label=self.lang.get('npi_dashboard_title', 'Dashboard Progetti...'),
            command=self._apri_dashboard_npi
        )
        self.npi_menu.add_command(
            label=self.lang.get('npi_view_gantt', 'Visualizza Gantt Progetto...'),
            command=self._seleziona_e_visualizza_gantt
        )
        self.npi_menu.add_separator()
        self.npi_menu.add_command(
            label=self.lang.get('npi_setup_tasks', 'Configura Catalogo Task...'),
            command=self._configura_catalogo_task_npi
        )

        # Disabilita tutto se il gestore NPI non è partito
        if self.npi_manager is None:
            # Itera sugli indici del menu per disabilitarli
            for i in range(self.npi_menu.index("end") + 1):
                try:
                    self.npi_menu.entryconfig(i, state="disabled")
                except tk.TclError:
                    pass  # Ignora errori su separatori

        # 6. Menu Materiali
        self.operations_menu.add_separator()
        
        materials_menu = tk.Menu(self.operations_menu, tearoff=0)
        self.operations_menu.add_cascade(
            label=self.lang.get('menu_materials', 'Materiali'),
            menu=materials_menu
        )
        
        # Sottomenu Etichette (voce diretta)
        materials_menu.add_command(
            label=self.lang.get('submenu_labels', 'Etichette'),
            command=self.open_label_print_with_login
        )
        
        # Sottomenu Configurazioni
        materials_config_menu = tk.Menu(materials_menu, tearoff=0)
        materials_menu.add_cascade(
            label=self.lang.get('submenu_configurations', 'Configurazioni'),
            menu=materials_config_menu
        )
        
        # Stampanti sotto Configurazioni
        materials_config_menu.add_command(
            label=self.lang.get('submenu_printer_settings', 'Stampanti'),
            command=self.open_printer_settings_with_login
        )
        
        # Etichetta sotto Configurazioni
        materials_config_menu.add_command(
            label=self.lang.get('submenu_label_config', 'Etichetta'),
            command=self._open_label_config_placeholder
        )


    def _update_production_submenu(self):
        """Aggiorna il sottomenu Produzione con tutte le sezioni"""
        self.production_submenu.delete(0, 'end')

        # 1. Dichiarazioni
        self.production_submenu.add_cascade(label=self.lang.get('submenu_declarations', "Dichiarazioni"),
                                            menu=self.declarations_submenu)
        self._update_declarations_submenu()

        # 2. KanBan
        self.production_submenu.add_cascade(label=self.lang.get('submenu_kanban', 'KanBan'),
                                            menu=self.kanban_root_submenu)
        self._update_kanban_submenus()

        # 3. TracciabilitÃ 
        self.production_submenu.add_cascade(label=self.lang.get('submenu_traceability', "TracciabilitÃ "),
                                            menu=self.traceability_submenu)
        self._update_traceability_submenu()

        # 4. Calibrazioni
        self.production_submenu.add_cascade(label=self.lang.get('submenu_calibrations', "Calibrazioni"),
                                            menu=self.calibrations_submenu)
        self._update_calibrations_submenu()

        # 5. Coating - Struttura organizzata
        self.production_submenu.add_cascade(label=self.lang.get('menu_coating', "Coating"),
                                            menu=self.coating_submenu)
        self._update_coating_submenu()

        # 6. Paste - Gestione Paste
        self.production_submenu.add_cascade(label=self.lang.get('menu_paste', "Paste"),
                                            menu=self.paste_submenu)
        self._update_paste_submenu()

        # 7. Verifiche Prodotti
        self.production_submenu.add_cascade(label=self.lang.get('menu_product_checks', "Verifiche"),
                                            menu=self.product_checks_submenu)
        self._update_product_checks_submenu()

        # 8. Rapporti
        self.production_submenu.add_cascade(label=self.lang.get('submenu_reports_prod', "Rapporti"),
                                            menu=self.reports_submenu)
        self._update_reports_submenu()

    def _update_declarations_submenu(self):
        """Aggiorna il sottomenu Dichiarazioni"""
        self.declarations_submenu.delete(0, 'end')
        self.declarations_submenu.add_command(
            label=self.lang.get('submenu_interruptions', "Interruzioni di produzione"),
            command=self.open_add_interruption_window_with_login
        )
        self.declarations_submenu.add_separator()
        self.declarations_submenu.add_command(
            label=self.lang.get('submenu_scrap_validation', "Validazione scarti"),
            command=self.open_scrap_validation_with_login
        )
        self.declarations_submenu.add_command(
            label=self.lang.get('submenu_scrap_declaration', "Dichiarazione scarti"),
            command=self.open_scrap_declaration_with_login
        )
        self.declarations_submenu.add_separator()
        self.declarations_submenu.add_cascade(
            label=self.lang.get('submenu_line_validation', "Validazione linea"),
            menu=self.line_validation_submenu
        )
        self._update_line_validation_submenu()

    def _update_line_validation_submenu(self):
        """Aggiorna il sottomenu Validazione linea"""
        self.line_validation_submenu.delete(0, 'end')
        self.line_validation_submenu.add_command(
            label=self.lang.get('gestione_template_fai', "Gestione Template"),
            command=self.open_fai_template_manager_with_login
        )
        self.line_validation_submenu.add_separator()
        self.line_validation_submenu.add_command(
            label=self.lang.get('validazione_line', "Validazioni"),
            command=self.open_line_validations_with_login
        )
        self.line_validation_submenu.add_command(
            label=self.lang.get('storico_validazioni_fai', "Storico Validazioni FAI"),
            command=self.open_fai_reports_viewer_with_login
        )
        self.line_validation_submenu.add_command(
            label=self.lang.get('rapporto_fai_fails', "Rapporto FAI fails"),
            command=self.open_fai_fails_report_with_login
        )


    def _update_kanban_submenus(self):
        """Aggiorna tutta la gerarchia KanBan"""
        self.kanban_root_submenu.delete(0, 'end')
        self.kanban_locations_submenu.delete(0, 'end')
        self.kanban_materials_submenu.delete(0, 'end')
        self.kanban_core_submenu.delete(0, 'end')

        # Locazioni KanBan
        self.kanban_locations_submenu.add_command(
            label=self.lang.get('action_create', 'Crea'),
            command=self.open_kanban_locations_create
        )
        self.kanban_locations_submenu.add_command(
            label=self.lang.get('menu_locations_modify', 'Modifica...'),
            command=self.open_kanban_locations_modify
        )
        self.kanban_locations_submenu.add_command(
            label=self.lang.get('action_labels', 'Etichette'),
            command=self.open_kanban_locations_labels
        )
        self.kanban_locations_submenu.add_separator()
        self.kanban_locations_submenu.add_command(
            label=self.lang.get('printer_setup_menu', 'Impostazioni stampante...'),
            command=self.open_printer_setup
        )

        # Materiali KanBan
        self.kanban_materials_submenu.add_command(
            label=self.lang.get('menu_materials_mgmt', 'Materiali - Gestione'),
            command=self.open_kanban_materials_management
        )
        self.kanban_materials_submenu.add_command(
            label=self.lang.get('menu_rules_mgmt', 'Gestione regole'),
            command=self.open_kanban_rules_management
        )
        self.kanban_materials_submenu.add_command(
            label=self.lang.get('submenu_report_single', 'Report'),
            command=self.open_kanban_materials_report
        )

        # Core KanBan
        self.kanban_core_submenu.add_command(
            label=self.lang.get('kanban_move', 'Movimenta'),
            command=self.open_kanban_move
        )
        self.kanban_core_submenu.add_command(
            label=self.lang.get('kanban_load', 'Ricarica (TEST)'),
            command=self.open_kanban_load
        )
        # self.kanban_core_submenu.add_command(
        #     label=self.lang.get('kanban_verify', 'Verifica'),
        #     command=self._schedule_kanban_refill_check
        # )
        self.kanban_core_submenu.add_command(
            label=self.lang.get('submenu_manage', 'Gestione'),
            command=self.open_kanban_manage
        )

        # Assembla gerarchia KanBan
        self.kanban_root_submenu.add_cascade(
            label=self.lang.get('submenu_kanban_locations', 'Locazioni'),
            menu=self.kanban_locations_submenu
        )
        self.kanban_root_submenu.add_cascade(
            label=self.lang.get('submenu_kanban_materials', 'Materiali'),
            menu=self.kanban_materials_submenu
        )
        self.kanban_root_submenu.add_cascade(
            label=self.lang.get('submenu_kanban_core', 'KanBan'),
            menu=self.kanban_core_submenu
        )

    def _update_traceability_submenu(self):
        """Aggiorna il sottomenu TracciabilitÃ """
        self.traceability_submenu.delete(0, 'end')

        # Clienti Finali
        customers_submenu = tk.Menu(self.traceability_submenu, tearoff=0)
        self.traceability_submenu.add_cascade(label=self.lang.get('submenu_final_customers', "Clienti Finali"),
                                              menu=customers_submenu)
        customers_submenu.add_command(label=self.lang.get('submenu_manage_customers', "Gestisci Clienti"),
                                      command=self.open_manage_customers_with_login)

        # Prodotti Finali
        products_submenu = tk.Menu(self.traceability_submenu, tearoff=0)
        self.traceability_submenu.add_cascade(label=self.lang.get('submenu_final_products', "Prodotti Finali"),
                                              menu=products_submenu)
        products_submenu.add_command(label=self.lang.get('submenu_define_products', "Definisci Prodotti"),
                                     command=self.open_define_products_with_login)

        # Collegamenti Prodotti
        links_submenu = tk.Menu(self.traceability_submenu, tearoff=0)
        self.traceability_submenu.add_cascade(label=self.lang.get('submenu_link_products', "Collega Prodotti"),
                                              menu=links_submenu)
        links_submenu.add_command(label=self.lang.get('submenu_manage_links', "Gestisci Collegamenti"),
                                  command=self.open_manage_links_with_login)

        # Verifica Associazione
        self.traceability_submenu.add_command(
            label=self.lang.get('verification_association', "Verifica Associazione"),
            command=self.open_verification_association_with_login
        )

        # FCT Transfer
        self.traceability_submenu.add_cascade(
            label=self.lang.get('menu_fct_transfer', "FCT Transfer"),
            menu=self.fct_transfer_submenu
        )
        self.fct_transfer_submenu.delete(0, 'end')
        self.fct_transfer_submenu.add_command(
            label=self.lang.get('menu_fct_settings', "Settings"),
            command=self._open_fct_settings
        )
        self.fct_transfer_submenu.add_command(
            label=self.lang.get('menu_fct_run', "Run"),
            command=self._toggle_fct_execution
        )

    def _update_calibrations_submenu(self):
        """Aggiorna il sottomenu Calibrazioni con tutte le voci"""
        self.calibrations_submenu.delete(0, 'end')

        # Aggiungi tutte le voci del menu Calibrazioni
        self.calibrations_submenu.add_command(
            label=self.lang.get('calibration_management', "Gestione Calibrazioni"),
           # command=self._open_calibration_management
            command=self.open_calibrations_manager_with_login
        )
        self.calibrations_submenu.add_command(
            label=self.lang.get('view_calibration_status', "Visualizza Situazione Calibrazioni"),
            command=self._open_calibration_status
        )
        self.calibrations_submenu.add_separator()
        self.calibrations_submenu.add_command(
            label=self.lang.get('generate_calibration_report', "Genera Report Calibrazioni"),
            command=self._generate_calibration_report
        )

    def _update_coating_submenu(self):
        """Aggiorna il sottomenu Coating con struttura organizzata"""
        self.coating_submenu.delete(0, 'end')

        # Gestione Materiali Coating
        self.coating_submenu.add_cascade(
            label=self.lang.get('submenu_coating_materials', "ðŸ§° Gestione Materiali"),
            menu=self.coating_materials_submenu
        )
        self.coating_materials_submenu.delete(0, 'end')

        # âœ… SEPARARE I DUE MENU
        self.coating_materials_submenu.add_command(
            label=self.lang.get('submenu_coating_types', "ðŸŽ¨ Gestione Tipi Vernice"),
            command=self.open_coating_types_with_login
        )
        self.coating_materials_submenu.add_command(
            label=self.lang.get('submenu_coating_thickness_specs', "ðŸ“ Gestione Specifiche Spessore"),
            command=self.open_coating_thickness_specs_with_login
        )

        # Controlli QualitÃ 
        self.coating_submenu.add_separator()
        self.coating_submenu.add_command(
            label=self.lang.get('submenu_coating_viscosity', "ðŸ§ª Controllo ViscositÃ "),
            command=self.open_coating_viscosity_with_login
        )
        self.coating_submenu.add_command(
            label=self.lang.get('submenu_coating_thickness', "ðŸ“ Controllo Spessore"),
            command=self.open_coating_thickness_with_login
        )

        # Rapporti
        self.coating_submenu.add_separator()
        self.coating_submenu.add_command(
            label=self.lang.get('submenu_coating_reports', "ðŸ“Š Rapporti"),
            command=self.open_coating_reports_with_login
        )


    def _update_paste_submenu(self):
        """Aggiorna il sottomenu Paste"""
        self.paste_submenu.delete(0, 'end')

        # 1. Configurazione
        self.paste_submenu.add_cascade(label=self.lang.get('submenu_paste_configuration', "Configurazione"),
                                       menu=self.paste_config_submenu)
        self._update_paste_config_submenu()

        # 2. Ricevimento
        self.paste_submenu.add_command(
            label=self.lang.get('submenu_paste_reception', "Ricevimento"),
            command=self.open_paste_reception_with_login
        )

        # 3. Trasferimento
        self.paste_submenu.add_cascade(label=self.lang.get('submenu_paste_transfer', "Trasferimento"),
                                       menu=self.paste_transfer_submenu)
        self._update_paste_transfer_submenu()

    def _update_paste_config_submenu(self):
        """Aggiorna il sottomenu Configurazione Paste"""
        self.paste_config_submenu.delete(0, 'end')

        # Paste
        self.paste_config_submenu.add_command(
            label=self.lang.get('submenu_paste_products', "Paste"),
            command=self.open_paste_configuration_with_login
        )

        # Frigoriferi
        self.paste_config_submenu.add_command(
            label=self.lang.get('submenu_paste_refrigerators', "Frigoriferi"),
            command=lambda: self._execute_authorized_action(
                'manage_paste_refrigerators',
                lambda: self._open_paste_refrigerators()
            )
        )

        # Locazioni Frigoriferi (sostituisce Allarmi)
        self.paste_config_submenu.add_command(
            label=self.lang.get('submenu_paste_locations', "Locazioni Frigoriferi"),
            command=self._open_paste_locations
        )
        
        # Produttori
        self.paste_config_submenu.add_command(
            label=self.lang.get('submenu_paste_producers', "Produttori"),
            command=lambda: self._open_producers()
        )
        
        # Stampanti
        self.paste_config_submenu.add_command(
            label=self.lang.get('submenu_paste_printers', "Stampanti"),
            command=lambda: self._execute_authorized_action(
                'config_printers',
                lambda: self._open_printer_config()
            )
        )

    def _open_printer_config(self):
        """Apre la finestra di configurazione stampanti"""
        try:
            user = getattr(self, 'last_authenticated_user_name', 'Unknown')
            printer_config_manager.open_printer_config(
                self,
                self.db,
                self.lang,
                user
            )
        except Exception as e:
            logger.error(f"Errore apertura configurazione stampanti: {e}")
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire la configurazione stampanti: {str(e)}"
            )

    def _update_paste_transfer_submenu(self):
        """Aggiorna il sottomenu Trasferimento Paste"""
        self.paste_transfer_submenu.delete(0, 'end')

        # In Produzione
        self.paste_transfer_submenu.add_command(
            label=self.lang.get('submenu_paste_to_production', "In Produzione"),
            command=self._open_paste_to_production
        )

        # Presa In Carico
        self.paste_transfer_submenu.add_command(
            label=self.lang.get('submenu_paste_take_charge', "Presa In Carico"),
            command=self._open_paste_take_charge
        )

        # Inizio Uso
        self.paste_transfer_submenu.add_command(
            label=self.lang.get('submenu_paste_start_use', "Inizio Uso"),
            command=lambda: messagebox.showinfo("Info", "Funzione Inizio Uso - Da implementare")
        )

        # Fine Uso
        self.paste_transfer_submenu.add_command(
            label=self.lang.get('submenu_paste_end_use', "Fine Uso"),
            command=lambda: messagebox.showinfo("Info", "Funzione Fine Uso - Da implementare")
        )

    def _open_paste_to_production(self):
        """Apre la finestra di trasferimento paste a produzione"""
        def action():
            import paste_transfer
            paste_transfer.open_paste_to_production(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        self._execute_authorized_action('submenu_paste_to_production', action)

    def _open_paste_take_charge(self):
        """Apre la finestra presa in carico paste"""
        def action():
            import paste_take_charge
            paste_take_charge.open_paste_take_charge(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        self._execute_authorized_action('submenu_paste_take_charge', action)

    def _update_product_checks_submenu(self):
        """Aggiorna il sottomenu Verifiche Prodotti"""
        self.product_checks_submenu.delete(0, 'end')

        self.product_checks_submenu.add_command(
            label=self.lang.get('submenu_manage_product_checks', "Gestione Prodotti"),
            command=lambda: self._execute_authorized_action(
                'manage_product_check',
                lambda: self._open_product_checks_management()
            )
        )

        self.product_checks_submenu.add_command(
            label=self.lang.get('submenu_manage_check_rules', "Gestione Regole"),
            command=lambda: self._execute_authorized_action(
                'manage_check_rules',
                lambda: self._open_check_tasks_management()
            )
        )

        self.product_checks_submenu.add_separator()
        self.product_checks_submenu.add_command(
            label=self.lang.get('submenu_verify_products', "Verifiche"),
            command=lambda: self.open_product_verification_with_login()
        )

        self.product_checks_submenu.add_separator()
        self.product_checks_submenu.add_command(
            label=self.lang.get('submenu_verification_reports', "Rapporti"),
            command=lambda: self._open_verification_reports()
        )

    def _open_verification_reports(self):
        """Apre la finestra dei rapporti di verifica"""
        logger.info("Apertura finestra dei rapporti di verifica")
        try:
            from product_checks_gui import VerificationReportsWindow
            user = getattr(self, 'last_authenticated_user_name', 'Unknown')
            VerificationReportsWindow(self, self.db, self.lang, user)
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire i rapporti: {e}")

    def _update_reports_submenu(self):
        """Aggiorna il sottomenu Rapporti"""
        self.reports_submenu.delete(0, 'end')

        # OperativitÃ 
        self.reports_submenu.add_cascade(label=self.lang.get('submenu_operativity', "Operativita'"),
                                         menu=self.operativity_submenu)
        self.operativity_submenu.delete(0, 'end')

        shipping_submenu = tk.Menu(self.operativity_submenu, tearoff=0)
        self.operativity_submenu.add_cascade(label=self.lang.get('submenu_shipping', "Spedizioni"),
                                             menu=shipping_submenu)
        shipping_submenu.delete(0, 'end')
        shipping_submenu.add_command(label=self.lang.get('upload_report_label', "Carica Report"),
                                     command=self.open_shipping_report_window_with_login)
        shipping_submenu.add_command(label=self.lang.get('submenu_shipping_settings', "Impostazioni Spedizione"),
                                     command=self.open_shipping_settings_with_login)
        shipping_submenu.add_command(label=self.lang.get('submenu_xls_settings', "Impostazioni XLS"),
                                     command=self.open_xls_settings_with_login)

        # Altri rapporti
        self.reports_submenu.add_command(
            label=self.lang.get('submenu_line_stoppage_reports', "Rapporti di fermo linea"),
            command=self.open_line_stoppage_report
        )

        self.reports_submenu.add_command(
            label=self.lang.get("scrap_reports", "Rapporto Scarti"),
            command=self.open_scrap_reports
        )

    def _update_maintenance_menu(self):
        """Aggiorna il menu Manutenzione"""
        self.maintenance_menu.delete(0, 'end')

        # Gestione Macchine
        machine_submenu = tk.Menu(self.maintenance_menu, tearoff=0)
        self.maintenance_menu.add_cascade(label=self.lang.get('submenu_machines'), menu=machine_submenu)
        machine_submenu.add_command(label=self.lang.get('submenu_add_machine'),
                                    command=self.open_add_machine_with_login)
        machine_submenu.add_command(label=self.lang.get('submenu_equipment_types', "Gestione Tipi Macchine"),
                                    command=self.open_equipment_types_manager_with_login)
        machine_submenu.add_command(label=self.lang.get('submenu_edit_machine'),
                                    command=self.open_edit_machine_with_login)
        machine_submenu.add_command(label=self.lang.get('submenu_view_machines'),
                                    command=lambda: maintenance_gui.open_view_machines(self, self.db, self.lang))
        machine_submenu.add_separator()

        # Fixture
        fixture_submenu = tk.Menu(machine_submenu, tearoff=0)
        machine_submenu.add_cascade(label=self.lang.get('submenu_fixture', 'Fixture'), menu=fixture_submenu)
        fixture_submenu.add_command(label=self.lang.get('submenu_fixture_rules', 'Regole Fixture'),
                                    command=self.open_fixture_rules_with_login)
        fixture_submenu.add_command(label=self.lang.get('submenu_assign_products_fixture', 'Assegnazione prodotti'),
                                    command=self.open_assign_products_fixture_with_login)
        
        machine_submenu.add_separator()
        machine_submenu.add_command(label=self.lang.get('submenu_brand_management', "Gestione Brand"),
                                    command=self.open_brand_manager_with_login)
        machine_submenu.add_command(label=self.lang.get('submenu_company_management', "Gestione Compagnie"),
                                    command=self.open_company_manager_with_login)
        # machine_submenu.add_command(label=self.lang.get('submenu_equipment_types', "Gestione Tipi Macchine"),
        #                             command=self.open_equipment_types_manager_with_login)

        # Task di Manutenzione
        tasks_submenu = tk.Menu(self.maintenance_menu, tearoff=0)
        self.maintenance_menu.add_cascade(
            label=self.lang.get('submenu_maintenance_tasks_header', 'Task di Manutenzione'), menu=tasks_submenu)
        tasks_submenu.add_command(label=self.lang.get('submenu_manage_maint_task', "Gestione Task di Manutenzione"),
                                  command=self.open_add_maintenance_tasks_with_login)
        tasks_submenu.add_command(label=self.lang.get('submenu_manage_task_cycles', "Gestione Voci Task"),
                                  command=self.open_task_cycles_manager_with_login)
        tasks_submenu.add_command(label=self.lang.get('submenu_assign_responsibles', "Assegna Responsabili"),
                                  command=self.open_assign_responsibles_with_login)

        # Voci principali
        self.maintenance_menu.add_command(label=self.lang.get('submenu_fill_templates'),
                                          command=self.open_fill_templates_with_login)
        self.maintenance_menu.add_separator()
        
        # Sottomenu Rapporti
        reports_submenu = tk.Menu(self.maintenance_menu, tearoff=0)
        self.maintenance_menu.add_cascade(
            label=self.lang.get('submenu_reports_header', 'Rapporti'), menu=reports_submenu)
        reports_submenu.add_command(label=self.lang.get('submenu_reports', "Report Panoramica"),
                                   command=lambda: maintenance_gui.open_reports(self, self.db, self.lang))
        reports_submenu.add_command(label=self.lang.get('submenu_fixtures_report', "Rapporti Fixtures"),
                                   command=lambda: self._open_fixtures_report())
        
        self.maintenance_menu.add_command(label=self.lang.get('submenu_missing_action', "Missing Action Report"),
                                          command=self.open_missing_action_report)

    def _update_submissions_menu(self):
        """Aggiorna il menu Segnalazioni"""
        self.submissions_menu.delete(0, 'end')
        self.submissions_menu.add_command(label=self.lang.get('submenu_new_submission', "Nuova Segnalazione"),
                                          command=self.open_new_submission_form)
        self.submissions_menu.add_command(
            label=self.lang.get('submenu_assign', 'Assegna'),
            command=self.open_assign_submissions_with_login
        )
        self.submissions_menu.add_command(
            label=self.lang.get('menu_submissions_management', "Gestione"),
            command=lambda: self._execute_authorized_action(
                'management_submissions',
                lambda: self._open_submissions_management()
            )
        )

    def _open_fixtures_report(self):
        """Apre la finestra dei rapporti fixtures (senza login)"""
        try:
            fixtures_report_window.open_fixtures_report(self, self.db, self.lang)
        except Exception as e:
            logger.error(f"Errore apertura rapporti fixtures: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore apertura finestra: {e}",
                parent=self
            )

    def _update_tools_menu(self):
        """Aggiorna il menu Strumenti"""
        self.tools_menu.delete(0, 'end')

        # Autorizzazioni
        self.tools_menu.add_cascade(label=self.lang.get('submenu_permissions', "Autorizzazioni"),
                                    menu=self.permissions_submenu)
        self.permissions_submenu.delete(0, 'end')
        self.permissions_submenu.add_command(
            label=self.lang.get('submenu_special_permissions', "Autorizzazioni Speciali"),
            command=self.open_manage_permissions_with_login)
        self.permissions_submenu.add_command(
            label=self.lang.get('submenu_view_permissions', "Visualizza Autorizzazioni"),
            command=self.open_view_permissions_with_login)

        self.tools_menu.add_separator()

        # Materiali
        self.tools_menu.add_cascade(label=self.lang.get('menu_materials', "Materiali"), menu=self.materials_submenu)
        self.materials_submenu.delete(0, 'end')
        self.materials_submenu.add_command(label=self.lang.get('submenu_manage', "Gestione"),
                                           command=self.open_manage_materials_with_login)
        self.materials_submenu.add_command(label=self.lang.get('submenu_view', "Visualizza"),
                                           command=self.open_view_materials)

        # Altri strumenti
        self.tools_menu.add_command(
            label=self.lang.get('submenu_scrap_types', 'Tipi scrap'),
            command=self.open_scrap_types_with_login
        )

        self.tools_menu.add_separator()

        # Room Booking
        self.tools_menu.add_cascade(label=self.lang.get('menu_room_booking', "Room Booking"), menu=self.room_booking_submenu)
        self.room_booking_submenu.delete(0, 'end')

        self.room_booking_submenu.add_command(
            label=self.lang.get('submenu_manage_rooms', "Manage Rooms"),
            command=self.open_manage_rooms_with_login
        )

        self.room_booking_submenu.add_command(
            label=self.lang.get('submenu_manage_booking', "Manage Booking"),
            command=self.open_manage_booking_with_login
        )

        self.tools_menu.add_separator()
        self.tools_menu.add_command(label=self.lang.get('submenu_suppliers', "Produttori"),
                                    command=self.open_suppliers_manager_with_login)
        self.tools_menu.add_command(label=self.lang.get('submenu_maint_cycles', "Cicli Manutenzione"),
                                    command=self.open_maint_cycles_manager_with_login)
        self.tools_menu.add_command(label=self.lang.get('submenu_doc_types', "Aggiungi Tipo Documento"),
                                    command=self.open_doc_types_manager_with_login)
        self.tools_menu.add_separator()
        self.tools_menu.add_command(label=self.lang.get('manage_translations', "Gestione Traduzioni"),
                                    command=self.open_translations_manager_with_login)
        self.tools_menu.add_separator()
        self.tools_menu.add_command(label=self.lang.get('submenu_maint_times', "Tempi Manutenzione"),
                                    command=self.open_maintenance_times_with_login)
        self.tools_menu.add_separator()
        self.tools_menu.add_command(label=self.lang.get('submenu_settings_email', "Manage setting emails"),
                                    command=self.open_settings_email_with_login)

    def _update_help_menu(self):
        """Aggiorna il menu Aiuto"""
        self.help_menu.delete(0, 'end')
        self.help_menu.add_cascade(label=self.lang.get('menu_language'), menu=self.language_menu)
        
        # Voce Cambia Password
        self.help_menu.add_command(
            label=self.lang.get('menu_change_password', 'Cambia Password'),
            command=self._open_change_password
        )
        
        # Voce Recupera Password
        self.help_menu.add_command(
            label=self.lang.get('menu_recover_password', 'Recupera Password'),
            command=self._open_password_recovery
        )
        
        self.help_menu.add_separator()
        about_menu_label = f"{self.lang.get('menu_about')} {APP_VERSION}"
        self.help_menu.add_command(label=about_menu_label, command=self._show_about)


    def _update_main_menubar(self):
        """Aggiorna le etichette della barra dei menu principale"""
        try:
            self.menubar.delete(0, 'end')
            self._add_main_menus_to_bar()
            self.config(menu=self.menubar)
            self.update_idletasks()
        except tk.TclError:
            pass

    def _open_calibration_status(self):
        """Apre la finestra di situazione calibrazioni"""
        try:
            from calibration_status_window import CalibrationStatusWindow
            CalibrationStatusWindow(self, self.db, self.lang)
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire la finestra situazione calibrazioni: {e}")

    def _open_calibration_management(self):
        """Apre la finestra di gestione calibrazioni esistente"""
        try:
            from calibration_gui import CalibrationsWindow
            CalibrationsWindow(self, self.db, self.lang)
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire la gestione calibrazioni: {e}")

    def _generate_calibration_report(self):
        """Genera direttamente il report Excel delle calibrazioni"""
        try:
            from calibration_status_window import CalibrationStatusWindow
            # Crea una finestra temporanea solo per generare il report
            temp_window = CalibrationStatusWindow(self, self.db, self.lang)
            temp_window._generate_excel_report()
            temp_window.destroy()  # Chiude la finestra dopo aver generato il report
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile generare il report: {e}")

    def open_npi_project_management(self):
        """Apre la finestra di gestione progetti NPI (protetta da autorizzazione)"""

        def authorized_action():
            """Azione da eseguire solo se l'utente Ã¨ autorizzato"""
            try:
                final_customers = self.db.get_suppliers()#fetch_final_customers()
                progetti = self.npi_manager.get_progetti_attivi()

                if not progetti:
                    messagebox.showinfo(
                        "Nessun progetto",
                        "Non ci sono progetti NPI attivi. Crea prima un progetto dalla Configurazione NPI.",
                        parent=self
                    )
                    return

                # Crea una finestra di selezione piÃ¹ larga
                selection_window = tk.Toplevel(self)
                selection_window.title("Seleziona Progetto NPI")
                selection_window.geometry("900x250")
                selection_window.transient(self)
                selection_window.grab_set()

                ttk.Label(selection_window, text="Seleziona il progetto da gestire:", 
                         font=('Helvetica', 10, 'bold')).pack(pady=10)

                # ðŸ†• Frame filtri
                filter_frame = ttk.LabelFrame(selection_window, text=self.lang.get('filters', 'Filtri'), padding=10)
                filter_frame.pack(pady=10, padx=20, fill=tk.X)

                # Filtro Cliente
                client_frame = ttk.Frame(filter_frame)
                client_frame.pack(fill=tk.X, pady=5)
                ttk.Label(client_frame, text=self.lang.get('npi_client_filter', 'Cliente:'), width=12).pack(side=tk.LEFT)
                
                client_var = tk.StringVar()
                client_combo = ttk.Combobox(client_frame, textvariable=client_var, width=50, state='readonly')
                
                # Estrai clienti unici dai progetti attivi
                clienti_set = set()
                for proj in progetti:
                    cliente = proj.get('Cliente', '').strip()
                    if cliente:
                        clienti_set.add(cliente)
                
                clienti_unici = sorted(clienti_set)
                client_values = [self.lang.get('npi_all_clients', 'Tutti i Clienti')] + clienti_unici
                client_combo['values'] = client_values
                client_combo.current(0)
                client_combo.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)

                # Frame per il combobox progetti
                combo_frame = ttk.Frame(selection_window)
                combo_frame.pack(pady=10, padx=20, fill=tk.X)
                
                ttk.Label(combo_frame, text=self.lang.get('npi_project_label', 'Progetto:'), width=12).pack(side=tk.LEFT)

                # Combobox editabile con i progetti - usa il campo ActiveNpi formattato
                combo_var = tk.StringVar()
                combo = ttk.Combobox(combo_frame, textvariable=combo_var, width=100)
                
                # Prepara lista completa dei progetti con formato ActiveNpi
                all_projects_list = [p['ActiveNpi'] for p in progetti]
                progetti_map = {p['ActiveNpi']: p['ProgettoId'] for p in progetti}
                # ðŸ†• Mappa inversa per accedere ai dati completi del progetto
                progetti_map_reverse = {p['ActiveNpi']: p for p in progetti}
                
                combo['values'] = all_projects_list
                combo.pack(fill=tk.X, expand=True)

                # ðŸ†• Funzione per filtrare il combobox mentre l'utente digita (con filtro cliente)
                def on_keyrelease(event):
                    typed_text = combo_var.get().lower()
                    selected_client = client_var.get()
                    all_clients_label = self.lang.get('npi_all_clients', 'Tutti i Clienti')
                    
                    # Filtra prima per cliente
                    if selected_client == all_clients_label:
                        filtered_by_client = all_projects_list
                    else:
                        filtered_by_client = [
                            p for p in all_projects_list 
                            if progetti_map_reverse[p].get('Cliente', '').strip() == selected_client
                        ]
                    
                    # Poi filtra per testo digitato
                    if typed_text == '':
                        # Se il campo Ã¨ vuoto, mostra tutti i progetti (filtrati per cliente)
                        combo['values'] = filtered_by_client
                    else:
                        # Filtra i progetti che contengono il testo digitato
                        filtered = [p for p in filtered_by_client if typed_text in p.lower()]
                        combo['values'] = filtered
                    
                    # Riapri il dropdown se ci sono risultati
                    if (filtered if typed_text else filtered_by_client):
                        combo.event_generate('<Down>')

                # Bind dell'evento di digitazione - solo su Enter
                combo.bind('<Return>', on_keyrelease)
                # Bind anche su click della freccia dropdown
                combo.bind('<Button-1>', lambda e: on_keyrelease(e) if combo_var.get() else None)
                
                # ðŸ†• Funzione per aggiornare la lista progetti quando cambia il cliente
                def on_client_change(event):
                    # Reset project combobox
                    combo_var.set('')
                    on_keyrelease(None)
                
                # Bind cambio cliente
                client_combo.bind('<<ComboboxSelected>>', on_client_change)

                def open_selected():
                    selected_text = combo_var.get()
                    if selected_text and selected_text in progetti_map:
                        project_id = progetti_map[selected_text]
                        selection_window.destroy()

                        # Recupera il nome dell'utente che ha appena effettuato l'autenticazione
                        logged_user = self.last_authenticated_user_name

                        # Passa il nome utente come argomento esplicito a ProjectWindow
                        ProjectWindow(self, self.npi_manager, self.lang, project_id, self,
                                      logged_user, final_customers)
                    else:
                        messagebox.showwarning(
                            "Selezione non valida",
                            "Seleziona un progetto valido dalla lista.",
                            parent=selection_window
                        )

                # Frame per i pulsanti
                btn_frame = ttk.Frame(selection_window)
                btn_frame.pack(pady=10)
                
                ttk.Button(btn_frame, text="Apri", command=open_selected, width=15).pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="Annulla", command=selection_window.destroy, width=15).pack(side=tk.LEFT, padx=5)

            except Exception as e:
                messagebox.showerror(
                    self.lang.get('error_title', 'Errore'),
                    f"Errore apertura gestione progetti: {str(e)}"
                )
                logger.error(f"Errore apertura ProjectWindow: {e}")

        # Usa il sistema di autorizzazione con la chiave corretta
        self._execute_authorized_action(
            menu_translation_key='project_window',  # o la chiave che usi nel DB
            action_callback=authorized_action
        )

    def _apri_dashboard_npi(self):
        """Callback per aprire la dashboard NPI (protetta da login semplice)."""
        logger.info("Tentativo di aprire la Dashboard NPI.")

        def action(user_name):
            try:
                # 'user_name' arriva dal login e non lo usiamo qui, ma Ã¨ richiesto dal callback
                logger.debug(f"Utente '{user_name}' autorizzato. Apertura Dashboard NPI.")
                dashboard = NpiDashboardWindow(master=self,
                                               npi_manager=self.npi_manager,
                                               lang=self.lang,
                                               logged_in_user=user_name)

                self.wait_window(dashboard)  # Rende la finestra modale

            except Exception as e:
                logger.error("Errore nell'apertura della NpiDashboardWindow: %s", e, exc_info=True)
                messagebox.showerror("Errore", f"Impossibile aprire la dashboard NPI: {e}")

        # Esegui l'azione solo dopo un login semplice (senza permessi specifici per ora)
        self._execute_simple_login(action_callback=action)

    def _seleziona_e_visualizza_gantt(self):
        """Callback per visualizzare il Gantt (protetta da login semplice)."""
        logger.info("Tentativo di visualizzare un Gantt NPI.")
        self._execute_simple_login(action_callback=lambda authorized_user:
            self._launch_gantt_window())

    def _configura_catalogo_task_npi(self):
        """Callback per configurare i task NPI (protetta da autorizzazione specifica)."""
        logger.info("Tentativo di accedere alla configurazione Task NPI.")
        self._execute_authorized_action(
            menu_translation_key='npi_setup_tasks',  # La chiave corrisponde al MenuValue nel DB
            action_callback=self._launch_config_window
        )

    def _launch_dashboard_window(self, username):
        """
        Crea e lancia la finestra della dashboard.
        'username' Ã¨ passato da _execute_simple_login.
        """
        logger.info(f"Utente '{username}' autorizzato. Apertura Dashboard NPI.")

        try:
            dashboard = NpiDashboardWindow(master=self, npi_manager=self.npi_manager, lang=self.lang)
            # Puoi rendere la finestra modale se necessario
            self.wait_window(dashboard)
            #self._load_npi_projects()  # Ricarica dopo la chiusura
            self.get_progetti_attivi()
        except Exception as e:
            logger.error("Errore nell'apertura della NpiDashboardWindow: %s", e, exc_info=True)
            messagebox.showerror("Errore", f"Impossibile aprire la dashboard NPI: {e}")

    def _launch_gantt_window(self):
        """
        Chiede all'utente di selezionare un progetto NPI attivo e poi lancia la
        finestra del Gantt per quel progetto, con filtri per cliente e prodotto.
        """
        if self.npi_manager is None:
            messagebox.showerror("Errore", "Il modulo NPI non Ã¨ inizializzato.")
            return

        try:
            # 1. Recupera la lista dei progetti attivi dal gestore NPI.
            progetti_attivi = self.npi_manager.get_progetti_attivi()
            if not progetti_attivi:
                messagebox.showinfo("Nessun Progetto", "Non ci sono progetti NPI attivi da visualizzare.", parent=self)
                return

            # 2. Recupera prodotti per il filtro
            prodotti = self.npi_manager.get_prodotti()

            # 3. Crea la finestra di dialogo
            scelta_dialog = tk.Toplevel(self)
            scelta_dialog.title(self.lang.get("gantt_select_project", "Seleziona Progetto"))
            scelta_dialog.geometry("700x500")
            scelta_dialog.transient(self)
            scelta_dialog.grab_set()

            # Frame principale
            main_frame = ttk.Frame(scelta_dialog, padding=10)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Label titolo
            ttk.Label(main_frame, text=self.lang.get("gantt_choose_project", "Scegli un progetto per visualizzare il Gantt:")).pack(pady=(0, 10))

            # Frame filtri
            filter_frame = ttk.LabelFrame(main_frame, text=self.lang.get("filters", "Filtri"), padding=10)
            filter_frame.pack(fill=tk.X, pady=(0, 10))

            # Filtro Cliente
            client_frame = ttk.Frame(filter_frame)
            client_frame.pack(fill=tk.X, pady=5)
            ttk.Label(client_frame, text=self.lang.get("gantt_select_client", "Cliente:"), width=12).pack(side=tk.LEFT)
            
            client_var = tk.StringVar()
            client_combo = ttk.Combobox(client_frame, textvariable=client_var, width=50)
            
            # Estrai clienti unici dai progetti attivi, rimuovendo spazi e duplicati
            clienti_set = set()
            for proj in progetti_attivi:
                cliente = proj.get('Cliente', '').strip()
                if cliente:
                    clienti_set.add(cliente)
            
            clienti_unici = sorted(clienti_set)
            client_values = [self.lang.get("gantt_all_clients", "Tutti i Clienti")] + clienti_unici
            
            client_combo['values'] = client_values
            client_combo.current(0)
            client_combo.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)

            # Filtro Prodotto
            product_frame = ttk.Frame(filter_frame)
            product_frame.pack(fill=tk.X, pady=5)
            ttk.Label(product_frame, text=self.lang.get("gantt_select_product", "Prodotto:"), width=12).pack(side=tk.LEFT)
            
            product_var = tk.StringVar()
            product_combo = ttk.Combobox(product_frame, textvariable=product_var, width=50)
            
            # Mappa completa prodotti (ID -> dati prodotto)
            all_products_map = {p.ProdottoID: p for p in prodotti}
            
            product_combo.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)

            # Frame listbox
            list_frame = ttk.Frame(main_frame)
            list_frame.pack(fill=tk.BOTH, expand=True)

            # Scrollbar
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Listbox progetti
            listbox = tk.Listbox(list_frame, exportselection=False, yscrollcommand=scrollbar.set)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=listbox.yview)

            # Mappa progetti: ActiveNpi -> dati completi
            progetti_map = {proj['ActiveNpi']: proj for proj in progetti_attivi}

            # Funzione per filtrare e aggiornare la listbox (DEFINITA PRIMA)
            def update_listbox(*args):
                listbox.delete(0, tk.END)
                
                selected_client_name = client_var.get()
                selected_product_id = None
                
                # Ottieni il product_map dal combobox (aggiornato dinamicamente)
                if hasattr(product_combo, 'product_map'):
                    selected_product_id = product_combo.product_map.get(product_var.get())
                
                # "Tutti i Clienti" ha valore None nella mappa
                filter_by_client = selected_client_name != self.lang.get("gantt_all_clients", "Tutti i Clienti")
                filter_by_product = selected_product_id is not None
                
                for active_npi, proj_data in progetti_map.items():
                    # Applica filtro cliente (basato sul nome stringa)
                    if filter_by_client and proj_data.get('Cliente', '').strip() != selected_client_name:
                        continue
                    # Applica filtro prodotto (basato su ProdottoID)
                    if filter_by_product and proj_data.get('ProdottoID') != selected_product_id:
                        continue
                    
                    listbox.insert(tk.END, active_npi)
            
            # Funzione per aggiornare la lista prodotti in base al cliente selezionato (DEFINITA DOPO)
            def update_product_combo(*args):
                selected_client = client_var.get()
                all_products_label = self.lang.get("gantt_all_products", "Tutti i Prodotti")
                
                if selected_client == self.lang.get("gantt_all_clients", "Tutti i Clienti"):
                    # Mostra tutti i prodotti
                    product_values = [all_products_label]
                    product_map = {all_products_label: None}
                    
                    for prodotto in prodotti:
                        nome = prodotto.NomeProdotto
                        product_values.append(nome)
                        product_map[nome] = prodotto.ProdottoID
                else:
                    # Filtra prodotti per il cliente selezionato
                    product_values = [all_products_label]
                    product_map = {all_products_label: None}
                    
                    # Trova i ProdottoID dei progetti di questo cliente
                    prodotti_cliente = set()
                    for proj in progetti_attivi:
                        if proj.get('Cliente', '').strip() == selected_client:
                            prodotti_cliente.add(proj.get('ProdottoID'))
                    
                    logger.info(f"Cliente selezionato: '{selected_client}', Prodotti trovati: {len(prodotti_cliente)} - IDs: {prodotti_cliente}")
                    
                    # Aggiungi solo i prodotti di questo cliente
                    for prod_id in sorted(prodotti_cliente):
                        if prod_id in all_products_map:
                            prodotto = all_products_map[prod_id]
                            nome = prodotto.NomeProdotto
                            product_values.append(nome)
                            product_map[nome] = prodotto.ProdottoID
                
                # Aggiorna il combobox
                product_combo['values'] = product_values
                product_var.set(all_products_label)
                
                # Salva la mappa per l'uso nel filtro
                product_combo.product_map = product_map
                
                # Aggiorna anche la listbox
                update_listbox()
            
            # Inizializza con tutti i prodotti
            update_product_combo()

            # Collega eventi combobox
            client_combo.bind('<<ComboboxSelected>>', update_product_combo)
            product_combo.bind('<<ComboboxSelected>>', update_listbox)

            # Popola listbox iniziale (non serve chiamare update_listbox qui, giÃ  chiamato da update_product_combo)

            scelta_utente = None

            def on_ok():
                nonlocal scelta_utente
                selezionato = listbox.curselection()
                if selezionato:
                    scelta_utente = listbox.get(selezionato[0])
                    scelta_dialog.destroy()
                else:
                    messagebox.showwarning(
                        self.lang.get("warning", "Attenzione"),
                        self.lang.get("gantt_select_project_warning", "Seleziona un progetto dalla lista."),
                        parent=scelta_dialog
                    )

            def on_double_click(event):
                nonlocal scelta_utente
                selezionato = listbox.curselection()
                if selezionato:
                    scelta_utente = listbox.get(selezionato[0])
                    scelta_dialog.destroy()

            # Bind doppio click
            listbox.bind('<Double-Button-1>', on_double_click)

            # Frame pulsanti
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(pady=(10, 0))
            ttk.Button(btn_frame, text=self.lang.get("button_ok", "OK"), command=on_ok).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text=self.lang.get("button_cancel", "Annulla"), command=scelta_dialog.destroy).pack(side=tk.LEFT, padx=5)

            # Centra la finestra di dialogo e attendi che venga chiusa
            scelta_dialog.update_idletasks()
            x = (scelta_dialog.winfo_screenwidth() // 2) - (scelta_dialog.winfo_width() // 2)
            y = (scelta_dialog.winfo_screenheight() // 2) - (scelta_dialog.winfo_height() // 2)
            scelta_dialog.geometry(f"+{x}+{y}")
            
            self.wait_window(scelta_dialog)

            # 4. Se l'utente ha fatto una scelta (e non ha annullato), procedi.
            if scelta_utente:
                progetto_selezionato_id = progetti_map[scelta_utente]['ProgettoId']
                logger.info(f"Lancio della finestra Gantt per il progetto ID: {progetto_selezionato_id}")

                # 5. Apri la finestra Gantt, passando l'ID del progetto richiesto.
                NpiGanttWindow(
                    master=self,
                    npi_manager=self.npi_manager,
                    lang=self.lang,
                    progetto_id=progetto_selezionato_id
                )
            else:
                logger.info("Apertura finestra Gantt annullata dall'utente.")

        except Exception as e:
            logger.error(f"Errore durante il lancio della finestra Gantt: {e}", exc_info=True)
            messagebox.showerror("Errore Imprevisto", f"Impossibile aprire la vista Gantt: {e}", parent=self)

    def _launch_config_window(self):
        """
        Crea e lancia la finestra di configurazione Task.
        `self._temp_authorized_user_id` Ã¨ disponibile qui.
        """
        authorized_user = self.last_authenticated_user_name
        logger.info(f"Utente '{authorized_user}' autorizzato per la Configurazione NPI.")
        try:
            config_win = NpiConfigWindow(master=self, npi_manager=self.npi_manager, lang=self.lang,
                                         authorized_user=authorized_user)
            self.wait_window(config_win)

        except Exception as e:
            logger.error("Errore nell'apertura della NpiConfigWindow: %s", e, exc_info=True)
            messagebox.showerror("Errore", f"Impossibile aprire la configurazione NPI: {e}")

    def _update_complaints_menu(self):
        """Aggiorna il sottomenu Reclami con tutte le sezioni"""
        self.complaints_menu.delete(0, 'end')

        # 1. Aggiungi Reclamo
        self.complaints_menu.add_command(
            label=self.lang.get('menu_add_complaint', 'Aggiungi Reclamo'),
            command=self._add_complaint
        )
        # 2. Gestisci Reclamo
        self.complaints_menu.add_command(
            label=self.lang.get('menu_manage_complaint', 'Gestisci Reclamo'),
            command=self._manage_complaint
        )

        # 3. Analisi Reclami
        self.complaints_menu.add_command(
            label=self.lang.get('menu_complaints_analysis', 'Analisi Reclami'),
            command=self._analyze_complaints
        )

        # 4. Report Reclami
        self.complaints_menu.add_command(
            label=self.lang.get('menu_complaints_report', 'Report Reclami'),
            command=self._complaints_report
        )

    def _add_complaint(self):
        """Apre la finestra per aggiungere un reclamo - con autorizzazione"""
        self._execute_authorized_action(
            'aggiungi_reclami',
            self._add_complaint_authorized
        )

    def _add_complaint_authorized(self):
        """
        Esegue l'aggiunta reclamo dopo autorizzazione
        Chiamato solo se l'utente Ã¨ autorizzato
        """
        try:
            title = self.lang.get('title_add_complaint', 'Aggiungi Reclamo')
            logger.info(f"[COMPLAINTS] Utente {self.last_authenticated_user_name} ha accesso a: {title}")

            #Apri finestra per aggiungere un nuovo reclamo
            AddComplaintWindow(
                self,
                self.db,
                self.lang,
                self.last_authenticated_user_name
            )

            logger.debug(f"[COMPLAINTS] Finestra aggiunta reclamo aperta")

        except Exception as e:
            logger.exception(f"[COMPLAINTS] Errore nell'apertura aggiunta reclamo: {e}")
            messagebox.showerror(
                "Errore",
                f"Errore nell'apertura della finestra: {str(e)}",
                parent=self
            )

    def _manage_complaint(self):
        """Apre la finestra per la gestione del reclamo"""
        self._execute_authorized_action(
            'gestici_reclami',
            self._manage_complaint_authorized
        )

    def _manage_complaint_authorized(self):
        """
                Esegue l'aggiunta reclamo dopo autorizzazione
                Chiamato solo se l'utente Ã¨ autorizzato
                """
        try:
            title = self.lang.get('title_add_complaint', 'Aggiungi Reclamo')
            logger.info(f"[COMPLAINTS] Utente {self.last_authenticated_user_name} ha accesso a: {title}")

            # TODO: Implementare la finestra di aggiunta reclamo
            messagebox.showinfo(
                title,
                f"Funzione in fase di sviluppo\nUtente autorizzato: {self.last_authenticated_user_name}",
                parent=self
            )
            logger.debug(f"[COMPLAINTS] Finestra aggiunta reclamo aperta")

        except Exception as e:
            logger.exception(f"[COMPLAINTS] Errore nell'apertura aggiunta reclamo: {e}")
            messagebox.showerror(
                "Errore",
                f"Errore nell'apertura della finestra: {str(e)}",
                parent=self
            )

    def _analyze_complaints(self):
        """Apre la finestra per analizzare reclami - con autorizzazione"""
        self._execute_authorized_action(
            'analizza_reclami',
            self._analyze_complaints_authorized
        )

    def _analyze_complaints_authorized(self):
        """
        Esegue l'analisi reclami dopo autorizzazione
        Chiamato solo se l'utente Ã¨ autorizzato
        """
        try:
            title = self.lang.get('title_analyze_complaints', 'Analisi Reclami')
            logger.info(f"[COMPLAINTS] Utente {self.last_authenticated_user_name} ha accesso a: {title}")

            # TODO: Implementare la finestra di analisi reclami
            messagebox.showinfo(
                title,
                f"Funzione in fase di sviluppo\nUtente autorizzato: {self.last_authenticated_user_name}",
                parent=self
            )
            logger.debug(f"[COMPLAINTS] Finestra analisi reclami aperta")

        except Exception as e:
            logger.exception(f"[COMPLAINTS] Errore nell'apertura analisi reclami: {e}")
            messagebox.showerror(
                "Errore",
                f"Errore nell'apertura della finestra: {str(e)}",
                parent=self
            )

    def _complaints_report(self):
        """Apre la finestra per il report reclami - con autorizzazione"""
        self._execute_authorized_action(
            'report_reclami',
            self._complaints_report_authorized
        )

    def _complaints_report_authorized(self):
        """
        Esegue il report reclami dopo autorizzazione
        Chiamato solo se l'utente Ã¨ autorizzato
        """
        try:
            title = self.lang.get('title_complaints_report', 'Report Reclami')
            logger.info(f"[COMPLAINTS] Utente {self.last_authenticated_user_name} ha accesso a: {title}")

            # TODO: Implementare la finestra di report reclami
            messagebox.showinfo(
                title,
                f"Funzione in fase di sviluppo\nUtente autorizzato: {self.last_authenticated_user_name}",
                parent=self
            )
            logger.debug(f"[COMPLAINTS] Finestra report reclami aperta")

        except Exception as e:
            logger.exception(f"[COMPLAINTS] Errore nell'apertura report reclami: {e}")
            messagebox.showerror(
                "Errore",
                f"Errore nell'apertura della finestra: {str(e)}",
                parent=self
            )

    def open_coating_thickness_with_login(self):
        """Apre la finestra di controllo spessore coating dopo login"""

        def action(user_name):
            try:
                from coating_gui import CoatingThicknessMeasurementWindow
                logger.debug(f"[main.py] Apertura CoatingThicknessMeasurementWindow con user_name: '{user_name}'")
                window = CoatingThicknessMeasurementWindow(
                    parent=self,
                    conn_str=self.db.conn_str,
                    username=user_name,
                    language_code=self.lang.current_language
                )
                window.show()
                logger.info(f"Aperta finestra Coating Thickness da utente: {user_name}")
            except Exception as e:
                logger.error(f"Errore apertura controllo spessore coating: {e}")
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Errore apertura controllo spessore: {e}"
                )

        self._execute_simple_login(action_callback=action)

    def open_scrap_reports(self):
        try:
            # Usa self direttamente come parent
            scrap_reports_gui.ScrapReportsWindow(self, self.db, self.lang)
        except Exception as e:
            logger.error(f"Errore apertura rapporti scarti: {e}")

    def _open_submissions_management(self, user_id=None):
        """Apre la finestra di gestione segnalazioni"""
        # ðŸ” DEBUG: Stampa tutti gli attributi che contengono "user" o "employee"
        print("\n=== ATTRIBUTI DISPONIBILI ===")
        for attr in dir(self):
            if not attr.startswith('_'):  # Escludi attributi privati
                if 'user' in attr.lower() or 'employee' in attr.lower() or 'logged' in attr.lower():
                    try:
                        value = getattr(self, attr)
                        if not callable(value):  # Escludi metodi
                            print(f"  {attr} = {value}")
                    except:
                        pass
        print("=============================\n")
        import submissions_management_gui
        user_id = getattr(self, '_temp_authorized_user_id', None)

        submissions_management_gui.SubmissionsManagementWindow(
            self, self.db, self.lang, user_id
        )

    def open_coating_types_with_login(self):
        """Apre la finestra di gestione tipi vernice con autenticazione autorizzata"""

        def action():
            try:
                from coating_gui import CoatingSettingsWindow
                window = CoatingSettingsWindow(
                    parent=self,
                    conn_str=self.db.conn_str,
                    language_code=self.lang.current_language
                )
                window.show()
                logger.info("Aperta finestra Gestione Tipi Vernice")
            except Exception as e:
                logger.error(f"Errore apertura Gestione Tipi Vernice: {e}")
                messagebox.showerror(
                    self.lang.get('error', "Errore"),
                    f"{self.lang.get('window_open_error', 'Impossibile aprire la finestra')}: {str(e)}"
                )

        self._execute_authorized_action('submenu_coating_types', action)

    def open_coating_thickness_specs_with_login(self):
        """Apre la finestra di gestione specifiche spessore con autenticazione autorizzata"""

        def action():
            try:
                from coating_gui import CoatingThicknessSettingsWindow
                window = CoatingThicknessSettingsWindow(
                    parent=self,
                    conn_str=self.db.conn_str,
                    language_code=self.lang.current_language
                )
                window.show()
                logger.info("Aperta finestra Gestione Specifiche Spessore")
            except Exception as e:
                logger.error(f"Errore apertura Gestione Specifiche Spessore: {e}")
                messagebox.showerror(
                    self.lang.get('error', "Errore"),
                    f"{self.lang.get('window_open_error', 'Impossibile aprire la finestra')}: {str(e)}"
                )

        self._execute_authorized_action('submenu_coating_thickness_specs', action)

    def open_coating_viscosity_with_login(self):
        """Apre la finestra di registrazione viscositÃ  con login semplice"""

        def action(user_name):
            try:
                from coating_gui import CoatingViscosityWindow
                logger.debug(f"[main.py] Apertura CoatingViscosityWindow con user_name: '{user_name}'")
                window = CoatingViscosityWindow(
                    parent=self,
                    conn_str=self.db.conn_str,
                    username=user_name,  # âœ… Parametro corretto: username (non user_name)
                    language_code=self.lang.current_language
                )
                window.show()
                logger.info(f"Aperta finestra Coating Viscosity da utente: {user_name}")
            except Exception as e:
                logger.error(f"Errore apertura Coating Viscosity: {e}")
                messagebox.showerror(
                    self.lang.get('error', "Errore"),
                    f"{self.lang.get('window_open_error', 'Impossibile aprire la finestra')}: {str(e)}"
                )

        self._execute_simple_login(action_callback=action)

    def open_coating_reports_with_login(self):
        """Apre la finestra dei report coating (senza login richiesto)"""
        try:
            from coating_gui import CoatingReportsWindow
            window = CoatingReportsWindow(
                parent=self,
                conn_str=self.db.conn_str,
                language_code=self.lang.current_language
            )
            window.show()
            logger.info("Aperta finestra Coating Reports")
        except Exception as e:
            logger.error(f"Errore apertura Coating Reports: {e}")
            messagebox.showerror(
                self.lang.get('error', "Errore"),
                f"{self.lang.get('window_open_error', 'Impossibile aprire la finestra')}: {str(e)}"
            )

    def stop_fct_loop(self):
        """Ferma il loop FCT e invia notifica email"""
        if self.fct_loop_running:
            self.fct_loop_running = False
            # Cancella il timer se esiste
            if hasattr(self, 'fct_timer_id') and self.fct_timer_id:
                self.after_cancel(self.fct_timer_id)
                self.fct_timer_id = None

            # Invia email in background
            self._send_fct_stop_notification()

            # Aggiorna UI
            if hasattr(self, 'fct_status_label'):
                self.fct_status_label.config(text=self.lang.get('fct_stopped', 'FCT Loop Fermato'))

    def _send_fct_stop_notification(self):
        """Invia notifica email in background quando FCT loop viene fermato"""
        import threading

        def send_email_thread():
            try:
                # Recupero dei destinatari dal database
                recipients = utils.get_email_recipients(self.db.conn, 'Sys_fct_settings')

                if not recipients:
                    logging.info("Nessun destinatario configurato per notifiche FCT")
                    return

                # Prepara il messaggio
                subject = self.lang.get('fct_loop_stopped_subject', 'FCT Transfer Loop Fermato')
                body = self.lang.get(
                    'fct_loop_stopped_body',
                    f'Il loop FCT Transfer Ã¨ stato fermato dall\'utente {getattr(self, "current_user", "N/A")} alle {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                )

                # Invio dell'email
                utils.send_email(
                    recipients=recipients,
                    subject=subject,
                    body=body
                )

                logging.info(f"Notifica FCT stop inviata a: {recipients}")

            except Exception as e:
                logging.error(f"Errore nell'invio della notifica FCT: {str(e)}")

        # Avvia il thread in background
        thread = threading.Thread(target=send_email_thread, daemon=True)
        thread.start()

    def open_scrap_types_with_login(self):
        # Usa il gate "per menÃ¹" con chiave di traduzione dedicata
        self._execute_authorized_action(
            menu_translation_key='submenu_scrap_types',
            action_callback=lambda: scarti_gui.open_scrap_reasons_manager(self, self.db, self.lang)
        )

    def open_manage_rooms_with_login(self):
        def open_room_manager():
            import room_booking_gui
            room_booking_gui.RoomManagerWindow(self, self.db, self.lang)
        
        self._execute_authorized_action(
            menu_translation_key='manage_room',
            action_callback=open_room_manager
        )

    def open_manage_booking_with_login(self):
        def open_booking_manager(user_name):
            import room_booking_gui
            room_booking_gui.BookingManagerWindow(self, self.db, self.lang, user_name)
        
        self._execute_simple_login(action_callback=open_booking_manager)

    def open_paste_reception_with_login(self):
        """Apre la finestra ricevimento paste con login"""
        def open_paste_reception(user_name):
            import paste_manager
            paste_manager.open_paste_reception(self, self.db, self.lang, user_name)
        
        self._execute_simple_login(action_callback=open_paste_reception)

    def _open_paste_locations(self):
        """Apre la finestra gestione locazioni frigoriferi"""
        def action():
            import paste_manager
            paste_manager.open_paste_locations(self, self.db, self.lang, self.last_authenticated_user_name)
        
        self._execute_authorized_action('submenu_paste_locations', action)

    def _open_producers(self):
        """Apre la finestra gestione produttori"""
        def action():
            from paste_manager import ProducersWindow
            ProducersWindow(self, self.db, self.lang)
        self._execute_authorized_action('submenu_paste_producers', action)

    def open_guest_registration_with_login(self):
        self._execute_authorized_action(
            menu_translation_key='manage_guests',
            action_callback=lambda: guests_gui.GuestRegistrationWindow(
                self, self.db, self.lang, self.last_authenticated_user_name
            )
        )

    def open_guest_report_with_login(self):
        self._execute_simple_login(
            action_callback=lambda user_name: guests_gui.GuestReportWindow(self, self.db, self.lang)
        )

    def generate_guests_pdf_report_with_login(self):
        """Genera il report PDF degli ospiti presenti in fabbrica"""
        logger.info("=== INIZIO generate_guests_pdf_report_with_login ===")
        
        def action():
            logger.info("Action callback chiamato, inizio generazione report...")
            try:
                success, message, pdf_path = guests_report_generator.generate_guests_pdf_report(self.db)
                logger.info(f"Risultato generazione: success={success}, message={message}, pdf_path={pdf_path}")
                
                if success:
                    messagebox.showinfo(
                        self.lang.get('success', 'Successo'),
                        message
                    )
                else:
                    messagebox.showerror(
                        self.lang.get('error', 'Errore'),
                        message
                    )
            except Exception as e:
                logger.error(f"Errore durante la generazione del report: {e}")
                import traceback
                logger.error(traceback.format_exc())
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Errore imprevisto: {str(e)}"
                )
        
        logger.info("Chiamata _execute_simple_login...")
        self._execute_simple_login(action_callback=lambda user_name: action())

    def open_absence_authorization_with_login(self):
        """Apre la finestra di autorizzazione assenze con login autorizzato"""
        logger.info("open_absence_authorization_with_login chiamata")
        
        def action():
            try:
                import absence_authorization
                logger.info(f"Apertura finestra autorizzazione assenze per utente: {self.last_authenticated_user_name} (ID: {self.last_authorized_user_id})")
                absence_authorization.AbsenceAuthorizationWindow(
                    self,
                    self.db,
                    self.lang,
                    self.last_authorized_user_id,  # Passa l'EmployeeHireHistoryId invece del nome
                    self.last_authenticated_user_name  # Passa anche il nome per visualizzazione
                )
            except Exception as e:
                logger.error(f"Errore apertura finestra autorizzazione assenze: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra di autorizzazione assenze: {str(e)}",
                    parent=self
                )
        
        self._execute_authorized_action(
            menu_translation_key='autorizza_assenze',
            action_callback=action
        )

    def open_absence_rules_with_login(self):
        """Apre la finestra di gestione regole assenze con login autorizzato"""
        logger.info("open_absence_rules_with_login chiamata")
        
        def action():
            try:
                import absence_rules
                logger.info(f"Apertura finestra regole assenze per utente: {self.last_authenticated_user_name} (ID: {self.last_authorized_user_id})")
                absence_rules.AbsenceRulesWindow(
                    self,
                    self.db,
                    self.lang,
                    self.last_authorized_user_id,
                    self.last_authenticated_user_name
                )
            except Exception as e:
                logger.error(f"Errore apertura finestra regole assenze: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra: {e}",
                    parent=self
                )
        
        self._execute_authorized_action(
            menu_translation_key='gestione_regole_assenze',
            action_callback=action
        )


    def open_news_management_with_login(self):
        """Apre la finestra di gestione messaggi con login"""
        logger.info("open_news_management_with_login chiamata")
        
        def action(user_name):
            try:
                import news_management
                logger.info(f"Apertura finestra gestione messaggi per utente: {user_name}")
                news_management.NewsManagementWindow(
                    self,
                    self.db,
                    self.lang,
                    user_name
                )
            except Exception as e:
                logger.error(f"Errore apertura finestra gestione messaggi: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra di gestione messaggi: {str(e)}",
                    parent=self
                )
        
        logger.info("Chiamata _execute_simple_login...")
        self._execute_simple_login(action_callback=lambda user_name: action(user_name))

    def _open_password_recovery(self):
        """Apre la finestra di recupero password (SENZA login)"""
        logger.info("_open_password_recovery chiamata")
        
        try:
            import password_recovery
            password_recovery.PasswordRecoveryWindow(
                self,
                self.db,
                self.lang
            )
        except Exception as e:
            logger.error(f"Errore apertura finestra recupero password: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Impossibile aprire la finestra di recupero password: {str(e)}",
                parent=self
            )

    def _change_language(self, lang_code):

        """Cambia la lingua, aggiorna la UI, salva l'impostazione e mostra una notifica."""
        self.lang.set_language(lang_code)
        self.update_texts()
        self._save_language_setting(lang_code)


    def _open_fct_settings(self):
        """Apre la finestra di configurazione FCT Transfer (con controllo login)"""
        ok = self._execute_authorized_action(
            menu_translation_key='menu_fct_settings',
            action_callback=lambda: fct_transfer.FCTTransferSettingsWindow(
                self,
                self.db,
                self.lang,
                self.fct_config
            )
        )
        if not ok:
            return

    def _toggle_fct_execution(self):
        """Avvia o ferma l'esecuzione FCT Transfer"""
        if not self.fct_config.bat_file_path:
            messagebox.showwarning(
                self.lang.get('warning', "Attenzione"),
                self.lang.get('fct_configure_first', "Configurare prima il file batch in Settings"),
                parent=self
            )
            return

        # Verifica stato attuale
        is_running, instances = self.fct_manager.check_bat_running_status()

        if not self.fct_manager.is_running:
            # Avvia esecuzione
            if is_running:
                # GiÃ  in esecuzione su altra istanza
                messagebox.showinfo(
                    self.lang.get('info', "Informazione"),
                    self.lang.get('fct_already_running', "Batch giÃ  in esecuzione su un'altra istanza"),
                    parent=self
                )
                # Cambia comunque il menu in Stop (per coerenza UI)
                self._update_fct_menu_label("Stop")
            else:
                # Avvia nuova esecuzione
                if self.fct_manager.start_bat_execution():
                    self._update_fct_menu_label("Stop")
                    messagebox.showinfo(
                        self.lang.get('success', "Successo"),
                        self.lang.get('fct_started', "Esecuzione FCT Transfer avviata"),
                        parent=self
                    )
                else:
                    messagebox.showerror(
                        self.lang.get('error', "Errore"),
                        self.lang.get('fct_start_error', "Errore durante l'avvio"),
                        parent=self
                    )
        else:
            # Ferma esecuzione
            _, active_instances = self.fct_manager.check_bat_running_status()

            if len(active_instances) == 0:
                # Nessuna istanza attiva (stato inconsistente)
                self.fct_manager.is_running = False
                self._update_fct_menu_label("Run")

            elif len(active_instances) == 1:
                # Una sola istanza: ferma direttamente
                if self.fct_manager.stop_bat_execution():
                    self._update_fct_menu_label("Run")
                    messagebox.showinfo(
                        self.lang.get('success', "Successo"),
                        self.lang.get('fct_stopped', "Esecuzione FCT Transfer fermata"),
                        parent=self
                    )

            else:
                # Multiple istanze: chiedi quale fermare
                dialog = fct_transfer.FCTTransferStopDialog(self, self.lang, active_instances)
                self.wait_window(dialog)

                if dialog.selected_instances:
                    # Ferma le istanze selezionate
                    for inst in dialog.selected_instances:
                        self.fct_manager.stop_bat_execution(inst['DateStart'])

                    # Verifica se ci sono ancora istanze attive
                    _, remaining = self.fct_manager.check_bat_running_status()
                    if not remaining:
                        self._update_fct_menu_label("Run")

                    messagebox.showinfo(
                        self.lang.get('success', "Successo"),
                        self.lang.get('fct_instances_stopped', f"{len(dialog.selected_instances)} istanze fermate"),
                        parent=self
                    )

    def _update_fct_menu_label(self, label_key: str):
        """Aggiorna l'etichetta del menu Run/Stop"""
        if hasattr(self, 'fct_menu') and self.fct_run_menu_index is not None:
            new_label = self.lang.get(f'menu_fct_{label_key.lower()}', label_key)
            self.fct_menu.entryconfig(
                self.fct_run_menu_index,
                label=new_label
            )

    def _show_about(self):
        """Mostra la finestra di dialogo 'About' con le informazioni del software."""
        about_title = f"{self.lang.get('about_title')} - v{APP_VERSION}"
        about_template = self.lang.get_raw('about_message')
        # Assicurati che il template nel DB usi {version} e {developer}
        about_message = about_template.replace('{version}', APP_VERSION).replace('{developer}', APP_DEVELOPER)

        messagebox.showinfo(
            about_title,
            about_message,
            parent=self
        )
    
    def _open_change_password(self):
        """Apre la finestra per il cambio password"""
        import change_password_gui
        change_password_gui.open_change_password_window(
            self, self.db, self.lang, user_id=None, force_change=False
        )

    def open_manage_materials_with_login(self):
        self._execute_simple_login(
            action_callback=lambda user_name: materials_gui.open_manage_materials(self, self.db, self.lang, user_name)
        )

    def open_translations_manager_with_login(self):
        """Apre la finestra di gestione traduzioni con login autorizzato"""
        def action():
            try:
                import translations_manager
                translations_manager.open_translations_manager(self, self.db, self.lang)
            except Exception as e:
                logger.error(f"Errore apertura finestra gestione traduzioni: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra di gestione traduzioni: {str(e)}",
                    parent=self
                )
        
        self._execute_authorized_action(
            menu_translation_key='manage_translations',
            action_callback=action
        )

    def open_view_materials(self):
        """Apre la finestra di visualizzazione materiali (senza login)."""
        # Passiamo 'None' come user_name perchÃ© non c'Ã¨ autenticazione
        materials_gui.open_view_materials(self, self.db, self.lang, user_name=None)

    def open_edit_machine_with_login(self):
        def action(user_name):
            self.authenticated_user_for_maintenance = user_name
            maintenance_gui.open_edit_machine(self, self.db, self.lang)

        self._execute_simple_login(action_callback=action)

    def open_add_machine_with_login(self):
        """Apre la finestra per aggiungere una macchina dopo un login semplice."""
        self._execute_simple_login(
            action_callback=lambda user_name: maintenance_gui.open_add_machine(self, self.db, self.lang)
        )

    # =========================================================================
    # METODI MENU ORDINI (PLACEHOLDER)
    # =========================================================================
    
    def _load_orders_placeholder(self):
        """Menu Carica Ordini - Protetto da autorizzazione"""
        def authorized_action():
            """Apre la finestra di caricamento ordini"""
            try:
                from orders.load_orders_window import open_load_orders_window
                
                # Recupera il nome dell'utente autenticato
                user_name = self.last_authenticated_user_name if hasattr(self, 'last_authenticated_user_name') else 'Unknown'
                
                open_load_orders_window(self, self.db, self.lang, user_name)
            except Exception as e:
                logger.error(f"Errore apertura finestra caricamento ordini: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"{self.lang.get('error_opening_window', 'Errore apertura finestra')}:\n{e}",
                    parent=self
                )
        
        self._execute_authorized_action(
            menu_translation_key='Aggiungi_Ordini',
            action_callback=authorized_action
        )
    
    
    def _orders_reports_placeholder(self):
        """Apre la finestra dei rapporti ordini con controllo autorizzazioni"""
        def authorized_action():
            try:
                from orders.orders_reports_window import open_orders_reports_window
                
                # Recupera il nome dell'utente autenticato
                user_name = self.last_authenticated_user_name if hasattr(self, 'last_authenticated_user_name') else 'Unknown'
                
                open_orders_reports_window(
                    master=self,
                    db=self.db,
                    lang=self.lang,
                    user_name=user_name
                )
            except Exception as e:
                logger.error(f"Errore apertura finestra rapporti ordini: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra rapporti ordini:\n{e}",
                    parent=self
                )
        
        # Esegue l'azione con controllo autorizzazioni
        self._execute_authorized_action(
            menu_translation_key='regole_di_spedizione',
            action_callback=authorized_action
        )

    def open_equipment_types_manager_with_login(self):
        """Apre la finestra di gestione tipi macchine con autorizzazione"""
        def authorized_action():
            try:
                user_name = self.last_authenticated_user_name if hasattr(self, 'last_authenticated_user_name') else 'Unknown'
                maintenance_gui.EquipmentTypesManagerWindow(self, self.db, self.lang, user_name)
            except Exception as e:
                logger.error(f"Errore apertura finestra gestione tipi macchine: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Impossibile aprire la finestra gestione tipi macchine:\n{e}",
                    parent=self
                )
        
        self._execute_authorized_action(
            menu_translation_key='gestione_tipi_macchine',
            action_callback=authorized_action
        )

    # =========================================================================
    # FINE METODI MENU ORDINI
    # =========================================================================

    def _on_closing(self, force_quit=False):
        """Gestisce la chiusura dell'applicazione."""
        # Ferma tutti i timer attivi
        if self.slideshow_job_id: self.after_cancel(self.slideshow_job_id)
        if self.birthday_flash_job_id: self.after_cancel(self.birthday_flash_job_id)
        if self.birthday_stop_job_id: self.after_cancel(self.birthday_stop_job_id)
        if self.periodic_check_job_id: self.after_cancel(self.periodic_check_job_id)

        self._stop_product_check_background_task()

        # Se force_quit Ã¨ True, chiudi senza chiedere conferma
        if force_quit:
            self.db.disconnect()
            self.destroy()
            return
        
        # Chiedi conferma all'utente con possibilitÃ  di annullare
        # askokcancel ritorna: True=OK (chiudi), False=Annulla (non chiudere)
        if messagebox.askokcancel(
            self.lang.get('quit_title', 'Chiudi Applicazione'),
            self.lang.get('quit_message', 'Sei sicuro di voler chiudere l\'applicazione?')
        ):
            self.db.disconnect()
            self.destroy()
        # else: l'utente ha cliccato Annulla, non fare nulla

class UpdateNotificationDialog(tk.Toplevel):
    """Dialogo per notificare un nuovo aggiornamento e chiedere l'azione desiderata."""

    def __init__(self, parent, lang, new_version, current_version):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title(lang.get('update_available_title', "Aggiornamento Disponibile"))
        self.result = None  # 'now', 'later', o 'ignore'

        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(expand=True, fill="both")

        message = lang.get('update_notification_message',
                           "Ãˆ stata rilasciata una nuova versione del programma ({0}).\n"
                           "La tua versione attuale Ã¨ la {1}.\n\n"
                           "Cosa vuoi fare?", new_version, current_version)

        ttk.Label(main_frame, text=message, justify=tk.LEFT).pack(pady=10)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)

        ttk.Button(btn_frame, text=lang.get('update_now_btn', "Aggiorna Ora"),
                   command=self._on_now).pack(side="left", padx=10)
        ttk.Button(btn_frame, text=lang.get('remind_later_btn', "Ricorda tra 15 min"),
                   command=self._on_later).pack(side="left", padx=10)
        ttk.Button(btn_frame, text=lang.get('ignore_session_btn', "Ignora per questa sessione"),
                   command=self._on_ignore).pack(side="left", padx=10)

    def _on_now(self):
        self.result = 'now'
        self.destroy()

    def _on_later(self):
        self.result = 'later'
        self.destroy()

    def _on_ignore(self):
        self.result = 'ignore'
        self.destroy()

if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except KeyboardInterrupt:
        # Gestione interruzione da tastiera (Ctrl+C)
        print("\n\nâš ï¸  Applicazione interrotta dall'utente (Ctrl+C)")
        print("Chiusura in corso...")
        try:
            app.destroy()
        except:
            pass
        import sys
        sys.exit(0)
    except Exception as e:
        # Gestione errori imprevisti
        print(f"\n\nâŒ Errore critico nell'applicazione: {e}")
        import traceback
        traceback.print_exc()
        
        # Mostra messaggio all'utente se possibile
        try:
            import tkinter.messagebox as mb
            mb.showerror(
                "Errore Critico",
                f"Si Ã¨ verificato un errore critico:\n\n{str(e)}\n\nL'applicazione verrÃ  chiusa."
            )
        except:
            pass
        
        import sys
        sys.exit(1)
    finally:
        # Cleanup finale
        print("Applicazione terminata.")
