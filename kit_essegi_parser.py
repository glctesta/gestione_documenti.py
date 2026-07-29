"""
kit_essegi_parser.py
Parser dei file XLSX "lista prelievo" generati da Essegi (report Reels
traceability) salvati in T:\\KITTING.

Spec: docs/PlanRespect_KitPreparation_Spec_v1.2.md §5.1.1
Rilevazioni Sprint 0:
  - le colonne NON hanno posizione fissa: vanno ricavate dalla riga di
    intestazione 'REEL CODE' (osservato: dati da col. B nel file reale)
  - gli ordini sono nell'intestazione in formato compatto 'PR554/553/552/551'
    e vanno normalizzati a 9 caratteri totali: 'PR554' -> 'PR0000554'
"""
import hashlib
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Optional

logger = logging.getLogger("PlanMonitor")

KITTING_DIR = r"T:\KITTING"

RE_UNIQUE_HU = re.compile(r'^HU\d{9}(_\d{2})?$')
RE_UNIQUE_SHORT = re.compile(r'^\d{6}$')
# Riconoscimento generico di un reel/unique code. I file reali contengono formati
# piu' vari dei due sopra: HU con numero di cifre diverso da 9 (es. HU0071143_05),
# altri prefissi (es. ES001098), oltre ai puri numerici. Accetta: prefisso di 0-3
# lettere + almeno 5 cifre + suffisso opzionale '_NN'. Restano esclusi (via
# RE_ORDERS_COMPACT) i numeri d'ordine 'PRnnn...', cosi' le testate ripetute a
# cambio pagina non vengono scambiate per righe materiale.
RE_UNIQUE_GENERIC = re.compile(r'^[A-Za-z]{0,3}\d{5,}(_\d{1,3})?$')
RE_ORDERS_COMPACT = re.compile(r'^PR\d+(?:/\d+)+$|^PR\d+$')
# Testata dei file D365 (kitting per singolo ordine): la prima cella e' del tipo
# 'PR0000821_PTH' (ordine + fase). Cattura la parte numerica dell'ordine.
RE_D365_ORDER = re.compile(r'^PR(\d+)')

ORDER_TOTAL_LEN = 9  # 'PR' + zeri di padding + numero = 9 caratteri


@dataclass
class EssegiRow:
    """Una riga materiale della lista prelievo."""
    row_number: int            # riga nel foglio Excel (per messaggi)
    unique_number: str         # REEL CODE
    material_code: str         # ITEM CODE
    quantity: float            # QT


@dataclass
class EssegiFile:
    """Risultato del parsing di un file lista prelievo."""
    file_path: str
    file_name: str
    file_hash: str             # SHA-256 del contenuto
    file_date: datetime        # LastWriteTime
    orders_compact: str        # es. 'PR554/553/552/551'
    orders: List[str]          # normalizzati, es. ['PR0000554', ...]
    rows: List[EssegiRow] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def distinct_materials(self) -> set:
        return {r.material_code for r in self.rows}


class EssegiParseError(Exception):
    """Errore bloccante nel parsing della lista prelievo."""


def normalize_order(num: str) -> str:
    """'554' -> 'PR0000554' (lunghezza totale 9, 'PR' incluso)."""
    return 'PR' + num.zfill(ORDER_TOTAL_LEN - 2)


def expand_orders(compact: str) -> List[str]:
    """'PR554/553/552/551' -> ['PR0000554', 'PR0000553', 'PR0000552', 'PR0000551']"""
    parts = compact.replace(' ', '').split('/')
    nums = [parts[0][2:]] + parts[1:]
    return [normalize_order(n) for n in nums if n.isdigit()]


def file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def list_kitting_files(directory: str = KITTING_DIR) -> List[dict]:
    """
    Elenca i file .xlsx in T:\\KITTING con anteprima ordini, ordinati per
    data modifica decrescente. Ogni voce: {path, name, date, orders_compact}.
    """
    if not os.path.isdir(directory):
        raise EssegiParseError(f"Directory non raggiungibile: {directory}")
    out = []
    for f in os.listdir(directory):
        if not f.lower().endswith('.xlsx') or f.startswith('~$'):
            continue
        path = os.path.join(directory, f)
        try:
            compact = _detect_orders_preview(path)
        except Exception as e:
            logger.warning("Anteprima ordini fallita per %s: %s", f, e)
            compact = None
        out.append({
            'path': path,
            'name': f,
            'date': datetime.fromtimestamp(os.path.getmtime(path)),
            'orders_compact': compact or '?',
        })
    out.sort(key=lambda d: d['date'], reverse=True)
    return out


def _detect_orders_preview(path: str):
    """Anteprima ordini per la finestra di scelta file: riconosce sia il formato
    multi-job (intestazione 'PRnnn/nnn/...') sia i file a singolo ordine con
    testata 'PR0000nnn_FASE' (D365 o multi-job a singolo ordine)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        compact, _ = _find_orders_header(ws)
        if compact:
            return compact
        title, digits = _find_d365_order(ws)
        if digits:
            return title or normalize_order(digits)
        return None
    finally:
        wb.close()


def _find_orders_header(ws, max_rows: int = 10):
    """Cerca la riga ordini compatta nelle prime righe (qualsiasi colonna)."""
    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        for cell in row:
            v = str(cell.value).strip() if cell.value is not None else ''
            if RE_ORDERS_COMPACT.match(v):
                return v, cell.row
    return None, None


def _find_d365_order(ws, max_rows: int = 8):
    """Trova numero ordine e testata a singolo ordine nelle prime righe.
    Ritorna (testo_testata, cifre_ordine) — es. ('PR0000821_PTH', '0000821')."""
    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        for cell in row:
            v = str(cell.value).strip() if cell.value is not None else ''
            m = RE_D365_ORDER.match(v)
            if m:
                return v, m.group(1)
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
#  Mappatura colonne guidata da dizionario (alias per campo logico)
# ─────────────────────────────────────────────────────────────────────────────
class UnmappedColumnsError(EssegiParseError):
    """Il file non ha (tutte) le colonne mappate dal dizionario. Porta con sé le
    informazioni per proporre all'operatore la mappatura mancante:
      - found     : {campo_logico: col_idx} già mappati
      - missing   : [campo_logico, ...] non trovati
      - header    : [(col_idx, valore_intestazione), ...] della riga candidata
      - suggestions: {campo: {'suggestion': valore, 'col': idx, 'ratio': r}}"""
    def __init__(self, found, missing, header, suggestions):
        self.found = found
        self.missing = missing
        self.header = header
        self.suggestions = suggestions
        super().__init__("Colonne non mappate: " + ", ".join(missing))


def _default_column_dict() -> dict:
    """Dizionario di default (senza DB): campo -> set di alias MAIUSCOLI."""
    try:
        import kit_column_dict as kcd
        return {k: set(a.upper() for a in v) for k, v in kcd.DEFAULT_ALIASES.items()}
    except Exception:
        return {
            'unique_number': {'REEL CODE'},
            'material_code': {'ITEM CODE'},
            'quantity': {'QT', 'QUANTITY', 'QTY'},
        }


def _normalize_column_dict(column_dict) -> dict:
    """Unisce il dizionario passato ai default; alias in MAIUSCOLO."""
    base = _default_column_dict()
    if column_dict:
        for k, v in column_dict.items():
            base.setdefault(k, set()).update(a.upper() for a in v)
    return base


def _suggest_columns(missing, header_cells, column_dict) -> dict:
    """Per ogni campo mancante trova la cella di intestazione più simile
    (difflib) agli alias del campo / al suo nome canonico."""
    from difflib import SequenceMatcher
    try:
        import kit_column_dict as kcd
        canon = kcd.FIELD_CANON
    except Exception:
        canon = {'unique_number': 'REEL CODE', 'material_code': 'ITEM CODE', 'quantity': 'QUANTITY'}
    out = {}
    for field in missing:
        targets = [a.upper() for a in column_dict.get(field, ())] + [canon.get(field, field).upper()]
        best_val, best_col, best_ratio = None, None, -1.0
        for col, val in header_cells:
            vu = val.upper()
            r = max((SequenceMatcher(None, vu, t).ratio() for t in targets), default=0.0)
            if r > best_ratio:
                best_ratio, best_val, best_col = r, val, col
        out[field] = {'suggestion': best_val, 'col': best_col, 'ratio': round(best_ratio, 2)}
    return out


def find_data_columns(ws, column_dict, max_rows: int = 30):
    """Individua la riga di intestazione e mappa i campi logici alle colonne,
    cercando gli alias del dizionario (indipendente dalla posizione).

    Ritorna (data_start_row, {campo: col_idx0}) oppure solleva UnmappedColumnsError
    con i suggerimenti per i campi non trovati."""
    required = REQUIRED_FIELDS_ORDER
    best_score, best_rownum, best_map, best_cells = -1, None, {}, []
    richest_n, richest_cells = -1, []

    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        cells = []
        for c in row:
            v = str(c.value).strip() if c.value is not None else ''
            if v:
                cells.append((c.column - 1, v))
        if not cells:
            continue
        if len(cells) > richest_n:
            richest_n, richest_cells = len(cells), cells
        upper = {}
        for idx, v in cells:
            upper.setdefault(v.upper(), idx)
        fmap = {}
        for field in required:
            for al in column_dict.get(field, ()):
                if al in upper:
                    fmap[field] = upper[al]
                    break
        score = len(fmap)
        if score > best_score:
            best_score, best_rownum, best_map, best_cells = score, _row_number(row), fmap, cells
        if score == len(required):
            break

    # Se non troviamo almeno 2 campi non abbiamo una vera riga di intestazione:
    # usiamo come candidata la riga con più celle di testo (per i suggerimenti).
    if best_score >= 2:
        header_cells, fmap, rownum = best_cells, best_map, best_rownum
    else:
        header_cells, fmap, rownum = richest_cells, {}, None

    missing = [f for f in required if f not in fmap]
    if missing:
        raise UnmappedColumnsError(fmap, missing, header_cells,
                                   _suggest_columns(missing, header_cells, column_dict))
    return rownum + 1, fmap


REQUIRED_FIELDS_ORDER = ('unique_number', 'material_code', 'quantity')


def _is_unique_code(s: str) -> bool:
    """True se `s` e' un reel/unique code plausibile (HU/ES/numerico, con o senza
    suffisso _NN) e NON un numero d'ordine 'PRnnn...'. Sostituisce il vecchio
    doppio controllo HU-a-9-cifre / 6-cifre, troppo rigido per i file reali."""
    return bool(RE_UNIQUE_GENERIC.match(s)) and not RE_ORDERS_COMPACT.match(s)


def _row_number(row) -> int:
    """Numero di riga (1-based) ricavato da una cella non vuota della riga. In
    openpyxl read-only le celle vuote iniziali (es. colonna A vuota nei file
    multi-job) sono 'EmptyCell' e NON hanno l'attributo .row: usare row[0].row
    solleverebbe AttributeError, quindi si cerca la prima cella con .row."""
    for c in row:
        if getattr(c, 'row', None) is not None:
            return c.row
    return 1


def _parse_qty(raw):
    """Converte la cella quantita' in float. Le celle numeriche (openpyxl) sono
    gia' numeriche; per le stringhe si assume il formato europeo (punto =
    migliaia, virgola = decimali). Solleva ValueError su valori negativi/invalidi."""
    qty = (float(str(raw).replace('.', '').replace(',', '.'))
           if isinstance(raw, str) else float(raw))
    if qty < 0:
        raise ValueError
    return qty


def _new_result(path, compact, orders) -> EssegiFile:
    return EssegiFile(
        file_path=path,
        file_name=os.path.basename(path),
        file_hash=file_sha256(path),
        file_date=datetime.fromtimestamp(os.path.getmtime(path)),
        orders_compact=compact,
        orders=orders,
    )


def parse_essegi_file(path: str, column_dict=None) -> EssegiFile:
    """
    Parsa un file lista prelievo di kitting.

    Ordine (INVARIATO): intestazione compatta 'PRnnn/nnn/...' (multi-job) oppure
    testata a singolo ordine 'PR0000nnn_FASE'.

    Colonne (DIZIONARIO): le colonne REEL CODE / ITEM CODE / QT-QUANTITY vengono
    cercate per nome tramite `column_dict` (campo logico -> alias), indipendente
    dalla posizione e dal formato. Così i vari tracciati (multi-job, D365,
    ibridi) sono gestiti dallo stesso codice.

    Se una colonna richiesta non è mappata, solleva UnmappedColumnsError (con i
    suggerimenti) invece di bloccare: il chiamante può proporre la mappatura.

    `column_dict`: dict campo -> iterabile di alias. Se None, usa i default.
    Le anomalie non bloccanti finiscono in .warnings.
    """
    from openpyxl import load_workbook
    if not os.path.isfile(path):
        raise EssegiParseError(f"File non trovato: {path}")

    cd = _normalize_column_dict(column_dict)
    # Valori di intestazione noti (tutti gli alias): per saltare in silenzio le
    # ripetizioni di testata a cambio pagina, qualunque sia il nome della colonna.
    header_values = set()
    for aliases in cd.values():
        header_values |= set(aliases)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active

        # ── Ordine (logica invariata) ────────────────────────────────────
        compact, _ = _find_orders_header(ws)
        if compact:
            orders = expand_orders(compact)
            orders_compact = compact
            if not orders:
                raise EssegiParseError(f"Nessun ordine valido nell'intestazione: '{compact}'")
        else:
            title, digits = _find_d365_order(ws)
            if not digits:
                raise EssegiParseError(
                    "Numero ordine non trovato (né 'PRnnn/nnn/...' né testata 'PR0000nnn...').")
            orders = [normalize_order(digits)]
            orders_compact = title or orders[0]

        # ── Colonne (dizionario) ─────────────────────────────────────────
        data_start, fmap = find_data_columns(ws, cd)   # può sollevare UnmappedColumnsError
        col_u = fmap['unique_number']
        col_c = fmap['material_code']
        col_q = fmap['quantity']

        result = _new_result(path, orders_compact, orders)

        def cell(row, idx):
            return row[idx].value if (idx is not None and len(row) > idx) else None

        for row in ws.iter_rows(min_row=data_start):
            raw = cell(row, col_u)
            unique = str(raw).strip() if raw is not None else ''
            if not unique:
                continue
            if not _is_unique_code(unique):
                # ripetizioni di intestazione a cambio pagina: ignora in silenzio
                if unique.upper() not in header_values and not RE_ORDERS_COMPACT.match(unique):
                    result.warnings.append(
                        f"Riga {_row_number(row)}: '{unique}' non riconosciuto come reel code, riga saltata")
                continue

            raw_code = cell(row, col_c)
            code = str(raw_code).strip() if raw_code is not None else ''
            if not code:
                result.warnings.append(f"Riga {_row_number(row)}: codice materiale mancante, riga saltata")
                continue

            raw_qty = cell(row, col_q)
            try:
                qty = _parse_qty(raw_qty)
            except (TypeError, ValueError):
                result.warnings.append(
                    f"Riga {_row_number(row)}: quantita' non valida ({raw_qty!r}), riga saltata")
                continue

            result.rows.append(EssegiRow(
                row_number=_row_number(row),
                unique_number=unique,
                material_code=code,
                quantity=qty,
            ))

        if not result.rows:
            raise EssegiParseError("Nessuna riga materiale valida nel file")

        logger.info("Parsed %s: ordini=%s righe=%d warnings=%d (col u=%s c=%s q=%s)",
                    result.file_name, result.orders, len(result.rows),
                    len(result.warnings), col_u, col_c, col_q)
        return result
    finally:
        wb.close()
