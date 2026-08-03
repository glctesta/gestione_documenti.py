# -*- coding: utf-8 -*-
"""
fai_autocheck.py — FAI Autocheck da PlanningMachine

Background worker che:
1. Ogni 30 minuti legge il file Excel più recente in T:\Planning\
2. Analizza tab PlanningMachine per produzioni nelle prossime 4 ore
3. Verifica corrispondenza con template FAI Autocheck=true
4. Se produzione non avviata → email preventiva ai responsabili in turno
5. Registra eventi in FaiAutocheckNotifications per anti-duplicazione
"""

import os
import glob
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional

logger = logging.getLogger("TraceabilityRS")

# ================================================================
# COSTANTI
# ================================================================

PLANNING_PATH = r"T:\Planning"
PLANNING_TAB = "PlanningMachine"
COL_PHASE = 4          # colonna E (0-based)
COL_ORDER_NUMBER = 10  # colonna K (0-based)
COL_PLANNED_START = 14 # colonna O (0-based)

LOOKAHEAD_HOURS = 4    # finestra di controllo
INTERVAL_MINUTES = 30  # intervallo tra i cicli


# ================================================================
# 1. TEMPLATE AUTOCHECK
# ================================================================

SQL_AUTOCHECK_TEMPLATES = """
    SELECT f.[FaiTemplateId], f.[NrDocument], f.[Revision],
           f.[FaiTitle], p.[PhaseName], p.[IdPhase]
    FROM [Traceability_RS].[fai].[FaiTemplates] f
    INNER JOIN [Traceability_RS].[dbo].[Phases] p
        ON p.[IdPhase] = f.[IdPhase]
    WHERE f.[Autocheck] = 1
"""


def get_autocheck_templates(conn) -> Dict[str, dict]:
    """
    Restituisce dict {PhaseName_upper: template_info}
    per match rapido con il file Excel.
    """
    templates = {}
    with conn.cursor() as cur:
        cur.execute(SQL_AUTOCHECK_TEMPLATES)
        for r in cur.fetchall():
            key = (r.PhaseName or '').strip().upper()
            templates[key] = {
                'FaiTemplateId': r.FaiTemplateId,
                'NrDocument': r.NrDocument,
                'Revision': r.Revision,
                'FaiTitle': r.FaiTitle,
                'PhaseName': r.PhaseName,
                'IdPhase': r.IdPhase
            }
    logger.info(f"FAI Autocheck: {len(templates)} template con Autocheck=1")
    return templates


# ================================================================
# 2. LETTURA FILE EXCEL
# ================================================================

def _find_latest_excel(folder: str) -> Optional[str]:
    """Trova il file Excel più recente nella cartella per data modifica.

    Esclude i file lock di Office (basename che inizia con '~$'): questi
    vengono creati quando un utente apre il file in Excel e causano
    PermissionError se aperti con openpyxl.
    """
    patterns = [os.path.join(folder, '*.xlsx'), os.path.join(folder, '*.xls')]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))

    # Escludi file lock di Office (~$filename.xlsx)
    files = [f for f in files if not os.path.basename(f).startswith('~$')]

    if not files:
        logger.warning(f"FAI Autocheck: nessun file Excel in {folder}")
        return None

    latest = max(files, key=os.path.getmtime)
    logger.info(f"FAI Autocheck: file selezionato: {os.path.basename(latest)}")
    return latest


def read_planning_excel(lookback_hours: int = 0,
                        lookahead_hours: int = None) -> List[dict]:
    """
    Legge il tab PlanningMachine dal file Excel più recente.
    Filtra righe con PlannedStart tra (now - lookback_hours) e (now + lookahead_hours).
    Restituisce lista di dict con phase, order_number, planned_start.

    Args:
        lookback_hours: ore nel passato da includere (default 0 = solo futuro).
                        Usato dall'enforcement per catturare ordini il cui
                        PlannedStart è appena passato (es. per L3 escalation).
        lookahead_hours: ore nel futuro (default LOOKAHEAD_HOURS = 4).
                        La form di compilazione usa una finestra più ampia:
                        il FAI si fa quando la linea viene preparata, anche
                        molte ore prima dell'avvio dell'ordine.
    """
    filepath = _find_latest_excel(PLANNING_PATH)
    if not filepath:
        return []

    if lookahead_hours is None:
        lookahead_hours = LOOKAHEAD_HOURS

    now = datetime.now()
    earliest = now - timedelta(hours=lookback_hours)
    cutoff = now + timedelta(hours=lookahead_hours)

    rows = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if PLANNING_TAB not in wb.sheetnames:
            logger.error(
                f"FAI Autocheck: tab '{PLANNING_TAB}' non trovato in {filepath}")
            wb.close()
            return []

        ws = wb[PLANNING_TAB]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Salta righe troppo corte
            if not row or len(row) <= COL_PLANNED_START:
                continue

            phase_raw = row[COL_PHASE]
            order_raw = row[COL_ORDER_NUMBER]
            start_raw = row[COL_PLANNED_START]

            if not phase_raw or not order_raw or not start_raw:
                continue

            # Parsing data/ora
            planned_start = None
            if isinstance(start_raw, datetime):
                planned_start = start_raw
            elif isinstance(start_raw, str):
                for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M',
                            '%Y-%m-%d %H:%M', '%m/%d/%Y %H:%M:%S'):
                    try:
                        planned_start = datetime.strptime(start_raw.strip(), fmt)
                        break
                    except ValueError:
                        continue
            if not planned_start:
                continue

            # Filtra per finestra temporale: earliest ≤ PlannedStart ≤ cutoff
            if not (earliest <= planned_start <= cutoff):
                continue

            rows.append({
                'phase': str(phase_raw).strip(),
                'order_number': str(order_raw).strip(),
                'planned_start': planned_start,
                'row_idx': row_idx
            })

        wb.close()

    except PermissionError as e:
        # Il file e' aperto in Excel da un utente (lock): log senza stacktrace
        # e riprova al prossimo ciclo.
        logger.warning(
            f"FAI Autocheck: file Excel temporaneamente in lock "
            f"(aperto da un utente): {os.path.basename(filepath)} — "
            f"riprovo al prossimo ciclo. Dettaglio: {e}"
        )
        return []
    except Exception as e:
        logger.error(f"FAI Autocheck: errore lettura Excel: {e}", exc_info=True)

    logger.info(f"FAI Autocheck: {len(rows)} righe valide in finestra "
                f"[-{lookback_hours}h, +{lookahead_hours}h]")
    return rows


# ================================================================
# 2-bis. CODA ORDINI IN ATTESA DELLA LINEA PTH
# ================================================================

# Fasi coinvolte (Traceability_RS.dbo.Phases)
PHASE_AOI = 2    # AOI dopo SMT
PHASE_PTHM = 4   # PTHM (montaggio manuale)

# Ordini per cui il FAI "3 ore prima" ha senso: la linea PTH deve essere
# preparata per riceverli.
#   1. almeno una scheda ha superato l'AOI dell'SMT  → il materiale esiste
#   2. la fase PTHM non e' ancora iniziata           → la linea li deve ancora ricevere
#   3. l'ordine non e' completato                    → versato a magazzino < qta ordine
# La quantita' versata arriva da D365 (LogApiDynamics / ProdFinishedGoods):
# in Traceability non esiste un campo di versamento valorizzato.
PTH_PENDING_ORDERS_QUERY = """
SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
WITH WH AS (
    SELECT JSON_VALUE(L.MessageSend,'$.Message.Reference') AS OrderNumber,
           SUM(TRY_CAST(JSON_VALUE(j.value,'$.RealValue') AS int)) AS QtyWarehouse
    FROM Traceability_RS.dbo.LogApiDynamics L
    CROSS APPLY OPENJSON(L.MessageSend,
        '$.Message.KeyValue.ListValue[0].ListValue[0].ListValue') j
    WHERE L.EndPointName = 'ProdFinishedGoods'
      AND JSON_VALUE(j.value,'$.Key') = 'GoodQty'
    GROUP BY JSON_VALUE(L.MessageSend,'$.Message.Reference')
)
SELECT o.IDOrder, o.OrderNumber, p.ProductCode, p.ProductName,
       o.OrderQuantity, ISNULL(wh.QtyWarehouse, 0) AS QtyWarehouse,
       aoi.BoardsAoi, aoi.LastAoi
FROM Traceability_RS.dbo.Orders o
INNER JOIN Traceability_RS.dbo.Products p ON p.IDProduct = o.IDProduct
CROSS APPLY (
    SELECT COUNT(DISTINCT s.IDBoard) AS BoardsAoi, MAX(s.ScanTimeFinish) AS LastAoi
    FROM Traceability_RS.dbo.Scannings s
    INNER JOIN Traceability_RS.dbo.OrderPhases op ON op.IDOrderPhase = s.IDOrderPhase
    WHERE op.IDOrder = o.IDOrder
      AND op.IDPhase = {aoi}
      AND s.ScanTimeFinish IS NOT NULL
) aoi
LEFT JOIN WH wh ON wh.OrderNumber = o.OrderNumber
WHERE aoi.BoardsAoi > 0
  AND NOT EXISTS (
        SELECT 1
        FROM Traceability_RS.dbo.Scannings s2
        INNER JOIN Traceability_RS.dbo.OrderPhases op2 ON op2.IDOrderPhase = s2.IDOrderPhase
        WHERE op2.IDOrder = o.IDOrder AND op2.IDPhase = {pthm})
  AND ISNULL(wh.QtyWarehouse, 0) < o.OrderQuantity
  {age_filter}
ORDER BY aoi.LastAoi DESC;
SET TRANSACTION ISOLATION LEVEL READ COMMITTED;
"""


def get_pth_pending_orders(conn, lookback_days: int = None) -> Dict[str, dict]:
    """Ordini in attesa della linea PTH (vedi PTH_PENDING_ORDERS_QUERY).

    Args:
        lookback_days: se valorizzato, considera solo gli ordini con AOI
                       completato negli ultimi N giorni (esclude i residui
                       fermi da mesi).
    Returns:
        {OrderNumber: {IDOrder, OrderNumber, ProductCode, ProductName,
                       OrderQuantity, QtyWarehouse, BoardsAoi, LastAoi}}
    """
    age_filter = (f"AND aoi.LastAoi >= DATEADD(day, -{int(lookback_days)}, GETDATE())"
                  if lookback_days else "")
    query = PTH_PENDING_ORDERS_QUERY.format(
        aoi=PHASE_AOI, pthm=PHASE_PTHM, age_filter=age_filter)
    orders = {}
    try:
        with conn.cursor() as cur:
            cur.execute(query)
            for r in cur.fetchall():
                orders[r.OrderNumber] = {
                    'IDOrder': r.IDOrder,
                    'OrderNumber': r.OrderNumber,
                    'ProductCode': r.ProductCode,
                    'ProductName': r.ProductName,
                    'OrderQuantity': r.OrderQuantity,
                    'QtyWarehouse': r.QtyWarehouse,
                    'BoardsAoi': r.BoardsAoi,
                    'LastAoi': r.LastAoi,
                }
        logger.info(f"FAI Autocheck: {len(orders)} ordini in attesa della linea PTH "
                    f"(AOI fatto, PTHM non iniziata, non completati)")
    except Exception as e:
        logger.error(f"FAI Autocheck: errore get_pth_pending_orders: {e}", exc_info=True)
    return orders


# ================================================================
# 3. VERIFICA PRODUZIONE AVVIATA
# ================================================================

SQL_CHECK_PRODUCTION = """
    SELECT COUNT(DISTINCT Traceability_RS.dbo.BoardLabels(Scannings.IDBoard)) AS Qty
    FROM Traceability_RS.dbo.Scannings
    INNER JOIN Traceability_RS.dbo.OrderPhases
        ON Scannings.IDOrderPhase = OrderPhases.IDOrderPhase
    INNER JOIN Traceability_RS.dbo.Orders
        ON OrderPhases.IDOrder = Orders.IDOrder
    INNER JOIN Traceability_RS.dbo.Phases
        ON OrderPhases.IDPhase = Phases.IDPhase
    INNER JOIN Traceability_RS.dbo.Boards
        ON Boards.IDBoard = Scannings.IDBoard
    WHERE Scannings.ScanTimeFinish BETWEEN GETDATE() - 500
        AND CAST(CAST(CAST(GETDATE() AS date) AS nvarchar(10))
            + ' 07:30:00' AS smalldatetime)
        AND Orders.OrderNumber = ?
        AND Phases.IdPhase = ?
"""


def check_production_started(conn, order_number: str, id_phase: int) -> int:
    """Verifica se la produzione è già avviata. Ritorna Qty (0 = non avviata)."""
    with conn.cursor() as cur:
        cur.execute(SQL_CHECK_PRODUCTION, (order_number, id_phase))
        row = cur.fetchone()
    return int(row.Qty or 0) if row else 0


# ================================================================
# 4. DESTINATARI EMAIL + VERIFICA PRESENZA
# ================================================================

SQL_RECIPIENTS = """
    SELECT e.EmployeeSurname + ' ' + e.EmployeeName AS Employee,
           a.WorkEmail,
           f.FunctionCode,
           ee.IDEmployee,
           cs.SubCdcDescription
    FROM Employee.dbo.EmployeeHireHistory h
    LEFT JOIN Employee.dbo.Employees e
        ON e.EmployeeId = h.EmployeeId
       AND h.EmployeerId = 2
       AND h.EndWorkDate IS NULL
    INNER JOIN Employee.dbo.EmployeeCdcStories ec
        ON h.EmployeeHireHistoryId = ec.EmployeeHireHistoryId
       AND ec.DateOut IS NULL
    INNER JOIN Employee.dbo.Functions f
        ON ec.FunctionId = f.FunctionId
    INNER JOIN Employee.dbo.CdcSub cs
        ON ec.SubCdcId = cs.SubCdcId
    INNER JOIN Employee.dbo.EmployeeAddress a
        ON e.EmployeeId = a.EmployeeId
       AND a.DateOut IS NULL
    INNER JOIN Timeclocking.dbo.Employee ee
        ON ee.UniqueID COLLATE database_default = e.EmployeeNID
       AND ee.DataStop IS NULL
    WHERE cs.SubCdcDescription = 'pthm'
      AND f.FunctionCode BETWEEN 21 AND 80
    ORDER BY f.FunctionCode
"""


def _check_presence(conn, id_employee: int) -> bool:
    """
    Verifica presenza in turno via SP GetEmployeesTimeclockReal.
    - Prima delle 15:30 → @from = oggi 06:40
    - Dopo le 15:30 → @from = oggi 16:20
    - @to = GETDATE()
    """
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')

    if now.hour < 15 or (now.hour == 15 and now.minute < 30):
        from_dt = f"{today_str} 06:40:00"
    else:
        from_dt = f"{today_str} 16:20:00"

    to_dt = now.strftime('%Y-%m-%d %H:%M:%S')

    try:
        with conn.cursor() as cur:
            cur.execute(
                "EXEC [Timeclocking].[dbo].[GetEmployeesTimeclockReal] ?, ?, ?",
                (from_dt, to_dt, id_employee))
            result = cur.fetchone()
        return result is not None
    except Exception as e:
        logger.warning(f"FAI Autocheck: errore verifica presenza ID {id_employee}: {e}")
        return False


def get_recipients_with_presence(conn) -> Tuple[List[str], List[str]]:
    """
    Restituisce (to_list, cc_list) di email.
    - FunctionCode < 60 e presente in turno → TO
    - FunctionCode >= 60 → sempre in CC
    """
    to_list = []
    cc_list = []

    with conn.cursor() as cur:
        cur.execute(SQL_RECIPIENTS)
        rows = cur.fetchall()

    for r in rows:
        email = (r.WorkEmail or '').strip()
        if not email or '@' not in email:
            continue

        fc = r.FunctionCode or 0
        if fc >= 60:
            cc_list.append(email)
        else:
            # Verifica presenza
            if _check_presence(conn, r.IDEmployee):
                to_list.append(email)
            else:
                logger.debug(
                    f"FAI Autocheck: {r.Employee} non in turno, escluso da TO")

    # Deduplica
    to_list = list(dict.fromkeys(to_list))
    cc_list = list(dict.fromkeys(cc_list))

    logger.info(f"FAI Autocheck: destinatari TO={len(to_list)}, CC={len(cc_list)}")
    return to_list, cc_list


# ================================================================
# 5. ANTI-DUPLICAZIONE
# ================================================================

SQL_CHECK_ALREADY_SENT = """
    SELECT 1
    FROM [Traceability_RS].[fai].[FaiAutocheckNotifications]
    WHERE OrderNumber = ?
      AND IdPhase = ?
      AND FaiTemplateId = ?
      AND PlannedStart = ?
      AND NotificationStatus IN ('SENT', 'SKIPPED_ALREADY_STARTED', 'SENDING')
"""


def check_already_notified(conn, order_number: str, id_phase: int,
                           template_id: int, planned_start: datetime) -> bool:
    """Verifica se esiste già una notifica per questa combinazione.

    E' solo un filtro rapido: la garanzia anti-duplicato e' il claim atomico
    (claim_notification), perche' il ciclo gira in parallelo su piu' PC.
    """
    with conn.cursor() as cur:
        cur.execute(SQL_CHECK_ALREADY_SENT,
                    (order_number, id_phase, template_id, planned_start))
        return cur.fetchone() is not None


# ── Claim atomico: prenota l'invio PRIMA di mandare l'email ──────────────────
SQL_CLAIM_NOTIFICATION = """
    INSERT INTO [Traceability_RS].[fai].[FaiAutocheckNotifications]
        (OrderNumber, IdPhase, PhaseName, FaiTemplateId, FaiTitle,
         NrDocument, Revision, PlannedStart, DetectionTime,
         EmailSentTime, EmailTo, EmailCc, ProductionQtyAtCheck,
         PresenceChecked, NotificationStatus)
    SELECT ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), NULL, '', '', 0, 1, 'SENDING'
    WHERE NOT EXISTS (
        SELECT 1 FROM [Traceability_RS].[fai].[FaiAutocheckNotifications]
        WITH (UPDLOCK, HOLDLOCK)
        WHERE OrderNumber = ?
          AND IdPhase = ?
          AND FaiTemplateId = ?
          AND PlannedStart = ?
          AND NotificationStatus IN ('SENT', 'SKIPPED_ALREADY_STARTED', 'SENDING')
    )
"""

SQL_FINALIZE_NOTIFICATION = """
    UPDATE [Traceability_RS].[fai].[FaiAutocheckNotifications]
    SET NotificationStatus = 'SENT',
        EmailSentTime = ?,
        EmailTo = ?,
        EmailCc = ?
    WHERE OrderNumber = ? AND IdPhase = ? AND FaiTemplateId = ?
      AND PlannedStart = ? AND NotificationStatus = 'SENDING'
"""

SQL_RELEASE_NOTIFICATION = """
    DELETE FROM [Traceability_RS].[fai].[FaiAutocheckNotifications]
    WHERE OrderNumber = ? AND IdPhase = ? AND FaiTemplateId = ?
      AND PlannedStart = ? AND NotificationStatus = 'SENDING'
"""


def claim_notification(conn, order_number: str, id_phase: int, template_id: int,
                       planned_start: datetime, template: dict) -> bool:
    """
    Prenota in modo ATOMICO l'invio della notifica FAI per questa combinazione
    (ordine + fase + template + inizio pianificato).

    Ritorna True solo se il claim e' stato ottenuto adesso: senza questo, tutte
    le istanze del ciclo in esecuzione sui vari PC superano insieme il controllo
    "gia' notificato?" e inviano una copia a testa della stessa segnalazione.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(SQL_CLAIM_NOTIFICATION, (
                order_number, id_phase, template.get('PhaseName'), template_id,
                template.get('FaiTitle'), template.get('NrDocument'),
                template.get('Revision'), planned_start,
                order_number, id_phase, template_id, planned_start
            ))
            claimed = cur.rowcount == 1
        conn.commit()
        return claimed
    except Exception as e:
        logger.info(
            f"FAI Autocheck: claim non ottenuto per {order_number}/{id_phase}: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return False


def finalize_notification(conn, order_number: str, id_phase: int, template_id: int,
                          planned_start: datetime, email_time, to_list, cc_list) -> None:
    """Chiude il claim come inviato, registrando orario e destinatari."""
    try:
        with conn.cursor() as cur:
            cur.execute(SQL_FINALIZE_NOTIFICATION, (
                email_time, '; '.join(to_list), '; '.join(cc_list),
                order_number, id_phase, template_id, planned_start
            ))
        conn.commit()
    except Exception as e:
        logger.error(
            f"FAI Autocheck: errore chiusura notifica {order_number}/{id_phase}: {e}",
            exc_info=True)
        try:
            conn.rollback()
        except Exception:
            pass


def release_notification_claim(conn, order_number: str, id_phase: int,
                               template_id: int, planned_start: datetime) -> None:
    """Rilascia il claim (invio fallito) cosi' il ciclo successivo ritenta."""
    try:
        with conn.cursor() as cur:
            cur.execute(SQL_RELEASE_NOTIFICATION,
                        (order_number, id_phase, template_id, planned_start))
        conn.commit()
    except Exception as e:
        logger.warning(
            f"FAI Autocheck: impossibile rilasciare il claim {order_number}/{id_phase}: {e}")
        try:
            conn.rollback()
        except Exception:
            pass


# ================================================================
# 6. REGISTRAZIONE TRACKING
# ================================================================

SQL_INSERT_NOTIFICATION = """
    INSERT INTO [Traceability_RS].[fai].[FaiAutocheckNotifications]
        (OrderNumber, IdPhase, PhaseName, FaiTemplateId, FaiTitle,
         NrDocument, Revision, PlannedStart, DetectionTime,
         EmailSentTime, EmailTo, EmailCc, ProductionQtyAtCheck,
         PresenceChecked, NotificationStatus)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), ?, ?, ?, ?, 1, ?)
"""


def record_notification(conn, data: dict):
    """Registra l'evento nella tabella tracking.

    Rollback difensivo se l'INSERT fallisce: evita che una transazione
    residua tenga lock esclusivi sulla tabella FaiAutocheckNotifications.
    Con autocommit=True il rollback e' un no-op ma serve da rete di sicurezza
    qualora la connessione venga aperta con autocommit=False.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(SQL_INSERT_NOTIFICATION, (
                data['order_number'],
                data['id_phase'],
                data['phase_name'],
                data['template_id'],
                data.get('fai_title'),
                data.get('nr_document'),
                data.get('revision'),
                data['planned_start'],
                data.get('email_sent_time'),      # NULL se non inviata
                data.get('email_to', ''),
                data.get('email_cc', ''),
                data.get('production_qty', 0),
                data['status']
            ))
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise


# ================================================================
# 7. INVIO EMAIL
# ================================================================

def send_fai_autocheck_email(to_emails: List[str], cc_emails: List[str],
                             order_data: dict, template_data: dict,
                             logo_path: str = "Logo.png"):
    """Compone e invia email professionale per FAI autocheck."""
    from email_connector import EmailSender

    order_number = order_data['order_number']
    phase_name = template_data['PhaseName']
    planned_start = order_data['planned_start']
    fai_title = template_data.get('FaiTitle', 'N/A')
    nr_doc = template_data.get('NrDocument', 'N/A')
    revision = template_data.get('Revision', 'N/A')

    planned_str = planned_start.strftime('%d/%m/%Y %H:%M')

    subject = (f"Azione richiesta — Esecuzione controllo FAI prima "
               f"dell'avvio produzione ordine {order_number}")

    html_body = f"""
    <html>
    <body style="font-family:'Segoe UI',Arial,sans-serif; color:#333; margin:0; padding:0;">
    <div style="max-width:700px; margin:0 auto; padding:20px;">

        <!-- Header -->
        <div style="border-bottom:3px solid #B71C1C; padding-bottom:15px; margin-bottom:20px;">
            <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
                <td style="font-size:20px; font-weight:bold; color:#B71C1C;">
                    ⚠️ CONTROLLO FAI RICHIESTO
                </td>
                <td style="text-align:right;">
                    <img src="cid:company_logo" alt="Vandewiele"
                         style="width:120px; height:auto;" />
                </td>
            </tr>
            </table>
        </div>

        <p style="font-size:14px;">Gentili Responsabili di Linea,</p>

        <p style="font-size:14px;">
            si comunica che per l'ordine <strong>{order_number}</strong>,
            fase <strong>{phase_name}</strong>, è previsto a breve l'avvio
            della produzione secondo pianificazione.
        </p>

        <p style="font-size:14px;">
            Dalle verifiche automatiche effettuate, il controllo FAI associato
            al template obbligatorio con gestione Autocheck
            <strong>non risulta ancora eseguito</strong>.
        </p>

        <p style="font-size:14px; font-weight:bold; color:#B71C1C;">
            Si richiede pertanto di provvedere con la massima urgenza
            all'esecuzione e registrazione del controllo FAI prima
            dell'inizio della produzione.
        </p>

        <!-- Dettagli -->
        <table style="border-collapse:collapse; width:100%; margin:20px 0;
                       background-color:#fff3cd; border:1px solid #ffc107;">
            <tr>
                <td style="padding:10px 14px; font-weight:bold; width:250px;
                           border-bottom:1px solid #ffeaa7;">
                    📅 Orario pianificato avvio fase
                </td>
                <td style="padding:10px 14px; border-bottom:1px solid #ffeaa7;
                           font-size:16px; font-weight:bold; color:#B71C1C;">
                    {planned_str}
                </td>
            </tr>
            <tr>
                <td style="padding:10px 14px; font-weight:bold;
                           border-bottom:1px solid #ffeaa7;">
                    📋 Template FAI applicabile
                </td>
                <td style="padding:10px 14px;
                           border-bottom:1px solid #ffeaa7;">{fai_title}</td>
            </tr>
            <tr>
                <td style="padding:10px 14px; font-weight:bold;">
                    📄 Documento / Revisione
                </td>
                <td style="padding:10px 14px;">{nr_doc} / {revision}</td>
            </tr>
        </table>

        <div style="background-color:#f8d7da; border-left:4px solid #B71C1C;
                    padding:12px 16px; margin:20px 0; border-radius:4px;">
            <p style="margin:0; font-size:13px; color:#721c24;">
                La presente comunicazione costituisce avviso operativo preventivo.
                L'eventuale mancata esecuzione del controllo sarà registrata
                ai fini di verifica del rispetto della procedura.
            </p>
        </div>

        <!-- Footer -->
        <div style="margin-top:30px; padding-top:15px; border-top:1px solid #dee2e6;">
            <p style="font-size:12px; color:#666;">
                Cordiali saluti,<br/>
                <strong>Sistema automatico di controllo FAI</strong>
            </p>
            <p style="font-size:10px; color:#aaa;">
                Questo messaggio è stato generato automaticamente dal sistema
                TraceabilityRS. Non rispondere a questa email.
            </p>
        </div>
    </div>
    </body>
    </html>
    """

    sender = EmailSender()
    sender.save_credentials("Accounting@Eutron.it", "9jHgFhSs7Vf+")

    attachments = []
    full_logo = os.path.join(os.path.dirname(__file__), logo_path)
    if os.path.exists(full_logo):
        attachments.append(('inline', full_logo, 'company_logo'))

    sender.send_email(
        to_email=', '.join(to_emails),
        subject=subject,
        body=html_body,
        is_html=True,
        attachments=attachments if attachments else None,
        cc_emails=cc_emails if cc_emails else None
    )

    logger.info(f"FAI Autocheck email inviata per ordine {order_number} "
                f"fase {phase_name} a {len(to_emails)} dest. TO, "
                f"{len(cc_emails)} CC")


# ================================================================
# 8. CICLO PRINCIPALE
# ================================================================

def _get_product_code(conn, order_number: str) -> Optional[str]:
    """Codice prodotto di un ordine (dbo.Orders -> dbo.Products), o None."""
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT TOP 1 p.ProductCode FROM dbo.Orders o "
                "INNER JOIN dbo.Products p ON p.IDProduct = o.IDProduct "
                "WHERE o.OrderNumber = ?", (order_number,))
            row = cur.fetchone()
            return row[0] if row and row[0] else None
    except Exception as e:
        logger.warning(f"FAI Autocheck: product code non risolto per {order_number}: {e}")
        return None


def run_autocheck_cycle(conn, logo_path: str = "Logo.png") -> int:
    """
    Esegue un ciclo completo di autocheck.
    Restituisce il numero di email inviate.
    """
    sent_count = 0

    # 1. Carica template autocheck
    templates = get_autocheck_templates(conn)
    if not templates:
        logger.info("FAI Autocheck: nessun template con Autocheck=1")
        return 0

    # 2. Leggi file Excel
    planning_rows = read_planning_excel()
    if not planning_rows:
        logger.info("FAI Autocheck: nessuna riga valida nel planning")
        return 0

    # 2b. Delay per codice prodotto (forno wave): per i codici presenti in
    #     fai.FaiCodeDelay la segnalazione FAI viene rimandata di N minuti a
    #     prescindere (grazia dal primo avvistamento). Se la lista è vuota,
    #     nessun effetto (comportamento invariato).
    try:
        import fai_code_delay
        delay_map = fai_code_delay.get_delay_map(conn)
        if delay_map:
            fai_code_delay.cleanup_pending(conn)
    except Exception as e:
        logger.warning(f"FAI Autocheck: delay codici non caricato: {e}")
        delay_map = {}
    product_code_cache = {}

    # 3. Per ogni riga valida
    # Cache destinatari (calcolati una volta sola per ciclo)
    recipients_cache = None

    for pr in planning_rows:
        phase_upper = pr['phase'].upper()

        # 3a. Match fase ← autocheck template
        template = templates.get(phase_upper)
        if not template:
            continue  # fase non soggetta ad autocheck

        order_number = pr['order_number']
        id_phase = template['IdPhase']
        template_id = template['FaiTemplateId']
        planned_start = pr['planned_start']

        # 3a-bis. Delay forno wave: se il codice prodotto dell'ordine è nella
        # lista dei delay, la segnalazione FAI è rimandata di N minuti a
        # prescindere (grazia dal primo avvistamento dell'ordine/fase). Nessuna
        # query se la lista è vuota (comportamento invariato).
        if delay_map:
            pcode = product_code_cache.get(order_number, False)
            if pcode is False:
                pcode = _get_product_code(conn, order_number)
                product_code_cache[order_number] = pcode
            delay_min = delay_map.get((pcode or '').strip().upper(), 0)
            if delay_min:
                try:
                    if fai_code_delay.check_delay_grace(
                            conn, order_number, id_phase, planned_start, delay_min):
                        logger.info(
                            "FAI Autocheck: %s (prodotto %s) in grazia %d min, "
                            "segnalazione rimandata", order_number, pcode, delay_min)
                        continue
                except Exception as e:
                    logger.warning(
                        f"FAI Autocheck: grazia delay non applicata per {order_number}: {e}")

        # 3b. Verifica anti-duplicazione
        try:
            if check_already_notified(conn, order_number, id_phase,
                                      template_id, planned_start):
                logger.debug(
                    f"FAI Autocheck: skip duplicato {order_number}/{phase_upper}")
                continue
        except Exception as e:
            logger.warning(f"FAI Autocheck: errore anti-dup check: {e}")

        # 3c. Verifica produzione avviata
        try:
            qty = check_production_started(conn, order_number, id_phase)
        except Exception as e:
            logger.error(
                f"FAI Autocheck: errore verifica produzione: {e}", exc_info=True)
            qty = 0

        if qty > 0:
            # Produzione già avviata → registra e skip
            try:
                record_notification(conn, {
                    'order_number': order_number,
                    'id_phase': id_phase,
                    'phase_name': template['PhaseName'],
                    'template_id': template_id,
                    'fai_title': template.get('FaiTitle'),
                    'nr_document': template.get('NrDocument'),
                    'revision': template.get('Revision'),
                    'planned_start': planned_start,
                    'email_sent_time': None,
                    'email_to': '',
                    'email_cc': '',
                    'production_qty': qty,
                    'status': 'SKIPPED_ALREADY_STARTED'
                })
            except Exception as e:
                logger.warning(
                    f"FAI Autocheck: errore registrazione skip: {e}")
            continue

        # 3d. Recupera destinatari (con cache)
        if recipients_cache is None:
            try:
                recipients_cache = get_recipients_with_presence(conn)
            except Exception as e:
                logger.error(
                    f"FAI Autocheck: errore recupero destinatari: {e}",
                    exc_info=True)
                recipients_cache = ([], [])

        to_list, cc_list = recipients_cache

        if not to_list:
            # Nessun destinatario in turno
            try:
                record_notification(conn, {
                    'order_number': order_number,
                    'id_phase': id_phase,
                    'phase_name': template['PhaseName'],
                    'template_id': template_id,
                    'fai_title': template.get('FaiTitle'),
                    'nr_document': template.get('NrDocument'),
                    'revision': template.get('Revision'),
                    'planned_start': planned_start,
                    'email_sent_time': None,
                    'email_to': '',
                    'email_cc': '; '.join(cc_list),
                    'production_qty': 0,
                    'status': 'SKIPPED_NO_RECIPIENT'
                })
            except Exception as e:
                logger.warning(
                    f"FAI Autocheck: errore registrazione no-recipient: {e}")
            logger.warning(
                f"FAI Autocheck: nessun responsabile in turno per "
                f"{order_number}/{phase_upper}")
            continue

        # 3e. Claim atomico PRIMA dell'invio: solo un PC manda questa notifica
        if not claim_notification(conn, order_number, id_phase, template_id,
                                  planned_start, template):
            logger.debug(
                f"FAI Autocheck: {order_number}/{phase_upper} gia' preso in "
                f"carico da un'altra istanza, skip")
            continue

        # 3f. Invia email
        try:
            send_fai_autocheck_email(
                to_list, cc_list, pr, template, logo_path)
            email_time = datetime.now()
            sent_count += 1
        except Exception as e:
            logger.error(
                f"FAI Autocheck: errore invio email per "
                f"{order_number}/{phase_upper}: {e}", exc_info=True)
            # Invio fallito: libera il claim, il ciclo successivo ritenta
            release_notification_claim(conn, order_number, id_phase,
                                       template_id, planned_start)
            continue

        # 3g. Chiude il claim come inviato
        finalize_notification(conn, order_number, id_phase, template_id,
                              planned_start, email_time, to_list, cc_list)

    logger.info(f"FAI Autocheck: ciclo completato, {sent_count} email inviate")
    return sent_count
