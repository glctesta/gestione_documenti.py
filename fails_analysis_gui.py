import tkinter as tk
from tkinter import ttk, messagebox
import threading, datetime, os, logging
import pyodbc
from collections import Counter, defaultdict

logger = logging.getLogger("TraceabilityRS")

# ======================================================================
# QUERY ALL-FAIL
# Una riga per ogni (scheda, fase) che ha avuto almeno un FAIL nel periodo,
# INDIPENDENTEMENTE dal fatto che sia stata poi riparata o meno.
#
# Per ogni (scheda, fase) calcola:
#   - FailEvents       : numero di scansioni FAIL nel periodo (re-fail inclusi)
#   - LastIsPass       : esito dell'ULTIMA scansione a quella fase
#                        (1 = ultima PASS -> "Repaired", 0 = ultima FAIL -> "Wait Repair")
#   - SdPass           : esito riparazione (ScanDefects) dell'ultima scansione FAIL
#                        (0 = SCRAP)
#   - RefailAfterRepair: 1 se la scheda e' stata RIPARATA (ScanDefects.IsPass=1)
#                        e poi e' RI-FALLITA alla stessa fase
#   - MultiFail        : 1 se ha fallito >= 2 volte alla stessa fase nel periodo
# ======================================================================
QUERY_ALLFAILS = """
DECLARE @from datetime=?, @to datetime=?;
WITH ScansInScope AS (
    SELECT s.IDScan, s.IDBoard, s.IsPass, s.ScanTimeFinish, op.IDPhase, op.IDOrder
    FROM Scannings s
    INNER JOIN OrderPhases op ON op.IDOrderPhase = s.IDOrderPhase
    WHERE s.ScanTimeFinish <= @to
),
FailedBP AS (
    SELECT DISTINCT IDBoard, IDPhase
    FROM ScansInScope
    WHERE IsPass = 0 AND ScanTimeFinish BETWEEN @from AND @to
),
PhaseScans AS (
    SELECT sc.IDBoard, sc.IDPhase, sc.IDScan, sc.IsPass, sc.ScanTimeFinish, sc.IDOrder
    FROM ScansInScope sc
    INNER JOIN FailedBP fb ON fb.IDBoard = sc.IDBoard AND fb.IDPhase = sc.IDPhase
),
RepairedFail AS (
    SELECT ps.IDBoard, ps.IDPhase, MIN(ps.IDScan) AS FirstRepairedScan
    FROM PhaseScans ps
    WHERE ps.IsPass = 0
      AND EXISTS (SELECT 1 FROM ScanDefects sd WHERE sd.IDScan = ps.IDScan AND sd.IsPass = 1)
    GROUP BY ps.IDBoard, ps.IDPhase
),
BP AS (
    SELECT IDBoard, IDPhase,
       MAX(IDOrder) AS IDOrder,
       SUM(CASE WHEN IsPass = 0 AND ScanTimeFinish BETWEEN @from AND @to THEN 1 ELSE 0 END) AS FailEvents,
       MAX(CASE WHEN IsPass = 0 THEN IDScan END) AS MaxFailScan,
       MAX(IDScan) AS LastScanId,
       MIN(CASE WHEN IsPass = 0 AND ScanTimeFinish BETWEEN @from AND @to THEN ScanTimeFinish END) AS FirstFailTime
    FROM PhaseScans
    GROUP BY IDBoard, IDPhase
)
SELECT
   bp.IDBoard,
   ph.PhaseName,
   o.OrderNumber,
   p.ProductCode,
   o.OrderQuantity,
   dbo.BoardLabels(bp.IDBoard) AS Labels,
   bp.FailEvents,
   ls.IsPass            AS LastIsPass,
   ls.ScanTimeFinish    AS LastScanTime,
   bp.FirstFailTime,
   sd.SdPass,
   CASE WHEN rf.FirstRepairedScan IS NOT NULL AND bp.MaxFailScan > rf.FirstRepairedScan
        THEN 1 ELSE 0 END AS RefailAfterRepair,
   CASE WHEN bp.FailEvents >= 2 THEN 1 ELSE 0 END AS MultiFail,
   CASE WHEN rf.FirstRepairedScan IS NOT NULL THEN 1 ELSE 0 END AS HasFormalRepair
FROM BP bp
INNER JOIN Phases   ph ON ph.IDPhase  = bp.IDPhase
INNER JOIN Boards   b  ON b.IDBoard   = bp.IDBoard
INNER JOIN Orders   o  ON o.IDOrder   = bp.IDOrder
INNER JOIN Products p  ON p.IDProduct = o.IDProduct
INNER JOIN Scannings ls ON ls.IDScan  = bp.LastScanId
LEFT JOIN RepairedFail rf ON rf.IDBoard = bp.IDBoard AND rf.IDPhase = bp.IDPhase
OUTER APPLY (
    SELECT MAX(CAST(sdd.IsPass AS INT)) AS SdPass
    FROM ScanDefects sdd WHERE sdd.IDScan = bp.MaxFailScan
) sd
ORDER BY ph.PhaseName, o.OrderNumber, bp.IDBoard
"""

# Indici di colonna del result set (per leggibilita')
C_IDBOARD, C_PHASE, C_ORDER, C_PRODUCT, C_QTY, C_LABELS, C_EVENTS, \
    C_LASTPASS, C_LASTTIME, C_FIRSTFAIL, C_SDPASS, C_REFAIL, C_MULTI, C_FORMAL = range(14)

# Fase considerata 'fine produzione' per il volume prodotto (report Tassi)
FINAL_PHASE = 'FINAL ASSEMBLY'

# ======================================================================
# REPORT #2 — TASSI SCRAP / REWORK rispetto al VOLUME PRODOTTO
# Base 'prodotto' = schede distinte (IDBoard) con almeno una scansione nel periodo.
# Scrap  = schede distinte con riparazione SCRAP    (ScanDefects.IsPass = 0)
# Rework = schede distinte con riparazione REPAIRED (ScanDefects.IsPass = 1)
# Aggregazione mensile.
# ======================================================================
QUERY_RATES_MONTHLY = f"""
DECLARE @from datetime=?, @to datetime=?;
WITH Produced AS (
    SELECT FORMAT(s.ScanTimeFinish,'yyyy-MM') AS Ym, COUNT(DISTINCT s.IDBoard) AS Produced
    FROM Scannings s
    INNER JOIN OrderPhases op ON op.IDOrderPhase = s.IDOrderPhase
    INNER JOIN Phases ph ON ph.IDPhase = op.IDPhase
    WHERE s.IsPass = 1 AND ph.PhaseName = N'{FINAL_PHASE}'
      AND s.ScanTimeFinish BETWEEN @from AND @to
    GROUP BY FORMAT(s.ScanTimeFinish,'yyyy-MM')
),
Scrap AS (
    SELECT FORMAT(sd.StopTime,'yyyy-MM') AS Ym, COUNT(DISTINCT sc.IDBoard) AS Scrap
    FROM ScanDefects sd INNER JOIN Scannings sc ON sc.IDScan=sd.IDScan
    WHERE sd.IsPass=0 AND sd.StopTime BETWEEN @from AND @to
    GROUP BY FORMAT(sd.StopTime,'yyyy-MM')
),
Rework AS (
    SELECT FORMAT(sd.StopTime,'yyyy-MM') AS Ym, COUNT(DISTINCT sc.IDBoard) AS Rework
    FROM ScanDefects sd INNER JOIN Scannings sc ON sc.IDScan=sd.IDScan
    WHERE sd.IsPass=1 AND sd.StopTime BETWEEN @from AND @to
    GROUP BY FORMAT(sd.StopTime,'yyyy-MM')
)
SELECT p.Ym, p.Produced, ISNULL(s.Scrap,0) AS Scrap, ISNULL(r.Rework,0) AS Rework
FROM Produced p LEFT JOIN Scrap s ON s.Ym=p.Ym LEFT JOIN Rework r ON r.Ym=p.Ym
ORDER BY p.Ym
"""

# Totale del periodo con conteggi DISTINTI reali (non somma dei mesi)
QUERY_RATES_TOTAL = f"""
DECLARE @from datetime=?, @to datetime=?;
SELECT
  (SELECT COUNT(DISTINCT s.IDBoard) FROM Scannings s
     INNER JOIN OrderPhases op ON op.IDOrderPhase = s.IDOrderPhase
     INNER JOIN Phases ph ON ph.IDPhase = op.IDPhase
     WHERE s.IsPass = 1 AND ph.PhaseName = N'{FINAL_PHASE}'
       AND s.ScanTimeFinish BETWEEN @from AND @to) AS Produced,
  (SELECT COUNT(DISTINCT sc.IDBoard) FROM ScanDefects sd INNER JOIN Scannings sc ON sc.IDScan=sd.IDScan
     WHERE sd.IsPass=0 AND sd.StopTime BETWEEN @from AND @to) AS Scrap,
  (SELECT COUNT(DISTINCT sc.IDBoard) FROM ScanDefects sd INNER JOIN Scannings sc ON sc.IDScan=sd.IDScan
     WHERE sd.IsPass=1 AND sd.StopTime BETWEEN @from AND @to) AS Rework
"""

ST_REPAIRED = 'REPAIRED'    # ultima scansione PASS e con riparazione formale (ScanDefects PASS)
ST_RECOVERED = 'RECOVERED'  # ultima scansione PASS ma senza record di riparazione formale (recuperata al retest)
ST_WAIT = 'WAIT'
ST_SCRAP = 'SCRAP'


def _status(row):
    """Determina lo stato finale della scheda/fase."""
    if row[C_SDPASS] == 0:
        return ST_SCRAP
    if row[C_LASTPASS] == 1:
        return ST_REPAIRED if row[C_FORMAL] == 1 else ST_RECOVERED
    return ST_WAIT


def _dt(v):
    return str(v)[:19] if v else ''


class FailsAnalysisWindow(tk.Toplevel):
    def __init__(self, parent, db, lang, user_name=''):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(lang.get('fa_title', 'Analisi Schede FAIL'))
        self.geometry('1320x780')
        self.minsize(960, 600)
        self._rows = []          # tutte le righe (scheda, fase)
        self._rates = []         # righe report tassi: (Ym, produced, scrap, rework)
        self._rates_total = None # (produced, scrap, rework) sul periodo
        self._loading = False
        self._rates_loading = False
        self.status_filter_var = tk.StringVar(value='ALL')
        self.refail_only_var = tk.BooleanVar(value=False)
        self.target_scrap_var = tk.StringVar(value='0.3')
        self.target_rework_var = tk.StringVar(value='5.0')
        self._build_ui()
        self.grab_set()
        # Auto-carica l'ultima settimana all'apertura
        self.after(250, self._on_load)

    # ------------------------------------------------------------------
    def _build_ui(self):
        hdr = tk.Frame(self, bg='#1F3864', padx=10, pady=8)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=self.lang.get('fa_title', 'Analisi Schede FAIL'),
                 bg='#1F3864', fg='white', font=('Segoe UI', 13, 'bold')).pack(side=tk.LEFT)
        tk.Label(hdr, text=self.lang.get('fa_subtitle', 'Tutti i FAIL di tutte le fasi'),
                 bg='#1F3864', fg='#BBD0F0', font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=12)

        bar = tk.Frame(self, padx=8, pady=6, bg='#F0F0F0')
        bar.pack(fill=tk.X)
        tk.Label(bar, text=self.lang.get('fa_date_from', 'Da:'), bg='#F0F0F0').pack(side=tk.LEFT)
        week_ago = (datetime.date.today() - datetime.timedelta(days=7)).strftime('%Y-%m-%d')
        self.from_var = tk.StringVar(value=week_ago)
        tk.Entry(bar, textvariable=self.from_var, width=12).pack(side=tk.LEFT, padx=4)
        tk.Label(bar, text=self.lang.get('fa_date_to', 'A:'), bg='#F0F0F0').pack(side=tk.LEFT)
        self.to_var = tk.StringVar(value=datetime.date.today().strftime('%Y-%m-%d'))
        tk.Entry(bar, textvariable=self.to_var, width=12).pack(side=tk.LEFT, padx=4)

        btn_style = dict(font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, padx=10, pady=3, cursor='hand2')
        self.load_btn = tk.Button(bar, text=self.lang.get('fa_load_btn', '🔄 Carica / Aggiorna'),
                                  command=self._on_load, bg='#2E75B6', fg='white', **btn_style)
        self.load_btn.pack(side=tk.LEFT, padx=10)
        self.export_btn = tk.Button(bar, text=self.lang.get('fa_export_btn', '📊 Esporta Excel'),
                                    command=self._export_excel, bg='#217346', fg='white',
                                    state=tk.DISABLED, **btn_style)
        self.export_btn.pack(side=tk.LEFT)
        self.status_var = tk.StringVar()
        tk.Label(bar, textvariable=self.status_var, bg='#F0F0F0', fg='#444',
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=14)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)

        tab_all = tk.Frame(self.nb)
        self.nb.add(tab_all, text=self.lang.get('fa_tab_allfail', 'Tutti i FAIL'))
        self._build_all(tab_all)

        tab_refail = tk.Frame(self.nb)
        self.nb.add(tab_refail, text=self.lang.get('fa_tab_refail', 'Riparate → ri-fallite'))
        self._build_refail(tab_refail)

        tab_st = tk.Frame(self.nb)
        self.nb.add(tab_st, text=self.lang.get('fa_tab_stats', 'Statistiche'))
        self._build_stats(tab_st)

        tab_rates = tk.Frame(self.nb)
        self.nb.add(tab_rates, text=self.lang.get('fa_tab_rates', 'Tassi Scrap/Rework'))
        self._build_rates(tab_rates)

    def _make_tree(self, parent, cols, widths, labels=None):
        f = tk.Frame(parent)
        f.pack(fill=tk.BOTH, expand=True)
        vb = ttk.Scrollbar(f, orient=tk.VERTICAL)
        hb = ttk.Scrollbar(f, orient=tk.HORIZONTAL)
        tree = ttk.Treeview(f, columns=cols, show='headings',
                            yscrollcommand=vb.set, xscrollcommand=hb.set)
        vb.config(command=tree.yview)
        hb.config(command=tree.xview)
        for i, (c, w) in enumerate(zip(cols, widths)):
            lbl = labels[i] if labels else c
            tree.heading(c, text=lbl)
            tree.column(c, width=w, minwidth=40, anchor=tk.CENTER if w <= 80 else tk.W)
        vb.pack(side=tk.RIGHT, fill=tk.Y)
        hb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(fill=tk.BOTH, expand=True)
        return tree

    def _build_all(self, parent):
        fbar = tk.Frame(parent, bg='#F8F8F8', padx=6, pady=4)
        fbar.pack(fill=tk.X)
        tk.Label(fbar, text=self.lang.get('fa_filter_status', 'Stato:'),
                 bg='#F8F8F8', font=('Segoe UI', 9)).pack(side=tk.LEFT)
        combo = ttk.Combobox(fbar, textvariable=self.status_filter_var, width=14, state='readonly',
                             values=['ALL', ST_WAIT, ST_REPAIRED, ST_RECOVERED, ST_SCRAP])
        combo.pack(side=tk.LEFT, padx=6)
        combo.bind('<<ComboboxSelected>>', lambda e: self._populate_all())
        tk.Checkbutton(fbar, text=self.lang.get('fa_filter_refail', 'Solo riparate → ri-fallite'),
                       variable=self.refail_only_var, command=self._populate_all,
                       bg='#F8F8F8', font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=12)
        self.all_count_lbl = tk.Label(fbar, text='', bg='#F8F8F8', font=('Segoe UI', 9), fg='#555')
        self.all_count_lbl.pack(side=tk.LEFT, padx=12)

        col_keys = [
            ('fa_col_order',     'N. Ordine',  110),
            ('fa_col_product',   'Prodotto',   120),
            ('fa_col_qty',       'Qty',         55),
            ('fa_col_phase',     'Fase',        100),
            ('fa_col_idboard',   'IDBoard',     70),
            ('fa_col_labels',    'Label',       170),
            ('fa_col_failcount', 'N. Fail',     60),
            ('fa_col_status',    'Stato',       90),
            ('fa_col_refail',    'Ri-fallita',  75),
            ('fa_col_firstfail', 'Primo Fail',  135),
            ('fa_col_lastscan',  'Ultima Scan', 135),
        ]
        cols = tuple(k for k, _, _ in col_keys)
        widths = tuple(w for _, _, w in col_keys)
        labels = tuple(self.lang.get(k, d) for k, d, _ in col_keys)
        self.tree_all = self._make_tree(parent, cols, widths, labels)
        self.tree_all.tag_configure('repaired',  background='#D4EDDA')
        self.tree_all.tag_configure('recovered', background='#CFE7F5')
        self.tree_all.tag_configure('wait',      background='#FADADD')
        self.tree_all.tag_configure('scrap',     background='#FFE5B4')
        self.tree_all.tag_configure('refail',    foreground='#A00000', font=('Segoe UI', 9, 'bold'))

    def _build_refail(self, parent):
        tk.Label(parent, bg='#FFF3CD', fg='#7A5C00', anchor='w', padx=8, pady=4,
                 font=('Segoe UI', 9),
                 text=self.lang.get('fa_refail_hint',
                                    'Schede RIPARATE (PASS in riparazione) e poi RI-FALLITE alla stessa fase — '
                                    'indicatore di inefficienza della riparazione.')).pack(fill=tk.X)
        col_keys = [
            ('fa_col_order',     'N. Ordine',  110),
            ('fa_col_product',   'Prodotto',   120),
            ('fa_col_phase',     'Fase',        100),
            ('fa_col_idboard',   'IDBoard',     70),
            ('fa_col_labels',    'Label',       180),
            ('fa_col_failcount', 'N. Fail',     60),
            ('fa_col_firstfail', 'Primo Fail',  135),
            ('fa_col_lastscan',  'Ultima Scan', 135),
            ('fa_col_status',    'Stato',       90),
        ]
        cols = tuple(k for k, _, _ in col_keys)
        widths = tuple(w for _, _, w in col_keys)
        labels = tuple(self.lang.get(k, d) for k, d, _ in col_keys)
        self.tree_refail = self._make_tree(parent, cols, widths, labels)
        self.tree_refail.tag_configure('wait',     background='#FADADD')
        self.tree_refail.tag_configure('repaired', background='#FFF3CD')

    def _build_stats(self, parent):
        self.stats_text = tk.Text(parent, wrap=tk.NONE, state=tk.DISABLED,
                                  font=('Consolas', 9), bg='#F8F9FA', padx=8, pady=6)
        sb = ttk.Scrollbar(parent, command=self.stats_text.yview)
        self.stats_text.config(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.stats_text.pack(fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------
    def _parse_dates(self):
        dt_from = datetime.datetime.strptime(self.from_var.get().strip(), '%Y-%m-%d')
        dt_to = datetime.datetime.strptime(self.to_var.get().strip(), '%Y-%m-%d')
        return dt_from, dt_to.replace(hour=23, minute=59, second=59)

    def _on_load(self):
        if self._loading:
            return
        try:
            dt_from, dt_to = self._parse_dates()
        except ValueError:
            messagebox.showerror('Errore', 'Formato data non valido (YYYY-MM-DD)', parent=self)
            return
        self._loading = True
        self.load_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.status_var.set(self.lang.get('fa_status_loading', 'Caricamento in corso...'))
        threading.Thread(target=self._bg_load, args=(dt_from, dt_to), daemon=True).start()

    def _bg_load(self, dt_from, dt_to):
        """Esegue QUERY_ALLFAILS su una connessione dedicata (thread background)."""
        try:
            conn = pyodbc.connect(self.db.conn_str, autocommit=True)
            cur = conn.cursor()
            cur.execute(QUERY_ALLFAILS, (dt_from, dt_to))
            rows = cur.fetchall()
            conn.close()
            self.after(0, lambda: self._on_loaded(rows))
        except Exception as e:
            logger.error(f"AnalisiFailsRS _bg_load: {e}", exc_info=True)
            err = str(e)
            self.after(0, lambda msg=err: (
                self.status_var.set(f'Errore: {msg}'),
                self.load_btn.config(state=tk.NORMAL),
                setattr(self, '_loading', False)
            ))

    def _on_loaded(self, rows):
        self._rows = rows
        self._loading = False
        self.load_btn.config(state=tk.NORMAL)
        self._populate_all()
        self._populate_refail()
        self._populate_stats()
        # Riepilogo nella status bar
        boards = len(self._rows)
        events = sum((r[C_EVENTS] or 0) for r in self._rows)
        rep = sum(1 for r in self._rows if _status(r) == ST_REPAIRED)
        wait = sum(1 for r in self._rows if _status(r) == ST_WAIT)
        scrap = sum(1 for r in self._rows if _status(r) == ST_SCRAP)
        recovered = sum(1 for r in self._rows if _status(r) == ST_RECOVERED)
        refail = sum(1 for r in self._rows if r[C_REFAIL] == 1)
        self.status_var.set(self.lang.get(
            'fa_status_ready2',
            '{0} schede/fase FAIL — {1} eventi — Wait {2} · Repaired {3} · Recovered {4} · Scrap {5} · Ri-fallite {6}'
        ).format(boards, events, wait, rep, recovered, scrap, refail))
        self.export_btn.config(state=tk.NORMAL if self._rows else tk.DISABLED)

    # ------------------------------------------------------------------
    def _populate_all(self):
        if not hasattr(self, 'tree_all'):
            return
        for i in self.tree_all.get_children():
            self.tree_all.delete(i)
        sf = self.status_filter_var.get()
        refail_only = self.refail_only_var.get()
        shown = 0
        for r in self._rows:
            st = _status(r)
            if sf != 'ALL' and st != sf:
                continue
            if refail_only and r[C_REFAIL] != 1:
                continue
            shown += 1
            tag = {ST_REPAIRED: 'repaired', ST_RECOVERED: 'recovered',
                   ST_SCRAP: 'scrap', ST_WAIT: 'wait'}.get(st, 'wait')
            tags = (tag, 'refail') if r[C_REFAIL] == 1 else (tag,)
            self.tree_all.insert('', tk.END, tags=tags, values=(
                r[C_ORDER], r[C_PRODUCT], r[C_QTY], r[C_PHASE], r[C_IDBOARD], r[C_LABELS],
                r[C_EVENTS], st, 'SI' if r[C_REFAIL] == 1 else '',
                _dt(r[C_FIRSTFAIL]), _dt(r[C_LASTTIME]),
            ))
        total = len(self._rows)
        self.all_count_lbl.config(
            text=self.lang.get('fa_all_count', '{0} di {1} righe').format(shown, total))

    def _populate_refail(self):
        for i in self.tree_refail.get_children():
            self.tree_refail.delete(i)
        for r in self._rows:
            if r[C_REFAIL] != 1:
                continue
            st = _status(r)
            tag = 'wait' if st == ST_WAIT else 'repaired'
            self.tree_refail.insert('', tk.END, tags=(tag,), values=(
                r[C_ORDER], r[C_PRODUCT], r[C_PHASE], r[C_IDBOARD], r[C_LABELS],
                r[C_EVENTS], _dt(r[C_FIRSTFAIL]), _dt(r[C_LASTTIME]), st,
            ))

    def _phase_aggregates(self):
        """Aggrega per fase. Ritorna (dict per fase, totale)."""
        agg = defaultdict(lambda: dict(boards=0, events=0, rep=0, recov=0, wait=0, scrap=0, refail=0, multi=0))
        for r in self._rows:
            a = agg[r[C_PHASE] or '?']
            a['boards'] += 1
            a['events'] += r[C_EVENTS] or 0
            st = _status(r)
            a['rep'] += (st == ST_REPAIRED)
            a['recov'] += (st == ST_RECOVERED)
            a['wait'] += (st == ST_WAIT)
            a['scrap'] += (st == ST_SCRAP)
            a['refail'] += (r[C_REFAIL] == 1)
            a['multi'] += (r[C_MULTI] == 1)
        return agg

    def _populate_stats(self):
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete('1.0', tk.END)
        T = self.lang.get
        agg = self._phase_aggregates()

        tot = dict(boards=0, events=0, rep=0, recov=0, wait=0, scrap=0, refail=0, multi=0)
        for a in agg.values():
            for k in tot:
                tot[k] += a[k]

        fmt = '  {0:<14}{1:>8}{2:>8}{3:>9}{4:>10}{5:>7}{6:>7}{7:>11}{8:>9}'
        lines = [
            '=' * 100,
            '  ' + T('fa_stats_title2', 'FAIL — RIEPILOGO PER FASE (tutti i FAIL, riparati e non)'),
            '=' * 100,
            fmt.format(
                T('fa_h_phase', 'Fase'), T('fa_h_boards', 'Schede'), T('fa_h_events', 'Eventi'),
                T('fa_h_repaired', 'Repaired'), T('fa_h_recovered', 'Recovered'),
                T('fa_h_wait', 'Wait'), T('fa_h_scrap', 'Scrap'),
                T('fa_h_refail', 'Ri-fallite'), T('fa_h_multi', 'Multi')),
            '  ' + '-' * 83,
        ]
        for ph in sorted(agg, key=lambda k: -agg[k]['boards']):
            a = agg[ph]
            lines.append(fmt.format(
                str(ph)[:14], a['boards'], a['events'], a['rep'], a['recov'],
                a['wait'], a['scrap'], a['refail'], a['multi']))
        lines.append('  ' + '-' * 83)
        lines.append(fmt.format(
            T('fa_h_total', 'TOTALE'), tot['boards'], tot['events'], tot['rep'], tot['recov'],
            tot['wait'], tot['scrap'], tot['refail'], tot['multi']))

        # Percentuali interne
        b = tot['boards'] or 1
        rep_b = tot['rep'] or 1
        lines += [
            '',
            '  ' + T('fa_stats_rates', 'Indicatori (interni al set FAIL):'),
            '  ' + '-' * 50,
            '  ' + T('fa_pct_wait',   '% ancora aperte (Wait)   :') + f" {tot['wait']/b*100:5.1f}%",
            '  ' + T('fa_pct_rep',    '% riparate (Repaired)    :') + f" {tot['rep']/b*100:5.1f}%",
            '  ' + T('fa_pct_recov',  '% recuperate (Recovered) :') + f" {tot['recov']/b*100:5.1f}%",
            '  ' + T('fa_pct_scrap',  '% scrap                  :') + f" {tot['scrap']/b*100:5.1f}%",
            '  ' + T('fa_pct_refail', '% ri-fallite su riparate :') + f" {tot['refail']/rep_b*100:5.1f}%"
            + '   ' + T('fa_pct_refail_note', '(efficienza riparazione)'),
        ]

        # Distribuzione per prodotto
        lines += ['', '  ' + T('fa_stats_top_products', 'Top 15 Prodotti con piu FAIL (eventi):'), '  ' + '-' * 50]
        prod = Counter()
        for r in self._rows:
            if r[C_PRODUCT]:
                prod[r[C_PRODUCT]] += (r[C_EVENTS] or 0)
        for pr, cnt in prod.most_common(15):
            lines.append(f'    {str(pr)[:40]:<40} {cnt:>6}')

        self.stats_text.insert(tk.END, '\n'.join(lines))
        self.stats_text.config(state=tk.DISABLED)

    # ============ REPORT #2 — TASSI SCRAP / REWORK ====================
    def _build_rates(self, parent):
        bar = tk.Frame(parent, bg='#F8F8F8', padx=6, pady=5)
        bar.pack(fill=tk.X)
        tk.Label(bar, text=self.lang.get('fa_target_scrap', 'Target Scrap %:'),
                 bg='#F8F8F8', font=('Segoe UI', 9)).pack(side=tk.LEFT)
        tk.Entry(bar, textvariable=self.target_scrap_var, width=6).pack(side=tk.LEFT, padx=4)
        tk.Label(bar, text=self.lang.get('fa_target_rework', 'Target Rework %:'),
                 bg='#F8F8F8', font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(10, 0))
        tk.Entry(bar, textvariable=self.target_rework_var, width=6).pack(side=tk.LEFT, padx=4)
        self.calc_btn = tk.Button(
            bar, text=self.lang.get('fa_calc_btn', '📈 Calcola tassi'),
            command=self._on_calc_rates, bg='#2E75B6', fg='white',
            font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, padx=10, pady=3, cursor='hand2')
        self.calc_btn.pack(side=tk.LEFT, padx=10)
        self.rates_status = tk.Label(bar, text='', bg='#F8F8F8', fg='#444', font=('Segoe UI', 9))
        self.rates_status.pack(side=tk.LEFT, padx=10)

        tk.Label(parent, bg='#EEF4FB', fg='#33506E', anchor='w', padx=8, pady=4,
                 font=('Segoe UI', 8),
                 text=self.lang.get('fa_rates_note',
                                    "Prodotto = schede distinte scansionate nel periodo. "
                                    "Scrap/Rework = schede distinte con esito riparazione SCRAP/REPAIRED. "
                                    "Verde = entro target, rosso = oltre target.")).pack(fill=tk.X)

        col_keys = [
            ('fa_rc_month',     'Mese',      90),
            ('fa_rc_produced',  'Prodotto', 100),
            ('fa_rc_scrap',     'Scrap',     80),
            ('fa_rc_rework',    'Rework',    80),
            ('fa_rc_scrappct',  'Scrap %',  100),
            ('fa_rc_reworkpct', 'Rework %', 100),
            ('fa_rc_failpct',   'Fail %',   100),
        ]
        cols = tuple(k for k, _, _ in col_keys)
        widths = tuple(w for _, _, w in col_keys)
        labels = tuple(self.lang.get(k, d) for k, d, _ in col_keys)
        self.tree_rates = self._make_tree(parent, cols, widths, labels)
        self.tree_rates.tag_configure('ok',    background='#D4EDDA')
        self.tree_rates.tag_configure('over',  background='#FADADD')
        self.tree_rates.tag_configure('total', background='#E2E8F0', font=('Segoe UI', 9, 'bold'))

    def _targets(self):
        try:
            ts = float(self.target_scrap_var.get().replace(',', '.'))
        except ValueError:
            ts = 0.3
        try:
            tr = float(self.target_rework_var.get().replace(',', '.'))
        except ValueError:
            tr = 5.0
        return ts, tr

    def _on_calc_rates(self):
        if self._rates_loading:
            return
        try:
            dt_from, dt_to = self._parse_dates()
        except ValueError:
            messagebox.showerror('Errore', 'Formato data non valido (YYYY-MM-DD)', parent=self)
            return
        self._rates_loading = True
        self.calc_btn.config(state=tk.DISABLED)
        self.rates_status.config(text=self.lang.get('fa_rates_loading', 'Calcolo tassi in corso...'))
        threading.Thread(target=self._bg_rates, args=(dt_from, dt_to), daemon=True).start()

    def _bg_rates(self, dt_from, dt_to):
        try:
            conn = pyodbc.connect(self.db.conn_str, autocommit=True)
            cur = conn.cursor()
            cur.execute(QUERY_RATES_MONTHLY, (dt_from, dt_to))
            monthly = cur.fetchall()
            cur.execute(QUERY_RATES_TOTAL, (dt_from, dt_to))
            total = cur.fetchone()
            conn.close()
            self.after(0, lambda: self._on_rates_loaded(monthly, total))
        except Exception as e:
            logger.error(f"AnalisiFailsRS _bg_rates: {e}", exc_info=True)
            err = str(e)
            self.after(0, lambda msg=err: (
                self.rates_status.config(text=f'Errore: {msg}'),
                self.calc_btn.config(state=tk.NORMAL),
                setattr(self, '_rates_loading', False)
            ))

    def _on_rates_loaded(self, monthly, total):
        self._rates = monthly
        self._rates_total = total
        self._rates_loading = False
        self.calc_btn.config(state=tk.NORMAL)
        self._populate_rates()
        self.rates_status.config(text=self.lang.get('fa_rates_done', 'Calcolo completato'))
        if self._rows or self._rates:
            self.export_btn.config(state=tk.NORMAL)

    def _populate_rates(self):
        for i in self.tree_rates.get_children():
            self.tree_rates.delete(i)
        ts, tr = self._targets()

        def add_row(month, prod, scrap, rework, is_total=False):
            sr = (scrap / prod * 100) if prod else 0
            rr = (rework / prod * 100) if prod else 0
            if is_total:
                tag = 'total'
            else:
                tag = 'over' if (sr > ts or rr > tr) else 'ok'
            self.tree_rates.insert('', tk.END, tags=(tag,), values=(
                month, prod, scrap, rework,
                f'{sr:.3f}%', f'{rr:.3f}%', f'{(sr + rr):.3f}%'))

        for r in self._rates:
            add_row(r[0], r[1], r[2], r[3])
        if self._rates_total:
            p, s, w = self._rates_total
            add_row(self.lang.get('fa_rates_total', 'TOTALE periodo'), p, s, w, is_total=True)

    # ------------------------------------------------------------------
    def _export_excel(self):
        if not self._rows and not self._rates:
            messagebox.showwarning('', self.lang.get('fa_no_data', 'Nessun dato'), parent=self)
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            HDR_FILL = PatternFill('solid', fgColor='1F3864')
            HDR_FONT = Font(bold=True, color='FFFFFF')
            THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            FILLS = {
                ST_REPAIRED:  PatternFill('solid', fgColor='D4EDDA'),
                ST_RECOVERED: PatternFill('solid', fgColor='CFE7F5'),
                ST_WAIT:      PatternFill('solid', fgColor='FADADD'),
                ST_SCRAP:     PatternFill('solid', fgColor='FFE5B4'),
            }
            REFAIL_FONT = Font(bold=True, color='A00000')

            def hdr(ws, headers):
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(1, c, h)
                    cell.fill = HDR_FILL
                    cell.font = HDR_FONT
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = THIN

            def autofit(ws):
                for col in ws.columns:
                    mx = max((len(str(ce.value or '')) for ce in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 50)

            wb = openpyxl.Workbook()

            # Sheet 1 — Tutti i FAIL
            ws1 = wb.active
            ws1.title = self.lang.get('fa_tab_allfail', 'Tutti i FAIL')[:31]
            hdr(ws1, ['OrderNumber', 'ProductCode', 'Qty', 'PhaseName', 'IDBoard', 'Labels',
                       'FailEvents', 'Status', 'RefailAfterRepair', 'FirstFail', 'LastScan'])
            ri = 2
            for r in self._rows:
                st = _status(r)
                fl = FILLS.get(st, PatternFill())
                vals = [r[C_ORDER], r[C_PRODUCT], r[C_QTY], r[C_PHASE], r[C_IDBOARD], r[C_LABELS],
                        r[C_EVENTS], st, 'YES' if r[C_REFAIL] == 1 else '',
                        _dt(r[C_FIRSTFAIL]), _dt(r[C_LASTTIME])]
                for c, v in enumerate(vals, 1):
                    ce = ws1.cell(ri, c, v)
                    ce.fill = fl
                    ce.border = THIN
                    if r[C_REFAIL] == 1 and c == 9:
                        ce.font = REFAIL_FONT
                ri += 1
            ws1.freeze_panes = 'A2'
            autofit(ws1)

            # Sheet 2 — Riparate poi ri-fallite
            ws2 = wb.create_sheet(self.lang.get('fa_tab_refail', 'Ri-fallite')[:31])
            hdr(ws2, ['OrderNumber', 'ProductCode', 'PhaseName', 'IDBoard', 'Labels',
                       'FailEvents', 'FirstFail', 'LastScan', 'Status'])
            ri = 2
            for r in self._rows:
                if r[C_REFAIL] != 1:
                    continue
                for c, v in enumerate([r[C_ORDER], r[C_PRODUCT], r[C_PHASE], r[C_IDBOARD], r[C_LABELS],
                                        r[C_EVENTS], _dt(r[C_FIRSTFAIL]), _dt(r[C_LASTTIME]), _status(r)], 1):
                    ce = ws2.cell(ri, c, v)
                    ce.border = THIN
                ri += 1
            ws2.freeze_panes = 'A2'
            autofit(ws2)

            # Sheet 3 — Statistiche per fase
            ws3 = wb.create_sheet(self.lang.get('fa_tab_stats', 'Statistiche')[:31])
            hdr(ws3, ['Fase', 'Schede', 'Eventi', 'Repaired', 'Recovered', 'Wait', 'Scrap',
                       'Ri-fallite', 'Multi-fail'])
            agg = self._phase_aggregates()
            ri = 2
            tot = dict(boards=0, events=0, rep=0, recov=0, wait=0, scrap=0, refail=0, multi=0)
            for ph in sorted(agg, key=lambda k: -agg[k]['boards']):
                a = agg[ph]
                for k in tot:
                    tot[k] += a[k]
                for c, v in enumerate([ph, a['boards'], a['events'], a['rep'], a['recov'], a['wait'],
                                        a['scrap'], a['refail'], a['multi']], 1):
                    ce = ws3.cell(ri, c, v)
                    ce.border = THIN
                ri += 1
            for c, v in enumerate(['TOTALE', tot['boards'], tot['events'], tot['rep'], tot['recov'],
                                    tot['wait'], tot['scrap'], tot['refail'], tot['multi']], 1):
                ce = ws3.cell(ri, c, v)
                ce.font = Font(bold=True)
                ce.border = THIN
            autofit(ws3)

            # Sheet 4 — Tassi Scrap/Rework (se calcolati)
            if self._rates:
                ws4 = wb.create_sheet(self.lang.get('fa_tab_rates', 'Tassi')[:31])
                hdr(ws4, ['Mese', 'Prodotto', 'Scrap', 'Rework', 'Scrap %', 'Rework %', 'Fail %'])
                tgt_s, tgt_r = self._targets()
                OK_FILL = PatternFill('solid', fgColor='D4EDDA')
                OVER_FILL = PatternFill('solid', fgColor='FADADD')
                ri = 2

                def rate_row(month, prod, scrap, rework, bold=False):
                    nonlocal ri
                    sr = (scrap / prod * 100) if prod else 0
                    rr = (rework / prod * 100) if prod else 0
                    fl = OVER_FILL if (sr > tgt_s or rr > tgt_r) else OK_FILL
                    vals = [month, prod, scrap, rework, round(sr, 3), round(rr, 3), round(sr + rr, 3)]
                    for c, v in enumerate(vals, 1):
                        ce = ws4.cell(ri, c, v)
                        ce.border = THIN
                        ce.fill = fl
                        if bold:
                            ce.font = Font(bold=True)
                    ri += 1

                for r in self._rates:
                    rate_row(r[0], r[1], r[2], r[3])
                if self._rates_total:
                    p, s, w = self._rates_total
                    rate_row(self.lang.get('fa_rates_total', 'TOTALE periodo'), p, s, w, bold=True)
                ws4.append([])
                ws4.append([self.lang.get('fa_target_scrap', 'Target Scrap %:'), tgt_s])
                ws4.append([self.lang.get('fa_target_rework', 'Target Rework %:'), tgt_r])
                ws4.freeze_panes = 'A2'
                autofit(ws4)

            os.makedirs(r'C:\Temp', exist_ok=True)
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            fp = rf'C:\Temp\AnalisiFailsRS_AllFail_{ts}.xlsx'
            wb.save(fp)
            logger.info(f"AnalisiFailsRS: Excel salvato {fp}")
            os.startfile(fp)
        except Exception as e:
            logger.error(f"AnalisiFailsRS export Excel: {e}", exc_info=True)
            messagebox.showerror('Errore', f'Impossibile generare Excel:\n{e}', parent=self)


def open_fails_analysis(parent, db, lang, user_name=''):
    FailsAnalysisWindow(parent, db, lang, user_name)
