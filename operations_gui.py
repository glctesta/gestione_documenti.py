# operations_gui.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
import shipping_processor  # Assicurati che il file shipping_processor.py esista

class AddInterruptionWindow(tk.Toplevel):
    """Finestra per l'inserimento di un fermo di produzione."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(lang.get('submenu_interruptions', "Registra Interruzione di Produzione"))
        self.geometry("800x650")
        self.transient(parent)
        self.grab_set()

        self.working_areas_data, self.sub_areas_data, self.lines_data = {}, {}, {}
        self.orders_data, self.issue_areas_data, self.problems_data = {}, {}, {}
        self.all_order_names = []

        self.working_area_var, self.sub_area_var, self.line_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
        self.order_var = tk.StringVar()
        self.time_choice_var = tk.StringVar(value="duration")
        self.from_hour_var, self.to_hour_var, self.duration_var = tk.StringVar(), tk.StringVar(), tk.StringVar()
        self.issue_area_var, self.problem_var = tk.StringVar(), tk.StringVar()

        self.action_plan_placeholder = self.lang.get('action_plan_placeholder', 'Inserire un piano d''azione...')
        self.comments_placeholder = self.lang.get('comments_placeholder',
                                                  'Digita una nota per spiegare meglio il problema...')
        
        # Lista per memorizzare i documenti caricati
        self.uploaded_documents = []  # Format: [{'filename': str, 'description': str, 'binary_data': bytes}, ...]

        self._create_widgets()
        self._load_initial_data()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill="both", expand=True)
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(3, weight=1)

        loc_frame = ttk.LabelFrame(main_frame, text=self.lang.get('loc_frame_title', "Localizzazione"))
        loc_frame.grid(row=0, column=0, columnspan=4, sticky="ew", pady=5)
        loc_frame.columnconfigure(1, weight=1)
        loc_frame.columnconfigure(3, weight=1)
        ttk.Label(loc_frame, text=self.lang.get('working_area_label', "Working Area (*):")).grid(row=0, column=0,
                                                                                                 sticky="w", padx=5,
                                                                                                 pady=3)
        self.wa_combo = ttk.Combobox(loc_frame, textvariable=self.working_area_var, state="readonly")
        self.wa_combo.grid(row=0, column=1, sticky="ew", padx=5)
        self.wa_combo.bind("<<ComboboxSelected>>", self._on_wa_select)
        ttk.Label(loc_frame, text=self.lang.get('sub_area_label', "Sub Area (*):")).grid(row=0, column=2, sticky="w",
                                                                                         padx=5, pady=3)
        self.sa_combo = ttk.Combobox(loc_frame, textvariable=self.sub_area_var, state="disabled")
        self.sa_combo.grid(row=0, column=3, sticky="ew", padx=5)
        self.sa_combo.bind("<<ComboboxSelected>>", self._on_sa_select)
        ttk.Label(loc_frame, text=self.lang.get('line_label', "Linea (*):")).grid(row=1, column=0, sticky="w", padx=5,
                                                                                  pady=3)
        self.line_combo = ttk.Combobox(loc_frame, textvariable=self.line_var, state="disabled")
        self.line_combo.grid(row=1, column=1, sticky="ew", padx=5)
        ttk.Label(loc_frame, text=self.lang.get('order_label', "Ordine di Produzione:")).grid(row=1, column=2,
                                                                                              sticky="w", padx=5,
                                                                                              pady=3)
        self.order_combo = ttk.Combobox(loc_frame, textvariable=self.order_var, state="normal")
        self.order_combo.grid(row=1, column=3, sticky="ew", padx=5)
        self.order_combo.bind("<KeyRelease>", self._filter_order_combo)

        time_frame = ttk.LabelFrame(main_frame, text=self.lang.get('time_frame_title', "Tempistiche"))
        time_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=5)
        time_frame.columnconfigure(1, weight=1)
        ttk.Label(time_frame, text=self.lang.get('date_label_req', "Data (*):")).grid(row=0, column=0, sticky="w",
                                                                                      padx=5, pady=3)
        self.date_entry = DateEntry(time_frame, date_pattern='dd/MM/yyyy')
        self.date_entry.grid(row=0, column=1, sticky="w")
        ttk.Radiobutton(time_frame, text=self.lang.get('time_radio_duration', "Durata (minuti)"),
                        variable=self.time_choice_var, value="duration", command=self._toggle_time_fields).grid(row=1,
                                                                                                                column=0,
                                                                                                                sticky="w",
                                                                                                                padx=5)
        self.duration_entry = ttk.Entry(time_frame, textvariable=self.duration_var)
        self.duration_entry.grid(row=1, column=1, sticky="w")
        ttk.Radiobutton(time_frame, text=self.lang.get('time_radio_range', "Intervallo Orario"),
                        variable=self.time_choice_var, value="range", command=self._toggle_time_fields).grid(row=2,
                                                                                                             column=0,
                                                                                                             sticky="w",
                                                                                                             padx=5)
        range_frame = ttk.Frame(time_frame)
        range_frame.grid(row=2, column=1, sticky="ew")
        ttk.Label(range_frame, text=self.lang.get('from_label_short', "Da:")).pack(side="left")
        self.from_entry = ttk.Entry(range_frame, textvariable=self.from_hour_var, width=8)
        self.from_entry.pack(side="left", padx=5)
        ttk.Label(range_frame, text=self.lang.get('to_label_short', "A:")).pack(side="left")
        self.to_entry = ttk.Entry(range_frame, textvariable=self.to_hour_var, width=8)
        self.to_entry.pack(side="left", padx=5)
        self.to_entry.bind("<FocusOut>", self._calculate_duration)

        cause_frame = ttk.LabelFrame(main_frame, text=self.lang.get('cause_frame_title', "Causa dell'Interruzione"))
        cause_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=5)
        cause_frame.columnconfigure(1, weight=1)
        cause_frame.columnconfigure(3, weight=1)
        ttk.Label(cause_frame, text=self.lang.get('issue_area_label', "Issue Area (*):")).grid(row=0, column=0,
                                                                                               sticky="w", padx=5,
                                                                                               pady=3)
        self.ia_combo = ttk.Combobox(cause_frame, textvariable=self.issue_area_var, state="readonly")
        self.ia_combo.grid(row=0, column=1, sticky="ew", padx=5)
        self.ia_combo.bind("<<ComboboxSelected>>", lambda e: self._update_problems_combo())
        ttk.Label(cause_frame, text=self.lang.get('event_label', "Evento/Problema (*):")).grid(row=0, column=2,
                                                                                               sticky="w", padx=5,
                                                                                               pady=3)
        self.problem_combo = ttk.Combobox(cause_frame, textvariable=self.problem_var, state="disabled")
        self.problem_combo.grid(row=0, column=3, sticky="ew", padx=5)

        notes_frame = ttk.LabelFrame(main_frame, text=self.lang.get('notes_frame_title', "Note e Piano d'Azione"))
        notes_frame.grid(row=3, column=0, columnspan=4, sticky="nsew", pady=5)
        main_frame.rowconfigure(3, weight=1)
        notes_frame.columnconfigure(0, weight=1)
        notes_frame.rowconfigure(1, weight=1)
        notes_frame.rowconfigure(3, weight=1)
        
        # Sezione Documenti Allegati
        docs_header_frame = ttk.Frame(notes_frame)
        docs_header_frame.pack(fill="x", padx=5, pady=(5, 2))
        ttk.Label(docs_header_frame, text=self.lang.get('documents_section_label', "Documenti Allegati:"), 
                  font=("Helvetica", 9, "bold")).pack(side="left")
        ttk.Button(docs_header_frame, text=self.lang.get('add_document_button', "Aggiungi Documento"),
                   command=self._add_document).pack(side="left", padx=10)
        ttk.Button(docs_header_frame, text=self.lang.get('remove_document_button', "Rimuovi Documento"),
                   command=self._remove_document).pack(side="left")
        
        # Lista documenti
        docs_list_frame = ttk.Frame(notes_frame)
        docs_list_frame.pack(fill="x", padx=5, pady=(0, 5))
        
        self.docs_listbox = tk.Listbox(docs_list_frame, height=3)
        self.docs_listbox.pack(side="left", fill="both", expand=True)
        
        docs_scrollbar = ttk.Scrollbar(docs_list_frame, orient="vertical", command=self.docs_listbox.yview)
        docs_scrollbar.pack(side="right", fill="y")
        self.docs_listbox.config(yscrollcommand=docs_scrollbar.set)
        
        ttk.Label(notes_frame, text=self.lang.get('comments_label', "Commenti (*):")).pack(anchor="w")
        self.notes_text = tk.Text(notes_frame, height=4)
        self.notes_text.insert("1.0", self.comments_placeholder)
        self.notes_text.config(foreground="grey")
        self.notes_text.bind("<FocusIn>", self._on_comments_focus_in)
        self.notes_text.bind("<FocusOut>", self._on_comments_focus_out)
        self.notes_text.pack(fill="x", expand=True, padx=5, pady=2)
        ttk.Label(notes_frame, text=self.lang.get('action_plan_label_req', "Piano d'Azione (*):")).pack(anchor="w")
        self.action_text = tk.Text(notes_frame, height=3)
        self.action_text.insert("1.0", self.action_plan_placeholder)
        self.action_text.config(foreground="grey")
        self.action_text.bind("<FocusIn>", self._on_action_plan_focus_in)
        self.action_text.bind("<FocusOut>", self._on_action_plan_focus_out)
        self.action_text.pack(fill="x", expand=True, padx=5, pady=2)

        ttk.Button(main_frame, text=self.lang.get('save_declaration_button', "Salva Dichiarazione"),
                   command=self._save_interruption).grid(row=4, column=3, sticky="e", pady=15)
        self._toggle_time_fields()

    def _load_initial_data(self):
        areas = self.db.fetch_working_areas()
        if areas:
            self.working_areas_data = {a.AreaName: a.WorkingAreaID for a in areas}
            self.wa_combo['values'] = sorted(list(self.working_areas_data.keys()))
        orders = self.db.fetch_production_orders_for_breakdown()
        if orders:
            self.orders_data = {o.OrderNumber: o.IdOrdine for o in orders}
            self.all_order_names = sorted(list(self.orders_data.keys()))
            self.order_combo['values'] = self.all_order_names
        issue_areas = self.db.fetch_issue_areas()
        if issue_areas:
            self.issue_areas_data = {ia.IssueArea: ia.IssueAreaId for ia in issue_areas}
            self.ia_combo['values'] = sorted(list(self.issue_areas_data.keys()))

    def _on_wa_select(self, event=None):
        self.sub_area_var.set('')
        self.sa_combo.config(state="disabled", values=[])
        self._on_sa_select()
        wa_id = self.working_areas_data.get(self.working_area_var.get())
        if not wa_id: return
        sub_areas = self.db.fetch_working_sub_areas(wa_id)
        if sub_areas:
            self.sub_areas_data = {sa.AreaSubName: sa.WorkingSubAreaID for sa in sub_areas}
            self.sa_combo['values'] = sorted(list(self.sub_areas_data.keys()))
            self.sa_combo.config(state="readonly")

    def _on_sa_select(self, event=None):
        self.line_var.set('')
        self.line_combo.config(state="disabled", values=[])
        self._update_problems_combo()
        wa_id = self.working_areas_data.get(self.working_area_var.get())
        sa_id = self.sub_areas_data.get(self.sub_area_var.get())
        if not wa_id or not sa_id: return
        lines = self.db.fetch_working_lines(wa_id, sa_id)
        if lines:
            self.lines_data = {l.WorkingLineName: l.WorkingLineID for l in lines}
            self.line_combo['values'] = sorted(list(self.lines_data.keys()))
            self.line_combo.config(state="readonly")

    def _update_problems_combo(self):
        self.problem_var.set('')
        self.problem_combo.config(state="disabled", values=[])
        wa_id = self.working_areas_data.get(self.working_area_var.get())
        sa_id = self.sub_areas_data.get(self.sub_area_var.get())
        ia_id = self.issue_areas_data.get(self.issue_area_var.get())
        if not all([wa_id, sa_id, ia_id]): return
        problems = self.db.fetch_issue_problems(wa_id, ia_id, sa_id)
        if problems:
            self.problems_data = {p.IssueDescription: p.IssueProblemId for p in problems}
            self.problem_combo['values'] = sorted(list(self.problems_data.keys()))
            self.problem_combo.config(state="readonly")

    def _filter_order_combo(self, event):
        typed_text = self.order_var.get().lower()
        if not typed_text:
            self.order_combo['values'] = self.all_order_names
            return
        filtered_list = [name for name in self.all_order_names if typed_text in name.lower()]
        self.order_combo['values'] = filtered_list

    def _toggle_time_fields(self):
        if self.time_choice_var.get() == "duration":
            self.duration_entry.config(state="normal")
            self.from_entry.config(state="disabled")
            self.to_entry.config(state="disabled")
        else:
            self.duration_entry.config(state="disabled")
            self.from_entry.config(state="normal")
            self.to_entry.config(state="normal")

    def _calculate_duration(self, event=None):
        try:
            start = datetime.strptime(self.from_hour_var.get(), "%H:%M")
            end = datetime.strptime(self.to_hour_var.get(), "%H:%M")
            if end < start: end += timedelta(days=1)
            duration = end - start
            self.duration_var.set(str(int(duration.total_seconds() / 60)))
        except ValueError:
            self.duration_var.set("")

    def _on_comments_focus_in(self, event):
        if self.notes_text.get("1.0", "end-1c") == self.comments_placeholder:
            self.notes_text.delete("1.0", tk.END)
            self.notes_text.config(foreground="black")

    def _on_comments_focus_out(self, event):
        if not self.notes_text.get("1.0", "end-1c"):
            self.notes_text.insert("1.0", self.comments_placeholder)
            self.notes_text.config(foreground="grey")

    def _on_action_plan_focus_in(self, event):
        if self.action_text.get("1.0", "end-1c") == self.action_plan_placeholder:
            self.action_text.delete("1.0", tk.END)
            self.action_text.config(foreground="black")

    def _on_action_plan_focus_out(self, event):
        if not self.action_text.get("1.0", "end-1c"):
            self.action_text.insert("1.0", self.action_plan_placeholder)
            self.action_text.config(foreground="grey")
    
    def _add_document(self):
        """Apre un file dialog per selezionare un documento da allegare."""
        file_path = filedialog.askopenfilename(
            title=self.lang.get('select_document_title', "Seleziona Documento"),
            filetypes=[
                ("All Files", "*.*"),
                ("PDF Files", "*.pdf"),
                ("Image Files", "*.png *.jpg *.jpeg *.bmp *.gif"),
                ("Word Documents", "*.doc *.docx"),
                ("Excel Files", "*.xls *.xlsx")
            ]
        )
        
        if not file_path:
            return
        
        # Chiedi la descrizione del documento
        description = simpledialog.askstring(
            self.lang.get('document_description_title', "Descrizione Documento"),
            self.lang.get('document_description_prompt', "Inserisci una breve descrizione del documento:"),
            parent=self
        )
        
        if not description:
            messagebox.showwarning(
                self.lang.get('warning_title', "Attenzione"),
                self.lang.get('description_required', "La descrizione è obbligatoria per allegare un documento."),
                parent=self
            )
            return
        
        # Leggi il file come dati binari
        try:
            with open(file_path, 'rb') as f:
                binary_data = f.read()
            
            filename = os.path.basename(file_path)
            
            # Aggiungi alla lista dei documenti
            self.uploaded_documents.append({
                'filename': filename,
                'description': description,
                'binary_data': binary_data
            })
            
            self._update_document_list()
            
        except Exception as e:
            messagebox.showerror(
                self.lang.get('error_title', "Errore"),
                self.lang.get('error_reading_file', f"Errore durante la lettura del file: {e}"),
                parent=self
            )
    
    def _remove_document(self):
        """Rimuove il documento selezionato dalla lista."""
        selection = self.docs_listbox.curselection()
        
        if not selection:
            messagebox.showwarning(
                self.lang.get('warning_title', "Attenzione"),
                self.lang.get('no_document_selected', "Nessun documento selezionato."),
                parent=self
            )
            return
        
        index = selection[0]
        del self.uploaded_documents[index]
        self._update_document_list()
    
    def _update_document_list(self):
        """Aggiorna la listbox con i documenti caricati."""
        self.docs_listbox.delete(0, tk.END)
        
        for doc in self.uploaded_documents:
            display_text = f"{doc['filename']} - {doc['description']}"
            self.docs_listbox.insert(tk.END, display_text)


    def _save_interruption(self):
        try:
            params = {}
            note_text = self.notes_text.get("1.0", tk.END).strip()
            action_plan_text = self.action_text.get("1.0", tk.END).strip()

            if not note_text or note_text == self.comments_placeholder:
                messagebox.showerror("Dati Mancanti", "Il campo Commenti è obbligatorio.", parent=self)
                return
            if not action_plan_text or action_plan_text == self.action_plan_placeholder:
                messagebox.showerror("Dati Mancanti", "Il campo Piano d'Azione è obbligatorio.", parent=self)
                return
            params['note'] = note_text
            params['action_plan'] = action_plan_text

            params['report_date'] = self.date_entry.get_date()
            params['user_name'] = self.user_name
            params['working_area_id'] = self.working_areas_data[self.working_area_var.get()]
            params['sub_area_id'] = self.sub_areas_data[self.sub_area_var.get()]
            params['line_id'] = self.lines_data[self.line_var.get()]
            params['issue_area_id'] = self.issue_areas_data[self.issue_area_var.get()]
            params['problem_id'] = self.problems_data[self.problem_var.get()]

            if self.time_choice_var.get() == "duration":
                params['lost_gain'] = int(self.duration_var.get())
                params['from_hour'] = None;
                params['to_hour'] = None
            else:
                self._calculate_duration(None)
                params['lost_gain'] = int(self.duration_var.get())
                params['from_hour'] = self.from_hour_var.get();
                params['to_hour'] = self.to_hour_var.get()

            order_name = self.order_var.get()
            params['po_number'] = order_name.split(' ')[0] if order_name and ' [' in order_name else order_name
            params['product_code'] = order_name.split('[')[1].replace(']',
                                                                      '') if order_name and '[' in order_name else None
            params['hours'] = round(params['lost_gain'] / 60.0, 2)

            # Salva la testata dell'interruzione
            success, msg, breakdown_id = self.db.add_production_interruption(params)

            if success:
                # Se ci sono documenti da salvare, salvali ora
                if self.uploaded_documents:
                    doc_success, doc_msg = self.db.add_interruption_documents(
                        breakdown_id, 
                        self.uploaded_documents, 
                        self.user_name
                    )
                    
                    if not doc_success:
                        messagebox.showwarning(
                            self.lang.get('warning_title', "Attenzione"),
                            self.lang.get('document_save_error', 
                                         f"Interruzione salvata, ma errore nel salvataggio documenti: {doc_msg}"),
                            parent=self
                        )
                
                messagebox.showinfo("Successo", msg, parent=self)
                self.destroy()
            else:
                messagebox.showerror("Errore", msg, parent=self)
        except (KeyError, ValueError) as e:
            messagebox.showerror("Dati Mancanti o non Validi",
                                 f"Assicurati che tutti i campi obbligatori (*) siano compilati correttamente.\n\nDettaglio: {e}",
                                 parent=self)


class ShippingReportWindow(tk.Toplevel):
    """Finestra per la gestione avanzata dei piani di spedizione."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(self.lang.get('shipping_management_title', "Gestione e Analisi Spedizioni"))
        self.geometry("1200x750")

        self.shipping_date = None
        self.plan_data = []
        self.initial_plan_state = ""
        self.summary_data = {}

        self._create_widgets()
        self._configure_styles()  # Ora questo metodo esiste

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(2, weight=1)
        main_frame.columnconfigure(0, weight=1)

        action_frame = ttk.LabelFrame(main_frame, text=self.lang.get('main_actions_label', "Azioni Principali"))
        action_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Button(action_frame, text=self.lang.get('load_plan_btn', "1. Carica Piano Spedizioni (.xlsx)"),
                   command=self._load_plan).pack(side="left", padx=10, pady=5)
        ttk.Button(action_frame, text=self.lang.get('load_actuals_btn', "2. Carica Consuntivo Spedizioni (.xlsx)"),
                   state="disabled").pack(side="left", padx=10, pady=5)
        self.analytics_button = ttk.Button(action_frame, text=self.lang.get('create_analytics_btn', "Crea Report Analitico"), command=self._generate_and_save_report, state="disabled")
        self.analytics_button.pack(side="left", padx=10, pady=5)

        summary_frame = ttk.LabelFrame(main_frame, text=self.lang.get('next_shipping_summary_label',
                                                                      "Riepilogo Prossima Spedizione"))
        summary_frame.grid(row=1, column=0, sticky="ew", pady=5)
        self.summary_label = ttk.Label(summary_frame, font=("Helvetica", 10, "bold"), justify=tk.CENTER)
        self.summary_label.pack(padx=10, pady=10, fill="x", expand=True)

        data_frame = ttk.LabelFrame(main_frame,
                                    text=self.lang.get('shipping_plan_details_label', "Dettaglio Piano di Spedizione"))
        data_frame.grid(row=2, column=0, sticky="nsew")
        data_frame.rowconfigure(0, weight=1)
        data_frame.columnconfigure(0, weight=1)

        cols = ('status', 'order', 'product', 'original', 'modified', 'note')
        self.tree = ttk.Treeview(data_frame, columns=cols, show="headings")
        self.tree.heading('status', text=self.lang.get('header_status', "Stato"))
        self.tree.heading('order', text=self.lang.get('header_order', "Ordine"))
        self.tree.heading('product', text=self.lang.get('header_product', "Prodotto"))
        self.tree.heading('original', text=self.lang.get('header_original_qty', "Q.tà Piano"))
        self.tree.heading('modified', text=self.lang.get('header_modified_qty', "Q.tà Modificata"))
        self.tree.heading('note', text=self.lang.get('header_note', "Nota"))
        self.tree.column('status', width=120)
        self.tree.pack(side="left", fill="both", expand=True)

        edit_frame = ttk.Frame(main_frame)
        edit_frame.grid(row=3, column=0, sticky="w", pady=10)
        self.edit_qty_btn = ttk.Button(edit_frame, text=self.lang.get('edit_qty_btn', "Modifica Q.tà"),
                                       command=self._edit_quantity, state="disabled")
        self.edit_qty_btn.pack(side="left")
        self.edit_note_btn = ttk.Button(edit_frame, text=self.lang.get('edit_note_btn', "Aggiungi/Mod. Nota"),
                                        command=self._edit_note, state="disabled")
        self.edit_note_btn.pack(side="left", padx=10)
        self.remove_row_btn = ttk.Button(edit_frame, text=self.lang.get('remove_row_btn', "Rimuovi Riga"),
                                         command=self._remove_row, state="disabled")
        self.remove_row_btn.pack(side="left")
        self.tree.bind("<<TreeviewSelect>>", self._on_row_select)

        self.save_db_button = ttk.Button(main_frame,
                                         text=self.lang.get('save_plan_changes_btn', "Salva Modifiche al Piano"),
                                         command=self._save_to_db, state="disabled")
        self.save_db_button.grid(row=4, column=0, sticky="e", pady=10)

    def _configure_styles(self):
        """Configura gli stili per colorare le righe della Treeview."""
        style = ttk.Style()
        style.map("Treeview")
        style.configure("modified.Treeview", background="#FFFACD")  # Giallo chiaro
        style.configure("removed.Treeview", background="#FFC0CB")  # Rosa
        style.configure("new.Treeview", background="#D4EDDA")  # Verde chiaro

    def _load_plan(self):
        # 1. Recupera le impostazioni di spedizione
        settings = self.db.fetch_shipping_settings()

        # 2. Recupera la configurazione del file Excel (nome del foglio, ecc.)
        mappings, sheet_name = self.db.fetch_xls_mappings(file_type_id=1)
        if not mappings or not sheet_name:
            messagebox.showerror(self.lang.get('error_config_title', "Errore Configurazione"),
                                 self.lang.get('error_config_message',
                                               "Mappatura del file di pianificazione non trovata."), parent=self)
            return
        xls_config = {'mappings': mappings, 'sheet_name': sheet_name}

        # 3. Chiede il file all'utente
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

        if not file_path: return

        # 4. Carica i dati GIA' SALVATI dal DB per il confronto
        # NOTA: Per ora, carichiamo tutti i dati. In futuro potremmo filtrare per data.
        db_data = self.db.fetch_shipping_plan_items(datetime.now().date())

        # 5. Chiama la funzione unificata del processore
        data, summary, error = shipping_processor.process_shipping_file(
            file_path, settings, db_data, self.user_name, xls_config
        )

        if error:
            messagebox.showerror(self.lang.get('error_processing_title', "Errore Elaborazione"), error, parent=self)
            return

        # 6. Salva i dati e aggiorna la UI
        self.shipping_date = datetime.strptime(summary['next_ship_date'], '%d/%m/%Y').date()
        self.plan_data = data
        self.summary_data = summary
        self.initial_plan_state = json.dumps(self.plan_data, sort_keys=True, default=str)
        self._update_ui()

    def _update_ui(self):
        """Aggiorna tutti i componenti della UI con i nuovi dati."""
        self._update_summary()
        self._update_treeview()
        # Abilita il pulsante Salva solo se ci sono dati caricati
        self.save_db_button.config(state="normal" if self.plan_data else "disabled")
        # Attiva il pulsante del report analitico
        self.analytics_button.config(state="normal" if self.plan_data else "disabled")

    def _update_summary(self):
        if not self.summary_data: self.summary_label.config(text=""); return
        num_orders = self.summary_data.get('total_orders', 0)
        num_items = len(
            [d for d in self.plan_data if d.get('status') not in ['status_removed_by_plan', 'status_removed_manually']])
        total_quantity = self.summary_data.get('total_quantity', 0)
        summary_text = (
            f"{self.lang.get('shipping_date_label', 'Data Spedizione')}: {self.summary_data.get('next_ship_date', 'N/D')}   |   "
            f"{self.lang.get('shipping_type_label', 'Tipo Spedizione')}: Normale   |   "
            f"{self.lang.get('orders_to_ship_label', 'Ordini da Spedire')}: {num_orders}   |   "
            f"{self.lang.get('items_to_ship_label', 'Schede da Spedire')}: {total_quantity}")
        self.summary_label.config(text=summary_text)

    def _update_treeview(self):
        self.tree.delete(*self.tree.get_children())
        if not self.plan_data: return
        for i, item in enumerate(self.plan_data):
            tag = ""
            status_key = item.get('status', '')
            if 'modified' in status_key:
                tag = "modified"
            elif 'removed' in status_key:
                tag = "removed"
            elif 'new' in status_key:
                tag = "new"
            translated_status = self.lang.get(status_key, status_key)
            self.tree.insert("", "end", iid=i, values=(
                translated_status, item.get('order', ''), item.get('product', ''),
                item.get('original_qty', ''), item.get('modified_qty', ''),
                item.get('note', '')), tags=(tag,))

    def _on_row_select(self, event=None):
        if self.tree.focus():
            self.edit_qty_btn.config(state="normal")
            self.edit_note_btn.config(state="normal")
            self.remove_row_btn.config(state="normal")
        else:
            self.edit_qty_btn.config(state="disabled")
            self.edit_note_btn.config(state="disabled")
            self.remove_row_btn.config(state="disabled")

    def _get_selected_item_index(self):
        selected_iid = self.tree.focus()
        if not selected_iid:
            messagebox.showwarning("Nessuna Selezione", "Selezionare una riga dalla lista.")
            return None
        return int(selected_iid)

    def _edit_quantity(self):
        index = self._get_selected_item_index()
        if index is None: return
        item = self.plan_data[index]
        prompt = f"Modifica quantità per:\n{item['product']}\n(Piano originale: {item['original_qty']})"
        new_qty_str = simpledialog.askstring("Modifica Quantità", prompt, parent=self)
        if new_qty_str:
            try:
                item['modified_qty'] = int(new_qty_str)
                item['status'] = 'status_modified_manually'
                self._update_ui()
            except ValueError:
                messagebox.showerror("Errore", "Inserire un valore numerico valido.", parent=self)

    def _edit_note(self):
        index = self._get_selected_item_index()
        if index is None: return
        item = self.plan_data[index]
        prompt = f"Nota per:\n{item['product']}"
        new_note = simpledialog.askstring("Modifica Nota", prompt, initialvalue=item['note'], parent=self)
        if new_note is not None:
            item['note'] = new_note
            self._update_treeview()

    def _remove_row(self):
        index = self._get_selected_item_index()
        if index is None: return
        if messagebox.askyesno("Conferma Rimozione", "Sei sicuro di voler rimuovere questa riga dal piano?"):
            self.plan_data[index]['status'] = 'status_removed_manually'
            self._update_ui()

    def _save_to_db(self):
        """Salva le modifiche al piano di spedizione nel database."""
        if not self.plan_data: return

        current_state = json.dumps(self.plan_data, sort_keys=True, default=str)

        # Procede solo se ci sono state modifiche
        if current_state != self.initial_plan_state:
            if not messagebox.askyesno("Conferma Salvataggio", "Salvare le modifiche apportate al piano nel database?"):
                return

            success, msg = self.db.save_shipping_plan_items(self.plan_data)
            if success:
                messagebox.showinfo("Successo", msg, parent=self)
                self._refresh_data_from_db()  # Ricarica per aggiornare lo stato
            else:
                messagebox.showerror("Errore", msg, parent=self)
        else:
            # Se non ci sono modifiche, informa l'utente
            messagebox.showinfo(
                self.lang.get('no_changes_title', "Nessuna Modifica"),
                self.lang.get('no_changes_message',
                              "Il piano non è stato modificato. Usa il pulsante 'Crea Report Analitico' per generare il file."),
                parent=self
            )

    def _refresh_data_from_db(self):
        if not self.shipping_date: return
        self.tree.delete(*self.tree.get_children())
        db_data = self.db.fetch_shipping_plan_items(self.shipping_date)
        self.plan_data = []
        for item in db_data:
            self.plan_data.append({
                'id': item.ItemId, 'shipping_date': item.ShippingDate.strftime('%Y-%m-%d'),
                'order': item.OrderNumber, 'product': item.ProductCode,
                'original_qty': item.OriginalQty, 'modified_qty': item.ModifiedQty,
                'note': item.Note, 'status': item.Status, 'user': item.UpdatedBy
            })
        self.initial_plan_state = json.dumps(self.plan_data, sort_keys=True, default=str)
        self._update_ui()

    def _generate_and_save_report(self):
        """Crea il file Excel del report e lo salva nel backlog."""
        if not self.plan_data:
            messagebox.showwarning("Dati Mancanti", "Nessun dato da inserire nel report.", parent=self)
            return

        # Controlla se ci sono modifiche non salvate
        current_state = json.dumps(self.plan_data, sort_keys=True, default=str)
        if current_state != self.initial_plan_state:
            if messagebox.askyesno("Modifiche non Salvate",
                                   "Ci sono modifiche non salvate. Vuoi salvarle prima di creare il report?"):
                self._save_to_db()
                # Se il salvataggio fallisce, non procedere
                if json.dumps(self.plan_data, sort_keys=True, default=str) != self.initial_plan_state:
                    return
            else:
                # L'utente non vuole salvare, quindi annulla la creazione del report
                return

        # 1. Crea il file Excel formattato in memoria
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.lang.get('excel_sheet_title', "Riepilogo Spedizione")

        headers = [
            self.lang.get('header_status', "Stato"), self.lang.get('header_order', "Ordine"),
            self.lang.get('header_product', "Prodotto"), self.lang.get('header_original_qty', "Q.tà Piano"),
            self.lang.get('header_modified_qty', "Q.tà Modificata"), self.lang.get('header_note', "Nota")
        ]
        ws.append(headers)

        # Applica stili all'header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Popola il file con i dati, usando le quantità modificate se presenti
        for item in self.plan_data:
            # Usa la quantità modificata se esiste, altrimenti quella originale
            final_qty = item.get('modified_qty') if item.get('modified_qty') is not None else item.get('original_qty')

            ws.append([
                self.lang.get(item.get('status', ''), item.get('status', '')),
                item.get('order'), item.get('product'),
                item.get('original_qty'), item.get('modified_qty'),
                item.get('note')
            ])

        ws.auto_filter.ref = ws.dimensions

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_binary_data = excel_buffer.getvalue()

        # 2. Salva il report nel database (tabelle BackLogs)
        general_note = f"Report generato per la spedizione del {self.summary_data['next_ship_date']}."
        notes_to_save = [(i, item['note']) for i, item in enumerate(self.plan_data) if item['note']]

        success, msg = self.db.save_backlog_report(self.user_name, general_note, excel_binary_data, notes_to_save)

        if success:
            messagebox.showinfo("Successo", msg, parent=self)
        else:
            messagebox.showerror("Errore", msg, parent=self)


class ShippingSettingsWindow(tk.Toplevel):
    """Finestra per la gestione delle impostazioni di spedizione."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.title(lang.get('shipping_settings_title', "Impostazioni di Spedizione"))
        self.geometry("600x400")
        self.transient(parent)
        self.grab_set()

        self.day_var = tk.StringVar()
        self.type_var = tk.StringVar()
        self.days_map = {
            "1 - Lunedì": 1, "2 - Martedì": 2, "3 - Mercoledì": 3,
            "4 - Giovedì": 4, "5 - Venerdì": 5, "6 - Sabato": 6, "7 - Domenica": 7
        }
        self.reverse_days_map = {v: k for k, v in self.days_map.items()}

        self._create_widgets()
        self._load_settings()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # Form di inserimento
        form_frame = ttk.LabelFrame(main_frame,
                                    text=self.lang.get('add_new_setting_label', "Aggiungi Nuova Impostazione"))
        form_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        form_frame.columnconfigure(1, weight=1)

        ttk.Label(form_frame, text=self.lang.get('day_of_week_label', "Giorno della Settimana:")).grid(row=0, column=0,
                                                                                                       padx=5, pady=5)
        self.day_combo = ttk.Combobox(form_frame, textvariable=self.day_var, state="readonly",
                                      values=list(self.days_map.keys()))
        self.day_combo.grid(row=0, column=1, sticky="ew", padx=5)

        ttk.Label(form_frame, text=self.lang.get('shipping_type_label', "Tipo di Trasporto:")).grid(row=1, column=0,
                                                                                                    padx=5, pady=5)
        self.type_combo = ttk.Combobox(form_frame, textvariable=self.type_var, state="readonly",
                                       values=["Normale", "Speciale"])
        self.type_combo.grid(row=1, column=1, sticky="ew", padx=5)

        ttk.Button(form_frame, text=self.lang.get('add_button', "Aggiungi"), command=self._add_setting).grid(row=1,
                                                                                                             column=2,
                                                                                                             padx=10)

        # Lista impostazioni attive
        list_frame = ttk.LabelFrame(main_frame, text=self.lang.get('active_settings_label', "Impostazioni Attive"))
        list_frame.grid(row=1, column=0, sticky="nsew")
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        cols = ('day', 'type')
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings")
        self.tree.heading('day', text=self.lang.get('header_day', "Giorno"))
        self.tree.heading('type', text=self.lang.get('header_type', "Tipo"))
        self.tree.pack(side="left", fill="both", expand=True)

        ttk.Button(main_frame, text=self.lang.get('delete_selected_button', "Cancella Selezionata"),
                   command=self._delete_setting).grid(row=2, column=0, sticky="e", pady=10)

    def _load_settings(self):
        self.tree.delete(*self.tree.get_children())
        settings = self.db.fetch_shipping_settings()
        for setting in settings:
            day_name = self.reverse_days_map.get(setting.DayOfWeek, "Sconosciuto")
            self.tree.insert("", "end", iid=setting.ShippingSettingId, values=(day_name, setting.ShippingType))

    def _add_setting(self):
        day_name = self.day_var.get()
        ship_type = self.type_var.get()

        if not day_name or not ship_type:
            messagebox.showwarning("Dati Mancanti", "Selezionare un giorno e un tipo di trasporto.", parent=self)
            return

        day_of_week = self.days_map.get(day_name)
        success, message = self.db.add_shipping_setting(day_of_week, ship_type)

        if success:
            messagebox.showinfo("Successo", message, parent=self)
            self._load_settings()
        else:
            messagebox.showerror("Errore", message, parent=self)

    def _delete_setting(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Nessuna Selezione", "Selezionare un'impostazione da cancellare.", parent=self)
            return

        setting_id = int(selected_item)
        if messagebox.askyesno("Conferma Cancellazione", "Sei sicuro di voler cancellare questa impostazione?"):
            success, message = self.db.delete_shipping_setting(setting_id)
            if success:
                messagebox.showinfo("Successo", message, parent=self)
                self._load_settings()
            else:
                messagebox.showerror("Errore", message, parent=self)


class XlsSettingsWindow(tk.Toplevel):
    """Finestra per configurare la mappatura delle colonne e del foglio dei file Excel."""

    def __init__(self, parent, db, lang, user_name):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.title(self.lang.get('xls_settings_title', "Impostazioni Import Excel"))
        self.geometry("600x500")
        self.transient(parent)
        self.grab_set()

        self.file_types_data = {}
        self.current_mappings = {}
        self.file_type_var = tk.StringVar()
        self.sheet_name_var = tk.StringVar()

        self._create_widgets()
        self._load_file_types()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        main_frame.rowconfigure(2, weight=1)
        main_frame.columnconfigure(0, weight=1)

        top_frame = ttk.Frame(main_frame)
        top_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Label(top_frame, text=self.lang.get('select_file_type_label', "Seleziona Tipo File da Configurare:")).pack(
            side="left")
        self.file_type_combo = ttk.Combobox(top_frame, textvariable=self.file_type_var, state="readonly")
        self.file_type_combo.pack(side="left", padx=10, expand=True, fill="x")
        self.file_type_combo.bind("<<ComboboxSelected>>", self._load_mappings)

        sheet_frame = ttk.Frame(main_frame)
        sheet_frame.grid(row=1, column=0, sticky="ew", pady=5)
        ttk.Label(sheet_frame, text=self.lang.get('sheet_name_label', "Nome Foglio di Lavoro:")).pack(side="left")
        self.sheet_name_entry = ttk.Entry(sheet_frame, textvariable=self.sheet_name_var)
        self.sheet_name_entry.pack(side="left", padx=10, expand=True, fill="x")

        self.grid_frame = ttk.LabelFrame(main_frame, text=self.lang.get('column_mappings_label', "Mappatura Colonne"))
        self.grid_frame.grid(row=2, column=0, sticky="nsew")
        self.grid_frame.columnconfigure(1, weight=1)
        self.grid_frame.columnconfigure(2, weight=1)

    def _load_file_types(self):
        types = self.db.fetch_xls_file_types()
        if types:
            self.file_types_data = {self.lang.get(t.TranslationKey, t.FileTypeName): t.FileTypeId for t in types}
            self.file_type_combo['values'] = sorted(list(self.file_types_data.keys()))

    def _load_mappings(self, event=None):
        for widget in self.grid_frame.winfo_children():
            widget.destroy()

        file_type_name = self.file_type_var.get()
        file_type_id = self.file_types_data.get(file_type_name)
        if not file_type_id: return

        self.current_mappings, sheet_name = self.db.fetch_xls_mappings(file_type_id)
        self.sheet_name_var.set(sheet_name or "")

        ttk.Label(self.grid_frame, text=self.lang.get('field_name_header', "Campo Dati"),
                  font=("Helvetica", 9, "bold")).grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(self.grid_frame, text=self.lang.get('column_letter_header', "Lettera Colonna"),
                  font=("Helvetica", 9, "bold")).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(self.grid_frame, text=self.lang.get('start_row_header', "Riga Inizio Dati"),
                  font=("Helvetica", 9, "bold")).grid(row=0, column=2, padx=5, pady=5)

        row_index = 1
        for field, details in self.current_mappings.items():
            ttk.Label(self.grid_frame, text=f"{field}:").grid(row=row_index, column=0, sticky="w", padx=5)
            col_var = tk.StringVar(value=details['col'])
            row_var = tk.StringVar(value=details['row'])
            ttk.Entry(self.grid_frame, textvariable=col_var, width=10).grid(row=row_index, column=1, sticky="w", padx=5)
            ttk.Entry(self.grid_frame, textvariable=row_var, width=10).grid(row=row_index, column=2, sticky="w", padx=5)
            details['col_var'] = col_var
            details['row_var'] = row_var
            row_index += 1

        ttk.Button(self.grid_frame, text=self.lang.get('save_mappings_button', "Salva Modifiche"),
                   command=self._save_mappings).grid(row=row_index, column=2, sticky="e", pady=15)

    def _save_mappings(self):
        try:
            file_type_id = self.file_types_data.get(self.file_type_var.get())
            if not file_type_id: return

            new_sheet_name = self.sheet_name_var.get().strip()
            self.db.update_xls_sheet_name(file_type_id, new_sheet_name)

            for field, details in self.current_mappings.items():
                mapping_id = details['id']
                new_col = details['col_var'].get().upper()
                new_row = int(details['row_var'].get())
                self.db.update_xls_mapping(mapping_id, new_col, new_row)

            messagebox.showinfo("Successo", "Impostazioni di mappatura salvate con successo.", parent=self)
        except Exception as e:
            messagebox.showerror("Errore",
                                 f"Impossibile salvare le modifiche. Controllare i valori inseriti.\n\nDettaglio: {e}",
                                 parent=self)


def open_add_interruption_window(parent, db, lang, user_name):
    AddInterruptionWindow(parent, db, lang, user_name)


def open_shipping_report_window(parent, db, lang, user_name):
    ShippingReportWindow(parent, db, lang, user_name)


def open_shipping_settings_window(parent, db, lang, user_name):
    ShippingSettingsWindow(parent, db, lang, user_name)


def open_xls_settings_window(parent, db, lang, user_name):
    XlsSettingsWindow(parent, db, lang, user_name)