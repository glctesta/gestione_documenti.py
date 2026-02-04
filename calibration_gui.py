
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import collections.abc
from tkinter import filedialog
import os, sys, tempfile, subprocess

import logging

logger = logging.getLogger(__name__)

class CalibrationsWindow(tk.Toplevel):
    def __init__(self, parent, db_object, language_manager):
        super().__init__(parent)
        self.parent = parent
        self.db = db_object
        self.lang = language_manager

        self.equipment_map = {}
        self.supplier_map = {}
        self.all_supplier_names = []
        self.all_equipment_names = []  # Lista completa equipment per filtro
        self.current_equipment_id = None

        self.title(self.lang.get('calibrations_title', "Gestione Calibrazioni"))
        self.geometry("650x600")  # Aumentata altezza per il nuovo bottone
        self.transient(parent)
        self.grab_set()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._create_widgets()
        self._load_initial_data()

        self.selected_pdf_path = None  # PDF scelto per nuovo inserimento
        self.selected_pdf_bytes = None  # bytes del PDF scelto per nuovo inserimento
        self.current_cert_bytes = None  # bytes del certificato dell'ultima calibrazione caricata

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # COMBOBOX MACCHINARI IN TESTA ALLA FORM
        select_frame = ttk.LabelFrame(main_frame, text=self.lang.get('select_equipment', "Seleziona Attrezzatura"),
                                      padding="10")
        select_frame.pack(fill=tk.X, expand=True, pady=(0, 10))
        self.combo_equipment = ttk.Combobox(select_frame, state="normal", font=('Segoe UI', 10))
        self.combo_equipment.pack(fill=tk.X, expand=True)
        self.combo_equipment.bind("<<ComboboxSelected>>", self._on_equipment_select)
        self.combo_equipment.bind('<KeyRelease>', self._on_equipment_search)

        self.details_frame = ttk.LabelFrame(main_frame,
                                            text=self.lang.get('calibration_details', "Dettagli Ultima Calibrazione"),
                                            padding="10")
        ttk.Label(self.details_frame, text=self.lang.get('last_calibration_date', "Data ultima calibrazione:")).grid(
            row=0, column=0, sticky="w", pady=2, padx=5)
        self.lbl_last_date = ttk.Label(self.details_frame, text="N/D", font=('Segoe UI', 10, 'bold'))
        self.lbl_last_date.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        ttk.Label(self.details_frame, text=self.lang.get('expiry_date', "Data di scadenza:")).grid(row=1, column=0,
                                                                                                   sticky="w", pady=2,
                                                                                                   padx=5)
        self.lbl_expiry_date = ttk.Label(self.details_frame, text="N/D", font=('Segoe UI', 10, 'bold'))
        self.lbl_expiry_date.grid(row=1, column=1, sticky="w", pady=2, padx=5)

        # Stato certificato + apri
        ttk.Label(self.details_frame, text=self.lang.get('certificate_status', "Stato certificato:")).grid(
            row=2, column=0, sticky="w", pady=2, padx=5)
        self.lbl_cert_status = ttk.Label(self.details_frame, text=self.lang.get('certificate_absent', "Assente"))
        self.lbl_cert_status.grid(row=2, column=1, sticky="w", pady=2, padx=5)
        self.btn_open_cert = ttk.Button(self.details_frame, text=self.lang.get('open_certificate', "Apri certificato"),
                                        command=self._open_certificate, state="disabled")
        self.btn_open_cert.grid(row=2, column=2, sticky="e", pady=2, padx=5)
        
        # Bottone Export Storico
        self.btn_export_history = ttk.Button(self.details_frame, 
                                             text=self.lang.get('export_history', "Esporta Storico"),
                                             command=self._export_calibration_history, 
                                             state="disabled")
        self.btn_export_history.grid(row=3, column=1, columnspan=2, sticky="e", pady=5, padx=5)

        self.insert_frame = ttk.LabelFrame(main_frame,
                                           text=self.lang.get('new_calibration_data', "Inserisci Nuova Calibrazione"),
                                           padding="10")
        ttk.Label(self.insert_frame, text=self.lang.get('new_expiry_date', "Nuova data di scadenza:")).grid(row=0,
                                                                                                            column=0,
                                                                                                            sticky="w",
                                                                                                            pady=5,
                                                                                                            padx=5)
        self.entry_new_expiry_date = DateEntry(self.insert_frame, width=18, background='darkblue', foreground='white',
                                               borderwidth=2, date_pattern='yyyy-mm-dd')
        self.entry_new_expiry_date.grid(row=0, column=1, sticky="w", pady=5, padx=5)

        ttk.Label(self.insert_frame, text=self.lang.get('certifying_body', "Ente certificatore:")).grid(row=1, column=0,
                                                                                                        sticky="w",
                                                                                                        pady=5, padx=5)
        self.combo_cert_body = ttk.Combobox(self.insert_frame, width=35)
        self.combo_cert_body.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        self.combo_cert_body.bind('<KeyRelease>', self._on_supplier_search)
        self.combo_cert_body.bind('<<ComboboxSelected>>', lambda e: self.focus())

        # Sostituisce "Numero certificato" con Upload PDF
        ttk.Label(self.insert_frame, text=self.lang.get('certificate_pdf', "Certificato (PDF):")).grid(row=2,
                                                                                                       column=0,
                                                                                                       sticky="w",
                                                                                                       pady=5, padx=5)
        self.lbl_cert_file = ttk.Label(self.insert_frame,
                                       text=self.lang.get('no_file_selected', "Nessun file selezionato"),
                                       foreground="gray")
        self.lbl_cert_file.grid(row=2, column=1, sticky="w", pady=5, padx=5)
        self.btn_upload_pdf = ttk.Button(self.insert_frame, text=self.lang.get('upload_pdf', "Carica PDF"),
                                         command=self._choose_pdf_file)
        self.btn_upload_pdf.grid(row=2, column=2, sticky="w", pady=5, padx=5)

        self.btn_save = ttk.Button(self.insert_frame, text=self.lang.get('save_button', "Salva"),
                                   command=self._save_calibration)
        self.btn_save.grid(row=3, column=1, sticky="e", pady=10, padx=5)
        self.btn_save.state(["disabled"])  # Disabilitato finché non viene caricato un PDF

    def _load_initial_data(self):
        self._load_equipment_list()
        self._load_suppliers()

    def _load_equipment_list(self):
        try:
            rows = self.db.get_calibratable_equipment()
            if not rows:
                self.combo_equipment['values'] = [
                    self.lang.get('no_equipment_found', "Nessuna attrezzatura da calibrare trovata.")]
                return
            equipment_display_list = [f"{row.InternalName} (Mat: {row.InventoryNumber}) - {row.Brand}" for row in rows]
            self.equipment_map = {f"{row.InternalName} (Mat: {row.InventoryNumber}) - {row.Brand}": row.EquipmentId for
                                  row in rows}
            self.all_equipment_names = equipment_display_list  # Salva lista completa per filtro
            self.combo_equipment['values'] = equipment_display_list
        except Exception as e:
            messagebox.showerror(self.lang.get('error', "Errore"), f"Impossibile caricare la lista attrezzature:\n{e}",
                                 parent=self)
    
    def _on_equipment_search(self, event=None):
        """Filtra la lista equipment mentre l'utente digita"""
        value = self.combo_equipment.get().lower()
        if value == '':
            self.combo_equipment['values'] = self.all_equipment_names
        else:
            filtered_data = [name for name in self.all_equipment_names if value in name.lower()]
            self.combo_equipment['values'] = filtered_data

    def _load_suppliers(self):
        try:
            rows = self.db.get_suppliers()
            if not rows: return
            self.all_supplier_names = sorted([row.SiteName for row in rows])
            self.supplier_map = {row.SiteName: row.IDSite for row in rows}
            self.combo_cert_body['values'] = self.all_supplier_names
        except Exception as e:
            messagebox.showerror(self.lang.get('error', "Errore"), f"Impossibile caricare la lista fornitori:\n{e}",
                                 parent=self)

    def _on_supplier_search(self, event=None):
        value = self.combo_cert_body.get().lower()
        if value == '':
            self.combo_cert_body['values'] = self.all_supplier_names
        else:
            filtered_data = [name for name in self.all_supplier_names if value in name.lower()]
            self.combo_cert_body['values'] = filtered_data

    def _on_equipment_select(self, event=None):
        selected_display_name = self.combo_equipment.get()
        if not selected_display_name: return
        equipment_id = self.equipment_map.get(selected_display_name)
        if equipment_id:
            self.current_equipment_id = equipment_id
            self._load_calibration_data(equipment_id)

    def _load_calibration_data(self, equipment_id):
        try:
            # reset stato certificati
            self.current_cert_bytes = None
            self.selected_pdf_path = None
            self.selected_pdf_bytes = None
            self.lbl_cert_file.config(text=self.lang.get('no_file_selected', "Nessun file selezionato"),
                                      foreground="gray")
            self.btn_save.state(["disabled"])

            row = self.db.get_last_calibration(equipment_id)

            if row:
                self.lbl_last_date.config(text=str(row.CalibratedOn) if row.CalibratedOn else "N/D")
                self.lbl_expiry_date.config(text=str(row.ExpireOn) if row.ExpireOn else "NESSUNA")
                # certificato
                cert = getattr(row, 'NrCertificate', None)
                if cert:
                    try:
                        self.current_cert_bytes = bytes(cert)
                    except Exception:
                        self.current_cert_bytes = cert  # pyodbc può restituire già bytes
                has_cert = bool(self.current_cert_bytes)
                self.lbl_cert_status.config(
                    text=self.lang.get('certificate_present', "Allegato") if has_cert
                    else self.lang.get('certificate_absent', "Assente")
                )
                self.btn_open_cert.configure(state="normal" if has_cert else "disabled")
                # Abilita export storico
                self.btn_export_history.configure(state="normal")
            else:
                self.lbl_last_date.config(text="Nessuna calibrazione registrata")
                self.lbl_expiry_date.config(text="N/D")
                self.lbl_cert_status.config(text=self.lang.get('certificate_absent', "Assente"))
                self.btn_open_cert.configure(state="disabled")
                # Abilita export storico anche se non ci sono calibrazioni (per mostrare vuoto)
                self.btn_export_history.configure(state="normal")

            self.details_frame.pack(fill=tk.X, expand=True, pady=10)

            # Mostra sempre il frame di inserimento per nuove calibrazioni
            self.insert_frame.pack(fill=tk.X, expand=True, pady=10)

        except Exception as e:
            messagebox.showerror(self.lang.get('error', "Errore"), f"Impossibile caricare i dati di calibrazione:\n{e}",
                                 parent=self)
    
    def _export_calibration_history(self):
        """Esporta lo storico calibrazioni dell'equipment selezionato in Excel"""
        if not self.current_equipment_id:
            messagebox.showwarning(
                self.lang.get('warning', "Attenzione"),
                "Selezionare un'attrezzatura prima di esportare lo storico.",
                parent=self
            )
            return
        
        try:
            # Recupera nome equipment per il filename
            selected_equipment_name = self.combo_equipment.get()
            equipment_safe_name = selected_equipment_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            
            # Recupera tutte le calibrazioni per questo equipment (anche quelle non valide)
            calibrations = self.db.get_all_calibrations_history(self.current_equipment_id)
            
            if not calibrations:
                messagebox.showinfo(
                    self.lang.get('info', "Informazione"),
                    "Nessuna calibrazione trovata per questa attrezzatura.",
                    parent=self
                )
                return
            
            # Crea directory C:\Temp se non esiste
            import os
            temp_dir = r"C:\Temp"
            os.makedirs(temp_dir, exist_ok=True)
            
            # Nome file con timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Calibrazioni_{equipment_safe_name}_{timestamp}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            # Crea Excel con openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Storico Calibrazioni"
            
            # Header
            headers = [
                "ID Calibrazione",
                "Data Calibrazione",
                "Data Scadenza",
                "Ente Certificatore",
                "Valido"
            ]
            
            # Stile header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dati
            for row_num, cal in enumerate(calibrations, 2):
                ws.cell(row=row_num, column=1, value=cal.CalibrationID)
                ws.cell(row=row_num, column=2, value=str(cal.CalibratedOn) if cal.CalibratedOn else "N/D")
                ws.cell(row=row_num, column=3, value=str(cal.ExpireOn) if cal.ExpireOn else "N/D")
                
                # Recupera nome fornitore
                supplier_name = "N/D"
                if hasattr(cal, 'SupplierId') and cal.SupplierId:
                    for name, id_val in self.supplier_map.items():
                        if id_val == cal.SupplierId:
                            supplier_name = name
                            break
                ws.cell(row=row_num, column=4, value=supplier_name)
                
                ws.cell(row=row_num, column=5, value="Sì" if getattr(cal, 'IsValid', 1) == 1 else "No")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salva file
            wb.save(filepath)
            
            # Apri file automaticamente
            if hasattr(os, "startfile"):  # Windows
                os.startfile(filepath)
            else:
                if sys.platform == "darwin":
                    subprocess.Popen(["open", filepath])
                else:
                    subprocess.Popen(["xdg-open", filepath])
            
            messagebox.showinfo(
                self.lang.get('success', "Successo"),
                f"Storico esportato con successo:\n{filepath}",
                parent=self
            )
            
        except Exception as e:
            messagebox.showerror(
                self.lang.get('error', "Errore"),
                f"Errore durante l'esportazione:\n{e}",
                parent=self
            )
            logger.error(f"Errore export calibrazioni: {e}", exc_info=True)

    def _choose_pdf_file(self):
        path = filedialog.askopenfilename(
            title=self.lang.get('select_pdf_title', "Seleziona file PDF"),
            filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            self.btn_save.state(["disabled"])
            return
        try:
            with open(path, "rb") as f:
                data = f.read()
            if not data:
                raise ValueError("Empty file")
            self.selected_pdf_path = path
            self.selected_pdf_bytes = data
            filename = os.path.basename(path)
            self.lbl_cert_file.config(text=filename, foreground="black")
            self.btn_save.state(["!disabled"])  # abilita il pulsante Salva
        except Exception as e:
            self.btn_save.state(["disabled"])
            messagebox.showerror(self.lang.get('error', "Errore"),
                                 self.lang.get('pdf_load_error', f"Impossibile caricare il PDF: {e}"),
                                 parent=self)

    def _open_certificate(self):
        if not self.current_cert_bytes:
            messagebox.showinfo(self.lang.get('info', "Informazione"),
                                self.lang.get('no_certificate_to_open', "Nessun certificato da aprire."),
                                parent=self)
            return
        try:
            fd, temp_path = tempfile.mkstemp(prefix="calibration_", suffix=".pdf")
            os.close(fd)
            with open(temp_path, "wb") as f:
                f.write(self.current_cert_bytes)
            # Apri con app predefinita
            if hasattr(os, "startfile"):  # Windows
                os.startfile(temp_path)
            else:
                if sys.platform == "darwin":
                    subprocess.Popen(["open", temp_path])
                else:
                    subprocess.Popen(["xdg-open", temp_path])
        except Exception as e:
            messagebox.showerror(self.lang.get('error', "Errore"),
                                 self.lang.get('pdf_open_error', f"Impossibile aprire il PDF: {e}"),
                                 parent=self)

    def _save_calibration(self):
        selected_equipment_name = self.combo_equipment.get()
        equipment_id = self.equipment_map.get(selected_equipment_name)
        selected_supplier_name = self.combo_cert_body.get().strip()
        supplier_id = self.supplier_map.get(selected_supplier_name)
        new_expiry_date = self.entry_new_expiry_date.get_date().strftime('%Y-%m-%d')

        # Recupera il nome utente loggato dal parent
        username = self._get_logged_in_username()

        if not equipment_id:
            messagebox.showwarning(self.lang.get('missing_data', "Dati Mancanti"),
                                   "Selezionare un'attrezzatura valida.", parent=self)
            return
        if not supplier_id:
            messagebox.showwarning(self.lang.get('missing_data', "Dati Mancanti"),
                                   self.lang.get('supplier_not_valid',
                                                 "Selezionare un ente certificatore valido dalla lista."),
                                   parent=self)
            return
        if not self.selected_pdf_bytes:
            messagebox.showwarning(
                self.lang.get('missing_data', "Dati Mancanti"),
                self.lang.get('certificate_required', "Caricare un certificato PDF prima di salvare."),
                parent=self
            )
            return

        try:
            # LOGICA SEMPLIFICATA: Sempre INSERT nuovo record
            # 1. Invalida tutte le calibrazioni precedenti per questo equipment (IsValid = 0)
            self.db.invalidate_previous_calibrations(equipment_id)
            
            # 2. Inserisce la nuova calibrazione (IsValid = 1 di default)
            self.db.add_new_calibration(
                equipment_id, 
                new_expiry_date, 
                supplier_id, 
                self.selected_pdf_bytes,
                username
            )
            
            messagebox.showinfo(
                self.lang.get('success', "Successo"),
                "Nuova calibrazione inserita correttamente.",
                parent=self
            )

            # Reset campi
            self.selected_pdf_path = None
            self.selected_pdf_bytes = None
            self.lbl_cert_file.config(
                text=self.lang.get('no_file_selected', "Nessun file selezionato"),
                foreground="gray"
            )
            self.btn_save.state(["disabled"])
            self.combo_cert_body.set('')
            
            # Ricarica dati per mostrare la nuova calibrazione
            self._load_calibration_data(equipment_id)

        except Exception as e:
            messagebox.showerror(
                self.lang.get('error', "Errore di Salvataggio"),
                f"Impossibile salvare i dati:\n{e}", 
                parent=self
            )

    def _get_logged_in_username(self):
        """Recupera il nome dell'utente loggato dalla finestra parent"""
        try:
            # Prova a recuperare l'utente dal parent (dove è stato salvato dal login)
            if hasattr(self.parent, 'current_user') and self.parent.current_user:
                return self.parent.current_user.name
            else:
                # Fallback: cerca di recuperare l'utente da qualsiasi metodo disponibile
                if hasattr(self.parent, 'get_current_username'):
                    return self.parent.get_current_username()
                else:
                    return 'Unknown'
        except Exception as e:
            logger.error(f"Errore nel recupero username: {e}")
            return 'Unknown'

    def destroy(self):
        """Override del metodo destroy per cleanup"""
        try:
            # Rilascia eventuali risorse
            self.grab_release()
            # Chiama il destroy della classe padre
            super().destroy()
        except Exception as e:
            logger.error(f"Errore durante il destroy: {e}")
            super().destroy()

    def _on_close(self):
        """Gestisce la chiusura della finestra"""
        try:
            self.grab_release()  # Rilascia il grab
            self.destroy()  # Distrugge la finestra
        except Exception as e:
            logger.error(f"Errore durante la chiusura della finestra: {e}")
            self.destroy()