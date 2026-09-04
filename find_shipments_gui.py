# -*- coding: utf-8 -*-
"""
find_shipments_gui.py
Form "Trova spedizioni": ricerca nei packing list di spedizione (database
WarehouseFinish) per ProductCode, CodePallet, CommercialNumber, UniqueNumber e
OrderNumberProduction. Con esportazione Excel formattata (intestazioni azzurro
pastello con filtri) nella directory temp.
"""
from __future__ import annotations
import logging
import os
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox

logger = logging.getLogger("TraceabilityRS")

# Tutte le tabelle risiedono in WarehouseFinish (PackingLists/Pallets/Packing/
# PackingDetail/Pick/OrderDet/IncomingDet/Item + schema ERP).
_Q_BASE_INNER = """
SELECT
    (SELECT Item.Code
       FROM WarehouseFinish.dbo.IncomingDet
       INNER JOIN WarehouseFinish.dbo.Item ON Item.ItemId = IncomingDet.ItemId
      WHERE IncomingDetId = Packing.IncomingDetId) AS ProductCode,
    OrderNumberTrac,
    ERPComm.UniqueID AS SalesOrder,
    CodePack AS PackingListCode,
    Pallets.CodePallet,
    ERPPick.CommercialNumber,
    ERPPick.UniqueNumber,
    COUNT(*) AS Quantity,
    FORMAT(CAST(DeclaredAllDateERP AS DATE), 'd', 'it-it') AS PkLIstDate
FROM WarehouseFinish.dbo.PackingLists
INNER JOIN WarehouseFinish.dbo.Pallets ON Pallets.IDPackingList = PackingLists.IDPackingList
INNER JOIN WarehouseFinish.dbo.Packing ON Packing.IDPallet = Pallets.IDPallet
INNER JOIN WarehouseFinish.dbo.PackingDetail ON PackingDetail.PackingId = Packing.PackingId
INNER JOIN WarehouseFinish.ERP.ERPCommercialOrder ERPComm
        ON Pallets.IdERPCommercialOrder = ERPComm.IdERPCommercialOrder
INNER JOIN WarehouseFinish.dbo.Pick ON Pick.PackingId = Packing.PackingId
INNER JOIN WarehouseFinish.dbo.OrderDet ON Pick.OrderDetId = OrderDet.OrderDetId
INNER JOIN WarehouseFinish.ERP.ERPPickingList ERPPick
        ON ERPPick.IdERPPickingList = OrderDet.IdERPPickingList
WHERE Pick.IsPicked = 1
  {inner_filters}
GROUP BY CodePack, OrderNumberTrac, ERPComm.UniqueID, Pallets.CodePallet,
         Packing.IncomingDetId, ERPPick.CommercialNumber, ERPPick.UniqueNumber,
         FORMAT(CAST(DeclaredAllDateERP AS DATE), 'd', 'it-it')
"""

_Q_OUTER = """
SELECT A.PkLIstDate, A.PackingListCode, A.ProductCode,
       A.OrderNumberTrac AS OrderNumberProduction,
       SUM(A.Quantity) AS Qty, A.CodePallet, A.SalesOrder,
       A.CommercialNumber, A.UniqueNumber
FROM ({inner}) A
{outer_where}
GROUP BY A.ProductCode, A.OrderNumberTrac, A.SalesOrder, A.PackingListCode,
         A.CodePallet, A.CommercialNumber, A.UniqueNumber, A.PkLIstDate
ORDER BY A.SalesOrder, A.CommercialNumber, A.CodePallet
"""

# Filtri applicabili nella query interna (colonne vere delle tabelle)
_INNER_FILTER_MAP = {
    'pallet': "Pallets.CodePallet = ?",
    'commercial': "ERPPick.CommercialNumber = ?",
    'unique': "ERPPick.UniqueNumber = ?",
    'prod_order': "OrderNumberTrac = ?",
}

_COLUMNS = [
    ('pklist_date', 'Data PL', 90),
    ('packing_list', 'Packing List', 110),
    ('product_code', 'Product Code', 160),
    ('prod_order', 'Ord. Produzione', 110),
    ('qty', 'Qty', 70),
    ('pallet', 'Pallet', 90),
    ('sales_order', 'Sales Order', 110),
    ('commercial', 'Commercial Number', 120),
    ('unique_no', 'Unique Number', 110),
]

_HEADER_FILL = 'BDD7EE'  # azzurro pastello


def open_find_shipments(master, db, lang, user_name: str):
    return FindShipmentsWindow(master, db, lang, user_name)


class FindShipmentsWindow(tk.Toplevel):
    """Ricerca spedizioni su WarehouseFinish con export Excel."""

    def __init__(self, parent, db, lang, user_name: str):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self._results = []

        L = lambda k, d: self.lang.get(k, d)  # noqa: E731
        self._L = L

        self.title(L('find_ship_title', 'Trova spedizioni'))
        self.geometry('1100x600')
        self.minsize(900, 450)
        self.grab_set()

        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        L = self._L

        # ── Barra ricerca ─────────────────────────────────────────────────────
        frm = tk.LabelFrame(self, text=L('find_ship_search', 'Criteri di ricerca (tutti opzionali, corrispondenza esatta)'),
                            padx=8, pady=8)
        frm.pack(fill=tk.X, padx=10, pady=8)

        self._vars = {}
        fields = [
            ('product', L('find_ship_product', 'Product Code')),
            ('pallet', L('find_ship_pallet', 'Pallet')),
            ('commercial', L('find_ship_commercial', 'Commercial Number')),
            ('unique', L('find_ship_unique', 'Unique Number')),
            ('prod_order', L('find_ship_prod_order', 'Ord. Produzione')),
        ]
        for i, (key, label) in enumerate(fields):
            tk.Label(frm, text=f'{label}:').grid(row=i // 3, column=(i % 3) * 2,
                                                 sticky=tk.W, padx=4, pady=4)
            var = tk.StringVar()
            entry = tk.Entry(frm, textvariable=var, width=32, font=('Segoe UI', 10))
            entry.grid(row=i // 3, column=(i % 3) * 2 + 1, sticky=tk.W, padx=4, pady=4)
            entry.bind('<Return>', lambda e: self._search())
            self._vars[key] = var

        tk.Button(frm, text=L('search', 'Cerca'), width=14,
                  bg='#2e86de', fg='#ffffff', relief=tk.FLAT,
                  command=self._search).grid(row=0, column=6, rowspan=2,
                                             padx=12, pady=4, sticky=tk.NS)

        # ── Risultati ─────────────────────────────────────────────────────────
        tbl = tk.Frame(self)
        tbl.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        self._tree = ttk.Treeview(tbl, columns=[c[0] for c in _COLUMNS],
                                  show='headings')
        for key, label, width in _COLUMNS:
            self._tree.heading(key, text=label)
            self._tree.column(key, width=width, anchor=tk.CENTER if key == 'qty' else tk.W)
        vsb = ttk.Scrollbar(tbl, orient=tk.VERTICAL, command=self._tree.yview)
        hsb = ttk.Scrollbar(tbl, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._count_lbl = tk.Label(self, text='', anchor=tk.W,
                                   font=('Segoe UI', 9, 'italic'), fg='#7f8c8d')
        self._count_lbl.pack(fill=tk.X, padx=10)

        # ── Bottoni ───────────────────────────────────────────────────────────
        btn = tk.Frame(self, padx=10, pady=8)
        btn.pack(fill=tk.X)
        tk.Button(btn, text=L('find_ship_export', '📊 Esporta Excel'), width=18,
                  command=self._export_excel).pack(side=tk.LEFT, padx=4)
        tk.Button(btn, text=L('close', 'Chiudi'), width=12,
                  command=self.destroy).pack(side=tk.RIGHT, padx=4)

    # ── Ricerca ───────────────────────────────────────────────────────────────

    def _search(self):
        L = self._L
        values = {k: v.get().strip() for k, v in self._vars.items()}

        if not any(values.values()):
            messagebox.showwarning(
                L('warning', 'Attenzione'),
                L('find_ship_no_criteria', 'Inserire almeno un criterio di ricerca.'),
                parent=self)
            return

        inner_filters = []
        params = []
        for key, condition in _INNER_FILTER_MAP.items():
            if values[key]:
                inner_filters.append(condition)
                params.append(values[key])
        inner_where = ('AND ' + ' AND '.join(inner_filters)) if inner_filters else ''

        outer_where = ''
        if values['product']:
            outer_where = 'WHERE A.ProductCode = ?'
            params.append(values['product'])

        query = _Q_OUTER.format(inner=_Q_BASE_INNER.format(inner_filters=inner_where),
                                outer_where=outer_where)

        self._tree.delete(*self._tree.get_children())
        self._results = []
        try:
            cur = self.db.conn.cursor()
            cur.execute(query, tuple(params))
            cols = [d[0] for d in cur.description]
            for row in cur.fetchall():
                rec = dict(zip(cols, row))
                self._results.append(rec)
                self._tree.insert('', tk.END, values=(
                    rec['PkLIstDate'] or '',
                    rec['PackingListCode'] or '',
                    rec['ProductCode'] or '',
                    rec['OrderNumberProduction'] or '',
                    rec['Qty'] or 0,
                    rec['CodePallet'] or '',
                    rec['SalesOrder'] or '',
                    rec['CommercialNumber'] or '',
                    rec['UniqueNumber'] or '',
                ))
            cur.close()
        except Exception as e:
            logger.error(f"FindShipments _search: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)
            return

        self._count_lbl.config(text=L('find_ship_found', '{n} righe trovate').format(n=len(self._results)))

    # ── Export Excel ──────────────────────────────────────────────────────────

    def _export_excel(self):
        L = self._L
        if not self._results:
            messagebox.showinfo(L('info', 'Informazione'),
                                L('find_ship_nothing', 'Nessun risultato da esportare.'),
                                parent=self)
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(L('error', 'Errore'), 'openpyxl non disponibile.', parent=self)
            return

        try:
            from datetime import datetime

            temp_dir = tempfile.gettempdir()
            os.makedirs(temp_dir, exist_ok=True)
            file_name = f"TrovaSpedizioni_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(temp_dir, file_name)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Spedizioni'

            headers = [label for _, label, _ in _COLUMNS]
            ws.append(headers)

            # Intestazioni: azzurro pastello, grassetto, filtri automatici
            fill = PatternFill(start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type='solid')
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_idx, (_, _, width) in enumerate(_COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(width / 7, 12)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            ws.freeze_panes = 'A2'

            for rec in self._results:
                ws.append([
                    rec['PkLIstDate'] or '',
                    rec['PackingListCode'] or '',
                    rec['ProductCode'] or '',
                    rec['OrderNumberProduction'] or '',
                    rec['Qty'] or 0,
                    rec['CodePallet'] or '',
                    rec['SalesOrder'] or '',
                    rec['CommercialNumber'] or '',
                    rec['UniqueNumber'] or '',
                ])

            wb.save(file_path)
            logger.info(f"Trova spedizioni: esportati {len(self._results)} righe in {file_path}")

            # Apri il file con l'applicazione predefinita
            os.startfile(file_path)
        except Exception as e:
            logger.error(f"FindShipments _export_excel: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), str(e), parent=self)
            return
