"""
indirect_materials_import.py
Finestra per importare codici materiali indiretti da file Excel
nella tabella ind.Materiali con logica soft-delete.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
import os
import socket
import openpyxl
from datetime import datetime

import indirect_materials_stock_data as stock_data

logger = logging.getLogger(__name__)


class ImportIndirectMaterialsWindow(tk.Toplevel):
    """Finestra per importare materiali indiretti da Excel."""

    def __init__(self, master, db, lang, user_name="Unknown"):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name
        self.import_data = []

        self.title(lang.get('ind_import_title', 'Allinea Codici - Import Materiali Indiretti'))
        self.geometry("800x550")
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ------------------------------------------------------------------ #
    #  UI                                                                  #
    # ------------------------------------------------------------------ #
    def _build_ui(self):
        main = ttk.Frame(self, padding=15)
        main.pack(expand=True, fill="both")

        # Header
        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 10))

        ttk.Label(
            header,
            text=self.lang.get('ind_import_header', 'Import Materiali Indiretti da Excel'),
            font=("Segoe UI", 13, "bold")
        ).pack(side="left")

        # Bottoni
        btn_frame = ttk.Frame(header)
        btn_frame.pack(side="right")

        ttk.Button(
            btn_frame,
            text=self.lang.get('ind_import_select_excel', 'Seleziona Excel'),
            command=self._select_excel
        ).pack(side="left", padx=(0, 5))

        self.btn_import = ttk.Button(
            btn_frame,
            text=self.lang.get('ind_import_btn', 'Importa'),
            command=self._import_data,
            state="disabled"
        )
        self.btn_import.pack(side="left")

        # Info file
        self.file_var = tk.StringVar(value=self.lang.get('ind_import_no_file', 'Nessun file selezionato'))
        ttk.Label(main, textvariable=self.file_var, foreground="gray").pack(fill="x", pady=(0, 5))

        # Treeview
        tree_frame = ttk.Frame(main)
        tree_frame.pack(fill="both", expand=True)

        columns = ('codice', 'descrizione', 'qta_stock', 'tipo')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='browse')

        self.tree.heading('codice', text=self.lang.get('ind_import_col_code', 'Codice Materiale'))
        self.tree.heading('descrizione', text=self.lang.get('ind_import_col_desc', 'Descrizione'))
        self.tree.heading('qta_stock', text=self.lang.get('ind_import_col_qty', 'Qtà Stock'))
        self.tree.heading('tipo', text=self.lang.get('ind_req_col_type', 'Tipo'))

        self.tree.column('codice', width=150)
        self.tree.column('descrizione', width=350)
        self.tree.column('qta_stock', width=100, anchor="e")
        self.tree.column('tipo', width=120)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_y.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")

        # Progress bar (nascosta inizialmente)
        self.progress = ttk.Progressbar(main, orient='horizontal', mode='determinate')
        self.progress.pack(fill="x", pady=(10, 0))
        self.progress.pack_forget()  # nascondi finché non serve

        # Status bar
        self.status_var = tk.StringVar(value="")
        ttk.Label(main, textvariable=self.status_var, relief="sunken", anchor="w").pack(fill="x", pady=(5, 0))

    # ------------------------------------------------------------------ #
    #  Selezione e lettura Excel                                           #
    # ------------------------------------------------------------------ #
    def _select_excel(self):
        file_path = filedialog.askopenfilename(
            title=self.lang.get('ind_import_select_excel', 'Seleziona file Excel'),
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')],
            parent=self
        )
        if not file_path:
            return

        self.file_path = file_path
        self.file_name = os.path.basename(file_path)

        try:
            self.import_data = []
            self.tree.delete(*self.tree.get_children())

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # --- Auto-detect colonne dall'intestazione (riga 1) ---
            # Mapping nome intestazione → chiave colonna
            HEADER_MAP = {
                'item number': 'code',
                'product name': 'desc',
                'name': 'desc',
                'description': 'desc',
                'physical inventory': 'stock',
                'physical on-hand inventory': 'stock',
                'warehouse': 'tipo',
            }
            # Default fisso (fallback se auto-detect non trova)
            COL_CODE = 0    # Colonna A
            COL_DESC = 1    # Colonna B
            COL_STOCK = 9   # Colonna J (Physical inventory)
            COL_TIPO = 16   # Colonna Q

            # Leggi intestazioni e tenta auto-detect
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if header_row:
                headers = header_row[0]
                for idx, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).strip().lower()
                    if h_lower in HEADER_MAP:
                        key = HEADER_MAP[h_lower]
                        if key == 'code':
                            COL_CODE = idx
                        elif key == 'desc':
                            COL_DESC = idx
                        elif key == 'stock':
                            COL_STOCK = idx
                        elif key == 'tipo':
                            COL_TIPO = idx
                logger.info(
                    f"Colonne rilevate: Code=[{chr(65+COL_CODE)}]({COL_CODE}), "
                    f"Desc=[{chr(65+COL_DESC)}]({COL_DESC}), "
                    f"Stock=[{chr(65+COL_STOCK)}]({COL_STOCK}), "
                    f"Tipo=[{chr(65+COL_TIPO)}]({COL_TIPO})"
                )
            else:
                logger.warning("Nessuna intestazione trovata, uso layout fisso di fallback")

            # Pre-carica i tipi materiale dal DB per lookup
            tipo_lookup = self._load_tipo_materiali_lookup()

            # Leggi righe (dalla riga 2, salta intestazione)
            first_row_logged = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                codice = row[COL_CODE] if len(row) > COL_CODE else None
                if not codice:
                    continue

                codice = str(codice).strip()
                descrizione = str(row[COL_DESC]).strip() if len(row) > COL_DESC and row[COL_DESC] else ''
                raw_stock = row[COL_STOCK] if len(row) > COL_STOCK else 0
                qta = self._safe_decimal(raw_stock)
                tipo_raw = str(row[COL_TIPO]).strip() if len(row) > COL_TIPO and row[COL_TIPO] else ''

                # Log diagnostico sulla prima riga per verifica colonne
                if not first_row_logged:
                    logger.info(
                        f"Prima riga Excel: codice={codice!r}, desc={descrizione[:30]!r}, "
                        f"raw_stock={raw_stock!r} (type={type(raw_stock).__name__}), "
                        f"qta_parsed={qta}, tipo={tipo_raw!r}"
                    )
                    first_row_logged = True

                # Lookup del tipo: se trovato usa l'ID, altrimenti 'Generico'
                tipo_id = tipo_lookup.get(tipo_raw.upper()) if tipo_raw else None
                tipo_display = tipo_raw if tipo_raw else 'Generico'
                if tipo_raw and tipo_id is None:
                    tipo_display = f"{tipo_raw} (→Generico)"

                self.import_data.append({
                    'codice': codice,
                    'descrizione': descrizione,
                    'qta_stock': qta,
                    'tipo_raw': tipo_raw,
                    'tipo_id': tipo_id
                })

                self.tree.insert('', 'end', values=(codice, descrizione, f"{qta:.2f}", tipo_display))

            self.file_var.set(f"📂 {file_path}")
            self.status_var.set(
                f"{len(self.import_data)} {self.lang.get('ind_import_rows_found', 'righe trovate')}"
            )

            if self.import_data:
                self.btn_import.state(["!disabled"])
            else:
                self.btn_import.state(["disabled"])
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('ind_import_no_data', 'Nessun dato valido trovato nel file Excel.'),
                    parent=self
                )

        except Exception as e:
            logger.error(f"Errore lettura Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"{self.lang.get('ind_import_read_error', 'Errore lettura file Excel')}:\n{e}",
                parent=self
            )

    # ------------------------------------------------------------------ #
    #  Import in database                                                  #
    # ------------------------------------------------------------------ #
    def _import_data(self):
        if not self.import_data:
            return

        # Conferma
        confirm_msg = self.lang.get(
            'ind_import_confirm',
            'Importare {0} codici materiali?\n\n'
            'I nuovi codici verranno aggiunti all\'anagrafica.\n'
            'Le giacenze attuali verranno storicizzate\n'
            'e sostituite con i nuovi dati.'
        ).format(len(self.import_data))

        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            confirm_msg,
            parent=self
        ):
            return

        self.status_var.set(self.lang.get('ind_import_importing', 'Importazione in corso...'))
        self.progress.pack(fill="x", pady=(10, 0))  # mostra la barra
        self.progress['value'] = 0
        total_items = len(self.import_data)
        total_steps = total_items * 2  # fase 1: upsert + fase 2: stock
        self.progress['maximum'] = total_steps
        self.btn_import.state(["disabled"])
        self.update()

        try:
            new_codes = 0
            updated_codes = 0
            stock_inserted = 0
            errors = 0
            processed = 0

            # ── Transazione atomica ──────────────────────────────────────
            self.db._ensure_connection()
            with self.db._lock:
                cursor = self.db.cursor

                try:
                    # 0. Recupera ID del tipo 'Generico' (default)
                    cursor.execute(
                        "SELECT TipoMaterialeId FROM ind.TipoMateriali WHERE UPPER(Tipo) = 'GENERICO'"
                    )
                    generico_row = cursor.fetchone()
                    if generico_row:
                        generico_tipo_id = generico_row[0]
                    else:
                        cursor.execute(
                            "INSERT INTO ind.TipoMateriali (Tipo, IsFrazionabile, QtaConfezione) VALUES ('Generico', 0, 1)"
                        )
                        cursor.execute("SELECT SCOPE_IDENTITY() as id")
                        generico_tipo_id = cursor.fetchone()[0]
                        logger.info(f"Tipo default 'Generico' creato con ID {generico_tipo_id}")

                    # 1. Upsert anagrafica ind.Materiali
                    for item in self.import_data:
                        try:
                            tipo_id = item.get('tipo_id') or generico_tipo_id

                            cursor.execute(
                                "SELECT MaterialeId, DescrizioneMateriale FROM ind.Materiali WHERE CodiceMateriale = ?",
                                (item['codice'],)
                            )
                            existing = cursor.fetchone()

                            if existing:
                                materiale_id = existing[0]
                                cursor.execute(
                                    "UPDATE ind.Materiali SET DescrizioneMateriale = ?, TipoMaterialeId = ?, IsActive = 1 WHERE MaterialeId = ?",
                                    (item['descrizione'], tipo_id, materiale_id)
                                )
                                updated_codes += 1
                            else:
                                cursor.execute(
                                    "INSERT INTO ind.Materiali (CodiceMateriale, DescrizioneMateriale, TipoMaterialeId, IsActive) "
                                    "VALUES (?, ?, ?, 1)",
                                    (item['codice'], item['descrizione'], tipo_id)
                                )
                                cursor.execute("SELECT SCOPE_IDENTITY() as id")
                                materiale_id = cursor.fetchone()[0]
                                new_codes += 1

                            item['materiale_id'] = materiale_id

                        except Exception as e:
                            errors += 1
                            logger.error(f"Errore upsert codice {item['codice']}: {e}")
                            item['materiale_id'] = None

                        processed += 1
                        self.progress['value'] = processed
                        self.status_var.set(
                            f"Anagrafica: {processed}/{total_items} — "
                            f"{new_codes} nuovi, {updated_codes} aggiornati"
                        )
                        self.update_idletasks()

                    # 2. Soft-close giacenze attive SOLO per i materiali importati
                    valid_ids = [item['materiale_id'] for item in self.import_data if item.get('materiale_id') is not None]
                    if valid_ids:
                        placeholders = ','.join('?' * len(valid_ids))
                        cursor.execute(
                            f"UPDATE ind.MaterialiStock SET DateOut = GETDATE() "
                            f"WHERE DateOut IS NULL AND MaterialeId IN ({placeholders})",
                            valid_ids
                        )
                        logger.info(f"Soft-close giacenze attive per {len(valid_ids)} materiali importati")

                    # Giacenza corrente PRIMA dell'import (dal libro movimenti) per calcolare il delta.
                    # NB: la giacenza visualizzata = SUM(ind.MaterialiMovimenti.Qty) via ind.vw_GiacenzaCorrente;
                    # ind.MaterialiStock NON entra nel calcolo, quindi va aggiornato anche il libro movimenti.
                    giac_map = {}
                    if valid_ids:
                        ph = ','.join('?' * len(valid_ids))
                        cursor.execute(
                            f"SELECT MaterialeId, Giacenza FROM ind.vw_GiacenzaCorrente WHERE MaterialeId IN ({ph})",
                            valid_ids
                        )
                        giac_map = {r[0]: float(r[1] or 0) for r in cursor.fetchall()}
                    hostname = socket.gethostname()

                    # 3. Insert nuove giacenze in ind.MaterialiStock + allineamento libro movimenti
                    for item in self.import_data:
                        if item.get('materiale_id') is None:
                            processed += 1
                            self.progress['value'] = processed
                            self.update_idletasks()
                            continue
                        try:
                            cursor.execute(
                                "INSERT INTO ind.MaterialiStock (MaterialeId, Qty, DateIn, DateOut, CaricatoDa) "
                                "VALUES (?, ?, GETDATE(), NULL, ?)",
                                (item['materiale_id'], item['qta_stock'], self.user_name)
                            )
                            # Allinea la giacenza (libro movimenti) al valore importato:
                            # movimento INVENTARIO con il delta necessario a raggiungere la quantita' importata.
                            target = float(item['qta_stock'] or 0)
                            current = giac_map.get(item['materiale_id'], 0.0)
                            delta = round(target - current, 4)
                            if delta != 0:
                                cursor.execute(
                                    "INSERT INTO ind.MaterialiMovimenti "
                                    "(MaterialeId, Qty, TipoMovimento, EseguitoDa, ComputerSrc, Note) "
                                    "VALUES (?, ?, 'INVENTARIO', ?, ?, ?)",
                                    (item['materiale_id'], delta, self.user_name, hostname,
                                     'Allineamento giacenza da import Excel')
                                )
                            stock_inserted += 1
                        except Exception as e:
                            errors += 1
                            logger.error(f"Errore insert stock per {item['codice']}: {e}")

                        processed += 1
                        self.progress['value'] = processed
                        self.status_var.set(
                            f"Giacenze: {processed - total_items}/{total_items} — "
                            f"{stock_inserted} caricate"
                        )
                        self.update_idletasks()

                    # 4. Azzera giacenze per materiali attivi presenti in DB ma non nel file Excel
                    imported_codes = {
                        str(item.get('codice') or '').strip().upper()
                        for item in self.import_data
                    }
                    zeroed_items = self._zero_missing_materials(cursor, imported_codes, hostname)

                    # ── COMMIT unico a fine transazione ──────────────────
                    self.db.conn.commit()
                    logger.info("Transazione import materiali committata con successo")

                except Exception as e:
                    self.db.conn.rollback()
                    logger.error(f"Rollback transazione import: {e}", exc_info=True)
                    raise
            # ── Fine transazione atomica ─────────────────────────────────

            # Invia notifica email per i materiali azzerati (dopo commit)
            if zeroed_items:
                self._send_missing_materials_email(zeroed_items)

            # Risultato
            msg = self.lang.get('ind_import_completed', 'Importazione completata') + ":\n\n"
            msg += f"🆕 {self.lang.get('ind_import_new_codes', 'Nuovi codici')}: {new_codes}\n"
            msg += f"🔄 {self.lang.get('ind_import_updated_codes', 'Codici aggiornati')}: {updated_codes}\n"
            msg += f"📦 {self.lang.get('ind_import_stock_loaded', 'Giacenze caricate')}: {stock_inserted}\n"
            if zeroed_items:
                msg += f"🅾️  {self.lang.get('ind_import_zeroed', 'Stock azzerati')}: {len(zeroed_items)}\n"
            if errors > 0:
                msg += f"❌ {self.lang.get('ind_import_errors', 'Errori')}: {errors}"

            messagebox.showinfo(
                self.lang.get('ind_import_result', 'Risultato Import'),
                msg,
                parent=self
            )

            self.progress['value'] = total_steps  # completa al 100%
            status_parts = [
                f"✅ {new_codes} nuovi, {updated_codes} aggiornati, "
                f"{stock_inserted} giacenze"
            ]
            if zeroed_items:
                status_parts.append(f"{len(zeroed_items)} azzerati")
            if errors > 0:
                status_parts.append(f"{errors} errori")
            self.status_var.set(", ".join(status_parts))
            self.btn_import.state(["disabled"])

            logger.info(
                f"Import completato: {new_codes} nuovi, {updated_codes} aggiornati, "
                f"{stock_inserted} stock, {len(zeroed_items or [])} azzerati, {errors} errori"
            )

        except Exception as e:
            logger.error(f"Errore import materiali indiretti: {e}", exc_info=True)
            err_msg = self.lang.get('ind_import_error', "Errore durante l'importazione")
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"{err_msg}:\n{e}",
                parent=self
            )

    # ------------------------------------------------------------------ #
    #  Azzeramento stock per codici assenti dall'Excel                     #
    # ------------------------------------------------------------------ #
    def _get_missing_active_materials(self, cursor):
        """Ritorna i materiali attivi del DB non presenti nel file Excel appena caricato.

        Ogni elemento e' un dict con: materiale_id, codice, descrizione, giacenza.
        """
        try:
            cursor.execute(
                """
                SELECT m.MaterialeId, m.CodiceMateriale, m.DescrizioneMateriale,
                       ISNULL(g.Giacenza, 0) AS Giacenza
                FROM ind.Materiali m
                LEFT JOIN ind.vw_GiacenzaCorrente g ON g.MaterialeId = m.MaterialeId
                WHERE m.IsActive = 1
                ORDER BY m.CodiceMateriale
                """
            )
            rows = cursor.fetchall()
        except Exception as e:
            logger.error(f"Errore lettura materiali attivi per azzeramento: {e}", exc_info=True)
            return []

        imported_codes = {
            str(item.get('codice') or '').strip().upper()
            for item in self.import_data
        }
        missing = []
        for row in rows:
            code = str(row[1] or '').strip().upper()
            if code in imported_codes:
                continue
            giacenza = float(row[3] or 0)
            if giacenza == 0:
                continue
            missing.append({
                'materiale_id': row[0],
                'codice': row[1] or '',
                'descrizione': row[2] or '',
                'giacenza': giacenza,
            })
        return missing

    def _zero_missing_materials(self, cursor, imported_codes, hostname):
        """Porta a zero le giacenze dei materiali attivi non presenti nel file Excel.

        Ritorna la lista dei materiali effettivamente azzerati.
        """
        missing = self._get_missing_active_materials(cursor)
        if not missing:
            return []

        zeroed = []
        file_name = getattr(self, 'file_name', 'D365 export')
        note = f"Azzeramento stock: codice assente nel file Excel {file_name}"

        for it in missing:
            current = it['giacenza']
            try:
                # Movimento INVENTARIO che porta la giacenza a zero
                cursor.execute(
                    """INSERT INTO ind.MaterialiMovimenti
                       (MaterialeId, Qty, TipoMovimento, EseguitoDa, ComputerSrc, Note)
                       VALUES (?, ?, 'INVENTARIO', ?, ?, ?)""",
                    (it['materiale_id'], round(-current, 4), self.user_name, hostname, note)
                )
                # Soft-close eventuale riga di stock aperta
                cursor.execute(
                    """UPDATE ind.MaterialiStock
                       SET DateOut = GETDATE()
                       WHERE DateOut IS NULL AND MaterialeId = ?""",
                    (it['materiale_id'],)
                )
                zeroed.append(it)
                logger.info(
                    f"Stock azzerato per {it['codice']} (giacenza precedente {current})"
                )
            except Exception as e:
                logger.error(f"Errore azzeramento stock per {it['codice']}: {e}", exc_info=True)

        return zeroed

    def _build_missing_materials_email(self, zeroed_items):
        """Costruisce (subject, html_body) per l'email di notifica azzeramento stock."""
        file_name = getattr(self, 'file_name', 'D365 export') or 'D365 export'
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')

        subject = "Indirect Materials Stock Reset Notification"

        rows_html = []
        for it in zeroed_items:
            rows_html.append(
                "<tr>"
                f"<td style='border:1px solid #ccc;padding:6px;'>{it['codice']}</td>"
                f"<td style='border:1px solid #ccc;padding:6px;'>{it['descrizione']}</td>"
                f"<td style='border:1px solid #ccc;padding:6px;text-align:right;'>"
                f"{it['giacenza']:.2f}</td>"
                "</tr>"
            )

        body = (
            "<p>Dear Purchasing Team,</p>"
            "<p>Following the import of the D365 export file "
            f"<strong>{file_name}</strong> on {timestamp}, the stock levels for the "
            "materials listed below have been reset to <strong>zero</strong>.</p>"
            "<p>These codes were not included in the downloaded file; therefore, "
            "their quantities have been cleared to ensure the system reflects the "
            "current inventory accurately.</p>"
            "<table style='border-collapse:collapse;font-family:Segoe UI,Arial;font-size:13px;'>"
            "<thead><tr style='background:#f0f0f0;'>"
            "<th style='border:1px solid #ccc;padding:6px;'>Material Code</th>"
            "<th style='border:1px solid #ccc;padding:6px;'>Description</th>"
            "<th style='border:1px solid #ccc;padding:6px;'>Previous Stock</th>"
            "</tr></thead><tbody>"
            + ''.join(rows_html) +
            "</tbody></table>"
            "<p style='color:#888;font-size:11px;margin-top:16px;'>"
            "This is an automatic notification sent by the Document Management system.</p>"
        )
        return subject, body

    def _send_missing_materials_email(self, zeroed_items):
        """Invia l'email di notifica azzeramento agli stessi destinatari del riordino."""
        try:
            recipients = stock_data._get_reorder_recipients(self.db)
            if not recipients:
                logger.warning(
                    "Nessun destinatario configurato per l'email di azzeramento stock"
                )
                return False

            subject, body = self._build_missing_materials_email(zeroed_items)
            try:
                from email_connector import EmailSender
                sender = EmailSender()
            except Exception as e:
                logger.error(f"EmailSender non inizializzato: {e}", exc_info=True)
                return False
            sender.send_email(
                to_email='; '.join(recipients),
                subject=subject,
                body=body,
                is_html=True
            )
            logger.info(
                f"Email azzeramento stock inviata a {len(recipients)} destinatari "
                f"per {len(zeroed_items)} codici"
            )
            return True
        except Exception as e:
            logger.error(f"Errore invio email azzeramento stock: {e}", exc_info=True)
            return False

    # ------------------------------------------------------------------ #
    #  Utility                                                             #
    # ------------------------------------------------------------------ #
    @staticmethod
    def _safe_decimal(value, default=0.0):
        """Converte un valore in float in modo sicuro."""
        if value is None or value == '':
            return default
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', '.'))
        except (ValueError, AttributeError):
            return default

    def _load_tipo_materiali_lookup(self):
        """Carica un dizionario {TIPO_UPPER: TipoMaterialeId} per lookup veloce."""
        lookup = {}
        try:
            self.db._ensure_connection()
            with self.db._lock:
                self.db.cursor.execute(
                    "SELECT TipoMaterialeId, Tipo FROM ind.TipoMateriali"
                )
                rows = self.db.cursor.fetchall()
            for row in (rows or []):
                if row[1]:
                    lookup[row[1].strip().upper()] = row[0]
            logger.info(f"Lookup tipi materiale caricato: {len(lookup)} tipi")
        except Exception as e:
            logger.error(f"Errore caricamento lookup tipi: {e}", exc_info=True)
        return lookup


def open_indirect_materials_import(master, db, lang, user_name="Unknown"):
    """Entry-point richiamabile da main.py."""
    ImportIndirectMaterialsWindow(master, db, lang, user_name)
