# -*- coding: utf-8 -*-
"""
indirect_materials_consumption_report.py
Report generale consumi materiali (di consumo) richiesti in un periodo A..B.

Considera TUTTI i materiali di consumo richiesti (ind.MaterialiRichieste) fra due
date, con la quantita' richiesta e, dove esiste, il materiale restituito quale
scrap (dbo.ReturnMaterials collegato tramite RichiestaId).

Filtri:
  - Periodo:   Data da (A) .. Data a (B)  (B inclusa)
  - Codice/Descrizione materiale (testo libero, LIKE su entrambi)
  - Gruppo materiali (ind.FamigliaMateriali)
  - Richiedente (RichiestoDa)
  - Scrap: tutti / solo con reso scrap / solo senza reso scrap

A monitor si mostrano SOLO valori sintetici (KPI + aggregato per codice).
L'export Excel contiene i dati analitici:
  - foglio "Sintesi"   : KPI + aggregato per codice
  - foglio "Analitico" : ogni singola richiesta con tutti i campi
  - fino a 10 fogli "T## <codice>" : i top 10 codici (per quantita' richiesta),
    con il dettaglio analitico delle richieste di quel codice.

Rapporto senza login.
"""

import os
import logging
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta

try:
    from tkcalendar import DateEntry
except Exception:
    DateEntry = None

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  Finestra principale
# ─────────────────────────────────────────────────────────────────────────────
class ConsumptionGeneralReportWindow(tk.Toplevel):

    # Indici colonne della riga di dettaglio (vedi _detail_query)
    C_RID, C_DATA, C_MATID, C_COD, C_DESC, C_FAM, C_TIPO, C_QTA, \
        C_STATO, C_RICH, C_PREP, C_DPREP, C_DPREL, C_SCRAP, C_SCRAPCONF, C_SCRAPN = range(16)

    def __init__(self, master, db, lang):
        super().__init__(master)
        self.db = db
        self.lang = lang
        L = self.lang.get

        self.title(L('cgr_title', 'Report Consumi Materiali - Generale'))
        self.geometry('1180x720')
        self.minsize(920, 560)
        self.resizable(True, True)
        self.transient(master)

        self._detail_rows = []   # righe analitiche (tuple)
        self._agg = []           # aggregato per materiale (dict)

        self._build_ui()
        self._populate_filters()
        self._load_data()
        self.protocol('WM_DELETE_WINDOW', self.destroy)

    # ------------------------------------------------------------------ #
    #  UI                                                                  #
    # ------------------------------------------------------------------ #
    def _build_ui(self):
        L = self.lang.get

        # ── Toolbar filtri ──────────────────────────────────────────────
        tb = ttk.LabelFrame(self, text=L('cgr_filters', 'Filtri'), padding=8)
        tb.pack(fill='x', padx=10, pady=(10, 0))

        # Periodo A
        ttk.Label(tb, text=L('cgr_date_from', 'Data da:')).grid(row=0, column=0, padx=(0, 4), pady=2, sticky='w')
        first_of_month = datetime.now().replace(day=1)
        if DateEntry:
            self.date_from = DateEntry(tb, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
            self.date_from.set_date(first_of_month)
        else:
            self.date_from = ttk.Entry(tb, width=12)
            self.date_from.insert(0, first_of_month.strftime('%d/%m/%Y'))
        self.date_from.grid(row=0, column=1, padx=(0, 12), pady=2)

        # Periodo B
        ttk.Label(tb, text=L('cgr_date_to', 'Data a:')).grid(row=0, column=2, padx=(0, 4), pady=2, sticky='w')
        if DateEntry:
            self.date_to = DateEntry(tb, width=12, date_pattern='dd/mm/yyyy', locale='it_IT')
            self.date_to.set_date(datetime.now())
        else:
            self.date_to = ttk.Entry(tb, width=12)
            self.date_to.insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.date_to.grid(row=0, column=3, padx=(0, 12), pady=2)

        # Codice/Descrizione
        ttk.Label(tb, text=L('cgr_text', 'Codice/Descrizione:')).grid(row=0, column=4, padx=(0, 4), pady=2, sticky='w')
        self.text_var = tk.StringVar()
        ent = ttk.Entry(tb, textvariable=self.text_var, width=26)
        ent.grid(row=0, column=5, padx=(0, 12), pady=2)
        ent.bind('<Return>', lambda e: self._load_data())

        # Gruppo materiali (Famiglia)
        ttk.Label(tb, text=L('cgr_group', 'Gruppo:')).grid(row=1, column=0, padx=(0, 4), pady=2, sticky='w')
        self.group_var = tk.StringVar()
        self.group_combo = ttk.Combobox(tb, textvariable=self.group_var, width=22, state='readonly')
        self.group_combo.grid(row=1, column=1, columnspan=2, padx=(0, 12), pady=2, sticky='w')

        # Richiedente
        ttk.Label(tb, text=L('cgr_requester', 'Richiedente:')).grid(row=1, column=3, padx=(0, 4), pady=2, sticky='w')
        self.req_var = tk.StringVar()
        self.req_combo = ttk.Combobox(tb, textvariable=self.req_var, width=24, state='readonly')
        self.req_combo.grid(row=1, column=4, padx=(0, 12), pady=2, sticky='w')

        # Scrap
        ttk.Label(tb, text=L('cgr_scrap', 'Scrap:')).grid(row=1, column=5, padx=(0, 4), pady=2, sticky='w')
        self.scrap_var = tk.StringVar()
        self.scrap_combo = ttk.Combobox(tb, textvariable=self.scrap_var, width=20, state='readonly',
                                        values=[L('cgr_scrap_all', 'Tutti'),
                                                L('cgr_scrap_only', 'Solo con reso scrap'),
                                                L('cgr_scrap_none', 'Solo senza reso scrap')])
        self.scrap_combo.current(0)
        self.scrap_combo.grid(row=1, column=6, padx=(0, 12), pady=2, sticky='w')

        # Pulsanti
        ttk.Button(tb, text=L('btn_refresh', '🔄 Aggiorna'),
                   command=self._load_data).grid(row=0, column=6, padx=6, pady=2)
        ttk.Button(tb, text=L('cgr_export', '📊 Export Excel analitico'),
                   command=self._export_excel).grid(row=0, column=7, rowspan=2, padx=6, pady=2)

        # ── KPI sintetici ───────────────────────────────────────────────
        kpi = ttk.LabelFrame(self, text=L('cgr_summary', 'Sintesi'), padding=8)
        kpi.pack(fill='x', padx=10, pady=(8, 0))
        self._kpi_vars = {k: tk.StringVar(value='0') for k in
                          ('period', 'n_req', 'n_mat', 'qty', 'scrap', 'n_scrap')}
        kpi_defs = [('period',  L('cgr_kpi_period', 'Periodo')),
                    ('n_req',   L('cgr_kpi_nreq', '# Richieste')),
                    ('n_mat',   L('cgr_kpi_nmat', '# Materiali')),
                    ('qty',     L('cgr_kpi_qty', 'Qta richiesta')),
                    ('scrap',   L('cgr_kpi_scrap', 'Scrap reso (kg)')),
                    ('n_scrap', L('cgr_kpi_nscrap', '# Rich. con scrap'))]
        for i, (k, lbl) in enumerate(kpi_defs):
            cell = ttk.Frame(kpi)
            cell.grid(row=0, column=i, padx=16, pady=2, sticky='w')
            tk.Label(cell, textvariable=self._kpi_vars[k], font=('Segoe UI', 15, 'bold'),
                     fg='#1F3864').pack(anchor='w')
            tk.Label(cell, text=lbl, font=('Segoe UI', 8), fg='#555').pack(anchor='w')

        # ── Tabella sintetica per codice ─────────────────────────────────
        body = ttk.LabelFrame(self, text=L('cgr_by_code', 'Aggregato per codice'), padding=6)
        body.pack(fill='both', expand=True, padx=10, pady=10)

        cols = ('cod', 'desc', 'gruppo', 'tipo', 'nreq', 'qta', 'scrap', 'nscrap')
        self.tree = ttk.Treeview(body, columns=cols, show='headings')
        headers = {
            'cod':    L('cgr_col_code', 'Codice'),
            'desc':   L('cgr_col_desc', 'Descrizione'),
            'gruppo': L('cgr_col_group', 'Gruppo'),
            'tipo':   L('cgr_col_type', 'Tipo'),
            'nreq':   L('cgr_col_nreq', '# Richieste'),
            'qta':    L('cgr_col_qty', 'Qta richiesta'),
            'scrap':  L('cgr_col_scrap', 'Scrap (kg)'),
            'nscrap': L('cgr_col_nscrap', '# con scrap'),
        }
        widths = {'cod': 110, 'desc': 340, 'gruppo': 120, 'tipo': 100,
                  'nreq': 90, 'qta': 110, 'scrap': 100, 'nscrap': 90}
        numeric = ('nreq', 'qta', 'scrap', 'nscrap')
        for c in cols:
            self.tree.heading(c, text=headers[c],
                              command=lambda cc=c: self._sort_by(cc))
            self.tree.column(c, width=widths[c], anchor='e' if c in numeric else 'w')

        sbv = ttk.Scrollbar(body, orient='vertical', command=self.tree.yview)
        sbh = ttk.Scrollbar(body, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=sbv.set, xscrollcommand=sbh.set)
        sbv.pack(side='right', fill='y')
        sbh.pack(side='bottom', fill='x')
        self.tree.pack(fill='both', expand=True)
        self.tree.tag_configure('has_scrap', background='#fff3cd')

        self._sort_state = {}

    # ------------------------------------------------------------------ #
    #  Helpers dati                                                        #
    # ------------------------------------------------------------------ #
    def _fetch(self, query, params=None):
        try:
            if hasattr(self.db, 'fetch_all'):
                return self.db.fetch_all(query, params) or []
            self.db._ensure_connection()
            with self.db._lock:
                self.db.cursor.execute(query, params or [])
                return self.db.cursor.fetchall() or []
        except Exception as e:
            logger.error(f"[CGR] Query error: {e}", exc_info=True)
            return []

    def _get_date(self, widget):
        if DateEntry and hasattr(widget, 'get_date'):
            return widget.get_date()
        try:
            return datetime.strptime(widget.get().strip(), '%d/%m/%Y').date()
        except Exception:
            return datetime.now().date()

    # ------------------------------------------------------------------ #
    #  Filtri                                                              #
    # ------------------------------------------------------------------ #
    def _populate_filters(self):
        L = self.lang.get

        # Gruppi materiali
        rows_g = self._fetch("SELECT Famiglia FROM ind.FamigliaMateriali ORDER BY Famiglia")
        all_groups = L('cgr_all_groups', 'Tutti i gruppi')
        self.group_combo['values'] = [all_groups] + [r[0] for r in rows_g if r[0]]
        self.group_combo.current(0)

        # Richiedenti
        rows_r = self._fetch("""
            SELECT DISTINCT RichiestoDa FROM ind.MaterialiRichieste
            WHERE RichiestoDa IS NOT NULL AND RichiestoDa <> '' ORDER BY RichiestoDa
        """)
        all_req = L('cgr_all_requesters', 'Tutti i richiedenti')
        self.req_combo['values'] = [all_req] + [r[0] for r in rows_r]
        self.req_combo.current(0)

    def _build_where(self):
        """WHERE + parametri in base ai filtri UI. Il collegamento agli scrap e'
        via dbo.ReturnMaterials.RichiestaId (DateOut IS NULL = riga viva)."""
        L = self.lang.get
        where = ['r.DataRichiesta >= ?', 'r.DataRichiesta < ?']
        d_from = self._get_date(self.date_from)
        d_to = self._get_date(self.date_to)
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        params = [datetime(d_from.year, d_from.month, d_from.day),
                  datetime(d_to.year, d_to.month, d_to.day) + timedelta(days=1)]

        # Codice o descrizione
        txt = self.text_var.get().strip()
        if txt:
            where.append('(m.CodiceMateriale LIKE ? OR m.DescrizioneMateriale LIKE ?)')
            params.extend([f'%{txt}%', f'%{txt}%'])

        # Gruppo
        grp = self.group_var.get()
        if grp and grp != L('cgr_all_groups', 'Tutti i gruppi'):
            where.append('fm.Famiglia = ?')
            params.append(grp)

        # Richiedente
        req = self.req_var.get()
        if req and req != L('cgr_all_requesters', 'Tutti i richiedenti'):
            where.append('r.RichiestoDa = ?')
            params.append(req)

        # Scrap
        scr = self.scrap_var.get()
        if scr == L('cgr_scrap_only', 'Solo con reso scrap'):
            where.append('EXISTS (SELECT 1 FROM dbo.ReturnMaterials x '
                         'WHERE x.RichiestaId = r.RichiestaId AND x.DateOut IS NULL)')
        elif scr == L('cgr_scrap_none', 'Solo senza reso scrap'):
            where.append('NOT EXISTS (SELECT 1 FROM dbo.ReturnMaterials x '
                         'WHERE x.RichiestaId = r.RichiestaId AND x.DateOut IS NULL)')

        return ' AND '.join(where), params

    def _detail_query(self, where):
        return f"""
            SELECT r.RichiestaId,
                   r.DataRichiesta,
                   m.MaterialeId,
                   m.CodiceMateriale,
                   m.DescrizioneMateriale,
                   ISNULL(fm.Famiglia, '—') AS Famiglia,
                   ISNULL(t.Tipo, 'Generico') AS Tipo,
                   r.QtaRichiesta,
                   r.Stato,
                   r.RichiestoDa,
                   ISNULL(r.PreparatoDa, '') AS PreparatoDa,
                   r.DataPreparazione,
                   r.DataPrelievo,
                   ISNULL(rm.ScrapKg, 0)      AS ScrapKg,
                   ISNULL(rm.ScrapConfKg, 0)  AS ScrapConfKg,
                   ISNULL(rm.ScrapCount, 0)   AS ScrapCount
            FROM ind.MaterialiRichieste r
            JOIN ind.Materiali m ON m.MaterialeId = r.MaterialeId
            LEFT JOIN ind.FamigliaMateriali fm ON fm.FamigliaMaterialiId = m.FamigliaMaterialiId
            LEFT JOIN ind.TipoMateriali t ON t.TipoMaterialeId = m.TipoMaterialeId
            OUTER APPLY (
                SELECT SUM(x.ReturWeight)     AS ScrapKg,
                       SUM(x.ConfirmedWeight) AS ScrapConfKg,
                       COUNT(*)               AS ScrapCount
                FROM dbo.ReturnMaterials x
                WHERE x.RichiestaId = r.RichiestaId AND x.DateOut IS NULL
            ) rm
            WHERE {where}
            ORDER BY r.DataRichiesta DESC
        """

    # ------------------------------------------------------------------ #
    #  Caricamento e aggregazione                                          #
    # ------------------------------------------------------------------ #
    def _load_data(self):
        where, params = self._build_where()
        self._detail_rows = self._fetch(self._detail_query(where), params or None)

        # Aggregazione per materiale (codice)
        agg = {}
        for row in self._detail_rows:
            mid = row[self.C_MATID]
            a = agg.get(mid)
            if a is None:
                a = {'cod': row[self.C_COD] or '', 'desc': row[self.C_DESC] or '',
                     'gruppo': row[self.C_FAM] or '', 'tipo': row[self.C_TIPO] or '',
                     'nreq': 0, 'qta': 0.0, 'scrap': 0.0, 'nscrap': 0}
                agg[mid] = a
            a['nreq'] += 1
            a['qta'] += float(row[self.C_QTA] or 0)
            a['scrap'] += float(row[self.C_SCRAP] or 0)
            if (row[self.C_SCRAPN] or 0) > 0:
                a['nscrap'] += 1

        self._agg = sorted(agg.values(), key=lambda x: x['qta'], reverse=True)

        # Popola tabella sintetica
        self.tree.delete(*self.tree.get_children())
        for a in self._agg:
            tag = ('has_scrap',) if a['nscrap'] > 0 else ()
            self.tree.insert('', 'end', values=(
                a['cod'], a['desc'], a['gruppo'], a['tipo'],
                a['nreq'], f"{a['qta']:.2f}", f"{a['scrap']:.2f}", a['nscrap']
            ), tags=tag)

        # KPI
        n_req = len(self._detail_rows)
        n_mat = len(self._agg)
        qty_tot = sum(float(r[self.C_QTA] or 0) for r in self._detail_rows)
        scrap_tot = sum(float(r[self.C_SCRAP] or 0) for r in self._detail_rows)
        n_scrap = sum(1 for r in self._detail_rows if (r[self.C_SCRAPN] or 0) > 0)
        d_from = self._get_date(self.date_from)
        d_to = self._get_date(self.date_to)
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        self._kpi_vars['period'].set(f"{d_from.strftime('%d/%m/%y')}–{d_to.strftime('%d/%m/%y')}")
        self._kpi_vars['n_req'].set(str(n_req))
        self._kpi_vars['n_mat'].set(str(n_mat))
        self._kpi_vars['qty'].set(f"{qty_tot:.0f}")
        self._kpi_vars['scrap'].set(f"{scrap_tot:.1f}")
        self._kpi_vars['n_scrap'].set(str(n_scrap))

    def _sort_by(self, col):
        """Ordina la tabella sintetica per la colonna scelta."""
        numeric = {'nreq', 'qta', 'scrap', 'nscrap'}
        rev = self._sort_state.get(col, False)
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        if col in numeric:
            items.sort(key=lambda t: float(str(t[0]).replace(',', '') or 0), reverse=not rev)
        else:
            items.sort(key=lambda t: str(t[0]).lower(), reverse=not rev)
        for idx, (_v, k) in enumerate(items):
            self.tree.move(k, '', idx)
        self._sort_state[col] = not rev

    # ------------------------------------------------------------------ #
    #  Export Excel analitico                                              #
    # ------------------------------------------------------------------ #
    @staticmethod
    def _safe_sheet(name):
        for ch in r'[]:*?/\\':
            name = name.replace(ch, '_')
        return name[:31]

    def _export_excel(self):
        L = self.lang.get
        if not self._detail_rows:
            messagebox.showinfo(L('info', 'Info'),
                                L('cgr_no_data', 'Nessun dato da esportare per i filtri correnti.'),
                                parent=self)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(L('error', 'Errore'),
                                 "La libreria 'openpyxl' non è installata.\nEsegui: pip install openpyxl",
                                 parent=self)
            return

        default_dir = r'C:\Temp'
        os.makedirs(default_dir, exist_ok=True)
        default_name = f"Report_Consumi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self, title=L('cgr_save_excel', 'Salva Report Excel'),
            defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')],
            initialdir=default_dir, initialfile=default_name)
        if not path:
            return

        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2F6DA4')
        title_font = Font(bold=True, size=12)
        scrap_fill = PatternFill('solid', fgColor='FFF3CD')

        def write_header(ws, headers, row=1):
            for j, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=j, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal='center')

        def autowidth(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        d_from = self._get_date(self.date_from)
        d_to = self._get_date(self.date_to)
        if d_from > d_to:
            d_from, d_to = d_to, d_from

        try:
            wb = openpyxl.Workbook()

            # ── Foglio Sintesi (KPI + aggregato per codice) ──────────────
            ws = wb.active
            ws.title = 'Sintesi'
            ws.cell(row=1, column=1, value=L('cgr_title', 'Report Consumi Materiali - Generale')).font = title_font
            ws.cell(row=2, column=1,
                    value=f"{L('cgr_kpi_period', 'Periodo')}: "
                          f"{d_from.strftime('%d/%m/%Y')} - {d_to.strftime('%d/%m/%Y')}")
            qty_tot = sum(float(r[self.C_QTA] or 0) for r in self._detail_rows)
            scrap_tot = sum(float(r[self.C_SCRAP] or 0) for r in self._detail_rows)
            n_scrap = sum(1 for r in self._detail_rows if (r[self.C_SCRAPN] or 0) > 0)
            ws.cell(row=3, column=1,
                    value=f"# {L('cgr_kpi_nreq', '# Richieste')}: {len(self._detail_rows)}   "
                          f"# {L('cgr_kpi_nmat', '# Materiali')}: {len(self._agg)}   "
                          f"{L('cgr_kpi_qty', 'Qta richiesta')}: {qty_tot:.2f}   "
                          f"{L('cgr_kpi_scrap', 'Scrap reso (kg)')}: {scrap_tot:.2f}   "
                          f"{L('cgr_kpi_nscrap', '# Rich. con scrap')}: {n_scrap}")

            agg_headers = ['Codice', 'Descrizione', 'Gruppo', 'Tipo',
                           '# Richieste', 'Qta richiesta', 'Scrap (kg)', '# con scrap']
            write_header(ws, agg_headers, row=5)
            r = 6
            for a in self._agg:
                ws.cell(row=r, column=1, value=a['cod'])
                ws.cell(row=r, column=2, value=a['desc'])
                ws.cell(row=r, column=3, value=a['gruppo'])
                ws.cell(row=r, column=4, value=a['tipo'])
                ws.cell(row=r, column=5, value=a['nreq'])
                ws.cell(row=r, column=6, value=round(a['qta'], 2))
                ws.cell(row=r, column=7, value=round(a['scrap'], 2))
                ws.cell(row=r, column=8, value=a['nscrap'])
                if a['nscrap'] > 0:
                    for j in range(1, 9):
                        ws.cell(row=r, column=j).fill = scrap_fill
                r += 1
            autowidth(ws, [14, 40, 16, 14, 12, 14, 12, 12])

            # ── Foglio Analitico (tutte le richieste) ────────────────────
            ws2 = wb.create_sheet('Analitico')
            det_headers = ['ID Rich.', 'Data Richiesta', 'Codice', 'Descrizione', 'Gruppo',
                           'Tipo', 'Qta Richiesta', 'Stato', 'Richiedente', 'Preparatore',
                           'Data Preparazione', 'Data Prelievo',
                           'Scrap Reso (kg)', 'Scrap Confermato (kg)', '# Resi Scrap']
            write_header(ws2, det_headers)
            for row in self._detail_rows:
                ws2.append(self._row_for_excel(row))
                if (row[self.C_SCRAPN] or 0) > 0:
                    for c in ws2[ws2.max_row]:
                        c.fill = scrap_fill
            autowidth(ws2, [9, 17, 14, 40, 16, 12, 12, 12, 20, 20, 17, 17, 14, 16, 11])

            # ── Fogli TOP 10 codici (per quantita' richiesta) ────────────
            for i, a in enumerate(self._agg[:10], 1):
                cod = a['cod'] or f'mat{i}'
                ws_t = wb.create_sheet(self._safe_sheet(f"T{i:02d} {cod}"))
                ws_t.cell(row=1, column=1, value=f"{cod} — {a['desc']}").font = title_font
                ws_t.cell(row=2, column=1,
                          value=f"{L('cgr_col_group', 'Gruppo')}: {a['gruppo'] or '—'}   "
                                f"{L('cgr_col_type', 'Tipo')}: {a['tipo'] or '—'}")
                ws_t.cell(row=3, column=1,
                          value=f"# {L('cgr_col_nreq', '# Richieste')}: {a['nreq']}   "
                                f"{L('cgr_col_qty', 'Qta richiesta')}: {a['qta']:.2f}   "
                                f"{L('cgr_col_scrap', 'Scrap (kg)')}: {a['scrap']:.2f}")
                write_header(ws_t, det_headers, row=5)
                rr = 6
                for row in self._detail_rows:
                    if row[self.C_COD] != a['cod']:
                        continue
                    for j, val in enumerate(self._row_for_excel(row), 1):
                        ws_t.cell(row=rr, column=j, value=val)
                    if (row[self.C_SCRAPN] or 0) > 0:
                        for j in range(1, len(det_headers) + 1):
                            ws_t.cell(row=rr, column=j).fill = scrap_fill
                    rr += 1
                autowidth(ws_t, [9, 17, 14, 40, 16, 12, 12, 12, 20, 20, 17, 17, 14, 16, 11])

            wb.save(path)
            logger.info(f"[CGR] Export Excel: {path}")
            if messagebox.askyesno(
                L('success', 'Successo'),
                L('cgr_export_ok', 'Report Excel salvato:\n{0}\n\nAprire il file?').format(path),
                parent=self
            ):
                import subprocess
                subprocess.Popen(['start', '', path], shell=True)

        except Exception as e:
            logger.error(f"[CGR] Export Excel error: {e}", exc_info=True)
            messagebox.showerror(L('error', 'Errore'), f"Errore export Excel:\n{e}", parent=self)

    def _row_for_excel(self, row):
        """Trasforma una riga di dettaglio nella lista di celle per Excel."""
        def d(v, fmt='%d/%m/%Y %H:%M'):
            return v.strftime(fmt) if v else ''
        return [
            row[self.C_RID],
            d(row[self.C_DATA]),
            row[self.C_COD] or '',
            row[self.C_DESC] or '',
            row[self.C_FAM] or '',
            row[self.C_TIPO] or '',
            float(row[self.C_QTA] or 0),
            row[self.C_STATO] or '',
            row[self.C_RICH] or '',
            row[self.C_PREP] or '',
            d(row[self.C_DPREP]),
            d(row[self.C_DPREL]),
            float(row[self.C_SCRAP] or 0),
            float(row[self.C_SCRAPCONF] or 0),
            int(row[self.C_SCRAPN] or 0),
        ]


# ─────────────────────────────────────────────────────────────────────────────
#  Entry-point (senza login)
# ─────────────────────────────────────────────────────────────────────────────
def open_consumption_general_report(master, db, lang):
    """Entry-point richiamabile da main.py (nessun login richiesto)."""
    ConsumptionGeneralReportWindow(master, db, lang)
