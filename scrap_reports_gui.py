"""
Modulo per la gestione dei report scarti con export Excel formattato.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pyodbc
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
import logging
from tkcalendar import DateEntry
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import subprocess
import platform

logger = logging.getLogger("TraceabilityRS")


class ScrapReportsWindow(tk.Toplevel):
    """Finestra per la generazione di report scarti."""

    def __init__(self, master, db_handler, translator):
        super().__init__(master)
        #self.conn_str = conn_str
        self.db = db_handler
        self.translator = translator
        self.report_data = []

        self.title(self.translator.get("scrap_reports_title", "Rapporti Scarti"))
        self.geometry("900x650")
        self.resizable(True, True)

        # Centra la finestra
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (900 // 2)
        y = (self.winfo_screenheight() // 2) - (650 // 2)
        self.geometry(f"900x650+{x}+{y}")

        self._build_ui()
        logger.info("Finestra Rapporti Scarti aperta")

    def _build_ui(self):
        """Costruisce l'interfaccia utente."""
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- FILTRI ---
        filter_frame = ttk.LabelFrame(
            main_frame,
            text=self.translator.get("filters", "Filtri"),
            padding=10
        )
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        # Riga 1: Date range
        date_frame = ttk.Frame(filter_frame)
        date_frame.pack(fill=tk.X, pady=5)

        ttk.Label(date_frame, text=self.translator.get("date_from", "Data Da:")).pack(side=tk.LEFT, padx=(0, 5))
        self.date_from = DateEntry(
            date_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='it_IT'
        )
        self.date_from.pack(side=tk.LEFT, padx=(0, 20))

        ttk.Label(date_frame, text=self.translator.get("date_to", "Data A:")).pack(side=tk.LEFT, padx=(0, 5))
        self.date_to = DateEntry(
            date_frame,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='it_IT'
        )
        self.date_to.pack(side=tk.LEFT, padx=(0, 20))

        # Pulsante Carica
        ttk.Button(
            date_frame,
            text=self.translator.get("load_report", "Carica Report"),
            command=self._load_report
        ).pack(side=tk.LEFT, padx=10)

        # Riga 2: Report rapidi
        quick_frame = ttk.Frame(filter_frame)
        quick_frame.pack(fill=tk.X, pady=5)

        ttk.Label(quick_frame, text=self.translator.get("quick_reports", "Report Rapidi:")).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Button(
            quick_frame,
            text=self.translator.get("current_month", "Mese Corrente"),
            command=self._load_current_month
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            quick_frame,
            text=self.translator.get("current_year", "Anno Corrente"),
            command=self._load_current_year
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            quick_frame,
            text=self.translator.get("last_month", "Mese Precedente"),
            command=self._load_last_month
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            quick_frame,
            text=self.translator.get("last_year", "Anno Precedente"),
            command=self._load_last_year
        ).pack(side=tk.LEFT, padx=5)

        # --- TREEVIEW ---
        tree_frame = ttk.LabelFrame(
            main_frame,
            text=self.translator.get("scrap_data", "Dati Scarti"),
            padding=10
        )
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Scrollbar
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Treeview
        columns = ("ID", "User", "LabelCode","Productcode", "Phase", "Reason", "Reference", "Note", "DateIn")
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=15,
            yscrollcommand=tree_scroll.set
        )
        tree_scroll.config(command=self.tree.yview)

        # Intestazioni
        self.tree.heading("ID", text="ID")
        self.tree.heading("User", text=self.translator.get("user", "Utente"))
        self.tree.heading("LabelCode", text=self.translator.get("label_code", "Codice Etichetta"))
        self.tree.heading("Productcode", text=self.translator.get("Productcode", "Productcode"))
        self.tree.heading("Phase", text=self.translator.get("phase", "Fase"))
        self.tree.heading("Reason", text=self.translator.get("reason", "Motivo"))
        self.tree.heading("Reference", text=self.translator.get("reference", "Riferimento"))
        self.tree.heading("Note", text=self.translator.get("note", "Note"))
        self.tree.heading("DateIn", text=self.translator.get("date", "Data"))

        # Larghezza colonne
        self.tree.column("ID", width=50, anchor=tk.CENTER)
        self.tree.column("User", width=100)
        self.tree.column("LabelCode", width=120)
        self.tree.column("Productcode", width=120)
        self.tree.column("Phase", width=120)
        self.tree.column("Reason", width=150)
        self.tree.column("Reference", width=100)
        self.tree.column("Note", width=150)
        self.tree.column("DateIn", width=130)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # --- PULSANTI AZIONE ---
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(
            button_frame,
            text=self.translator.get("export_excel", "Esporta Excel"),
            command=self._export_excel
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text=self.translator.get("close", "Chiudi"),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)

        # Label contatore record
        self.count_label = ttk.Label(button_frame, text="")
        self.count_label.pack(side=tk.LEFT, padx=20)

    def _load_report(self):
        """Carica i dati del report per il periodo selezionato."""
        try:
            #conn = pyodbc.connect(self.conn_str)
            #cursor = conn.cursor()

            date_from = self.date_from.get_date()
            date_to = self.date_to.get_date()

            query = """SELECT 
                    s.[ScrapDeclarationId],
                    [User],
                    l.LabelCod   as LabelCode,
                    p.ProductCode as Productcode,
                    pp.AreaName as ParentPhaseName,
                    sr.Reason,
                    [Riferiments],
                    [Note] + iif(Accepted =1, ' (Scrapt accepted by quality)','') as [Note],
                    [DateIn],
                    [Picture]
                FROM [Traceability_RS].[dbo].[ScarpDeclarations] as S 
                INNER JOIN [Traceability_RS].[dbo].[ScrapResons] as sr 
                    ON s.ScrapReasonId = sr.ScrapReasonId
                inner join[Traceability_RS].[dbo].[Areas] pp on pp.idArea=s.[IDParentPhase]                 
                inner join traceability_rs.dbo.LabelCodes as L on s.IdLabelCode=l.IDLabelCode
                inner join traceability_rs.dbo.boards B on b.IDBoard=l.IDBoard
                inner join traceability_rs.dbo.orders o on b.IDOrder=o.idorder
                inner join traceability_rs.dbo.products p on p.idproduct=o.idproduct
                WHERE CAST(DateIn as date) BETWEEN ? AND ?
                ORDER BY DateIn DESC;
            """

            self.db.cursor.execute(query, date_from, date_to)
            self.report_data = self.db.cursor.fetchall()

            # Pulisci treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Inserisci dati
            for row in self.report_data:
                date_str = row.DateIn.strftime('%d/%m/%Y %H:%M') if row.DateIn else ""
                self.tree.insert("", tk.END, values=(
                    row.ScrapDeclarationId,
                    row.User or "",
                    row.LabelCode or "",
                    row.Productcode or "",
                    row.ParentPhaseName or "",
                    row.Reason or "",
                    row.Riferiments or "",
                    row.Note or "",
                    date_str
                ))

            #conn.close()

            # Aggiorna contatore
            self.count_label.config(
                text=self.translator.get("records_found", f"Record trovati: {len(self.report_data)}")
            )

            logger.info(f"Caricati {len(self.report_data)} record scarti dal {date_from} al {date_to}")

        except Exception as e:
            logger.error(f"Errore caricamento report scarti: {e}")
            messagebox.showerror(
                self.translator.get("error", "Errore"),
                f"{self.translator.get('load_error', 'Errore caricamento')}: {str(e)}"
            )

    def _load_current_month(self):
        """Carica il report per il mese corrente."""
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = (first_day + relativedelta(months=1)) - timedelta(days=1)

        self.date_from.set_date(first_day)
        self.date_to.set_date(last_day)
        self._load_report()

    def _load_current_year(self):
        """Carica il report per l'anno corrente."""
        today = datetime.now()
        first_day = today.replace(month=1, day=1)
        last_day = today.replace(month=12, day=31)

        self.date_from.set_date(first_day)
        self.date_to.set_date(last_day)
        self._load_report()

    def _load_last_month(self):
        """Carica il report per il mese precedente."""
        today = datetime.now()
        first_day = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_day = today.replace(day=1) - timedelta(days=1)

        self.date_from.set_date(first_day)
        self.date_to.set_date(last_day)
        self._load_report()

    def _load_last_year(self):
        """Carica il report per l'anno precedente."""
        today = datetime.now()
        year = today.year - 1
        first_day = datetime(year, 1, 1)
        last_day = datetime(year, 12, 31)

        self.date_from.set_date(first_day)
        self.date_to.set_date(last_day)
        self._load_report()

    def _export_excel(self):
        """Esporta il report in Excel formattato professionalmente."""
        if not self.report_data:
            messagebox.showwarning(
                self.translator.get("warning", "Attenzione"),
                self.translator.get("no_data_to_export", "Nessun dato da esportare")
            )
            return

        try:
            # Crea directory C:\Temp se non esiste
            output_dir = r"C:\Temp"
            os.makedirs(output_dir, exist_ok=True)

            # Nome file con timestamp
            date_from_str = self.date_from.get_date().strftime('%Y%m%d')
            date_to_str = self.date_to.get_date().strftime('%Y%m%d')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = os.path.join(
                output_dir,
                f"Scrap_Report_{date_from_str}_{date_to_str}_{timestamp}.xlsx"
            )

            # Crea DataFrame
            data = []
            for row in self.report_data:
                date_str = row.DateIn.strftime('%d/%m/%Y %H:%M') if row.DateIn else ""
                data.append({
                    'ID': row.ScrapDeclarationId,
                    'Utente': row.User or "",
                    'Codice Etichetta': row.LabelCode or "",
                    'Codice prodotto:': row.Productcode or "",
                    'Fase': row.ParentPhaseName or "",
                    'Motivo': row.Reason or "",
                    'Riferimento': row.Riferiments or "",
                    'Note': row.Note or "",
                    'Data': date_str
                })

            df = pd.DataFrame(data)

            # Scrivi Excel base
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Scarti')

            # Applica formattazione professionale
            self._format_excel(filename, date_from_str, date_to_str)

            messagebox.showinfo(
                self.translator.get("success", "Successo"),
                f"{self.translator.get('export_success', 'Esportazione completata')}\n{filename}"
            )

            # Apri il file
            self._open_file(filename)

            logger.info(f"Report scarti esportato: {filename}")

        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}")
            messagebox.showerror(
                self.translator.get("error", "Errore"),
                f"{self.translator.get('export_error', 'Errore esportazione')}: {str(e)}"
            )

    def _format_excel(self, filename, date_from_str, date_to_str):
        """Applica formattazione professionale al file Excel."""
        wb = load_workbook(filename)
        ws = wb.active

        # Colori aziendali
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        title_font = Font(name='Calibri', size=14, bold=True, color="1F4788")
        cell_font = Font(name='Calibri', size=10)

        # Bordi
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        from datetime import datetime

        # Inserisci riga titolo
        ws.insert_rows(1, 2)
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        date_from_formatted = datetime.strptime(date_from_str, "%Y%m%d").strftime("%d-%m-%Y")
        date_to_formatted = datetime.strptime(date_to_str, "%Y%m%d").strftime("%d-%m-%Y")
        title_cell.value = f"REPORT SCARTI - Periodo: {date_from_formatted} / {date_to_formatted}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Formatta header (ora riga 3)
        for col_num, cell in enumerate(ws[3], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Formatta dati
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Larghezza colonne
        column_widths = {
            'A': 8,   # ID
            'B': 15,  # Utente
            'C': 20,  # Codice Etichetta
            'D': 20,  # ProductCode
            'E': 20,  # Fase
            'F': 25,  # Motivo
            'G': 20,  # Riferimento
            'H': 30,  # Note
            'I': 18   # Data
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Altezza riga titolo
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 20

        # Freeze panes (blocca intestazioni)
        ws.freeze_panes = 'A4'

        # Salva
        wb.save(filename)

    def _open_file(self, filename):
        """Apre il file con l'applicazione predefinita."""
        try:
            if platform.system() == 'Windows':
                os.startfile(filename)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', filename])
            else:  # Linux
                subprocess.call(['xdg-open', filename])
        except Exception as e:
            logger.warning(f"Impossibile aprire automaticamente il file: {e}")


def show(master, conn_str, translator):
    """Funzione helper per aprire la finestra."""
    window = ScrapReportsWindow(master, conn_str, translator)
    window.transient(master)
    window.grab_set()
    master.wait_window(window)
