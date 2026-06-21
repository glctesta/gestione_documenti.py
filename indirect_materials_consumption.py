"""
indirect_materials_consumption.py
Analisi consumi materiali indiretti su base settimanale / mensile / annuale
e motore di proposta budget per l'anno successivo.

I consumi si basano sui movimenti di SCARICO (ind.MaterialiMovimenti):
  consumo = -SUM(Qty)  (Qty degli scarichi e' negativa).
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    1: 'Gen', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mag', 6: 'Giu',
    7: 'Lug', 8: 'Ago', 9: 'Set', 10: 'Ott', 11: 'Nov', 12: 'Dic'
}


# ----------------------------------------------------------------------------
#  Query consumi (riutilizzabili)
# ----------------------------------------------------------------------------
def get_weekly_consumption(db, weeks=26):
    """Consumo totale per settimana ISO nelle ultime `weeks` settimane."""
    query = """
        SELECT DATEPART(YEAR, mv.DataMovimento)     AS Anno,
               DATEPART(ISO_WEEK, mv.DataMovimento) AS Settimana,
               -SUM(mv.Qty)                          AS Consumo,
               COUNT(*)                              AS NMovimenti
        FROM ind.MaterialiMovimenti mv
        WHERE mv.TipoMovimento = 'SCARICO'
          AND mv.DataMovimento >= DATEADD(WEEK, -?, GETDATE())
        GROUP BY DATEPART(YEAR, mv.DataMovimento), DATEPART(ISO_WEEK, mv.DataMovimento)
        ORDER BY Anno, Settimana
    """
    rows = db.fetch_all(query, (weeks,))
    return [{'anno': r[0], 'settimana': r[1], 'consumo': float(r[2] or 0),
             'n_mov': r[3]} for r in (rows or [])]


def get_monthly_consumption(db, months=24):
    """Consumo totale per mese negli ultimi `months` mesi."""
    query = """
        SELECT YEAR(mv.DataMovimento)  AS Anno,
               MONTH(mv.DataMovimento) AS Mese,
               -SUM(mv.Qty)            AS Consumo,
               COUNT(*)                AS NMovimenti
        FROM ind.MaterialiMovimenti mv
        WHERE mv.TipoMovimento = 'SCARICO'
          AND mv.DataMovimento >= DATEADD(MONTH, -?, GETDATE())
        GROUP BY YEAR(mv.DataMovimento), MONTH(mv.DataMovimento)
        ORDER BY Anno, Mese
    """
    rows = db.fetch_all(query, (months,))
    return [{'anno': r[0], 'mese': r[1], 'consumo': float(r[2] or 0),
             'n_mov': r[3]} for r in (rows or [])]


def get_yearly_consumption(db):
    """Consumo totale per anno (tutto lo storico movimenti)."""
    query = """
        SELECT YEAR(mv.DataMovimento) AS Anno,
               -SUM(mv.Qty)           AS Consumo,
               COUNT(*)               AS NMovimenti
        FROM ind.MaterialiMovimenti mv
        WHERE mv.TipoMovimento = 'SCARICO'
        GROUP BY YEAR(mv.DataMovimento)
        ORDER BY Anno
    """
    rows = db.fetch_all(query)
    return [{'anno': r[0], 'consumo': float(r[1] or 0), 'n_mov': r[2]}
            for r in (rows or [])]


def get_budget_proposal(db, growth_pct=0.0):
    """Proposta budget per l'anno successivo, per materiale, basata sul
    consumo degli ultimi 12 mesi. growth_pct = % di crescita da applicare.

    Ritorna lista di dict con: codice, descrizione, tipo,
    consumo_12m, budget_annuo, budget_mensile.
    """
    query = """
        SELECT m.CodiceMateriale, m.DescrizioneMateriale,
               ISNULL(t.Tipo, 'Generico') AS Tipo,
               -SUM(mv.Qty) AS Consumo12m
        FROM ind.MaterialiMovimenti mv
        JOIN ind.Materiali m ON m.MaterialeId = mv.MaterialeId
        LEFT JOIN ind.TipoMateriali t ON t.TipoMaterialeId = m.TipoMaterialeId
        WHERE mv.TipoMovimento = 'SCARICO'
          AND mv.DataMovimento >= DATEADD(MONTH, -12, GETDATE())
        GROUP BY m.CodiceMateriale, m.DescrizioneMateriale, ISNULL(t.Tipo, 'Generico')
        HAVING -SUM(mv.Qty) > 0
        ORDER BY Consumo12m DESC
    """
    rows = db.fetch_all(query)
    factor = 1.0 + (growth_pct / 100.0)
    result = []
    for r in (rows or []):
        consumo = float(r[3] or 0)
        budget_annuo = consumo * factor
        result.append({
            'codice': r[0] or '',
            'descrizione': r[1] or '',
            'tipo': r[2] or 'Generico',
            'consumo_12m': consumo,
            'budget_annuo': budget_annuo,
            'budget_mensile': budget_annuo / 12.0,
        })
    return result


# ============================================================================
#  Finestra di analisi consumi e budget
# ============================================================================
class ConsumptionAnalysisWindow(tk.Toplevel):
    """Finestra a tab: consumi settimanali / mensili / annuali + budget."""

    def __init__(self, master, db, lang, user_name="Unknown"):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name

        self.title(lang.get('ind_cons_title', 'Analisi Consumi & Budget Materiali Indiretti'))
        self.geometry("1000x640")
        self.resizable(True, True)
        self.transient(master)

        self._build_ui()
        self._load_all()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(expand=True, fill="both")

        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text=self.lang.get('ind_cons_header', 'Analisi Consumi'),
                  font=("Segoe UI", 12, "bold")).pack(side="left")
        ttk.Button(header, text=self.lang.get('btn_refresh', 'Aggiorna'),
                   command=self._load_all).pack(side="right", padx=4)
        ttk.Button(header, text=self.lang.get('ind_cons_export', '📊 Esporta Excel'),
                   command=self._export_excel).pack(side="right", padx=4)

        self.nb = ttk.Notebook(main)
        self.nb.pack(expand=True, fill="both")

        # --- Tab settimanale ---
        self.tab_week = ttk.Frame(self.nb)
        self.nb.add(self.tab_week, text=self.lang.get('ind_cons_tab_week', 'Settimanale'))
        self.tree_week = self._make_tree(
            self.tab_week,
            [('periodo', self.lang.get('ind_cons_col_period', 'Periodo'), 160, 'w'),
             ('consumo', self.lang.get('ind_cons_col_consumption', 'Consumo'), 140, 'e'),
             ('nmov', self.lang.get('ind_cons_col_moves', 'N. movimenti'), 120, 'e')])

        # --- Tab mensile ---
        self.tab_month = ttk.Frame(self.nb)
        self.nb.add(self.tab_month, text=self.lang.get('ind_cons_tab_month', 'Mensile'))
        self.tree_month = self._make_tree(
            self.tab_month,
            [('periodo', self.lang.get('ind_cons_col_period', 'Periodo'), 160, 'w'),
             ('consumo', self.lang.get('ind_cons_col_consumption', 'Consumo'), 140, 'e'),
             ('nmov', self.lang.get('ind_cons_col_moves', 'N. movimenti'), 120, 'e')])

        # --- Tab annuale ---
        self.tab_year = ttk.Frame(self.nb)
        self.nb.add(self.tab_year, text=self.lang.get('ind_cons_tab_year', 'Annuale'))
        self.tree_year = self._make_tree(
            self.tab_year,
            [('periodo', self.lang.get('ind_cons_col_year', 'Anno'), 160, 'w'),
             ('consumo', self.lang.get('ind_cons_col_consumption', 'Consumo'), 140, 'e'),
             ('nmov', self.lang.get('ind_cons_col_moves', 'N. movimenti'), 120, 'e')])

        # --- Tab budget ---
        self.tab_budget = ttk.Frame(self.nb)
        self.nb.add(self.tab_budget, text=self.lang.get('ind_cons_tab_budget', 'Budget anno prossimo'))

        bctrl = ttk.Frame(self.tab_budget)
        bctrl.pack(fill="x", pady=(6, 6))
        ttk.Label(bctrl, text=self.lang.get('ind_cons_growth', 'Crescita % attesa:')).pack(side="left", padx=(4, 4))
        self.growth_var = tk.StringVar(value="0")
        ttk.Entry(bctrl, textvariable=self.growth_var, width=8).pack(side="left")
        ttk.Button(bctrl, text=self.lang.get('ind_cons_recalc', 'Ricalcola budget'),
                   command=self._load_budget).pack(side="left", padx=8)

        self.tree_budget = self._make_tree(
            self.tab_budget,
            [('codice', self.lang.get('ind_import_col_code', 'Codice'), 110, 'w'),
             ('descrizione', self.lang.get('ind_import_col_desc', 'Descrizione'), 280, 'w'),
             ('tipo', self.lang.get('ind_req_col_type', 'Tipo'), 100, 'w'),
             ('consumo12m', self.lang.get('ind_cons_col_12m', 'Consumo 12 mesi'), 130, 'e'),
             ('budgeta', self.lang.get('ind_cons_col_budget_year', 'Budget annuo'), 120, 'e'),
             ('budgetm', self.lang.get('ind_cons_col_budget_month', 'Budget mensile'), 120, 'e')],
            pack=False)
        self.tree_budget.pack(in_=self.tab_budget, side="left", fill="both", expand=True)

    def _make_tree(self, parent, columns, pack=True):
        frame = ttk.Frame(parent)
        if pack:
            frame.pack(fill="both", expand=True)
        cols = [c[0] for c in columns]
        tree = ttk.Treeview(frame, columns=cols, show='headings', selectmode='browse')
        for key, label, width, anchor in columns:
            tree.heading(key, text=label)
            tree.column(key, width=width, anchor=anchor)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        if pack:
            tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
        return tree

    # ------------------------------------------------------------------ #
    def _load_all(self):
        self._load_weekly()
        self._load_monthly()
        self._load_yearly()
        self._load_budget()

    def _load_weekly(self):
        self.tree_week.delete(*self.tree_week.get_children())
        try:
            self._week_data = get_weekly_consumption(self.db)
        except Exception as e:
            logger.error(f"Errore consumi settimanali: {e}", exc_info=True)
            self._week_data = []
        for r in self._week_data:
            self.tree_week.insert('', 'end', values=(
                f"{r['anno']} - W{r['settimana']:02d}", f"{r['consumo']:.2f}", r['n_mov']))

    def _load_monthly(self):
        self.tree_month.delete(*self.tree_month.get_children())
        try:
            self._month_data = get_monthly_consumption(self.db)
        except Exception as e:
            logger.error(f"Errore consumi mensili: {e}", exc_info=True)
            self._month_data = []
        for r in self._month_data:
            self.tree_month.insert('', 'end', values=(
                f"{r['anno']} - {MONTH_NAMES.get(r['mese'], r['mese'])}",
                f"{r['consumo']:.2f}", r['n_mov']))

    def _load_yearly(self):
        self.tree_year.delete(*self.tree_year.get_children())
        try:
            self._year_data = get_yearly_consumption(self.db)
        except Exception as e:
            logger.error(f"Errore consumi annuali: {e}", exc_info=True)
            self._year_data = []
        for r in self._year_data:
            self.tree_year.insert('', 'end', values=(
                str(r['anno']), f"{r['consumo']:.2f}", r['n_mov']))

    def _load_budget(self):
        self.tree_budget.delete(*self.tree_budget.get_children())
        try:
            growth = float(self.growth_var.get().replace(',', '.'))
        except ValueError:
            growth = 0.0
        try:
            self._budget_data = get_budget_proposal(self.db, growth_pct=growth)
        except Exception as e:
            logger.error(f"Errore budget: {e}", exc_info=True)
            self._budget_data = []
        for r in self._budget_data:
            self.tree_budget.insert('', 'end', values=(
                r['codice'], r['descrizione'], r['tipo'],
                f"{r['consumo_12m']:.2f}", f"{r['budget_annuo']:.2f}",
                f"{r['budget_mensile']:.2f}"))

    # ------------------------------------------------------------------ #
    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "La libreria 'openpyxl' non è installata.\nEsegui: pip install openpyxl",
                parent=self)
            return

        default_name = f"Consumi_Materiali_Indiretti_{datetime.now().strftime('%Y%m%d')}.xlsx"
        default_dir = r'C:\Temp'
        os.makedirs(default_dir, exist_ok=True)
        path = filedialog.asksaveasfilename(
            parent=self,
            title=self.lang.get('ind_cons_save_excel', 'Salva Consumi Excel'),
            defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')],
            initialdir=default_dir, initialfile=default_name)
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            hdr_font = Font(bold=True, color='FFFFFF')
            hdr_fill = PatternFill('solid', fgColor='2F6DA4')

            def fill_sheet(ws, headers, rows):
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                for row in rows:
                    ws.append(row)

            ws1 = wb.active
            ws1.title = 'Settimanale'
            fill_sheet(ws1, ['Anno', 'Settimana', 'Consumo', 'N. movimenti'],
                       [[r['anno'], r['settimana'], r['consumo'], r['n_mov']] for r in self._week_data])

            ws2 = wb.create_sheet('Mensile')
            fill_sheet(ws2, ['Anno', 'Mese', 'Consumo', 'N. movimenti'],
                       [[r['anno'], MONTH_NAMES.get(r['mese'], r['mese']), r['consumo'], r['n_mov']]
                        for r in self._month_data])

            ws3 = wb.create_sheet('Annuale')
            fill_sheet(ws3, ['Anno', 'Consumo', 'N. movimenti'],
                       [[r['anno'], r['consumo'], r['n_mov']] for r in self._year_data])

            ws4 = wb.create_sheet('Budget')
            fill_sheet(ws4, ['Codice', 'Descrizione', 'Tipo', 'Consumo 12 mesi',
                             'Budget annuo', 'Budget mensile'],
                       [[r['codice'], r['descrizione'], r['tipo'], r['consumo_12m'],
                         r['budget_annuo'], r['budget_mensile']] for r in self._budget_data])

            wb.save(path)
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('ind_cons_export_ok', 'Export completato:\n{0}').format(path),
                parent=self)
        except Exception as e:
            logger.error(f"Errore export consumi: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)


def open_consumption_analysis(master, db, lang, user_name="Unknown"):
    ConsumptionAnalysisWindow(master, db, lang, user_name)
