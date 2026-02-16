import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import logging
import os
from .project_window import ProjectWindow
from .gantt_window import NpiGanttWindow
from .analysis_window import ProjectAnalysisWindow

# Import per PDF
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)


class NpiDashboardWindow(tk.Toplevel):
    def __init__(self, master, npi_manager, lang, logged_in_user):
        super().__init__(master)
        self.npi_manager = npi_manager
        self.lang = lang
        self.master_app = master
        self.logged_in_user = logged_in_user

        self.title(self.lang.get('npi_dashboard_title', "Dashboard Progetti NPI"))
        self.geometry("1200x700")
        self.transient(master)
        self.grab_set()

        self._create_widgets()
        self.load_npi_projects()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- SEZIONE STATISTICHE ---
        stats_frame = ttk.LabelFrame(main_frame, text="Riepilogo NPI", padding=10)
        stats_frame.pack(fill=tk.X, pady=(0, 10))

        # Frame per i contatori principali (Sinistra)
        counters_frame = ttk.Frame(stats_frame)
        counters_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 20))

        # Helper per creare box statistica
        def create_stat_box(parent, label_text, color='black'):
            f = ttk.Frame(parent)
            f.pack(fill=tk.X, pady=2)
            ttk.Label(f, text=label_text, width=25).pack(side=tk.LEFT)
            l = ttk.Label(f, text="0", font=('Helvetica', 10, 'bold'), foreground=color)
            l.pack(side=tk.RIGHT)
            return l

        self.lbl_total = create_stat_box(counters_frame, "Totale NPI Caricati:")
        self.lbl_completed = create_stat_box(counters_frame, "Completati (Chiusi):", 'green')
        self.lbl_delayed = create_stat_box(counters_frame, "In Ritardo (vs Scadenza):", 'red')

        # Stato notifiche automatiche NPI (check visivo)
        self.lbl_notifications_status = create_stat_box(counters_frame, "Notifiche Automatiche:")
        if hasattr(self.lbl_notifications_status, "bind"):
            self.lbl_notifications_status.bind("<Button-1>", self._on_notifications_status_click)
            self.lbl_notifications_status.configure(cursor="hand2")
        self._update_notifications_status()

        # Separatore verticale
        ttk.Separator(stats_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        # Frame per statistiche clienti (Destra)
        customer_frame = ttk.Frame(stats_frame)
        customer_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Header con filtro anno
        header_frame = ttk.Frame(customer_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(header_frame, text="Dettaglio per Cliente:", font=('Helvetica', 9, 'underline')).pack(side=tk.LEFT)
        
        # Filtro Anno
        ttk.Label(header_frame, text="Anno:", font=('Helvetica', 8)).pack(side=tk.LEFT, padx=(20, 5))
        self.year_filter_var = tk.StringVar()
        self.year_combo = ttk.Combobox(header_frame, textvariable=self.year_filter_var, 
                                       state="readonly", width=12)
        self.year_combo.pack(side=tk.LEFT)
        self.year_combo.bind('<<ComboboxSelected>>', lambda e: self.load_npi_projects())
        
        # Popola anni disponibili (anno corrente e precedenti)
        current_year = datetime.now().year
        years = ["Tutti gli anni"] + [str(year) for year in range(current_year, current_year - 10, -1)]
        self.year_combo['values'] = years
        self.year_combo.set(str(current_year))  # Default: anno corrente
        
        # Filtro Cliente
        ttk.Label(header_frame, text="Cliente:", font=('Helvetica', 8)).pack(side=tk.LEFT, padx=(20, 5))
        self.client_filter_var = tk.StringVar()
        self.client_combo = ttk.Combobox(header_frame, textvariable=self.client_filter_var,
                                        state="readonly", width=20)
        self.client_combo.pack(side=tk.LEFT)
        self.client_combo.bind('<<ComboboxSelected>>', lambda e: self.load_npi_projects())
        
        # Popola clienti disponibili
        try:
            clients = self.npi_manager.get_all_clients()
            client_values = ["Tutti i clienti"] + sorted(clients, key=lambda c: (c or "").lower())
            self.client_combo['values'] = client_values
            self.client_combo.set("Tutti i clienti")  # Default: tutti i clienti
        except Exception as e:
            logger.error(f"Errore nel caricamento clienti: {e}")
            self.client_combo['values'] = ["Tutti i clienti"]
            self.client_combo.set("Tutti i clienti")
        
        # Filtro Stato
        ttk.Label(header_frame, text="Stato:", font=('Helvetica', 8)).pack(side=tk.LEFT, padx=(20, 5))
        self.status_filter_var = tk.StringVar()
        self.status_combo = ttk.Combobox(header_frame, textvariable=self.status_filter_var,
                                        state="readonly", width=15)
        self.status_combo.pack(side=tk.LEFT)
        self.status_combo.bind('<<ComboboxSelected>>', lambda e: self.load_npi_projects())
        
        status_options = ["Tutti i progetti", "Solo in ritardo", "Solo chiusi", "Solo on track"]
        self.status_combo['values'] = status_options
        self.status_combo.set("Tutti i progetti")
        
        # Canvas scrollabile orizzontalmente per le statistiche clienti
        canvas_frame = ttk.Frame(customer_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar orizzontale
        h_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Canvas
        self.customer_canvas = tk.Canvas(canvas_frame, height=100, 
                                         xscrollcommand=h_scrollbar.set)
        self.customer_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        h_scrollbar.config(command=self.customer_canvas.xview)
        
        # Frame interno per i clienti
        self.customer_stats_inner_frame = ttk.Frame(self.customer_canvas)
        self.customer_canvas_window = self.customer_canvas.create_window((0, 0), 
                                                                          window=self.customer_stats_inner_frame, 
                                                                          anchor='nw')
        
        # Aggiorna scroll region quando il frame cambia dimensione
        def on_frame_configure(event):
            self.customer_canvas.configure(scrollregion=self.customer_canvas.bbox("all"))
        
        self.customer_stats_inner_frame.bind('<Configure>', on_frame_configure)
        # --- FINE SEZIONE STATISTICHE ---

        list_frame = ttk.LabelFrame(main_frame, text=self.lang.get('active_npi_projects', "Progetti NPI Attivi"),
                                    padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True)

        # --- **MODIFICA QUI: NUOVE COLONNE E TAG** ---
        cols = ('status_icon', 'project_name', 'product_code', 'customer', 'project_end_date',
                'total_tasks', 'completed_on_time', 'completed_late', 'pending_late', 'completion_pct')
        self.project_tree = ttk.Treeview(list_frame, columns=cols, show='headings', selectmode='browse')

        # Colonna per l'icona di stato 'X' (senza header)
        self.project_tree.heading('status_icon', text='')
        self.project_tree.column('status_icon', width=30, anchor=tk.CENTER, stretch=tk.NO)

        # Configurazioni colonne esistenti
        self.project_tree.heading('project_name', text=self.lang.get('col_project_name', 'Nome Progetto'))
        self.project_tree.column('project_name', width=250)
        self.project_tree.heading('product_code', text=self.lang.get('col_product_code', 'Codice Prodotto'))
        self.project_tree.column('product_code', width=120)
        self.project_tree.heading('customer', text=self.lang.get('col_customer', 'Cliente'))
        self.project_tree.column('customer', width=150)
        self.project_tree.heading('project_end_date',
                                  text=self.lang.get('col_project_end_date', 'Data Fine Progetto'))
        self.project_tree.column('project_end_date', width=120, anchor=tk.CENTER)
        
        # Nuove colonne per statistiche task
        self.project_tree.heading('total_tasks', text='Tot Task')
        self.project_tree.column('total_tasks', width=70, anchor=tk.CENTER)
        self.project_tree.heading('completed_on_time', text='✅ In Tempo')
        self.project_tree.column('completed_on_time', width=80, anchor=tk.CENTER)
        self.project_tree.heading('completed_late', text='⏰ In Ritardo')
        self.project_tree.column('completed_late', width=90, anchor=tk.CENTER)
        self.project_tree.heading('pending_late', text='⚠️ Da Fare (Ritardo)')
        self.project_tree.column('pending_late', width=120, anchor=tk.CENTER)
        self.project_tree.heading('completion_pct', text='% Completamento')
        self.project_tree.column('completion_pct', width=110, anchor=tk.CENTER)

        # Definizione dei tag
        self.project_tree.tag_configure('overdue', foreground='red', font=('Helvetica', 9, 'bold'))
        self.project_tree.tag_configure('closed', foreground='green', font=('Helvetica', 9, 'bold'))  # Verde per progetti chiusi
        self.project_tree.tag_configure('selected_client', background='#E3F2FD')  # Light blue per cliente selezionato
        self.project_tree.tag_configure('high_completion', foreground='green', font=('Helvetica', 9, 'bold'))
        self.project_tree.tag_configure('medium_completion', foreground='orange', font=('Helvetica', 9, 'bold'))
        self.project_tree.tag_configure('low_completion', foreground='red', font=('Helvetica', 9, 'bold'))
        # --- **FINE MODIFICA** ---

        self.project_tree.pack(fill=tk.BOTH, expand=True)

        self.project_tree.bind('<Double-1>', self._on_double_click)
        self.project_tree.bind("<Button-3>", self._show_project_context_menu)

        # Pulsanti
        button_frame = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        button_frame.pack(fill=tk.X)
        ttk.Button(button_frame, text=self.lang.get('btn_refresh', 'Aggiorna'), command=self.load_npi_projects).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.lang.get('btn_analyze', 'Analisi'),
                   command=self._launch_analysis_window).pack(side=tk.LEFT, padx=5)
        self.export_pdf_button = ttk.Button(button_frame, text=self.lang.get('btn_export_pdf', 'Export PDF a C:\\Temp'),
                                            command=self._export_summary_pdf)
        self.export_pdf_button.pack(side=tk.LEFT, padx=5)
        if not REPORTLAB_AVAILABLE:
            self.export_pdf_button.config(state=tk.DISABLED)
        
        # Bottone Export Excel
        self.export_excel_button = ttk.Button(button_frame, text=self.lang.get('btn_export_excel', 'Export Excel a C:\\Temp'),
                                              command=self._export_to_excel_new)
        self.export_excel_button.pack(side=tk.LEFT, padx=5)
        
        # Bottone Export Report Panoramico
        self.export_overview_button = ttk.Button(
            button_frame,
            text=self.lang.get('btn_export_overview', 'Export Rapporto Panoramico'),
            command=self._export_overview_report
        )
        self.export_overview_button.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text=self.lang.get('btn_close', 'Chiudi'), command=self.destroy).pack(side=tk.RIGHT)

    def _update_statistics(self):
        """Aggiorna la sezione delle statistiche."""
        try:
            self._update_notifications_status()
            # Ottieni anno selezionato
            year_str = self.year_filter_var.get()
            year_filter = None if year_str == "Tutti gli anni" else int(year_str)
            
            # Ottieni cliente selezionato
            client_str = self.client_filter_var.get()
            client_filter = None if client_str == "Tutti i clienti" else client_str
            
            stats = self.npi_manager.get_npi_statistics(year_filter=year_filter, client_filter=client_filter)
            if not stats:
                return

            # Aggiorna etichette totali
            self.lbl_total.config(text=str(stats['total_projects']))
            self.lbl_completed.config(text=str(stats['completed_projects']))
            self.lbl_delayed.config(text=str(stats['delayed_projects']))

            # Aggiorna statistiche clienti (pulisci frame prima)
            for widget in self.customer_stats_inner_frame.winfo_children():
                widget.destroy()

            # Disponi i clienti ORIZZONTALMENTE (senza limite)
            # Header
            ttk.Label(self.customer_stats_inner_frame, text="Cliente", 
                     font=('Helvetica', 8, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=2)
            ttk.Label(self.customer_stats_inner_frame, text="N°", 
                     font=('Helvetica', 8, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=2)
            ttk.Label(self.customer_stats_inner_frame, text="%", 
                     font=('Helvetica', 8, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=2)

            # Dati clienti (disposti in colonne)
            for i, item in enumerate(stats['customer_stats']):
                col = i + 1
                ttk.Label(self.customer_stats_inner_frame, text=item['client']).grid(
                    row=0, column=col, sticky='w', padx=10, pady=2)
                ttk.Label(self.customer_stats_inner_frame, text=str(item['count'])).grid(
                    row=1, column=col, sticky='e', padx=10, pady=2)
                ttk.Label(self.customer_stats_inner_frame, text=f"{item['percentage']}%").grid(
                    row=2, column=col, sticky='e', padx=10, pady=2)
                    
        except Exception as e:
            logger.error(f"Errore aggiornamento statistiche: {e}")

    def _update_notifications_status(self):
        """Aggiorna indicatore visivo dello stato notifiche automatiche."""
        status_text = "Sconosciuto"
        status_color = "black"
        try:
            config_path = "npi_notifications_config.json"
            if os.path.exists(config_path):
                import json
                with open(config_path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                enabled = bool(config.get("notification_settings", {}).get("enabled", False))
                if enabled:
                    status_text = "Attive"
                    status_color = "green"
                else:
                    status_text = "Disattivate"
                    status_color = "red"
            else:
                status_text = "Config assente"
                status_color = "orange"
        except Exception as e:
            logger.error(f"Errore stato notifiche automatiche: {e}")
            status_text = "Errore"
            status_color = "red"

        if hasattr(self, "lbl_notifications_status"):
            self.lbl_notifications_status.config(text=status_text, foreground=status_color)

    def _on_notifications_status_click(self, event=None):
        """Apre il file di configurazione notifiche NPI con autorizzazione."""
        try:
            if hasattr(self.master_app, '_execute_authorized_action'):
                self.master_app._execute_authorized_action(
                    menu_translation_key='Gestisci json NPI',
                    action_callback=self._open_npi_notifications_config
                )
            else:
                self._open_npi_notifications_config()
        except Exception as e:
            logger.error(f"Errore apertura config notifiche NPI: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile aprire la configurazione NPI:\n{e}", parent=self)

    def _open_npi_notifications_config(self):
        config_path = "npi_notifications_config.json"
        if not os.path.exists(config_path):
            messagebox.showwarning("Config assente", "File di configurazione non trovato.", parent=self)
            return
        try:
            warning_text = self.lang.get(
                'npi_notifications_config_warning',
                "Attenzione: la modifica di questo file può compromettere la funzionalità delle notifiche automatiche."
            )
            messagebox.showwarning("Attenzione", warning_text, parent=self)
            import subprocess
            subprocess.Popen(["notepad.exe", config_path])
        except Exception as e:
            logger.error(f"Errore apertura file config NPI: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile aprire il file:\n{e}", parent=self)

    def load_npi_projects(self):
        self._update_statistics()
        for i in self.project_tree.get_children():
            self.project_tree.delete(i)

        try:
            # Ottieni anno selezionato
            year_str = self.year_filter_var.get()
            year_filter = None if year_str == "Tutti gli anni" else int(year_str)
            
            # Ottieni cliente selezionato
            client_str = self.client_filter_var.get()
            client_filter = None if client_str == "Tutti i clienti" else client_str
            
            progetti = self.npi_manager.get_dashboard_projects(year_filter=year_filter, client_filter=client_filter)
            if not progetti:
                self.project_tree.insert('', tk.END, text="-1", values=('', self.lang.get('no_active_projects',
                                                                                          "Nessun progetto attivo trovato."),
                                                                        "", "", ""))
                return

            # Applica filtro stato
            status_filter = self.status_filter_var.get() if hasattr(self, 'status_filter_var') else "Tutti i progetti"
            
            filtered_count = 0
            for proj in progetti:
                # Recupera statistiche task PRIMA di applicare il filtro
                # (necessario per determinare se il progetto è in ritardo)
                try:
                    stats = self.npi_manager.get_project_task_statistics(proj.ProgettoId)
                    # I progetti chiusi hanno sempre 100% di completamento
                    if proj.StatoProgetto == 'Chiuso':
                        stats['completion_percentage'] = 100
                except Exception as e:
                    logger.error(f"Errore nel recupero statistiche task per progetto {proj.ProgettoId}: {e}")
                    stats = {
                        'total_tasks': 0,
                        'completed_on_time': 0,
                        'completed_late': 0,
                        'pending_late': 0,
                        'completion_percentage': 0
                    }
                
                # Applica filtro stato basato sui task
                if status_filter == "Solo chiusi":
                    if proj.StatoProgetto != 'Chiuso':
                        continue
                elif status_filter == "Solo in ritardo":
                    # Progetti con task in ritardo (pending_late > 0)
                    if proj.StatoProgetto == 'Chiuso':
                        continue
                    if stats['pending_late'] == 0:
                        logger.debug(f"Progetto {proj.ProgettoId} '{proj.NomeProgetto}' saltato: nessun task in ritardo")
                        continue
                    logger.debug(f"Progetto {proj.ProgettoId} '{proj.NomeProgetto}' IN RITARDO: {stats['pending_late']} task in ritardo")
                elif status_filter == "Solo on track":
                    # Progetti attivi senza task in ritardo
                    if proj.StatoProgetto == 'Chiuso':
                        continue
                    if stats['pending_late'] > 0:
                        continue
                # "Tutti i progetti" - no filtering
                
                filtered_count += 1
                
                # Recupera la data fine progetto dall'oggetto progetto
                project_end_date = getattr(proj, 'ScadenzaProgetto', None)
                display_date = project_end_date.strftime('%d/%m/%Y') if project_end_date else "N/D"
                status_icon = ''
                row_tags = []

                # Gestisce sia datetime.date che datetime.datetime
                # I progetti chiusi non sono mai "in ritardo"
                if project_end_date and proj.StatoProgetto != 'Chiuso':
                    # Converti in date se necessario
                    end_date = project_end_date.date() if hasattr(project_end_date, 'date') else project_end_date
                    is_overdue = end_date < datetime.now().date()
                else:
                    is_overdue = False
                    
                if is_overdue:
                    status_icon = 'X'
                    row_tags.append('overdue')
                
                # Progetti chiusi in verde
                if proj.StatoProgetto == 'Chiuso':
                    row_tags.append('closed')
                
                # Le statistiche sono già state recuperate sopra per il filtro
                
                # Evidenzia cliente selezionato
                if client_filter and proj.Cliente == client_filter:
                    row_tags.append('selected_client')
                
                # Tag per percentuale completamento (solo se non in ritardo)
                if not is_overdue:
                    pct = stats['completion_percentage']
                    # Non aggiungiamo tag colore qui, lo gestiamo con item configuration

                self.project_tree.insert(
                    '', tk.END,
                    text=str(proj.ProgettoId),
                    values=(
                        status_icon,
                        proj.NomeProgetto,
                        proj.CodiceProdotto or "",
                        proj.Cliente or "",
                        display_date,
                        stats['total_tasks'],
                        stats['completed_on_time'],
                        stats['completed_late'],
                        stats['pending_late'],
                        f"{stats['completion_percentage']}%"
                    ),
                    tags=tuple(row_tags)
                )
            
            # Log risultati filtro
            if status_filter != "Tutti i progetti":
                logger.info(f"Filtro '{status_filter}': {filtered_count} progetti su {len(progetti)} totali")
        except Exception as e:
            logger.error(f"Errore nel caricamento progetti dashboard: {e}", exc_info=True)
            self.project_tree.insert('', tk.END, text="-1",
                                     values=('', "Errore", "Impossibile caricare i dati.", str(e), ""))

    def _on_double_click(self, event):
        """Il doppio click ora lancia la finestra di analisi."""
        self._launch_analysis_window()

    def _get_selected_project_info(self):
        """Helper per recuperare ID e nome del progetto selezionato."""
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showwarning("Selezione", "Seleziona un progetto dalla lista.", parent=self)
            return None, None

        item_iid = selection[0]
        project_id = self.project_tree.item(item_iid, 'text')
        values = self.project_tree.item(item_iid, 'values')

        try:
            project_id = int(project_id)
            project_name = values[0]  # Il nome progetto è il primo valore della riga
            return project_id, project_name
        except (ValueError, IndexError):
            messagebox.showwarning("Selezione", "Seleziona un progetto valido dalla lista.", parent=self)
            return None, None

    def _launch_analysis_window(self):
        """Lancia la nuova finestra di analisi."""
        project_id, project_name = self._get_selected_project_info()
        if project_id is None:
            return

        ProjectAnalysisWindow(self, self.npi_manager, self.lang, project_id, project_name)

    def _show_project_context_menu(self, event):
        iid = self.project_tree.identify_row(event.y)
        if not iid: return

        # Seleziona la riga cliccata
        self.project_tree.selection_set(iid)

        project_id, _ = self._get_selected_project_info()
        if project_id is None: return

        # Recupera info progetto per determinare se è chiuso
        try:
            from npi.data_models import ProgettoNPI
            session = self.npi_manager._get_session()
            try:
                project_obj = session.get(ProgettoNPI, project_id)
                is_closed = project_obj and project_obj.StatoProgetto == 'Chiuso'
                logger.debug(f"Progetto {project_id}: StatoProgetto = {project_obj.StatoProgetto if project_obj else 'NOT FOUND'}, is_closed = {is_closed}")
            finally:
                session.close()
        except Exception as e:
            logger.error(f"Errore nel recupero stato progetto {project_id}: {e}")
            is_closed = False
        
        context_menu = tk.Menu(self, tearoff=0)
        context_menu.add_command(label=self.lang.get('manage_project', "Gestisci Dettagli Task..."),
                                 command=self._launch_project_window)
        context_menu.add_command(label=self.lang.get('npi_view_gantt', "Visualizza Gantt..."),
                                 command=self._launch_gantt_window)
        
        # Aggiungi opzione "Riapri Progetto" solo per progetti chiusi
        if is_closed:
            logger.debug(f"Aggiunta opzione 'Riapri Progetto' al menu per progetto {project_id}")
            context_menu.add_separator()
            context_menu.add_command(
                label=self.lang.get('npi_reopen_project', "Riapri Progetto"),
                command=self._reopen_project_from_dashboard
            )
        else:
            logger.debug(f"Opzione 'Riapri Progetto' NON aggiunta (progetto non chiuso)")
        
        context_menu.post(event.x_root, event.y_root)

    def _launch_project_window(self):
        """Apre la finestra di gestione dettaglio task."""
        project_id, _ = self._get_selected_project_info()
        if project_id is None: return

        logged_user = self.logged_in_user

        # Usa l'helper di autorizzazione del master se esiste
        if hasattr(self.master_app, '_execute_authorized_action'):
            self.master_app._execute_authorized_action(
                menu_translation_key='project_window',
                action_callback=lambda: ProjectWindow(self, self.npi_manager, self.lang, project_id, self.master_app,
                                                      logged_user)
            )
        else:  # Fallback
            ProjectWindow(self, self.npi_manager, self.lang, project_id, self.master_app, logged_user)

    def _reopen_project_from_dashboard(self):
        """Riapre un progetto chiuso dal menu contestuale del dashboard."""
        project_id, project_name = self._get_selected_project_info()
        if project_id is None:
            return
        
        # Conferma dall'utente
        confirm = messagebox.askyesno(
            self.lang.get('npi_reopen_project', "Riapri Progetto"),
            self.lang.get('npi_reopen_project_confirm', 
                         f"Sei sicuro di voler riaprire il progetto '{project_name}'?\n\n"
                         f"Il progetto tornerà allo stato 'Attivo'."),
            parent=self
        )
        
        if not confirm:
            return
        
        try:
            success = self.npi_manager.reopen_project(project_id)
            if success:
                messagebox.showinfo(
                    self.lang.get('success', "Successo"),
                    self.lang.get('npi_reopen_project_success', 
                                 f"Progetto '{project_name}' riaperto con successo."),
                    parent=self
                )
                # Ricarica la lista progetti per aggiornare la visualizzazione
                self.load_npi_projects()
            else:
                messagebox.showerror(
                    self.lang.get('error', "Errore"),
                    self.lang.get('npi_reopen_project_failed', 
                                 "Impossibile riaprire il progetto."),
                    parent=self
                )
        except Exception as e:
            logger.error(f"Errore nella riapertura del progetto {project_id}: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', "Errore"),
                f"Errore: {str(e)}",
                parent=self
            )

    def _launch_gantt_window(self):
        """Apre la finestra Gantt per il progetto selezionato."""
        project_id, _ = self._get_selected_project_info()
        if project_id is None: return

        NpiGanttWindow(self, self.npi_manager, self.lang, project_id)

    def _export_to_excel_new(self):
        """Esporta i progetti NPI in un file Excel completo con statistiche task e tab per cliente."""
        self.export_excel_button.config(state=tk.DISABLED)
        self.update_idletasks()
        
        try:
            # Ottieni filtri correnti
            year_str = self.year_filter_var.get()
            year_filter = None if year_str == "Tutti gli anni" else int(year_str)
            
            client_str = self.client_filter_var.get()
            client_filter = None if client_str == "Tutti i clienti" else client_str
            
            # Chiama il nuovo metodo di export completo
            file_path = self.npi_manager.export_npi_to_excel_comprehensive(
                year_filter=year_filter,
                client_filter=client_filter
            )
            
            if messagebox.askyesno("Successo", 
                                  f"Report Excel salvato con successo in:\n{file_path}\n\nVuoi aprirlo ora?",
                                  parent=self):
                os.startfile(file_path)
                
        except ImportError as e:
            messagebox.showerror("Libreria Mancante",
                               f"{str(e)}",
                               parent=self)
        except ValueError as e:
            messagebox.showinfo("Info", str(e), parent=self)
        except Exception as e:
            logger.error(f"Errore durante l'export Excel: {e}", exc_info=True)
            messagebox.showerror("Errore Export", f"Impossibile generare il file Excel:\n{e}", parent=self)
        finally:
            self.export_excel_button.config(state=tk.NORMAL)

    def _export_to_excel(self):
        """Esporta i progetti NPI in un file Excel formattato con logo."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            messagebox.showerror("Libreria Mancante",
                               "La libreria 'openpyxl' è necessaria per l'export Excel.\nInstallala con: pip install openpyxl",
                               parent=self)
            return
        
        self.export_excel_button.config(state=tk.DISABLED)
        self.update_idletasks()
        
        try:
            # Ottieni anno selezionato
            year_str = self.year_filter_var.get()
            year_filter = None if year_str == "Tutti gli anni" else int(year_str)
            year_label = year_str if year_str != "Tutti gli anni" else "All Years"
            
            # Recupera progetti
            progetti = self.npi_manager.get_dashboard_projects(year_filter=year_filter)
            if not progetti:
                messagebox.showinfo("Info", "Nessun progetto da esportare.", parent=self)
                return
            
            # Crea directory C:\Temp se non esiste
            temp_dir = "C:\\Temp"
            os.makedirs(temp_dir, exist_ok=True)
            
            # Crea workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NPI Projects Report"
            
            # Stili
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=16, color="0078D4")
            subtitle_font = Font(size=10, color="666666")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row = 1
            
            # Logo (se esiste)
            logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "logo.png")
            if os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    img.width = 120
                    img.height = 40
                    ws.add_image(img, 'A1')
                    current_row = 4
                except Exception as e:
                    logger.warning(f"Impossibile caricare logo: {e}")
            
            # Titolo
            ws.merge_cells(f'A{current_row}:H{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "NPI Projects Report - Active and Closed Projects"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Sottotitolo
            ws.merge_cells(f'A{current_row}:H{current_row}')
            subtitle_cell = ws[f'A{current_row}']
            subtitle_cell.value = f"Year: {year_label} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='center')
            current_row += 2
            
            # Header tabella
            headers = ["Project ID", "Project Name", "Product Code", "Customer", "Status", 
                      "Start Date", "End Date (Deadline)", "Days Overdue"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            current_row += 1
            
            # Dati progetti
            today = datetime.now().date()
            for proj in progetti:
                # Status
                status = "Closed" if proj.StatoProgetto == "Chiuso" else "Active"
                
                # Date
                start_date = proj.DataInizio.strftime('%Y-%m-%d') if proj.DataInizio else "N/A"
                end_date = proj.ScadenzaProgetto.strftime('%Y-%m-%d') if proj.ScadenzaProgetto else "N/A"
                
                # Days overdue
                days_overdue = ""
                if status == "Active" and proj.ScadenzaProgetto:
                    deadline = proj.ScadenzaProgetto.date() if hasattr(proj.ScadenzaProgetto, 'date') else proj.ScadenzaProgetto
                    if deadline < today:
                        days_overdue = (today - deadline).days
                
                # Scrivi riga
                row_data = [
                    proj.ProgettoId,
                    proj.NomeProgetto,
                    proj.CodiceProdotto or "",
                    proj.Cliente or "",
                    status,
                    start_date,
                    end_date,
                    days_overdue
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left' if col_idx <= 4 else 'center')
                    
                    # Formattazione condizionale
                    if status == "Closed":
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif days_overdue:
                        cell.font = Font(color="FF0000", bold=True)
                
                current_row += 1
            
            # Auto-width colonne
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Salva file
            file_path = os.path.join(temp_dir, f"NPI_Projects_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(file_path)
            
            if messagebox.askyesno("Successo", 
                                  f"Report Excel salvato con successo in:\n{file_path}\n\nVuoi aprirlo ora?",
                                  parent=self):
                os.startfile(file_path)
                
        except Exception as e:
            logger.error(f"Errore durante l'export Excel: {e}", exc_info=True)
            messagebox.showerror("Errore Export", f"Impossibile generare il file Excel:\n{e}", parent=self)
        finally:
            self.export_excel_button.config(state=tk.NORMAL)

    def _export_overview_report(self):
        """Esporta un report panoramico NPI (senza dettaglio task)."""
        self.export_overview_button.config(state=tk.DISABLED)
        self.update_idletasks()
        
        try:
            # Ottieni filtri correnti
            year_str = self.year_filter_var.get()
            year_filter = None if year_str == "Tutti gli anni" else int(year_str)
            
            client_str = self.client_filter_var.get()
            client_filter = None if client_str == "Tutti i clienti" else client_str
            
            file_path = self.npi_manager.export_npi_overview_report(
                year_filter=year_filter,
                client_filter=client_filter
            )
            
            if messagebox.askyesno("Successo",
                                   f"Report panoramico salvato con successo in:\n{file_path}\n\nVuoi aprirlo ora?",
                                   parent=self):
                os.startfile(file_path)
                
        except ImportError as e:
            messagebox.showerror("Libreria Mancante",
                               f"{str(e)}",
                               parent=self)
        except ValueError as e:
            messagebox.showinfo("Info", str(e), parent=self)
        except Exception as e:
            logger.error(f"Errore durante l'export report panoramico: {e}", exc_info=True)
            messagebox.showerror("Errore Export", f"Impossibile generare il report:\n{e}", parent=self)
        finally:
            self.export_overview_button.config(state=tk.NORMAL)

    def _export_summary_pdf(self):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Libreria Mancante",
                                 "La libreria 'reportlab' è necessaria per l'export PDF.\nInstallala con: pip install reportlab",
                                 parent=self)
            return
        self.export_pdf_button.config(state=tk.DISABLED)
        self.update_idletasks()
        try:
            full_report_data = self.npi_manager.get_full_projects_report_data()
            if not full_report_data:
                messagebox.showinfo("Info", "Nessun progetto da includere nel report.", parent=self)
                return
            temp_dir = "C:\\Temp"
            os.makedirs(temp_dir, exist_ok=True)
            file_path = os.path.join(temp_dir, f"NPI_Projects_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
            c = canvas.Canvas(file_path, pagesize=A4)
            width, height = A4

            def draw_project_section(y_start, project_data):
                y = y_start
                c.setFont("Helvetica-Bold", 12)
                is_overdue = project_data['info'].ScadenzaMilestoneFinale and project_data[
                    'info'].ScadenzaMilestoneFinale.date() < datetime.now().date()
                if is_overdue:
                    c.setFillColor(colors.red)
                c.drawString(inch, y, f"Project: {project_data['info'].NomeProgetto}")
                c.setFillColor(colors.black)
                y -= 20
                c.setFont("Helvetica", 9)
                c.drawString(inch, y,
                             f"Product Code: {project_data['info'].CodiceProdotto or 'N/A'}  |  Customer: {project_data['info'].Cliente or 'N/A'}")
                y -= 15
                due_date_str = project_data['info'].ScadenzaMilestoneFinale.strftime('%Y-%m-%d') if project_data[
                    'info'].ScadenzaMilestoneFinale else "Not Set"
                c.drawString(inch, y, f"Final Milestone Due Date: {due_date_str}")
                y -= 25
                c.setFont("Helvetica-Bold", 10)
                c.drawString(inch, y, "Overdue Task Analysis:")
                y -= 20
                analysis = project_data['analysis']
                if not analysis:
                    c.setFont("Helvetica-Oblique", 9)
                    c.setFillColor(colors.green)
                    c.drawString(inch, y, "No overdue tasks found.")
                    c.setFillColor(colors.black)
                    y -= 20
                else:
                    table_data = [['Owner', '# Late Tasks']]
                    for item in analysis:
                        table_data.append([item['owner_name'], item['late_count']])
                    table = Table(table_data, colWidths=[3 * inch, 1.5 * inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))

                    # --- ** CORREZIONE QUI ** ---
                    # Cattura l'altezza reale calcolata dal metodo wrapOn
                    _w, _h = table.wrapOn(c, width, height)

                    # Usa l'altezza catturata _h per il posizionamento e l'aggiornamento della coordinata y
                    table.drawOn(c, inch, y - _h)
                    y -= (_h + 20)
                    # --- ** FINE CORREZIONE ** ---

                c.line(inch, y, width - inch, y)
                y -= 15
                return y

            y_pos = height - inch
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(width / 2.0, y_pos, "NPI Projects Summary Report")
            y_pos -= 20
            c.setFont("Helvetica", 10)
            c.drawCentredString(width / 2.0, y_pos, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            y_pos -= 40
            for project in full_report_data:
                estimated_height = 120 + (len(project['analysis']) * 20 if project['analysis'] else 0)
                if y_pos < estimated_height:
                    c.showPage()
                    y_pos = height - inch
                y_pos = draw_project_section(y_pos, project)
            c.save()
            if messagebox.askyesno("Successo", f"Report PDF salvato con successo in:\n{file_path}\n\nVuoi aprirlo ora?",
                                   parent=self):
                os.startfile(file_path)
        except Exception as e:
            logger.error(f"Errore durante l'export PDF: {e}", exc_info=True)
            messagebox.showerror("Errore Export", f"Impossibile generare il PDF:\n{e}", parent=self)
        finally:
            self.export_pdf_button.config(state=tk.NORMAL)
