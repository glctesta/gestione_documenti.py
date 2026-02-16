# File: npi/npi_manager.py
import logging
import os
from datetime import datetime
from sqlalchemy.orm import sessionmaker, joinedload, subqueryload, selectinload, scoped_session
from sqlalchemy.engine import Engine
from sqlalchemy import select, update, delete, func, and_, or_, event, text, case
from sqlalchemy.pool import QueuePool
from npi.data_models import (ProgettoNPI, Base, Prodotto, Soggetto, TaskCatalogo,
                             Categoria, WaveNPI, TaskProdotto, NpiDocument, NpiDocumentType, TaskDependency,
                             FamilyNpi, FamilyNpiLog)
import urllib

logger = logging.getLogger(__name__)

class GestoreNPI:
    """
    Classe principale che orchestra la gestione dei dati NPI tramite il DB.
    """

    def __init__(self, engine: Engine):
        if not engine:
            raise ValueError("SQLAlchemy Engine non può essere None.")

        self.engine = engine

        # Crea le tabelle se non esistono
        Base.metadata.create_all(self.engine, checkfirst=True)
        # Assicura colonne default per categorie/task
        self._ensure_default_flags_schema()

        # Configura sessionmaker - NON usare scoped_session che causa problemi
        self.session_factory = sessionmaker(
            bind=self.engine,
            expire_on_commit=False,
            autoflush=True,
            autocommit=False
        )

        logger.info("GestoreNPI pronto e tabelle NPI verificate/create.")

    def _ensure_default_flags_schema(self):
        """Assicura che le colonne DefaultCategory e DefaultTask esistano."""
        try:
            with self.engine.begin() as conn:
                conn.execute(text("""
                IF COL_LENGTH('dbo.Categories', 'DefaultCategory') IS NULL
                BEGIN
                    ALTER TABLE dbo.Categories
                    ADD DefaultCategory bit NOT NULL CONSTRAINT DF_Categories_DefaultCategory DEFAULT(0);
                END
                """))
                conn.execute(text("""
                IF COL_LENGTH('dbo.TaskCatalogo', 'DefaultTask') IS NULL
                BEGIN
                    ALTER TABLE dbo.TaskCatalogo
                    ADD DefaultTask bit NOT NULL CONSTRAINT DF_TaskCatalogo_DefaultTask DEFAULT(0);
                END
                """))
        except Exception as e:
            logger.error(f"Errore aggiornamento schema colonne default: {e}", exc_info=True)

    def _setup_connection_events(self):
        """Setup event listeners per gestione connessioni (opzionale per debugging)."""

        @event.listens_for(self.engine, "connect")
        def receive_connect(dbapi_conn, connection_record):
            """Chiamato quando viene creata una nuova connessione."""
            logger.debug("Nuova connessione database stabilita")

        @event.listens_for(self.engine, "close")
        def receive_close(dbapi_conn, connection_record):
            """Chiamato quando una connessione viene chiusa."""
            logger.debug("Connessione database chiusa")

    # Sostituisci il metodo _get_session con questo:

    def _get_session(self):
        """Crea una nuova sessione per ogni operazione."""
        try:
            session = self.session_factory()
            # Test semplice della connessione
            session.execute(text("SELECT 1"))
            return session
        except Exception as e:
            logger.error(f"Errore nella creazione della sessione: {e}")
            raise

    def _detach_object(self, session, obj):
        """Helper per rendere un oggetto indipendente dalla sessione."""
        if obj:
            session.expunge(obj)
        return obj

    def _detach_list(self, session, obj_list):
        """Helper per rendere una lista di oggetti indipendente dalla sessione."""
        for obj in obj_list:
            session.expunge(obj)
        return obj_list

    def get_soggetti(self):
        """Recupera tutti i soggetti ordinati per nome."""
        session = self._get_session()
        try:
            soggetti = session.scalars(
                select(Soggetto).order_by(Soggetto.Nome)
            ).all()
            return self._detach_list(session, soggetti)
        except Exception as e:
            logger.error(f"Errore in get_soggetti: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_soggetto_by_id(self, soggetto_id):
        """Recupera un soggetto per ID."""
        session = self._get_session()
        try:
            soggetto = session.scalars(
                select(Soggetto).where(Soggetto.SoggettoId == soggetto_id)
            ).first()
            return self._detach_object(session, soggetto)
        except Exception as e:
            logger.error(f"Errore in get_soggetto_by_id: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def create_soggetto(self, data):
        """Crea un nuovo soggetto."""
        session = self._get_session()
        try:
            nuovo_soggetto = Soggetto(**data)
            session.add(nuovo_soggetto)
            session.commit()
            session.refresh(nuovo_soggetto)
            return self._detach_object(session, nuovo_soggetto)
        except Exception as e:
            logger.error(f"Errore in create_soggetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_soggetto(self, soggetto_id, data):
        """Aggiorna un soggetto esistente."""
        session = self._get_session()
        try:
            soggetto = session.get(Soggetto, soggetto_id)
            if soggetto:
                for key, value in data.items():
                    setattr(soggetto, key, value)
                session.commit()
                session.refresh(soggetto)
                return self._detach_object(session, soggetto)
            return None
        except Exception as e:
            logger.error(f"Errore in update_soggetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def delete_soggetto(self, soggetto_id):
        """Elimina un soggetto."""
        session = self._get_session()
        try:
            soggetto = session.get(Soggetto, soggetto_id)
            if soggetto:
                session.delete(soggetto)
                session.commit()
                return True
            return False
        except Exception as e:
            logger.error(f"Errore in delete_soggetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()


    def get_prodotti(self):
        """Recupera tutti i prodotti ordinati per nome."""
        session = self._get_session()
        try:
            prodotti = session.scalars(
                select(Prodotto).order_by(Prodotto.NomeProdotto)
            ).all()

            return self._detach_list(session, prodotti)
        except Exception as e:
            logger.error(f"Errore in get_prodotti: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_prodotto_by_id(self, prodotto_id):
        """Recupera un prodotto per ID."""
        session = self._get_session()
        try:
            prodotto = session.scalars(
                select(Prodotto).where(Prodotto.ProdottoID == prodotto_id)
            ).first()

            return self._detach_object(session, prodotto)
        except Exception as e:
            logger.error(f"Errore in get_prodotto_by_id: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def create_prodotto(self, data):
        """Crea un nuovo prodotto."""
        session = self._get_session()
        try:
            nuovo_prodotto = Prodotto(**data)
            session.add(nuovo_prodotto)
            session.commit()
            session.refresh(nuovo_prodotto)

            return self._detach_object(session, nuovo_prodotto)
        except Exception as e:
            logger.error(f"Errore in create_prodotto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_prodotto(self, prodotto_id, data):
        """Aggiorna un prodotto esistente."""
        session = self._get_session()
        try:
            prodotto = session.get(Prodotto, prodotto_id)
            if prodotto:
                for key, value in data.items():
                    setattr(prodotto, key, value)
                session.commit()
                session.refresh(prodotto)
                return self._detach_object(session, prodotto)
            return None
        except Exception as e:
            logger.error(f"Errore in update_prodotto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def delete_prodotto(self, prodotto_id):
        """Elimina un prodotto."""
        session = self._get_session()
        try:
            prodotto = session.get(Prodotto, prodotto_id)
            if prodotto:
                session.delete(prodotto)
                session.commit()
                return True
            return False
        except Exception as e:
            logger.error(f"Errore in delete_prodotto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

            # --- METODI CRUD PER CATEGORIE ---

    def get_categories(self):
        """Recupera tutte le categorie ordinate per NrOrdin."""
        session = self._get_session()
        try:
            categories = session.scalars(
                select(Categoria).order_by(Categoria.NrOrdin)
            ).all()

            return self._detach_list(session, categories)
        except Exception as e:
            logger.error(f"Errore in get_categories: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_category_default_flag(self, category_id: int, is_default: bool):
        """Aggiorna il flag DefaultCategory per una categoria."""
        session = self._get_session()
        try:
            category = session.get(Categoria, category_id)
            if not category:
                return False
            category.DefaultCategory = bool(is_default)
            session.commit()
            return True
        except Exception as e:
            logger.error(f"Errore in update_category_default_flag: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_category_by_id(self, category_id):
        """Recupera una categoria per ID."""
        session = self._get_session()
        try:
            category = session.scalars(
                select(Categoria).where(Categoria.CategoryId == category_id)
            ).first()

            return self._detach_object(session, category)
        except Exception as e:
            logger.error(f"Errore in get_category_by_id: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_tasks_by_category(self, category_id):
        """
        Recupera tutti i task del catalogo per una specifica categoria.
        
        Args:
            category_id: ID della categoria
            
        Returns:
            Lista di TaskCatalogo ordinati per NrOrdin
        """
        from sqlalchemy.orm import selectinload
        
        session = self._get_session()
        try:
            tasks = session.scalars(
                select(TaskCatalogo).options(
                    selectinload(TaskCatalogo.categoria)
                ).where(
                    TaskCatalogo.CategoryId == category_id
                ).order_by(TaskCatalogo.NrOrdin)
            ).all()
            
            return self._detach_list(session, tasks)
        except Exception as e:
            logger.error(f"Errore in get_tasks_by_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def is_category_used_in_projects(self, category_id):
        """
        Verifica se una categoria è stata utilizzata in almeno un progetto.
        
        Args:
            category_id: ID della categoria
            
        Returns:
            True se la categoria è usata, False altrimenti
        """
        session = self._get_session()
        try:
            # Query per verificare se esistono TaskProdotto con task di questa categoria
            count = session.scalar(
                select(func.count(TaskProdotto.TaskProdottoID)).join(
                    TaskCatalogo, TaskProdotto.TaskID == TaskCatalogo.TaskID
                ).where(
                    TaskCatalogo.CategoryId == category_id
                )
            )
            
            return count > 0
        except Exception as e:
            logger.error(f"Errore in is_category_used_in_projects: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def create_category(self, data):
        """Crea una nuova categoria con controllo duplicati NrOrdin."""
        session = self._get_session()
        try:
            # Controllo duplicati di NrOrdin prima di inserire
            nr_ordin_value = data.get('NrOrdin')
            if nr_ordin_value is not None:
                existing = session.scalars(
                    select(Categoria).where(Categoria.NrOrdin == nr_ordin_value)
                ).first()
                if existing:
                    raise ValueError(
                        f"Il numero d'ordine {nr_ordin_value} è già utilizzato "
                        f"dalla categoria '{existing.Category}'."
                    )

            new_category = Categoria(**data)
            session.add(new_category)
            session.commit()
            session.refresh(new_category)

            return self._detach_object(session, new_category)
        except Exception as e:
            logger.error(f"Errore in create_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_category(self, category_id, data):
        """Aggiorna una categoria con controllo duplicati NrOrdin."""
        session = self._get_session()
        try:
            # Controllo duplicati di NrOrdin, escludendo la categoria corrente
            nr_ordin_value = data.get('NrOrdin')
            if nr_ordin_value is not None:
                existing = session.scalars(
                    select(Categoria).where(
                        and_(
                            Categoria.NrOrdin == nr_ordin_value,
                            Categoria.CategoryId != category_id
                        )
                    )
                ).first()
                if existing:
                    raise ValueError(
                        f"Il numero d'ordine {nr_ordin_value} è già utilizzato "
                        f"dalla categoria '{existing.Category}'."
                    )

            category = session.get(Categoria, category_id)
            if category:
                for key, value in data.items():
                    setattr(category, key, value)
                session.commit()
                session.refresh(category)
                return self._detach_object(session, category)
            return None
        except Exception as e:
            logger.error(f"Errore in update_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def delete_category(self, category_id):
        """Elimina una categoria con controllo utilizzo."""
        session = self._get_session()
        try:
            # Controlla se la categoria è associata a qualche task
            num_tasks = session.scalar(
                select(func.count(TaskCatalogo.TaskID)).where(
                    TaskCatalogo.CategoryId == category_id
                )
            )
            if num_tasks > 0:
                raise Exception(
                    f"Impossibile eliminare la categoria: è associata a {num_tasks} "
                    f"task nel catalogo."
                )

            category = session.get(Categoria, category_id)
            if category:
                session.delete(category)
                session.commit()
                return True
            return False
        except Exception as e:
            logger.error(f"Errore in delete_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_catalogo_task(self):
        """Recupera tutti i task del catalogo con categoria."""
        session = self._get_session()
        try:
            stmt = select(TaskCatalogo).options(
                selectinload(TaskCatalogo.categoria)
            ).order_by(TaskCatalogo.NrOrdin)

            tasks = session.scalars(stmt).all()

            return self._detach_list(session, tasks)
        except Exception as e:
            logger.error(f"Errore in get_catalogo_task: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_catalogo_task_by_category(self, category_id: int):
        """Recupera i task del catalogo filtrati per categoria."""
        session = self._get_session()
        try:
            tasks = session.scalars(
                select(TaskCatalogo)
                .where(TaskCatalogo.CategoryId == category_id)
                .order_by(TaskCatalogo.NrOrdin)
            ).all()
            return self._detach_list(session, tasks)
        except Exception as e:
            logger.error(f"Errore in get_catalogo_task_by_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_task_default_flag(self, task_id: int, is_default: bool):
        """Aggiorna il flag DefaultTask per un task del catalogo."""
        session = self._get_session()
        try:
            task = session.get(TaskCatalogo, task_id)
            if not task:
                return False
            task.DefaultTask = bool(is_default)
            session.commit()
            return True
        except Exception as e:
            logger.error(f"Errore in update_task_default_flag: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def set_default_tasks_for_category(self, category_id: int, is_default: bool):
        """Imposta DefaultTask per tutti i task della categoria."""
        session = self._get_session()
        try:
            session.execute(
                update(TaskCatalogo)
                .where(TaskCatalogo.CategoryId == category_id)
                .values(DefaultTask=bool(is_default))
            )
            session.commit()
            return True
        except Exception as e:
            logger.error(f"Errore in set_default_tasks_for_category: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_catalogo_task_by_id(self, task_id):
        """Recupera un task del catalogo per ID."""
        session = self._get_session()
        try:
            stmt = select(TaskCatalogo).options(
                selectinload(TaskCatalogo.categoria)
            ).where(TaskCatalogo.TaskID == task_id)

            task = session.scalars(stmt).first()

            return self._detach_object(session, task)
        except Exception as e:
            logger.error(f"Errore in get_catalogo_task_by_id: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_catalogo_task_by_itemid(self, item_id: str):
        """Recupera un task del catalogo per ItemID."""
        session = self._get_session()
        try:
            task = session.scalars(
                select(TaskCatalogo).where(TaskCatalogo.ItemID == item_id)
            ).first()

            return self._detach_object(session, task)
        except Exception as e:
            logger.error(f"Errore in get_catalogo_task_by_itemid: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def _renumber_all_tasks(self, session):
        """Rinumera tutti i task con un passo di 10."""
        logger.info("Rinumerazione in corso per tutti i task del catalogo...")
        all_tasks = session.scalars(
            select(TaskCatalogo).order_by(TaskCatalogo.NrOrdin, TaskCatalogo.ItemID)
        ).all()

        for i, task in enumerate(all_tasks):
            task.NrOrdin = (i + 1) * 10

        session.flush()
        logger.info("Rinumerazione completata.")

    def create_catalogo_task(self, data, anchor_task_id=None):
        """
        Crea un nuovo task nel catalogo con numerazione gerarchica.
        
        La numerazione segue il pattern: (NrOrdin_Categoria * 100) + numero_task
        I task nella stessa categoria avanzano di 5 in 5.
        
        Esempio:
        - Categoria con NrOrdin=10: task avranno 1005, 1010, 1015, ...
        - Categoria con NrOrdin=20: task avranno 2005, 2010, 2015, ...
        """
        session = self._get_session()
        try:
            category_id = data.get('CategoryId')
            if not category_id:
                raise ValueError("CategoryId è obbligatorio per creare un task")
            
            # Ottieni il NrOrdin della categoria
            cat_ordin = session.scalar(
                select(Categoria.NrOrdin).where(
                    Categoria.CategoryId == category_id
                )
            )
            if cat_ordin is None:
                raise ValueError(f"Categoria con ID {category_id} non trovata")
            
            # Calcola la base per questa categoria (es: 10 -> 1000, 20 -> 2000)
            category_base = cat_ordin * 100
            
            # Trova il massimo NrOrdin esistente per questa categoria
            max_ordin_in_category = session.scalar(
                select(func.max(TaskCatalogo.NrOrdin)).where(
                    TaskCatalogo.CategoryId == category_id
                )
            )
            
            # Calcola il nuovo NrOrdin
            if max_ordin_in_category is None:
                # Primo task di questa categoria
                new_ordin = category_base + 5
            else:
                # Incrementa di 5 rispetto all'ultimo task della categoria
                new_ordin = max_ordin_in_category + 5
            
            data['NrOrdin'] = new_ordin
            
            logger.info(f"Creazione task in categoria {category_id} (NrOrdin cat: {cat_ordin}): nuovo NrOrdin = {new_ordin}")

            nuovo = TaskCatalogo(**data)
            session.add(nuovo)
            session.commit()
            session.refresh(nuovo)

            return self._detach_object(session, nuovo)
        except Exception as e:
            logger.error(f"Errore in create_catalogo_task: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def update_catalogo_task(self, task_id, data):
        """
        Aggiorna un task del catalogo.
        
        Se viene modificato il NrOrdin, verifica che non esista già 
        un altro task con lo stesso NrOrdin nella stessa categoria.
        """
        session = self._get_session()
        try:
            task = session.get(TaskCatalogo, task_id)
            if not task:
                return None
            
            # Se viene modificato il NrOrdin, verifica duplicati nella categoria
            if 'NrOrdin' in data and data['NrOrdin'] != task.NrOrdin:
                new_ordin = data['NrOrdin']
                category_id = data.get('CategoryId', task.CategoryId)
                
                # Verifica se esiste già un task con questo NrOrdin nella stessa categoria
                existing_task = session.scalar(
                    select(TaskCatalogo).where(
                        and_(
                            TaskCatalogo.CategoryId == category_id,
                            TaskCatalogo.NrOrdin == new_ordin,
                            TaskCatalogo.TaskID != task_id
                        )
                    )
                )
                
                if existing_task:
                    raise ValueError(
                        f"Il numero d'ordine {new_ordin} è già utilizzato "
                        f"dal task '{existing_task.NomeTask}' nella stessa categoria."
                    )
            
            # Applica le modifiche
            for key, value in data.items():
                setattr(task, key, value)
            
            session.commit()
            session.refresh(task)
            return self._detach_object(session, task)
            
        except Exception as e:
            logger.error(f"Errore in update_catalogo_task: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def delete_catalogo_task(self, task_id):
        """Elimina un task del catalogo."""
        session = self._get_session()
        try:
            task = session.get(TaskCatalogo, task_id)
            if task:
                session.delete(task)
                session.commit()
                return True
            return False
        except Exception as e:
            logger.error(f"Errore in delete_catalogo_task: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_progetti_attivi(self):
        """
        Restituisce una lista di tutti i progetti NPI con stato 'Attivo'.
        Utilizza una query SQL custom per formattare i dati nel formato richiesto:
        Cliente -> CodiceProdotto (Version: X) NomeProdotto [Status]
        Include il conteggio dei task assegnati.
        """
        session = self._get_session()
        try:
            # Query SQL personalizzata con conteggio task
            sql_query = """
                SELECT n.[ProgettoID],
                       p.Cliente + ' -> ' + p.CodiceProdotto + ' (Version: ' + ISNULL(n.[version],'#N/D') + ') ' + p.NomeProdotto
                       + ' [' + IIF(COUNT(t.[TaskProdottoID])=0, 'Not Started', 'Assigned task #' + CAST(COUNT(t.[TaskProdottoID]) AS NVARCHAR)) + ']' AS ActiveNpi,
                       p.Cliente,
                       p.CodiceProdotto,
                       p.NomeProdotto,
                       n.[Version],
                       n.ProdottoID,
                       n.OwnerID
                FROM [Traceability_RS].[dbo].[ProgettiNPI] n 
                INNER JOIN Prodotti p ON p.ProdottoID = n.ProdottoID
                INNER JOIN WaveNPI w ON w.ProgettoID = n.ProgettoID
                LEFT JOIN [Traceability_RS].[dbo].[TaskProdotto] t ON w.WaveID = t.WaveID AND t.[stato] IS NOT NULL
                WHERE StatoProgetto = 'Attivo'
                GROUP BY n.[ProgettoID], 
                         p.Cliente + ' -> ' + p.CodiceProdotto + ' (Version: ' + ISNULL(n.[version],'#N/D') + ') ' + p.NomeProdotto,
                         p.Cliente,
                         p.CodiceProdotto,
                         p.NomeProdotto,
                         n.[Version],
                         n.ProdottoID,
                         n.OwnerID
                ORDER BY p.Cliente + ' -> ' + p.CodiceProdotto + ' (Version: ' + ISNULL(n.[version],'#N/D') + ') ' + p.NomeProdotto, n.[Version]
            """
            
            # Esegui la query raw
            result = session.execute(text(sql_query))
            
            # Converti i risultati in una lista di dizionari
            progetti = []
            for row in result:
                progetti.append({
                    'ProgettoId': row[0],
                    'ActiveNpi': row[1],
                    'Cliente': row[2],
                    'CodiceProdotto': row[3],
                    'NomeProdotto': row[4],
                    'Version': row[5],
                    'ProdottoID': row[6],
                    'OwnerID': row[7]
                })
            
            return progetti
        except Exception as e:
            logger.error(f"Errore in get_progetti_attivi: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def create_progetto_npi_for_prodotto(self, prodotto_id, version=None, owner_id=None, descrizione=None, add_defaults: bool = False):
        """Crea un progetto NPI, la sua Wave 1.0 e i task associati.
        
        Se add_defaults=True, aggiunge solo i task di default appartenenti a categorie DefaultCategory=1
        e TaskCatalogo.DefaultTask=1.
        """
        session = self._get_session()
        try:
            # Verifica se esiste già un progetto per questo prodotto
            existing = session.scalars(
                select(ProgettoNPI).where(ProgettoNPI.ProdottoID == prodotto_id)
            ).first()
            if existing:
                return None

            # Crea il progetto
            nuovo_progetto = ProgettoNPI(
                ProdottoID=prodotto_id,
                StatoProgetto='Attivo',
                DataInizio=datetime.now(),
                Version=version,
                OwnerID=owner_id,
                Descrizione=descrizione
            )
            session.add(nuovo_progetto)
            session.flush()

            # Crea la Wave 1.0
            nuova_wave = WaveNPI(
                ProgettoID=nuovo_progetto.ProgettoId,
                WaveIdentifier=1.0
            )
            session.add(nuova_wave)
            session.flush()

            # Crea i task prodotto da catalogo
            if add_defaults:
                tasks_catalogo = session.scalars(
                    select(TaskCatalogo)
                    .join(Categoria, TaskCatalogo.CategoryId == Categoria.CategoryId)
                    .where(
                        and_(
                            Categoria.DefaultCategory == True,
                            TaskCatalogo.DefaultTask == True
                        )
                    )
                ).all()
            else:
                tasks_catalogo = session.scalars(select(TaskCatalogo)).all()
            for task_cat in tasks_catalogo:
                nuovo_task_prodotto = TaskProdotto(
                    WaveID=nuova_wave.WaveID,
                    TaskID=task_cat.TaskID  # Corretto: TaskID non TaskId
                )
                session.add(nuovo_task_prodotto)

            session.commit()

            # Recupera il progetto completo
            progetto_completo = session.scalars(
                select(ProgettoNPI)
                .options(joinedload(ProgettoNPI.prodotto))
                .where(ProgettoNPI.ProgettoId == nuovo_progetto.ProgettoId)
            ).one()

            return self._detach_object(session, progetto_completo)
        except Exception as e:
            logger.error(f"Errore in create_progetto_npi_for_prodotto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_progetto_by_prodotto(self, prodotto_id):
        """Recupera il progetto NPI associato a un prodotto."""
        session = self._get_session()
        try:
            progetto = session.scalars(
                select(ProgettoNPI)
                .options(joinedload(ProgettoNPI.prodotto))
                .where(ProgettoNPI.ProdottoID == prodotto_id)
            ).first()
            return self._detach_object(session, progetto) if progetto else None
        finally:
            session.close()

    def update_progetto_npi(self, progetto_id, data):
        """Aggiorna i dati di un progetto NPI."""
        session = self._get_session()
        try:
            progetto = session.get(ProgettoNPI, progetto_id)
            if not progetto:
                raise ValueError(f"Progetto {progetto_id} non trovato")
            
            # Aggiorna i campi
            for key, value in data.items():
                if hasattr(progetto, key):
                    setattr(progetto, key, value)
            
            session.commit()
            logger.info(f"Progetto {progetto_id} aggiornato")
            return self._detach_object(session, progetto)
        except Exception as e:
            logger.error(f"Errore aggiornamento progetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def add_default_tasks_to_project(self, progetto_id: int):
        """
        Aggiunge al progetto i task di default mancanti:
        Categoria.DefaultCategory=1 e TaskCatalogo.DefaultTask=1.
        """
        session = self._get_session()
        try:
            wave = session.scalars(
                select(WaveNPI).where(WaveNPI.ProgettoID == progetto_id)
            ).first()
            if not wave:
                return 0

            # Task default da catalogo
            default_tasks = session.scalars(
                select(TaskCatalogo)
                .join(Categoria, TaskCatalogo.CategoryId == Categoria.CategoryId)
                .where(
                    and_(
                        Categoria.DefaultCategory == True,
                        TaskCatalogo.DefaultTask == True
                    )
                )
            ).all()

            if not default_tasks:
                return 0

            # Task già presenti nel progetto
            existing_task_ids = set(
                session.scalars(
                    select(TaskProdotto.TaskID)
                    .where(TaskProdotto.WaveID == wave.WaveID)
                ).all()
            )

            added = 0
            for task_cat in default_tasks:
                if task_cat.TaskID in existing_task_ids:
                    continue
                session.add(TaskProdotto(WaveID=wave.WaveID, TaskID=task_cat.TaskID))
                added += 1

            if added:
                session.commit()
            return added
        except Exception as e:
            logger.error(f"Errore add_default_tasks_to_project: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def has_default_categories_and_tasks(self) -> bool:
        """Verifica se esistono categorie e task marcati come default."""
        session = self._get_session()
        try:
            count = session.scalar(
                select(func.count(TaskCatalogo.TaskID))
                .join(Categoria, TaskCatalogo.CategoryId == Categoria.CategoryId)
                .where(
                    and_(
                        Categoria.DefaultCategory == True,
                        TaskCatalogo.DefaultTask == True
                    )
                )
            )
            return (count or 0) > 0
        except Exception as e:
            logger.error(f"Errore has_default_categories_and_tasks: {e}")
            return False
        finally:
            session.close()

    def validate_final_milestone_completion(self, task_prodotto_id: int) -> tuple:
        """
        Valida che il task finale (IsFinalMilestone=True) possa essere marcato come 'Completato'.

        REGOLA: Un task finale può essere completato SOLO se TUTTI i task con NrOrdin inferiore
        sono già stati marcati come 'Completato'.

        Args:
            task_prodotto_id: ID del task da validare

        Returns:
            tuple: (is_valid: bool, message: str)
                - (True, "") se la validazione passa
                - (False, messaggio_errore) se la validazione fallisce
        """
        session = self._get_session()
        try:
            # Recupera il task
            task = session.get(TaskProdotto, task_prodotto_id)
            if not task:
                return False, "Task non trovato."

            # Verifica se è il task finale
            if not task.IsPostFinalMilestone:
                # Non è un task finale, niente da validare
                return True, ""

            logger.info(
                f"Validazione task finale {task_prodotto_id} "
                f"({task.task_catalogo.NomeTask if task.task_catalogo else 'Unknown'})"
            )
            
            # Ottieni il ProgettoID dalla wave del task
            project_id = task.wave.ProgettoID if task.wave else None
            if not project_id:
                return False, "Impossibile determinare il progetto del task."
            
            # Recupera SOLO i task della stessa wave E dello stesso progetto
            # che sono stati effettivamente assegnati (OwnerID non null O Stato non null)
            wave_tasks = session.scalars(
                select(TaskProdotto)
                .join(TaskCatalogo, TaskProdotto.TaskID == TaskCatalogo.TaskID)
                .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                .where(
                    and_(
                        TaskProdotto.WaveID == task.WaveID,
                        WaveNPI.ProgettoID == project_id,
                        or_(
                            TaskProdotto.OwnerID.isnot(None),
                            TaskProdotto.Stato.isnot(None)
                        )
                    )
                )
                .order_by(TaskCatalogo.NrOrdin.asc())
            ).all()
            
            logger.info(
                f"Validazione progetto {project_id}: trovati {len(wave_tasks)} task "
                f"(filtrati per WaveID={task.WaveID}, ProgettoID={project_id}, con Owner/Stato assegnato)"
            )

            # Trova il task finale nella lista
            final_task_index = None
            for idx, t in enumerate(wave_tasks):
                if t.TaskProdottoID == task_prodotto_id:
                    final_task_index = idx
                    break

            if final_task_index is None:
                return False, "Task finale non trovato nella wave."

            # Recupera TUTTI i task PRECEDENTI (NrOrdin inferiore)
            preceding_tasks = wave_tasks[:final_task_index]

            if not preceding_tasks:
                # Non ci sono task precedenti, il finale può essere completato
                logger.info("Nessun task precedente trovato. Task finale può essere completato.")
                return True, ""

            # Controlla che TUTTI i task precedenti siano 'Completato'
            incomplete_tasks = [
                t for t in preceding_tasks
                if t.Stato != 'Completato'
            ]

            if incomplete_tasks:
                # Costruisci il messaggio con i task incompleti
                incomplete_names = []
                for t in incomplete_tasks:
                    name = (t.task_catalogo.NomeTask
                            if t.task_catalogo
                            else f"Task {t.TaskProdottoID}")
                    stato = t.Stato or "Non assegnato"
                    incomplete_names.append(f"  • {name} (Stato: {stato})")

                error_message = (
                        "Non è possibile completare il task finale.\n\n"
                        "I seguenti task precedenti devono essere completati prima:\n\n" +
                        "\n".join(incomplete_names)
                )

                logger.warning(
                    f"Validazione fallita: {len(incomplete_tasks)} task precedenti non completati"
                )
                return False, error_message

            logger.info("Validazione superata: tutti i task precedenti sono completati.")
            return True, ""

        except Exception as e:
            logger.error(
                f"Errore durante la validazione del task finale {task_prodotto_id}: {e}",
                exc_info=True
            )
            return False, f"Errore durante la validazione: {str(e)}"
        finally:
            session.close()

    def debug_get_progetto(self, project_id):
        """Metodo di debug per verificare se il progetto esiste."""
        session = self._get_session()
        try:
            progetto = session.scalars(
                select(ProgettoNPI).where(ProgettoNPI.ProgettoID == project_id)
            ).first()

            if progetto:
                logger.info(f"Progetto trovato: ID={progetto.ProgettoID}, ProdottoID={progetto.ProdottoID}")
                prodotto = session.get(Prodotto, progetto.ProdottoID)
                if prodotto:
                    logger.info(f"Prodotto associato: {prodotto.NomeProdotto}")
                else:
                    logger.info("Nessun prodotto associato trovato")
            else:
                logger.info(f"Nessun progetto trovato con ID {project_id}")

            return self._detach_object(session, progetto)
        except Exception as e:
            logger.error(f"Errore in debug_get_progetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_dettagli_progetto(self, project_id):
        """Recupera UN progetto NPI con tutte le sue wave e i suoi task."""
        session = self._get_session()
        try:
            logger.debug(f"Cerco progetto con ID {project_id}")

            stmt = (
                select(ProgettoNPI)
                .options(
                    # Caricamento delle relazioni esistenti...
                    joinedload(ProgettoNPI.prodotto),
                    subqueryload(ProgettoNPI.waves).subqueryload(WaveNPI.tasks).options(
                        joinedload(TaskProdotto.owner),
                        joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria),

                        # --- CORREZIONE QUI ---
                        # Aggiungiamo questa riga per dire a SQLAlchemy:
                        # "Quando carichi i task, carica subito anche la lista dei loro documenti".
                        selectinload(TaskProdotto.documents).joinedload(NpiDocument.document_type)
                    )
                )
                .where(ProgettoNPI.ProgettoId == project_id)
            )

            progetto = session.scalars(stmt).first()

            if progetto:
                logger.info(f"Progetto trovato - {progetto.prodotto.NomeProdotto}")
                if progetto.waves:
                    # Questo log ora funzionerà senza errori
                    for task in progetto.waves[0].tasks:
                        logger.debug(f"Task '{task.task_catalogo.NomeTask}' ha {len(task.documents)} documenti.")
            else:
                logger.info(f"Nessun progetto trovato con ID {project_id}")

            # Quando l'oggetto viene scollegato, ora avrà già la lista .documents popolata!
            return self._detach_object(session, progetto)
        except Exception as e:
            logger.error(f"Errore durante il recupero del progetto: {e}")
            session.rollback()
            return None
        finally:
            session.close()

    def update_project_dates(self, project_id: int, start_date: datetime, due_date: datetime, lang):
        """
        Aggiorna le date di un progetto con logica di business per il task finale (IsPostFinalMilestone=True).
        Restituisce una tupla (progetto_aggiornato, messaggio_informativo).
        """
        session = self._get_session()
        try:
            # 1. VALIDAZIONE INIZIALE
            if due_date and start_date and due_date < start_date:  # Aggiunto controllo su start_date not None
                raise ValueError(lang.get('validation_error_end_before_start'))
            if not start_date:
                raise ValueError(lang.get('validation_error_start_date_required'))

            override_message = None
            progetto = session.get(ProgettoNPI, project_id)
            if not progetto:
                raise ValueError(f"Progetto con ID {project_id} non trovato.")

            # 1.1 VALIDAZIONE: Data fine progetto >= ultima data task
            if due_date:
                # Trova l'ultima data di scadenza tra tutti i task del progetto
                max_task_due_date = session.scalar(
                    select(func.max(TaskProdotto.DataScadenza))
                    .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                    .where(
                        and_(
                            WaveNPI.ProgettoID == project_id,
                            TaskProdotto.DataScadenza.isnot(None)
                        )
                    )
                )
                
                if max_task_due_date:
                    # Converti in date se necessario
                    max_date = max_task_due_date.date() if hasattr(max_task_due_date, 'date') else max_task_due_date
                    due_date_check = due_date.date() if hasattr(due_date, 'date') else due_date
                    
                    if due_date_check < max_date:
                        error_msg = lang.get(
                            'error_project_end_before_tasks',
                            f'La data fine progetto ({due_date_check.strftime("%d/%m/%Y")}) non può essere precedente '
                            f'all\'ultima data di scadenza dei task ({max_date.strftime("%d/%m/%Y")}).'
                        )
                        raise ValueError(error_msg)

            # 2. CERCA IL TASK SPECIALE (con IsFinalMilestone=True)
            # --- MODIFICA LOGICA QUI ---
            # Query per trovare il TaskProdotto con IsPostFinalMilestone=True
            final_milestone_task = session.scalars(
                select(TaskProdotto)
                .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                .filter(
                    WaveNPI.ProgettoID == project_id,
                    TaskProdotto.IsPostFinalMilestone == True
                )
            ).first()
            # --- FINE MODIFICA ---

            # 3. APPLICA LOGICA DI BUSINESS SUL TASK FINALE
            if final_milestone_task:
                # Caso A: Se il task finale è già assegnato e ha una data...
                if final_milestone_task.OwnerID is not None and final_milestone_task.DataScadenza is not None:
                    # La scadenza del progetto DEVE essere uguale a quella del task.
                    due_date = final_milestone_task.DataScadenza
                    base_msg = lang.get('info_project_due_date_aligned')
                    task_name = final_milestone_task.task_catalogo.NomeTask
                    override_message = f"{base_msg} (Task: '{task_name}')"
                # Caso B: Il task finale non è assegnato (o non ha una data)...
                else:
                    # La sua scadenza viene impostata uguale a quella del progetto.
                    final_milestone_task.DataScadenza = due_date

            # 4. AGGIORNA LE DATE DEL PROGETTO
            progetto.DataInizio = start_date
            progetto.ScadenzaProgetto = due_date

            session.commit()

            # 5. RESTITUISCI I DATI
            # Non serve fare il refresh, basta restituire l'oggetto con le date aggiornate
            return progetto, override_message

        except Exception as e:
            logger.error(f"Errore durante l'aggiornamento delle date del progetto {project_id}: {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()

    def update_progetto_npi(self, project_id: int, data: dict):
        """Aggiorna i campi di un progetto NPI."""
        session = self._get_session()
        try:
            progetto = session.get(ProgettoNPI, project_id)
            if not progetto:
                raise ValueError(f"Progetto con ID {project_id} non trovato.")
            
            # Aggiorna i campi forniti
            for key, value in data.items():
                if hasattr(progetto, key):
                    setattr(progetto, key, value)
            
            session.commit()
            return self._detach_object(session, progetto)
        except Exception as e:
            logger.error(f"Errore durante l'aggiornamento del progetto {project_id}: {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()

    # ========================================
    # Task Dependencies Management
    # ========================================
    
    def get_task_dependencies(self, task_prodotto_id):
        """Recupera tutte le dipendenze di un task."""
        session = self._get_session()
        try:
            dependencies = session.scalars(
                select(TaskDependency)
                .options(
                    joinedload(TaskDependency.depends_on_task).joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria)
                )
                .where(TaskDependency.TaskProdottoID == task_prodotto_id)
            ).all()
            return self._detach_list(session, dependencies)
        except Exception as e:
            logger.error(f"Errore nel recupero dipendenze task {task_prodotto_id}: {e}", exc_info=True)
            return []
        finally:
            session.close()
    
    def get_available_predecessor_tasks(self, task_id, wave_id):
        """
        Restituisce i task disponibili come predecessori per un dato task.
        Esclude il task stesso e i task che creerebbero dipendenze circolari.
        Mostra solo i task già assegnati (con OwnerID non null).
        """
        session = self._get_session()
        try:
            # Ottieni tutti i task della stessa wave che sono già assegnati
            all_tasks = session.scalars(
                select(TaskProdotto)
                .options(
                    joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria)
                )
                .where(TaskProdotto.WaveID == wave_id)
                .where(TaskProdotto.TaskProdottoID != task_id)
                .where(TaskProdotto.OwnerID.isnot(None))  # Solo task assegnati
            ).all()
            
            # Filtra i task che creerebbero dipendenze circolari
            available_tasks = []
            for task in all_tasks:
                if not self._would_create_circular_dependency(session, task_id, task.TaskProdottoID):
                    available_tasks.append(task)
            
            return self._detach_list(session, available_tasks)
        except Exception as e:
            logger.error(f"Errore nel recupero predecessori disponibili: {e}", exc_info=True)
            return []
        finally:
            session.close()
    
    def add_task_dependency(self, task_id, depends_on_task_id, dependency_type='FinishToStart'):
        """
        Aggiunge una dipendenza tra task.
        Valida che non si creino dipendenze circolari.
        """
        session = self._get_session()
        try:
            # Verifica dipendenza circolare
            if self._would_create_circular_dependency(session, task_id, depends_on_task_id):
                return False, "Impossibile aggiungere: creerebbe una dipendenza circolare"
            
            # Verifica che la dipendenza non esista già
            existing = session.scalars(
                select(TaskDependency).where(
                    and_(
                        TaskDependency.TaskProdottoID == task_id,
                        TaskDependency.DependsOnTaskProdottoID == depends_on_task_id
                    )
                )
            ).first()
            
            if existing:
                return False, "Questa dipendenza esiste già"
            
            # Crea la dipendenza
            new_dependency = TaskDependency(
                TaskProdottoID=task_id,
                DependsOnTaskProdottoID=depends_on_task_id,
                DependencyType=dependency_type
            )
            session.add(new_dependency)
            session.commit()
            
            return True, "Dipendenza aggiunta con successo"
        except Exception as e:
            logger.error(f"Errore nell'aggiunta dipendenza: {e}", exc_info=True)
            session.rollback()
            return False, f"Errore: {str(e)}"
        finally:
            session.close()
    
    def remove_task_dependency(self, dependency_id):
        """Rimuove una dipendenza tra task."""
        session = self._get_session()
        try:
            dependency = session.get(TaskDependency, dependency_id)
            if dependency:
                session.delete(dependency)
                session.commit()
                return True, "Dipendenza rimossa con successo"
            return False, "Dipendenza non trovata"
        except Exception as e:
            logger.error(f"Errore nella rimozione dipendenza: {e}", exc_info=True)
            session.rollback()
            return False, f"Errore: {str(e)}"
        finally:
            session.close()
    
    def validate_task_dependencies(self, task_id, lang=None):
        """
        Valida che tutte le dipendenze di un task siano soddisfatte.
        Restituisce (is_valid, error_message).
        """
        session = self._get_session()
        try:
            dependencies = session.scalars(
                select(TaskDependency)
                .options(
                    joinedload(TaskDependency.depends_on_task).joinedload(TaskProdotto.task_catalogo)
                )
                .where(TaskDependency.TaskProdottoID == task_id)
            ).all()
            
            if not dependencies:
                return True, ""
            
            unsatisfied = []
            for dep in dependencies:
                predecessor = dep.depends_on_task
                if dep.DependencyType == 'FinishToStart':
                    if predecessor.Stato != 'Completato':
                        task_name = predecessor.task_catalogo.NomeTask if predecessor.task_catalogo else f"Task {predecessor.TaskProdottoID}"
                        unsatisfied.append(f"  • {task_name} (Stato: {predecessor.Stato or 'Non assegnato'})")
            
            if unsatisfied:
                if lang:
                    error_msg = lang.get('error_dependency_not_satisfied', 
                                        "Impossibile completare il task. Le seguenti dipendenze non sono soddisfatte:")
                else:
                    error_msg = "Impossibile completare il task. Le seguenti dipendenze non sono soddisfatte:"
                error_msg += "\n\n" + "\n".join(unsatisfied)
                return False, error_msg
            
            return True, ""
        except Exception as e:
            logger.error(f"Errore nella validazione dipendenze: {e}", exc_info=True)
            return False, f"Errore durante la validazione: {str(e)}"
        finally:
            session.close()
    
    def _would_create_circular_dependency(self, session, task_id, depends_on_task_id):
        """
        Verifica ricorsivamente se aggiungere una dipendenza creerebbe un ciclo.
        Restituisce True se si creerebbe un ciclo, False altrimenti.
        """
        # Se depends_on_task_id dipende da task_id (direttamente o indirettamente), abbiamo un ciclo
        visited = set()
        
        def has_path(from_task, to_task):
            if from_task == to_task:
                return True
            if from_task in visited:
                return False
            visited.add(from_task)
            
            # Trova tutti i task da cui from_task dipende
            dependencies = session.scalars(
                select(TaskDependency.DependsOnTaskProdottoID)
                .where(TaskDependency.TaskProdottoID == from_task)
            ).all()
            
            for dep_task_id in dependencies:
                if has_path(dep_task_id, to_task):
                    return True
            return False
        
        return has_path(depends_on_task_id, task_id)

    def set_target_npi_task(self, task_id, project_id):
        """
        Imposta un task come Target NPI (IsPostFinalMilestone=True) e rimuove il flag da tutti gli altri task del progetto.
        """
        session = self._get_session()
        try:
            # 1. Trova il task selezionato
            target_task = session.get(TaskProdotto, task_id)
            if not target_task:
                raise ValueError(f"Task {task_id} non trovato.")

            # 2. Trova tutti i task del progetto (Wave)
            # Assumiamo che il progetto abbia una sola wave attiva o che vogliamo l'unicità per progetto
            # Recuperiamo la wave del task
            wave_id = target_task.WaveID
            
            # 3. Reset di tutti i task della stessa wave
            session.query(TaskProdotto).filter(
                TaskProdotto.WaveID == wave_id
            ).update({TaskProdotto.IsPostFinalMilestone: False})

            # 4. Imposta il flag sul task selezionato
            target_task.IsPostFinalMilestone = True
            
            session.commit()
            return True
        except Exception as e:
            logger.error(f"Errore impostazione Target NPI per task {task_id}: {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()

    def update_task_prodotto(self, task_prodotto_id, data):
        """Aggiorna i dettagli di un singolo TaskProdotto."""
        session = self._get_session()
        try:
            task = session.get(TaskProdotto, task_prodotto_id)
            if not task:
                raise ValueError(f"Nessun TaskProdotto trovato con ID {task_prodotto_id}")

            # Gestione campi data
            date_fields = ['DataScadenza', 'DataInizio', 'DataCompletamento']
            for field in date_fields:
                # Controlla se la chiave esiste nel dizionario 'data' prima di accedervi
                if field in data:
                    field_value = data.get(field)

                    if field_value in ('', None):
                        data[field] = None  # Imposta esplicitamente a None
                    elif isinstance(field_value, str):
                        try:
                            # Converte la stringa in un oggetto datetime completo
                            data[field] = datetime.strptime(field_value, '%d/%m/%Y')
                        except (ValueError, TypeError):
                            # Se la conversione fallisce, imposta a None
                            data[field] = None

            # Applica tutti gli aggiornamenti
            for key, value in data.items():
                setattr(task, key, value)

            session.commit()
            session.refresh(task)
            
            # Auto-chiudi il progetto se tutti i task sono completati
            if 'DataCompletamento' in data and data['DataCompletamento'] is not None:
                # Recupera il ProgettoID dal task
                wave = session.get(WaveNPI, task.WaveID)
                if wave:
                    try:
                        self.auto_update_project_status(wave.ProgettoID)
                    except Exception as e:
                        logger.warning(f"Errore auto-chiusura progetto {wave.ProgettoID}: {e}")

            return self._detach_object(session, task)
        except Exception as e:
            logger.error(f"Errore in update_task_prodotto: {e.__class__.__name__} - {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()

    def invia_notifiche_task(self, task: TaskProdotto, conferma_utente: bool = True, lang: dict = None):
        """Invia le notifiche email per un task assegnato."""
        session = self._get_session()
        try:
            # Ricarica il task con tutte le relazioni necessarie
            task_completo = session.scalars(
                select(TaskProdotto)
                .options(
                    joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria),
                    joinedload(TaskProdotto.wave).joinedload(WaveNPI.progetto).joinedload(ProgettoNPI.prodotto),
                    joinedload(TaskProdotto.wave).joinedload(WaveNPI.progetto).joinedload(ProgettoNPI.owner),
                    joinedload(TaskProdotto.owner),
                    joinedload(TaskProdotto.predecessors),
                    joinedload(TaskProdotto.successors)
                )
                .where(TaskProdotto.TaskProdottoID == task.TaskProdottoID)
            ).first()
            
            if not task_completo or not task_completo.owner or not task_completo.owner.Email:
                logger.warning(f"Task {task.TaskProdottoID}: owner o email non disponibili")
                return False
            
            # Ottieni il progetto e l'owner del progetto
            progetto = task_completo.wave.progetto if task_completo.wave else None
            if not progetto:
                logger.warning(f"Task {task.TaskProdottoID}: progetto non trovato")
                return False
            
            project_owner = progetto.owner if progetto.owner else None
            
            # Prepara i dati per l'email
            email_data = {
                'assignee_name': task_completo.owner.Nome,
                'assignee_email': task_completo.owner.Email,
                'project_name': progetto.prodotto.NomeProdotto if progetto.prodotto else "N/A",
                'project_code': progetto.prodotto.CodiceProdotto if progetto.prodotto else "N/A",
                'project_owner_name': project_owner.Nome if project_owner else "N/A",
                'project_owner_email': project_owner.Email if project_owner else None,
                'project_description': progetto.Descrizione or "Nessuna descrizione disponibile",
                'project_start': progetto.DataInizio.strftime('%d/%m/%Y') if progetto.DataInizio else "N/A",
                'project_due': progetto.ScadenzaProgetto.strftime('%d/%m/%Y') if progetto.ScadenzaProgetto else "N/A",
                'project_version': progetto.Version or "N/A",
                'task_item_id': task_completo.task_catalogo.ItemID if task_completo.task_catalogo else "N/A",
                'task_name': task_completo.task_catalogo.NomeTask if task_completo.task_catalogo else "N/A",
                'task_category': task_completo.task_catalogo.categoria.Category if task_completo.task_catalogo and task_completo.task_catalogo.categoria else "N/A",
                'task_description': task_completo.task_catalogo.Descrizione if task_completo.task_catalogo else "N/A",
                'task_due_date': task_completo.DataScadenza.strftime('%d/%m/%Y') if task_completo.DataScadenza else "N/A",
                'task_status': task_completo.Stato or "Non Iniziato",
                'predecessors': self._format_task_dependencies(task_completo.predecessors, session),
                'successors': self._format_task_dependencies(task_completo.successors, session, is_successor=True)
            }
            
            # Genera HTML email
            email_html = self._generate_task_assignment_email_html(email_data)
            
            # Invia email usando utils.send_email
            from utils import send_email
            
            subject = f"{email_data['project_name']} - Assegnazione Task NPI"
            
            try:
                send_email(
                    recipients=[task_completo.owner.Email],
                    subject=subject,
                    body=email_html,
                    is_html=True
                )
                success = True
                logger.info(f"Notifica inviata con successo per task {task.TaskProdottoID} a {task_completo.owner.Email}")
            except Exception as email_error:
                success = False
                logger.error(f"Invio notifica fallito per task {task.TaskProdottoID}: {email_error}")
            
            return success
            
        except Exception as e:
            logger.error(f"Errore nell'invio notifiche per task {task.TaskProdottoID}: {e}", exc_info=True)
            return False
        finally:
            session.close()
    
    def _format_task_dependencies(self, dependencies, session, is_successor=False):
        """Formatta le dipendenze del task per l'email."""
        if not dependencies:
            return []
        
        formatted = []
        for dep in dependencies:
            # Ottieni l'ID del task dipendente
            # Se is_successor=True, questo task è il "depends_on" e vogliamo il task principale
            # Se is_successor=False, questo task è il principale e vogliamo il "depends_on"
            dep_task_id = dep.TaskProdottoID if is_successor else dep.DependsOnTaskProdottoID
            
            # Carica il task dipendente
            dep_task = session.get(TaskProdotto, dep_task_id)
            if dep_task:
                # Carica le relazioni necessarie
                session.refresh(dep_task, ['task_catalogo', 'owner'])
                
                formatted.append({
                    'item_id': dep_task.task_catalogo.ItemID if dep_task.task_catalogo else "N/A",
                    'name': dep_task.task_catalogo.NomeTask if dep_task.task_catalogo else "N/A",
                    'owner': dep_task.owner.Nome if dep_task.owner else "Non Assegnato",
                    'due_date': dep_task.DataScadenza.strftime('%d/%m/%Y') if dep_task.DataScadenza else "N/A"
                })
        return formatted
    
    def _generate_task_assignment_email_html(self, data):
        """Genera HTML professionale multilingua (IT, EN, RO) per email assegnazione task."""
        
        # Traduzioni
        translations = {
            'it': {
                'title': 'Assegnazione Task NPI',
                'subtitle': 'Sistema di Gestione Progetti NPI',
                'greeting': f"Gentile <strong>{data['assignee_name']}</strong>,",
                'intro': 'Ti è stato assegnato il seguente task per il progetto NPI:',
                'project_details': 'DETTAGLI PROGETTO',
                'project_name': 'Nome Progetto:',
                'product_code': 'Codice Prodotto:',
                'project_owner': 'Responsabile Progetto:',
                'start_date': 'Data Inizio:',
                'due_date': 'Scadenza Progetto:',
                'version': 'Versione:',
                'description': 'Descrizione:',
                'your_task': 'IL TUO TASK ASSEGNATO',
                'category': 'Categoria:',
                'task_due': 'Scadenza Task:',
                'status': 'Stato Attuale:',
                'predecessors': '⚠ Questo task dipende da:',
                'successors': 'ℹ Altri task dipendono da questo:',
                'assigned_to': 'Assegnato a:',
                'important_notes': '⚠ NOTE IMPORTANTI',
                'note1': 'Rivedi attentamente le dipendenze del task',
                'note2': 'Coordina con i membri del team per i task dipendenti',
                'note3': 'Aggiorna regolarmente lo stato del task nel sistema',
                'note4': 'Contatta il responsabile del progetto per qualsiasi domanda',
                'regards': 'Cordiali saluti,',
                'project_manager': 'Responsabile Progetto',
                'disclaimer': 'Questa è una notifica automatica dal Sistema di Gestione Progetti NPI.<br>Si prega di non rispondere a questa email.'
            },
            'en': {
                'title': 'NPI Task Assignment',
                'subtitle': 'NPI Project Management System',
                'greeting': f"Dear <strong>{data['assignee_name']}</strong>,",
                'intro': 'You have been assigned the following task for the NPI project:',
                'project_details': 'PROJECT DETAILS',
                'project_name': 'Project Name:',
                'product_code': 'Product Code:',
                'project_owner': 'Project Manager:',
                'start_date': 'Start Date:',
                'due_date': 'Project Deadline:',
                'version': 'Version:',
                'description': 'Description:',
                'your_task': 'YOUR ASSIGNED TASK',
                'category': 'Category:',
                'task_due': 'Task Deadline:',
                'status': 'Current Status:',
                'predecessors': '⚠ This task depends on:',
                'successors': 'ℹ Other tasks depend on this:',
                'assigned_to': 'Assigned to:',
                'important_notes': '⚠ IMPORTANT NOTES',
                'note1': 'Carefully review task dependencies',
                'note2': 'Coordinate with team members for dependent tasks',
                'note3': 'Regularly update task status in the system',
                'note4': 'Contact the project manager for any questions',
                'regards': 'Best regards,',
                'project_manager': 'Project Manager',
                'disclaimer': 'This is an automatic notification from the NPI Project Management System.<br>Please do not reply to this email.'
            },
            'ro': {
                'title': 'Atribuire Task NPI',
                'subtitle': 'Sistem de Management Proiecte NPI',
                'greeting': f"Stimate <strong>{data['assignee_name']}</strong>,",
                'intro': 'Ți-a fost atribuit următorul task pentru proiectul NPI:',
                'project_details': 'DETALII PROIECT',
                'project_name': 'Nume Proiect:',
                'product_code': 'Cod Produs:',
                'project_owner': 'Responsabil Proiect:',
                'start_date': 'Data Început:',
                'due_date': 'Termen Limită Proiect:',
                'version': 'Versiune:',
                'description': 'Descriere:',
                'your_task': 'TASK-UL TĂU ATRIBUIT',
                'category': 'Categorie:',
                'task_due': 'Termen Limită Task:',
                'status': 'Status Curent:',
                'predecessors': '⚠ Acest task depinde de:',
                'successors': 'ℹ Alte task-uri depind de acesta:',
                'assigned_to': 'Atribuit către:',
                'important_notes': '⚠ NOTE IMPORTANTE',
                'note1': 'Revizuiește cu atenție dependențele task-ului',
                'note2': 'Coordonează-te cu membrii echipei pentru task-urile dependente',
                'note3': 'Actualizează regulat statusul task-ului în sistem',
                'note4': 'Contactează responsabilul de proiect pentru orice întrebare',
                'regards': 'Cu stimă,',
                'project_manager': 'Responsabil Proiect',
                'disclaimer': 'Aceasta este o notificare automată de la Sistemul de Management Proiecte NPI.<br>Vă rugăm să nu răspundeți la acest email.'
            }
        }
        
        # Genera sezioni per ogni lingua
        sections_html = ""
        for lang_code in ['it', 'en', 'ro']:
            t = translations[lang_code]
            
            # Formatta predecessori
            predecessors_html = ""
            if data['predecessors']:
                predecessors_html = f"<h4 style='color: #d9534f; margin-top: 15px;'>{t['predecessors']}</h4><ul style='margin: 5px 0;'>"
                for pred in data['predecessors']:
                    predecessors_html += f"<li><strong>{pred['item_id']}</strong> - {pred['name']} - {t['assigned_to']} {pred['owner']} - {t['task_due']} {pred['due_date']}</li>"
                predecessors_html += "</ul>"
            
            # Formatta successori
            successors_html = ""
            if data['successors']:
                successors_html = f"<h4 style='color: #5bc0de; margin-top: 15px;'>{t['successors']}</h4><ul style='margin: 5px 0;'>"
                for succ in data['successors']:
                    successors_html += f"<li><strong>{succ['item_id']}</strong> - {succ['name']} - {t['assigned_to']} {succ['owner']} - {t['task_due']} {succ['due_date']}</li>"
                successors_html += "</ul>"
            
            # Aggiungi sezione lingua
            lang_flag = {'it': '🇮🇹', 'en': '🇬🇧', 'ro': '🇷🇴'}[lang_code]
            sections_html += f"""
            <div class="lang-section" style="margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 8px;">
                <div style="text-align: center; margin-bottom: 20px;">
                    <span style="font-size: 24px;">{lang_flag}</span>
                    <h2 style="color: #0078d4; margin: 10px 0;">{t['title']}</h2>
                </div>
                
                <p style="font-size: 16px; margin-bottom: 20px;">{t['greeting']}</p>
                <p>{t['intro']}</p>
                
                <div class="section">
                    <h2>📊 {t['project_details']}</h2>
                    <div class="info-row"><span class="label">{t['project_name']}</span> <span class="value">{data['project_name']}</span></div>
                    <div class="info-row"><span class="label">{t['product_code']}</span> <span class="value">{data['project_code']}</span></div>
                    <div class="info-row"><span class="label">{t['project_owner']}</span> <span class="value">{data['project_owner_name']}</span></div>
                    <div class="info-row"><span class="label">{t['start_date']}</span> <span class="value">{data['project_start']}</span></div>
                    <div class="info-row"><span class="label">{t['due_date']}</span> <span class="value">{data['project_due']}</span></div>
                    <div class="info-row"><span class="label">{t['version']}</span> <span class="value">{data['project_version']}</span></div>
                    <div class="info-row" style="margin-top: 15px;">
                        <span class="label">{t['description']}</span><br>
                        <span class="value" style="display: block; margin-top: 5px; padding: 10px; background: white; border-radius: 4px;">{data['project_description']}</span>
                    </div>
                </div>
                
                <div class="section">
                    <h2>✅ {t['your_task']}</h2>
                    <div class="task-box">
                        <h3>{data['task_item_id']} - {data['task_name']}</h3>
                        <div class="info-row"><span class="label">{t['category']}</span> <span class="value">{data['task_category']}</span></div>
                        <div class="info-row"><span class="label">{t['description']}</span> <span class="value">{data['task_description']}</span></div>
                        <div class="info-row"><span class="label">{t['task_due']}</span> <span class="value" style="color: #d9534f; font-weight: bold;">{data['task_due_date']}</span></div>
                        <div class="info-row"><span class="label">{t['status']}</span> <span class="value">{data['task_status']}</span></div>
                        
                        {predecessors_html}
                        {successors_html}
                    </div>
                </div>
                
                <div class="important-notes">
                    <h3>{t['important_notes']}</h3>
                    <ul>
                        <li>{t['note1']}</li>
                        <li>{t['note2']}</li>
                        <li>{t['note3']}</li>
                        <li>{t['note4']}</li>
                    </ul>
                </div>
                
                <div class="footer">
                    <p><strong>{t['regards']}</strong></p>
                    <p><strong>{data['project_owner_name']}</strong><br>{t['project_manager']}</p>
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 15px 0;">
                    <p style="font-size: 11px; color: #999;">{t['disclaimer']}</p>
                </div>
            </div>
            {"<hr style='border: none; border-top: 3px solid #0078d4; margin: 40px 0;'>" if lang_code != 'ro' else ""}
            """
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; background-color: #f9f9f9; }}
                .header {{ background: linear-gradient(135deg, #0078d4 0%, #005a9e 100%); color: #1a3a52; padding: 30px 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 28px; color: #1a3a52; }}
                .header p {{ margin: 5px 0 0 0; font-size: 14px; opacity: 0.9; color: #1a3a52; }}
                .content {{ background: white; padding: 30px; border-radius: 0 0 8px 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .section {{ margin: 25px 0; padding: 20px; background: white; border-left: 4px solid #0078d4; border-radius: 4px; }}
                .section h2 {{ color: #0078d4; margin: 0 0 15px 0; font-size: 20px; border-bottom: 2px solid #0078d4; padding-bottom: 10px; }}
                .task-box {{ background: white; padding: 20px; margin: 15px 0; border: 2px solid #0078d4; border-radius: 6px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
                .task-box h3 {{ color: #0078d4; margin: 0 0 15px 0; font-size: 18px; }}
                .label {{ font-weight: bold; color: #555; display: inline-block; min-width: 150px; }}
                .value {{ color: #333; }}
                .info-row {{ margin: 8px 0; }}
                .footer {{ background: white; padding: 20px; margin-top: 20px; border-top: 3px solid #0078d4; border-radius: 4px; text-align: center; }}
                .footer p {{ margin: 5px 0; color: #666; font-size: 13px; }}
                .important-notes {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 4px; }}
                .important-notes h3 {{ color: #856404; margin: 0 0 10px 0; font-size: 16px; }}
                .important-notes ul {{ margin: 10px 0; padding-left: 20px; color: #856404; }}
                ul {{ line-height: 1.8; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📋 NPI Task Assignment / Assegnazione Task NPI / Atribuire Task NPI</h1>
                    <p>🇮🇹 Italiano | 🇬🇧 English | 🇷🇴 Română</p>
                </div>
                
                <div class="content">
                    {sections_html}
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def send_project_completion_email(self, project_id: int, lang: dict = None):
        """
        Invia email di completamento progetto a tutti i partecipanti con statistiche complete.
        
        Args:
            project_id: ID del progetto completato
            lang: Dizionario traduzioni (opzionale)
        
        Returns:
            tuple: (success: bool, message: str)
        """
        session = self._get_session()
        try:
            # 1. Carica progetto con tutte le relazioni
            progetto = session.scalars(
                select(ProgettoNPI)
                .options(
                    joinedload(ProgettoNPI.prodotto),
                    joinedload(ProgettoNPI.owner),
                    subqueryload(ProgettoNPI.waves).subqueryload(WaveNPI.tasks).options(
                        joinedload(TaskProdotto.owner),
                        joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria),
                        selectinload(TaskProdotto.documents).joinedload(NpiDocument.document_type)
                    )
                )
                .where(ProgettoNPI.ProgettoId == project_id)
            ).first()
            
            if not progetto or not progetto.waves:
                return False, "Progetto non trovato"
            
            # 2. Raccogli tutti i partecipanti (owner unici dei task)
            participants = {}
            wave = progetto.waves[0]
            
            for task in wave.tasks:
                if task.owner and task.owner.Email and task.OwnerID:
                    participants[task.owner.Email] = task.owner.Nome
            
            if not participants:
                return False, "Nessun partecipante trovato"
            
            logger.info(f"Trovati {len(participants)} partecipanti per progetto {project_id}")
            
            # 3. Calcola statistiche task
            completed_tasks = [t for t in wave.tasks if t.Stato == 'Completato' and t.OwnerID]
            total_tasks = len(completed_tasks)
            
            if total_tasks == 0:
                return False, "Nessun task completato"
            
            on_time_tasks = []
            late_tasks = []
            
            for task in completed_tasks:
                if task.DataCompletamento and task.DataScadenza:
                    if task.DataCompletamento <= task.DataScadenza:
                        on_time_tasks.append(task)
                    else:
                        late_tasks.append(task)
            
            on_time_count = len(on_time_tasks)
            late_count = len(late_tasks)
            on_time_percent = round((on_time_count / total_tasks) * 100, 1) if total_tasks > 0 else 0
            late_percent = round((late_count / total_tasks) * 100, 1) if total_tasks > 0 else 0
            
            # 4. Calcola costi progetto (da documenti NPI)
            total_cost = 0
            cost_breakdown = {}
            
            for task in wave.tasks:
                for doc in task.documents:
                    if doc.DocValue and not doc.DateOut:
                        total_cost += doc.DocValue
                        doc_type = doc.document_type.TypeName if doc.document_type else "Other"
                        cost_breakdown[doc_type] = cost_breakdown.get(doc_type, 0) + doc.DocValue
            
            # 5. Prepara dati email
            email_data = {
                'project_name': progetto.prodotto.NomeProdotto if progetto.prodotto else "N/A",
                'product_code': progetto.prodotto.CodiceProdotto if progetto.prodotto else "N/A",
                'version': progetto.Version or "N/A",
                'start_date': progetto.DataInizio.strftime('%d/%m/%Y') if progetto.DataInizio else "N/A",
                'completion_date': datetime.now().strftime('%d/%m/%Y'),
                'owner_name': progetto.owner.Nome if progetto.owner else "N/A",
                'total_tasks': total_tasks,
                'on_time_tasks': on_time_count,
                'on_time_percent': on_time_percent,
                'late_tasks': late_count,
                'late_percent': late_percent,
                'total_cost': total_cost,
                'cost_breakdown': cost_breakdown,
                'tasks': completed_tasks,
                'on_time_task_list': on_time_tasks,
                'late_task_list': late_tasks
            }
            
            # 6. Genera HTML email
            email_html = self._generate_project_completion_email_html(email_data)
            
            # 7. Invia email a tutti i partecipanti
            from utils import send_email
            
            recipient_list = list(participants.keys())
            subject = f"🎉 NPI Project Completed: {email_data['project_name']}"
            
            try:
                send_email(
                    recipients=recipient_list,
                    subject=subject,
                    body=email_html,
                    is_html=True
                )
                logger.info(f"Email completamento progetto {project_id} inviata a {len(recipient_list)} partecipanti")
                return True, f"Email inviata a {len(recipient_list)} partecipanti"
            except Exception as email_error:
                logger.error(f"Errore invio email completamento progetto: {email_error}")
                return False, f"Errore invio email: {str(email_error)}"
                
        except Exception as e:
            logger.error(f"Errore in send_project_completion_email: {e}", exc_info=True)
            return False, f"Errore: {str(e)}"
        finally:
            session.close()
    
    def _generate_project_completion_email_html(self, data):
        """Genera HTML email di completamento progetto in inglese con logo aziendale."""
        import base64
        
        # Carica logo e converti in base64 per embedding
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Logo.png')
        logo_base64 = ""
        
        try:
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except Exception as e:
            logger.warning(f"Impossibile caricare logo: {e}")
        
        # Genera righe tabella task
        task_rows = ""
        for task in data['tasks']:
            is_on_time = task in data.get('on_time_task_list', [])
            status_icon = "✅" if is_on_time else "⚠️"
            status_text = "On-Time" if is_on_time else "Late"
            status_color = "#28a745" if is_on_time else "#dc3545"
            
            task_rows += f"""
            <tr>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{task.task_catalogo.NomeTask if task.task_catalogo else 'N/A'}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{task.task_catalogo.categoria.Category if task.task_catalogo and task.task_catalogo.categoria else 'N/A'}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{task.owner.Nome if task.owner else 'N/A'}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{task.DataScadenza.strftime('%d/%m/%Y') if task.DataScadenza else 'N/A'}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{task.DataCompletamento.strftime('%d/%m/%Y') if task.DataCompletamento else 'N/A'}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6; color: {status_color}; font-weight: bold;">{status_icon} {status_text}</td>
            </tr>
            """
        
        # Genera righe breakdown costi
        cost_rows = ""
        for doc_type, cost in data['cost_breakdown'].items():
            cost_rows += f"""
            <tr>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{doc_type}</td>
                <td style="padding: 8px; border-bottom: 1px solid #dee2e6; text-align: right;">€ {cost:,.2f}</td>
            </tr>
            """
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px 10px 0 0; }}
                .logo {{ width: 80px; height: auto; margin-bottom: 15px; }}
                .content {{ background: #f8f9fa; padding: 30px; }}
                .section {{ background: white; padding: 20px; margin-bottom: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                h2 {{ color: #667eea; margin-top: 0; }}
                h3 {{ color: #764ba2; border-bottom: 2px solid #667eea; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
                th {{ background: #667eea; color: white; padding: 12px; text-align: left; }}
                .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; }}
                .stat-box {{ background: #e9ecef; padding: 15px; border-radius: 8px; text-align: center; }}
                .stat-value {{ font-size: 32px; font-weight: bold; color: #667eea; }}
                .stat-label {{ font-size: 14px; color: #6c757d; margin-top: 5px; }}
                .footer {{ text-align: center; padding: 20px; color: #6c757d; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    {f'<img src="data:image/png;base64,{logo_base64}" class="logo" alt="Company Logo">' if logo_base64 else ''}
                    <h1 style="margin: 0;">🎉 NPI Project Completed</h1>
                    <p style="margin: 10px 0 0 0; opacity: 0.9;">Project Management System</p>
                </div>
                
                <div class="content">
                    <!-- Project Summary -->
                    <div class="section">
                        <h3>📊 Project Summary</h3>
                        <table style="border: none;">
                            <tr><td style="padding: 8px; font-weight: bold; width: 200px;">Product Name:</td><td style="padding: 8px;">{data['project_name']}</td></tr>
                            <tr><td style="padding: 8px; font-weight: bold;">Product Code:</td><td style="padding: 8px;">{data['product_code']}</td></tr>
                            <tr><td style="padding: 8px; font-weight: bold;">Version:</td><td style="padding: 8px;">{data['version']}</td></tr>
                            <tr><td style="padding: 8px; font-weight: bold;">Start Date:</td><td style="padding: 8px;">{data['start_date']}</td></tr>
                            <tr><td style="padding: 8px; font-weight: bold;">Completion Date:</td><td style="padding: 8px;">{data['completion_date']}</td></tr>
                            <tr><td style="padding: 8px; font-weight: bold;">Project Owner:</td><td style="padding: 8px;">{data['owner_name']}</td></tr>
                        </table>
                    </div>
                    
                    <!-- Task Statistics -->
                    <div class="section">
                        <h3>📈 Task Statistics</h3>
                        <div class="stats-grid">
                            <div class="stat-box">
                                <div class="stat-value">{data['total_tasks']}</div>
                                <div class="stat-label">Total Tasks Completed</div>
                            </div>
                            <div class="stat-box" style="background: #d4edda;">
                                <div class="stat-value" style="color: #28a745;">{data['on_time_percent']}%</div>
                                <div class="stat-label">On-Time Completion ({data['on_time_tasks']} tasks)</div>
                            </div>
                            <div class="stat-box" style="background: #f8d7da;">
                                <div class="stat-value" style="color: #dc3545;">{data['late_percent']}%</div>
                                <div class="stat-label">Late Completion ({data['late_tasks']} tasks)</div>
                            </div>
                            <div class="stat-box">
                                <div class="stat-value">€ {data['total_cost']:,.2f}</div>
                                <div class="stat-label">Total Project Cost</div>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Cost Breakdown -->
                    {f'''
                    <div class="section">
                        <h3>💰 Cost Breakdown</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>Category</th>
                                    <th style="text-align: right;">Amount</th>
                                </tr>
                            </thead>
                            <tbody>
                                {cost_rows}
                                <tr style="background: #e9ecef; font-weight: bold;">
                                    <td style="padding: 12px;">TOTAL</td>
                                    <td style="padding: 12px; text-align: right;">€ {data['total_cost']:,.2f}</td>
                                </tr>
                            </tbody>
                        </table>
                    </div>
                    ''' if data['cost_breakdown'] else ''}
                    
                    <!-- Task Details -->
                    <div class="section">
                        <h3>📋 Task Details</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>Task Name</th>
                                    <th>Category</th>
                                    <th>Assigned To</th>
                                    <th>Due Date</th>
                                    <th>Completed</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody>
                                {task_rows}
                            </tbody>
                        </table>
                    </div>
                    
                    <div class="footer">
                        <p>This is an automated notification from the NPI Project Management System.</p>
                        <p>© {datetime.now().year} Vandewiele. All rights reserved.</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html


    def get_gantt_data(self, progetto_id: int):
        """Recupera i dati di un progetto formattati per Plotly per il diagramma di Gantt."""
        session = self._get_session()
        try:
            progetto = session.scalars(
                select(ProgettoNPI)
                .options(
                    joinedload(ProgettoNPI.prodotto),
                    subqueryload(ProgettoNPI.waves).subqueryload(WaveNPI.tasks).options(
                        joinedload(TaskProdotto.owner),
                        joinedload(TaskProdotto.task_catalogo).joinedload(TaskCatalogo.categoria),
                        subqueryload(TaskProdotto.dependencies)
                    )
                )
                .where(ProgettoNPI.ProgettoId == progetto_id)
            ).first()

            if not progetto or not progetto.waves:
                return None, None

            # Filtra i task validi (con owner e date)
            valid_tasks = []
            for task in progetto.waves[0].tasks:
                if (task.OwnerID is not None and
                        task.DataInizio and
                        task.DataScadenza):
                    valid_tasks.append(task)
            
            # Applica ordinamento topologico
            sorted_tasks = self._topological_sort_tasks(valid_tasks)
            
            # Crea i dati per il Gantt usando i task ordinati
            df_data = []
            for task in sorted_tasks:
                # Assicurati che la data di fine sia >= alla data di inizio
                data_inizio = min(task.DataInizio, task.DataScadenza)
                data_fine = max(task.DataInizio, task.DataScadenza)

                # Recupera le dipendenze
                deps = [d.DependsOnTaskProdottoID for d in task.dependencies]

                # 🆕 Calcolo dinamico percentuale di completamento
                completion_percentage = 0
                
                # Normalizza lo stato per il confronto (rimuove spazi e rende case-insensitive)
                stato_normalized = task.Stato.strip() if task.Stato else ""
                logger.info(f"🔍 Task {task.TaskProdottoID} ({task.task_catalogo.NomeTask if task.task_catalogo else 'N/A'}): Stato raw='{task.Stato}', normalized='{stato_normalized}'")
                
                if stato_normalized == 'Completato':
                    # Task completato = 100%
                    completion_percentage = 100
                    logger.info(f"✅ Task {task.TaskProdottoID}: Completato -> 100%")
                elif stato_normalized == 'In Lavorazione':
                    # Task in lavorazione = calcola in base alle date
                    try:
                        today = datetime.now().date()
                        start_date = data_inizio.date() if hasattr(data_inizio, 'date') else data_inizio
                        end_date = data_fine.date() if hasattr(data_fine, 'date') else data_fine
                        
                        logger.info(f"📅 Task {task.TaskProdottoID}: Oggi={today}, Inizio={start_date}, Fine={end_date}")
                        
                        if today < start_date:
                            # Non ancora iniziato (data futura)
                            completion_percentage = 0
                            logger.info(f"⏸️ Task {task.TaskProdottoID}: In Lavorazione ma non iniziato (futuro) -> 0%")
                        elif today > end_date:
                            # 🆕 Scadenza superata ma NON completato = 50% (default sconosciuto)
                            completion_percentage = 50
                            logger.info(f"⏰ Task {task.TaskProdottoID}: In Lavorazione scaduto -> 50% (default ritardo)")
                        else:
                            # Task in corso: oggi è tra start_date e end_date (inclusi)
                            total_duration = (end_date - start_date).days
                            
                            if total_duration == 0:
                                # Task con durata 0 giorni (inizio == fine == oggi)
                                completion_percentage = 50
                                logger.info(f"⚡ Task {task.TaskProdottoID}: In Lavorazione stesso giorno (inizio=fine=oggi) -> 50%")
                            else:
                                # Calcola percentuale in base al progresso temporale
                                elapsed_duration = (today - start_date).days
                                completion_percentage = int((elapsed_duration / total_duration) * 100)
                                # Assicura che sia almeno 1% se è iniziato
                                completion_percentage = max(1, completion_percentage)
                                logger.info(f"📊 Task {task.TaskProdottoID}: In Lavorazione {elapsed_duration}/{total_duration} giorni -> {completion_percentage}%")
                    except Exception as e:
                        logger.warning(f"❌ Errore calcolo percentuale per task {task.TaskProdottoID}: {e}")
                        completion_percentage = 0
                else:
                    # 🆕 Task in altri stati (Da Fare, Bloccato, ecc.)
                    # Se la scadenza è superata, 50% default, altrimenti 0%
                    try:
                        today = datetime.now().date()
                        end_date = data_fine.date() if hasattr(data_fine, 'date') else data_fine
                        
                        if today > end_date:
                            # Scadenza superata ma non completato
                            completion_percentage = 50
                            logger.info(f"🔴 Task {task.TaskProdottoID}: Stato '{stato_normalized}' scaduto -> 50% (default ritardo)")
                        else:
                            # Non ancora scaduto
                            completion_percentage = 0
                            logger.info(f"🔴 Task {task.TaskProdottoID}: Stato '{stato_normalized}' -> 0%")
                    except Exception as e:
                        logger.warning(f"❌ Errore verifica scadenza per task {task.TaskProdottoID}: {e}")
                        completion_percentage = 0

                df_data.append({
                    'Task': task.task_catalogo.NomeTask if task.task_catalogo else "Task non definito",
                    'Category': task.task_catalogo.categoria.Category if (task.task_catalogo and task.task_catalogo.categoria) else "Nessuna Categoria",
                    'Start': data_inizio,
                    'Finish': data_fine,
                    'Owner': task.owner.Nome if task.owner else 'Non Assegnato',
                    'Status': task.Stato,
                    'TaskProdottoID': task.TaskProdottoID,
                    'Dependencies': deps,
                    'Completion': completion_percentage,  # 🆕 Percentuale calcolata dinamicamente
                    'IsTargetNPI': task.IsPostFinalMilestone if task.IsPostFinalMilestone is not None else False  # 🆕 Flag Target NPI
                })

            logger.debug(f"Totale task nel progetto: {len(progetto.waves[0].tasks)}")
            logger.debug(f"Task mostrati nel Gantt: {len(df_data)}")

            return df_data, progetto.prodotto.NomeProdotto

        except Exception as e:
            logger.error(f"Errore in get_gantt_data: {e}")
            raise
        finally:
            session.close()
    
    def _topological_sort_tasks(self, tasks):
        """Ordina i task rispettando le dipendenze con ordinamento topologico.
        
        I task senza dipendenze sono ordinati per data di inizio.
        I task con dipendenze appaiono sempre dopo i loro predecessori.
        Metodo condiviso tra project_window e generazione Gantt.
        """
        from collections import defaultdict, deque
        from datetime import datetime
        
        # Crea una mappa task_id -> task per accesso rapido
        task_map = {t.TaskProdottoID: t for t in tasks}
        task_ids = set(task_map.keys())
        
        # Costruisci il grafo delle dipendenze
        # graph[task_id] = lista di task che dipendono da task_id
        graph = defaultdict(list)
        # in_degree[task_id] = numero di dipendenze (predecessori)
        in_degree = defaultdict(int)
        
        # Inizializza tutti i task con in_degree 0
        for task_id in task_ids:
            in_degree[task_id] = 0
        
        # Popola il grafo e calcola gli in_degree
        for task in tasks:
            # Usa get_task_dependencies se disponibile, altrimenti usa dependencies
            if hasattr(task, 'dependencies'):
                dependencies = task.dependencies
            else:
                dependencies = self.get_task_dependencies(task.TaskProdottoID)
            
            for dep in dependencies:
                # L'attributo corretto è sempre DependsOnTaskProdottoID
                predecessor_id = dep.DependsOnTaskProdottoID
                # Considera solo dipendenze tra task nel set corrente
                if predecessor_id in task_ids:
                    graph[predecessor_id].append(task.TaskProdottoID)
                    in_degree[task.TaskProdottoID] += 1
        
        # Ottieni tutti i task senza dipendenze (in_degree == 0)
        # e ordinali per data di inizio
        queue = []
        for task_id in task_ids:
            if in_degree[task_id] == 0:
                task = task_map[task_id]
                # Usa la data di inizio per l'ordinamento, o una data molto lontana se None
                start_date = task.DataInizio if task.DataInizio else datetime(2099, 12, 31)
                queue.append((start_date, task_id))
        
        # Ordina i task iniziali per data di inizio
        queue.sort(key=lambda x: x[0])
        queue = deque([task_id for _, task_id in queue])
        
        # Algoritmo di Kahn per ordinamento topologico
        result = []
        
        while queue:
            # Prendi il prossimo task dalla coda
            current_id = queue.popleft()
            result.append(task_map[current_id])
            
            # Per ogni task che dipende dal task corrente
            successors_to_add = []
            for successor_id in graph[current_id]:
                in_degree[successor_id] -= 1
                
                # Se tutte le dipendenze del successor sono state soddisfatte
                if in_degree[successor_id] == 0:
                    task = task_map[successor_id]
                    start_date = task.DataInizio if task.DataInizio else datetime(2099, 12, 31)
                    successors_to_add.append((start_date, successor_id))
            
            # Ordina i successori per data di inizio prima di aggiungerli alla coda
            successors_to_add.sort(key=lambda x: x[0])
            for _, successor_id in successors_to_add:
                queue.append(successor_id)
        
        # Verifica che tutti i task siano stati processati (no cicli)
        if len(result) != len(tasks):
            logger.warning("Rilevato ciclo nelle dipendenze dei task. Alcuni task potrebbero non essere visualizzati correttamente.")
            # Aggiungi i task rimanenti alla fine (fallback)
            processed_ids = {t.TaskProdottoID for t in result}
            remaining = [t for t in tasks if t.TaskProdottoID not in processed_ids]
            result.extend(remaining)
        
        return result

    def aggiorna_date_task_from_gantt(self, task_prodotto_id: int, nuova_data_inizio: datetime,
                                      nuova_data_scadenza: datetime, user: str):
        """Aggiorna le date di un task specifico dal Gantt."""
        session = self._get_session()
        try:
            task = session.get(TaskProdotto, task_prodotto_id)
            if not task:
                raise ValueError(f"Task non trovato con ID {task_prodotto_id}")

            task.DataInizio = nuova_data_inizio.date()
            task.DataScadenza = nuova_data_scadenza.date()

            current_note = task.Note or ''
            log_note = f"\nDate aggiornate da {user} il {datetime.now().strftime('%d/%m/%Y')}."
            task.Note = current_note.strip() + log_note

            session.commit()
            session.refresh(task)

            logger.info(f"Date del TaskProdotto {task_prodotto_id} aggiornate da {user}.")

            return self._detach_object(session, task)
        except Exception as e:
            logger.error(f"Errore nell'aggiornamento del task {task_prodotto_id}: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_dashboard_projects(self, year_filter=None, client_filter=None):
        """
        Recupera i dati dei progetti per la dashboard principale.
        
        Args:
            year_filter: Anno per filtrare i progetti (None = tutti gli anni)
            client_filter: Nome cliente per filtrare i progetti (None = tutti i clienti)
        
        Logica di filtraggio:
        - Progetti attivi (StatoProgetto != 'Chiuso'): sempre inclusi
        - Progetti chiusi: solo se DataInizio nell'anno selezionato
        - Se client_filter specificato: solo progetti del cliente selezionato
        """
        session = self._get_session()
        try:
            # Query base
            stmt = select(
                ProgettoNPI.ProgettoId,
                func.coalesce(ProgettoNPI.NomeProgetto, Prodotto.NomeProdotto).label("NomeProgetto"),
                Prodotto.CodiceProdotto,
                Prodotto.Cliente,
                ProgettoNPI.ScadenzaProgetto.label("ScadenzaProgetto"),
                ProgettoNPI.StatoProgetto,
                ProgettoNPI.DataInizio
            ).join(
                Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID
            )
            
            # Applica filtro anno se specificato
            if year_filter:
                stmt = stmt.where(
                    or_(
                        # Progetti attivi: sempre inclusi
                        ProgettoNPI.StatoProgetto != 'Chiuso',
                        # Progetti chiusi: solo se DataInizio nell'anno selezionato
                        and_(
                            ProgettoNPI.StatoProgetto == 'Chiuso',
                            func.year(ProgettoNPI.DataInizio) == year_filter
                        )
                    )
                )
            
            # Applica filtro cliente se specificato
            if client_filter:
                stmt = stmt.where(Prodotto.Cliente == client_filter)
            
            stmt = stmt.order_by(ProgettoNPI.DataInizio.desc())

            result = session.execute(stmt).all()
            return result
        finally:
            session.close()

    def get_npi_statistics(self, year_filter=None, client_filter=None):
        """
        Calcola le statistiche per la dashboard:
        - Totale progetti NPI
        - Progetti completati (Stato = 'Chiuso')
        - Progetti in ritardo (Stato != 'Chiuso' e ScadenzaProgetto < Oggi)
        - Statistiche per cliente
        
        Args:
            year_filter: Anno per filtrare i progetti (None = tutti gli anni)
            client_filter: Nome cliente per filtrare i progetti (None = tutti i clienti)
        """
        session = self._get_session()
        try:
            stats = {
                'total_projects': 0,
                'completed_projects': 0,
                'delayed_projects': 0,
                'customer_stats': []
            }
            
            # Crea la condizione di filtro anno (riutilizzabile)
            year_condition = None
            if year_filter:
                year_condition = or_(
                    # Progetti attivi: sempre inclusi
                    ProgettoNPI.StatoProgetto != 'Chiuso',
                    # Progetti chiusi: solo se DataInizio nell'anno selezionato
                    and_(
                        ProgettoNPI.StatoProgetto == 'Chiuso',
                        func.year(ProgettoNPI.DataInizio) == year_filter
                    )
                )

            # 1. Totale Progetti
            query = select(func.count(ProgettoNPI.ProgettoId)).join(
                Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID
            )
            if year_condition is not None:
                query = query.where(year_condition)
            if client_filter:
                query = query.where(Prodotto.Cliente == client_filter)
            stats['total_projects'] = session.scalar(query)

            # 2. Progetti Completati
            query = select(func.count(ProgettoNPI.ProgettoId)).join(
                Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID
            ).where(ProgettoNPI.StatoProgetto == 'Chiuso')
            if year_filter:
                # Per progetti chiusi, aggiungi filtro anno (solo DataInizio)
                query = query.where(func.year(ProgettoNPI.DataInizio) == year_filter)
            if client_filter:
                query = query.where(Prodotto.Cliente == client_filter)
            stats['completed_projects'] = session.scalar(query)

            # 3. Progetti in Ritardo
            query = select(func.count(ProgettoNPI.ProgettoId)).join(
                Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID
            ).where(
                and_(
                    ProgettoNPI.StatoProgetto != 'Chiuso',
                    ProgettoNPI.ScadenzaProgetto < datetime.now()
                )
            )
            if client_filter:
                query = query.where(Prodotto.Cliente == client_filter)
            # Progetti in ritardo sono sempre attivi, quindi nessun filtro anno aggiuntivo necessario
            stats['delayed_projects'] = session.scalar(query)

            # 4. Statistiche per Cliente
            customer_query = (
                select(
                    Prodotto.Cliente,
                    func.count(ProgettoNPI.ProgettoId).label('count')
                )
                .join(Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID)
            )
            
            if year_condition is not None:
                customer_query = customer_query.where(year_condition)
            
            if client_filter:
                customer_query = customer_query.where(Prodotto.Cliente == client_filter)
            
            customer_query = customer_query.group_by(Prodotto.Cliente).order_by(func.count(ProgettoNPI.ProgettoId).desc())
            
            customer_results = session.execute(customer_query).all()
            
            total = stats['total_projects'] or 1
            
            for client, count in customer_results:
                client_name = client or "N/A"
                percentage = (count / total) * 100
                stats['customer_stats'].append({
                    'client': client_name,
                    'count': count,
                    'percentage': round(percentage, 1)
                })

            return stats
        except Exception as e:
            logger.error(f"Errore nel calcolo statistiche NPI: {e}", exc_info=True)
            return None
        finally:
            session.close()

    def get_project_analysis(self, project_id: int):
        """
        Analizza un progetto per trovare i task in ritardo per ogni owner.
        """
        session = self._get_session()
        try:
            from collections import defaultdict
            today = datetime.now().date()

            # Carica tutti i task del progetto con owner e catalogo
            tasks = session.scalars(
                select(TaskProdotto)
                .join(WaveNPI)
                .options(
                    joinedload(TaskProdotto.owner),
                    joinedload(TaskProdotto.task_catalogo)
                )
                .where(WaveNPI.ProgettoID == project_id)
            ).all()

            late_tasks_by_owner = defaultdict(list)
            final_milestone_task = None

            for task in tasks:
                # Trova la milestone finale per il reminder
                if task.IsPostFinalMilestone:
                    final_milestone_task = task

                # Controlla se il task è in ritardo
                if (task.owner and task.DataScadenza and
                        task.DataScadenza.date() < today and
                        task.Stato not in ['Completato']):
                    late_tasks_by_owner[task.owner].append(task)

            return late_tasks_by_owner, final_milestone_task
        finally:
            session.close()

    def send_project_reminders(self, late_tasks_by_owner, final_milestone_task):
        """
        Invia email di sollecito (in inglese) agli owner con task in ritardo.
        """
        from email_connector import EmailSender  # Import locale
        total_sent = 0
        total_failed = 0

        # Prepara il corpo base del reminder sulla milestone finale
        final_milestone_info = "No final milestone date set for this project."
        if final_milestone_task and final_milestone_task.DataScadenza:
            final_milestone_info = (f"Please remember the final project milestone is on: "
                                    f"<strong>{final_milestone_task.DataScadenza.strftime('%Y-%m-%d')}</strong>.")

        # Inizializza il sender
        try:
            email_sender = EmailSender()
            # Le credenziali sono caricate da file dentro EmailSender
        except Exception as e:
            logger.error(f"Failed to initialize EmailSender: {e}")
            raise Exception(f"Failed to initialize EmailSender: {e}")

        for owner, tasks in late_tasks_by_owner.items():
            if not owner.Email:
                logger.warning(f"Owner {owner.Nome} has no email address. Skipping reminder.")
                total_failed += 1
                continue

            # Costruisci l'elenco dei task in ritardo
            tasks_html_list = "<ul>"
            for task in tasks:
                tasks_html_list += (f"<li><strong>{task.task_catalogo.NomeTask}</strong> "
                                    f"(Due Date: {task.DataScadenza.strftime('%Y-%m-%d')})</li>")
            tasks_html_list += "</ul>"

            # Corpo e oggetto dell'email
            subject = "NPI Project - Overdue Task Reminder"
            body = f"""
            <html>
            <body>
                <p>Dear {owner.Nome},</p>
                <p>This is an automatic reminder. The following tasks assigned to you for the NPI project are overdue:</p>
                {tasks_html_list}
                <p>Please update their status or complete them as soon as possible.</p>
                <br>
                <p><em>{final_milestone_info}</em></p>
                <br>
                <p>Thank you,</p>
                <p>NPI Management System</p>
            </body>
            </html>
            """

            # Invia email
            try:
                email_sender.send_email(
                    to_email=owner.Email,
                    subject=subject,
                    body=body,
                    is_html=True
                )
                logger.info(f"Reminder sent successfully to {owner.Email}")
                total_sent += 1
            except Exception as e:
                logger.error(f"Failed to send reminder to {owner.Email}: {e}")
                total_failed += 1

        return total_sent, total_failed

    def get_full_projects_report_data(self):
        """
        Prepara i dati completi per il report PDF, includendo l'analisi dei ritardi
        per ogni progetto attivo. Restituisce una lista di dizionari.
        """
        report_data = []

        # 1. Recupera la lista dei progetti attivi dalla dashboard
        active_projects = self.get_dashboard_projects()
        if not active_projects:
            return []

        for proj_summary in active_projects:
            project_id = proj_summary.ProgettoId

            # 2. Per ogni progetto, esegui l'analisi dei ritardi
            late_tasks_by_owner, _ = self.get_project_analysis(project_id)

            # 3. Formatta l'analisi dei ritardi
            analysis_summary = []
            if late_tasks_by_owner:
                for owner, tasks in late_tasks_by_owner.items():
                    analysis_summary.append({'owner_name': owner.Nome, 'late_count': len(tasks)})

            # 4. Aggrega tutti i dati per questo progetto
            report_data.append({
                'info': proj_summary,
                'analysis': analysis_summary
            })

        return report_data

    def get_all_npi_projects(self):
        """Recupera tutti i progetti NPI con le informazioni del prodotto associato."""
        session = self._get_session()
        try:
            progetti = session.scalars(
                select(ProgettoNPI)
                .options(joinedload(ProgettoNPI.prodotto))
                .order_by(ProgettoNPI.DataInizio.desc())
            ).all()
            return progetti
        finally:
            session.close()

    def get_all_clients(self):
        """
        Recupera tutti i nomi dei clienti unici dal database.
        
        Returns:
            Lista di nomi clienti ordinati alfabeticamente (esclude NULL/vuoti)
        """
        session = self._get_session()
        try:
            clients = session.scalars(
                select(Prodotto.Cliente)
                .distinct()
                .where(Prodotto.Cliente.isnot(None))
                .where(Prodotto.Cliente != '')
                .order_by(Prodotto.Cliente)
            ).all()
            return list(clients)
        except Exception as e:
            logger.error(f"Errore nel recupero clienti: {e}", exc_info=True)
            return []
        finally:
            session.close()

    def get_project_task_statistics(self, project_id: int):
        """
        Calcola le statistiche dei task per un progetto specifico.
        
        Args:
            project_id: ID del progetto NPI
            
        Returns:
            Dizionario con:
            - total_tasks: numero totale di task
            - completed_on_time: task completati entro la scadenza
            - completed_late: task completati dopo la scadenza
            - pending_late: task non completati con scadenza superata
            - completion_percentage: percentuale di completamento (0-100)
        """
        session = self._get_session()
        try:
            today = datetime.now().date()
            
            # Recupera tutti i task ASSEGNATI del progetto (OwnerID NOT NULL)
            tasks = session.scalars(
                select(TaskProdotto)
                .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                .where(WaveNPI.ProgettoID == project_id)
                .where(TaskProdotto.OwnerID.isnot(None))  # Solo task assegnati
            ).all()
            
            if not tasks:
                return {
                    'total_tasks': 0,
                    'completed_on_time': 0,
                    'completed_late': 0,
                    'pending_late': 0,
                    'completion_percentage': 0
                }
            
            total_tasks = len(tasks)
            completed_on_time = 0
            completed_late = 0
            pending_late = 0
            
            for task in tasks:
                # Task completati
                if task.Stato == 'Completato':
                    if task.DataCompletamento and task.DataScadenza:
                        # Normalizza le date
                        completion_date = task.DataCompletamento.date() if hasattr(task.DataCompletamento, 'date') else task.DataCompletamento
                        due_date = task.DataScadenza.date() if hasattr(task.DataScadenza, 'date') else task.DataScadenza
                        
                        if completion_date <= due_date:
                            completed_on_time += 1
                        else:
                            completed_late += 1
                    else:
                        # Se non ha date, consideriamo completato in tempo
                        completed_on_time += 1
                # Task non completati con scadenza superata
                elif task.DataScadenza:
                    due_date = task.DataScadenza.date() if hasattr(task.DataScadenza, 'date') else task.DataScadenza
                    if due_date < today:
                        pending_late += 1
            
            # Calcola percentuale di completamento
            completed_total = completed_on_time + completed_late
            completion_percentage = int((completed_total / total_tasks) * 100) if total_tasks > 0 else 0
            
            return {
                'total_tasks': total_tasks,
                'completed_on_time': completed_on_time,
                'completed_late': completed_late,
                'pending_late': pending_late,
                'completion_percentage': completion_percentage
            }
            
        except Exception as e:
            logger.error(f"Errore nel calcolo statistiche task per progetto {project_id}: {e}", exc_info=True)
            return {
                'total_tasks': 0,
                'completed_on_time': 0,
                'completed_late': 0,
                'pending_late': 0,
                'completion_percentage': 0
            }
        finally:
            session.close()

    def get_npi_overview_report_data(self, year_filter=None, client_filter=None):
        """
        Prepara i dati per un report di panoramica NPI (senza dettaglio task).
        
        Include:
        - Stato sintetico progetto (Attivo / In Completamento / In Ritardo / Completato)
        - Risorse coinvolte (owner progetto + owner task)
        - Tempo impiegato per progetti chiusi (da DataInizio a ultimo completamento task)
        """
        progetti = self.get_dashboard_projects(year_filter=year_filter, client_filter=client_filter)
        if not progetti:
            return None
        
        project_ids = [p.ProgettoId for p in progetti]
        owners_by_project = {}
        session = self._get_session()
        try:
            from collections import defaultdict
            owners_by_project = defaultdict(set)
            
            # Owner progetto
            owner_rows = session.execute(
                select(ProgettoNPI.ProgettoId, Soggetto.Nome)
                .join(Soggetto, ProgettoNPI.OwnerID == Soggetto.SoggettoId, isouter=True)
                .where(ProgettoNPI.ProgettoId.in_(project_ids))
            ).all()
            for proj_id, nome in owner_rows:
                if nome:
                    owners_by_project[proj_id].add(nome)
            
            # Owner task
            task_owner_rows = session.execute(
                select(WaveNPI.ProgettoID, Soggetto.Nome)
                .join(TaskProdotto, TaskProdotto.WaveID == WaveNPI.WaveID)
                .join(Soggetto, TaskProdotto.OwnerID == Soggetto.SoggettoId, isouter=True)
                .where(WaveNPI.ProgettoID.in_(project_ids))
            ).all()
            for proj_id, nome in task_owner_rows:
                if nome:
                    owners_by_project[proj_id].add(nome)
            
        except Exception as e:
            logger.error(f"Errore preparazione dati report panoramico NPI: {e}", exc_info=True)
        finally:
            session.close()
        
        today = datetime.now().date()
        summary = {
            'total': 0,
            'active': 0,
            'in_completion': 0,
            'completed': 0,
            'overdue': 0
        }
        
        projects_data = []
        
        for proj in progetti:
            stats = self.get_project_task_statistics(proj.ProgettoId)
            completion_pct = stats.get('completion_percentage', 0)
            
            is_closed = proj.StatoProgetto == 'Chiuso'
            is_overdue = False
            
            if not is_closed and proj.ScadenzaProgetto:
                deadline = proj.ScadenzaProgetto.date() if hasattr(proj.ScadenzaProgetto, 'date') else proj.ScadenzaProgetto
                is_overdue = deadline < today
            
            if is_closed:
                category = "Completato"
            elif is_overdue:
                category = "In Ritardo"
            elif completion_pct > 0:
                category = "In Completamento"
            else:
                category = "Attivo"
            
            # Tempo impiegato (giorni) per progetti chiusi (usa ScadenzaProgetto)
            duration_days = None
            if is_closed:
                start_date = proj.DataInizio.date() if hasattr(proj.DataInizio, 'date') else proj.DataInizio
                end_date = proj.ScadenzaProgetto
                end_date = end_date.date() if hasattr(end_date, 'date') else end_date
                if start_date and end_date and end_date >= start_date:
                    duration_days = (end_date - start_date).days
            
            resources = sorted(owners_by_project.get(proj.ProgettoId, set()))
            
            projects_data.append({
                'project_id': proj.ProgettoId,
                'project_name': proj.NomeProgetto,
                'product_code': proj.CodiceProdotto or "",
                'customer': proj.Cliente or "",
                'status': "Chiuso" if is_closed else "Attivo",
                'category': category,
                'start_date': proj.DataInizio,
                'end_date': proj.ScadenzaProgetto,
                'completion_pct': completion_pct,
                'resources': resources,
                'duration_days': duration_days
            })
            
            summary['total'] += 1
            if is_closed:
                summary['completed'] += 1
            else:
                summary['active'] += 1
                if category == "In Completamento":
                    summary['in_completion'] += 1
                if category == "In Ritardo":
                    summary['overdue'] += 1
        
        return {
            'summary': summary,
            'projects': projects_data
        }

    def export_npi_overview_report(self, year_filter=None, client_filter=None):
        """
        Esporta un report di panoramica generale NPI (senza dettaglio task).
        
        Include per progetto:
        - Stato sintetico
        - Risorse coinvolte
        - Tempo impiegato (solo per chiusi)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("La libreria 'openpyxl' Ã¨ necessaria per l'export Excel. Installala con: pip install openpyxl")
        
        report_data = self.get_npi_overview_report_data(year_filter=year_filter, client_filter=client_filter)
        if not report_data:
            raise ValueError("Nessun progetto da esportare.")
        
        temp_dir = "C:\\Temp"
        os.makedirs(temp_dir, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NPI Overview"
        
        # Stili
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16, color="0078D4")
        subtitle_font = Font(size=10, color="666666")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = "Rapporto Panoramico NPI"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"Generato il: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{row}'].font = subtitle_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        year_label = "Tutti gli anni" if not year_filter else str(year_filter)
        client_label = "Tutti i clienti" if not client_filter else str(client_filter)
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"Filtri - Anno: {year_label} | Cliente: {client_label}"
        ws[f'A{row}'].font = subtitle_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Riepilogo
        summary = report_data['summary']
        summary_labels = [
            ("Totale Progetti", summary['total']),
            ("Attivi", summary['active']),
            ("In Completamento", summary['in_completion']),
            ("Completati", summary['completed']),
            ("In Ritardo", summary['overdue'])
        ]
        
        ws[f'A{row}'] = "Riepilogo"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for label, value in summary_labels:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        headers = [
            "Cliente", "Nome Progetto", "Codice Prodotto", "Stato", "Categoria",
            "Data Inizio", "Data Fine (Scadenza)", "% Completamento",
            "Risorse Coinvolte", "Tempo Impiegato (gg)"
        ]
        header_row = row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        row += 1
        
        projects_sorted = sorted(
            report_data['projects'],
            key=lambda p: ((p.get('customer') or '').lower(), (p.get('project_name') or '').lower())
        )
        for proj in projects_sorted:
            start_date = proj['start_date'].strftime('%Y-%m-%d') if proj['start_date'] else "N/A"
            end_date = proj['end_date'].strftime('%Y-%m-%d') if proj['end_date'] else "N/A"
            resources = ", ".join(proj['resources']) if proj['resources'] else "N/A"
            duration = proj['duration_days'] if proj['duration_days'] is not None else ""
            
            row_data = [
                proj['customer'],
                proj['project_name'],
                proj['product_code'],
                proj['status'],
                proj['category'],
                start_date,
                end_date,
                f"{proj['completion_pct']}%",
                resources,
                duration
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
        
        # Applica filtro sulle intestazioni
        ws.auto_filter.ref = f"A{header_row}:J{header_row}"
        
        # Auto-width colonne (evita celle merge)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell is None or cell.value is None:
                    continue
                try:
                    value_len = len(str(cell.value))
                    if value_len > max_length:
                        max_length = value_len
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # --- TAB per ogni progetto con dettaglio task ---
        def _safe_sheet_name(name, used):
            base = (name or "Progetto").strip()
            invalid = ['\\', '/', '*', '[', ']', ':', '?']
            for ch in invalid:
                base = base.replace(ch, ' ')
            base = base[:31] if len(base) > 31 else base
            if not base:
                base = "Progetto"
            candidate = base
            idx = 1
            while candidate in used:
                suffix = f"_{idx}"
                candidate = (base[:31 - len(suffix)] + suffix) if len(base) > len(suffix) else (base + suffix)[:31]
                idx += 1
            used.add(candidate)
            return candidate
        
        used_names = set()
        today = datetime.now().date()
        session = self._get_session()
        try:
            for proj in projects_sorted:
                sheet_name = _safe_sheet_name(proj['project_name'], used_names)
                ws_proj = wb.create_sheet(title=sheet_name)
                
                ws_proj.merge_cells("A1:E1")
                ws_proj["A1"] = f"Dettaglio Task - {proj['project_name']}"
                ws_proj["A1"].font = title_font
                ws_proj["A1"].alignment = Alignment(horizontal='center')
                
                ws_proj.merge_cells("A2:E2")
                ws_proj["A2"] = f"Cliente: {proj['customer']} | Codice: {proj['product_code']}"
                ws_proj["A2"].font = subtitle_font
                ws_proj["A2"].alignment = Alignment(horizontal='center')
                
                header_row_proj = 4
                headers_proj = ["Task", "Responsabile", "Data Scadenza", "Stato", "Ritardo (gg)", "Data Completamento"]
                for col_idx, header in enumerate(headers_proj, start=1):
                    cell = ws_proj.cell(row=header_row_proj, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                
                # Recupera task ordinati per data scadenza (solo con responsabile)
                tasks = session.execute(
                    select(
                        TaskCatalogo.NomeTask,
                        Soggetto.Nome,
                        TaskProdotto.DataScadenza,
                        TaskProdotto.Stato,
                        TaskProdotto.DataCompletamento
                    )
                    .join(TaskProdotto, TaskProdotto.TaskID == TaskCatalogo.TaskID)
                    .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                    .join(Soggetto, TaskProdotto.OwnerID == Soggetto.SoggettoId)
                    .where(WaveNPI.ProgettoID == proj['project_id'])
                    .where(TaskProdotto.OwnerID.isnot(None))
                    # SQL Server: simula NULLS LAST
                    .order_by(
                        case((TaskProdotto.DataScadenza.is_(None), 1), else_=0),
                        TaskProdotto.DataScadenza.asc()
                    )
                ).all()
                
                row_idx = header_row_proj + 1
                for task_name, owner_name, due_dt, stato, completed_dt in tasks:
                    due_date = due_dt.date() if hasattr(due_dt, 'date') else due_dt
                    completed_date = completed_dt.date() if hasattr(completed_dt, 'date') else completed_dt
                    
                    days_late = ""
                    is_late = False
                    if due_date:
                        if completed_date and stato == 'Completato':
                            if completed_date > due_date:
                                days_late = (completed_date - due_date).days
                                is_late = True
                        elif due_date < today:
                            days_late = (today - due_date).days
                            is_late = True
                    
                    row_values = [
                        task_name or "",
                        owner_name or "",
                        due_date.strftime('%Y-%m-%d') if due_date else "N/A",
                        stato or "",
                        days_late,
                        completed_date.strftime('%Y-%m-%d') if completed_date else ""
                    ]
                    
                    for col_idx, value in enumerate(row_values, start=1):
                        cell = ws_proj.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        if is_late:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    row_idx += 1
                
                ws_proj.auto_filter.ref = f"A{header_row_proj}:F{header_row_proj}"
                
                # Auto-width colonne tab progetto
                for col_idx in range(1, ws_proj.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for r in range(1, ws_proj.max_row + 1):
                        cell = ws_proj.cell(row=r, column=col_idx)
                        if cell is None or cell.value is None:
                            continue
                        try:
                            value_len = len(str(cell.value))
                            if value_len > max_length:
                                max_length = value_len
                        except Exception:
                            pass
                    ws_proj.column_dimensions[column_letter].width = min(max_length + 2, 60)
        finally:
            session.close()
        
        file_path = os.path.join(temp_dir, f"NPI_Overview_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(file_path)
        return file_path

    def export_npi_to_excel_comprehensive(self, year_filter=None, client_filter=None):
        """
        Genera un file Excel professionale con statistiche NPI complete.
        
        Args:
            year_filter: Anno per filtrare i progetti (None = tutti gli anni)
            client_filter: Nome cliente per filtrare i progetti (None = tutti i clienti)
            
        Returns:
            Path del file Excel generato
            
        Struttura del file:
        - Tab "Overview": Tutti i progetti con statistiche task aggregate
        - Tab per ogni cliente: Progetti del cliente con dettaglio task
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            raise ImportError("La libreria 'openpyxl' è necessaria per l'export Excel. Installala con: pip install openpyxl")
        
        try:
            # Crea directory C:\Temp se non esiste
            temp_dir = "C:\\Temp"
            os.makedirs(temp_dir, exist_ok=True)
            
            # Recupera progetti
            progetti = self.get_dashboard_projects(year_filter=year_filter, client_filter=client_filter)
            if not progetti:
                raise ValueError("Nessun progetto da esportare.")
            
            # Crea workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Rimuovi il foglio di default
            
            # Stili comuni
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=16, color="0078D4")
            subtitle_font = Font(size=10, color="666666")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Logo path
            logo_path = os.path.join(os.path.dirname(__file__), "..", "logo.png")
            
            # ===== TAB OVERVIEW =====
            ws_overview = wb.create_sheet("Overview")
            current_row = 1
            
            # Logo
            if os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    img.width = 120
                    img.height = 40
                    ws_overview.add_image(img, 'A1')
                    current_row = 4
                except Exception as e:
                    logger.warning(f"Impossibile caricare logo: {e}")
            
            # Titolo
            ws_overview.merge_cells(f'A{current_row}:K{current_row}')
            title_cell = ws_overview[f'A{current_row}']
            title_cell.value = "NPI Projects - Overview Report"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Sottotitolo
            ws_overview.merge_cells(f'A{current_row}:K{current_row}')
            subtitle_cell = ws_overview[f'A{current_row}']
            year_label = str(year_filter) if year_filter else "All Years"
            client_label = client_filter if client_filter else "All Clients"
            subtitle_cell.value = f"Year: {year_label} | Client: {client_label} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='center')
            current_row += 2
            
            # Header tabella Overview
            headers = ["Project ID", "Project Name", "Product Code", "Customer", "Status", 
                      "End Date", "Total Tasks", "✅ On Time", "⏰ Late", "⚠️ Pending Late", "% Complete"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_overview.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            current_row += 1
            
            # Dati progetti con statistiche
            today = datetime.now().date()
            for proj in progetti:
                # Recupera statistiche task
                stats = self.get_project_task_statistics(proj.ProgettoId)
                
                # Status
                status = "Closed" if proj.StatoProgetto == "Chiuso" else "Active"
                
                # Date
                end_date = proj.ScadenzaProgetto.strftime('%Y-%m-%d') if proj.ScadenzaProgetto else "N/A"
                
                # Scrivi riga
                row_data = [
                    proj.ProgettoId,
                    proj.NomeProgetto,
                    proj.CodiceProdotto or "",
                    proj.Cliente or "",
                    status,
                    end_date,
                    stats['total_tasks'],
                    stats['completed_on_time'],
                    stats['completed_late'],
                    stats['pending_late'],
                    f"{stats['completion_percentage']}%"
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_overview.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left' if col_idx <= 4 else 'center')
                    
                    # Formattazione condizionale
                    if status == "Closed":
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif col_idx == 11:  # Colonna percentuale
                        pct = stats['completion_percentage']
                        if pct >= 80:
                            cell.font = Font(color="008000", bold=True)  # Verde
                        elif pct >= 50:
                            cell.font = Font(color="FF8C00", bold=True)  # Arancione
                        else:
                            cell.font = Font(color="FF0000", bold=True)  # Rosso
                
                current_row += 1
            
            # Auto-width colonne Overview
            for column_cells in ws_overview.columns:
                max_length = 0
                column_letter = None
                for cell in column_cells:
                    # Skip merged cells
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws_overview.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes
            ws_overview.freeze_panes = 'A7' if os.path.exists(logo_path) else 'A4'
            
            # ===== TAB PER CLIENTE =====
            # Raggruppa progetti per cliente
            projects_by_client = {}
            for proj in progetti:
                client_name = proj.Cliente or "N/A"
                if client_name not in projects_by_client:
                    projects_by_client[client_name] = []
                projects_by_client[client_name].append(proj)
            
            # Crea un tab per ogni cliente
            for client_name, client_projects in sorted(projects_by_client.items()):
                # Sanitizza nome tab (max 31 caratteri, no caratteri speciali)
                safe_sheet_name = client_name[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '').replace(':', '').replace('?', '')
                ws_client = wb.create_sheet(safe_sheet_name)
                
                current_row = 1
                
                # Titolo cliente
                ws_client.merge_cells(f'A{current_row}:J{current_row}')
                title_cell = ws_client[f'A{current_row}']
                title_cell.value = f"Client: {client_name}"
                title_cell.font = Font(bold=True, size=14, color="0078D4")
                title_cell.alignment = Alignment(horizontal='center')
                current_row += 2
                
                # Header
                client_headers = ["Project ID", "Project Name", "Product Code", "Status", 
                                "End Date", "Total Tasks", "✅ On Time", "⏰ Late", "⚠️ Pending Late", "% Complete"]
                for col_idx, header in enumerate(client_headers, start=1):
                    cell = ws_client.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                current_row += 1
                
                # Dati progetti del cliente
                for proj in client_projects:
                    stats = self.get_project_task_statistics(proj.ProgettoId)
                    status = "Closed" if proj.StatoProgetto == "Chiuso" else "Active"
                    end_date = proj.ScadenzaProgetto.strftime('%Y-%m-%d') if proj.ScadenzaProgetto else "N/A"
                    
                    row_data = [
                        proj.ProgettoId,
                        proj.NomeProgetto,
                        proj.CodiceProdotto or "",
                        status,
                        end_date,
                        stats['total_tasks'],
                        stats['completed_on_time'],
                        stats['completed_late'],
                        stats['pending_late'],
                        f"{stats['completion_percentage']}%"
                    ]
                    
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws_client.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left' if col_idx <= 3 else 'center')
                        
                        # Formattazione condizionale
                        if status == "Closed":
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        elif col_idx == 10:  # Colonna percentuale
                            pct = stats['completion_percentage']
                            if pct >= 80:
                                cell.font = Font(color="008000", bold=True)
                            elif pct >= 50:
                                cell.font = Font(color="FF8C00", bold=True)
                            else:
                                cell.font = Font(color="FF0000", bold=True)
                    
                    current_row += 1
                
                # Auto-width colonne
                for column_cells in ws_client.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column_cells:
                        # Skip merged cells
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            continue
                        if column_letter is None:
                            column_letter = cell.column_letter
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 50)
                        ws_client.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze panes
                ws_client.freeze_panes = 'A4'
            
            # Salva file
            file_path = os.path.join(temp_dir, f"NPI_Comprehensive_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(file_path)
            
            logger.info(f"Report Excel completo salvato in: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Errore durante l'export Excel completo: {e}", exc_info=True)
            raise


    def copy_tasks_from_project(self, source_project_id: int, target_project_id: int):
        """
        Copia gli assegnamenti Owner dai task di un progetto sorgente a uno di destinazione.
        Per i task copiati, imposta lo Stato a 'Da Fare' e la DataScadenza a NULL.
        """
        if source_project_id == target_project_id:
            raise ValueError("Il progetto sorgente e quello di destinazione non possono essere gli stessi.")

        session = self._get_session()
        try:
            # 1. Recupera i task di origine (Source) e mettili in una mappa per un accesso rapido
            source_tasks_query = (
                select(TaskProdotto)
                .join(WaveNPI)
                .where(WaveNPI.ProgettoID == source_project_id)
            )
            source_tasks = session.scalars(source_tasks_query).all()
            source_task_map = {task.TaskID: task for task in source_tasks}  # Mappa TaskCatalogoID -> TaskProdotto

            # 2. Recupera i task di destinazione (Target)
            target_tasks_query = (
                select(TaskProdotto)
                .join(WaveNPI)
                .where(WaveNPI.ProgettoID == target_project_id)
            )
            target_tasks = session.scalars(target_tasks_query).all()

            if not target_tasks:
                raise ValueError("Il progetto di destinazione non ha task da aggiornare.")

            # 3. Itera sui task di destinazione e aggiornali
            updated_count = 0
            for target_task in target_tasks:
                source_task = source_task_map.get(target_task.TaskID)

                # Se esiste un task corrispondente nell'origine e ha un owner...
                if source_task and source_task.OwnerID is not None:
                    target_task.OwnerID = source_task.OwnerID
                    target_task.Stato = 'Da Fare'  # Reset dello stato
                    target_task.DataScadenza = None  # Reset della data
                    target_task.DataInizio = None
                    target_task.DataCompletamento = None
                    updated_count += 1

            session.commit()
            logger.info(
                f"Copiati {updated_count} assegnamenti dal progetto {source_project_id} al {target_project_id}.")
            return updated_count

        except Exception as e:
            logger.error(f"Errore durante la copia dei task: {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()


    def get_npi_document_types(self):
        """Recupera tutti i tipi di documenti NPI."""
        session = self._get_session()
        try:
            doc_types = session.scalars(
                select(NpiDocumentType).order_by(NpiDocumentType.NpiDocumentDescription)
            ).all()
            return self._detach_list(session, doc_types)
        finally:
            session.close()


    def get_documents_for_task(self, task_prodotto_id: int):
        """Recupera tutti i documenti per un dato task, ordinati per data decrescente."""
        session = self._get_session()
        try:
            documents = session.scalars(
                select(NpiDocument)
                .options(joinedload(NpiDocument.document_type))
                .where(NpiDocument.TaskProdottoId == task_prodotto_id)
                .order_by(NpiDocument.DateIn.desc())
            ).all()
            return self._detach_list(session, documents)
        finally:
            session.close()


    def get_document_body(self, document_id: int):
        """Recupera solo il corpo binario di un documento per l'apertura."""
        session = self._get_session()
        try:
            # Seleziona solo la colonna DocumentBody per efficienza
            body = session.scalar(
                select(NpiDocument.DocumentBody).where(NpiDocument.NpiDocumentId == document_id)
            )
            return body
        finally:
            session.close()

    def add_npi_document(self, task_prodotto_id, doc_type_id, title, body, user,
                         note=None, replaces_doc_id=None, doc_value=None,
                         due_date=None, IDSite=None):  # ← Aggiungi questo parametro
        """
        Aggiunge un nuovo documento NPI al database.
        """
        try:
            with self._get_session() as session:
                # Gestisci la sostituzione del documento precedente
                if replaces_doc_id:
                    old_doc = session.query(NpiDocument).filter_by(
                        NpiDocumentId=replaces_doc_id
                    ).first()
                    if old_doc:
                        old_doc.DateOut = datetime.now()
                        version_number = old_doc.VersionNumber + 1
                    else:
                        version_number = 0
                else:
                    version_number = 0

                # Crea il nuovo documento
                new_doc = NpiDocument(
                    TaskProdottoId=task_prodotto_id,
                    NpiDocumentTypeId=doc_type_id,
                    DocumentTitle=title,
                    DocumentBody=body,
                    DateIn=datetime.now(),
                    User=user,
                    NewVersionOf=replaces_doc_id,
                    ValueInEur=doc_value,
                    DateOut=due_date,
                    VersionNumber=version_number,
                    Note=note,
                    IDSite=IDSite
                )

                session.add(new_doc)
                session.commit()

                logger.info(f"Documento '{title}' aggiunto con successo per task {task_prodotto_id}")
                return new_doc

        except Exception as e:
            logger.error(f"Errore durante l'aggiunta del documento: {e}", exc_info=True)
            raise

    def check_and_notify_document_deadlines(self):
        """
        Controlla tutti i documenti con una DueDate e invia notifiche se sono scaduti.
        """
        session = self._get_session()
        try:
            today = datetime.utcnow().date()

            # Seleziona documenti non obsoleti, con una DueDate passata e associati a un task con un owner
            overdue_docs = session.scalars(
                select(NpiDocument)
                .join(NpiDocument.task_prodotto)
                .options(
                    joinedload(NpiDocument.task_prodotto).joinedload(TaskProdotto.owner),
                    joinedload(NpiDocument.task_prodotto).joinedload(TaskProdotto.task_catalogo)
                )
                .where(
                    NpiDocument.DateOut == None,  # Documento attivo
                    NpiDocument.DueDate != None,  # Ha una data di scadenza
                    NpiDocument.DueDate < today,  # La data è passata
                    TaskProdotto.OwnerID != None  # Il task ha un owner
                )
            ).all()

            if not overdue_docs:
                logger.info("Nessuna scadenza documento da notificare.")
                return 0

            # Qui dovresti implementare la logica di invio email/notifica Teams.
            # Per semplicità, logghiamo solo i documenti trovati.
            # Puoi raggrupparli per owner e inviare una singola notifica riepilogativa.

            logger.warning(f"Trovati {len(overdue_docs)} documenti con scadenza superata.")
            for doc in overdue_docs:
                owner = doc.task_prodotto.owner
                task = doc.task_prodotto
                logger.info(f"  -> Documento '{doc.DocumentTitle}' (ID: {doc.NpiDocumentId}) "
                            f"del task '{task.task_catalogo.NomeTask}' "
                            f"assegnato a '{owner.Nome}' è SCADUTO il {doc.DueDate.strftime('%Y-%m-%d')}.")

            # TODO: Implementare qui la logica di invio notifiche raggruppate per owner
            # Esempio: email_sender.send_overdue_document_summary(owner, lista_documenti_scaduti)

            return len(overdue_docs)

        finally:
            session.close()


    def authorize_document(self, document_id: int, authorizer_name: str):
        """Autorizza un documento, impostando nome e data."""
        session = self._get_session()
        try:
            doc = session.get(NpiDocument, document_id)
            if not doc:
                raise ValueError(f"Documento con ID {document_id} non trovato.")

            doc.AutorizedBy = authorizer_name
            doc.AuthorizedOn = datetime.utcnow()

            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def get_all_catalog_tasks(self):
        """Ottiene tutti i task del catalogo."""
        session = self._get_session()
        try:
            tasks = session.scalars(
                select(TaskCatalogo)
                .options(joinedload(TaskCatalogo.categoria))
                .order_by(TaskCatalogo.NrOrdin)
            ).all()
            return [self._detach_object(session, t) for t in tasks]
        finally:
            session.close()

    def add_catalog_tasks_to_project(self, wave_id, catalog_tasks):
        """Aggiunge i task del catalogo mancanti a una wave del progetto."""
        session = self._get_session()
        try:
            added_count = 0
            for catalog_task in catalog_tasks:
                # Crea un nuovo TaskProdotto per ogni task del catalogo
                new_task = TaskProdotto(
                    WaveID=wave_id,
                    TaskID=catalog_task.TaskID
                )
                session.add(new_task)
                added_count += 1
            
            session.commit()
            logger.info(f"Aggiunti {added_count} task alla wave {wave_id}")
            return added_count
        except Exception as e:
            logger.error(f"Errore aggiunta task al progetto: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    # ========================================
    # GESTIONE DOCUMENTI PROGETTO
    # ========================================

    def add_progetto_documento(self, progetto_id, nome_file, tipo_file, dimensione, 
                               contenuto, descrizione=None, caricato_da=None):
        """Aggiunge un documento al progetto."""
        session = self._get_session()
        try:
            from .data_models import ProgettoDocumento
            
            documento = ProgettoDocumento(
                ProgettoID=progetto_id,
                NomeFile=nome_file,
                TipoFile=tipo_file,
                Dimensione=dimensione,
                Contenuto=contenuto,
                Descrizione=descrizione,
                CaricatoDa=caricato_da
            )
            session.add(documento)
            session.commit()
            logger.info(f"Documento {nome_file} aggiunto al progetto {progetto_id}")
            return True
        except Exception as e:
            logger.error(f"Errore aggiunta documento: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    def get_progetto_documenti(self, progetto_id):
        """Recupera tutti i documenti di un progetto."""
        session = self._get_session()
        try:
            from .data_models import ProgettoDocumento
            
            documenti = session.scalars(
                select(ProgettoDocumento)
                .where(ProgettoDocumento.ProgettoID == progetto_id)
                .order_by(ProgettoDocumento.DataCaricamento.desc())
            ).all()
            return [self._detach_object(session, d) for d in documenti]
        finally:
            session.close()

    def get_progetto_documento(self, documento_id):
        """Recupera un singolo documento."""
        session = self._get_session()
        try:
            from .data_models import ProgettoDocumento
            
            documento = session.get(ProgettoDocumento, documento_id)
            return self._detach_object(session, documento) if documento else None
        finally:
            session.close()

    def delete_progetto_documento(self, documento_id):
        """Elimina un documento."""
        session = self._get_session()
        try:
            from .data_models import ProgettoDocumento
            
            documento = session.get(ProgettoDocumento, documento_id)
            if documento:
                session.delete(documento)
                session.commit()
                logger.info(f"Documento {documento_id} eliminato")
                return True
            return False
        except Exception as e:
            logger.error(f"Errore eliminazione documento: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def get_all_project_documents(self, project_id):
        """
        Recupera tutti i documenti di un progetto NPI (documenti task).
        Restituisce una lista di dizionari con informazioni complete.
        """
        session = self._get_session()
        try:
            from .data_models import NpiDocument, TaskProdotto, WaveNPI, TaskCatalogo, NpiDocumentType
            
            # Query per ottenere tutti i documenti dei task del progetto
            task_documents = session.scalars(
                select(NpiDocument)
                .join(TaskProdotto, NpiDocument.TaskProdottoId == TaskProdotto.TaskProdottoID)
                .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                .where(WaveNPI.ProgettoID == project_id)
                .options(
                    joinedload(NpiDocument.task_prodotto).joinedload(TaskProdotto.task_catalogo),
                    joinedload(NpiDocument.document_type)
                )
                .order_by(NpiDocument.DateIn.desc())
            ).all()
            
            # Formatta i risultati
            documents = []
            for doc in task_documents:
                task_name = doc.task_prodotto.task_catalogo.NomeTask if doc.task_prodotto and doc.task_prodotto.task_catalogo else "N/A"
                doc_type = doc.document_type.NpiDocumentDescription if doc.document_type else "N/A"
                
                documents.append({
                    'doc_id': doc.NpiDocumentId,
                    'task_name': task_name,
                    'doc_type': doc_type,
                    'doc_title': doc.DocumentTitle or "",
                    'date_in': doc.DateIn,
                    'date_in_formatted': doc.DateIn.strftime('%d/%m/%Y %H:%M') if doc.DateIn else "N/A",
                    'user': doc.User or "N/A",
                    'version': doc.VersionNumber or 0,
                    'value': doc.ValueInEur,
                    'note': doc.Note
                })
            
            return documents
            
        except Exception as e:
            logger.error(f"Errore recupero documenti progetto {project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_npi_document(self, document_id):
        """Recupera un singolo documento NPI con il suo contenuto."""
        session = self._get_session()
        try:
            from .data_models import NpiDocument
            
            documento = session.get(NpiDocument, document_id)
            return documento
            
        except Exception as e:
            logger.error(f"Errore recupero documento {document_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    # ========================================
    # ðŸ†• GESTIONE GERARCHIA PROGETTI NPI
    # ========================================
    
    def get_child_projects(self, project_id):
        """
        Recupera tutti i progetti figli di un progetto.
        
        Args:
            project_id: ID del progetto padre
            
        Returns:
            Lista di ProgettoNPI figli ordinati per HierarchyLevel e NomeProgetto
        """
        session = self._get_session()
        try:
            from sqlalchemy.orm import joinedload
            
            children = session.query(ProgettoNPI)\
                .options(joinedload(ProgettoNPI.prodotto))\
                .filter(ProgettoNPI.ParentProjectID == project_id)\
                .order_by(ProgettoNPI.HierarchyLevel, ProgettoNPI.NomeProgetto)\
                .all()
            
            # Detach per usare fuori dalla sessione
            result = [self._detach_object(session, child) for child in children]
            return result
            
        except Exception as e:
            logger.error(f"Errore recupero progetti figli per {project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_parent_project(self, project_id):
        """
        Recupera il progetto padre di un progetto.
        
        Args:
            project_id: ID del progetto figlio
            
        Returns:
            ProgettoNPI padre o None se non ha padre
        """
        from sqlalchemy.orm import joinedload
        
        session = self._get_session()
        try:
            progetto = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ProgettoId == project_id)\
                .first()
            
            if progetto and progetto.ParentProjectID:
                parent = session.query(ProgettoNPI)\
                    .options(joinedload(ProgettoNPI.prodotto))\
                    .filter(ProgettoNPI.ProgettoId == progetto.ParentProjectID)\
                    .first()
                
                if parent:
                    return self._detach_object(session, parent)
            
            return None
            
        except Exception as e:
            logger.error(f"Errore recupero progetto padre per {project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_available_child_projects(self, parent_project_id, parent_client):
        """
        Recupera progetti che possono diventare figli del progetto specificato.
        
        Criteri:
        - Stesso cliente del padre
        - Non il padre stesso
        - Non progetti che sono già padri di altri progetti
        - Non già figli di questo padre (ma OK se non hanno padre)
        
        Args:
            parent_project_id: ID del progetto padre
            parent_client: Cliente del progetto padre
            
        Returns:
            Lista di ProgettoNPI disponibili come figli
        """
        session = self._get_session()
        try:
            from sqlalchemy import and_, or_
            
            # 🐛 DEBUG: Log parametri
            logger.info(f"🔍 get_available_child_projects: parent_id={parent_project_id}, client='{parent_client}'")
            
            # Subquery per trovare progetti che sono già padri
            is_parent_subq = session.query(ProgettoNPI.ParentProjectID)\
                .filter(ProgettoNPI.ParentProjectID.isnot(None))\
                .distinct()\
                .subquery()
            
            
            # Query principale
            from sqlalchemy.orm import joinedload
            
            available = session.query(ProgettoNPI)\
                .join(Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID)\
                .options(joinedload(ProgettoNPI.prodotto))\
                .filter(
                    and_(
                        # Stesso cliente
                        Prodotto.Cliente == parent_client,
                        # Non il padre stesso
                        ProgettoNPI.ProgettoId != parent_project_id,
                        # Non già figlio di questo padre (ma OK se ParentProjectID è NULL)
                        or_(
                            ProgettoNPI.ParentProjectID.is_(None),
                            ProgettoNPI.ParentProjectID != parent_project_id
                        ),
                        # Non già padre di altri progetti
                        ~ProgettoNPI.ProgettoId.in_(is_parent_subq)
                    )
                )\
                .order_by(ProgettoNPI.DataInizio.desc())\
                .all()
            
            # 🐛 DEBUG: Log risultati
            logger.info(f"🔍 Trovati {len(available)} progetti disponibili")
            if len(available) > 0:
                logger.info(f"🔍 Primi 3 progetti: {[(p.ProgettoId, p.NomeProgetto) for p in available[:3]]}")
            else:
                # Verifica quanti progetti totali ci sono per questo cliente
                total_for_client = session.query(ProgettoNPI)\
                    .join(Prodotto, ProgettoNPI.ProdottoID == Prodotto.ProdottoID)\
                    .filter(Prodotto.Cliente == parent_client)\
                    .count()
                logger.warning(f"⚠️ Nessun progetto disponibile, ma ci sono {total_for_client} progetti totali per cliente '{parent_client}'")
            
            # Detach per usare fuori dalla sessione
            result = [self._detach_object(session, proj) for proj in available]
            return result
            
        except Exception as e:
            logger.error(f"Errore recupero progetti disponibili come figli: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def assign_child_project(self, parent_project_id, child_project_id):
        """
        Assegna un progetto come figlio di un altro progetto.
        
        Args:
            parent_project_id: ID del progetto padre
            child_project_id: ID del progetto da assegnare come figlio
        """
        session = self._get_session()
        try:
            # Verifica che il figlio non sia già padre di altri
            existing_children = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ParentProjectID == child_project_id)\
                .count()
            
            if existing_children > 0:
                raise ValueError(f"Il progetto {child_project_id} è già padre di altri progetti e non può diventare figlio.")
            
            # Recupera il progetto figlio
            child = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ProgettoId == child_project_id)\
                .first()
            
            if not child:
                raise ValueError(f"Progetto {child_project_id} non trovato.")
            
            # Recupera il progetto padre per determinare il livello
            parent = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ProgettoId == parent_project_id)\
                .first()
            
            if not parent:
                raise ValueError(f"Progetto padre {parent_project_id} non trovato.")
            
            # Assegna il padre e aggiorna i metadati
            child.ParentProjectID = parent_project_id
            child.HierarchyLevel = (parent.HierarchyLevel or 0) + 1
            child.ProjectType = 'Child'
            
            # Aggiorna il padre come 'Parent' se non lo è già
            if parent.ProjectType != 'Parent':
                parent.ProjectType = 'Parent'
            
            session.commit()
            logger.info(f"Progetto {child_project_id} assegnato come figlio di {parent_project_id}")
            
        except Exception as e:
            session.rollback()
            logger.error(f"Errore assegnazione figlio {child_project_id} a {parent_project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def remove_child_project(self, child_project_id):
        """
        Rimuove un progetto dalla gerarchia (setta ParentProjectID = NULL).
        
        Args:
            child_project_id: ID del progetto figlio da rimuovere
        """
        session = self._get_session()
        try:
            child = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ProgettoId == child_project_id)\
                .first()
            
            if not child:
                raise ValueError(f"Progetto {child_project_id} non trovato.")
            
            # Salva l'ID del padre per verificare dopo
            old_parent_id = child.ParentProjectID
            
            # Rimuovi dalla gerarchia
            child.ParentProjectID = None
            child.HierarchyLevel = 0
            child.ProjectType = 'Standard'
            
            session.commit()
            
            # Verifica se il padre ha ancora altri figli
            if old_parent_id:
                remaining_children = session.query(ProgettoNPI)\
                    .filter(ProgettoNPI.ParentProjectID == old_parent_id)\
                    .count()
                
                # Se non ha più figli, riporta il padre a 'Standard'
                if remaining_children == 0:
                    parent = session.query(ProgettoNPI)\
                        .filter(ProgettoNPI.ProgettoId == old_parent_id)\
                        .first()
                    if parent:
                        parent.ProjectType = 'Standard'
                        session.commit()
            
            logger.info(f"Progetto {child_project_id} rimosso dalla gerarchia")
            
        except Exception as e:
            session.rollback()
            logger.error(f"Errore rimozione figlio {child_project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_project_hierarchy(self, root_project_id, max_depth=10):
        """
        Recupera l'intera gerarchia di un progetto (padre e tutti i figli ricorsivamente).
        
        Args:
            root_project_id: ID del progetto root
            max_depth: ProfonditÃ  massima della ricorsione (default 10)
            
        Returns:
            Dizionario con struttura gerarchica:
            {
                'project': ProgettoNPI,
                'children': [
                    {'project': ProgettoNPI, 'children': [...]},
                    ...
                ]
            }
        """
        session = self._get_session()
        try:
            def build_tree(project_id, current_depth=0):
                if current_depth >= max_depth:
                    logger.warning(f"ProfonditÃ  massima ({max_depth}) raggiunta per progetto {project_id}")
                    return None
                
                project = session.query(ProgettoNPI)\
                    .options(joinedload(ProgettoNPI.prodotto))\
                    .filter(ProgettoNPI.ProgettoId == project_id)\
                    .first()
                
                if not project:
                    return None
                
                children = session.query(ProgettoNPI)\
                    .options(joinedload(ProgettoNPI.prodotto))\
                    .filter(ProgettoNPI.ParentProjectID == project_id)\
                    .order_by(ProgettoNPI.HierarchyLevel, ProgettoNPI.NomeProgetto)\
                    .all()
                
                tree = {
                    'project': self._detach_object(session, project),
                    'children': []
                }
                
                for child in children:
                    child_tree = build_tree(child.ProgettoId, current_depth + 1)
                    if child_tree:
                        tree['children'].append(child_tree)
                
                return tree
            
            return build_tree(root_project_id)
            
        except Exception as e:
            logger.error(f"Errore recupero gerarchia per progetto {root_project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def can_complete_project(self, project_id):
        """
        Verifica se un progetto puÃ² essere completato.
        Un progetto puÃ² essere completato solo se tutti i suoi progetti figli sono completati.
        
        Args:
            project_id: ID del progetto da verificare
            
        Returns:
            Tupla (can_complete: bool, message: str, incomplete_children: list)
        """
        session = self._get_session()
        try:
            children = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ParentProjectID == project_id)\
                .all()
            
            if not children:
                return True, "Il progetto puÃ² essere completato (nessun progetto figlio)", []
            
            incomplete = [
                self._detach_object(session, child) 
                for child in children 
                if child.StatoProgetto != 'Completato'
            ]
            
            if not incomplete:
                return True, f"Il progetto puÃ² essere completato (tutti i {len(children)} progetti figli sono completati)", []
            else:
                message = f"Il progetto NON puÃ² essere completato: {len(incomplete)} progetti figli su {len(children)} non sono ancora completati"
                return False, message, incomplete
            
        except Exception as e:
            logger.error(f"Errore verifica completamento progetto {project_id}: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def update_hierarchy_levels(self, project_id=None):
        """
        Aggiorna il campo HierarchyLevel per un progetto e tutti i suoi discendenti.
        Se project_id Ã¨ None, aggiorna tutti i progetti.
        
        Args:
            project_id: ID del progetto root (opzionale)
        """
        session = self._get_session()
        try:
            def update_tree(proj_id, level=0):
                project = session.query(ProgettoNPI)\
                    .filter(ProgettoNPI.ProgettoId == proj_id)\
                    .first()
                
                if project:
                    project.HierarchyLevel = level
                    
                    children = session.query(ProgettoNPI)\
                        .filter(ProgettoNPI.ParentProjectID == proj_id)\
                        .all()
                    
                    for child in children:
                        update_tree(child.ProgettoId, level + 1)
            
            if project_id:
                # Aggiorna solo la gerarchia specificata
                update_tree(project_id)
            else:
                # Aggiorna tutte le gerarchie (progetti root)
                root_projects = session.query(ProgettoNPI)\
                    .filter(ProgettoNPI.ParentProjectID.is_(None))\
                    .all()
                
                for root in root_projects:
                    update_tree(root.ProgettoId)
            
            session.commit()
            logger.info(f"HierarchyLevel aggiornati per progetto {project_id if project_id else 'tutti'}")
            
        except Exception as e:
            session.rollback()
            logger.error(f"Errore aggiornamento HierarchyLevel: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def validate_no_circular_dependency(self, child_id, new_parent_id):
        """
        Verifica che l'assegnazione di un progetto padre non crei un ciclo.
        
        Args:
            child_id: ID del progetto che diventerÃ  figlio
            new_parent_id: ID del nuovo progetto padre
            
        Returns:
            Tupla (is_valid: bool, message: str)
        """
        if child_id == new_parent_id:
            return False, "Un progetto non puÃ² essere padre di se stesso"
        
        session = self._get_session()
        try:
            # Verifica se new_parent_id Ã¨ giÃ  un discendente di child_id
            def is_descendant(ancestor_id, descendant_id, visited=None):
                if visited is None:
                    visited = set()
                
                if descendant_id in visited:
                    return False  # GiÃ  visitato, evita loop infiniti
                
                visited.add(descendant_id)
                
                children = session.query(ProgettoNPI.ProgettoId)\
                    .filter(ProgettoNPI.ParentProjectID == descendant_id)\
                    .all()
                
                for (child_proj_id,) in children:
                    if child_proj_id == ancestor_id:
                        return True
                    if is_descendant(ancestor_id, child_proj_id, visited):
                        return True
                
                return False
            
            if is_descendant(child_id, new_parent_id):
                return False, "Operazione non permessa: creerebbe un ciclo di dipendenze (il nuovo padre Ã¨ giÃ  un discendente di questo progetto)"
            
            return True, "Validazione OK: nessun ciclo rilevato"
            
        except Exception as e:
            logger.error(f"Errore validazione dipendenze circolari: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_root_projects(self):
        """
        Recupera tutti i progetti root (senza padre).
        
        Returns:
            Lista di ProgettoNPI ordinati per NomeProgetto
        """
        session = self._get_session()
        try:
            projects = session.query(ProgettoNPI)\
                .filter(ProgettoNPI.ParentProjectID.is_(None))\
                .order_by(ProgettoNPI.NomeProgetto)\
                .all()
            
            result = [self._detach_object(session, p) for p in projects]
            return result
            
        except Exception as e:
            logger.error(f"Errore recupero progetti root: {e}", exc_info=True)
            raise
        finally:
            session.close()
    
    def get_gantt_hierarchy_data(self, root_project_id):
        """
        Recupera dati Gantt per un progetto e tutti i suoi discendenti (gerarchia completa).
        
        Args:
            root_project_id: ID del progetto root da cui partire
            
        Returns:
            {
                'root_project_id': int,
                'root_project_name': str,
                'has_hierarchy': bool,
                'projects': [
                    {
                        'project_id': int,
                        'project_name': str,
                        'is_root': bool,
                        'is_parent': bool,  # Ha figli?
                        'level': int,  # 0=root, 1=figlio diretto, etc.
                        'parent_id': int or None,
                        'tasks': [...]  # Dati Gantt standard
                    },
                    ...
                ]
            }
        """
        try:
            # Recupera la gerarchia completa
            hierarchy = self.get_project_hierarchy(root_project_id)
            
            if not hierarchy:
                # Nessun progetto trovato
                return {
                    'root_project_id': root_project_id,
                    'root_project_name': 'Unknown',
                    'has_hierarchy': False,
                    'projects': []
                }
            
            # Costruisci lista piatta di progetti con livelli
            projects_data = []
            
            def process_node(node, level=0, parent_id=None):
                """Processa ricorsivamente i nodi della gerarchia."""
                project = node['project']
                children = node.get('children', [])
                
                logger.info(f"🔍 Processando progetto {project.ProgettoId} - {project.NomeProgetto} (Level {level}, {len(children)} figli)")
                
                # Recupera i dati Gantt per questo progetto
                gantt_data, product_name = self.get_gantt_data(project.ProgettoId)
                
                logger.info(f"🔍 Progetto {project.ProgettoId}: {len(gantt_data) if gantt_data else 0} task trovati")
                
                project_info = {
                    'project_id': project.ProgettoId,
                    'project_name': (project.prodotto.CodiceProdotto if project.prodotto else None) or f"Progetto {project.ProgettoId}",  # 🆕 Usa CodiceProdotto invece di NomeProgetto
                    'is_root': (level == 0),
                    'is_parent': len(children) > 0,
                    'level': level,
                    'parent_id': parent_id,
                    'tasks': gantt_data if gantt_data else [],
                    'product_name': product_name or (project.prodotto.NomeProdotto if project.prodotto else None),
                    'status': project.StatoProgetto,
                    'owner': project.Responsabile if hasattr(project, 'Responsabile') else 'N/A'  # 🆕 Responsabile progetto
                }
                
                projects_data.append(project_info)
                
                # Processa ricorsivamente i figli
                for child_node in children:
                    process_node(child_node, level + 1, project.ProgettoId)
            
            # Avvia processamento dalla root
            process_node(hierarchy)
            
            # Determina se c'Ã¨ una vera gerarchia (piÃ¹ di un progetto)
            has_hierarchy = len(projects_data) > 1
            
            root_project = hierarchy['project']
            
            return {
                'root_project_id': root_project_id,
                'root_project_name': (root_project.prodotto.CodiceProdotto if root_project.prodotto else None) or f"Progetto {root_project_id}",  # 🆕 Usa CodiceProdotto
                'has_hierarchy': has_hierarchy,
                'projects': projects_data
            }
            
        except Exception as e:
            logger.error(f"Errore recupero gerarchia Gantt per progetto {root_project_id}: {e}", exc_info=True)
            return {
                'root_project_id': root_project_id,
                'root_project_name': 'Error',
                'has_hierarchy': False,
                'projects': []
            }

    # ========================================
    # FAMILY MANAGEMENT METHODS
    # ========================================
    
    def get_families(self, filter_text=None):
        """
        Recupera tutte le famiglie attive (DateEnd IS NULL).
        
        Args:
            filter_text: Filtro opzionale sul nome della famiglia
            
        Returns:
            Lista di FamilyNpi ordinati per NpiFamily
        """
        from npi.data_models import FamilyNpi
        from sqlalchemy import select
        
        session = self._get_session()
        try:
            stmt = select(FamilyNpi).where(FamilyNpi.DateEnd.is_(None))
            
            if filter_text:
                stmt = stmt.where(FamilyNpi.NpiFamily.like(f'%{filter_text}%'))
            
            stmt = stmt.order_by(FamilyNpi.NpiFamily)
            families = session.scalars(stmt).all()
            
            return self._detach_list(session, families)
        except Exception as e:
            logger.error(f"Errore in get_families: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def get_family_by_id(self, family_id):
        """Recupera una famiglia per ID."""
        from npi.data_models import FamilyNpi
        
        session = self._get_session()
        try:
            family = session.get(FamilyNpi, family_id)
            return self._detach_object(session, family)
        except Exception as e:
            logger.error(f"Errore in get_family_by_id: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def is_family_name_unique(self, name, exclude_id=None):
        """
        Verifica se il nome della famiglia è unico tra le famiglie attive.
        
        Args:
            name: Nome della famiglia da verificare
            exclude_id: ID famiglia da escludere (per update)
            
        Returns:
            True se il nome è unico, False altrimenti
        """
        from npi.data_models import FamilyNpi
        from sqlalchemy import select, and_
        
        session = self._get_session()
        try:
            stmt = select(FamilyNpi).where(
                and_(
                    FamilyNpi.NpiFamily == name,
                    FamilyNpi.DateEnd.is_(None)
                )
            )
            
            if exclude_id:
                stmt = stmt.where(FamilyNpi.FamilyNpiID != exclude_id)
            
            existing = session.scalars(stmt).first()
            return existing is None
        except Exception as e:
            logger.error(f"Errore in is_family_name_unique: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def create_family(self, data):
        """
        Crea una nuova famiglia con validazione unicità nome.
        
        Args:
            data: Dizionario con i dati della famiglia (NpiFamily)
            
        Returns:
            FamilyNpi creato
        """
        from npi.data_models import FamilyNpi
        
        session = self._get_session()
        try:
            # Validazione unicità nome
            family_name = data.get('NpiFamily', '').strip()
            if not family_name:
                raise ValueError("Il nome della famiglia è obbligatorio")
            
            if not self.is_family_name_unique(family_name):
                raise ValueError(f"Esiste già una famiglia con il nome '{family_name}'")
            
            new_family = FamilyNpi(NpiFamily=family_name, DateEnd=None)
            session.add(new_family)
            session.commit()
            session.refresh(new_family)
            
            logger.info(f"Famiglia creata: {family_name} (ID: {new_family.FamilyNpiID})")
            return self._detach_object(session, new_family)
        except Exception as e:
            logger.error(f"Errore in create_family: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def update_family(self, family_id, data):
        """
        Aggiorna una famiglia esistente con validazione unicità nome.
        
        Args:
            family_id: ID della famiglia da aggiornare
            data: Dizionario con i nuovi dati
            
        Returns:
            FamilyNpi aggiornato o None se non trovato
        """
        from npi.data_models import FamilyNpi
        
        session = self._get_session()
        try:
            family = session.get(FamilyNpi, family_id)
            if not family or family.DateEnd is not None:
                return None
            
            # Validazione unicità nome
            new_name = data.get('NpiFamily', '').strip()
            if not new_name:
                raise ValueError("Il nome della famiglia è obbligatorio")
            
            if not self.is_family_name_unique(new_name, exclude_id=family_id):
                raise ValueError(f"Esiste già una famiglia con il nome '{new_name}'")
            
            family.NpiFamily = new_name
            session.commit()
            session.refresh(family)
            
            logger.info(f"Famiglia aggiornata: ID {family_id} -> '{new_name}'")
            return self._detach_object(session, family)
        except Exception as e:
            logger.error(f"Errore in update_family: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def delete_family(self, family_id):
        """
        Soft delete di una famiglia (imposta DateEnd).
        Rimuove anche tutti i collegamenti attivi con i task.
        
        Args:
            family_id: ID della famiglia da eliminare
            
        Returns:
            True se eliminata, False se non trovata
        """
        from npi.data_models import FamilyNpi, FamilyNpiLog
        from datetime import datetime
        from sqlalchemy import update, and_
        
        session = self._get_session()
        try:
            family = session.get(FamilyNpi, family_id)
            if not family or family.DateEnd is not None:
                return False
            
            # Soft delete della famiglia
            now = datetime.now()
            family.DateEnd = now
            
            # Soft delete di tutti i collegamenti attivi
            session.execute(
                update(FamilyNpiLog)
                .where(
                    and_(
                        FamilyNpiLog.FamilyNpiID == family_id,
                        FamilyNpiLog.DateEnd.is_(None)
                    )
                )
                .values(DateEnd=now)
            )
            
            session.commit()
            logger.info(f"Famiglia eliminata (soft delete): ID {family_id}")
            return True
        except Exception as e:
            logger.error(f"Errore in delete_family: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def get_family_links(self, family_id=None):
        """
        Recupera i collegamenti famiglia-task attivi.
        
        Args:
            family_id: Filtra per famiglia specifica (opzionale)
            
        Returns:
            Lista di FamilyNpiLog con task e categoria caricati
        """
        from npi.data_models import FamilyNpiLog, TaskCatalogo
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        
        session = self._get_session()
        try:
            stmt = select(FamilyNpiLog).options(
                joinedload(FamilyNpiLog.task).joinedload(TaskCatalogo.categoria),
                joinedload(FamilyNpiLog.family)
            ).where(FamilyNpiLog.DateEnd.is_(None))
            
            if family_id:
                stmt = stmt.where(FamilyNpiLog.FamilyNpiID == family_id)
            
            links = session.scalars(stmt).all()
            return self._detach_list(session, links)
        except Exception as e:
            logger.error(f"Errore in get_family_links: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def get_available_tasks_for_family(self, family_id):
        """
        Recupera i task del catalogo non ancora collegati alla famiglia specificata.
        
        Args:
            family_id: ID della famiglia
            
        Returns:
            Lista di TaskCatalogo non collegati
        """
        from npi.data_models import FamilyNpiLog, TaskCatalogo
        from sqlalchemy import select, and_
        from sqlalchemy.orm import joinedload
        
        session = self._get_session()
        try:
            # Ottieni gli ID dei task già collegati
            linked_task_ids = session.scalars(
                select(FamilyNpiLog.TaskID).where(
                    and_(
                        FamilyNpiLog.FamilyNpiID == family_id,
                        FamilyNpiLog.DateEnd.is_(None)
                    )
                )
            ).all()
            
            # Recupera tutti i task non collegati
            stmt = select(TaskCatalogo).options(
                joinedload(TaskCatalogo.categoria)
            ).order_by(TaskCatalogo.NrOrdin)
            
            if linked_task_ids:
                stmt = stmt.where(TaskCatalogo.TaskID.not_in(linked_task_ids))
            
            tasks = session.scalars(stmt).all()
            return self._detach_list(session, tasks)
        except Exception as e:
            logger.error(f"Errore in get_available_tasks_for_family: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def is_family_link_duplicate(self, family_id, task_id):
        """
        Verifica se esiste già un collegamento attivo tra famiglia e task.
        
        Args:
            family_id: ID della famiglia
            task_id: ID del task
            
        Returns:
            True se il collegamento esiste già, False altrimenti
        """
        from npi.data_models import FamilyNpiLog
        from sqlalchemy import select, and_
        
        session = self._get_session()
        try:
            existing = session.scalars(
                select(FamilyNpiLog).where(
                    and_(
                        FamilyNpiLog.FamilyNpiID == family_id,
                        FamilyNpiLog.TaskID == task_id,
                        FamilyNpiLog.DateEnd.is_(None)
                    )
                )
            ).first()
            
            return existing is not None
        except Exception as e:
            logger.error(f"Errore in is_family_link_duplicate: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def create_family_link(self, family_id, task_id):
        """
        Crea un collegamento tra famiglia e task con validazione duplicati.
        
        Args:
            family_id: ID della famiglia
            task_id: ID del task
            
        Returns:
            FamilyNpiLog creato
        """
        from npi.data_models import FamilyNpiLog
        
        session = self._get_session()
        try:
            # Validazione duplicati
            if self.is_family_link_duplicate(family_id, task_id):
                raise ValueError("Questo task è già collegato alla famiglia")
            
            new_link = FamilyNpiLog(
                FamilyNpiID=family_id,
                TaskID=task_id,
                DateEnd=None
            )
            session.add(new_link)
            session.commit()
            session.refresh(new_link)
            
            logger.info(f"Collegamento creato: Famiglia {family_id} <-> Task {task_id}")
            return self._detach_object(session, new_link)
        except Exception as e:
            logger.error(f"Errore in create_family_link: {e}")
            session.rollback()
            raise
        finally:
            session.close()
    
    def delete_family_link(self, link_id):
        """
        Soft delete di un collegamento famiglia-task.
        
        Args:
            link_id: ID del FamilyNpiLog
            
        Returns:
            True se eliminato, False se non trovato
        """
        from npi.data_models import FamilyNpiLog
        from datetime import datetime
        
        session = self._get_session()
        try:
            link = session.get(FamilyNpiLog, link_id)
            if not link or link.DateEnd is not None:
                return False
            
            link.DateEnd = datetime.now()
            session.commit()
            
            logger.info(f"Collegamento eliminato (soft delete): ID {link_id}")
            return True
        except Exception as e:
            logger.error(f"Errore in delete_family_link: {e}")
            session.rollback()
            raise
        finally:
            session.close()

    # ========================================
    # Auto-Close Project Methods
    # ========================================
    
    def auto_update_project_status(self, project_id):
        """
        Auto-chiude un progetto se tutti i task sono completati.
        
        Args:
            project_id: ID del progetto da verificare
            
        Returns:
            True se il progetto è stato chiuso, False altrimenti
        """
        from npi.data_models import ProgettoNPI, TaskProdotto, WaveNPI
        from sqlalchemy import select
        from datetime import datetime
        
        session = self._get_session()
        try:
            # Recupera tutti i task ASSEGNATI del progetto (OwnerID NOT NULL)
            stmt = (
                select(TaskProdotto)
                .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                .where(WaveNPI.ProgettoID == project_id)
                .where(TaskProdotto.OwnerID.isnot(None))  # Solo task assegnati
            )
            
            # Log della query SQL per debug
            logger.info(f"🔍 Progetto {project_id}: Esecuzione query SQL (solo task assegnati):")
            logger.info(f"  {stmt}")
            
            tasks = session.scalars(stmt).all()
            
            if not tasks:
                logger.info(f"Progetto {project_id}: nessun task trovato")
                return False
            
            # Debug: mostra stato di ogni task
            logger.info(f"🔍 Progetto {project_id}: Analisi {len(tasks)} task")
            for i, task in enumerate(tasks, 1):
                wave_id = task.WaveID
                # Recupera ProgettoID dalla wave per verifica
                wave = session.get(WaveNPI, wave_id) if wave_id else None
                prog_id = wave.ProgettoID if wave else None
                logger.info(f"  Task {i}/{len(tasks)}: TaskID={task.TaskProdottoID}, "
                           f"WaveID={wave_id}, ProgettoID={prog_id}, "
                           f"DataCompletamento={task.DataCompletamento}, "
                           f"Completato={'SI' if task.DataCompletamento is not None else 'NO'}")
            
            # Verifica se tutti i task sono completati
            all_completed = all(task.DataCompletamento is not None for task in tasks)
            
            logger.info(f"Progetto {project_id}: Tutti completati = {all_completed}")
            
            if all_completed:
                # Aggiorna lo stato del progetto
                project = session.get(ProgettoNPI, project_id)
                if project:
                    if project.StatoProgetto != 'Chiuso':
                        logger.info(f"Progetto {project_id}: Cambio stato da '{project.StatoProgetto}' a 'Chiuso'")
                        project.StatoProgetto = 'Chiuso'
                        session.commit()
                        logger.info(f"✅ Progetto {project_id} auto-chiuso: tutti i {len(tasks)} task completati")
                    else:
                        logger.debug(f"Progetto {project_id}: già chiuso (StatoProgetto='{project.StatoProgetto}')")
                    
                    # Ritorna True se tutti i task sono completati, anche se già chiuso
                    return True
                else:
                    logger.warning(f"Progetto {project_id}: NON TROVATO nel database!")
                    return False
            else:
                completed_count = sum(1 for task in tasks if task.DataCompletamento is not None)
                logger.info(f"Progetto {project_id}: {completed_count}/{len(tasks)} task completati - NON CHIUDO")
            
            return False
        except Exception as e:
            logger.error(f"Errore in auto_update_project_status per progetto {project_id}: {e}", exc_info=True)
            session.rollback()
            raise
        finally:
            session.close()
    
    def reopen_project(self, project_id):
        """
        Riapre manualmente un progetto chiuso.
        
        Args:
            project_id: ID del progetto da riaprire
            
        Returns:
            True se riaperto con successo, False altrimenti
        """
        from npi.data_models import ProgettoNPI
        
        session = self._get_session()
        try:
            project = session.get(ProgettoNPI, project_id)
            if not project:
                logger.warning(f"Progetto {project_id} non trovato")
                return False
            
            if project.StatoProgetto == 'Chiuso':
                project.StatoProgetto = 'Attivo'
                session.commit()
                
                logger.info(f"Progetto {project_id} riaperto manualmente: '{project.NomeProgetto}'")
                return True
            else:
                logger.debug(f"Progetto {project_id} non è chiuso (stato: {project.StatoProgetto})")
                return False
        except Exception as e:
            logger.error(f"Errore in reopen_project per progetto {project_id}: {e}")
            session.rollback()
            raise
        finally:
            session.close()
