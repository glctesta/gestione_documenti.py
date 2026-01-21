# File: npi/windows/project_window.py (VERSION: 20251219_1040)
import logging
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl.styles import Alignment
from tkcalendar import DateEntry
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

# Assicurati che i percorsi di importazione siano corretti per la tua struttura
from .import_tasks_window import ImportTasksWindow
from .task_documents_window import TaskDocumentsWindow
from .task_dependencies_window import TaskDependenciesWindow

logger = logging.getLogger(__name__)


class ProjectWindow(tk.Toplevel):

    def __init__(self, master, npi_manager, lang, project_id, master_app, logged_in_user: str,
                 final_customers: list = None):
        super().__init__(master)
        if project_id is None:
            self.destroy()
            return

        if openpyxl is None:
            messagebox.showerror("Libreria Mancante",
                                 "La libreria 'openpyxl' è necessaria per l'esportazione Excel.\n"
                                 "Installala con: pip install openpyxl", parent=self)
            # Non blocchiamo la finestra, ma l'esportazione non funzionerà

        # 1. Imposta gli attributi di base
        self.npi_manager = npi_manager
        self.lang = lang
        self.project_id = project_id
        self.master_app = master_app
        self.logged_in_user = logged_in_user

        if not self.logged_in_user:
            logger.error("ProjectWindow aperta senza un nome utente valido.")
            messagebox.showerror("Errore Critico", "Utente non valido.", parent=master)
            self.destroy()
            return

        # 2. Inizializza le variabili e i dizionari di stato
        self.progetto = None
        self.current_task_id = None
        self.soggetti_map = {}
        self.soggetti_map_rev = {}
        self.doc_types_map = {}
        self.doc_types_map_rev = {}
        self.selected_file_path = None
        self.selected_file_data = None
        self.active_docs_for_task = {}
        self.doc_types_properties = {}
        

        # Mappe per clienti finali
        self.final_customers = final_customers if final_customers else []

        try:
            self.final_clients_map = {fc[1]: fc[0] for fc in self.final_customers}
            self.final_clients_map_rev = {v: k for k, v in self.final_clients_map.items()}
        except AttributeError as e:
            messagebox.showerror("Errore di Configurazione",
                                 f"Impossibile mappare i clienti finali. Attributo non trovato: {e}\nControlla i nomi delle colonne nel modello dati.",
                                 parent=self)
            self.final_clients_map = {}
            self.final_clients_map_rev = {}

        self.fields = {}
        self.doc_widgets = {}

        self.status_map_display = {
            'Da Fare': self.lang.get('status_todo', 'Da Fare'),
            'In Lavorazione': self.lang.get('status_wip', 'In Lavorazione'),
            'Completato': self.lang.get('status_done', 'Completato'),
            'Bloccato': self.lang.get('status_blocked', 'Bloccato'),
        }
        self.status_map_db = {v: k for k, v in self.status_map_display.items()}

        self.geometry("1500x900")
        self.title(self.lang.get('project_window_title', 'Gestione Progetto NPI'))
        self.transient(master)
        self.grab_set()

        self.show_assigned_var = tk.BooleanVar(value=True)
        self.category_filter_var = tk.StringVar(value='')  # Filtro categoria
        self.owner_filter_var = tk.StringVar(value='')     # Filtro persone assegnate
        self.dep_category_var = tk.StringVar(value='')     # Filtro categoria dipendenze
        self.category_map = {}  # Mappa nome categoria -> ID
        self.owner_filter_map = {}  # Mappa nome persona -> SoggettoId per filtro
        self.is_project_owner = False  # Flag per verificare se l'utente loggato è l'owner del progetto
        self.all_available_tasks = [] # Cache per task disponibili (dipendenze)
        self._create_widgets()

    def _create_widgets(self):
        # Main Layout
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Frame per il titolo del progetto con supporto multi-riga
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
        
        self.header_label = ttk.Label(title_frame, text="", font=('Segoe UI', 16, 'bold'), 
                                       wraplength=600, justify=tk.LEFT, anchor='nw')
        self.header_label.pack(side=tk.LEFT, anchor='nw')

        # Buttons in Header - PRIMA RIGA
        toolbar_row1 = ttk.Frame(header_frame)
        toolbar_row1.pack(side=tk.TOP, fill=tk.X, pady=(0, 5))
        
        self.import_button = ttk.Button(toolbar_row1, text=self.lang.get('btn_import_tasks', 'Importa Task'), command=self._launch_import_tasks_window)
        self.import_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # Bottone per sincronizzare task dal catalogo
        self.sync_button = ttk.Button(toolbar_row1, text=self.lang.get('btn_sync_catalog', 'Sincronizza Catalogo'), command=self._sync_catalog_tasks)
        self.sync_button.pack(side=tk.LEFT, padx=5)
        
        self.export_button = ttk.Button(toolbar_row1, text=self.lang.get('btn_export_excel', 'Export Excel'), command=self._export_cost_report, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT, padx=5)

        # SECONDA RIGA - Filtri e checkbox
        toolbar_row2 = ttk.Frame(header_frame)
        toolbar_row2.pack(side=tk.TOP, fill=tk.X)
        
        ttk.Checkbutton(toolbar_row2, text=self.lang.get('show_assigned', 'Mostra Assegnati'), variable=self.show_assigned_var, command=self._populate_treeview).pack(side=tk.LEFT, padx=(0, 10))

        # Filtro per categoria
        ttk.Label(toolbar_row2, text=self.lang.get('filter_category', 'Categoria:')).pack(side=tk.LEFT, padx=(0, 5))
        self.category_filter_combo = ttk.Combobox(toolbar_row2, textvariable=self.category_filter_var, state='readonly', width=20)
        self.category_filter_combo.pack(side=tk.LEFT, padx=5)
        self.category_filter_combo.bind('<<ComboboxSelected>>', lambda e: self._populate_treeview())
        
        # Filtro per persona assegnata
        ttk.Label(toolbar_row2, text=self.lang.get('filter_owner', 'Assegnato a:')).pack(side=tk.LEFT, padx=(10, 5))
        self.owner_filter_combo = ttk.Combobox(toolbar_row2, textvariable=self.owner_filter_var, state='readonly', width=20)
        self.owner_filter_combo.pack(side=tk.LEFT, padx=5)
        self.owner_filter_combo.bind('<<ComboboxSelected>>', lambda e: self._populate_treeview())

        # Project Dates
        dates_frame = ttk.LabelFrame(header_frame, text=self.lang.get('project_dates_title', 'Date Progetto'))
        dates_frame.pack(side=tk.RIGHT, padx=10)
        
        # Prima riga: Date
        dates_row1 = ttk.Frame(dates_frame)
        dates_row1.pack(fill=tk.X, padx=2, pady=2)
        
        ttk.Label(dates_row1, text=self.lang.get('start_date', 'Inizio:')).pack(side=tk.LEFT, padx=5)
        self.project_start_date_entry = DateEntry(dates_row1, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.project_start_date_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(dates_row1, text=self.lang.get('due_date', 'Scadenza:')).pack(side=tk.LEFT, padx=5)
        self.project_due_date_entry = DateEntry(dates_row1, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.project_due_date_entry.pack(side=tk.LEFT, padx=5)
        
        # Seconda riga: Versione e Salva
        dates_row2 = ttk.Frame(dates_frame)
        dates_row2.pack(fill=tk.X, padx=2, pady=2)
        
        ttk.Label(dates_row2, text=self.lang.get('label_version', 'Versione:')).pack(side=tk.LEFT, padx=5)
        self.project_version_entry = ttk.Entry(dates_row2, width=10)
        self.project_version_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(dates_row2, text=self.lang.get('save_dates', 'Salva Date'), command=self._save_project_dates).pack(side=tk.LEFT, padx=5)
        
        # Terza riga: Avviso Milestone
        dates_row3 = ttk.Frame(dates_frame)
        dates_row3.pack(fill=tk.X, padx=2, pady=(2, 0))
        
        self.milestone_warning_label = ttk.Label(
            dates_row3,
            text="",
            foreground="orange",
            font=('Helvetica', 9, 'bold')  # Cambiato da italic a bold
        )
        self.milestone_warning_label.pack(side=tk.LEFT, padx=5)
        
        # 🆕 GERARCHIA PROGETTI (Parent-Child)
        hierarchy_frame = ttk.LabelFrame(header_frame, text="🔗 Gerarchia Progetti")
        hierarchy_frame.pack(side=tk.RIGHT, padx=10)
        
        # Riga 1: Progetto Padre
        hierarchy_row1 = ttk.Frame(hierarchy_frame)
        hierarchy_row1.pack(fill=tk.X, padx=2, pady=2)
        
        ttk.Label(hierarchy_row1, text="Progetto Padre:").pack(side=tk.LEFT, padx=5)
        self.parent_project_combo = ttk.Combobox(hierarchy_row1, state='readonly', width=25)
        self.parent_project_combo.pack(side=tk.LEFT, padx=5)
        self.parent_project_combo.bind('<<ComboboxSelected>>', self._on_parent_project_selected)
        
        # Riga 2: Info progetto padre corrente
        hierarchy_row2 = ttk.Frame(hierarchy_frame)
        hierarchy_row2.pack(fill=tk.X, padx=2, pady=2)
        
        self.current_parent_label = ttk.Label(
            hierarchy_row2,
            text="",
            foreground="blue",
            font=('Helvetica', 9)
        )
        self.current_parent_label.pack(side=tk.LEFT, padx=5)
        
        # Riga 3: Progetti figli count
        hierarchy_row3 = ttk.Frame(hierarchy_frame)
        hierarchy_row3.pack(fill=tk.X, padx=2, pady=2)
        
        self.children_projects_label = ttk.Label(
            hierarchy_row3,
            text="",
            foreground="green",
            font=('Helvetica', 9)
        )
        self.children_projects_label.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            hierarchy_row3,
            text="📋 Mostra Figli",
            command=self._show_child_projects_dialog
        ).pack(side=tk.LEFT, padx=5)


        # Content Split
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        # Left Panel: Treeview
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)

        columns = ('Category', 'Name', 'Owner', 'Status', 'StartDate', 'DueDate')
        self.tree = ttk.Treeview(left_frame, columns=columns, show='headings')
        self.tree.heading('Category', text=self.lang.get('col_category', 'Categoria'))
        self.tree.heading('Name', text=self.lang.get('col_task', 'Task'))
        self.tree.heading('Owner', text=self.lang.get('col_owner', 'Assegnato a'))
        self.tree.heading('Status', text=self.lang.get('col_status', 'Stato'))
        self.tree.heading('StartDate', text=self.lang.get('col_start_date', 'Inizio'))
        self.tree.heading('DueDate', text=self.lang.get('col_due_date', 'Scadenza'))
        
        self.tree.column('Category', width=120)
        self.tree.column('Name', width=200)
        self.tree.column('Owner', width=100)
        self.tree.column('Status', width=100)
        self.tree.column('StartDate', width=100)
        self.tree.column('DueDate', width=100)
        
        self.tree.tag_configure('special_task', foreground='#0078d4')  # BLU per Target NPI
        self.tree.tag_configure('late_task', foreground='red')  # ROSSO per task in ritardo
        self.tree.tag_configure('bold_task', font=('Segoe UI', 9, 'bold'))
        self.tree.bind('<<TreeviewSelect>>', self._on_task_select)
        
        scrollbar = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Right Panel: Details
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)

        # Task Details Form
        details_frame = ttk.LabelFrame(right_frame, text=self.lang.get('task_details', 'Dettagli Task'))
        details_frame.pack(fill=tk.X, padx=5, pady=5)

        # Fields
        grid_frame = ttk.Frame(details_frame)
        grid_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Name & Category (Read-only)
        ttk.Label(grid_frame, text=self.lang.get('label_task', 'Task:')).grid(row=0, column=0, sticky=tk.W)
        self.fields['task_name'] = ttk.Label(grid_frame, text="", font=('Segoe UI', 10, 'bold'))
        self.fields['task_name'].grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(grid_frame, text=self.lang.get('label_category', 'Categoria:')).grid(row=1, column=0, sticky=tk.W)
        self.fields['task_category'] = ttk.Label(grid_frame, text="")
        self.fields['task_category'].grid(row=1, column=1, sticky=tk.W)

        # Editable Fields
        labels_config = [
            ('OwnerID', self.lang.get('label_owner', 'Assegnato a'), 'combo'),
            ('Stato', self.lang.get('label_status', 'Stato'), 'combo'),
            ('DataInizio', self.lang.get('label_start_date', 'Data Inizio'), 'date'),
            ('DataScadenza', self.lang.get('label_due_date', 'Data Scadenza'), 'date'),
            ('DataCompletamento', self.lang.get('label_completion_date', 'Data Completamento'), 'date'),
        ]

        r = 2
        for field_name, label_text, widget_type in labels_config:
            ttk.Label(grid_frame, text=label_text).grid(row=r, column=0, sticky=tk.W, pady=2)
            if widget_type == 'combo':
                w = ttk.Combobox(grid_frame, state='readonly')
                # Aggiungi binding per il campo Stato per gestire DataCompletamento
                if field_name == 'Stato':
                    w.bind('<<ComboboxSelected>>', self._on_status_change)
            elif widget_type == 'date':
                w = DateEntry(grid_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
            self.fields[field_name] = w
            w.grid(row=r, column=1, sticky=tk.EW, pady=2)
            r += 1

        # Checkbox IsPostFinalMilestone
        self.fields['IsPostFinalMilestone'] = ttk.Checkbutton(grid_frame, text="Target NPI", command=self._on_target_npi_change)
        self.fields['IsPostFinalMilestone'].var = tk.BooleanVar()
        self.fields['IsPostFinalMilestone'].config(variable=self.fields['IsPostFinalMilestone'].var)
        self.fields['IsPostFinalMilestone'].grid(row=r, column=1, sticky=tk.W, pady=2)
        r += 1

        # Notes
        ttk.Label(grid_frame, text=self.lang.get('label_notes', 'Note')).grid(row=r, column=0, sticky=tk.NW, pady=2)
        self.fields['Note'] = tk.Text(grid_frame, height=4, width=40)
        self.fields['Note'].grid(row=r, column=1, sticky=tk.EW, pady=2)
        r += 1

        # Dependencies Section
        deps_label_frame = ttk.LabelFrame(details_frame, text=self.lang.get('dependencies_summary', 'Sommario Dipendenze'))
        deps_label_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.deps_listbox = tk.Listbox(deps_label_frame, height=3, state=tk.DISABLED)
        self.deps_listbox.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_manage_deps = ttk.Button(deps_label_frame, text=self.lang.get('btn_manage_dependencies', 'Gestionează Dipendenze'), 
                                         command=self._launch_task_dependencies_window)
        self.btn_manage_deps.pack(pady=5)

        # Bottoni di azione
        action_buttons_frame = ttk.Frame(details_frame)
        action_buttons_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(action_buttons_frame, text=self.lang.get('btn_save', 'Salva Modifiche'), command=self._save_task_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_buttons_frame, text=self.lang.get('btn_delete_task', 'Elimina Task'), command=self._delete_task_prodotto).pack(side=tk.LEFT, padx=5)

        # Documents Section
        doc_frame = ttk.LabelFrame(right_frame, text=self.lang.get('documents_title', 'Documenti'))
        doc_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Doc Form
        df = ttk.Frame(doc_frame)
        df.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(df, text=self.lang.get('doc_type', 'Tipo:')).grid(row=0, column=0, sticky=tk.W)
        self.doc_widgets['type'] = ttk.Combobox(df, state='readonly')
        self.doc_widgets['type'].grid(row=0, column=1, columnspan=2, sticky=tk.EW)
        self.doc_widgets['type'].bind('<<ComboboxSelected>>', self._on_doc_type_selected)

        ttk.Label(df, text=self.lang.get('doc_title', 'Titolo:')).grid(row=1, column=0, sticky=tk.W)
        self.doc_widgets['title'] = ttk.Entry(df)
        self.doc_widgets['title'].grid(row=1, column=1, columnspan=2, sticky=tk.EW)

        ttk.Label(df, text=self.lang.get('final_client', 'Cliente Finale:')).grid(row=2, column=0, sticky=tk.W)
        self.doc_widgets['final_client'] = ttk.Combobox(df, values=[fc[1] for fc in self.final_customers])
        self.doc_widgets['final_client'].grid(row=2, column=1, columnspan=2, sticky=tk.EW)

        # Dynamic fields
        self.doc_widgets['value_label'] = ttk.Label(df, text=self.lang.get('doc_value', 'Valore (€):'))
        self.doc_widgets['value_entry'] = ttk.Entry(df)
        
        self.doc_widgets['due_date_label'] = ttk.Label(df, text=self.lang.get('doc_due_date', 'Scadenza Doc:'))
        self.doc_widgets['due_date_entry'] = DateEntry(df, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')

        # Replacement logic
        self.doc_widgets['is_replacement_var'] = tk.BooleanVar()
        cb = ttk.Checkbutton(df, text=self.lang.get('is_replacement', 'Sostituisce doc esistente'), variable=self.doc_widgets['is_replacement_var'], command=self._toggle_replacement_combo)
        cb.grid(row=5, column=0, columnspan=3, sticky=tk.W)
        
        self.doc_widgets['replaces_combo'] = ttk.Combobox(df, state='disabled')
        self.doc_widgets['replaces_combo'].grid(row=6, column=0, columnspan=3, sticky=tk.EW)

        # File selection
        ttk.Button(df, text=self.lang.get('select_file', 'Seleziona File...'), command=self._select_file).grid(row=7, column=0, sticky=tk.W)
        self.doc_widgets['file_label'] = ttk.Label(df, text=self.lang.get('no_file_selected', 'Nessun file'))
        self.doc_widgets['file_label'].grid(row=7, column=1, columnspan=2, sticky=tk.W)

        # Note Doc
        ttk.Label(df, text=self.lang.get('notes', 'Note:')).grid(row=8, column=0, sticky=tk.NW)
        self.doc_widgets['note_text'] = tk.Text(df, height=3, width=30)
        self.doc_widgets['note_text'].grid(row=8, column=1, columnspan=2, sticky=tk.EW)

        ttk.Button(df, text=self.lang.get('save_doc', 'Carica Documento'), command=self._save_document).grid(row=9, column=1, pady=5)
        
        self.view_docs_button = ttk.Button(doc_frame, text=self.lang.get('view_docs', 'Vedi Documenti Caricati'), command=self._launch_view_documents_window, state=tk.DISABLED)
        self.view_docs_button.pack(pady=5)

        # Status Bar
        self.status_bar = ttk.Label(self, text="Inizializzazione...", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Initial call to load data
        self.after(100, self._load_data_and_populate_ui)

    def log_status(self, message):
        """Aggiorna la barra di stato e il log di sistema."""
        try:
            if hasattr(self, 'status_bar'):
                self.status_bar.config(text=message)
            logger.info(message)
        except Exception as e:
            print(f"Error logging status: {e}")

    def _on_doc_type_selected(self, event=None):
        doc_type_desc = self.doc_widgets['type'].get()
        if not doc_type_desc: return
        doc_type_id = self.doc_types_map[doc_type_desc]
        props = self.doc_types_properties[doc_type_id]
        if props['has_value']:
            self.doc_widgets['value_label'].grid(row=3, column=0, sticky=tk.W, pady=2)
            self.doc_widgets['value_entry'].grid(row=3, column=1, columnspan=2, sticky=tk.EW, pady=2)
        else:
            self.doc_widgets['value_label'].grid_remove()
            self.doc_widgets['value_entry'].grid_remove()
        if props['check_date']:
            self.doc_widgets['due_date_label'].grid(row=4, column=0, sticky=tk.W, pady=2)
            self.doc_widgets['due_date_entry'].grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=2)
        else:
            self.doc_widgets['due_date_label'].grid_remove()
            self.doc_widgets['due_date_entry'].grid_remove()

    def _load_data_and_populate_ui(self):
        try:
            soggetti = self.npi_manager.get_soggetti()
            self.soggetti_map = {s.Nome: s.SoggettoId for s in soggetti}
            self.soggetti_map_rev = {v: k for k, v in self.soggetti_map.items()}
            self.fields['OwnerID']['values'] = [''] + list(self.soggetti_map.keys())
            self.fields['Stato']['values'] = list(self.status_map_display.values())

            # Carica le categorie per il filtro
            categories = self.npi_manager.get_categories()
            self.category_map = {cat.Category: cat.CategoryId for cat in categories}
            all_categories_label = self.lang.get('all_categories', 'Tutte le categorie')
            category_values = [all_categories_label] + list(self.category_map.keys())
            self.category_filter_combo['values'] = category_values
            self.category_filter_var.set(all_categories_label)  # Seleziona "Tutte" di default

            doc_types = self.npi_manager.get_npi_document_types()
            self.doc_types_map = {dt.NpiDocumentDescription: dt.NpiDocumentTypeId for dt in doc_types}
            self.doc_types_map_rev = {v: k for k, v in self.doc_types_map.items()}
            self.doc_types_properties = {dt.NpiDocumentTypeId: {'has_value': dt.HasValue, 'check_date': dt.CheckDate}
                                         for dt in doc_types}
            self.doc_widgets['type']['values'] = list(self.doc_types_map.keys())

            self.progetto = self.npi_manager.get_dettagli_progetto(self.project_id)
            if not self.progetto: raise ValueError("Progetto non trovato.")

            # Costruisci il titolo con versione se presente
            version_str = f" (v{self.progetto.Version})" if self.progetto.Version else ""
            title = f"{self.progetto.prodotto.CodiceProdotto or 'N/A'} - {self.progetto.prodotto.NomeProdotto}{version_str}"
            self.header_label.config(text=title)
            self.project_start_date_entry.set_date(self.progetto.DataInizio)
            self.project_due_date_entry.set_date(self.progetto.ScadenzaProgetto)
            
            # Popola il campo Version
            if self.progetto.Version:
                self.project_version_entry.delete(0, tk.END)
                self.project_version_entry.insert(0, self.progetto.Version)

            # --- LOGICA FILTRO PERSONE ---
            # Determina se l'utente loggato è l'owner del progetto
            project_owner_id = self.progetto.OwnerID if self.progetto else None
            project_owner_name = None
            
            if project_owner_id and project_owner_id in self.soggetti_map_rev:
                project_owner_name = self.soggetti_map_rev[project_owner_id]
                self.is_project_owner = (project_owner_name == self.logged_in_user)
            else:
                # Nessun owner assegnato al progetto, l'utente ha accesso completo
                self.is_project_owner = True
            
            # Raccoglie tutte le persone che hanno task assegnati
            assigned_owners = set()
            if self.progetto and self.progetto.waves:
                for wave in self.progetto.waves:
                    for task in wave.tasks:
                        if task.OwnerID and task.OwnerID in self.soggetti_map_rev:
                            assigned_owners.add(self.soggetti_map_rev[task.OwnerID])
            
            # Popola il filtro persone solo con le persone che hanno task assegnati
            all_owners_label = self.lang.get('all_owners', 'Tutte le persone')
            self.owner_filter_map = {nome: self.soggetti_map[nome] for nome in assigned_owners if nome in self.soggetti_map}
            owner_values = [all_owners_label] + sorted(list(self.owner_filter_map.keys()))
            self.owner_filter_combo['values'] = owner_values
            
            # Applica la logica di auto-selezione e blocco
            if not self.is_project_owner:
                # L'utente NON è l'owner del progetto
                if self.logged_in_user in self.owner_filter_map:
                    # L'utente loggato ha task assegnati, imposta il filtro su di lui
                    self.owner_filter_var.set(self.logged_in_user)
                    self.owner_filter_combo.config(state='disabled')
                    logger.info(f"Filtro persone impostato su '{self.logged_in_user}' e bloccato (non è l'owner del progetto)")
                else:
                    # L'utente loggato non ha task assegnati, mostra tutti ma blocca il combo
                    self.owner_filter_var.set(all_owners_label)
                    self.owner_filter_combo.config(state='disabled')
                    logger.info(f"Utente '{self.logged_in_user}' non ha task assegnati. Filtro bloccato su 'Tutte le persone'")
            else:
                # L'utente È l'owner del progetto, combo abilitato
                self.owner_filter_var.set(all_owners_label)
                self.owner_filter_combo.config(state='readonly')
                logger.info(f"Utente '{self.logged_in_user}' è l'owner del progetto. Filtro persone abilitato")

            self._populate_treeview()
            
            # Controlla se esiste un task milestone finale
            self._check_milestone_status()
            
            # 🆕 Popola gerarchia progetti
            self._populate_project_hierarchy()

            if openpyxl: self.export_button.config(state=tk.NORMAL)

        except Exception as e:
            logger.error(f"Errore caricamento dati: {e}", exc_info=True)
            messagebox.showerror("Errore Caricamento", f"Impossibile caricare i dati:\n{e}", parent=self)
            self.destroy()

        self._disable_form()
        self._disable_doc_form()
        
        # Abilita bottone visualizzazione documenti
        self.view_docs_button.config(state=tk.NORMAL)

    def _export_cost_report(self):
        if not self.progetto:
            messagebox.showwarning("Warning", "No project loaded.", parent=self)
            return

        temp_dir = "C:\\temp"
        try:
            os.makedirs(temp_dir, exist_ok=True)
        except OSError as e:
            messagebox.showerror("Folder Creation Error", f"Could not create folder {temp_dir}:\n{e}", parent=self)
            return

        product_code_file = self.progetto.prodotto.CodiceProdotto.replace(" ",
                                                                          "_") if self.progetto.prodotto.CodiceProdotto else "NO_CODE"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(temp_dir, f"Cost_Report_{product_code_file}_{timestamp}.xlsx")

        messagebox.showinfo("Exporting", f"Creating the report...\nThe file will be saved in:\n{file_path}",
                            parent=self)
        self.update_idletasks()

        report_rows = []
        grand_total = 0.0

        for task in self.progetto.waves[0].tasks:
            if task.documents:
                task_total = 0.0
                task_docs_data = []
                for doc in task.documents:
                    if doc.ValueInEur and doc.ValueInEur > 0:
                        doc_value = doc.ValueInEur
                        task_docs_data.append([
                            doc.document_type.NpiDocumentDescription if doc.document_type else "N/A",
                            doc.DocumentTitle,
                            doc_value
                        ])
                        task_total += doc_value

                if task_docs_data:
                    task_name = task.task_catalogo.NomeTask if task.task_catalogo else "Unknown Task"
                    report_rows.append({'type': 'task', 'data': [task_name]})
                    for doc_data in task_docs_data:
                        report_rows.append({'type': 'doc', 'data': doc_data})
                    report_rows.append({'type': 'subtotal', 'data': ["Task Subtotal", "", task_total]})
                    grand_total += task_total

        if not report_rows:
            messagebox.showinfo("No Data", "No documents with costs found for this project.", parent=self)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cost Report"

            bold_font = Font(bold=True)
            header_font = Font(bold=True, size=14)
            currency_format = '#,##0.00 "€"'  # Formato valuta

            ws.merge_cells('A1:C1');
            cell = ws['A1']
            cell.value = f"Project Cost Report: {self.progetto.prodotto.NomeProdotto}"
            cell.font = header_font;
            cell.alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:C2');
            cell = ws['A2']
            cell.value = f"Product Code: {self.progetto.prodotto.CodiceProdotto or 'N/A'}"
            cell.font = bold_font

            headers = ["Task / Document Type", "Document Title", "Cost (€)"]
            ws.append([])
            ws.append(headers)
            for col_letter in ['A', 'B', 'C']: ws[f'{col_letter}4'].font = bold_font

            for row_info in report_rows:
                row_num = ws.max_row + 1
                if row_info['type'] == 'task':
                    ws.cell(row_num, 1).value = row_info['data'][0]
                    ws.cell(row_num, 1).font = bold_font
                elif row_info['type'] == 'doc':
                    ws.cell(row_num, 1).value = f"  {row_info['data'][0]}"
                    ws.cell(row_num, 2).value = row_info['data'][1]
                    cell = ws.cell(row_num, 3)
                    cell.value = row_info['data'][2]
                    cell.number_format = currency_format
                elif row_info['type'] == 'subtotal':
                    ws.cell(row_num, 1).value = row_info['data'][0]
                    ws.cell(row_num, 1).font = bold_font
                    ws.cell(row_num, 1).alignment = Alignment(horizontal='right')
                    cell = ws.cell(row_num, 3)
                    cell.value = row_info['data'][2]
                    cell.font = bold_font
                    cell.number_format = currency_format
                    ws.append([])

            ws.append([])
            total_row = ws.max_row + 1
            ws.cell(total_row, 1).value = "PROJECT GRAND TOTAL"
            ws.cell(total_row, 1).font = header_font
            cell = ws.cell(total_row, 3)
            cell.value = grand_total
            cell.font = header_font
            cell.number_format = currency_format

            for col in ['A', 'B']: ws.column_dimensions[col].width = 50
            ws.column_dimensions['C'].width = 20

            wb.save(file_path)

            if messagebox.askyesno("Export Complete",
                                   f"Cost report saved successfully to:\n{file_path}\n\nDo you want to open the file now?",
                                   parent=self):
                try:
                    os.startfile(file_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not open the file:\n{e}", parent=self)

        except Exception as e:
            logger.error(f"Error creating the Excel file: {e}", exc_info=True)
            messagebox.showerror("Export Error", f"Could not create the Excel file:\n{e}", parent=self)

    def _populate_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        if not self.progetto or not self.progetto.waves: return
        wave = self.progetto.waves[0]
        self.import_button.config(state=tk.DISABLED if any(t.OwnerID is not None for t in wave.tasks) else tk.NORMAL)

        # Ottieni la categoria selezionata per il filtro
        selected_category = self.category_filter_var.get()
        all_categories_label = self.lang.get('all_categories', 'Tutte le categorie')
        filter_category_id = None
        
        if selected_category and selected_category != all_categories_label:
            filter_category_id = self.category_map.get(selected_category)

        # Ottieni la persona selezionata per il filtro
        selected_owner = self.owner_filter_var.get()
        all_owners_label = self.lang.get('all_owners', 'Tutte le persone')
        filter_owner_id = None
        
        if selected_owner and selected_owner != all_owners_label:
            filter_owner_id = self.owner_filter_map.get(selected_owner)

        # Filtra i task
        filtered_tasks = []
        for task in wave.tasks:
            if self.show_assigned_var.get() != (task.OwnerID is not None): 
                continue
            
            # Applica filtro categoria se selezionato
            if filter_category_id is not None:
                task_category_id = task.task_catalogo.CategoryId if task.task_catalogo else None
                if task_category_id != filter_category_id:
                    continue
            
            # Applica filtro persona se selezionato
            if filter_owner_id is not None:
                if task.OwnerID != filter_owner_id:
                    continue
            
            filtered_tasks.append(task)
        
        # Ordina i task con ordinamento topologico rispettando le dipendenze
        sorted_tasks = self._topological_sort_tasks(filtered_tasks)
        
        # Visualizza i task ordinati
        for task in sorted_tasks:
            owner = task.owner.Nome if task.owner else ""
            start_date = task.DataInizio.strftime('%d/%m/%Y') if task.DataInizio else ""
            due_date = task.DataScadenza.strftime('%d/%m/%Y') if task.DataScadenza else ""

            # Se la lista dei documenti del task non è vuota, aggiungi un asterisco.
            if task.documents:
                due_date+= " *"

            status = self.status_map_display.get(task.Stato, task.Stato)
            
            # 🆕 Logica tag con priorità: Target NPI > Task in Ritardo
            tags = []
            
            # Verifica se in ritardo (data scadenza passata e non completato)
            from datetime import datetime
            today = datetime.now().date()
            
            # Converti DataScadenza a date per il confronto
            if task.DataScadenza:
                task_due_date = task.DataScadenza.date() if hasattr(task.DataScadenza, 'date') else task.DataScadenza
                is_late = (task_due_date < today and task.Stato != 'Completato')
            else:
                is_late = False
            
            if task.IsPostFinalMilestone:
                # Task Target NPI = BLU (anche se in ritardo)
                tags.append('special_task')
                tags.append('bold_task')
            elif is_late:
                # Task in ritardo = ROSSO
                tags.append('late_task')
                tags.append('bold_task')
            
            cat = task.task_catalogo.categoria.Category if task.task_catalogo and task.task_catalogo.categoria else ""
            name = task.task_catalogo.NomeTask if task.task_catalogo else "Catalogo non trovato"
            self.tree.insert('', tk.END, text=task.TaskProdottoID, values=(cat, name, owner, status, start_date, due_date),
                             tags=tuple(tags))
    
    def _check_milestone_status(self):
        """Controlla se esiste un task definito come milestone finale e mostra avviso se mancante."""
        try:
            if not self.progetto or not self.progetto.waves:
                return
            
            wave = self.progetto.waves[0]
            
            # Cerca task con IsPostFinalMilestone = True
            has_milestone = any(task.IsPostFinalMilestone for task in wave.tasks)
            
            if has_milestone:
                # Milestone definita, nascondi warning
                self.milestone_warning_label.config(text="")
            else:
                # Nessuna milestone definita, mostra warning
                warning_text = self.lang.get(
                    'warning_no_milestone_defined',
                    '⚠ Nessun task definito come milestone finale'
                )
                self.milestone_warning_label.config(text=warning_text)
                
        except Exception as e:
            logger.error(f"Errore controllo milestone: {e}", exc_info=True)
    
    def _topological_sort_tasks(self, tasks):
        """Ordina i task rispettando le dipendenze con ordinamento topologico.
        
        I task senza dipendenze sono ordinati per data di inizio.
        I task con dipendenze appaiono sempre dopo i loro predecessori.
        """
        from collections import defaultdict, deque
        
        # Crea una mappa task_id -> task per accesso rapido
        task_map = {t.TaskProdottoID: t for t in tasks}
        task_ids = set(task_map.keys())
        
        # Costruisci il grafo delle dipendenze
        # graph[task_id] = lista di task che dipendono da task_id
        graph = defaultdict(list)
        # in_degree[task_id] = numero di dipendenze (predecessori)
        in_degree = defaultdict(int)
        
        # Inizializza tutti i task con in_degree 0
        for task_id in task_ids:
            in_degree[task_id] = 0
        
        # Popola il grafo e calcola gli in_degree
        for task in tasks:
            dependencies = self.npi_manager.get_task_dependencies(task.TaskProdottoID)
            for dep in dependencies:
                predecessor_id = dep.DependsOnTaskProdottoID
                # Considera solo dipendenze tra task nel set corrente
                if predecessor_id in task_ids:
                    graph[predecessor_id].append(task.TaskProdottoID)
                    in_degree[task.TaskProdottoID] += 1
        
        # Ottieni tutti i task senza dipendenze (in_degree == 0)
        # e ordinali per data di inizio
        queue = []
        for task_id in task_ids:
            if in_degree[task_id] == 0:
                task = task_map[task_id]
                # Usa la data di inizio per l'ordinamento, o una data molto lontana se None
                start_date = task.DataInizio if task.DataInizio else datetime(2099, 12, 31)
                queue.append((start_date, task_id))
        
        # Ordina i task iniziali per data di inizio
        queue.sort(key=lambda x: x[0])
        queue = deque([task_id for _, task_id in queue])
        
        # Algoritmo di Kahn per ordinamento topologico
        result = []
        
        while queue:
            # Prendi il prossimo task dalla coda
            current_id = queue.popleft()
            result.append(task_map[current_id])
            
            # Per ogni task che dipende dal task corrente
            successors_to_add = []
            for successor_id in graph[current_id]:
                in_degree[successor_id] -= 1
                
                # Se tutte le dipendenze del successor sono state soddisfatte
                if in_degree[successor_id] == 0:
                    task = task_map[successor_id]
                    start_date = task.DataInizio if task.DataInizio else datetime(2099, 12, 31)
                    successors_to_add.append((start_date, successor_id))
            
            # Ordina i successori per data di inizio prima di aggiungerli alla coda
            successors_to_add.sort(key=lambda x: x[0])
            for _, successor_id in successors_to_add:
                queue.append(successor_id)
        
        # Verifica che tutti i task siano stati processati (no cicli)
        if len(result) != len(tasks):
            logger.warning("Rilevato ciclo nelle dipendenze dei task. Alcuni task potrebbero non essere visualizzati correttamente.")
            # Aggiungi i task rimanenti alla fine (fallback)
            processed_ids = {t.TaskProdottoID for t in result}
            remaining = [t for t in tasks if t.TaskProdottoID not in processed_ids]
            result.extend(remaining)
        
        return result

    def _on_target_npi_change(self):
        """Gestisce il cambio immediato del flag Target NPI."""
        if not self.current_task_id: return
        
        is_target = self.fields['IsPostFinalMilestone'].var.get()
        
        # Se l'utente ha deselezionato, permettiamo (nessun target)
        # Se ha selezionato, chiamiamo il manager per l'esclusività
        
        if is_target:
            try:
                self.npi_manager.set_target_npi_task(self.current_task_id, self.project_id)
                # Aggiorna il modello locale (opzionale se ricarichiamo tutto)
                # Ricarica tutto per aggiornare la treeview e gli altri task
                self._load_data_and_populate_ui()
                # Ripristina la selezione
                # (omesso per semplicità, l'utente vedrà il refresh)
            except Exception as e:
                logger.error(f"Errore impostazione Target NPI: {e}")
                messagebox.showerror("Errore", f"Impossibile impostare il Target NPI:\n{e}", parent=self)
                # Revert checkbox
                self.fields['IsPostFinalMilestone'].var.set(False)
        else:
             # Se deseleziona, dobbiamo aggiornare solo questo task nel DB?
             # La richiesta dice "solo 1 task... deve essere marchato". 
             # Se deseleziono, nessuno è target.
             # Implementiamo un update semplice per questo caso o usiamo set_target_npi_task?
             # set_target_npi_task imposta a True.
             # Per impostare a False, usiamo update_task_prodotto standard o una nuova chiamata.
             # Usiamo update standard per semplicità.
             try:
                 self.npi_manager.update_task_prodotto(self.current_task_id, {'IsPostFinalMilestone': False})
                 self._load_data_and_populate_ui()
             except Exception as e:
                 logger.error(f"Errore rimozione Target NPI: {e}")
                 messagebox.showerror("Errore", f"Impossibile rimuovere il Target NPI:\n{e}", parent=self)
                 self.fields['IsPostFinalMilestone'].var.set(True)

    def _on_task_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            self._disable_form();
            self._disable_doc_form()
            self.view_docs_button.config(state=tk.DISABLED);
            return
        self.current_task_id = int(self.tree.item(sel[0], 'text'))
        task = self._get_task_by_id(self.current_task_id)
        if task:
            self._populate_form(task);
            self._enable_form();
            self._enable_doc_form()
            docs = self.npi_manager.get_documents_for_task(self.current_task_id)
            self.active_docs_for_task = {doc.NpiDocumentId: doc.DocumentTitle for doc in docs if not doc.DateOut}
            self.doc_widgets['replaces_combo']['values'] = list(self.active_docs_for_task.values())
            self.view_docs_button.config(state=tk.NORMAL if docs else tk.DISABLED)
            self._reset_doc_form()
            
            # Load dependencies summary
            self._load_task_dependencies()
        else:
            self._disable_form();
            self._disable_doc_form()

    def _load_task_dependencies(self):
        """Carica e visualizza il sommario delle dipendenze del task corrente."""
        if not self.current_task_id:
            return
        
        self.deps_listbox.config(state=tk.NORMAL)
        self.deps_listbox.delete(0, tk.END)
        dependencies = self.npi_manager.get_task_dependencies(self.current_task_id)
        
        if not dependencies:
            self.deps_listbox.insert(tk.END, self.lang.get('no_dependencies', 'Nessuna dipendenza definita'))
        else:
            for dep in dependencies:
                pred = dep.depends_on_task
                if pred and pred.task_catalogo:
                    task_name = pred.task_catalogo.NomeTask
                    self.deps_listbox.insert(tk.END, f"• {task_name}")
        
        self.deps_listbox.config(state=tk.DISABLED)

    def _launch_task_dependencies_window(self):
        """Apre la finestra dedicata per la gestione delle dipendenze."""
        if not self.current_task_id:
            return
        
        task = self._get_task_by_id(self.current_task_id)
        if not task: return
        
        task_name = task.task_catalogo.NomeTask if task.task_catalogo else "N/A"
        wave_id = self.progetto.waves[0].WaveID if self.progetto and self.progetto.waves else None
        
        def on_close():
            self._load_task_dependencies()
            win.destroy()

        win = TaskDependenciesWindow(
            self, self.npi_manager, self.lang, 
            self.current_task_id, self.project_id, wave_id, task_name
        )
        # Quando la finestra si chiude, ricarichiamo il sommario
        win.protocol("WM_DELETE_WINDOW", on_close)
    
    def _reset_doc_form(self):
        self.doc_widgets['type'].set('')
        self.doc_widgets['title'].delete(0, tk.END)
        self.doc_widgets['final_client'].set('')
        self.doc_widgets['value_entry'].delete(0, tk.END)
        self.doc_widgets['due_date_entry'].set_date(None)
        self.doc_widgets['replaces_combo'].set('')
        self.doc_widgets['is_replacement_var'].set(False)
        self.doc_widgets['note_text'].delete('1.0', tk.END)
        self.selected_file_path = None
        self.selected_file_data = None
        self.doc_widgets['file_label'].config(text=self.lang.get('no_file_selected'))
        self._toggle_replacement_combo()

    def _populate_form(self, task):
        name = task.task_catalogo.NomeTask if task.task_catalogo else "Task non definito"
        cat = task.task_catalogo.categoria.Category if task.task_catalogo and task.task_catalogo.categoria else ""
        self.fields['task_name'].config(text=name);
        self.fields['task_category'].config(text=cat)
        # Mostra il nome dell'owner, non l'ID
        owner_name = task.owner.Nome if task.owner else ""
        self.fields['OwnerID'].set(owner_name)
        # Imposta lo stato - assicurati che sia nella lista dei valori
        stato_display = self.status_map_display.get(task.Stato, task.Stato)
        if stato_display not in self.fields['Stato']['values']:
            # Se lo stato non è nella lista, usa il primo valore o vuoto
            stato_display = self.fields['Stato']['values'][0] if self.fields['Stato']['values'] else ""
        self.fields['Stato'].set(stato_display)
        self.fields['Note'].delete('1.0', tk.END);
        self.fields['Note'].insert('1.0', task.Note or "")
        
        # Gestisci le date in modo sicuro - se None, usa una data default
        from datetime import date
        if task.DataScadenza:
            self.fields['DataScadenza'].set_date(task.DataScadenza);
        else:
            self.fields['DataScadenza'].set_date(date.today())
            
        if task.DataInizio:
            self.fields['DataInizio'].set_date(task.DataInizio)
        else:
            self.fields['DataInizio'].set_date(date.today())
            
        if task.DataCompletamento:
            self.fields['DataCompletamento'].set_date(task.DataCompletamento)
        else:
            self.fields['DataCompletamento'].set_date(None)  # Può rimanere None
            
        if 'IsPostFinalMilestone' in self.fields:
             self.fields['IsPostFinalMilestone'].var.set(task.IsPostFinalMilestone or False)

    def _on_status_change(self, event=None):
        """Gestisce il cambio di stato per abilitare/disabilitare il campo DataCompletamento."""
        stato_display = self.fields['Stato'].get()
        stato_db = self.status_map_db.get(stato_display, 'Da Fare')
        
        # Abilita DataCompletamento solo se lo stato è "Completato"
        if stato_db == 'Completato':
            self.fields['DataCompletamento'].config(state='readonly')
        else:
            self.fields['DataCompletamento'].config(state='disabled')
            # Resetta la data se non è completato
            self.fields['DataCompletamento'].set_date(None)

    def _disable_form(self):
        self.current_task_id = None
        for child in self.fields.values(): child.config(state='disabled')

    def _enable_form(self):
        for name, child in self.fields.items():
            if name in ['task_name', 'task_category']: continue
            # DataCompletamento viene gestito separatamente in base allo stato
            if name == 'DataCompletamento':
                # Lascia disabilitato, verrà abilitato da _on_status_change se necessario
                child.config(state='disabled')
            else:
                state = 'readonly' if isinstance(child, (DateEntry, ttk.Combobox)) else 'normal'
                child.config(state=state)
        # Aggiorna lo stato del campo DataCompletamento in base allo stato corrente
        self._on_status_change()

    def _disable_doc_form(self):
        for name, widget in self.doc_widgets.items():
            if isinstance(widget, tk.BooleanVar): continue
            widget.config(state=tk.DISABLED)
        self._reset_doc_form()

    def _enable_doc_form(self):
        for name, widget in self.doc_widgets.items():
            if isinstance(widget, tk.BooleanVar): continue
            if name == 'replaces_combo': continue  # Gestito dal toggle
            state = 'readonly' if isinstance(widget, ttk.Combobox) else 'normal'
            widget.config(state=state)

    def _toggle_replacement_combo(self):
        is_replacement = self.doc_widgets['is_replacement_var'].get()
        self.doc_widgets['replaces_combo'].config(state='readonly' if is_replacement else 'disabled')
        if not is_replacement: self.doc_widgets['replaces_combo'].set('')

    def _select_file(self):
        path = filedialog.askopenfilename(title=self.lang.get('select_file_title'), filetypes=[('All', '*.*')])
        if not path: return
        try:
            with open(path, 'rb') as f:
                self.selected_file_data = f.read()
            self.selected_file_path = path
            self.doc_widgets['file_label'].config(text=os.path.basename(path))
        except Exception as e:
            messagebox.showerror("Errore Lettura", f"Impossibile leggere il file:\n{e}", parent=self)
            self.selected_file_path = None;
            self.selected_file_data = None
            self.doc_widgets['file_label'].config(text=self.lang.get('no_file_selected'))

    def _save_document(self):
        if not self.current_task_id: return

        doc_type_desc = self.doc_widgets['type'].get()
        title = self.doc_widgets['title'].get().strip()
        final_client_name = self.doc_widgets['final_client'].get()

        if not all([doc_type_desc, title]):
            messagebox.showwarning("Dati Mancanti", "Tipo Documento e Titolo sono obbligatori.", parent=self)
            return
        if not self.selected_file_data:
            messagebox.showwarning("Dati Mancanti", "Selezionare un file da allegare.", parent=self)
            return

        doc_type_id = self.doc_types_map[doc_type_desc]
        props = self.doc_types_properties[doc_type_id]

        doc_value = float(self.doc_widgets['value_entry'].get() or 0) if props['has_value'] else None
        due_date_str = self.doc_widgets['due_date_entry'].get()
        due_date = datetime.strptime(due_date_str, '%d/%m/%Y') if props['check_date'] and due_date_str else None

        if props['has_value'] and (doc_value is None or doc_value <= 0):
            messagebox.showwarning("Dati non validi", "Il Valore deve essere maggiore di zero.", parent=self)
            return
        if props['check_date'] and (not due_date or due_date.date() <= datetime.now().date()):
            messagebox.showwarning("Data non valida", "La Scadenza deve essere una data futura.", parent=self)
            return

        replaces_doc_id = None
        if self.doc_widgets['is_replacement_var'].get():
            doc_to_replace_title = self.doc_widgets['replaces_combo'].get()
            if not doc_to_replace_title:
                messagebox.showwarning("Dati Mancanti", "Selezionare un documento da sostituire.", parent=self)
                return
            rev_map = {title: doc_id for doc_id, title in self.active_docs_for_task.items()}
            replaces_doc_id = rev_map.get(doc_to_replace_title)

        try:
            self.npi_manager.add_npi_document(
                task_prodotto_id=self.current_task_id,
                doc_type_id=doc_type_id,
                title=title,
                body=self.selected_file_data,
                user=self.logged_in_user,
                note=self.doc_widgets['note_text'].get('1.0', tk.END).strip(),
                replaces_doc_id=replaces_doc_id,
                doc_value=doc_value,
                due_date=due_date,
                # --- CORREZIONE QUI ---
                # Il nome del parametro deve corrispondere a quello atteso dalla funzione del gestore,
                # che è molto probabilmente lo stesso della colonna nel DB.
                #IdFinalClient=self.final_clients_map.get(final_client_name)
                IDSite=self.final_clients_map.get(final_client_name)
            )
            messagebox.showinfo("Successo", "Documento salvato.", parent=self)
            self._on_task_select()  # Ricarica i dati del task
        except Exception as e:
            logger.error(f"Errore salvataggio documento: {e}", exc_info=True)
            messagebox.showerror("Errore Database", f"Impossibile salvare il documento:\n{e}", parent=self)

    def _launch_view_documents_window(self):
        if not self.current_task_id: return
        task = self._get_task_by_id(self.current_task_id)
        name = task.task_catalogo.NomeTask if task and task.task_catalogo else "Task"
        # Passa la mappa INVERSA alla finestra dei documenti
        TaskDocumentsWindow(
            self, self.npi_manager, self.lang, self.current_task_id,
            self.master_app, name, self.final_clients_map_rev
        )

    def _get_task_by_id(self, task_id):
        return next((t for t in self.progetto.waves[0].tasks if t.TaskProdottoID == task_id),
                    None) if self.progetto and self.progetto.waves else None

    def _launch_import_tasks_window(self):
        win = ImportTasksWindow(self, self.npi_manager, self.lang, self.project_id)
        self.wait_window(win)
        self._load_data_and_populate_ui()

    def _save_project_dates(self):
        try:
            start = datetime.strptime(self.project_start_date_entry.get(), '%d/%m/%Y')
            due = datetime.strptime(self.project_due_date_entry.get(), '%d/%m/%Y')
            version = self.project_version_entry.get().strip() or None
            
            # Salva le date
            updated, msg = self.npi_manager.update_project_dates(self.project_id, start, due, self.lang)
            
            # Salva la versione
            self.npi_manager.update_progetto_npi(self.project_id, {'Version': version})
            messagebox.showinfo("Successo", "Date e versione aggiornate.", parent=self)
            if msg: messagebox.showinfo('Informazione', msg, parent=self)
            self._load_data_and_populate_ui()
        except Exception as e:
            logger.error(f"Errore salvataggio date progetto: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile salvare le date:\n{e}", parent=self)

    def _save_task_details(self):
        """Salva i dettagli del task con validazione per milestone finale."""
        if not self.current_task_id:
            return

        # Recupera il nuovo stato
        nuovo_stato_display = self.fields['Stato'].get()
        nuovo_stato_db = self.status_map_db.get(nuovo_stato_display, 'Da Fare')

        # ===== VALIDAZIONE DATE =====
        try:
            data_inizio_str = self.fields['DataInizio'].get()
            data_scadenza_str = self.fields['DataScadenza'].get()
            
            if data_inizio_str and data_scadenza_str:
                data_inizio = datetime.strptime(data_inizio_str, '%d/%m/%Y')
                data_scadenza = datetime.strptime(data_scadenza_str, '%d/%m/%Y')
                
                # Validazione 1: Data scadenza >= Data inizio
                if data_scadenza < data_inizio:
                    messagebox.showwarning(
                        self.lang.get('validation_error_title', 'Errore Validazione'),
                        self.lang.get('error_due_date_before_start', 'La data di scadenza non può essere precedente alla data di inizio.'),
                        parent=self
                    )
                    # Sposta il focus sul campo DataScadenza per permettere la correzione
                    self.fields['DataScadenza'].focus()
                    return
                
                # Validazione 2: Verifica date rispetto alla data fine progetto
                if self.progetto and self.progetto.ScadenzaProgetto:
                    data_fine_progetto = self.progetto.ScadenzaProgetto
                    if isinstance(data_fine_progetto, str):
                        data_fine_progetto = datetime.strptime(data_fine_progetto, '%d/%m/%Y')
                    
                    # Converti a date object se è datetime, altrimenti usa così com'è
                    if hasattr(data_fine_progetto, 'date'):
                        data_fine_progetto_date = data_fine_progetto.date()
                    else:
                        data_fine_progetto_date = data_fine_progetto
                    
                    if data_inizio.date() > data_fine_progetto_date:
                        messagebox.showwarning(
                            self.lang.get('validation_error_title', 'Errore Validazione'),
                            self.lang.get('error_start_after_project_end', 'La data di inizio non può essere successiva alla data fine progetto.'),
                            parent=self
                        )
                        self.fields['DataInizio'].focus()
                        return
                    
                    if data_scadenza.date() > data_fine_progetto_date:
                        messagebox.showwarning(
                            self.lang.get('validation_error_title', 'Errore Validazione'),
                            self.lang.get('error_due_after_project_end', 'La data di scadenza non può essere successiva alla data fine progetto.'),
                            parent=self
                        )
                        self.fields['DataScadenza'].focus()
                        return
                
                # Validazione 3: Verifica dipendenze
                dependencies = self.npi_manager.get_task_dependencies(self.current_task_id)
                if dependencies:
                    # Il task ha dipendenze - verifica le date dei task predecessori
                    for dep in dependencies:
                        pred_task = dep.depends_on_task
                        if pred_task and pred_task.DataScadenza:
                            pred_due_date = pred_task.DataScadenza
                            if isinstance(pred_due_date, str):
                                pred_due_date = datetime.strptime(pred_due_date, '%d/%m/%Y')
                            
                            # Converti a date object se è datetime, altrimenti usa così com'è
                            if hasattr(pred_due_date, 'date'):
                                pred_due_date_date = pred_due_date.date()
                            else:
                                pred_due_date_date = pred_due_date
                            
                            pred_task_name = pred_task.task_catalogo.NomeTask if pred_task.task_catalogo else "Task"
                            
                            # Data inizio task corrente >= Data fine task predecessore
                            if data_inizio.date() < pred_due_date_date:
                                messagebox.showwarning(
                                    self.lang.get('validation_error_title', 'Errore Validazione'),
                                    self.lang.get('error_start_before_dependency', 
                                                 f'La data di inizio non può essere precedente alla data fine del task "{pred_task_name}" ({pred_due_date.strftime("%d/%m/%Y") if hasattr(pred_due_date, "strftime") else str(pred_due_date)}).'),
                                    parent=self
                                )
                                self.fields['DataInizio'].focus()
                                return
                            
                            # Data fine task corrente >= Data fine task predecessore
                            if data_scadenza.date() < pred_due_date_date:
                                messagebox.showwarning(
                                    self.lang.get('validation_error_title', 'Errore Validazione'),
                                    self.lang.get('error_due_before_dependency', 
                                                 f'La data di scadenza non può essere precedente alla data fine del task "{pred_task_name}" ({pred_due_date.strftime("%d/%m/%Y") if hasattr(pred_due_date, "strftime") else str(pred_due_date)}).'),
                                    parent=self
                                )
                                self.fields['DataScadenza'].focus()
                                return
                
        except ValueError as e:
            messagebox.showerror(
                self.lang.get('validation_error_title', 'Errore Validazione'),
                f"Errore nel formato delle date: {e}",
                parent=self
            )
            return

        # ===== VALIDAZIONE DIPENDENZE E MILESTONE FINALE =====
        # Se il nuovo stato è 'Completato', verifica prima le dipendenze poi la milestone
        if nuovo_stato_db == 'Completato':
            # 1. Valida dipendenze task
            is_valid_deps, dep_error = self.npi_manager.validate_task_dependencies(
                self.current_task_id,
                self.lang
            )
            
            if not is_valid_deps:
                messagebox.showwarning(
                    self.lang.get('validation_error_title', 'Errore Validazione'),
                    dep_error,
                    parent=self
                )
                return
            
            # 2. Valida milestone finale
            is_valid, error_msg = self.npi_manager.validate_final_milestone_completion(
                self.current_task_id
            )

            if not is_valid:
                messagebox.showwarning(
                    self.lang.get('validation_error_title', 'Errore Validazione'),
                    error_msg,
                    parent=self
                )
                # Non procede con il salvataggio
                return

        # ===== SALVATAGGIO =====
        data = {
            'OwnerID': self.soggetti_map.get(self.fields['OwnerID'].get()),
            'Stato': nuovo_stato_db,
            'Note': self.fields['Note'].get('1.0', tk.END).strip(),
            'DataScadenza': self.fields['DataScadenza'].get(),
            'DataInizio': self.fields['DataInizio'].get(),
            'DataCompletamento': self.fields['DataCompletamento'].get(),
            'IsPostFinalMilestone': self.fields['IsPostFinalMilestone'].var.get()
        }

        try:
            task = self.npi_manager.update_task_prodotto(self.current_task_id, data)
            if messagebox.askyesno(
                    self.lang.get('notification_send_title', 'Conferma Notifiche'),
                    self.lang.get('notification_send_prompt', 'Inviare notifiche per questo task?'),
                    parent=self
            ):
                self.npi_manager.invia_notifiche_task(
                    task,
                    conferma_utente=False,
                    lang=self.lang
                )

            messagebox.showinfo(
                self.lang.get('success_title', 'Successo'),
                self.lang.get('success_task_updated', 'Task aggiornato con successo.'),
                parent=self
            )

            # Ricarica la UI
            self._load_data_and_populate_ui()

        except Exception as e:
            logger.error(f"Errore salvataggio task: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error_title', 'Errore'),
                f"Impossibile aggiornare il task:\n{e}",
                parent=self
            )
    
    def _delete_task_prodotto(self):
        """Elimina un task dal progetto se non ha task dipendenti."""
        if not self.current_task_id:
            messagebox.showwarning(
                self.lang.get('warning_title', 'Attenzione'),
                self.lang.get('warning_no_task_selected', 'Seleziona un task da eliminare.'),
                parent=self
            )
            return
        
        # Ottieni il task corrente
        task = self._get_task_by_id(self.current_task_id)
        if not task:
            return
        
        task_name = task.task_catalogo.NomeTask if task.task_catalogo else "Task"
        
        # Verifica se ci sono task che dipendono da questo
        try:
            # Controlla se ci sono dipendenze inverse (altri task che dipendono da questo)
            from sqlalchemy import select
            from ..models import TaskProductDependency
            
            with self.npi_manager.session_scope() as session:
                # Cerca task che hanno questo task come dipendenza
                stmt = select(TaskProductDependency).where(
                    TaskProductDependency.DependsOnTaskID == self.current_task_id
                )
                dependent_tasks = session.execute(stmt).scalars().all()
                
                if dependent_tasks:
                    # Ci sono task che dipendono da questo - non può essere eliminato
                    dependent_names = []
                    for dep in dependent_tasks:
                        dep_task = session.get(self.npi_manager.TaskProdotto, dep.TaskProdottoID)
                        if dep_task and dep_task.task_catalogo:
                            dependent_names.append(dep_task.task_catalogo.NomeTask)
                    
                    messagebox.showwarning(
                        self.lang.get('warning_title', 'Attenzione'),
                        self.lang.get('error_cannot_delete_task_has_dependents', 
                                     f'Impossibile eliminare il task "{task_name}".\n\n'
                                     f'I seguenti task dipendono da esso:\n' +
                                     '\n'.join(f'• {name}' for name in dependent_names)),
                        parent=self
                    )
                    return
        
        except Exception as e:
            logger.error(f"Errore verifica dipendenze: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error_title', 'Errore'),
                f"Errore durante la verifica delle dipendenze:\\n{e}",
                parent=self
            )
            return
        
        # Conferma eliminazione
        if not messagebox.askyesno(
                self.lang.get('confirm_delete_title', 'Conferma Eliminazione'),
                self.lang.get('confirm_delete_task', 
                             f'Sei sicuro di voler eliminare il task "{task_name}" dal progetto?\\n\\n'
                             f'Questa operazione non può essere annullata.'),
                parent=self
        ):
            return
        
        try:
            # Elimina il task dal progetto
            with self.npi_manager.session_scope() as session:
                task_to_delete = session.get(self.npi_manager.TaskProdotto, self.current_task_id)
                if task_to_delete:
                    session.delete(task_to_delete)
            
            messagebox.showinfo(
                self.lang.get('success_title', 'Successo'),
                self.lang.get('success_task_deleted', 'Task eliminato con successo.'),
                parent=self
            )
            
            # Ricarica la UI
            self._load_data_and_populate_ui()
            self._disable_form()
            self._disable_doc_form()
            
        except Exception as e:
            logger.error(f"Errore eliminazione task: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error_title', 'Errore'),
                f"Impossibile eliminare il task:\\n{e}",
                parent=self
            )
    
    def _sync_catalog_tasks(self):
        """Sincronizza i task del catalogo con il progetto corrente."""
        try:
            # Conferma con l'utente
            response = messagebox.askyesno(
                self.lang.get('confirm_title', 'Conferma'),
                self.lang.get('msg_sync_catalog', 
                             'Questa operazione aggiungerà al progetto tutti i task del catalogo che non sono ancora presenti.\n\n'
                             'I task già esistenti non verranno modificati.\n\n'
                             'Vuoi continuare?'),
                parent=self
            )
            
            if not response:
                return
            
            # Ottieni tutti i task del catalogo
            all_catalog_tasks = self.npi_manager.get_all_catalog_tasks()
            
            # Ottieni i task già presenti nel progetto
            wave = self.progetto.waves[0] if self.progetto and self.progetto.waves else None
            if not wave:
                messagebox.showerror(
                    self.lang.get('error_title', 'Errore'),
                    'Nessuna wave trovata per questo progetto',
                    parent=self
                )
                return
            
            existing_task_ids = {t.TaskID for t in wave.tasks}
            
            # Trova i task mancanti
            missing_tasks = [t for t in all_catalog_tasks if t.TaskID not in existing_task_ids]
            
            if not missing_tasks:
                messagebox.showinfo(
                    self.lang.get('info_title', 'Informazione'),
                    self.lang.get('msg_no_missing_tasks', 'Tutti i task del catalogo sono già presenti nel progetto.'),
                    parent=self
                )
                return
            
            # Aggiungi i task mancanti
            added_count = self.npi_manager.add_catalog_tasks_to_project(wave.WaveID, missing_tasks)
            
            messagebox.showinfo(
                self.lang.get('success_title', 'Successo'),
                self.lang.get('msg_tasks_synced', f'Aggiunti {added_count} task dal catalogo.'),
                parent=self
            )
            
            # Ricarica la UI
            self._load_data_and_populate_ui()
            self.log_status(f"Sincronizzati {added_count} task dal catalogo")
            
        except Exception as e:
            logger.error(f"Errore sincronizzazione catalogo: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error_title', 'Errore'),
                f"Errore durante la sincronizzazione:\n{e}",
                parent=self
            )
    
    def _launch_view_documents_window(self):
        """Lancia la finestra per visualizzare tutti i documenti del progetto."""
        try:
            from .project_documents_window import ProjectDocumentsWindow
            
            project_name = self.progetto.prodotto.NomeProdotto if self.progetto and self.progetto.prodotto else "Progetto"
            
            ProjectDocumentsWindow(self, self.npi_manager, self.lang, self.project_id, project_name)
            
        except Exception as e:
            logger.error(f"Errore apertura finestra documenti: {e}", exc_info=True)
            messagebox.showerror('Errore', f"Impossibile aprire la finestra documenti:\n{e}", parent=self)
    
    # ========================================
    # ðŸ†• FUNZIONI GERARCHIA PROGETTI
    # ========================================
    
    def _populate_project_hierarchy(self):
        """Popola i widget della gerarchia progetti."""
        try:
            # Carica tutti i progetti disponibili come possibili padri
            all_projects = self.npi_manager.get_progetti_npi()
            
            # Escludi il progetto corrente dalla lista dei padri possibili
            available_parents = [
                p for p in all_projects 
                if p.ProgettoId != self.project_id
            ]
            
            # ðŸ›¡ï¸ Validazione: Escludi anche i discendenti (evita cicli)
            # Un progetto non puÃ² avere come padre uno dei suoi discendenti
            child_ids = self._get_all_descendant_ids(self.project_id)
            available_parents = [
                p for p in available_parents
                if p.ProgettoId not in child_ids
            ]
            
            # Popola combo con "(Nessuno)" + lista progetti
            parent_names = ["(Nessuno - Progetto Root)"]
            self.parent_projects_map = {}
            
            for p in available_parents:
                display_name = f"{p.NomeProgetto} (ID: {p.ProgettoId})"
                parent_names.append(display_name)
                self.parent_projects_map[display_name] = p.ProgettoId
            
            self.parent_project_combo['values'] = parent_names
            
            # Imposta il valore corrente
            if self.progetto.ParentProjectID:
                parent = self.npi_manager.get_parent_project(self.project_id)
                if parent:
                    display_name = f"{parent.NomeProgetto} (ID: {parent.ProgettoId})"
                    if display_name in parent_names:
                        self.parent_project_combo.set(display_name)
                        self.current_parent_label.config(
                            text=f"ðŸ“¦ Progetto Padre: {parent.NomeProgetto}",
                            foreground="blue"
                        )
                    else:
                        self.parent_project_combo.set(parent_names[0])
                        self.current_parent_label.config(text="âœ… Progetto Root (nessun padre)")
                else:
                    self.parent_project_combo.set(parent_names[0])
                    self.current_parent_label.config(text="âœ… Progetto Root (nessun padre)")
            else:
                self.parent_project_combo.set(parent_names[0])
                self.current_parent_label.config(text="âœ… Progetto Root (nessun padre)")
            
            # Conta e mostra progetti figli
            children = self.npi_manager.get_child_projects(self.project_id)
            if children:
                completed_count = sum(1 for c in children if c.StatoProgetto == 'Completato')
                self.children_projects_label.config(
                    text=f"ðŸ“„ {len(children)} progetti figli ({completed_count} completati)",
                    foreground="green" if completed_count == len(children) else "orange"
                )
            else:
                self.children_projects_label.config(
                    text="Nessun progetto figlio",
                    foreground="gray"
                )
                
        except Exception as e:
            logger.error(f"Errore popolamento gerarchia: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Errore caricamento gerarchia:\\n{e}", parent=self)
    
    def _get_all_descendant_ids(self, project_id, visited=None):
        """
        Recupera ricorsivamente tutti gli ID dei progetti discendenti.
        Utilizzato per evitare cicli nella selezione del padre.
        """
        if visited is None:
            visited = set()
        
        if project_id in visited:
            return visited
        
        visited.add(project_id)
        
        try:
            children = self.npi_manager.get_child_projects(project_id)
            for child in children:
                self._get_all_descendant_ids(child.ProgettoId, visited)
        except Exception as e:
            logger.warning(f"Errore recupero discendenti per {project_id}: {e}")
        
        return visited
    
    def _on_parent_project_selected(self, event=None):
        """Gestisce la selezione di un nuovo progetto padre."""
        selected = self.parent_project_combo.get()
        
        if selected == "(Nessuno - Progetto Root)":
            new_parent_id = None
        else:
            new_parent_id = self.parent_projects_map.get(selected)
        
        # Conferma cambio
        if new_parent_id == self.progetto.ParentProjectID:
            return  # Nessun cambio
        
        parent_name = selected if selected != "(Nessuno - Progetto Root)" else "nessun padre"
        response = messagebox.askyesno(
            "Conferma Modifica",
            f"Vuoi modificare il progetto padre a:\\n{parent_name}?",
            parent=self
        )
        
        if not response:
            # Ripristina valore originale
            self._populate_project_hierarchy()
            return
        
        # Valida contro cicli
        if new_parent_id:
            is_valid, msg = self.npi_manager.validate_no_circular_dependency(
                self.project_id,
                new_parent_id
            )
            
            if not is_valid:
                messagebox.showerror("Errore Validazione", msg, parent=self)
                self._populate_project_hierarchy()
                return
        
        # Salva modifica
        try:
            self.progetto.ParentProjectID = new_parent_id
            
            # Aggiorna anche ProjectType
            if new_parent_id:
                self.progetto.ProjectType = 'Child'
            else:
                # Verifica se ha figli per determinare se Ã¨ Parent o Standard
                children = self.npi_manager.get_child_projects(self.project_id)
                self.progetto.ProjectType = 'Parent' if children else 'Standard'
            
            # Salva nel database
            self.npi_manager.update_progetto_npi(self.project_id, {
                'ParentProjectID': new_parent_id,
                'ProjectType': self.progetto.ProjectType
            })
            
            # Aggiorna i livelli gerarchici
            if new_parent_id:
                self.npi_manager.update_hierarchy_levels(new_parent_id)
            else:
                self.npi_manager.update_hierarchy_levels(self.project_id)
            
            messagebox.showinfo("Successo", "Gerarchia progetti aggiornata con successo!", parent=self)
            
            # Ricarica i dati
            self._populate_project_hierarchy()
            
        except Exception as e:
            logger.error(f"Errore salvataggio padre: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile salvare:\\n{e}", parent=self)
            self._populate_project_hierarchy()
    
    def _show_child_projects_dialog(self):
        """Mostra un dialog con la lista dei progetti figli e i loro stati."""
        try:
            children = self.npi_manager.get_child_projects(self.project_id)
            
            if not children:
                messagebox.showinfo("Progetti Figli", "Questo progetto non ha progetti figli.", parent=self)
                return
            
            # Crea dialog
            dialog = tk.Toplevel(self)
            dialog.title("Progetti Figli")
            dialog.geometry("600x400")
            dialog.transient(self)
            dialog.grab_set()
            
            # Titolo
            ttk.Label(
                dialog,
                text=f"Progetti Figli di: {self.progetto.NomeProgetto}",
                font=('Helvetica', 12, 'bold')
            ).pack(pady=10)
            
            # Treeview
            tree_frame = ttk.Frame(dialog)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            tree = ttk.Treeview(
                tree_frame,
                columns=('Nome', 'Stato', 'Livello', 'Tipo'),
                show='tree headings'
            )
            
            tree.heading('#0', text='ID')
            tree.heading('Nome', text='Nome Progetto')
            tree.heading('Stato', text='Stato')
            tree.heading('Livello', text='Livello')
            tree.heading('Tipo', text='Tipo')
            
            tree.column('#0', width=60)
            tree.column('Nome', width=300)
            tree.column('Stato', width=100)
            tree.column('Livello', width=60)
            tree.column('Tipo', width=80)
            
            # Scrollbar
            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Popola
            for child in children:
                tags = []
                if child.StatoProgetto == 'Completato':
                    tags.append('completed')
                elif child.StatoProgetto == 'In Lavorazione':
                    tags.append('in_progress')
                
                tree.insert('', tk.END,
                    text=str(child.ProgettoId),
                    values=(
                        child.NomeProgetto,
                        child.StatoProgetto,
                        child.HierarchyLevel or 0,
                        child.ProjectType or 'Standard'
                    ),
                    tags=tags
                )
            
            # Tag styling
            tree.tag_configure('completed', foreground='green')
            tree.tag_configure('in_progress', foreground='orange')
            
            # Pulsanti
            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)
            
            ttk.Button(button_frame, text="Chiudi", command=dialog.destroy).pack()
            
        except Exception as e:
            logger.error(f"Errore visualizzazione figli: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile visualizzare i progetti figli:\\n{e}", parent=self)
