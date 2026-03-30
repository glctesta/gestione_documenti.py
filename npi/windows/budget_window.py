# File: npi/windows/budget_window.py
"""
Finestra per la gestione del Budget per progetti NPI.
CRUD completo, import/export Excel, approvazione integrale e per riga.
"""

import logging
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime

logger = logging.getLogger(__name__)

BUDGET_STATUSES = ['Bozza', 'InLavorazione', 'Terminato']
APPROVAL_STATUSES = ['DaApprovare', 'Approvato', 'Rifiutato']
ITEM_STATUSES = ['InLavorazione', 'Terminato']


class BudgetWindow(tk.Toplevel):
    """Gestione budget per un progetto NPI."""

    def __init__(self, master, npi_manager, lang, project_id, project_name='', logged_in_user=''):
        super().__init__(master)
        self.npi_manager = npi_manager
        self.lang = lang
        self.project_id = project_id
        self.project_name = project_name
        self.logged_in_user = logged_in_user

        self.title(f"{lang.get('budget_title', 'Budget Progetto')} - {project_name}")
        self.geometry("1400x800")
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()

        self._selected_budget_id = None
        self._selected_item_id = None
        self._categories_map = {}     # name -> id
        self._categories_map_rev = {} # id -> name
        self._budgets_list = []

        self._create_widgets()
        self._load_categories()
        self._load_budgets()

        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ------------------------------------------------------------------ #
    #  UI Construction                                                      #
    # ------------------------------------------------------------------ #
    def _create_widgets(self):
        main = ttk.Frame(self, padding=10)
        main.pack(expand=True, fill="both")

        # --- Header ---
        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 10))

        ttk.Label(header, text=f"💰 {self.lang.get('budget_header', 'Budget Progetto NPI')}",
                  font=("Segoe UI", 14, "bold")).pack(side="left")

        ttk.Label(header, text=self.project_name, font=("Segoe UI", 11),
                  foreground="blue").pack(side="left", padx=20)

        # --- Budget selector + toolbar ---
        selector_frame = ttk.LabelFrame(main, text=self.lang.get('budget_selector', 'Seleziona Budget'), padding=5)
        selector_frame.pack(fill="x", pady=(0, 5))

        ttk.Label(selector_frame, text=self.lang.get('budget_name', 'Budget:')).pack(side="left", padx=5)
        self.budget_combo = ttk.Combobox(selector_frame, state='readonly', width=40)
        self.budget_combo.pack(side="left", padx=5)
        self.budget_combo.bind('<<ComboboxSelected>>', self._on_budget_selected)

        # Toolbar buttons
        ttk.Button(selector_frame, text=self.lang.get('budget_btn_new', '➕ Nuovo Budget'),
                   command=self._new_budget).pack(side="left", padx=3)
        ttk.Button(selector_frame, text=self.lang.get('budget_btn_delete', '🗑️ Elimina Budget'),
                   command=self._delete_budget).pack(side="left", padx=3)

        self.active_btn = ttk.Button(selector_frame, text=self.lang.get('budget_btn_activate', '✅ Imposta Attivo'),
                                     command=self._set_active)
        self.active_btn.pack(side="left", padx=3)

        # Status label
        self.budget_info_var = tk.StringVar(value="")
        ttk.Label(selector_frame, textvariable=self.budget_info_var,
                  foreground="gray", font=("Segoe UI", 9)).pack(side="right", padx=10)

        # --- Budget status + actions toolbar ---
        action_frame = ttk.Frame(main)
        action_frame.pack(fill="x", pady=(0, 5))

        ttk.Label(action_frame, text=self.lang.get('budget_status_label', 'Stato:')).pack(side="left", padx=5)
        self.status_combo = ttk.Combobox(action_frame, values=BUDGET_STATUSES, state='readonly', width=15)
        self.status_combo.pack(side="left", padx=3)
        ttk.Button(action_frame, text=self.lang.get('budget_btn_save_status', '💾 Salva Stato'),
                   command=self._save_budget_status).pack(side="left", padx=3)

        ttk.Separator(action_frame, orient="vertical").pack(side="left", fill="y", padx=10)

        ttk.Button(action_frame, text=self.lang.get('budget_btn_template', '📥 Scarica Template'),
                   command=self._download_template).pack(side="left", padx=3)
        ttk.Button(action_frame, text=self.lang.get('budget_btn_import', '📤 Importa Excel'),
                   command=self._import_excel).pack(side="left", padx=3)
        ttk.Button(action_frame, text=self.lang.get('budget_btn_export', '📊 Esporta Excel'),
                   command=self._export_excel).pack(side="left", padx=3)

        ttk.Separator(action_frame, orient="vertical").pack(side="left", fill="y", padx=10)

        ttk.Button(action_frame, text=self.lang.get('budget_btn_approve_rows', '🔔 Approva Righe'),
                   command=self._request_approval_rows).pack(side="left", padx=3)
        ttk.Button(action_frame, text=self.lang.get('budget_btn_approve_all', '📋 Approva Budget'),
                   command=self._request_approval_all).pack(side="left", padx=3)

        # --- Items Treeview ---
        tree_frame = ttk.Frame(main)
        tree_frame.pack(fill="both", expand=True, pady=(0, 5))

        columns = ('desc', 'category', 'qty', 'unit_price', 'total', 'status', 'approval', 'note')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='extended')

        self.tree.heading('desc', text=self.lang.get('budget_col_desc', 'Descrizione'))
        self.tree.heading('category', text=self.lang.get('budget_col_category', 'Categoria'))
        self.tree.heading('qty', text=self.lang.get('budget_col_qty', 'Qtà'))
        self.tree.heading('unit_price', text=self.lang.get('budget_col_unit_price', 'Prezzo Unit.'))
        self.tree.heading('total', text=self.lang.get('budget_col_total', 'Totale'))
        self.tree.heading('status', text=self.lang.get('budget_col_status', 'Stato'))
        self.tree.heading('approval', text=self.lang.get('budget_col_approval', 'Approvazione'))
        self.tree.heading('note', text=self.lang.get('budget_col_note', 'Note'))

        self.tree.column('desc', width=250)
        self.tree.column('category', width=120)
        self.tree.column('qty', width=70, anchor="e")
        self.tree.column('unit_price', width=100, anchor="e")
        self.tree.column('total', width=100, anchor="e")
        self.tree.column('status', width=100, anchor="center")
        self.tree.column('approval', width=100, anchor="center")
        self.tree.column('note', width=200)

        # Tags per colori
        self.tree.tag_configure('approved', background='#d4edda')
        self.tree.tag_configure('rejected', background='#f8d7da')
        self.tree.tag_configure('pending', background='#fff3cd')

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind('<<TreeviewSelect>>', self._on_item_selected)

        # --- Detail form ---
        form_frame = ttk.LabelFrame(main, text=self.lang.get('budget_detail', 'Dettaglio Riga'), padding=10)
        form_frame.pack(fill="x")

        # Row 0: Desc + Category
        ttk.Label(form_frame, text=self.lang.get('budget_col_desc', 'Descrizione:')).grid(row=0, column=0, sticky="w", padx=5)
        self.desc_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=self.desc_var, width=40).grid(row=0, column=1, padx=5, sticky="ew")

        ttk.Label(form_frame, text=self.lang.get('budget_col_category', 'Categoria:')).grid(row=0, column=2, sticky="w", padx=5)
        self.cat_combo = ttk.Combobox(form_frame, state='readonly', width=20)
        self.cat_combo.grid(row=0, column=3, padx=5)

        # Row 1: Qty + UnitPrice + Note
        ttk.Label(form_frame, text=self.lang.get('budget_col_qty', 'Qtà:')).grid(row=1, column=0, sticky="w", padx=5)
        self.qty_var = tk.StringVar(value="1")
        ttk.Entry(form_frame, textvariable=self.qty_var, width=10).grid(row=1, column=1, padx=5, sticky="w")

        ttk.Label(form_frame, text=self.lang.get('budget_col_unit_price', 'Prezzo Unit.:')).grid(row=1, column=2, sticky="w", padx=5)
        self.price_var = tk.StringVar(value="0.00")
        ttk.Entry(form_frame, textvariable=self.price_var, width=15).grid(row=1, column=3, padx=5, sticky="w")

        ttk.Label(form_frame, text=self.lang.get('budget_col_note', 'Note:')).grid(row=1, column=4, sticky="w", padx=5)
        self.note_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=self.note_var, width=30).grid(row=1, column=5, padx=5, sticky="ew")

        form_frame.columnconfigure(1, weight=1)
        form_frame.columnconfigure(5, weight=1)

        # Row 2: Buttons
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=6, pady=10, sticky="e")

        ttk.Button(btn_frame, text=self.lang.get('budget_btn_add_item', '➕ Aggiungi Riga'),
                   command=self._add_item).pack(side="left", padx=3)
        ttk.Button(btn_frame, text=self.lang.get('budget_btn_update_item', '💾 Aggiorna Riga'),
                   command=self._update_item).pack(side="left", padx=3)
        ttk.Button(btn_frame, text=self.lang.get('budget_btn_delete_item', '🗑️ Elimina Riga'),
                   command=self._delete_item).pack(side="left", padx=3)
        ttk.Button(btn_frame, text=self.lang.get('budget_btn_clear', '🧹 Pulisci'),
                   command=self._clear_form).pack(side="left", padx=3)

        # --- Documenti per riga ---
        docs_frame = ttk.LabelFrame(main, text=self.lang.get('budget_docs_panel', '📎 Documenti Riga'), padding=5)
        docs_frame.pack(fill="both", expand=True, pady=(5, 0))

        docs_btn_row = ttk.Frame(docs_frame)
        docs_btn_row.pack(fill="x", pady=(0, 3))

        ttk.Button(docs_btn_row, text=self.lang.get('budget_btn_upload_doc', '📎 Carica Doc.'),
                   command=self._upload_doc_for_item).pack(side="left", padx=3)
        ttk.Button(docs_btn_row, text=self.lang.get('budget_btn_link_doc', '🔗 Collega Esistente'),
                   command=self._link_existing_doc).pack(side="left", padx=3)
        ttk.Button(docs_btn_row, text=self.lang.get('budget_btn_unlink_doc', '❌ Scollega'),
                   command=self._unlink_doc).pack(side="left", padx=3)
        ttk.Button(docs_btn_row, text=self.lang.get('budget_btn_open_doc', '👁️ Apri'),
                   command=self._open_doc).pack(side="left", padx=3)

        doc_tree_frame = ttk.Frame(docs_frame)
        doc_tree_frame.pack(fill="both", expand=True)

        doc_cols = ('doc_title', 'doc_type', 'doc_date', 'doc_user')
        self.doc_tree = ttk.Treeview(doc_tree_frame, columns=doc_cols, show='headings',
                                     height=4, selectmode='browse')
        self.doc_tree.heading('doc_title', text=self.lang.get('budget_doc_col_title', 'Titolo'))
        self.doc_tree.heading('doc_type', text=self.lang.get('budget_doc_col_type', 'Tipo'))
        self.doc_tree.heading('doc_date', text=self.lang.get('budget_doc_col_date', 'Data'))
        self.doc_tree.heading('doc_user', text=self.lang.get('budget_doc_col_user', 'Utente'))
        self.doc_tree.column('doc_title', width=250)
        self.doc_tree.column('doc_type', width=120)
        self.doc_tree.column('doc_date', width=130, anchor="center")
        self.doc_tree.column('doc_user', width=100)

        doc_scroll = ttk.Scrollbar(doc_tree_frame, orient="vertical", command=self.doc_tree.yview)
        self.doc_tree.configure(yscrollcommand=doc_scroll.set)
        self.doc_tree.pack(side="left", fill="both", expand=True)
        doc_scroll.pack(side="right", fill="y")
        self.doc_tree.bind('<Double-1>', lambda e: self._open_doc())

        # Totale footer
        footer = ttk.Frame(main)
        footer.pack(fill="x", pady=(5, 0))
        self.total_var = tk.StringVar(value="Totale: € 0.00")
        ttk.Label(footer, textvariable=self.total_var, font=("Segoe UI", 12, "bold"),
                  foreground="darkgreen").pack(side="right", padx=10)

    # ------------------------------------------------------------------ #
    #  Data Loading                                                         #
    # ------------------------------------------------------------------ #
    def _load_categories(self):
        """Carica le categorie budget dal DB."""
        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetCategory
                cats = session.query(NpiBudgetCategory).filter(
                    NpiBudgetCategory.DateOut.is_(None)
                ).order_by(NpiBudgetCategory.CategoryName).all()
                self._categories_map = {c.CategoryName: c.CategoryId for c in cats}
                self._categories_map_rev = {c.CategoryId: c.CategoryName for c in cats}
                self.cat_combo['values'] = [''] + list(self._categories_map.keys())
            finally:
                session.close()
        except Exception as e:
            logger.error(f"Errore caricamento categorie budget: {e}", exc_info=True)

    def _load_budgets(self):
        """Carica la lista budget per il progetto corrente."""
        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget
                budgets = session.query(NpiBudget).filter(
                    NpiBudget.ProgettoID == self.project_id,
                    NpiBudget.DateOut.is_(None)
                ).order_by(NpiBudget.CreatedDate.desc()).all()

                self._budgets_list = []
                display_values = []
                for b in budgets:
                    active_marker = " ✅" if b.IsActive else ""
                    display = f"{b.BudgetName} [{b.BudgetStatus}]{active_marker}"
                    display_values.append(display)
                    self._budgets_list.append({
                        'id': b.BudgetId,
                        'name': b.BudgetName,
                        'status': b.BudgetStatus,
                        'approval': b.ApprovalStatus,
                        'is_active': b.IsActive,
                        'display': display
                    })

                self.budget_combo['values'] = display_values

                # Se c'era un budget selezionato, ri-selezionalo
                if self._selected_budget_id:
                    for i, b in enumerate(self._budgets_list):
                        if b['id'] == self._selected_budget_id:
                            self.budget_combo.current(i)
                            self._on_budget_selected()
                            break
                elif display_values:
                    self.budget_combo.current(0)
                    self._on_budget_selected()
                else:
                    self._selected_budget_id = None
                    self.tree.delete(*self.tree.get_children())
                    self.budget_info_var.set("")
                    self.total_var.set("Totale: € 0.00")
            finally:
                session.close()
        except Exception as e:
            logger.error(f"Errore caricamento budgets: {e}", exc_info=True)

    def _load_items(self):
        """Carica le righe del budget selezionato."""
        self.tree.delete(*self.tree.get_children())
        if not self._selected_budget_id:
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetItem
                items = session.query(NpiBudgetItem).filter(
                    NpiBudgetItem.BudgetId == self._selected_budget_id,
                    NpiBudgetItem.DateOut.is_(None)
                ).order_by(NpiBudgetItem.BudgetItemId).all()

                total = 0.0
                for item in items:
                    cat_name = self._categories_map_rev.get(item.CategoryId, '')
                    total_price = float(item.Quantity or 0) * float(item.UnitPrice or 0)
                    total += total_price

                    # Tag per colore
                    tag = 'pending'
                    if item.ItemApproval == 'Approvato':
                        tag = 'approved'
                    elif item.ItemApproval == 'Rifiutato':
                        tag = 'rejected'

                    self.tree.insert('', 'end', iid=str(item.BudgetItemId), values=(
                        item.ItemDescription or '',
                        cat_name,
                        f"{float(item.Quantity or 0):.2f}",
                        f"€ {float(item.UnitPrice or 0):.2f}",
                        f"€ {total_price:.2f}",
                        item.ItemStatus or '',
                        item.ItemApproval or '',
                        item.Note or ''
                    ), tags=(tag,))

                self.total_var.set(f"Totale: € {total:,.2f}")
            finally:
                session.close()
        except Exception as e:
            logger.error(f"Errore caricamento righe budget: {e}", exc_info=True)

    # ------------------------------------------------------------------ #
    #  Event Handlers                                                       #
    # ------------------------------------------------------------------ #
    def _on_budget_selected(self, event=None):
        idx = self.budget_combo.current()
        if idx < 0 or idx >= len(self._budgets_list):
            return

        budget = self._budgets_list[idx]
        self._selected_budget_id = budget['id']

        # Update info
        active_text = "✅ ATTIVO" if budget['is_active'] else "⚪ Non attivo"
        self.budget_info_var.set(
            f"Stato: {budget['status']} | Approvazione: {budget['approval']} | {active_text}")

        # Set status combo
        if budget['status'] in BUDGET_STATUSES:
            self.status_combo.set(budget['status'])

        self._load_items()
        self._clear_form()

    def _on_item_selected(self, event=None):
        selection = self.tree.selection()
        if not selection:
            self._selected_item_id = None
            return

        self._selected_item_id = int(selection[0])
        values = self.tree.item(selection[0])['values']

        self.desc_var.set(values[0] or '')
        self.cat_combo.set(values[1] or '')
        self.qty_var.set(str(values[2]).replace(',', '') if values[2] else '1')
        self.price_var.set(str(values[3]).replace('€ ', '').replace(',', '') if values[3] else '0.00')
        self.note_var.set(values[7] or '')
        self._load_docs_for_item()

    def _clear_form(self):
        self._selected_item_id = None
        self.desc_var.set('')
        self.cat_combo.set('')
        self.qty_var.set('1')
        self.price_var.set('0.00')
        self.note_var.set('')
        if self.tree.selection():
            self.tree.selection_remove(self.tree.selection())
        self.doc_tree.delete(*self.doc_tree.get_children())

    # ------------------------------------------------------------------ #
    #  Budget CRUD                                                          #
    # ------------------------------------------------------------------ #
    def _new_budget(self):
        """Crea un nuovo budget per il progetto."""
        name = simpledialog.askstring(
            self.lang.get('budget_new_title', 'Nuovo Budget'),
            self.lang.get('budget_new_prompt', 'Nome del budget:'),
            parent=self
        )
        if not name or not name.strip():
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget
                new_budget = NpiBudget(
                    ProgettoID=self.project_id,
                    BudgetName=name.strip(),
                    BudgetStatus='Bozza',
                    ApprovalStatus='DaApprovare',
                    IsActive=False,
                    CreatedBy=self.logged_in_user,
                    CreatedDate=datetime.now()
                )
                session.add(new_budget)
                session.commit()
                self._selected_budget_id = new_budget.BudgetId
                logger.info(f"Budget creato: {name} per progetto {self.project_id}")
            finally:
                session.close()

            self._load_budgets()
        except Exception as e:
            logger.error(f"Errore creazione budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 f"Errore creazione budget:\n{e}", parent=self)

    def _delete_budget(self):
        """Soft-delete del budget selezionato."""
        if not self._selected_budget_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_first', 'Selezionare un budget.'),
                                   parent=self)
            return

        if not messagebox.askyesno(self.lang.get('confirm', 'Conferma'),
                                    self.lang.get('budget_confirm_delete', 'Eliminare questo budget?'),
                                    parent=self):
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget
                budget = session.query(NpiBudget).get(self._selected_budget_id)
                if budget:
                    budget.DateOut = datetime.now()
                    session.commit()
                    logger.info(f"Budget eliminato (soft): {self._selected_budget_id}")
            finally:
                session.close()

            self._selected_budget_id = None
            self._load_budgets()
        except Exception as e:
            logger.error(f"Errore eliminazione budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _set_active(self):
        """Imposta il budget corrente come attivo (disattiva gli altri)."""
        if not self._selected_budget_id:
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget
                # Disattiva tutti i budget del progetto
                session.query(NpiBudget).filter(
                    NpiBudget.ProgettoID == self.project_id,
                    NpiBudget.DateOut.is_(None)
                ).update({NpiBudget.IsActive: False})

                # Attiva il budget selezionato
                budget = session.query(NpiBudget).get(self._selected_budget_id)
                if budget:
                    budget.IsActive = True
                session.commit()
                logger.info(f"Budget {self._selected_budget_id} impostato come attivo")
            finally:
                session.close()

            self._load_budgets()
        except Exception as e:
            logger.error(f"Errore attivazione budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _save_budget_status(self):
        """Salva il cambio di stato del budget."""
        if not self._selected_budget_id:
            return

        new_status = self.status_combo.get()
        if not new_status:
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget
                budget = session.query(NpiBudget).get(self._selected_budget_id)
                if budget:
                    budget.BudgetStatus = new_status
                    session.commit()
                    logger.info(f"Budget {self._selected_budget_id} stato → {new_status}")
            finally:
                session.close()

            self._load_budgets()
        except Exception as e:
            logger.error(f"Errore aggiornamento stato budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    # ------------------------------------------------------------------ #
    #  Item CRUD                                                            #
    # ------------------------------------------------------------------ #
    def _validate_item_form(self):
        """Valida i campi della form riga."""
        desc = self.desc_var.get().strip()
        if not desc:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('budget_desc_required', 'La descrizione è obbligatoria.'),
                                 parent=self)
            return None

        try:
            qty = float(self.qty_var.get().replace(',', '.'))
            if qty <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('budget_qty_invalid', 'Quantità non valida.'), parent=self)
            return None

        try:
            price = float(self.price_var.get().replace(',', '.').replace('€', '').strip())
            if price < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror(self.lang.get('error', 'Errore'),
                                 self.lang.get('budget_price_invalid', 'Prezzo non valido.'), parent=self)
            return None

        cat_name = self.cat_combo.get()
        cat_id = self._categories_map.get(cat_name) if cat_name else None

        return {
            'description': desc,
            'category_id': cat_id,
            'quantity': qty,
            'unit_price': price,
            'note': self.note_var.get().strip()
        }

    def _add_item(self):
        """Aggiunge una nuova riga al budget."""
        if not self._selected_budget_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_first', 'Selezionare un budget.'),
                                   parent=self)
            return

        data = self._validate_item_form()
        if not data:
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetItem
                item = NpiBudgetItem(
                    BudgetId=self._selected_budget_id,
                    ItemDescription=data['description'],
                    CategoryId=data['category_id'],
                    Quantity=data['quantity'],
                    UnitPrice=data['unit_price'],
                    ItemStatus='InLavorazione',
                    ItemApproval='DaApprovare',
                    Note=data['note']
                )
                session.add(item)
                session.commit()
                logger.info(f"Riga budget aggiunta: {data['description']}")
            finally:
                session.close()

            self._load_items()
            self._clear_form()
        except Exception as e:
            logger.error(f"Errore aggiunta riga: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _update_item(self):
        """Aggiorna la riga selezionata."""
        if not self._selected_item_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_item', 'Selezionare una riga.'),
                                   parent=self)
            return

        data = self._validate_item_form()
        if not data:
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetItem
                item = session.query(NpiBudgetItem).get(self._selected_item_id)
                if item:
                    item.ItemDescription = data['description']
                    item.CategoryId = data['category_id']
                    item.Quantity = data['quantity']
                    item.UnitPrice = data['unit_price']
                    item.Note = data['note']
                    session.commit()
                    logger.info(f"Riga budget aggiornata: {self._selected_item_id}")
            finally:
                session.close()

            self._load_items()
        except Exception as e:
            logger.error(f"Errore aggiornamento riga: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _delete_item(self):
        """Soft-delete della riga selezionata."""
        if not self._selected_item_id:
            return

        if not messagebox.askyesno(self.lang.get('confirm', 'Conferma'),
                                    self.lang.get('budget_confirm_delete_item', 'Eliminare questa riga?'),
                                    parent=self):
            return

        try:
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetItem
                item = session.query(NpiBudgetItem).get(self._selected_item_id)
                if item:
                    item.DateOut = datetime.now()
                    session.commit()
            finally:
                session.close()

            self._selected_item_id = None
            self._load_items()
            self._clear_form()
        except Exception as e:
            logger.error(f"Errore eliminazione riga: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    # ------------------------------------------------------------------ #
    #  Excel Operations                                                     #
    # ------------------------------------------------------------------ #
    def _download_template(self):
        """Genera e salva un template Excel per l'import del budget."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Budget Template"

            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            headers = ['Descrizione', 'Categoria', 'Quantità', 'Prezzo Unitario', 'Note']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Categorie disponibili come nota
            cat_list = ', '.join(self._categories_map.keys()) or 'Nessuna categoria configurata'
            ws.cell(row=3, column=1, value=f"Categorie valide: {cat_list}").font = Font(italic=True, color="666666")

            # Example row
            ws.cell(row=2, column=1, value="Esempio - Stampo prova")
            ws.cell(row=2, column=2, value="Attrezzatura")
            ws.cell(row=2, column=3, value=1)
            ws.cell(row=2, column=4, value=1500.00)
            ws.cell(row=2, column=5, value="Stampo per prototipo")

            # Auto-width
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 30

            # Save
            save_path = filedialog.asksaveasfilename(
                parent=self,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="Budget_Template.xlsx"
            )
            if save_path:
                wb.save(save_path)
                messagebox.showinfo(self.lang.get('info', 'Info'),
                                    f"Template salvato in:\n{save_path}", parent=self)
                os.startfile(save_path)
        except Exception as e:
            logger.error(f"Errore generazione template: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _import_excel(self):
        """Importa righe budget da un file Excel."""
        if not self._selected_budget_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_first', 'Selezionare un budget.'),
                                   parent=self)
            return

        file_path = filedialog.askopenfilename(
            parent=self,
            filetypes=[("Excel", "*.xlsx *.xls")],
            title=self.lang.get('budget_import_title', 'Seleziona file Excel')
        )
        if not file_path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            imported = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0] or str(row[0]).strip() == '':
                    continue

                desc = str(row[0]).strip()
                cat_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                cat_id = self._categories_map.get(cat_name) if cat_name else None

                try:
                    qty = float(row[2]) if len(row) > 2 and row[2] else 1.0
                except (ValueError, TypeError):
                    qty = 1.0

                try:
                    price = float(row[3]) if len(row) > 3 and row[3] else 0.0
                except (ValueError, TypeError):
                    price = 0.0

                note = str(row[4]).strip() if len(row) > 4 and row[4] else ''

                # Skip "Esempio" or category info rows
                if desc.startswith('Esempio') or desc.startswith('Categorie valide'):
                    continue

                try:
                    session = self.npi_manager._get_session()
                    try:
                        from ..data_models import NpiBudgetItem
                        item = NpiBudgetItem(
                            BudgetId=self._selected_budget_id,
                            ItemDescription=desc,
                            CategoryId=cat_id,
                            Quantity=qty,
                            UnitPrice=price,
                            ItemStatus='InLavorazione',
                            ItemApproval='DaApprovare',
                            Note=note
                        )
                        session.add(item)
                        session.commit()
                        imported += 1
                    finally:
                        session.close()
                except Exception as row_err:
                    errors.append(f"Riga {row_idx}: {row_err}")

            msg = f"Importate {imported} righe."
            if errors:
                msg += f"\n\nErrori ({len(errors)}):\n" + "\n".join(errors[:5])
            messagebox.showinfo(self.lang.get('info', 'Info'), msg, parent=self)
            self._load_items()

        except Exception as e:
            logger.error(f"Errore importazione Excel: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _export_excel(self):
        """Esporta il budget corrente in Excel."""
        if not self._selected_budget_id:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active

            # Budget info
            budget_info = None
            for b in self._budgets_list:
                if b['id'] == self._selected_budget_id:
                    budget_info = b
                    break

            ws.title = (budget_info['name'] if budget_info else 'Budget')[:31]

            # Header row 1: Project info
            ws.merge_cells('A1:G1')
            cell = ws.cell(row=1, column=1, value=f"Budget: {budget_info['name'] if budget_info else ''} — {self.project_name}")
            cell.font = Font(bold=True, size=13, color="0078D4")

            # Headers
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            headers = ['Descrizione', 'Categoria', 'Quantità', 'Prezzo Unit.', 'Totale', 'Stato', 'Approvazione', 'Note']

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center')

            # Data
            row_num = 4
            grand_total = 0.0

            for child in self.tree.get_children():
                vals = self.tree.item(child)['values']
                ws.cell(row=row_num, column=1, value=vals[0])
                ws.cell(row=row_num, column=2, value=vals[1])
                ws.cell(row=row_num, column=3, value=float(str(vals[2]).replace(',', '')) if vals[2] else 0)
                ws.cell(row=row_num, column=4, value=float(str(vals[3]).replace('€ ', '').replace(',', '')) if vals[3] else 0)
                total_val = float(str(vals[4]).replace('€ ', '').replace(',', '')) if vals[4] else 0
                ws.cell(row=row_num, column=5, value=total_val)
                ws.cell(row=row_num, column=6, value=vals[5])
                ws.cell(row=row_num, column=7, value=vals[6])
                ws.cell(row=row_num, column=8, value=vals[7])

                # Format currency
                ws.cell(row=row_num, column=4).number_format = '#,##0.00 €'
                ws.cell(row=row_num, column=5).number_format = '#,##0.00 €'

                grand_total += total_val
                row_num += 1

            # Total row
            ws.cell(row=row_num + 1, column=4, value="TOTALE:").font = Font(bold=True, size=11)
            total_cell = ws.cell(row=row_num + 1, column=5, value=grand_total)
            total_cell.font = Font(bold=True, size=11, color="006100")
            total_cell.number_format = '#,##0.00 €'

            # Auto-width
            for col_idx, width in enumerate([35, 18, 10, 15, 15, 14, 14, 25], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

            # Save
            safe_name = (budget_info['name'] if budget_info else 'Budget').replace(' ', '_')[:20]
            save_path = filedialog.asksaveasfilename(
                parent=self,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"Budget_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            if save_path:
                wb.save(save_path)
                messagebox.showinfo(self.lang.get('info', 'Info'),
                                    f"Budget esportato:\n{save_path}", parent=self)
                os.startfile(save_path)

        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    # ------------------------------------------------------------------ #
    #  Budget Item Documents                                                 #
    # ------------------------------------------------------------------ #
    def _load_docs_for_item(self):
        """Carica i documenti collegati alla riga selezionata."""
        self.doc_tree.delete(*self.doc_tree.get_children())
        if not self._selected_item_id:
            return
        try:
            docs = self.npi_manager.get_documents_for_budget_item(self._selected_item_id)
            for doc in docs:
                self.doc_tree.insert('', 'end', iid=str(doc['doc_id']), values=(
                    doc['doc_title'],
                    doc['doc_type'],
                    doc['date_in'],
                    doc['user']
                ))
        except Exception as e:
            logger.error(f"Errore caricamento documenti riga budget: {e}", exc_info=True)

    def _upload_doc_for_item(self):
        """Carica un nuovo documento e lo associa alla riga di budget selezionata."""
        if not self._selected_item_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_item', 'Selezionare una riga.'),
                                   parent=self)
            return

        filepath = filedialog.askopenfilename(parent=self)
        if not filepath:
            return

        try:
            with open(filepath, 'rb') as f:
                file_data = f.read()
            filename = os.path.basename(filepath)

            # Recupera il primo task_prodotto_id del progetto
            from ..data_models import TaskProdotto, WaveNPI
            session = self.npi_manager._get_session()
            try:
                from sqlalchemy import select
                task_id = session.scalar(
                    select(TaskProdotto.TaskProdottoID)
                    .join(WaveNPI, TaskProdotto.WaveID == WaveNPI.WaveID)
                    .where(WaveNPI.ProgettoID == self.project_id)
                    .limit(1)
                )
            finally:
                session.close()

            # Chiedi tipo documento
            doc_types = self.npi_manager.get_npi_document_types()
            if not doc_types:
                messagebox.showerror(self.lang.get('error', 'Errore'),
                                     'Nessun tipo documento configurato.', parent=self)
                return

            type_map = {dt.NpiDocumentDescription: dt.NpiDocumentTypeId for dt in doc_types}

            # Dialog semplice per tipo
            type_win = tk.Toplevel(self)
            type_win.title(self.lang.get('budget_doc_select_type', 'Tipo Documento'))
            type_win.geometry("350x150")
            type_win.transient(self)
            type_win.grab_set()

            ttk.Label(type_win, text=self.lang.get('budget_doc_select_type', 'Tipo Documento:'),
                       font=("Segoe UI", 10)).pack(pady=(15, 5))
            type_var = tk.StringVar()
            type_combo = ttk.Combobox(type_win, textvariable=type_var, state='readonly',
                                      values=list(type_map.keys()), width=35)
            type_combo.pack(padx=20)
            if type_map:
                type_combo.current(0)

            def on_confirm():
                sel = type_var.get()
                if not sel:
                    return
                doc_type_id = type_map[sel]
                type_win.destroy()

                self.npi_manager.add_npi_document(
                    task_prodotto_id=task_id,
                    doc_type_id=doc_type_id,
                    title=filename,
                    body=file_data,
                    user=self.logged_in_user,
                    budget_item_id=self._selected_item_id
                )
                logger.info(f"Documento '{filename}' caricato per budget item {self._selected_item_id}")
                self._load_docs_for_item()

            ttk.Button(type_win, text=self.lang.get('btn_save', 'Salva'),
                       command=on_confirm).pack(pady=15)

        except Exception as e:
            logger.error(f"Errore upload documento budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _link_existing_doc(self):
        """Mostra documenti progetto non collegati per associarli alla riga selezionata."""
        if not self._selected_item_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_item', 'Selezionare una riga.'),
                                   parent=self)
            return

        try:
            docs = self.npi_manager.get_unlinked_project_documents(self.project_id)
            if not docs:
                messagebox.showinfo(self.lang.get('info', 'Info'),
                                    self.lang.get('budget_doc_none',
                                                  'Nessun documento disponibile da collegare.'),
                                    parent=self)
                return

            # Finestra selezione
            sel_win = tk.Toplevel(self)
            sel_win.title(self.lang.get('budget_doc_select_existing', 'Seleziona Documento'))
            sel_win.geometry("650x400")
            sel_win.transient(self)
            sel_win.grab_set()

            cols = ('doc_title', 'doc_type', 'doc_date', 'doc_user')
            tree = ttk.Treeview(sel_win, columns=cols, show='headings', selectmode='browse')
            tree.heading('doc_title', text=self.lang.get('budget_doc_col_title', 'Titolo'))
            tree.heading('doc_type', text=self.lang.get('budget_doc_col_type', 'Tipo'))
            tree.heading('doc_date', text=self.lang.get('budget_doc_col_date', 'Data'))
            tree.heading('doc_user', text=self.lang.get('budget_doc_col_user', 'Utente'))
            tree.column('doc_title', width=250)
            tree.column('doc_type', width=120)
            tree.column('doc_date', width=130, anchor="center")
            tree.column('doc_user', width=100)

            for doc in docs:
                tree.insert('', 'end', iid=str(doc['doc_id']), values=(
                    doc['doc_title'], doc['doc_type'], doc['date_in'], doc['doc_user'] if 'doc_user' in doc else doc.get('user', '')
                ))

            scroll = ttk.Scrollbar(sel_win, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scroll.set)
            tree.pack(side="left", fill="both", expand=True)
            scroll.pack(side="right", fill="y")

            def on_link():
                sel = tree.selection()
                if not sel:
                    return
                doc_id = int(sel[0])
                self.npi_manager.link_document_to_budget_item(doc_id, self._selected_item_id)
                logger.info(f"Documento {doc_id} collegato a budget item {self._selected_item_id}")
                sel_win.destroy()
                self._load_docs_for_item()

            btn_frame = ttk.Frame(sel_win)
            btn_frame.pack(fill="x", pady=5, side="bottom")
            ttk.Button(btn_frame, text=self.lang.get('budget_btn_link_doc', '🔗 Collega'),
                       command=on_link).pack(side="left", padx=10)
            ttk.Button(btn_frame, text=self.lang.get('btn_cancel', 'Annulla'),
                       command=sel_win.destroy).pack(side="left", padx=10)

        except Exception as e:
            logger.error(f"Errore collegamento documento: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _unlink_doc(self):
        """Scollega il documento selezionato dalla riga di budget."""
        sel = self.doc_tree.selection()
        if not sel:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_doc_select', 'Selezionare un documento.'),
                                   parent=self)
            return

        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            self.lang.get('budget_doc_unlink_confirm',
                          'Scollegare il documento dalla riga di budget?\n(Il documento non verrà eliminato)'),
            parent=self
        ):
            return

        doc_id = int(sel[0])
        try:
            self.npi_manager.unlink_document_from_budget_item(doc_id)
            self._load_docs_for_item()
        except Exception as e:
            logger.error(f"Errore scollegamento documento: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _open_doc(self):
        """Apre il documento selezionato dal pannello documenti."""
        sel = self.doc_tree.selection()
        if not sel:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_doc_select', 'Selezionare un documento.'),
                                   parent=self)
            return

        doc_id = int(sel[0])
        try:
            import tempfile
            documento = self.npi_manager.get_npi_document(doc_id)
            if not documento or not documento.DocumentBody:
                messagebox.showerror(self.lang.get('error', 'Errore'),
                                     self.lang.get('document_not_found', 'Documento non trovato o vuoto.'),
                                     parent=self)
                return

            suffix = os.path.splitext(documento.DocumentTitle)[1] or '.bin'
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(documento.DocumentBody)
                tmp_path = tmp.name
            os.startfile(tmp_path)
            logger.info(f"Documento aperto: {documento.DocumentTitle}")
        except Exception as e:
            logger.error(f"Errore apertura documento: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    # ------------------------------------------------------------------ #
    #  Approval Requests                                                    #
    # ------------------------------------------------------------------ #
    def _request_approval_rows(self):
        """Richiede l'approvazione per le righe selezionate."""
        if not self._selected_budget_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_first', 'Selezionare un budget.'),
                                   parent=self)
            return

        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_approve_select_rows',
                                                 'Selezionare una o più righe da approvare.'),
                                   parent=self)
            return

        count = len(selection)
        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            self.lang.get('budget_approve_rows_confirm',
                          'Inviare richiesta di approvazione per {count} riga/e?').format(count=count),
            parent=self
        ):
            return

        try:
            import socket
            hostname = socket.gethostname()
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudgetItem
                item_ids = [int(iid) for iid in selection]
                for item_id in item_ids:
                    item = session.query(NpiBudgetItem).get(item_id)
                    if item:
                        item.ItemApproval = 'DaApprovare'

                # Inserisci richiesta approvazione
                self._insert_approval_request(
                    session, self._selected_budget_id, item_ids,
                    'Righe', hostname
                )
                session.commit()
                logger.info(f"Richiesta approvazione inviata per {count} righe del budget {self._selected_budget_id}")
            finally:
                session.close()

            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('budget_approve_sent',
                              'Richiesta di approvazione inviata con successo.'),
                parent=self
            )
            self._load_items()
        except Exception as e:
            logger.error(f"Errore richiesta approvazione righe: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _request_approval_all(self):
        """Richiede l'approvazione per l'intero budget."""
        if not self._selected_budget_id:
            messagebox.showwarning(self.lang.get('warning', 'Attenzione'),
                                   self.lang.get('budget_select_first', 'Selezionare un budget.'),
                                   parent=self)
            return

        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            self.lang.get('budget_approve_all_confirm',
                          'Inviare richiesta di approvazione per l\'intero budget?'),
            parent=self
        ):
            return

        try:
            import socket
            hostname = socket.gethostname()
            session = self.npi_manager._get_session()
            try:
                from ..data_models import NpiBudget, NpiBudgetItem

                # Aggiorna stato approvazione budget
                budget = session.query(NpiBudget).get(self._selected_budget_id)
                if budget:
                    budget.ApprovalStatus = 'DaApprovare'

                # Aggiorna tutte le righe non ancora approvate
                items = session.query(NpiBudgetItem).filter(
                    NpiBudgetItem.BudgetId == self._selected_budget_id,
                    NpiBudgetItem.DateOut.is_(None),
                    NpiBudgetItem.ItemApproval != 'Approvato'
                ).all()

                item_ids = [item.BudgetItemId for item in items]
                for item in items:
                    item.ItemApproval = 'DaApprovare'

                # Inserisci richiesta approvazione
                self._insert_approval_request(
                    session, self._selected_budget_id, item_ids,
                    'Budget', hostname
                )
                session.commit()
                logger.info(f"Richiesta approvazione intero budget {self._selected_budget_id}")
            finally:
                session.close()

            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('budget_approve_sent',
                              'Richiesta di approvazione inviata con successo.'),
                parent=self
            )
            self._load_budgets()
        except Exception as e:
            logger.error(f"Errore richiesta approvazione budget: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)

    def _insert_approval_request(self, session, budget_id, item_ids, request_type, hostname):
        """Inserisce una richiesta di approvazione nel DB (per monitoring)."""
        try:
            # Uso raw SQL per inserire nella tabella NpiBudgetApprovalRequests
            item_ids_str = ','.join(str(i) for i in item_ids) if item_ids else ''
            session.execute(
                """INSERT INTO dbo.NpiBudgetApprovalRequests
                   (BudgetId, ItemIds, RequestType, RequestedBy, RequestedDate,
                    ComputerRichiedente, Status)
                   VALUES (:budget_id, :item_ids, :req_type, :req_by, GETDATE(),
                           :hostname, 'Pending')""",
                {
                    'budget_id': budget_id,
                    'item_ids': item_ids_str,
                    'req_type': request_type,
                    'req_by': self.logged_in_user,
                    'hostname': hostname
                }
            )
        except Exception as e:
            logger.error(f"Errore inserimento richiesta approvazione: {e}", exc_info=True)

