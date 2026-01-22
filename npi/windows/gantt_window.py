# npi/windows/gantt_window.py

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import timedelta, datetime
import webbrowser
import os
import logging
import base64
import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.dates as mdates

logger = logging.getLogger(__name__)


class NpiGanttWindow(tk.Toplevel):
    """
    Finestra che genera e visualizza un diagramma di Gantt per un progetto NPI.
    Con funzionalità avanzate di export Excel e PDF.
    """

    def __init__(self, master, npi_manager, lang, progetto_id: int, **kwargs):
        super().__init__(master, **kwargs)
        self.npi_manager = npi_manager
        self.lang = lang
        self.progetto_id = progetto_id
        self.df = None
        self.product_name = ""
        self.gantt_data_raw = None  # Dati originali per statistiche

        self.title(f"Gantt Progetto NPI - ID: {progetto_id}")
        self.geometry("900x650")

        self.create_widgets()
        self.generate_gantt()

    def create_widgets(self):
        """Crea i widget della finestra."""
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # Titolo
        title_label = ttk.Label(
            main_frame,
            text="Generazione Diagramma di Gantt",
            font=("Calibri", 14, "bold")
        )
        title_label.pack(pady=10)

        # Frame pulsanti principali
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)

        ttk.Button(
            button_frame,
            text="🔄 Rigenera Gantt",
            command=self.generate_gantt
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="📊 Esporta Excel Completo",
            command=self.export_to_excel_advanced
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="📄 Esporta PDF",
            command=self.export_to_pdf
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="❌ Chiudi",
            command=self.destroy
        ).pack(side=tk.LEFT, padx=5)

        # Separator
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=10)
        
        # 🆕 TABS PER GERARCHIA PROGETTI
        # Controlla se il progetto ha una gerarchia (padre o figli)
        try:
            self.hierarchy_data = self.npi_manager.get_gantt_hierarchy_data(self.progetto_id)
            self.has_hierarchy = self.hierarchy_data.get('has_hierarchy', False)
        except Exception as e:
            logger.warning(f"Errore recupero gerarchia Gantt: {e}")
            self.hierarchy_data = None
            self.has_hierarchy = False
        
        if self.has_hierarchy:
            # Crea Notebook con tabs
            tabs_frame = ttk.LabelFrame(main_frame, text="🗂️ Viste Gantt", padding="10")
            tabs_frame.pack(fill='x', pady=10)
            
            self.gantt_notebook = ttk.Notebook(tabs_frame)
            self.gantt_notebook.pack(fill='both', expand=True)
            
            # Tab 1: Progetto Corrente (sempre presente)
            tab_current = ttk.Frame(self.gantt_notebook)
            self.gantt_notebook.add(tab_current, text="📋 Progetto Corrente")
            
            current_label = ttk.Label(
                tab_current, 
                text=f"Genera Gantt per: {self.hierarchy_data.get('root_project_name', 'Progetto')}",
                font=("Calibri", 10)
            )
            current_label.pack(pady=10)
            
            # Tab 2: Vista Consolidata (se ha gerarchia)
            tab_consolidated = ttk.Frame(self.gantt_notebook)
            self.gantt_notebook.add(tab_consolidated, text="🔗 Vista Consolidata")
            
            consolidated_label = ttk.Label(
                tab_consolidated,
                text=f"Genera Gantt Consolidato con tutti i {len(self.hierarchy_data['projects'])} progetti",
                font=("Calibri", 10)
            )
            consolidated_label.pack(pady=10)
            
            # Tab 3+: Un tab per ogni progetto figlio
            for proj_data in self.hierarchy_data['projects']:
                if not proj_data['is_root']:  # Solo figli, non il root
                    tab_child = ttk.Frame(self.gantt_notebook)
                    tab_name = proj_data['project_name'][:30]  # Limita lunghezza
                    self.gantt_notebook.add(tab_child, text=f"📄 {tab_name}")
                    
                    child_label = ttk.Label(
                        tab_child,
                        text=f"Genera Gantt per: {proj_data['project_name']}",
                        font=("Calibri", 10)
                    )
                    child_label.pack(pady=10)
            
            # Bind evento cambio tab
            self.gantt_notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)
            
            # Variabile per tracciare tab selezionato
            self.current_gantt_mode = 'current'  # 'current', 'consolidated', 'child_X'
        else:
            # Nessuna gerarchia - comportamento normale (retrocompatibile)
            self.gantt_notebook = None
            self.current_gantt_mode = 'current'

        # Frame statistiche
        stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistiche Progetto", padding="10")
        stats_frame.pack(fill='x', pady=10)

        self.stats_label = ttk.Label(
            stats_frame,
            text="Genera il Gantt per visualizzare le statistiche",
            font=("Calibri", 10)
        )
        self.stats_label.pack()

        # Area di stato
        status_frame = ttk.LabelFrame(main_frame, text="📋 Log Operazioni", padding="10")
        status_frame.pack(fill=tk.BOTH, expand=True)

        self.status_text = tk.Text(status_frame, height=15, width=80, wrap=tk.WORD)
        status_scrollbar = ttk.Scrollbar(status_frame, orient="vertical", command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)

        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        status_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def log_status(self, message):
        """Aggiunge un messaggio allo status text."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.update_idletasks()

    def generate_gantt(self):
        """Genera il diagramma di Gantt."""
        try:
            self.log_status("=== INIZIO GENERAZIONE GANTT ===")
            
            # 🆕 Gestisci diverse modalità di visualizzazione
            if self.current_gantt_mode == 'current':
                # Modalità normale (progetto corrente)
                self._generate_standard_gantt()
            elif self.current_gantt_mode == 'consolidated':
                # Modalità consolidata (gerarchia completa)
                self._generate_consolidated_gantt()
            elif self.current_gantt_mode.startswith('child_'):
                # Modalità singolo progetto figlio
                child_project_id = int(self.current_gantt_mode.split('_')[1])
                self._generate_child_gantt(child_project_id)
            else:
                self.log_status(f"⚠️ Modalità sconosciuta: {self.current_gantt_mode}")
                self._generate_standard_gantt()  # Fallback

        except Exception as e:
            error_msg = f"❌ Errore: {str(e)}"
            self.log_status(error_msg)
            logger.error(f"Errore Gantt: {e}", exc_info=True)
            messagebox.showerror("Errore", f"Impossibile generare il Gantt:\n{str(e)}", parent=self)
    
    def _generate_standard_gantt(self):
        """Genera Gantt standard per il progetto corrente (comportamento originale)."""
        try:
            self.log_status("📋 Modalità: Progetto Corrente")

            # Recupera i dati
            self.log_status("Recupero dati dal database...")
            gantt_data, product_name = self.npi_manager.get_gantt_data(self.progetto_id)

            if not gantt_data:
                self.log_status("❌ Nessun task disponibile per il Gantt")
                messagebox.showinfo("Info", "Nessun task assegnato con date valide.", parent=self)
                return

            self.log_status(f"✅ Ricevuti {len(gantt_data)} task")

            # Salva dati raw
            self.gantt_data_raw = gantt_data

            # Crea DataFrame
            df = pd.DataFrame(gantt_data)

            # Converti date
            df['Start'] = pd.to_datetime(df['Start'])
            df['Finish'] = pd.to_datetime(df['Finish'])

            # Aggiusta durate
            df = self.adjust_durations(df)

            # Calcola statistiche
            self.calculate_statistics(df)

            # Salva per export
            self.df = df.copy()
            self.product_name = product_name

            # Crea grafico
            self.create_chart(df, product_name)
            
        except Exception as e:
            raise  # Rilancia per gestione in generate_gantt()
    
    def _generate_consolidated_gantt(self):
        """Genera Gantt consolidato con gerarchia progetti (Step 2)."""
        try:
            self.log_status("🔗 Modalità: Vista Consolidata")
            
            if not self.hierarchy_data:
                self.log_status("❌ Nessun dato gerarchia disponibile")
                messagebox.showwarning("Avviso", "Dati gerarchia non disponibili", parent=self)
                return
            
            projects = self.hierarchy_data.get('projects', [])
            if not projects:
                self.log_status("❌ Nessun progetto nella gerarchia")
                return
            
            self.log_status(f"📊 Processa {len(projects)} progetti nella gerarchia...")
            
            # Costruisci lista di task consolidata
            consolidated_tasks = []
            
            for proj_data in projects:
                project_name = proj_data['project_name']
                level = proj_data['level']
                is_root = proj_data['is_root']
                
                # Indentazione visuale basata sul livello
                indent = "  " * level
                prefix = "📦" if is_root else "📄"
                
                self.log_status(f"{indent}{prefix} {project_name} (Level {level})")
                
                # Per ogni task del progetto
                tasks = proj_data.get('tasks', [])
                for task in tasks:
                    # Aggiungi indicatore livello al nome task
                    task_modified = task.copy()
                    task_modified['Task'] = f"{indent}  └─ {task['Task']}"
                    task_modified['_project_name'] = project_name
                    task_modified['_level'] = level
                    consolidated_tasks.append(task_modified)
            
            if not consolidated_tasks:
                self.log_status("❌ Nessun task trovato nella gerarchia")
                messagebox.showinfo("Info", "Nessun task disponibile", parent=self)
                return
            
            self.log_status(f"✅ Totale {len(consolidated_tasks)} task consolidati")
            
            # Salva dati raw
            self.gantt_data_raw = consolidated_tasks
            
            # Crea DataFrame
            df = pd.DataFrame(consolidated_tasks)
            
            # Converti date
            df['Start'] = pd.to_datetime(df['Start'])
            df['Finish'] = pd.to_datetime(df['Finish'])
            
            # Aggiusta durate
            df = self.adjust_durations(df)
            
            # Calcola statistiche
            self.calculate_statistics(df)
            
            # Salva per export
            self.df = df.copy()
            self.product_name = f"Consolidato: {self.hierarchy_data.get('root_project_name', 'Progetti')}"
            
            # Crea grafico
            self.create_chart(df, self.product_name)
            
        except Exception as e:
            raise  # Rilancia per gestione in generate_gantt()
    
    def _generate_child_gantt(self, child_project_id):
        """Genera Gantt per un singolo progetto figlio."""
        try:
            self.log_status(f"📄 Modalità: Progetto Figlio (ID: {child_project_id})")
            
            # Trova i dati del progetto figlio
            child_data = None
            for proj in self.hierarchy_data.get('projects', []):
                if proj['project_id'] == child_project_id:
                    child_data = proj
                    break
            
            if not child_data:
                self.log_status(f"❌ Progetto {child_project_id} non trovato")
                return
            
            project_name = child_data['project_name']
            tasks = child_data.get('tasks', [])
            
            if not tasks:
                self.log_status(f"❌ Nessun task per progetto {project_name}")
                messagebox.showinfo("Info", f"Nessun task per {project_name}", parent=self)
                return
            
            self.log_status(f"✅ Ricevuti {len(tasks)} task per {project_name}")
            
            # Salva dati raw
            self.gantt_data_raw = tasks
            
            # Crea DataFrame
            df = pd.DataFrame(tasks)
            
            # Converti date
            df['Start'] = pd.to_datetime(df['Start'])
            df['Finish'] = pd.to_datetime(df['Finish'])
            
            # Aggiusta durate
            df = self.adjust_durations(df)
            
            # Calcola statistiche
            self.calculate_statistics(df)
            
            # Salva per export
            self.df = df.copy()
            self.product_name = project_name
            
            # Crea grafico
            self.create_chart(df, project_name)
            
        except Exception as e:
            raise  # Rilancia per gestione in generate_gantt()

    def adjust_durations(self, df):
        """Aggiusta le durate dei task."""
        self.log_status("Aggiustamento durate task...")

        for idx, row in df.iterrows():
            duration = (row['Finish'] - row['Start']).days
            if duration <= 0:
                df.at[idx, 'Finish'] = row['Start'] + timedelta(days=1)
                self.log_status(f"  Task '{row['Task']}': durata impostata a 1 giorno")

        return df

    def calculate_statistics(self, df):
        """Calcola e visualizza statistiche del progetto."""
        try:
            total_tasks = len(df)

            # Calcola durata totale progetto
            project_start = df['Start'].min()
            project_end = df['Finish'].max()
            project_duration = (project_end - project_start).days

            # Task per owner
            tasks_per_owner = df['Owner'].value_counts().to_dict()
            
            # Task per category
            tasks_per_category = df.get('Category', pd.Series(['N/A']*len(df))).value_counts().to_dict()

            # Durata media task
            df['Duration'] = (df['Finish'] - df['Start']).dt.days
            avg_duration = df['Duration'].mean()

            # Task critici (durata > media)
            critical_tasks = len(df[df['Duration'] > avg_duration])

            # Verifica ritardi (se oggi è oltre la data di fine)
            today = pd.Timestamp.now()
            late_tasks = len(df[df['Finish'] < today])

            # Percentuale completamento (stima basata su date)
            if today < project_start:
                completion = 0
            elif today > project_end:
                completion = 100
            else:
                elapsed = (today - project_start).days
                completion = min(100, (elapsed / project_duration) * 100)

            # Aggiorna label statistiche
            stats_text = f"""
📊 Task Totali: {total_tasks} | ⏱️ Durata Progetto: {project_duration} giorni
📅 Inizio: {project_start.strftime('%d/%m/%Y')} | 🏁 Fine: {project_end.strftime('%d/%m/%Y')}
📈 Completamento Stimato: {completion:.1f}% | ⚠️ Task in Ritardo: {late_tasks}
🎯 Task Critici (>media): {critical_tasks} | ⏳ Durata Media: {avg_duration:.1f} giorni
👥 Assegnazioni: {', '.join([f'{k}({v})' for k, v in tasks_per_owner.items()])}
📂 Categorie: {', '.join([f'{k}({v})' for k, v in tasks_per_category.items()])}
            """.strip()

            self.stats_label.config(text=stats_text)
            self.log_status("✅ Statistiche calcolate")

            # Salva per export
            self.statistics = {
                'total_tasks': total_tasks,
                'project_start': project_start,
                'project_end': project_end,
                'project_duration': project_duration,
                'avg_duration': avg_duration,
                'critical_tasks': critical_tasks,
                'late_tasks': late_tasks,
                'completion': completion,
                'tasks_per_owner': tasks_per_owner,
                'tasks_per_category': tasks_per_category
            }

        except Exception as e:
            logger.error(f"Errore calcolo statistiche: {e}")
            self.stats_label.config(text="⚠️ Errore nel calcolo delle statistiche")

    def create_chart(self, df, product_name):
        """Crea il diagramma di Gantt professionale con Logo, Dipendenze e Milestones."""
        try:
            self.log_status("Generazione visualizzazione grafica premium...")

            if df is None or df.empty:
                self.log_status("⚠️ Nessun dato da visualizzare")
                return

            # 0. Gestione Logo (Ricerca robusta)
            logo_base64 = ""
            try:
                # Cerca in: 1. CWD, 2. Script Dir, 3. Root Dir
                possible_paths = [
                    os.path.join(os.getcwd(), "Logo.png"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "Logo.png"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logo.png")
                ]
                
                logo_path = None
                for p in possible_paths:
                    if os.path.exists(p):
                        logo_path = p
                        break
                
                if logo_path:
                    with open(logo_path, "rb") as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode()
                        logo_base64 = f"data:image/png;base64,{encoded_string}"
                    logger.info(f"Logo caricato con successo da: {logo_path}")
                else:
                    logger.warning("File Logo.png non trovato nei percorsi previsti.")
                    self.log_status("⚠️ Logo non trovato")
            except Exception as e:
                logger.warning(f"Errore caricamento logo: {e}")

            # 1. Preparazione Dati
            df['Start'] = pd.to_datetime(df['Start'])
            df['Finish'] = pd.to_datetime(df['Finish'])
            
            # Ordinamento per Categoria e poi Data Inizio (molto importante per il raggruppamento visuale)
            # I task sono mostrati dal più vecchio al più nuovo (ordine cronologico)
            df = df.sort_values(by=['Category', 'Start'], ascending=[True, True])
            
            # Etichetta asse Y: Categoria + Nome Task + ID per unicità
            df['Label'] = df.apply(lambda r: f"<b>[{r['Category'].upper()}]</b><br>{r['Task']} (ID:{r['TaskProdottoID']})", axis=1)
            
            # Crea mapping ID -> Label per le dipendenze
            id_to_label = {row['TaskProdottoID']: row['Label'] for _, row in df.iterrows()}

            fig = go.Figure()

            # 2. Rendering Task
            for i, row in df.iterrows():
                # DURATA IN MILLISECONDI (fondamentale per Plotly go.Bar su date)
                diff = row['Finish'] - row['Start']
                duration_ms = diff.total_seconds() * 1000
                is_milestone = duration_ms <= 0
                
                # 🆕 Logica Colori con Task in Ritardo e Target NPI
                today = pd.Timestamp.now()
                is_late = (row['Finish'] < today) and (row['Status'] != 'Completato')
                is_target_npi = row.get('IsTargetNPI', False)
                
                # Priorità: Target NPI > In Ritardo > Stato normale
                if is_target_npi:
                    # Task Target NPI = BLU (anche se in ritardo)
                    color = "#0078d4"  # Blu Microsoft
                elif is_late:
                    # Task in ritardo = ROSSO
                    color = "#e74c3c"  # Rosso
                elif row['Status'] == 'Completato':
                    color = "#2ecc71"  # Verde
                elif row['Status'] == 'In Lavorazione':
                    color = "#f1c40f"  # Giallo
                else:
                    color = "#3498db"  # Blu chiaro (default)

                status_text = row.get('Status', 'N/A')
                completion_val = row.get('Completion', 0)

                if is_milestone:
                    fig.add_trace(go.Scatter(
                        x=[row['Start']], y=[row['Label']],
                        mode='markers',
                        marker=dict(symbol='diamond', size=14, color='#8e44ad', line=dict(width=1, color='black')),
                        name="Milestone",
                        hovertext=f"<b>{row['Task']}</b><br>Data: {row['Start'].strftime('%d/%m/%Y')}",
                        hoverinfo='text', showlegend=False
                    ))
                else:
                    # Barra Task
                    fig.add_trace(go.Bar(
                        base=[row['Start']],
                        x=[duration_ms],
                        y=[row['Label']],
                        orientation='h',
                        marker=dict(color=color, line=dict(color='black', width=0.5)),
                        hovertext=f"<b>{row['Task']}</b><br>Responsabile: {row['Owner']}<br>Fine prevista: {row['Finish'].strftime('%d/%m/%Y')}<br>Completamento: {completion_val}%",
                        hoverinfo='text',
                        text=f" {completion_val}% ",
                        textposition='auto',
                        showlegend=False
                    ))
                    
                    # Overlay Progresso
                    if 0 < completion_val < 100:
                        fig.add_trace(go.Bar(
                            base=[row['Start']],
                            x=[duration_ms * (completion_val / 100)],
                            y=[row['Label']],
                            orientation='h',
                            marker=dict(color='rgba(0,0,0,0.2)'),
                            showlegend=False, hoverinfo='skip'
                        ))

            # 3. Dipendenze
            for _, row in df.iterrows():
                deps = row.get('Dependencies', [])
                for dep_id in deps:
                    if dep_id in id_to_label:
                        pred_label = id_to_label[dep_id]
                        # Utilizziamo ID per il matching robusto
                        pred_rows = df[df['TaskProdottoID'] == dep_id]
                        if not pred_rows.empty:
                            p_row = pred_rows.iloc[0]
                            # Colore più scuro e visibile per le linee di collegamento
                            link_color = "rgba(44, 62, 80, 0.7)" # Blu-grigio scuro
                            
                            fig.add_trace(go.Scatter(
                                x=[p_row['Finish'], p_row['Finish'], row['Start']],
                                y=[pred_label, row['Label'], row['Label']],
                                mode='lines+markers',
                                line=dict(color=link_color, width=1.4),
                                marker=dict(
                                    size=[0, 0, 8], 
                                    symbol="triangle-right",
                                    color=link_color
                                ),
                                showlegend=False, 
                                hoverinfo='text',
                                hovertext=f"Dipendenza: {p_row['Task']} ➔ {row['Task']}"
                            ))

            # 4. Layout
            fig.update_layout(
                title=dict(
                    text=f"<b>PROJECT GANTT: {product_name}</b>",
                    x=0.5, y=0.96, font=dict(size=24, color='#2c3e50')
                ),
                xaxis=dict(type='date', side='top', gridcolor='lightgrey', tickformat='%d/%m/%y'),
                yaxis=dict(autorange="reversed", gridcolor='whitesmoke', automargin=True),
                plot_bgcolor='white', paper_bgcolor='white',
                height=max(600, len(df) * 45 + 160),
                margin=dict(l=20, r=40, t=120, b=50),
                barmode='overlay'
            )

            if logo_base64:
                fig.add_layout_image(
                    dict(
                        source=logo_base64,
                        xref="paper", yref="paper",
                        x=0, y=1.15,
                        sizex=0.18, sizey=0.18,
                        xanchor="left", yanchor="top",
                        layer="above"
                    )
                )

            # Oggi
            today = datetime.now()
            fig.add_shape(
                type="line", x0=today, x1=today, y0=0, y1=1,
                xref="x", yref="paper", line=dict(color="red", width=1.5, dash="dash")
            )
            fig.add_annotation(
                x=today, y=1.02, yref="paper", text="OGGI",
                showarrow=False, font=dict(color="red", size=11, family="Arial Black")
            )

            # 5. Export
            temp_file = os.path.join(tempfile.gettempdir(), f"Gantt_NPI_{self.progetto_id}.html")
            fig.write_html(temp_file, auto_open=False)
            webbrowser.open(f'file://{temp_file}')
            
            self.log_status("✅ Gantt grafico rigenerato con successo!")

        except Exception as e:
            self.log_status(f"❌ Errore grafico: {str(e)}")
            logger.error(f"Gantt creation error: {e}", exc_info=True)
            raise

    # ========== EXPORT EXCEL AVANZATO ==========

    def export_to_excel_advanced(self):
        """Esporta il Gantt in formato Excel con tutte le funzionalità avanzate."""
        if self.df is None or self.df.empty:
            messagebox.showwarning(
                "Attenzione",
                "Genera prima il Gantt prima di esportare!",
                parent=self
            )
            return

        try:
            # Chiedi dove salvare
            default_filename = f"Gantt_NPI_{self.progetto_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                parent=self,
                title="Salva Gantt Excel",
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if not file_path:
                return

            self.log_status("=== EXPORT EXCEL AVANZATO ===")
            self.log_status(f"Salvataggio in: {file_path}")

            # Crea workbook
            wb = Workbook()
            wb.remove(wb.active)  # Rimuovi foglio default

            # 1️⃣ Foglio Riepilogo
            self.log_status("Creazione foglio Riepilogo...")
            self._create_summary_sheet(wb)

            # 2️⃣ Foglio Gantt Chart con tabella e filtri
            self.log_status("Creazione foglio Gantt Chart...")
            self._create_gantt_sheet(wb)

            # 3️⃣ Foglio Timeline Visuale
            self.log_status("Creazione foglio Timeline Visuale...")
            self._create_timeline_sheet(wb)

            # 4️⃣ Foglio Statistiche con grafici
            self.log_status("Creazione foglio Statistiche...")
            self._create_statistics_sheet(wb)

            # 5️⃣ Foglio Task Critici
            self.log_status("Creazione foglio Task Critici...")
            self._create_critical_tasks_sheet(wb)

            # Salva
            wb.save(file_path)

            self.log_status(f"✅ Excel salvato con successo!")
            self.log_status(f"📁 {file_path}")
            self.log_status(f"📊 Fogli creati: {len(wb.sheetnames)}")

            # Chiedi se aprire
            if messagebox.askyesno(
                    "Successo",
                    f"File Excel creato con successo!\n\nFogli inclusi:\n{chr(10).join(['• ' + s for s in wb.sheetnames])}\n\nVuoi aprirlo ora?",
                    parent=self
            ):
                os.startfile(file_path)

        except Exception as e:
            error_msg = f"Errore export Excel: {str(e)}"
            self.log_status(f"❌ {error_msg}")
            logger.error(error_msg, exc_info=True)
            messagebox.showerror("Errore", f"Impossibile esportare:\n{str(e)}", parent=self)

    def _create_summary_sheet(self, wb):
        """Crea foglio riepilogo progetto."""
        ws = wb.create_sheet("📋 Riepilogo", 0)

        # Titolo
        ws['B2'] = f"Gantt Chart - {self.product_name}"
        ws['B2'].font = Font(size=20, bold=True, color="1F4E78")
        ws.merge_cells('B2:F2')

        ws['B3'] = f"Progetto NPI ID: {self.progetto_id}"
        ws['B3'].font = Font(size=14, italic=True, color="666666")

        ws['B4'] = f"Generato il: {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
        ws['B4'].font = Font(size=10, color="999999")

        # Box statistiche
        row = 7
        stats = self.statistics

        # Stile box
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        value_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        value_font = Font(size=11)

        stats_data = [
            ("📊 Informazioni Generali", ""),
            ("Task Totali", stats['total_tasks']),
            ("Data Inizio Progetto", stats['project_start'].strftime('%d/%m/%Y')),
            ("Data Fine Progetto", stats['project_end'].strftime('%d/%m/%Y')),
            ("Durata Totale (giorni)", stats['project_duration']),
            ("", ""),
            ("📈 Metriche Performance", ""),
            ("Completamento Stimato", f"{stats['completion']:.1f}%"),
            ("Durata Media Task (giorni)", f"{stats['avg_duration']:.1f}"),
            ("Task Critici (>media)", stats['critical_tasks']),
            ("Task in Ritardo", stats['late_tasks']),
            ("", ""),
            ("👥 Distribuzione Carico", ""),
        ]

        # Aggiungi task per owner
        for owner, count in stats['tasks_per_owner'].items():
            stats_data.append((f"  • {owner}", count))

        for idx, (label, value) in enumerate(stats_data, start=row):
            if label.startswith(("📊", "📈", "👥")):  # Headers
                cell = ws.cell(row=idx, column=2, value=label)
                cell.font = header_font
                cell.fill = header_fill
                ws.merge_cells(f'B{idx}:C{idx}')
            elif label == "":  # Spacer
                continue
            else:
                ws.cell(row=idx, column=2, value=label).font = Font(size=10)
                cell = ws.cell(row=idx, column=3, value=value)
                cell.font = value_font
                cell.fill = value_fill

        # Larghezza colonne
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

        # Nota finale
        note_row = row + len(stats_data) + 3
        ws[f'B{note_row}'] = "💡 Nota: Consulta i fogli successivi per dettagli e visualizzazioni"
        ws[f'B{note_row}'].font = Font(italic=True, color="666666", size=9)

    def _create_gantt_sheet(self, wb):
        """Crea foglio Gantt con tabella, filtri e formattazione condizionale."""
        ws = wb.create_sheet("📊 Gantt Chart")

        # Prepara dati
        export_df = self.df.copy()
        
        # Ordinamento per Categoria e poi Data Inizio
        export_df = export_df.sort_values(by=['Category', 'Start'], ascending=[True, True])
        
        export_df['Start'] = export_df['Start'].dt.strftime('%d/%m/%Y')
        export_df['Finish'] = export_df['Finish'].dt.strftime('%d/%m/%Y')
        export_df['Durata (gg)'] = (
                pd.to_datetime(self.df['Finish']) - pd.to_datetime(self.df['Start'])
        ).dt.days

        # Calcola giorni rimanenti
        today = datetime.now()
        export_df['Giorni Rimanenti'] = (
                pd.to_datetime(self.df['Finish']) - today
        ).dt.days

        # Status
        export_df['Status'] = export_df['Giorni Rimanenti'].apply(
            lambda x: '✅ Completato' if x < 0 else ('⚠️ In Ritardo' if x < 7 else '🟢 In Corso')
        )

        # Riordina colonne includendo CATEGORY
        column_order = ['Category', 'Task', 'Owner', 'Start', 'Finish', 'Durata (gg)', 'Giorni Rimanenti', 'Status']
        export_df = export_df[column_order]

        # Scrivi header
        ws['A1'] = f"Gantt Chart - {self.product_name}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:H1')

        # Scrivi dati (inizia dalla riga 3)
        start_row = 3

        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # ⭐ APPLICA TABELLA CON FILTRI
        tab = Table(displayName="GanttTable", ref=f"A{start_row}:H{start_row + len(export_df)}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # ⭐ FORMATTAZIONE CONDIZIONALE
        # Giorni Rimanenti (colonna G ora, dopo Category e Task)
        col_letter = 'G'
        first_row = start_row + 1
        last_row = start_row + len(export_df)

        # Rosso per negativi (ritardo)
        ws.conditional_formatting.add(
            f'{col_letter}{first_row}:{col_letter}{last_row}',
            CellIsRule(
                operator='lessThan',
                formula=['0'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006", bold=True)
            )
        )

        # Giallo per urgenti (< 7 giorni)
        ws.conditional_formatting.add(
            f'{col_letter}{first_row}:{col_letter}{last_row}',
            CellIsRule(
                operator='between',
                formula=['0', '7'],
                fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                font=Font(color="9C6500")
            )
        )

        # Verde per OK (> 7 giorni)
        ws.conditional_formatting.add(
            f'{col_letter}{first_row}:{col_letter}{last_row}',
            CellIsRule(
                operator='greaterThan',
                formula=['7'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                font=Font(color="006100")
            )
        )

        # ⭐ COLOR SCALE per Durata (colonna F ora)
        ws.conditional_formatting.add(
            f'F{first_row}:F{last_row}',
            ColorScaleRule(
                start_type='min',
                start_color='63BE7B',
                mid_type='percentile',
                mid_value=50,
                mid_color='FFEB84',
                end_type='max',
                end_color='F8696B'
            )
        )

        # Larghezza colonne
        ws.column_dimensions['A'].width = 25 # Category
        ws.column_dimensions['B'].width = 40 # Task
        ws.column_dimensions['C'].width = 20 # Owner
        ws.column_dimensions['D'].width = 12 # Start
        ws.column_dimensions['E'].width = 12 # Finish
        ws.column_dimensions['F'].width = 12 # Duration
        ws.column_dimensions['G'].width = 18 # Days Left
        ws.column_dimensions['H'].width = 18 # Status

        # Freeze panes
        ws.freeze_panes = 'A4'

    def _create_timeline_sheet(self, wb):
        """Crea foglio con timeline visuale del Gantt."""
        ws = wb.create_sheet("📅 Timeline Visuale")

        # Calcola range date
        min_date = pd.to_datetime(self.df['Start']).min()
        max_date = pd.to_datetime(self.df['Finish']).max()
        
        # Crea lista di settimane
        date_range = pd.date_range(start=min_date, end=max_date, freq='W-MON')

        # Ordinamento per Categoria e poi Data Inizio
        df_sortedRequested = self.df.sort_values(by=['Category', 'Start'], ascending=[True, True])

        # Header
        ws['A1'] = "Categoria"
        ws['B1'] = "Attività / Settimana"
        for cell_ref in ['A1', 'B1']:
            cell = ws[cell_ref]
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Date header (Inizia da colonna C ora)
        for col_idx, date in enumerate(date_range, start=3):
            cell = ws.cell(row=1, column=col_idx, value=date.strftime('%d/%m'))
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
            ws.column_dimensions[cell.column_letter].width = 3.5

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.row_dimensions[1].height = 60

        # Righe task
        task_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

        for row_idx, (_, row) in enumerate(df_sortedRequested.iterrows(), start=2):
            # Categoria
            cell_cat = ws.cell(row=row_idx, column=1, value=row['Category'])
            cell_cat.font = Font(size=9, bold=True)
            cell_cat.alignment = Alignment(vertical="center")
            
            # Nome task
            cell_task = ws.cell(row=row_idx, column=2, value=row['Task'])
            cell_task.font = Font(size=9)
            cell_task.alignment = Alignment(vertical="center")

            task_start = pd.to_datetime(row['Start'])
            task_end = pd.to_datetime(row['Finish'])

            # Colora celle nel range
            for col_idx, date in enumerate(date_range, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if task_start <= date <= task_end:
                    cell.fill = task_fill
                    cell.border = border

        # Freeze panes
        ws.freeze_panes = 'C2'

        # Legenda
        legend_row = len(self.df) + 4
        ws.cell(row=legend_row, column=1, value="Legenda:")
        ws.cell(row=legend_row, column=1).font = Font(bold=True, size=10)

        ws.cell(row=legend_row + 1, column=1, value="■ Task attivo")
        ws.cell(row=legend_row + 1, column=2).fill = task_fill

    def _create_statistics_sheet(self, wb):
        """Crea foglio statistiche con grafici Excel."""
        ws = wb.create_sheet("📈 Statistiche")

        # Titolo
        ws['B2'] = "Statistiche e Analisi Progetto"
        ws['B2'].font = Font(size=16, bold=True, color="1F4E78")
        ws.merge_cells('B2:F2')

        # ⭐ GRAFICO 1: Task per Owner (Bar Chart)
        row = 5
        ws.cell(row=row, column=2, value="Owner").font = Font(bold=True)
        ws.cell(row=row, column=3, value="N. Task").font = Font(bold=True)

        tasks_per_owner = self.statistics['tasks_per_owner']
        for idx, (owner, count) in enumerate(tasks_per_owner.items(), start=row + 1):
            ws.cell(row=idx, column=2, value=owner)
            ws.cell(row=idx, column=3, value=count)

        # Crea grafico a barre
        chart1 = BarChart()
        chart1.title = "Distribuzione Task per Owner"
        chart1.x_axis.title = "Owner"
        chart1.y_axis.title = "Numero Task"

        data1 = Reference(ws, min_col=3, min_row=row, max_row=row + len(tasks_per_owner))
        cats1 = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(tasks_per_owner))

        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.height = 10
        chart1.width = 15
        ws.add_chart(chart1, f"E{row}")

        # ⭐ GRAFICO 2: Task per Categoria (Bar Chart)
        row_cat = row
        ws.cell(row=row_cat, column=8, value="Categoria").font = Font(bold=True)
        ws.cell(row=row_cat, column=9, value="N. Task").font = Font(bold=True)

        tasks_per_cat = self.statistics['tasks_per_category']
        for idx, (cat, count) in enumerate(tasks_per_cat.items(), start=row_cat + 1):
            ws.cell(row=idx, column=8, value=cat)
            ws.cell(row=idx, column=9, value=count)

        chart_cat = BarChart()
        chart_cat.title = "Task per Categoria"
        chart_cat.x_axis.title = "Categoria"
        chart_cat.y_axis.title = "Numero Task"

        data_cat = Reference(ws, min_col=9, min_row=row_cat, max_row=row_cat + len(tasks_per_cat))
        cats_cat = Reference(ws, min_col=8, min_row=row_cat + 1, max_row=row_cat + len(tasks_per_cat))

        chart_cat.add_data(data_cat, titles_from_data=True)
        chart_cat.set_categories(cats_cat)
        chart_cat.height = 10
        chart_cat.width = 15
        ws.add_chart(chart_cat, f"K{row}")

        # ⭐ GRAFICO 2: Durata Task (Bar Chart)
        row = row + len(tasks_per_owner) + 5

        ws.cell(row=row, column=2, value="Task").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Durata (gg)").font = Font(bold=True)

        df_sorted = self.df.sort_values('Start')
        for idx, (_, task_row) in enumerate(df_sorted.iterrows(), start=row + 1):
            duration = (task_row['Finish'] - task_row['Start']).days
            ws.cell(row=idx, column=2, value=task_row['Task'][:30])  # Tronca nome
            ws.cell(row=idx, column=3, value=duration)

        # Crea grafico durate
        chart2 = BarChart()
        chart2.type = "bar"  # Orizzontale
        chart2.title = "Durata Task (giorni)"
        chart2.x_axis.title = "Giorni"
        chart2.y_axis.title = "Task"

        data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df_sorted))
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df_sorted))

        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = max(10, len(df_sorted) * 0.5)
        chart2.width = 20

        ws.add_chart(chart2, f"E{row}")

        # Larghezza colonne
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15

    def _create_critical_tasks_sheet(self, wb):
        """Crea foglio con task critici e in ritardo."""
        ws = wb.create_sheet("⚠️ Task Critici")

        # Titolo
        ws['B2'] = "Task Critici e in Ritardo"
        ws['B2'].font = Font(size=14, bold=True, color="C00000")
        ws.merge_cells('B2:F2')

        ws['B3'] = "Task con durata superiore alla media o scaduti"
        ws['B3'].font = Font(size=10, italic=True, color="666666")

        # Calcola task critici
        df = self.df.copy()
        df['Duration'] = (df['Finish'] - df['Start']).dt.days
        avg_duration = df['Duration'].mean()

        today = pd.Timestamp.now()
        df['Days_Remaining'] = (df['Finish'] - today).dt.days
        df['Is_Late'] = df['Days_Remaining'] < 0
        df['Is_Critical'] = df['Duration'] > avg_duration

        critical_df = df[df['Is_Critical'] | df['Is_Late']].copy()

        if critical_df.empty:
            ws['B6'] = "✅ Nessun task critico rilevato!"
            ws['B6'].font = Font(size=12, color="00B050", bold=True)
            return

        # Header
        row = 6
        headers = ['Task', 'Owner', 'Inizio', 'Fine', 'Durata (gg)', 'Giorni Rimanenti', 'Criticità']
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dati
        for idx, (_, task) in enumerate(critical_df.iterrows(), start=row + 1):
            ws.cell(row=idx, column=2, value=task['Task'])
            ws.cell(row=idx, column=3, value=task['Owner'])
            ws.cell(row=idx, column=4, value=task['Start'].strftime('%d/%m/%Y'))
            ws.cell(row=idx, column=5, value=task['Finish'].strftime('%d/%m/%Y'))
            ws.cell(row=idx, column=6, value=task['Duration'])
            ws.cell(row=idx, column=7, value=task['Days_Remaining'])

            # Criticità
            criticality = []
            if task['Is_Late']:
                criticality.append("🔴 IN RITARDO")
            if task['Is_Critical']:
                criticality.append("⚠️ DURATA ELEVATA")

            cell = ws.cell(row=idx, column=8, value=" | ".join(criticality))
            cell.font = Font(bold=True, color="C00000")

            # Colora riga se in ritardo
            if task['Is_Late']:
                for col in range(2, 9):
                    ws.cell(row=idx, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

        # Larghezza colonne
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 30

        # Nota
        note_row = row + len(critical_df) + 3
        ws[f'B{note_row}'] = f"📊 Durata media task: {avg_duration:.1f} giorni"
        ws[f'B{note_row}'].font = Font(italic=True, size=9, color="666666")

    # ========== EXPORT PDF ==========

    def export_to_pdf(self):
        """Esporta il Gantt in formato PDF."""
        if self.df is None or self.df.empty:
            messagebox.showwarning(
                "Attenzione",
                "Genera prima il Gantt prima di esportare!",
                parent=self
            )
            return

        try:
            # Chiedi dove salvare
            default_filename = f"Gantt_NPI_{self.progetto_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = filedialog.asksaveasfilename(
                parent=self,
                title="Salva Gantt PDF",
                defaultextension=".pdf",
                initialfile=default_filename,
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )

            if not file_path:
                return

            self.log_status("=== EXPORT PDF ===")
            self.log_status(f"Salvataggio in: {file_path}")

            with PdfPages(file_path) as pdf:
                # Pagina 1: Riepilogo
                self.log_status("Creazione pagina riepilogo...")
                self._create_pdf_summary_page(pdf)

                # Pagina 2: Gantt Chart
                self.log_status("Creazione pagina Gantt...")
                self._create_pdf_gantt_page(pdf)

                # Pagina 3: Statistiche
                self.log_status("Creazione pagina statistiche...")
                self._create_pdf_statistics_page(pdf)

                # Metadata
                d = pdf.infodict()
                d['Title'] = f'Gantt Chart - {self.product_name}'
                d['Author'] = 'NPI Manager'
                d['Subject'] = f'Progetto NPI ID: {self.progetto_id}'
                d['Keywords'] = 'Gantt, NPI, Project Management'
                d['CreationDate'] = datetime.now()

            self.log_status(f"✅ PDF salvato con successo!")
            self.log_status(f"📁 {file_path}")

            if messagebox.askyesno(
                    "Successo",
                    "File PDF creato con successo!\n\nVuoi aprirlo ora?",
                    parent=self
            ):
                os.startfile(file_path)

        except Exception as e:
            error_msg = f"Errore export PDF: {str(e)}"
            self.log_status(f"❌ {error_msg}")
            logger.error(error_msg, exc_info=True)
            messagebox.showerror("Errore", f"Impossibile esportare PDF:\n{str(e)}", parent=self)

    def _create_pdf_summary_page(self, pdf):
        """Crea pagina riepilogo nel PDF."""
        fig, ax = plt.subplots(figsize=(11, 8.5))
        ax.axis('off')

        # Titolo
        ax.text(0.5, 0.95, f"Gantt Chart - {self.product_name}",
                ha='center', va='top', fontsize=20, fontweight='bold')

        ax.text(0.5, 0.90, f"Progetto NPI ID: {self.progetto_id}",
                ha='center', va='top', fontsize=14, style='italic', color='gray')

        ax.text(0.5, 0.87, f"Generato il: {datetime.now().strftime('%d/%m/%Y alle %H:%M')}",
                ha='center', va='top', fontsize=10, color='gray')

        # Box statistiche
        stats = self.statistics
        y_pos = 0.75

        stats_text = f"""
📊 INFORMAZIONI GENERALI
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Task Totali: {stats['total_tasks']}
Data Inizio: {stats['project_start'].strftime('%d/%m/%Y')}
Data Fine: {stats['project_end'].strftime('%d/%m/%Y')}
Durata Totale: {stats['project_duration']} giorni

📈 METRICHE PERFORMANCE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Completamento Stimato: {stats['completion']:.1f}%
Durata Media Task: {stats['avg_duration']:.1f} giorni
Task Critici (>media): {stats['critical_tasks']}
Task in Ritardo: {stats['late_tasks']}

👥 DISTRIBUZIONE CARICO
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        for owner, count in stats['tasks_per_owner'].items():
            stats_text += f"  • {owner}: {count} task\n"

        ax.text(0.1, y_pos, stats_text,
                ha='left', va='top', fontsize=11, family='monospace',
                bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.3))

        pdf.savefig(fig, bbox_inches='tight')
        plt.close()

    def _create_pdf_gantt_page(self, pdf):
        """Crea pagina Gantt nel PDF con raggruppamento per categoria."""
        fig, ax = plt.subplots(figsize=(11, 8.5))

        # Prepara dati
        df = self.df.copy()
        # Ordinamento per Categoria e poi Data Inizio (dal più vecchio al più nuovo)
        df = df.sort_values(by=['Category', 'Start'], ascending=[True, True])
        
        # Inverti l'ordine per matplotlib (matplotlib mette indice 0 in basso, indici alti in alto)
        # Così i task più vecchi (che erano all'inizio) finiscono con indici alti e appaiono in alto
        df = df.iloc[::-1].reset_index(drop=True)
        
        # Mappa ID -> Indice per tracciare dipendenze nell'asse Y
        id_to_idx = {row['TaskProdottoID']: i for i, (_, row) in enumerate(df.iterrows())}

        # Colori per owner
        owners = df['Owner'].unique()
        colors = plt.cm.Set3(range(len(owners)))
        owner_colors = dict(zip(owners, colors))

        # 1. Disegna Linee Dipendenze (prima delle barre così stanno "sotto")
        for i, (_, row) in enumerate(df.iterrows()):
            deps = row.get('Dependencies', [])
            for dep_id in deps:
                if dep_id in id_to_idx:
                    pred_idx = id_to_idx[dep_id]
                    p_row = df.iloc[pred_idx]
                    
                    # Coord X: fine predecessore -> inizio successore
                    x_start = mdates.date2num(p_row['Finish'])
                    x_end = mdates.date2num(row['Start'])
                    
                    # Disegna linea spezzata (verticale a x_start, orizzontale a y=i)
                    ax.plot([x_start, x_start, x_end], [pred_idx, i, i], 
                            color='#7f8c8d', linestyle='-', linewidth=0.8, alpha=0.5, zorder=1)
                    
                    # Aggiungi freccia di direzione
                    if x_end > x_start:
                        ax.annotate('', xy=(x_end, i), xytext=(x_end - 0.05, i),
                                    arrowprops=dict(arrowstyle="->", color='#7f8c8d', lw=0.8, alpha=0.5))

        # 2. Disegna barre task
        for idx, (_, row) in enumerate(df.iterrows()):
            start = mdates.date2num(row['Start'])
            end = mdates.date2num(row['Finish'])
            duration = end - start
            
            # Impedisce durate zero per visibilità (milestones)
            plot_width = max(duration, 0.1) 

            ax.barh(idx, plot_width, left=start, height=0.6,
                    color=owner_colors[row['Owner']],
                    edgecolor='black', linewidth=0.5,
                    zorder=2,
                    label=row['Owner'] if row['Owner'] not in ax.get_legend_handles_labels()[1] else "")

        # Etichette con CATEGORIA
        ax.set_yticks(range(len(df)))
        labels = [f"[{row['Category']}] {row['Task']}" for _, row in df.iterrows()]
        ax.set_yticklabels(labels, fontsize=7)
        ax.set_xlabel('Timeline', fontsize=12)
        ax.set_title(f'Gantt Chart - {self.product_name}', fontsize=14, fontweight='bold')

        # Formatta asse X
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')

        # Griglia
        ax.grid(True, axis='x', alpha=0.3, zorder=0)

        # Legenda
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax.legend(by_label.values(), by_label.keys(), loc='upper right', fontsize=8)

        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()

    def _create_pdf_statistics_page(self, pdf):
        """Crea pagina statistiche nel PDF."""
        fig = plt.figure(figsize=(11, 8.5))

        # Subplot 1: Task per Owner
        ax1 = plt.subplot(2, 2, 1)
        tasks_per_owner = self.statistics['tasks_per_owner']
        ax1.bar(tasks_per_owner.keys(), tasks_per_owner.values(), color='skyblue', edgecolor='black')
        ax1.set_title('Task per Owner', fontweight='bold')
        ax1.set_ylabel('Numero Task')
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')

        # Subplot 2: Durata Task
        ax2 = plt.subplot(2, 2, 2)
        durations = (self.df['Finish'] - self.df['Start']).dt.days
        ax2.hist(durations, bins=10, color='lightgreen', edgecolor='black')
        ax2.set_title('Distribuzione Durata Task', fontweight='bold')
        ax2.set_xlabel('Durata (giorni)')
        ax2.set_ylabel('Frequenza')
        ax2.axvline(durations.mean(), color='red', linestyle='--', label=f'Media: {durations.mean():.1f}')
        ax2.legend()

        # Subplot 3: Completamento
        ax3 = plt.subplot(2, 2, 3)
        completion = self.statistics['completion']
        colors_pie = ['#70AD47', '#D9D9D9']
        ax3.pie([completion, 100 - completion], labels=['Completato', 'Rimanente'],
                autopct='%1.1f%%', colors=colors_pie, startangle=90)
        ax3.set_title('Completamento Progetto', fontweight='bold')

        # Subplot 4: Task Critici
        ax4 = plt.subplot(2, 2, 4)
        critical = self.statistics['critical_tasks']
        late = self.statistics['late_tasks']
        normal = self.statistics['total_tasks'] - critical - late

        categories = ['Normali', 'Critici', 'In Ritardo']
        values = [normal, critical, late]
        colors_bar = ['green', 'orange', 'red']

        ax4.bar(categories, values, color=colors_bar, edgecolor='black')
        ax4.set_title('Stato Task', fontweight='bold')
        ax4.set_ylabel('Numero Task')

        plt.suptitle('Statistiche Progetto', fontsize=16, fontweight='bold', y=0.98)
        plt.tight_layout(rect=[0, 0, 1, 0.96])

        pdf.savefig(fig, bbox_inches='tight')
        plt.close()
    
    def _on_tab_changed(self, event=None):
        """Gestisce il cambio di tab nel Gantt gerarchico."""
        if not self.gantt_notebook:
            return
        
        # Determina quale tab è selezionato
        current_tab_index = self.gantt_notebook.index(self.gantt_notebook.select())
        
        if current_tab_index == 0:
            # Tab "Progetto Corrente"
            self.current_gantt_mode = 'current'
            self.log_status("📋 Selezionato: Progetto Corrente")
        elif current_tab_index == 1:
            # Tab "Vista Consolidata"
            self.current_gantt_mode = 'consolidated'
            self.log_status("🔗 Selezionato: Vista Consolidata")
        else:
            # Tab di un progetto figlio
            child_index = current_tab_index - 2  # Sottrai i primi 2 tabs
           # Trova il progetto figlio corrispondente
            children = [p for p in self.hierarchy_data['projects'] if not p['is_root']]
            if child_index < len(children):
                selected_child = children[child_index]
                self.current_gantt_mode = f"child_{selected_child['project_id']}"
                self.log_status(f"📄 Selezionato: {selected_child['project_name']}")
