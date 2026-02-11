# npi/npi_auto_notifications.py
"""
Sistema di notifiche automatiche giornaliere per task NPI in scadenza o scaduti.
Invia email professionali ai responsabili dei task e ai responsabili dei progetti.
"""

import logging
import json
import os
from datetime import datetime, timedelta, date
from pathlib import Path
import threading
import time
import tempfile
from typing import Dict, List, Tuple, Optional

from sqlalchemy import select, and_, or_, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, subqueryload

logger = logging.getLogger(__name__)


class NpiAutoNotificationService:
    """
    Servizio per notifiche automatiche giornaliere per task NPI.
    Controlla task in scadenza (domani) e scaduti, inviando email ai responsabili.
    """
    
    def __init__(self, npi_manager, config_path='npi_notifications_config.json'):
        self.npi_manager = npi_manager
        self.config_path = config_path
        self.config = self._load_config()
        self.running = False
        self.thread = None
        
    def _load_config(self) -> dict:
        """Carica la configurazione dal file JSON."""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                logger.info(f"Configurazione notifiche caricata da {self.config_path}")
                return config
            else:
                logger.warning(f"File configurazione {self.config_path} non trovato. Uso configurazione default.")
                return self._get_default_config()
        except Exception as e:
            logger.error(f"Errore caricamento configurazione: {e}")
            return self._get_default_config()
    
    def _get_default_config(self) -> dict:
        """Restituisce configurazione default."""
        return {
            "notification_settings": {
                "enabled": True,
                "check_time": "08:00",
                "email_sender_name": "NPI Project Management System",
                "include_logo": True,
                "logo_path": "logo.png"
            },
            "recipient_types": {
                "Interno": {"send_email": True},
                "Cliente": {"send_email": False},
                "Fornitore": {"send_email": False}
            },
            "notification_types": {
                "task_due_tomorrow": {"enabled": True, "days_before": 1},
                "task_overdue": {"enabled": True}
            }
        }
    
    def start(self):
        """Avvia il servizio di notifiche in background."""
        if self.running:
            logger.warning("Servizio notifiche già in esecuzione")
            return
        
        if not self.config['notification_settings']['enabled']:
            logger.info("Servizio notifiche disabilitato nella configurazione")
            return

        if self.npi_manager is None:
            logger.error("Servizio notifiche NPI non avviato: npi_manager non disponibile")
            return
        
        self.running = True
        self.thread = threading.Thread(target=self._run_service, daemon=True, name="NPINotificationService")
        self.thread.start()
        logger.info("Servizio notifiche automatiche NPI avviato")
    
    def stop(self):
        """Ferma il servizio di notifiche."""
        self.running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=5)
        logger.info("Servizio notifiche automatiche NPI fermato")
    
    def _run_service(self):
        """Loop principale del servizio (esegue in thread separato)."""
        logger.info("Thread servizio notifiche in esecuzione")
        
        # Esegui immediatamente al primo avvio
        self._check_and_send_notifications()
        
        # Poi controlla ogni ora se è il momento di inviare
        check_time_str = self.config['notification_settings']['check_time']
        target_hour, target_minute = map(int, check_time_str.split(':'))
        
        while self.running:
            try:
                now = datetime.now()
                
                # Controlla se è l'ora configurata
                if now.hour == target_hour and now.minute == target_minute:
                    logger.info(f"Ora configurata raggiunta ({check_time_str}), avvio controllo notifiche")
                    self._check_and_send_notifications()
                    # Dormi un minuto per evitare esecuzioni multiple
                    time.sleep(60)
                else:
                    # Controlla ogni 30 secondi
                    time.sleep(30)
                    
            except Exception as e:
                logger.error(f"Errore nel loop del servizio notifiche: {e}", exc_info=True)
                time.sleep(300)  # Attendi 5 minuti in caso di errore

    def _to_date(self, value):
        """Converte datetime/date in date; ritorna None se valore non valido."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if hasattr(value, "date"):
            try:
                return value.date()
            except Exception:
                return None
        return None
    
    def _check_and_send_notifications(self):
        """Controlla task e invia notifiche necessarie."""
        logger.info("=== INIZIO CONTROLLO NOTIFICHE NPI ===")
        
        try:
            # Ottieni tutti i progetti attivi
            progetti = self.npi_manager.get_all_npi_projects()
            
            if not progetti:
                logger.info("Nessun progetto NPI trovato")
                return
            
            total_sent = 0
            total_skipped = 0
            total_failed = 0
            
            for progetto in progetti:
                # Salta progetti chiusi
                if progetto.StatoProgetto == 'Chiuso':
                    continue
                
                sent, skipped, failed = self._process_project_notifications(progetto.ProgettoId)
                total_sent += sent
                total_skipped += skipped
                total_failed += failed

            # Notifica aggregata progetti in ritardo/potenziale ritardo
            sent, skipped, failed = self._check_project_delay_and_notify_owners()
            total_sent += sent
            total_skipped += skipped
            total_failed += failed
            
            logger.info(f"=== FINE CONTROLLO NOTIFICHE NPI ===")
            logger.info(f"Totale email inviate: {total_sent}")
            logger.info(f"Totale email saltate: {total_skipped}")
            logger.info(f"Totale email fallite: {total_failed}")
            
        except Exception as e:
            logger.error(f"Errore durante controllo notifiche: {e}", exc_info=True)

    def _check_project_delay_and_notify_owners(self) -> Tuple[int, int, int]:
        """
        Valuta ritardi progetto a livello aggregato e invia email:
        - TO: owner progetto
        - CC: destinatari da setting Sys_email_potential_npi_delay
        Allega un Excel con 2 tab: Overdue e Potential Delay.
        """
        sent = 0
        skipped = 0
        failed = 0

        try:
            if not self.npi_manager:
                logger.warning("NPI manager non disponibile per notifica ritardi progetto")
                return 0, 1, 0

            overdue_projects, potential_delay_projects = self._collect_project_delay_data()
            total_overdue = len(overdue_projects)
            total_potential = len(potential_delay_projects)

            if total_overdue == 0 and total_potential == 0:
                logger.info("Nessun progetto in ritardo/potenziale ritardo da notificare")
                return 0, 1, 0

            owners_map = {}
            for row in overdue_projects:
                owner_email = row.get('OwnerEmail')
                if not owner_email:
                    continue
                owners_map.setdefault(owner_email, {"owner_name": row.get('OwnerName') or "Project Owner",
                                                    "overdue": [], "potential": []})
                owners_map[owner_email]["overdue"].append(row)

            for row in potential_delay_projects:
                owner_email = row.get('OwnerEmail')
                if not owner_email:
                    continue
                owners_map.setdefault(owner_email, {"owner_name": row.get('OwnerName') or "Project Owner",
                                                    "overdue": [], "potential": []})
                owners_map[owner_email]["potential"].append(row)

            if not owners_map:
                logger.warning("Nessun owner con email valida per notifica ritardi progetto")
                return 0, 1, 0

            cc_recipients = self._get_setting_email_list('Sys_email_potential_npi_delay')
            report_path = self._build_project_delay_excel_report(overdue_projects, potential_delay_projects)
            self._ensure_project_delay_email_log_table()

            try:
                for owner_email, owner_payload in owners_map.items():
                    if self._was_project_delay_email_sent_today(owner_email):
                        logger.info("Email ritardi progetto già inviata oggi a %s, skip", owner_email)
                        skipped += 1
                        continue

                    owner_cc = [e for e in cc_recipients if e.lower() != owner_email.lower()]
                    result = self._send_project_delay_email(
                        owner_email=owner_email,
                        owner_name=owner_payload["owner_name"],
                        owner_overdue=owner_payload["overdue"],
                        owner_potential=owner_payload["potential"],
                        total_overdue=total_overdue,
                        total_potential=total_potential,
                        cc_recipients=owner_cc,
                        attachment_path=report_path
                    )
                    if result:
                        sent += 1
                        self._mark_project_delay_email_sent(owner_email, total_overdue, total_potential)
                    else:
                        failed += 1
            finally:
                try:
                    if report_path and os.path.exists(report_path):
                        os.remove(report_path)
                except Exception:
                    pass

            return sent, skipped, failed
        except Exception as e:
            logger.error(f"Errore notifica ritardi progetto NPI: {e}", exc_info=True)
            return sent, skipped, failed + 1

    def _collect_project_delay_data(self) -> Tuple[List[dict], List[dict]]:
        """Raccoglie i progetti in ritardo e in potenziale ritardo."""
        from .data_models import ProgettoNPI, WaveNPI, TaskProdotto

        session = self.npi_manager._get_session()
        overdue_projects: List[dict] = []
        potential_delay_projects: List[dict] = []
        today = date.today()

        try:
            projects = session.scalars(
                select(ProgettoNPI)
                .options(
                    joinedload(ProgettoNPI.prodotto),
                    joinedload(ProgettoNPI.owner),
                    subqueryload(ProgettoNPI.waves).subqueryload(WaveNPI.tasks)
                    .options(joinedload(TaskProdotto.task_catalogo), joinedload(TaskProdotto.owner))
                )
                .where(ProgettoNPI.StatoProgetto != 'Chiuso')
            ).all()

            for project in projects:
                deadline = self._to_date(getattr(project, 'ScadenzaProgetto', None))
                tasks = []
                for wave in (project.waves or []):
                    tasks.extend(wave.tasks or [])

                incomplete_tasks = [t for t in tasks if (t.Stato or '').strip().lower() != 'completato']
                if not incomplete_tasks:
                    continue

                due_dates = []
                overdue_task_count = 0
                for task in incomplete_tasks:
                    if task.DataScadenza:
                        task_due = self._to_date(task.DataScadenza)
                        if not task_due:
                            continue
                        due_dates.append(task_due)
                        if task_due < today:
                            overdue_task_count += 1

                latest_task_due = max(due_dates) if due_dates else None
                product_name = project.prodotto.NomeProdotto if project.prodotto else "N/A"
                product_code = project.prodotto.CodiceProdotto if project.prodotto else "N/A"
                owner_name = project.owner.Nome if project.owner else ""
                owner_email = project.owner.Email if project.owner else ""

                base_row = {
                    "ProjectId": project.ProgettoId,
                    "ProjectName": project.NomeProgetto or product_name,
                    "ProductCode": product_code,
                    "OwnerName": owner_name,
                    "OwnerEmail": owner_email,
                    "ProjectDeadline": deadline,
                    "RemainingTasks": len(incomplete_tasks),
                    "OverdueOpenTasks": overdue_task_count,
                    "LatestRemainingTaskDue": latest_task_due
                }

                if deadline and deadline < today:
                    row = dict(base_row)
                    row["DelayDays"] = (today - deadline).days
                    row["RiskReason"] = "Project deadline is already expired"
                    overdue_projects.append(row)
                    continue

                is_potential_delay = False
                reason_parts = []
                delay_days = 0

                if deadline and latest_task_due and latest_task_due > deadline:
                    is_potential_delay = True
                    delay_days = (latest_task_due - deadline).days
                    reason_parts.append("Latest open-task due date exceeds project deadline")

                if deadline and overdue_task_count > 0:
                    is_potential_delay = True
                    reason_parts.append("There are open tasks already overdue")

                if is_potential_delay:
                    row = dict(base_row)
                    row["DelayDays"] = delay_days
                    row["RiskReason"] = "; ".join(reason_parts) if reason_parts else "Potential schedule slippage detected"
                    potential_delay_projects.append(row)

            return overdue_projects, potential_delay_projects
        finally:
            session.close()

    def _get_setting_email_list(self, setting_key: str) -> List[str]:
        """Legge una lista email da dbo.settings.VALUE separata da ; o ,."""
        session = self.npi_manager._get_session()
        emails: List[str] = []
        try:
            rows = session.execute(
                text("SELECT [VALUE] FROM [dbo].[settings] WHERE atribute = :attr"),
                {"attr": setting_key}
            ).all()
            for row in rows:
                raw = row[0]
                if not raw:
                    continue
                for token in str(raw).replace(';', ',').split(','):
                    em = token.strip()
                    if em and '@' in em:
                        emails.append(em)
            # Dedup preservando ordine
            deduped = []
            seen = set()
            for em in emails:
                key = em.lower()
                if key in seen:
                    continue
                deduped.append(em)
                seen.add(key)
            return deduped
        except Exception as e:
            logger.error(f"Errore lettura setting email {setting_key}: {e}", exc_info=True)
            return []
        finally:
            session.close()

    def _build_notification_cc_list(self, project_owner_email: Optional[str], to_email: Optional[str] = None) -> List[str]:
        """
        Costruisce la lista CC per notifiche:
        - owner progetto
        - email da setting Sys_email_general_napi
        Deduplica e rimuove eventuale duplicato con TO.
        """
        cc_raw: List[str] = []
        if project_owner_email and '@' in project_owner_email:
            cc_raw.append(project_owner_email.strip())
        cc_raw.extend(self._get_setting_email_list('Sys_email_general_napi'))

        to_key = (to_email or '').strip().lower()
        cc_clean: List[str] = []
        seen = set()
        for email in cc_raw:
            key = (email or '').strip().lower()
            if not key or key == to_key or key in seen:
                continue
            seen.add(key)
            cc_clean.append(email.strip())
        return cc_clean

    def _ensure_project_delay_email_log_table(self) -> None:
        """Crea la tabella log invii (se assente) per garantire una sola email al giorno per owner."""
        session = self.npi_manager._get_session()
        try:
            session.execute(text("""
                IF OBJECT_ID(N'[dbo].[NpiProjectDelayEmailLog]', N'U') IS NULL
                BEGIN
                    CREATE TABLE [dbo].[NpiProjectDelayEmailLog](
                        [Id] INT IDENTITY(1,1) PRIMARY KEY,
                        [NotificationDate] DATE NOT NULL,
                        [RecipientEmail] NVARCHAR(255) NOT NULL,
                        [OverdueCount] INT NULL,
                        [PotentialCount] INT NULL,
                        [SentDateTime] DATETIME NOT NULL DEFAULT(GETDATE())
                    );
                    CREATE UNIQUE INDEX [UX_NpiProjectDelayEmailLog_DateRecipient]
                        ON [dbo].[NpiProjectDelayEmailLog]([NotificationDate], [RecipientEmail]);
                END
            """))
            session.commit()
        except Exception as e:
            session.rollback()
            logger.error(f"Errore creazione tabella NpiProjectDelayEmailLog: {e}", exc_info=True)
        finally:
            session.close()

    def _was_project_delay_email_sent_today(self, owner_email: str) -> bool:
        """Ritorna True se oggi è già stata inviata email ritardi a questo owner."""
        session = self.npi_manager._get_session()
        try:
            today = date.today()
            row = session.execute(
                text("""
                    SELECT TOP 1 1
                    FROM [dbo].[NpiProjectDelayEmailLog]
                    WHERE [NotificationDate] = :today
                      AND [RecipientEmail] = :email
                """),
                {"today": today, "email": owner_email}
            ).first()
            return row is not None
        except Exception as e:
            logger.error(f"Errore verifica log invio NPI project delay per {owner_email}: {e}", exc_info=True)
            return False
        finally:
            session.close()

    def _mark_project_delay_email_sent(self, owner_email: str, overdue_count: int, potential_count: int) -> None:
        """Registra l'invio email ritardi progetto per owner/data odierna."""
        session = self.npi_manager._get_session()
        try:
            today = date.today()
            session.execute(
                text("""
                    INSERT INTO [dbo].[NpiProjectDelayEmailLog]
                        ([NotificationDate], [RecipientEmail], [OverdueCount], [PotentialCount])
                    VALUES
                        (:today, :email, :overdue, :potential)
                """),
                {
                    "today": today,
                    "email": owner_email,
                    "overdue": int(overdue_count or 0),
                    "potential": int(potential_count or 0)
                }
            )
            session.commit()
        except Exception as e:
            session.rollback()
            logger.error(f"Errore registrazione log invio NPI project delay per {owner_email}: {e}", exc_info=True)
        finally:
            session.close()

    def _resolve_logo_path(self) -> Optional[str]:
        """Trova il logo da allegare inline all'email."""
        candidates = [
            os.path.join(os.getcwd(), "Logo.png"),
            os.path.join(os.getcwd(), "logo.png"),
            os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Logo.png")),
            os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "logo.png")),
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        return None

    def _build_project_delay_excel_report(self, overdue_projects: List[dict], potential_delay_projects: List[dict]) -> str:
        """Crea il file Excel con 2 tab: progetti in ritardo e progetti in potenziale ritardo."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception as e:
            raise RuntimeError(f"openpyxl non disponibile per export report ritardi progetto: {e}") from e

        wb = Workbook()
        ws_overdue = wb.active
        ws_overdue.title = "Overdue Projects"
        ws_potential = wb.create_sheet("Potential Delay Projects")

        headers = [
            "Project ID", "Project Name", "Product Code", "Owner", "Owner Email",
            "Project Deadline", "Remaining Tasks", "Open Tasks Overdue",
            "Latest Remaining Task Due", "Delay Days", "Risk Reason"
        ]

        def _write_sheet(ws, rows, header_fill_color):
            ws.append(headers)
            for h_cell in ws[1]:
                h_cell.font = Font(bold=True, color="FFFFFF")
                h_cell.fill = PatternFill("solid", fgColor=header_fill_color)
            for row in rows:
                ws.append([
                    row.get("ProjectId"),
                    row.get("ProjectName"),
                    row.get("ProductCode"),
                    row.get("OwnerName"),
                    row.get("OwnerEmail"),
                    row.get("ProjectDeadline").strftime("%Y-%m-%d") if row.get("ProjectDeadline") else "",
                    row.get("RemainingTasks"),
                    row.get("OverdueOpenTasks"),
                    row.get("LatestRemainingTaskDue").strftime("%Y-%m-%d") if row.get("LatestRemainingTaskDue") else "",
                    row.get("DelayDays"),
                    row.get("RiskReason"),
                ])
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        _write_sheet(ws_overdue, overdue_projects, "C00000")
        _write_sheet(ws_potential, potential_delay_projects, "CC8400")

        fd, report_path = tempfile.mkstemp(prefix="npi_project_delay_", suffix=".xlsx")
        os.close(fd)
        wb.save(report_path)
        return report_path

    def _send_project_delay_email(
        self,
        owner_email: str,
        owner_name: str,
        owner_overdue: List[dict],
        owner_potential: List[dict],
        total_overdue: int,
        total_potential: int,
        cc_recipients: List[str],
        attachment_path: str
    ) -> bool:
        """Invia email professionale di alert ritardi progetto NPI."""
        from email_connector import EmailSender

        if not owner_email:
            return False

        logo_path = self._resolve_logo_path()
        logo_html = ""
        attachments: List = [attachment_path]
        if logo_path:
            attachments = [('inline', logo_path, 'company_logo')] + attachments
            logo_html = '<img src="cid:company_logo" alt="Company Logo" style="max-width: 180px; height: auto;"/>'

        def _rows_html(rows: List[dict], empty_msg: str) -> str:
            if not rows:
                return f"<p style='margin:8px 0;color:#666;'>{empty_msg}</p>"

            lines = []
            for r in rows:
                delay_days = r.get("DelayDays") or 0
                deadline = r.get("ProjectDeadline")
                latest_due = r.get("LatestRemainingTaskDue")
                deadline_txt = deadline.strftime("%Y-%m-%d") if deadline else "N/A"
                latest_due_txt = latest_due.strftime("%Y-%m-%d") if latest_due else "N/A"
                lines.append(
                    f"<tr>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{r.get('ProjectId')}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{r.get('ProjectName')}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{r.get('ProductCode') or ''}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{deadline_txt}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{r.get('RemainingTasks') or 0}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{latest_due_txt}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;text-align:center;font-weight:bold;color:#b30000;'>{delay_days}</td>"
                    f"</tr>"
                )
            return (
                "<table style='border-collapse:collapse;width:100%;margin:10px 0;'>"
                "<thead><tr style='background:#1F4E78;color:#fff;'>"
                "<th style='padding:8px;border:1px solid #ddd;'>Project ID</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Project</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Product Code</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Project Deadline</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Remaining Tasks</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Latest Open Task Due</th>"
                "<th style='padding:8px;border:1px solid #ddd;'>Delay (days)</th>"
                "</tr></thead><tbody>"
                + "".join(lines) +
                "</tbody></table>"
            )

        owner_overdue_count = len(owner_overdue)
        owner_potential_count = len(owner_potential)

        html_body = f"""
        <html>
        <body style="font-family:Segoe UI,Arial,sans-serif;color:#1f2937;background:#f5f7fb;padding:20px;">
            <div style="max-width:980px;margin:0 auto;background:#ffffff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
                <div style="background:#0B3A66;color:#fff;padding:20px 24px;">
                    {logo_html}
                    <h2 style="margin:12px 0 4px 0;">NPI Project Schedule Risk Alert</h2>
                    <p style="margin:0;opacity:0.92;">Automatic monitoring report for project deadlines and open-task schedule risk.</p>
                </div>
                <div style="padding:22px 24px;">
                    <p>Dear {owner_name or 'Project Owner'},</p>
                    <p>
                        The monitoring engine detected schedule risks in NPI projects.
                        <strong>{total_overdue}</strong> projects are overdue and
                        <strong>{total_potential}</strong> projects are in potential delay.
                    </p>
                    <div style="background:#EEF4FB;border-left:4px solid #1F4E78;padding:12px 14px;margin:14px 0;">
                        <p style="margin:0;"><strong>Your scope:</strong> {owner_overdue_count} overdue project(s), {owner_potential_count} potential-delay project(s).</p>
                    </div>

                    <h3 style="margin:18px 0 8px 0;color:#8B0000;">Projects Overdue (your ownership)</h3>
                    {_rows_html(owner_overdue, "No overdue projects found for your ownership.")}

                    <h3 style="margin:18px 0 8px 0;color:#9A6B00;">Projects in Potential Delay (your ownership)</h3>
                    {_rows_html(owner_potential, "No potential-delay projects found for your ownership.")}

                    <p style="margin-top:18px;">
                        Please review the attached Excel report with the complete project list:
                        <strong>Overdue Projects</strong> and <strong>Potential Delay Projects</strong>.
                    </p>
                    <p style="color:#6b7280;font-size:12px;margin-top:24px;">
                        This is an automated notification from the NPI Project Management System. Please do not reply to this email.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        subject = (
            f"NPI Delay Alert - {total_overdue} overdue / {total_potential} potential delay projects"
        )

        try:
            sender = EmailSender()
            sender.save_credentials("Accounting@Eutron.it", "9jHgFhSs7Vf+")
            sender.send_email(
                to_email=owner_email,
                subject=subject,
                body=html_body,
                is_html=True,
                attachments=attachments,
                cc_emails=cc_recipients
            )
            logger.info(
                "Email ritardi progetto inviata a %s (CC=%d, overdue=%d, potential=%d)",
                owner_email, len(cc_recipients), owner_overdue_count, owner_potential_count
            )
            return True
        except Exception as e:
            logger.error(f"Errore invio email ritardi progetto a {owner_email}: {e}", exc_info=True)
            return False
    
    def _process_project_notifications(self, project_id: int) -> Tuple[int, int, int]:
        """Processa notifiche per un singolo progetto."""
        sent = 0
        skipped = 0
        failed = 0
        
        try:
            from .data_models import TaskProdotto, WaveNPI, ProgettoNPI
            
            # Ottieni tutti i task del progetto
            session = self.npi_manager._get_session()
            try:
                progetto = session.scalars(
                    select(ProgettoNPI)
                    .options(
                        joinedload(ProgettoNPI.owner),
                        joinedload(ProgettoNPI.prodotto)
                    )
                    .where(ProgettoNPI.ProgettoId == project_id)
                ).first()

                if not progetto:
                    logger.warning("Progetto NPI %s non trovato durante processamento notifiche", project_id)
                    return sent, skipped + 1, failed

                tasks = session.scalars(
                    select(TaskProdotto)
                    .join(WaveNPI)
                    .where(
                        and_(
                            WaveNPI.ProgettoID == project_id,
                            TaskProdotto.OwnerID.isnot(None),
                            TaskProdotto.Stato != 'Completato'
                        )
                    )
                    .options(
                        joinedload(TaskProdotto.owner),
                        joinedload(TaskProdotto.task_catalogo),
                        joinedload(TaskProdotto.wave).joinedload(WaveNPI.progetto).joinedload(ProgettoNPI.owner),
                        subqueryload(TaskProdotto.dependencies)
                    )
                ).all()
                
                for task in tasks:
                    s, sk, f = self._process_task_notifications(task, progetto, session)
                    sent += s
                    skipped += sk
                    failed += f
                
                # NUOVO: Analisi rischio progetto e notifica Project Owner
                s, sk, f = self._assess_project_risk_and_notify(progetto, tasks, session)
                sent += s
                skipped += sk
                failed += f
            finally:
                session.close()
                    
        except Exception as e:
            logger.error(f"Errore processamento progetto {project_id}: {e}", exc_info=True)
            failed += 1
        
        return sent, skipped, failed
    
    def _assess_project_risk_and_notify(self, progetto, tasks, session) -> Tuple[int, int, int]:
        """
        Analizza il rischio del progetto basandosi su task in ritardo.
        Invia email al Project Owner se il progetto è a rischio.
        
        Logica intelligente:
        - Conta task scaduti
        - Verifica task scaduti nel critical path (con dipendenti)
        - Calcola buffer temporale vs giorni necessari
        - Determina livello di rischio (Low/Medium/High/Critical)
        """
        from .data_models_notification import NpiTaskNotification
        from .data_models import TaskDependency, TaskProdotto
        
        try:
            if not progetto.owner or not progetto.owner.Email:
                return 0, 0, 0
            
            # Verifica tipo soggetto
            recipient_category = progetto.owner.Tipo or 'Interno'
            recipient_config = self.config['recipient_types'].get(recipient_category, {"send_email": False})
            if not recipient_config['send_email']:
                return 0, 1, 0
            
            today = date.today()
            
            # 1. Identifica task in ritardo
            overdue_tasks = []
            for task in tasks:
                if task.DataScadenza:
                    due_date = self._to_date(task.DataScadenza)
                    if not due_date:
                        continue
                    if due_date < today and task.Stato != 'Completato':
                        overdue_tasks.append(task)
            
            # Se nessun task in ritardo, progetto ok
            if not overdue_tasks:
                return 0, 0, 0
            
            # 2. Analizza impatto task in ritardo
            critical_path_tasks = []  # Task in ritardo che bloccano altri task
            blocking_info = {}  # task_id -> lista di task bloccati
            
            for task in overdue_tasks:
                # Trova task che dipendono da questo task in ritardo
                dependent_tasks = session.scalars(
                    select(TaskProdotto)
                    .join(TaskDependency, TaskDependency.TaskProdottoID == TaskProdotto.TaskProdottoID)
                    .where(TaskDependency.DependsOnTaskProdottoID == task.TaskProdottoID)
                    .options(joinedload(TaskProdotto.task_catalogo), joinedload(TaskProdotto.owner))
                ).all()
                
                if dependent_tasks:
                    critical_path_tasks.append(task)
                    blocking_info[task.TaskProdottoID] = dependent_tasks
            
            # 3. Calcola buffer temporale
            project_end = progetto.ScadenzaProgetto
            if not project_end:
                buffer_days = None
            else:
                end_date = self._to_date(project_end)
                if not end_date:
                    buffer_days = None
                else:
                    buffer_days = (end_date - today).days
            
            # 4. Determina livello di rischio
            risk_level, risk_color, risk_icon = self._calculate_risk_level(
                len(overdue_tasks),
                len(critical_path_tasks),
                buffer_days,
                len(tasks)
            )
            
            # 5. Verifica se già inviata oggi (evita duplicati)
            notification_type = f'ProjectRisk_{risk_level}'
            existing = session.scalars(
                select(NpiTaskNotification).where(
                    and_(
                        NpiTaskNotification.ProgettoID == progetto.ProgettoId,
                        NpiTaskNotification.RecipientSoggettoID == progetto.owner.SoggettoId,
                        NpiTaskNotification.NotificationType == notification_type,
                        NpiTaskNotification.NotificationDate == today
                    )
                )
            ).first()
            
            if existing:
                logger.debug(f"Risk notification già inviata oggi al Project Owner per progetto {progetto.ProgettoId}")
                return 0, 1, 0
            
            # 6. Genera e invia email di rischio progetto
            email_html = self._generate_project_risk_email(
                progetto, overdue_tasks, critical_path_tasks, blocking_info,
                risk_level, risk_color, risk_icon, buffer_days
            )
            
            from utils import send_email
            
            subject = f"🚨 NPI Project Risk Alert: {progetto.prodotto.NomeProdotto if progetto.prodotto else 'Project'} - {risk_level.upper()} RISK"
            
            cc_list = self._build_notification_cc_list(
                project_owner_email=(progetto.owner.Email if progetto.owner else None),
                to_email=(progetto.owner.Email if progetto.owner else None)
            )
            send_email(
                recipients=[progetto.owner.Email],
                subject=subject,
                body=email_html,
                is_html=True,
                cc_emails=cc_list
            )
            
            # 7. Registra notifica
            notification = NpiTaskNotification(
                TaskProdottoID=None,  # Non specifico di un task
                ProgettoID=progetto.ProgettoId,
                RecipientSoggettoID=progetto.owner.SoggettoId,
                NotificationType=notification_type,
                NotificationDate=today,
                RecipientEmail=progetto.owner.Email,
                RecipientName=progetto.owner.Nome,
                RecipientType='ProjectOwner',
                DeliveryStatus='Sent'
            )
            session.add(notification)
            session.commit()
            
            logger.info(f"Project Risk Alert ({risk_level}) inviato a {progetto.owner.Nome} per progetto {progetto.ProgettoId}")
            return 1, 0, 0
            
        except Exception as e:
            logger.error(f"Errore valutazione rischio progetto: {e}", exc_info=True)
            return 0, 0, 1
    
    def _calculate_risk_level(self, overdue_count, critical_path_count, buffer_days, total_tasks) -> Tuple[str, str, str]:
        """
        Calcola il livello di rischio del progetto basandosi su metriche.
        
        Returns:
            (risk_level, risk_color, risk_icon)
        """
        # Fattori di rischio
        overdue_ratio = overdue_count / max(total_tasks, 1)
        
        # CRITICAL: Task nel critical path in ritardo + poco buffer
        if critical_path_count > 0 and (buffer_days is None or buffer_days <= 7):
            return 'Critical', '#8B0000', '🔴'  # Dark Red
        
        # HIGH: Molti task in ritardo o critical path
        if overdue_ratio > 0.3 or critical_path_count >= 3:
            return 'High', '#DC143C', '🚨'  # Crimson
        
        # MEDIUM: Alcuni task in ritardo
        if overdue_ratio > 0.15 or critical_path_count >= 1:
            return 'Medium', '#FFA500', '⚠️'  # Orange
        
        # LOW: Pochi task in ritardo, molto buffer
        return 'Low', '#FFD700', '⚡'  # Gold
    
    def _process_task_notifications(self, task, progetto, session) -> Tuple[int, int, int]:
        """Processa notifiche per un singolo task."""
        sent = 0
        skipped = 0
        failed = 0
        
        if not task.DataScadenza:
            return sent, skipped, failed
        
        today = date.today()
        due_date = self._to_date(task.DataScadenza)
        if not due_date:
            return sent, skipped + 1, failed
        
        # Determina il tipo di notifica
        notification_type = None
        
        if self.config['notification_types']['task_due_tomorrow']['enabled']:
            days_before = self.config['notification_types']['task_due_tomorrow']['days_before']
            if due_date == today + timedelta(days=days_before):
                notification_type = 'TaskDueTomorrow'
        
        if self.config['notification_types']['task_overdue']['enabled']:
            if due_date < today:
                notification_type = 'TaskOverdue'
        
        if not notification_type:
            return sent, skipped, failed
        
        # Invia email al task owner
        if task.owner:
            s, sk, f = self._send_notification_email(
                task, progetto, task.owner, 'TaskOwner', notification_type, session
            )
            sent += s
            skipped += sk
            failed += f
        
        # Invia email al project owner se diverso dal task owner
        if progetto.owner and progetto.owner.SoggettoId != (task.OwnerID if task.owner else None):
            s, sk, f = self._send_notification_email(
                task, progetto, progetto.owner, 'ProjectOwner', notification_type, session
            )
            sent += s
            skipped += sk
            failed += f
        
        return sent, skipped, failed
    
    def _send_notification_email(
        self, task, progetto, recipient, recipient_type, notification_type, session
    ) -> Tuple[int, int, int]:
        """Invia una email di notifica ed registra nel database."""
        from .data_models_notification import NpiTaskNotification
        
        # Verifica se il tipo di soggetto può ricevere email
        recipient_category = recipient.Tipo or 'Interno'
        recipient_config = self.config['recipient_types'].get(recipient_category, {"send_email": False})
        
        if not recipient_config['send_email']:
            logger.debug(f"Email saltata per {recipient.Nome} (tipo: {recipient_category})")
            return 0, 1, 0
        
        if not recipient.Email:
            logger.warning(f"Nessuna email configurata per {recipient.Nome}")
            return 0, 1, 0
        
        # Controllo duplicati
        today = date.today()
        existing = session.scalars(
            select(NpiTaskNotification).where(
                and_(
                    NpiTaskNotification.TaskProdottoID == task.TaskProdottoID,
                    NpiTaskNotification.RecipientSoggettoID == recipient.SoggettoId,
                    NpiTaskNotification.NotificationType == notification_type,
                    NpiTaskNotification.NotificationDate == today
                )
            )
        ).first()
        
        if existing:
            logger.debug(f"Notifica già inviata oggi a {recipient.Nome} per task {task.TaskProdottoID}")
            return 0, 1, 0
        
        # Genera email HTML
        email_html = self._generate_notification_email(task, progetto, recipient, recipient_type, notification_type)
        
        # Invia email
        try:
            from utils import send_email
            
            subject = f"NPI Task Alert: {task.task_catalogo.NomeTask if task.task_catalogo else 'Task'}"
            if notification_type == 'TaskDueTomorrow':
                subject += " - Due Tomorrow"
            else:
                subject += " - OVERDUE"
            
            cc_list = self._build_notification_cc_list(
                project_owner_email=(progetto.owner.Email if progetto and progetto.owner else None),
                to_email=recipient.Email
            )
            send_email(
                recipients=[recipient.Email],
                subject=subject,
                body=email_html,
                is_html=True,
                cc_emails=cc_list
            )
            
            # Registra notifica
            notification = NpiTaskNotification(
                TaskProdottoID=task.TaskProdottoID,
                ProgettoID=progetto.ProgettoId,
                RecipientSoggettoID=recipient.SoggettoId,
                NotificationType=notification_type,
                NotificationDate=today,
                RecipientEmail=recipient.Email,
                RecipientName=recipient.Nome,
                RecipientType=recipient_type,
                DeliveryStatus='Sent'
            )
            session.add(notification)
            session.commit()
            
            logger.info(f"Notifica {notification_type} inviata a {recipient.Nome} per task {task.TaskProdottoID}")
            return 1, 0, 0
            
        except Exception as e:
            logger.error(f"Errore invio email a {recipient.Nome}: {e}", exc_info=True)
            
            # Registra errore
            try:
                notification = NpiTaskNotification(
                    TaskProdottoID=task.TaskProdottoID,
                    ProgettoID=progetto.ProgettoId,
                    RecipientSoggettoID=recipient.SoggettoId,
                    NotificationType=notification_type,
                    NotificationDate=today,
                    RecipientEmail=recipient.Email,
                    RecipientName=recipient.Nome,
                    RecipientType=recipient_type,
                    DeliveryStatus='Failed',
                    ErrorMessage=str(e)
                )
                session.add(notification)
                session.commit()
            except:
                pass
            
            return 0, 0, 1
    
    def _generate_notification_email(
        self, task, progetto, recipient, recipient_type, notification_type
    ) -> str:
        """Genera HTML email professionale con logo."""
        
        # Determina il titolo e colore in base al tipo
        if notification_type == 'TaskDueTomorrow':
            alert_title = "⚠️ TASK DUE TOMORROW"
            alert_color = "#FFA500"  # Orange
            message = "This task is due tomorrow"
        else:  # TaskOverdue
            alert_title = "🚨 TASK OVERDUE"
            alert_color = "#DC143C"  # Crimson
            message = "This task is overdue and requires immediate attention"
        
        # Dati task
        task_name = task.task_catalogo.NomeTask if task.task_catalogo else "N/A"
        task_owner = task.owner.Nome if task.owner else "Unassigned"
        task_due_date = task.DataScadenza.strftime('%d/%m/%Y') if task.DataScadenza else "N/A"
        task_start_date = task.DataInizio.strftime('%d/%m/%Y') if task.DataInizio else "N/A"
        task_status = task.Stato or "Not Started"
        
        # Dati progetto
        project_name = progetto.prodotto.NomeProdotto if progetto.prodotto else "N/A"
        project_code = progetto.prodotto.CodiceProdotto if progetto.prodotto else "N/A"
        project_owner = progetto.owner.Nome if progetto.owner else "N/A"
        
        # Dipendenze bloccate
        blocked_tasks_html = ""
        dependencies = self.npi_manager.get_task_dependencies(task.TaskProdottoID)
        if dependencies:
            blocked_list = []
            session = self.npi_manager._get_session()
            try:
                from .data_models import TaskProdotto, TaskDependency
                
                # Trova task che dipendono da questo
                dependent_tasks = session.scalars(
                    select(TaskProdotto)
                    .join(TaskDependency, TaskDependency.TaskProdottoID == TaskProdotto.TaskProdottoID)
                    .where(TaskDependency.DependsOnTaskProdottoID == task.TaskProdottoID)
                    .options(joinedload(TaskProdotto.task_catalogo), joinedload(TaskProdotto.owner))
                ).all()
                
                for dep_task in dependent_tasks:
                    dep_name = dep_task.task_catalogo.NomeTask if dep_task.task_catalogo else "Task"
                    dep_owner = dep_task.owner.Nome if dep_task.owner else "Unassigned"
                    blocked_list.append(f"<li><strong>{dep_name}</strong> (Owner: {dep_owner})</li>")
            finally:
                session.close()
            
            if blocked_list:
                blocked_tasks_html = f"""
                <div style="background: #FFF3CD; border-left: 4px solid #FFC107; padding: 15px; margin: 20px 0; border-radius: 4px;">
                    <h4 style="color: #856404; margin: 0 0 10px 0;">⚠️ BLOCKED TASKS</h4>
                    <p style="margin: 0 0 10px 0; color: #856404;">The following tasks depend on this task and are currently blocked:</p>
                    <ul style="margin: 0; padding-left: 20px; color: #856404;">
                        {''.join(blocked_list)}
                    </ul>
                </div>
                """
        
        # Logo (base64 embedded o path)
        logo_html = ""
        if self.config['notification_settings']['include_logo']:
            logo_path = self.config['notification_settings']['logo_path']
            if os.path.exists(logo_path):
                # Per semplicità, usiamo un path relativo
                # In produzione, considerare di embeddare come base64
                logo_html = f'<img src="cid:logo" style="max-width: 150px; height: auto;" alt="Company Logo">'
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                .container {{ max-width: 700px; margin: 0 auto; padding: 20px; background-color: #f9f9f9; }}
                .header {{ background: linear-gradient(135deg, #0078d4 0%, #005a9e 100%); color: white; padding: 30px 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                .header h1 {{ margin: 10px 0 0 0; font-size: 24px; }}
                .alert-box {{ background: {alert_color}; color: white; padding: 20px; text-align: center; font-size: 20px; font-weight: bold; }}
                .content {{ background: white; padding: 30px; border-radius: 0 0 8px 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .section {{ margin: 25px 0; padding: 20px; background: #f8f9fa; border-left: 4px solid #0078d4; border-radius: 4px; }}
                .section h2 {{ color: #0078d4; margin: 0 0 15px 0; font-size: 18px; border-bottom: 2px solid #0078d4; padding-bottom: 10px; }}
                .info-row {{ margin: 8px 0; }}
                .label {{ font-weight: bold; color: #555; display: inline-block; min-width: 120px; }}
                .value {{ color: #333; }}
                .footer {{ background: #f1f1f1; padding: 20px; margin-top: 20px; border-top: 3px solid #0078d4; border-radius: 4px; text-align: center; }}
                .footer p {{ margin: 5px 0; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    {logo_html}
                    <h1>NPI Project Management System</h1>
                    <p>Automated Task Notification</p>
                </div>
                
                <div class="alert-box">
                    {alert_title}
                </div>
                
                <div class="content">
                    <p style="font-size: 16px; margin-bottom: 20px;">Dear <strong>{recipient.Nome}</strong>,</p>
                    <p>{message}:</p>
                    
                    <div class="section">
                        <h2>📋 PROJECT DETAILS</h2>
                        <div class="info-row"><span class="label">Project Name:</span> <span class="value">{project_name}</span></div>
                        <div class="info-row"><span class="label">Product Code:</span> <span class="value">{project_code}</span></div>
                        <div class="info-row"><span class="label">Project Owner:</span> <span class="value">{project_owner}</span></div>
                    </div>
                    
                    <div class="section">
                        <h2>⚠️ TASK DETAILS</h2>
                        <div class="info-row"><span class="label">Task Name:</span> <span class="value" style="font-weight: bold; color: {alert_color};">{task_name}</span></div>
                        <div class="info-row"><span class="label">Task Owner:</span> <span class="value">{task_owner}</span></div>
                        <div class="info-row"><span class="label">Start Date:</span> <span class="value">{task_start_date}</span></div>
                        <div class="info-row"><span class="label">Due Date:</span> <span class="value" style="font-weight: bold; color: {alert_color};">{task_due_date}</span></div>
                        <div class="info-row"><span class="label">Status:</span> <span class="value">{task_status}</span></div>
                    </div>
                    
                    {blocked_tasks_html}
                    
                    <div style="background: #e7f3ff; border-left: 4px solid #0078d4; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <h4 style="color: #0078d4; margin: 0 0 10px 0;">📌 ACTION REQUIRED</h4>
                        <ul style="margin: 0; padding-left: 20px; color: #333;">
                            <li>Please review and update the task status in the NPI system</li>
                            <li>Coordinate with team members if dependencies exist</li>
                            <li>Contact the project owner if you need assistance</li>
                        </ul>
                    </div>
                </div>
                
                <div class="footer">
                    <p><strong>This is an automated notification from the NPI Project Management System.</strong></p>
                    <p>Please do not reply to this email.</p>
                    <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generate_project_risk_email(
        self, progetto, overdue_tasks, critical_path_tasks, blocking_info,
        risk_level, risk_color, risk_icon, buffer_days
    ) -> str:
        """Genera email HTML di Project Risk Alert per il Project Owner."""
        
        # Dati progetto
        project_name = progetto.prodotto.NomeProdotto if progetto.prodotto else "N/A"
        project_code = progetto.prodotto.CodiceProdotto if progetto.prodotto else "N/A"
        project_owner = progetto.owner.Nome if progetto.owner else "N/A"
        project_end = progetto.ScadenzaProgetto.strftime('%d/%m/%Y') if progetto.ScadenzaProgetto else "Not Set"
        
        # Buffer info
        buffer_html = ""
        if buffer_days is not None:
            buffer_color = "#DC143C" if buffer_days <= 7 else "#FFA500" if buffer_days <= 14 else "#28a745"
            buffer_html = f'<div class="info-row"><span class="label">Days to Deadline:</span> <span class="value" style="font-weight: bold; color: {buffer_color};">{buffer_days} days</span></div>'
        
        # Lista task in ritardo
        overdue_list_html = ""
        for task in overdue_tasks:
            task_name = task.task_catalogo.NomeTask if task.task_catalogo else "Task"
            task_owner = task.owner.Nome if task.owner else "Unassigned"
            due_date = task.DataScadenza.strftime('%d/%m/%Y') if task.DataScadenza else "N/A"
            due_date = self._to_date(task.DataScadenza)
            if not due_date:
                continue
            days_overdue = (date.today() - due_date).days
            
            # Icon se critical path
            critical_icon = " 🔗" if task in critical_path_tasks else ""
            
            overdue_list_html += f"""
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;">{task_name}{critical_icon}</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{task_owner}</td>
                <td style="padding: 8px; border: 1px solid #ddd; color: #DC143C; font-weight: bold;">{due_date}</td>
                <td style="padding: 8px; border: 1px solid #ddd; color: #DC143C;">{days_overdue} days</td>
            </tr>
            """
        
        # Blocchi critici
        critical_path_html = ""
        if critical_path_tasks:
            blocked_sections = []
            for task in critical_path_tasks:
                task_name = task.task_catalogo.NomeTask if task.task_catalogo else "Task"
                blocked_tasks = blocking_info.get(task.TaskProdottoID, [])
                
                blocked_list = []
                for dep_task in blocked_tasks:
                    dep_name = dep_task.task_catalogo.NomeTask if dep_task.task_catalogo else "Task"
                    dep_owner = dep_task.owner.Nome if dep_task.owner else "Unassigned"
                    blocked_list.append(f"<li><strong>{dep_name}</strong> (Owner: {dep_owner})</li>")
                
                if blocked_list:
                    blocked_sections.append(f"""
                    <div style="margin: 10px 0; padding: 10px; background: #fff; border-left: 3px solid #DC143C;">
                        <strong style="color: #DC143C;">🔗 {task_name}</strong> is blocking:
                        <ul style="margin: 5px 0 0 20px; padding: 0;">
                            {''.join(blocked_list)}
                        </ul>
                    </div>
                    """)
            
            if blocked_sections:
                critical_path_html = f"""
                <div style="background: #FFF3CD; border-left: 4px solid #FFC107; padding: 15px; margin: 20px 0; border-radius: 4px;">
                    <h4 style="color: #856404; margin: 0 0 10px 0;">🔗 CRITICAL PATH IMPACT</h4>
                    <p style="margin: 0 0 10px 0; color: #856404;">The following overdue tasks are blocking other tasks, creating a critical path delay:</p>
                    {''.join(blocked_sections)}
                </div>
                """
        
        # Risk summary
        total_overdue = len(overdue_tasks)
        total_critical = len(critical_path_tasks)
        
        # Logo
        logo_html = ""
        if self.config['notification_settings']['include_logo']:
            logo_path = self.config['notification_settings']['logo_path']
            if os.path.exists(logo_path):
                logo_html = f'<img src="cid:logo" style="max-width: 150px; height: auto;" alt="Company Logo">'
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; background-color: #f9f9f9; }}
                .header {{ background: linear-gradient(135deg, #0078d4 0%, #005a9e 100%); color: white; padding: 30px 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                .header h1 {{ margin: 10px 0 0 0; font-size: 24px; }}
                .alert-box {{ background: {risk_color}; color: white; padding: 25px; text-align: center; font-size: 24px; font-weight: bold; }}
                .content {{ background: white; padding: 30px; border-radius: 0 0 8px 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .section {{ margin: 25px 0; padding: 20px; background: #f8f9fa; border-left: 4px solid #0078d4; border-radius: 4px; }}
                .section h2 {{ color: #0078d4; margin: 0 0 15px 0; font-size: 18px; border-bottom: 2px solid #0078d4; padding-bottom: 10px; }}
                .info-row {{ margin: 8px 0; }}
                .label {{ font-weight: bold; color: #555; display: inline-block; min-width: 150px; }}
                .value {{ color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
                th {{ background: #0078d4; color: white; padding: 10px; text-align: left; }}
                .footer {{ background: #f1f1f1; padding: 20px; margin-top: 20px; border-top: 3px solid #0078d4; border-radius: 4px; text-align: center; }}
                .footer p {{ margin: 5px 0; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    {logo_html}
                    <h1>NPI Project Management System</h1>
                    <p>Project Risk Alert</p>
                </div>
                
                <div class="alert-box">
                    {risk_icon} {risk_level.upper()} RISK - PROJECT AT RISK
                </div>
                
                <div class="content">
                    <p style="font-size: 16px; margin-bottom: 20px;">Dear <strong>{project_owner}</strong>,</p>
                    <p><strong>This automated alert notifies you that your project is at risk due to overdue tasks.</strong></p>
                    <p>Immediate action is required to prevent project delays and ensure timely completion.</p>
                    
                    <div class="section">
                        <h2>📋 PROJECT DETAILS</h2>
                        <div class="info-row"><span class="label">Project Name:</span> <span class="value">{project_name}</span></div>
                        <div class="info-row"><span class="label">Product Code:</span> <span class="value">{project_code}</span></div>
                        <div class="info-row"><span class="label">Project Owner:</span> <span class="value">{project_owner}</span></div>
                        <div class="info-row"><span class="label">Project End Date:</span> <span class="value">{project_end}</span></div>
                        {buffer_html}
                    </div>
                    
                    <div class="section">
                        <h2>⚠️ RISK ANALYSIS</h2>
                        <div class="info-row"><span class="label">Risk Level:</span> <span class="value" style="font-weight: bold; color: {risk_color};">{risk_icon} {risk_level.upper()}</span></div>
                        <div class="info-row"><span class="label">Total Overdue Tasks:</span> <span class="value" style="font-weight: bold; color: #DC143C;">{total_overdue}</span></div>
                        <div class="info-row"><span class="label">Critical Path Tasks:</span> <span class="value" style="font-weight: bold; color: #DC143C;">{total_critical}</span></div>
                        
                        <div style="background: #f8d7da; border-left: 4px solid #DC143C; padding: 15px; margin: 15px 0; border-radius: 4px;">
                            <h4 style="color: #721c24; margin: 0 0 10px 0;">⚠️ IMPACT ASSESSMENT</h4>
                            <p style="margin: 0; color: #721c24;">
                                Due to the overdue tasks, <strong>it is difficult for this project to be completed on time</strong>.
                                {f'With only {buffer_days} days remaining, ' if buffer_days is not None else ''}
                                {'Critical path tasks are blocking other team members, creating cascading delays.' if total_critical > 0 else 'Immediate corrective action is essential to recover the schedule.'}
                            </p>
                        </div>
                    </div>
                    
                    <div class="section">
                        <h2>📝 OVERDUE TASKS DETAILS</h2>
                        <table>
                            <thead>
                                <tr>
                                    <th>Task Name</th>
                                    <th>Owner</th>
                                    <th>Due Date</th>
                                    <th>Days Overdue</th>
                                </tr>
                            </thead>
                            <tbody>
                                {overdue_list_html}
                            </tbody>
                        </table>
                        <p style="font-size: 12px; color: #666; margin-top: 10px;">🔗 = Task on critical path (blocking other tasks)</p>
                    </div>
                    
                    {critical_path_html}
                    
                    <div style="background: #d1ecf1; border-left: 4px solid #0078d4; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <h4 style="color: #0c5460; margin: 0 0 10px 0;">📌 RECOMMENDED ACTIONS</h4>
                        <ul style="margin: 0; padding-left: 20px; color: #0c5460;">
                            <li><strong>Immediate Review:</strong> Contact task owners to understand delays and obstacles</li>
                            <li><strong>Resource Allocation:</strong> Consider reassigning resources to critical path tasks</li>
                            <li><strong>Timeline Adjustment:</strong> Evaluate if project timeline needs revision</li>
                            <li><strong>Stakeholder Communication:</strong> Inform stakeholders about project status and risks</li>
                            <li><strong>Daily Monitoring:</strong> Implement daily standup to track progress on overdue tasks</li>
                        </ul>
                    </div>
                </div>
                
                <div class="footer">
                    <p><strong>This is an automated risk assessment from the NPI Project Management System.</strong></p>
                    <p>Please do not reply to this email.</p>
                    <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html


# Istanza globale del servizio
_notification_service = None


def start_notification_service(npi_manager, config_path='npi_notifications_config.json'):
    """Avvia il servizio di notifiche automatiche."""
    global _notification_service
    
    if _notification_service is None:
        _notification_service = NpiAutoNotificationService(npi_manager, config_path)
        _notification_service.start()
        logger.info("Servizio notifiche automatiche NPI inizializzato")
    else:
        logger.warning("Servizio notifiche già attivo")
    
    return _notification_service


def ensure_notification_config(config_path='npi_notifications_config.json') -> dict:
    """
    Garantisce l'esistenza del file di configurazione notifiche.
    Se assente, crea il file con valori di default e ritorna la config.
    """
    try:
        service = NpiAutoNotificationService(npi_manager=None, config_path=config_path)
        config = service._load_config()
        if not os.path.exists(config_path):
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            logger.info(f"Creato file configurazione notifiche NPI: {config_path}")
        return config
    except Exception as e:
        logger.error(f"Errore ensure config notifiche NPI: {e}", exc_info=True)
        return {}


def stop_notification_service():
    """Ferma il servizio di notifiche automatiche."""
    global _notification_service
    
    if _notification_service:
        _notification_service.stop()
        _notification_service = None
        logger.info("Servizio notifiche automatiche NPI terminato")
