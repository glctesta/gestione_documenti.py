"""
indirect_materials_stock.py
GUI per la gestione giacenze Materiali Indiretti:
  - StockCheckWindow      : Verifica Giacenze (con riordino manuale + export)
  - MinStockConfigWindow  : Configurazione scorte minime / riordino

Usa il layer dati indirect_materials_stock_data.py.
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from datetime import datetime

import indirect_materials_stock_data as stock_data

logger = logging.getLogger(__name__)


# ============================================================================
#  Verifica Giacenze
# ============================================================================
class StockCheckWindow(tk.Toplevel):
    """Finestra di verifica delle giacenze correnti con evidenza sotto-soglia,
    dettaglio movimenti, riordino manuale ed export Excel."""

    def __init__(self, master, db, lang, user_name="Unknown"):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name

        self.title(lang.get('ind_stock_title', 'Verifica Giacenze Materiali Indiretti'))
        self.geometry("1050x680")
        self.resizable(True, True)
        self.transient(master)

        self._all_rows = []        # tutte le giacenze caricate
        self._filtered_rows = []   # giacenze attualmente visualizzate

        self._build_ui()
        self._load_giacenze()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ------------------------------------------------------------------ #
    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(expand=True, fill="both")

        # Header
        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(
            header,
            text=self.lang.get('ind_stock_header', 'Giacenze Materiali Indiretti'),
            font=("Segoe UI", 12, "bold")
        ).pack(side="left")

        ttk.Button(header, text=self.lang.get('btn_refresh', 'Aggiorna'),
                   command=self._load_giacenze).pack(side="right", padx=4)
        ttk.Button(header, text=self.lang.get('ind_stock_export', '📊 Esporta Excel'),
                   command=self._export_excel).pack(side="right", padx=4)
        self.btn_reorder = ttk.Button(
            header,
            text=self.lang.get('ind_stock_send_reorder', '📤 Invia riordino ora'),
            command=self._send_reorder_now
        )
        self.btn_reorder.pack(side="right", padx=4)

        # Filtri
        filt = ttk.Frame(main)
        filt.pack(fill="x", pady=(0, 8))
        ttk.Label(filt, text=self.lang.get('ind_req_filter_code', 'Codice:')).pack(side="left")
        self.filter_code_var = tk.StringVar()
        e1 = ttk.Entry(filt, textvariable=self.filter_code_var, width=18)
        e1.pack(side="left", padx=(2, 12))
        e1.bind('<KeyRelease>', lambda ev: self._apply_filter())

        ttk.Label(filt, text=self.lang.get('ind_req_filter_desc', 'Descrizione:')).pack(side="left")
        self.filter_desc_var = tk.StringVar()
        e2 = ttk.Entry(filt, textvariable=self.filter_desc_var, width=28)
        e2.pack(side="left", padx=(2, 12))
        e2.bind('<KeyRelease>', lambda ev: self._apply_filter())

        self.only_below_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            filt,
            text=self.lang.get('ind_stock_only_below', 'Solo sotto scorta minima'),
            variable=self.only_below_var,
            command=self._apply_filter
        ).pack(side="left", padx=(4, 0))

        # Tabella giacenze
        tree_frame = ttk.Frame(main)
        tree_frame.pack(fill="both", expand=True)

        cols = ('codice', 'descrizione', 'tipo', 'giacenza', 'minimo', 'lotto', 'stato')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode='browse')
        headings = {
            'codice':      self.lang.get('ind_import_col_code', 'Codice'),
            'descrizione': self.lang.get('ind_import_col_desc', 'Descrizione'),
            'tipo':        self.lang.get('ind_req_col_type', 'Tipo'),
            'giacenza':    self.lang.get('ind_stock_col_stock', 'Giacenza'),
            'minimo':      self.lang.get('ind_min_col_min', 'Scorta minima'),
            'lotto':       self.lang.get('ind_min_col_lot', 'Lotto riordino'),
            'stato':       self.lang.get('ind_stock_col_status', 'Stato'),
        }
        widths = {'codice': 110, 'descrizione': 300, 'tipo': 110,
                  'giacenza': 90, 'minimo': 100, 'lotto': 100, 'stato': 130}
        for c in cols:
            self.tree.heading(c, text=headings[c])
            anchor = 'e' if c in ('giacenza', 'minimo', 'lotto') else 'w'
            self.tree.column(c, width=widths[c], anchor=anchor)

        self.tree.tag_configure('below', background='#F8D7DA')   # rosso chiaro
        self.tree.tag_configure('ok', background='#D4EDDA')      # verde chiaro

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind('<<TreeviewSelect>>', self._on_select)

        # Dettaglio movimenti del materiale selezionato
        mov_lbl = ttk.Label(main, text=self.lang.get('ind_stock_movements', 'Ultimi movimenti'),
                            font=("Segoe UI", 10, "bold"))
        mov_lbl.pack(anchor="w", pady=(10, 2))

        mov_frame = ttk.Frame(main)
        mov_frame.pack(fill="both", expand=False)
        mcols = ('data', 'tipo', 'qty', 'utente', 'note')
        self.mov_tree = ttk.Treeview(mov_frame, columns=mcols, show='headings',
                                     selectmode='none', height=6)
        mheads = {
            'data':   self.lang.get('ind_req_col_date', 'Data'),
            'tipo':   self.lang.get('ind_stock_col_movtype', 'Tipo'),
            'qty':    self.lang.get('ind_req_qty', 'Quantità'),
            'utente': self.lang.get('ind_req_col_requester', 'Utente'),
            'note':   self.lang.get('ind_req_col_note', 'Note'),
        }
        mwidths = {'data': 140, 'tipo': 110, 'qty': 90, 'utente': 140, 'note': 360}
        for c in mcols:
            self.mov_tree.heading(c, text=mheads[c])
            self.mov_tree.column(c, width=mwidths[c],
                                 anchor='e' if c == 'qty' else 'w')
        mvsb = ttk.Scrollbar(mov_frame, orient="vertical", command=self.mov_tree.yview)
        self.mov_tree.configure(yscrollcommand=mvsb.set)
        self.mov_tree.pack(side="left", fill="both", expand=True)
        mvsb.pack(side="right", fill="y")

        # Barra di stato
        self.status_var = tk.StringVar(value="")
        ttk.Label(main, textvariable=self.status_var,
                  foreground="#555").pack(anchor="w", pady=(6, 0))

    # ------------------------------------------------------------------ #
    def _load_giacenze(self):
        try:
            self._all_rows = stock_data.get_giacenze(self.db)
        except Exception as e:
            logger.error(f"Errore caricamento giacenze: {e}", exc_info=True)
            self._all_rows = []
        self._apply_filter()
        below = sum(1 for r in self._all_rows if r['sotto_soglia'])
        self.status_var.set(
            self.lang.get('ind_stock_status', '{0} materiali · {1} sotto scorta minima')
            .format(len(self._all_rows), below)
        )

    def _apply_filter(self):
        code_f = self.filter_code_var.get().strip().lower()
        desc_f = self.filter_desc_var.get().strip().lower()
        only_below = self.only_below_var.get()

        self.tree.delete(*self.tree.get_children())
        self._filtered_rows = []
        for r in self._all_rows:
            if only_below and not r['sotto_soglia']:
                continue
            if code_f and code_f not in r['codice'].lower():
                continue
            if desc_f and desc_f not in r['descrizione'].lower():
                continue
            idx = len(self._filtered_rows)
            min_str = f"{r['livello_minimo']:.2f}" if r['livello_minimo'] is not None else '-'
            lot_str = f"{r['lotto_riordino']:.2f}" if r['lotto_riordino'] is not None else '-'
            if r['sotto_soglia']:
                stato_txt = self.lang.get('ind_stock_below', '⚠ Sotto minimo')
                tag = ('below',)
            elif r['is_riordino_attivo'] and r['livello_minimo'] is not None:
                stato_txt = self.lang.get('ind_stock_ok', 'OK')
                tag = ('ok',)
            else:
                stato_txt = self.lang.get('ind_stock_unmanaged', 'Non gestito')
                tag = ()
            self.tree.insert('', 'end', iid=str(idx), tags=tag, values=(
                r['codice'], r['descrizione'], r['tipo'],
                f"{r['giacenza']:.2f}", min_str, lot_str, stato_txt
            ))
            self._filtered_rows.append(r)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        self.mov_tree.delete(*self.mov_tree.get_children())
        if not sel:
            return
        idx = int(sel[0])
        if idx >= len(self._filtered_rows):
            return
        r = self._filtered_rows[idx]
        try:
            movs = stock_data.get_movimenti(self.db, r['materiale_id'], limit=100)
        except Exception as e:
            logger.error(f"Errore caricamento movimenti: {e}", exc_info=True)
            movs = []
        for m in movs:
            data_str = m['data'].strftime('%d/%m/%Y %H:%M') if m['data'] else ''
            self.mov_tree.insert('', 'end', values=(
                data_str, m['tipo'], f"{m['qty']:+.2f}", m['eseguito_da'], m['note']
            ))

    # ------------------------------------------------------------------ #
    def _send_reorder_now(self):
        """Invio riordino manuale, previa autorizzazione
        (chiave permesso: riordine_materiali_indiretti)."""
        auth = getattr(self.master, '_execute_authorized_action', None)
        if callable(auth):
            auth(menu_translation_key='riordine_materiali_indiretti',
                 action_callback=self._do_send_reorder)
        else:
            # Fallback (finestra non agganciata alla form principale): invio diretto
            self._do_send_reorder()

    def _do_send_reorder(self):
        if not messagebox.askyesno(
            self.lang.get('confirm', 'Conferma'),
            self.lang.get('ind_stock_reorder_confirm',
                          'Inviare ora la richiesta di riordino per i materiali sotto scorta minima?'),
            parent=self
        ):
            return
        try:
            res = stock_data.check_and_send_reorder(self.db, self.lang, force=True)
        except Exception as e:
            logger.error(f"Errore invio riordino manuale: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)
            return

        reason = res.get('reason')
        if res.get('sent'):
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('ind_stock_reorder_sent',
                              'Riordino inviato per {0} materiali a {1} destinatari.')
                .format(res['count'], len(res['recipients'])),
                parent=self
            )
        elif reason == 'no_items':
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('ind_stock_reorder_none', 'Nessun materiale sotto scorta minima.'),
                parent=self
            )
        elif reason == 'no_recipients':
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('ind_stock_reorder_no_recipients',
                              "Nessun destinatario configurato (Settings: {0}).")
                .format(stock_data.REORDER_EMAIL_ATTRIBUTE),
                parent=self
            )
        else:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                self.lang.get('ind_stock_reorder_error', 'Invio riordino non riuscito: {0}')
                .format(reason),
                parent=self
            )

    # ------------------------------------------------------------------ #
    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                "La libreria 'openpyxl' non è installata.\nEsegui: pip install openpyxl",
                parent=self
            )
            return

        default_name = f"Giacenze_Materiali_Indiretti_{datetime.now().strftime('%Y%m%d')}.xlsx"
        default_dir = r'C:\Temp'
        os.makedirs(default_dir, exist_ok=True)
        path = filedialog.asksaveasfilename(
            parent=self,
            title=self.lang.get('ind_stock_save_excel', 'Salva Giacenze Excel'),
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialdir=default_dir,
            initialfile=default_name
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Giacenze'
            hdr_font = Font(bold=True, color='FFFFFF')
            hdr_fill = PatternFill('solid', fgColor='2F6DA4')
            red_fill = PatternFill('solid', fgColor='F8D7DA')

            headers = ['Codice', 'Descrizione', 'Tipo', 'Giacenza',
                       'Scorta minima', 'Lotto riordino', 'Riordino attivo', 'Sotto minimo']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            for r in self._filtered_rows:
                ws.append([
                    r['codice'], r['descrizione'], r['tipo'], r['giacenza'],
                    r['livello_minimo'] if r['livello_minimo'] is not None else '',
                    r['lotto_riordino'] if r['lotto_riordino'] is not None else '',
                    'Sì' if r['is_riordino_attivo'] else 'No',
                    'Sì' if r['sotto_soglia'] else 'No',
                ])
                if r['sotto_soglia']:
                    for c in ws[ws.max_row]:
                        c.fill = red_fill

            for i, w in enumerate([14, 38, 14, 12, 14, 14, 14, 12], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            wb.save(path)
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('ind_stock_export_ok', 'Export completato:\n{0}').format(path),
                parent=self
            )
        except Exception as e:
            logger.error(f"Errore export giacenze: {e}", exc_info=True)
            messagebox.showerror(self.lang.get('error', 'Errore'), str(e), parent=self)


# ============================================================================
#  Configurazione scorte minime
# ============================================================================
class MinStockConfigWindow(tk.Toplevel):
    """Form per impostare scorta minima / lotto riordino per i codici dove
    ha senso (riordino attivabile per codice)."""

    def __init__(self, master, db, lang, user_name="Unknown"):
        super().__init__(master)
        self.db = db
        self.lang = lang
        self.user_name = user_name

        self.title(lang.get('ind_min_title', 'Configura Scorte Minime'))
        self.geometry("900x600")
        self.resizable(True, True)
        self.transient(master)
        self.grab_set()

        self._rows = []
        self._selected = None

        self._build_ui()
        self._load()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(expand=True, fill="both")

        ttk.Label(main, text=self.lang.get('ind_min_header', 'Scorte minime materiali indiretti'),
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

        # Filtri: codice e descrizione separati, come nella form Verifica Giacenze.
        # Prima c'era una casella sola, etichettata "Codice" ma che cercava anche
        # nella descrizione: non si poteva restringere per descrizione senza
        # trascinarsi dentro i codici che contenevano le stesse cifre.
        filt = ttk.Frame(main)
        filt.pack(fill="x", pady=(0, 6))
        ttk.Label(filt, text=self.lang.get('ind_req_filter_code', 'Codice:')).pack(side="left")
        self.filter_var = tk.StringVar()
        e = ttk.Entry(filt, textvariable=self.filter_var, width=18)
        e.pack(side="left", padx=(2, 12))
        e.bind('<KeyRelease>', lambda ev: self._apply_filter())

        ttk.Label(filt, text=self.lang.get('ind_req_filter_desc', 'Descrizione:')).pack(side="left")
        self.filter_desc_var = tk.StringVar()
        e_desc = ttk.Entry(filt, textvariable=self.filter_desc_var, width=30)
        e_desc.pack(side="left", padx=(2, 12))
        e_desc.bind('<KeyRelease>', lambda ev: self._apply_filter())

        ttk.Button(filt, text=self.lang.get('btn_clear_filters', 'Azzera filtri'),
                   command=self._clear_filters).pack(side="left")

        # Tabella
        tree_frame = ttk.Frame(main)
        tree_frame.pack(fill="both", expand=True)
        cols = ('codice', 'descrizione', 'giacenza', 'minimo', 'racc', 'lotto', 'attivo')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode='browse')
        heads = {
            'codice':      self.lang.get('ind_import_col_code', 'Codice'),
            'descrizione': self.lang.get('ind_import_col_desc', 'Descrizione'),
            'giacenza':    self.lang.get('ind_stock_col_stock', 'Giacenza'),
            'minimo':      self.lang.get('ind_min_col_min', 'Scorta minima'),
            'racc':        self.lang.get('ind_min_col_recommended', 'Scorta raccomandata'),
            'lotto':       self.lang.get('ind_min_col_lot', 'Lotto riordino'),
            'attivo':      self.lang.get('ind_min_col_active', 'Riordino attivo'),
        }
        widths = {'codice': 110, 'descrizione': 280, 'giacenza': 90,
                  'minimo': 100, 'racc': 120, 'lotto': 100, 'attivo': 110}
        for c in cols:
            self.tree.heading(c, text=heads[c])
            self.tree.column(c, width=widths[c],
                             anchor='e' if c in ('giacenza', 'minimo', 'racc', 'lotto') else 'w')
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind('<<TreeviewSelect>>', self._on_select)

        # Editor
        ed = ttk.LabelFrame(main, text=self.lang.get('ind_min_editor', 'Configurazione codice selezionato'),
                            padding=10)
        ed.pack(fill="x", pady=(10, 0))

        self.sel_label_var = tk.StringVar(value=self.lang.get('ind_min_no_sel', 'Nessun codice selezionato'))
        ttk.Label(ed, textvariable=self.sel_label_var, font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 8))

        ttk.Label(ed, text=self.lang.get('ind_min_col_min', 'Scorta minima') + ':').grid(
            row=1, column=0, sticky="e", padx=4, pady=3)
        self.min_var = tk.StringVar()
        ttk.Entry(ed, textvariable=self.min_var, width=12).grid(row=1, column=1, sticky="w", padx=4)

        ttk.Label(ed, text=self.lang.get('ind_min_col_recommended', 'Scorta raccomandata') + ':').grid(
            row=1, column=2, sticky="e", padx=4, pady=3)
        self.recc_var = tk.StringVar()
        ttk.Entry(ed, textvariable=self.recc_var, width=12).grid(row=1, column=3, sticky="w", padx=4)

        ttk.Label(ed, text=self.lang.get('ind_min_col_lot', 'Lotto riordino') + ':').grid(
            row=2, column=0, sticky="e", padx=4, pady=3)
        self.lot_var = tk.StringVar()
        ttk.Entry(ed, textvariable=self.lot_var, width=12).grid(row=2, column=1, sticky="w", padx=4)

        self.active_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(ed, text=self.lang.get('ind_min_col_active', 'Riordino attivo'),
                        variable=self.active_var).grid(row=2, column=2, sticky="w", padx=4, pady=6)

        self.btn_save = ttk.Button(ed, text=self.lang.get('btn_save', 'Salva'),
                                   command=self._save, state='disabled')
        self.btn_save.grid(row=2, column=3, sticky="e", padx=4, pady=6)

    def _load(self):
        try:
            self._rows = stock_data.get_giacenze(self.db)
        except Exception as e:
            logger.error(f"Errore caricamento config minimi: {e}", exc_info=True)
            self._rows = []
        self._apply_filter()

    def _clear_filters(self):
        self.filter_var.set('')
        self.filter_desc_var.set('')
        self._apply_filter()

    def _apply_filter(self):
        code_f = self.filter_var.get().strip().lower()
        desc_f = self.filter_desc_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        self._visible = []
        for r in self._rows:
            if code_f and code_f not in (r['codice'] or '').lower():
                continue
            if desc_f and desc_f not in (r['descrizione'] or '').lower():
                continue
            idx = len(self._visible)
            min_str = f"{r['livello_minimo']:.2f}" if r['livello_minimo'] is not None else '-'
            racc_str = (f"{r.get('livello_raccomandato'):.2f}"
                        if r.get('livello_raccomandato') is not None else '-')
            lot_str = f"{r['lotto_riordino']:.2f}" if r['lotto_riordino'] is not None else '-'
            att_str = (self.lang.get('yes', 'Sì') if r['is_riordino_attivo']
                       else self.lang.get('no', 'No')) if r['livello_minimo'] is not None else '-'
            self.tree.insert('', 'end', iid=str(idx), values=(
                r['codice'], r['descrizione'], f"{r['giacenza']:.2f}",
                min_str, racc_str, lot_str, att_str
            ))
            self._visible.append(r)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            self._selected = None
            self.btn_save.state(['disabled'])
            return
        idx = int(sel[0])
        if idx >= len(self._visible):
            return
        r = self._visible[idx]
        self._selected = r
        self.sel_label_var.set(f"{r['codice']} - {r['descrizione']}")
        self.min_var.set(f"{r['livello_minimo']:.2f}" if r['livello_minimo'] is not None else '')
        self.recc_var.set(f"{r.get('livello_raccomandato'):.2f}"
                          if r.get('livello_raccomandato') is not None else '')
        self.lot_var.set(f"{r['lotto_riordino']:.2f}" if r['lotto_riordino'] is not None else '')
        self.active_var.set(r['is_riordino_attivo'] if r['livello_minimo'] is not None else True)
        self.btn_save.state(['!disabled'])

    def _save(self):
        if not self._selected:
            return
        try:
            min_val = float(self.min_var.get().replace(',', '.'))
        except ValueError:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('ind_min_invalid_min', 'Scorta minima non valida.'),
                parent=self)
            return
        if min_val < 0:
            messagebox.showwarning(
                self.lang.get('warning', 'Attenzione'),
                self.lang.get('ind_min_invalid_min', 'Scorta minima non valida.'),
                parent=self)
            return
        lot_raw = self.lot_var.get().strip().replace(',', '.')
        lot_val = None
        if lot_raw:
            try:
                lot_val = float(lot_raw)
            except ValueError:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('ind_min_invalid_lot', 'Lotto riordino non valido.'),
                    parent=self)
                return

        recc_raw = self.recc_var.get().strip().replace(',', '.')
        recc_val = None
        if recc_raw:
            try:
                recc_val = float(recc_raw)
            except ValueError:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('ind_min_invalid_recommended', 'Scorta raccomandata non valida.'),
                    parent=self)
                return
            if recc_val < min_val:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('ind_min_recommended_lt_min',
                                  'La scorta raccomandata deve essere ≥ della scorta minima.'),
                    parent=self)
                return

        ok, msg = stock_data.upsert_min_config(
            self.db, self._selected['materiale_id'], min_val, lot_val,
            self.active_var.get(), self.user_name, livello_raccomandato=recc_val
        )
        if ok:
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('ind_min_saved', 'Configurazione salvata.'),
                parent=self)
            self._load()
        else:
            messagebox.showerror(self.lang.get('error', 'Errore'), msg, parent=self)


# ============================================================================
#  Entry-point richiamabili da main.py
# ============================================================================
def open_stock_check(master, db, lang, user_name="Unknown"):
    StockCheckWindow(master, db, lang, user_name)


def open_min_stock_config(master, db, lang, user_name="Unknown"):
    MinStockConfigWindow(master, db, lang, user_name)
