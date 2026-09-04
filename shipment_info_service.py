# -*- coding: utf-8 -*-
"""
shipment_info_service.py
Servizio "Info Spedizioni": monitora le directory \\192.168.10.110\Shipping\<sito>
configurate in ShipmentEmailConfig; quando trova un file Excel di export D365
lo legge, invia una email professionale ai destinatari configurati (TO/CC),
allega il file e rinomina l'originale con prefisso 'Executed_'.

I file che iniziano per 'Executed_' non vengono mai ripresi in considerazione.
Il coordinamento multi-PC e' gestito da email_job_coordinator (claim_job_run).
"""
import logging
import os

logger = logging.getLogger("TraceabilityRS")

JOB_NAME = 'shipment_info_emails'

# Radice share spedizioni (indirizzo assoluto, non lettera mappata:
# T: = \\192.168.10.110\InternalApplications, quindi T:\Shipping diventa...)
SHIPPING_ROOT = r"\\192.168.10.110\InternalApplications\Shipping"

EXECUTED_PREFIX = 'Executed_'

CREATE_TABLE_SQL = """
IF OBJECT_ID('Traceability_RS.dbo.ShipmentEmailConfig', 'U') IS NULL
CREATE TABLE Traceability_RS.dbo.ShipmentEmailConfig (
    ConfigId      INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
    IDSite        INT               NOT NULL,
    DirectoryName NVARCHAR(200)     NOT NULL,
    ToEmails      NVARCHAR(MAX)     NULL,
    CcEmails      NVARCHAR(MAX)     NULL,
    IsActive      BIT               NOT NULL DEFAULT 1,
    [User]        NVARCHAR(100)     NULL,
    DateIn        DATETIME          NOT NULL DEFAULT GETDATE(),
    DateOut       DATETIME          NULL
)
"""

# File gia' elaborati: la presenza della riga (qualsiasi Status) impedisce per
# sempre il re-invio per quel sito/file. La rename con prefisso Executed_
# resta solo un aiuto visivo, non e' piu' il meccanismo anti-duplicato.
CREATE_PROCESSED_TABLE_SQL = """
IF OBJECT_ID('Traceability_RS.dbo.ShipmentProcessedFiles', 'U') IS NULL
CREATE TABLE Traceability_RS.dbo.ShipmentProcessedFiles (
    ProcessedFileId INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
    IDSite          INT               NOT NULL,
    FileName        NVARCHAR(300)     NOT NULL,
    Status          VARCHAR(20)       NOT NULL DEFAULT 'SENT',
    SentTo          NVARCHAR(MAX)     NULL,
    ErrorMessage    NVARCHAR(500)     NULL,
    ProcessedAt     DATETIME          NOT NULL DEFAULT GETDATE(),
    CONSTRAINT UQ_ShipmentProcessedFiles_Site_File UNIQUE (IDSite, FileName)
)
"""

_Q_PROCESSED_NAMES = """
SELECT FileName
FROM Traceability_RS.dbo.ShipmentProcessedFiles
WHERE IDSite = ?
"""

_Q_MARK_PROCESSED = """
INSERT INTO Traceability_RS.dbo.ShipmentProcessedFiles
    (IDSite, FileName, Status, SentTo)
VALUES (?, ?, ?, ?)
"""

_Q_MARK_ERROR = """
INSERT INTO Traceability_RS.dbo.ShipmentProcessedFiles
    (IDSite, FileName, Status, ErrorMessage)
VALUES (?, ?, 'ERROR', ?)
"""

_Q_CONFIGS = """
SELECT c.ConfigId, c.IDSite, s.SiteName, c.DirectoryName,
       c.ToEmails, c.CcEmails, c.IsActive
FROM Traceability_RS.dbo.ShipmentEmailConfig c
INNER JOIN Traceability_RS.dbo.Sites s ON s.IDSite = c.IDSite
WHERE c.DateOut IS NULL
  AND c.IsActive = 1
  AND LEN(LTRIM(RTRIM(ISNULL(c.ToEmails, '')))) > 0
ORDER BY s.SiteName
"""

# Colonne attese nell'export D365 (DynamicsExport_*.xlsx)
_COL_PHYSICAL_DATE = 'physical date'
_COL_ITEM_NUMBER = 'item number'
_COL_NUMBER = 'number'
_COL_BATCH_NUMBER = 'batch number'
_COL_QUANTITY = 'quantity'


def ensure_config_table(db):
    """Crea le tabelle ShipmentEmailConfig e ShipmentProcessedFiles se mancano (best-effort)."""
    try:
        db._ensure_connection()
        with db._lock:
            cur = db.cursor
            cur.execute(CREATE_TABLE_SQL)
            cur.execute(CREATE_PROCESSED_TABLE_SQL)
            db.conn.commit()
    except Exception as e:
        logger.warning("Impossibile assicurare tabelle spedizioni: %s", e)


def load_processed_file_names(db, id_site):
    """Nomi dei file gia' registrati per il sito: non vanno mai riproposti."""
    try:
        db._ensure_connection()
        with db._lock:
            cur = db.cursor
            cur.execute(_Q_PROCESSED_NAMES, (id_site,))
            return {r[0] for r in cur.fetchall()}
    except Exception as e:
        logger.error("shipment_info_service: errore lettura file processati: %s", e)
        return set()


def mark_file_processed(db, id_site, file_name, status, info=''):
    """Registra il file nella tabella anti-duplicati (fallisce in silenzio solo in caso di errore DB)."""
    try:
        db._ensure_connection()
        with db._lock:
            cur = db.cursor
            if status == 'SENT':
                cur.execute(_Q_MARK_PROCESSED, (id_site, file_name, 'SENT', info))
            else:
                cur.execute(_Q_MARK_ERROR, (id_site, file_name, str(info)[:450]))
            db.conn.commit()
            return True
    except Exception as e:
        logger.error("shipment_info_service: errore registrazione file processato %s: %s",
                     file_name, e)
        return False


def load_active_configs(db):
    """Carica le configurazioni attive con destinatari valorizzati."""
    try:
        db._ensure_connection()
        with db._lock:
            cur = db.cursor
            cur.execute(_Q_CONFIGS)
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, r)) for r in cur.fetchall()]
    except Exception as e:
        logger.error("shipment_info_service: errore caricamento config: %s", e, exc_info=True)
        return []


def _split_addresses(raw):
    if not raw:
        return []
    return [a.strip() for a in raw.replace(',', ';').split(';') if a.strip()]


def _find_col(header_map, wanted):
    """Trova l'indice colonna cercando per nome esatto (normalizzato) o per prefisso."""
    if wanted in header_map:
        return header_map[wanted]
    for name, idx in header_map.items():
        if name.startswith(wanted):
            return idx
    return None


def _parse_excel(file_path):
    """
    Legge l'export D365 e ritorna dict con:
    ship_date (str dd/mm/YYYY o None), item_count, total_qty, rows [(item, number, batch, qty)]
    Ritorna None se il file non e' leggibile o mancano le colonne chiave.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("shipment_info_service: openpyxl non disponibile")
        return None

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return None

        header_map = {}
        for idx, cell in enumerate(header):
            if cell is None:
                continue
            name = str(cell).strip().lower()
            if name and name not in header_map:
                header_map[name] = idx

        idx_item = _find_col(header_map, _COL_ITEM_NUMBER)
        idx_num = _find_col(header_map, _COL_NUMBER)
        idx_batch = _find_col(header_map, _COL_BATCH_NUMBER)
        idx_qty = _find_col(header_map, _COL_QUANTITY)
        idx_date = _find_col(header_map, _COL_PHYSICAL_DATE)

        if idx_item is None or idx_qty is None:
            logger.warning("shipment_info_service: colonne Item number/Quantity non trovate in %s",
                           os.path.basename(file_path))
            wb.close()
            return None

        parsed_rows = []
        ship_date = None
        total_qty = 0.0
        for r in rows_iter:
            if r is None:
                continue
            item = str(r[idx_item]).strip() if idx_item is not None and r[idx_item] is not None else ''
            qty_raw = r[idx_qty] if idx_qty is not None else None
            try:
                qty = abs(float(qty_raw)) if qty_raw is not None else 0.0
            except (TypeError, ValueError):
                qty = 0.0
            number = str(r[idx_num]).strip() if idx_num is not None and r[idx_num] is not None else ''
            batch = str(r[idx_batch]).strip() if idx_batch is not None and r[idx_batch] is not None else ''

            if not item and qty == 0.0:
                continue

            if ship_date is None and idx_date is not None and r[idx_date] is not None:
                ship_date = _format_date(r[idx_date])

            total_qty += qty
            parsed_rows.append((item, number, batch, qty))

        wb.close()

        if not parsed_rows:
            return None

        item_count = len({row[0] for row in parsed_rows if row[0]})
        return {
            'ship_date': ship_date or '',
            'item_count': item_count,
            'total_qty': total_qty,
            'rows': parsed_rows,
        }
    except Exception as e:
        logger.error("shipment_info_service: errore lettura %s: %s", file_path, e, exc_info=True)
        return None


def _format_date(value):
    """Normalizza la Physical date (datetime o stringa) in dd/mm/YYYY."""
    try:
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y')
        s = str(value).strip()
        # Prova i formati piu' comuni
        from datetime import datetime
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(s[:10], fmt).strftime('%d/%m/%Y')
            except ValueError:
                continue
        return s[:10]
    except Exception:
        return str(value)[:10]


def _build_html(site_name, file_name, data, greeting):
    from html import escape
    date_txt = data['ship_date'] or '—'
    total = data['total_qty']
    total_fmt = f"{total:,.0f}" if total == int(total) else f"{total:,.2f}"

    table_rows = []
    for item, number, batch, qty in data['rows']:
        qty_fmt = f"{qty:,.0f}" if qty == int(qty) else f"{qty:,.2f}"
        table_rows.append(
            f"<tr>"
            f"<td style='padding:6px 10px;border:1px solid #dde1e7;'>{escape(item)}</td>"
            f"<td style='padding:6px 10px;border:1px solid #dde1e7;'>{escape(number)}</td>"
            f"<td style='padding:6px 10px;border:1px solid #dde1e7;'>{escape(batch)}</td>"
            f"<td style='padding:6px 10px;border:1px solid #dde1e7;text-align:right;'>{qty_fmt}</td>"
            f"</tr>"
        )

    body_date = escape(date_txt)
    body_file = escape(file_name)

    return f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background-color:#f4f6f8;font-family:Segoe UI,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f6f8;">
    <tr><td align="center" style="padding:24px 12px;">
      <table role="presentation" width="720" cellpadding="0" cellspacing="0"
             style="background-color:#ffffff;border:1px solid #dde1e7;">
        <tr>
          <td style="background-color:#1f3864;padding:16px 24px;">
            <img src="cid:company_logo" alt="Logo" style="max-height:48px;"/>
          </td>
        </tr>
        <tr>
          <td style="padding:24px 24px 8px 24px;color:#2c3e50;font-size:15px;">
            <p style="margin:0 0 12px 0;"><strong>{greeting},</strong></p>
            <p style="margin:0 0 12px 0;line-height:1.5;">
              in allegato trovate i dati relativi alla spedizione del giorno
              <strong>{body_date}</strong> (file <em>{body_file}</em>):
              <strong>{data['item_count']}</strong> codici per un totale di
              <strong>{total_fmt}</strong> pezzi.
            </p>
          </td>
        </tr>
        <tr>
          <td style="padding:8px 24px 16px 24px;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                   style="border-collapse:collapse;font-size:13px;color:#2c3e50;">
              <tr style="background-color:#1f3864;color:#ffffff;">
                <th style="padding:8px 10px;border:1px solid #1f3864;text-align:left;">Item number</th>
                <th style="padding:8px 10px;border:1px solid #1f3864;text-align:left;">Sales order</th>
                <th style="padding:8px 10px;border:1px solid #1f3864;text-align:left;">Production order</th>
                <th style="padding:8px 10px;border:1px solid #1f3864;text-align:right;">Quantity</th>
              </tr>
              {''.join(table_rows)}
            </table>
          </td>
        </tr>
        <tr>
          <td style="padding:8px 24px 24px 24px;color:#7f8c8d;font-size:12px;">
            <p style="margin:0;">Messaggio generato automaticamente dal sistema Traceability RS —
            si prega di non rispondere.</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""


def _greeting():
    from datetime import datetime
    return 'Buonasera' if datetime.now().hour >= 18 else 'Buongiorno'


def _send_for_file(db, config, file_path):
    """Elabora un singolo file: parse, email, registrazione anti-duplicati, rename best-effort.
    Ritorna 'sent' oppure 'skipped'."""
    from datetime import datetime
    from email_connector import EmailSender

    file_name = os.path.basename(file_path)
    data = _parse_excel(file_path)
    if data is None:
        logger.warning("shipment_info_service: file %s saltato (vuoto o colonne mancanti)", file_name)
        return 'skipped'

    to_list = _split_addresses(config['ToEmails'])
    cc_list = _split_addresses(config.get('CcEmails'))
    if not to_list:
        logger.warning("shipment_info_emails: nessun destinatario TO per sito %s", config['SiteName'])
        return 'skipped'

    date_part = data['ship_date'].replace('/', '-') if data['ship_date'] else datetime.now().strftime('%d-%m-%Y')
    if data['total_qty'] == int(data['total_qty']):
        total_fmt = f"{data['total_qty']:,.0f}".replace(',', ' ')
    else:
        total_fmt = f"{data['total_qty']:,.2f}"
    subject = (f"Spedizione {config['SiteName']} del {date_part} — "
               f"{data['item_count']} codici / {total_fmt} pz")

    html = _build_html(config['SiteName'], file_name, data, _greeting())

    attachments = [file_path]
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Logo.png')
    if os.path.exists(logo_path):
        attachments.append(('inline', logo_path, 'company_logo'))
    else:
        logger.warning("Logo non trovato in %s, email inviata senza logo", logo_path)

    try:
        sender = EmailSender()
        sender.send_email(
            to_email=';'.join(to_list),
            subject=subject,
            body=html,
            is_html=True,
            attachments=attachments,
            cc_emails=cc_list or None,
        )
    except Exception as e:
        # Registra l'errore cosi' il file non viene mai piu' riproposto (log per controllo manuale).
        mark_file_processed(db, config['IDSite'], file_name, 'ERROR', str(e))
        raise

    # Invio riuscito: registra il file PRIMA di qualsiasi altra operazione, cosi'
    # nessun ciclo successivo (su questo o su altri PC) potra' mai reinviarlo.
    mark_file_processed(db, config['IDSite'], file_name, 'SENT',
                        f"TO={';'.join(to_list)};CC={';'.join(cc_list)}")

    # Rename con prefisso Executed_: solo aiuto visivo, mai bloccante.
    executed_path = os.path.join(os.path.dirname(file_path), EXECUTED_PREFIX + file_name)
    try:
        os.replace(file_path, executed_path)
    except OSError as e:
        logger.warning(
            "shipment_info_emails: file %s NON rinominato (%s) — resta con il nome "
            "originale ma NON verra' mai riprocessato grazie alla tabella",
            file_name, e)

    logger.info(
        "shipment_info_emails: inviata email spedizione %s | file=%s | TO=%s | CC=%s",
        config['SiteName'], file_name, ';'.join(to_list), ';'.join(cc_list) or '—'
    )
    return 'sent'


def run_shipment_info_check(db):
    """
    Punto di ingresso del job: scansione directory, invio email, rename file.
    Coordinato multi-PC da email_job_coordinator (claim_job_run).
    Ritorna dict riepilogativo.
    """
    from email_job_coordinator import claim_job_run, release_job_lock, log_job_run

    summary = {'sent': 0, 'errors': 0, 'scanned_dirs': 0, 'details': []}

    if not claim_job_run(db, JOB_NAME, lock_minutes=2):
        summary['skipped'] = 'locked_or_disabled'
        return summary

    delivered = False
    try:
        ensure_config_table(db)
        configs = load_active_configs(db)
        if not configs:
            logger.info("shipment_info_emails: nessuna configurazione attiva")
            log_job_run(db, JOB_NAME, 'SKIPPED', 'Nessuna configurazione attiva')
            return summary

        for config in configs:
            directory = os.path.join(SHIPPING_ROOT, str(config['DirectoryName']).strip())
            summary['scanned_dirs'] += 1
            if not os.path.isdir(directory):
                logger.warning("shipment_info_emails: directory non raggiungibile: %s", directory)
                continue

            try:
                files = sorted(
                    f for f in os.listdir(directory)
                    if f.lower().endswith('.xlsx')
                    and not f.startswith(EXECUTED_PREFIX)
                    and not f.startswith('~$')
                )
            except Exception as e:
                logger.error("shipment_info_emails: errore lettura %s: %s", directory, e)
                summary['errors'] += 1
                continue

            # File gia' registrati in tabella: non verranno MAI riproposti.
            processed_names = load_processed_file_names(db, config['IDSite'])
            skipped_processed = [f for f in files if f in processed_names]
            if skipped_processed:
                logger.info(
                    "shipment_info_emails: %d file gia' elaborati saltati per sito %s (%s...)",
                    len(skipped_processed), config['SiteName'], skipped_processed[0])
            files = [f for f in files if f not in processed_names]

            for f in files:
                file_path = os.path.join(directory, f)
                try:
                    if _send_for_file(db, config, file_path) == 'sent':
                        summary['sent'] += 1
                        delivered = True
                except Exception as e:
                    summary['errors'] += 1
                    logger.error("shipment_info_emails: errore invio %s: %s", f, e, exc_info=True)

        status = 'OK' if summary['errors'] == 0 else 'ERROR'
        log_job_run(db, JOB_NAME, status,
                    f"Inviate {summary['sent']} email, {summary['errors']} errori")
        return summary

    except Exception as e:
        summary['errors'] += 1
        logger.error("shipment_info_emails: %s", e, exc_info=True)
        log_job_run(db, JOB_NAME, 'ERROR', str(e)[:400])
        return summary
    finally:
        if not delivered and summary['errors'] > 0:
            # Invio non consegnato: rilascia il lock cosi' un altro PC (o il giro
            # successivo) riprova subito invece di aspettare la scadenza del lock.
            release_job_lock(db, JOB_NAME)
