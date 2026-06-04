# -*- coding: utf-8 -*-
"""
material_consumption_report_gui.py

Interactive search form for alloy/flux consumption reports (PTHM phase).
- Filter by date range and optional product code
- Results grouped by production day
- Products WITHOUT weight data shown in RED with zero values
- Summary footer: codes with weight / without weight / difference
- Export to Excel (c:\\temp\\Raport_consum_aliaj_flux_per_perioada_<start>-<end>.xlsx)
- Email scheduling info: key 'Sys_missing_data_alloy' in settings, daily at 08:05 Mon-Sat
"""
from __future__ import annotations

import datetime
import logging
import os
import tkinter as tk
from tkinter import messagebox, ttk

logger = logging.getLogger("TraceabilityRS")

# ─── SQL ──────────────────────────────────────────────────────────────────────

_Q_CONSUMPTION_RANGE = """
SELECT
    CAST(s.ScanTimeFinish AS DATE)              AS ProductionDay,
    o.IDProduct,
    p.ProductCode,
    ISNULL(SUM(CASE WHEN s.IsPass = 1 THEN 1 ELSE 0 END), 0) AS QtyProcessed,
    pc.MaterialConsumptionGR,
    CASE WHEN pc.ProductConsumptionId IS NULL THEN 1 ELSE 0 END AS MissingAlloyConsumption
FROM Traceability_RS.dbo.Scannings s
INNER JOIN Traceability_RS.dbo.OrderPhases op ON s.IDOrderPhase = op.IDOrderPhase
INNER JOIN Traceability_RS.dbo.Orders      o  ON op.IDOrder     = o.IDOrder
INNER JOIN Traceability_RS.dbo.Phases      ph ON op.IDPhase     = ph.IDPhase
INNER JOIN Traceability_RS.dbo.Products    p  ON o.IDProduct    = p.IDProduct
LEFT  JOIN Traceability_RS.dbo.ProductConsumptions pc
       ON  pc.IdProduct          = o.IDProduct
      AND  pc.MaterialConsumption = 'Alloy'
      AND  pc.DateOut             IS NULL
WHERE ph.IDPhase        = 107
  AND s.ScanTimeFinish >= ?
  AND s.ScanTimeFinish  < ?
{code_filter}
GROUP BY
    CAST(s.ScanTimeFinish AS DATE),
    o.IDProduct, p.ProductCode,
    pc.ProductConsumptionId, pc.MaterialConsumption, pc.MaterialConsumptionGR
ORDER BY ProductionDay, p.ProductCode
"""


def _run_query(conn, date_from: datetime.date, date_to: datetime.date,
               product_code: str = '') -> list[dict]:
    """Execute the range query.  date_from/date_to are inclusive calendar dates."""
    start = datetime.datetime.combine(date_from, datetime.time(7, 30, 0))
    end   = datetime.datetime.combine(date_to + datetime.timedelta(days=1),
                                      datetime.time(7, 30, 0))
    params: list = [start, end]
    if product_code.strip():
        code_filter = "AND p.ProductCode LIKE ?"
        params.append(f'%{product_code.strip()}%')
    else:
        code_filter = ''

    sql = _Q_CONSUMPTION_RANGE.format(code_filter=code_filter)
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


# ─── Excel export ─────────────────────────────────────────────────────────────

def _export_excel(rows: list[dict], date_from: datetime.date,
                  date_to: datetime.date) -> str:
    """Build Excel file, save to C:\\temp, return path."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export.") from exc

    out_dir = r'C:\temp'
    os.makedirs(out_dir, exist_ok=True)
    filename = (
        f"Raport_consum_aliaj_flux_per_perioada_"
        f"{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    )
    path = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alloy Flux Consumption'

    H_FILL   = PatternFill('solid', fgColor='1F3864')
    H_FONT   = Font(bold=True, color='FFFFFF', size=10)
    DAY_FILL = PatternFill('solid', fgColor='D6E4F0')
    DAY_FONT = Font(bold=True, size=10, color='1F3864')
    RED_FONT = Font(color='CC0000', size=9)
    TOT_FILL = PatternFill('solid', fgColor='FFF3CD')
    TOT_FONT = Font(bold=True, size=10)
    ALT_FILL = PatternFill('solid', fgColor='F4F6F8')
    THIN_SIDE = Side(style='thin', color='C0C0C0')
    THIN = Border(left=THIN_SIDE, right=THIN_SIDE,
                  top=THIN_SIDE,  bottom=THIN_SIDE)

    # Title rows
    ws['A1'] = 'Alloy / Flux Consumption Report — PTHM Phase'
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:F1')
    ws['A2'] = (f'Period: {date_from.strftime("%d/%m/%Y")} → '
                f'{date_to.strftime("%d/%m/%Y")}')
    ws['A2'].font = Font(size=9, color='7F8C8D')
    ws.merge_cells('A2:F2')

    # Column headers (row 4)
    headers = ['Day', 'Product Code', 'Qty Processed',
               'Unit Weight GR (Alloy)', 'Partial Total GR', 'Status']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(4, ci, h)
        c.fill, c.font, c.border = H_FILL, H_FONT, THIN
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[4].height = 22

    # Group by day
    grouped: dict[str, list[dict]] = {}
    for r in rows:
        key = str(r['ProductionDay'])
        grouped.setdefault(key, []).append(r)

    excel_row = 5
    total_with    = 0
    total_without = 0

    for day_str in sorted(grouped.keys()):
        day_rows = grouped[day_str]
        try:
            day_obj = datetime.date.fromisoformat(day_str)
            day_label = day_obj.strftime('%d/%m/%Y')
        except Exception:
            day_label = day_str

        # Day separator row
        for ci in range(1, 7):
            c = ws.cell(excel_row, ci)
            c.fill = DAY_FILL
            if ci == 1:
                c.value = day_label
                c.font  = DAY_FONT
            c.border = THIN
        ws.merge_cells(
            start_row=excel_row, start_column=1,
            end_row=excel_row,   end_column=6
        )
        excel_row += 1

        day_total_gr = 0.0
        for i, r in enumerate(day_rows):
            missing = bool(r.get('MissingAlloyConsumption', 0))
            unit_gr = r.get('MaterialConsumptionGR') or 0.0
            qty     = r.get('QtyProcessed') or 0
            partial = qty * unit_gr
            day_total_gr += partial

            if missing:
                total_without += 1
                status_val = 'MISSING'
                row_font   = RED_FONT
            else:
                total_with += 1
                status_val = 'OK'
                row_font   = Font(size=9)

            fill = ALT_FILL if i % 2 == 0 else None

            vals = [
                '',           # day column blank on data rows
                r['ProductCode'],
                qty,
                unit_gr if not missing else 0,
                partial if not missing else 0,
                status_val,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(excel_row, ci, v)
                cell.font   = row_font
                cell.border = THIN
                if fill:
                    cell.fill = fill
                if ci in (3, 4, 5):
                    cell.alignment = Alignment(horizontal='center')
                if ci in (4, 5) and isinstance(v, float):
                    cell.number_format = '0.00'
            excel_row += 1

        # Day total row
        ws.cell(excel_row, 4, 'Day total GR:').font = TOT_FONT
        ws.cell(excel_row, 4).fill   = TOT_FILL
        ws.cell(excel_row, 4).border = THIN
        ws.cell(excel_row, 5, round(day_total_gr, 2)).font = TOT_FONT
        ws.cell(excel_row, 5).fill   = TOT_FILL
        ws.cell(excel_row, 5).number_format = '0.00'
        ws.cell(excel_row, 5).border = THIN
        excel_row += 1

    # Summary section
    excel_row += 1
    SUMM_FILL = PatternFill('solid', fgColor='E8F5E9')
    SUMM_FONT = Font(bold=True, size=10, color='1B5E20')

    for label, val in [
        ('Product codes WITH weight data:',    total_with),
        ('Product codes WITHOUT weight data:', total_without),
        ('Difference (with − without):',       total_with - total_without),
    ]:
        ws.cell(excel_row, 4, label).font  = SUMM_FONT
        ws.cell(excel_row, 4).fill  = SUMM_FILL
        ws.cell(excel_row, 4).border = THIN
        ws.cell(excel_row, 5, val).font   = SUMM_FONT
        ws.cell(excel_row, 5).fill  = SUMM_FILL
        ws.cell(excel_row, 5).border = THIN
        ws.cell(excel_row, 5).alignment = Alignment(horizontal='center')
        excel_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.freeze_panes = 'A5'

    wb.save(path)
    return path


# ─── GUI ──────────────────────────────────────────────────────────────────────

class MaterialConsumptionReportWindow(tk.Toplevel):
    """Interactive consumption report form."""

    _COL_IDS   = ('day', 'code', 'qty', 'unit_gr', 'total_gr', 'status')
    _COL_HEADS = ('Day', 'Product Code', 'Qty Processed',
                  'Unit Weight (gr)', 'Total Weight (gr)', 'Status')
    _COL_WIDTHS = (90, 200, 90, 110, 110, 80)

    def __init__(self, master, db, lang: dict):
        super().__init__(master)
        self.db   = db
        self.lang = lang
        self._rows: list[dict] = []

        self.title(self.lang.get('mcr_window_title', 'Alloy / Flux Consumption Report'))
        self.resizable(True, True)
        self.geometry('950x680')
        self.minsize(800, 500)

        self._build_ui()
        self.grab_set()

    # ── UI construction ──────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header with logo ──
        header = tk.Frame(self, bg='#1F3864')
        header.pack(fill=tk.X)

        logo_path = os.path.join(os.path.dirname(__file__), 'Logo.png')
        try:
            from PIL import Image as PILImage, ImageTk
            img = PILImage.open(logo_path)
            img.thumbnail((120, 40))
            self._logo_img = ImageTk.PhotoImage(img)
            tk.Label(header, image=self._logo_img,
                     bg='#1F3864').pack(side=tk.LEFT, padx=14, pady=8)
        except Exception:
            pass

        tk.Label(
            header,
            text=self.lang.get('mcr_window_title', 'Alloy / Flux Consumption Report'),
            bg='#1F3864', fg='white',
            font=('Helvetica', 13, 'bold')
        ).pack(side=tk.LEFT, padx=8, pady=12)

        # ── Email scheduling info label ──
        info_bar = tk.Frame(self, bg='#FFF8E1', bd=0)
        info_bar.pack(fill=tk.X)
        tk.Label(
            info_bar,
            text=self.lang.get(
                'mcr_email_info',
                'Automated daily email: key "Sys_missing_data_alloy" in Settings  |  '
                'Schedule: Mon–Sat at 08:05'
            ),
            bg='#FFF8E1', fg='#795548',
            font=('Helvetica', 8), anchor='w', padx=10
        ).pack(fill=tk.X, pady=2)

        # ── Filter panel ──
        filter_frame = ttk.LabelFrame(
            self,
            text=self.lang.get('mcr_filter_title', 'Search filters')
        )
        filter_frame.pack(fill=tk.X, padx=10, pady=6)

        # Date from
        ttk.Label(filter_frame,
                  text=self.lang.get('mcr_date_from', 'From (dd/mm/yyyy):')).grid(
            row=0, column=0, sticky='w', padx=8, pady=6)
        self._var_date_from = tk.StringVar(
            value=(datetime.date.today() - datetime.timedelta(days=7)).strftime('%d/%m/%Y')
        )
        ttk.Entry(filter_frame, textvariable=self._var_date_from, width=13).grid(
            row=0, column=1, padx=4, pady=6)

        # Date to
        ttk.Label(filter_frame,
                  text=self.lang.get('mcr_date_to', 'To (dd/mm/yyyy):')).grid(
            row=0, column=2, sticky='w', padx=8, pady=6)
        self._var_date_to = tk.StringVar(
            value=datetime.date.today().strftime('%d/%m/%Y')
        )
        ttk.Entry(filter_frame, textvariable=self._var_date_to, width=13).grid(
            row=0, column=3, padx=4, pady=6)

        # Product code filter
        ttk.Label(filter_frame,
                  text=self.lang.get('mcr_product_code', 'Product code (optional):')).grid(
            row=0, column=4, sticky='w', padx=8, pady=6)
        self._var_code = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self._var_code, width=18).grid(
            row=0, column=5, padx=4, pady=6)

        # Search button
        ttk.Button(
            filter_frame,
            text=self.lang.get('mcr_btn_search', '🔍 Search'),
            command=self._do_search
        ).grid(row=0, column=6, padx=12, pady=6)

        # ── Treeview ──
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        self._tree = ttk.Treeview(
            tree_frame,
            columns=self._COL_IDS,
            show='headings',
            selectmode='browse'
        )
        for cid, head, width in zip(self._COL_IDS, self._COL_HEADS, self._COL_WIDTHS):
            self._tree.heading(cid, text=head)
            anchor = 'w' if cid == 'code' else 'center'
            self._tree.column(cid, width=width, anchor=anchor, stretch=(cid == 'code'))

        # Tag for missing rows
        self._tree.tag_configure('missing', foreground='#CC0000')
        self._tree.tag_configure('day_header',
                                 background='#D6E4F0',
                                 font=('Helvetica', 9, 'bold'))
        self._tree.tag_configure('day_total',
                                 background='#FFF9C4',
                                 font=('Helvetica', 9, 'bold'))

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',
                             command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal',
                             command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # ── Summary bar ──
        self._summary_frame = tk.Frame(self, bg='#E8F5E9', bd=1, relief='groove')
        self._summary_frame.pack(fill=tk.X, padx=10, pady=2)
        self._lbl_with    = tk.Label(self._summary_frame, bg='#E8F5E9',
                                     font=('Helvetica', 9, 'bold'), fg='#1B5E20',
                                     anchor='w')
        self._lbl_without = tk.Label(self._summary_frame, bg='#E8F5E9',
                                     font=('Helvetica', 9, 'bold'), fg='#CC0000',
                                     anchor='w')
        self._lbl_diff    = tk.Label(self._summary_frame, bg='#E8F5E9',
                                     font=('Helvetica', 9, 'bold'), fg='#1F3864',
                                     anchor='w')
        for lbl in (self._lbl_with, self._lbl_without, self._lbl_diff):
            lbl.pack(side=tk.LEFT, padx=18, pady=4)

        # ── Bottom button bar ──
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill=tk.X, padx=10, pady=6)

        self._btn_export = ttk.Button(
            btn_bar,
            text=self.lang.get('mcr_btn_export_excel', '📊 Export Excel'),
            command=self._do_export,
            state='disabled'
        )
        self._btn_export.pack(side=tk.LEFT, padx=4)

        ttk.Button(
            btn_bar,
            text=self.lang.get('btn_close', 'Close'),
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=4)

    # ── Search logic ─────────────────────────────────────────────────────────

    def _parse_date(self, text: str) -> datetime.date | None:
        text = text.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        return None

    def _do_search(self):
        date_from = self._parse_date(self._var_date_from.get())
        date_to   = self._parse_date(self._var_date_to.get())

        if not date_from or not date_to:
            messagebox.showwarning(
                self.lang.get('warning', 'Warning'),
                self.lang.get('mcr_invalid_dates',
                               'Please enter valid dates in dd/mm/yyyy format.'),
                parent=self
            )
            return
        if date_from > date_to:
            messagebox.showwarning(
                self.lang.get('warning', 'Warning'),
                self.lang.get('mcr_date_order', '"From" date must be ≤ "To" date.'),
                parent=self
            )
            return

        try:
            self._rows = _run_query(
                self.db.conn, date_from, date_to,
                self._var_code.get()
            )
        except Exception as exc:
            logger.error(f'MaterialConsumptionReportWindow search: {exc}', exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Error'),
                f'Query error:\n{exc}',
                parent=self
            )
            return

        self._date_from = date_from
        self._date_to   = date_to
        self._populate_tree()

    # ── Tree population ──────────────────────────────────────────────────────

    def _populate_tree(self):
        for item in self._tree.get_children():
            self._tree.delete(item)

        if not self._rows:
            self._tree.insert('', 'end', values=(
                '', self.lang.get('mcr_no_results', 'No data found for the selected period.'),
                '', '', '', ''
            ))
            self._update_summary(0, 0)
            self._btn_export.config(state='disabled')
            return

        # Group by day
        grouped: dict[str, list[dict]] = {}
        for r in self._rows:
            key = str(r['ProductionDay'])
            grouped.setdefault(key, []).append(r)

        total_with    = 0
        total_without = 0

        for day_str in sorted(grouped.keys()):
            day_rows = grouped[day_str]
            try:
                day_obj   = datetime.date.fromisoformat(day_str)
                day_label = day_obj.strftime('%d/%m/%Y — %A')
            except Exception:
                day_label = day_str

            # Day separator
            self._tree.insert('', 'end',
                               values=(day_label, '', '', '', '', ''),
                               tags=('day_header',))

            day_total = 0.0
            for r in day_rows:
                missing = bool(r.get('MissingAlloyConsumption', 0))
                unit_gr = r.get('MaterialConsumptionGR') or 0.0
                qty     = r.get('QtyProcessed') or 0
                partial = qty * unit_gr if not missing else 0.0
                day_total += partial

                if missing:
                    total_without += 1
                    tag    = 'missing'
                    u_disp = '0.00'
                    t_disp = '0.00'
                    status = 'MISSING'
                else:
                    total_with += 1
                    tag    = ''
                    u_disp = f'{unit_gr:.2f}'
                    t_disp = f'{partial:.2f}'
                    status = 'OK'

                self._tree.insert('', 'end', values=(
                    '',
                    r['ProductCode'],
                    qty,
                    u_disp,
                    t_disp,
                    status,
                ), tags=(tag,) if tag else ())

            # Day total row
            self._tree.insert('', 'end', values=(
                '', '', '',
                self.lang.get('mcr_day_total', 'Day total (gr):'),
                f'{day_total:.2f}', ''
            ), tags=('day_total',))

        self._update_summary(total_with, total_without)
        self._btn_export.config(state='normal')

    # ── Summary update ───────────────────────────────────────────────────────

    def _update_summary(self, with_w: int, without_w: int):
        diff = with_w - without_w
        self._lbl_with.config(
            text=self.lang.get('mcr_summary_with',
                                'Codes WITH weight: {0}').format(with_w)
        )
        self._lbl_without.config(
            text=self.lang.get('mcr_summary_without',
                                'Codes WITHOUT weight: {0}').format(without_w)
        )
        self._lbl_diff.config(
            text=self.lang.get('mcr_summary_diff',
                                'Difference (with − without): {0}').format(diff)
        )

    # ── Excel export ─────────────────────────────────────────────────────────

    def _do_export(self):
        if not self._rows:
            messagebox.showinfo(
                self.lang.get('info', 'Info'),
                self.lang.get('mcr_no_data_export',
                               'No data to export. Run a search first.'),
                parent=self
            )
            return
        try:
            path = _export_excel(self._rows, self._date_from, self._date_to)
        except Exception as exc:
            logger.error(f'MaterialConsumptionReportWindow export: {exc}', exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Error'),
                f'Export error:\n{exc}',
                parent=self
            )
            return

        messagebox.showinfo(
            self.lang.get('success', 'Success'),
            self.lang.get('mcr_export_ok', 'File saved:\n{0}').format(path),
            parent=self
        )
        try:
            os.startfile(path)
        except Exception:
            pass
