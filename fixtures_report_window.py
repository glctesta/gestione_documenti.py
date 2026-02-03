# -*- coding: utf-8 -*-
"""
Finestra per i rapporti sullo stato dei fixtures
"""
import tkinter as tk
from tkinter import ttk, messagebox
import logging
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class FixturesReportWindow(tk.Toplevel):
    """Finestra per visualizzare i rapporti sullo stato dei fixtures"""
    
    def __init__(self, parent, db, lang):
        super().__init__(parent)
        self.db = db
        self.lang = lang
        self.parent = parent
        
        # Variabili per ordinamento multiplo
        self.sort_columns = []  # Lista di tuple (column, reverse)
        
        self.title(self.lang.get('fixtures_report_title', 'Rapporti Fixtures'))
        self.geometry('1400x800')
        self.transient(parent)
        
        self._create_widgets()
        
        # Carica dati in background con progress bar
        self.after(100, self._load_data_with_progress)
    
    def _create_widgets(self):
        """Crea i widget della finestra"""
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame filtri
        self._create_filters_frame(main_frame)
        
        # Frame treeview
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Treeview
        columns = ('Equipment', 'QtyScan', 'CriticalOverEndOfLife', 'MaintenanceMessage', 
                   'LastMaintenanceDate', 'NoCycle')
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='browse')
        
        # Intestazioni con binding per ordinamento
        headers = {
            'Equipment': self.lang.get('col_equipment', 'Equipaggiamento'),
            'QtyScan': self.lang.get('col_qty_scan', 'Cicli Totali'),
            'CriticalOverEndOfLife': self.lang.get('col_critical_status', 'Stato Critico'),
            'MaintenanceMessage': self.lang.get('col_maintenance_msg', 'Messaggio Manutenzione'),
            'LastMaintenanceDate': self.lang.get('col_last_maint', 'Ultima Manutenzione'),
            'NoCycle': self.lang.get('col_no_cycle', 'Cicli Previsti')
        }
        
        for col, header_text in headers.items():
            self.tree.heading(col, text=header_text, 
                            command=lambda c=col: self._on_column_click(c))
        
        # Larghezze colonne
        self.tree.column('Equipment', width=250, anchor=tk.W)
        self.tree.column('QtyScan', width=100, anchor=tk.CENTER)
        self.tree.column('CriticalOverEndOfLife', width=350, anchor=tk.W)
        self.tree.column('MaintenanceMessage', width=250, anchor=tk.W)
        self.tree.column('LastMaintenanceDate', width=150, anchor=tk.CENTER)
        self.tree.column('NoCycle', width=120, anchor=tk.CENTER)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Tag per colorazione
        self.tree.tag_configure('danger', background='#f8d7da')
        self.tree.tag_configure('critical', background='#fff3cd')
        self.tree.tag_configure('normal', background='#d4edda')
        
        # Frame info ordinamento
        self.sort_info_label = ttk.Label(main_frame, text='', foreground='blue')
        self.sort_info_label.pack(pady=(5, 0))
    
    def _create_filters_frame(self, parent):
        """Crea il frame dei filtri"""
        filters_frame = ttk.LabelFrame(parent, text=self.lang.get('filters', 'Filtri'), padding=10)
        filters_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Prima riga
        row1 = ttk.Frame(filters_frame)
        row1.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(row1, text=self.lang.get('equipment_filter', 'Equipaggiamento:')).pack(side=tk.LEFT, padx=(0, 5))
        self.equipment_filter = ttk.Entry(row1, width=30)
        self.equipment_filter.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(row1, text=self.lang.get('no_cycle_filter', 'Cicli Previsti:')).pack(side=tk.LEFT, padx=(0, 5))
        self.no_cycle_filter = ttk.Entry(row1, width=15)
        self.no_cycle_filter.pack(side=tk.LEFT, padx=(0, 15))
        
        # Bottoni
        ttk.Button(row1, text=self.lang.get('btn_apply_filters', 'Applica Filtri'), 
                  command=self._apply_filters).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text=self.lang.get('btn_reset', 'Reset'), 
                  command=self._reset_filters_and_sort).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text=self.lang.get('btn_export_excel', 'Esporta Excel'), 
                  command=self._export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text=self.lang.get('btn_close', 'Chiudi'), 
                  command=self.destroy).pack(side=tk.RIGHT, padx=5)
    
    def _load_data_with_progress(self):
        """Carica i dati mostrando una progress bar"""
        # Crea finestra di progresso
        progress_window = tk.Toplevel(self)
        progress_window.title(self.lang.get('loading', 'Caricamento...'))
        progress_window.geometry('400x120')
        progress_window.transient(self)
        progress_window.grab_set()
        
        # Centra la finestra
        progress_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - progress_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - progress_window.winfo_height()) // 2
        progress_window.geometry(f'+{x}+{y}')
        
        # Frame contenuto
        frame = ttk.Frame(progress_window, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Label
        label = ttk.Label(frame, text=self.lang.get('loading_fixtures_data', 'Caricamento dati fixtures in corso...'))
        label.pack(pady=(0, 10))
        
        # Progress bar indeterminata
        progress = ttk.Progressbar(frame, mode='indeterminate', length=350)
        progress.pack(pady=10)
        progress.start(10)
        
        # Esegui caricamento dopo un breve delay per permettere alla progress bar di apparire
        def do_load():
            try:
                # Esegui query (deve essere nel thread principale per pyodbc)
                rows = self._fetch_data()
                
                # Salva dati
                self.original_data = rows
                
                # Popola tree
                self._populate_tree(rows)
                
                logger.info(f"Caricati {len(rows)} record per fixtures report")
                
            except Exception as e:
                logger.error(f"Errore caricamento dati fixtures report: {e}", exc_info=True)
                messagebox.showerror(
                    self.lang.get('error', 'Errore'),
                    f"Errore caricamento dati: {e}",
                    parent=self
                )
            finally:
                # Chiudi finestra progresso
                try:
                    progress_window.destroy()
                except:
                    pass
        
        # Avvia caricamento dopo 200ms per permettere alla UI di aggiornarsi
        self.after(200, do_load)
    
    def _fetch_data(self, filters=None):
        """Esegue la query e restituisce i dati (da chiamare in thread)"""
        query = """
        WITH EquipmentBaseData AS (
            -- Questa CTE calcola QtyScan e CylesOverEquipmentLife una sola volta per ogni EquipmentId e CompitoId
            SELECT
                K.EquipmentId,
                K.Equipment,
                K.QtyScan,
                K.ProductionDate,
                K.EndOfLifeCycle,
                K.QtyScan - K.EndOfLifeCycle AS CylesOverEquipmentLife,
                K.CompitoId
            FROM
                (
                    SELECT
                        A.EquipmentId,
                        A.Equipment + ' [' + A.EquipmentType + ']' AS Equipment,
                        A.ProductionDate,
                        A.EndOfLifeCycle,
                        SUM(CASE WHEN A.IsPass = 0 THEN A.QtyBoards ELSE 0 END) + SUM(CASE WHEN A.IsPass = 1 THEN A.QtyBoards ELSE 0 END) AS QtyScan,
                        CM.CompitoId
                    FROM
                        (
                            SELECT DISTINCT
                                PR.ProductCode,
                                P.PhaseName,
                                S.IsPass,
                                COUNT(*) AS QtyBoards,
                                E.InternalName + ' S/N:' + E.SerialNumber AS Equipment,
                                ET.EquipmentType,
                                E.EquipmentId,
                                ISNULL(ET.EndOfLifeCycle, 100000) AS EndOfLifeCycle,
                                ISNULL(E.ProductionYear, 2015) AS ProductionDate
                            FROM
                                Traceability_rs.dbo.Scannings S
                            INNER JOIN Traceability_rs.dbo.OrderPhases OP ON OP.IDOrderPhase = S.IDOrderPhase
                            INNER JOIN Traceability_rs.dbo.Phases P ON P.IDPhase = OP.IDPhase
                            INNER JOIN Traceability_rs.dbo.Boards B ON B.IDBoard = S.IDBoard
                            INNER JOIN Traceability_rs.dbo.Orders O ON O.IDOrder = B.IDOrder
                            INNER JOIN Traceability_rs.dbo.Products PR ON PR.IDProduct = O.IDProduct
                            INNER JOIN Traceability_rs.eqp.EquipmentFixtures EF ON EF.IdProduct = PR.IDProduct
                            INNER JOIN Traceability_rs.eqp.Equipments E ON E.EquipmentId = EF.EquipmentId
                            INNER JOIN Traceability_rs.eqp.EquipmentBrands EB ON EB.EquipmentBrandId = E.BrandId
                            INNER JOIN Traceability_rs.eqp.EquipmentTypes ET ON ET.EquipmentTypeId = E.EquipmentTypeId
                            INNER JOIN Traceability_RS.eqp.EquipmentFixtureRules EFR ON EFR.EquipmentTypeId = ET.EquipmentTypeId
                            WHERE
                                S.ScanTimeFinish BETWEEN DATEFROMPARTS(E.ProductionYear, 1, 1) AND GETDATE()
                                AND P.IDPhase IN (102, 103)
                            GROUP BY
                                PR.ProductCode,
                                P.PhaseName,
                                S.IsPass,
                                E.InternalName + ' S/N:' + E.SerialNumber,
                                ISNULL(ET.EndOfLifeCycle, 100000),
                                ET.EquipmentType,
                                E.EquipmentId,
                                ISNULL(E.ProductionYear, 2015)
                        ) A
                    INNER JOIN Traceability_rs.eqp.CompitiManutenzione CM ON CM.EquipmentId = A.EquipmentId
                    GROUP BY
                        A.Equipment,
                        A.EquipmentType,
                        A.EquipmentId,
                        A.EndOfLifeCycle,
                        A.ProductionDate,
                        CM.CompitoId
                ) K
        ),
        LogAndCompitiData AS (
            SELECT
                EBD.EquipmentId,
                EBD.CompitoId,
                C.Categoria,
                C.DescrizioneCompito,
                LM.DateStop,
                LM.UserName,
                LM.NoteGenerali,
                P.NoCycle
            FROM
                EquipmentBaseData EBD
            LEFT JOIN Traceability_RS.eqp.LogManutenzioni LM
                ON LM.EquipmentID = EBD.EquipmentId
                AND LM.CompitoId = EBD.CompitoId
            LEFT JOIN Traceability_RS.eqp.CompitiManutenzione C
                ON C.CompitoId = EBD.CompitoId
                AND C.EquipmentId = EBD.EquipmentId
            LEFT JOIN Traceability_RS.eqp.ProgrammedInterventions P
                ON P.ProgrammedInterventionId = C.ProgrammedInterventionId
        )
        SELECT DISTINCT
            EBD.Equipment,
            EBD.QtyScan,
            case    
                when QtyScan < 0.9 * EndOfLifeCycle
                    then 'Missing ' + FORMAT((0.9*EndOfLifeCycle) -QtyScan ,'#,#') + ' to hit End of life 95% [End of life set at:' + format(endoflifecycle,'#,#]')
                When QtyScan > 0.9 * EndOfLifeCycle AND QtyScan<=EndOfLifeCycle
                    then 'Critical ' + format(EndOfLifeCycle-QtyScan,'#,#') + ', equipment on RED ZONE [End of life set at:'+ format(endoflifecycle,'#,#]')
                when QtyScan > EndOfLifeCycle
                then 'Danger ZONE ' + cast(qtyscan-EndOfLifeCycle as nvarchar) + ' OVER End Of Life Tool [End of life set at:' + format(endoflifecycle,'#,#]')
            END AS CriticalOverEndOfLife,
            case
                when QtyScan >= Nocycle
                    then lcd.categoria + ' MUST DONE'
                when Qtyscan <Nocycle
                    then lcd.Categoria + ' Doesn''t need'
            END AS MaintenanceMessage,
            LCD.DateStop AS LastMaintenanceDate,
            LCD.NoCycle
        FROM
            EquipmentBaseData EBD
        LEFT JOIN LogAndCompitiData LCD
            ON LCD.EquipmentId = EBD.EquipmentId
            AND LCD.CompitoId = EBD.CompitoId
        """
        
        # Aggiungi filtri WHERE se presenti
        where_clauses = []
        params = []
        
        if filters:
            if filters.get('equipment'):
                where_clauses.append("EBD.Equipment LIKE ?")
                params.append(f"%{filters['equipment']}%")
            
            if filters.get('no_cycle'):
                where_clauses.append("LCD.NoCycle = ?")
                params.append(filters['no_cycle'])
        
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        
        query += " ORDER BY EBD.QtyScan DESC, EBD.Equipment, LCD.DateStop DESC, LCD.NoCycle"
        
        self.db.cursor.execute(query, params)
        return self.db.cursor.fetchall()
    
    def _load_data(self, filters=None):
        """Carica i dati dalla query"""
        try:
            query = """
            WITH EquipmentBaseData AS (
                -- Questa CTE calcola QtyScan e CylesOverEquipmentLife una sola volta per ogni EquipmentId e CompitoId
                SELECT
                    K.EquipmentId,
                    K.Equipment,
                    K.QtyScan,
                    K.ProductionDate,
                    K.EndOfLifeCycle,
                    K.QtyScan - K.EndOfLifeCycle AS CylesOverEquipmentLife,
                    K.CompitoId
                FROM
                    (
                        SELECT
                            A.EquipmentId,
                            A.Equipment + ' [' + A.EquipmentType + ']' AS Equipment,
                            A.ProductionDate,
                            A.EndOfLifeCycle,
                            SUM(CASE WHEN A.IsPass = 0 THEN A.QtyBoards ELSE 0 END) + SUM(CASE WHEN A.IsPass = 1 THEN A.QtyBoards ELSE 0 END) AS QtyScan,
                            CM.CompitoId
                        FROM
                            (
                                SELECT DISTINCT
                                    PR.ProductCode,
                                    P.PhaseName,
                                    S.IsPass,
                                    COUNT(*) AS QtyBoards,
                                    E.InternalName + ' S/N:' + E.SerialNumber AS Equipment,
                                    ET.EquipmentType,
                                    E.EquipmentId,
                                    ISNULL(ET.EndOfLifeCycle, 100000) AS EndOfLifeCycle,
                                    ISNULL(E.ProductionYear, 2015) AS ProductionDate
                                FROM
                                    Traceability_rs.dbo.Scannings S
                                INNER JOIN Traceability_rs.dbo.OrderPhases OP ON OP.IDOrderPhase = S.IDOrderPhase
                                INNER JOIN Traceability_rs.dbo.Phases P ON P.IDPhase = OP.IDPhase
                                INNER JOIN Traceability_rs.dbo.Boards B ON B.IDBoard = S.IDBoard
                                INNER JOIN Traceability_rs.dbo.Orders O ON O.IDOrder = B.IDOrder
                                INNER JOIN Traceability_rs.dbo.Products PR ON PR.IDProduct = O.IDProduct
                                INNER JOIN Traceability_rs.eqp.EquipmentFixtures EF ON EF.IdProduct = PR.IDProduct
                                INNER JOIN Traceability_rs.eqp.Equipments E ON E.EquipmentId = EF.EquipmentId
                                INNER JOIN Traceability_rs.eqp.EquipmentBrands EB ON EB.EquipmentBrandId = E.BrandId
                                INNER JOIN Traceability_rs.eqp.EquipmentTypes ET ON ET.EquipmentTypeId = E.EquipmentTypeId
                                INNER JOIN Traceability_RS.eqp.EquipmentFixtureRules EFR ON EFR.EquipmentTypeId = ET.EquipmentTypeId
                                WHERE
                                    S.ScanTimeFinish BETWEEN DATEFROMPARTS(E.ProductionYear, 1, 1) AND GETDATE()
                                    AND P.IDPhase IN (102, 103)
                                GROUP BY
                                    PR.ProductCode,
                                    P.PhaseName,
                                    S.IsPass,
                                    E.InternalName + ' S/N:' + E.SerialNumber,
                                    ISNULL(ET.EndOfLifeCycle, 100000),
                                    ET.EquipmentType,
                                    E.EquipmentId,
                                    ISNULL(E.ProductionYear, 2015)
                            ) A
                        INNER JOIN Traceability_rs.eqp.CompitiManutenzione CM ON CM.EquipmentId = A.EquipmentId
                        GROUP BY
                            A.Equipment,
                            A.EquipmentType,
                            A.EquipmentId,
                            A.EndOfLifeCycle,
                            A.ProductionDate,
                            CM.CompitoId
                    ) K
            ),
            LogAndCompitiData AS (
                SELECT
                    EBD.EquipmentId,
                    EBD.CompitoId,
                    C.Categoria,
                    C.DescrizioneCompito,
                    LM.DateStop,
                    LM.UserName,
                    LM.NoteGenerali,
                    P.NoCycle
                FROM
                    EquipmentBaseData EBD
                LEFT JOIN Traceability_RS.eqp.LogManutenzioni LM
                    ON LM.EquipmentID = EBD.EquipmentId
                    AND LM.CompitoId = EBD.CompitoId
                LEFT JOIN Traceability_RS.eqp.CompitiManutenzione C
                    ON C.CompitoId = EBD.CompitoId
                    AND C.EquipmentId = EBD.EquipmentId
                LEFT JOIN Traceability_RS.eqp.ProgrammedInterventions P
                    ON P.ProgrammedInterventionId = C.ProgrammedInterventionId
            )
            SELECT DISTINCT
                EBD.Equipment,
                EBD.QtyScan,
                case    
                    when QtyScan < 0.9 * EndOfLifeCycle
                        then 'Missing ' + FORMAT((0.9*EndOfLifeCycle) -QtyScan ,'#,#') + ' to hit End of life 95% [End of life set at:' + format(endoflifecycle,'#,#]')
                    When QtyScan > 0.9 * EndOfLifeCycle AND QtyScan<=EndOfLifeCycle
                        then 'Critical ' + format(EndOfLifeCycle-QtyScan,'#,#') + ', equipment on RED ZONE [End of life set at:'+ format(endoflifecycle,'#,#]')
                    when QtyScan > EndOfLifeCycle
                    then 'Danger ZONE ' + cast(qtyscan-EndOfLifeCycle as nvarchar) + ' OVER End Of Life Tool [End of life set at:' + format(endoflifecycle,'#,#]')
                END AS CriticalOverEndOfLife,
                case
                    when QtyScan >= Nocycle
                        then lcd.categoria + ' MUST DONE'
                    when Qtyscan <Nocycle
                        then lcd.Categoria + ' Doesn''t need'
                END AS MaintenanceMessage,
                LCD.DateStop AS LastMaintenanceDate,
                LCD.NoCycle
            FROM
                EquipmentBaseData EBD
            LEFT JOIN LogAndCompitiData LCD
                ON LCD.EquipmentId = EBD.EquipmentId
                AND LCD.CompitoId = EBD.CompitoId
            """
            
            # Aggiungi filtri WHERE se presenti
            where_clauses = []
            params = []
            
            if filters:
                if filters.get('equipment'):
                    where_clauses.append("EBD.Equipment LIKE ?")
                    params.append(f"%{filters['equipment']}%")
                
                if filters.get('no_cycle'):
                    where_clauses.append("LCD.NoCycle = ?")
                    params.append(filters['no_cycle'])
            
            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)
            
            query += " ORDER BY EBD.QtyScan DESC, EBD.Equipment, LCD.DateStop DESC, LCD.NoCycle"
            
            self.db.cursor.execute(query, params)
            rows = self.db.cursor.fetchall()
            
            # Salva i dati originali
            self.original_data = rows
            
            # Popola il tree
            self._populate_tree(rows)
            
            logger.info(f"Caricati {len(rows)} record per fixtures report")
            
        except Exception as e:
            logger.error(f"Errore caricamento dati fixtures report: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore caricamento dati: {e}",
                parent=self
            )
    
    def _populate_tree(self, rows):
        """Popola il treeview con i dati"""
        # Pulisci tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Popola
        for row in rows:
            last_maint = row.LastMaintenanceDate.strftime('%d/%m/%Y') if row.LastMaintenanceDate else 'N/A'
            no_cycle = row.NoCycle if row.NoCycle else 'N/A'
            
            # Determina tag per colore
            tag = 'normal'
            if row.CriticalOverEndOfLife and 'Danger ZONE' in row.CriticalOverEndOfLife:
                tag = 'danger'
            elif row.CriticalOverEndOfLife and 'Critical' in row.CriticalOverEndOfLife:
                tag = 'critical'
            
            self.tree.insert('', tk.END, values=(
                row.Equipment,
                row.QtyScan,
                row.CriticalOverEndOfLife or '',
                row.MaintenanceMessage or '',
                last_maint,
                no_cycle
            ), tags=(tag,))
    
    def _on_column_click(self, column):
        """Gestisce il click su una colonna per ordinamento multiplo"""
        # Verifica se la colonna è già nell'elenco di ordinamento
        existing = None
        for i, (col, reverse) in enumerate(self.sort_columns):
            if col == column:
                existing = i
                break
        
        if existing is not None:
            # Inverti l'ordinamento per questa colonna
            col, reverse = self.sort_columns[existing]
            self.sort_columns[existing] = (col, not reverse)
        else:
            # Aggiungi nuova colonna all'ordinamento
            self.sort_columns.append((column, False))
        
        # Applica ordinamento
        self._apply_sorting()
        
        # Aggiorna label info
        self._update_sort_info()
    
    def _apply_sorting(self):
        """Applica l'ordinamento multiplo ai dati"""
        if not self.sort_columns:
            return
        
        # Ottieni tutti gli item
        items = [(self.tree.item(item)['values'], item) for item in self.tree.get_children()]
        
        # Mappa indici colonne
        col_map = {
            'Equipment': 0,
            'QtyScan': 1,
            'CriticalOverEndOfLife': 2,
            'MaintenanceMessage': 3,
            'LastMaintenanceDate': 4,
            'NoCycle': 5
        }
        
        # Ordina in base alle colonne selezionate (in ordine inverso per priorità)
        for column, reverse in reversed(self.sort_columns):
            col_idx = col_map[column]
            
            # Funzione di sort che gestisce valori numerici e stringhe
            def sort_key(item):
                val = item[0][col_idx]
                # Prova a convertire in numero se possibile
                try:
                    return (0, int(val))  # Numeri prima
                except (ValueError, TypeError):
                    try:
                        return (0, float(val))
                    except (ValueError, TypeError):
                        return (1, str(val).lower())  # Stringhe dopo
            
            items.sort(key=sort_key, reverse=reverse)
        
        # Riordina nel tree
        for index, (values, item) in enumerate(items):
            self.tree.move(item, '', index)
    
    def _update_sort_info(self):
        """Aggiorna il label con le info sull'ordinamento"""
        if not self.sort_columns:
            self.sort_info_label.config(text='')
            return
        
        sort_text = self.lang.get('sorted_by', 'Ordinato per: ')
        sort_parts = []
        for col, reverse in self.sort_columns:
            direction = '↓' if reverse else '↑'
            sort_parts.append(f"{col} {direction}")
        
        self.sort_info_label.config(text=sort_text + ', '.join(sort_parts))
    
    def _apply_filters(self):
        """Applica i filtri"""
        filters = {}
        
        equipment = self.equipment_filter.get().strip()
        if equipment:
            filters['equipment'] = equipment
        
        no_cycle = self.no_cycle_filter.get().strip()
        if no_cycle:
            try:
                filters['no_cycle'] = int(no_cycle)
            except ValueError:
                messagebox.showwarning(
                    self.lang.get('warning', 'Attenzione'),
                    self.lang.get('invalid_no_cycle', 'Cicli Previsti deve essere un numero'),
                    parent=self
                )
                return
        
        self._load_data(filters)
    
    def _reset_filters_and_sort(self):
        """Reset filtri e ordinamenti"""
        self.equipment_filter.delete(0, tk.END)
        self.no_cycle_filter.delete(0, tk.END)
        self.sort_columns = []
        self._update_sort_info()
        self._load_data()
    
    def _export_to_excel(self):
        """Esporta i dati in Excel"""
        try:
            # Crea directory se non esiste
            temp_dir = r'c:\Temp'
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            # Nome file con timestamp
            filename = f"FixturesReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            # Crea workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapporti Fixtures"
            
            # Stili
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            danger_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            critical_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            normal_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Intestazioni
            headers = [
                self.lang.get('col_equipment', 'Equipaggiamento'),
                self.lang.get('col_qty_scan', 'Cicli Totali'),
                self.lang.get('col_critical_status', 'Stato Critico'),
                self.lang.get('col_maintenance_msg', 'Messaggio Manutenzione'),
                self.lang.get('col_last_maint', 'Ultima Manutenzione'),
                self.lang.get('col_no_cycle', 'Cicli Previsti')
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Dati
            for item_id in self.tree.get_children():
                values = self.tree.item(item_id, 'values')
                tags = self.tree.item(item_id, 'tags')
                
                row_idx = len(ws['A']) + 1
                
                # Determina colore riga
                if 'danger' in tags:
                    row_fill = danger_fill
                elif 'critical' in tags:
                    row_fill = critical_fill
                else:
                    row_fill = normal_fill
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.fill = row_fill
                    
                    # Allineamento
                    if col_idx in [2, 5, 6]:  # Colonne numeriche
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Larghezze colonne
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            
            # Salva
            wb.save(filepath)
            
            # Chiedi se aprire il file
            response = messagebox.askyesno(
                self.lang.get('success', 'Successo'),
                f"{self.lang.get('export_success', 'Esportazione completata')}:\n{filepath}\n\n{self.lang.get('open_file_question', 'Vuoi aprire il file?')}",
                parent=self
            )
            
            if response:
                # Apri il file con l'applicazione predefinita
                try:
                    os.startfile(filepath)
                    logger.info(f"File Excel aperto: {filepath}")
                except Exception as e:
                    logger.error(f"Errore apertura file Excel: {e}", exc_info=True)
                    messagebox.showwarning(
                        self.lang.get('warning', 'Attenzione'),
                        f"{self.lang.get('cannot_open_file', 'Impossibile aprire il file')}:\n{e}",
                        parent=self
                    )
            
            logger.info(f"Esportazione Excel completata: {filepath}")
            
        except Exception as e:
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror(
                self.lang.get('error', 'Errore'),
                f"Errore esportazione: {e}",
                parent=self
            )


def open_fixtures_report(parent, db, lang):
    """Apre la finestra dei rapporti fixtures"""
    try:
        FixturesReportWindow(parent, db, lang)
    except Exception as e:
        logger.error(f"Errore apertura finestra rapporti fixtures: {e}", exc_info=True)
        messagebox.showerror(
            lang.get('error', 'Errore'),
            f"Errore apertura finestra: {e}",
            parent=parent
        )
