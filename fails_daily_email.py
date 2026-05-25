# -*- coding: utf-8 -*-
"""
fails_daily_email.py
Daily email report for FAIL boards analysis.

Sent every working day (Mon–Sat) at application startup.
Anti-duplication: INSERT WHERE NOT EXISTS on settings table key 'SentFailsReport_YYYYMMDD'.

Email contains:
  - New FAILs opened yesterday
  - Boards repaired / scrapped yesterday
  - Rolling month (1st of month → yesterday) and YTD (1 Jan → yesterday) stats
  - Professional HTML with Logo.png embedded as base64
  - Excel attachment with raw + summary data
"""
from __future__ import annotations
import base64
import datetime
import io
import logging
import os
from collections import Counter
from typing import Any

logger = logging.getLogger("TraceabilityRS")

# ─── SQL queries (all datetime-parameterized, no DECLARE trick needed here) ───

# New FAILs: boards with a FAIL scan result in the period and NO repair scan yet
_Q_NEW_FAILS = """
SELECT DISTINCT
    Boards.IDBoard,
    dbo.BoardLabels(Boards.IDBoard) AS Labels,
    Orders.OrderNumber,
    Products.ProductCode,
    Phases.PhaseName,
    LastScan.ScanTimeFinish AS ScanTime
FROM (
    SELECT Scannings.*
    FROM Scannings
    INNER JOIN Boards ON Boards.IDBoard = Scannings.IDBoard
    WHERE ScanTimeFinish BETWEEN ? AND ?
) A
CROSS APPLY (
    SELECT TOP 1 * FROM Scannings s
    WHERE s.IDBoard = A.IDBoard
    ORDER BY s.IDScan DESC
) LastScan
INNER JOIN Boards  ON Boards.IDBoard  = LastScan.IDBoard
INNER JOIN Orders  ON Orders.IDOrder  = Boards.IDOrder
INNER JOIN Products ON Products.IDProduct = Orders.IDProduct
INNER JOIN OrderPhases ON LastScan.IDOrderPhase = OrderPhases.IDOrderPhase
INNER JOIN Phases  ON Phases.IDPhase  = OrderPhases.IDPhase
LEFT JOIN  ScanDefects ON ScanDefects.IDScan = LastScan.IDScan
WHERE LastScan.IsPass = 0
  AND ScanDefects.IsPass IS NULL
ORDER BY Orders.OrderNumber, Products.ProductCode, Boards.IDBoard
"""

# Repairs & scraps in period (reuses same join logic as QUERY_REPAIRS in fails_analysis_gui)
_Q_REPAIRS = """
SELECT DISTINCT
    Boards.IDBoard,
    dbo.BoardLabels(Boards.IDBoard) AS Labels,
    Products.ProductCode,
    Orders.OrderProduction,
    Phases.PhaseName,
    CASE WHEN ScanDefects.IsPass = 1 THEN 'REPAIRED' ELSE 'SCRAP' END AS ResultRepair,
    ScanDefects.StopTime AS DateRepair,
    Defects.DefectNameRO AS Defect
FROM ScanDefects
INNER JOIN ScanDefectDetails  ON ScanDefects.IDScanDefect       = ScanDefectDetails.IDScanDefect
INNER JOIN DefectsRiferiments ON DefectsRiferiments.IDScanDefectDet = ScanDefectDetails.IDScanDefectDet
INNER JOIN Riferiments        ON Riferiments.IDDibaRiferimento   = DefectsRiferiments.IDDibaRiferimento
INNER JOIN Defects            ON ScanDefectDetails.IDDefect       = Defects.IDDefect
INNER JOIN Scannings          ON Scannings.IDScan                 = ScanDefects.IDScan
INNER JOIN OrderPhases        ON OrderPhases.IDOrderPhase         = Scannings.IDOrderPhase
INNER JOIN Orders             ON OrderPhases.IDOrder              = Orders.IDOrder
INNER JOIN Products           ON Products.IDProduct               = Orders.IDProduct
INNER JOIN Phases             ON OrderPhases.IDPhase              = Phases.IDPhase
INNER JOIN Boards             ON Scannings.IDBoard                = Boards.IDBoard
WHERE ScanDefects.StopTime BETWEEN ? AND ?
ORDER BY ScanDefects.StopTime
"""

# Total boards produced in a period (boxes closed, not re-work lines)
_Q_PRODUCED = """
SELECT ISNULL(SUM(t.cantitate), 0) AS Prodotto
FROM (
    SELECT Products.ProductCode, Orders.OrderProduction, COUNT(*) AS cantitate
    FROM Boxes
    INNER JOIN BoxDetails ON BoxDetails.IDBox     = Boxes.IDBox
    INNER JOIN Boards     ON Boards.IDBoard       = BoxDetails.IDBoard
    INNER JOIN Orders     ON Orders.IDOrder       = Boards.IDOrder
    INNER JOIN Products   ON Products.IDProduct   = Orders.IDProduct
    WHERE Boxes.DateTimeStop BETWEEN ? AND ?
      AND Boxes.IDWorkLine IS NULL
    GROUP BY Products.ProductCode, Orders.OrderProduction
) AS T
"""

# FAIL boards cross-tab: one row per (ProductCode, PhaseName) with count
_Q_FAIL_BREAKDOWN = """
SELECT
    Products.ProductCode,
    Phases.PhaseName,
    COUNT(DISTINCT LastScan.IDBoard) AS FailCount
FROM (
    SELECT DISTINCT IDBoard
    FROM Scannings
    WHERE ScanTimeFinish BETWEEN ? AND ?
) A
CROSS APPLY (
    SELECT TOP 1 IDBoard, IDScan, IDOrderPhase, IsPass
    FROM Scannings s
    WHERE s.IDBoard = A.IDBoard
    ORDER BY s.IDScan DESC
) LastScan
INNER JOIN Boards      ON Boards.IDBoard           = LastScan.IDBoard
INNER JOIN Orders      ON Orders.IDOrder            = Boards.IDOrder
INNER JOIN Products    ON Products.IDProduct        = Orders.IDProduct
INNER JOIN OrderPhases ON OrderPhases.IDOrderPhase  = LastScan.IDOrderPhase
INNER JOIN Phases      ON Phases.IDPhase            = OrderPhases.IDPhase
LEFT  JOIN ScanDefects ON ScanDefects.IDScan        = LastScan.IDScan
WHERE LastScan.IsPass = 0
  AND ScanDefects.IsPass IS NULL
GROUP BY Products.ProductCode, Phases.PhaseName
ORDER BY FailCount DESC, Products.ProductCode, Phases.PhaseName
"""

# ─── Data helpers ─────────────────────────────────────────────────────────────

def _fetch(conn, sql: str, params: tuple) -> list:
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur.fetchall()
    except Exception as e:
        logger.error(f"fails_daily_email _fetch: {e}", exc_info=True)
        return []


def _period_stats(rows: list) -> dict:
    """
    Summarise a list of repair rows.
    row layout: 0=IDBoard,1=Labels,2=ProductCode,3=OrderNumber,
                4=PhaseName,5=ResultRepair,6=DateRepair,7=Defect
    """
    boards     = {r[0] for r in rows}
    repaired   = {r[0] for r in rows if r[5] == 'REPAIRED'}
    scrapped   = {r[0] for r in rows if r[5] == 'SCRAP'}
    phases     = Counter(r[4] for r in rows if r[4])
    products   = Counter(r[2] for r in rows if r[2])
    return {
        'total':    len(boards),
        'repaired': len(repaired),
        'scrap':    len(scrapped),
        'phases':   phases.most_common(10),
        'products': products.most_common(15),
    }


def _fail_stats(rows: list) -> dict:
    """
    Summarise a list of FAIL rows.
    row layout: 0=IDBoard,1=Labels,2=OrderNumber,3=ProductCode,
                4=PhaseName,5=ScanTime
    Deduplicates by IDBoard so totals match the card values.
    """
    # One entry per board — eliminates duplicates caused by multiple
    # scannings per board in the query period.
    seen: dict = {}
    for r in rows:
        if r[0] not in seen:
            seen[r[0]] = r
    unique_rows = list(seen.values())

    phases   = Counter(r[4] for r in unique_rows if r[4])
    products = Counter(r[3] for r in unique_rows if r[3])

    # Top-3 fail phases per product
    prod_phase: dict = {}
    for r in unique_rows:
        prod, phase = r[3], r[4]
        if prod and phase:
            if prod not in prod_phase:
                prod_phase[prod] = Counter()
            prod_phase[prod][phase] += 1
    product_top_phases = {p: c.most_common(3) for p, c in prod_phase.items()}

    return {
        'total':    len(unique_rows),
        'phases':   phases.most_common(10),
        'products': products.most_common(15),
        'product_top_phases': product_top_phases,
    }


# ─── HTML builder ─────────────────────────────────────────────────────────────

def _logo_base64() -> str:
    """Return Logo.png as base64 data URI, or empty string if not found."""
    logo_path = os.path.join(os.path.dirname(__file__), 'Logo.png')
    if not os.path.isfile(logo_path):
        return ''
    try:
        with open(logo_path, 'rb') as f:
            return 'data:image/png;base64,' + base64.b64encode(f.read()).decode()
    except Exception as e:
        logger.warning(f"fails_daily_email: cannot read Logo.png: {e}")
        return ''


def _top_table(items: list, col1: str, col2: str) -> str:
    if not items:
        return '<p style="color:#777;font-size:12px">No data</p>'
    rows = ''.join(
        f'<tr style="background:{"#f9f9f9" if i%2==0 else "#ffffff"}">'
        f'<td style="padding:5px 10px;border:1px solid #e0e0e0">{name}</td>'
        f'<td style="padding:5px 10px;border:1px solid #e0e0e0;text-align:center;font-weight:bold">{cnt}</td>'
        f'</tr>'
        for i, (name, cnt) in enumerate(items)
    )
    return (
        f'<table style="border-collapse:collapse;font-size:11px;width:100%">'
        f'<tr style="background:#1f3864;color:#fff">'
        f'<th style="padding:5px 10px;text-align:left">{col1}</th>'
        f'<th style="padding:5px 10px;text-align:center">{col2}</th>'
        f'</tr>{rows}</table>'
    )


def _product_phase_table(items: list, product_top_phases: dict) -> str:
    """Products table with top-3 fail phases inline per product row."""
    if not items:
        return '<p style="color:#777;font-size:12px">No data</p>'
    rows_html = ''
    for i, (name, cnt) in enumerate(items):
        phases = product_top_phases.get(name, [])
        phase_parts = [
            f'<span style="color:#666">{ph}</span>:<b>{c}</b>'
            for ph, c in phases
        ]
        phase_text = ' &nbsp;·&nbsp; '.join(phase_parts) if phase_parts else '—'
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:4px 10px;border:1px solid #e0e0e0">{name}</td>'
            f'<td style="padding:4px 10px;border:1px solid #e0e0e0;'
            f'text-align:center;font-weight:bold">{cnt}</td>'
            f'<td style="padding:4px 10px;border:1px solid #e0e0e0;'
            f'font-size:10px;white-space:nowrap">{phase_text}</td>'
            f'</tr>'
        )
    return (
        f'<table style="border-collapse:collapse;font-size:11px;width:100%">'
        f'<tr style="background:#1f3864;color:#fff">'
        f'<th style="padding:5px 10px;text-align:left">Product</th>'
        f'<th style="padding:5px 10px;text-align:center">FAILs</th>'
        f'<th style="padding:5px 10px;text-align:left">Top Phases (3)</th>'
        f'</tr>{rows_html}</table>'
    )


def _fail_breakdown_section(rows: list, title: str) -> str:
    """
    Render a product × phase cross-table.
    rows: list of (ProductCode, PhaseName, FailCount)
    """
    if not rows:
        return ''
    phases = list(dict.fromkeys(r[1] for r in rows if r[1]))
    prod_totals: Counter = Counter()
    matrix: dict = {}
    for prod, phase, cnt in rows:
        matrix[(prod, phase)] = cnt
        prod_totals[prod] += cnt
    products_sorted = [p for p, _ in prod_totals.most_common()]

    ph_headers = ''.join(
        f'<th style="padding:4px 8px;background:#1f3864;color:#fff;font-size:10px;'
        f'white-space:nowrap;text-align:center">{ph}</th>'
        for ph in phases
    )
    header = (
        f'<tr>'
        f'<th style="padding:4px 10px;background:#1f3864;color:#fff;font-size:10px;'
        f'text-align:left">Product</th>'
        f'{ph_headers}'
        f'<th style="padding:4px 8px;background:#c0392b;color:#fff;font-size:10px">Total</th>'
        f'</tr>'
    )
    data_rows = ''
    for i, prod in enumerate(products_sorted):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        cells = ''
        for ph in phases:
            cnt = matrix.get((prod, ph), 0)
            if cnt == 0:
                style, val = 'color:#ccc', '—'
            elif cnt >= 10:
                style, val = 'color:#c0392b;font-weight:bold', str(cnt)
            elif cnt >= 5:
                style, val = 'color:#e67e22;font-weight:bold', str(cnt)
            else:
                style, val = 'color:#27ae60', str(cnt)
            cells += (
                f'<td style="padding:3px 8px;text-align:center;border:1px solid #eee;'
                f'{style}">{val}</td>'
            )
        total_cnt = prod_totals[prod]
        data_rows += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:3px 10px;border:1px solid #eee;font-size:11px">{prod}</td>'
            f'{cells}'
            f'<td style="padding:3px 8px;text-align:center;border:1px solid #eee;'
            f'font-weight:bold;color:#c0392b">{total_cnt}</td>'
            f'</tr>'
        )
    return (
        f'<tr><td colspan="2" style="padding:12px 0 4px">'
        f'<b style="font-size:13px;color:#1f3864">{title}</b></td></tr>'
        f'<tr><td colspan="2" style="overflow-x:auto">'
        f'<table style="border-collapse:collapse;font-size:11px;width:100%">'
        f'{header}{data_rows}'
        f'</table></td></tr>'
    )


def _stat_card(title: str, total: int, repaired: int, scrap: int,
               color: str = '#1f3864', produced: int | None = None) -> str:
    pct = f'{repaired/total*100:.1f}' if total else '0.0'
    if produced:
        fail_rate = f'{total/produced*100:.2f}%'
        produced_line = (
            f'<br><span style="color:#888">📦 Produced: <b>{produced:,}</b>'
            f' &nbsp;| FAIL rate: <b style="color:#e67e22">{fail_rate}</b></span>'
        )
    else:
        produced_line = ''
    return f"""
    <td style="padding:8px">
      <div style="background:#fff;border:1px solid #ddd;border-radius:8px;
                  padding:14px 18px;min-width:160px;border-top:4px solid {color}">
        <div style="font-size:11px;color:#666;font-weight:bold;text-transform:uppercase;
                    letter-spacing:0.5px;margin-bottom:8px">{title}</div>
        <div style="font-size:26px;font-weight:bold;color:{color}">{total}</div>
        <div style="font-size:11px;color:#444;margin-top:4px">
          ✅ Repaired: <b>{repaired}</b> &nbsp;|&nbsp;
          🗑️ Scrap: <b>{scrap}</b><br>
          <span style="color:#2ecc71">% Repaired: {pct}%</span>{produced_line}
        </div>
      </div>
    </td>"""


def _build_html(
    yesterday: datetime.date,
    fail_yesterday:   dict,
    repair_yesterday: dict,
    repair_month:     dict,
    repair_ytd:       dict,
    month_label:      str,
    year_label:       str,
    produced_yesterday: int = 0,
    produced_month:     int = 0,
    produced_ytd:       int = 0,
) -> str:
    logo_uri = _logo_base64()
    logo_html = (
        f'<img src="{logo_uri}" alt="Logo" style="height:52px;margin-bottom:4px">'
        if logo_uri else ''
    )
    date_str = yesterday.strftime('%d %B %Y')

    cards = (
        _stat_card(f'New FAILs – {yesterday.strftime("%d/%m")}',
                   fail_yesterday['total'], 0, 0, '#e74c3c',
                   produced=produced_yesterday or None) +
        _stat_card(f'Repairs \u2013 {yesterday.strftime("%d/%m")}',
                   repair_yesterday['total'],
                   repair_yesterday['repaired'],
                   repair_yesterday['scrap'], '#2980b9') +
        _stat_card(f'Month ({month_label})',
                   repair_month['total'],
                   repair_month['repaired'],
                   repair_month['scrap'], '#8e44ad',
                   produced=produced_month or None) +
        _stat_card(f'YTD ({year_label})',
                   repair_ytd['total'],
                   repair_ytd['repaired'],
                   repair_ytd['scrap'], '#27ae60',
                   produced=produced_ytd or None)
    )

    def top_section(title: str, stats: dict, show_repairs: bool) -> str:
        phase_col = 'Repairs' if show_repairs else 'FAILs'
        if not show_repairs and stats.get('product_top_phases'):
            prod_table = _product_phase_table(
                stats['products'], stats['product_top_phases']
            )
        else:
            prod_table = _top_table(stats['products'], 'Product', phase_col)
        return f"""
        <tr><td colspan="2" style="padding:12px 0 4px">
          <b style="font-size:13px;color:#1f3864">{title}</b>
        </td></tr>
        <tr>
          <td style="padding:4px 8px 4px 0;vertical-align:top;width:40%">
            <div style="font-size:11px;color:#555;margin-bottom:4px">
              Top Phases — {phase_col}
            </div>
            {_top_table(stats['phases'], 'Phase', phase_col)}
          </td>
          <td style="padding:4px 0 4px 8px;vertical-align:top;width:60%">
            <div style="font-size:11px;color:#555;margin-bottom:4px">
              Top Products — {phase_col}
            </div>
            {prod_table}
          </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f6f8;font-family:Arial,Helvetica,sans-serif">
<div style="max-width:860px;margin:24px auto;background:#fff;border-radius:10px;
            box-shadow:0 2px 12px rgba(0,0,0,0.10);overflow:hidden">

  <!-- Header -->
  <div style="background:#1f3864;padding:20px 28px;display:flex;align-items:center">
    {logo_html}
    <div style="margin-left:16px">
      <div style="color:#fff;font-size:20px;font-weight:bold">
        FAIL Boards — Daily Report
      </div>
      <div style="color:#a8c4e0;font-size:13px">{date_str}</div>
    </div>
  </div>

  <!-- KPI cards -->
  <div style="padding:20px 24px 4px">
    <div style="font-size:12px;color:#888;margin-bottom:10px;font-weight:bold;
                text-transform:uppercase;letter-spacing:0.5px">Key Indicators</div>
    <table style="border-collapse:collapse"><tr>{cards}</tr></table>
  </div>

  <!-- Detail tables -->
  <div style="padding:10px 24px 20px">
    <table style="width:100%;border-collapse:collapse">
      {top_section(f'Yesterday — New FAILs ({yesterday.strftime("%d/%m/%Y")})',
                   fail_yesterday, False)}
      {top_section(f'Yesterday — Repairs &amp; Scraps ({yesterday.strftime("%d/%m/%Y")})',
                   repair_yesterday, True)}
      {top_section(f'Month to Date ({month_label})', repair_month, True)}
      {top_section(f'Year to Date ({year_label})',   repair_ytd,   True)}
    </table>
  </div>

  <!-- Footer -->
  <div style="background:#f4f6f8;padding:12px 24px;border-top:1px solid #e0e0e0">
    <div style="font-size:10px;color:#999">
      See the attached Excel file for full raw data and pivot summaries.<br>
      Generated automatically by <b>TraceabilityRS</b> —
      {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}
    </div>
  </div>

</div>
</body>
</html>"""


# ─── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(
    yesterday:            datetime.date,
    fail_rows:            list,
    repair_rows:          list,
    month_repair:         list,
    ytd_repair:           list,
    month_label:          str,
    year_label:           str,
    fail_breakdown_ytd:   list | None = None,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not available — skipping Excel attachment")
        return b''

    THIN  = Border(*(Side(style='thin'),)*4)
    H_FILL = PatternFill('solid', fgColor='1F3864')
    H_FONT = Font(bold=True, color='FFFFFF')
    R_FILL = PatternFill('solid', fgColor='FFF2CC')
    S_FILL = PatternFill('solid', fgColor='FADADD')
    F_FILL = PatternFill('solid', fgColor='FCE4D6')

    wb = openpyxl.Workbook()

    def _header(ws, headers: list):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill, cell.font, cell.alignment = H_FILL, H_FONT, Alignment(horizontal='center')
            cell.border = THIN
        ws.row_dimensions[1].height = 18

    def _autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 60)

    # ── Sheet 1: New FAILs yesterday ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f'New FAILs {yesterday.strftime("%d-%m")}'
    _header(ws1, ['Board ID','Labels','Order','Product','Phase','Scan Time'])
    for ri, r in enumerate(fail_rows, 2):
        for ci, v in enumerate([r[0], r[1], r[2], r[3], r[4],
                                 str(r[5])[:19] if r[5] else ''], 1):
            cell = ws1.cell(ri, ci, v)
            cell.fill, cell.border = F_FILL, THIN
    ws1.freeze_panes = 'A2'
    _autofit(ws1)

    # ── Sheet 2: Repairs & Scraps yesterday ───────────────────────────────────
    ws2 = wb.create_sheet(f'Repairs {yesterday.strftime("%d-%m")}')
    _header(ws2, ['Board ID','Labels','Product','Order','Phase','Result','Date Repair','Defect'])
    for ri, r in enumerate(repair_rows, 2):
        result = r[5] or ''
        fill = R_FILL if result == 'REPAIRED' else S_FILL
        for ci, v in enumerate([r[0], r[1], r[2], r[3], r[4], result,
                                 str(r[6])[:19] if r[6] else '', r[7] or ''], 1):
            cell = ws2.cell(ri, ci, v)
            cell.fill, cell.border = fill, THIN
    ws2.freeze_panes = 'A2'
    _autofit(ws2)

    def _pivot_sheet(name, rows):
        """Phase/Product pivot summary."""
        ws = wb.create_sheet(name)
        phases   = Counter(r[4] for r in rows if r[4])
        products = Counter(r[2] for r in rows if r[2])
        repaired = Counter(r[4] for r in rows if r[4] and r[5] == 'REPAIRED')
        scrapped = Counter(r[4] for r in rows if r[4] and r[5] == 'SCRAP')

        ws.cell(1, 1, 'Phases — Summary').font = Font(bold=True, size=12, color='1F3864')
        _header_row = ['Phase', 'Total', 'Repaired', 'Scrap']
        for ci, h in enumerate(_header_row, 1):
            cell = ws.cell(2, ci, h)
            cell.fill, cell.font, cell.border = H_FILL, H_FONT, THIN
        ri = 3
        for ph, cnt in sorted(phases.items(), key=lambda x: -x[1]):
            for ci, v in enumerate([ph, cnt, repaired[ph], scrapped[ph]], 1):
                ws.cell(ri, ci, v).border = THIN
            ri += 1
        ri += 2
        ws.cell(ri, 1, 'Products — Summary').font = Font(bold=True, size=12, color='1F3864')
        ri += 1
        for ci, h in enumerate(['Product', 'Total'], 1):
            cell = ws.cell(ri, ci, h)
            cell.fill, cell.font, cell.border = H_FILL, H_FONT, THIN
        ri += 1
        for pr, cnt in products.most_common(30):
            ws.cell(ri, 1, pr).border = THIN
            ws.cell(ri, 2, cnt).border = THIN
            ri += 1
        ws.freeze_panes = 'A3'
        _autofit(ws)

    _pivot_sheet(f'Month Summary ({month_label})', month_repair)
    _pivot_sheet(f'YTD Summary ({year_label})',    ytd_repair)

    # ── Sheet 5: FAILs by Product × Phase (YTD) ──────────────────────────────
    if fail_breakdown_ytd:
        bd_rows = fail_breakdown_ytd
        phases = list(dict.fromkeys(r[1] for r in bd_rows if r[1]))
        prod_totals_bd: Counter = Counter()
        matrix_bd: dict = {}
        for prod, phase, cnt in bd_rows:
            matrix_bd[(prod, phase)] = cnt
            prod_totals_bd[prod] += cnt
        products_sorted_bd = [p for p, _ in prod_totals_bd.most_common()]

        ws_bd = wb.create_sheet(f'FAIL Breakdown YTD ({year_label})')
        # header row: Product | phase1 | phase2 | ... | Total
        headers_bd = ['Product'] + phases + ['Total']
        _header(ws_bd, headers_bd)
        RED_FILL   = PatternFill('solid', fgColor='FADADD')
        ORG_FILL   = PatternFill('solid', fgColor='FCE4D6')
        GRN_FILL   = PatternFill('solid', fgColor='E2EFDA')
        for ri_bd, prod in enumerate(products_sorted_bd, 2):
            ws_bd.cell(ri_bd, 1, prod).border = THIN
            for ci_bd, ph in enumerate(phases, 2):
                cnt_bd = matrix_bd.get((prod, ph), 0)
                cell = ws_bd.cell(ri_bd, ci_bd, cnt_bd if cnt_bd else None)
                cell.border = THIN
                cell.alignment = Alignment(horizontal='center')
                if cnt_bd >= 10:
                    cell.fill = RED_FILL
                elif cnt_bd >= 5:
                    cell.fill = ORG_FILL
                elif cnt_bd > 0:
                    cell.fill = GRN_FILL
            tot_cell = ws_bd.cell(ri_bd, len(headers_bd), prod_totals_bd[prod])
            tot_cell.border = THIN
            tot_cell.font   = Font(bold=True)
            tot_cell.alignment = Alignment(horizontal='center')
        ws_bd.freeze_panes = 'B2'
        _autofit(ws_bd)

    # ── Sheet 6: Top FAILs by Product × Phase (yesterday) ────────────────────
    if fail_rows:
        seen_fp: dict = {}
        for r in fail_rows:
            if r[0] not in seen_fp:
                seen_fp[r[0]] = r
        fail_rows_uniq = list(seen_fp.values())
        fail_by_prod: dict = {}
        for r in fail_rows_uniq:
            prod, phase = r[3], r[4]
            if prod and phase:
                if prod not in fail_by_prod:
                    fail_by_prod[prod] = {'total': 0, 'phases': Counter()}
                fail_by_prod[prod]['total'] += 1
                fail_by_prod[prod]['phases'][phase] += 1
        prod_sorted_pp = sorted(
            fail_by_prod.items(), key=lambda x: -x[1]['total']
        )
        ws_pp = wb.create_sheet(
            f'Top FAILs by Product {yesterday.strftime("%d-%m")}'
        )
        hdrs_pp = [
            'Product', 'Total FAILs',
            'Phase 1', 'Count 1', 'Phase 2', 'Count 2', 'Phase 3', 'Count 3',
        ]
        _header(ws_pp, hdrs_pp)
        for ri_pp, (prod, data) in enumerate(prod_sorted_pp, 2):
            top3 = data['phases'].most_common(3)
            row_vals: list = [prod, data['total']]
            for ph, c in top3:
                row_vals.extend([ph, c])
            while len(row_vals) < len(hdrs_pp):
                row_vals.append(None)
            for ci_pp, v in enumerate(row_vals, 1):
                cell = ws_pp.cell(ri_pp, ci_pp, v)
                cell.border = THIN
                if ci_pp == 2:
                    cell.font = Font(bold=True)
        ws_pp.freeze_panes = 'A2'
        _autofit(ws_pp)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Anti-duplication claim ────────────────────────────────────────────────────

def _claim_send_slot(conn, setting_key: str) -> bool:
    """
    Returns True if this PC won the race to send today's email
    (INSERT WHERE NOT EXISTS atomic pattern).
    """
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO traceability_rs.dbo.settings (atribute, [value])
            SELECT ?, ?
            WHERE NOT EXISTS (
                SELECT 1 FROM traceability_rs.dbo.settings
                WITH (UPDLOCK, HOLDLOCK) WHERE atribute = ?
            )
        """, (setting_key,
              datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              setting_key))
        claimed = cur.rowcount > 0
        conn.commit()
        cur.close()
        return claimed
    except Exception as e:
        logger.error(f"fails_daily_email _claim_send_slot: {e}")
        return False


# ─── Public entry point ────────────────────────────────────────────────────────

def run_fails_daily_email(db: Any) -> None:
    """
    Main entry point. Call from main.py in a daemon thread.

    Parameters
    ----------
    db : database wrapper with .conn (pyodbc.Connection) and .conn_str (str)
    """
    import utils  # local import to avoid circular imports at module load

    try:
        today     = datetime.date.today()
        yesterday = today - datetime.timedelta(days=1)

        # Mon=0 … Sat=5, Sun=6  → skip Sunday
        if today.weekday() == 6:
            logger.debug("fails_daily_email: Sunday — skip")
            return

        setting_key = f'SentFailsReport_{today.strftime("%Y%m%d")}'  # 24 chars

        conn = db.conn

        # ── Anti-duplication ──────────────────────────────────────────────────
        if not _claim_send_slot(conn, setting_key):
            logger.info(f"fails_daily_email: già inviata per {today} (altra istanza), skip")
            return
        logger.info(f"fails_daily_email: claim acquisito per {today}")

        # ── Recipients ────────────────────────────────────────────────────────
        try:
            recipients = utils.get_email_recipients(conn, 'Sys_Fail_report')
        except Exception as e:
            logger.error(f"fails_daily_email: errore recupero destinatari: {e}")
            return
        if not recipients:
            logger.warning("fails_daily_email: nessun destinatario Sys_fail_report configurato")
            return

        # ── Date ranges ───────────────────────────────────────────────────────
        yest_start = datetime.datetime.combine(yesterday, datetime.time.min)
        yest_end   = datetime.datetime.combine(yesterday, datetime.time.max)
        month_start = datetime.datetime(today.year, today.month, 1)
        ytd_start   = datetime.datetime(today.year, 1, 1)
        month_label = yesterday.strftime('%B %Y')
        year_label  = str(today.year)

        logger.info(f"fails_daily_email: query DB per yesterday={yesterday}")

        # ── Query data ────────────────────────────────────────────────────────
        fail_rows_yest    = _fetch(conn, _Q_NEW_FAILS, (yest_start, yest_end))
        repair_rows_yest  = _fetch(conn, _Q_REPAIRS,  (yest_start, yest_end))
        repair_rows_month = _fetch(conn, _Q_REPAIRS,  (month_start, yest_end))
        repair_rows_ytd   = _fetch(conn, _Q_REPAIRS,  (ytd_start,   yest_end))

        # Boards produced (yesterday, month, YTD) — scalar
        def _produced(start, end) -> int:
            rows = _fetch(conn, _Q_PRODUCED, (start, end))
            return int(rows[0][0]) if rows and rows[0][0] else 0
        produced_yesterday = _produced(yest_start, yest_end)
        produced_month     = _produced(month_start, yest_end)
        produced_ytd       = _produced(ytd_start, yest_end)

        # FAILs by Product × Phase (YTD)
        fail_breakdown_ytd = _fetch(conn, _Q_FAIL_BREAKDOWN, (ytd_start, yest_end))

        # ── Build stats ───────────────────────────────────────────────────────
        fail_stats_y   = _fail_stats(fail_rows_yest)
        repair_stats_y = _period_stats(repair_rows_yest)
        repair_stats_m = _period_stats(repair_rows_month)
        repair_stats_d = _period_stats(repair_rows_ytd)

        logger.info(
            f"fails_daily_email: "
            f"yesterday FAILs={fail_stats_y['total']}, "
            f"repairs={repair_stats_y['repaired']}, "
            f"scrap={repair_stats_y['scrap']}"
        )

        # ── Build HTML ────────────────────────────────────────────────────────
        html = _build_html(
            yesterday,
            fail_stats_y,
            repair_stats_y,
            repair_stats_m,
            repair_stats_d,
            month_label,
            year_label,
            produced_yesterday=produced_yesterday,
            produced_month=produced_month,
            produced_ytd=produced_ytd,
        )

        # ── Build Excel ───────────────────────────────────────────────────────
        excel_bytes = _build_excel(
            yesterday,
            fail_rows_yest,
            repair_rows_yest,
            repair_rows_month,
            repair_rows_ytd,
            month_label,
            year_label,
            fail_breakdown_ytd=fail_breakdown_ytd,
        )

        # ── Send email ────────────────────────────────────────────────────────
        subject = (
            f"FAIL Boards Daily Report — "
            f"{yesterday.strftime('%d/%m/%Y')} — "
            f"FAILs: {fail_stats_y['total']} | "
            f"Repaired: {repair_stats_y['repaired']} | "
            f"Scrap: {repair_stats_y['scrap']}"
        )

        # EmailSender expects file paths, not bytes → write to a temp file
        attachments = []
        tmp_path = None
        if excel_bytes:
            import tempfile
            tmp_fd, tmp_path = tempfile.mkstemp(
                suffix='.xlsx',
                prefix=f'FailBoards_{yesterday.strftime("%Y%m%d")}_'
            )
            try:
                with os.fdopen(tmp_fd, 'wb') as fh:
                    fh.write(excel_bytes)
                attachments = [tmp_path]
            except Exception as ex:
                logger.warning(f"fails_daily_email: cannot write temp Excel: {ex}")
                tmp_path = None

        try:
            utils.send_email(
                recipients, subject, html,
                is_html=True,
                attachments=attachments if attachments else None
            )
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

        logger.info(
            f"fails_daily_email: inviata a {recipients} — "
            f"FAILs={fail_stats_y['total']}, "
            f"Repairs={repair_stats_y['repaired']}, "
            f"Scrap={repair_stats_y['scrap']}"
        )

    except Exception as e:
        logger.error(f"fails_daily_email run_fails_daily_email: {e}", exc_info=True)
